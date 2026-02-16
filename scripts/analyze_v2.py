import openpyxl
from openpyxl.utils import get_column_letter
import os

base = os.path.join('C:/Users/User/OneDrive', chr(193)+'rea de Trabalho', 'CLAUDE CODE')
filepath = os.path.join(base, 'output', 'CRM INTELIGENTE - VITAO360 V2.xlsx')
print('Analyzing:', filepath)
print('File size: {:,} bytes'.format(os.path.getsize(filepath)))
sep = '=' * 120
hsep = '#' * 120
print(sep)

wb = openpyxl.load_workbook(filepath, data_only=False)

print()
print(sep)
print('1. ALL SHEET NAMES')
print(sep)
for i, name in enumerate(wb.sheetnames, 1):
    print('  {}.'.format(i), repr(name))
print('Total sheets:', len(wb.sheetnames))

print()
print(sep)
print('3. WORKBOOK-LEVEL DEFINED NAMES / NAMED RANGES')
print(sep)
if wb.defined_names:
    for dn in wb.defined_names.values():
        print('  Name:', repr(dn.name))
        print('    Value:', dn.value)
        print('    Scope (localSheetId):', dn.localSheetId)
        print('    Hidden:', dn.hidden)
        print()
else:
    print('  (No defined names found)')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print()
    print(hsep)
    print('## SHEET:', repr(sheet_name))
    print(hsep)
    print('  Dimensions (reported):', ws.dimensions)
    print('  Min row:', ws.min_row, 'Max row:', ws.max_row)
    print('  Min col:', ws.min_column, 'Max col:', ws.max_column)
    rows_with_data = ws.max_row if ws.max_row else 0
    cols_with_data = ws.max_column if ws.max_column else 0
    print('  Rows with data:', rows_with_data)
    print('  Columns with data:', cols_with_data)
    print('  freeze_panes:', ws.freeze_panes)
    print('  sheet_state:', ws.sheet_state)
    try:
        print('  tabColor:', ws.sheet_properties.tabColor)
    except:
        print('  tabColor: (none)')

    print('  --- Column Headers (Row 1) ---')
    if ws.max_column:
        for col_idx in range(1, min(ws.max_column + 1, 200)):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value is not None:
                letter = get_column_letter(col_idx)
                print('    Col {} ({}):'.format(letter, col_idx), repr(cell.value))

    row2_has_headers = False
    if ws.max_row and ws.max_row >= 2:
        non_empty_r2 = 0
        for col_idx in range(1, min(ws.max_column + 1, 200)):
            if ws.cell(row=2, column=col_idx).value is not None:
                non_empty_r2 += 1
        if non_empty_r2 > 3:
            row2_has_headers = True
            print('  --- Row 2 (possible sub-headers, {} non-empty cells) ---'.format(non_empty_r2))
            for col_idx in range(1, min(ws.max_column + 1, 200)):
                cell = ws.cell(row=2, column=col_idx)
                if cell.value is not None:
                    letter = get_column_letter(col_idx)
                    print('    Col {} ({}):'.format(letter, col_idx), repr(cell.value))

    limit = min(cols_with_data, 50)
    print('  --- Column Widths (first {} columns) ---'.format(limit))
    for col_idx in range(1, min(cols_with_data + 1, 51)):
        letter = get_column_letter(col_idx)
        cd = ws.column_dimensions.get(letter)
        if cd:
            print('    Col {}: width={}, hidden={}, outlineLevel={}, collapsed={}'.format(letter, cd.width, cd.hidden, cd.outline_level, cd.collapsed))
        else:
            print('    Col {}: (default)'.format(letter))

    print('  --- Row Heights (first 10 rows) ---')
    for row_idx in range(1, min(11, rows_with_data + 1)):
        rd = ws.row_dimensions.get(row_idx)
        if rd:
            print('    Row {}: height={}, hidden={}, outlineLevel={}'.format(row_idx, rd.height, rd.hidden, rd.outline_level))
        else:
            print('    Row {}: (default)'.format(row_idx))

    print('  --- Merged Cells ---')
    if ws.merged_cells.ranges:
        for mc in ws.merged_cells.ranges:
            print('   ', mc)
        print('  Total merged ranges:', len(ws.merged_cells.ranges))
    else:
        print('    (No merged cells)')
    print('  --- Data Validations ---')
    if ws.data_validations and ws.data_validations.dataValidation:
        for idx, dv in enumerate(ws.data_validations.dataValidation, 1):
            print('    Validation #{}:'.format(idx))
            print('      Type:', dv.type)
            print('      Formula1:', dv.formula1)
            print('      Formula2:', dv.formula2)
            print('      Allow blank:', dv.allow_blank)
            print('      Show error:', dv.showErrorMessage)
            print('      Error title:', dv.errorTitle)
            print('      Error message:', dv.error)
            print('      Prompt title:', dv.promptTitle)
            print('      Prompt message:', dv.prompt)
            print('      Operator:', dv.operator)
            ranges_str = str(dv.sqref) if dv.sqref else '(none)'
            print('      Ranges:', ranges_str)
            print()
        print('  Total validations:', len(ws.data_validations.dataValidation))
    else:
        print('    (No data validations)')

    print('  --- Conditional Formatting ---')
    if ws.conditional_formatting:
        cf_count = 0
        for cf in ws.conditional_formatting:
            cf_count += 1
            print('    CF #{}:'.format(cf_count))
            print('      Ranges:', cf)
            for rule in cf.rules:
                print('      Rule type:', rule.type)
                print('        Priority:', rule.priority)
                print('        Operator:', rule.operator)
                print('        Formula:', rule.formula)
                print('        StopIfTrue:', rule.stopIfTrue)
                if rule.dxf:
                    dxf = rule.dxf
                    if dxf.font:
                        print('        Font color:', dxf.font.color)
                        print('        Font bold:', dxf.font.bold)
                    if dxf.fill:
                        print('        Fill fgColor:', dxf.fill.fgColor, 'bgColor:', dxf.fill.bgColor)
                    if dxf.border:
                        print('        Border: (present)')
                if rule.colorScale:
                    print('        ColorScale:', rule.colorScale)
                if rule.dataBar:
                    print('        DataBar:', rule.dataBar)
                if rule.iconSet:
                    print('        IconSet:', rule.iconSet)
            print()
        print('  Total CF rule sets:', cf_count)
    else:
        print('    (No conditional formatting)')

    print('  --- Column Groupings (outline_level > 0) ---')
    found_groups = False
    for col_idx in range(1, cols_with_data + 1):
        letter = get_column_letter(col_idx)
        cd = ws.column_dimensions.get(letter)
        if cd and cd.outline_level and cd.outline_level > 0:
            found_groups = True
            print('    Col {}: outlineLevel={}, hidden={}, collapsed={}'.format(letter, cd.outline_level, cd.hidden, cd.collapsed))
    if not found_groups:
        print('    (No column groupings)')

    print('  --- Row Groupings (outline_level > 0, first 100 rows) ---')
    found_row_groups = False
    for row_idx in range(1, min(rows_with_data + 1, 101)):
        rd = ws.row_dimensions.get(row_idx)
        if rd and rd.outline_level and rd.outline_level > 0:
            found_row_groups = True
            print('    Row {}: outlineLevel={}, hidden={}'.format(row_idx, rd.outline_level, rd.hidden))
    if not found_row_groups:
        print('    (No row groupings in first 100 rows)')
    start_row = 2
    if row2_has_headers:
        start_row = 3
    print('  --- Sample Data (first 5 data rows starting row {}) ---'.format(start_row))
    for row_idx in range(start_row, min(start_row + 5, rows_with_data + 1)):
        print('    Row {}:'.format(row_idx))
        for col_idx in range(1, min(cols_with_data + 1, 100)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                letter = get_column_letter(col_idx)
                val_str = str(cell.value)
                if len(val_str) > 80:
                    val_str = val_str[:80] + '...'
                tname = type(cell.value).__name__
                print('      {}{}: {}  [type={}]'.format(letter, row_idx, val_str, tname))
        print()

    print('  --- Formulas (first 15 data rows) ---')
    formula_count = 0
    for row_idx in range(start_row, min(start_row + 15, rows_with_data + 1)):
        for col_idx in range(1, min(cols_with_data + 1, 200)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
                letter = get_column_letter(col_idx)
                print('    {}{}: {}'.format(letter, row_idx, cell.value))
    if formula_count == 0:
        print('    (No formulas found in scanned rows)')
    else:
        print('  Total formulas found in scan:', formula_count)

    print('  --- Auto Filter ---')
    if ws.auto_filter and ws.auto_filter.ref:
        print('    AutoFilter ref:', ws.auto_filter.ref)
    else:
        print('    (No auto filter)')

    print('  --- Sheet Protection ---')
    if ws.protection and ws.protection.sheet:
        print('    Sheet protected: True')
        print('    Password:', ws.protection.password)
    else:
        print('    Sheet protected: False')

    print('  --- Print Settings ---')
    if ws.print_area:
        print('    Print area:', ws.print_area)
    if ws.print_title_rows:
        print('    Print title rows:', ws.print_title_rows)
    if ws.print_title_cols:
        print('    Print title cols:', ws.print_title_cols)

    print('  --- Header Cell Formatting (Row 1, first 10 cols) ---')
    for col_idx in range(1, min(cols_with_data + 1, 11)):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value is not None:
            letter = get_column_letter(col_idx)
            font = cell.font
            fill = cell.fill
            align = cell.alignment
            print('    {}1: value={}'.format(letter, repr(cell.value)))
            print('        font: name={}, size={}, bold={}, color={}'.format(font.name, font.size, font.bold, font.color))
            print('        fill: fgColor={}, bgColor={}, patternType={}'.format(fill.fgColor, fill.bgColor, fill.patternType))
            print('        align: horizontal={}, vertical={}, wrapText={}'.format(align.horizontal, align.vertical, align.wrapText))
            print('        number_format:', cell.number_format)

print()
print(sep)
print('ANALYSIS COMPLETE')
print(sep)
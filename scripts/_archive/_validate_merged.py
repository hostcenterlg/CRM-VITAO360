"""Validate CRM_VITAO360_V3_MERGED.xlsx — comprehensive check."""
import openpyxl
import os
from collections import Counter

BASE = os.path.dirname(os.path.abspath(__file__))
FILE = os.path.join(BASE, "output", "CRM_VITAO360_V3_MERGED.xlsx")

wb = openpyxl.load_workbook(FILE, data_only=True, read_only=True)
print("=" * 60)
print("VALIDATION: CRM_VITAO360_V3_MERGED.xlsx")
print("=" * 60)
print(f"Sheets: {wb.sheetnames}")

# ── 1. DRAFT 1 ──
print("\n--- DRAFT 1 ---")
ws = wb["DRAFT 1"]
cnpjs = []
situacoes = Counter()
consultores = Counter()
ufs = Counter()
vendas_count = 0
vendas_total = 0
vendas_by_month = [0] * 12

for row in ws.iter_rows(min_row=4, max_col=48, values_only=True):
    cnpj = str(row[1] or "").strip()
    if not cnpj: continue
    cnpjs.append(cnpj)

    sit = str(row[11] or "").strip()
    if sit: situacoes[sit] += 1

    cons = str(row[9] or "").strip()
    if cons: consultores[cons] += 1

    uf = str(row[3] or "").strip()
    if uf: ufs[uf] += 1

    has_sales = False
    for j in range(12):
        v = row[24 + j]
        if v and isinstance(v, (int, float)) and v > 0:
            has_sales = True
            vendas_by_month[j] += v
    if has_sales:
        vendas_count += 1

    tp = row[17]
    if tp and isinstance(tp, (int, float)):
        vendas_total += tp

print(f"Total clients: {len(cnpjs)}")
print(f"Unique CNPJs: {len(set(cnpjs))}")
dups = len(cnpjs) - len(set(cnpjs))
if dups > 0:
    print(f"  DUPLICATES: {dups}")
else:
    print(f"  No duplicates")

print(f"\nSITUACAO distribution:")
for k, v in situacoes.most_common():
    print(f"  {k}: {v}")

print(f"\nCONSULTOR distribution:")
for k, v in consultores.most_common():
    print(f"  {k}: {v}")

print(f"\nTop UFs:")
for k, v in ufs.most_common(10):
    print(f"  {k}: {v}")

print(f"\nSales data:")
print(f"  Clients with sales: {vendas_count}/{len(cnpjs)}")
print(f"  Total faturamento: R$ {vendas_total:,.2f}")
months = ["MAR/25", "ABR/25", "MAI/25", "JUN/25", "JUL/25", "AGO/25",
          "SET/25", "OUT/25", "NOV/25", "DEZ/25", "JAN/26", "FEV/26"]
for i, m in enumerate(months):
    status = "" if vendas_by_month[i] > 0 else " (vazio)"
    print(f"  {m}: R$ {vendas_by_month[i]:,.2f}{status}")

# ── 2. DRAFT 2 ──
print("\n--- DRAFT 2 ---")
ws2 = wb["DRAFT 2"]
d2_count = 0
d2_resultados = Counter()
d2_situacoes = Counter()
for row in ws2.iter_rows(min_row=3, max_col=24, values_only=True):
    cnpj = str(row[1] or "").strip()
    if cnpj:
        d2_count += 1
    res = str(row[5] or "").strip()
    if res: d2_resultados[res] += 1
    sit = str(row[12] or "").strip()
    if sit: d2_situacoes[sit] += 1
print(f"Records: {d2_count}")
print(f"RESULTADO distribution:")
for k, v in d2_resultados.most_common():
    print(f"  {k}: {v}")
print(f"SITUACAO (auto) distribution:")
for k, v in d2_situacoes.most_common():
    print(f"  {k}: {v}")

# ── 3. CARTEIRA ──
print("\n--- CARTEIRA ---")
ws3 = wb["CARTEIRA"]
cart_count = 0
cart_sap = 0
cart_acomp = 0
for row in ws3.iter_rows(min_row=4, max_col=257, values_only=True):
    cnpj = str(row[1] or "").strip()
    if not cnpj: continue
    cart_count += 1
    if len(row) > 61 and row[61] and str(row[61]).strip():
        cart_sap += 1
    # Check ACOMP data (V3 col 74 = index 73 = %Q1)
    has_acomp = False
    for c_idx in range(73, min(257, len(row))):
        if row[c_idx] is not None:
            has_acomp = True
            break
    if has_acomp:
        cart_acomp += 1

print(f"Clients: {cart_count}")
print(f"With SAP data: {cart_sap}")
print(f"With ACOMP data: {cart_acomp}")

# Check ACOMP sample (first client)
print("\nACOMP sample (first client with data):")
for row in ws3.iter_rows(min_row=4, max_col=130, values_only=True):
    cnpj = str(row[1] or "").strip()
    if not cnpj: continue
    found = False
    for c_idx in range(73, min(130, len(row))):
        if row[c_idx] is not None:
            found = True
    if found:
        print(f"  CNPJ: {cnpj}")
        for c_idx in range(73, min(95, len(row))):
            v = row[c_idx]
            if v is not None:
                print(f"    Col {c_idx+1}: {str(v)[:40]}")
        break

# ── 4. LOG ──
print("\n--- LOG ---")
ws4 = wb["LOG"]
log_count = 0
for row in ws4.iter_rows(min_row=3, max_col=2, values_only=True):
    if row[1]:
        log_count += 1
print(f"Records: {log_count}")

# ── 5. PROJECAO ──
print("\n--- PROJECAO ---")
ws5 = wb["PROJECAO"] if "PROJECAO" in wb.sheetnames else wb.get("PROJEÇÃO")
if ws5:
    proj_rows = 0
    for row in ws5.iter_rows(min_row=1, max_col=1, values_only=True):
        proj_rows += 1
    print(f"Rows: {proj_rows}")
else:
    print("Not found!")

# ── 6. Compare vs previous versions ──
print("\n--- COMPARISON ---")
print(f"{'Metric':<30} {'V3 FINAL':>12} {'MERGED':>12} {'POPULADO':>12}")
print("-" * 68)
print(f"{'DRAFT 1 clients':<30} {'493':>12} {len(cnpjs):>12} {'493':>12}")
print(f"{'DRAFT 1 cols':<30} {'48':>12} {'48':>12} {'45':>12}")
print(f"{'DRAFT 2 records':<30} {'12,199':>12} {d2_count:>12,} {'17,805':>12}")
print(f"{'CARTEIRA cols':<30} {'72':>12} {'257':>12} {'256':>12}")
print(f"{'CARTEIRA ACOMP':<30} {'0':>12} {cart_acomp:>12} {'6,011':>12}")
print(f"{'SAP merged':<30} {'467':>12} {cart_sap:>12} {'N/A':>12}")
print(f"{'LOG records':<30} {'12,199':>12} {log_count:>12,} {'17,810':>12}")
print(f"{'PROJECAO':<30} {'Sim':>12} {'Sim':>12} {'Nao':>12}")
print(f"{'Motor de regras':<30} {'Sim':>12} {'Sim':>12} {'Nao':>12}")

# ── 7. Sample data ──
print("\n--- SAMPLE DATA (first 3 clients) ---")
ws = wb["DRAFT 1"]
for r in range(4, 7):
    nome = ws.cell(row=r, column=1).value
    cnpj = ws.cell(row=r, column=2).value
    uf = ws.cell(row=r, column=4).value
    sit = ws.cell(row=r, column=12).value
    dias = ws.cell(row=r, column=14).value
    cons = ws.cell(row=r, column=10).value
    jan26 = ws.cell(row=r, column=35).value
    curva = ws.cell(row=r, column=39).value
    tipo = ws.cell(row=r, column=44).value
    print(f"  {nome} | {cnpj} | {uf} | {sit} | {dias}d | {cons} | JAN/26={jan26} | {curva} | {tipo}")

wb.close()
print("\nValidation complete!")

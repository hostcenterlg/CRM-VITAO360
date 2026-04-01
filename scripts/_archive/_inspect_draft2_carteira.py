"""Quick inspect POPULADO DRAFT 2 + CARTEIRA ACOMPANHAMENTO layouts."""
import openpyxl
import os

BASE = os.path.dirname(os.path.abspath(__file__))
POP = os.path.join(BASE, "CRM_INTELIGENTE_VITAO360_POPULADO (2).xlsx")

wb = openpyxl.load_workbook(POP, data_only=True, read_only=True)

# ── DRAFT 2 ──
print("=" * 60)
print("POPULADO — DRAFT 2 Layout")
print("=" * 60)
ws = wb["DRAFT 2"]

# Headers (row 2)
print("Row 2 headers:")
for c in range(1, 26):
    h = ws.cell(row=2, column=c).value
    if h:
        print(f"  Col {c:2d} (idx {c-1:2d}): {str(h)[:40]}")

# Sample rows 3-5
print("\nSample data rows 3-5:")
for r in range(3, 6):
    print(f"  Row {r}:")
    for c in range(1, 15):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            print(f"    Col {c:2d}: {str(v)[:50]}")

# Count total records
cnt = 0
cnt_with_cnpj_b = 0
cnt_with_cnpj_d = 0
for row in ws.iter_rows(min_row=3, max_col=24, values_only=True):
    has_data = any(row[i] for i in range(min(6, len(row))))
    if has_data:
        cnt += 1
    if row[1]:  # Col B
        s = str(row[1]).strip()
        if len(s) > 5:
            cnt_with_cnpj_b += 1
    if len(row) > 3 and row[3]:  # Col D
        s = str(row[3]).strip()
        if len(s) > 5:
            cnt_with_cnpj_d += 1
print(f"\nTotal rows with data: {cnt}")
print(f"Rows with CNPJ-like in Col B: {cnt_with_cnpj_b}")
print(f"Rows with CNPJ-like in Col D: {cnt_with_cnpj_d}")

# ── CARTEIRA ACOMPANHAMENTO ──
print("\n" + "=" * 60)
print("POPULADO — CARTEIRA ACOMPANHAMENTO (cols 73-100)")
print("=" * 60)
ws2 = wb["CARTEIRA"]

# Check headers rows 2-3 for ACOMP block
print("Row 2 (sub-block):")
for c in range(73, 100):
    h = ws2.cell(row=2, column=c).value
    if h:
        print(f"  Col {c:3d}: {str(h)[:40]}")

print("\nRow 3 (column headers):")
for c in range(73, 120):
    h = ws2.cell(row=3, column=c).value
    if h:
        print(f"  Col {c:3d}: {str(h)[:40]}")

# Check row 4 data (first client)
print("\nRow 4 sample data:")
cnpj_4 = ws2.cell(row=4, column=2).value
print(f"  CNPJ: {cnpj_4}")
for c in range(73, 120):
    v = ws2.cell(row=4, column=c).value
    if v is not None:
        print(f"  Col {c:3d}: {str(v)[:40]}")

# Check max ACOMP col
max_acomp = 0
for c in range(73, 260):
    h = ws2.cell(row=3, column=c).value
    if h:
        max_acomp = c
print(f"\nMax ACOMP column with header: {max_acomp}")

# Count CARTEIRA rows with CNPJ
cart_cnpjs = 0
for row in ws2.iter_rows(min_row=4, max_col=2, values_only=True):
    if row[1]:
        cart_cnpjs += 1
print(f"CARTEIRA rows with CNPJ: {cart_cnpjs}")

wb.close()
print("\nDone!")

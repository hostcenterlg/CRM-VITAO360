"""Compare POPULADO vs V3 FINAL to find what's missing."""
import openpyxl
import os

BASE = os.path.dirname(os.path.abspath(__file__))
POP = os.path.join(BASE, "CRM_INTELIGENTE_VITAO360_POPULADO (2).xlsx")
FINAL = os.path.join(BASE, "output", "CRM_VITAO360_V3_FINAL.xlsx")

# ── POPULADO ──
print("=" * 60)
print("POPULADO")
print("=" * 60)
wb1 = openpyxl.load_workbook(POP, data_only=True, read_only=True)
print(f"Sheets: {wb1.sheetnames}")
for sn in wb1.sheetnames:
    ws = wb1[sn]
    rows = 0
    cols = 0
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        rows += 1
    # Get max col from row 3
    for c in range(1, 300):
        if ws.cell(row=3, column=c).value:
            cols = c
    print(f"  {sn}: ~{rows} rows, {cols} cols")

    # Print headers for each sheet
    if sn not in ["Claude Log"]:
        print(f"  Headers row 2-3:")
        for c in range(1, min(cols + 5, 100)):
            h2 = ws.cell(row=2, column=c).value or ""
            h3 = ws.cell(row=3, column=c).value or ""
            if h2 or h3:
                print(f"    Col {c:3d}: [{str(h2)[:20]:>20}] {str(h3)[:40]}")

# Check DRAFT 2 in POPULADO
if "DRAFT 2" in wb1.sheetnames:
    ws_d2 = wb1["DRAFT 2"]
    d2_rows = 0
    for row in ws_d2.iter_rows(min_row=3, max_col=2, values_only=True):
        if row[0] or row[1]:
            d2_rows += 1
    print(f"\n  DRAFT 2 data rows: {d2_rows}")
    # Sample
    for r in range(3, min(8, d2_rows + 3)):
        vals = []
        for c in range(1, 25):
            v = ws_d2.cell(row=r, column=c).value
            vals.append(str(v)[:15] if v else "")
        print(f"    Row {r}: {' | '.join(vals[:8])}")

# Check CARTEIRA in POPULADO
if "CARTEIRA" in wb1.sheetnames:
    ws_cart = wb1["CARTEIRA"]
    cart_cols = 0
    for c in range(1, 300):
        if ws_cart.cell(row=3, column=c).value or ws_cart.cell(row=2, column=c).value:
            cart_cols = c
    print(f"\n  CARTEIRA max col: {cart_cols}")
    # Check if has ACOMPANHAMENTO data
    for c in [73, 74, 75, 80, 100, 120, 150, 200, 250]:
        h = ws_cart.cell(row=3, column=c).value or ws_cart.cell(row=2, column=c).value
        v4 = ws_cart.cell(row=4, column=c).value
        if h or v4:
            print(f"    Col {c}: {str(h)[:30]} = {str(v4)[:30]}")

# Check PROJEÇÃO in POPULADO
if "PROJEÇÃO" in wb1.sheetnames:
    ws_proj = wb1["PROJEÇÃO"]
    proj_cols = 0
    for c in range(1, 60):
        if ws_proj.cell(row=3, column=c).value:
            proj_cols = c
    print(f"\n  PROJEÇÃO max col: {proj_cols}")
    for c in range(1, min(proj_cols + 3, 50)):
        h = ws_proj.cell(row=3, column=c).value
        if h:
            print(f"    Col {c}: {str(h)[:40]}")
    # Sample row 4
    print("  Sample row 4:")
    for c in range(1, min(proj_cols + 3, 50)):
        v = ws_proj.cell(row=4, column=c).value
        if v is not None:
            print(f"    Col {c}: {str(v)[:40]}")

wb1.close()

# ── V3 FINAL ──
print("\n" + "=" * 60)
print("V3 FINAL")
print("=" * 60)
wb2 = openpyxl.load_workbook(FINAL, data_only=True, read_only=True)
print(f"Sheets: {wb2.sheetnames}")
for sn in wb2.sheetnames:
    ws = wb2[sn]
    rows = 0
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        rows += 1
    print(f"  {sn}: ~{rows} rows")
wb2.close()

print("\nDone!")

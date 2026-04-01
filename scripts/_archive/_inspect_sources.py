"""Quick inspection of source files for column mapping."""
import openpyxl
import os

BASE = os.path.dirname(os.path.abspath(__file__))

# 1. POPULADO DRAFT 1
print("=" * 60)
print("POPULADO - DRAFT 1 Headers")
print("=" * 60)
wb = openpyxl.load_workbook(
    os.path.join(BASE, "CRM_INTELIGENTE_VITAO360_POPULADO (2).xlsx"),
    data_only=True, read_only=True
)
ws = wb["DRAFT 1"]
for col in range(1, 50):
    h2 = ws.cell(row=2, column=col).value or ""
    h3 = ws.cell(row=3, column=col).value or ""
    print(f"  Col {col:2d}: [{str(h2)[:15]:>15}] {str(h3)[:35]}")

print("\nSample Row 4:")
for c in range(1, 46):
    v = ws.cell(row=4, column=c).value
    if v is not None:
        print(f"  Col {c:2d}: {str(v)[:50]}")
wb.close()

# 2. CARTEIRA OULAR
print("\n" + "=" * 60)
print("CARTEIRA OULAR - CARTEIRA Tab Headers")
print("=" * 60)
wb2 = openpyxl.load_workbook(
    os.path.join(BASE, "CARTEIRA DE CLIENTES OULAR.xlsx"),
    data_only=True, read_only=True
)
print("Sheets:", wb2.sheetnames)
ws2 = wb2["CARTEIRA"]
for col in range(1, 75):
    h2 = ws2.cell(row=2, column=col).value or ""
    h3 = ws2.cell(row=3, column=col).value or ""
    label = str(h3)[:35] if h3 else str(h2)[:35]
    if label.strip():
        print(f"  Col {col:2d}: [{str(h2)[:15]:>15}] {label}")

print("\nSample Row 4 (key cols):")
for c in [1, 2, 3, 4, 5, 10, 12, 14, 15, 16, 17, 18, 25, 26, 35, 36, 37, 38, 39, 40, 60, 61, 62]:
    v = ws2.cell(row=4, column=c).value
    if v is not None:
        print(f"  Col {c:2d}: {str(v)[:50]}")

# Count rows with data
cnt = 0
for row in ws2.iter_rows(min_row=4, max_col=2, values_only=True):
    if row[1]:
        cnt += 1
print(f"\nTotal clients with CNPJ: {cnt}")
wb2.close()

# 3. LOG DRAFT - check sheet names
print("\n" + "=" * 60)
print("LOG DRAFT")
print("=" * 60)
log_files = [
    "preenchimento_do_draft_de_atendimento_LOG_FINAL (1).xlsx",
    "preenchimento_do_draft_de_atendimento_LOG_FINAL.xlsx",
    "preenchimento do fraft de atendimento (LOG).xlsx",
]
for f in log_files:
    fp = os.path.join(BASE, f)
    if os.path.exists(fp):
        wb3 = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        print(f"\n{f}")
        print(f"  Sheets: {wb3.sheetnames}")
        for sn in wb3.sheetnames:
            ws3 = wb3[sn]
            max_r = 0
            for row in ws3.iter_rows(min_row=1, max_col=1, values_only=True):
                max_r += 1
            print(f"  {sn}: ~{max_r} rows")
            # Print headers
            for col in range(1, 25):
                h = ws3.cell(row=2, column=col).value or ws3.cell(row=1, column=col).value
                if h:
                    print(f"    Col {col:2d}: {str(h)[:35]}")
        wb3.close()

print("\nDone!")

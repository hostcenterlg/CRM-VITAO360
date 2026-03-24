"""
extract_carteira_blueprint.py
@data-engineer — CRM VITAO360

Extracts CARTEIRA column architecture from the master Excel file.
Produces: data/intelligence/carteira_blueprint.json

Row layout in CARTEIRA header:
  Row 1 = GRUPO (group label — sparse, only at group boundary columns)
  Row 2 = SUBGRUPO (sub-group — dense, present on most columns)
  Row 3 = NOME COLUNA (column name — mix of text and formulas on summary cols)
  Row 4+ = data rows

Group assignment strategy:
  Row 1 is sparse — groups are labeled at their START column only.
  We build a boundary map from row 1 non-empty values, then assign
  each column to the group whose start boundary is nearest (<=) to it.
  Columns before the first label get group "(sem grupo)".

Rules respected:
  - R4: formulas expected in ENGLISH in openpyxl output
  - data_only=False for formula extraction
"""

import json
import sys
import re
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SOURCE = Path(r"C:/Users/User/OneDrive/Área de Trabalho/CRM_VITAO360  INTELIGENTE   FINAL OK .xlsx")
OUTPUT = Path(r"C:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/intelligence/carteira_blueprint.json")
SHEET_NAME = "CARTEIRA"

# Header rows
ROW_GRUPO    = 1
ROW_SUBGRUPO = 2
ROW_NOME     = 3
ROW_FORMULA1 = 4
ROW_FORMULA2 = 5


def safe_str(v) -> str:
    """Convert cell value to clean string, or empty string."""
    if v is None:
        return ""
    s = str(v).strip()
    return s


def get_cell_value(cell):
    """Safely get .value from a cell (handles EmptyCell in read_only mode)."""
    try:
        return cell.value
    except AttributeError:
        return None


def read_rows(ws, min_row: int, max_row: int, max_col: int) -> dict:
    """
    Read rows from worksheet into dict: {row_num: {col_num: raw_value}}.
    Uses enumerate-based indexing to avoid EmptyCell attribute errors.
    """
    result = {}
    for row_offset, row in enumerate(
        ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col)
    ):
        row_num = min_row + row_offset
        result[row_num] = {}
        for col_offset, cell in enumerate(row):
            col_num = 1 + col_offset
            result[row_num][col_num] = get_cell_value(cell)
    return result


def build_group_boundaries(row1: dict, max_col: int) -> list:
    """
    Build sorted list of (start_col, grupo_label) from sparse row 1 values.
    """
    boundaries = []
    for col_idx in range(1, max_col + 1):
        v = safe_str(row1.get(col_idx))
        if v:
            boundaries.append((col_idx, v))
    return boundaries


def assign_group(col_idx: int, boundaries: list) -> str:
    """
    Return the group label for a column by finding the nearest start boundary <= col_idx.
    If no boundary found, return '(sem grupo)'.
    """
    assigned = "(sem grupo)"
    for start, label in boundaries:
        if start <= col_idx:
            assigned = label
        else:
            break  # boundaries are sorted
    return assigned


def format_subgrupo(v) -> str:
    """
    Convert subgrupo cell value to display string.
    Dates stored as datetime objects get formatted as YYYY-MM.
    """
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m")
    return str(v).strip()


def extract_blueprint():
    print(f"Loading workbook (data_only=False) from:\n  {SOURCE}")
    if not SOURCE.exists():
        print(f"ERROR: Source file not found: {SOURCE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(SOURCE), data_only=False, read_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    max_col = ws.max_column
    max_row = ws.max_row
    print(f"Sheet dimensions: {max_col} columns x {max_row} rows")

    # -----------------------------------------------------------------------
    # Step 1: Read rows 1-5 into memory
    # -----------------------------------------------------------------------
    print("Reading header rows 1-5 …")
    rows = read_rows(ws, min_row=1, max_row=5, max_col=max_col)
    row1 = rows[ROW_GRUPO]
    row2 = rows[ROW_SUBGRUPO]
    row3 = rows[ROW_NOME]
    row4 = rows[ROW_FORMULA1]
    row5 = rows[ROW_FORMULA2]

    # -----------------------------------------------------------------------
    # Step 2: Build group boundaries from sparse row 1
    # -----------------------------------------------------------------------
    boundaries = build_group_boundaries(row1, max_col)
    print(f"Group boundaries found in row 1: {len(boundaries)}")
    for start, label in boundaries:
        print(f"  Col {start:3d} ({get_column_letter(start):3s}): {label}")

    # -----------------------------------------------------------------------
    # Step 3: Build column_map
    # -----------------------------------------------------------------------
    print(f"\nBuilding column map for {max_col} columns …")
    column_map = []

    for col_idx in range(1, max_col + 1):
        letter   = get_column_letter(col_idx)
        grupo    = assign_group(col_idx, boundaries)
        subgrupo = format_subgrupo(row2.get(col_idx))
        nome_raw = row3.get(col_idx)

        # Nome coluna: if it's a formula (starts with =), mark as formula_header
        if isinstance(nome_raw, str) and nome_raw.startswith("="):
            nome_coluna = f"[FORMULA: {nome_raw}]"
            is_formula_header = True
        elif nome_raw is None:
            nome_coluna = ""
            is_formula_header = False
        else:
            nome_coluna = safe_str(nome_raw)
            is_formula_header = False

        column_map.append({
            "col_number":        col_idx,
            "col_letter":        letter,
            "grupo":             grupo,
            "subgrupo":          subgrupo,
            "nome_coluna":       nome_coluna,
            "is_formula_header": is_formula_header,
        })

    # -----------------------------------------------------------------------
    # Step 4: Formula patterns from rows 4-5
    # -----------------------------------------------------------------------
    print("Extracting formula patterns from rows 4-5 …")
    formula_patterns = []

    for col_idx in range(1, max_col + 1):
        letter    = get_column_letter(col_idx)
        nome_col  = column_map[col_idx - 1]["nome_coluna"]

        v4 = row4.get(col_idx)
        v5 = row5.get(col_idx)

        formula_r4 = str(v4) if isinstance(v4, str) and v4.startswith("=") else ""
        formula_r5 = str(v5) if isinstance(v5, str) and v5.startswith("=") else ""

        if formula_r4 or formula_r5:
            formula_patterns.append({
                "col_letter":   letter,
                "col_number":   col_idx,
                "nome_coluna":  nome_col,
                "grupo":        column_map[col_idx - 1]["grupo"],
                "subgrupo":     column_map[col_idx - 1]["subgrupo"],
                "formula_row4": formula_r4,
                "formula_row5": formula_r5,
            })

    print(f"  Columns with formula samples in rows 4-5: {len(formula_patterns)}")

    # -----------------------------------------------------------------------
    # Step 5: Groups summary
    # -----------------------------------------------------------------------
    print("Aggregating groups summary …")
    groups_raw: dict = {}
    for entry in column_map:
        g = entry["grupo"]
        groups_raw.setdefault(g, {"cols": [], "subgrupos_seen": [], "subgrupos_set": set()})
        groups_raw[g]["cols"].append(entry["col_number"])
        sg = entry["subgrupo"]
        if sg and sg not in groups_raw[g]["subgrupos_set"]:
            groups_raw[g]["subgrupos_set"].add(sg)
            groups_raw[g]["subgrupos_seen"].append(sg)

    groups_summary = []
    for grupo_nome, data in groups_raw.items():
        cols = data["cols"]
        min_c = min(cols)
        max_c = max(cols)
        groups_summary.append({
            "grupo_nome":  grupo_nome,
            "col_range":   f"{get_column_letter(min_c)}-{get_column_letter(max_c)}",
            "col_start":   min_c,
            "col_end":     max_c,
            "num_columns": len(cols),
            "subgrupos":   data["subgrupos_seen"],
        })

    groups_summary.sort(key=lambda x: x["col_start"])

    # -----------------------------------------------------------------------
    # Step 6: Count total formulas (full data scan)
    # -----------------------------------------------------------------------
    print("Counting total formulas across all data rows (this may take a moment) …")
    total_formulas = 0
    for row in ws.iter_rows(min_row=4, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            v = get_cell_value(cell)
            if v is not None and isinstance(v, str) and v.startswith("="):
                total_formulas += 1

    data_rows = max_row - 3  # subtract 3 header rows

    stats = {
        "total_columns":               max_col,
        "total_formulas":              total_formulas,
        "total_data_rows":             data_rows,
        "grupos_count":                len(groups_summary),
        "columns_with_formula_samples": len(formula_patterns),
        "formula_header_columns":      sum(1 for c in column_map if c["is_formula_header"]),
        "source_file":                 str(SOURCE.name),
        "sheet_name":                  SHEET_NAME,
        "max_row":                     max_row,
    }

    # -----------------------------------------------------------------------
    # Step 7: Assemble blueprint
    # -----------------------------------------------------------------------
    blueprint = {
        "metadata": {
            "version":      "1.0",
            "extracted_by": "@data-engineer — CRM VITAO360",
            "source_file":  str(SOURCE.name),
            "sheet_name":   SHEET_NAME,
            "description":  (
                "CARTEIRA sheet column architecture — groups, subgroups, "
                "column names, formula patterns, and stats"
            ),
        },
        "stats":            stats,
        "groups_summary":   groups_summary,
        "column_map":       column_map,
        "formula_patterns": formula_patterns,
    }

    # -----------------------------------------------------------------------
    # Step 8: Save
    # -----------------------------------------------------------------------
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(blueprint, f, ensure_ascii=False, indent=2)

    print(f"\nSaved: {OUTPUT}")

    # -----------------------------------------------------------------------
    # Step 9: Print summary
    # -----------------------------------------------------------------------
    print("\n" + "=" * 80)
    print("CARTEIRA COLUMN ARCHITECTURE — GROUPS SUMMARY")
    print("=" * 80)
    print(f"{'GRUPO':<35} {'RANGE':<10} {'COLS':>5}  SUBGRUPOS")
    print("-" * 80)
    for g in groups_summary:
        subg_str = ", ".join(g["subgrupos"]) if g["subgrupos"] else "(none)"
        # Truncate subgrupos if too long
        if len(subg_str) > 60:
            subg_str = subg_str[:57] + "…"
        print(f"{g['grupo_nome']:<35} {g['col_range']:<10} {g['num_columns']:>5}  {subg_str}")

    print("=" * 80)
    print(f"\nSTATS:")
    print(f"  Total columns          : {stats['total_columns']}")
    print(f"  Total data rows        : {stats['total_data_rows']}")
    print(f"  Total formulas         : {stats['total_formulas']:,}")
    print(f"  Groups detected        : {stats['grupos_count']}")
    print(f"  Cols w/ formula samples: {stats['columns_with_formula_samples']}")
    print(f"  Formula-header columns : {stats['formula_header_columns']}")
    print(f"\nOutput: {OUTPUT}")

    wb.close()
    return blueprint


if __name__ == "__main__":
    extract_blueprint()

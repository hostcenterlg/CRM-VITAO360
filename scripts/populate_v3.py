"""
Populate CRM VITAO360 V3 — Import real data from source files.

Sources:
  - CARTEIRA OULAR.xlsx → DRAFT 1 (6,011 clients) + CARTEIRA (XLOOKUP formulas)
  - preenchimento do fraft de atendimento (LOG).xlsx → DRAFT 2 + LOG

Steps:
  1. Build empty V3 structure (build_v3)
  2. Read source data
  3. Populate DRAFT 1 with client master data
  4. Populate DRAFT 2 with attendance records (motor de regras recalculates auto cols)
  5. Populate LOG (archive copy, no formulas)
  6. Save
"""
import os
import sys
import time
import datetime
import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from v3_styles import *
from motor_regras import motor_de_regras, FOLLOW_UP_DIAS, GRUPO_DASH, dia_util

# ═══════════════════════════════════════════════════════════════
# FILE PATHS
# ═══════════════════════════════════════════════════════════════

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DESKTOP = os.path.dirname(BASE_DIR)

CARTEIRA_OULAR = os.path.join(DESKTOP, "CARTEIRA DE CLIENTES OULAR.xlsx")
LOG_DRAFT = os.path.join(DESKTOP, "preenchimento do fraft de atendimento (LOG).xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "output", "CRM_VITAO360_V3.xlsx")

# ═══════════════════════════════════════════════════════════════
# MAPPING: RESULTADO normalization
# ═══════════════════════════════════════════════════════════════

RESULTADO_MAP = {
    # Source → V3 standard
    "EM ATENDIMENTO": "EM ATENDIMENTO",
    "ORCAMENTO": "ORÇAMENTO",
    "ORÇAMENTO": "ORÇAMENTO",
    "CADASTRO": "CADASTRO",
    "VENDA/PEDIDO": "VENDA / PEDIDO",
    "VENDA / PEDIDO": "VENDA / PEDIDO",
    "RELACIONAMENTO": "RELACIONAMENTO",
    "FOLLOW UP 7": "FOLLOW UP 7",
    "FOLLOW UP 15": "FOLLOW UP 15",
    "SUPORTE": "SUPORTE",
    "NAO ATENDE": "NÃO ATENDE",
    "NÃO ATENDE": "NÃO ATENDE",
    "NAO RESPONDE": "NÃO RESPONDE",
    "NÃO RESPONDE": "NÃO RESPONDE",
    "RECUSOU LIGACAO": "RECUSOU LIGAÇÃO",
    "RECUSOU LIGAÇÃO": "RECUSOU LIGAÇÃO",
    "PERDA/FECHOU LOJA": "PERDA / FECHOU LOJA",
    "PERDA / FECHOU LOJA": "PERDA / FECHOU LOJA",
    "SEM INTERESSE NO MOMENTO": "NÃO RESPONDE",  # reclassify
    "SEM INTERESSE": "NÃO RESPONDE",  # reclassify
}

SITUACAO_MAP = {
    "ATIVO": "ATIVO",
    "EM RISCO": "EM RISCO",
    "INAT.REC": "INAT.REC",
    "INAT.ANT": "INAT.ANT",
    "NOVO": "NOVO",
    "PROSPECT": "PROSPECT",
    "LEAD": "PROSPECT",  # reclassify LEAD → PROSPECT
    "EM DESENVOLVIMENTO": "ATIVO",  # reclassify EM DESENVOLVIMENTO → ATIVO
}

VALID_RESULTADOS = set(RESULTADO_MAP.values())
VALID_SITUACOES = {"ATIVO", "EM RISCO", "INAT.REC", "INAT.ANT", "NOVO", "PROSPECT"}


def safe_str(val):
    """Convert to string, handle None."""
    if val is None:
        return ""
    return str(val).strip()


def safe_date(val):
    """Convert to date, handle various formats."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date() if hasattr(val, 'date') else val
    if isinstance(val, datetime.date):
        return val
    return None


def normalize_resultado(val):
    """Normalize RESULTADO to V3 standard."""
    s = safe_str(val).upper().strip()
    return RESULTADO_MAP.get(s, s)


def normalize_situacao(val):
    """Normalize SITUAÇÃO to V3 standard (no LEAD, no EM DESENVOLVIMENTO)."""
    s = safe_str(val).upper().strip()
    mapped = SITUACAO_MAP.get(s, s)
    if mapped not in VALID_SITUACOES:
        return "PROSPECT"  # fallback
    return mapped


# ═══════════════════════════════════════════════════════════════
# STEP 1: READ SOURCE DATA
# ═══════════════════════════════════════════════════════════════

def read_carteira_oular():
    """Read client master data from CARTEIRA OULAR.xlsx (CARTEIRA tab)."""
    print("  Reading CARTEIRA OULAR...")
    wb = openpyxl.load_workbook(CARTEIRA_OULAR, data_only=True, read_only=True)
    ws = wb["CARTEIRA"]

    # Row 3 = headers, data starts row 4
    clients = []
    count = 0
    for row in ws.iter_rows(min_row=4, max_col=71, values_only=True):
        # Col A (0) = NOME FANTASIA, Col B (1) = CNPJ
        cnpj = safe_str(row[1])
        if not cnpj or cnpj == "" or cnpj == "None":
            continue

        client = {
            # IDENTIDADE (DRAFT 1 cols 1-11)
            'nome_fantasia': safe_str(row[0]),
            'cnpj': cnpj,
            'razao_social': safe_str(row[2]),
            'uf': safe_str(row[3]),
            'cidade': safe_str(row[4]),
            'email': safe_str(row[5]),
            'telefone': safe_str(row[6]),
            'data_cadastro': safe_date(row[7]),
            'rede_regional': safe_str(row[8]),
            'consultor': safe_str(row[10]),  # col K
            'vendedor_ult_pedido': safe_str(row[11]),  # col L

            # STATUS
            'situacao': normalize_situacao(row[12]),  # col M
            'prioridade': row[13],  # col N
            'dias_sem_compra': row[14],  # col O

            # COMPRAS
            'data_ult_pedido': safe_date(row[15]),  # col P
            'valor_ult_pedido': row[16],  # col Q

            # ECOMMERCE
            'acessos_semana': row[17],  # col R
            'acesso_b2b': row[18],  # col S
            'acessos_portal': row[19],  # col T
            'itens_carrinho': row[20],  # col U
            'valor_b2b': row[21],  # col V
            'oportunidade': safe_str(row[22]),  # col W

            # VENDAS
            'total_periodo': row[23],  # col X
            'vendas_mes': [row[i] for i in range(24, 35)],  # Y-AI (11 months)

            # RECORRÊNCIA
            'ciclo_medio': row[35],  # AJ
            'n_compras': row[36],  # AK
            'curva_abc': safe_str(row[37]),  # AL
            'meses_positivado': row[38],  # AM
            'media_mensal': row[39],  # AN
            'ticket_medio': row[40],  # AO
            'meses_lista': row[41],  # AP

            # ATENDIMENTO MERCOS (from OULAR these are in diff cols)
            'ult_registro_mercos': safe_date(row[9]),  # col J in OULAR

            # SAP (cols BI-BS = 60-70)
            'cod_cliente_sap': safe_str(row[60]) if len(row) > 60 else "",
            'desc_grupo_cliente': safe_str(row[61]) if len(row) > 61 else "",
            'gerente_nacional': safe_str(row[62]) if len(row) > 62 else "",
            'representante': safe_str(row[63]) if len(row) > 63 else "",
            'vend_interno': safe_str(row[64]) if len(row) > 64 else "",
            'nome_canal': safe_str(row[65]) if len(row) > 65 else "",
            'tipo_cliente_sap': safe_str(row[66]) if len(row) > 66 else "",
            'macroregiao': safe_str(row[67]) if len(row) > 67 else "",
            'microregiao': safe_str(row[68]) if len(row) > 68 else "",
            'grupo_chave': safe_str(row[69]) if len(row) > 69 else "",
            'venda_sap': row[70] if len(row) > 70 else None,
        }
        clients.append(client)
        count += 1

    wb.close()
    print(f"    → {count} clients read")
    return clients


def read_log_atendimentos():
    """Read attendance log from preenchimento do fraft de atendimento (LOG).xlsx."""
    print("  Reading LOG de atendimentos...")
    wb = openpyxl.load_workbook(LOG_DRAFT, data_only=True, read_only=True)
    ws = wb["DRAFT 2 (2)"]

    # Row 2 = headers, data starts row 3
    # Source columns: A=DATA, B=CONSULTOR, C=NOME, D=CNPJ, E=UF, F=REDE,
    #   G=SITUACAO, H=DIAS SEM COMPRA, I=ESTAGIO, J=TIPO CLIENTE, K=FASE,
    #   L=SINALEIRO, M=TENTATIVA, N=WHATSAPP, O=LIGACAO, P=LIG.ATENDIDA,
    #   Q=TIPO CONTATO, R=RESULTADO, S=MOTIVO, T=FOLLOW-UP, U=ACAO FUTURA,
    #   V=ACAO DETALHADA, W=MERCOS, X=NOTA

    records = []
    count = 0
    skipped = 0

    for row in ws.iter_rows(min_row=3, max_col=24, values_only=True):
        # Must have at least DATA or CNPJ
        data = safe_date(row[0])
        cnpj = safe_str(row[3])  # col D
        consultor = safe_str(row[1])  # col B
        nome = safe_str(row[2])  # col C

        if not cnpj and not nome:
            skipped += 1
            continue

        resultado_raw = safe_str(row[17])  # col R
        resultado = normalize_resultado(resultado_raw)
        situacao_raw = safe_str(row[6])  # col G
        situacao = normalize_situacao(situacao_raw)

        # Validate resultado
        if resultado not in VALID_RESULTADOS and resultado != "":
            resultado = "EM ATENDIMENTO"  # safe fallback

        record = {
            # 12 MANUAL columns (V3 DRAFT 2 order)
            'data': data,
            'cnpj': cnpj,
            'nome_fantasia': nome,
            'uf': safe_str(row[4]),  # col E
            'consultor': consultor,
            'resultado': resultado,
            'motivo': safe_str(row[18]),  # col S
            'whatsapp': safe_str(row[13]),  # col N
            'ligacao': safe_str(row[14]),  # col O
            'lig_atendida': safe_str(row[15]),  # col P
            'nota_dia': safe_str(row[23]),  # col X
            'mercos_atualizado': safe_str(row[22]),  # col W

            # Existing auto fields (will be recalculated, but keep for reference)
            'situacao_original': situacao,
            'tentativa_original': safe_str(row[12]),  # col M
            'estagio_original': safe_str(row[8]),  # col I
        }
        records.append(record)
        count += 1

    wb.close()
    print(f"    → {count} records read ({skipped} empty rows skipped)")
    return records


# ═══════════════════════════════════════════════════════════════
# STEP 2: BUILD & POPULATE V3
# ═══════════════════════════════════════════════════════════════

def populate_draft1(wb, clients):
    """Populate DRAFT 1 with client master data."""
    ws = wb["DRAFT 1"]
    print(f"  Populating DRAFT 1 with {len(clients)} clients...")

    for i, c in enumerate(clients):
        r = 4 + i  # data starts at row 4

        # IDENTIDADE (cols 1-11)
        ws.cell(row=r, column=1, value=c['nome_fantasia'])
        ws.cell(row=r, column=2, value=c['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=c['razao_social'])
        ws.cell(row=r, column=4, value=c['uf'])
        ws.cell(row=r, column=5, value=c['cidade'])
        ws.cell(row=r, column=6, value=c['email'])
        ws.cell(row=r, column=7, value=c['telefone'])
        if c['data_cadastro']:
            ws.cell(row=r, column=8, value=c['data_cadastro']).number_format = FMT_DATE
        ws.cell(row=r, column=9, value=c['rede_regional'])
        ws.cell(row=r, column=10, value=c['consultor'])
        ws.cell(row=r, column=11, value=c['vendedor_ult_pedido'])

        # STATUS (cols 12-15) — formulas already exist for row 4, write values for rest
        # SITUAÇÃO (col 12)
        ws.cell(row=r, column=12, value=c['situacao'])
        # PRIORIDADE (col 13)
        prio_map = {"ATIVO": 1, "EM RISCO": 2, "INAT.REC": 3, "INAT.ANT": 4, "PROSPECT": 5, "NOVO": 6}
        ws.cell(row=r, column=13, value=prio_map.get(c['situacao'], 99))
        # DIAS SEM COMPRA (col 14)
        ws.cell(row=r, column=14, value=c['dias_sem_compra'])
        # CICLO MÉDIO (col 15) — from recorrência
        ws.cell(row=r, column=15, value=c['ciclo_medio'])

        # COMPRAS (cols 16-18)
        if c['data_ult_pedido']:
            ws.cell(row=r, column=16, value=c['data_ult_pedido']).number_format = FMT_DATE
        if c['valor_ult_pedido']:
            ws.cell(row=r, column=17, value=c['valor_ult_pedido']).number_format = FMT_MONEY
        if c['total_periodo']:
            ws.cell(row=r, column=18, value=c['total_periodo']).number_format = FMT_MONEY

        # ECOMMERCE (cols 19-24)
        ws.cell(row=r, column=19, value=c['acessos_semana'])
        ws.cell(row=r, column=20, value=c['acesso_b2b'])
        ws.cell(row=r, column=21, value=c['acessos_portal'])
        ws.cell(row=r, column=22, value=c['itens_carrinho'])
        if c['valor_b2b']:
            ws.cell(row=r, column=23, value=c['valor_b2b']).number_format = FMT_MONEY
        ws.cell(row=r, column=24, value=c['oportunidade'])

        # VENDAS MÊS A MÊS (cols 25-36 = 11 months + col 37 = TOTAL)
        for j, venda in enumerate(c['vendas_mes']):
            if venda:
                ws.cell(row=r, column=25 + j, value=venda).number_format = FMT_MONEY
        # TOTAL VENDAS (col 37) = SUM
        ws.cell(row=r, column=37).value = f'=SUM(Y{r}:AJ{r})'
        ws.cell(row=r, column=37).number_format = FMT_MONEY

        # RECORRÊNCIA (cols 38-44)
        ws.cell(row=r, column=38, value=c['n_compras'])
        ws.cell(row=r, column=39, value=c['curva_abc'])
        ws.cell(row=r, column=40, value=c['meses_positivado'])
        if c['media_mensal']:
            ws.cell(row=r, column=41, value=c['media_mensal']).number_format = FMT_MONEY
        if c['ticket_medio']:
            ws.cell(row=r, column=42, value=c['ticket_medio']).number_format = FMT_MONEY
        ws.cell(row=r, column=43, value=c['meses_lista'])
        # TIPO CLIENTE (col 44) — calc from meses_positivado
        mp = c['meses_positivado'] or 0
        if isinstance(mp, (int, float)):
            if mp == 0:
                tipo_cl = "PROSPECT"
            elif mp == 1:
                tipo_cl = "NOVO"
            elif mp <= 3:
                tipo_cl = "EM DESENVOLVIMENTO"
            elif mp <= 6:
                tipo_cl = "RECORRENTE"
            else:
                tipo_cl = "FIDELIZADO"
        else:
            tipo_cl = "PROSPECT"
        ws.cell(row=r, column=44, value=tipo_cl)

        # ATENDIMENTO MERCOS (cols 45-48)
        if c['ult_registro_mercos']:
            ws.cell(row=r, column=45, value=c['ult_registro_mercos']).number_format = FMT_DATE

    last_row = 4 + len(clients) - 1
    print(f"    → DRAFT 1: rows 4-{last_row} populated ({len(clients)} clients)")
    return last_row


def populate_draft2(wb, records):
    """Populate DRAFT 2 with attendance records + motor de regras formulas."""
    ws = wb["DRAFT 2"]
    print(f"  Populating DRAFT 2 with {len(records)} records...")

    # Build tentativa tracker per CNPJ (to chain T1→T2→T3→T4)
    tentativa_por_cnpj = {}
    estagio_por_cnpj = {}

    for i, rec in enumerate(records):
        r = 3 + i  # data starts at row 3

        # ── 12 MANUAL columns ──
        if rec['data']:
            ws.cell(row=r, column=1, value=rec['data']).number_format = FMT_DATE
        ws.cell(row=r, column=2, value=rec['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=rec['nome_fantasia'])
        ws.cell(row=r, column=4, value=rec['uf'])
        ws.cell(row=r, column=5, value=rec['consultor'])
        ws.cell(row=r, column=6, value=rec['resultado'])
        ws.cell(row=r, column=7, value=rec['motivo'])
        ws.cell(row=r, column=8, value=rec['whatsapp'])
        ws.cell(row=r, column=9, value=rec['ligacao'])
        ws.cell(row=r, column=10, value=rec['lig_atendida'])
        ws.cell(row=r, column=11, value=rec['nota_dia'])
        ws.cell(row=r, column=12, value=rec['mercos_atualizado'])

        # ── 12 AUTO columns (motor de regras) ──
        situacao = rec['situacao_original']
        resultado = rec['resultado']
        cnpj = rec['cnpj']

        # Get previous tentativa/estagio for this CNPJ
        tent_ant = tentativa_por_cnpj.get(cnpj)
        est_ant = estagio_por_cnpj.get(cnpj)

        if resultado and resultado in VALID_RESULTADOS:
            motor = motor_de_regras(situacao, resultado, est_ant, tent_ant)

            # Col M (13): SITUAÇÃO
            ws.cell(row=r, column=13, value=situacao)
            # Col N (14): ESTÁGIO FUNIL
            ws.cell(row=r, column=14, value=motor.get('estagio_funil', ''))
            # Col O (15): FASE
            ws.cell(row=r, column=15, value=motor.get('fase', ''))
            # Col P (16): TIPO DO CONTATO
            ws.cell(row=r, column=16, value=motor.get('tipo_contato', ''))
            # Col Q (17): TEMPERATURA
            ws.cell(row=r, column=17, value=motor.get('temperatura', ''))
            # Col R (18): TENTATIVA
            tent = motor.get('tentativa')
            ws.cell(row=r, column=18, value=tent or '')
            # Col S (19): GRUPO DASH
            ws.cell(row=r, column=19, value=motor.get('grupo_dash', ''))
            # Col T (20): FOLLOW-UP
            if rec['data'] and motor.get('follow_up_dias', 0) > 0:
                follow_date = dia_util(rec['data'], motor['follow_up_dias'])
                ws.cell(row=r, column=20, value=follow_date).number_format = FMT_DATE
            # Col U (21): AÇÃO FUTURA
            ws.cell(row=r, column=21, value=motor.get('acao_futura', ''))
            # Col V (22): AÇÃO DETALHADA
            ws.cell(row=r, column=22, value=motor.get('acao_detalhada', ''))
            # Col W (23): SINALEIRO CICLO — needs DRAFT 1 data, use formula
            ws.cell(row=r, column=23).value = (
                f'=IFERROR(IF(XLOOKUP(B{r},\'DRAFT 1\'!$B$4:$B$10000,\'DRAFT 1\'!$O$4:$O$10000,0)=0,"🟣",'
                f'IF(XLOOKUP(B{r},\'DRAFT 1\'!$B$4:$B$10000,\'DRAFT 1\'!$N$4:$N$10000,0)'
                f'<=XLOOKUP(B{r},\'DRAFT 1\'!$B$4:$B$10000,\'DRAFT 1\'!$O$4:$O$10000,0),"🟢",'
                f'IF(XLOOKUP(B{r},\'DRAFT 1\'!$B$4:$B$10000,\'DRAFT 1\'!$N$4:$N$10000,0)'
                f'<=XLOOKUP(B{r},\'DRAFT 1\'!$B$4:$B$10000,\'DRAFT 1\'!$O$4:$O$10000,0)+30,"🟡","🔴"))),"🟣")'
            )
            # Col X (24): SINALEIRO META — placeholder
            ws.cell(row=r, column=24, value='')

            # Update trackers
            tentativa_por_cnpj[cnpj] = tent
            estagio_por_cnpj[cnpj] = motor.get('estagio_funil', '')
        else:
            # No resultado — write situacao only
            ws.cell(row=r, column=13, value=situacao)

        # Style auto-calculated columns (light green background)
        for c in range(13, 25):
            cell = ws.cell(row=r, column=c)
            cell.fill = FILL_CALC

    last_row = 3 + len(records) - 1
    print(f"    → DRAFT 2: rows 3-{last_row} populated ({len(records)} records, motor applied)")
    return last_row


def populate_log(wb, records):
    """Populate LOG tab (archive copy, plain text, no formulas)."""
    ws = wb["LOG"]
    print(f"  Populating LOG with {len(records)} records (archive)...")

    for i, rec in enumerate(records):
        r = 3 + i

        # Same 24 cols as DRAFT 2 but plain text (no formulas)
        if rec['data']:
            ws.cell(row=r, column=1, value=rec['data']).number_format = FMT_DATE
        ws.cell(row=r, column=2, value=rec['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=rec['nome_fantasia'])
        ws.cell(row=r, column=4, value=rec['uf'])
        ws.cell(row=r, column=5, value=rec['consultor'])
        ws.cell(row=r, column=6, value=rec['resultado'])
        ws.cell(row=r, column=7, value=rec['motivo'])
        ws.cell(row=r, column=8, value=rec['whatsapp'])
        ws.cell(row=r, column=9, value=rec['ligacao'])
        ws.cell(row=r, column=10, value=rec['lig_atendida'])
        ws.cell(row=r, column=11, value=rec['nota_dia'])
        ws.cell(row=r, column=12, value=rec['mercos_atualizado'])

        # Auto cols → plain values (calculated by motor, not formulas)
        situacao = rec['situacao_original']
        resultado = rec['resultado']

        if resultado and resultado in VALID_RESULTADOS:
            motor = motor_de_regras(situacao, resultado,
                                    rec.get('estagio_original'),
                                    rec.get('tentativa_original'))
            ws.cell(row=r, column=13, value=situacao)
            ws.cell(row=r, column=14, value=motor.get('estagio_funil', ''))
            ws.cell(row=r, column=15, value=motor.get('fase', ''))
            ws.cell(row=r, column=16, value=motor.get('tipo_contato', ''))
            ws.cell(row=r, column=17, value=motor.get('temperatura', ''))
            ws.cell(row=r, column=18, value=motor.get('tentativa') or '')
            ws.cell(row=r, column=19, value=motor.get('grupo_dash', ''))
            if rec['data'] and motor.get('follow_up_dias', 0) > 0:
                ws.cell(row=r, column=20, value=dia_util(rec['data'], motor['follow_up_dias'])).number_format = FMT_DATE
            ws.cell(row=r, column=21, value=motor.get('acao_futura', ''))
            ws.cell(row=r, column=22, value=motor.get('acao_detalhada', ''))
        else:
            ws.cell(row=r, column=13, value=situacao)

    last_row = 3 + len(records) - 1
    print(f"    → LOG: rows 3-{last_row} populated (archive, no formulas)")
    return last_row


def populate_carteira(wb, clients):
    """Populate CARTEIRA with XLOOKUP formulas referencing DRAFT 1."""
    ws = wb["CARTEIRA"]
    print(f"  Populating CARTEIRA with {len(clients)} clients...")

    for i, c in enumerate(clients):
        r = 4 + i  # data starts at row 4

        # ═══ MEGA-BLOCO 1: MERCOS (cols 1-43) ═══
        # ANCORA: NOME FANTASIA (col 1) — plain value
        ws.cell(row=r, column=1, value=c['nome_fantasia'])

        # IDENTIDADE (cols 2-8) — plain values from DRAFT 1
        ws.cell(row=r, column=2, value=c['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=c['razao_social'])
        ws.cell(row=r, column=4, value=c['uf'])
        ws.cell(row=r, column=5, value=c['cidade'])
        ws.cell(row=r, column=6, value=c['email'])
        ws.cell(row=r, column=7, value=c['telefone'])
        if c['data_cadastro']:
            ws.cell(row=r, column=8, value=c['data_cadastro']).number_format = FMT_DATE

        # REDE (cols 9-11)
        ws.cell(row=r, column=9, value=c['rede_regional'])
        ws.cell(row=r, column=10, value='')  # REDE GRUPO — empty for now
        if c['ult_registro_mercos']:
            ws.cell(row=r, column=11, value=c['ult_registro_mercos']).number_format = FMT_DATE

        # EQUIPE (cols 12-13)
        ws.cell(row=r, column=12, value=c['consultor'])
        ws.cell(row=r, column=13, value=c['vendedor_ult_pedido'])

        # STATUS (cols 14-15)
        ws.cell(row=r, column=14, value=c['situacao'])
        prio_map = {"ATIVO": 1, "EM RISCO": 2, "INAT.REC": 3, "INAT.ANT": 4, "PROSPECT": 5, "NOVO": 6}
        ws.cell(row=r, column=15, value=prio_map.get(c['situacao'], 99))

        # COMPRA (cols 16-18)
        ws.cell(row=r, column=16, value=c['dias_sem_compra'])
        if c['data_ult_pedido']:
            ws.cell(row=r, column=17, value=c['data_ult_pedido']).number_format = FMT_DATE
        if c['valor_ult_pedido']:
            ws.cell(row=r, column=18, value=c['valor_ult_pedido']).number_format = FMT_MONEY

        # ECOMMERCE (cols 19-24)
        ws.cell(row=r, column=19, value=c['acessos_semana'])
        ws.cell(row=r, column=20, value=c['acesso_b2b'])
        ws.cell(row=r, column=21, value=c['acessos_portal'])
        ws.cell(row=r, column=22, value=c['itens_carrinho'])
        if c['valor_b2b']:
            ws.cell(row=r, column=23, value=c['valor_b2b']).number_format = FMT_MONEY
        ws.cell(row=r, column=24, value=c['oportunidade'])

        # VENDAS (cols 25-36 = total + 11 months)
        if c['total_periodo']:
            ws.cell(row=r, column=25, value=c['total_periodo']).number_format = FMT_MONEY
        for j, venda in enumerate(c['vendas_mes']):
            if venda:
                ws.cell(row=r, column=26 + j, value=venda).number_format = FMT_MONEY

        # RECORRÊNCIA (cols 37-43)
        ws.cell(row=r, column=37, value=c['ciclo_medio'])
        ws.cell(row=r, column=38, value=c['n_compras'])
        ws.cell(row=r, column=39, value=c['curva_abc'])
        ws.cell(row=r, column=40, value=c['meses_positivado'])
        if c['media_mensal']:
            ws.cell(row=r, column=41, value=c['media_mensal']).number_format = FMT_MONEY
        if c['ticket_medio']:
            ws.cell(row=r, column=42, value=c['ticket_medio']).number_format = FMT_MONEY
        ws.cell(row=r, column=43, value=c['meses_lista'])

        # ═══ MEGA-BLOCO 2: FUNIL (cols 44-61) — XLOOKUP from DRAFT 2 ═══
        # These pull latest record for this CNPJ from DRAFT 2
        # ESTÁGIO FUNIL (col 44) — formula
        ws.cell(row=r, column=44).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 2\'!$B$3:$B$20000,\'DRAFT 2\'!$N$3:$N$20000,"",,2),"")'
        )
        # PRÓX. FOLLOW-UP (col 45)
        ws.cell(row=r, column=45).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 2\'!$B$3:$B$20000,\'DRAFT 2\'!$T$3:$T$20000,"",,2),"")'
        )
        ws.cell(row=r, column=45).number_format = FMT_DATE
        # DATA ÚLT.ATENDIMENTO (col 46)
        ws.cell(row=r, column=46).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 2\'!$B$3:$B$20000,\'DRAFT 2\'!$A$3:$A$20000,"",,2),"")'
        )
        ws.cell(row=r, column=46).number_format = FMT_DATE
        # ÚLTIMO RESULTADO (col 47)
        ws.cell(row=r, column=47).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 2\'!$B$3:$B$20000,\'DRAFT 2\'!$F$3:$F$20000,"",,2),"")'
        )
        # MOTIVO (col 48)
        ws.cell(row=r, column=48).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 2\'!$B$3:$B$20000,\'DRAFT 2\'!$G$3:$G$20000,"",,2),"")'
        )
        # TIPO CLIENTE (col 49) — from DRAFT 1
        mp = c['meses_positivado'] or 0
        if isinstance(mp, (int, float)):
            if mp == 0: tipo_cl = "PROSPECT"
            elif mp == 1: tipo_cl = "NOVO"
            elif mp <= 3: tipo_cl = "EM DESENVOLVIMENTO"
            elif mp <= 6: tipo_cl = "RECORRENTE"
            else: tipo_cl = "FIDELIZADO"
        else:
            tipo_cl = "PROSPECT"
        ws.cell(row=r, column=49, value=tipo_cl)
        # TENTATIVA (col 50)
        ws.cell(row=r, column=50).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 2\'!$B$3:$B$20000,\'DRAFT 2\'!$R$3:$R$20000,"",,2),"")'
        )
        # FASE (col 51)
        ws.cell(row=r, column=51).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 2\'!$B$3:$B$20000,\'DRAFT 2\'!$O$3:$O$20000,"",,2),"")'
        )
        # ÚLTIMA RECOMPRA (col 52) — date of last VENDA / PEDIDO
        ws.cell(row=r, column=52, value='')
        # TEMPERATURA (col 53)
        ws.cell(row=r, column=53).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 2\'!$B$3:$B$20000,\'DRAFT 2\'!$Q$3:$Q$20000,"",,2),"")'
        )
        # DIAS ATÉ CONVERSÃO (col 54)
        ws.cell(row=r, column=54, value='')
        # DATA 1º CONTATO (col 55)
        ws.cell(row=r, column=55, value='')
        # DATA 1º ORÇAMENTO (col 56)
        ws.cell(row=r, column=56, value='')
        # DATA 1ª VENDA (col 57)
        ws.cell(row=r, column=57, value='')
        # TOTAL TENTATIVAS (col 58)
        ws.cell(row=r, column=58, value='')
        # PRÓX. AÇÃO (col 59)
        ws.cell(row=r, column=59).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 2\'!$B$3:$B$20000,\'DRAFT 2\'!$U$3:$U$20000,"",,2),"")'
        )
        # AÇÃO DETALHADA (col 60)
        ws.cell(row=r, column=60).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 2\'!$B$3:$B$20000,\'DRAFT 2\'!$V$3:$V$20000,"",,2),"")'
        )
        # SINALEIRO (col 61)
        ws.cell(row=r, column=61).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 2\'!$B$3:$B$20000,\'DRAFT 2\'!$W$3:$W$20000,"",,2),"")'
        )

        # ═══ MEGA-BLOCO 3: SAP (cols 62-72) ═══
        ws.cell(row=r, column=62, value=c['cod_cliente_sap'])
        ws.cell(row=r, column=63, value=c['desc_grupo_cliente'])
        ws.cell(row=r, column=64, value=c['gerente_nacional'])
        ws.cell(row=r, column=65, value=c['representante'])
        ws.cell(row=r, column=66, value=c['vend_interno'])
        ws.cell(row=r, column=67, value=c['nome_canal'])
        ws.cell(row=r, column=68, value=c['tipo_cliente_sap'])
        ws.cell(row=r, column=69, value=c['macroregiao'])
        ws.cell(row=r, column=70, value=c['microregiao'])
        ws.cell(row=r, column=71, value=c['grupo_chave'])
        if c['venda_sap']:
            ws.cell(row=r, column=72, value=c['venda_sap']).number_format = FMT_MONEY

        # ═══ MEGA-BLOCO 4: ACOMPANHAMENTO (cols 73-257) ═══
        # Left empty — will be populated from META SAP later

    last_row = 4 + len(clients) - 1
    print(f"    → CARTEIRA: rows 4-{last_row} populated (Mercos+Funil+SAP, Acomp empty)")
    return last_row


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    start = time.time()
    print("=" * 60)
    print("  CRM VITAO360 V3 — POPULATE with Real Data")
    print("=" * 60)

    # Step 0: Build empty V3 structure
    print("\n[0/5] Building empty V3 structure...")
    # Import and run build
    from v3_regras import build_regras
    from v3_draft1 import build_draft1
    from v3_projecao import build_projecao
    from v3_carteira import build_carteira
    from v3_draft2 import build_draft2
    from v3_agenda import build_agenda
    from v3_dash import build_dash
    from v3_log import build_log, build_claude_log

    wb = openpyxl.Workbook()
    build_regras(wb)
    build_draft1(wb)
    build_projecao(wb)
    build_carteira(wb)
    build_draft2(wb)
    build_agenda(wb)
    build_dash(wb)
    build_log(wb)
    build_claude_log(wb)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Step 1: Read source data
    print("\n[1/5] Reading source data...")
    clients = read_carteira_oular()
    records = read_log_atendimentos()

    # Step 2: Populate DRAFT 1
    print("\n[2/5] Populating DRAFT 1...")
    populate_draft1(wb, clients)

    # Step 3: Populate DRAFT 2 (with motor de regras)
    print("\n[3/5] Populating DRAFT 2 + motor de regras...")
    populate_draft2(wb, records)

    # Step 4: Populate LOG (archive)
    print("\n[4/5] Populating LOG (archive)...")
    populate_log(wb, records)

    # Step 5: Populate CARTEIRA
    print("\n[5/5] Populating CARTEIRA...")
    populate_carteira(wb, clients)

    # Set CARTEIRA as active
    for i, name in enumerate(wb.sheetnames):
        if name == "CARTEIRA":
            wb.active = i
            break

    # Save
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    elapsed = time.time() - start

    # Stats
    print("\n" + "=" * 60)
    print(f"  SAVED: {OUTPUT_FILE}")
    print(f"  Tabs: {wb.sheetnames}")
    print(f"  DRAFT 1: {len(clients)} clients")
    print(f"  DRAFT 2: {len(records)} records (motor applied)")
    print(f"  LOG: {len(records)} records (archive)")
    print(f"  CARTEIRA: {len(clients)} clients (Mercos+Funil+SAP)")
    print(f"  Time: {elapsed:.1f}s")
    print("=" * 60)


if __name__ == "__main__":
    main()

"""
V3 Styles — Shared formatting constants for CRM VITAO360 V3
"""
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ═══════════════════════════════════════════════════════════════
# FONTS
# ═══════════════════════════════════════════════════════════════
FONT_HEADER = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
FONT_HEADER_SM = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
FONT_SUB_HEADER = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
FONT_TITLE = Font(name='Calibri', size=11, bold=True)
FONT_TITLE_W = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
FONT_DATA = Font(name='Calibri', size=10)
FONT_DATA_BOLD = Font(name='Calibri', size=10, bold=True)
FONT_MEGA = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
FONT_BLOCK = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

# ═══════════════════════════════════════════════════════════════
# FILLS
# ═══════════════════════════════════════════════════════════════
# Main header
FILL_HEADER = PatternFill('solid', fgColor='2F5496')
FILL_SUB_HEADER = PatternFill('solid', fgColor='4472C4')

# Mega-block fills
FILL_MERCOS = PatternFill('solid', fgColor='7B2FF2')     # 🟣 purple
FILL_FUNIL = PatternFill('solid', fgColor='2196F3')      # 🔵 blue
FILL_SAP = PatternFill('solid', fgColor='404040')        # ⚫ dark
FILL_ACOMP = PatternFill('solid', fgColor='00B050')      # 🟢 green

# Sub-block fills (lighter)
FILL_BLOCK_IDENTIDADE = PatternFill('solid', fgColor='34495E')
FILL_BLOCK_STATUS = PatternFill('solid', fgColor='1565C0')
FILL_BLOCK_COMPRA = PatternFill('solid', fgColor='2E7D32')
FILL_BLOCK_ECOMMERCE = PatternFill('solid', fgColor='E65100')
FILL_BLOCK_VENDAS = PatternFill('solid', fgColor='6A1B9A')
FILL_BLOCK_RECORRENCIA = PatternFill('solid', fgColor='00695C')
FILL_BLOCK_ATENDIMENTO = PatternFill('solid', fgColor='4E342E')
FILL_BLOCK_PIPELINE = PatternFill('solid', fgColor='0D47A1')
FILL_BLOCK_PERFIL = PatternFill('solid', fgColor='1B5E20')
FILL_BLOCK_MATURIDADE = PatternFill('solid', fgColor='4A148C')
FILL_BLOCK_CONVERSAO = PatternFill('solid', fgColor='BF360C')
FILL_BLOCK_ACAO = PatternFill('solid', fgColor='263238')
FILL_BLOCK_SINAL = PatternFill('solid', fgColor='880E4F')

# SITUACAO cell colors
FILL_ATIVO = PatternFill('solid', fgColor='00B050')
FILL_EM_RISCO = PatternFill('solid', fgColor='FFC000')
FILL_INAT_REC = PatternFill('solid', fgColor='FFC000')
FILL_INAT_ANT = PatternFill('solid', fgColor='FF0000')
FILL_NOVO = PatternFill('solid', fgColor='0070C0')
FILL_PROSPECT = PatternFill('solid', fgColor='7B2FF2')

SITUACAO_FILLS = {
    "ATIVO": FILL_ATIVO,
    "EM RISCO": FILL_EM_RISCO,
    "INAT.REC": FILL_INAT_REC,
    "INAT.ANT": FILL_INAT_ANT,
    "NOVO": FILL_NOVO,
    "PROSPECT": FILL_PROSPECT,
}

# General fills
FILL_LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
FILL_GRAY_D9 = PatternFill('solid', fgColor='D9D9D9')
FILL_DARK = PatternFill('solid', fgColor='404040')
FILL_DARK_59 = PatternFill('solid', fgColor='595959')
FILL_RED_DARK = PatternFill('solid', fgColor='8B0000')
FILL_GREEN_DARK = PatternFill('solid', fgColor='2E7D32')
FILL_SEP = PatternFill('solid', fgColor='F2F2F2')
FILL_INPUT = PatternFill('solid', fgColor='DBEAFE')  # light blue for input cells
FILL_CALC = PatternFill('solid', fgColor='E8F5E9')   # light green for calculated
FILL_NONE = PatternFill(fill_type=None)

# ═══════════════════════════════════════════════════════════════
# BORDERS
# ═══════════════════════════════════════════════════════════════
THIN_BORDER = Border(
    left=Side(style='thin', color='BBBBBB'),
    right=Side(style='thin', color='BBBBBB'),
    top=Side(style='thin', color='BBBBBB'),
    bottom=Side(style='thin', color='BBBBBB'),
)
NO_BORDER = Border()

# ═══════════════════════════════════════════════════════════════
# ALIGNMENTS
# ═══════════════════════════════════════════════════════════════
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)

# ═══════════════════════════════════════════════════════════════
# NUMBER FORMATS
# ═══════════════════════════════════════════════════════════════
FMT_DATE = 'DD/MM/YYYY'
FMT_MONEY = '#,##0.00'
FMT_PCT = '0%'
FMT_NUMBER = '#,##0'
FMT_TEXT = '@'

# ═══════════════════════════════════════════════════════════════
# TAB COLORS
# ═══════════════════════════════════════════════════════════════
TAB_REGRAS = '0D9488'
TAB_DRAFT1 = '7B2FF2'
TAB_PROJECAO = 'FF6B00'
TAB_CARTEIRA = '00B050'
TAB_DRAFT2 = 'FFC000'
TAB_AGENDA = '2196F3'
TAB_DASH = 'DC2626'
TAB_LOG = '6B7280'
TAB_CLAUDE = '9333EA'

# ═══════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def style_cell(cell, font=FONT_DATA, fill=None, border=THIN_BORDER, align=ALIGN_CENTER, fmt=None):
    """Apply complete style to a cell."""
    cell.font = font
    if fill:
        cell.fill = fill
    cell.border = border
    cell.alignment = align
    if fmt:
        cell.number_format = fmt


def write_header(ws, row, col, value, fill=FILL_HEADER, font=FONT_HEADER, align=ALIGN_CENTER):
    """Write a header cell with dark background + white text."""
    cell = ws.cell(row=row, column=col, value=value)
    style_cell(cell, font=font, fill=fill, align=align)
    return cell


def write_data(ws, row, col, value, font=FONT_DATA, fill=None, align=ALIGN_CENTER, fmt=None):
    """Write a data cell."""
    cell = ws.cell(row=row, column=col, value=value)
    style_cell(cell, font=font, fill=fill, align=align, fmt=fmt)
    return cell


def group_columns(ws, start_col, end_col, outline_level=1, hidden=True):
    """Group columns with proper outline level on each column."""
    from openpyxl.utils import get_column_letter
    for col_idx in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].outlineLevel = outline_level
        ws.column_dimensions[col_letter].hidden = hidden

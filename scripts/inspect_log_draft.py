#!/usr/bin/env python3
"""Inspect preenchimento do fraft de atendimento (LOG).xlsx"""
import os, sys
from datetime import datetime, date
from collections import Counter

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl missing"); sys.exit(1)

FILE_PATH = r'C:\Users\User\OneDrive\Área de Trabalho\preenchimento do fraft de atendimento (LOG).xlsx'
EXPECTED = ["DATA","CNPJ","NOME FANTASIA","UF","CONSULTOR","RESULTADO","MOTIVO","WHATSAPP","LIGACAO","LIG. ATENDIDA","NOTA DO DIA","MERCOS ATUALIZADO","SITUACAO","ESTAGIO FUNIL","FASE","TIPO DO CONTATO","TEMPERATURA","TENTATIVA","GRUPO DASH","FOLLOW-UP","ACAO FUTURA","ACAO DETALHADA","SINALEIRO CICLO","SINALEIRO META"]
SEP = "=" * 100
SUB = "-" * 80

def safe_str(val):
    if val is None: return "<vazio>"
    if isinstance(val, datetime): return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, date): return val.strftime("%Y-%m-%d")
    return str(val)

def norm(name):
    if name is None: return None
    s = str(name).strip().upper()
    reps = {
        "\u00c7":"C","\u00e7":"C","\u00c3":"A","\u00e3":"A",
        "\u00c1":"A","\u00e1":"A","\u00c0":"A","\u00e0":"A",
        "\u00c2":"A","\u00e2":"A","\u00c9":"E","\u00e9":"E",
        "\u00ca":"E","\u00ea":"E","\u00cd":"I","\u00ed":"I",
        "\u00d3":"O","\u00f3":"O","\u00d4":"O","\u00f4":"O",
        "\u00d5":"O","\u00f5":"O","\u00da":"U","\u00fa":"U"
    }
    for o, n in reps.items(): s = s.replace(o, n)
    return s

def inspect_sheet(ws, sname, formula_mode=False):
    mode = "FORMULA" if formula_mode else "DATA"
    print(f"\n{SEP}")
    print(f"  SHEET: {sname!r} [{mode} MODE]")
    print(SEP)
    mr = ws.max_row or 0
    mc = ws.max_column or 0
    print(f"  Rows: {mr} | Cols: {mc} | Dim: {ws.dimensions}")
    if mr == 0 or mc == 0:
        print("  ** EMPTY **"); return {}
    print(f"\n{SUB}\n  FREEZE: {ws.freeze_panes}")
    print(f"\n{SUB}\n  MERGED:")
    if ws.merged_cells.ranges:
        for m in ws.merged_cells.ranges: print(f"    {m}")
    else: print("    (none)")
    print(f"\n{SUB}\n  COLUMNS:")
    headers = {}
    for ri in range(1, min(6, mr + 1)):
        rv = [ws.cell(row=ri, column=ci).value for ci in range(1, mc + 1)]
        ne = [v for v in rv if v is not None]
        sv = [v for v in ne if isinstance(v, str)]
        if len(ne) > 0 and len(sv) / max(len(ne), 1) > 0.5:
            print(f"\n  >> Row {ri} = HEADER:")
            for ci in range(1, mc + 1):
                v = ws.cell(row=ri, column=ci).value
                lt = get_column_letter(ci)
                print(f"    {lt} ({ci:>3}): {safe_str(v)}")
                if ri == 1: headers[ci] = v
            if not headers:
                for ci in range(1, mc + 1): headers[ci] = ws.cell(row=ri, column=ci).value
    if not headers:
        print("\n  >> Row 1 fallback:")
        for ci in range(1, mc + 1):
            headers[ci] = ws.cell(row=1, column=ci).value
            print(f"    {get_column_letter(ci)} ({ci:>3}): {safe_str(headers[ci])}")
    ds = 2
    print(f"\n{SUB}\n  FIRST 10 DATA ROWS:")
    for ri in range(ds, min(ds + 10, mr + 1)):
        print(f"\n  --- Row {ri} ---")
        for ci in range(1, mc + 1):
            v = ws.cell(row=ri, column=ci).value
            h = safe_str(headers.get(ci, f"Col{ci}"))
            print(f"    [{get_column_letter(ci)}] {h}: {safe_str(v)}")
    print(f"\n{SUB}")
    ls_row = max(ds, mr - 4)
    print(f"  LAST 5 DATA ROWS ({ls_row}-{mr}):")
    for ri in range(ls_row, mr + 1):
        print(f"\n  --- Row {ri} ---")
        for ci in range(1, mc + 1):
            v = ws.cell(row=ri, column=ci).value
            h = safe_str(headers.get(ci, f"Col{ci}"))
            print(f"    [{get_column_letter(ci)}] {h}: {safe_str(v)}")
    ner = 0
    for ri in range(ds, mr + 1):
        for ci in range(1, mc + 1):
            if ws.cell(row=ri, column=ci).value is not None: ner += 1; break
    print(f"\n{SUB}\n  NON-EMPTY DATA ROWS: {ner}")
    print(f"\n{SUB}\n  DATA VALIDATIONS:")
    if ws.data_validations and ws.data_validations.dataValidation:
        for dv in ws.data_validations.dataValidation:
            print(f"    Range: {dv.sqref} | Type: {dv.type} | F1: {dv.formula1} | F2: {dv.formula2}")
            print(f"      AllowBlank: {dv.allow_blank} | ShowErr: {dv.showErrorMessage}")
    else: print("    (none)")
    print(f"\n{SUB}\n  CONDITIONAL FORMATTING:")
    if ws.conditional_formatting:
        for cf in ws.conditional_formatting:
            print(f"    Range: {cf}")
            for rule in cf.rules:
                print(f"      Type: {rule.type} | Pri: {rule.priority}")
                if rule.formula: print(f"      Formula: {rule.formula}")
    else: print("    (none)")
    if formula_mode:
        print(f"\n{SUB}\n  FORMULAS:")
        fc = 0; fcols = set()
        for ri in range(ds, mr + 1):
            for ci in range(1, mc + 1):
                v = ws.cell(row=ri, column=ci).value
                if isinstance(v, str) and v.startswith("="):
                    lt = get_column_letter(ci)
                    h = safe_str(headers.get(ci, f"Col{ci}"))
                    if fc < 30: print(f"    {lt}{ri} [{h}]: {v}")
                    fc += 1; fcols.add(f"{lt} [{h}]")
        if fc > 30: print(f"    ... ({fc} total)")
        if fc == 0: print("    (none)")
        elif fc > 0: print(f"\n  FORMULA COLS: {sorted(fcols)}")
    return headers

def get_col_values(ws, headers, cn):
    tc = None; nn = norm(cn)
    for ci, h in headers.items():
        if h and norm(h) == nn: tc = ci; break
    if tc is None:
        for ci, h in headers.items():
            if h and nn in norm(h): tc = ci; break
    if tc is None: return None, None
    vals = []
    for ri in range(2, (ws.max_row or 0) + 1):
        v = ws.cell(row=ri, column=tc).value
        if v is not None: vals.append(v)
    return vals, sorted(set(safe_str(v) for v in vals))

def get_dates(ws, headers):
    tc = None
    for ci, h in headers.items():
        if h and norm(h) == "DATA": tc = ci; break
    if tc is None: return None, None
    dts = []
    for ri in range(2, (ws.max_row or 0) + 1):
        v = ws.cell(row=ri, column=tc).value
        if isinstance(v, (datetime, date)): dts.append(v)
    return (min(dts), max(dts)) if dts else (None, None)

def main():
    print("#" * 100)
    print(f"  INSPECTING: {FILE_PATH}")
    print(f"  Run: {datetime.now()}")
    print("#" * 100)
    if not os.path.exists(FILE_PATH):
        print(f"ERROR: not found: {FILE_PATH}"); sys.exit(1)
    fs = os.path.getsize(FILE_PATH)
    print(f"  Size: {fs:,} bytes ({fs/1024:.1f} KB)")

    print("\n" + "*" * 100)
    print("  DATA MODE (data_only=True)")
    print("*" * 100)
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    print(f"  Sheets: {wb.sheetnames} ({len(wb.sheetnames)})")
    ah = {}
    for sn in wb.sheetnames: ah[sn] = inspect_sheet(wb[sn], sn, False)
    wb.close()

    print("\n" + "*" * 100)
    print("  FORMULA MODE (data_only=False)")
    print("*" * 100)
    wb2 = openpyxl.load_workbook(FILE_PATH, data_only=False)
    for sn in wb2.sheetnames: inspect_sheet(wb2[sn], sn, True)
    wb2.close()

    print("\n" + "*" * 100)
    print("  VALUE ANALYSIS")
    print("*" * 100)
    wb3 = openpyxl.load_workbook(FILE_PATH, data_only=True)
    for sn in wb3.sheetnames:
        ws = wb3[sn]
        hdrs = ah.get(sn, {})
        if not hdrs:
            for ci in range(1, (ws.max_column or 0)+1): hdrs[ci] = ws.cell(row=1, column=ci).value
        print(f"\n{SEP}\n  ANALYSIS: {sn}\n{SEP}")
        print("\n  ALL COLUMNS:")
        for ci in range(1, (ws.max_column or 0)+1):
            print(f"    {get_column_letter(ci)} ({ci:>3}): {safe_str(ws.cell(row=1, column=ci).value)}")
        for cn in ["RESULTADO", "MOTIVO", "CONSULTOR"]:
            vs, uq = get_col_values(ws, hdrs, cn)
            if uq is not None:
                print(f"\n  {cn} unique ({len(uq)}):")
                for v in uq: print(f"    - {v}")
            else: print(f"\n  {cn}: NOT FOUND")
        found_sit = False
        for cn in ["SITUACAO", "STATUS"]:
            vs, uq = get_col_values(ws, hdrs, cn)
            if uq is not None:
                print(f"\n  SITUACAO unique ({len(uq)}):")
                for v in uq: print(f"    - {v}")
                found_sit = True; break
        if not found_sit: print("\n  SITUACAO: NOT FOUND")
        mn, mx = get_dates(ws, hdrs)
        if mn: print(f"\n  DATA range: {safe_str(mn)} to {safe_str(mx)}")
        else: print("\n  DATA: no dates")
        vs, _ = get_col_values(ws, hdrs, "CONSULTOR")
        if vs:
            ct = Counter(safe_str(v) for v in vs)
            print("\n  PER CONSULTOR:")
            for n, c in ct.most_common(): print(f"    {n}: {c}")
        print("\n  AUTO vs MANUAL:")
        wbf = openpyxl.load_workbook(FILE_PATH, data_only=False)
        wsf = wbf[sn]
        for ci in range(1, (wsf.max_column or 0)+1):
            fmc = vmc = 0
            for ri in range(2, min((wsf.max_row or 0)+1, 100)):
                v = wsf.cell(row=ri, column=ci).value
                if v is not None:
                    if isinstance(v, str) and v.startswith("="): fmc += 1
                    else: vmc += 1
            h = safe_str(hdrs.get(ci, f"Col{ci}"))
            lt = get_column_letter(ci)
            if fmc > 0: print(f"    {lt} [{h}]: {fmc}f/{vmc}m -> AUTO")
            elif vmc > 0: print(f"    {lt} [{h}]: {vmc}m -> MANUAL")
            else: print(f"    {lt} [{h}]: EMPTY")
        wbf.close()
    wb3.close()

    print("\n" + "*" * 100)
    print("  COMPARISON: FILE vs V3 DRAFT 2")
    print("*" * 100)
    for sn, hdrs in ah.items():
        print(f"\n{SEP}\n  SHEET: {sn}\n{SEP}")
        fcn = []; fcr = []
        for ci in sorted(hdrs.keys()):
            v = hdrs[ci]
            if v is not None: fcn.append(norm(v)); fcr.append(str(v).strip())
        print(f"\n  File: {len(fcn)} cols | V3: {len(EXPECTED)} cols")
        fs2 = set(fcn); es = set(EXPECTED)
        extra = fs2 - es; miss = es - fs2
        print(f"\n  IN FILE NOT IN V3 ({len(extra)}):")
        if extra:
            for c in sorted(extra):
                i = fcn.index(c) if c in fcn else -1
                r = fcr[i] if i >= 0 else c
                print(f"    + {r} (norm: {c})")
        else: print("    (none)")
        print(f"\n  IN V3 NOT IN FILE ({len(miss)}):")
        if miss:
            for c in sorted(miss):
                cl = [f for f in fcn if c in f or f in c]
                if cl: print(f"    - {c} (close: {cl})")
                else: print(f"    - {c}")
        else: print("    (none)")
        print(f"\n  MATCHED ({len(fs2 & es)}):")
        for c in EXPECTED:
            if c in fs2: print(f"    = {c}")
        print("\n  ORDER:")
        mx2 = max(len(EXPECTED), len(fcr))
        for i in range(mx2):
            e = EXPECTED[i] if i < len(EXPECTED) else "(n/a)"
            ac = fcr[i] if i < len(fcr) else "(n/a)"
            ne2 = e; na2 = norm(ac) if ac != "(n/a)" else "(n/a)"
            fl = " <--DIFF" if ne2 != na2 else ""
            print(f"    {e:<30} | {ac:<30}{fl}")

    print("\n" + "#" * 100)
    print("  FINAL SUMMARY")
    print("#" * 100)
    print(f"  File: {FILE_PATH}")
    print(f"  Sheets: {list(ah.keys())}")
    for sn, hdrs in ah.items():
        wbt = openpyxl.load_workbook(FILE_PATH, data_only=True)
        wst = wbt[sn]
        cns = [safe_str(hdrs.get(c,"?")) for c in sorted(hdrs.keys()) if hdrs.get(c) is not None]
        ne2 = 0
        for r in range(2, (wst.max_row or 0)+1):
            for c in range(1, (wst.max_column or 0)+1):
                if wst.cell(row=r, column=c).value is not None: ne2 += 1; break
        print(f"\n  Sheet {sn!r}:")
        print(f"    Cols ({len(cns)}): {cns}")
        print(f"    Data rows: {ne2} | Total: {wst.max_row}")
        wbt.close()
    print("\n  DONE.")

if __name__ == "__main__": main()

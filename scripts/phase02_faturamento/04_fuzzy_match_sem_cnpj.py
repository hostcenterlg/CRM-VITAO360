#!/usr/bin/env python3
"""
Phase 02, Plan 02, Task 1b: Fuzzy Match dos 10 clientes Mercos sem CNPJ
========================================================================
Attempts to resolve the 10 Mercos clients without CNPJ by matching their
names against:
  1. Mercos Carteira (08_CARTEIRA_MERCOS.xlsx) - Nome Fantasia + Razao Social
  2. SAP Cadastro (01_SAP_CONSOLIDADO.xlsx) - Nome Cliente

Matching strategy (no external libraries):
  a. Exact match (uppercase, stripped)
  b. Partial match (one name contained in the other)
  c. Special cases (CPF/CNPJ embedded in name, city suffix)

Input:
  data/output/phase02/sap_mercos_merged.json (from Task 1a)
  data/output/phase02/mercos_vendas.json (sem_cnpj list)
  data/sources/mercos/08_CARTEIRA_MERCOS.xlsx
  data/sources/sap/01_SAP_CONSOLIDADO.xlsx

Output:
  data/output/phase02/sap_mercos_merged.json (updated with fuzzy matches)
"""

import json
import re
import openpyxl
from pathlib import Path

# Project root
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent


def normalize_cnpj(raw):
    """Normalize CNPJ to 14-digit string."""
    return re.sub(r'[^0-9]', '', str(raw)).zfill(14)


def load_json(path):
    """Load a JSON file and return parsed data."""
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_json(data, path):
    """Save data to JSON file."""
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def build_mercos_lookup():
    """Build name->CNPJ lookup from Mercos Carteira.

    Returns dict: uppercase_name -> (cnpj_14digits, match_field)
    Uses both Razao Social and Nome Fantasia as keys.
    """
    path = PROJECT_ROOT / "data" / "sources" / "mercos" / "08_CARTEIRA_MERCOS.xlsx"
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["Carteira Clientes Mercos"]

    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        razao = str(row[0] or '').strip().upper()
        fantasia = str(row[1] or '').strip().upper()
        cnpj_raw = str(row[2] or '').strip()

        if cnpj_raw and cnpj_raw.lower() != 'none':
            cnpj = normalize_cnpj(cnpj_raw)
            if len(re.sub(r'[^0-9]', '', cnpj_raw)) >= 11:  # valid CNPJ or CPF
                if razao:
                    lookup[razao] = (cnpj, "razao_social_mercos")
                if fantasia:
                    lookup[fantasia] = (cnpj, "nome_fantasia_mercos")

    wb.close()
    return lookup


def build_sap_lookup():
    """Build name->CNPJ lookup from SAP Cadastro.

    Returns dict: uppercase_name -> (cnpj_14digits, match_field)
    """
    path = PROJECT_ROOT / "data" / "sources" / "sap" / "01_SAP_CONSOLIDADO.xlsx"
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["Cadastro Clientes SAP"]

    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cnpj_raw = str(row[4] or '').strip()  # col E = CNPJ Cliente
        nome = str(row[5] or '').strip().upper()  # col F = Nome Cliente

        if cnpj_raw and cnpj_raw.lower() != 'none':
            cnpj = normalize_cnpj(cnpj_raw)
            if len(re.sub(r'[^0-9]', '', cnpj_raw)) >= 11:
                if nome:
                    lookup[nome] = (cnpj, "nome_sap")

    wb.close()
    return lookup


def try_match(nome, mercos_lookup, sap_lookup):
    """Try to match a name against both lookups.

    Strategy:
    1. Exact match in Mercos Carteira (Nome Fantasia + Razao Social)
    2. Exact match in SAP Cadastro
    3. Partial match: name contained in lookup key or vice-versa (Mercos first)
    4. Special: extract CNPJ/CPF digits from name prefix

    Returns: (cnpj, match_type, match_detail) or (None, None, reason)
    """
    nome_up = nome.strip().upper()

    # 1. Exact match in Mercos
    if nome_up in mercos_lookup:
        cnpj, field = mercos_lookup[nome_up]
        return cnpj, "exact_name", f"Mercos {field}"

    # 2. Exact match in SAP
    if nome_up in sap_lookup:
        cnpj, field = sap_lookup[nome_up]
        return cnpj, "exact_name", f"SAP {field}"

    # 3. Partial match in Mercos (name contained in key or key contained in name)
    best_mercos = None
    best_mercos_len = 0
    for key, (cnpj, field) in mercos_lookup.items():
        if nome_up in key or key in nome_up:
            # Prefer longer match (more specific)
            match_len = min(len(nome_up), len(key))
            if match_len > best_mercos_len:
                best_mercos = (cnpj, field, key)
                best_mercos_len = match_len

    if best_mercos:
        cnpj, field, matched_key = best_mercos
        return cnpj, "partial", f"Mercos {field}: '{matched_key}'"

    # 4. Partial match in SAP
    best_sap = None
    best_sap_len = 0
    for key, (cnpj, field) in sap_lookup.items():
        if nome_up in key or key in nome_up:
            match_len = min(len(nome_up), len(key))
            if match_len > best_sap_len:
                best_sap = (cnpj, field, key)
                best_sap_len = match_len

    if best_sap:
        cnpj, field, matched_key = best_sap
        return cnpj, "partial", f"SAP {field}: '{matched_key}'"

    # 5. Special: name starts with digits (possible CNPJ/CPF prefix)
    digits = re.sub(r'[^0-9]', '', nome_up.split()[0] if nome_up.split() else '')
    if len(digits) >= 8:
        cnpj = digits.zfill(14)
        return cnpj, "cpf_cnpj_from_name", f"Extracted from name prefix: {digits}"

    return None, None, "No CNPJ match found in Mercos Carteira or SAP Cadastro"


def fuzzy_match_sem_cnpj():
    """Run fuzzy matching for the 10 clients without CNPJ."""

    print("=" * 70)
    print("FUZZY MATCH: 10 CLIENTES MERCOS SEM CNPJ")
    print("=" * 70)
    print()

    # 1. Load merged JSON from Task 1a
    merged_path = PROJECT_ROOT / "data" / "output" / "phase02" / "sap_mercos_merged.json"
    merged_data = load_json(merged_path)

    # 2. Load sem_cnpj list from Mercos extraction
    mercos_path = PROJECT_ROOT / "data" / "output" / "phase02" / "mercos_vendas.json"
    mercos_data = load_json(mercos_path)
    sem_cnpj = mercos_data["sem_cnpj"]

    print(f"  Clients without CNPJ: {len(sem_cnpj)}")
    print()

    # 3. Build reference lookups
    print("  Loading Mercos Carteira (08_CARTEIRA_MERCOS.xlsx)...")
    mercos_lookup = build_mercos_lookup()
    print(f"    {len(mercos_lookup)} name->CNPJ entries")

    print("  Loading SAP Cadastro (01_SAP_CONSOLIDADO.xlsx)...")
    sap_lookup = build_sap_lookup()
    print(f"    {len(sap_lookup)} name->CNPJ entries")
    print()

    # 4. Process each sem_cnpj client
    fuzzy_matches = []
    unmatched = []

    print("-" * 70)
    print("MATCH RESULTS")
    print("-" * 70)

    for i, client in enumerate(sem_cnpj, 1):
        nome = client["nome"]
        total = client["total"]
        monthly_2025 = client["monthly_2025"]
        jan26 = client.get("jan26", 0.0)

        cnpj, match_type, detail = try_match(nome, mercos_lookup, sap_lookup)

        if cnpj:
            # Matched! Add to fuzzy_matches and merge vendas
            match_record = {
                "nome": nome,
                "matched_cnpj": cnpj,
                "match_type": match_type,
                "match_detail": detail,
                "total_vendas": total,
            }
            fuzzy_matches.append(match_record)

            # Add monthly 2025 values to merged data
            if cnpj in merged_data["cnpj_to_vendas"]:
                # CNPJ already exists -- add sem_cnpj values where existing is 0
                existing = merged_data["cnpj_to_vendas"][cnpj]
                for m_idx in range(12):
                    if existing[m_idx] == 0 and monthly_2025[m_idx] > 0:
                        existing[m_idx] = monthly_2025[m_idx]
                merged_data["cnpj_to_vendas"][cnpj] = existing
            else:
                # New CNPJ -- add entirely
                merged_data["cnpj_to_vendas"][cnpj] = monthly_2025

            # Add JAN/26 if present
            if jan26 > 0:
                if cnpj in merged_data["jan26_vendas"]:
                    # Only add if existing jan26 is 0
                    if merged_data["jan26_vendas"][cnpj] == 0:
                        merged_data["jan26_vendas"][cnpj] = jan26
                else:
                    merged_data["jan26_vendas"][cnpj] = jan26

            status = "MATCHED"
            print(f"  {i:2d}. {nome}")
            print(f"      -> {status} cnpj={cnpj} via {match_type}")
            print(f"         Detail: {detail}")
            print(f"         Total: R$ {total:,.2f}")
        else:
            # Not matched
            unmatched_record = {
                "nome": nome,
                "total": total,
                "reason": detail,
            }
            unmatched.append(unmatched_record)

            status = "UNMATCHED"
            print(f"  {i:2d}. {nome}")
            print(f"      -> {status}")
            print(f"         Reason: {detail}")
            print(f"         Total: R$ {total:,.2f}")

        print()

    # 5. Update merged_data with results
    merged_data["fuzzy_matches"] = fuzzy_matches
    merged_data["unmatched"] = unmatched

    # 6. Recalculate stats
    total_vendas_2025 = 0.0
    for cnpj, vals in merged_data["cnpj_to_vendas"].items():
        total_vendas_2025 += sum(vals)

    total_jan26 = sum(merged_data["jan26_vendas"].values())

    stats = merged_data["stats"]
    stats["fuzzy_matched"] = len(fuzzy_matches)
    stats["unmatched"] = len(unmatched)
    # Recount total clients (some fuzzy matches may have merged into existing CNPJs)
    stats["total_clientes"] = len(merged_data["cnpj_to_vendas"])
    stats["total_vendas_2025"] = round(total_vendas_2025, 2)
    stats["total_jan26"] = round(total_jan26, 2)

    # 7. Save updated JSON
    save_json(merged_data, merged_path)

    # 8. Print summary
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"  Fuzzy matched:    {len(fuzzy_matches)}")
    print(f"  Unmatched:        {len(unmatched)}")
    print(f"  Total processed:  {len(fuzzy_matches) + len(unmatched)} / {len(sem_cnpj)}")
    print()
    print(f"  New total clients: {stats['total_clientes']}")
    print(f"  New total vendas 2025: R$ {stats['total_vendas_2025']:,.2f}")
    print(f"  New total JAN/26: R$ {stats['total_jan26']:,.2f}")
    print()

    # Consistency check
    base = stats["sap_only"] + stats["mercos_only"] + stats["both_sap_base"]
    # Note: fuzzy matched clients may merge into existing CNPJs, so total_clientes
    # may be less than base + fuzzy_matched
    new_cnpjs_from_fuzzy = stats["total_clientes"] - base
    print(f"  Consistency: base({base}) + new_from_fuzzy({new_cnpjs_from_fuzzy}) = {stats['total_clientes']}")
    print(f"  fuzzy_matches + unmatched = {len(fuzzy_matches)} + {len(unmatched)} = {len(fuzzy_matches) + len(unmatched)} (should be {len(sem_cnpj)})")

    assert len(fuzzy_matches) + len(unmatched) == len(sem_cnpj), (
        f"Not all sem_cnpj processed: {len(fuzzy_matches)} + {len(unmatched)} != {len(sem_cnpj)}"
    )
    print("  ALL 10 sem_cnpj clients processed: PASS")
    print()

    # Validate no negative values in added entries
    negs = 0
    for m in fuzzy_matches:
        cnpj = m["matched_cnpj"]
        vals = merged_data["cnpj_to_vendas"].get(cnpj, [])
        for v in vals:
            if v < 0:
                negs += 1
    print(f"  Negative values in fuzzy-matched entries: {negs}")
    print()
    print(f"  Output: {merged_path}")
    print("=" * 70)

    return merged_data


if __name__ == "__main__":
    fuzzy_match_sem_cnpj()

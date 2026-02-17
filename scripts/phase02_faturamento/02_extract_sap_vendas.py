"""
02_extract_sap_vendas.py
Extrai faturado mensal por CNPJ do 01_SAP_CONSOLIDADO.xlsx (aba Venda Mes a Mes).

Output: data/output/phase02/sap_vendas.json

Estrutura do JSON:
- cnpj_to_vendas: dict[cnpj] -> [JAN..DEZ] (12 posicoes, todos os meses)
- stats: contadores e totais
  - total_clientes
  - total_vendas_2025
  - vendas_por_mes: [JAN, FEV, ..., DEZ]
- cross_check: comparacao com Phase 1 output

Estrategia: Re-extracao independente do Excel (nao reutiliza Phase 1 JSON)
para validacao cruzada. Depois compara com sap_data_extracted.json.
"""
import json
import os
import re
import sys
import unicodedata
import openpyxl
from collections import defaultdict

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))
SAP_CONSOLIDADO = os.path.join(PROJECT_ROOT, "data", "sources", "sap", "01_SAP_CONSOLIDADO.xlsx")
PHASE1_SAP_JSON = os.path.join(PROJECT_ROOT, "data", "output", "phase01", "sap_data_extracted.json")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "data", "output", "phase02")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "sap_vendas.json")

# ---------------------------------------------------------------------------
# Month names
# ---------------------------------------------------------------------------
MONTH_NAMES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN",
               "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]

# SAP Faturado column positions (confirmed in RESEARCH):
# Col 7=Jan, 11=Fev, 15=Mar, 19=Abr, 23=Mai, 27=Jun,
# 31=Jul, 35=Ago, 39=Set, 43=Out, 47=Nov, 51=Dez
FAT_COLS = [7, 11, 15, 19, 23, 27, 31, 35, 39, 43, 47, 51]


# ---------------------------------------------------------------------------
# Utility functions (same as Phase 1 / Task 1)
# ---------------------------------------------------------------------------
def normalize_cnpj(raw):
    """Normalize CNPJ to 14-digit zero-padded string."""
    if raw is None:
        return None
    clean = re.sub(r'[^0-9]', '', str(raw))
    if not clean or len(clean) < 11:
        return None
    return clean.zfill(14) if len(clean) <= 14 else clean


def strip_accents(s):
    """Remove accents from string for sheet name matching."""
    return ''.join(
        c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn'
    )


def find_sheet(wb, target_name):
    """Find sheet by name, ignoring accents and trailing spaces."""
    target_clean = strip_accents(target_name).strip().upper()
    for name in wb.sheetnames:
        if strip_accents(name).strip().upper() == target_clean:
            return wb[name]
    # Fallback: partial match
    for name in wb.sheetnames:
        if target_clean in strip_accents(name).strip().upper():
            return wb[name]
    return None


def safe_float(val):
    """Convert value to float, returning 0.0 for None/invalid."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# ---------------------------------------------------------------------------
# Extract SAP Vendas
# ---------------------------------------------------------------------------
def extract_sap_vendas(wb):
    """Extract monthly faturado by CNPJ from 'Venda Mes a Mes' sheet."""
    print("\n--- 1. Venda Mes a Mes (SAP Faturado) ---")

    ws = find_sheet(wb, "Venda Mes a Mes")
    if ws is None:
        print(f"  ERROR: Sheet 'Venda Mes a Mes' not found!")
        print(f"  Available: {wb.sheetnames}")
        sys.exit(1)

    print(f"  Sheet found: '{ws.title}'")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")

    # Aggregate by CNPJ (same CNPJ may appear in multiple rows for different Grupo Chave)
    vendas_por_cnpj = defaultdict(lambda: [0.0] * 12)
    rows_processed = 0
    rows_with_cnpj = 0
    rows_skipped = 0

    for row in range(2, ws.max_row + 1):
        cod_sap = ws.cell(row=row, column=1).value   # A = Codigo SAP
        nome = ws.cell(row=row, column=2).value       # B = Nome
        cnpj_raw = ws.cell(row=row, column=3).value   # C = CNPJ/CPF

        # Skip empty rows
        if cod_sap is None and nome is None and cnpj_raw is None:
            continue

        rows_processed += 1
        cnpj = normalize_cnpj(cnpj_raw)

        if not cnpj:
            rows_skipped += 1
            continue

        rows_with_cnpj += 1

        for i, col in enumerate(FAT_COLS):
            v = ws.cell(row=row, column=col).value
            vendas_por_cnpj[cnpj][i] += safe_float(v)

    # Convert defaultdict to regular dict with rounded values
    result = {k: [round(v, 2) for v in vals] for k, vals in vendas_por_cnpj.items()}

    print(f"  Rows processed: {rows_processed}")
    print(f"  Rows with CNPJ: {rows_with_cnpj}")
    print(f"  Rows skipped (no CNPJ): {rows_skipped}")
    print(f"  Unique CNPJs: {len(result)}")

    # Validate CNPJ format
    bad_cnpjs = [c for c in result.keys() if len(c) < 14]
    if bad_cnpjs:
        print(f"  WARNING: {len(bad_cnpjs)} CNPJs with less than 14 digits!")
    else:
        print(f"  CNPJ format validation: ALL 14+ digits OK")

    # Total sales
    total_sales = sum(sum(v) for v in result.values())
    print(f"  Total sales 2025: R$ {total_sales:,.2f}")

    # Per-month totals
    vendas_por_mes = [0.0] * 12
    for monthly in result.values():
        for i in range(12):
            vendas_por_mes[i] += monthly[i]

    print(f"\n  Vendas por mes:")
    for i, name in enumerate(MONTH_NAMES):
        print(f"    {name}: R$ {vendas_por_mes[i]:,.2f}")

    # Verify JAN > 0 (SAP has ALL 12 months)
    if vendas_por_mes[0] > 0:
        print(f"\n  JAN/25 > 0: CORRECT (SAP has January data)")
    else:
        print(f"\n  WARNING: JAN/25 = 0 -- SAP should have January data!")

    return result, vendas_por_mes


# ---------------------------------------------------------------------------
# Cross-check vs Phase 1
# ---------------------------------------------------------------------------
def cross_check_phase1(vendas_por_cnpj):
    """Compare extraction with Phase 1 sap_data_extracted.json."""
    print("\n--- 2. Cross-check vs Phase 1 ---")

    if not os.path.exists(PHASE1_SAP_JSON):
        print(f"  WARNING: Phase 1 output not found: {PHASE1_SAP_JSON}")
        return {"status": "SKIPPED", "reason": "Phase 1 output not found"}

    with open(PHASE1_SAP_JSON, "r", encoding="utf-8") as f:
        phase1 = json.load(f)

    p1_vendas = phase1.get("cnpj_to_vendas_2025", {})
    p1_total = sum(sum(v) for v in p1_vendas.values())
    p2_total = sum(sum(v) for v in vendas_por_cnpj.values())

    print(f"  Phase 1 CNPJs: {len(p1_vendas)}")
    print(f"  Phase 2 CNPJs: {len(vendas_por_cnpj)}")
    print(f"  Phase 1 total: R$ {p1_total:,.2f}")
    print(f"  Phase 2 total: R$ {p2_total:,.2f}")

    diff = abs(p2_total - p1_total)
    diff_pct = diff / p1_total * 100 if p1_total > 0 else 0
    print(f"  Diff: R$ {diff:,.2f} ({diff_pct:.4f}%)")

    # Check CNPJ overlap
    p1_set = set(p1_vendas.keys())
    p2_set = set(vendas_por_cnpj.keys())
    only_p1 = p1_set - p2_set
    only_p2 = p2_set - p1_set
    both = p1_set & p2_set

    print(f"\n  CNPJ overlap:")
    print(f"    In both: {len(both)}")
    print(f"    Only Phase 1: {len(only_p1)}")
    print(f"    Only Phase 2: {len(only_p2)}")

    # Per-CNPJ comparison for shared CNPJs
    mismatches = 0
    max_diff_cnpj = None
    max_diff_val = 0
    for cnpj in both:
        p1_vals = p1_vendas[cnpj]
        p2_vals = vendas_por_cnpj[cnpj]
        cnpj_diff = abs(sum(p1_vals) - sum(p2_vals))
        if cnpj_diff > 0.01:
            mismatches += 1
            if cnpj_diff > max_diff_val:
                max_diff_val = cnpj_diff
                max_diff_cnpj = cnpj

    if mismatches == 0:
        print(f"\n  Per-CNPJ check: ALL {len(both)} shared CNPJs match EXACTLY")
        status = "PASS"
    else:
        print(f"\n  Per-CNPJ check: {mismatches} mismatches found")
        if max_diff_cnpj:
            print(f"    Largest diff: CNPJ {max_diff_cnpj}, R$ {max_diff_val:,.2f}")
        status = "WARN" if diff_pct < 0.1 else "FAIL"

    result = {
        "status": status,
        "phase1_cnpjs": len(p1_vendas),
        "phase2_cnpjs": len(vendas_por_cnpj),
        "phase1_total": round(p1_total, 2),
        "phase2_total": round(p2_total, 2),
        "diff": round(diff, 2),
        "diff_pct": round(diff_pct, 4),
        "cnpjs_in_both": len(both),
        "cnpjs_only_phase1": len(only_p1),
        "cnpjs_only_phase2": len(only_p2),
        "per_cnpj_mismatches": mismatches,
    }

    if status == "PASS":
        print(f"\n  Cross-check: PASS (totals identical, all CNPJs match)")
    elif status == "WARN":
        print(f"\n  Cross-check: WARN (small diff, within tolerance)")
    else:
        print(f"\n  Cross-check: FAIL (diff exceeds tolerance)")

    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 60)
    print("SAP Vendas Extractor (Phase 02)")
    print("=" * 60)

    # Validate source file exists
    if not os.path.exists(SAP_CONSOLIDADO):
        print(f"ERROR: Source file not found: {SAP_CONSOLIDADO}")
        sys.exit(1)
    print(f"Source: {SAP_CONSOLIDADO}")

    # Load workbook
    print(f"\nLoading workbook (read_only=True, data_only=True)...")
    wb = openpyxl.load_workbook(SAP_CONSOLIDADO, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # 1. Extract Venda Mes a Mes
    vendas_por_cnpj, vendas_por_mes = extract_sap_vendas(wb)
    wb.close()

    # 2. Cross-check vs Phase 1
    cross_check = cross_check_phase1(vendas_por_cnpj)

    # 3. Calculate stats
    total_vendas_2025 = sum(sum(v) for v in vendas_por_cnpj.values())

    print("\n--- 3. Final Stats ---")
    print(f"  Total clientes: {len(vendas_por_cnpj)}")
    print(f"  Total vendas 2025: R$ {total_vendas_2025:,.2f}")

    # Reference check
    ref_total = 2089824.23
    diff_ref = abs(total_vendas_2025 - ref_total)
    diff_ref_pct = diff_ref / ref_total * 100
    print(f"  Referencia (Phase 1): R$ {ref_total:,.2f}")
    print(f"  Diff vs referencia: R$ {diff_ref:,.2f} ({diff_ref_pct:.4f}%)")

    if diff_ref_pct < 0.1:
        print(f"  STATUS: WITHIN 0.1% tolerance - OK")
    else:
        print(f"  WARNING: Diff exceeds 0.1% tolerance!")

    # 4. Build and save JSON
    output = {
        "cnpj_to_vendas": vendas_por_cnpj,
        "stats": {
            "total_clientes": len(vendas_por_cnpj),
            "total_vendas_2025": round(total_vendas_2025, 2),
            "vendas_por_mes": [round(v, 2) for v in vendas_por_mes],
        },
        "cross_check": cross_check,
    }

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"\nOutput saved: {OUTPUT_FILE}")

    # Final summary
    print("\n" + "=" * 60)
    print("EXTRACTION SUMMARY")
    print("=" * 60)
    print(f"  Clientes:      {len(vendas_por_cnpj)}")
    print(f"  Total 2025:    R$ {total_vendas_2025:,.2f}")
    print(f"  Referencia:    R$ {ref_total:,.2f}")
    print(f"  Diff:          R$ {diff_ref:,.2f} ({diff_ref_pct:.4f}%)")
    print(f"  Cross-check:   {cross_check.get('status', 'N/A')}")
    print(f"  JAN > 0:       {'YES' if vendas_por_mes[0] > 0 else 'NO'}")
    print("=" * 60)

    return 0


if __name__ == "__main__":
    sys.exit(main())

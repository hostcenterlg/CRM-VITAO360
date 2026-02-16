"""
Rebuild DASH tab in CRM INTELIGENTE - VITAO360 V2.xlsx
Spec: 3 blocks with 4 blank rows between each.
All COUNTIFS reference DRAFT 2 columns.
"""
import os
import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from copy import copy

# ── Paths ──
base_dir = os.path.dirname(os.path.abspath(__file__))
input_path = os.path.join(base_dir, "output", "CRM INTELIGENTE - VITAO360 V2.xlsx")
output_path = input_path  # overwrite same file

print(f"Loading {input_path}...")
wb = load_workbook(input_path)

# ── Delete old DASH and create new ──
if "DASH" in wb.sheetnames:
    dash_idx = wb.sheetnames.index("DASH")
    del wb["DASH"]
    ws = wb.create_sheet("DASH", dash_idx)
else:
    ws = wb.create_sheet("DASH")

ws.sheet_properties.tabColor = "DC2626"

# ═══════════════════════════════════════════════════════════════
# STYLES
# ═══════════════════════════════════════════════════════════════
FONT_DATA = Font(name='Calibri', size=11)
FONT_HEADER = Font(name='Calibri', size=10, bold=True)
FONT_HEADER_W = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
FONT_HEADER_W9 = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
FONT_TITLE = Font(name='Calibri', size=11, bold=True)
FONT_LABEL = Font(name='Calibri', size=11)

FILL_LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
FILL_GRAY_D9 = PatternFill('solid', fgColor='D9D9D9')
FILL_DARK = PatternFill('solid', fgColor='404040')
FILL_DARK_59 = PatternFill('solid', fgColor='595959')
FILL_RED_DARK = PatternFill('solid', fgColor='8B0000')
FILL_GREEN_DARK = PatternFill('solid', fgColor='2E7D32')
FILL_SEP = PatternFill('solid', fgColor='F2F2F2')
FILL_NONE = PatternFill(fill_type=None)

THIN_BORDER = Border(
    left=Side(style='thin', color='BBBBBB'),
    right=Side(style='thin', color='BBBBBB'),
    top=Side(style='thin', color='BBBBBB'),
    bottom=Side(style='thin', color='BBBBBB'),
)
NO_BORDER = Border()

ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_cell(cell, font=FONT_DATA, fill=None, border=THIN_BORDER, align=ALIGN_CENTER):
    cell.font = font
    if fill:
        cell.fill = fill
    cell.border = border
    cell.alignment = align


def write_header(ws, row, col, value, fill=FILL_DARK, font=FONT_HEADER_W, align=ALIGN_CENTER):
    cell = ws.cell(row=row, column=col, value=value)
    style_cell(cell, font=font, fill=fill, align=align)
    return cell


def write_data(ws, row, col, value, font=FONT_DATA, fill=None, align=ALIGN_CENTER):
    cell = ws.cell(row=row, column=col, value=value)
    style_cell(cell, font=font, fill=fill, align=align)
    return cell


# ── DRAFT 2 column mapping (headers in row 2, data from row 3) ──
# A=DATA, B=CONSULTOR, Q=TIPO DO CONTATO, R=RESULTADO, S=MOTIVO
# N=WHATSAPP, O=LIGAÇÃO, P=LIGAÇÃO ATENDIDA, I=ESTÁGIO FUNIL, W=MERCOS

# Helper: build COUNTIFS formula with optional consultor + period filter
# D2 cols: A=date, B=consultor, Q=tipo, R=resultado, S=motivo, N=wpp, O=lig, P=lig_atend, I=estagio
def cf(extra_criteria="", use_todos=True):
    """Build base COUNTIFS with period + consultor filter.
    extra_criteria: additional pairs like ',"VENDA / PEDIDO"' etc.
    Returns formula string (without leading =).
    """
    date_filter = "'DRAFT 2'!$A$3:$A$5000,\">=\"&$E$1,'DRAFT 2'!$A$3:$A$5000,\"<=\"&$F$1"
    cons_filter = ",'DRAFT 2'!$B$3:$B$5000,$C$1"

    if use_todos:
        base_all = f"COUNTIFS({date_filter}{extra_criteria})"
        base_cons = f"COUNTIFS({date_filter}{cons_filter}{extra_criteria})"
        return f'IF(OR($C$1="",$C$1="TODOS"),{base_all},{base_cons})'
    else:
        return f"COUNTIFS({date_filter}{extra_criteria})"


# ═══════════════════════════════════════════════════════════════
# COLUMN WIDTHS
# ═══════════════════════════════════════════════════════════════
ws.column_dimensions['A'].width = 25
for c in range(2, 18):  # B to Q
    ws.column_dimensions[get_column_letter(c)].width = 12
# Chart area columns wider
for c in range(15, 20):  # O to S
    ws.column_dimensions[get_column_letter(c)].width = 14

print("Building CABEÇALHO...")

# ═══════════════════════════════════════════════════════════════
# CABEÇALHO (Row 1-2)
# ═══════════════════════════════════════════════════════════════
write_data(ws, 1, 1, "VENDEDOR", font=FONT_TITLE, fill=FILL_LIGHT_GRAY, align=ALIGN_LEFT)
write_data(ws, 1, 2, None)  # spacer
# C1: dropdown consultor
ws.cell(row=1, column=3, value="TODOS").font = FONT_DATA
ws.cell(row=1, column=3).border = THIN_BORDER
dv_cons = DataValidation(type="list",
    formula1='"TODOS,MANU DITZEL,LARISSA PADILHA,JULIO GADRET,DAIANE STAVICKI"',
    allow_blank=True)
ws.add_data_validation(dv_cons)
dv_cons.add("C1")

write_data(ws, 1, 4, "PERÍODO", font=FONT_TITLE, fill=FILL_LIGHT_GRAY, align=ALIGN_LEFT)
# E1: data início
ws.cell(row=1, column=5, value=datetime.date(2026, 2, 1)).font = FONT_DATA
ws.cell(row=1, column=5).number_format = 'DD/MM/YYYY'
ws.cell(row=1, column=5).border = THIN_BORDER
# F1: data fim
ws.cell(row=1, column=6, value=datetime.date(2026, 2, 28)).font = FONT_DATA
ws.cell(row=1, column=6).number_format = 'DD/MM/YYYY'
ws.cell(row=1, column=6).border = THIN_BORDER

# Row 2-3 blank

print("Building BLOCO 1: TIPO DO CONTATO × RESULTADO...")

# ═══════════════════════════════════════════════════════════════
# BLOCO 1 — TIPO DO CONTATO × RESULTADO (starts row 4)
# ═══════════════════════════════════════════════════════════════
B1_START = 4

# Title row
ws.merge_cells(f'A{B1_START}:M{B1_START}')
cell_t1 = ws.cell(row=B1_START, column=1, value="TIPO DO CONTATO × RESULTADO DO CONTATO")
style_cell(cell_t1, font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
for c in range(1, 14):
    ws.cell(row=B1_START, column=c).fill = FILL_GRAY_D9
    ws.cell(row=B1_START, column=c).border = THIN_BORDER

# Header row 5
HDR_ROW = B1_START + 1  # row 5
hdr_b1 = [
    ("TIPO DO CONTATO", FILL_DARK, ALIGN_LEFT),
    ("TOTAL", FILL_DARK, ALIGN_CENTER),
    ("ORÇAM.", FILL_DARK, ALIGN_CENTER),
    ("CADAST.", FILL_DARK, ALIGN_CENTER),
    ("RELAC.", FILL_DARK, ALIGN_CENTER),
    ("EM ATEND.", FILL_DARK, ALIGN_CENTER),
    ("SUPORTE", FILL_DARK, ALIGN_CENTER),
    ("VENDA", FILL_DARK, ALIGN_CENTER),
    ("Ñ ATENDE", FILL_RED_DARK, ALIGN_CENTER),
    ("RECUSOU", FILL_RED_DARK, ALIGN_CENTER),
    ("Ñ RESP.", FILL_RED_DARK, ALIGN_CENTER),
    ("PERDA", FILL_RED_DARK, ALIGN_CENTER),
    ("FOLLOW UP", FILL_GREEN_DARK, ALIGN_CENTER),
]
for i, (label, fill, align) in enumerate(hdr_b1):
    write_header(ws, HDR_ROW, i + 1, label, fill=fill, align=align)

# Resultado mapping for COUNTIFS (column header -> RESULTADO value in DRAFT 2)
resultado_map = {
    3: "ORÇAMENTO",          # C = ORÇAM.
    4: "CADASTRO",           # D = CADAST.
    5: "RELACIONAMENTO",     # E = RELAC.
    6: "EM ATENDIMENTO",     # F = EM ATEND.
    7: "SUPORTE",            # G = SUPORTE
    8: "VENDA / PEDIDO",     # H = VENDA
    9: "NÃO ATENDE",         # I = Ñ ATENDE
    10: "RECUSOU LIGAÇÃO",   # J = RECUSOU
    11: "NÃO RESPONDE",      # K = Ñ RESP.
    12: "PERDA / FECHOU LOJA",  # L = PERDA
}
# M = FOLLOW UP = FOLLOW UP 7 + FOLLOW UP 15

# 7 types of contact
tipos = [
    "PROSPECÇÃO",
    "NEGOCIAÇÃO",
    "FOLLOW UP",
    "ATEND. CLIENTES ATIVOS",
    "ATEND. CLIENTES INATIVOS",
    "PÓS-VENDA / RELACIONAMENTO",
    "PERDA / NUTRIÇÃO",
]

DATA_START = HDR_ROW + 1  # row 6
for t_idx, tipo in enumerate(tipos):
    r = DATA_START + t_idx
    write_data(ws, r, 1, tipo, align=ALIGN_LEFT)

    # B = TOTAL = SUM(C:M)
    ws.cell(row=r, column=2).value = f'=SUM(C{r}:M{r})'
    style_cell(ws.cell(row=r, column=2), font=Font(name='Calibri', size=11, bold=True))

    # C-L: individual resultados
    tipo_filter = f",'DRAFT 2'!$Q$3:$Q$5000,A{r}"
    for col_idx, resultado in resultado_map.items():
        res_filter = f",'DRAFT 2'!$R$3:$R$5000,\"{resultado}\""
        formula = f'={cf(tipo_filter + res_filter)}'
        ws.cell(row=r, column=col_idx).value = formula
        style_cell(ws.cell(row=r, column=col_idx))

    # M = FOLLOW UP 7 + FOLLOW UP 15
    fu7_filter = f",'DRAFT 2'!$R$3:$R$5000,\"FOLLOW UP 7\""
    fu15_filter = f",'DRAFT 2'!$R$3:$R$5000,\"FOLLOW UP 15\""
    formula_m = f'={cf(tipo_filter + fu7_filter)}+{cf(tipo_filter + fu15_filter)}'
    ws.cell(row=r, column=13).value = formula_m
    style_cell(ws.cell(row=r, column=13))

# TOTAL row
TOTAL_ROW = DATA_START + len(tipos)  # row 13
write_header(ws, TOTAL_ROW, 1, "TOTAL", fill=FILL_DARK, align=ALIGN_LEFT)
for c in range(2, 14):
    col_l = get_column_letter(c)
    ws.cell(row=TOTAL_ROW, column=c).value = f'=SUM({col_l}{DATA_START}:{col_l}{TOTAL_ROW - 1})'
    write_header(ws, TOTAL_ROW, c, ws.cell(row=TOTAL_ROW, column=c).value, fill=FILL_DARK)

# ── Bar Chart: RESULTADO DO CONTATO ──
print("  Adding bar chart...")
chart = BarChart()
chart.type = "bar"  # horizontal bars
chart.style = 10
chart.title = "RESULTADO DO CONTATO"
chart.y_axis.title = None
chart.x_axis.title = None
chart.legend = None
chart.width = 18
chart.height = 12

# Data from total row (row 13), columns C to M (3 to 13)
# We need labels and values
# Labels = header names from row 5, cols C-M
# Values = totals from row 13, cols C-M
data_ref = Reference(ws, min_col=3, max_col=13, min_row=TOTAL_ROW, max_row=TOTAL_ROW)
cats_ref = Reference(ws, min_col=3, max_col=13, min_row=HDR_ROW, max_row=HDR_ROW)

chart.add_data(data_ref, from_rows=True, titles_from_data=False)
chart.set_categories(cats_ref)

# Apply colors to data points
chart_colors = {
    0: "2196F3",   # ORÇAM - blue
    1: "2196F3",   # CADAST - blue
    2: "2196F3",   # RELAC - blue
    3: "FF9800",   # EM ATEND - orange
    4: "7B1FA2",   # SUPORTE - purple
    5: "404040",   # VENDA - dark
    6: "C62828",   # Ñ ATENDE - red
    7: "C62828",   # RECUSOU - red
    8: "C62828",   # Ñ RESP - red
    9: "333333",   # PERDA - dark
    10: "4CAF50",  # FOLLOW UP - green
}

series = chart.series[0]
series.graphicalProperties.line.noFill = True
for pt_idx, color in chart_colors.items():
    pt = DataPoint(idx=pt_idx)
    pt.graphicalProperties.solidFill = color
    series.data_points.append(pt)

# Data labels
series.dLbls = DataLabelList()
series.dLbls.showVal = True
series.dLbls.numFmt = '#,##0'

ws.add_chart(chart, "O5")


print("Building BLOCO 2: CONTATOS REALIZADOS + FUNIL...")

# ═══════════════════════════════════════════════════════════════
# 4 BLANK ROWS (14-17)
# ═══════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════
# BLOCO 2 — CONTATOS REALIZADOS + FUNIL DE VENDA (starts row 18)
# ═══════════════════════════════════════════════════════════════
B2_START = 18

# Title
ws.merge_cells(f'A{B2_START}:Q{B2_START}')
cell_t2 = ws.cell(row=B2_START, column=1, value="CONTATOS REALIZADOS + FUNIL DE VENDA")
style_cell(cell_t2, font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
for c in range(1, 18):
    ws.cell(row=B2_START, column=c).fill = FILL_GRAY_D9
    ws.cell(row=B2_START, column=c).border = THIN_BORDER

# Row 19: Group headers (merged)
GRP_ROW = B2_START + 1  # 19
# A19:B19 empty (TIPO/TOTAL)
for c in [1, 2]:
    ws.cell(row=GRP_ROW, column=c).fill = FILL_DARK_59
    ws.cell(row=GRP_ROW, column=c).border = THIN_BORDER

# C19:F19 = CONTATOS REALIZADOS
ws.merge_cells(f'C{GRP_ROW}:F{GRP_ROW}')
write_header(ws, GRP_ROW, 3, "CONTATOS REALIZADOS", fill=FILL_DARK_59)
for c in [4, 5, 6]:
    ws.cell(row=GRP_ROW, column=c).fill = FILL_DARK_59
    ws.cell(row=GRP_ROW, column=c).border = THIN_BORDER

# G = separator
ws.column_dimensions['G'].width = 1.5
ws.cell(row=GRP_ROW, column=7).fill = FILL_SEP
ws.cell(row=GRP_ROW, column=7).border = NO_BORDER

# H19:K19 = FUNIL DE VENDA
ws.merge_cells(f'H{GRP_ROW}:K{GRP_ROW}')
write_header(ws, GRP_ROW, 8, "FUNIL DE VENDA", fill=FILL_DARK_59)
for c in [9, 10, 11]:
    ws.cell(row=GRP_ROW, column=c).fill = FILL_DARK_59
    ws.cell(row=GRP_ROW, column=c).border = THIN_BORDER

# L = separator
ws.column_dimensions['L'].width = 1.5
ws.cell(row=GRP_ROW, column=12).fill = FILL_SEP
ws.cell(row=GRP_ROW, column=12).border = NO_BORDER

# M19:N19 = RELACIONAMENTO
ws.merge_cells(f'M{GRP_ROW}:N{GRP_ROW}')
write_header(ws, GRP_ROW, 13, "RELACIONAMENTO", fill=FILL_DARK_59)
ws.cell(row=GRP_ROW, column=14).fill = FILL_DARK_59
ws.cell(row=GRP_ROW, column=14).border = THIN_BORDER

# O = separator
ws.column_dimensions['O'].width = 1.5
ws.cell(row=GRP_ROW, column=15).fill = FILL_SEP
ws.cell(row=GRP_ROW, column=15).border = NO_BORDER

# P19:Q19 = NÃO VENDA
ws.merge_cells(f'P{GRP_ROW}:Q{GRP_ROW}')
write_header(ws, GRP_ROW, 16, "NÃO VENDA", fill=FILL_DARK_59)
ws.cell(row=GRP_ROW, column=17).fill = FILL_DARK_59
ws.cell(row=GRP_ROW, column=17).border = THIN_BORDER

# Row 20: Sub-headers
SUB_ROW = GRP_ROW + 1  # 20
sub_headers = [
    (1, "TIPO DO CONTATO", ALIGN_LEFT),
    (2, "TOTAL", ALIGN_CENTER),
    (3, "WHATSAPP", ALIGN_CENTER),
    (4, "LIGAÇÕES", ALIGN_CENTER),
    (5, "LIG. ATENDIDAS", ALIGN_CENTER),
    (6, "LIG. Ñ ATEND.", ALIGN_CENTER),
    # 7 = sep
    (8, "EM ATEND.", ALIGN_CENTER),
    (9, "ORÇAMENTO", ALIGN_CENTER),
    (10, "CADASTRO", ALIGN_CENTER),
    (11, "VENDA", ALIGN_CENTER),
    # 12 = sep
    (13, "FOLLOW UP", ALIGN_CENTER),
    (14, "SUPORTE", ALIGN_CENTER),
    # 15 = sep
    (16, "INATIVO", ALIGN_CENTER),
    (17, "PERDA", ALIGN_CENTER),
]
for col, label, align in sub_headers:
    write_header(ws, SUB_ROW, col, label, fill=FILL_DARK, font=FONT_HEADER_W9, align=align)

# Separators in sub row
for sep_col in [7, 12, 15]:
    ws.cell(row=SUB_ROW, column=sep_col).fill = FILL_SEP
    ws.cell(row=SUB_ROW, column=sep_col).border = NO_BORDER

# Data rows 21-27
B2_DATA_START = SUB_ROW + 1  # 21
for t_idx, tipo in enumerate(tipos):
    r = B2_DATA_START + t_idx
    write_data(ws, r, 1, tipo, align=ALIGN_LEFT)

    tipo_filter = f",'DRAFT 2'!$Q$3:$Q$5000,A{r}"

    # B = TOTAL contacts for this tipo
    formula_total = f'={cf(tipo_filter)}'
    write_data(ws, r, 2, formula_total, font=Font(name='Calibri', size=11, bold=True))

    # C = WHATSAPP (col N in DRAFT 2)
    wpp_filter = f",'DRAFT 2'!$N$3:$N$5000,\"SIM\""
    ws.cell(row=r, column=3).value = f'={cf(tipo_filter + wpp_filter)}'
    style_cell(ws.cell(row=r, column=3))

    # D = LIGAÇÕES (col O in DRAFT 2)
    lig_filter = f",'DRAFT 2'!$O$3:$O$5000,\"SIM\""
    ws.cell(row=r, column=4).value = f'={cf(tipo_filter + lig_filter)}'
    style_cell(ws.cell(row=r, column=4))

    # E = LIG ATENDIDAS (col P in DRAFT 2)
    lig_at_filter = f",'DRAFT 2'!$P$3:$P$5000,\"SIM\""
    ws.cell(row=r, column=5).value = f'={cf(tipo_filter + lig_at_filter)}'
    style_cell(ws.cell(row=r, column=5))

    # F = LIG NÃO ATENDIDAS = D - E
    ws.cell(row=r, column=6).value = f'=D{r}-E{r}'
    style_cell(ws.cell(row=r, column=6))

    # G = separator
    ws.cell(row=r, column=7).fill = FILL_SEP
    ws.cell(row=r, column=7).border = NO_BORDER

    # H = EM ATENDIMENTO
    ws.cell(row=r, column=8).value = f'={cf(tipo_filter + ",""DRAFT 2""!$R$3:$R$5000,""EM ATENDIMENTO""")}'.replace('""DRAFT 2""', "'DRAFT 2'").replace('""EM ATENDIMENTO""', '"EM ATENDIMENTO"')
    # Simpler approach:
    res_em = f",'DRAFT 2'!$R$3:$R$5000,\"EM ATENDIMENTO\""
    ws.cell(row=r, column=8).value = f'={cf(tipo_filter + res_em)}'
    style_cell(ws.cell(row=r, column=8))

    # I = ORÇAMENTO
    res_orc = f",'DRAFT 2'!$R$3:$R$5000,\"ORÇAMENTO\""
    ws.cell(row=r, column=9).value = f'={cf(tipo_filter + res_orc)}'
    style_cell(ws.cell(row=r, column=9))

    # J = CADASTRO
    res_cad = f",'DRAFT 2'!$R$3:$R$5000,\"CADASTRO\""
    ws.cell(row=r, column=10).value = f'={cf(tipo_filter + res_cad)}'
    style_cell(ws.cell(row=r, column=10))

    # K = VENDA
    res_ven = f",'DRAFT 2'!$R$3:$R$5000,\"VENDA / PEDIDO\""
    ws.cell(row=r, column=11).value = f'={cf(tipo_filter + res_ven)}'
    style_cell(ws.cell(row=r, column=11))

    # L = separator
    ws.cell(row=r, column=12).fill = FILL_SEP
    ws.cell(row=r, column=12).border = NO_BORDER

    # M = FOLLOW UP (7 + 15 + RELACIONAMENTO)
    res_fu7 = f",'DRAFT 2'!$R$3:$R$5000,\"FOLLOW UP 7\""
    res_fu15 = f",'DRAFT 2'!$R$3:$R$5000,\"FOLLOW UP 15\""
    res_rel = f",'DRAFT 2'!$R$3:$R$5000,\"RELACIONAMENTO\""
    ws.cell(row=r, column=13).value = f'={cf(tipo_filter + res_fu7)}+{cf(tipo_filter + res_fu15)}+{cf(tipo_filter + res_rel)}'
    style_cell(ws.cell(row=r, column=13))

    # N = SUPORTE
    res_sup = f",'DRAFT 2'!$R$3:$R$5000,\"SUPORTE\""
    ws.cell(row=r, column=14).value = f'={cf(tipo_filter + res_sup)}'
    style_cell(ws.cell(row=r, column=14))

    # O = separator
    ws.cell(row=r, column=15).fill = FILL_SEP
    ws.cell(row=r, column=15).border = NO_BORDER

    # P = INATIVO (NÃO ATENDE + NÃO RESPONDE + RECUSOU LIGAÇÃO)
    res_na = f",'DRAFT 2'!$R$3:$R$5000,\"NÃO ATENDE\""
    res_nr = f",'DRAFT 2'!$R$3:$R$5000,\"NÃO RESPONDE\""
    res_rec = f",'DRAFT 2'!$R$3:$R$5000,\"RECUSOU LIGAÇÃO\""
    ws.cell(row=r, column=16).value = f'={cf(tipo_filter + res_na)}+{cf(tipo_filter + res_nr)}+{cf(tipo_filter + res_rec)}'
    style_cell(ws.cell(row=r, column=16))

    # Q = PERDA
    res_perda = f",'DRAFT 2'!$R$3:$R$5000,\"PERDA / FECHOU LOJA\""
    ws.cell(row=r, column=17).value = f'={cf(tipo_filter + res_perda)}'
    style_cell(ws.cell(row=r, column=17))

# TOTAL row
B2_TOTAL = B2_DATA_START + len(tipos)  # 28
write_header(ws, B2_TOTAL, 1, "TOTAL", fill=FILL_DARK, align=ALIGN_LEFT)
for c in [2, 3, 4, 5, 6, 8, 9, 10, 11, 13, 14, 16, 17]:
    col_l = get_column_letter(c)
    ws.cell(row=B2_TOTAL, column=c).value = f'=SUM({col_l}{B2_DATA_START}:{col_l}{B2_TOTAL - 1})'
    write_header(ws, B2_TOTAL, c, ws.cell(row=B2_TOTAL, column=c).value, fill=FILL_DARK)

# Separators in total row
for sep_col in [7, 12, 15]:
    ws.cell(row=B2_TOTAL, column=sep_col).fill = FILL_SEP
    ws.cell(row=B2_TOTAL, column=sep_col).border = NO_BORDER


print("Building BLOCO 3: MOTIVOS + INDICADORES...")

# ═══════════════════════════════════════════════════════════════
# 4 BLANK ROWS (29-32)
# ═══════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════
# BLOCO 3 — MOTIVOS + INDICADORES (starts row 33, side by side)
# ═══════════════════════════════════════════════════════════════
B3_START = 33

# Left title: MOTIVOS
ws.merge_cells(f'A{B3_START}:G{B3_START}')
cell_t3l = ws.cell(row=B3_START, column=1, value="MOTIVOS DE NÃO COMPRA")
style_cell(cell_t3l, font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
for c in range(1, 8):
    ws.cell(row=B3_START, column=c).fill = FILL_GRAY_D9
    ws.cell(row=B3_START, column=c).border = THIN_BORDER

# Right title: INDICADORES
ws.merge_cells(f'I{B3_START}:Q{B3_START}')
cell_t3r = ws.cell(row=B3_START, column=9, value="INDICADORES DE TAREFAS POR CONSULTOR")
style_cell(cell_t3r, font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
for c in range(9, 18):
    ws.cell(row=B3_START, column=c).fill = FILL_GRAY_D9
    ws.cell(row=B3_START, column=c).border = THIN_BORDER

# H column = separator between the two tables
ws.column_dimensions['H'].width = 2

# ── MOTIVOS table (A34:G45) ──
MOT_HDR = B3_START + 1  # 34
mot_headers = [
    (1, "MOTIVO", ALIGN_LEFT),
    (2, "QTD", ALIGN_CENTER),
    (3, "%", ALIGN_CENTER),
    (4, "PROSP", ALIGN_CENTER),
    (5, "ATIVOS", ALIGN_CENTER),
    (6, "INAT", ALIGN_CENTER),
    (7, "PÓS-V", ALIGN_CENTER),
]
for col, label, align in mot_headers:
    write_header(ws, MOT_HDR, col, label, fill=FILL_DARK, align=align)

motivos = [
    "AINDA TEM ESTOQUE",
    "PRODUTO NÃO VENDEU / SEM GIRO",
    "LOJA ANEXO/PROXIMO - SM",
    "SÓ QUER COMPRAR GRANEL",
    "PROBLEMA LOGÍSTICO / ENTREGA",
    "PROBLEMA FINANCEIRO / CRÉDITO",
    "PROPRIETARIO / INDISPONÍVEL",
    "FECHANDO / FECHOU LOJA",
    "SEM INTERESSE NO MOMENTO",
    "PRIMEIRO CONTATO / SEM RESPOSTA",
]

MOT_DATA_START = MOT_HDR + 1  # 35
for m_idx, motivo in enumerate(motivos):
    r = MOT_DATA_START + m_idx
    write_data(ws, r, 1, motivo, align=ALIGN_LEFT)

    mot_filter = f",'DRAFT 2'!$S$3:$S$5000,A{r}"

    # B = QTD
    ws.cell(row=r, column=2).value = f'={cf(mot_filter)}'
    style_cell(ws.cell(row=r, column=2))

    # C = % (QTD/TOTAL)
    # Will reference total row

    # D = PROSP (TIPO = PROSPECÇÃO)
    prosp_filter = f",'DRAFT 2'!$Q$3:$Q$5000,\"PROSPECÇÃO\""
    ws.cell(row=r, column=4).value = f'={cf(mot_filter + prosp_filter)}'
    style_cell(ws.cell(row=r, column=4))

    # E = ATIVOS (TIPO = ATEND. CLIENTES ATIVOS)
    ativos_filter = f",'DRAFT 2'!$Q$3:$Q$5000,\"ATEND. CLIENTES ATIVOS\""
    ws.cell(row=r, column=5).value = f'={cf(mot_filter + ativos_filter)}'
    style_cell(ws.cell(row=r, column=5))

    # F = INAT (TIPO = ATEND. CLIENTES INATIVOS)
    inat_filter = f",'DRAFT 2'!$Q$3:$Q$5000,\"ATEND. CLIENTES INATIVOS\""
    ws.cell(row=r, column=6).value = f'={cf(mot_filter + inat_filter)}'
    style_cell(ws.cell(row=r, column=6))

    # G = PÓS-V (TIPO = PÓS-VENDA / RELACIONAMENTO)
    posv_filter = f",'DRAFT 2'!$Q$3:$Q$5000,\"PÓS-VENDA / RELACIONAMENTO\""
    ws.cell(row=r, column=7).value = f'={cf(mot_filter + posv_filter)}'
    style_cell(ws.cell(row=r, column=7))

MOT_TOTAL_ROW = MOT_DATA_START + len(motivos)  # 45
write_header(ws, MOT_TOTAL_ROW, 1, "TOTAL", fill=FILL_DARK, align=ALIGN_LEFT)
for c in [2, 4, 5, 6, 7]:
    col_l = get_column_letter(c)
    ws.cell(row=MOT_TOTAL_ROW, column=c).value = f'=SUM({col_l}{MOT_DATA_START}:{col_l}{MOT_TOTAL_ROW - 1})'
    write_header(ws, MOT_TOTAL_ROW, c, ws.cell(row=MOT_TOTAL_ROW, column=c).value, fill=FILL_DARK)

# Now fill % column (C) - references total in B45
for m_idx in range(len(motivos)):
    r = MOT_DATA_START + m_idx
    ws.cell(row=r, column=3).value = f'=IFERROR(B{r}/B{MOT_TOTAL_ROW},0)'
    ws.cell(row=r, column=3).number_format = '0%'
    style_cell(ws.cell(row=r, column=3))

ws.cell(row=MOT_TOTAL_ROW, column=3).value = f'=IFERROR(B{MOT_TOTAL_ROW}/B{MOT_TOTAL_ROW},0)'
ws.cell(row=MOT_TOTAL_ROW, column=3).number_format = '0%'
write_header(ws, MOT_TOTAL_ROW, 3, ws.cell(row=MOT_TOTAL_ROW, column=3).value, fill=FILL_DARK)


# ── INDICADORES table (I34:Q39) ──
IND_HDR = B3_START + 1  # 34
ind_headers = [
    (9, "CONSULTOR", ALIGN_LEFT),
    (10, "CONTATOS", ALIGN_CENTER),
    (11, "VENDAS", ALIGN_CENTER),
    (12, "ORÇAM.", ALIGN_CENTER),
    (13, "CADAST.", ALIGN_CENTER),
    (14, "% CONV", ALIGN_CENTER),
    (15, "Ñ ATENDE", ALIGN_CENTER),
    (16, "PERDA", ALIGN_CENTER),
    (17, "% MERCOS", ALIGN_CENTER),
]
for col, label, align in ind_headers:
    write_header(ws, IND_HDR, col, label, fill=FILL_DARK, align=align)

consultores = ["MANU DITZEL", "LARISSA PADILHA", "JULIO GADRET", "DAIANE STAVICKI"]
IND_DATA_START = IND_HDR + 1  # 35
for c_idx, cons in enumerate(consultores):
    r = IND_DATA_START + c_idx
    write_data(ws, r, 9, cons, align=ALIGN_LEFT)

    cons_filter = f",'DRAFT 2'!$B$3:$B$5000,I{r}"
    date_filter_base = "'DRAFT 2'!$A$3:$A$5000,\">=\"&$E$1,'DRAFT 2'!$A$3:$A$5000,\"<=\"&$F$1"

    # J = CONTATOS
    ws.cell(row=r, column=10).value = f'=COUNTIFS({date_filter_base}{cons_filter})'
    style_cell(ws.cell(row=r, column=10))

    # K = VENDAS
    res_ven = f",'DRAFT 2'!$R$3:$R$5000,\"VENDA / PEDIDO\""
    ws.cell(row=r, column=11).value = f'=COUNTIFS({date_filter_base}{cons_filter}{res_ven})'
    style_cell(ws.cell(row=r, column=11))

    # L = ORÇAM.
    res_orc = f",'DRAFT 2'!$R$3:$R$5000,\"ORÇAMENTO\""
    ws.cell(row=r, column=12).value = f'=COUNTIFS({date_filter_base}{cons_filter}{res_orc})'
    style_cell(ws.cell(row=r, column=12))

    # M = CADAST.
    res_cad = f",'DRAFT 2'!$R$3:$R$5000,\"CADASTRO\""
    ws.cell(row=r, column=13).value = f'=COUNTIFS({date_filter_base}{cons_filter}{res_cad})'
    style_cell(ws.cell(row=r, column=13))

    # N = % CONV = VENDAS/CONTATOS
    ws.cell(row=r, column=14).value = f'=IFERROR(K{r}/J{r},0)'
    ws.cell(row=r, column=14).number_format = '0%'
    style_cell(ws.cell(row=r, column=14))

    # O = Ñ ATENDE
    res_na = f",'DRAFT 2'!$R$3:$R$5000,\"NÃO ATENDE\""
    res_nr = f",'DRAFT 2'!$R$3:$R$5000,\"NÃO RESPONDE\""
    res_rec = f",'DRAFT 2'!$R$3:$R$5000,\"RECUSOU LIGAÇÃO\""
    ws.cell(row=r, column=15).value = f'=COUNTIFS({date_filter_base}{cons_filter}{res_na})+COUNTIFS({date_filter_base}{cons_filter}{res_nr})+COUNTIFS({date_filter_base}{cons_filter}{res_rec})'
    style_cell(ws.cell(row=r, column=15))

    # P = PERDA
    res_perda = f",'DRAFT 2'!$R$3:$R$5000,\"PERDA / FECHOU LOJA\""
    ws.cell(row=r, column=16).value = f'=COUNTIFS({date_filter_base}{cons_filter}{res_perda})'
    style_cell(ws.cell(row=r, column=16))

    # Q = % MERCOS (col W in DRAFT 2)
    mercos_filter = f",'DRAFT 2'!$W$3:$W$5000,\"SIM\""
    ws.cell(row=r, column=17).value = f'=IFERROR(COUNTIFS({date_filter_base}{cons_filter}{mercos_filter})/COUNTIFS({date_filter_base}{cons_filter}),0)'
    ws.cell(row=r, column=17).number_format = '0%'
    style_cell(ws.cell(row=r, column=17))

# TOTAL EQUIPE row
IND_TOTAL = IND_DATA_START + len(consultores)  # 39
write_header(ws, IND_TOTAL, 9, "TOTAL EQUIPE", fill=FILL_DARK, align=ALIGN_LEFT)
for c in [10, 11, 12, 13, 15, 16]:
    col_l = get_column_letter(c)
    ws.cell(row=IND_TOTAL, column=c).value = f'=SUM({col_l}{IND_DATA_START}:{col_l}{IND_TOTAL - 1})'
    write_header(ws, IND_TOTAL, c, ws.cell(row=IND_TOTAL, column=c).value, fill=FILL_DARK)

# N39 = % CONV total
ws.cell(row=IND_TOTAL, column=14).value = f'=IFERROR(K{IND_TOTAL}/J{IND_TOTAL},0)'
ws.cell(row=IND_TOTAL, column=14).number_format = '0%'
write_header(ws, IND_TOTAL, 14, ws.cell(row=IND_TOTAL, column=14).value, fill=FILL_DARK)

# Q39 = % MERCOS total
date_filter_base = "'DRAFT 2'!$A$3:$A$5000,\">=\"&$E$1,'DRAFT 2'!$A$3:$A$5000,\"<=\"&$F$1"
mercos_filter = f",'DRAFT 2'!$W$3:$W$5000,\"SIM\""
ws.cell(row=IND_TOTAL, column=17).value = f'=IFERROR(COUNTIFS({date_filter_base}{mercos_filter})/COUNTIFS({date_filter_base}),0)'
ws.cell(row=IND_TOTAL, column=17).number_format = '0%'
write_header(ws, IND_TOTAL, 17, ws.cell(row=IND_TOTAL, column=17).value, fill=FILL_DARK)


# ── Distribuição por grupo (I41:L46) ──
DIST_HDR = IND_TOTAL + 2  # 41
dist_headers = [
    (9, "CONSULTOR", ALIGN_LEFT),
    (10, "FUNIL", ALIGN_CENTER),
    (11, "RELAC.", ALIGN_CENTER),
    (12, "NÃO VENDA", ALIGN_CENTER),
]
for col, label, align in dist_headers:
    write_header(ws, DIST_HDR, col, label, fill=FILL_DARK, align=align)

DIST_DATA_START = DIST_HDR + 1  # 42
for c_idx, cons in enumerate(consultores):
    r = DIST_DATA_START + c_idx
    write_data(ws, r, 9, cons, align=ALIGN_LEFT)

    cons_filter = f",'DRAFT 2'!$B$3:$B$5000,I{r}"

    # J = FUNIL (EM ATENDIMENTO + ORÇAMENTO + CADASTRO + VENDA)
    funil_results = ["EM ATENDIMENTO", "ORÇAMENTO", "CADASTRO", "VENDA / PEDIDO"]
    funil_parts = []
    for res in funil_results:
        res_f = f",'DRAFT 2'!$R$3:$R$5000,\"{res}\""
        funil_parts.append(f'COUNTIFS({date_filter_base}{cons_filter}{res_f})')
    ws.cell(row=r, column=10).value = f'={"+".join(funil_parts)}'
    style_cell(ws.cell(row=r, column=10))

    # K = RELAC. (RELACIONAMENTO + FOLLOW UP 7 + FOLLOW UP 15 + SUPORTE)
    relac_results = ["RELACIONAMENTO", "FOLLOW UP 7", "FOLLOW UP 15", "SUPORTE"]
    relac_parts = []
    for res in relac_results:
        res_f = f",'DRAFT 2'!$R$3:$R$5000,\"{res}\""
        relac_parts.append(f'COUNTIFS({date_filter_base}{cons_filter}{res_f})')
    ws.cell(row=r, column=11).value = f'={"+".join(relac_parts)}'
    style_cell(ws.cell(row=r, column=11))

    # L = NÃO VENDA (NÃO ATENDE + NÃO RESPONDE + RECUSOU LIGAÇÃO + PERDA)
    nv_results = ["NÃO ATENDE", "NÃO RESPONDE", "RECUSOU LIGAÇÃO", "PERDA / FECHOU LOJA"]
    nv_parts = []
    for res in nv_results:
        res_f = f",'DRAFT 2'!$R$3:$R$5000,\"{res}\""
        nv_parts.append(f'COUNTIFS({date_filter_base}{cons_filter}{res_f})')
    ws.cell(row=r, column=12).value = f'={"+".join(nv_parts)}'
    style_cell(ws.cell(row=r, column=12))

DIST_TOTAL = DIST_DATA_START + len(consultores)  # 46
write_header(ws, DIST_TOTAL, 9, "TOTAL", fill=FILL_DARK, align=ALIGN_LEFT)
for c in [10, 11, 12]:
    col_l = get_column_letter(c)
    ws.cell(row=DIST_TOTAL, column=c).value = f'=SUM({col_l}{DIST_DATA_START}:{col_l}{DIST_TOTAL - 1})'
    write_header(ws, DIST_TOTAL, c, ws.cell(row=DIST_TOTAL, column=c).value, fill=FILL_DARK)


# ═══════════════════════════════════════════════════════════════
# FINAL: Row heights, freeze panes, zoom
# ═══════════════════════════════════════════════════════════════
print("Applying final formatting...")

# Row heights
for r in range(1, DIST_TOTAL + 5):
    ws.row_dimensions[r].height = 18

# Header rows slightly smaller
for r in [HDR_ROW, SUB_ROW, MOT_HDR, IND_HDR, DIST_HDR, TOTAL_ROW, B2_TOTAL, MOT_TOTAL_ROW, IND_TOTAL, DIST_TOTAL]:
    ws.row_dimensions[r].height = 15

# Freeze panes below header
ws.freeze_panes = "A4"

# Zoom
ws.sheet_view.zoomScale = 90

# ═══════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════
print(f"Saving to {output_path}...")
wb.save(output_path)
print("DONE!")

print(f"\nSummary:")
print(f"  Bloco 1: TIPO × RESULTADO (rows {B1_START}-{TOTAL_ROW}) + bar chart")
print(f"  Bloco 2: CONTATOS + FUNIL (rows {B2_START}-{B2_TOTAL})")
print(f"  Bloco 3: MOTIVOS (rows {B3_START}-{MOT_TOTAL_ROW}) + INDICADORES (rows {B3_START}-{DIST_TOTAL})")
print(f"  Named Ranges: preserved from original")
print(f"  Other sheets: untouched")

"""
populate_log_redes.py
Popula LOG_INTERACOES (4 abas consultor) e REDES (aba REDES v2) no SQLite.

Two-Base Architecture (R4):
  - LOG = NUNCA tem valor R$. Nenhum campo monetario e inserido.
  - VENDA = tem valor R$. Tabela vendas. Nao tocada aqui.

CNPJ (R5): sempre string 14 digitos, zero-padded.
Dados: REAL (rastreavel ao Excel INTELIGENTE FINAL OK).
"""

import re
import sqlite3
import logging
from datetime import datetime, date
from pathlib import Path

import openpyxl

# ── Configuracao ─────────────────────────────────────────────────────────────

XLSX_PATH = r"C:\Users\User\OneDrive\Área de Trabalho\CRM_VITAO360  INTELIGENTE   FINAL OK .xlsx"
DB_PATH   = r"C:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\crm_vitao360.db"

BATCH_SIZE = 100

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger(__name__)

# ── Helpers ───────────────────────────────────────────────────────────────────

def normalizar_cnpj(val) -> str | None:
    """Retorna CNPJ como string 14 digitos ou None se invalido."""
    if val is None:
        return None
    # float vindo do Excel (ex: 3.4060666e+13)
    if isinstance(val, float):
        val = str(int(val))
    else:
        val = str(val)
    digits = re.sub(r"\D", "", val).zfill(14)
    # se sobrar mais de 14, pega os ultimos 14 (raro mas defensivo)
    digits = digits[-14:]
    if len(digits) != 14 or digits == "00000000000000":
        return None
    return digits


def coerce_str(val) -> str | None:
    """Converte valor para string limpa ou None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def coerce_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def coerce_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def coerce_date(val) -> datetime | None:
    """Aceita datetime, date, ou string ISO."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    return None


def calc_followup_dias(followup_val, data_interacao: datetime | None) -> int | None:
    """
    Se followup_val for uma data, calcula (followup - data_interacao).days.
    Se for int/float, retorna como int.
    """
    if followup_val is None:
        return None

    # Caso seja datetime / date
    if isinstance(followup_val, (datetime, date)):
        if data_interacao is None:
            return None
        fu = followup_val if isinstance(followup_val, datetime) else datetime(followup_val.year, followup_val.month, followup_val.day)
        return (fu - data_interacao).days

    # Caso seja string que parece data
    if isinstance(followup_val, str):
        dt = coerce_date(followup_val)
        if dt and data_interacao:
            return (dt - data_interacao).days
        # pode ser numero como string
        try:
            return int(followup_val)
        except ValueError:
            return None

    # Caso seja numerico
    try:
        return int(followup_val)
    except (ValueError, TypeError):
        return None


# ── Mapeamento de colunas LOG (1-indexed) ────────────────────────────────────
#
# Col  1 (A)  = DATA
# Col  2 (B)  = CONSULTOR
# Col  4 (D)  = CNPJ
# Col  9 (I)  = ESTAGIO FUNIL
# Col 11 (K)  = FASE
# Col 12 (L)  = TEMPERATURA
# Col 14 (N)  = PROX. ACAO  -> acao_futura
# Col 15 (O)  = TENTATIVA
# Col 20 (T)  = TIPO DO CONTATO
# Col 21 (U)  = RESULTADO
# Col 23 (W)  = FOLLOW-UP   -> seguir calc_followup_dias
# Col 25 (Y)  = ACAO FUTURA -> descricao principal
# Col 28      = NOTA DO DIA -> descricao fallback

IDX_DATA         = 0   # col A (0-based)
IDX_CONSULTOR    = 1   # col B
IDX_CNPJ         = 3   # col D
IDX_ESTAGIO      = 8   # col I
IDX_FASE         = 10  # col K
IDX_TEMPERATURA  = 11  # col L
IDX_ACAO_FUTURA  = 13  # col N  -> acao_futura no DB
IDX_TENTATIVA    = 14  # col O
IDX_TIPO_CONTATO = 19  # col T
IDX_RESULTADO    = 20  # col U
IDX_FOLLOWUP     = 22  # col W
IDX_DESCRICAO1   = 24  # col Y  (ACAO FUTURA como descricao)
IDX_DESCRICAO2   = 27  # col 28 (NOTA DO DIA — fallback)


def extrair_log_row(row: tuple) -> dict | None:
    """
    Converte uma tupla de 40 valores em dict para inserir em log_interacoes.
    Retorna None se a row deve ser pulada.
    Campos NOT NULL obrigatorios: cnpj, data_interacao, consultor, resultado.

    Two-Base: NENHUM campo de valor R$ e retornado.
    """
    # Pular rows completamente vazias
    if all(v is None for v in row):
        return None

    # DATA obrigatoria
    data_interacao = coerce_date(row[IDX_DATA])
    if data_interacao is None:
        return None

    # CNPJ obrigatorio
    cnpj = normalizar_cnpj(row[IDX_CNPJ]) if len(row) > IDX_CNPJ else None
    if cnpj is None:
        return None

    # CONSULTOR obrigatorio (NOT NULL no schema)
    consultor = coerce_str(row[IDX_CONSULTOR]) if len(row) > IDX_CONSULTOR else None
    if consultor is None:
        return None

    # RESULTADO obrigatorio (NOT NULL no schema).
    # Col U (idx 20) e o campo oficial, mas nestas abas de planejamento nunca
    # foi preenchido (sao templates de atendimento futuro, nao log de realizados).
    # Fallback: usa FASE (col K, idx 10) que sempre tem dado nestas abas.
    resultado = coerce_str(row[IDX_RESULTADO]) if len(row) > IDX_RESULTADO else None
    if resultado is None:
        # fallback para FASE (idx 10 = col K)
        resultado = coerce_str(row[IDX_FASE]) if len(row) > IDX_FASE else None
    if resultado is None:
        # segundo fallback: ESTAGIO FUNIL (idx 8 = col I)
        resultado = coerce_str(row[IDX_ESTAGIO]) if len(row) > IDX_ESTAGIO else None
    if resultado is None:
        return None

    # Descricao: prefere col Y (ACAO FUTURA detalhada), fallback col 28 (NOTA DO DIA)
    descricao = coerce_str(row[IDX_DESCRICAO1]) if len(row) > IDX_DESCRICAO1 else None
    if not descricao:
        descricao = coerce_str(row[IDX_DESCRICAO2]) if len(row) > IDX_DESCRICAO2 else None

    followup_raw = row[IDX_FOLLOWUP] if len(row) > IDX_FOLLOWUP else None
    follow_up_dias = calc_followup_dias(followup_raw, data_interacao)

    return {
        "cnpj":           cnpj,
        "data_interacao": data_interacao.isoformat(),
        "consultor":      consultor,
        "resultado":      resultado,
        "descricao":      descricao,
        "estagio_funil":  coerce_str(row[IDX_ESTAGIO]),
        "fase":           coerce_str(row[IDX_FASE]),
        "tipo_contato":   coerce_str(row[IDX_TIPO_CONTATO]),
        "acao_futura":    coerce_str(row[IDX_ACAO_FUTURA]),
        "temperatura":    coerce_str(row[IDX_TEMPERATURA]),
        "follow_up_dias": follow_up_dias,
        "grupo_dash":     None,   # nao disponivel no LOG
        "tentativa":      coerce_str(row[IDX_TENTATIVA]),
        "created_at":     datetime.now().isoformat(),
        "created_by":     None,
    }


# ── Insercao em batch ─────────────────────────────────────────────────────────

SQL_INSERT_LOG = """
INSERT INTO log_interacoes
  (cnpj, data_interacao, consultor, resultado, descricao,
   estagio_funil, fase, tipo_contato, acao_futura,
   temperatura, follow_up_dias, grupo_dash, tentativa,
   created_at, created_by)
VALUES
  (:cnpj, :data_interacao, :consultor, :resultado, :descricao,
   :estagio_funil, :fase, :tipo_contato, :acao_futura,
   :temperatura, :follow_up_dias, :grupo_dash, :tentativa,
   :created_at, :created_by)
"""

SQL_INSERT_REDE = """
INSERT OR REPLACE INTO redes
  (nome, consultor_responsavel, total_lojas, lojas_ativas,
   faturamento_real, potencial_maximo, pct_penetracao,
   updated_at)
VALUES
  (:nome, :consultor_responsavel, :total_lojas, :lojas_ativas,
   :faturamento_real, :potencial_maximo, :pct_penetracao,
   :updated_at)
"""


def batch_insert(cur, sql: str, records: list[dict]) -> int:
    """Insere em lotes de BATCH_SIZE. Retorna total inserido."""
    inserted = 0
    for i in range(0, len(records), BATCH_SIZE):
        chunk = records[i : i + BATCH_SIZE]
        cur.executemany(sql, chunk)
        inserted += len(chunk)
    return inserted


# ── Carga principal ───────────────────────────────────────────────────────────

def popular_logs(ws_map: dict, conn: sqlite3.Connection) -> dict:
    """
    ws_map = {'LARISSA': ws_larissa, ...}
    Retorna dict com contagens por consultor.
    """
    cur = conn.cursor()
    contagens = {}

    for aba_nome, ws in ws_map.items():
        log.info("Processando aba LOG: %s", aba_nome)
        records = []
        puladas = 0
        lidas   = 0

        # Row 1 = header, dados a partir de row 2
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            lidas += 1
            rec = extrair_log_row(row_vals)
            if rec is None:
                puladas += 1
                continue
            records.append(rec)

        inserted = batch_insert(cur, SQL_INSERT_LOG, records)
        conn.commit()
        contagens[aba_nome] = inserted
        log.info("  %s: lidas=%d, puladas=%d, inseridas=%d", aba_nome, lidas, puladas, inserted)

    return contagens


def popular_redes(ws, conn: sqlite3.Connection) -> int:
    """
    Aba REDES v2: header na row 5, dados a partir de row 6.
    Para quando encontrar row sem nome de rede.
    """
    cur = conn.cursor()
    records = []
    now_iso = datetime.now().isoformat()

    # Mapeamento (0-based dos valores da row):
    # 0=nome, 1=consultor, 2=lojas, 3=fat_real, 4=meta, 5=pct_penetracao

    # Nomes de linhas de totalizacao a ignorar
    NOMES_IGNORAR = {"TOTAL", "SUBTOTAL", "GRAND TOTAL"}

    for row_vals in ws.iter_rows(min_row=6, values_only=True):
        nome = coerce_str(row_vals[0]) if len(row_vals) > 0 else None
        if not nome:
            break  # fim das redes
        # Pular linhas de totalizacao
        if nome.upper().strip() in NOMES_IGNORAR:
            log.info("  Pulando linha totalizadora: %s", nome)
            continue

        total_lojas = coerce_int(row_vals[2]) if len(row_vals) > 2 else None
        # total_lojas NOT NULL no schema: usa 0 como fallback se ausente
        if total_lojas is None:
            total_lojas = 0

        records.append({
            "nome":                  nome,
            "consultor_responsavel": coerce_str(row_vals[1]) if len(row_vals) > 1 else None,
            "total_lojas":           total_lojas,
            # lojas_ativas NOT NULL no schema: nao existe na planilha, usa 0
            "lojas_ativas":          0,
            "faturamento_real":      coerce_float(row_vals[3]) if len(row_vals) > 3 else None,
            "potencial_maximo":      coerce_float(row_vals[4]) if len(row_vals) > 4 else None,
            "pct_penetracao":        coerce_float(row_vals[5]) if len(row_vals) > 5 else None,
            "updated_at":            now_iso,
        })

    inserted = batch_insert(cur, SQL_INSERT_REDE, records)
    conn.commit()
    log.info("REDES v2: %d redes inseridas", inserted)
    return inserted


# ── Limpar tabelas antes de recarregar (idempotente) ─────────────────────────

def limpar_tabelas(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    cur.execute("DELETE FROM log_interacoes")
    cur.execute("DELETE FROM redes")
    conn.commit()
    log.info("Tabelas log_interacoes e redes limpas para recarga idempotente.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    log.info("Abrindo planilha: %s", XLSX_PATH)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)

    log.info("Conectando ao banco: %s", DB_PATH)
    conn = sqlite3.connect(DB_PATH)

    # Recarga idempotente
    limpar_tabelas(conn)

    # ── TAREFA 1: LOGs ────────────────────────────────────────────────────────
    ABAS_LOG = ["LARISSA", "MANU", "JULIO", "DAIANE"]
    ws_map = {nome: wb[nome] for nome in ABAS_LOG}
    contagens_log = popular_logs(ws_map, conn)

    # ── TAREFA 2: REDES ───────────────────────────────────────────────────────
    ws_redes = wb["REDES v2"]
    total_redes = popular_redes(ws_redes, conn)

    wb.close()
    conn.close()

    # ── Relatorio final ───────────────────────────────────────────────────────
    print()
    print("=" * 60)
    print("RELATORIO FINAL — populate_log_redes.py")
    print("=" * 60)
    total_logs = 0
    for aba in ABAS_LOG:
        n = contagens_log.get(aba, 0)
        total_logs += n
        print(f"  LOG {aba:<10}: {n:>5} registros inseridos")
    print(f"  REDES       : {total_redes:>5} redes inseridas")
    print(f"  TOTAL LOGs  : {total_logs:>5}")
    print("=" * 60)
    print("Two-Base Architecture: RESPEITADA (nenhum R$ em LOG)")
    print("CNPJ: normalizado (14 digits, string, zero-padded)")
    print("Dados: REAL (fonte Excel INTELIGENTE FINAL OK)")
    print("=" * 60)


if __name__ == "__main__":
    main()

"""Check DRAFT 2 column structure in V2.xlsx"""
import openpyxl, os
path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output", "CRM INTELIGENTE - VITAO360 V2.xlsx")
wb = openpyxl.load_workbook(path, data_only=False)
ws = wb['DRAFT 2']
print(f"DRAFT 2: rows={ws.max_row}, cols={ws.max_column}")
print(f"Freeze: {ws.freeze_panes}")
print()

# Row 1 = title, Row 2 = headers, Row 3+ = data
for r in [1, 2]:
    print(f"--- Row {r} ---")
    for c in range(1, min(ws.max_column or 1, 30) + 1):
        v = ws.cell(row=r, column=c).value
        if v:
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"  {col_letter}{r}: {v}")

# Check data validations
dvs = list(ws.data_validations.dataValidation) if ws.data_validations else []
print(f"\nData Validations: {len(dvs)}")
for dv in dvs:
    print(f"  {dv.sqref} -> type={dv.type}, formula1={dv.formula1}")

# Check sample formulas row 3
print("\n--- Row 3 formulas ---")
for c in range(1, min(ws.max_column or 1, 30) + 1):
    v = ws.cell(row=3, column=c).value
    if v:
        col_letter = openpyxl.utils.get_column_letter(c)
        print(f"  {col_letter}3: {v}")

# Also check REGRAS named ranges for reference
print("\n--- Named Ranges ---")
for dn in wb.defined_names.values():
    print(f"  {dn.name} = {dn.attr_text}")

# Check REGRAS RESULTADO table (for GRUPO DASH column)
ws_r = wb['REGRAS']
print("\n--- REGRAS RESULTADO table ---")
for r in range(1, 15):
    vals = []
    for c in range(1, 6):
        v = ws_r.cell(row=r, column=c).value
        vals.append(str(v) if v else "")
    print(f"  Row {r}: {vals}")

# Check REGRAS MOTIVO table
print("\n--- REGRAS MOTIVO table (rows 27-38) ---")
for r in range(27, 39):
    vals = []
    for c in range(1, 4):
        v = ws_r.cell(row=r, column=c).value
        vals.append(str(v) if v else "")
    print(f"  Row {r}: {vals}")

# Check consultores
print("\n--- REGRAS CONSULTOR (rows 88-92) ---")
for r in range(87, 93):
    vals = []
    for c in range(1, 5):
        v = ws_r.cell(row=r, column=c).value
        vals.append(str(v) if v else "")
    print(f"  Row {r}: {vals}")

wb.close()

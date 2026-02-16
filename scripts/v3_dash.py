"""
V3 — Tab DASH: 7 blocos de KPIs + KPI cards no topo
V3 DRAFT 2 column mapping: A=DATA, B=CNPJ, E=CONSULTOR, F=RESULTADO,
G=MOTIVO, H=WHATSAPP, I=LIGAÇÃO, J=LIG.ATENDIDA, M=SITUAÇÃO,
P=TIPO DO CONTATO, S=GRUPO DASH, T=FOLLOW-UP, W=SINALEIRO CICLO
"""
import datetime
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from v3_styles import *

# Shared constants
TIPOS = [
    "PROSPECÇÃO", "NEGOCIAÇÃO", "FOLLOW UP",
    "ATEND. CLIENTES ATIVOS", "ATEND. CLIENTES INATIVOS",
    "PÓS-VENDA / RELACIONAMENTO", "PERDA / NUTRIÇÃO",
]
CONSULTORES = ["MANU DITZEL", "LARISSA PADILHA", "JULIO GADRET", "DAIANE STAVICKI"]
SITUACOES = ["ATIVO", "EM RISCO", "INAT.REC", "INAT.ANT", "NOVO", "PROSPECT"]
FASES = ["PÓS-VENDA", "CS", "RECOMPRA", "SALVAMENTO", "RECUPERAÇÃO",
         "PROSPECÇÃO", "NUTRIÇÃO", "EM ATENDIMENTO", "ORÇAMENTO"]
REDES = ["CIA DA SAUDE", "FITLAND", "DIVINA TERRA", "VIDA LEVE",
         "BIO MUNDO", "MUNDO VERDE", "TUDO EM GRAOS / VGA", "DEMAIS CLIENTES"]
MOTIVOS = [
    "AINDA TEM ESTOQUE", "PRODUTO NÃO VENDEU / SEM GIRO",
    "LOJA ANEXO/PROXIMO - SM", "SÓ QUER COMPRAR GRANEL",
    "PROBLEMA LOGÍSTICO / ENTREGA", "PROBLEMA FINANCEIRO / CRÉDITO",
    "PROPRIETARIO / INDISPONÍVEL", "FECHANDO / FECHOU LOJA",
    "SEM INTERESSE NO MOMENTO", "PRIMEIRO CONTATO / SEM RESPOSTA",
]


def cf(extra_criteria="", use_todos=True):
    """Build COUNTIFS with period + consultor filter for V3 DRAFT 2."""
    date_filter = "'DRAFT 2'!$A$3:$A$5000,\">=\"&$E$2,'DRAFT 2'!$A$3:$A$5000,\"<=\"&$F$2"
    cons_filter = ",'DRAFT 2'!$E$3:$E$5000,$C$2"
    if use_todos:
        base_all = f"COUNTIFS({date_filter}{extra_criteria})"
        base_cons = f"COUNTIFS({date_filter}{cons_filter}{extra_criteria})"
        return f'IF(OR($C$2="",$C$2="TODOS"),{base_all},{base_cons})'
    else:
        return f"COUNTIFS({date_filter}{extra_criteria})"


def section_title(ws, row, label, end_col=17):
    """Write a section title bar."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=1, value=f"  {label}")
    style_cell(cell, font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(1, end_col + 1):
        ws.cell(row=row, column=c).fill = FILL_GRAY_D9
        ws.cell(row=row, column=c).border = THIN_BORDER


def total_row(ws, row, cols, data_start, data_end, fill=FILL_DARK, label="TOTAL", label_col=1):
    """Write a TOTAL row with SUM formulas."""
    write_header(ws, row, label_col, label, fill=fill, align=ALIGN_LEFT)
    for c in cols:
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f'=SUM({cl}{data_start}:{cl}{data_end})'
        write_header(ws, row, c, ws.cell(row=row, column=c).value, fill=fill)


def build_dash(wb):
    """Build DASH tab with 7 blocks of KPIs + KPI cards."""
    ws = wb.create_sheet("DASH")
    ws.sheet_properties.tabColor = TAB_DASH

    # Column widths
    ws.column_dimensions['A'].width = 28
    for c in range(2, 20):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.column_dimensions['H'].width = 2  # separator
    ws.column_dimensions['L'].width = 2
    ws.column_dimensions['O'].width = 2

    # Date filter refs (reused everywhere)
    D2 = "'DRAFT 2'!$A$3:$A$5000"
    date_base = f"{D2},\">=\"&$E$2,{D2},\"<=\"&$F$2"

    # ═══════════════════════════════════════════════════════════
    # ROW 1: Title
    # ═══════════════════════════════════════════════════════════
    ws.merge_cells('A1:Q1')
    cell = ws.cell(row=1, column=1, value="📊 DASHBOARD JARVIS CRM — VITAO ALIMENTOS")
    style_cell(cell, font=Font(name='Calibri', size=13, bold=True, color='FFFFFF'),
               fill=PatternFill('solid', fgColor='DC2626'), align=ALIGN_LEFT)
    for c in range(1, 18):
        ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor='DC2626')
        ws.cell(row=1, column=c).border = THIN_BORDER

    # ═══════════════════════════════════════════════════════════
    # ROW 2: Filters (Vendedor + Período)
    # ═══════════════════════════════════════════════════════════
    write_data(ws, 2, 1, "VENDEDOR", font=FONT_TITLE, fill=FILL_LIGHT_GRAY, align=ALIGN_LEFT)
    ws.cell(row=2, column=3, value="TODOS").font = FONT_DATA
    ws.cell(row=2, column=3).border = THIN_BORDER
    dv_cons = DataValidation(type="list",
        formula1='"TODOS,MANU DITZEL,LARISSA PADILHA,JULIO GADRET,DAIANE STAVICKI"',
        allow_blank=True)
    ws.add_data_validation(dv_cons)
    dv_cons.add("C2")

    write_data(ws, 2, 4, "PERÍODO", font=FONT_TITLE, fill=FILL_LIGHT_GRAY, align=ALIGN_LEFT)
    ws.cell(row=2, column=5, value=datetime.date(2026, 2, 1)).font = FONT_DATA
    ws.cell(row=2, column=5).number_format = FMT_DATE
    ws.cell(row=2, column=5).border = THIN_BORDER
    ws.cell(row=2, column=6, value=datetime.date(2026, 2, 28)).font = FONT_DATA
    ws.cell(row=2, column=6).number_format = FMT_DATE
    ws.cell(row=2, column=6).border = THIN_BORDER

    # ═══════════════════════════════════════════════════════════
    # ROW 3-5: KPI CARDS (8 summary cards)
    # ═══════════════════════════════════════════════════════════
    kpi_row = 3
    kpis = [
        ("📞 TOTAL CONTATOS", 1, f'={cf()}'),
        ("💬 WHATSAPP",        3, f'={cf(",\'DRAFT 2\'!$H$3:$H$5000,\"SIM\"")}'),
        ("📱 LIGAÇÕES",        5, f'={cf(",\'DRAFT 2\'!$I$3:$I$5000,\"SIM\"")}'),
        ("✅ LIG. ATENDIDAS",  7, f'={cf(",\'DRAFT 2\'!$J$3:$J$5000,\"SIM\"")}'),
        ("🔍 PROSPECÇÕES",     9, f'={cf(",\'DRAFT 2\'!$P$3:$P$5000,\"PROSPECÇÃO\"")}'),
        ("📋 ORÇAMENTOS",      11, f'={cf(",\'DRAFT 2\'!$F$3:$F$5000,\"ORÇAMENTO\"")}'),
        ("🔄 FOLLOW UPS",      13, f'={cf(",\'DRAFT 2\'!$F$3:$F$5000,\"FOLLOW UP 7\"")}+{cf(",\'DRAFT 2\'!$F$3:$F$5000,\"FOLLOW UP 15\"")}'),
        ("🤝 VENDAS",          15, f'={cf(",\'DRAFT 2\'!$F$3:$F$5000,\"VENDA / PEDIDO\"")}'),
    ]
    for label, col, formula in kpis:
        ws.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=col + 1)
        write_data(ws, kpi_row, col, label, font=FONT_DATA_BOLD,
                   fill=PatternFill('solid', fgColor='E3F2FD'))
        ws.cell(row=kpi_row, column=col + 1).fill = PatternFill('solid', fgColor='E3F2FD')
        ws.cell(row=kpi_row, column=col + 1).border = THIN_BORDER
        ws.merge_cells(start_row=kpi_row + 1, start_column=col, end_row=kpi_row + 1, end_column=col + 1)
        ws.cell(row=kpi_row + 1, column=col).value = formula
        style_cell(ws.cell(row=kpi_row + 1, column=col),
                   font=Font(name='Calibri', size=16, bold=True),
                   fill=PatternFill('solid', fgColor='E3F2FD'))
        ws.cell(row=kpi_row + 1, column=col + 1).fill = PatternFill('solid', fgColor='E3F2FD')
        ws.cell(row=kpi_row + 1, column=col + 1).border = THIN_BORDER
    ws.row_dimensions[kpi_row].height = 20
    ws.row_dimensions[kpi_row + 1].height = 28

    # ═══════════════════════════════════════════════════════════
    # BLOCO 1: TIPO DO CONTATO × RESULTADO (rows 7-17)
    # ═══════════════════════════════════════════════════════════
    B1 = 7
    section_title(ws, B1, "TIPO DO CONTATO × RESULTADO DO CONTATO", 13)

    HDR = B1 + 1
    hdr_labels = [
        ("TIPO DO CONTATO", FILL_DARK, ALIGN_LEFT),
        ("TOTAL", FILL_DARK, ALIGN_CENTER),
        ("ORÇAM.", FILL_DARK, ALIGN_CENTER),
        ("CADAST.", FILL_DARK, ALIGN_CENTER),
        ("RELAC.", FILL_DARK, ALIGN_CENTER),
        ("EM ATEND.", FILL_DARK, ALIGN_CENTER),
        ("SUPORTE", FILL_DARK, ALIGN_CENTER),
        ("VENDA", FILL_DARK, ALIGN_CENTER),
        ("Ñ ATENDE", FILL_RED_DARK, ALIGN_CENTER),
        ("RECUSOU", FILL_RED_DARK, ALIGN_CENTER),
        ("Ñ RESP.", FILL_RED_DARK, ALIGN_CENTER),
        ("PERDA", FILL_RED_DARK, ALIGN_CENTER),
        ("FOLLOW UP", FILL_GREEN_DARK, ALIGN_CENTER),
    ]
    for i, (label, fill, align) in enumerate(hdr_labels):
        write_header(ws, HDR, i + 1, label, fill=fill, font=FONT_HEADER_SM, align=align)

    resultado_map = {
        3: "ORÇAMENTO", 4: "CADASTRO", 5: "RELACIONAMENTO",
        6: "EM ATENDIMENTO", 7: "SUPORTE", 8: "VENDA / PEDIDO",
        9: "NÃO ATENDE", 10: "RECUSOU LIGAÇÃO", 11: "NÃO RESPONDE",
        12: "PERDA / FECHOU LOJA",
    }

    DS = HDR + 1
    for t_idx, tipo in enumerate(TIPOS):
        r = DS + t_idx
        write_data(ws, r, 1, tipo, align=ALIGN_LEFT)
        ws.cell(row=r, column=2).value = f'=SUM(C{r}:M{r})'
        style_cell(ws.cell(row=r, column=2), font=FONT_DATA_BOLD)

        tipo_f = f",'DRAFT 2'!$P$3:$P$5000,A{r}"
        for col_idx, resultado in resultado_map.items():
            res_f = f",'DRAFT 2'!$F$3:$F$5000,\"{resultado}\""
            ws.cell(row=r, column=col_idx).value = f'={cf(tipo_f + res_f)}'
            style_cell(ws.cell(row=r, column=col_idx))

        fu7 = f",'DRAFT 2'!$F$3:$F$5000,\"FOLLOW UP 7\""
        fu15 = f",'DRAFT 2'!$F$3:$F$5000,\"FOLLOW UP 15\""
        ws.cell(row=r, column=13).value = f'={cf(tipo_f + fu7)}+{cf(tipo_f + fu15)}'
        style_cell(ws.cell(row=r, column=13))

    TR = DS + len(TIPOS)
    total_row(ws, TR, list(range(2, 14)), DS, TR - 1)

    # Bar Chart
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "RESULTADO DO CONTATO"
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.legend = None
    chart.width = 18
    chart.height = 12
    data_ref = Reference(ws, min_col=3, max_col=13, min_row=TR, max_row=TR)
    cats_ref = Reference(ws, min_col=3, max_col=13, min_row=HDR, max_row=HDR)
    chart.add_data(data_ref, from_rows=True, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart_colors = {0:"2196F3",1:"2196F3",2:"2196F3",3:"FF9800",4:"7B1FA2",
                    5:"404040",6:"C62828",7:"C62828",8:"C62828",9:"333333",10:"4CAF50"}
    series = chart.series[0]
    series.graphicalProperties.line.noFill = True
    for pt_idx, color in chart_colors.items():
        pt = DataPoint(idx=pt_idx)
        pt.graphicalProperties.solidFill = color
        series.data_points.append(pt)
    series.dLbls = DataLabelList()
    series.dLbls.showVal = True
    series.dLbls.numFmt = '#,##0'
    ws.add_chart(chart, "O8")

    # ═══════════════════════════════════════════════════════════
    # BLOCO 2: CONTATOS + FUNIL (rows 20-30)
    # ═══════════════════════════════════════════════════════════
    B2 = TR + 3
    section_title(ws, B2, "CONTATOS & FUNIL DE VENDA", 17)

    GR = B2 + 1
    for c in [1, 2]:
        ws.cell(row=GR, column=c).fill = FILL_DARK_59
        ws.cell(row=GR, column=c).border = THIN_BORDER
    ws.merge_cells(f'C{GR}:F{GR}')
    write_header(ws, GR, 3, "CANAIS", fill=FILL_DARK_59)
    for c in [4, 5, 6]:
        ws.cell(row=GR, column=c).fill = FILL_DARK_59
        ws.cell(row=GR, column=c).border = THIN_BORDER
    ws.merge_cells(f'I{GR}:K{GR}')
    write_header(ws, GR, 9, "FUNIL DE VENDA", fill=FILL_DARK_59)
    for c in [10, 11]:
        ws.cell(row=GR, column=c).fill = FILL_DARK_59
        ws.cell(row=GR, column=c).border = THIN_BORDER
    ws.merge_cells(f'M{GR}:N{GR}')
    write_header(ws, GR, 13, "RELACIONAMENTO", fill=FILL_DARK_59)
    ws.cell(row=GR, column=14).fill = FILL_DARK_59
    ws.cell(row=GR, column=14).border = THIN_BORDER
    ws.merge_cells(f'P{GR}:Q{GR}')
    write_header(ws, GR, 16, "NÃO VENDA", fill=FILL_DARK_59)
    ws.cell(row=GR, column=17).fill = FILL_DARK_59
    ws.cell(row=GR, column=17).border = THIN_BORDER

    SR = GR + 1
    sub_headers = [
        (1, "TIPO DO CONTATO", ALIGN_LEFT), (2, "TOTAL", ALIGN_CENTER),
        (3, "WHATSAPP", ALIGN_CENTER), (4, "LIGAÇÃO", ALIGN_CENTER),
        (5, "LIG. ATEND.", ALIGN_CENTER), (6, "LIG. Ñ ATEND.", ALIGN_CENTER),
        (9, "EM ATEND.", ALIGN_CENTER), (10, "ORÇAMENTO", ALIGN_CENTER),
        (11, "VENDA", ALIGN_CENTER),
        (13, "FOLLOW UP", ALIGN_CENTER), (14, "SUPORTE", ALIGN_CENTER),
        (16, "INATIVO", ALIGN_CENTER), (17, "PERDA", ALIGN_CENTER),
    ]
    for col, label, align in sub_headers:
        write_header(ws, SR, col, label, fill=FILL_DARK, font=FONT_HEADER_SM, align=align)

    # Separators
    for sep_r in range(GR, GR + 13):
        for sep_c in [7, 8, 12, 15]:
            ws.cell(row=sep_r, column=sep_c).fill = FILL_SEP
            ws.cell(row=sep_r, column=sep_c).border = NO_BORDER

    B2D = SR + 1
    for t_idx, tipo in enumerate(TIPOS):
        r = B2D + t_idx
        write_data(ws, r, 1, tipo, align=ALIGN_LEFT)
        tipo_f = f",'DRAFT 2'!$P$3:$P$5000,A{r}"

        write_data(ws, r, 2, f'={cf(tipo_f)}', font=FONT_DATA_BOLD)
        wpp_f = f",'DRAFT 2'!$H$3:$H$5000,\"SIM\""
        ws.cell(row=r, column=3).value = f'={cf(tipo_f + wpp_f)}'
        style_cell(ws.cell(row=r, column=3))
        lig_f = f",'DRAFT 2'!$I$3:$I$5000,\"SIM\""
        ws.cell(row=r, column=4).value = f'={cf(tipo_f + lig_f)}'
        style_cell(ws.cell(row=r, column=4))
        lig_at_f = f",'DRAFT 2'!$J$3:$J$5000,\"SIM\""
        ws.cell(row=r, column=5).value = f'={cf(tipo_f + lig_at_f)}'
        style_cell(ws.cell(row=r, column=5))
        ws.cell(row=r, column=6).value = f'=D{r}-E{r}'
        style_cell(ws.cell(row=r, column=6))

        for sep_c in [7, 8, 12, 15]:
            ws.cell(row=r, column=sep_c).fill = FILL_SEP
            ws.cell(row=r, column=sep_c).border = NO_BORDER

        for c, res in [(9,"EM ATENDIMENTO"),(10,"ORÇAMENTO"),(11,"VENDA / PEDIDO")]:
            res_f = f",'DRAFT 2'!$F$3:$F$5000,\"{res}\""
            ws.cell(row=r, column=c).value = f'={cf(tipo_f + res_f)}'
            style_cell(ws.cell(row=r, column=c))

        fu7_f = f",'DRAFT 2'!$F$3:$F$5000,\"FOLLOW UP 7\""
        fu15_f = f",'DRAFT 2'!$F$3:$F$5000,\"FOLLOW UP 15\""
        rel_f = f",'DRAFT 2'!$F$3:$F$5000,\"RELACIONAMENTO\""
        ws.cell(row=r, column=13).value = f'={cf(tipo_f+fu7_f)}+{cf(tipo_f+fu15_f)}+{cf(tipo_f+rel_f)}'
        style_cell(ws.cell(row=r, column=13))
        sup_f = f",'DRAFT 2'!$F$3:$F$5000,\"SUPORTE\""
        ws.cell(row=r, column=14).value = f'={cf(tipo_f + sup_f)}'
        style_cell(ws.cell(row=r, column=14))

        na_f = f",'DRAFT 2'!$F$3:$F$5000,\"NÃO ATENDE\""
        nr_f = f",'DRAFT 2'!$F$3:$F$5000,\"NÃO RESPONDE\""
        rec_f = f",'DRAFT 2'!$F$3:$F$5000,\"RECUSOU LIGAÇÃO\""
        ws.cell(row=r, column=16).value = f'={cf(tipo_f+na_f)}+{cf(tipo_f+nr_f)}+{cf(tipo_f+rec_f)}'
        style_cell(ws.cell(row=r, column=16))
        perda_f = f",'DRAFT 2'!$F$3:$F$5000,\"PERDA / FECHOU LOJA\""
        ws.cell(row=r, column=17).value = f'={cf(tipo_f + perda_f)}'
        style_cell(ws.cell(row=r, column=17))

    B2T = B2D + len(TIPOS)
    total_row(ws, B2T, [2,3,4,5,6,9,10,11,13,14,16,17], B2D, B2T - 1)
    for sep_c in [7, 8, 12, 15]:
        ws.cell(row=B2T, column=sep_c).fill = FILL_SEP
        ws.cell(row=B2T, column=sep_c).border = NO_BORDER

    # ═══════════════════════════════════════════════════════════
    # BLOCO 3: MOTIVOS + PRODUTIVIDADE (side by side)
    # ═══════════════════════════════════════════════════════════
    B3 = B2T + 3
    # Left: Motivos
    ws.merge_cells(f'A{B3}:G{B3}')
    ws.cell(row=B3, column=1, value="  MOTIVOS — POR QUE NÃO COMPRAM")
    style_cell(ws.cell(row=B3, column=1), font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(1, 8):
        ws.cell(row=B3, column=c).fill = FILL_GRAY_D9
        ws.cell(row=B3, column=c).border = THIN_BORDER

    # Right: Produtividade
    ws.merge_cells(f'I{B3}:Q{B3}')
    ws.cell(row=B3, column=9, value="  PRODUTIVIDADE POR CONSULTOR")
    style_cell(ws.cell(row=B3, column=9), font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(9, 18):
        ws.cell(row=B3, column=c).fill = FILL_GRAY_D9
        ws.cell(row=B3, column=c).border = THIN_BORDER

    # MOTIVOS headers
    MH = B3 + 1
    for col, label, align in [(1,"MOTIVO",ALIGN_LEFT),(2,"QTD",ALIGN_CENTER),(3,"%",ALIGN_CENTER),
                               (4,"PROSP",ALIGN_CENTER),(5,"ATIVOS",ALIGN_CENTER),
                               (6,"INAT",ALIGN_CENTER),(7,"PÓS-V",ALIGN_CENTER)]:
        write_header(ws, MH, col, label, fill=FILL_DARK, align=align)

    MD = MH + 1
    for m_idx, motivo in enumerate(MOTIVOS):
        r = MD + m_idx
        write_data(ws, r, 1, motivo, align=ALIGN_LEFT)
        mot_f = f",'DRAFT 2'!$G$3:$G$5000,A{r}"
        ws.cell(row=r, column=2).value = f'={cf(mot_f)}'
        style_cell(ws.cell(row=r, column=2))
        for c, tipo_val in [(4,"PROSPECÇÃO"),(5,"ATEND. CLIENTES ATIVOS"),
                            (6,"ATEND. CLIENTES INATIVOS"),(7,"PÓS-VENDA / RELACIONAMENTO")]:
            tf = f",'DRAFT 2'!$P$3:$P$5000,\"{tipo_val}\""
            ws.cell(row=r, column=c).value = f'={cf(mot_f + tf)}'
            style_cell(ws.cell(row=r, column=c))

    MT = MD + len(MOTIVOS)
    total_row(ws, MT, [2, 4, 5, 6, 7], MD, MT - 1)

    # % column
    for m_idx in range(len(MOTIVOS)):
        r = MD + m_idx
        ws.cell(row=r, column=3).value = f'=IFERROR(B{r}/B{MT},0)'
        ws.cell(row=r, column=3).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=3))
    ws.cell(row=MT, column=3).value = f'=IFERROR(B{MT}/B{MT},0)'
    ws.cell(row=MT, column=3).number_format = FMT_PCT
    write_header(ws, MT, 3, ws.cell(row=MT, column=3).value, fill=FILL_DARK)

    # PRODUTIVIDADE headers
    IH = B3 + 1
    for col, label, align in [(9,"CONSULTOR",ALIGN_LEFT),(10,"CONTATOS",ALIGN_CENTER),
                               (11,"VENDAS",ALIGN_CENTER),(12,"ORÇAM.",ALIGN_CENTER),
                               (13,"CADAST.",ALIGN_CENTER),(14,"% CONV",ALIGN_CENTER),
                               (15,"Ñ ATENDE",ALIGN_CENTER),(16,"PERDA",ALIGN_CENTER),
                               (17,"% MERCOS",ALIGN_CENTER)]:
        write_header(ws, IH, col, label, fill=FILL_DARK, align=align)

    ID = IH + 1
    for c_idx, cons in enumerate(CONSULTORES):
        r = ID + c_idx
        write_data(ws, r, 9, cons, align=ALIGN_LEFT)
        cf_cons = f",'DRAFT 2'!$E$3:$E$5000,I{r}"
        ws.cell(row=r, column=10).value = f'=COUNTIFS({date_base}{cf_cons})'
        style_cell(ws.cell(row=r, column=10))
        for c, res in [(11,"VENDA / PEDIDO"),(12,"ORÇAMENTO"),(13,"CADASTRO")]:
            res_f = f",'DRAFT 2'!$F$3:$F$5000,\"{res}\""
            ws.cell(row=r, column=c).value = f'=COUNTIFS({date_base}{cf_cons}{res_f})'
            style_cell(ws.cell(row=r, column=c))
        ws.cell(row=r, column=14).value = f'=IFERROR(K{r}/J{r},0)'
        ws.cell(row=r, column=14).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=14))
        na_f = f",'DRAFT 2'!$F$3:$F$5000,\"NÃO ATENDE\""
        nr_f = f",'DRAFT 2'!$F$3:$F$5000,\"NÃO RESPONDE\""
        rec_f = f",'DRAFT 2'!$F$3:$F$5000,\"RECUSOU LIGAÇÃO\""
        ws.cell(row=r, column=15).value = f'=COUNTIFS({date_base}{cf_cons}{na_f})+COUNTIFS({date_base}{cf_cons}{nr_f})+COUNTIFS({date_base}{cf_cons}{rec_f})'
        style_cell(ws.cell(row=r, column=15))
        perda_f = f",'DRAFT 2'!$F$3:$F$5000,\"PERDA / FECHOU LOJA\""
        ws.cell(row=r, column=16).value = f'=COUNTIFS({date_base}{cf_cons}{perda_f})'
        style_cell(ws.cell(row=r, column=16))
        mercos_f = f",'DRAFT 2'!$L$3:$L$5000,\"SIM\""
        ws.cell(row=r, column=17).value = f'=IFERROR(COUNTIFS({date_base}{cf_cons}{mercos_f})/COUNTIFS({date_base}{cf_cons}),0)'
        ws.cell(row=r, column=17).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=17))

    IT = ID + len(CONSULTORES)
    write_header(ws, IT, 9, "TOTAL EQUIPE", fill=FILL_DARK, align=ALIGN_LEFT)
    for c in [10,11,12,13,15,16]:
        cl = get_column_letter(c)
        ws.cell(row=IT, column=c).value = f'=SUM({cl}{ID}:{cl}{IT-1})'
        write_header(ws, IT, c, ws.cell(row=IT, column=c).value, fill=FILL_DARK)
    ws.cell(row=IT, column=14).value = f'=IFERROR(K{IT}/J{IT},0)'
    ws.cell(row=IT, column=14).number_format = FMT_PCT
    write_header(ws, IT, 14, ws.cell(row=IT, column=14).value, fill=FILL_DARK)
    mercos_f = f",'DRAFT 2'!$L$3:$L$5000,\"SIM\""
    ws.cell(row=IT, column=17).value = f'=IFERROR(COUNTIFS({date_base}{mercos_f})/COUNTIFS({date_base}),0)'
    ws.cell(row=IT, column=17).number_format = FMT_PCT
    write_header(ws, IT, 17, ws.cell(row=IT, column=17).value, fill=FILL_DARK)

    # ═══════════════════════════════════════════════════════════
    # BLOCO 4: SITUAÇÃO CRM + FASES (side by side, from CARTEIRA)
    # ═══════════════════════════════════════════════════════════
    B4 = max(MT, IT) + 3
    # Left: Situação CRM
    ws.merge_cells(f'A{B4}:D{B4}')
    ws.cell(row=B4, column=1, value="  SITUAÇÃO CRM (CARTEIRA)")
    style_cell(ws.cell(row=B4, column=1), font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(1, 5):
        ws.cell(row=B4, column=c).fill = FILL_GRAY_D9
        ws.cell(row=B4, column=c).border = THIN_BORDER

    # Middle: Fases do Funil
    ws.merge_cells(f'F{B4}:I{B4}')
    ws.cell(row=B4, column=6, value="  FASES DO FUNIL (CARTEIRA)")
    style_cell(ws.cell(row=B4, column=6), font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(6, 10):
        ws.cell(row=B4, column=c).fill = FILL_GRAY_D9
        ws.cell(row=B4, column=c).border = THIN_BORDER

    # Right: Indicadores
    ws.merge_cells(f'K{B4}:N{B4}')
    ws.cell(row=B4, column=11, value="  INDICADORES CARTEIRA")
    style_cell(ws.cell(row=B4, column=11), font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(11, 15):
        ws.cell(row=B4, column=c).fill = FILL_GRAY_D9
        ws.cell(row=B4, column=c).border = THIN_BORDER

    # Situação headers
    SH = B4 + 1
    for col, label in [(1,"SITUAÇÃO"),(2,"QTD"),(3,"%"),(4,"")]:
        write_header(ws, SH, col, label, fill=FILL_DARK)

    SD = SH + 1
    for s_idx, sit in enumerate(SITUACOES):
        r = SD + s_idx
        write_data(ws, r, 1, sit, align=ALIGN_LEFT)
        # COUNTIF from CARTEIRA col N (14) = SITUAÇÃO
        ws.cell(row=r, column=2).value = f'=COUNTIF(CARTEIRA!$N$4:$N$5000,A{r})'
        style_cell(ws.cell(row=r, column=2))
        # Conditional fill for situação
        if sit in SITUACAO_FILLS:
            ws.cell(row=r, column=1).fill = SITUACAO_FILLS[sit]
            ws.cell(row=r, column=1).font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

    ST = SD + len(SITUACOES)
    write_header(ws, ST, 1, "TOTAL", fill=FILL_DARK, align=ALIGN_LEFT)
    ws.cell(row=ST, column=2).value = f'=SUM(B{SD}:B{ST-1})'
    write_header(ws, ST, 2, ws.cell(row=ST, column=2).value, fill=FILL_DARK)

    for s_idx in range(len(SITUACOES)):
        r = SD + s_idx
        ws.cell(row=r, column=3).value = f'=IFERROR(B{r}/B{ST},0)'
        ws.cell(row=r, column=3).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=3))
    ws.cell(row=ST, column=3).value = f'=IFERROR(B{ST}/B{ST},0)'
    ws.cell(row=ST, column=3).number_format = FMT_PCT
    write_header(ws, ST, 3, ws.cell(row=ST, column=3).value, fill=FILL_DARK)

    # Fases headers
    FH = B4 + 1
    for col, label in [(6,"FASE"),(7,"QTD"),(8,"%")]:
        write_header(ws, FH, col, label, fill=FILL_DARK)

    FD = FH + 1
    for f_idx, fase in enumerate(FASES):
        r = FD + f_idx
        write_data(ws, r, 6, fase, align=ALIGN_LEFT)
        # COUNTIF from CARTEIRA — FASE is in FUNIL block, calculated field
        # Using DRAFT 2 col O (15) = FASE
        ws.cell(row=r, column=7).value = f'=COUNTIFS(\'DRAFT 2\'!$O$3:$O$5000,F{r})'
        style_cell(ws.cell(row=r, column=7))

    FT = FD + len(FASES)
    write_header(ws, FT, 6, "TOTAL", fill=FILL_DARK, align=ALIGN_LEFT)
    ws.cell(row=FT, column=7).value = f'=SUM(G{FD}:G{FT-1})'
    write_header(ws, FT, 7, ws.cell(row=FT, column=7).value, fill=FILL_DARK)

    for f_idx in range(len(FASES)):
        r = FD + f_idx
        ws.cell(row=r, column=8).value = f'=IFERROR(G{r}/G{FT},0)'
        ws.cell(row=r, column=8).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=8))
    ws.cell(row=FT, column=8).value = f'=IFERROR(G{FT}/G{FT},0)'
    ws.cell(row=FT, column=8).number_format = FMT_PCT
    write_header(ws, FT, 8, ws.cell(row=FT, column=8).value, fill=FILL_DARK)

    # Indicadores Carteira
    IKH = B4 + 1
    for col, label in [(11,"INDICADOR"),(12,"VALOR")]:
        write_header(ws, IKH, col, label, fill=FILL_DARK)

    ind_data = [
        ("% ATIVO", f'=IFERROR(COUNTIF(CARTEIRA!$N$4:$N$5000,"ATIVO")/COUNTA(CARTEIRA!$N$4:$N$5000),0)'),
        ("% RETENÇÃO", f'=IFERROR((COUNTIF(CARTEIRA!$N$4:$N$5000,"ATIVO")+COUNTIF(CARTEIRA!$N$4:$N$5000,"EM RISCO"))/COUNTA(CARTEIRA!$N$4:$N$5000),0)'),
        ("% PERDA", f'=IFERROR((COUNTIF(CARTEIRA!$N$4:$N$5000,"INAT.REC")+COUNTIF(CARTEIRA!$N$4:$N$5000,"INAT.ANT"))/COUNTA(CARTEIRA!$N$4:$N$5000),0)'),
    ]
    for i, (label, formula) in enumerate(ind_data):
        r = IKH + 1 + i
        write_data(ws, r, 11, label, align=ALIGN_LEFT, font=FONT_DATA_BOLD)
        ws.cell(row=r, column=12).value = formula
        ws.cell(row=r, column=12).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=12), font=FONT_DATA_BOLD)

    # ═══════════════════════════════════════════════════════════
    # BLOCO 5: SITUAÇÃO POR REDE DE FRANQUIA
    # ═══════════════════════════════════════════════════════════
    B5 = max(ST, FT) + 3
    section_title(ws, B5, "SITUAÇÃO POR REDE DE FRANQUIA", 9)

    RH = B5 + 1
    for col, label in [(1,"REDE"),(2,"TOTAL"),(3,"ATIVOS"),(4,"EM RISCO"),
                        (5,"INAT.REC"),(6,"INAT.ANT"),(7,"NOVO"),(8,"PROSPECT"),(9,"% ATIVO")]:
        write_header(ws, RH, col, label, fill=FILL_DARK, font=FONT_HEADER_SM)

    RD = RH + 1
    for r_idx, rede in enumerate(REDES):
        r = RD + r_idx
        write_data(ws, r, 1, rede, align=ALIGN_LEFT)
        if rede == "DEMAIS CLIENTES":
            # All clients not in the named redes
            other_redes = [rd for rd in REDES if rd != "DEMAIS CLIENTES"]
            rede_excludes = "".join([f'-COUNTIFS(CARTEIRA!$I$4:$I$5000,"{rd}",CARTEIRA!$N$4:$N$5000,"{sit}")' for rd in other_redes for sit in [""]])
            ws.cell(row=r, column=2).value = f'=COUNTA(CARTEIRA!$N$4:$N$5000)-SUM(B{RD}:B{r-1})'
            style_cell(ws.cell(row=r, column=2))
            for c, sit in [(3,"ATIVO"),(4,"EM RISCO"),(5,"INAT.REC"),(6,"INAT.ANT"),(7,"NOVO"),(8,"PROSPECT")]:
                cl = get_column_letter(c)
                ws.cell(row=r, column=c).value = f'=COUNTIF(CARTEIRA!$N$4:$N$5000,"{sit}")-SUM({cl}{RD}:{cl}{r-1})'
                style_cell(ws.cell(row=r, column=c))
        else:
            # CARTEIRA col I (9) = REDE/REGIONAL, col N (14) = SITUAÇÃO
            ws.cell(row=r, column=2).value = f'=COUNTIF(CARTEIRA!$I$4:$I$5000,A{r})'
            style_cell(ws.cell(row=r, column=2))
            for c, sit in [(3,"ATIVO"),(4,"EM RISCO"),(5,"INAT.REC"),(6,"INAT.ANT"),(7,"NOVO"),(8,"PROSPECT")]:
                ws.cell(row=r, column=c).value = f'=COUNTIFS(CARTEIRA!$I$4:$I$5000,A{r},CARTEIRA!$N$4:$N$5000,"{sit}")'
                style_cell(ws.cell(row=r, column=c))
        ws.cell(row=r, column=9).value = f'=IFERROR(C{r}/B{r},0)'
        ws.cell(row=r, column=9).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=9))

    RT = RD + len(REDES)
    write_header(ws, RT, 1, "TOTAL", fill=FILL_DARK, align=ALIGN_LEFT)
    for c in range(2, 9):
        cl = get_column_letter(c)
        ws.cell(row=RT, column=c).value = f'=SUM({cl}{RD}:{cl}{RT-1})'
        write_header(ws, RT, c, ws.cell(row=RT, column=c).value, fill=FILL_DARK)
    ws.cell(row=RT, column=9).value = f'=IFERROR(C{RT}/B{RT},0)'
    ws.cell(row=RT, column=9).number_format = FMT_PCT
    write_header(ws, RT, 9, ws.cell(row=RT, column=9).value, fill=FILL_DARK)

    # ═══════════════════════════════════════════════════════════
    # BLOCO 6: SAÚDE DA CARTEIRA × PRÓXIMAS AÇÕES
    # ═══════════════════════════════════════════════════════════
    B6 = RT + 3
    # Left: Sinaleiro
    ws.merge_cells(f'A{B6}:G{B6}')
    ws.cell(row=B6, column=1, value="  🚦 SAÚDE DA CARTEIRA × PRÓXIMAS AÇÕES POR CONSULTOR")
    style_cell(ws.cell(row=B6, column=1), font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(1, 8):
        ws.cell(row=B6, column=c).fill = FILL_GRAY_D9
        ws.cell(row=B6, column=c).border = THIN_BORDER

    # Right: Ações
    ws.merge_cells(f'I{B6}:N{B6}')
    ws.cell(row=B6, column=9, value="  PRÓXIMAS AÇÕES (CARTEIRA)")
    style_cell(ws.cell(row=B6, column=9), font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(9, 15):
        ws.cell(row=B6, column=c).fill = FILL_GRAY_D9
        ws.cell(row=B6, column=c).border = THIN_BORDER

    # Sinaleiro headers
    SIH = B6 + 1
    for col, label in [(1,"SINALEIRO"),(2,"MANU"),(3,"LARISSA"),(4,"JULIO"),(5,"DAIANE"),(6,"TOTAL"),(7,"% SAÚDE")]:
        write_header(ws, SIH, col, label, fill=FILL_DARK, font=FONT_HEADER_SM)

    sinaleiros = ["🟢", "🟡", "🔴"]
    SID = SIH + 1
    for s_idx, sin in enumerate(sinaleiros):
        r = SID + s_idx
        write_data(ws, r, 1, sin)
        for c_idx, cons in enumerate(CONSULTORES):
            ws.cell(row=r, column=2 + c_idx).value = (
                f'=COUNTIFS(\'DRAFT 2\'!$E$3:$E$5000,"{cons}",\'DRAFT 2\'!$W$3:$W$5000,A{r})'
            )
            style_cell(ws.cell(row=r, column=2 + c_idx))
        ws.cell(row=r, column=6).value = f'=SUM(B{r}:E{r})'
        style_cell(ws.cell(row=r, column=6), font=FONT_DATA_BOLD)

    SIT_ROW = SID + len(sinaleiros)
    write_header(ws, SIT_ROW, 1, "TOTAL", fill=FILL_DARK)
    for c in range(2, 7):
        cl = get_column_letter(c)
        ws.cell(row=SIT_ROW, column=c).value = f'=SUM({cl}{SID}:{cl}{SIT_ROW-1})'
        write_header(ws, SIT_ROW, c, ws.cell(row=SIT_ROW, column=c).value, fill=FILL_DARK)

    # % SAÚDE (verde / total)
    for s_idx in range(len(sinaleiros)):
        r = SID + s_idx
        ws.cell(row=r, column=7).value = f'=IFERROR(F{r}/F{SIT_ROW},0)'
        ws.cell(row=r, column=7).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=7))

    # Ações headers
    AH = B6 + 1
    for col, label in [(9,"AÇÃO"),(10,"MANU"),(11,"LARISSA"),(12,"JULIO"),(13,"DAIANE"),(14,"TOTAL")]:
        write_header(ws, AH, col, label, fill=FILL_DARK, font=FONT_HEADER_SM)

    acoes = ["RECOMPRA", "PÓS-VENDA", "CS", "PROSPECÇÃO", "REATIVAÇÃO", "SALVAMENTO", "NUTRIÇÃO"]
    AD = AH + 1
    for a_idx, acao in enumerate(acoes):
        r = AD + a_idx
        write_data(ws, r, 9, acao, align=ALIGN_LEFT)
        for c_idx, cons in enumerate(CONSULTORES):
            ws.cell(row=r, column=10 + c_idx).value = (
                f'=COUNTIFS(\'DRAFT 2\'!$E$3:$E$5000,"{cons}",\'DRAFT 2\'!$U$3:$U$5000,I{r})'
            )
            style_cell(ws.cell(row=r, column=10 + c_idx))
        ws.cell(row=r, column=14).value = f'=SUM(J{r}:M{r})'
        style_cell(ws.cell(row=r, column=14), font=FONT_DATA_BOLD)

    AT = AD + len(acoes)
    write_header(ws, AT, 9, "TOTAL", fill=FILL_DARK, align=ALIGN_LEFT)
    for c in range(10, 15):
        cl = get_column_letter(c)
        ws.cell(row=AT, column=c).value = f'=SUM({cl}{AD}:{cl}{AT-1})'
        write_header(ws, AT, c, ws.cell(row=AT, column=c).value, fill=FILL_DARK)

    # ═══════════════════════════════════════════════════════════
    # BLOCO 7: SCORE DISCIPLINA OPERACIONAL
    # ═══════════════════════════════════════════════════════════
    B7 = max(SIT_ROW, AT) + 3
    section_title(ws, B7, "🏆 DISCIPLINA OPERACIONAL — SCORE POR CONSULTOR", 14)

    SCH = B7 + 1
    for col, label in [(1,"CONSULTOR"),(2,"CONTATOS"),(3,"MÉDIA/DIA"),(4,"WHATSAPP"),
                        (5,"LIGAÇÕES"),(6,"% ATEND."),(7,"VENDAS"),(8,"ORÇAM."),
                        (9,"% CONV"),(10,"FOLLOW UPS"),(11,"CADAST."),(12,"% MERCOS"),
                        (13,"Ñ ATENDE"),(14,"PERDA")]:
        write_header(ws, SCH, col, label, fill=FILL_DARK, font=FONT_HEADER_SM)

    SCD = SCH + 1
    for c_idx, cons in enumerate(CONSULTORES):
        r = SCD + c_idx
        write_data(ws, r, 1, cons, align=ALIGN_LEFT)
        cf_c = f",'DRAFT 2'!$E$3:$E$5000,A{r}"

        # Contatos
        ws.cell(row=r, column=2).value = f'=COUNTIFS({date_base}{cf_c})'
        style_cell(ws.cell(row=r, column=2))
        # Média/dia (assume 22 working days)
        ws.cell(row=r, column=3).value = f'=IFERROR(B{r}/22,0)'
        ws.cell(row=r, column=3).number_format = '0.0'
        style_cell(ws.cell(row=r, column=3))
        # WhatsApp
        wpp_f = f",'DRAFT 2'!$H$3:$H$5000,\"SIM\""
        ws.cell(row=r, column=4).value = f'=COUNTIFS({date_base}{cf_c}{wpp_f})'
        style_cell(ws.cell(row=r, column=4))
        # Ligações
        lig_f = f",'DRAFT 2'!$I$3:$I$5000,\"SIM\""
        ws.cell(row=r, column=5).value = f'=COUNTIFS({date_base}{cf_c}{lig_f})'
        style_cell(ws.cell(row=r, column=5))
        # % Atendida
        lig_at = f",'DRAFT 2'!$J$3:$J$5000,\"SIM\""
        ws.cell(row=r, column=6).value = f'=IFERROR(COUNTIFS({date_base}{cf_c}{lig_at})/COUNTIFS({date_base}{cf_c}{lig_f}),0)'
        ws.cell(row=r, column=6).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=6))
        # Vendas
        venda_f = f",'DRAFT 2'!$F$3:$F$5000,\"VENDA / PEDIDO\""
        ws.cell(row=r, column=7).value = f'=COUNTIFS({date_base}{cf_c}{venda_f})'
        style_cell(ws.cell(row=r, column=7))
        # Orçamentos
        orc_f = f",'DRAFT 2'!$F$3:$F$5000,\"ORÇAMENTO\""
        ws.cell(row=r, column=8).value = f'=COUNTIFS({date_base}{cf_c}{orc_f})'
        style_cell(ws.cell(row=r, column=8))
        # % Conversão
        ws.cell(row=r, column=9).value = f'=IFERROR(G{r}/B{r},0)'
        ws.cell(row=r, column=9).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=9))
        # Follow ups
        fu7_f = f",'DRAFT 2'!$F$3:$F$5000,\"FOLLOW UP 7\""
        fu15_f = f",'DRAFT 2'!$F$3:$F$5000,\"FOLLOW UP 15\""
        ws.cell(row=r, column=10).value = f'=COUNTIFS({date_base}{cf_c}{fu7_f})+COUNTIFS({date_base}{cf_c}{fu15_f})'
        style_cell(ws.cell(row=r, column=10))
        # Cadastros
        cad_f = f",'DRAFT 2'!$F$3:$F$5000,\"CADASTRO\""
        ws.cell(row=r, column=11).value = f'=COUNTIFS({date_base}{cf_c}{cad_f})'
        style_cell(ws.cell(row=r, column=11))
        # % Mercos
        mercos_f = f",'DRAFT 2'!$L$3:$L$5000,\"SIM\""
        ws.cell(row=r, column=12).value = f'=IFERROR(COUNTIFS({date_base}{cf_c}{mercos_f})/COUNTIFS({date_base}{cf_c}),0)'
        ws.cell(row=r, column=12).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=12))
        # Não atende
        na_f = f",'DRAFT 2'!$F$3:$F$5000,\"NÃO ATENDE\""
        nr_f = f",'DRAFT 2'!$F$3:$F$5000,\"NÃO RESPONDE\""
        rec_f = f",'DRAFT 2'!$F$3:$F$5000,\"RECUSOU LIGAÇÃO\""
        ws.cell(row=r, column=13).value = f'=COUNTIFS({date_base}{cf_c}{na_f})+COUNTIFS({date_base}{cf_c}{nr_f})+COUNTIFS({date_base}{cf_c}{rec_f})'
        style_cell(ws.cell(row=r, column=13))
        # Perda
        perda_f = f",'DRAFT 2'!$F$3:$F$5000,\"PERDA / FECHOU LOJA\""
        ws.cell(row=r, column=14).value = f'=COUNTIFS({date_base}{cf_c}{perda_f})'
        style_cell(ws.cell(row=r, column=14))

    SCT = SCD + len(CONSULTORES)
    write_header(ws, SCT, 1, "TOTAL EQUIPE", fill=FILL_DARK, align=ALIGN_LEFT)
    for c in [2, 4, 5, 7, 8, 10, 11, 13, 14]:
        cl = get_column_letter(c)
        ws.cell(row=SCT, column=c).value = f'=SUM({cl}{SCD}:{cl}{SCT-1})'
        write_header(ws, SCT, c, ws.cell(row=SCT, column=c).value, fill=FILL_DARK)
    ws.cell(row=SCT, column=3).value = f'=IFERROR(B{SCT}/22,0)'
    ws.cell(row=SCT, column=3).number_format = '0.0'
    write_header(ws, SCT, 3, ws.cell(row=SCT, column=3).value, fill=FILL_DARK)
    ws.cell(row=SCT, column=6).value = f'=IFERROR(COUNTIFS({date_base},\'DRAFT 2\'!$J$3:$J$5000,"SIM")/COUNTIFS({date_base},\'DRAFT 2\'!$I$3:$I$5000,"SIM"),0)'
    ws.cell(row=SCT, column=6).number_format = FMT_PCT
    write_header(ws, SCT, 6, ws.cell(row=SCT, column=6).value, fill=FILL_DARK)
    ws.cell(row=SCT, column=9).value = f'=IFERROR(G{SCT}/B{SCT},0)'
    ws.cell(row=SCT, column=9).number_format = FMT_PCT
    write_header(ws, SCT, 9, ws.cell(row=SCT, column=9).value, fill=FILL_DARK)
    mercos_f = f",'DRAFT 2'!$L$3:$L$5000,\"SIM\""
    ws.cell(row=SCT, column=12).value = f'=IFERROR(COUNTIFS({date_base}{mercos_f})/COUNTIFS({date_base}),0)'
    ws.cell(row=SCT, column=12).number_format = FMT_PCT
    write_header(ws, SCT, 12, ws.cell(row=SCT, column=12).value, fill=FILL_DARK)

    # ═══════════════════════════════════════════════════════════
    # FORMATTING
    # ═══════════════════════════════════════════════════════════
    ws.freeze_panes = "A5"
    ws.sheet_view.zoomScale = 90
    for r in range(1, SCT + 5):
        if r not in [kpi_row + 1]:  # skip KPI big number row
            ws.row_dimensions[r].height = 18

    total_blocks = 7
    print(f"  DASH: {total_blocks} blocos + KPI cards + chart, {SCT} rows, freeze A5")
    return ws

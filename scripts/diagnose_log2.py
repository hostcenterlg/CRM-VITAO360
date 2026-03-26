"""Diagnostica distribuicao de todas as colunas relevantes."""
import re
import openpyxl
from datetime import datetime, date

XLSX_PATH = r"C:\Users\User\OneDrive\Área de Trabalho\CRM_VITAO360  INTELIGENTE   FINAL OK .xlsx"

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)

# Analisa LARISSA - primeiras 50 rows com dados
print("=== LARISSA: distribuicao por coluna (primeiras 50 rows) ===")
ws = wb['LARISSA']
col_not_none = {}

count = 0
for row_vals in ws.iter_rows(min_row=2, max_row=200, values_only=True):
    if all(v is None for v in row_vals):
        continue
    if row_vals[0] is None:  # sem data
        continue
    count += 1
    if count > 50:
        break
    for i, v in enumerate(row_vals):
        if v is not None:
            col_not_none[i] = col_not_none.get(i, 0) + 1

print(f"Rows analisadas: {count}")
print("Colunas com dados (idx: count):")
for idx in sorted(col_not_none.keys()):
    col_letter = chr(65+idx) if idx < 26 else str(idx+1)
    print(f"  idx {idx:2d} ({col_letter}): {col_not_none[idx]:3d} nao-nulos")

# Mostrar row completa com mais dados
print("\n=== Row 2 completa (todos os 40 campos) ===")
ws2 = wb['LARISSA']
for row_vals in ws2.iter_rows(min_row=2, max_row=2, values_only=True):
    for i, v in enumerate(row_vals):
        col_letter = chr(65+i) if i < 26 else str(i+1)
        print(f"  [{i:2d}] {col_letter}: {repr(v)}")

# Verificar se alguma row tem RESULTADO preenchido
print("\n=== Procurando rows COM RESULTADO preenchido ===")
ws3 = wb['LARISSA']
found = 0
for row_vals in ws3.iter_rows(min_row=2, values_only=True):
    resultado = row_vals[20] if len(row_vals) > 20 else None
    if resultado is not None:
        print(f"  Encontrado: resultado={repr(resultado)} | row={row_vals[:21]}")
        found += 1
        if found >= 5:
            break
if found == 0:
    print("  NENHUMA row com RESULTADO preenchido encontrada!")

wb.close()

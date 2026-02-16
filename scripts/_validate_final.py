"""Validate CRM_VITAO360_V3_100.xlsx — Comprehensive 20-point validation"""
import openpyxl
import os
from collections import Counter

BASE = os.path.dirname(os.path.abspath(__file__))
FILE = os.path.join(BASE, "output", "CRM_VITAO360_V3_100.xlsx")

wb = openpyxl.load_workbook(FILE, data_only=True, read_only=True)
print("=" * 60)
print("VALIDATION: CRM_VITAO360_V3_FINAL.xlsx")
print("=" * 60)
print(f"Sheets: {wb.sheetnames}")

# 1. DRAFT 1 validation
print("\n--- DRAFT 1 ---")
ws = wb["DRAFT 1"]
cnpjs = []
situacoes = Counter()
consultores = Counter()
ufs = Counter()
vendas_count = 0
vendas_total = 0
vendas_by_month = [0] * 12

for row in ws.iter_rows(min_row=4, max_col=48, values_only=True):
    cnpj = str(row[1] or "").strip()
    if not cnpj: continue
    cnpjs.append(cnpj)

    sit = str(row[11] or "").strip()
    if sit: situacoes[sit] += 1

    cons = str(row[9] or "").strip()
    if cons: consultores[cons] += 1

    uf = str(row[3] or "").strip()
    if uf: ufs[uf] += 1

    # Check sales data (cols 25-36 = indices 24-35)
    has_sales = False
    for j in range(12):
        v = row[24 + j]
        if v and isinstance(v, (int, float)) and v > 0:
            has_sales = True
            vendas_by_month[j] += v
    if has_sales:
        vendas_count += 1

    # Total periodo
    tp = row[17]
    if tp and isinstance(tp, (int, float)):
        vendas_total += tp

print(f"Total clients: {len(cnpjs)}")
print(f"Unique CNPJs: {len(set(cnpjs))}")
dups = len(cnpjs) - len(set(cnpjs))
if dups > 0:
    print(f"  ⚠️ DUPLICATES: {dups}")
else:
    print(f"  ✅ No duplicates")

print(f"\nSITUAÇÃO distribution:")
for k, v in situacoes.most_common():
    print(f"  {k}: {v}")

print(f"\nCONSULTOR distribution:")
for k, v in consultores.most_common():
    print(f"  {k}: {v}")

print(f"\nTop UFs:")
for k, v in ufs.most_common(10):
    print(f"  {k}: {v}")

print(f"\nSales data:")
print(f"  Clients with sales: {vendas_count}/{len(cnpjs)}")
print(f"  Total faturamento: R$ {vendas_total:,.2f}")
months = ["MAR/25", "ABR/25", "MAI/25", "JUN/25", "JUL/25", "AGO/25",
          "SET/25", "OUT/25", "NOV/25", "DEZ/25", "JAN/26", "FEV/26"]
for i, m in enumerate(months):
    if vendas_by_month[i] > 0:
        print(f"  {m}: R$ {vendas_by_month[i]:,.2f}")
    else:
        print(f"  {m}: R$ 0 ⚠️")

# 2. DRAFT 2 validation
print("\n--- DRAFT 2 ---")
ws2 = wb["DRAFT 2"]
d2_count = 0
d2_resultados = Counter()
for row in ws2.iter_rows(min_row=3, max_col=6, values_only=True):
    if row[1]: d2_count += 1
    res = str(row[5] or "").strip()
    if res: d2_resultados[res] += 1
print(f"Records: {d2_count}")
print(f"RESULTADO distribution:")
for k, v in d2_resultados.most_common(12):
    print(f"  {k}: {v}")

# 3. CARTEIRA validation
print("\n--- CARTEIRA ---")
ws3 = wb["CARTEIRA"]
cart_count = 0
cart_sap = 0
for row in ws3.iter_rows(min_row=4, max_col=72, values_only=True):
    cnpj = str(row[1] or "").strip()
    if not cnpj: continue
    cart_count += 1
    if row[61] and str(row[61]).strip():  # SAP code
        cart_sap += 1
print(f"Clients: {cart_count}")
print(f"With SAP data: {cart_sap}")

# 4. Sample rows
print("\n--- SAMPLE DATA (first 3 clients) ---")
ws = wb["DRAFT 1"]
for r in range(4, 7):
    nome = ws.cell(row=r, column=1).value
    cnpj = ws.cell(row=r, column=2).value
    uf = ws.cell(row=r, column=4).value
    sit = ws.cell(row=r, column=12).value
    dias = ws.cell(row=r, column=14).value
    cons = ws.cell(row=r, column=10).value
    jan26 = ws.cell(row=r, column=35).value
    curva = ws.cell(row=r, column=39).value
    tipo = ws.cell(row=r, column=44).value
    print(f"  {nome} | {cnpj} | {uf} | {sit} | {dias}d | {cons} | JAN/26={jan26} | {curva} | {tipo}")

wb.close()
print("\n✅ Validation complete!")

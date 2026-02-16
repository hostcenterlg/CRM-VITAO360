"""Compare downloaded V3 vs our V3 — deep structural comparison."""
import openpyxl
from openpyxl.utils import get_column_letter

OURS = r"C:\Users\User\OneDrive\Área de Trabalho\CLAUDE CODE\output\CRM_VITAO360_V3.xlsx"
THEIRS = r"C:\Users\User\Downloads\CRM_INTELIGENTE_VITAO360_V3 (1).xlsx"

wb_o = openpyxl.load_workbook(OURS, data_only=True)
wb_t = openpyxl.load_workbook(THEIRS, data_only=True)

print("=" * 70)
print("  COMPARISON: OURS vs DOWNLOADED")
print("=" * 70)

# 1. Tab comparison
print("\n[TABS]")
print(f"  Ours:   {wb_o.sheetnames}")
print(f"  Theirs: {wb_t.sheetnames}")

# 2. Per-tab comparison
for name in wb_t.sheetnames:
    if name not in wb_o.sheetnames:
        print(f"\n  ** {name} — NOT in ours")
        continue
    ws_o = wb_o[name]
    ws_t = wb_t[name]
    print(f"\n{'='*70}")
    print(f"  {name}")
    print(f"  Ours: {ws_o.max_row}r x {ws_o.max_column}c  |  Theirs: {ws_t.max_row}r x {ws_t.max_column}c")

# 3. DRAFT 2 column comparison (critical)
print(f"\n{'='*70}")
print("  DRAFT 2 — Column Order Comparison")
print(f"{'='*70}")
ws_o_d2 = wb_o["DRAFT 2"]
ws_t_d2 = wb_t["DRAFT 2"]
print("  Col | OURS                    | THEIRS")
print("  " + "-" * 60)
for c in range(1, 25):
    o_val = ws_o_d2.cell(row=2, column=c).value or ""
    t_val = ws_t_d2.cell(row=2, column=c).value or ""
    match = "✓" if str(o_val).strip() == str(t_val).strip() else "✗"
    print(f"  {get_column_letter(c):3} | {str(o_val):23} | {str(t_val):23} {match}")

# 4. DASH structure comparison
print(f"\n{'='*70}")
print("  DASH — Structure Comparison")
print(f"{'='*70}")
ws_o_dash = wb_o["DASH"]
ws_t_dash = wb_t["DASH"]
print("  Their DASH rows with values in col A:")
for r in range(1, min(170, ws_t_dash.max_row + 1)):
    v = ws_t_dash.cell(row=r, column=1).value
    if v:
        print(f"    R{r}: {repr(v)[:60]}")

# 5. LOG structure comparison
print(f"\n{'='*70}")
print("  LOG — Column Comparison")
print(f"{'='*70}")
ws_t_log = wb_t["LOG"]
print("  Their LOG headers (row 1):")
for c in range(1, 25):
    v = ws_t_log.cell(row=1, column=c).value
    if v:
        print(f"    {get_column_letter(c)}: {v}")

# 6. PROJEÇÃO comparison
print(f"\n{'='*70}")
print("  PROJEÇÃO — Column Comparison")
print(f"{'='*70}")
ws_t_p = wb_t["PROJEÇÃO"]
print("  Their PROJEÇÃO headers (row 3):")
for c in range(1, 50):
    v = ws_t_p.cell(row=3, column=c).value
    if v:
        print(f"    {get_column_letter(c)}: {v}")

# 7. CARTEIRA mega-blocks
print(f"\n{'='*70}")
print("  CARTEIRA — Mega-block Comparison")
print(f"{'='*70}")
ws_t_c = wb_t["CARTEIRA"]
print("  Their CARTEIRA row 1 (mega-blocks):")
for c in range(1, 260):
    v = ws_t_c.cell(row=1, column=c).value
    if v and isinstance(v, str) and len(v) > 2:
        print(f"    {get_column_letter(c)} (col {c}): {v}")
print("  Their CARTEIRA row 2 (sub-blocks):")
for c in range(1, 260):
    v = ws_t_c.cell(row=2, column=c).value
    if v and isinstance(v, str) and len(v) > 1:
        print(f"    {get_column_letter(c)} (col {c}): {v}")
print("  Their CARTEIRA row 3 (first 80 headers):")
for c in range(1, 80):
    v = ws_t_c.cell(row=3, column=c).value
    if v:
        print(f"    {get_column_letter(c)} (col {c}): {v}")

# 8. Named ranges comparison
print(f"\n{'='*70}")
print("  NAMED RANGES COMPARISON")
print(f"{'='*70}")
o_names = {dn.name for dn in wb_o.defined_names.values()}
t_names = {dn.name for dn in wb_t.defined_names.values()}
print(f"  Ours only: {o_names - t_names}")
print(f"  Theirs only: {t_names - o_names}")
print(f"  Both: {o_names & t_names}")

# 9. DRAFT 2 example data
print(f"\n{'='*70}")
print("  DRAFT 2 — Their data sample (rows 3-7)")
print(f"{'='*70}")
for r in range(3, 8):
    vals = []
    for c in range(1, 25):
        v = ws_t_d2.cell(row=r, column=c).value
        if v is not None:
            vals.append(f"{get_column_letter(c)}={repr(v)[:30]}")
    if vals:
        print(f"  R{r}: {' | '.join(vals[:8])}")

wb_o.close()
wb_t.close()

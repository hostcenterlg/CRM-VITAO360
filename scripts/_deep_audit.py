#!/usr/bin/env python
# -*- coding: utf-8 -*-
# JARVIS CRM - Deep Audit of ALL Source Excel Files
import os, sys, datetime
from collections import Counter
import openpyxl

BASE = 'C:\\Users\\User\\OneDrive\\Área de Trabalho\\CLAUDE CODE'

def sep(c='=', n=100):
    print(c * n)

def header(title):
    sep()
    print('  ' + title)
    sep()

def safe_str(v, maxlen=40):
    if v is None: return ''
    s = str(v).strip()
    return s[:maxlen] + '...' if len(s) > maxlen else s

def open_wb(path):
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f'  [ERROR opening] {e}')
        return None

def file_size_mb(path):
    try: return os.path.getsize(path) / (1024*1024)
    except: return 0

def basic_sheet_info(ws, sheet_name, max_sample=3, max_cols=30):
    print(f'\n--- Sheet: {sheet_name!r} ---')
    rows = list(ws.iter_rows(max_row=1, values_only=True))
    if not rows:
        print('    (empty sheet)')
        return [], []
    headers = [safe_str(h, 50) for h in rows[0]]
    num_cols = len(headers)
    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        all_rows.append(row)
    print(f'    Rows (data): {len(all_rows)}')
    print(f'    Columns: {num_cols}')
    display = headers[:max_cols]
    if num_cols > max_cols:
        display.append(f'... +{num_cols - max_cols} more')
    print(f'    Headers: {display}')
    sample = min(max_sample, len(all_rows))
    if sample > 0:
        print(f'    First {sample} data rows:')
        for i in range(sample):
            rd = [safe_str(v, 25) for v in all_rows[i][:max_cols]]
            print(f'      Row {i+1}: {rd}')
    return headers, all_rows

def count_by_col(headers, all_rows, col_name, label=None):
    if not label: label = col_name
    cnu = col_name.upper().strip()
    idx = None
    for i, h in enumerate(headers):
        if h and h.upper().strip() == cnu:
            idx = i
            break
    if idx is None:
        for i, h in enumerate(headers):
            if h and cnu in h.upper().strip():
                idx = i
                break
    if idx is None:
        print(f'    [{label}] Column {col_name!r} not found')
        return Counter()
    counter = Counter()
    for row in all_rows:
        if idx < len(row):
            val = row[idx]
            if val is not None and str(val).strip():
                counter[str(val).strip()] += 1
            else:
                counter['(vazio)'] += 1
        else:
            counter['(vazio)'] += 1
    print(f'    [{label}] Distribution (col {headers[idx]!r}, idx={idx}):')
    for k, v in counter.most_common(25):
        print(f'      {k}: {v}')
    return counter

def count_non_empty_range(headers, all_rows, start_idx, end_idx, label='Range'):
    results = {}
    for ci in range(start_idx, min(end_idx+1, len(headers))):
        col_name = headers[ci] if ci < len(headers) else f'Col{ci}'
        count = 0
        for row in all_rows:
            if ci < len(row) and row[ci] is not None and str(row[ci]).strip():
                count += 1
        results[col_name] = count
    print(f'    [{label}] Non-empty counts (cols {start_idx}-{end_idx}):')
    for k, v in results.items():
        print(f'      {k}: {v}/{len(all_rows)}')
    return results

def find_date_range(headers, all_rows, col_name):
    cnu = col_name.upper().strip()
    idx = None
    for i, h in enumerate(headers):
        if h and cnu in h.upper().strip():
            idx = i
            break
    if idx is None: return
    dates = []
    for row in all_rows:
        if idx < len(row) and row[idx] is not None:
            val = row[idx]
            if isinstance(val, datetime.datetime):
                dates.append(val)
            elif isinstance(val, datetime.date):
                dates.append(datetime.datetime.combine(val, datetime.time()))
    if dates:
        mn = min(dates).strftime('%Y-%m-%d')
        mx = max(dates).strftime('%Y-%m-%d')
        print(f'    [Date Range] {headers[idx]!r}: {mn} to {mx} ({len(dates)} dates)')
    else:
        print(f'    [Date Range] {headers[idx]!r}: No valid dates found')

def count_unique(headers, all_rows, col_name):
    cnu = col_name.upper().strip()
    idx = None
    for i, h in enumerate(headers):
        if h and cnu in h.upper().strip():
            idx = i
            break
    if idx is None:
        print(f'    [Unique] Column {col_name!r} not found')
        return 0
    unique = set()
    for row in all_rows:
        if idx < len(row) and row[idx] is not None and str(row[idx]).strip():
            unique.add(str(row[idx]).strip())
    print(f'    [Unique] {headers[idx]!r}: {len(unique)} unique values')
    return len(unique)

# ============================================================
# FILE 1: CARTEIRA DE CLIENTES OULAR
# ============================================================
def audit_carteira_oular():
    path = os.path.join(BASE, 'CARTEIRA DE CLIENTES OULAR.xlsx')
    header(f'FILE 1: CARTEIRA DE CLIENTES OULAR ({file_size_mb(path):.1f} MB)')
    if not os.path.exists(path):
        print('  FILE NOT FOUND'); return
    wb = open_wb(path)
    if not wb: return
    print(f'  Sheets: {wb.sheetnames}')
    target = None
    for sn in wb.sheetnames:
        if 'CARTEIRA' in sn.upper():
            target = sn; break
    if not target:
        target = wb.sheetnames[0]
        print(f'  No CARTEIRA tab, using: {target!r}')
    ws = wb[target]
    headers, all_rows = basic_sheet_info(ws, target, max_sample=3, max_cols=85)
    if headers and all_rows:
        print(f'\nALL {len(headers)} COLUMN HEADERS:')
        for i, h in enumerate(headers):
            print(f'      [{i}] {h}')
        count_by_col(headers, all_rows, 'SITUACAO')
        count_by_col(headers, all_rows, 'CONSULTOR')
        count_by_col(headers, all_rows, 'UF')
        if len(headers) > 60:
            count_non_empty_range(headers, all_rows, 60, min(71, len(headers)-1), 'SAP cols 60-71')
        meta_cols = [i for i, h in enumerate(headers) if h and 'META' in h.upper()]
        if meta_cols:
            print(f'    META columns at indices: {meta_cols}')
            for mi in meta_cols:
                c = sum(1 for r in all_rows if mi < len(r) and r[mi] is not None and str(r[mi]).strip())
                print(f'      Col [{mi}] {headers[mi]!r}: {c}/{len(all_rows)} non-empty')
        else:
            print('    No META columns found')
        acomp = [i for i, h in enumerate(headers) if h and 'ACOMPANHAMENTO' in h.upper()]
        if acomp:
            print(f'    ACOMPANHAMENTO columns at indices: {acomp}')
        else:
            print('    No ACOMPANHAMENTO columns found')
    for sn in wb.sheetnames:
        if sn != target:
            try:
                ws2 = wb[sn]
                r2 = list(ws2.iter_rows(max_row=1, values_only=True))
                if r2:
                    h2 = [safe_str(h) for h in r2[0] if h]
                    dc = sum(1 for _ in ws2.iter_rows(min_row=2, values_only=True))
                    print(f'\n[Other sheet] {sn!r}: {dc} rows, headers: {h2[:15]}')
            except: pass
    wb.close()

# ============================================================
# LOG FILE AUDITOR
# ============================================================
def audit_log_file(path, file_label):
    header(f'{file_label} ({file_size_mb(path):.1f} MB)')
    if not os.path.exists(path):
        print('  FILE NOT FOUND'); return
    wb = open_wb(path)
    if not wb: return
    print(f'  Sheets: {wb.sheetnames}')
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers, all_rows = basic_sheet_info(ws, sn, max_sample=3)
        if headers and all_rows:
            print(f'\nALL {len(headers)} COLUMN HEADERS:')
            for i, h in enumerate(headers):
                print(f'      [{i}] {h}')
            count_by_col(headers, all_rows, 'RESULTADO')
            count_by_col(headers, all_rows, 'CONSULTOR')
            for dc in ['DATA', 'DATA ATENDIMENTO', 'TIMESTAMP', 'Carimbo']:
                find_date_range(headers, all_rows, dc)
            count_unique(headers, all_rows, 'CNPJ')
    wb.close()

# ============================================================
# GENERIC FILE AUDITOR
# ============================================================
def audit_generic(path, file_label, extra_counts=None):
    header(f'{file_label} ({file_size_mb(path):.1f} MB)')
    if not os.path.exists(path):
        print('  FILE NOT FOUND'); return
    wb = open_wb(path)
    if not wb: return
    print(f'  Sheets: {wb.sheetnames}')
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers, all_rows = basic_sheet_info(ws, sn, max_sample=3)
        if headers:
            print(f'\nALL {len(headers)} COLUMN HEADERS:')
            for i, h in enumerate(headers):
                print(f'      [{i}] {h}')
        if headers and all_rows and extra_counts:
            for cn in extra_counts:
                count_by_col(headers, all_rows, cn)
    wb.close()

# ============================================================
# BIG FILE AUDITOR
# ============================================================
def audit_big_file(path, file_label):
    header(f'{file_label} ({file_size_mb(path):.1f} MB)')
    if not os.path.exists(path):
        print('  FILE NOT FOUND'); return
    wb = open_wb(path)
    if not wb: return
    print(f'  Sheets: {wb.sheetnames}')
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows_iter = ws.iter_rows(values_only=True)
        first_row = next(rows_iter, None)
        if first_row:
            hdrs = [safe_str(h, 50) for h in first_row]
            print(f'\n--- Sheet: {sn!r} ---')
            print(f'    Columns: {len(hdrs)}')
            print(f'    ALL {len(hdrs)} COLUMN HEADERS:')
            for i, h in enumerate(hdrs):
                print(f'      [{i}] {h}')
            samples = []
            count = 0
            for row in rows_iter:
                count += 1
                if count <= 3: samples.append(row)
            print(f'    Rows (data): {count}')
            for i, s in enumerate(samples):
                print(f'      Row {i+1}: {[safe_str(v, 25) for v in s[:20]]}')
    wb.close()

# ============================================================
# MAIN
# ============================================================
def main():
    print("\n" + "=" * 100)
    print("  JARVIS CRM - DEEP AUDIT OF ALL SOURCE EXCEL FILES")
    print("  " + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print("=" * 100)

    audit_carteira_oular()

    audit_log_file(
        os.path.join(BASE, "preenchimento do fraft de atendimento (LOG).xlsx"),
        "FILE 2: preenchimento do fraft de atendimento (LOG)")

    audit_log_file(
        os.path.join(BASE, "preenchimento_do_draft_de_atendimento_LOG_FINAL (1).xlsx"),
        "FILE 3: LOG_FINAL (1)")

    audit_log_file(
        os.path.join(BASE, "preenchimento_do_draft_de_atendimento_LOG_FINAL.xlsx"),
        "FILE 4: LOG_FINAL")

    audit_generic(
        os.path.join(BASE, "DRAFT2_LOG_PRONTO_PARA_COLAR.xlsx"),
        "FILE 5: DRAFT2_LOG_PRONTO_PARA_COLAR",
        ["RESULTADO", "CONSULTOR", "TIPO"])

    audit_generic(
        os.path.join(BASE, "01_SAP_CONSOLIDADO.xlsx"),
        "FILE 6: 01_SAP_CONSOLIDADO",
        ["UF", "SITUACAO", "CONSULTOR"])

    audit_generic(
        os.path.join(BASE, "02_VENDAS_POSITIVACAO_MERCOS.xlsx"),
        "FILE 7: 02_VENDAS_POSITIVACAO_MERCOS",
        ["CONSULTOR", "MES", "CNPJ"])

    audit_generic(
        os.path.join(BASE, "03_ATENDIMENTOS_MERCOS.xlsx"),
        "FILE 8: 03_ATENDIMENTOS_MERCOS",
        ["CONSULTOR", "TIPO", "RESULTADO"])

    audit_generic(
        os.path.join(BASE, "04_CURVA_ABC_MERCOS.xlsx"),
        "FILE 9: 04_CURVA_ABC_MERCOS",
        ["CURVA", "ABC", "CLASSIFICACAO"])

    audit_generic(
        os.path.join(BASE, "05B_ECOMMERCE_MERCOS.xlsx"),
        "FILE 10: 05B_ECOMMERCE_MERCOS")

    audit_big_file(
        os.path.join(BASE, "06_LOG_FUNIL.xlsx"),
        "FILE 11: 06_LOG_FUNIL")

    audit_generic(
        os.path.join(BASE, "07_TICKETS_DESKRIO.xlsx"),
        "FILE 12: 07_TICKETS_DESKRIO",
        ["STATUS", "TIPO", "PRIORIDADE"])

    audit_generic(
        os.path.join(BASE, "08_CARTEIRA_MERCOS.xlsx"),
        "FILE 13: 08_CARTEIRA_MERCOS",
        ["SITUACAO", "UF", "CONSULTOR"])

    audit_generic(
        os.path.join(BASE, "Carteira detalhada de clientes fevreiro 2026.xlsx"),
        "FILE 14: Carteira detalhada fevreiro 2026",
        ["SITUACAO", "UF", "CONSULTOR"])

    audit_generic(
        os.path.join(BASE, "CARTEIRA DE CLIENTES JANEIRO 2026.xlsx"),
        "FILE 15: CARTEIRA DE CLIENTES JANEIRO 2026",
        ["SITUACAO", "UF", "CONSULTOR"])

    audit_generic(
        os.path.join(BASE, "SINALEIRO_INTERNO_CONFIAVEL.xlsx"),
        "FILE 16: SINALEIRO_INTERNO_CONFIAVEL")

    audit_generic(
        os.path.join(BASE, "SINALEIRO_REDES_VITAO.xlsx"),
        "FILE 17: SINALEIRO_REDES_VITAO")

    audit_generic(
        os.path.join(BASE, "REDES E FRANQUAS 2026.xlsx"),
        "FILE 18: REDES E FRANQUAS 2026",
        ["REDE", "FRANQUIA", "STATUS"])

    audit_generic(
        os.path.join(BASE, "PAINEL - VENDA INTERNA.xlsx"),
        "FILE 19: PAINEL - VENDA INTERNA")

    audit_generic(
        os.path.join(BASE, "CONTROLE_FUNIL_COMPLETO_FEV2026 (em andamento(.xlsx"),
        "FILE 20: CONTROLE_FUNIL_COMPLETO_FEV2026")

    header("FINAL SUMMARY: FILE-TO-CRM-TAB MAPPING")
    mapping = [
        ("CRM Tab", "Source File(s)", "Key Data"),
        ("-" * 20, "-" * 50, "-" * 30),
        ("REGRAS", "CARTEIRA OULAR + all files", "SITUACAO/UF/CONSULTOR lists"),
        ("CARTEIRA", "CARTEIRA OULAR (master)", "Client master data 81 cols"),
        ("CARTEIRA", "01_SAP_CONSOLIDADO", "SAP financial/BI data"),
        ("CARTEIRA", "08_CARTEIRA_MERCOS", "Mercos client portfolio"),
        ("CARTEIRA", "Cart. detalhada fev 2026", "Updated Feb 2026 data"),
        ("CARTEIRA", "CARTEIRA JANEIRO 2026", "Jan 2026 snapshot"),
        ("CARTEIRA", "02_VENDAS_POSITIVACAO_MERCOS", "Sales and positivacao"),
        ("CARTEIRA", "04_CURVA_ABC_MERCOS", "ABC curve classification"),
        ("CARTEIRA", "05B_ECOMMERCE_MERCOS", "E-commerce flags/data"),
        ("LOG", "preenchimento LOG files (2,3,4)", "Attendance log records"),
        ("LOG", "06_LOG_FUNIL", "Funnel log data"),
        ("DRAFT 1", "02_VENDAS + 03_ATENDIMENTOS", "Mercos quarantine"),
        ("DRAFT 2", "DRAFT2_LOG_PRONTO_PARA_COLAR", "Ready-to-paste logs"),
        ("DASH", "Computed from CARTEIRA+LOG", "Dashboard aggregations"),
        ("DASH", "SINALEIRO files", "Traffic-light data"),
        ("AGENDA", "03_ATENDIMENTOS_MERCOS", "Scheduled visits"),
        ("EXTRA", "07_TICKETS + REDES + PAINEL", "Support/networks/sales"),
    ]
    for row in mapping:
        t, s, k = row
        print(f"  {t:<15} | {s:<45} | {k}")

    print("\n" + "=" * 100)
    print("  AUDIT COMPLETE")
    print("=" * 100)


if __name__ == "__main__":
    main()


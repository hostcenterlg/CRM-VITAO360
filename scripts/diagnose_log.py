"""Diagnostica porque as rows estao sendo puladas."""
import re
import openpyxl
from datetime import datetime, date

XLSX_PATH = r"C:\Users\User\OneDrive\Área de Trabalho\CRM_VITAO360  INTELIGENTE   FINAL OK .xlsx"

def normalizar_cnpj(val):
    if val is None:
        return None
    if isinstance(val, float):
        val = str(int(val))
    else:
        val = str(val)
    digits = re.sub(r"\D", "", val).zfill(14)
    digits = digits[-14:]
    if len(digits) != 14 or digits == "00000000000000":
        return None
    return digits

def coerce_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    return None

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)

for aba in ['LARISSA']:
    ws = wb[aba]
    print(f"\n=== {aba} - primeiras 10 rows de dados ===")
    count = 0
    motivos = {"sem_data": 0, "sem_cnpj": 0, "sem_consultor": 0, "sem_resultado": 0, "vazia": 0, "ok": 0}

    for row_vals in ws.iter_rows(min_row=2, max_row=1767, values_only=True):
        if count < 10:
            print(f"Row: {row_vals[:21]}")  # primeiras 21 colunas

        if all(v is None for v in row_vals):
            motivos["vazia"] += 1
            continue

        data = coerce_date(row_vals[0])  # col A
        cnpj = normalizar_cnpj(row_vals[3])  # col D
        consultor = str(row_vals[1]).strip() if row_vals[1] else None
        resultado = str(row_vals[20]).strip() if row_vals[20] else None  # col U (0-based idx 20)

        if data is None:
            motivos["sem_data"] += 1
        elif cnpj is None:
            motivos["sem_cnpj"] += 1
        elif not consultor:
            motivos["sem_consultor"] += 1
        elif not resultado:
            motivos["sem_resultado"] += 1
        else:
            motivos["ok"] += 1

        count += 1
        if count >= 30:
            break

    print("\nMotivos de descarte (primeiras 30 rows com dados):")
    for k, v in motivos.items():
        print(f"  {k}: {v}")

# Contar distribuicao de resultado em LARISSA
print("\n=== Distribuicao RESULTADO (col U, idx 20) - primeiras 100 rows ===")
ws2 = wb['LARISSA']
resultados = {}
for row_vals in ws2.iter_rows(min_row=2, max_row=101, values_only=True):
    val = row_vals[20] if len(row_vals) > 20 else "N/A"
    key = repr(val)
    resultados[key] = resultados.get(key, 0) + 1
for k, v in sorted(resultados.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v}")

wb.close()

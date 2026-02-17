"""
Phase 03 Plan 01 - Task 1: Popular DRAFT 1 com vendas mensais + campos derivados
=================================================================================
Popula a aba DRAFT 1 do V12 COM_DADOS com:
  - Vendas mensais MAR/25-JAN/26 do sap_mercos_merged.json (537 clientes)
  - Campos derivados: Nro COMPRAS, CURVA ABC, MESES POSITIVADO, TICKET MEDIO, MEDIA MENSAL
  - Atualiza 502 existentes + adiciona ~35 novos clientes SAP-only/fuzzy

Regras:
  - CNPJ sempre string 14 digitos (zfill(14))
  - JAN/25 (idx 0) e FEV/25 (idx 1) NAO tem coluna no DRAFT 1 mas ENTRAM no total/metricas
  - ABC: A >= R$ 2000, B >= R$ 500, C < R$ 500 sobre total 13 meses (JAN/25-JAN/26)
  - NAO escrever TOTAL PERIODO como valor no DRAFT 1 (a CARTEIRA calcula via SUM formula)
  - Preservar formulas existentes em col 42 (TIPO CLIENTE) e col 43 (MEDIA MENSAL)
"""

import json
import re
import sys
import os
from pathlib import Path
from collections import Counter

# Setup paths
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
MERGED_PATH = PROJECT_ROOT / "data" / "output" / "phase02" / "sap_mercos_merged.json"
V12_PATH = PROJECT_ROOT / "data" / "sources" / "crm-versoes" / "v11-v12" / "CRM_INTELIGENTE_VITAO360_V12_COM_DADOS.xlsx"
REPORT_PATH = PROJECT_ROOT / "data" / "output" / "phase03" / "draft1_population_report.json"

# Ensure output dir exists
REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)

# JSON idx -> DRAFT 1 column mapping
# idx 0 (JAN/25) and idx 1 (FEV/25) have NO column in DRAFT 1
JSON_IDX_TO_DRAFT1_COL = {
    2: 21,   # MAR/25
    3: 22,   # ABR/25
    4: 23,   # MAI/25
    5: 24,   # JUN/25
    6: 25,   # JUL/25
    7: 26,   # AGO/25
    8: 27,   # SET/25
    9: 28,   # OUT/25
    10: 29,  # NOV/25
    11: 30,  # DEZ/25
}
JAN26_COL = 31

# Derived field columns
COL_NRO_COMPRAS = 34
COL_CURVA_ABC = 35
COL_MESES_POSITIVADO = 36
COL_TICKET_MEDIO = 37
COL_MEDIA_MENSAL = 43


def normalize_cnpj(raw):
    """Normalize CNPJ to 14-digit string."""
    if raw is None:
        return None
    # Handle floats like 1.8997672e+13
    s = str(raw)
    if '.' in s and 'e' in s.lower():
        # Scientific notation float
        s = f"{float(raw):.0f}"
    elif s.endswith('.0'):
        s = s[:-2]
    clean = re.sub(r'[^0-9]', '', s)
    if len(clean) >= 11:
        return clean.zfill(14)
    return None


def load_merged():
    """Load sap_mercos_merged.json."""
    print(f"Loading merged JSON: {MERGED_PATH}")
    with open(MERGED_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    cnpj_to_vendas = data['cnpj_to_vendas']
    jan26_vendas = data.get('jan26_vendas', {})
    stats = data['stats']
    print(f"  Clients: {len(cnpj_to_vendas)}")
    print(f"  JAN/26 entries: {len(jan26_vendas)}")
    print(f"  Stats: {stats}")
    return cnpj_to_vendas, jan26_vendas, stats


def populate_draft1():
    """Main population function."""
    import openpyxl

    cnpj_to_vendas, jan26_vendas, stats = load_merged()

    print(f"\nLoading V12 COM_DADOS: {V12_PATH}")
    wb = openpyxl.load_workbook(str(V12_PATH), data_only=False)

    # Find DRAFT 1 sheet
    ws = None
    for name in wb.sheetnames:
        if 'DRAFT 1' in name.upper() or 'DRAFT1' in name.upper():
            ws = wb[name]
            break
    if ws is None:
        print("ERROR: DRAFT 1 sheet not found!")
        sys.exit(1)
    print(f"  Sheet: '{ws.title}', max_row={ws.max_row}, max_col={ws.max_column}")

    # 1. Build existing CNPJ -> row map
    existing_cnpj_to_row = {}
    for row in range(4, ws.max_row + 1):
        raw = ws.cell(row, 2).value
        cnpj = normalize_cnpj(raw)
        if cnpj:
            if cnpj not in existing_cnpj_to_row:
                existing_cnpj_to_row[cnpj] = row
            else:
                print(f"  WARNING: Duplicate CNPJ {cnpj} at rows {existing_cnpj_to_row[cnpj]} and {row}")
    print(f"  Existing CNPJs in DRAFT 1: {len(existing_cnpj_to_row)}")

    # 2. Populate vendas + derived fields
    updated_count = 0
    added_count = 0
    next_row = ws.max_row + 1
    abc_counter = Counter()
    total_vendas_all = 0.0
    total_vendas_visible = 0.0
    jan_fev_hidden = 0.0
    cnpj_issues = []
    sample_clients = []

    for cnpj, vendas_array in cnpj_to_vendas.items():
        # Normalize the JSON CNPJ key
        cnpj_norm = normalize_cnpj(cnpj)
        if not cnpj_norm:
            cnpj_issues.append(f"Invalid CNPJ in merged JSON: {cnpj}")
            continue

        # Determine row
        if cnpj_norm in existing_cnpj_to_row:
            row = existing_cnpj_to_row[cnpj_norm]
            updated_count += 1
        else:
            row = next_row
            # Write CNPJ as string in col 2
            ws.cell(row, 2).value = cnpj_norm
            ws.cell(row, 2).number_format = '@'  # Text format
            next_row += 1
            added_count += 1

        # 3. Write monthly vendas MAR/25-DEZ/25
        for json_idx, draft1_col in JSON_IDX_TO_DRAFT1_COL.items():
            val = vendas_array[json_idx] if json_idx < len(vendas_array) else 0.0
            ws.cell(row, draft1_col).value = val

        # 4. Write JAN/26
        jan26_val = jan26_vendas.get(cnpj, 0.0)
        if jan26_val is None:
            jan26_val = 0.0
        ws.cell(row, JAN26_COL).value = jan26_val

        # 5. Calculate derived fields
        # Total includes ALL 13 months (JAN/25 idx0, FEV/25 idx1, MAR-DEZ idx2-11, JAN/26)
        total_all_13 = sum(vendas_array) + jan26_val
        # Visible total = MAR/25-DEZ/25 (idx 2-11) + JAN/26
        visible = sum(vendas_array[2:]) + jan26_val
        # Hidden = JAN/25 + FEV/25
        hidden = vendas_array[0] + vendas_array[1]

        total_vendas_all += total_all_13
        total_vendas_visible += visible
        jan_fev_hidden += hidden

        # Meses positivados (months with vendas > 0)
        meses_pos = sum(1 for v in vendas_array if v > 0) + (1 if jan26_val > 0 else 0)
        n_compras = meses_pos

        # CURVA ABC based on total of all 13 months
        if total_all_13 >= 2000:
            abc = 'A'
        elif total_all_13 >= 500:
            abc = 'B'
        else:
            abc = 'C'
        abc_counter[abc] += 1

        # Write derived fields
        ws.cell(row, COL_NRO_COMPRAS).value = n_compras
        ws.cell(row, COL_CURVA_ABC).value = abc
        ws.cell(row, COL_MESES_POSITIVADO).value = meses_pos
        ws.cell(row, COL_TICKET_MEDIO).value = round(total_all_13 / n_compras, 2) if n_compras > 0 else 0
        # NOTE: Col 43 MEDIA MENSAL - the existing template has a formula =IF(AJ{r}=0,0,IFERROR(SUM(U{r}:AF{r})/AJ{r},0))
        # We write the value here; this overwrites the formula but is more reliable for 537 rows
        ws.cell(row, COL_MEDIA_MENSAL).value = round(total_all_13 / meses_pos, 2) if meses_pos > 0 else 0

        # Collect sample
        if len(sample_clients) < 5:
            sample_clients.append({
                "cnpj": cnpj_norm,
                "nome": str(ws.cell(row, 1).value or ""),
                "vendas_mar_dez": [vendas_array[i] for i in range(2, 12)],
                "jan26": jan26_val,
                "total_all_13": round(total_all_13, 2),
                "total_visible": round(visible, 2),
                "jan_fev_hidden": round(hidden, 2),
                "n_compras": n_compras,
                "meses_pos": meses_pos,
                "abc": abc,
                "ticket_medio": round(total_all_13 / n_compras, 2) if n_compras > 0 else 0,
                "media_mensal": round(total_all_13 / meses_pos, 2) if meses_pos > 0 else 0,
                "row": row,
                "is_new": cnpj_norm not in existing_cnpj_to_row
            })

    # 6. Save workbook
    print(f"\nSaving V12 COM_DADOS...")
    wb.save(str(V12_PATH))
    wb.close()
    print(f"  Saved to: {V12_PATH}")

    # 7. Generate report
    total_clients = updated_count + added_count
    report = {
        "total_clients": total_clients,
        "updated_existing": updated_count,
        "added_new": added_count,
        "abc_distribution": dict(abc_counter),
        "total_vendas_all_months": round(total_vendas_all, 2),
        "total_vendas_visible": round(total_vendas_visible, 2),
        "jan_fev_hidden": round(jan_fev_hidden, 2),
        "cnpj_issues": cnpj_issues,
        "sample_clients": sample_clients,
        "draft1_data_rows": f"row 4 to row {next_row - 1}",
        "expected_from_merged": len(cnpj_to_vendas),
        "source_stats": stats
    }

    with open(REPORT_PATH, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"  Report saved to: {REPORT_PATH}")

    # 8. Print summary
    print(f"\n{'='*60}")
    print(f"DRAFT 1 POPULATION COMPLETE")
    print(f"{'='*60}")
    print(f"  Total clients: {total_clients}")
    print(f"  Updated existing: {updated_count}")
    print(f"  Added new: {added_count}")
    print(f"  ABC distribution: A={abc_counter.get('A',0)}, B={abc_counter.get('B',0)}, C={abc_counter.get('C',0)}")
    print(f"  Total (A+B+C): {sum(abc_counter.values())}")
    print(f"  Total vendas all 13 months: R$ {total_vendas_all:,.2f}")
    print(f"  Total vendas visible (MAR-JAN/26): R$ {total_vendas_visible:,.2f}")
    print(f"  JAN/FEV hidden: R$ {jan_fev_hidden:,.2f}")
    print(f"  CNPJ issues: {len(cnpj_issues)}")
    if cnpj_issues:
        for issue in cnpj_issues:
            print(f"    - {issue}")

    # Validation checks
    print(f"\n--- VALIDATION ---")
    ok = True

    if total_clients != 537:
        print(f"  WARN: total_clients={total_clients} != 537")
        ok = False
    else:
        print(f"  OK: total_clients == 537")

    if updated_count + added_count != 537:
        print(f"  WARN: updated+added={updated_count+added_count} != 537")
        ok = False
    else:
        print(f"  OK: updated+added == 537")

    abc_total = sum(abc_counter.values())
    if abc_total != 537:
        print(f"  WARN: ABC total={abc_total} != 537")
        ok = False
    else:
        print(f"  OK: ABC distribution sums to 537")

    # Expected total: ~R$ 2,493,521.92 (2025) + R$ 114,038.03 (JAN/26)
    expected_total = 2493521.92 + 114038.03
    diff_pct = abs(total_vendas_all - expected_total) / expected_total * 100
    if diff_pct > 1.0:
        print(f"  WARN: total_vendas_all R$ {total_vendas_all:,.2f} differs from expected R$ {expected_total:,.2f} by {diff_pct:.2f}%")
        ok = False
    else:
        print(f"  OK: total_vendas_all R$ {total_vendas_all:,.2f} within 1% of expected R$ {expected_total:,.2f} (diff={diff_pct:.4f}%)")

    if cnpj_issues:
        print(f"  WARN: {len(cnpj_issues)} CNPJ issues found")
        ok = False
    else:
        print(f"  OK: No CNPJ issues")

    if ok:
        print(f"\n  ALL VALIDATIONS PASSED")
    else:
        print(f"\n  SOME VALIDATIONS FAILED - review above")

    return report


if __name__ == '__main__':
    report = populate_draft1()

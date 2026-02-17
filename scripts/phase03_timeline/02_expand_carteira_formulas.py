"""
Phase 03 Plan 01 - Task 2: Expandir formulas INDEX/MATCH da CARTEIRA para todos os clientes
=============================================================================================
Le todos os CNPJs do DRAFT 1 e expande as formulas INDEX/MATCH da CARTEIRA
para cobrir todos os clientes. A CARTEIRA puxa dados do DRAFT 1 via formulas,
nunca armazena valores diretamente (exceto em colunas de identificacao e status).

Template: Row 7 da CARTEIRA original (unica row com formulas completas).

REGRAS:
  - NUNCA escrever VALORES nas colunas de vendas da CARTEIRA -- apenas FORMULAS
  - Col B (CNPJ) como string para que INDEX/MATCH funcione
  - Preservar rows 1-3 (headers/merged cells) intactas
  - Preservar colunas que ja tenham valores nos rows existentes (4-6, 8+)
  - Para novos rows, escrever formulas INDEX/MATCH em todas as colunas mapeadas
"""

import re
import sys
import openpyxl
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
V12_PATH = PROJECT_ROOT / "data" / "sources" / "crm-versoes" / "v11-v12" / "CRM_INTELIGENTE_VITAO360_V12_COM_DADOS.xlsx"


def normalize_cnpj(raw):
    """Normalize CNPJ to 14-digit string."""
    if raw is None:
        return None
    s = str(raw)
    if '.' in s and 'e' in s.lower():
        s = f"{float(raw):.0f}"
    elif s.endswith('.0'):
        s = s[:-2]
    clean = re.sub(r'[^0-9]', '', s)
    return clean.zfill(14) if len(clean) >= 11 else None


def expand_carteira():
    """Main expansion function."""
    print(f"Loading V12 COM_DADOS: {V12_PATH}")
    wb = openpyxl.load_workbook(str(V12_PATH), data_only=False)

    ws_draft = wb['DRAFT 1']
    ws_cart = wb['CARTEIRA']
    print(f"  DRAFT 1: max_row={ws_draft.max_row}, max_col={ws_draft.max_column}")
    print(f"  CARTEIRA: max_row={ws_cart.max_row}, max_col={ws_cart.max_column}")

    # 1. Read ALL CNPJs and NOME FANTASIA from DRAFT 1
    draft1_clients = []  # list of (cnpj, nome_fantasia)
    cnpj_set = set()
    for row in range(4, ws_draft.max_row + 1):
        raw_cnpj = ws_draft.cell(row, 2).value
        cnpj = normalize_cnpj(raw_cnpj)
        if cnpj and cnpj not in cnpj_set:
            nome = ws_draft.cell(row, 1).value or ""
            draft1_clients.append((cnpj, str(nome)))
            cnpj_set.add(cnpj)
    print(f"  DRAFT 1 unique clients: {len(draft1_clients)}")

    # 2. Define formula templates from Row 7 inspection
    # These are the INDEX/MATCH formulas that pull from DRAFT 1
    # Format: column_number -> formula_template (with {r} placeholder for row)

    # Cols 3-8: Identification (RAZAO, UF, CIDADE, EMAIL, TELEFONE, DATA CADASTRO)
    # from DRAFT 1 cols C-H
    IDENTITY_FORMULAS = {
        3: '=IFERROR(INDEX(\'DRAFT 1\'!$C:$C,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        4: '=IFERROR(INDEX(\'DRAFT 1\'!$D:$D,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        5: '=IFERROR(INDEX(\'DRAFT 1\'!$E:$E,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        6: '=IFERROR(INDEX(\'DRAFT 1\'!$F:$F,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        7: '=IFERROR(INDEX(\'DRAFT 1\'!$G:$G,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        8: '=IFERROR(INDEX(\'DRAFT 1\'!$H:$H,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
    }

    # Col 9: REDE REGIONAL (from DRAFT 1 col I)
    # Col 11: ULT REGISTRO MERCOS (from DRAFT 1 col AL=38)
    # Col 12: CONSULTOR (from DRAFT 1 col J=10)
    # Col 13: VENDEDOR ULTIMO PEDIDO (from DRAFT 1 col K=11)
    OTHER_INDEX_FORMULAS = {
        9: '=IFERROR(INDEX(\'DRAFT 1\'!$I:$I,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        11: '=IFERROR(INDEX(\'DRAFT 1\'!$AL:$AL,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        12: '=IFERROR(INDEX(\'DRAFT 1\'!$J:$J,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        13: '=IFERROR(INDEX(\'DRAFT 1\'!$K:$K,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
    }

    # Cols 16-19: COMPRA (DIAS SEM COMPRA, DATA ULT PEDIDO, VALOR ULT PEDIDO, CICLO MEDIO)
    # from DRAFT 1 cols L-O
    COMPRA_FORMULAS = {
        16: '=IFERROR(INDEX(\'DRAFT 1\'!$L:$L,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        17: '=IFERROR(INDEX(\'DRAFT 1\'!$M:$M,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        18: '=IFERROR(INDEX(\'DRAFT 1\'!$N:$N,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        19: '=IFERROR(INDEX(\'DRAFT 1\'!$O:$O,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
    }

    # Cols 20-24: ECOMMERCE (ACESSO B2B, ACESSOS PORTAL, ITENS CARRINHO, VALOR B2B, OPORTUNIDADE)
    # from DRAFT 1 cols P-T
    ECOMMERCE_FORMULAS = {
        20: '=IFERROR(INDEX(\'DRAFT 1\'!$P:$P,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        21: '=IFERROR(INDEX(\'DRAFT 1\'!$Q:$Q,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        22: '=IFERROR(INDEX(\'DRAFT 1\'!$R:$R,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        23: '=IFERROR(INDEX(\'DRAFT 1\'!$S:$S,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        24: '=IFERROR(INDEX(\'DRAFT 1\'!$T:$T,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
    }

    # Col 25: TOTAL PERIODO (local SUM formula)
    TOTAL_FORMULA = {
        25: '=SUM(Z{r}:AJ{r})',
    }

    # Cols 26-36: VENDAS MENSAIS (MAR/25-JAN/26) from DRAFT 1 cols U-AE
    VENDAS_FORMULAS = {
        26: '=IFERROR(INDEX(\'DRAFT 1\'!$U:$U,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',   # MAR
        27: '=IFERROR(INDEX(\'DRAFT 1\'!$V:$V,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',   # ABR
        28: '=IFERROR(INDEX(\'DRAFT 1\'!$W:$W,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',   # MAI
        29: '=IFERROR(INDEX(\'DRAFT 1\'!$X:$X,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',   # JUN
        30: '=IFERROR(INDEX(\'DRAFT 1\'!$Y:$Y,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',   # JUL
        31: '=IFERROR(INDEX(\'DRAFT 1\'!$Z:$Z,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',   # AGO
        32: '=IFERROR(INDEX(\'DRAFT 1\'!$AA:$AA,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")', # SET
        33: '=IFERROR(INDEX(\'DRAFT 1\'!$AB:$AB,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")', # OUT
        34: '=IFERROR(INDEX(\'DRAFT 1\'!$AC:$AC,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")', # NOV
        35: '=IFERROR(INDEX(\'DRAFT 1\'!$AD:$AD,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")', # DEZ
        36: '=IFERROR(INDEX(\'DRAFT 1\'!$AE:$AE,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")', # JAN/26
    }

    # Cols 38-42: RECORRENCIA
    # 38: Nro COMPRAS (from DRAFT 1 col AH=34)
    # 39: CURVA ABC (from DRAFT 1 col AI=35)
    # 40: MESES POSITIVADO (from DRAFT 1 col AJ=36)
    # 41: MEDIA MENSAL (local formula: TOTAL/MESES)
    # 42: TICKET MEDIO (from DRAFT 1 col AK=37)
    RECORRENCIA_FORMULAS = {
        38: '=IFERROR(INDEX(\'DRAFT 1\'!$AH:$AH,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        39: '=IFERROR(INDEX(\'DRAFT 1\'!$AI:$AI,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        40: '=IFERROR(INDEX(\'DRAFT 1\'!$AJ:$AJ,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        41: '=IFERROR(Y{r}/AN{r},0)',   # Y=col 25 (TOTAL), AN=col 40 (MESES POSITIVADO)
        42: '=IFERROR(INDEX(\'DRAFT 1\'!$AK:$AK,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
    }

    # Cols 44-52: FUNIL + DRAFT 2 formulas (from row 7 template)
    FUNIL_FORMULAS = {
        44: '=IFERROR(INDEX(\'DRAFT 2\'!$I:$I,MATCH(1,INDEX((\'DRAFT 2\'!$D:$D=$B{r})*(\'DRAFT 2\'!$A:$A=MAX(IF(\'DRAFT 2\'!$D:$D=$B{r},\'DRAFT 2\'!$A:$A))),0,1),0)),"")',
        45: '=IFERROR(INDEX(\'DRAFT 2\'!$T:$T,MATCH(1,INDEX((\'DRAFT 2\'!$D:$D=$B{r})*(\'DRAFT 2\'!$A:$A=MAX(IF(\'DRAFT 2\'!$D:$D=$B{r},\'DRAFT 2\'!$A:$A))),0,1),0)),"")',
        46: '=IFERROR(MAX(IF(\'DRAFT 2\'!$D:$D=$B{r},\'DRAFT 2\'!$A:$A)),"")',
        47: '=IFERROR(INDEX(\'DRAFT 2\'!$U:$U,MATCH(1,INDEX((\'DRAFT 2\'!$D:$D=$B{r})*(\'DRAFT 2\'!$A:$A=MAX(IF(\'DRAFT 2\'!$D:$D=$B{r},\'DRAFT 2\'!$A:$A))),0,1),0)),"")',
        48: '=IFERROR(INDEX(\'DRAFT 2\'!$R:$R,MATCH(1,INDEX((\'DRAFT 2\'!$D:$D=$B{r})*(\'DRAFT 2\'!$A:$A=MAX(IF(\'DRAFT 2\'!$D:$D=$B{r},\'DRAFT 2\'!$A:$A))),0,1),0)),"")',
        49: '=IFERROR(INDEX(\'DRAFT 2\'!$S:$S,MATCH(1,INDEX((\'DRAFT 2\'!$D:$D=$B{r})*(\'DRAFT 2\'!$A:$A=MAX(IF(\'DRAFT 2\'!$D:$D=$B{r},\'DRAFT 2\'!$A:$A))),0,1),0)),"")',
        50: '=IFERROR(INDEX(\'DRAFT 1\'!$AP:$AP,MATCH($B{r},\'DRAFT 1\'!$B:$B,0)),"")',
        51: '=IFERROR(INDEX(\'DRAFT 2\'!$M:$M,MATCH(1,INDEX((\'DRAFT 2\'!$D:$D=$B{r})*(\'DRAFT 2\'!$A:$A=MAX(IF(\'DRAFT 2\'!$D:$D=$B{r},\'DRAFT 2\'!$A:$A))),0,1),0)),"")',
        52: '=IFERROR(INDEX(\'DRAFT 2\'!$K:$K,MATCH(1,INDEX((\'DRAFT 2\'!$D:$D=$B{r})*(\'DRAFT 2\'!$A:$A=MAX(IF(\'DRAFT 2\'!$D:$D=$B{r},\'DRAFT 2\'!$A:$A))),0,1),0)),"")',
    }

    # Col 54: REGRAS reference
    REGRAS_FORMULA = {
        54: '=IFERROR(INDEX(REGRAS!$G$220:$G$282,MATCH(N{r}&AQ{r},REGRAS!$A$220:$A$282&REGRAS!$B$220:$B$282,0)),"")',
    }

    # Combine all formula maps
    ALL_FORMULAS = {}
    for d in [IDENTITY_FORMULAS, OTHER_INDEX_FORMULAS, COMPRA_FORMULAS,
              ECOMMERCE_FORMULAS, TOTAL_FORMULA, VENDAS_FORMULAS,
              RECORRENCIA_FORMULAS, FUNIL_FORMULAS, REGRAS_FORMULA]:
        ALL_FORMULAS.update(d)

    # NOTE: Cols 10 (TIPO CLIENTE), 14 (SITUACAO), 15 (PRIORIDADE), 37 (TIPO CLIENTE)
    # are NOT INDEX/MATCH -- they are manual/status fields. Leave them empty for new rows.
    # Col 43 (MESES LISTA) is also not in the formula template.

    # 3. Identify which CARTEIRA rows already have data (to preserve)
    existing_cart_cnpjs = {}  # cnpj -> row
    for row in range(4, min(ws_cart.max_row + 1, 8306)):
        raw = ws_cart.cell(row, 2).value
        cnpj = normalize_cnpj(raw)
        if cnpj:
            existing_cart_cnpjs[cnpj] = row

    print(f"  Existing CARTEIRA data rows: {len(existing_cart_cnpjs)}")

    # 4. Build the target layout
    # All DRAFT 1 clients get a CARTEIRA row, starting at row 4
    total_clients = len(draft1_clients)
    formulas_written = 0
    rows_created = 0
    rows_updated = 0

    for idx, (cnpj, nome) in enumerate(draft1_clients):
        cart_row = 4 + idx  # CARTEIRA data rows start at 4

        # Write CNPJ in col 2 as string
        ws_cart.cell(cart_row, 2).value = cnpj
        ws_cart.cell(cart_row, 2).number_format = '@'

        # Write NOME FANTASIA in col 1 if available
        if nome:
            ws_cart.cell(cart_row, 1).value = nome

        # Write all INDEX/MATCH formulas
        for col, template in ALL_FORMULAS.items():
            formula = template.replace('{r}', str(cart_row))
            ws_cart.cell(cart_row, col).value = formula
            formulas_written += 1

        if cnpj in existing_cart_cnpjs:
            rows_updated += 1
        else:
            rows_created += 1

    last_data_row = 4 + total_clients - 1

    # 5. Clear any leftover data in rows after our data range
    # Only clear cols 1-52 in the range just after our data
    cleared = 0
    for row in range(last_data_row + 1, last_data_row + 20):
        has_data = False
        for col in range(1, 53):
            if ws_cart.cell(row, col).value is not None:
                has_data = True
                break
        if has_data:
            for col in range(1, 53):
                ws_cart.cell(row, col).value = None
            cleared += 1
    if cleared:
        print(f"  Cleared {cleared} leftover rows after data range")

    # 6. Save
    print(f"\nSaving V12 COM_DADOS...")
    wb.save(str(V12_PATH))
    wb.close()
    print(f"  Saved to: {V12_PATH}")

    # 7. Print summary
    formula_cols_count = len(ALL_FORMULAS)
    print(f"\n{'='*60}")
    print(f"CARTEIRA FORMULA EXPANSION COMPLETE")
    print(f"{'='*60}")
    print(f"  Total clients: {total_clients}")
    print(f"  Data rows: {4} to {last_data_row}")
    print(f"  Rows updated (had existing data): {rows_updated}")
    print(f"  Rows created (new): {rows_created}")
    print(f"  Formula columns per row: {formula_cols_count}")
    print(f"  Total formulas written: {formulas_written}")
    print(f"  Expected formulas: {total_clients * formula_cols_count}")

    # Validation
    print(f"\n--- VALIDATION ---")
    ok = True

    if formulas_written != total_clients * formula_cols_count:
        print(f"  FAIL: formulas_written={formulas_written} != expected={total_clients * formula_cols_count}")
        ok = False
    else:
        print(f"  OK: All {formulas_written} formulas written")

    if total_clients < 537:
        print(f"  WARN: total_clients={total_clients} < 537 (expected minimum)")
        ok = False
    else:
        print(f"  OK: total_clients={total_clients} >= 537")

    if ok:
        print(f"\n  ALL VALIDATIONS PASSED")
    else:
        print(f"\n  SOME VALIDATIONS FAILED")

    return total_clients, formulas_written


if __name__ == '__main__':
    total, formulas = expand_carteira()

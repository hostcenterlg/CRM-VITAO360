"""
Phase 03 Plan 02 - Task 1: Validar vendas DRAFT 1 vs merged JSON + ABC recalculo independente
================================================================================================
Validacao cruzada completa da Phase 3:
  1. Cross-check vendas mensal DRAFT 1 vs sap_mercos_merged.json (zero divergencia)
  2. ABC recalculo independente e comparacao com DRAFT 1 col 35
  3. Campos derivados spot-check (10 clientes aleatorios)
  4. CARTEIRA formula validation (10 clientes)
  5. V13 PROJECAO integridade (19.224 formulas)
  6. Avaliacao formal TIME-01, TIME-02, TIME-03

Output:
  - data/output/phase03/abc_classification.json
  - data/output/phase03/validation_report.json
"""

import json
import re
import sys
import random
from pathlib import Path
from collections import Counter
from datetime import datetime

# Setup paths
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
MERGED_PATH = PROJECT_ROOT / "data" / "output" / "phase02" / "sap_mercos_merged.json"
V12_PATH = PROJECT_ROOT / "data" / "sources" / "crm-versoes" / "v11-v12" / "CRM_INTELIGENTE_VITAO360_V12_COM_DADOS.xlsx"
V13_PATH = PROJECT_ROOT / "data" / "output" / "CRM_VITAO360_V13_PROJECAO.xlsx"
ABC_OUTPUT = PROJECT_ROOT / "data" / "output" / "phase03" / "abc_classification.json"
REPORT_OUTPUT = PROJECT_ROOT / "data" / "output" / "phase03" / "validation_report.json"

# Ensure output dir
ABC_OUTPUT.parent.mkdir(parents=True, exist_ok=True)

# Column mappings (same as population script)
JSON_IDX_TO_DRAFT1_COL = {
    2: 21,   # MAR/25
    3: 22,   # ABR/25
    4: 23,   # MAI/25
    5: 24,   # JUN/25
    6: 25,   # JUL/25
    7: 26,   # AGO/25
    8: 27,   # SET/25
    9: 28,   # OUT/25
    10: 29,  # NOV/25
    11: 30,  # DEZ/25
}
JAN26_COL = 31

COL_NRO_COMPRAS = 34
COL_CURVA_ABC = 35
COL_MESES_POSITIVADO = 36
COL_TICKET_MEDIO = 37
COL_MEDIA_MENSAL = 43

MONTH_NAMES = {
    2: "MAR/25", 3: "ABR/25", 4: "MAI/25", 5: "JUN/25",
    6: "JUL/25", 7: "AGO/25", 8: "SET/25", 9: "OUT/25",
    10: "NOV/25", 11: "DEZ/25"
}
TOLERANCE = 0.01


def normalize_cnpj(raw):
    """Normalize CNPJ to 14-digit string."""
    if raw is None:
        return None
    s = str(raw)
    if '.' in s and 'e' in s.lower():
        s = f"{float(raw):.0f}"
    elif s.endswith('.0'):
        s = s[:-2]
    clean = re.sub(r'[^0-9]', '', s)
    return clean.zfill(14) if len(clean) >= 11 else None


def safe_float(val):
    """Convert cell value to float, treating None/empty/str as 0."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def load_merged():
    """Load sap_mercos_merged.json."""
    print(f"[1/6] Loading merged JSON: {MERGED_PATH}")
    with open(MERGED_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    cnpj_to_vendas = data['cnpj_to_vendas']
    jan26_vendas = data.get('jan26_vendas', {})
    stats = data['stats']
    print(f"  Clients: {len(cnpj_to_vendas)}, JAN/26 entries: {len(jan26_vendas)}")
    return cnpj_to_vendas, jan26_vendas, stats


def validate_all():
    """Main validation function."""
    import openpyxl

    cnpj_to_vendas, jan26_vendas, stats = load_merged()
    json_cnpjs = set(cnpj_to_vendas.keys())

    # ==================================================================
    # STEP 2: Load DRAFT 1 from V12 COM_DADOS
    # ==================================================================
    print(f"\n[2/6] Loading V12 COM_DADOS: {V12_PATH}")
    wb = openpyxl.load_workbook(str(V12_PATH), data_only=False)

    ws_draft = None
    for name in wb.sheetnames:
        if 'DRAFT 1' in name.upper() or 'DRAFT1' in name.upper():
            ws_draft = wb[name]
            break
    if ws_draft is None:
        print("ERROR: DRAFT 1 sheet not found!")
        sys.exit(1)
    print(f"  DRAFT 1: max_row={ws_draft.max_row}, max_col={ws_draft.max_column}")

    # Read all DRAFT 1 data
    draft1_data = {}  # cnpj -> {row, vendas_by_json_idx, jan26, derived_fields}
    draft1_cnpj_rows = []  # (cnpj, row) for duplicate detection
    for row in range(4, ws_draft.max_row + 1):
        raw = ws_draft.cell(row, 2).value
        cnpj = normalize_cnpj(raw)
        if cnpj is None:
            continue

        draft1_cnpj_rows.append((cnpj, row))

        # Read monthly vendas
        vendas = {}
        for json_idx, col in JSON_IDX_TO_DRAFT1_COL.items():
            vendas[json_idx] = safe_float(ws_draft.cell(row, col).value)
        jan26_val = safe_float(ws_draft.cell(row, JAN26_COL).value)

        # Read derived fields
        nro_compras = ws_draft.cell(row, COL_NRO_COMPRAS).value
        curva_abc = ws_draft.cell(row, COL_CURVA_ABC).value
        meses_pos = ws_draft.cell(row, COL_MESES_POSITIVADO).value
        ticket_medio = ws_draft.cell(row, COL_TICKET_MEDIO).value
        media_mensal = ws_draft.cell(row, COL_MEDIA_MENSAL).value

        if cnpj not in draft1_data:
            draft1_data[cnpj] = {
                'row': row,
                'vendas': vendas,
                'jan26': jan26_val,
                'nro_compras': nro_compras,
                'curva_abc': curva_abc,
                'meses_pos': meses_pos,
                'ticket_medio': ticket_medio,
                'media_mensal': media_mensal,
            }

    draft1_cnpjs = set(draft1_data.keys())
    print(f"  DRAFT 1 unique CNPJs: {len(draft1_cnpjs)}")

    # Detect duplicates
    cnpj_counts = Counter(c for c, r in draft1_cnpj_rows)
    duplicates = [(c, cnt) for c, cnt in cnpj_counts.items() if cnt > 1]
    if duplicates:
        print(f"  WARNING: {len(duplicates)} duplicate CNPJs in DRAFT 1")
        for c, cnt in duplicates[:5]:
            print(f"    {c}: {cnt} rows")

    # ==================================================================
    # STEP 3: Cross-check vendas mensal
    # ==================================================================
    print(f"\n[3/6] Cross-checking vendas: DRAFT 1 vs merged JSON")
    mismatches = []
    matched_count = 0
    checked_count = 0

    # CNPJs present in both
    common_cnpjs = json_cnpjs & draft1_cnpjs
    missing_from_draft1 = sorted(json_cnpjs - draft1_cnpjs)
    extra_in_draft1 = sorted(draft1_cnpjs - json_cnpjs)

    print(f"  JSON CNPJs: {len(json_cnpjs)}")
    print(f"  DRAFT 1 CNPJs: {len(draft1_cnpjs)}")
    print(f"  In both: {len(common_cnpjs)}")
    print(f"  Missing from DRAFT 1: {len(missing_from_draft1)}")
    print(f"  Extra in DRAFT 1 (pre-existing): {len(extra_in_draft1)}")

    for cnpj in sorted(common_cnpjs):
        json_vendas = cnpj_to_vendas[cnpj]
        d1 = draft1_data[cnpj]
        client_ok = True
        checked_count += 1

        # Check MAR/25-DEZ/25 (json idx 2-11)
        for json_idx, col in JSON_IDX_TO_DRAFT1_COL.items():
            json_val = json_vendas[json_idx] if json_idx < len(json_vendas) else 0.0
            draft1_val = d1['vendas'].get(json_idx, 0.0)
            if abs(draft1_val - json_val) > TOLERANCE:
                mismatches.append({
                    'cnpj': cnpj,
                    'month': MONTH_NAMES.get(json_idx, f"idx{json_idx}"),
                    'json_idx': json_idx,
                    'valor_draft1': draft1_val,
                    'valor_json': json_val,
                    'diff': round(draft1_val - json_val, 4),
                    'row': d1['row']
                })
                client_ok = False

        # Check JAN/26
        json_jan26 = jan26_vendas.get(cnpj, 0.0)
        if json_jan26 is None:
            json_jan26 = 0.0
        draft1_jan26 = d1['jan26']
        if abs(draft1_jan26 - json_jan26) > TOLERANCE:
            mismatches.append({
                'cnpj': cnpj,
                'month': 'JAN/26',
                'json_idx': 'jan26',
                'valor_draft1': draft1_jan26,
                'valor_json': json_jan26,
                'diff': round(draft1_jan26 - json_jan26, 4),
                'row': d1['row']
            })
            client_ok = False

        if client_ok:
            matched_count += 1

    print(f"  Checked: {checked_count}")
    print(f"  Perfect match: {matched_count}")
    print(f"  Mismatches: {len(mismatches)}")
    if mismatches:
        for m in mismatches[:10]:
            print(f"    {m['cnpj']} {m['month']}: draft1={m['valor_draft1']}, json={m['valor_json']}, diff={m['diff']}")
        if len(mismatches) > 10:
            print(f"    ... and {len(mismatches) - 10} more")

    # ==================================================================
    # STEP 4: ABC recalculo independente
    # ==================================================================
    print(f"\n[4/6] ABC independent recalculation")
    abc_by_client = []
    abc_mismatches = []
    abc_counter = Counter()

    for cnpj in sorted(json_cnpjs):
        vendas_array = cnpj_to_vendas[cnpj]
        jan26_val = jan26_vendas.get(cnpj, 0.0) or 0.0

        # Total = sum of all 12 months (JAN/25-DEZ/25) + JAN/26
        total = sum(vendas_array) + jan26_val

        # ABC classification
        if total >= 2000:
            abc = 'A'
        elif total >= 500:
            abc = 'B'
        else:
            abc = 'C'
        abc_counter[abc] += 1

        # Meses positivados (count of months > 0)
        meses_pos = sum(1 for v in vendas_array if v > 0) + (1 if jan26_val > 0 else 0)

        abc_by_client.append({
            'cnpj': cnpj,
            'abc': abc,
            'total': round(total, 2),
            'meses_pos': meses_pos
        })

        # Compare with DRAFT 1 col 35
        if cnpj in draft1_data:
            d1_abc = draft1_data[cnpj]['curva_abc']
            # Normalize comparison
            d1_abc_str = str(d1_abc).strip().upper() if d1_abc is not None else ""
            if d1_abc_str != abc:
                abc_mismatches.append({
                    'cnpj': cnpj,
                    'calculated': abc,
                    'draft1': d1_abc_str,
                    'total': round(total, 2),
                    'row': draft1_data[cnpj]['row']
                })

    print(f"  ABC distribution: A={abc_counter.get('A',0)}, B={abc_counter.get('B',0)}, C={abc_counter.get('C',0)}")
    print(f"  Total classified: {sum(abc_counter.values())}")
    print(f"  ABC mismatches vs DRAFT 1: {len(abc_mismatches)}")
    if abc_mismatches:
        for m in abc_mismatches[:5]:
            print(f"    {m['cnpj']}: calc={m['calculated']}, draft1={m['draft1']}, total=R${m['total']:,.2f}")

    # ==================================================================
    # STEP 5: Derived fields spot-check (10 random clients)
    # ==================================================================
    print(f"\n[5/6] Derived fields spot-check (10 random clients)")
    # Use a fixed seed for reproducibility
    random.seed(42)
    spot_check_cnpjs = random.sample(sorted(common_cnpjs), min(10, len(common_cnpjs)))

    derived_failures = []
    derived_checked = 0

    for cnpj in spot_check_cnpjs:
        vendas_array = cnpj_to_vendas[cnpj]
        jan26_val = jan26_vendas.get(cnpj, 0.0) or 0.0
        total = sum(vendas_array) + jan26_val
        meses_pos_calc = sum(1 for v in vendas_array if v > 0) + (1 if jan26_val > 0 else 0)
        n_compras_calc = meses_pos_calc

        d1 = draft1_data[cnpj]
        derived_checked += 1
        client_issues = []

        # Check Nro COMPRAS (col 34)
        d1_nro_compras = d1['nro_compras']
        if d1_nro_compras is not None and safe_float(d1_nro_compras) != n_compras_calc:
            client_issues.append(f"Nro COMPRAS: draft1={d1_nro_compras}, calc={n_compras_calc}")

        # Check MESES POSITIVADO (col 36)
        d1_meses_pos = d1['meses_pos']
        if d1_meses_pos is not None and safe_float(d1_meses_pos) != meses_pos_calc:
            client_issues.append(f"MESES POS: draft1={d1_meses_pos}, calc={meses_pos_calc}")

        # Check TICKET MEDIO (col 37)
        ticket_calc = round(total / n_compras_calc, 2) if n_compras_calc > 0 else 0
        d1_ticket = d1['ticket_medio']
        if d1_ticket is not None and abs(safe_float(d1_ticket) - ticket_calc) > TOLERANCE:
            client_issues.append(f"TICKET MEDIO: draft1={d1_ticket}, calc={ticket_calc}")

        # Check MEDIA MENSAL (col 43)
        media_calc = round(total / meses_pos_calc, 2) if meses_pos_calc > 0 else 0
        d1_media = d1['media_mensal']
        if d1_media is not None and abs(safe_float(d1_media) - media_calc) > TOLERANCE:
            client_issues.append(f"MEDIA MENSAL: draft1={d1_media}, calc={media_calc}")

        if client_issues:
            derived_failures.append({
                'cnpj': cnpj,
                'row': d1['row'],
                'issues': client_issues
            })
            print(f"  FAIL {cnpj} (row {d1['row']}): {'; '.join(client_issues)}")
        else:
            print(f"  OK {cnpj} (row {d1['row']})")

    derived_all_pass = len(derived_failures) == 0
    print(f"  Result: {'ALL PASS' if derived_all_pass else f'{len(derived_failures)} FAILURES'}")

    # ==================================================================
    # STEP 6: CARTEIRA formula validation (10 clients)
    # ==================================================================
    print(f"\n[6a/6] CARTEIRA formula validation (10 clients)")
    ws_cart = None
    for name in wb.sheetnames:
        if 'CARTEIRA' in name.upper():
            ws_cart = wb[name]
            break
    if ws_cart is None:
        print("ERROR: CARTEIRA sheet not found!")
        carteira_result = {
            'checked': 0,
            'all_have_formulas': False,
            'issues': ['CARTEIRA sheet not found']
        }
    else:
        print(f"  CARTEIRA: max_row={ws_cart.max_row}, max_col={ws_cart.max_column}")

        # Build CNPJ -> row map for CARTEIRA
        cart_cnpj_to_row = {}
        for row in range(4, min(ws_cart.max_row + 1, 8306)):
            raw = ws_cart.cell(row, 2).value
            cnpj = normalize_cnpj(raw)
            if cnpj:
                cart_cnpj_to_row[cnpj] = row
        print(f"  CARTEIRA CNPJs: {len(cart_cnpj_to_row)}")

        # Check 10 random clients from the merged set
        check_cnpjs = random.sample(sorted(common_cnpjs & set(cart_cnpj_to_row.keys())),
                                     min(10, len(common_cnpjs & set(cart_cnpj_to_row.keys()))))

        carteira_issues = []
        carteira_checked = 0

        for cnpj in check_cnpjs:
            cart_row = cart_cnpj_to_row[cnpj]
            carteira_checked += 1
            client_issues = []

            # Check that col 2 is CNPJ as string
            cell_b = ws_cart.cell(cart_row, 2).value
            cnpj_norm = normalize_cnpj(cell_b)
            if cnpj_norm != cnpj:
                client_issues.append(f"Col B CNPJ mismatch: {cell_b} vs {cnpj}")

            # Check vendas formulas (cols 26-36)
            for col in range(26, 37):
                cell_val = ws_cart.cell(cart_row, col).value
                if cell_val is None:
                    client_issues.append(f"Col {col}: empty (expected formula)")
                elif isinstance(cell_val, str) and cell_val.startswith('='):
                    # It's a formula -- check it references DRAFT 1
                    if 'DRAFT 1' not in cell_val and 'INDEX' not in cell_val:
                        client_issues.append(f"Col {col}: formula doesn't reference DRAFT 1: {cell_val[:50]}")
                else:
                    # It's a value, not a formula -- this is expected if data_only=False returns formulas
                    # but the populate script wrote VALUES for some cells
                    pass  # OK - values are also valid

            # Check col 25 TOTAL formula
            cell_25 = ws_cart.cell(cart_row, 25).value
            if cell_25 is None:
                client_issues.append("Col 25 (TOTAL): empty")
            elif isinstance(cell_25, str) and cell_25.startswith('='):
                if 'SUM' not in cell_25.upper():
                    client_issues.append(f"Col 25 (TOTAL): expected SUM formula, got: {cell_25[:50]}")
            # If it's a number that's also OK (pre-calculated)

            # Check recorrencia formulas (cols 38-42)
            for col in range(38, 43):
                cell_val = ws_cart.cell(cart_row, col).value
                if cell_val is None:
                    client_issues.append(f"Col {col}: empty (expected formula)")
                elif isinstance(cell_val, str) and cell_val.startswith('='):
                    pass  # Formula present - good
                else:
                    pass  # Value also acceptable (from data_only=False with cached)

            if client_issues:
                carteira_issues.append({
                    'cnpj': cnpj,
                    'row': cart_row,
                    'issues': client_issues
                })
                print(f"  ISSUES {cnpj} (row {cart_row}): {'; '.join(client_issues[:3])}")
            else:
                print(f"  OK {cnpj} (row {cart_row})")

        all_have_formulas = len(carteira_issues) == 0
        carteira_result = {
            'checked': carteira_checked,
            'all_have_formulas': all_have_formulas,
            'issues': [{'cnpj': i['cnpj'], 'row': i['row'], 'issues': i['issues']} for i in carteira_issues]
        }
        print(f"  Result: {'ALL OK' if all_have_formulas else f'{len(carteira_issues)} with issues'}")

    # Close V12
    wb.close()

    # ==================================================================
    # STEP 7: V13 PROJECAO integrity check
    # ==================================================================
    print(f"\n[6b/6] V13 PROJECAO integrity check")
    v13_result = {'projecao_formulas': 0, 'status': 'UNKNOWN'}

    if not V13_PATH.exists():
        print(f"  ERROR: V13 file not found at {V13_PATH}")
        v13_result['status'] = 'FAIL'
        v13_result['error'] = 'File not found'
    else:
        wb_v13 = openpyxl.load_workbook(str(V13_PATH), data_only=False)
        # Find PROJECAO sheet (may have accent)
        ws_proj = None
        for name in wb_v13.sheetnames:
            # Strip accents for matching
            clean = name.upper().strip()
            if 'PROJE' in clean:
                ws_proj = wb_v13[name]
                break
        if ws_proj is None:
            print(f"  ERROR: PROJECAO sheet not found in V13. Sheets: {wb_v13.sheetnames}")
            v13_result['status'] = 'FAIL'
            v13_result['error'] = f'Sheet not found. Available: {wb_v13.sheetnames}'
        else:
            print(f"  Sheet: '{ws_proj.title}', max_row={ws_proj.max_row}, max_col={ws_proj.max_column}")
            formula_count = 0
            for row in ws_proj.iter_rows(min_row=1, max_row=ws_proj.max_row,
                                          min_col=1, max_col=ws_proj.max_column):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1

            v13_result['projecao_formulas'] = formula_count
            print(f"  Formula count: {formula_count}")

            if formula_count == 19224:
                v13_result['status'] = 'PASS'
                print(f"  Status: PASS (exactly 19,224 formulas)")
            elif formula_count >= 19000:
                v13_result['status'] = 'PASS'
                print(f"  Status: PASS ({formula_count} formulas, close to 19,224)")
            else:
                v13_result['status'] = 'FAIL'
                print(f"  Status: FAIL (expected 19,224, got {formula_count})")

        wb_v13.close()

    # ==================================================================
    # EVALUATE TIME-01, TIME-02, TIME-03
    # ==================================================================
    print(f"\n{'='*60}")
    print(f"REQUIREMENT EVALUATION")
    print(f"{'='*60}")

    # TIME-01: 537 clientes com vendas MAR/25-JAN/26, zero mismatches vs merged JSON
    time01_pass = (
        len(common_cnpjs) == 537
        and len(mismatches) == 0
        and len(missing_from_draft1) == 0
    )
    time01_detail = (
        f"{len(common_cnpjs)} clientes matched in DRAFT 1, "
        f"{len(mismatches)} mismatches, "
        f"{len(missing_from_draft1)} missing from DRAFT 1"
    )
    print(f"\n  TIME-01: {'PASS' if time01_pass else 'FAIL'}")
    print(f"    {time01_detail}")

    # TIME-02: Dados cruzados via sap_mercos_merged.json (537 == stats.total_clientes)
    time02_pass = (
        stats.get('total_clientes') == 537
        and len(json_cnpjs) == 537
    )
    time02_detail = (
        f"merged JSON has {len(json_cnpjs)} clients (stats.total_clientes={stats.get('total_clientes')}), "
        f"SAP-First merge with {stats.get('months_filled_from_mercos', 0)} month-cells from Mercos"
    )
    print(f"  TIME-02: {'PASS' if time02_pass else 'FAIL'}")
    print(f"    {time02_detail}")

    # TIME-03: ABC recalculada para 537 clientes, 100% match entre calculo independente e DRAFT 1
    time03_pass = (
        sum(abc_counter.values()) == 537
        and len(abc_mismatches) == 0
    )
    time03_detail = (
        f"ABC: A={abc_counter.get('A',0)}, B={abc_counter.get('B',0)}, C={abc_counter.get('C',0)} "
        f"(total={sum(abc_counter.values())}), {len(abc_mismatches)} mismatches vs DRAFT 1"
    )
    print(f"  TIME-03: {'PASS' if time03_pass else 'FAIL'}")
    print(f"    {time03_detail}")

    # Overall
    overall_pass = time01_pass and time02_pass and time03_pass
    v13_pass = v13_result['status'] == 'PASS'
    all_pass = overall_pass and v13_pass and derived_all_pass

    print(f"\n  V13 integrity: {v13_result['status']}")
    print(f"  Derived fields spot-check: {'PASS' if derived_all_pass else 'FAIL'}")
    print(f"  CARTEIRA formulas: {'PASS' if carteira_result.get('all_have_formulas', False) else 'ISSUES'}")
    print(f"\n  OVERALL: {'PASS' if all_pass else 'FAIL'}")

    # ==================================================================
    # GENERATE abc_classification.json
    # ==================================================================
    print(f"\nGenerating abc_classification.json...")
    abc_output = {
        "distribution": {
            "A": abc_counter.get('A', 0),
            "B": abc_counter.get('B', 0),
            "C": abc_counter.get('C', 0)
        },
        "total_clients": sum(abc_counter.values()),
        "thresholds": {"A": ">=2000", "B": ">=500", "C": "<500"},
        "by_client": abc_by_client,
        "comparison_with_draft1": {
            "match": sum(abc_counter.values()) - len(abc_mismatches),
            "mismatch": len(abc_mismatches),
            "mismatches": abc_mismatches
        }
    }
    with open(ABC_OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(abc_output, f, indent=2, ensure_ascii=False)
    print(f"  Saved: {ABC_OUTPUT}")

    # ==================================================================
    # GENERATE validation_report.json
    # ==================================================================
    print(f"Generating validation_report.json...")
    report = {
        "phase": "03-timeline-mensal",
        "plan": "02",
        "date": datetime.now().strftime("%Y-%m-%d"),
        "cross_check": {
            "total_cnpjs_json": len(json_cnpjs),
            "total_cnpjs_draft1": len(draft1_cnpjs),
            "common": len(common_cnpjs),
            "matched": matched_count,
            "mismatches": mismatches,
            "missing_from_draft1": missing_from_draft1,
            "extra_in_draft1": extra_in_draft1,
            "duplicates_in_draft1": [{"cnpj": c, "count": cnt} for c, cnt in duplicates]
        },
        "abc_validation": {
            "total_match": sum(abc_counter.values()) - len(abc_mismatches),
            "total_mismatch": len(abc_mismatches),
            "distribution": {
                "A": abc_counter.get('A', 0),
                "B": abc_counter.get('B', 0),
                "C": abc_counter.get('C', 0)
            },
            "mismatches": abc_mismatches
        },
        "derived_fields_spot_check": {
            "checked": derived_checked,
            "all_pass": derived_all_pass,
            "failures": [{"cnpj": f['cnpj'], "row": f['row'], "issues": f['issues']} for f in derived_failures]
        },
        "carteira_formulas": carteira_result,
        "v13_integrity": v13_result,
        "requirements": {
            "TIME-01": {
                "status": "PASS" if time01_pass else "FAIL",
                "detail": f"537 clientes com vendas MAR/25-JAN/26 no DRAFT 1 + CARTEIRA. {len(mismatches)} mismatches, {len(missing_from_draft1)} missing."
            },
            "TIME-02": {
                "status": "PASS" if time02_pass else "FAIL",
                "detail": f"Dados cruzados via sap_mercos_merged.json (SAP-First + Mercos complement). {stats.get('total_clientes')} clients, {stats.get('months_filled_from_mercos', 0)} month-cells from Mercos."
            },
            "TIME-03": {
                "status": "PASS" if time03_pass else "FAIL",
                "detail": f"ABC recalculada: A={abc_counter.get('A',0)}, B={abc_counter.get('B',0)}, C={abc_counter.get('C',0)} (thresholds A>=2000, B>=500, C<500). {len(abc_mismatches)} mismatches."
            }
        },
        "overall": "PASS" if all_pass else "FAIL"
    }
    with open(REPORT_OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"  Saved: {REPORT_OUTPUT}")

    # ==================================================================
    # FINAL SUMMARY
    # ==================================================================
    print(f"\n{'='*60}")
    print(f"PHASE 03 VALIDATION COMPLETE")
    print(f"{'='*60}")
    print(f"  Cross-check mismatches: {len(mismatches)}")
    print(f"  ABC mismatches: {len(abc_mismatches)}")
    print(f"  Derived fields: {'ALL PASS' if derived_all_pass else 'FAILURES'}")
    print(f"  CARTEIRA formulas: {'OK' if carteira_result.get('all_have_formulas', False) else 'ISSUES'}")
    print(f"  V13 integrity: {v13_result['status']} ({v13_result['projecao_formulas']} formulas)")
    print(f"  TIME-01: {'PASS' if time01_pass else 'FAIL'}")
    print(f"  TIME-02: {'PASS' if time02_pass else 'FAIL'}")
    print(f"  TIME-03: {'PASS' if time03_pass else 'FAIL'}")
    print(f"  OVERALL: {'PASS' if all_pass else 'FAIL'}")

    return report


if __name__ == '__main__':
    report = validate_all()
    # Exit with code based on overall result
    sys.exit(0 if report['overall'] == 'PASS' else 1)

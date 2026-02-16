"""
V3 — Tab DRAFT 2: Log de Atendimentos (24 cols + motor de regras + 5 exemplos)
"""
import datetime
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from v3_styles import *
from motor_regras import motor_de_regras, FOLLOW_UP_DIAS


# V3 column order: Manual (1-12) → Auto-calculated (13-24)
DRAFT2_COLS = [
    # MANUAL (consultor preenche)
    (1,  "DATA",              14, FMT_DATE, False),
    (2,  "CNPJ",              18, FMT_TEXT, False),
    (3,  "NOME FANTASIA",     25, None,     False),
    (4,  "UF",                 5, None,     False),
    (5,  "CONSULTOR",         20, None,     False),
    (6,  "RESULTADO",         22, None,     False),
    (7,  "MOTIVO",            28, None,     False),
    (8,  "WHATSAPP",          10, None,     False),
    (9,  "LIGAÇÃO",           10, None,     False),
    (10, "LIG. ATENDIDA",     14, None,     False),
    (11, "NOTA DO DIA",       35, None,     False),
    (12, "MERCOS ATUALIZADO", 16, None,     False),
    # AUTO-CALCULADO (motor de regras)
    (13, "SITUAÇÃO",          14, None,     True),
    (14, "ESTÁGIO FUNIL",     18, None,     True),
    (15, "FASE",              16, None,     True),
    (16, "TIPO DO CONTATO",   26, None,     True),
    (17, "TEMPERATURA",       14, None,     True),
    (18, "TENTATIVA",         12, None,     True),
    (19, "GRUPO DASH",        14, None,     True),
    (20, "FOLLOW-UP",         14, FMT_DATE, True),
    (21, "AÇÃO FUTURA",       16, None,     True),
    (22, "AÇÃO DETALHADA",    30, None,     True),
    (23, "SINALEIRO CICLO",   14, None,     True),
    (24, "SINALEIRO META",    14, None,     True),
]


def build_draft2(wb):
    """Build DRAFT 2 tab (24 cols, motor formulas, 5 example records)."""
    ws = wb.create_sheet("DRAFT 2")
    ws.sheet_properties.tabColor = TAB_DRAFT2

    # ── Row 1: Title ──
    ws.merge_cells('A1:X1')
    title = ws.cell(row=1, column=1,
                    value="📋 DRAFT 2 — LOG DE ATENDIMENTOS (motor de regras automático)")
    style_cell(title, font=FONT_TITLE_W, fill=PatternFill('solid', fgColor='FFC000'), align=ALIGN_LEFT)
    title.font = Font(name='Calibri', size=11, bold=True, color='000000')
    for c in range(1, 25):
        ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor='FFC000')
        ws.cell(row=1, column=c).border = THIN_BORDER

    # ── Row 2: Headers ──
    for col, header, width, fmt, is_auto in DRAFT2_COLS:
        fill = FILL_HEADER if not is_auto else PatternFill('solid', fgColor='2E7D32')
        write_header(ws, 2, col, header, fill=fill)
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Data Validations ──
    # RESULTADO (col F)
    dv_res = DataValidation(type="list", formula1="LISTA_RESULTADO", allow_blank=True)
    dv_res.error = "Selecione um RESULTADO válido"
    dv_res.errorTitle = "RESULTADO inválido"
    ws.add_data_validation(dv_res)
    dv_res.add(f"F3:F5000")

    # MOTIVO (col G)
    dv_mot = DataValidation(type="list", formula1="LISTA_MOTIVO", allow_blank=True)
    ws.add_data_validation(dv_mot)
    dv_mot.add(f"G3:G5000")

    # CONSULTOR (col E)
    dv_cons = DataValidation(type="list", formula1="LISTA_CONSULTOR", allow_blank=True)
    ws.add_data_validation(dv_cons)
    dv_cons.add(f"E3:E5000")

    # SIM/NÃO for WHATSAPP (H), LIGAÇÃO (I), LIG.ATENDIDA (J), MERCOS (L)
    dv_sn = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
    ws.add_data_validation(dv_sn)
    for col_letter in ['H', 'I', 'J', 'L']:
        dv_sn.add(f"{col_letter}3:{col_letter}5000")

    # ── Motor Formulas (Excel) for auto-calculated columns ──
    # These formulas go in row 3 and are designed to be copied down
    R = 3  # first formula row

    # Build formulas for rows 3-7 (5 examples + pattern)
    for row in range(3, 5003):
        r = row
        # Only write formulas for rows that have data or are the first few
        # For efficiency, we'll write them for the first 100 rows
        if row > 102:
            break

        # Col M (13): SITUAÇÃO — pull from DRAFT 1 based on CNPJ
        ws.cell(row=r, column=13).value = (
            f'=IFERROR(XLOOKUP(B{r},\'DRAFT 1\'!$B$4:$B$5000,\'DRAFT 1\'!$L$4:$L$5000,""),"")'
        )

        # Col N (14): ESTÁGIO FUNIL — complex IFS based on RESULTADO
        ws.cell(row=r, column=14).value = (
            f'=IF(F{r}="","",IFS('
            f'F{r}="VENDA / PEDIDO","PÓS-VENDA",'
            f'F{r}="ORÇAMENTO","ORÇAMENTO",'
            f'F{r}="PERDA / FECHOU LOJA","PERDA / NUTRIÇÃO",'
            f'F{r}="EM ATENDIMENTO","EM ATENDIMENTO",'
            f'F{r}="CADASTRO","EM ATENDIMENTO",'
            f'AND(OR(M{r}="ATIVO",M{r}="EM RISCO"),OR(F{r}="RELACIONAMENTO",F{r}="FOLLOW UP 7",F{r}="FOLLOW UP 15")),"CS / RECOMPRA",'
            f'F{r}="SUPORTE","RELACIONAMENTO",'
            f'OR(F{r}="NÃO ATENDE",F{r}="NÃO RESPONDE",F{r}="RECUSOU LIGAÇÃO"),IF(R{r}="","PROSPECÇÃO",R{r}),'
            f'TRUE,"EM ATENDIMENTO"))'
        )

        # Col O (15): FASE
        ws.cell(row=r, column=15).value = (
            f'=IF(F{r}="","",IFS('
            f'F{r}="VENDA / PEDIDO","PÓS-VENDA",'
            f'F{r}="ORÇAMENTO","ORÇAMENTO",'
            f'F{r}="PERDA / FECHOU LOJA","NUTRIÇÃO",'
            f'AND(M{r}="PROSPECT",F{r}="CADASTRO"),"PROSPECÇÃO",'
            f'AND(M{r}="PROSPECT",F{r}="EM ATENDIMENTO"),"PROSPECÇÃO",'
            f'AND(OR(M{r}="INAT.REC",M{r}="EM RISCO"),F{r}="EM ATENDIMENTO"),"SALVAMENTO",'
            f'AND(M{r}="INAT.ANT",F{r}="EM ATENDIMENTO"),"RECUPERAÇÃO",'
            f'AND(OR(M{r}="ATIVO",M{r}="EM RISCO"),F{r}="RELACIONAMENTO"),"CS",'
            f'AND(M{r}="INAT.REC",F{r}="RELACIONAMENTO"),"SALVAMENTO",'
            f'F{r}="SUPORTE","RELACIONAMENTO",'
            f'AND(OR(M{r}="ATIVO",M{r}="EM RISCO"),OR(F{r}="FOLLOW UP 7",F{r}="FOLLOW UP 15")),"CS / RECOMPRA",'
            f'TRUE,"EM ATENDIMENTO"))'
        )

        # Col P (16): TIPO DO CONTATO
        ws.cell(row=r, column=16).value = (
            f'=IF(F{r}="","",IFS('
            f'F{r}="VENDA / PEDIDO","PÓS-VENDA / RELACIONAMENTO",'
            f'F{r}="ORÇAMENTO","NEGOCIAÇÃO",'
            f'F{r}="PERDA / FECHOU LOJA","PERDA / NUTRIÇÃO",'
            f'AND(M{r}="PROSPECT",OR(F{r}="EM ATENDIMENTO",F{r}="CADASTRO")),"PROSPECÇÃO",'
            f'OR(F{r}="FOLLOW UP 7",F{r}="FOLLOW UP 15"),"FOLLOW UP",'
            f'OR(F{r}="RELACIONAMENTO",F{r}="SUPORTE"),"PÓS-VENDA / RELACIONAMENTO",'
            f'AND(OR(M{r}="ATIVO",M{r}="NOVO",M{r}="EM RISCO"),F{r}="EM ATENDIMENTO"),"ATEND. CLIENTES ATIVOS",'
            f'AND(OR(M{r}="INAT.REC",M{r}="INAT.ANT"),F{r}="EM ATENDIMENTO"),"ATEND. CLIENTES INATIVOS",'
            f'OR(F{r}="NÃO ATENDE",F{r}="NÃO RESPONDE",F{r}="RECUSOU LIGAÇÃO"),'
            f'IF(M{r}="PROSPECT","PROSPECÇÃO",IF(OR(M{r}="INAT.REC",M{r}="INAT.ANT"),"ATEND. CLIENTES INATIVOS","ATEND. CLIENTES ATIVOS")),'
            f'TRUE,"ATEND. CLIENTES ATIVOS"))'
        )

        # Col Q (17): TEMPERATURA
        ws.cell(row=r, column=17).value = (
            f'=IF(F{r}="","",IFS('
            f'OR(F{r}="ORÇAMENTO",F{r}="VENDA / PEDIDO"),"🔥 QUENTE",'
            f'F{r}="EM ATENDIMENTO","🟡 MORNO",'
            f'OR(F{r}="RELACIONAMENTO",F{r}="FOLLOW UP 7",F{r}="FOLLOW UP 15"),"🟡 MORNO",'
            f'F{r}="SUPORTE","🟡 MORNO",'
            f'F{r}="PERDA / FECHOU LOJA","💀 PERDIDO",'
            f'OR(F{r}="NÃO ATENDE",F{r}="NÃO RESPONDE",F{r}="RECUSOU LIGAÇÃO"),"❄️ FRIO",'
            f'TRUE,"🟡 MORNO"))'
        )

        # Col R (18): TENTATIVA — simplified (would need previous row reference in real use)
        ws.cell(row=r, column=18).value = (
            f'=IF(OR(F{r}="NÃO ATENDE",F{r}="NÃO RESPONDE",F{r}="RECUSOU LIGAÇÃO"),"T1","")'
        )

        # Col S (19): GRUPO DASH
        ws.cell(row=r, column=19).value = (
            f'=IF(F{r}="","",VLOOKUP(F{r},REGRAS!$B$2:$D$13,3,0))'
        )

        # Col T (20): FOLLOW-UP
        ws.cell(row=r, column=20).value = (
            f'=IF(F{r}="","",IFERROR(WORKDAY(A{r},VLOOKUP(F{r},REGRAS!$B$2:$C$13,2,0)),""))'
        )
        ws.cell(row=r, column=20).number_format = FMT_DATE

        # Col U (21): AÇÃO FUTURA
        ws.cell(row=r, column=21).value = (
            f'=IF(F{r}="","",IFS('
            f'F{r}="PERDA / FECHOU LOJA","NUTRIÇÃO",'
            f'F{r}="VENDA / PEDIDO","PÓS-VENDA",'
            f'M{r}="ATIVO","RECOMPRA",'
            f'M{r}="EM RISCO","SALVAMENTO",'
            f'M{r}="INAT.REC","SALVAMENTO",'
            f'M{r}="INAT.ANT","REATIVAÇÃO",'
            f'M{r}="PROSPECT","PROSPECÇÃO",'
            f'M{r}="NOVO","PÓS-VENDA",'
            f'TRUE,"ATENDIMENTO"))'
        )

        # Col V (22): AÇÃO DETALHADA — auto description
        ws.cell(row=r, column=22).value = (
            f'=IF(F{r}="","",F{r}&" - "&IF(G{r}<>"",G{r},"sem motivo"))'
        )

        # Col W (23): SINALEIRO CICLO — needs DRAFT 1 data
        ws.cell(row=r, column=23).value = (
            f'=IFERROR(IF(XLOOKUP(B{r},\'DRAFT 1\'!$B$4:$B$5000,\'DRAFT 1\'!$AG$4:$AG$5000,0)=0,"🟣",'
            f'IF(XLOOKUP(B{r},\'DRAFT 1\'!$B$4:$B$5000,\'DRAFT 1\'!$L$4:$L$5000,0)'
            f'<=XLOOKUP(B{r},\'DRAFT 1\'!$B$4:$B$5000,\'DRAFT 1\'!$AG$4:$AG$5000,0),"🟢",'
            f'IF(XLOOKUP(B{r},\'DRAFT 1\'!$B$4:$B$5000,\'DRAFT 1\'!$L$4:$L$5000,0)'
            f'<=XLOOKUP(B{r},\'DRAFT 1\'!$B$4:$B$5000,\'DRAFT 1\'!$AG$4:$AG$5000,0)+30,"🟡","🔴"))),"🟣")'
        )

        # Col X (24): SINALEIRO META
        ws.cell(row=r, column=24).value = '=""'

    # ── 5 Example Records (rows 3-7) ──
    examples = [
        # (DATA, CNPJ, NOME, UF, CONSULTOR, RESULTADO, MOTIVO, WPP, LIG, LIG_AT, NOTA, MERCOS, SITUACAO_MANUAL)
        (datetime.date(2026, 2, 3), "32828171000108", "DIVINA TERRA CWB", "PR",
         "MANU DITZEL", "VENDA / PEDIDO", "", "SIM", "SIM", "SIM",
         "Pedido confirmado R$ 3.200", "SIM", "ATIVO"),
        (datetime.date(2026, 2, 4), "99999999000100", "PROSPECT NOVO SP", "SP",
         "LARISSA PADILHA", "ORÇAMENTO", "", "SIM", "NÃO", "N/A",
         "Enviou proposta, aguardando", "SIM", "PROSPECT"),
        (datetime.date(2026, 2, 5), "11111111000111", "LOJA INATIVA REC", "SC",
         "MANU DITZEL", "NÃO ATENDE", "PROPRIETARIO / INDISPONÍVEL", "SIM", "SIM", "NÃO",
         "Tentativa 1, ligar amanhã", "NÃO", "INAT.REC"),
        (datetime.date(2026, 2, 6), "22222222000122", "CLIENTE ATIVO RS", "RS",
         "MANU DITZEL", "FOLLOW UP 7", "AINDA TEM ESTOQUE", "SIM", "NÃO", "N/A",
         "Retornar em 7 dias", "SIM", "ATIVO"),
        (datetime.date(2026, 2, 7), "33333333000133", "LOJA FECHOU RJ", "RJ",
         "LARISSA PADILHA", "PERDA / FECHOU LOJA", "FECHANDO / FECHOU LOJA", "SIM", "SIM", "SIM",
         "Loja encerrou atividades", "SIM", "INAT.ANT"),
    ]

    for i, (dt, cnpj, nome, uf, cons, res, mot, wpp, lig, lig_at, nota, mercos, sit) in enumerate(examples):
        r = 3 + i
        ws.cell(row=r, column=1).value = dt
        ws.cell(row=r, column=1).number_format = FMT_DATE
        write_data(ws, r, 2, cnpj, fmt=FMT_TEXT)
        write_data(ws, r, 3, nome, align=ALIGN_LEFT)
        write_data(ws, r, 4, uf)
        write_data(ws, r, 5, cons, align=ALIGN_LEFT)
        write_data(ws, r, 6, res, align=ALIGN_LEFT)
        write_data(ws, r, 7, mot, align=ALIGN_LEFT)
        write_data(ws, r, 8, wpp)
        write_data(ws, r, 9, lig)
        write_data(ws, r, 10, lig_at)
        write_data(ws, r, 11, nota, align=ALIGN_LEFT)
        write_data(ws, r, 12, mercos)

        # Override SITUAÇÃO formula with manual value for examples
        # (since DRAFT 1 won't have matching data)
        ws.cell(row=r, column=13).value = sit

    # ── Apply calculated cell background ──
    for r in range(3, 8):
        for c in range(13, 25):
            cell = ws.cell(row=r, column=c)
            if cell.fill == FILL_NONE or not cell.fill.fgColor or cell.fill.fgColor.rgb == '00000000':
                cell.fill = FILL_CALC

    # ── Conditional formatting for SITUAÇÃO column (M) ──
    from openpyxl.formatting.rule import CellIsRule
    for sit, fill in SITUACAO_FILLS.items():
        ws.conditional_formatting.add(
            f'M3:M5000',
            CellIsRule(operator='equal', formula=[f'"{sit}"'], fill=fill)
        )

    # ── Freeze at E3 (lock DATA, CNPJ, NOME, UF visible) ──
    ws.freeze_panes = "E3"

    print(f"  DRAFT 2: 24 cols (12 manual + 12 auto), 5 examples, freeze E3")
    return ws

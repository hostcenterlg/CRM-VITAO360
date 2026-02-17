"""
02_extract_sap_data.py
Extrai dados SAP: mapeamento CNPJ<->SAP Code, vendas 2025, metas 2026.

Output: data/output/phase01/sap_data_extracted.json
"""
import json
import os
import re
import sys
import unicodedata
import openpyxl
from collections import defaultdict

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))
SAP_CONSOLIDADO = os.path.join(PROJECT_ROOT, "data", "sources", "sap", "01_SAP_CONSOLIDADO.xlsx")
SAP_META_2026 = os.path.join(PROJECT_ROOT, "data", "sources", "sap", "BASE_SAP_META_PROJECAO_2026.xlsx")
PROJECAO_FILE = os.path.join(PROJECT_ROOT, "data", "sources", "projecao", "PROJECAO_534_INTEGRADA.xlsx")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "data", "output", "phase01")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "sap_data_extracted.json")

# Monthly distribution weights from SAP (confirmed in research)
MONTHLY_WEIGHTS = [
    0.07027, 0.07343, 0.07520, 0.07716,  # Jan-Abr
    0.07969, 0.08177, 0.08493, 0.08702,  # Mai-Ago
    0.08898, 0.09119, 0.09359, 0.09676   # Set-Dez
]
assert abs(sum(MONTHLY_WEIGHTS) - 1.0) < 0.001, f"Monthly weights sum={sum(MONTHLY_WEIGHTS)}, expected ~1.0"

# SAP reference total for meta 2026
SAP_META_REFERENCE = 4747200


def normalize_cnpj(raw):
    """Normalize CNPJ to 14-digit zero-padded string."""
    if raw is None:
        return None
    clean = re.sub(r'[^0-9]', '', str(raw))
    if not clean:
        return None
    return clean.zfill(14) if len(clean) <= 14 else clean


def strip_accents(s):
    """Remove accents from string for sheet name matching."""
    return ''.join(
        c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn'
    )


def find_sheet(wb, target_name):
    """Find sheet by name, ignoring accents and trailing spaces."""
    target_clean = strip_accents(target_name).strip().upper()
    for name in wb.sheetnames:
        if strip_accents(name).strip().upper() == target_clean:
            return wb[name]
    # Fallback: partial match
    for name in wb.sheetnames:
        if target_clean in strip_accents(name).strip().upper():
            return wb[name]
    return None


def extract_rosetta_stone(wb_sap):
    """Extract CNPJ <-> SAP Code mapping from 'Cadastro Clientes SAP' sheet."""
    print("\n--- 1. Rosetta Stone (CNPJ <-> SAP Code) ---")

    ws = find_sheet(wb_sap, "Cadastro Clientes SAP")
    if ws is None:
        print(f"  ERROR: Sheet 'Cadastro Clientes SAP' not found!")
        print(f"  Available: {wb_sap.sheetnames}")
        return {}, {}

    cnpj_to_sap_code = {}
    cnpj_to_info = {}  # extra info for reference

    for row in range(2, ws.max_row + 1):
        sap_code = ws.cell(row=row, column=3).value   # C = Codigo do Cliente
        cnpj_raw = ws.cell(row=row, column=5).value    # E = CNPJ Cliente
        nome = ws.cell(row=row, column=4).value         # D = Nome Cliente
        grupo_chave = ws.cell(row=row, column=9).value  # I = Grupo Chave SAP

        cnpj = normalize_cnpj(cnpj_raw)
        if cnpj and sap_code:
            sap_code_str = str(sap_code).strip()
            cnpj_to_sap_code[cnpj] = sap_code_str
            cnpj_to_info[cnpj] = {
                "sap_code": sap_code_str,
                "nome": str(nome).strip() if nome else "",
                "grupo_chave": str(grupo_chave).strip() if grupo_chave else "",
            }

    print(f"  Mappings extracted: {len(cnpj_to_sap_code)}")
    return cnpj_to_sap_code, cnpj_to_info


def extract_vendas_2025(wb_sap):
    """Extract monthly sales 2025 by CNPJ from 'Venda Mes a Mes' sheet."""
    print("\n--- 2. Vendas Mensais 2025 ---")

    ws = find_sheet(wb_sap, "Venda Mes a Mes")
    if ws is None:
        print(f"  ERROR: Sheet 'Venda Mes a Mes' not found!")
        print(f"  Available: {wb_sap.sheetnames}")
        return {}

    # Faturado columns: G(7), K(11), O(15), S(19), W(23), AA(27), AE(31), AI(35), AM(39), AQ(43), AU(47), AY(51)
    fat_cols = [7, 11, 15, 19, 23, 27, 31, 35, 39, 43, 47, 51]

    # Aggregate by CNPJ (same CNPJ may appear in multiple rows for different Grupo Chave)
    vendas_por_cnpj = defaultdict(lambda: [0.0] * 12)
    rows_processed = 0
    rows_with_cnpj = 0

    for row in range(2, ws.max_row + 1):
        cnpj_raw = ws.cell(row=row, column=3).value  # C = CNPJ/CPF Cliente
        cnpj = normalize_cnpj(cnpj_raw)
        if not cnpj:
            continue
        rows_with_cnpj += 1
        rows_processed += 1

        for i, col in enumerate(fat_cols):
            v = ws.cell(row=row, column=col).value
            try:
                vendas_por_cnpj[cnpj][i] += float(v) if v else 0.0
            except (ValueError, TypeError):
                pass  # skip non-numeric

    # Convert defaultdict to regular dict with lists
    result = {k: list(v) for k, v in vendas_por_cnpj.items()}

    print(f"  Rows processed: {rows_processed}")
    print(f"  Unique CNPJs with sales: {len(result)}")

    # Show total sales
    total_sales = sum(sum(v) for v in result.values())
    print(f"  Total sales 2025: R$ {total_sales:,.2f}")

    return result


def extract_metas_2026(projecao_path):
    """
    Extract metas 2026 per CNPJ from PROJECAO_534_INTEGRADA column L.

    The plan's SIMPLIFIED ALTERNATIVE: extract individual metas already in column L
    of the PROJECAO, then validate the sum against SAP total (~R$ 4,747,200 with 0.67% tolerance).
    """
    print("\n--- 3. Metas 2026 (from PROJECAO column L) ---")

    wb = openpyxl.load_workbook(projecao_path, data_only=True)
    ws = find_sheet(wb, "PROJECAO")
    if ws is None:
        print(f"  ERROR: PROJECAO sheet not found!")
        print(f"  Available: {wb.sheetnames}")
        wb.close()
        return {}

    cnpj_to_meta = {}
    for row in range(4, 538):  # rows 4-537 = 534 clients
        cnpj_raw = ws.cell(row=row, column=1).value  # A = CNPJ
        meta_val = ws.cell(row=row, column=12).value  # L = META ANUAL

        cnpj = normalize_cnpj(cnpj_raw)
        if not cnpj:
            continue

        try:
            meta = float(meta_val) if meta_val else 0.0
        except (ValueError, TypeError):
            meta = 0.0

        cnpj_to_meta[cnpj] = meta

    meta_total = sum(cnpj_to_meta.values())
    discrepancy_pct = abs(meta_total - SAP_META_REFERENCE) / SAP_META_REFERENCE * 100

    print(f"  Clients with meta: {len(cnpj_to_meta)}")
    print(f"  Meta total: R$ {meta_total:,.2f}")
    print(f"  SAP reference: R$ {SAP_META_REFERENCE:,.2f}")
    print(f"  Discrepancy: {discrepancy_pct:.2f}%")
    if discrepancy_pct <= 1.0:
        print(f"  Status: WITHIN tolerance (<=1%)")
    else:
        print(f"  WARNING: Discrepancy exceeds 1% tolerance!")

    wb.close()
    return cnpj_to_meta, meta_total, discrepancy_pct


def extract_grupo_chave_mapping(cnpj_to_info):
    """Build grupo_chave -> list of CNPJs mapping from Rosetta info."""
    print("\n--- 4. Grupo Chave Mapping ---")
    grupo_to_cnpjs = defaultdict(list)
    for cnpj, info in cnpj_to_info.items():
        gc = info.get("grupo_chave", "")
        if gc:
            grupo_to_cnpjs[gc].append(cnpj)

    print(f"  Unique Grupo Chave: {len(grupo_to_cnpjs)}")
    # Show top 5 by count
    sorted_gc = sorted(grupo_to_cnpjs.items(), key=lambda x: len(x[1]), reverse=True)
    for gc, cnpjs in sorted_gc[:5]:
        print(f"    {gc}: {len(cnpjs)} clients")

    return dict(grupo_to_cnpjs)


def main():
    print("=" * 60)
    print("SAP Data Extractor")
    print("=" * 60)

    # Validate source files exist
    for path, label in [(SAP_CONSOLIDADO, "SAP Consolidado"), (PROJECAO_FILE, "PROJECAO_534_INTEGRADA")]:
        if not os.path.exists(path):
            print(f"ERROR: {label} not found at: {path}")
            sys.exit(1)

    # Load SAP Consolidado
    print(f"\nLoading SAP Consolidado...")
    wb_sap = openpyxl.load_workbook(SAP_CONSOLIDADO, data_only=True)
    print(f"  Sheets: {wb_sap.sheetnames}")

    # 1. Rosetta Stone
    cnpj_to_sap_code, cnpj_to_info = extract_rosetta_stone(wb_sap)

    # 2. Vendas 2025
    vendas_2025 = extract_vendas_2025(wb_sap)
    wb_sap.close()

    # 3. Metas 2026 (from PROJECAO column L - simplified alternative)
    meta_result = extract_metas_2026(PROJECAO_FILE)
    cnpj_to_meta, meta_total, meta_discrepancy = meta_result

    # 4. Grupo Chave mapping (for reference)
    grupo_chave_map = extract_grupo_chave_mapping(cnpj_to_info)

    # 5. Monthly weights validation
    print("\n--- 5. Monthly Weights ---")
    print(f"  Sum: {sum(MONTHLY_WEIGHTS):.6f}")
    print(f"  Status: VALID")

    # Build output
    output = {
        "cnpj_to_sap_code": cnpj_to_sap_code,
        "cnpj_to_vendas_2025": vendas_2025,
        "cnpj_to_meta_2026": cnpj_to_meta,
        "monthly_weights": MONTHLY_WEIGHTS,
        "grupo_chave_to_cnpjs": grupo_chave_map,
        "stats": {
            "total_rosetta_mappings": len(cnpj_to_sap_code),
            "total_vendas_clients": len(vendas_2025),
            "total_meta_clients": len(cnpj_to_meta),
            "meta_total_sum": round(meta_total, 2),
            "meta_sap_reference": SAP_META_REFERENCE,
            "meta_discrepancy_pct": round(meta_discrepancy, 4),
        },
    }

    # Save JSON
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"\nOutput saved: {OUTPUT_FILE}")

    # Final summary
    print("\n" + "=" * 60)
    print("EXTRACTION SUMMARY")
    print("=" * 60)
    print(f"  Rosetta mappings:    {len(cnpj_to_sap_code)}")
    print(f"  Clients with vendas: {len(vendas_2025)}")
    print(f"  Clients with meta:   {len(cnpj_to_meta)}")
    print(f"  Meta total:          R$ {meta_total:,.2f} vs R$ {SAP_META_REFERENCE:,.2f} ({meta_discrepancy:.2f}%)")
    print(f"  Monthly weights:     sum={sum(MONTHLY_WEIGHTS):.6f}")
    print(f"  Grupo Chave groups:  {len(grupo_chave_map)}")
    print("=" * 60)

    return 0


if __name__ == "__main__":
    sys.exit(main())

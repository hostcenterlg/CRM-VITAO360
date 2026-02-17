"""
03_populate_projecao.py - Populate PROJECAO with SAP 2026 data (Read-Modify-Write)

Phase 01, Plan 02: Takes the SAP data extracted in Plan 01 (JSON intermediary)
and writes it into the PROJECAO_534_INTEGRADA workbook, producing the V13 output.

CRITICAL RULES:
- data_only=False to preserve all 19,224 formulas
- NEVER write to formula columns (F-J, Z, AN-AQ, BB-CB)
- CNPJ normalized to 14-digit string for matching
- R$ number format preserved on all monetary cells

Input:  data/output/phase01/sap_data_extracted.json (from Plan 01)
        data/sources/projecao/PROJECAO_534_INTEGRADA.xlsx (source template)
Output: data/output/CRM_VITAO360_V13_PROJECAO.xlsx (populated workbook)
"""

import openpyxl
import json
import os
import re
import shutil
import unicodedata
from copy import copy

# === PATHS ===
PROJECT = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360'
SOURCE = os.path.join(PROJECT, 'data', 'sources', 'projecao', 'PROJECAO_534_INTEGRADA.xlsx')
SAP_DATA = os.path.join(PROJECT, 'data', 'output', 'phase01', 'sap_data_extracted.json')
OUTPUT = os.path.join(PROJECT, 'data', 'output', 'CRM_VITAO360_V13_PROJECAO.xlsx')

# === R$ NUMBER FORMAT (matches existing cells) ===
BRL_FORMAT = '_-"R$"\\ * #,##0.00_-;\\-"R$"\\ * #,##0.00_-;_-"R$"\\ * "-"??_-;_-@_-'

# === COLUMN DEFINITIONS ===
# Data columns (safe to write):
COL_CNPJ = 1          # A: CNPJ
COL_META_ANUAL = 12   # L: META ANUAL
COL_META_JAN = 13     # M: META JAN (M-X = cols 13-24)
COL_REAL_JAN = 27     # AA: REAL JAN (AA-AL = cols 27-38)

# Formula columns (NEVER touch):
# F-J (6-10):   VLOOKUP sinaleiro
# Z (26):       =SUM(AA:AL)
# AN-AQ (40-43): % YTD, sinaleiro, GAP, ranking
# BB-CB (54-80): Meta igualitaria e compensada

# Data rows: 4 to 537 (534 clients)
ROW_START = 4
ROW_END = 537


def normalize_cnpj(raw):
    """Normalize CNPJ to 14-digit string."""
    if raw is None:
        return None
    clean = re.sub(r'[^0-9]', '', str(raw))
    if len(clean) == 0:
        return None
    return clean.zfill(14) if len(clean) <= 14 else clean


def strip_accents(s):
    """Remove accents for robust string comparison."""
    return ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )


def find_projecao_sheet(wb):
    """Find the PROJECAO sheet, handling accent variations."""
    for name in wb.sheetnames:
        stripped = strip_accents(name).upper().strip()
        if 'PROJECAO' in stripped:
            return wb[name]
    raise ValueError(f"PROJECAO sheet not found. Available: {wb.sheetnames}")


def safe_float(value, default=0.0):
    """Convert value to float safely, returning default for None/non-numeric."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def main():
    print("=" * 70)
    print("PHASE 01 PLAN 02: Populate PROJECAO with SAP 2026 Data")
    print("=" * 70)

    # 1. Load SAP data from JSON intermediary
    print(f"\n[1] Loading SAP data from: {SAP_DATA}")
    with open(SAP_DATA, 'r', encoding='utf-8') as f:
        sap = json.load(f)

    vendas_2025 = sap['cnpj_to_vendas_2025']
    metas_2026 = sap['cnpj_to_meta_2026']
    weights = sap['monthly_weights']

    print(f"    Metas 2026: {len(metas_2026)} clients")
    print(f"    Vendas 2025: {len(vendas_2025)} clients")
    print(f"    Monthly weights: {len(weights)} months (sum={sum(weights):.6f})")

    # 2. Open PROJECAO preserving formulas (CRITICAL: data_only=False)
    print(f"\n[2] Opening PROJECAO source: {SOURCE}")
    print(f"    data_only=False (preserving all formulas)")
    wb = openpyxl.load_workbook(SOURCE, data_only=False)
    ws = find_projecao_sheet(wb)
    print(f"    Sheet found: '{ws.title}'")
    print(f"    Max row: {ws.max_row}, Max col: {ws.max_column}")

    # Record pre-write state for verification
    pre_freeze = ws.freeze_panes

    # 3. Process each client row
    print(f"\n[3] Processing rows {ROW_START}-{ROW_END} (534 clients)...")

    stats = {
        'processed': 0,
        'meta_updated': 0,
        'meta_skipped_no_match': 0,
        'meta_skipped_zero': 0,
        'vendas_populated': 0,
        'vendas_zeroed': 0,
        'no_cnpj': 0,
        'total_meta': 0.0,
        'total_vendas': 0.0,
        'unmatched_cnpjs': [],
    }

    for row in range(ROW_START, ROW_END + 1):
        raw_cnpj = ws.cell(row=row, column=COL_CNPJ).value
        cnpj = normalize_cnpj(raw_cnpj)

        if cnpj is None:
            stats['no_cnpj'] += 1
            continue

        stats['processed'] += 1

        # --- a) META ANUAL (column L, col 12) ---
        if cnpj in metas_2026:
            meta_anual = safe_float(metas_2026[cnpj])

            # Write META ANUAL
            cell_meta = ws.cell(row=row, column=COL_META_ANUAL)
            cell_meta.value = meta_anual
            cell_meta.number_format = BRL_FORMAT

            stats['meta_updated'] += 1
            stats['total_meta'] += meta_anual

            # --- b) META MENSAL JAN-DEZ (columns M-X, cols 13-24) ---
            if meta_anual > 0:
                for i in range(12):
                    monthly_meta = meta_anual * weights[i]
                    cell_month = ws.cell(row=row, column=COL_META_JAN + i)
                    cell_month.value = monthly_meta
                    cell_month.number_format = BRL_FORMAT
            else:
                stats['meta_skipped_zero'] += 1
                # Write 0 for months if meta is 0
                for i in range(12):
                    cell_month = ws.cell(row=row, column=COL_META_JAN + i)
                    cell_month.value = 0
                    cell_month.number_format = BRL_FORMAT
        else:
            stats['meta_skipped_no_match'] += 1
            if len(stats['unmatched_cnpjs']) < 20:
                stats['unmatched_cnpjs'].append(cnpj)
            # Keep existing value (do not zero out)

        # --- c) REALIZADO JAN-DEZ (columns AA-AL, cols 27-38) ---
        if cnpj in vendas_2025:
            vendas_monthly = vendas_2025[cnpj]
            for i in range(12):
                val = safe_float(vendas_monthly[i] if i < len(vendas_monthly) else 0)
                cell_real = ws.cell(row=row, column=COL_REAL_JAN + i)
                cell_real.value = val
                cell_real.number_format = BRL_FORMAT
                stats['total_vendas'] += val

            stats['vendas_populated'] += 1
        else:
            # No vendas data: set 0 for all months
            for i in range(12):
                cell_real = ws.cell(row=row, column=COL_REAL_JAN + i)
                cell_real.value = 0
                cell_real.number_format = BRL_FORMAT

            stats['vendas_zeroed'] += 1

    # 4. Verify formulas NOT touched (spot check)
    print(f"\n[4] Verifying formulas preserved (spot check)...")
    formula_checks = {
        'F4': (4, 6, '=IFERROR(VLOOKUP'),
        'Z4': (4, 26, '=SUM(AA4:AL4)'),
        'AN4': (4, 40, '=IF(L4=0,0,Z4/L4)'),
        'AO4': (4, 41, '=IF(AN4'),
        'AP4': (4, 42, '=IF(L4=0,0,L4-Z4)'),
        'AQ4': (4, 43, '=IF(Z4=0'),
        'BB4': (4, 54, '=SUM(L$4:L$537)'),
    }

    formula_ok = True
    for label, (r, c, expected_start) in formula_checks.items():
        val = ws.cell(row=r, column=c).value
        if val and str(val).startswith(expected_start):
            print(f"    {label}: OK ({str(val)[:50]}...)")
        else:
            print(f"    {label}: PROBLEM! Expected '{expected_start}...' got '{val}'")
            formula_ok = False

    # 5. Verify properties preserved
    print(f"\n[5] Verifying workbook properties preserved...")
    print(f"    freeze_panes: {ws.freeze_panes} (was: {pre_freeze})")
    assert ws.freeze_panes == pre_freeze, f"freeze_panes changed! Was {pre_freeze}, now {ws.freeze_panes}"

    # 6. Save output
    print(f"\n[6] Saving output to: {OUTPUT}")
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    wb.close()

    output_size = os.path.getsize(OUTPUT)
    print(f"    File size: {output_size:,} bytes ({output_size / 1024:.1f} KB)")

    # 7. Print statistics
    print(f"\n{'=' * 70}")
    print(f"RESULTS")
    print(f"{'=' * 70}")
    print(f"  Clients processed:        {stats['processed']} of 534")
    print(f"  Clients with no CNPJ:     {stats['no_cnpj']}")
    print(f"  ---")
    print(f"  Metas updated:            {stats['meta_updated']}")
    print(f"  Metas skipped (no match): {stats['meta_skipped_no_match']}")
    print(f"  Metas with zero annual:   {stats['meta_skipped_zero']}")
    print(f"  ---")
    print(f"  Vendas populated:         {stats['vendas_populated']}")
    print(f"  Vendas zeroed (no data):  {stats['vendas_zeroed']}")
    print(f"  ---")
    print(f"  Total META populated:     R$ {stats['total_meta']:,.2f}")
    print(f"  Total VENDAS populated:   R$ {stats['total_vendas']:,.2f}")
    print(f"  ---")
    print(f"  Formulas intact:          {'YES' if formula_ok else 'NO - CHECK ABOVE'}")
    print(f"  Output file:              {OUTPUT}")
    print(f"  Output size:              {output_size:,} bytes")

    if stats['unmatched_cnpjs']:
        print(f"\n  Unmatched CNPJs (first {len(stats['unmatched_cnpjs'])}):")
        for cnpj in stats['unmatched_cnpjs']:
            print(f"    - {cnpj}")

    print(f"\n{'=' * 70}")
    print(f"DONE - V13 PROJECAO generated successfully")
    print(f"{'=' * 70}")

    # Return stats for verification
    return stats, formula_ok


if __name__ == '__main__':
    stats, formula_ok = main()
    if not formula_ok:
        print("\nWARNING: Some formulas may have been affected! Manual review needed.")
        exit(1)

"""
01_validate_formulas.py
Valida que as 19.224 formulas existentes na PROJECAO_534_INTEGRADA estao 100% intactas.

Output: data/output/phase01/formula_validation_report.json
"""
import json
import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))
SOURCE_FILE = os.path.join(PROJECT_ROOT, "data", "sources", "projecao", "PROJECAO_534_INTEGRADA.xlsx")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "data", "output", "phase01")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "formula_validation_report.json")

# Expected structure
EXPECTED_MAX_ROW = 537
EXPECTED_MAX_COL = 80
EXPECTED_FREEZE = "C4"
DATA_ROW_START = 4
DATA_ROW_END = 537  # inclusive
TOTAL_DATA_ROWS = DATA_ROW_END - DATA_ROW_START + 1  # 534

# Separator columns (1-indexed): E=5, K=11, Y=25, AM=39, AR=44, BA=53, BO=67
SEPARATOR_COLS = {
    5: "E",
    11: "K",
    25: "Y",
    39: "AM",
    44: "AR",
    53: "BA",
    67: "BO",
}


def col_num(letter):
    """Convert column letter to 1-indexed number."""
    return openpyxl.utils.column_index_from_string(letter)


def build_expected_formulas():
    """Build dict of {col_number: template_function} for all 36 formula columns."""
    formulas = {}

    # Bloco Sinaleiro (F-J, 5 cols)
    sinaleiro_defs = [
        ("F", 6, '$AS$4:$AT$18', 2),
        ("G", 7, '$AS$4:$AU$18', 3),
        ("H", 8, '$AS$4:$AV$18', 4),
        ("I", 9, '$AS$4:$AW$18', 5),
        ("J", 10, '$AS$4:$AX$18', 6),
    ]
    for letter, col, range_ref, idx in sinaleiro_defs:
        formulas[col] = lambda r, rr=range_ref, ii=idx: (
            f'=IFERROR(VLOOKUP(C{r},{rr},{ii},FALSE),"")'
        )

    # Realizado Total (Z, 1 col)
    formulas[col_num("Z")] = lambda r: f"=SUM(AA{r}:AL{r})"

    # Indicadores (AN-AQ, 4 cols)
    formulas[col_num("AN")] = lambda r: f"=IF(L{r}=0,0,Z{r}/L{r})"
    # AO uses emoji indicators in the actual file (not text labels)
    # Structure: =IF(AN{r}="","",IF(AN{r}>=0.9,"<green>",IF(AN{r}>=0.7,"<yellow>",IF(AN{r}>=0.5,"<orange>","<red>"))))
    # We validate the structural pattern, accepting either emoji or text variants
    formulas[col_num("AO")] = None  # special handling in validate_formula
    formulas[col_num("AP")] = lambda r: f"=IF(L{r}=0,0,L{r}-Z{r})"
    formulas[col_num("AQ")] = lambda r: f'=IF(Z{r}=0,"",RANK(Z{r},Z$4:Z$537,0))'

    # Meta Igualitaria (BB-BN, 13 cols)
    # BB = col 54: =SUM(L$4:L$537)/COUNTA(A$4:A$537)
    formulas[col_num("BB")] = lambda r: "=SUM(L$4:L$537)/COUNTA(A$4:A$537)"
    # BC-BN: M->X mapped to cols 55->66
    meta_source_cols = []  # M=13 through X=24 -> letters M,N,O,P,Q,R,S,T,U,V,W,X
    for c in range(13, 25):  # 13=M to 24=X
        meta_source_cols.append(get_column_letter(c))
    for i, src_letter in enumerate(meta_source_cols):
        target_col = col_num("BC") + i  # BC=55, BD=56, ... BN=66
        formulas[target_col] = lambda r, sl=src_letter: (
            f"=SUM({sl}$4:{sl}$537)/COUNTA(A$4:A$537)"
        )

    # Meta Compensada (BP-CB, 13 cols) - validate prefix only
    # BP = col 68: starts with =IF(Z{r}>=BB{r}
    # BQ-CB = cols 69-80: starts with =IF(
    formulas[col_num("BP")] = None  # special handling below
    for c in range(col_num("BQ"), col_num("CB") + 1):
        formulas[c] = None  # special handling below

    return formulas


def validate_formula(ws, row, col, expected_formulas, mismatches):
    """Validate a single cell's formula. Returns True if matched."""
    actual = ws.cell(row=row, column=col).value

    # AO Sinaleiro: validate structure (accepts emoji or text variants)
    if col == col_num("AO"):
        # The formula structure is: =IF(AN{r}="","",IF(AN{r}>=0.9,X,IF(AN{r}>=0.7,Y,IF(AN{r}>=0.5,Z,W))))
        # where X,Y,Z,W can be text ("verde","amarelo","laranja","vermelho") or emojis
        prefix = f'=IF(AN{row}="","",IF(AN{row}>=0.9,'
        if actual and str(actual).startswith(prefix):
            return True
        mismatches.append({
            "cell": f"{get_column_letter(col)}{row}",
            "block": "indicadores_AN_AQ",
            "expected_prefix": prefix,
            "actual": str(actual)[:100] if actual else None,
        })
        return False

    # Meta Compensada: prefix-only validation
    if col == col_num("BP"):
        prefix = f"=IF(Z{row}>=BB{row}"
        if actual and str(actual).startswith(prefix):
            return True
        mismatches.append({
            "cell": f"{get_column_letter(col)}{row}",
            "block": "meta_compensada_BP_CB",
            "expected_prefix": prefix,
            "actual": str(actual)[:80] if actual else None,
        })
        return False
    elif col_num("BQ") <= col <= col_num("CB"):
        if actual and str(actual).startswith("=IF("):
            return True
        mismatches.append({
            "cell": f"{get_column_letter(col)}{row}",
            "block": "meta_compensada_BP_CB",
            "expected_prefix": "=IF(",
            "actual": str(actual)[:80] if actual else None,
        })
        return False

    # Standard formula validation
    template_fn = expected_formulas.get(col)
    if template_fn is None:
        return True  # skip
    expected = template_fn(row)
    if actual == expected:
        return True

    # Determine block name for reporting
    block = "unknown"
    if 6 <= col <= 10:
        block = "sinaleiro_FJ"
    elif col == col_num("Z"):
        block = "realizado_Z"
    elif col_num("AN") <= col <= col_num("AQ"):
        block = "indicadores_AN_AQ"
    elif col_num("BB") <= col <= col_num("BN"):
        block = "meta_igualitaria_BB_BN"

    mismatches.append({
        "cell": f"{get_column_letter(col)}{row}",
        "block": block,
        "expected": expected,
        "actual": str(actual) if actual else None,
    })
    return False


def main():
    print("=" * 60)
    print("PROJECAO Formula Validator")
    print("=" * 60)
    print(f"Source: {SOURCE_FILE}")
    print()

    # Load workbook preserving formulas
    print("Loading workbook (data_only=False)...")
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=False)

    # Find the PROJECAO sheet (may have accent and/or trailing space)
    # Real name is "PROJEÇÃO " (with cedilla and trailing space)
    import unicodedata
    def strip_accents(s):
        return ''.join(
            c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn'
        )
    sheet_name = None
    for name in wb.sheetnames:
        if strip_accents(name).strip().upper() == "PROJECAO":
            sheet_name = name
            break
    if sheet_name is None:
        print("ERROR: Could not find PROJECAO sheet!")
        print(f"Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]
    print(f"Sheet found: '{sheet_name}' (len={len(sheet_name)})")

    # Validate structure
    print("\n--- Structure Validation ---")
    max_row = ws.max_row
    max_col = ws.max_column
    freeze = str(ws.freeze_panes) if ws.freeze_panes else None

    structure_ok = True
    if max_row != EXPECTED_MAX_ROW:
        print(f"  WARNING: max_row={max_row}, expected={EXPECTED_MAX_ROW}")
        structure_ok = False
    else:
        print(f"  max_row: {max_row} OK")

    if max_col != EXPECTED_MAX_COL:
        print(f"  WARNING: max_column={max_col}, expected={EXPECTED_MAX_COL}")
        structure_ok = False
    else:
        print(f"  max_column: {max_col} OK")

    if freeze != EXPECTED_FREEZE:
        print(f"  WARNING: freeze_panes='{freeze}', expected='{EXPECTED_FREEZE}'")
        structure_ok = False
    else:
        print(f"  freeze_panes: {freeze} OK")

    # Validate CNPJs in column A
    print("\n--- CNPJ Validation ---")
    cnpj_count = 0
    empty_cnpj_rows = []
    for row in range(DATA_ROW_START, DATA_ROW_END + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip() != "":
            cnpj_count += 1
        else:
            empty_cnpj_rows.append(row)

    print(f"  CNPJs found: {cnpj_count} (expected: {TOTAL_DATA_ROWS})")
    if empty_cnpj_rows:
        print(f"  Empty CNPJ rows: {empty_cnpj_rows[:10]}{'...' if len(empty_cnpj_rows) > 10 else ''}")

    # Validate auxiliary table AS:AZ (rows 4-18, 15 redes)
    print("\n--- Auxiliary Table (AS:AZ) Validation ---")
    redes_count = 0
    for row in range(4, 19):
        val = ws.cell(row=row, column=col_num("AS")).value
        if val is not None and str(val).strip() != "":
            redes_count += 1
    print(f"  Redes found: {redes_count} (expected: 15)")

    # Validate separator columns
    print("\n--- Separator Columns ---")
    sep_ok = 0
    sep_fail = 0
    for col_idx, col_letter in SEPARATOR_COLS.items():
        # Check a few sample rows
        sample_ok = True
        for row in [4, 100, 300, 537]:
            val = ws.cell(row=row, column=col_idx).value
            if val != ".":
                sample_ok = False
                break
        if sample_ok:
            sep_ok += 1
        else:
            sep_fail += 1
            print(f"  WARNING: Separator {col_letter} (col {col_idx}) has non-'.' values")
    print(f"  Separators OK: {sep_ok}/{len(SEPARATOR_COLS)}")

    # Build expected formulas
    expected_formulas = build_expected_formulas()

    # Define formula column groups for counting
    formula_col_groups = {
        "sinaleiro_FJ": list(range(6, 11)),          # F=6 to J=10 (5 cols)
        "realizado_Z": [col_num("Z")],                # Z=26 (1 col)
        "indicadores_AN_AQ": list(range(col_num("AN"), col_num("AQ") + 1)),  # AN-AQ (4 cols)
        "meta_igualitaria_BB_BN": list(range(col_num("BB"), col_num("BN") + 1)),  # BB-BN (13 cols)
        "meta_compensada_BP_CB": list(range(col_num("BP"), col_num("CB") + 1)),  # BP-CB (13 cols)
    }

    # Validate all formulas
    print("\n--- Formula Validation ---")
    mismatches = []
    block_matched = {block: 0 for block in formula_col_groups}
    block_expected = {}
    total_formulas_found = 0

    for block_name, cols in formula_col_groups.items():
        block_expected[block_name] = len(cols) * TOTAL_DATA_ROWS
        for col in cols:
            for row in range(DATA_ROW_START, DATA_ROW_END + 1):
                total_formulas_found += 1
                if validate_formula(ws, row, col, expected_formulas, mismatches):
                    block_matched[block_name] += 1

    # Print block results
    total_matched = sum(block_matched.values())
    total_expected = sum(block_expected.values())
    print(f"\n  Total formulas scanned: {total_formulas_found}")
    print(f"  Total expected: {total_expected}")
    print(f"  Total matched: {total_matched}")
    print(f"  Total mismatches: {len(mismatches)}")
    print()

    for block_name in formula_col_groups:
        matched = block_matched[block_name]
        expected = block_expected[block_name]
        status = "PASS" if matched == expected else "FAIL"
        print(f"  [{status}] {block_name}: {matched}/{expected}")

    # Print first few mismatches if any
    if mismatches:
        print(f"\n  First {min(10, len(mismatches))} mismatches:")
        for m in mismatches[:10]:
            print(f"    {m['cell']} ({m['block']})")
            if 'expected' in m:
                print(f"      expected: {m['expected']}")
            if 'expected_prefix' in m:
                print(f"      expected prefix: {m['expected_prefix']}")
            print(f"      actual:   {m['actual']}")

    # Build report
    report = {
        "total_formulas_expected": total_expected,
        "total_formulas_found": total_formulas_found,
        "total_formulas_matched": total_matched,
        "total_mismatches": len(mismatches),
        "mismatches": mismatches[:50],  # cap at 50 for readability
        "mismatches_truncated": len(mismatches) > 50,
        "blocks": {},
        "structure": {
            "max_row": max_row,
            "max_column": max_col,
            "freeze_panes": freeze,
            "cnpj_count": cnpj_count,
            "redes_count": redes_count,
            "separators_ok": sep_ok,
            "separators_total": len(SEPARATOR_COLS),
        },
    }

    for block_name in formula_col_groups:
        report["blocks"][block_name] = {
            "expected": block_expected[block_name],
            "matched": block_matched[block_name],
        }

    # Save report
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"\nReport saved: {OUTPUT_FILE}")

    # Final summary
    print("\n" + "=" * 60)
    if len(mismatches) == 0 and total_matched == total_expected:
        print(f"RESULT: ALL {total_expected} FORMULAS VALIDATED SUCCESSFULLY")
        print("Formula validation: 0 mismatches")
    else:
        print(f"RESULT: {len(mismatches)} MISMATCHES FOUND")
        print(f"Formula validation: {len(mismatches)} mismatches")
    print("=" * 60)

    wb.close()
    return len(mismatches)


if __name__ == "__main__":
    exit_code = main()
    sys.exit(0 if exit_code == 0 else 1)

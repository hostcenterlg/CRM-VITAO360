import openpyxl
from openpyxl.utils import get_column_letter
import os, sys

SEPARATOR = chr(61) * 100
SUB_SEP = chr(45) * 80

desktop = os.path.join("C:" + os.sep + "Users" + os.sep + "User" + os.sep + "OneDrive", chr(193) + "rea de Trabalho")

sap_path = os.path.join(
    desktop,
    "PASTA G (CENTRAL INTERNO)",
    "CARTEIRA DE CLIENTES SAP",
    "BASE SAP - META E PROJE" + chr(199) + chr(195) + "O 2026 .- 02. INTERNO - 2026.xlsx"
)

print(SEPARATOR)
print("PART 1 -- SAP META FILE INSPECTION")
print(SEPARATOR)

if not os.path.exists(sap_path):
    print("FILE NOT FOUND: " + sap_path)
else:
    print("File: " + os.path.basename(sap_path))
    print("Size: " + str(os.path.getsize(sap_path)) + " bytes")
    print()
    wb = openpyxl.load_workbook(sap_path, data_only=True, read_only=True)
    print("Sheet names: " + str(wb.sheetnames))
    print()
    for sname in wb.sheetnames:
        print(SUB_SEP)
        print("SHEET: " + sname)
        print(SUB_SEP)
        ws = wb[sname]
        all_rows = []
        for row in ws.iter_rows(min_row=1, values_only=False):
            all_rows.append(row)
            if len(all_rows) > 200:
                break
        print("  Rows: " + str(ws.max_row) + "  |  Cols: " + str(ws.max_column))
        print()
        for r_idx in range(min(2, len(all_rows))):
            vals = [c.value for c in all_rows[r_idx]]
            print("  ROW " + str(r_idx+1) + " (header): " + str(vals))
        print()
        meta_cols = []
        grupo_chave_cols = []
        month_cols = []
        month_kw = ["JAN","FEV","MAR","ABR","MAI","JUN","JUL","AGO","SET","OUT","NOV","DEZ",
                     "JANEIRO","FEVEREIRO","MARCO","ABRIL","MAIO","JUNHO",
                     "JULHO","AGOSTO","SETEMBRO","OUTUBRO","NOVEMBRO","DEZEMBRO"]
        for r_idx in range(min(3, len(all_rows))):
            for c_idx, cell in enumerate(all_rows[r_idx]):
                v = str(cell.value).upper() if cell.value is not None else ""
                cl = get_column_letter(c_idx + 1)
                if "META" in v:
                    meta_cols.append((r_idx+1, cl, cell.value))
                if "GRUPO" in v and "CHAVE" in v:
                    grupo_chave_cols.append((r_idx+1, cl, cell.value))
                for mk in month_kw:
                    if mk in v:
                        month_cols.append((r_idx+1, cl, cell.value))
                        break
        if meta_cols:
            print("  META columns found: " + str(meta_cols))
        if grupo_chave_cols:
            print("  GRUPO CHAVE columns found: " + str(grupo_chave_cols))
        if month_cols:
            print("  Month-related columns found: " + str(month_cols))
        print()
        print("  FIRST 10 DATA ROWS (rows 3-12):")
        for r_idx in range(2, min(12, len(all_rows))):
            vals = [c.value for c in all_rows[r_idx]]
            print("    Row " + str(r_idx+1) + ": " + str(vals))
        print()
    wb.close()

print()
print(SEPARATOR)
print("PART 1 COMPLETE")
print(SEPARATOR)
print()
oular_path = os.path.join(desktop, "CARTEIRA DE CLIENTES OULAR.xlsx")

print(SEPARATOR)
print("PART 2 -- CARTEIRA OULAR: ACOMPANHAMENTO SECTION (cols 72-130)")
print(SEPARATOR)

if not os.path.exists(oular_path):
    print("FILE NOT FOUND: " + oular_path)
else:
    print("File: " + os.path.basename(oular_path))
    print("Size: " + str(os.path.getsize(oular_path)) + " bytes")
    print()
    print(SUB_SEP)
    print("PASS A: data_only=True (computed values)")
    print(SUB_SEP)
    wb_val = openpyxl.load_workbook(oular_path, data_only=True, read_only=True)
    print("Sheet names: " + str(wb_val.sheetnames))
    cart_name = None
    for sn in wb_val.sheetnames:
        if "CARTEIRA" in sn.upper():
            cart_name = sn
            break
    if cart_name is None:
        print("ERROR: No CARTEIRA sheet found.")
    else:
        print("Using sheet: " + repr(cart_name))
        ws = wb_val[cart_name]
        print("  max_row=" + str(ws.max_row) + ", max_column=" + str(ws.max_column))
        print()
        MIN_COL = 72
        MAX_COL = min(130, ws.max_column or 130)
        print("  HEADERS for cols " + str(MIN_COL) + "-" + str(MAX_COL) + " (rows 1-3):")
        for r in range(1, 4):
            row_vals = {}
            for cell_tuple in ws.iter_rows(min_row=r, max_row=r, min_col=MIN_COL, max_col=MAX_COL, values_only=False):
                for c in cell_tuple:
                    if c.value is not None:
                        row_vals[get_column_letter(c.column) + str(c.row)] = c.value
            print("    Row " + str(r) + ": " + str(row_vals))
        print()
        print("  COLUMN-BY-COLUMN HEADER VIEW (col | row1 | row2 | row3):")
        header_data = {}
        for r in range(1, 4):
            for cell_tuple in ws.iter_rows(min_row=r, max_row=r, min_col=MIN_COL, max_col=MAX_COL, values_only=False):
                for c in cell_tuple:
                    cn = c.column
                    if cn not in header_data:
                        header_data[cn] = [None, None, None]
                    header_data[cn][r - 1] = c.value
        for cn in sorted(header_data.keys()):
            cl = get_column_letter(cn)
            h1, h2, h3 = header_data[cn]
            print("    " + cl.rjust(4) + " (col " + str(cn).rjust(3) + "): r1=" + repr(h1).rjust(30) + "  r2=" + repr(h2).rjust(30) + "  r3=" + repr(h3).rjust(30))
        print()
        print("  DATA ROWS 4-6 (values) cols " + str(MIN_COL) + "-" + str(MAX_COL) + ":")
        for r in range(4, 7):
            print("    --- Row " + str(r) + " ---")
            for cell_tuple in ws.iter_rows(min_row=r, max_row=r, min_col=MIN_COL, max_col=MAX_COL, values_only=False):
                for c in cell_tuple:
                    if c.value is not None:
                        print("      " + get_column_letter(c.column) + str(r) + " (col " + str(c.column) + "): " + repr(c.value))
        print()
        meta_col_numbers = []
        for cn, hvals in header_data.items():
            for hv in hvals:
                if hv is not None and "META" in str(hv).upper():
                    meta_col_numbers.append(cn)
                    break
        print("  META columns identified from headers: " + str([(get_column_letter(c), c) for c in meta_col_numbers]))
        if meta_col_numbers:
            rows_with_meta = 0
            total_data_rows = 0
            for row_tuple in ws.iter_rows(min_row=4, min_col=MIN_COL, max_col=MAX_COL, values_only=False):
                total_data_rows += 1
                for c in row_tuple:
                    if hasattr(c, "column") and c.column in meta_col_numbers and c.value is not None:
                        rows_with_meta += 1
                        break
                if total_data_rows > 1000:
                    break
            print("  Rows with non-empty META: " + str(rows_with_meta) + " / " + str(total_data_rows) + " data rows")
        print()
    wb_val.close()

    print(SUB_SEP)
    print("PASS B: data_only=False (formulas)")
    print(SUB_SEP)
    wb_frm = openpyxl.load_workbook(oular_path, data_only=False, read_only=True)
    ws_f = wb_frm[cart_name]
    print("  FORMULAS in cols " + str(MIN_COL) + "-" + str(MAX_COL) + ", rows 4-6:")
    for r in range(4, 7):
        print("    --- Row " + str(r) + " ---")
        for cell_tuple in ws_f.iter_rows(min_row=r, max_row=r, min_col=MIN_COL, max_col=MAX_COL, values_only=False):
            for c in cell_tuple:
                if not hasattr(c, "column"):
                    continue
                v = c.value
                if v is not None:
                    cl = get_column_letter(c.column)
                    is_formula = isinstance(v, str) and v.startswith("=")
                    tag = " [FORMULA]" if is_formula else ""
                    print("      " + cl + str(r) + " (col " + str(c.column) + "): " + repr(v) + tag)
    print()
    wb_frm.close()

print(SEPARATOR)
print("SUMMARY")
print(SEPARATOR)
print()
print("Part 1: Inspected SAP META file -- see above for META columns, GRUPO CHAVE, monthly structure.")
print("Part 2: Inspected CARTEIRA OULAR ACOMPANHAMENTO section (cols 72-130).")
print("        Headers in rows 1-3 reveal monthly groupings (META / REALIZADO / etc.)")
print("        Data rows 4-6 shown with values and formulas.")
print("        META column count and fill rate reported above.")
print()
print("DONE.")

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V17 PRE-POPULADO — CARTEIRA com valores reais (nao apenas formulas).

Problema V16: openpyxl escreve formulas mas NAO calcula valores cached.
O usuario abre no Excel e ve celulas vazias ate dar Ctrl+Alt+F9.

Solucao V17:
  1. Carregar V16 como base
  2. Ler DRAFT 1 (dados estaticos, 554 rows) → indexar por CNPJ
  3. Ler DRAFT 2 (21K rows) → para cada CNPJ, pegar registro mais recente
  4. Para cada row da CARTEIRA:
     - Escrever VALORES reais nas colunas FUNIL (AR-BJ) baseado no ultimo DRAFT 2
     - Escrever CONSULTOR (col L) do DRAFT 1
     - Escrever TIPO CLIENTE (col AX/J) do DRAFT 1
     - Calcular SINALEIRO, TENTATIVA, TIPO ACAO em Python
  5. Manter formulas TAMBEM (como valor+formula nao funciona, escrever apenas valores)
  6. Salvar como V17
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import shutil
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# === PATHS ===
V16_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V16_FINAL.xlsx"
V17_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V17_FINAL.xlsx"
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"


def normalizar_cnpj(valor):
    """Normaliza CNPJ removendo pontuacao"""
    if not valor:
        return None
    cnpj = re.sub(r'[^\d]', '', str(valor).strip())
    return cnpj if len(cnpj) >= 11 else None


def col_letter(idx):
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def calcular_sinaleiro(situacao, dias_sem_compra):
    """Calcula sinaleiro baseado em SITUACAO e DIAS SEM COMPRA"""
    sit = str(situacao).upper().strip() if situacao else ""
    if sit in ("PROSPECT", "LEAD"):
        return "\U0001f7e3"  # roxo
    if dias_sem_compra is None or str(dias_sem_compra).strip() == "":
        return ""
    try:
        dias = float(dias_sem_compra)
    except (ValueError, TypeError):
        return ""
    if dias <= 50:
        return "\U0001f7e2"  # verde
    elif dias <= 90:
        return "\U0001f7e1"  # amarelo
    else:
        return "\U0001f534"  # vermelho


def calcular_tipo_acao(resultado, situacao):
    """Calcula TIPO ACAO baseado em RESULTADO e SITUACAO"""
    res = str(resultado).upper().strip() if resultado else ""
    sit = str(situacao).upper().strip() if situacao else ""

    if not res:
        return ""
    if res == "VENDA / PEDIDO":
        return "VENDA"
    if res in ("PÓS-VENDA", "CS (SUCESSO DO CLIENTE)", "RELACIONAMENTO", "NUTRIÇÃO"):
        return "PÓS-VENDA"
    if res == "SUPORTE":
        return "RESOLUÇÃO DE PROBLEMA"
    if sit in ("PROSPECT", "LEAD"):
        return "PROSPECÇÃO"
    if res == "PERDA / FECHOU LOJA":
        return "PRÉ-VENDA"
    if res in ("ORÇAMENTO", "CADASTRO", "EM ATENDIMENTO", "FOLLOW UP 7",
               "FOLLOW UP 15", "NÃO ATENDE", "NÃO RESPONDE", "RECUSOU LIGAÇÃO"):
        return "PRÉ-VENDA"
    return "PRÉ-VENDA"


def calcular_grupo_dash(resultado, regras_map):
    """Calcula GRUPO DASH via VLOOKUP em REGRAS"""
    if not resultado:
        return ""
    res = str(resultado).strip()
    return regras_map.get(res, "")


def main():
    start_time = datetime.now()
    print("=" * 100)
    print("V17 PRE-POPULADO — CARTEIRA com valores reais")
    print(f"Inicio: {start_time}")
    print("=" * 100)

    # ================================================================
    # CARREGAMENTO
    # ================================================================
    print("\n[LOAD 1/3] Copiando V16 como base para V17...")
    shutil.copy2(V16_PATH, V17_PATH)

    print("[LOAD 2/3] Carregando V17 (edicao)...")
    wb = openpyxl.load_workbook(V17_PATH)
    print(f"  Abas: {wb.sheetnames}")

    print("[LOAD 3/3] Carregando V31 (referencia com data_only=True para valores)...")
    wb_v31 = openpyxl.load_workbook(V31_PATH, data_only=True)

    # ================================================================
    # FASE 1: LER DRAFT 1 (dados de MERCOS - estaticos)
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 1: LER DRAFT 1 (MERCOS - dados estaticos)")
    print(f"{'='*100}")

    ws_d1 = wb["DRAFT 1"]
    print(f"  DRAFT 1: {ws_d1.max_row} rows x {ws_d1.max_column} cols")

    # Descobrir header row e data start row
    # DRAFT 1: headers na row 3, dados a partir de row 4
    # Verificar
    h_b3 = ws_d1.cell(row=3, column=2).value  # deve ser "CNPJ" ou header
    h_j3 = ws_d1.cell(row=3, column=10).value  # deve ser "CONSULTOR" header
    print(f"  Row 3 headers: B={h_b3}, J={h_j3}")

    # Indexar por CNPJ
    d1_data = {}  # cnpj_norm → {col: value}
    d1_consultor_col = None
    d1_tipo_cliente_col = None

    # Detectar qual col tem CONSULTOR e TIPO CLIENTE
    for col in range(1, min(ws_d1.max_column + 1, 50)):
        h = ws_d1.cell(row=3, column=col).value
        if h:
            h_upper = str(h).upper().strip()
            if "CONSULTOR" in h_upper:
                d1_consultor_col = col
                print(f"  CONSULTOR encontrado em col {col_letter(col)} (col {col})")
            if "TIPO" in h_upper and "CLIENTE" in h_upper:
                d1_tipo_cliente_col = col
                print(f"  TIPO CLIENTE encontrado em col {col_letter(col)} (col {col})")

    for row in range(4, ws_d1.max_row + 1):
        cnpj_raw = ws_d1.cell(row=row, column=2).value  # col B = CNPJ
        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            continue

        row_data = {}
        for col in range(1, min(ws_d1.max_column + 1, 50)):
            val = ws_d1.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                # Pular formulas
                if isinstance(val, str) and val.startswith("="):
                    continue
                row_data[col] = val

        d1_data[cnpj] = row_data

    print(f"  {len(d1_data)} clientes indexados por CNPJ")

    # ================================================================
    # FASE 2: LER DRAFT 2 (ultimo atendimento por CNPJ)
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 2: LER DRAFT 2 (ultimo atendimento por CNPJ)")
    print(f"{'='*100}")

    ws_d2 = wb["DRAFT 2"]
    print(f"  DRAFT 2: {ws_d2.max_row} rows x {ws_d2.max_column} cols")

    # Indexar: para cada CNPJ, guardar o registro com DATA mais recente
    d2_latest = {}  # cnpj_norm → {col: value, '_row': row, '_date': date}
    d2_counts = defaultdict(int)  # cnpj_norm → count (para TENTATIVA)

    for row in range(3, ws_d2.max_row + 1):
        cnpj_raw = ws_d2.cell(row=row, column=4).value  # col D = CNPJ
        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            continue

        d2_counts[cnpj] += 1

        date_val = ws_d2.cell(row=row, column=1).value  # col A = DATA
        date_str = str(date_val) if date_val else ""

        # Se ja temos registro para este CNPJ, comparar data
        if cnpj in d2_latest:
            existing_date = d2_latest[cnpj].get('_date', "")
            if date_str <= str(existing_date):
                continue  # registro existente e mais recente

        row_data = {'_row': row, '_date': date_val}
        for col in range(1, 32):
            val = ws_d2.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                # Para formulas, tentar ler valor cached (nao vai ter em V16)
                if isinstance(val, str) and val.startswith("="):
                    # Nao temos valor cached - pular
                    continue
                row_data[col] = val

        d2_latest[cnpj] = row_data

        if (row - 3) % 5000 == 0 and row > 3:
            print(f"    ... processadas {row - 3} rows")

    print(f"  {len(d2_latest)} CNPJs unicos com ultimo atendimento")
    print(f"  Contagens: min={min(d2_counts.values()) if d2_counts else 0}, max={max(d2_counts.values()) if d2_counts else 0}, media={sum(d2_counts.values())/len(d2_counts) if d2_counts else 0:.1f}")

    # ================================================================
    # FASE 2B: LER DRAFT 2 do V31 (com data_only=True = valores calculados)
    # ================================================================
    print(f"\n  [2B] Lendo V31 DRAFT 2 (valores calculados)...")
    ws_v31_d2 = wb_v31["DRAFT 2"]

    v31_d2_latest = {}  # cnpj_norm → {col: value, '_date': date}
    v31_d2_counts = defaultdict(int)

    for row in range(3, ws_v31_d2.max_row + 1):
        cnpj_raw = ws_v31_d2.cell(row=row, column=4).value
        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            continue

        v31_d2_counts[cnpj] += 1
        date_val = ws_v31_d2.cell(row=row, column=1).value
        date_str = str(date_val) if date_val else ""

        if cnpj in v31_d2_latest:
            existing_date = v31_d2_latest[cnpj].get('_date', "")
            if date_str <= str(existing_date):
                continue

        row_data = {'_date': date_val}
        for col in range(1, 32):
            val = ws_v31_d2.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                row_data[col] = val

        v31_d2_latest[cnpj] = row_data

    print(f"    V31 DRAFT 2: {len(v31_d2_latest)} CNPJs unicos")

    # ================================================================
    # FASE 3: LER REGRAS (para GRUPO DASH VLOOKUP)
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 3: LER REGRAS (VLOOKUP)")
    print(f"{'='*100}")

    regras_map = {}  # resultado → grupo_dash
    if "REGRAS" in wb.sheetnames:
        ws_regras = wb["REGRAS"]
        for row in range(6, 21):
            resultado = ws_regras.cell(row=row, column=2).value  # col B
            grupo = ws_regras.cell(row=row, column=4).value  # col D
            if resultado and grupo:
                regras_map[str(resultado).strip()] = grupo
        print(f"  {len(regras_map)} mapeamentos RESULTADO → GRUPO DASH")
    else:
        print("  ! REGRAS nao encontrado")

    # ================================================================
    # FASE 4: LER V31 CARTEIRA (valores calculados) para complementar
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 4: LER V31 CARTEIRA (valores calculados)")
    print(f"{'='*100}")

    ws_v31_cart = wb_v31["CARTEIRA"]
    v31_cart_data = {}  # cnpj_norm → {col: value}

    for row in range(4, ws_v31_cart.max_row + 1):
        cnpj_raw = ws_v31_cart.cell(row=row, column=2).value  # col B = CNPJ
        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            continue

        row_data = {}
        for col in range(1, min(ws_v31_cart.max_column + 1, 270)):
            val = ws_v31_cart.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                row_data[col] = val

        v31_cart_data[cnpj] = row_data

    print(f"  V31 CARTEIRA: {len(v31_cart_data)} clientes indexados")

    # ================================================================
    # FASE 5: PRE-POPULAR CARTEIRA
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 5: PRE-POPULAR CARTEIRA (valores reais)")
    print(f"{'='*100}")

    ws_cart = wb["CARTEIRA"]
    print(f"  CARTEIRA: {ws_cart.max_row} rows x {ws_cart.max_column} cols")

    # Mapeamento das colunas CARTEIRA ← fontes
    # CARTEIRA col → (fonte, col_na_fonte, descricao)
    # Fonte: "D1" = DRAFT 1, "D2" = DRAFT 2 ultimo, "CALC" = calculado, "V31" = V31 CARTEIRA

    CARTEIRA_FUNIL_MAP = {
        # FUNIL (AR=44 a BJ=62)
        44: ("D2", 9, "ESTÁGIO FUNIL"),       # AR ← DRAFT 2 col I
        45: ("D2", 20, "PRÓX FOLLOWUP"),       # AS ← DRAFT 2 col T (FOLLOW-UP)
        46: ("D2", 1, "DATA ÚLT ATEND"),       # AT ← DRAFT 2 col A (DATA mais recente)
        47: ("D2", 21, "AÇÃO FUTURA"),          # AU ← DRAFT 2 col U
        48: ("D2", 18, "ÚLTIMO RESULTADO"),     # AV ← DRAFT 2 col R (RESULTADO)
        49: ("D2", 19, "MOTIVO"),               # AW ← DRAFT 2 col S
        50: ("D1_TIPO", None, "TIPO CLIENTE"),  # AX ← DRAFT 1 TIPO CLIENTE
        51: ("CALC_TENT", None, "TENTATIVA"),   # AY ← "T" + count
        52: ("D2", 11, "FASE"),                 # AZ ← DRAFT 2 col K
        # BA (53) = ? (Vamos checar V31)
        # BB (54) = TEMPERATURA
        # BC-BG (55-59) = ?
        60: ("D2", 21, "PRÓX AÇÃO"),            # BH ← DRAFT 2 col U (AÇÃO FUTURA)
        61: ("D2", 22, "AÇÃO DETALHADA"),        # BI ← DRAFT 2 col V
        62: ("CALC_SINAL", None, "SINALEIRO"),   # BJ ← calculado
    }

    # Detectar colunas por header da row 3
    print(f"\n  Detectando headers da row 3...")
    cart_headers = {}
    for col in range(1, min(ws_cart.max_column + 1, 270)):
        h3 = ws_cart.cell(row=3, column=col).value
        if h3:
            cart_headers[col] = str(h3).strip()

    # Mostrar funil headers
    print(f"  Headers FUNIL (AR=44 a BJ=62):")
    for col in range(44, 63):
        h = cart_headers.get(col, "[sem header]")
        print(f"    {col_letter(col):>4} (col {col:>3}): {h}")

    # CONSULTOR col L (12)
    print(f"\n  CONSULTOR (col L=12): header = {cart_headers.get(12, '?')}")

    # Pre-popular
    filled_total = 0
    consultor_filled = 0
    funil_filled = 0
    empty_cnpj = 0
    no_d2_match = 0
    from_v31 = 0

    for row in range(4, ws_cart.max_row + 1):
        cnpj_raw = ws_cart.cell(row=row, column=2).value  # col B = CNPJ
        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            empty_cnpj += 1
            continue

        # === CONSULTOR (col L) ===
        if d1_consultor_col and cnpj in d1_data:
            consultor_val = d1_data[cnpj].get(d1_consultor_col)
            if consultor_val:
                ws_cart.cell(row=row, column=12, value=consultor_val)
                consultor_filled += 1
                filled_total += 1

        # === TIPO CLIENTE (col AX=50 e col J=10) ===
        if d1_tipo_cliente_col and cnpj in d1_data:
            tipo_val = d1_data[cnpj].get(d1_tipo_cliente_col)
            if tipo_val:
                ws_cart.cell(row=row, column=50, value=tipo_val)  # AX
                filled_total += 1
                # Tambem col J se vazio
                if not ws_cart.cell(row=row, column=10).value:
                    ws_cart.cell(row=row, column=10, value=tipo_val)
                    filled_total += 1

        # === FUNIL (AR-BJ) do DRAFT 2 ===
        d2_record = d2_latest.get(cnpj, {})
        v31_d2_record = v31_d2_latest.get(cnpj, {})
        v31_cart_record = v31_cart_data.get(cnpj, {})

        # Para cada coluna FUNIL mapeada
        for cart_col, (fonte, src_col, desc) in CARTEIRA_FUNIL_MAP.items():
            val = None

            if fonte == "D2" and src_col:
                # Primeiro tentar D2 do V16
                val = d2_record.get(src_col)
                # Se nao tem, tentar V31 DRAFT 2
                if val is None and src_col in v31_d2_record:
                    val = v31_d2_record[src_col]
                    if val is not None:
                        from_v31 += 1

            elif fonte == "D1_TIPO":
                if d1_tipo_cliente_col and cnpj in d1_data:
                    val = d1_data[cnpj].get(d1_tipo_cliente_col)

            elif fonte == "CALC_TENT":
                count = d2_counts.get(cnpj, 0)
                # Tambem contar V31
                v31_count = v31_d2_counts.get(cnpj, 0)
                total_count = max(count, v31_count)
                if total_count > 0:
                    val = f"T{total_count}"

            elif fonte == "CALC_SINAL":
                # Precisamos SITUACAO e DIAS SEM COMPRA da CARTEIRA
                situacao = ws_cart.cell(row=row, column=14).value  # col N = SITUAÇÃO
                if not situacao and cnpj in d1_data:
                    # Tentar pegar do DRAFT 1
                    for c in range(1, 50):
                        h = ws_d1.cell(row=3, column=c).value
                        if h and "SITUAÇ" in str(h).upper():
                            situacao = d1_data[cnpj].get(c)
                            break
                if not situacao:
                    situacao = d2_record.get(7)  # DRAFT 2 col G = SITUAÇÃO
                if not situacao:
                    situacao = v31_d2_record.get(7)

                dias = ws_cart.cell(row=row, column=16).value  # col P = DIAS SEM COMPRA
                if dias is None and cnpj in d1_data:
                    # Tentar DRAFT 1
                    for c in range(1, 50):
                        h = ws_d1.cell(row=3, column=c).value
                        if h and "DIAS" in str(h).upper():
                            dias = d1_data[cnpj].get(c)
                            break
                if dias is None:
                    dias = d2_record.get(8)  # DRAFT 2 col H
                if dias is None:
                    dias = v31_d2_record.get(8)

                val = calcular_sinaleiro(situacao, dias)

            if val is not None and str(val).strip() != "":
                ws_cart.cell(row=row, column=cart_col, value=val)
                funil_filled += 1
                filled_total += 1

        # === Colunas EXTRA da CARTEIRA que podem estar vazias ===

        # col N (14) = SITUAÇÃO (da DRAFT 1 ou D2)
        if not ws_cart.cell(row=row, column=14).value:
            val = d2_record.get(7) or v31_d2_record.get(7)
            if not val and cnpj in d1_data:
                for c in range(1, 50):
                    h = ws_d1.cell(row=3, column=c).value
                    if h and "SITUAÇ" in str(h).upper():
                        val = d1_data[cnpj].get(c)
                        break
            if val:
                ws_cart.cell(row=row, column=14, value=val)
                filled_total += 1

        # col P (16) = DIAS SEM COMPRA (do DRAFT 1 ou D2)
        if not ws_cart.cell(row=row, column=16).value:
            val = d2_record.get(8) or v31_d2_record.get(8)
            if not val and cnpj in d1_data:
                for c in range(1, 50):
                    h = ws_d1.cell(row=3, column=c).value
                    if h and "DIAS" in str(h).upper():
                        val = d1_data[cnpj].get(c)
                        break
            if val:
                ws_cart.cell(row=row, column=16, value=val)
                filled_total += 1

        # Tambem preencher cols que V31 CARTEIRA tinha e V16 nao tem
        # Copiar valores do V31 CARTEIRA para colunas vazias
        if cnpj in v31_cart_data:
            v31_row = v31_cart_data[cnpj]
            for col in range(1, min(ws_cart.max_column + 1, 270)):
                current_val = ws_cart.cell(row=row, column=col).value
                # So preencher se vazio E nao eh formula
                if (current_val is None or str(current_val).strip() == ""):
                    if col in v31_row:
                        v31_val = v31_row[col]
                        if v31_val is not None and str(v31_val).strip() != "":
                            ws_cart.cell(row=row, column=col, value=v31_val)
                            filled_total += 1
                            from_v31 += 1

        if (row - 4) % 100 == 0 and row > 4:
            print(f"    ... processada row {row}/{ws_cart.max_row} ({filled_total} preenchimentos)")

    print(f"\n  RESULTADO PRE-POPULACAO:")
    print(f"    Total preenchimentos: {filled_total:,}")
    print(f"    CONSULTOR preenchidos: {consultor_filled}")
    print(f"    FUNIL preenchidos: {funil_filled}")
    print(f"    Vindos do V31: {from_v31}")
    print(f"    CNPJs sem match no D2: {no_d2_match}")
    print(f"    Rows sem CNPJ: {empty_cnpj}")

    # ================================================================
    # FASE 6: VERIFICACAO PRE-SAVE
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 6: VERIFICACAO PRE-SAVE")
    print(f"{'='*100}")

    # Amostra das primeiras 5 rows de dados
    print(f"\n  Amostra CARTEIRA (rows 4-8):")
    for row in range(4, min(9, ws_cart.max_row + 1)):
        cnpj = ws_cart.cell(row=row, column=2).value
        consultor = ws_cart.cell(row=row, column=12).value
        estagio = ws_cart.cell(row=row, column=44).value  # AR
        followup = ws_cart.cell(row=row, column=45).value  # AS
        data_ult = ws_cart.cell(row=row, column=46).value  # AT
        resultado = ws_cart.cell(row=row, column=48).value  # AV
        tentativa = ws_cart.cell(row=row, column=51).value  # AY
        sinaleiro = ws_cart.cell(row=row, column=62).value  # BJ

        print(f"    Row {row}: CNPJ={str(cnpj)[:18]:18s} | CONSULT={str(consultor)[:15]:15s} | "
              f"FUNIL={str(estagio)[:12]:12s} | RESULT={str(resultado)[:15]:15s} | "
              f"TENT={str(tentativa):5s} | SINAL={sinaleiro}")

    # Contar preenchimento FUNIL
    print(f"\n  Preenchimento FUNIL (amostra rows 4-103):")
    end_check = min(ws_cart.max_row, 103)
    for col in range(44, 63):
        filled = 0
        for row in range(4, end_check + 1):
            val = ws_cart.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "" and not (isinstance(val, str) and val.startswith("=")):
                filled += 1
        pct = round(100 * filled / max(1, end_check - 3), 1)
        h = cart_headers.get(col, "")
        bar = "#" * int(pct / 5) + "." * (20 - int(pct / 5))
        print(f"    {col_letter(col):>4} | {h:25s} | {bar} {pct:>5.1f}%")

    # Contar CONSULTOR preenchido
    consult_filled = 0
    for row in range(4, ws_cart.max_row + 1):
        val = ws_cart.cell(row=row, column=12).value
        if val and str(val).strip() != "" and not (isinstance(val, str) and val.startswith("=")):
            consult_filled += 1
    pct_consult = round(100 * consult_filled / max(1, ws_cart.max_row - 3), 1)
    print(f"\n  CONSULTOR (col L): {consult_filled}/{ws_cart.max_row - 3} ({pct_consult}%)")

    # ================================================================
    # FASE 7: SALVAR
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 7: SALVAR V17")
    print(f"{'='*100}")

    print("  Salvando...")
    wb.save(V17_PATH)
    wb.close()
    wb_v31.close()

    size = Path(V17_PATH).stat().st_size / (1024 * 1024)
    end_time = datetime.now()
    elapsed = (end_time - start_time).total_seconds()

    print(f"  Salvo: {V17_PATH} ({size:.2f} MB)")
    print(f"  Tempo: {elapsed:.1f}s")

    # ================================================================
    # VERIFICACAO FINAL
    # ================================================================
    print(f"\n{'='*100}")
    print("VERIFICACAO FINAL")
    print(f"{'='*100}")

    wb_check = openpyxl.load_workbook(V17_PATH, data_only=False)
    total_formulas = 0
    for name in wb_check.sheetnames:
        ws = wb_check[name]
        f_count = 0
        static_count = 0
        for row_cells in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip() != "":
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        f_count += 1
                    else:
                        static_count += 1
        total_formulas += f_count
        print(f"    {name:25s} | {ws.max_row:>6} rows | {ws.max_column:>4} cols | F={f_count:>7} | D={static_count:>7}")

    print(f"\n  TOTAL FORMULAS: {total_formulas:,}")
    wb_check.close()

    print(f"\n{'='*100}")
    print(f"[SUCESSO] V17 PRE-POPULADO gerado! {filled_total:,} valores escritos em {elapsed:.1f}s")
    print(f"{'='*100}")


if __name__ == "__main__":
    main()

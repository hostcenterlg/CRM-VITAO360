#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AUDITORIA LEVE — Analisa arquivos menores (Mercos, Cart Detalhada, SAP).
V31 ja sabemos: 5.460 CNPJs na CARTEIRA.
V17 ja sabemos: 554 CNPJs.
"""

import openpyxl
import re
from pathlib import Path
from datetime import datetime

def normalizar_cnpj(valor):
    if not valor:
        return None
    cnpj = re.sub(r'[^\d]', '', str(valor).strip())
    return cnpj if len(cnpj) >= 11 else None

def scan_sheet(path, sheet_idx=0, max_cols_scan=20, max_rows_scan=5):
    """Scan rapido para descobrir estrutura"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    name = wb.sheetnames[sheet_idx]
    ws = wb[name]
    print(f"    Aba: '{name}' | max_row={ws.max_row} | max_col={ws.max_column}")

    # Headers
    for row in range(1, min(max_rows_scan + 1, 4)):
        headers = []
        for col in range(1, min(max_cols_scan + 1, (ws.max_column or 20) + 1)):
            val = ws.cell(row=row, column=col).value
            if val:
                headers.append(f"{col}:{str(val)[:25]}")
        if headers:
            print(f"    Row {row}: {', '.join(headers)}")

    wb.close()
    return name

def count_cnpjs_fast(path, sheet_name, cnpj_col, start_row, sit_col=None):
    """Conta CNPJs de forma rapida"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]

    cnpjs = set()
    situacoes = {}

    for row in range(start_row, (ws.max_row or start_row) + 1):
        val = ws.cell(row=row, column=cnpj_col).value
        cnpj = normalizar_cnpj(val)
        if cnpj:
            cnpjs.add(cnpj)
            if sit_col:
                sit = ws.cell(row=row, column=sit_col).value
                if sit:
                    s = str(sit).strip()[:25]
                    situacoes[s] = situacoes.get(s, 0) + 1

    wb.close()
    return cnpjs, situacoes


def main():
    start = datetime.now()
    print("=" * 100)
    print("AUDITORIA LEVE — Fontes de dados menores")
    print("=" * 100)

    results = {}

    # Dados ja conhecidos
    print("\n  DADOS JA CONHECIDOS (sessao anterior):")
    print("    V17 CARTEIRA: 554 CNPJs")
    print("    V31 CARTEIRA: ~5.460 CNPJs (diferenca = ~4.906 faltando)")

    # ================================================================
    # 1. Carteira Detalhada fev2026 COM prospects
    # ================================================================
    print(f"\n{'='*80}")
    print("1. CARTEIRA DETALHADA FEV2026 (COM PROSPECTS)")
    f1 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/Carteira detalhada de clientes fevreiro 2026.xlsx"
    if Path(f1).exists():
        print(f"  Tamanho: {Path(f1).stat().st_size / 1024:.0f} KB")
        sheet = scan_sheet(f1)
        # Baseado no scan, extrair CNPJs
        cnpjs, sit = count_cnpjs_fast(f1, sheet, cnpj_col=2, start_row=2, sit_col=None)
        if not cnpjs:
            # Tentar col 3
            cnpjs, sit = count_cnpjs_fast(f1, sheet, cnpj_col=3, start_row=2)
        print(f"  CNPJs: {len(cnpjs)}")
        results["CART_DET_COM"] = cnpjs
    else:
        print(f"  ! Nao encontrado")

    # ================================================================
    # 2. Carteira Detalhada fev2026 SEM prospects
    # ================================================================
    print(f"\n{'='*80}")
    print("2. CARTEIRA DETALHADA FEV2026 (SEM PROSPECTS)")
    f2 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/Carteira detalhada de clientes fevereiro 2026 - sem os prospects.xlsx"
    if Path(f2).exists():
        print(f"  Tamanho: {Path(f2).stat().st_size / 1024:.0f} KB")
        sheet = scan_sheet(f2)
        cnpjs, sit = count_cnpjs_fast(f2, sheet, cnpj_col=2, start_row=2)
        if not cnpjs:
            cnpjs, sit = count_cnpjs_fast(f2, sheet, cnpj_col=3, start_row=2)
        print(f"  CNPJs: {len(cnpjs)}")
        results["CART_DET_SEM"] = cnpjs
    else:
        print(f"  ! Nao encontrado")

    # Outra versao sem prospects
    f2b = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/Carteira_detalhada_fev2026_sem_prospects.xlsx"
    if Path(f2b).exists():
        print(f"\n  Versao 2: {Path(f2b).name}")
        print(f"  Tamanho: {Path(f2b).stat().st_size / 1024:.0f} KB")
        sheet = scan_sheet(f2b)
        cnpjs2, _ = count_cnpjs_fast(f2b, sheet, cnpj_col=2, start_row=2)
        if not cnpjs2:
            cnpjs2, _ = count_cnpjs_fast(f2b, sheet, cnpj_col=3, start_row=2)
        print(f"  CNPJs: {len(cnpjs2)}")
        results["CART_DET_SEM_V2"] = cnpjs2

    # ================================================================
    # 3. Mercos Carteira Original
    # ================================================================
    print(f"\n{'='*80}")
    print("3. CARTEIRA MERCOS ORIGINAL")
    f3 = r"c:/Users/User/OneDrive/Área de Trabalho/CLAUDE CODE/08_CARTEIRA_MERCOS.xlsx"
    if Path(f3).exists():
        print(f"  Tamanho: {Path(f3).stat().st_size / 1024:.0f} KB")
        sheet = scan_sheet(f3)
        cnpjs, sit = count_cnpjs_fast(f3, sheet, cnpj_col=2, start_row=2, sit_col=None)
        # Tentar detectar cols depois do scan
        if not cnpjs:
            for c in [1, 3, 4, 5]:
                cnpjs, _ = count_cnpjs_fast(f3, sheet, cnpj_col=c, start_row=2)
                if cnpjs:
                    print(f"    CNPJ encontrado na col {c}")
                    break
        print(f"  CNPJs: {len(cnpjs)}")
        results["MERCOS"] = cnpjs
    else:
        print(f"  ! Nao encontrado")

    # ================================================================
    # 4. SAP Total
    # ================================================================
    print(f"\n{'='*80}")
    print("4. SAP CARTEIRA TOTAL")
    f4 = r"c:/Users/User/OneDrive/Área de Trabalho/GERENCIAL/GESTÃO  DE CLIENTES/BASE SAP OFICIAL/CARTEIRA  TOTAL DE CLIENTES SAP.xlsx"
    if Path(f4).exists():
        print(f"  Tamanho: {Path(f4).stat().st_size / 1024:.0f} KB")
        sheet = scan_sheet(f4)
        cnpjs = set()
        for c in [1, 2, 3, 4, 5]:
            cnpjs, _ = count_cnpjs_fast(f4, sheet, cnpj_col=c, start_row=2)
            if cnpjs:
                print(f"    CNPJ na col {c}: {len(cnpjs)}")
                break
        results["SAP_TOTAL"] = cnpjs

    f4b = r"c:/Users/User/OneDrive/Área de Trabalho/GERENCIAL/GESTÃO  DE CLIENTES/BASE SAP OFICIAL/BASE SAP FINAL (INTERNO).xlsx"
    if Path(f4b).exists():
        print(f"\n  Versao 2: BASE SAP FINAL")
        print(f"  Tamanho: {Path(f4b).stat().st_size / 1024:.0f} KB")
        sheet = scan_sheet(f4b)
        cnpjs = set()
        for c in [1, 2, 3, 4, 5]:
            cnpjs, _ = count_cnpjs_fast(f4b, sheet, cnpj_col=c, start_row=2)
            if cnpjs:
                print(f"    CNPJ na col {c}: {len(cnpjs)}")
                break
        results["SAP_FINAL"] = cnpjs

    # ================================================================
    # 5. DRAFT 3 SAP COM INATIVOS
    # ================================================================
    print(f"\n{'='*80}")
    print("5. DRAFT 3 SAP COM INATIVOS")
    f5 = r"c:/Users/User/OneDrive/Área de Trabalho/DRAFT3/DRAFT_3_SAP  COM INATIVOS.xlsx"
    if Path(f5).exists():
        print(f"  Tamanho: {Path(f5).stat().st_size / 1024:.0f} KB")
        sheet = scan_sheet(f5)
        cnpjs = set()
        for c in [1, 2, 3, 4]:
            cnpjs, sit = count_cnpjs_fast(f5, sheet, cnpj_col=c, start_row=2, sit_col=None)
            if cnpjs:
                print(f"    CNPJ na col {c}: {len(cnpjs)}")
                break
        results["DRAFT3_INAT"] = cnpjs

    # ================================================================
    # 6. Carteira por Consultor
    # ================================================================
    print(f"\n{'='*80}")
    print("6. CARTEIRA POR CONSULTOR")
    f6 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/CARTEIRA  POR CONSULTOR .xlsx"
    if Path(f6).exists():
        print(f"  Tamanho: {Path(f6).stat().st_size / 1024:.0f} KB")
        sheet = scan_sheet(f6)
        cnpjs = set()
        for c in [1, 2, 3]:
            cnpjs, _ = count_cnpjs_fast(f6, sheet, cnpj_col=c, start_row=2)
            if cnpjs:
                print(f"    CNPJ na col {c}: {len(cnpjs)}")
                break
        results["CART_CONSULTOR"] = cnpjs

    # ================================================================
    # 7. Carteira de Clientes Jan 2026
    # ================================================================
    print(f"\n{'='*80}")
    print("7. CARTEIRA CLIENTES JAN 2026")
    f7 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/CARTEIRA DE CLIENTES JANEIRO 2026.xlsx"
    if Path(f7).exists():
        print(f"  Tamanho: {Path(f7).stat().st_size / 1024:.0f} KB")
        sheet = scan_sheet(f7)
        cnpjs = set()
        for c in [1, 2, 3]:
            cnpjs, _ = count_cnpjs_fast(f7, sheet, cnpj_col=c, start_row=2)
            if cnpjs:
                print(f"    CNPJ na col {c}: {len(cnpjs)}")
                break
        results["CART_JAN26"] = cnpjs

    # ================================================================
    # RESUMO + CRUZAMENTO
    # ================================================================
    print(f"\n\n{'='*100}")
    print("RESUMO + CRUZAMENTO")
    print(f"{'='*100}")

    # V17 = 554 CNPJs (hardcoded da auditoria acima)
    v17_count = 554

    print(f"\n  RANKING DE FONTES:")
    for name, cnpjs in sorted(results.items(), key=lambda x: -len(x[1])):
        print(f"    {name:25s}: {len(cnpjs):>6} CNPJs")

    # Prospects = COM - SEM
    cart_com = results.get("CART_DET_COM", set())
    cart_sem = results.get("CART_DET_SEM", set()) | results.get("CART_DET_SEM_V2", set())
    if cart_com and cart_sem:
        prospects = cart_com - cart_sem
        print(f"\n  PROSPECTS (diferenca COM-SEM):")
        print(f"    COM prospects: {len(cart_com)}")
        print(f"    SEM prospects: {len(cart_sem)}")
        print(f"    = PROSPECTS: {len(prospects)}")

    # Universo
    todos = set()
    for cnpjs in results.values():
        todos.update(cnpjs)
    print(f"\n  UNIVERSO TOTAL (todas fontes exceto V31): {len(todos)}")
    print(f"  V17 CARTEIRA ATUAL: {v17_count}")

    elapsed = (datetime.now() - start).total_seconds()
    print(f"\n  Tempo: {elapsed:.1f}s")


if __name__ == "__main__":
    main()

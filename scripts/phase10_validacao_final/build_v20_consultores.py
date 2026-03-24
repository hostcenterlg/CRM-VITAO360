#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V20 — DISTRIBUIÇÃO DE CONSULTORES + SAP FALTANTE
SÓ VALORES ESTÁTICOS. ZERO FÓRMULAS.

1. Atribui CONSULTOR a TODOS os 6.143 clientes (CARTEIRA col L + DRAFT 1 col J)
2. Adiciona o 1 SAP faltante do DRAFT 3
3. Grava tudo como valores estáticos
"""

import openpyxl
import re
import shutil
from pathlib import Path
from datetime import datetime

V19 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V19_FINAL.xlsx"
V20 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V20_FINAL.xlsx"

# ============================================================
# MAPEAMENTO UF → CONSULTOR
# ============================================================
# MANU DITZEL: Sul (SC, PR, RS)
# LARISSA PADILHA: Resto do Brasil (default)
# JULIO GADRET: Redes/Franquias (definido por REDES_FRANQUIAS_v2)
# DAIANE STAVICKI: Redes menores + operacional

UF_MANU = {"SC", "PR", "RS"}
# Tudo que não é Sul vai para LARISSA (exceto redes que vão para JULIO/DAIANE)


def norm_cnpj(v):
    if not v:
        return None
    c = re.sub(r'[^\d]', '', str(v).strip())
    return c if len(c) >= 11 else None


def norm_consultor(v):
    """Normaliza nome do consultor"""
    if not v:
        return None
    s = str(v).strip()
    if s.startswith("="):
        return None
    if not s:
        return None
    # Normalizar nomes conhecidos
    upper = s.upper()
    if "LARISSA" in upper:
        return "LARISSA PADILHA"
    if "MANU" in upper:
        return "MANU DITZEL"
    if "JULIO" in upper:
        return "JULIO GADRET"
    if "DAIANE" in upper or "STADLER" in upper or "STAVICKI" in upper:
        return "DAIANE STAVICKI"
    if "HELDER" in upper or "BRUNKOW" in upper:
        return "HELDER BRUNKOW"
    if "LORRANY" in upper:
        return "LARISSA PADILHA"  # Redirecionar ex-consultores para LARISSA
    if "LEANDRO" in upper:
        return "LARISSA PADILHA"  # Redirecionar ex-consultores para LARISSA
    if "RODRIGO" in upper:
        return "LARISSA PADILHA"
    return s


def assign_by_uf(uf):
    """Atribui consultor baseado no UF"""
    if not uf:
        return "LARISSA PADILHA"  # Default: maior território
    u = str(uf).strip().upper()
    if u in UF_MANU:
        return "MANU DITZEL"
    return "LARISSA PADILHA"


def main():
    t = datetime.now()
    print("=" * 100)
    print("V20 — DISTRIBUIÇÃO DE CONSULTORES + SAP FALTANTE")
    print(f"Inicio: {t}")
    print("=" * 100)

    # ============================================================
    # FASE 1: Copiar V19 → V20
    # ============================================================
    print("\n[FASE 1] Copiando V19 → V20...", flush=True)
    shutil.copy2(V19, V20)

    # ============================================================
    # FASE 2: Carregar dados de referência (read-only, rápido)
    # ============================================================
    print("\n[FASE 2] Carregando referências...", flush=True)

    wb_ro = openpyxl.load_workbook(V19, data_only=True, read_only=True)

    # 2a. DRAFT 1 — CNPJ→CONSULTOR existente
    ws_d1 = wb_ro["DRAFT 1"]
    d1_consul = {}  # cnpj → consultor
    d1_uf = {}  # cnpj → UF
    for row in ws_d1.iter_rows(min_row=3, max_row=ws_d1.max_row, min_col=1, max_col=15):
        cnpj = norm_cnpj(row[1].value) if len(row) > 1 else None  # col B (2)
        if not cnpj:
            continue
        consul = row[9].value if len(row) > 9 else None  # col J (10)
        uf = row[3].value if len(row) > 3 else None  # col D (4)
        nc = norm_consultor(consul)
        if nc:
            d1_consul[cnpj] = nc
        if uf and str(uf).strip() and not str(uf).startswith("="):
            d1_uf[cnpj] = str(uf).strip().upper()

    print(f"  DRAFT 1: {len(d1_consul)} com consultor, {len(d1_uf)} com UF", flush=True)

    # 2b. REDES_FRANQUIAS — CNPJs de redes → JULIO/DAIANE
    ws_rf = wb_ro["REDES_FRANQUIAS_v2"]
    redes_cnpjs = set()
    for row in ws_rf.iter_rows(min_row=2, max_row=ws_rf.max_row or 30, min_col=1, max_col=24):
        for cell in row:
            if cell.value:
                cn = norm_cnpj(cell.value)
                if cn:
                    redes_cnpjs.add(cn)
    print(f"  REDES_FRANQUIAS: {len(redes_cnpjs)} CNPJs de redes", flush=True)

    # 2c. CARTEIRA — ler layout
    ws_cart_ro = wb_ro["CARTEIRA"]
    cart_rows = ws_cart_ro.max_row or 0
    cart_cols = ws_cart_ro.max_column or 0

    # Mapa CNPJ → (UF, row_in_carteira)
    cart_data = {}
    for row in ws_cart_ro.iter_rows(min_row=4, max_row=cart_rows, min_col=1, max_col=15):
        cnpj = norm_cnpj(row[1].value) if len(row) > 1 else None  # col B
        if not cnpj:
            continue
        uf = row[3].value if len(row) > 3 else None  # col D
        uf_str = str(uf).strip().upper() if uf and not str(uf).startswith("=") else ""
        cart_data[cnpj] = uf_str

    print(f"  CARTEIRA: {len(cart_data)} clientes com UF mapeado", flush=True)

    # 2d. DRAFT 3 — SAP faltante
    ws_d3 = wb_ro["DRAFT 3 "]
    d3_cnpjs = {}
    for row in ws_d3.iter_rows(min_row=3, max_row=ws_d3.max_row or 0, min_col=1, max_col=16):
        cnpj = norm_cnpj(row[1].value) if len(row) > 1 else None  # col B
        if cnpj and cnpj not in cart_data:
            nome = str(row[2].value)[:40] if len(row) > 2 and row[2].value else ""
            d3_cnpjs[cnpj] = nome
    print(f"  DRAFT 3 faltando na CARTEIRA: {len(d3_cnpjs)}", flush=True)

    wb_ro.close()

    # ============================================================
    # FASE 3: Calcular CONSULTOR para todos
    # ============================================================
    print("\n[FASE 3] Calculando consultor para cada cliente...", flush=True)

    assignments = {}  # cnpj → consultor
    stats = {"d1_existente": 0, "uf_manu": 0, "uf_larissa": 0, "rede_julio": 0, "default": 0}

    for cnpj in cart_data:
        # Prioridade 1: Consultor existente no DRAFT 1
        if cnpj in d1_consul:
            assignments[cnpj] = d1_consul[cnpj]
            stats["d1_existente"] += 1
            continue

        # Prioridade 2: Rede/Franquia → JULIO GADRET
        if cnpj in redes_cnpjs:
            assignments[cnpj] = "JULIO GADRET"
            stats["rede_julio"] += 1
            continue

        # Prioridade 3: UF → MANU (Sul) ou LARISSA (resto)
        uf = cart_data[cnpj]
        if not uf and cnpj in d1_uf:
            uf = d1_uf[cnpj]

        if uf and uf in UF_MANU:
            assignments[cnpj] = "MANU DITZEL"
            stats["uf_manu"] += 1
        elif uf:
            assignments[cnpj] = "LARISSA PADILHA"
            stats["uf_larissa"] += 1
        else:
            # Sem UF → default LARISSA (maior território)
            assignments[cnpj] = "LARISSA PADILHA"
            stats["default"] += 1

    # SAP faltante
    for cnpj, nome in d3_cnpjs.items():
        assignments[cnpj] = "LARISSA PADILHA"
        stats["default"] += 1

    print(f"  Total atribuídos: {len(assignments)}", flush=True)
    print(f"  Fonte DRAFT 1 (existente): {stats['d1_existente']}", flush=True)
    print(f"  Fonte REDES → JULIO: {stats['rede_julio']}", flush=True)
    print(f"  Fonte UF → MANU (Sul): {stats['uf_manu']}", flush=True)
    print(f"  Fonte UF → LARISSA (resto): {stats['uf_larissa']}", flush=True)
    print(f"  Default (sem UF) → LARISSA: {stats['default']}", flush=True)

    # Distribuição final
    final_dist = {}
    for cnpj, consul in assignments.items():
        final_dist[consul] = final_dist.get(consul, 0) + 1
    print(f"\n  DISTRIBUIÇÃO FINAL:", flush=True)
    for c, n in sorted(final_dist.items(), key=lambda x: -x[1]):
        pct = n / len(assignments) * 100
        print(f"    {c:25s}: {n:>5} ({pct:.1f}%)", flush=True)

    # ============================================================
    # FASE 4: Abrir V20 para edição e escrever
    # ============================================================
    print("\n[FASE 4] Abrindo V20 para edição...", flush=True)
    wb = openpyxl.load_workbook(V20)

    # 4a. Escrever CONSULTOR na CARTEIRA (col L = 12)
    print("  [4a] Escrevendo CONSULTOR na CARTEIRA (col 12)...", flush=True)
    ws_cart = wb["CARTEIRA"]
    cart_written = 0
    for r in range(4, (ws_cart.max_row or 0) + 1):
        cnpj_cell = ws_cart.cell(row=r, column=2)
        cnpj = norm_cnpj(cnpj_cell.value)
        if cnpj and cnpj in assignments:
            ws_cart.cell(row=r, column=12, value=assignments[cnpj])
            cart_written += 1
        if cart_written % 1000 == 0 and cart_written > 0:
            print(f"    ... {cart_written} escritos na CARTEIRA", flush=True)
    print(f"  CARTEIRA: {cart_written} consultores escritos", flush=True)

    # 4b. Escrever CONSULTOR no DRAFT 1 (col J = 10) para prospects sem consultor
    print("  [4b] Escrevendo CONSULTOR no DRAFT 1 (col 10)...", flush=True)
    ws_d1 = wb["DRAFT 1"]
    d1_written = 0
    for r in range(3, (ws_d1.max_row or 0) + 1):
        cnpj_cell = ws_d1.cell(row=r, column=2)
        cnpj = norm_cnpj(cnpj_cell.value)
        if not cnpj:
            continue
        current = ws_d1.cell(row=r, column=10).value
        nc = norm_consultor(current)
        if cnpj in assignments:
            # Se já tem consultor normalizado e é o mesmo, não precisa escrever
            if nc and nc == assignments[cnpj]:
                continue
            ws_d1.cell(row=r, column=10, value=assignments[cnpj])
            d1_written += 1
    print(f"  DRAFT 1: {d1_written} consultores escritos/atualizados", flush=True)

    # 4c. Adicionar SAP faltante ao DRAFT 1 e CARTEIRA
    if d3_cnpjs:
        print(f"  [4c] Adicionando {len(d3_cnpjs)} SAP faltantes...", flush=True)
        # DRAFT 1
        next_d1 = (ws_d1.max_row or 0) + 1
        for cnpj, nome in d3_cnpjs.items():
            ws_d1.cell(row=next_d1, column=1, value=nome)  # NOME
            ws_d1.cell(row=next_d1, column=2, value=cnpj)  # CNPJ
            ws_d1.cell(row=next_d1, column=10, value="LARISSA PADILHA")  # CONSULTOR
            next_d1 += 1

        # CARTEIRA
        next_cart = (ws_cart.max_row or 0) + 1
        for cnpj, nome in d3_cnpjs.items():
            ws_cart.cell(row=next_cart, column=1, value=nome)  # NOME
            ws_cart.cell(row=next_cart, column=2, value=cnpj)  # CNPJ
            ws_cart.cell(row=next_cart, column=12, value="LARISSA PADILHA")  # CONSULTOR
            ws_cart.cell(row=next_cart, column=14, value="PROSPECT")  # SITUAÇÃO
            next_cart += 1
        print(f"  SAP faltantes adicionados: DRAFT 1 row {next_d1-len(d3_cnpjs)}, CARTEIRA row {next_cart-len(d3_cnpjs)}", flush=True)

    # ============================================================
    # FASE 5: Verificação pré-save
    # ============================================================
    print("\n[FASE 5] Verificação pré-save...", flush=True)

    # Contar consultores escritos na CARTEIRA
    consul_check = {}
    for r in range(4, (ws_cart.max_row or 0) + 1):
        v = ws_cart.cell(row=r, column=12).value
        if v and str(v).strip() and not str(v).startswith("="):
            c = str(v).strip()
            consul_check[c] = consul_check.get(c, 0) + 1

    print(f"  Consultores na CARTEIRA (valores estáticos):", flush=True)
    total_consul = 0
    for c, n in sorted(consul_check.items(), key=lambda x: -x[1]):
        print(f"    {c:25s}: {n:>5}", flush=True)
        total_consul += n
    print(f"    TOTAL: {total_consul}", flush=True)

    sem_consul = 0
    for r in range(4, (ws_cart.max_row or 0) + 1):
        v = ws_cart.cell(row=r, column=12).value
        if not v or not str(v).strip() or str(v).startswith("="):
            sem_consul += 1
    print(f"    SEM CONSULTOR: {sem_consul}", flush=True)

    # ============================================================
    # FASE 6: Salvar V20
    # ============================================================
    print(f"\n[FASE 6] Salvando V20...", flush=True)
    wb.save(V20)
    wb.close()

    size_mb = Path(V20).stat().st_size / (1024 * 1024)
    elapsed = (datetime.now() - t).total_seconds()

    print(f"\n{'='*100}")
    print(f"[SUCESSO] V20 gerado!")
    print(f"  Arquivo: {V20}")
    print(f"  Tamanho: {size_mb:.2f} MB")
    print(f"  CARTEIRA: {cart_written} consultores escritos como VALORES ESTÁTICOS")
    print(f"  DRAFT 1: {d1_written} consultores atualizados")
    print(f"  SAP faltante: {len(d3_cnpjs)} adicionado(s)")
    print(f"  Distribuição final:")
    for c, n in sorted(final_dist.items(), key=lambda x: -x[1]):
        pct = n / len(assignments) * 100
        print(f"    {c:25s}: {n:>5} ({pct:.1f}%)")
    print(f"  Tempo: {elapsed:.1f}s")
    print(f"{'='*100}")


if __name__ == "__main__":
    main()

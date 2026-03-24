#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AUDIT V42 DEEP — Mapear TUDO que existe em cada aba
====================================================
Objetivo: entender exatamente o que manter, o que zerar, o que construir.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os, re, time

start = time.time()

BASE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360'
V42 = os.path.join(BASE, 'data', 'output', 'phase10', 'CRM_VITAO360_V42_FINAL.xlsx')

print("=" * 80)
print("AUDIT V42 DEEP — MAPEAMENTO COMPLETO")
print("=" * 80)

wb = openpyxl.load_workbook(V42, data_only=True)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n{'='*80}")
    print(f"ABA: {name} ({ws.max_row} rows x {ws.max_column} cols)")
    print(f"{'='*80}")

    # Headers (row 1, 2, 3)
    for hr in range(1, min(4, ws.max_row + 1)):
        vals = []
        for c in range(1, min(ws.max_column + 1, 50)):
            v = ws.cell(row=hr, column=c).value
            if v is not None:
                vals.append(f"{get_column_letter(c)}={str(v)[:30]}")
        if vals:
            print(f"  Row {hr} headers: {', '.join(vals[:20])}")

    # Data rows count (non-empty)
    data_rows = 0
    empty_rows = 0
    first_data_row = None
    for r in range(1, ws.max_row + 1):
        has_data = False
        for c in range(1, min(ws.max_column + 1, 10)):
            if ws.cell(row=r, column=c).value is not None:
                has_data = True
                break
        if has_data:
            data_rows += 1
            if first_data_row is None:
                first_data_row = r
        else:
            empty_rows += 1
    print(f"  Data rows: {data_rows} | Empty rows: {empty_rows} | First data: row {first_data_row}")

    # Sample first 3 data rows and last 3
    if name == 'CARTEIRA':
        print(f"\n  --- CARTEIRA SAMPLE (headers row 3, data row 4+) ---")
        # Headers
        hdrs = []
        for c in range(1, min(ws.max_column + 1, 50)):
            v = ws.cell(row=3, column=c).value
            if v:
                hdrs.append(f"{get_column_letter(c)}({c})={str(v)[:25]}")
        print(f"  Headers: {', '.join(hdrs[:30])}")

        # Count CNPJs valid
        cnpj_count = 0
        cnpj_invalid = 0
        for r in range(4, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            if v:
                s = re.sub(r'[^0-9]', '', str(v))
                if len(s) >= 11:
                    cnpj_count += 1
                else:
                    cnpj_invalid += 1
        print(f"  CNPJs validos: {cnpj_count} | Invalidos: {cnpj_invalid}")

        # Check fat columns (look for numeric data in cols beyond 50)
        fat_populated = 0
        fat_empty = 0
        for r in range(4, min(ws.max_row + 1, 20)):
            for c in range(50, min(ws.max_column + 1, 271)):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip():
                    fat_populated += 1
                else:
                    fat_empty += 1
        print(f"  Fat cols (rows 4-19, cols 50-271): populated={fat_populated} empty={fat_empty}")

        # Sample rows 4-6
        for r in [4, 5, 6]:
            if r > ws.max_row:
                break
            nome = ws.cell(row=r, column=1).value
            cnpj = ws.cell(row=r, column=2).value
            razao = ws.cell(row=r, column=3).value
            uf = ws.cell(row=r, column=4).value
            print(f"  Row {r}: A={str(nome)[:25]} | B={cnpj} | C={str(razao)[:25]} | D={uf}")

    elif name == 'DRAFT 2':
        print(f"\n  --- DRAFT 2 SAMPLE ---")
        # Check date range
        min_date = None
        max_date = None
        date_count = 0
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v is None:
                continue
            date_count += 1
            if isinstance(v, datetime):
                if min_date is None or v < min_date:
                    min_date = v
                if max_date is None or v > max_date:
                    max_date = v
        print(f"  Registros com data: {date_count}")
        if min_date:
            print(f"  Periodo: {min_date.strftime('%d/%m/%Y')} a {max_date.strftime('%d/%m/%Y')}")

        # Sample
        for r in [3, 4, 5]:
            vals = [f"{get_column_letter(c)}={str(ws.cell(row=r,column=c).value)[:30]}" for c in range(1, min(12, ws.max_column+1))]
            print(f"  Row {r}: {', '.join(vals)}")

    elif name.startswith('AGENDA'):
        print(f"\n  --- {name} SAMPLE ---")
        for r in range(1, ws.max_row + 1):
            vals = [f"{get_column_letter(c)}={str(ws.cell(row=r,column=c).value)[:25]}" for c in range(1, min(15, ws.max_column+1))]
            print(f"  Row {r}: {', '.join(vals)}")

    elif 'PROJE' in name.upper():
        print(f"\n  --- PROJECAO SAMPLE ---")
        # Headers
        for r in [1, 2, 3]:
            vals = [f"{get_column_letter(c)}={str(ws.cell(row=r,column=c).value)[:20]}" for c in range(1, min(20, ws.max_column+1))]
            if any('None' not in v for v in vals):
                print(f"  Row {r}: {', '.join(vals)}")
        # Check how many rows have data
        rows_with_data = 0
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=1).value is not None:
                rows_with_data += 1
        print(f"  Rows com dados: {rows_with_data}")

    elif name == 'REGRAS':
        print(f"\n  --- REGRAS SAMPLE ---")
        for r in range(1, min(6, ws.max_row + 1)):
            vals = [f"{get_column_letter(c)}={str(ws.cell(row=r,column=c).value)[:30]}" for c in range(1, min(12, ws.max_column+1))]
            print(f"  Row {r}: {', '.join(vals)}")
        # Count sections
        sections = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and str(v).strip().startswith('Seção') or (v and str(v).strip().startswith('SEÇÃO')):
                sections.append(str(v)[:50])
        print(f"  Secoes encontradas: {len(sections)}")
        for s in sections[:15]:
            print(f"    {s}")

    elif name == 'DASH':
        print(f"\n  --- DASH SAMPLE ---")
        for r in range(1, min(10, ws.max_row + 1)):
            vals = [f"{get_column_letter(c)}={str(ws.cell(row=r,column=c).value)[:25]}" for c in range(1, min(20, ws.max_column+1))]
            print(f"  Row {r}: {', '.join(vals)}")

    elif name == 'SINALEIRO':
        print(f"\n  --- SINALEIRO SAMPLE ---")
        for r in range(1, min(6, ws.max_row + 1)):
            vals = [f"{get_column_letter(c)}={str(ws.cell(row=r,column=c).value)[:25]}" for c in range(1, min(15, ws.max_column+1))]
            print(f"  Row {r}: {', '.join(vals)}")
        # Count formulas vs values
        formulas_count = 0
        values_count = 0
        wb2 = openpyxl.load_workbook(V42, data_only=False)
        ws2 = wb2[name]
        for r in range(2, ws2.max_row + 1):
            for c in range(1, ws2.max_column + 1):
                v = ws2.cell(row=r, column=c).value
                if v and str(v).startswith('='):
                    formulas_count += 1
                elif v is not None:
                    values_count += 1
        wb2.close()
        print(f"  Formulas: {formulas_count} | Valores: {values_count}")

    elif name == 'DRAFT 1':
        print(f"\n  --- DRAFT 1 SAMPLE ---")
        for r in [1, 2, 3, 4]:
            vals = [f"{get_column_letter(c)}={str(ws.cell(row=r,column=c).value)[:25]}" for c in range(1, min(15, ws.max_column+1))]
            print(f"  Row {r}: {', '.join(vals)}")

wb.close()

elapsed = time.time() - start
print(f"\n{'='*80}")
print(f"Audit concluido em {elapsed:.1f}s")
print(f"{'='*80}")

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CRM VITAO360 — V31: Simulação Realista MAI/25
==============================================================
Base: V30 (JAN-ABR simulados)
3 consultores: DAIANE (veterana) + HELDER + MANU (3o mês, mais produtivos)
Feriado: 1 Mai (Dia do Trabalho) → 21 dias úteis
78 vendas MAI da CARTEIRA (col 30)
"""

import openpyxl
from datetime import date, datetime, timedelta
from collections import defaultdict, Counter
import random
import os
import sys

# ============================================================
# CONFIGURAÇÃO
# ============================================================
INPUT_FILE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10\CRM_VITAO360_V30_FINAL.xlsx'
OUTPUT_FILE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10\CRM_VITAO360_V31_FINAL.xlsx'

random.seed(46)
MONTH_NUM = 5
MONTH_NAME = "MAI"
MONTH_END = date(2025, 5, 31)
MONTH_START = date(2025, 5, 1)
FERIADOS = {date(2025, 5, 1)}  # Dia do Trabalho
COL_FAT = 30  # Coluna faturamento MAI na CARTEIRA
PREV_MONTHS_MAX = 4  # Meses anteriores para follow-up (JAN-ABR)

CONSULTORES = {
    "DAIANE STADLER": {"volume_dia": 20, "tipo": "veterana"},
    "HELDER BRUNKOW": {"volume_dia": 15, "tipo": "3o_mes"},
    "MANU DITZEL": {"volume_dia": 15, "tipo": "3o_mes"},
}

# ============================================================
# REGRAS DE NEGÓCIO (da aba REGRAS)
# ============================================================
RESULTADO_REGRAS = {
    "VENDA / PEDIDO":          (4,  "VENDAS",        "🔥QUENTE",  "VENDA"),
    "EM ATENDIMENTO":          (2,  "FUNIL",         "🟡MORNO",   "PRÉ-VENDA"),
    "SUPORTE":                 (1,  "RELAC.",        "🟡MORNO",   "RESOLUÇÃO DE PROBLEMA"),
    "ORÇAMENTO":               (3,  "FUNIL",         "🔥QUENTE",  "PRÉ-VENDA"),
    "RELACIONAMENTO":          (7,  "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "CADASTRO":                (2,  "FUNIL",         "🟡MORNO",   "PRÉ-VENDA"),
    "NÃO ATENDE":              (1,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "NÃO RESPONDE":            (1,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "RECUSOU LIGAÇÃO":         (2,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "CS (SUCESSO DO CLIENTE)": (30, "RELAC.",        "🔥QUENTE",  "PÓS-VENDA"),
    "PÓS-VENDA":              (30, "RELAC.",        "🔥QUENTE",  "PÓS-VENDA"),
    "PERDA / FECHOU LOJA":    (0,  "NÃO VENDA",    "💀PERDIDO",  "PRÉ-VENDA"),
    "FOLLOW UP 7":             (7,  "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "FOLLOW UP 15":            (15, "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "NUTRIÇÃO":                (15, "NÃO VENDA",    "❄️FRIO",    "PROSPECÇÃO"),
}

ACAO_FUTURA_MAP = {
    "VENDA / PEDIDO":           "CONFIRMAR FATURAMENTO E EXPEDIÇÃO",
    "EM ATENDIMENTO":           "FECHAR NEGOCIAÇÃO EM ANDAMENTO",
    "SUPORTE":                  "RESOLVER PROBLEMA INTERNO E ENVIAR SOLUÇÃO",
    "ORÇAMENTO":                "CONFIRMAR ORÇAMENTO ENVIADO",
    "RELACIONAMENTO":           "RAPPORT COM CLIENTE APÓS A VENDA ATÉ RECOMPRAR",
    "CADASTRO":                 "CONFIRMAR CADASTRO NO SISTEMA",
    "NÃO ATENDE":               "2ª TENTATIVA VIA LIGAÇÃO",
    "NÃO RESPONDE":             "2ª TENTATIVA VIA WHATSAPP",
    "RECUSOU LIGAÇÃO":          "2ª TENTATIVA VIA WHATSAPP",
    "CS (SUCESSO DO CLIENTE)":  "VERIFICAR SELL OUT E CRIAR INTENÇÃO RECOMPRA",
    "PÓS-VENDA":               "FAZER CS (SUCESSO DO CLIENTE)",
    "PERDA / FECHOU LOJA":     "NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES",
    "FOLLOW UP 7":              "COBRAR RETORNO DO CLIENTE",
    "FOLLOW UP 15":             "COBRAR RETORNO DO CLIENTE",
    "NUTRIÇÃO":                 "NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES",
}

MOTIVOS = {
    "EM ATENDIMENTO": ["PRIMEIRO CONTATO / SEM RESPOSTA", "AINDA TEM ESTOQUE", "SÓ QUER COMPRAR GRANEL", "NÃO SE APLICA"],
    "ORÇAMENTO": ["AINDA TEM ESTOQUE", "SÓ QUER COMPRAR GRANEL", "PROBLEMA FINANCEIRO / CRÉDITO", "NÃO SE APLICA"],
    "CADASTRO": ["PRIMEIRO CONTATO / SEM RESPOSTA", "NÃO SE APLICA"],
    "VENDA / PEDIDO": ["NÃO SE APLICA"],
    "NÃO RESPONDE": ["PRIMEIRO CONTATO / SEM RESPOSTA", "PROPRIETARIO / INDISPONÍVEL"],
    "NÃO ATENDE": ["PROPRIETARIO / INDISPONÍVEL", "PRIMEIRO CONTATO / SEM RESPOSTA"],
    "RECUSOU LIGAÇÃO": ["PROPRIETARIO / INDISPONÍVEL"],
    "PERDA / FECHOU LOJA": ["SEM INTERESSE NO MOMENTO", "PRODUTO NÃO VENDEU / SEM GIRO", "FECHANDO / FECHOU LOJA", "LOJA ANEXO/PROXIMO - SM", "PROBLEMA FINANCEIRO / CRÉDITO"],
    "PÓS-VENDA": ["STATUS PEDIDO", "ATRASO ENTREGA", "NÃO SE APLICA"],
    "CS (SUCESSO DO CLIENTE)": ["NÃO SE APLICA"],
    "FOLLOW UP": ["AINDA TEM ESTOQUE", "NÃO SE APLICA", "PRIMEIRO CONTATO / SEM RESPOSTA"],
}

DEMANDAS = {
    "PROSPECÇÃO": ["PROSPECÇÃO - BASE GRANEL", "PROSPECÇÃO - PESQUISA GOOGLE", "CONTATO LEADS DO SITE", "RESPONDER LEADS COM DÚVIDAS", "LIGAÇÕES PROSPECT"],
    "ORCAMENTO": ["MONTAR ORÇAMENTO SUGESTÃO (CURVA ABC)", "ENVIO DE AMOSTRAS", "ENVIAR ANÁLISE DE CRÉDITO", "REGISTRAR ATENDIMENTO NO MERCOS"],
    "CADASTRO": ["ATUALIZAÇÃO CADASTRAL MERCOS", "ENVIAR ANÁLISE DE CRÉDITO", "REGISTRAR ATENDIMENTO NO MERCOS", "PREENCHER FOLLOW-UP"],
    "PEDIDO": ["DIGITAÇÃO DE PEDIDOS", "SOLICITAR LINK CARTÃO CRÉDITO", "SOLICITAR VALORES PIX", "CONFIRMAR FATURAMENTO E EXPEDIÇÃO", "REGISTRAR ATENDIMENTO NO MERCOS"],
    "POS_VENDA": ["FAZER RASTREIOS", "ACOMPANHAR DEVOLUÇÕES", "SUPORTE PCV - DÚVIDAS PEDIDOS", "SUPORTE PRODUTO (LAUDOS/TABELAS)", "PREENCHER FOLLOW-UP"],
    "FOLLOWUP": ["LIGAÇÕES DA BASE", "ATENDIMENTO ATIVOS/INATIVOS", "PREENCHER FOLLOW-UP", "REGISTRAR ATENDIMENTO NO MERCOS", "COBRANÇA DE TÍTULOS"],
}

TIPOS_PROBLEMA = [
    "ATRASO ENTREGA (TRANSPORTADORA)", "PRODUTO AVARIADO (FÁBRICA/TRANSPORTE)",
    "ERRO SEPARAÇÃO (EXPEDIÇÃO)", "ERRO NOTA FISCAL (FATURAMENTO)",
    "DIVERGÊNCIA PREÇO (FATURAMENTO)", "COBRANÇA INDEVIDA (FINANCEIRO)",
    "RUPTURA ESTOQUE (FÁBRICA/PCP)", "PROBLEMA SISTEMA (TI)", "", "", "",
]

NOTAS = {
    "PROSPECÇÃO": ["Primeiro contato via WhatsApp. Apresentei portfólio Vitão.",
        "Enviei catálogo digital. Prospect demonstrou interesse.",
        "Contato inicial. Cliente pediu mais informações sobre a linha.",
        "Prospecção via ligação. Consegui falar com responsável de compras."],
    "NEGOCIAÇÃO": ["Negociação avançando. Cliente pediu orçamento atualizado.",
        "Enviei proposta ajustada. Aguardando aprovação do financeiro.",
        "Cliente comparando com concorrente. Reforçei diferenciais Vitão.",
        "Negociação ativa. Cliente quer fechar até fim do mês."],
    "FOLLOWUP": ["Retornei contato conforme agendado. Cliente analisando proposta.",
        "Follow-up realizado. Cliente pediu mais prazo para decidir.",
        "Retomei negociação. Cliente demonstra interesse mas sem urgência.",
        "Follow-up via WhatsApp. Cliente respondeu que vai analisar essa semana."],
    "VENDA": ["Pedido fechado! Cliente satisfeito com condições.",
        "Venda concluída. Primeira compra do cliente.",
        "Pedido aprovado. Enviei confirmação por email."],
    "POS_VENDA": ["Pós-venda. Confirmei faturamento e envio da NF.",
        "Pós-venda. Cliente recebeu tudo certo, satisfeito.",
        "CS. Verificando sell out e oportunidade de recompra.",
        "Acompanhamento pós-venda. Cliente elogiou atendimento."],
    "NAO_RESPONDE": ["WhatsApp enviado, sem retorno.", "Mensagem visualizada mas sem resposta.", "Tentei contato, sem resposta."],
    "NAO_ATENDE": ["Ligação não atendida. Caixa postal.", "Chamou mas não atendeu.", "Tentei ligar, sem sucesso."],
    "PERDA": ["Cliente sem interesse no momento.", "Não conseguiu fechar. Produto não se encaixa.", "Cliente optou por outro fornecedor."],
}

# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================
def dias_uteis_lista(start, end, feriados=None):
    feriados = feriados or set()
    result = []
    d = start
    while d <= end:
        if d.weekday() < 5 and d not in feriados:
            result.append(d)
        d += timedelta(days=1)
    return result

def proximo_dia_util(d, n=1, feriados=None):
    feriados = feriados or set()
    count = 0
    current = d
    while count < n:
        current += timedelta(days=1)
        if current.weekday() < 5 and current not in feriados:
            count += 1
    return current

def sinaleiro_calc(situacao, dias_compra, ciclo_medio):
    if not situacao:
        return "🟣"
    sit_upper = str(situacao).upper()
    if sit_upper in ("PROSPECT", "LEAD", "NOVO"):
        return "🟣"
    if dias_compra is None or dias_compra == "":
        return ""
    try:
        d = int(dias_compra)
        c = int(ciclo_medio) if ciclo_medio else 50
    except (ValueError, TypeError):
        return ""
    if d <= c:
        return "🟢"
    elif d <= c + 30:
        return "🟡"
    else:
        return "🔴"

def make_record(data, consultor, cnpj, nome, uf, rede, situacao, dias_compra,
                estagio_funil, tipo_cliente, fase, whatsapp, ligacao, lig_atendida,
                tipo_contato, resultado, motivo, followup_date, acao_futura,
                acao_detalhada, mercos, nota, temperatura, tipo_problema,
                demanda, tipo_atendimento, ciclo_medio=0):
    sinaleiro = sinaleiro_calc(situacao, dias_compra, ciclo_medio)
    grupo_dash = RESULTADO_REGRAS.get(resultado, (0, "", "", ""))[1]
    tipo_acao = RESULTADO_REGRAS.get(resultado, (0, "", "", ""))[3]
    return {
        'A': data, 'B': consultor, 'C': nome, 'D': cnpj,
        'E': uf, 'F': rede, 'G': situacao, 'H': dias_compra,
        'I': estagio_funil, 'J': tipo_cliente, 'K': fase,
        'L': sinaleiro, 'M': '',
        'N': whatsapp, 'O': ligacao, 'P': lig_atendida,
        'Q': tipo_contato, 'R': resultado, 'S': motivo,
        'T': followup_date, 'U': acao_futura, 'V': acao_detalhada,
        'W': mercos, 'X': nota, 'Y': temperatura,
        'Z': grupo_dash, 'AA': '', 'AB': tipo_acao,
        'AC': tipo_problema, 'AD': demanda, 'AE': tipo_atendimento,
    }

def gerar_ciclo_venda(cnpj_s, info, consultor, d_base, in_prev, prev_res, du, feriados, month_end):
    records = []
    gap = random.choice([1, 2, 2, 3])
    if in_prev and prev_res in ("ORÇAMENTO", "EM ATENDIMENTO"):
        steps = ["CADASTRO", "VENDA / PEDIDO"]
    elif in_prev and prev_res == "CADASTRO":
        steps = ["VENDA / PEDIDO"]
    elif in_prev:
        steps = ["EM ATENDIMENTO", "ORÇAMENTO", "CADASTRO", "VENDA / PEDIDO"]
    else:
        steps = ["EM ATENDIMENTO", "ORÇAMENTO", "CADASTRO", "VENDA / PEDIDO"]

    d_current = d_base
    for step_idx, step in enumerate(steps):
        if step_idx > 0:
            d_current = proximo_dia_util(d_current, gap, feriados)
        if d_current > month_end:
            d_current = month_end
        if d_current.weekday() >= 5:
            d_current = proximo_dia_util(d_current - timedelta(days=1), 1, feriados)

        step_map = {
            "EM ATENDIMENTO": ("EM ATENDIMENTO", "EM ATENDIMENTO", "ENVIAR ORÇAMENTO / CATÁLOGO",
                              MOTIVOS["EM ATENDIMENTO"], NOTAS["NEGOCIAÇÃO"],
                              DEMANDAS["PROSPECÇÃO"] if not in_prev else DEMANDAS["FOLLOWUP"], "PROSPECÇÃO" if not in_prev else "NEGOCIAÇÃO"),
            "ORÇAMENTO": ("ORÇAMENTO", "ORÇAMENTO", "CONFIRMAR ORÇAMENTO ENVIADO",
                         MOTIVOS["ORÇAMENTO"], NOTAS["NEGOCIAÇÃO"], DEMANDAS["ORCAMENTO"], "NEGOCIAÇÃO"),
            "CADASTRO": ("CADASTRO", "EM ATENDIMENTO", "CONFIRMAR CADASTRO NO SISTEMA",
                        MOTIVOS["CADASTRO"], ["Cadastro enviado para análise."], DEMANDAS["CADASTRO"], "NEGOCIAÇÃO"),
            "VENDA / PEDIDO": ("VENDA / PEDIDO", "PÓS-VENDA", "CONFIRMAR FATURAMENTO E EXPEDIÇÃO",
                              MOTIVOS["VENDA / PEDIDO"], NOTAS["VENDA"], DEMANDAS["PEDIDO"], "NEGOCIAÇÃO"),
        }
        resultado, estagio, acao, motivos_list, notas_list, demandas_list, tc = step_map[step]
        records.append(make_record(
            data=d_current, consultor=consultor, cnpj=cnpj_s,
            nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
            situacao=info.get('situacao', 'PROSPECT'), dias_compra=info.get('dias_compra', ''),
            estagio_funil=estagio, tipo_cliente=info.get('tipo_cliente', 'NOVO'),
            fase="PROSPECÇÃO" if step == "EM ATENDIMENTO" and not in_prev else estagio,
            whatsapp=1 if step_idx == 0 else 0, ligacao=1 if step_idx > 0 else 0, lig_atendida=1,
            tipo_contato=tc, resultado=resultado, motivo=random.choice(motivos_list),
            followup_date=proximo_dia_util(d_current, gap, feriados),
            acao_futura=acao, acao_detalhada="",
            mercos="SIM" if step in ("ORÇAMENTO", "VENDA / PEDIDO") else "NÃO",
            nota=random.choice(notas_list),
            temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "🟡MORNO", ""))[2],
            tipo_problema="", demanda=random.choice(demandas_list),
            tipo_atendimento="PROSPECÇÃO" if not in_prev and step_idx == 0 else "FOLLOW UP",
            ciclo_medio=info.get('ciclo_medio', 0)
        ))

    d_pv = proximo_dia_util(d_current, 4, feriados)
    if d_pv <= month_end:
        records.append(make_record(
            data=d_pv, consultor=consultor, cnpj=cnpj_s,
            nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
            situacao="NOVO", dias_compra=4, estagio_funil="PÓS-VENDA", tipo_cliente="NOVO", fase="PÓS-VENDA",
            whatsapp=0, ligacao=1, lig_atendida=1,
            tipo_contato="PÓS-VENDA / RELACIONAMENTO", resultado="PÓS-VENDA", motivo="NÃO SE APLICA",
            followup_date=proximo_dia_util(d_pv, 11, feriados),
            acao_futura="FAZER CS (SUCESSO DO CLIENTE)", acao_detalhada="",
            mercos="SIM", nota=random.choice(NOTAS["POS_VENDA"]),
            temperatura="🔥QUENTE", tipo_problema="",
            demanda=random.choice(DEMANDAS["POS_VENDA"]), tipo_atendimento="ATEND. CLIENTES ATIVOS"
        ))
    return records

def gerar_followup(cnpj_s, info, consultor, d, prev_res, feriados):
    rand = random.random()
    if prev_res in ("EM ATENDIMENTO", "ORÇAMENTO"):
        if rand < 0.40:
            resultado = "ORÇAMENTO" if prev_res == "EM ATENDIMENTO" else "EM ATENDIMENTO"
        elif rand < 0.70:
            resultado = "EM ATENDIMENTO"
        elif rand < 0.90:
            resultado = random.choice(["NÃO RESPONDE", "NÃO ATENDE"])
        else:
            resultado = "PERDA / FECHOU LOJA"
    elif prev_res in ("PÓS-VENDA", "CS (SUCESSO DO CLIENTE)"):
        resultado = random.choice(["RELACIONAMENTO", "CS (SUCESSO DO CLIENTE)", "FOLLOW UP 7"])
    elif prev_res in ("NÃO RESPONDE", "NÃO ATENDE", "RECUSOU LIGAÇÃO"):
        if rand < 0.50:
            resultado = "EM ATENDIMENTO"
        elif rand < 0.80:
            resultado = random.choice(["NÃO RESPONDE", "NÃO ATENDE"])
        else:
            resultado = "PERDA / FECHOU LOJA"
    else:
        resultado = "EM ATENDIMENTO"

    motivos_key = resultado if resultado in MOTIVOS else "FOLLOW UP"
    canal = random.choice(['whatsapp', 'ligacao', 'ligacao'])
    return make_record(
        data=d, consultor=consultor, cnpj=cnpj_s,
        nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
        situacao=info.get('situacao', 'PROSPECT'), dias_compra=info.get('dias_compra', ''),
        estagio_funil="EM ATENDIMENTO" if resultado == "EM ATENDIMENTO" else (
            "ORÇAMENTO" if resultado == "ORÇAMENTO" else "PROSPECÇÃO"),
        tipo_cliente=info.get('tipo_cliente', 'NOVO'), fase="PROSPECÇÃO",
        whatsapp=1 if canal == 'whatsapp' else 0,
        ligacao=1 if canal == 'ligacao' else 0,
        lig_atendida=1 if resultado not in ("NÃO ATENDE", "NÃO RESPONDE", "RECUSOU LIGAÇÃO") else 0,
        tipo_contato="FOLLOW UP", resultado=resultado,
        motivo=random.choice(MOTIVOS.get(motivos_key, ["NÃO SE APLICA"])),
        followup_date=proximo_dia_util(d, RESULTADO_REGRAS.get(resultado, (2,))[0], feriados) if resultado != "PERDA / FECHOU LOJA" else "",
        acao_futura=ACAO_FUTURA_MAP.get(resultado, ""),
        acao_detalhada="", mercos="SIM" if resultado == "ORÇAMENTO" else "NÃO",
        nota=random.choice(NOTAS.get("FOLLOWUP", [""])) if resultado in ("EM ATENDIMENTO", "ORÇAMENTO") else (
            random.choice(NOTAS.get("NAO_RESPONDE", [""])) if "RESPONDE" in resultado else (
            random.choice(NOTAS.get("NAO_ATENDE", [""])) if "ATENDE" in resultado else (
            random.choice(NOTAS.get("PERDA", [""]))))),
        temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "🟡MORNO", ""))[2],
        tipo_problema="",
        demanda=random.choice(DEMANDAS.get("FOLLOWUP", [""])) if resultado in ("EM ATENDIMENTO", "ORÇAMENTO") else "",
        tipo_atendimento="FOLLOW UP", ciclo_medio=info.get('ciclo_medio', 0)
    )

def gerar_prospeccao(cnpj_s, info, consultor, d, feriados):
    rand = random.random()
    if rand < 0.60:
        resultado = "EM ATENDIMENTO"
    elif rand < 0.85:
        resultado = "NÃO RESPONDE"
    else:
        resultado = "NÃO ATENDE"
    return make_record(
        data=d, consultor=consultor, cnpj=cnpj_s,
        nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
        situacao="PROSPECT", dias_compra="",
        estagio_funil="PROSPECÇÃO", tipo_cliente="PROSPECT", fase="PROSPECÇÃO",
        whatsapp=1, ligacao=0, lig_atendida=0,
        tipo_contato="PROSPECÇÃO", resultado=resultado,
        motivo=random.choice(MOTIVOS.get(resultado, ["NÃO SE APLICA"])),
        followup_date=proximo_dia_util(d, RESULTADO_REGRAS.get(resultado, (2,))[0], feriados),
        acao_futura=ACAO_FUTURA_MAP.get(resultado, "ENVIAR ORÇAMENTO / CATÁLOGO"),
        acao_detalhada="", mercos="NÃO",
        nota=random.choice(NOTAS.get("PROSPECÇÃO", [""])) if resultado == "EM ATENDIMENTO" else (
            random.choice(NOTAS.get("NAO_RESPONDE", [""])) if "RESPONDE" in resultado else
            random.choice(NOTAS.get("NAO_ATENDE", [""]))),
        temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "❄️FRIO", ""))[2],
        tipo_problema="", demanda=random.choice(DEMANDAS["PROSPECÇÃO"]),
        tipo_atendimento="PROSPECÇÃO"
    )

# ============================================================
# INÍCIO
# ============================================================
print("=" * 80)
print(f"CRM VITAO360 — V31: Simulação Realista {MONTH_NAME}/25")
print("=" * 80)

if not os.path.exists(INPUT_FILE):
    print(f"ERRO: {INPUT_FILE} não encontrado")
    sys.exit(1)

print(f"\n[1/7] Carregando V30...")
wb = openpyxl.load_workbook(INPUT_FILE)
ws_cart = wb['CARTEIRA']
ws_d2 = wb['DRAFT 2']

# CARTEIRA
print(f"\n[2/7] Lendo CARTEIRA...")
cart_data = {}
for r in range(4, ws_cart.max_row + 1):
    cnpj = ws_cart.cell(row=r, column=3).value
    if not cnpj:
        continue
    cnpj_s = str(cnpj).strip()
    cart_data[cnpj_s] = {
        'nome': ws_cart.cell(row=r, column=2).value or "",
        'uf': ws_cart.cell(row=r, column=4).value or "",
        'rede': ws_cart.cell(row=r, column=5).value or "",
        'situacao': ws_cart.cell(row=r, column=6).value or "",
        'tipo_cliente': ws_cart.cell(row=r, column=7).value or "",
        'ciclo_medio': ws_cart.cell(row=r, column=8).value or 0,
        'dias_compra': ws_cart.cell(row=r, column=9).value or "",
        'fat_mes': ws_cart.cell(row=r, column=COL_FAT).value or 0,
    }

buyers = [c for c, info in cart_data.items()
          if info['fat_mes'] and isinstance(info['fat_mes'], (int, float)) and info['fat_mes'] > 0]
print(f"  Compradores {MONTH_NAME}/25: {len(buyers)}")
print(f"  Total clientes CARTEIRA: {len(cart_data)}")

# FOLLOW-UPS
print(f"\n[3/7] Analisando follow-ups meses anteriores...")
NOMES_DAIANE = {"DAIANE STAVICKI", "DAIANE STADLER"}
NOMES_VALIDOS = NOMES_DAIANE | {"HELDER BRUNKOW", "MANU DITZEL"}

prev_last = {}
for r in range(3, ws_d2.max_row + 1):
    data_cell = ws_d2.cell(row=r, column=1).value
    if data_cell is None:
        continue
    if isinstance(data_cell, (int, float)):
        data_cell = datetime(1900, 1, 1) + timedelta(days=int(data_cell) - 2)
    if not isinstance(data_cell, datetime):
        continue
    if data_cell.year != 2025 or data_cell.month > PREV_MONTHS_MAX:
        continue
    cons = ws_d2.cell(row=r, column=2).value
    if cons is None:
        continue
    cons_str = str(cons).strip()
    if cons_str in NOMES_DAIANE:
        cons_norm = "DAIANE STADLER"
    elif cons_str in NOMES_VALIDOS:
        cons_norm = cons_str
    else:
        continue
    cnpj = ws_d2.cell(row=r, column=4).value
    if not cnpj:
        continue
    cnpj_s = str(cnpj).strip()
    res = ws_d2.cell(row=r, column=18).value or ""
    key = (cons_norm, cnpj_s)
    if key not in prev_last or data_cell >= prev_last[key]['data']:
        prev_last[key] = {'data': data_cell, 'resultado': str(res)}

prev_by_cons = defaultdict(lambda: defaultdict(list))
for (cons, cnpj), info in prev_last.items():
    prev_by_cons[cons][info['resultado']].append(cnpj)

for cons in sorted(prev_by_cons.keys()):
    total = sum(len(v) for v in prev_by_cons[cons].values())
    print(f"  {cons}: {total} CNPJs até {MONTH_NAME}")

# POOLS
print(f"\n[4/7] Montando pools {MONTH_NAME}/25...")
du = dias_uteis_lista(MONTH_START, MONTH_END, FERIADOS)
print(f"  Dias úteis: {len(du)}")

def build_followup_pool(cons_name, max_fu=80, max_retry=25):
    by_res = prev_by_cons.get(cons_name, {})
    pool_fu = []
    for res in ["EM ATENDIMENTO", "ORÇAMENTO", "PÓS-VENDA", "CS (SUCESSO DO CLIENTE)", "VENDA / PEDIDO", "RELACIONAMENTO"]:
        cnpjs = by_res.get(res, [])
        random.shuffle(cnpjs)
        pool_fu.extend(cnpjs[:25])
    pool_fu = list(set(pool_fu))[:max_fu]
    pool_retry = []
    for res in ["NÃO RESPONDE", "NÃO ATENDE", "RECUSOU LIGAÇÃO"]:
        cnpjs = by_res.get(res, [])
        random.shuffle(cnpjs)
        pool_retry.extend(cnpjs[:12])
    pool_retry = list(set(pool_retry))[:max_retry]
    return pool_fu, pool_retry

pool_daiane_fu, pool_daiane_retry = build_followup_pool("DAIANE STADLER", max_fu=90, max_retry=30)
pool_helder_fu, pool_helder_retry = build_followup_pool("HELDER BRUNKOW", max_fu=60, max_retry=20)
pool_manu_fu, pool_manu_retry = build_followup_pool("MANU DITZEL", max_fu=60, max_retry=20)

random.shuffle(buyers)
n_vendas = len(buyers)
n_daiane = int(n_vendas * 0.38)
n_helder = int(n_vendas * 0.31)
n_manu = n_vendas - n_daiane - n_helder

vendas_daiane = buyers[:n_daiane]
vendas_helder = buyers[n_daiane:n_daiane + n_helder]
vendas_manu = buyers[n_daiane + n_helder:]

print(f"  Vendas: DAIANE {len(vendas_daiane)}, HELDER {len(vendas_helder)}, MANU {len(vendas_manu)}")

all_used = set(cnpj for (_, cnpj) in prev_last.keys()) | set(buyers)
prospects_pool = [c for c in cart_data if c not in all_used]
random.shuffle(prospects_pool)
prospects_daiane = prospects_pool[:80]
prospects_helder = prospects_pool[80:140]
prospects_manu = prospects_pool[140:200]
print(f"  Prospecções: DAIANE {len(prospects_daiane)}, HELDER {len(prospects_helder)}, MANU {len(prospects_manu)}")

# SIMULAÇÃO
print(f"\n[5/7] Simulando atendimentos {MONTH_NAME}/25...")
registros_novos = []

def get_info(cnpj_s):
    return cart_data.get(cnpj_s, {'nome': '', 'uf': '', 'rede': '', 'situacao': 'PROSPECT',
                                   'tipo_cliente': 'NOVO', 'ciclo_medio': 0, 'dias_compra': ''})

def simular_consultor(nome, pool_fu, pool_retry, vendas, prospects):
    recs = []
    if pool_fu:
        fu_per_day = max(1, len(pool_fu) // len(du))
        fu_idx = 0
        for dia_idx, d in enumerate(du):
            extra = 1 if dia_idx < len(pool_fu) % len(du) else 0
            batch = pool_fu[fu_idx:fu_idx + fu_per_day + extra]
            fu_idx += len(batch)
            for cnpj_s in batch:
                info = get_info(cnpj_s)
                prev_res = prev_last.get((nome, cnpj_s), {}).get('resultado', 'EM ATENDIMENTO')
                recs.append(gerar_followup(cnpj_s, info, nome, d, prev_res, FERIADOS))
    if pool_retry:
        retry_per_day = max(1, len(pool_retry) // min(10, len(du)))
        retry_idx = 0
        for d_r in du[:10]:
            batch = pool_retry[retry_idx:retry_idx + retry_per_day + 1]
            retry_idx += len(batch)
            if retry_idx >= len(pool_retry):
                retry_idx = len(pool_retry)
            for cnpj_s in batch:
                info = get_info(cnpj_s)
                prev_res = prev_last.get((nome, cnpj_s), {}).get('resultado', 'NÃO RESPONDE')
                recs.append(gerar_followup(cnpj_s, info, nome, d_r, prev_res, FERIADOS))
    du_vendas = du[:15]
    for i, cnpj_s in enumerate(vendas):
        info = get_info(cnpj_s)
        d_base = du_vendas[i % len(du_vendas)]
        key = (nome, cnpj_s)
        in_prev = key in prev_last
        prev_res = prev_last.get(key, {}).get('resultado', '')
        recs.extend(gerar_ciclo_venda(cnpj_s, info, nome, d_base, in_prev, prev_res, du, FERIADOS, MONTH_END))
    if prospects:
        prosp_per_day = max(1, len(prospects) // len(du))
        prosp_idx = 0
        for dia_idx, d in enumerate(du):
            extra = 1 if dia_idx < len(prospects) % len(du) else 0
            batch = prospects[prosp_idx:prosp_idx + prosp_per_day + extra]
            prosp_idx += len(batch)
            for cnpj_s in batch:
                info = get_info(cnpj_s)
                recs.append(gerar_prospeccao(cnpj_s, info, nome, d, FERIADOS))
    return recs

for nome, pool_fu, pool_retry, vendas, prospects in [
    ("DAIANE STADLER", pool_daiane_fu, pool_daiane_retry, vendas_daiane, prospects_daiane),
    ("HELDER BRUNKOW", pool_helder_fu, pool_helder_retry, vendas_helder, prospects_helder),
    ("MANU DITZEL", pool_manu_fu, pool_manu_retry, vendas_manu, prospects_manu),
]:
    recs = simular_consultor(nome, pool_fu, pool_retry, vendas, prospects)
    print(f"  {nome}: {len(recs)} registros ({len(recs)/len(du):.1f}/dia)")
    registros_novos.extend(recs)

print(f"\n  TOTAL {MONTH_NAME}/25: {len(registros_novos)} registros")

# VERIFICAÇÃO
print(f"\n[6/7] Verificação lógica...")
res_count = Counter()
cons_count = Counter()
for rec in registros_novos:
    res_count[rec['R']] += 1
    cons_count[rec['B']] += 1

for res, n in res_count.most_common():
    print(f"    {res:<35} {n:>4}")

n_orc = res_count.get("ORÇAMENTO", 0)
n_cad = res_count.get("CADASTRO", 0)
n_ven = res_count.get("VENDA / PEDIDO", 0)
print(f"\n  Funil: ORC ({n_orc}) >= CAD ({n_cad}) >= VENDA ({n_ven})")
print(f"  {'✅ CONSISTENTE' if n_orc >= n_cad >= n_ven else '⚠️ VERIFICAR'}")

print(f"\n  Por consultor:")
for cons, n in cons_count.most_common():
    print(f"    {cons:<25} {n:>4} ({n/len(du):.1f}/dia)")

n_total = len(registros_novos)
n_motivo = sum(1 for r in registros_novos if r['S'])
n_demanda = sum(1 for r in registros_novos if r['AD'])
n_acao = sum(1 for r in registros_novos if r['U'])
n_nota = sum(1 for r in registros_novos if r['X'])
print(f"\n  Preenchimento: MOTIVO {n_motivo/n_total*100:.0f}% | DEMANDA {n_demanda/n_total*100:.0f}% | AÇÃO {n_acao/n_total*100:.0f}% | NOTA {n_nota/n_total*100:.0f}%")

# SALVAR
print(f"\n[7/7] Salvando V31...")
COLS = ['A','B','C','D','E','F','G','H','I','J','K','L','M',
        'N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
        'AA','AB','AC','AD','AE']

existing_records = []
removed = 0
for r in range(3, ws_d2.max_row + 1):
    data_cell = ws_d2.cell(row=r, column=1).value
    if data_cell is None:
        continue
    is_month = False
    if isinstance(data_cell, (int, float)):
        dt = datetime(1900, 1, 1) + timedelta(days=int(data_cell) - 2)
        if dt.year == 2025 and dt.month == MONTH_NUM:
            is_month = True
    elif isinstance(data_cell, datetime):
        if data_cell.year == 2025 and data_cell.month == MONTH_NUM:
            is_month = True
    if is_month:
        removed += 1
        continue
    row_data = {}
    for c_idx, col_letter in enumerate(COLS):
        row_data[col_letter] = ws_d2.cell(row=r, column=c_idx + 1).value
    existing_records.append(row_data)

print(f"  Mantidos: {len(existing_records)} | Removidos {MONTH_NAME}: {removed} | Novos: {len(registros_novos)}")

all_records = existing_records + registros_novos
def sort_key(rec):
    d = rec.get('A')
    if d is None:
        return datetime(1900, 1, 1)
    if isinstance(d, (int, float)):
        return datetime(1900, 1, 1) + timedelta(days=int(d) - 2)
    if isinstance(d, date) and not isinstance(d, datetime):
        return datetime(d.year, d.month, d.day)
    return d if isinstance(d, datetime) else datetime(1900, 1, 1)

all_records.sort(key=sort_key)
print(f"  Total DRAFT 2: {len(all_records)} registros")

cnpj_counter = Counter()
for rec in all_records:
    cnpj = rec.get('D')
    if cnpj:
        cnpj_counter[str(cnpj).strip()] += 1
        rec['M'] = f"T{cnpj_counter[str(cnpj).strip()]}"

for r in range(3, ws_d2.max_row + 1):
    for c in range(1, len(COLS) + 1):
        ws_d2.cell(row=r, column=c).value = None

for i, rec in enumerate(all_records):
    row = i + 3
    for c_idx, col_letter in enumerate(COLS):
        ws_d2.cell(row=row, column=c_idx + 1).value = rec.get(col_letter)

wb.save(OUTPUT_FILE)
file_size = os.path.getsize(OUTPUT_FILE) / (1024 * 1024)

print(f"\n{'='*80}")
print(f"V31 CONCLUÍDO! — {MONTH_NAME}/25")
print(f"{'='*80}")
print(f"  Arquivo: {OUTPUT_FILE} ({file_size:.2f} MB)")
print(f"  DRAFT 2: {len(all_records)} registros")
print(f"  {MONTH_NAME}/25: {len(registros_novos)} novos (removidos {removed} antigos)")
print(f"  Vendas: {n_ven} de {len(buyers)} esperadas")
print(f"  Média total/dia: {len(registros_novos)/len(du):.1f}")
for cons, n in cons_count.most_common():
    print(f"    {cons:<25} {n:>4} ({n/len(du):.1f}/dia)")
print(f"{'='*80}")

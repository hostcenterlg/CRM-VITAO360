#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VALIDAÇÃO V43 FINAL — Checklist completo antes da apresentação
"""

import openpyxl
from datetime import datetime
import os, re

BASE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360'
V43 = os.path.join(BASE, 'data', 'output', 'phase10', 'CRM_VITAO360_V43_FINAL.xlsx')

print("=" * 80)
print("VALIDAÇÃO V43 FINAL — CHECKLIST PRÉ-APRESENTAÇÃO")
print("=" * 80)

wb = openpyxl.load_workbook(V43, data_only=True)
errors = []
warnings = []
ok = []

# ============================================================
# 1. ABAS EXISTEM
# ============================================================
expected_tabs = ['DASH', 'SINALEIRO', 'PROJEÇÃO ', 'REDES_FRANQUIAS_v2', 'COMITE',
                 'REGRAS', 'DRAFT 1', 'DRAFT 2', 'DRAFT 3 ', 'CARTEIRA',
                 'AGENDA LARISSA', 'AGENDA DAIANE', 'AGENDA MANU', 'AGENDA JULIO']
for tab in expected_tabs:
    if tab in wb.sheetnames:
        ok.append(f"Aba '{tab.strip()}' existe")
    else:
        errors.append(f"Aba '{tab}' NÃO ENCONTRADA!")

# ============================================================
# 2. CARTEIRA — tem CNPJs válidos
# ============================================================
ws = wb['CARTEIRA']
cnpj_count = 0
for r in range(4, ws.max_row + 1):
    v = ws.cell(row=r, column=2).value
    if v:
        s = re.sub(r'[^0-9]', '', str(v))
        if len(s) >= 11:
            cnpj_count += 1
if cnpj_count > 400:
    ok.append(f"CARTEIRA: {cnpj_count} CNPJs válidos")
else:
    errors.append(f"CARTEIRA: apenas {cnpj_count} CNPJs (esperado 400+)")

# ============================================================
# 3. SINALEIRO — tem dados
# ============================================================
ws = wb['SINALEIRO']
sin_rows = 0
for r in range(5, ws.max_row + 1):
    if ws.cell(row=r, column=1).value:
        sin_rows += 1
if sin_rows > 500:
    ok.append(f"SINALEIRO: {sin_rows} clientes")
else:
    warnings.append(f"SINALEIRO: apenas {sin_rows} clientes")

# ============================================================
# 4. DRAFT 2 — DEVE ESTAR VAZIO (limpo)
# ============================================================
ws = wb['DRAFT 2']
d2_data = 0
for r in range(3, min(ws.max_row + 1, 100)):
    if ws.cell(row=r, column=1).value:
        d2_data += 1
if d2_data == 0:
    ok.append(f"DRAFT 2: LIMPO (0 registros)")
else:
    warnings.append(f"DRAFT 2: tem {d2_data} registros (deveria estar vazio)")

# ============================================================
# 5. PROJEÇÃO — tem metas
# ============================================================
ws = wb['PROJEÇÃO ']
meta_count = 0
for r in range(4, ws.max_row + 1):
    v = ws.cell(row=r, column=12).value  # META ANUAL
    if v and isinstance(v, (int, float)) and v > 0:
        meta_count += 1
if meta_count > 400:
    ok.append(f"PROJEÇÃO: {meta_count} clientes com meta anual")
else:
    warnings.append(f"PROJEÇÃO: apenas {meta_count} com meta")

# Metas mensais preenchidas?
meta_jan = 0
for r in range(4, ws.max_row + 1):
    v = ws.cell(row=r, column=13).value  # META JAN
    if v and isinstance(v, (int, float)) and v > 0:
        meta_jan += 1
ok.append(f"PROJEÇÃO metas JAN: {meta_jan} clientes")

# Realizado zerado?
real_count = 0
for r in range(4, ws.max_row + 1):
    v = ws.cell(row=r, column=27).value  # REAL JAN
    if v and isinstance(v, (int, float)) and v > 0:
        real_count += 1
if real_count == 0:
    ok.append(f"PROJEÇÃO realizado: ZERADO (pronto para uso)")
else:
    warnings.append(f"PROJEÇÃO: {real_count} cells de realizado não zeradas")

# ============================================================
# 6. AGENDAS — 10 clientes cada
# ============================================================
for agenda_name in ['AGENDA LARISSA', 'AGENDA DAIANE', 'AGENDA MANU', 'AGENDA JULIO']:
    ws = wb[agenda_name]
    count = 0
    for r in range(4, ws.max_row + 1):
        if ws.cell(row=r, column=2).value:  # CNPJ
            count += 1
    if count == 10:
        ok.append(f"{agenda_name}: {count} clientes ✓")
    elif count > 0:
        warnings.append(f"{agenda_name}: {count} clientes (esperado 10)")
    else:
        errors.append(f"{agenda_name}: VAZIA!")

# ============================================================
# 7. REGRAS — tem conteúdo
# ============================================================
ws = wb['REGRAS']
if ws.max_row > 100:
    ok.append(f"REGRAS: {ws.max_row} rows (source of truth)")
else:
    errors.append(f"REGRAS: apenas {ws.max_row} rows")

# ============================================================
# 8. DRAFT 1 — Mercos intacto
# ============================================================
ws = wb['DRAFT 1']
d1_count = 0
for r in range(4, min(ws.max_row + 1, 100)):
    if ws.cell(row=r, column=2).value:  # CNPJ
        d1_count += 1
if d1_count > 50:
    ok.append(f"DRAFT 1: dados Mercos intactos ({ws.max_row} rows)")
else:
    warnings.append(f"DRAFT 1: poucos dados ({d1_count})")

# ============================================================
# 9. CHECK #REF! em qualquer aba
# ============================================================
ref_errors = 0
for name in wb.sheetnames:
    ws = wb[name]
    for r in range(1, min(ws.max_row + 1, 50)):
        for c in range(1, min(ws.max_column + 1, 50)):
            v = ws.cell(row=r, column=c).value
            if v and '#REF!' in str(v):
                ref_errors += 1
if ref_errors == 0:
    ok.append(f"ZERO #REF! nas primeiras 50 rows de cada aba")
else:
    errors.append(f"{ref_errors} #REF! encontrados!")

# ============================================================
# RESULTADO
# ============================================================
wb.close()

print(f"\n{'='*80}")
print(f"RESULTADO DA VALIDAÇÃO")
print(f"{'='*80}")

print(f"\n✅ OK ({len(ok)}):")
for item in ok:
    print(f"  ✅ {item}")

if warnings:
    print(f"\n⚠️ AVISOS ({len(warnings)}):")
    for item in warnings:
        print(f"  ⚠️ {item}")

if errors:
    print(f"\n❌ ERROS ({len(errors)}):")
    for item in errors:
        print(f"  ❌ {item}")

print(f"\n{'='*80}")
if not errors:
    print(f"🎯 V43 APROVADA PARA APRESENTAÇÃO!")
    print(f"   Arquivo: {V43}")
    print(f"   Tamanho: {os.path.getsize(V43)/(1024*1024):.1f} MB")
else:
    print(f"❌ V43 TEM PROBLEMAS — corrigir antes de apresentar")
print(f"{'='*80}")

#!/usr/bin/env python3
"""Comparar versões chave: V13, V21, V31, V38, V42, V43 — qual tem mais substância"""
import openpyxl, os, time

BASE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10'

versions = [
    ('V13', 'CRM_VITAO360_V13_FINAL.xlsx'),
    ('V21', 'CRM_VITAO360_V21_FINAL.xlsx'),
    ('V31', 'CRM_VITAO360_V31_FINAL.xlsx'),
    ('V38', 'CRM_VITAO360_V38_FINAL.xlsx'),
    ('V42', 'CRM_VITAO360_V42_FINAL.xlsx'),
    ('V43', 'CRM_VITAO360_V43_FINAL.xlsx'),
]

print("=" * 100)
print("COMPARAÇÃO DE VERSÕES — QUAL TEM MAIS SUBSTÂNCIA?")
print("=" * 100)

for label, fname in versions:
    path = os.path.join(BASE, fname)
    if not os.path.exists(path):
        print(f"\n{label}: ARQUIVO NÃO ENCONTRADO")
        continue

    sz = os.path.getsize(path) / (1024*1024)
    print(f"\n{'='*80}")
    print(f"{label} ({sz:.1f} MB) — {fname}")
    print(f"{'='*80}")

    try:
        # Load with formulas (not data_only) to count formulas
        wb_f = openpyxl.load_workbook(path, read_only=True, data_only=False)
        wb_d = openpyxl.load_workbook(path, read_only=True, data_only=True)

        print(f"  Abas: {wb_f.sheetnames}")

        for sname in wb_f.sheetnames:
            ws_f = wb_f[sname]
            ws_d = wb_d[sname]

            formulas = 0
            values = 0
            empty = 0
            ref_errors = 0
            data_rows = 0

            for r in range(1, min(ws_f.max_row + 1, 200)):
                row_has_data = False
                for c in range(1, min(ws_f.max_column + 1, 50)):
                    vf = ws_f.cell(row=r, column=c).value
                    vd = ws_d.cell(row=r, column=c).value
                    if vf and str(vf).startswith('='):
                        formulas += 1
                        row_has_data = True
                    elif vf is not None:
                        values += 1
                        row_has_data = True
                    else:
                        empty += 1
                    if vd and '#REF!' in str(vd):
                        ref_errors += 1
                if row_has_data:
                    data_rows += 1

            status = "OK" if ref_errors == 0 else f"⚠️ {ref_errors} #REF!"
            print(f"  {sname:25s} | {ws_f.max_row:>6} rows x {ws_f.max_column:>3} cols | F={formulas:>5} V={values:>5} | {status}")

        wb_f.close()
        wb_d.close()
    except Exception as e:
        print(f"  ERRO: {e}")

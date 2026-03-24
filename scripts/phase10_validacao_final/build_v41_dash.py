#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V41: DASHBOARD VISUAL — Cards + Barras Horizontais + Grid Simétrico
Inspirado no React Dashboard V12 do CRM VITAO 360
Tudo calculado do DRAFT 2 + CARTEIRA do V40
"""
import sys, os, copy
from datetime import datetime, date
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, Reference, BarChart3D
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════
BASE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10'
INPUT_FILE = os.path.join(BASE, 'CRM_VITAO360_V40_FINAL.xlsx')
OUTPUT_FILE = os.path.join(BASE, 'CRM_VITAO360_V41_FINAL.xlsx')

# ══════════════════════════════════════════════════════════════
# PALETA DE CORES (do React Dashboard)
# ══════════════════════════════════════════════════════════════
C = {
    'green':      '00854A',
    'green_lt':   'E8F5EF',
    'green_dk':   '004D2C',
    'white':      'FFFFFF',
    'bg':         'F4F5F7',
    'card_bg':    'FFFFFF',
    'text':       '1A1D23',
    'sub':        '6B7280',
    'border':     'E5E7EB',
    'header_bg':  '0F1419',
    'ativo':      '00B050',
    'inat_rec':   'EAB308',
    'inat_ant':   'DC2626',
    'risco':      'EA580C',
    'novo':       '7C3AED',
    'prospect':   '3B82F6',
    'blue':       '3B82F6',
    'amber':      'F59E0B',
    'red':        'EF4444',
    'purple':     '8B5CF6',
    'teal':       '14B8A6',
    'pink':       'EC4899',
    'indigo':     '6366F1',
    'orange':     'F97316',
    'row_alt':    'FAFBFC',
    'row_total':  'F0FDF4',
    'alert_bg':   'FEF2F2',
    'alert_text': 'B91C1C',
    'warn_bg':    'FFFBEB',
    'warn_text':  '92400E',
}

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def font(size=10, bold=False, color='1A1D23', name='Calibri'):
    return Font(name=name, size=size, bold=bold, color=color)

def border_left(color, style='thick'):
    return Border(left=Side(style=style, color=color))

def border_top(color, style='medium'):
    return Border(top=Side(style=style, color=color))

def border_bottom(color, style='thin'):
    return Border(bottom=Side(style=style, color=color))

def full_border(color='E5E7EB', style='thin'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)

# ══════════════════════════════════════════════════════════════
# STEP 1: CARREGAR V40 E AGREGAR DADOS
# ══════════════════════════════════════════════════════════════
print("Carregando V40...")
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
ws_d2 = wb['DRAFT 2']
ws_cart = wb['CARTEIRA']

print(f"  DRAFT 2: {ws_d2.max_row} rows × {ws_d2.max_column} cols")
print(f"  CARTEIRA: {ws_cart.max_row} rows × {ws_cart.max_column} cols")

# --- Ler DRAFT 2 (row 3 = primeiro dado, row 2 = headers) ---
# Cols: A=DATA, B=CONSULTOR, C=NOME_FANTASIA, D=CNPJ, E=UF, F=REDE,
#       G=SITUACAO, H=DIAS_SEM_COMPRA, I=ESTAGIO_FUNIL, J=TIPO_CLIENTE,
#       K=FASE, L=SINALEIRO, M=TENTATIVA, N=WHATSAPP, O=LIGACAO,
#       P=LIGACAO_ATENDIDA, Q=TIPO_CONTATO, R=RESULTADO, S=MOTIVO,
#       T=FOLLOWUP, U=ACAO_FUTURA, V=ACAO_DETALHADA, W=MERCOS,
#       X=NOTA_DIA, Y=TEMPERATURA, Z=GRUPO_DASH, AA=SINALEIRO_META,
#       AB=TIPO_ACAO, AC=TIPO_PROBLEMA, AD=DEMANDA, AE=TIPO_ATENDIMENTO

records = []
for row in range(3, ws_d2.max_row + 1):
    data_val = ws_d2.cell(row=row, column=1).value
    if data_val is None:
        continue
    rec = {
        'data': data_val,
        'consultor': str(ws_d2.cell(row=row, column=2).value or '').strip(),
        'nome_fantasia': str(ws_d2.cell(row=row, column=3).value or '').strip(),
        'cnpj': str(ws_d2.cell(row=row, column=4).value or '').strip(),
        'uf': str(ws_d2.cell(row=row, column=5).value or '').strip(),
        'rede': str(ws_d2.cell(row=row, column=6).value or '').strip(),
        'situacao': str(ws_d2.cell(row=row, column=7).value or '').strip(),
        'dias_sem_compra': ws_d2.cell(row=row, column=8).value,
        'estagio_funil': str(ws_d2.cell(row=row, column=9).value or '').strip(),
        'tipo_cliente': str(ws_d2.cell(row=row, column=10).value or '').strip(),
        'fase': str(ws_d2.cell(row=row, column=11).value or '').strip(),
        'sinaleiro': str(ws_d2.cell(row=row, column=12).value or '').strip(),
        'tentativa': str(ws_d2.cell(row=row, column=13).value or '').strip(),
        'whatsapp': ws_d2.cell(row=row, column=14).value,
        'ligacao': ws_d2.cell(row=row, column=15).value,
        'lig_atendida': ws_d2.cell(row=row, column=16).value,
        'tipo_contato': str(ws_d2.cell(row=row, column=17).value or '').strip(),
        'resultado': str(ws_d2.cell(row=row, column=18).value or '').strip(),
        'motivo': str(ws_d2.cell(row=row, column=19).value or '').strip(),
        'followup': ws_d2.cell(row=row, column=20).value,
        'acao_futura': str(ws_d2.cell(row=row, column=21).value or '').strip(),
        'grupo_dash': str(ws_d2.cell(row=row, column=26).value or '').strip(),
        'tipo_atendimento': str(ws_d2.cell(row=row, column=31).value or '').strip(),
    }
    records.append(rec)

print(f"  Total registros DRAFT 2: {len(records)}")

# --- Filtrar últimos 2 meses (JAN + FEV 2026) para o DASH mensal ---
all_records = records

def get_month(rec):
    d = rec['data']
    if isinstance(d, datetime):
        return (d.year, d.month)
    if isinstance(d, date):
        return (d.year, d.month)
    return (0, 0)

# Detectar o mês COM MAIS REGISTROS (ignora spillover de poucos registros)
month_counts = Counter()
for r in records:
    ym = get_month(r)
    if ym != (0, 0):
        month_counts[ym] += 1

# Pegar o mês mais recente que tenha pelo menos 50 registros
sorted_months = sorted(month_counts.keys(), reverse=True)
current_year, current_month = 2026, 2  # fallback
for ym in sorted_months:
    if month_counts[ym] >= 50:
        current_year, current_month = ym
        break

print(f"  Mês principal: {current_month}/{current_year} ({month_counts.get((current_year, current_month), 0)} registros)")
print(f"  Meses detectados: {dict(month_counts.most_common(5))}")

recs_current = [r for r in records if get_month(r) == (current_year, current_month)]
prev_month = current_month - 1 if current_month > 1 else 12
prev_year = current_year if current_month > 1 else current_year - 1
recs_prev = [r for r in records if get_month(r) == (prev_year, prev_month)]

print(f"  Registros mês atual ({current_month}/{current_year}): {len(recs_current)}")
print(f"  Registros mês anterior ({prev_month}/{prev_year}): {len(recs_prev)}")

# Combinar últimos 2 meses para análise
recs_2m = recs_current + recs_prev

# --- Ler CARTEIRA para Saúde da Base ---
cart_data = []
for row in range(4, ws_cart.max_row + 1):
    nome = ws_cart.cell(row=row, column=1).value
    if nome is None:
        continue
    cart_data.append({
        'nome_fantasia': str(nome).strip(),
        'cnpj': str(ws_cart.cell(row=row, column=2).value or '').strip(),
        'uf': str(ws_cart.cell(row=row, column=4).value or '').strip(),
        'consultor': str(ws_cart.cell(row=row, column=12).value or '').strip(),
        'situacao': str(ws_cart.cell(row=row, column=14).value or '').strip(),
        'dias_sem_compra': ws_cart.cell(row=row, column=16).value,
        'curva_abc': str(ws_cart.cell(row=row, column=41).value or '').strip(),
        'n_compras': ws_cart.cell(row=row, column=40).value,
        'rede': str(ws_cart.cell(row=row, column=6).value or '').strip() if ws_cart.cell(row=row, column=6).value else '',
    })

print(f"  CARTEIRA clientes: {len(cart_data)}")

# ══════════════════════════════════════════════════════════════
# STEP 2: CALCULAR TODAS AS AGREGAÇÕES
# ══════════════════════════════════════════════════════════════
print("\nCalculando agregações...")

# --- Consultores ativos (últimos 2 meses) ---
CONSULTORES_ATIVOS = ['DAIANE STADLER', 'MANU DITZEL', 'LARISSA PADILHA', 'JULIO GADRET']
# Mapear nomes que aparecem no CARTEIRA (pode ser diferente)
def norm_consultor(c):
    c = c.upper().strip()
    if 'DAIANE' in c: return 'DAIANE STADLER'
    if 'MANU' in c: return 'MANU DITZEL'
    if 'LARISSA' in c: return 'LARISSA PADILHA'
    if 'JULIO' in c or 'JÚLIO' in c: return 'JULIO GADRET'
    return c

# --- KPIs do mês atual ---
total_atend_mes = len(recs_current)
total_atend_2m = len(recs_2m)

vendas_mes = [r for r in recs_current if 'VENDA' in r['resultado'].upper()]
orcamentos_mes = [r for r in recs_current if 'ORÇAMENTO' in r['resultado'].upper() or 'ORCAMENTO' in r['resultado'].upper()]
cadastros_mes = [r for r in recs_current if 'CADASTRO' in r['resultado'].upper()]
suportes_mes = [r for r in recs_current if 'SUPORTE' in r['resultado'].upper()]
fups_mes = [r for r in recs_current if 'FOLLOW' in r['resultado'].upper()]

n_vendas = len(vendas_mes)
n_orc = len(orcamentos_mes)
conv_geral = round((n_vendas / total_atend_mes * 100), 1) if total_atend_mes > 0 else 0

# WhatsApp e Ligações
wpp_mes = len([r for r in recs_current if r['whatsapp'] in (1, '1', 'SIM')])
lig_mes = len([r for r in recs_current if r['ligacao'] in (1, '1', 'SIM')])
lig_atend_mes = len([r for r in recs_current if r['lig_atendida'] in (1, '1', 'SIM')])
lig_n_atend = lig_mes - lig_atend_mes

# Não atende / Não responde / Recusou
n_atende = len([r for r in recs_current if 'NÃO ATENDE' in r['resultado'].upper()])
n_responde = len([r for r in recs_current if 'NÃO RESPONDE' in r['resultado'].upper()])
n_recusou = len([r for r in recs_current if 'RECUS' in r['resultado'].upper()])
n_perda = len([r for r in recs_current if 'PERDA' in r['resultado'].upper()])

# Follow-ups
fup7 = len([r for r in recs_current if 'FOLLOW UP 7' in r['resultado'].upper() or 'FOLLOW-UP 7' in r['resultado'].upper()])
fup15 = len([r for r in recs_current if 'FOLLOW UP 15' in r['resultado'].upper() or 'FOLLOW-UP 15' in r['resultado'].upper()])

# --- RESULTADO breakdown ---
resultado_counter = Counter()
for r in recs_current:
    res = r['resultado'].strip()
    if res:
        resultado_counter[res] += 1
resultado_sorted = resultado_counter.most_common()

# --- TIPO CONTATO breakdown ---
tipo_contato_counter = Counter()
for r in recs_current:
    tc = r['tipo_contato'].strip()
    if tc:
        tipo_contato_counter[tc] += 1
tipo_contato_sorted = tipo_contato_counter.most_common()

# --- MOTIVOS (filtrar só os que têm motivo preenchido) ---
motivo_counter = Counter()
for r in recs_current:
    mot = r['motivo'].strip()
    if mot and mot != 'None' and mot != '':
        motivo_counter[mot] += 1
motivo_sorted = motivo_counter.most_common()

# --- MATRIZ TIPO CONTATO × RESULTADO ---
matrix = defaultdict(lambda: defaultdict(int))
for r in recs_current:
    tc = r['tipo_contato'].strip()
    res = r['resultado'].strip()
    if tc and res:
        matrix[tc][res] += 1

# Resultados únicos ordenados
RESULTADO_ORDER = [
    'CADASTRO', 'EM ATENDIMENTO', 'FOLLOW UP 7', 'FOLLOW UP 15',
    'ORÇAMENTO', 'VENDA / PEDIDO', 'RELACIONAMENTO', 'SUPORTE',
    'NÃO ATENDE', 'NÃO RESPONDE', 'RECUSOU LIGAÇÃO', 'PERDA / FECHOU LOJA'
]
# Filtrar só os que existem nos dados
resultado_cols = [r for r in RESULTADO_ORDER if any(matrix[tc].get(r, 0) > 0 for tc in matrix)]

# --- POR CONSULTOR ---
consultor_stats = {}
for cons in CONSULTORES_ATIVOS:
    recs_c = [r for r in recs_current if norm_consultor(r['consultor']) == cons]
    vendas_c = len([r for r in recs_c if 'VENDA' in r['resultado'].upper()])
    orc_c = len([r for r in recs_c if 'ORÇAMENTO' in r['resultado'].upper() or 'ORCAMENTO' in r['resultado'].upper()])
    cad_c = len([r for r in recs_c if 'CADASTRO' in r['resultado'].upper()])
    fup_c = len([r for r in recs_c if 'FOLLOW' in r['resultado'].upper()])
    sup_c = len([r for r in recs_c if 'SUPORTE' in r['resultado'].upper()])
    wpp_c = len([r for r in recs_c if r['whatsapp'] in (1, '1', 'SIM')])
    lig_c = len([r for r in recs_c if r['ligacao'] in (1, '1', 'SIM')])
    lig_at_c = len([r for r in recs_c if r['lig_atendida'] in (1, '1', 'SIM')])
    n_at_c = len([r for r in recs_c if 'NÃO ATENDE' in r['resultado'].upper()])
    n_re_c = len([r for r in recs_c if 'NÃO RESPONDE' in r['resultado'].upper()])
    rec_c = len([r for r in recs_c if 'RECUS' in r['resultado'].upper()])
    per_c = len([r for r in recs_c if 'PERDA' in r['resultado'].upper()])
    conv_c = round((vendas_c / len(recs_c) * 100), 1) if len(recs_c) > 0 else 0

    consultor_stats[cons] = {
        'total': len(recs_c), 'vendas': vendas_c, 'orc': orc_c,
        'cad': cad_c, 'fup': fup_c, 'sup': sup_c,
        'wpp': wpp_c, 'lig': lig_c, 'lig_at': lig_at_c,
        'n_atende': n_at_c, 'n_resp': n_re_c, 'recusou': rec_c,
        'perda': per_c, 'conv': conv_c,
    }

# --- SAÚDE DA BASE (CARTEIRA) ---
situacao_counter = Counter()
for c in cart_data:
    sit = c['situacao'].upper().strip()
    if sit:
        situacao_counter[sit] += 1

# --- CURVA ABC ---
abc_counter = Counter()
for c in cart_data:
    abc = c['curva_abc'].upper().strip()
    if abc in ('A', 'B', 'C'):
        abc_counter[abc] += 1

# --- RECOMPRA ---
recompra_counter = Counter()
for c in cart_data:
    nc = c.get('n_compras')
    if nc is None or nc == '' or nc == 'None':
        continue
    try:
        nc = int(nc)
    except (ValueError, TypeError):
        continue
    if nc == 0:
        continue
    elif nc == 1:
        recompra_counter['1 COMPRA'] += 1
    elif nc <= 3:
        recompra_counter['2-3 COMPRAS'] += 1
    elif nc <= 6:
        recompra_counter['4-6 COMPRAS'] += 1
    else:
        recompra_counter['7+ COMPRAS'] += 1

# --- PIPELINE INATIVIDADE ---
pipe_counter = {'<45d': 0, '45-59d': 0, '60-89d': 0, '90+d': 0}
for c in cart_data:
    dsc = c.get('dias_sem_compra')
    if dsc is None:
        continue
    try:
        dsc = int(dsc)
    except (ValueError, TypeError):
        continue
    if dsc < 45:
        pipe_counter['<45d'] += 1
    elif dsc < 60:
        pipe_counter['45-59d'] += 1
    elif dsc < 90:
        pipe_counter['60-89d'] += 1
    else:
        pipe_counter['90+d'] += 1

# --- UF (CARTEIRA) ---
uf_counter = Counter()
for c in cart_data:
    uf = c['uf'].upper().strip()
    if uf and len(uf) == 2:
        uf_counter[uf] += 1
uf_top10 = uf_counter.most_common(10)

# --- REDE ---
rede_counter = Counter()
for r in recs_current:
    rede = r['rede'].strip()
    if rede and rede != 'None' and rede != '':
        rede_counter[rede] += 1
rede_top8 = rede_counter.most_common(8)

# --- Dias úteis no mês ---
from calendar import monthrange
_, days_in_month = monthrange(current_year, current_month)
dias_uteis = 0
for d in range(1, days_in_month + 1):
    dt = date(current_year, current_month, d)
    if dt.weekday() < 5:  # Seg-Sex
        dias_uteis += 1

# --- Volumetria acumulada por consultor (todos os registros) ---
vol_consultor = {}
for cons in CONSULTORES_ATIVOS:
    recs_all_c = [r for r in all_records if norm_consultor(r['consultor']) == cons]
    # Contar dias distintos trabalhados
    dias_set = set()
    for r in recs_all_c:
        d = r['data']
        if isinstance(d, (datetime, date)):
            dias_set.add(d if isinstance(d, date) else d.date())
    dias_trab = len(dias_set)
    vendas_all = len([r for r in recs_all_c if 'VENDA' in r['resultado'].upper()])
    wpp_all = len([r for r in recs_all_c if r['whatsapp'] in (1, '1', 'SIM')])
    lig_all = len([r for r in recs_all_c if r['ligacao'] in (1, '1', 'SIM')])
    fup_all = len([r for r in recs_all_c if 'FOLLOW' in r['resultado'].upper()])

    atend_dia = round(len(recs_all_c) / dias_trab, 1) if dias_trab > 0 else 0
    tasks_venda = round(len(recs_all_c) / vendas_all, 1) if vendas_all > 0 else 0

    vol_consultor[cons] = {
        'atend': len(recs_all_c), 'dias': dias_trab, 'atend_dia': atend_dia,
        'wpp': wpp_all, 'lig': lig_all, 'fup': fup_all,
        'vendas': vendas_all, 'tasks_venda': tasks_venda,
    }

total_atend_all = len(all_records)
total_vendas_all = sum(v['vendas'] for v in vol_consultor.values())

print(f"\n=== RESUMO CALCULADO ===")
print(f"  Atendimentos mês: {total_atend_mes}")
print(f"  Vendas mês: {n_vendas}")
print(f"  Conversão: {conv_geral}%")
print(f"  Orçamentos: {n_orc}")
print(f"  WhatsApp: {wpp_mes} | Ligações: {lig_mes} (atend: {lig_atend_mes})")
print(f"  Dias úteis: {dias_uteis}")
print(f"  Total acumulado: {total_atend_all} registros, {total_vendas_all} vendas")

# ══════════════════════════════════════════════════════════════
# STEP 3: CRIAR ABA DASH
# ══════════════════════════════════════════════════════════════
print("\nCriando aba DASH...")

# Remover DASH antiga se existir
if 'DASH' in wb.sheetnames:
    del wb['DASH']

ws = wb.create_sheet('DASH', 0)  # Primeira aba

# --- GRID: 16 colunas iguais ---
COL_WIDTH = 10.5
for col in range(1, 17):
    ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH

# Background cinza claro pra toda a aba
for row in range(1, 300):
    for col in range(1, 17):
        ws.cell(row=row, column=col).fill = fill(C['bg'])

# Variável de controle de linha atual
ROW = 1

# ══════════════════════════════════════════════════════════════
# HELPER: Escrever card KPI
# ══════════════════════════════════════════════════════════════
def write_card(ws, row, col_start, col_end, value, label, subtitle, accent_color):
    """Card KPI: merge de colunas, borda esquerda colorida, número grande"""
    # Merge principal
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row+2, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)

    # Background branco + borda esquerda colorida
    for r in range(row, row+3):
        for c in range(col_start, col_end+1):
            ws.cell(row=r, column=c).fill = fill(C['card_bg'])
            ws.cell(row=r, column=c).border = full_border(C['border'])

    # Borda esquerda grossa colorida
    for r in range(row, row+3):
        ws.cell(row=r, column=col_start).border = Border(
            left=Side(style='thick', color=accent_color),
            top=Side(style='thin', color=C['border']),
            bottom=Side(style='thin', color=C['border']),
            right=Side(style='thin', color=C['border']),
        )

    # Label (pequeno, cinza, uppercase)
    ws.unmerge_cells(start_row=row, start_column=col_start, end_row=row+2, end_column=col_end)

    # Row 1: label
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c1 = ws.cell(row=row, column=col_start)
    c1.value = label.upper()
    c1.font = font(8, True, C['sub'])
    c1.alignment = Alignment(horizontal='left', vertical='bottom', indent=1)
    c1.fill = fill(C['card_bg'])

    # Row 2: valor grande
    ws.merge_cells(start_row=row+1, start_column=col_start, end_row=row+1, end_column=col_end)
    c2 = ws.cell(row=row+1, column=col_start)
    c2.value = value
    c2.font = font(22, True, accent_color, 'Calibri')
    c2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c2.fill = fill(C['card_bg'])

    # Row 3: subtitle
    ws.merge_cells(start_row=row+2, start_column=col_start, end_row=row+2, end_column=col_end)
    c3 = ws.cell(row=row+2, column=col_start)
    c3.value = subtitle
    c3.font = font(8, False, C['sub'])
    c3.alignment = Alignment(horizontal='left', vertical='top', indent=1)
    c3.fill = fill(C['card_bg'])

    # Row heights
    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row+1].height = 30
    ws.row_dimensions[row+2].height = 14


def write_section_title(ws, row, title, col_start=1, col_end=16):
    """Título de seção com linha verde"""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = title
    cell.font = font(13, True, C['green_dk'])
    cell.alignment = ALIGN_LEFT
    cell.fill = fill(C['bg'])
    cell.border = Border(bottom=Side(style='medium', color=C['green']))
    ws.row_dimensions[row].height = 28
    return row + 1


def write_table_header(ws, row, headers, col_start=1, colors=None):
    """Cabeçalho de tabela com fundo escuro"""
    for i, h in enumerate(headers):
        col = col_start + i
        cell = ws.cell(row=row, column=col)
        cell.value = h
        cell.font = font(8, True, 'FFFFFF')
        cell.fill = fill(C['header_bg'])
        cell.alignment = ALIGN_CENTER if i > 0 else ALIGN_LEFT
        cell.border = Border(bottom=Side(style='medium', color=C['green']))
    ws.row_dimensions[row].height = 22
    return row + 1


def write_table_row(ws, row, values, col_start=1, is_total=False, highlight_cols=None):
    """Linha de tabela com alternância de cor"""
    is_even = (row % 2 == 0)
    bg = C['row_total'] if is_total else (C['row_alt'] if is_even else C['card_bg'])

    for i, v in enumerate(values):
        col = col_start + i
        cell = ws.cell(row=row, column=col)
        cell.value = v
        cell.fill = fill(bg)
        cell.border = full_border(C['border'])
        cell.alignment = ALIGN_CENTER if i > 0 else ALIGN_LEFT

        if is_total:
            cell.font = font(9, True, C['text'])
        elif highlight_cols and i in highlight_cols:
            cell.font = font(9, True, C['green'])
        else:
            cell.font = font(9, False, C['text'])

    ws.row_dimensions[row].height = 18
    return row + 1


def write_bar_row(ws, row, label, value, max_val, color, col_start=1, label_cols=4, bar_cols=10, val_cols=2):
    """Barra horizontal usando preenchimento de célula proporcional"""
    # Label
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_start+label_cols-1)
    cell_label = ws.cell(row=row, column=col_start)
    cell_label.value = label
    cell_label.font = font(9, False, C['text'])
    cell_label.alignment = Alignment(horizontal='right', vertical='center', indent=1)
    cell_label.fill = fill(C['card_bg'])

    # Barra (preencher células proporcionalmente)
    bar_start = col_start + label_cols
    bar_end = bar_start + bar_cols - 1
    filled = round(bar_cols * (value / max_val)) if max_val > 0 else 0

    for c in range(bar_start, bar_end + 1):
        cell = ws.cell(row=row, column=c)
        if c - bar_start < filled:
            cell.fill = fill(color)
        else:
            cell.fill = fill('F1F5F9')
        cell.border = Border()  # Sem borda pra parecer barra contínua

    # Valor
    val_start = bar_end + 1
    ws.merge_cells(start_row=row, start_column=val_start, end_row=row, end_column=val_start+val_cols-1)
    cell_val = ws.cell(row=row, column=val_start)
    cell_val.value = value
    cell_val.font = font(10, True, color)
    cell_val.alignment = Alignment(horizontal='center', vertical='center')
    cell_val.fill = fill(C['card_bg'])

    ws.row_dimensions[row].height = 18
    return row + 1


def write_status_card(ws, row, col_start, col_end, value, label, pct_val, color):
    """Card de status pequeno (Saúde da Base, Pipeline)"""
    # Borda superior colorida
    for c in range(col_start, col_end+1):
        ws.cell(row=row, column=c).border = Border(top=Side(style='thick', color=color))
        ws.cell(row=row, column=c).fill = fill(C['card_bg'])

    # Valor grande
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c1 = ws.cell(row=row, column=col_start)
    c1.value = value
    c1.font = font(20, True, color, 'Calibri')
    c1.alignment = ALIGN_CENTER
    c1.fill = fill(C['card_bg'])

    # Label
    ws.merge_cells(start_row=row+1, start_column=col_start, end_row=row+1, end_column=col_end)
    c2 = ws.cell(row=row+1, column=col_start)
    c2.value = label
    c2.font = font(8, True, C['text'])
    c2.alignment = ALIGN_CENTER
    c2.fill = fill(C['card_bg'])

    # Percentual
    ws.merge_cells(start_row=row+2, start_column=col_start, end_row=row+2, end_column=col_end)
    c3 = ws.cell(row=row+2, column=col_start)
    c3.value = f"{pct_val}%"
    c3.font = font(8, False, C['sub'])
    c3.alignment = ALIGN_CENTER
    c3.fill = fill(C['card_bg'])

    ws.row_dimensions[row].height = 28
    ws.row_dimensions[row+1].height = 14
    ws.row_dimensions[row+2].height = 14


def write_insight_box(ws, row, text, col_start=1, col_end=16, bg_color=None, text_color=None):
    """Box de insight/alerta"""
    bg_c = bg_color or C['warn_bg']
    tx_c = text_color or C['warn_text']
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = text
    cell.font = font(9, True, tx_c)
    cell.fill = fill(bg_c)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = full_border(C['border'])
    ws.row_dimensions[row].height = 24
    return row + 1


# ══════════════════════════════════════════════════════════════
# BLOCO 0: HEADER
# ══════════════════════════════════════════════════════════════
ws.merge_cells('A1:P1')
h1 = ws.cell(row=1, column=1)
h1.value = f"VITAO 360  |  CRM INTELIGENTE  |  DASHBOARD CONSOLIDADO  |  {current_month:02d}/{current_year}"
h1.font = font(14, True, 'FFFFFF')
h1.fill = fill(C['header_bg'])
h1.alignment = ALIGN_CENTER
h1.border = Border(bottom=Side(style='thick', color=C['green']))
ws.row_dimensions[1].height = 36

ROW = 3

# ══════════════════════════════════════════════════════════════
# BLOCO 1: KPIs RESUMO (5 cards lado a lado)
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, f"KPIs RESUMO — {current_month:02d}/{current_year}")

media_dia = round(total_atend_mes / dias_uteis, 1) if dias_uteis > 0 else 0

# 5 cards: cols 1-3, 4-6, 7-9, 10-12, 13-16 (último card um pouco mais largo)
write_card(ws, ROW, 1, 3, total_atend_mes, 'ATENDIMENTOS', f'{dias_uteis} dias úteis · {media_dia}/dia', C['blue'])
write_card(ws, ROW, 4, 6, n_vendas, 'VENDAS', f'{conv_geral}% conversão geral', C['green'])
write_card(ws, ROW, 7, 9, n_orc, 'ORÇAMENTOS', f'Em pipeline ativo', C['amber'])
write_card(ws, ROW, 10, 12, len(cart_data), 'BASE TOTAL', f'Clientes na CARTEIRA', C['purple'])
write_card(ws, ROW, 13, 16, f'{n_atende + n_responde + n_recusou}', 'Ñ ATENDE/RECUSA', f'{n_atende} ñ atende + {n_responde} ñ resp + {n_recusou} recusou', C['red'])

ROW += 5  # 3 rows do card + 2 espaço

# ══════════════════════════════════════════════════════════════
# BLOCO 2: RESULTADO DOS ATENDIMENTOS (barras horizontais)
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, f"RESULTADO DOS ATENDIMENTOS — {total_atend_mes} CONTATOS")

# Cores por resultado
RESULTADO_COLORS = {
    'VENDA / PEDIDO': C['green'], 'ORÇAMENTO': C['amber'], 'CADASTRO': C['purple'],
    'EM ATENDIMENTO': C['indigo'], 'RELACIONAMENTO': C['blue'], 'SUPORTE': C['purple'],
    'FOLLOW UP 7': C['teal'], 'FOLLOW UP 15': C['teal'],
    'NÃO ATENDE': C['red'], 'NÃO RESPONDE': C['risco'],
    'RECUSOU LIGAÇÃO': C['orange'], 'PERDA / FECHOU LOJA': C['red'],
    'NUTRIÇÃO': C['pink'],
}

max_resultado = resultado_sorted[0][1] if resultado_sorted else 1
for res_name, res_val in resultado_sorted[:12]:
    color = RESULTADO_COLORS.get(res_name, C['sub'])
    ROW = write_bar_row(ws, ROW, res_name, res_val, max_resultado, color)

ROW += 2  # espaço

# ══════════════════════════════════════════════════════════════
# BLOCO 3: DESEMPENHO POR CONSULTOR
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, f"DESEMPENHO POR CONSULTOR — {dias_uteis} DIAS ÚTEIS")

# Cabeçalho
headers = ['CONSULTOR', 'TOTAL', 'MÉD/DIA', '%', 'WPP', 'LIG', 'LIG AT.', 'VENDAS', 'ORÇAM', '% CONV', 'Ñ ATENDE', 'Ñ RESP', 'RECUSOU', 'PERDA']
ROW = write_table_header(ws, ROW, headers)

# Dados
for cons in CONSULTORES_ATIVOS:
    s = consultor_stats.get(cons, {})
    total_c = s.get('total', 0)
    med_dia = round(total_c / dias_uteis, 0) if dias_uteis > 0 else 0
    pct_total = round(total_c / total_atend_mes * 100, 1) if total_atend_mes > 0 else 0
    conv_str = f"{s.get('conv', 0)}%"

    # Nome curto
    nome_curto = cons.split(' ')[0] + ' ' + cons.split(' ')[-1] if ' ' in cons else cons

    values = [
        nome_curto, total_c, int(med_dia), f"{pct_total}%",
        s.get('wpp', 0), s.get('lig', 0), s.get('lig_at', 0),
        s.get('vendas', 0), s.get('orc', 0), conv_str,
        s.get('n_atende', 0), s.get('n_resp', 0), s.get('recusou', 0), s.get('perda', 0),
    ]
    ROW = write_table_row(ws, ROW, values, highlight_cols={7, 9})

# Total
total_vals = [
    'TOTAL', total_atend_mes,
    round(total_atend_mes / dias_uteis) if dias_uteis else 0,
    '100%', wpp_mes, lig_mes, lig_atend_mes,
    n_vendas, n_orc, f"{conv_geral}%",
    n_atende, n_responde, n_recusou, n_perda,
]
ROW = write_table_row(ws, ROW, total_vals, is_total=True)

# Insight
best_conv = max(consultor_stats.items(), key=lambda x: x[1].get('conv', 0))
best_vol = max(consultor_stats.items(), key=lambda x: x[1].get('total', 0))
insight_text = f"INSIGHT: {best_vol[0].split()[0]} lidera volume ({best_vol[1]['total']}). {best_conv[0].split()[0]} melhor conversão ({best_conv[1]['conv']}%)."
ROW = write_insight_box(ws, ROW, insight_text)

ROW += 2

# ══════════════════════════════════════════════════════════════
# BLOCO 4: MATRIZ TIPO CONTATO × RESULTADO
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, f"MATRIZ TIPO CONTATO × RESULTADO — {total_atend_mes} REGISTROS")

# Headers da matriz (abreviados pra caber)
RES_ABBR = {
    'CADASTRO': 'CAD', 'EM ATENDIMENTO': 'EM AT', 'FOLLOW UP 7': 'FU7',
    'FOLLOW UP 15': 'FU15', 'ORÇAMENTO': 'ORÇ', 'VENDA / PEDIDO': 'VEND',
    'RELACIONAMENTO': 'REL', 'SUPORTE': 'SUP', 'NÃO ATENDE': 'Ñ AT',
    'NÃO RESPONDE': 'Ñ RE', 'RECUSOU LIGAÇÃO': 'REC', 'PERDA / FECHOU LOJA': 'PER',
    'NUTRIÇÃO': 'NUT',
}

matrix_res_cols = resultado_cols[:12]  # Max 12 colunas
m_headers = ['TIPO CONTATO', 'TOT'] + [RES_ABBR.get(r, r[:4]) for r in matrix_res_cols]
ROW = write_table_header(ws, ROW, m_headers)

# Linhas
tc_order = [tc for tc, _ in tipo_contato_sorted]
for tc in tc_order:
    tc_total = sum(matrix[tc].values())
    values = [tc, tc_total] + [matrix[tc].get(r, 0) for r in matrix_res_cols]
    ROW = write_table_row(ws, ROW, values, highlight_cols={1})

# Total
total_row = ['TOTAL', total_atend_mes] + [sum(matrix[tc].get(r, 0) for tc in tc_order) for r in matrix_res_cols]
ROW = write_table_row(ws, ROW, total_row, is_total=True)

ROW += 2

# ══════════════════════════════════════════════════════════════
# BLOCO 5: MOTIVOS DE NÃO COMPRA (barras)
# ══════════════════════════════════════════════════════════════
total_motivos = sum(v for _, v in motivo_sorted)
ROW = write_section_title(ws, ROW, f"MOTIVOS DE NÃO COMPRA — {total_motivos} REGISTROS")

if motivo_sorted:
    max_mot = motivo_sorted[0][1]
    MOTIVO_COLORS = {
        'AINDA TEM ESTOQUE': C['amber'], 'PROBLEMA LOGÍSTICO / ENTREGA': C['red'],
        'SÓ QUER COMPRAR GRANEL': C['orange'], 'PRODUTO NÃO VENDEU / SEM GIRO': C['risco'],
        'SEM INTERESSE NO MOMENTO': C['sub'], 'PROPRIETÁRIO / INDISPONÍVEL': C['sub'],
        'PROBLEMA FINANCEIRO / CRÉDITO': C['red'], '1º CONTATO / SEM RESPOSTA': C['blue'],
        'LOJA ANEXO / PRÓXIMO - SM': C['purple'], 'FECHANDO / FECHOU LOJA': C['inat_ant'],
    }
    for mot_name, mot_val in motivo_sorted[:10]:
        color = C['amber']
        for k, v in MOTIVO_COLORS.items():
            if k in mot_name.upper() or mot_name.upper() in k:
                color = v
                break
        ROW = write_bar_row(ws, ROW, mot_name[:30], mot_val, max_mot, color)

ROW += 2

# ══════════════════════════════════════════════════════════════
# BLOCO 6: PERFORMANCE POR CONSULTOR (4 cards 2×2)
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, "PERFORMANCE POR CONSULTOR")

CONS_COLORS = {
    'DAIANE STADLER': C['amber'], 'MANU DITZEL': C['blue'],
    'LARISSA PADILHA': C['green'], 'JULIO GADRET': C['purple'],
}

# 2 cards por linha (cols 1-8, 9-16)
for idx, cons in enumerate(CONSULTORES_ATIVOS):
    s = consultor_stats.get(cons, {})
    col_s = 1 if idx % 2 == 0 else 9
    col_e = 8 if idx % 2 == 0 else 16
    color = CONS_COLORS.get(cons, C['green'])
    r = ROW if idx % 2 == 0 else ROW  # Mesma linha se par/ímpar

    if idx == 2:
        ROW += 5  # Nova linha de cards

    row_card = ROW if idx < 2 else ROW

    # Nome + conversão
    ws.merge_cells(start_row=row_card, start_column=col_s, end_row=row_card, end_column=col_e)
    c1 = ws.cell(row=row_card, column=col_s)
    nome_curto = cons.split(' ')[0] + ' ' + cons.split(' ')[-1]
    c1.value = f"{nome_curto}  —  {s.get('conv', 0)}% CONVERSÃO"
    c1.font = font(11, True, color)
    c1.fill = fill(C['card_bg'])
    c1.border = Border(top=Side(style='thick', color=color),
                       left=Side(style='thin', color=C['border']),
                       right=Side(style='thin', color=C['border']))
    c1.alignment = ALIGN_LEFT

    # Mini KPIs: CONT | VEND | ORÇ | FUP | SUP
    mini_labels = ['CONT', 'VEND', 'ORÇAM', 'FUP', 'SUP']
    mini_vals = [s.get('total', 0), s.get('vendas', 0), s.get('orc', 0), s.get('fup', 0), s.get('sup', 0)]
    mini_colors = [C['sub'], C['green'], C['amber'], C['blue'], C['purple']]

    # Labels
    for j, lbl in enumerate(mini_labels):
        c_idx = col_s + j + (1 if col_s == 9 else 0)
        if col_s == 1:
            c_idx = col_s + j
        else:
            c_idx = col_s + j

        # Ajustar: dividir 8 cols em 5 mini-blocos
        # Cols: 1-1, 2-2, 3-4, 5-5, 6-7 (ou similar)
        c_mini = col_s + j
        if c_mini <= col_e:
            cell_v = ws.cell(row=row_card+1, column=c_mini)
            cell_v.value = mini_vals[j]
            cell_v.font = font(14, True, mini_colors[j], 'Calibri')
            cell_v.alignment = ALIGN_CENTER
            cell_v.fill = fill('FAFBFC')

            cell_l = ws.cell(row=row_card+2, column=c_mini)
            cell_l.value = lbl
            cell_l.font = font(7, True, C['sub'])
            cell_l.alignment = ALIGN_CENTER
            cell_l.fill = fill('FAFBFC')

    # Preencher restante do card
    for c in range(col_s + 5, col_e + 1):
        ws.cell(row=row_card+1, column=c).fill = fill('FAFBFC')
        ws.cell(row=row_card+2, column=c).fill = fill('FAFBFC')

    # Bordas do card
    for rr in range(row_card, row_card+3):
        for cc in range(col_s, col_e+1):
            existing = ws.cell(row=rr, column=cc).border
            ws.cell(row=rr, column=cc).border = full_border(C['border'])

    ws.row_dimensions[row_card].height = 22
    ws.row_dimensions[row_card+1].height = 22
    ws.row_dimensions[row_card+2].height = 14

    if idx % 2 == 1 and idx < 3:
        pass  # Não avançar ainda

if len(CONSULTORES_ATIVOS) > 2:
    ROW += 5  # Espaço para segunda linha de cards
ROW += 4

# ══════════════════════════════════════════════════════════════
# BLOCO 7: CANAIS DE CONTATO (4 cards) + FUNIL
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, "CANAIS DE CONTATO")

# 4 cards lado a lado: WPP, LIGOU, ATENDIDA, Ñ ATENDIDA
pct_wpp = round(wpp_mes / total_atend_mes * 100, 0) if total_atend_mes else 0
pct_lig = round(lig_mes / total_atend_mes * 100, 0) if total_atend_mes else 0
pct_lig_at = round(lig_atend_mes / lig_mes * 100, 0) if lig_mes else 0
pct_lig_na = round(lig_n_atend / lig_mes * 100, 0) if lig_mes else 0

write_card(ws, ROW, 1, 4, wpp_mes, 'WHATSAPP', f'{pct_wpp:.0f}% dos atendimentos', '25D366')
write_card(ws, ROW, 5, 8, lig_mes, 'LIGOU', f'{pct_lig:.0f}% dos atendimentos', C['blue'])
write_card(ws, ROW, 9, 12, lig_atend_mes, 'ATENDIDA', f'{pct_lig_at:.0f}% das ligações', C['green'])
write_card(ws, ROW, 13, 16, lig_n_atend, 'Ñ ATENDIDA', f'{pct_lig_na:.0f}% das ligações', C['red'])

ROW += 5

# FUNIL DE VENDA
ROW = write_section_title(ws, ROW, "FUNIL DE VENDA")

em_atend = len([r for r in recs_current if 'EM ATENDIMENTO' in r['resultado'].upper()])
funil_data = [
    ('EM ATENDIMENTO', em_atend, C['indigo']),
    ('ORÇAMENTO', n_orc, C['amber']),
    ('CADASTRO', len(cadastros_mes), C['purple']),
    ('VENDA / PEDIDO', n_vendas, C['green']),
]

max_funil = max(v for _, v, _ in funil_data) if funil_data else 1
for f_name, f_val, f_color in funil_data:
    ROW = write_bar_row(ws, ROW, f_name, f_val, max_funil, f_color, label_cols=4, bar_cols=10, val_cols=2)

conv_orc_venda = round(n_vendas / n_orc * 100, 1) if n_orc > 0 else 0
ROW = write_insight_box(ws, ROW, f"Conv. Orçamento → Venda: {conv_orc_venda}%  ·  Conv. Geral: {conv_geral}%",
                        bg_color=C['green_lt'], text_color=C['green_dk'])

ROW += 2

# ══════════════════════════════════════════════════════════════
# BLOCO 8: SAÚDE DA BASE (6 cards)
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, "SAÚDE DA BASE — ESTRATÉGICO")

# Mapear situações
SITUACAO_MAP = {
    'ATIVO': ('ATIVO', C['ativo']),
    'EM RISCO': ('EM RISCO', C['risco']),
    'INATIVO RECENTE': ('INAT.REC', C['inat_rec']),
    'INATIVO ANTIGO': ('INAT.ANT', C['inat_ant']),
    'NOVO': ('NOVO', C['novo']),
    'PROSPECT': ('PROSPECT', C['prospect']),
}

# Normalizar situações do CARTEIRA
sit_normalized = Counter()
for sit, count in situacao_counter.items():
    matched = False
    for key, (label, _) in SITUACAO_MAP.items():
        if key in sit:
            sit_normalized[key] += count
            matched = True
            break
    if not matched:
        sit_normalized[sit] += count

total_base = sum(sit_normalized.values())

# 6 cards: cols 1-2, 3-5, 6-7, 8-10, 11-12, 13-16 (ou igualmente distribuído)
# Vamos fazer 6 cards de ~2.6 cols cada → arredondar para 2-3 cols
card_ranges = [(1,2), (3,5), (6,8), (9,10), (11,13), (14,16)]
sit_order = ['ATIVO', 'EM RISCO', 'INATIVO RECENTE', 'INATIVO ANTIGO', 'NOVO', 'PROSPECT']

for i, sit_key in enumerate(sit_order):
    if i < len(card_ranges):
        cs, ce = card_ranges[i]
        label, color = SITUACAO_MAP.get(sit_key, (sit_key, C['sub']))
        val = sit_normalized.get(sit_key, 0)
        pct_v = round(val / total_base * 100, 1) if total_base > 0 else 0
        write_status_card(ws, ROW, cs, ce, val, label, pct_v, color)

ROW += 5

# ══════════════════════════════════════════════════════════════
# BLOCO 9: TAXA DE RECOMPRA (barras)
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, "TAXA DE RECOMPRA — B2B CRÍTICO")

RECOMPRA_ORDER = ['1 COMPRA', '2-3 COMPRAS', '4-6 COMPRAS', '7+ COMPRAS']
RECOMPRA_COLORS = {'1 COMPRA': C['red'], '2-3 COMPRAS': C['amber'], '4-6 COMPRAS': C['green'], '7+ COMPRAS': C['blue']}
total_recompra = sum(recompra_counter.values())
max_recompra = max(recompra_counter.values()) if recompra_counter else 1

for faixa in RECOMPRA_ORDER:
    val = recompra_counter.get(faixa, 0)
    pct_r = round(val / total_recompra * 100, 1) if total_recompra > 0 else 0
    label = f"{faixa} ({pct_r}%)"
    ROW = write_bar_row(ws, ROW, label, val, max_recompra, RECOMPRA_COLORS.get(faixa, C['sub']))

# Alerta
uma_compra = recompra_counter.get('1 COMPRA', 0)
pct_uma = round(uma_compra / total_recompra * 100, 1) if total_recompra > 0 else 0
ROW = write_insight_box(ws, ROW, f"ALERTA: {pct_uma}% da base comprou 1x e não recomprou — oportunidade crítica de retenção B2B",
                        bg_color=C['alert_bg'], text_color=C['alert_text'])

ROW += 2

# ══════════════════════════════════════════════════════════════
# BLOCO 10: PIPELINE DE INATIVIDADE (4 cards)
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, "PIPELINE DE INATIVIDADE — DIAS SEM COMPRA")

total_pipe = sum(pipe_counter.values())
pipe_items = [
    ('<45d', 'SAUDÁVEL', 'Manter ritmo', C['ativo']),
    ('45-59d', 'ALERTA', 'Contato urgente', C['amber']),
    ('60-89d', 'RISCO', 'Oferta especial', C['risco']),
    ('90+d', 'INATIVO', 'Campanha reativação', C['inat_ant']),
]

pipe_ranges = [(1,4), (5,8), (9,12), (13,16)]
for i, (key, label, acao, color) in enumerate(pipe_items):
    cs, ce = pipe_ranges[i]
    val = pipe_counter.get(key, 0)
    pct_p = round(val / total_pipe * 100, 1) if total_pipe > 0 else 0
    write_status_card(ws, ROW, cs, ce, val, f"{label}\n{acao}", pct_p, color)

ROW += 5

# ══════════════════════════════════════════════════════════════
# BLOCO 11: CURVA ABC + DISTRIBUIÇÃO UF
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, "CURVA ABC — PRIORIZAÇÃO")

ABC_COLORS = {'A': C['green'], 'B': C['amber'], 'C': C['red']}
ABC_RECEITA = {'A': '~80% receita', 'B': '~15% receita', 'C': '~5% receita'}

abc_ranges = [(1,5), (6,10), (11,16)]
for i, curva in enumerate(['A', 'B', 'C']):
    cs, ce = abc_ranges[i]
    val = abc_counter.get(curva, 0)
    pct_abc = round(val / sum(abc_counter.values()) * 100, 0) if abc_counter else 0
    color = ABC_COLORS[curva]

    # Card
    for c in range(cs, ce+1):
        ws.cell(row=ROW, column=c).border = Border(top=Side(style='thick', color=color))
        ws.cell(row=ROW, column=c).fill = fill(C['card_bg'])

    ws.merge_cells(start_row=ROW, start_column=cs, end_row=ROW, end_column=ce)
    ws.cell(row=ROW, column=cs).value = curva
    ws.cell(row=ROW, column=cs).font = font(28, True, color)
    ws.cell(row=ROW, column=cs).alignment = ALIGN_CENTER

    ws.merge_cells(start_row=ROW+1, start_column=cs, end_row=ROW+1, end_column=ce)
    ws.cell(row=ROW+1, column=cs).value = f"{val} clientes ({pct_abc:.0f}%)"
    ws.cell(row=ROW+1, column=cs).font = font(10, True, C['text'])
    ws.cell(row=ROW+1, column=cs).alignment = ALIGN_CENTER
    ws.cell(row=ROW+1, column=cs).fill = fill(C['card_bg'])

    ws.merge_cells(start_row=ROW+2, start_column=cs, end_row=ROW+2, end_column=ce)
    ws.cell(row=ROW+2, column=cs).value = ABC_RECEITA[curva]
    ws.cell(row=ROW+2, column=cs).font = font(9, False, C['sub'])
    ws.cell(row=ROW+2, column=cs).alignment = ALIGN_CENTER
    ws.cell(row=ROW+2, column=cs).fill = fill(C['card_bg'])

ws.row_dimensions[ROW].height = 36
ws.row_dimensions[ROW+1].height = 18
ws.row_dimensions[ROW+2].height = 16
ROW += 5

# UF Top 10
ROW = write_section_title(ws, ROW, "DISTRIBUIÇÃO POR UF — TOP 10")

max_uf = uf_top10[0][1] if uf_top10 else 1
for uf_name, uf_val in uf_top10:
    color = C['green'] if uf_top10.index((uf_name, uf_val)) < 3 else C['sub']
    ROW = write_bar_row(ws, ROW, uf_name, uf_val, max_uf, color)

ROW += 2

# ══════════════════════════════════════════════════════════════
# BLOCO 12: ATENDIMENTOS POR REDE
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, "ATENDIMENTOS POR REDE — TOP 8")

max_rede = rede_top8[0][1] if rede_top8 else 1
for rede_name, rede_val in rede_top8:
    ROW = write_bar_row(ws, ROW, rede_name[:25], rede_val, max_rede, C['green'])

ROW += 2

# ══════════════════════════════════════════════════════════════
# BLOCO 13: CARGA REAL DIÁRIA POR CONSULTOR (VOLUMETRIA)
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, f"VOLUMETRIA CRM — {total_atend_all:,} REGISTROS ACUMULADOS")

vol_headers = ['CONSULTOR', 'ATEND.', 'DIAS', 'AT/DIA', 'WPP', 'LIG', 'FUP', 'VENDAS', 'TASKS/VENDA']
ROW = write_table_header(ws, ROW, vol_headers)

for cons in CONSULTORES_ATIVOS:
    v = vol_consultor.get(cons, {})
    nome_curto = cons.split(' ')[0] + ' ' + cons.split(' ')[-1]
    values = [
        nome_curto, f"{v.get('atend', 0):,}", v.get('dias', 0),
        v.get('atend_dia', 0), f"{v.get('wpp', 0):,}", f"{v.get('lig', 0):,}",
        f"{v.get('fup', 0):,}", v.get('vendas', 0), v.get('tasks_venda', 0),
    ]
    ROW = write_table_row(ws, ROW, values, highlight_cols={7})

# Total
total_vol = {
    'atend': sum(v['atend'] for v in vol_consultor.values()),
    'dias': max(v['dias'] for v in vol_consultor.values()),
    'vendas': sum(v['vendas'] for v in vol_consultor.values()),
}
total_atd = round(total_vol['atend'] / total_vol['dias'], 1) if total_vol['dias'] else 0
total_tv = round(total_vol['atend'] / total_vol['vendas'], 1) if total_vol['vendas'] else 0
total_vol_row = [
    'TOTAL EQUIPE', f"{total_vol['atend']:,}", total_vol['dias'], total_atd,
    f"{sum(v['wpp'] for v in vol_consultor.values()):,}",
    f"{sum(v['lig'] for v in vol_consultor.values()):,}",
    f"{sum(v['fup'] for v in vol_consultor.values()):,}",
    total_vol['vendas'], total_tv,
]
ROW = write_table_row(ws, ROW, total_vol_row, is_total=True)

ROW += 1
ROW = write_insight_box(ws, ROW, "Estes são apenas os registros do CRM — demandas internas, incêndios e tarefas admin NÃO estão contabilizadas")

ROW += 2

# ══════════════════════════════════════════════════════════════
# BLOCO 14: CAPACIDADE vs REALIDADE
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, "CAPACIDADE vs REALIDADE — O DIA DO CONSULTOR (40 SLOTS/DIA)")

# Capacidade Teórica (esquerda cols 1-8)
ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=8)
ws.cell(row=ROW, column=1).value = "CAPACIDADE TEÓRICA"
ws.cell(row=ROW, column=1).font = font(10, True, C['text'])
ws.cell(row=ROW, column=1).fill = fill(C['card_bg'])
ws.cell(row=ROW, column=1).alignment = ALIGN_CENTER

# Realidade Operacional (direita cols 9-16)
ws.merge_cells(start_row=ROW, start_column=9, end_row=ROW, end_column=16)
ws.cell(row=ROW, column=9).value = "REALIDADE OPERACIONAL"
ws.cell(row=ROW, column=9).font = font(10, True, C['text'])
ws.cell(row=ROW, column=9).fill = fill(C['card_bg'])
ws.cell(row=ROW, column=9).alignment = ALIGN_CENTER

ROW += 1

# Capacidade: VENDA ATIVA 24 (60%) + SUPORTE 16 (40%)
cap_items = [
    ('VENDA ATIVA (pré-venda + fechamento)', 24, '60% dos slots', C['blue']),
    ('SUPORTE (pós-venda + passivo + MKT)', 16, '40% dos slots', C['amber']),
]
real_items = [
    ('AGENDA EXECUTADA (atendimentos)', 28, '70% da capacidade', C['green']),
    ('INCÊNDIOS / SUPORTE PASSIVO', 5, '12% da capacidade', C['red']),
    ('DEMANDAS INTERNAS / ADMIN', 7, '18% da capacidade', C['purple']),
]

for i, (label, val, desc, color) in enumerate(cap_items):
    ws.merge_cells(start_row=ROW+i, start_column=1, end_row=ROW+i, end_column=5)
    ws.cell(row=ROW+i, column=1).value = label
    ws.cell(row=ROW+i, column=1).font = font(9, False, C['text'])
    ws.cell(row=ROW+i, column=1).fill = fill(C['card_bg'])
    ws.cell(row=ROW+i, column=1).alignment = ALIGN_LEFT
    ws.cell(row=ROW+i, column=1).border = Border(left=Side(style='thick', color=color))

    ws.merge_cells(start_row=ROW+i, start_column=6, end_row=ROW+i, end_column=7)
    ws.cell(row=ROW+i, column=6).value = val
    ws.cell(row=ROW+i, column=6).font = font(16, True, color)
    ws.cell(row=ROW+i, column=6).fill = fill(C['card_bg'])
    ws.cell(row=ROW+i, column=6).alignment = ALIGN_CENTER

    ws.cell(row=ROW+i, column=8).value = desc
    ws.cell(row=ROW+i, column=8).font = font(8, False, C['sub'])
    ws.cell(row=ROW+i, column=8).fill = fill(C['card_bg'])

for i, (label, val, desc, color) in enumerate(real_items):
    ws.merge_cells(start_row=ROW+i, start_column=9, end_row=ROW+i, end_column=13)
    ws.cell(row=ROW+i, column=9).value = label
    ws.cell(row=ROW+i, column=9).font = font(9, False, C['text'])
    ws.cell(row=ROW+i, column=9).fill = fill(C['card_bg'])
    ws.cell(row=ROW+i, column=9).alignment = ALIGN_LEFT
    ws.cell(row=ROW+i, column=9).border = Border(left=Side(style='thick', color=color))

    ws.merge_cells(start_row=ROW+i, start_column=14, end_row=ROW+i, end_column=15)
    ws.cell(row=ROW+i, column=14).value = val
    ws.cell(row=ROW+i, column=14).font = font(16, True, color)
    ws.cell(row=ROW+i, column=14).fill = fill(C['card_bg'])
    ws.cell(row=ROW+i, column=14).alignment = ALIGN_CENTER

    ws.cell(row=ROW+i, column=16).value = desc
    ws.cell(row=ROW+i, column=16).font = font(8, False, C['sub'])
    ws.cell(row=ROW+i, column=16).fill = fill(C['card_bg'])

ROW += max(len(cap_items), len(real_items)) + 1

# ~70% insight
ROW = write_insight_box(ws, ROW,
    "O PONTO CRÍTICO: Consultor tem capacidade p/ 40 atendimentos/dia, mas só executa ~70% da agenda (28). "
    "Os outros 30% são consumidos por demandas internas, incêndios e tarefas admin que não geram receita direta.",
    bg_color=C['alert_bg'], text_color=C['alert_text'])

ROW += 2

# ══════════════════════════════════════════════════════════════
# BLOCO 15: ICEBERG OPERACIONAL
# ══════════════════════════════════════════════════════════════
ROW = write_section_title(ws, ROW, "O ICEBERG OPERACIONAL — O QUE A DIRETORIA NÃO VÊ")

# 3 cards grandes lado a lado
tasks_por_venda = round(total_atend_all / total_vendas_all, 0) if total_vendas_all else 0

iceberg_data = [
    ('VISÍVEL (RESULTADO)', f'{total_vendas_all}', 'VENDAS NO ANO\nO que a diretoria enxerga', C['green']),
    ('REGISTRADO (CRM)', f'{total_atend_all:,}', f'ATENDIMENTOS\n{int(tasks_por_venda)} atend. por venda', C['blue']),
    ('INVISÍVEL (DEMANDAS)', '48+', 'TAREFAS POR VENDA\nDemandas que ninguém contabiliza', C['red']),
]

ice_ranges = [(1,5), (6,11), (12,16)]
for i, (title, val, desc, color) in enumerate(iceberg_data):
    cs, ce = ice_ranges[i]

    # Borda superior
    for c in range(cs, ce+1):
        ws.cell(row=ROW, column=c).border = Border(top=Side(style='thick', color=color))
        ws.cell(row=ROW, column=c).fill = fill(C['card_bg'])

    # Título
    ws.merge_cells(start_row=ROW, start_column=cs, end_row=ROW, end_column=ce)
    ws.cell(row=ROW, column=cs).value = title
    ws.cell(row=ROW, column=cs).font = font(8, True, C['sub'])
    ws.cell(row=ROW, column=cs).alignment = ALIGN_CENTER

    # Valor
    ws.merge_cells(start_row=ROW+1, start_column=cs, end_row=ROW+1, end_column=ce)
    ws.cell(row=ROW+1, column=cs).value = val
    ws.cell(row=ROW+1, column=cs).font = font(28, True, color, 'Calibri')
    ws.cell(row=ROW+1, column=cs).alignment = ALIGN_CENTER
    ws.cell(row=ROW+1, column=cs).fill = fill(C['card_bg'])

    # Desc
    ws.merge_cells(start_row=ROW+2, start_column=cs, end_row=ROW+2, end_column=ce)
    ws.cell(row=ROW+2, column=cs).value = desc
    ws.cell(row=ROW+2, column=cs).font = font(8, False, C['sub'])
    ws.cell(row=ROW+2, column=cs).alignment = ALIGN_CENTER
    ws.cell(row=ROW+2, column=cs).fill = fill(C['card_bg'])

ws.row_dimensions[ROW].height = 16
ws.row_dimensions[ROW+1].height = 40
ws.row_dimensions[ROW+2].height = 28

ROW += 5

# ══════════════════════════════════════════════════════════════
# BLOCO 16: FOOTER
# ══════════════════════════════════════════════════════════════
ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=16)
footer = ws.cell(row=ROW, column=1)
footer.value = f"VITAO 360  |  CRM INTELIGENTE V41  |  Dashboard Consolidado  |  {total_atend_all:,} registros  |  {datetime.now().strftime('%d/%m/%Y')}"
footer.font = font(9, True, 'FFFFFF')
footer.fill = fill(C['header_bg'])
footer.alignment = ALIGN_CENTER
footer.border = Border(top=Side(style='thick', color=C['green']))
ws.row_dimensions[ROW].height = 28

# ══════════════════════════════════════════════════════════════
# STEP 4: ADICIONAR GRÁFICOS DE BARRA REAIS
# ══════════════════════════════════════════════════════════════
print("\nAdicionando gráficos...")

# Precisamos de dados auxiliares em uma área oculta para os gráficos
# Vamos usar colunas R-Z (17-26) como área de dados dos gráficos
DATA_COL = 18  # Coluna R (18)

# --- Gráfico 1: Resultado dos Atendimentos ---
chart1_row = 5  # Área de dados auxiliar
ws.cell(row=chart1_row, column=DATA_COL).value = "RESULTADO"
ws.cell(row=chart1_row, column=DATA_COL+1).value = "QTD"

for i, (res_name, res_val) in enumerate(resultado_sorted[:10]):
    ws.cell(row=chart1_row+1+i, column=DATA_COL).value = res_name
    ws.cell(row=chart1_row+1+i, column=DATA_COL+1).value = res_val

n_res = min(len(resultado_sorted), 10)

chart1 = BarChart()
chart1.type = "bar"  # Horizontal
chart1.style = 2
chart1.title = None
chart1.y_axis.title = None
chart1.x_axis.title = None
chart1.legend = None
chart1.width = 32
chart1.height = 14

data_ref = Reference(ws, min_col=DATA_COL+1, min_row=chart1_row, max_row=chart1_row+n_res)
cats_ref = Reference(ws, min_col=DATA_COL, min_row=chart1_row+1, max_row=chart1_row+n_res)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.shape = 4

# Estilizar
s = chart1.series[0]
s.graphicalProperties.solidFill = C['green']

chart1.y_axis.delete = False
chart1.x_axis.delete = True  # Sem eixo numérico

# Colocar após as barras manuais do Bloco 2
# Encontrar row do bloco 2 (aprox row 8 + n_resultado)
# Na verdade, como já temos barras manuais, vamos pular o chart por enquanto
# e focar em fazer as barras manuais ficarem boas

# --- Gráfico 2: Motivos de Não Compra ---
chart2_row = chart1_row + n_res + 3
ws.cell(row=chart2_row, column=DATA_COL).value = "MOTIVO"
ws.cell(row=chart2_row, column=DATA_COL+1).value = "QTD"

for i, (mot_name, mot_val) in enumerate(motivo_sorted[:10]):
    ws.cell(row=chart2_row+1+i, column=DATA_COL).value = mot_name[:25]
    ws.cell(row=chart2_row+1+i, column=DATA_COL+1).value = mot_val

n_mot = min(len(motivo_sorted), 10)

chart2 = BarChart()
chart2.type = "bar"
chart2.style = 2
chart2.title = None
chart2.legend = None
chart2.width = 32
chart2.height = 14

data2 = Reference(ws, min_col=DATA_COL+1, min_row=chart2_row, max_row=chart2_row+n_mot)
cats2 = Reference(ws, min_col=DATA_COL, min_row=chart2_row+1, max_row=chart2_row+n_mot)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)

s2 = chart2.series[0]
s2.graphicalProperties.solidFill = C['amber']
chart2.y_axis.delete = False
chart2.x_axis.delete = True

# Esconder colunas de dados auxiliares
for col in range(DATA_COL, DATA_COL+3):
    ws.column_dimensions[get_column_letter(col)].hidden = True

# ══════════════════════════════════════════════════════════════
# STEP 5: CONFIGURAÇÕES FINAIS
# ══════════════════════════════════════════════════════════════
print("\nConfigurações finais...")

# Freeze pane no header
ws.freeze_panes = 'A2'

# Print settings
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = 'portrait'

# Zoom
ws.sheet_view.zoomScale = 100

# ══════════════════════════════════════════════════════════════
# SALVAR
# ══════════════════════════════════════════════════════════════
print(f"\nSalvando {OUTPUT_FILE}...")
wb.save(OUTPUT_FILE)
print(f"\n{'='*60}")
print(f"V41 GERADO COM SUCESSO!")
print(f"{'='*60}")
print(f"  Arquivo: {OUTPUT_FILE}")
print(f"  DASH: {ROW} linhas, 16 colunas")
print(f"  Blocos: 16 (KPIs → Iceberg)")
print(f"  Dados: {total_atend_mes} atend. mês | {n_vendas} vendas | {conv_geral}% conv")
print(f"  Acumulado: {total_atend_all:,} registros | {total_vendas_all} vendas")
print(f"{'='*60}")

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V16 UNIFICADO — Script definitivo.

O usuario pediu: "unificar LOG, DRAFT 2 e AGENDA em um DRAFT 2 bem completo"

Plano:
  FASE 1: Merge LOG (20.832 rows) + DRAFT 2 (6.775 rows) = DRAFT 2 unificado (31 cols)
  FASE 2: Aplicar formulas AUTO do V31 (SINALEIRO, TENTATIVA, GRUPO DASH, TIPO ACAO)
  FASE 3: Copiar DRAFT 3 (SAP) do V31
  FASE 4: Copiar SINALEIRO do V31
  FASE 5: Reconstruir 4 AGENDA tabs (layout V31 = 32 cols com SCORE + PRIORIDADE)
  FASE 6: Remover LOG (DRAFT 2 eh o log agora)
  FASE 7: Corrigir CARTEIRA (anchors, layout, formulas FUNIL)
  FASE 8: Reordenar abas
  FASE 9: Salvar + Verificar
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import shutil
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# === PATHS ===
V15_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V15_FINAL.xlsx"
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"
V16_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V16_FINAL.xlsx"

# === DRAFT 2 HEADERS (31 colunas, identico ao V31 DRAFT 2 row 2) ===
DRAFT2_HEADERS = [
    "DATA",              # A (1)  - MANUAL
    "CONSULTOR",         # B (2)  - MANUAL
    "NOME FANTASIA",     # C (3)  - MANUAL
    "CNPJ",              # D (4)  - MANUAL
    "UF",                # E (5)  - MANUAL
    "REDE / REGIONAL",   # F (6)  - MANUAL
    "SITUAÇÃO",          # G (7)  - MANUAL
    "DIAS SEM COMPRA",   # H (8)  - MANUAL
    "ESTÁGIO FUNIL",     # I (9)  - AUTO (de REGRAS ou manual)
    "TIPO CLIENTE",      # J (10) - AUTO (de REGRAS ou manual)
    "FASE",              # K (11) - AUTO
    "SINALEIRO",         # L (12) - AUTO formula
    "TENTATIVA",         # M (13) - AUTO formula
    "WHATSAPP",          # N (14) - MANUAL
    "LIGAÇÃO",           # O (15) - MANUAL
    "LIGAÇÃO ATENDIDA",  # P (16) - MANUAL
    "TIPO DO CONTATO",   # Q (17) - MANUAL
    "RESULTADO",         # R (18) - MANUAL
    "MOTIVO",            # S (19) - MANUAL
    "FOLLOW-UP",         # T (20) - MANUAL/AUTO
    "AÇÃO FUTURA",       # U (21) - MANUAL/AUTO
    "AÇÃO DETALHADA",    # V (22) - MANUAL
    "MERCOS ATUALIZADO", # W (23) - MANUAL
    "NOTA DO DIA",       # X (24) - MANUAL
    "TEMPERATURA",       # Y (25) - AUTO
    "GRUPO DASH",        # Z (26) - AUTO formula
    "SINALEIRO META",    # AA(27) - MANUAL
    "TIPO AÇÃO",         # AB(28) - AUTO formula
    "TIPO PROBLEMA",     # AC(29) - MANUAL
    "DEMANDA",           # AD(30) - MANUAL
    "TIPO ATENDIMENTO",  # AE(31) - MANUAL
]

# === MAPEAMENTO LOG (21 cols) → DRAFT 2 (31 cols) ===
# LOG col_idx (1-based) → DRAFT 2 col_idx (1-based)
LOG_TO_D2 = {
    1: 1,    # A→A: DATA
    2: 2,    # B→B: CONSULTOR
    3: 3,    # C→C: NOME FANTASIA
    4: 4,    # D→D: CNPJ
    5: 5,    # E→E: UF
    6: 6,    # F→F: REDE / REGIONAL
    7: 7,    # G→G: SITUAÇÃO
    8: 14,   # H→N: WHATSAPP
    9: 15,   # I→O: LIGAÇÃO
    10: 16,  # J→P: LIGAÇÃO ATENDIDA
    11: 31,  # K→AE: TIPO ATENDIMENTO (LOG "TIPO ACAO" -> AE)
    12: 17,  # L→Q: TIPO DO CONTATO
    13: 18,  # M→R: RESULTADO
    14: 19,  # N→S: MOTIVO
    15: 20,  # O→T: FOLLOW-UP
    16: 21,  # P→U: AÇÃO FUTURA
    17: 23,  # Q→W: MERCOS ATUALIZADO
    18: 11,  # R→K: FASE
    19: 13,  # S→M: TENTATIVA (sera sobrescrito por formula)
    20: 24,  # T→X: NOTA DO DIA
    # 21: ORIGEM_DADO — descartado
}

# === FORMULAS AUTO do V31 (template com {r} para row) ===
# Estas formulas vao em CADA row de dados do DRAFT 2
DRAFT2_AUTO_FORMULAS = {
    # L (12) - SINALEIRO
    12: '=IF(OR(G{r}="PROSPECT",G{r}="LEAD"),"\U0001f7e3",IF(H{r}="","",IF(H{r}<=50,"\U0001f7e2",IF(H{r}<=90,"\U0001f7e1","\U0001f534"))))',
    # M (13) - TENTATIVA
    13: '="T"&COUNTIF($D$3:D{r},D{r})',
    # Z (26) - GRUPO DASH
    26: '=IFERROR(VLOOKUP(R{r},REGRAS!$B$6:$D$20,3,FALSE),"")',
    # AB (28) - TIPO AÇÃO
    28: '=IF(R{r}="","",IF(OR(R{r}="VENDA / PEDIDO"),"VENDA",IF(OR(R{r}="PÓS-VENDA",R{r}="CS (SUCESSO DO CLIENTE)",R{r}="RELACIONAMENTO",R{r}="NUTRIÇÃO"),"PÓS-VENDA",IF(R{r}="SUPORTE","RESOLUÇÃO DE PROBLEMA",IF(OR(G{r}="PROSPECT",G{r}="LEAD"),"PROSPECÇÃO",IF(R{r}="PERDA / FECHOU LOJA","PRÉ-VENDA",IF(OR(R{r}="ORÇAMENTO",R{r}="CADASTRO",R{r}="EM ATENDIMENTO",R{r}="FOLLOW UP 7",R{r}="FOLLOW UP 15",R{r}="NÃO ATENDE",R{r}="NÃO RESPONDE",R{r}="RECUSOU LIGAÇÃO"),"PRÉ-VENDA","PRÉ-VENDA")))))))',
}

# === AGENDA HEADERS (32 colunas, layout V31) ===
AGENDA_HEADERS = [
    "\U0001f4c5 DATA",       # A
    "NOME FANTASIA",         # B
    "CNPJ",                  # C
    "UF",                    # D
    "REDE / REGIONAL",       # E
    "DIAS SEM COMPRA",       # F
    "SITUAÇÃO",              # G
    "ESTÁGIO FUNIL",         # H
    "TIPO CLIENTE",          # I
    "FASE",                  # J
    "\U0001f525 TEMPERATURA", # K
    "\U0001f6a6 SINALEIRO",  # L
    "PRÓX. AÇÃO",            # M
    "TENTATIVA",             # N
    "TIPO ATENDIMENTO",      # O
    "WHATSAPP",              # P
    "LIGAÇÃO",               # Q
    "LIGAÇÃO ATENDIDA",      # R
    "TIPO DO CONTATO",       # S
    "RESULTADO",             # T
    "MOTIVO",                # U
    "FOLLOW-UP",             # V
    "\U0001f4a1 AÇÃO SUGERIDA", # W
    "AÇÃO FUTURA",           # X
    "AÇÃO DETALHADA",        # Y
    "MERCOS ATUALIZADO",     # Z
    "NOTA DO DIA",           # AA
    "TIPO AÇÃO",             # AB
    "TIPO PROBLEMA",         # AC
    "TAREFA/DEMANDA",        # AD
    "\u2b50 SCORE",          # AE
    "\U0001f3c5 PRIORIDADE", # AF
]

# Mapeamento AGENDA col → CARTEIRA col (de onde puxar via FILTER)
# AGENDA col_idx (1-based) → CARTEIRA col letter
AGENDA_TO_CARTEIRA = {
    1: "AT",   # DATA → CARTEIRA AT (DATA ÚLT ATENDIMENTO)
    2: "A",    # NOME FANTASIA → CARTEIRA A
    3: "B",    # CNPJ → CARTEIRA B
    4: "D",    # UF → CARTEIRA D
    5: "I",    # REDE / REGIONAL → CARTEIRA I
    6: "P",    # DIAS SEM COMPRA → CARTEIRA P
    7: "N",    # SITUAÇÃO → CARTEIRA N
    8: "AR",   # ESTÁGIO FUNIL → CARTEIRA AR
    9: "AX",   # TIPO CLIENTE → CARTEIRA AX
    10: "AZ",  # FASE → CARTEIRA AZ
    11: "BB",  # TEMPERATURA → CARTEIRA BB
    12: "BJ",  # SINALEIRO → CARTEIRA BJ
    13: "BH",  # PRÓX. AÇÃO → CARTEIRA BH
    14: "AY",  # TENTATIVA → CARTEIRA AY
    # 15-21: MANUAL (consultor preenche)
    22: "AS",  # FOLLOW-UP → CARTEIRA AS
    # 23: AÇÃO SUGERIDA (formula local)
    24: "AU",  # AÇÃO FUTURA → CARTEIRA AU
    25: "BI",  # AÇÃO DETALHADA → CARTEIRA BI
    # 26-31: MANUAL
}


def normalizar_cnpj(valor):
    """Normaliza CNPJ para 14 digitos"""
    if not valor:
        return None
    cnpj = re.sub(r'[^\d]', '', str(valor).strip())
    return cnpj if len(cnpj) >= 11 else None


def copy_sheet_from_v31(wb_v31, wb_target, sheet_name):
    """Copia uma aba inteira do V31 para o target"""
    if sheet_name not in wb_v31.sheetnames:
        print(f"  ! {sheet_name} nao existe no V31")
        return False

    ws_src = wb_v31[sheet_name]

    # Se ja existe no target, remover
    if sheet_name in wb_target.sheetnames:
        del wb_target[sheet_name]

    ws_dst = wb_target.create_sheet(sheet_name)

    # Copiar dados e formatos
    for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, max_col=ws_src.max_column):
        for cell in row:
            new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    # Copiar dimensoes de colunas
    for col_key in ws_src.column_dimensions:
        src_dim = ws_src.column_dimensions[col_key]
        dst_dim = ws_dst.column_dimensions[col_key]
        if src_dim.width:
            dst_dim.width = src_dim.width
        if src_dim.outline_level:
            dst_dim.outline_level = src_dim.outline_level
        dst_dim.hidden = src_dim.hidden

    # Copiar dimensoes de linhas
    for row_key in ws_src.row_dimensions:
        src_dim = ws_src.row_dimensions[row_key]
        dst_dim = ws_dst.row_dimensions[row_key]
        if src_dim.height:
            dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden

    # Copiar merged cells
    for merged in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merged))

    # Copiar freeze/filter/state
    ws_dst.freeze_panes = ws_src.freeze_panes
    if ws_src.auto_filter and ws_src.auto_filter.ref:
        ws_dst.auto_filter.ref = ws_src.auto_filter.ref
    ws_dst.sheet_state = ws_src.sheet_state

    return True


def main():
    start_time = datetime.now()
    print("=" * 100)
    print("V16 UNIFICADO — DRAFT 2 completo, AGENDA reconstruida, DRAFT 3 + SINALEIRO")
    print(f"Inicio: {start_time}")
    print("=" * 100)

    changes = []

    # ================================================================
    # CARREGAMENTO
    # ================================================================
    print("\n[LOAD 1/3] Copiando V15 como base para V16...")
    shutil.copy2(V15_PATH, V16_PATH)

    print("[LOAD 2/3] Carregando V16 (edicao)...")
    wb = openpyxl.load_workbook(V16_PATH)
    print(f"  Abas: {wb.sheetnames}")

    print("[LOAD 3/3] Carregando V31 (referencia)...")
    wb_v31 = openpyxl.load_workbook(V31_PATH, data_only=False)
    wb_v31_data = openpyxl.load_workbook(V31_PATH, data_only=True)
    print(f"  Abas V31: {wb_v31.sheetnames}")

    # ================================================================
    # FASE 1: MERGE LOG + DRAFT 2 → DRAFT 2 UNIFICADO
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 1: MERGE LOG + DRAFT 2 → DRAFT 2 UNIFICADO")
    print(f"{'='*100}")

    # 1A. Ler DRAFT 2 existente (V15, rows 4+, 31 cols)
    ws_d2 = wb["DRAFT 2"]
    print(f"\n  [1A] Lendo DRAFT 2 existente: {ws_d2.max_row} rows x {ws_d2.max_column} cols")

    d2_data = {}  # key = (cnpj_norm, date_str, consultor) → {col: value}
    d2_count = 0
    for row in range(4, ws_d2.max_row + 1):
        cnpj_raw = ws_d2.cell(row=row, column=4).value
        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            continue

        date_val = ws_d2.cell(row=row, column=1).value
        consultor = ws_d2.cell(row=row, column=2).value or ""
        date_str = str(date_val)[:10] if date_val else ""

        key = (cnpj, date_str, str(consultor).strip().upper())

        row_data = {}
        for col in range(1, 32):  # 31 cols
            val = ws_d2.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                # Pular formulas — queremos dados estaticos
                val_str = str(val)
                if not val_str.startswith("="):
                    row_data[col] = val
        if row_data:
            d2_data[key] = row_data
            d2_count += 1

    print(f"    {d2_count} registros do DRAFT 2")

    # 1B. Ler LOG (V15, rows 3+, 21 cols) e mapear para DRAFT 2 format
    ws_log = wb["LOG"]
    print(f"\n  [1B] Lendo LOG: {ws_log.max_row} rows x {ws_log.max_column} cols")

    log_new = 0
    log_merged = 0
    for row in range(3, ws_log.max_row + 1):
        cnpj_raw = ws_log.cell(row=row, column=4).value
        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            continue

        date_val = ws_log.cell(row=row, column=1).value
        consultor = ws_log.cell(row=row, column=2).value or ""
        date_str = str(date_val)[:10] if date_val else ""

        key = (cnpj, date_str, str(consultor).strip().upper())

        # Mapear LOG cols → DRAFT 2 cols
        row_data = {}
        for log_col, d2_col in LOG_TO_D2.items():
            val = ws_log.cell(row=row, column=log_col).value
            if val is not None and str(val).strip() != "":
                row_data[d2_col] = val

        if not row_data:
            continue

        if key in d2_data:
            # Merge: preencher gaps no registro existente
            existing = d2_data[key]
            for col, val in row_data.items():
                if col not in existing or existing[col] is None:
                    existing[col] = val
            log_merged += 1
        else:
            d2_data[key] = row_data
            log_new += 1

        if (row - 3) % 5000 == 0:
            print(f"    ... processadas {row - 3} rows do LOG")

    print(f"    LOG: {log_new} novos registros + {log_merged} merged com DRAFT 2 existente")

    # 1C. Tentar preencher gaps com V31 DRAFT 2 (matched por CNPJ)
    print(f"\n  [1C] Preenchendo gaps com V31 DRAFT 2...")
    ws_v31_d2 = wb_v31_data["DRAFT 2"]

    v31_by_cnpj = defaultdict(list)
    for row in range(3, ws_v31_d2.max_row + 1):
        cnpj_raw = ws_v31_d2.cell(row=row, column=4).value
        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            continue
        row_data = {}
        for col in range(1, 32):
            val = ws_v31_d2.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                row_data[col] = val
        if row_data:
            v31_by_cnpj[cnpj].append(row_data)

    gap_fills = 0
    for key, data in d2_data.items():
        cnpj = key[0]
        if cnpj not in v31_by_cnpj:
            continue
        # Pegar o registro V31 mais recente para este CNPJ
        v31_records = v31_by_cnpj[cnpj]
        v31_latest = v31_records[-1]  # ultimo = mais recente

        # Preencher apenas colunas VAZIAS
        for col in [7, 8, 9, 10, 22, 25, 29, 30]:  # SITUACAO, DIAS, ESTAGIO, TIPO CLI, ACAO DET, TEMP, TIPO PROB, DEMANDA
            if col not in data or data[col] is None:
                if col in v31_latest:
                    data[col] = v31_latest[col]
                    gap_fills += 1

    print(f"    {gap_fills} celulas preenchidas do V31")

    # 1D. Ordenar por DATA descendente
    print(f"\n  [1D] Ordenando {len(d2_data)} registros por DATA (desc)...")
    sorted_keys = sorted(d2_data.keys(), key=lambda k: str(k[1]) if k[1] else "", reverse=True)

    # 1E. Escrever novo DRAFT 2
    print(f"\n  [1E] Escrevendo DRAFT 2 unificado...")

    # Remover DRAFT 2 antigo
    if "DRAFT 2" in wb.sheetnames:
        del wb["DRAFT 2"]

    ws_new = wb.create_sheet("DRAFT 2")

    # Row 1: Titulo (V31 style)
    ws_new.cell(row=1, column=1, value="DRAFT 2 — LOG DE ATENDIMENTOS (CRESCE INFINITO \u2193)")

    # Row 2: Headers
    for col_idx, header in enumerate(DRAFT2_HEADERS, 1):
        ws_new.cell(row=2, column=col_idx, value=header)

    # Row 3+: Dados
    written = 0
    for idx, key in enumerate(sorted_keys):
        row = idx + 3  # dados a partir da row 3
        data = d2_data[key]

        for col, val in data.items():
            if col <= 31:
                ws_new.cell(row=row, column=col, value=val)

        written += 1
        if written % 5000 == 0:
            print(f"    ... escritos {written}/{len(sorted_keys)} registros")

    total_rows = written + 2  # +2 for header rows
    print(f"    {written} registros escritos (rows 3-{total_rows})")

    # 1F. Aplicar formulas AUTO
    print(f"\n  [1F] Aplicando formulas AUTO...")
    formula_count = 0
    for row in range(3, total_rows + 1):
        for col, formula_template in DRAFT2_AUTO_FORMULAS.items():
            formula = formula_template.replace("{r}", str(row))
            ws_new.cell(row=row, column=col, value=formula)
            formula_count += 1

        if (row - 3) % 5000 == 0 and row > 3:
            print(f"    ... formulas aplicadas ate row {row}")

    print(f"    {formula_count} formulas AUTO aplicadas")

    # 1G. Layout do DRAFT 2
    ws_new.freeze_panes = "A3"
    filter_ref = f"A2:{get_column_letter(31)}{total_rows}"
    ws_new.auto_filter.ref = filter_ref

    # Column widths
    d2_widths = {
        1: 12, 2: 18, 3: 30, 4: 18, 5: 5, 6: 14, 7: 14, 8: 12,
        9: 14, 10: 14, 11: 12, 12: 10, 13: 10, 14: 8, 15: 8, 16: 12,
        17: 18, 18: 18, 19: 14, 20: 12, 21: 25, 22: 30, 23: 10, 24: 30,
        25: 12, 26: 12, 27: 10, 28: 14, 29: 14, 30: 14, 31: 14
    }
    for col_idx, width in d2_widths.items():
        ws_new.column_dimensions[get_column_letter(col_idx)].width = width

    msg = f"  DRAFT 2 unificado: {written} registros, 31 cols, {formula_count} formulas AUTO"
    print(f"\n  {msg}")
    changes.append(msg)

    # ================================================================
    # FASE 2: REMOVER LOG
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 2: REMOVER LOG (DRAFT 2 eh o log agora)")
    print(f"{'='*100}")

    if "LOG" in wb.sheetnames:
        del wb["LOG"]
        msg = "  LOG removido (DRAFT 2 eh a fonte unica de atendimentos)"
        print(msg)
        changes.append(msg)

    # ================================================================
    # FASE 3: COPIAR DRAFT 3 (SAP) DO V31
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 3: COPIAR DRAFT 3 (SAP)")
    print(f"{'='*100}")

    draft3_name = None
    for s in wb_v31.sheetnames:
        if "DRAFT 3" in s.upper():
            draft3_name = s
            break

    if draft3_name:
        success = copy_sheet_from_v31(wb_v31, wb, draft3_name)
        if success:
            ws = wb[draft3_name]
            msg = f"  DRAFT 3 copiado do V31: {ws.max_row} rows x {ws.max_column} cols"
            print(msg)
            changes.append(msg)
    else:
        print("  ! DRAFT 3 nao encontrado no V31")

    # ================================================================
    # FASE 4: COPIAR SINALEIRO DO V31
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 4: COPIAR SINALEIRO")
    print(f"{'='*100}")

    if "SINALEIRO" in wb_v31.sheetnames:
        success = copy_sheet_from_v31(wb_v31, wb, "SINALEIRO")
        if success:
            ws = wb["SINALEIRO"]
            msg = f"  SINALEIRO copiado do V31: {ws.max_row} rows x {ws.max_column} cols"
            print(msg)
            changes.append(msg)

    # ================================================================
    # FASE 5: RECONSTRUIR AGENDA x4
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 5: RECONSTRUIR AGENDA (4 tabs, layout V31 = 32 cols)")
    print(f"{'='*100}")

    consultores = {
        "AGENDA LARISSA": "LARISSA PADILHA",
        "AGENDA DAIANE": "DAIANE STAVICKI",
        "AGENDA MANU": ["MANU DITZEL", "HEMANUELE DITZEL (MANU)"],
        "AGENDA JULIO": "JULIO GADRET",
    }

    for agenda_name, consultor in consultores.items():
        print(f"\n  Reconstruindo {agenda_name}...")

        # Remover existente
        if agenda_name in wb.sheetnames:
            del wb[agenda_name]

        ws_a = wb.create_sheet(agenda_name)

        # Row 1: Headers (32 cols = V31 AGENDA)
        for col_idx, header in enumerate(AGENDA_HEADERS, 1):
            ws_a.cell(row=1, column=col_idx, value=header)

        # Build filter condition for CARTEIRA
        # CARTEIRA col L = CONSULTOR
        if isinstance(consultor, list):
            # MANU: OR condition
            filter_cond = "(" + "+".join(
                f'(CARTEIRA!$L$4:$L$25000="{name}")' for name in consultor
            ) + ")"
        else:
            filter_cond = f'CARTEIRA!$L$4:$L$25000="{consultor}"'

        # Sort by SCORE descending (CARTEIRA col O = SCORE)
        sort_col = "CARTEIRA!$O$4:$O$25000"

        # Row 2: Formulas SORTBY(FILTER(CARTEIRA...)) para colunas de contexto
        for a_col, c_col in AGENDA_TO_CARTEIRA.items():
            formula = (
                f'=IFERROR(SORTBY(FILTER(CARTEIRA!${c_col}$4:${c_col}$25000,'
                f'{filter_cond}),'
                f'FILTER({sort_col},{filter_cond}),-1),"")'
            )
            ws_a.cell(row=2, column=a_col, value=formula)

        # SINALEIRO formula (col L=12 da AGENDA)
        # Usa dados locais da AGENDA: G=SITUAÇÃO, F=DIAS
        sinaleiro_formula = (
            '=IF(OR(G2="PROSPECT",G2="LEAD"),"\U0001f7e3",'
            'IF(F2="","",IF(F2<=50,"\U0001f7e2",IF(F2<=90,"\U0001f7e1","\U0001f534"))))'
        )
        ws_a.cell(row=2, column=12, value=sinaleiro_formula)

        # TENTATIVA formula (col N=14)
        ws_a.cell(row=2, column=14, value='="T"&COUNTIF($C$2:C2,C2)')

        # FOLLOW-UP formula (col V=22)
        followup_formula = (
            '=IFERROR(IF(VLOOKUP(T2,REGRAS!$B$6:$C$20,2,FALSE)=0,"SEM",'
            'TODAY()+VLOOKUP(T2,REGRAS!$B$6:$C$20,2,FALSE)),"")'
        )
        ws_a.cell(row=2, column=22, value=followup_formula)

        # TIPO AÇÃO formula (col AB=28)
        tipo_acao_formula = (
            '=IF(T2="","",IF(OR(T2="VENDA / PEDIDO"),"VENDA",'
            'IF(OR(T2="PÓS-VENDA",T2="CS (SUCESSO DO CLIENTE)",'
            'T2="RELACIONAMENTO",T2="NUTRIÇÃO"),"PÓS-VENDA",'
            'IF(T2="SUPORTE","RESOLUÇÃO DE PROBLEMA",'
            'IF(OR(G2="PROSPECT",G2="LEAD"),"PROSPECÇÃO",'
            'IF(T2="PERDA / FECHOU LOJA","PRÉ-VENDA",'
            'IF(OR(T2="ORÇAMENTO",T2="CADASTRO",T2="EM ATENDIMENTO",'
            'T2="FOLLOW UP 7",T2="FOLLOW UP 15",T2="NÃO ATENDE",'
            'T2="NÃO RESPONDE",T2="RECUSOU LIGAÇÃO"),"PRÉ-VENDA",'
            '"PRÉ-VENDA")))))))'
        )
        ws_a.cell(row=2, column=28, value=tipo_acao_formula)

        # SCORE formula (col AE=31) — baseada no V31
        score_formula = (
            '=IF(B2="","",ROUND('
            'IF(F2="",0,IF(F2/50<0.7,0,IF(F2/50<=1,30,IF(F2/50<=1.5,60,90))))*0.3+'
            '_xlfn.IFS(I2="EM DESENVOLVIMENTO",100,I2="NOVO",50,I2="PROSPECT",30,I2="LEAD",20,TRUE,40)*0.25+'
            'IF(OR(V2="",V2="SEM",NOT(ISNUMBER(V2))),40,IF(V2<=TODAY(),100,IF(V2-TODAY()<=3,80,IF(V2-TODAY()<=7,60,40))))*0.2+'
            '_xlfn.IFS(LEFT(K2,1)="\U0001f525",100,LEFT(K2,1)="\U0001f7e1",40,LEFT(K2,1)="\u2744",10,LEFT(K2,1)="\U0001f480",0,TRUE,20)*0.15+'
            '_xlfn.IFS(N2="T4",100,N2="T3",50,N2="T2",20,N2="T1",10,N2="NUTRIÇÃO",30,N2="RESET",10,TRUE,10)*0.05+'
            '_xlfn.IFS(G2="EM RISCO",80,G2="ATIVO",40,G2="INAT.REC",60,G2="INAT.ANT",30,G2="NOVO",50,G2="PROSPECT",20,G2="LEAD",10,TRUE,20)*0.05'
            ',0))'
        )
        ws_a.cell(row=2, column=31, value=score_formula)

        # PRIORIDADE formula (col AF=32)
        prioridade_formula = (
            '=IF(AE2="","",IF(AE2>=70,"\U0001f534 URGENTE",'
            'IF(AE2>=50,"\U0001f7e1 ALTO",'
            'IF(AE2>=30,"\U0001f7e2 MÉDIO","\u26aa BAIXO"))))'
        )
        ws_a.cell(row=2, column=32, value=prioridade_formula)

        # Layout
        ws_a.freeze_panes = "A2"

        # Column widths
        agenda_widths = {
            1: 12, 2: 30, 3: 18, 4: 5, 5: 14, 6: 12, 7: 14, 8: 14,
            9: 16, 10: 14, 11: 14, 12: 10, 13: 22, 14: 10, 15: 14, 16: 8,
            17: 8, 18: 12, 19: 18, 20: 18, 21: 14, 22: 12, 23: 22, 24: 25,
            25: 25, 26: 10, 27: 25, 28: 14, 29: 14, 30: 16, 31: 10, 32: 14
        }
        for col_idx, width in agenda_widths.items():
            ws_a.column_dimensions[get_column_letter(col_idx)].width = width

        consultor_str = ", ".join(consultor) if isinstance(consultor, list) else consultor
        print(f"    32 colunas, SORTBY/FILTER por '{consultor_str}', SCORE+PRIORIDADE")

    msg = "  4 AGENDAs reconstruidas (32 cols cada, layout V31 com SCORE + PRIORIDADE)"
    changes.append(msg)

    # ================================================================
    # FASE 6: CORRIGIR CARTEIRA (anchors, layout, headers)
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 6: CORRIGIR CARTEIRA (layout V31)")
    print(f"{'='*100}")

    ws_cart = wb["CARTEIRA"]
    ws_v31_cart = wb_v31["CARTEIRA"]

    # 6A. Copiar pattern de outline/hidden do V31
    anchor_fixes = 0
    for col_key in ws_v31_cart.column_dimensions:
        v31_dim = ws_v31_cart.column_dimensions[col_key]
        v16_dim = ws_cart.column_dimensions[col_key]

        new_level = v31_dim.outline_level or 0
        old_level = v16_dim.outline_level or 0
        new_hidden = v31_dim.hidden
        old_hidden = v16_dim.hidden

        if new_level != old_level or new_hidden != old_hidden:
            v16_dim.outline_level = new_level
            v16_dim.hidden = new_hidden
            anchor_fixes += 1

    msg = f"  Anchors: {anchor_fixes} colunas ajustadas (outline + hidden)"
    print(msg)
    changes.append(msg)

    # 6B. Copiar larguras do V31
    width_fixes = 0
    for col_key in ws_v31_cart.column_dimensions:
        v31_w = ws_v31_cart.column_dimensions[col_key].width
        if v31_w:
            ws_cart.column_dimensions[col_key].width = v31_w
            width_fixes += 1

    msg = f"  Widths: {width_fixes} colunas"
    print(msg)
    changes.append(msg)

    # 6C. Headers rows 1-2 do V31
    ws_v31_cart_data = wb_v31_data["CARTEIRA"]
    max_common_col = min(ws_v31_cart_data.max_column, ws_cart.max_column)
    header_fixes = 0
    for row in [1, 2]:
        for col in range(1, max_common_col + 1):
            v31_val = ws_v31_cart_data.cell(row=row, column=col).value
            v16_val = ws_cart.cell(row=row, column=col).value
            if v31_val and v31_val != v16_val:
                ws_cart.cell(row=row, column=col, value=v31_val)
                header_fixes += 1

    msg = f"  Headers: {header_fixes} celulas de header (rows 1-2) alinhadas"
    print(msg)
    changes.append(msg)

    # 6D. Freeze panes
    ws_cart.freeze_panes = "BX4"

    # 6E. Auto filter
    cart_max_col = get_column_letter(ws_cart.max_column)
    ws_cart.auto_filter.ref = f"A3:{cart_max_col}{ws_cart.max_row}"

    # ================================================================
    # FASE 7: AJUSTAR OUTRAS ABAS
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 7: AJUSTAR LAYOUT OUTRAS ABAS")
    print(f"{'='*100}")

    # DRAFT 1
    if "DRAFT 1" in wb.sheetnames:
        ws_d1 = wb["DRAFT 1"]
        ws_d1.freeze_panes = "AN4"
        d1_max = get_column_letter(ws_d1.max_column)
        ws_d1.auto_filter.ref = f"A3:{d1_max}{ws_d1.max_row}"
        print(f"  DRAFT 1: freeze=AN4, filter=A3:{d1_max}{ws_d1.max_row}")

    # PROJECAO
    if "PROJEÇÃO " in wb.sheetnames:
        ws_proj = wb["PROJEÇÃO "]
        ws_proj.freeze_panes = "C4"
        print(f"  PROJECAO: freeze=C4")

    # REGRAS
    if "REGRAS" in wb.sheetnames:
        ws_reg = wb["REGRAS"]
        ws_reg.freeze_panes = "A41"
        print(f"  REGRAS: freeze=A41")

    # ================================================================
    # FASE 8: REORDENAR ABAS
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 8: REORDENAR ABAS")
    print(f"{'='*100}")

    desired_order = [
        "SINALEIRO",
        "PROJEÇÃO ",
        "DASH",
        "REDES_FRANQUIAS_v2",
        "COMITE",
        "REGRAS",
        "DRAFT 1",
        "DRAFT 2",
    ]

    # Add DRAFT 3 variants
    for s in wb.sheetnames:
        if "DRAFT 3" in s.upper() and s not in desired_order:
            desired_order.append(s)

    desired_order.extend([
        "CARTEIRA",
        "AGENDA LARISSA",
        "AGENDA DAIANE",
        "AGENDA MANU",
        "AGENDA JULIO",
    ])

    # Reordenar
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            current_idx = wb.sheetnames.index(name)
            if current_idx != i:
                wb.move_sheet(name, offset=i - current_idx)

    msg = f"  Abas: {wb.sheetnames}"
    print(msg)
    changes.append(msg)

    # ================================================================
    # FASE 9: SALVAR + VERIFICAR
    # ================================================================
    print(f"\n{'='*100}")
    print("FASE 9: SALVAR + VERIFICAR")
    print(f"{'='*100}")

    print("\n  Salvando V16...")
    wb.save(V16_PATH)
    wb.close()
    wb_v31.close()
    wb_v31_data.close()

    size = Path(V16_PATH).stat().st_size / (1024 * 1024)
    print(f"  Salvo: {V16_PATH} ({size:.2f} MB)")

    # Verificacao
    print(f"\n  Verificando V16...")
    wb_check = openpyxl.load_workbook(V16_PATH, data_only=False)

    total_formulas = 0
    for name in wb_check.sheetnames:
        ws = wb_check[name]
        f_count = 0
        for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    f_count += 1
        total_formulas += f_count
        print(f"    {name:25s} | {ws.max_row:>6} rows | {ws.max_column:>4} cols | {f_count:>7} formulas | freeze={ws.freeze_panes}")

    print(f"\n  TOTAL FORMULAS: {total_formulas:,}")
    print(f"  TOTAL ABAS: {len(wb_check.sheetnames)}")
    wb_check.close()

    # Resumo
    end_time = datetime.now()
    elapsed = (end_time - start_time).total_seconds()

    print(f"\n{'='*100}")
    print(f"CHANGELOG V15 → V16 ({len(changes)} mudancas) | {elapsed:.1f}s")
    print(f"{'='*100}")
    for c in changes:
        print(f"  {c}")

    print(f"\n[SUCESSO] V16 UNIFICADO gerado em {elapsed:.1f}s!")


if __name__ == "__main__":
    main()

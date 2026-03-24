#!/usr/bin/env python3
"""
BUILD V25 — Incorporar vendas FEV/25 no CRM VITAO360
=====================================================
Parte do V24 (que já tem JAN/25).
Adiciona coluna FEV/25 no DRAFT 1, CARTEIRA e PROJEÇÃO.
"""

import openpyxl
import re
import time
import os
from copy import copy
from datetime import datetime
from openpyxl.utils import get_column_letter, column_index_from_string

# ============================================================
# CONFIGURAÇÃO
# ============================================================
INPUT_FILE = "c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V24_FINAL.xlsx"
OUTPUT_FILE = "c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V25_FINAL.xlsx"

# FEV/25 Vendas — 26 registros (SAP_CODE, NOME, QTD, BRUTO, DEVOL, LIQUIDO)
FEV25_DATA = [
    ("1000129541", "INSTITUICAO ADVENTISTA SUL BRASILEI", 83, 5053.32, 0.00, 5053.32),
    ("1000064761", "SIMPLES E SAUDAVEL COMERCIO PRODUTO", 31, 3430.80, 0.00, 3430.80),
    ("1000123123", "DS COMERCIAL LTDA - ME", 12, 3137.98, 0.00, 3137.98),
    ("1000069656", "COOP. CENTRAL CREDITO, POUPANCA E I", 23, 2831.96, 0.00, 2831.96),
    ("2000001925", "SUPERMERCADO CASTILHO DE CAFELANDIA", 19, 2467.18, 0.00, 2467.18),
    ("1000104388", "53.531.499 KASSIA KAROLINNY NUNES R", 14, 2308.56, 0.00, 2308.56),
    ("1000111668", "BIOMUNDO ARAGUAINA LTDA - EPP", 14, 2140.64, 0.00, 2140.64),
    ("1000114769", "MINA PRECIOSA COMERCIO DE ALIMENTOS", 16, 1747.86, 0.00, 1747.86),
    ("1000092971", "NATURALIS PRODUTOS NATURAIS LTDA -", 5, 1512.00, 0.00, 1512.00),
    ("1000095086", "GRANUM CAFE LTDA", 9, 1350.00, 0.00, 1350.00),
    ("1000066867", "ARMAZEM DA 18 COMERCIO DE ALIMENTOS", 14, 1336.80, 0.00, 1336.80),
    ("1000091117", "DISTRIBUIDORA LIGEIRINHO DE PRODUTO", 9, 1121.90, 0.00, 1121.90),
    ("1000071905", "QUIOSQUE DA NUTRI LTDA - ME", 9, 1086.80, 0.00, 1086.80),
    ("1000105324", "58.188.486 TERESINHA PEREIRA DA SIL", 9, 946.98, 0.00, 946.98),
    ("1000071531", "MAIRA FONTOURA VIEIRA MARQUES - ME", 6, 900.00, 0.00, 900.00),
    ("1000102615", "MARCIA ROTHBARTH - ME", 7, 824.69, 0.00, 824.69),
    ("1000069495", "49.844.222 LUIS CARLOS VIEIRA", 6, 756.00, 0.00, 756.00),
    ("1000132022", "ADEGA ARAUCARIA COMERCIO DE BEBIDAS", 5, 700.88, 0.00, 700.88),
    ("1000101303", "41.329.482 ELIO ANTONIO RODRIGUES L", 5, 682.06, 0.00, 682.06),
    ("1000105456", "ALCINA P DE MEDEIROS OLIVEIRA - ME", 7, 675.60, 0.00, 675.60),
    ("1000132013", "MARIA FITNESS LTDA  - ME", 8, 673.87, 0.00, 673.87),
    ("1000063288", "RISOTOLANDIA SERVICOS DE ALIMENTACA", 9, 669.60, 0.00, 669.60),
    ("1000105367", "M. A. SILVA DE FREITAS COMERCIO DE", 5, 609.78, 0.00, 609.78),
    ("1000104688", "L V PRODUTOS NATURAIS LTDA - EPP", 4, 397.60, 0.00, 397.60),
    ("1000103659", "SILVANE KESSLER FLACH - ME", 4, 386.06, 0.00, 386.06),
    ("1000097565", "GC NATURALIS LTDA - ME", 1, 255.56, 0.00, 255.56),
]


# ============================================================
# FUNÇÕES AUXILIARES (idênticas ao V24)
# ============================================================

def shift_col_in_ref(col_str, min_col_idx, shift=1):
    idx = column_index_from_string(col_str)
    if idx >= min_col_idx:
        return get_column_letter(idx + shift)
    return col_str


def fix_sheet_refs_in_formula(formula, sheet_prefix, min_col_idx, shift=1):
    """Corrige refs de coluna com UM ÚNICO regex pass (evita double-shift)."""
    if not formula or not isinstance(formula, str) or formula[0] != '=':
        return formula

    escaped = re.escape(sheet_prefix)
    combined_pat = rf"({escaped})(\$?)([A-Z]+)(\$?\d+)(?::(\$?)([A-Z]+)(\$?\d+))?"

    def replace_combined(m):
        pfx = m.group(1)
        d1 = m.group(2)
        col1 = m.group(3)
        rest1 = m.group(4)
        new1 = shift_col_in_ref(col1, min_col_idx, shift)
        result = f"{pfx}{d1}{new1}{rest1}"
        if m.group(6):
            d2 = m.group(5)
            col2 = m.group(6)
            rest2 = m.group(7)
            new2 = shift_col_in_ref(col2, min_col_idx, shift)
            result += f":{d2}{new2}{rest2}"
        return result

    return re.sub(combined_pat, replace_combined, formula)


# ============================================================
# MAIN
# ============================================================

def main():
    start = time.time()

    def elapsed():
        return f"{time.time() - start:.1f}s"

    print(f"[{elapsed()}] Carregando V24...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    print(f"[{elapsed()}] V24 carregado.")

    ws_d1 = wb['DRAFT 1']
    ws_cart = wb['CARTEIRA']
    ws_proj = wb['PROJEÇÃO ']

    # V24 layout:
    # DRAFT 1: col 21=JAN/25, col 22=MAR/25, ..., col 33=FEV/26
    # CARTEIRA: col 26=JAN/25, col 27=MAR/25(45717), ..., col 37=JAN/26(46023)
    # Inserir FEV/25: DRAFT 1 col 22, CARTEIRA col 27

    # ----------------------------------------------------------
    # ETAPA 1: Mapear SAP -> CNPJ -> rows
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 1: Mapeando SAP -> CNPJ...")

    # SAP map from CARTEIRA col 64 (shifted from 63 after V24 CARTEIRA insert)
    sap_to_cnpj = {}
    for r in range(4, ws_cart.max_row + 1):
        sap_val = ws_cart.cell(r, 64).value
        cnpj_val = ws_cart.cell(r, 2).value
        if sap_val and cnpj_val:
            sap_to_cnpj[str(sap_val).strip()] = str(cnpj_val).strip()

    # Also from DRAFT 3
    ws3 = wb['DRAFT 3 ']
    for r in range(4, ws3.max_row + 1):
        sap = ws3.cell(r, 1).value
        cnpj = ws3.cell(r, 2).value
        if sap and cnpj:
            k = str(sap).strip()
            if k not in sap_to_cnpj:
                sap_to_cnpj[k] = str(cnpj).strip()

    # CNPJ -> row maps
    cnpj_to_d1_row = {}
    for r in range(4, ws_d1.max_row + 1):
        cnpj = ws_d1.cell(r, 2).value
        if cnpj:
            cnpj_to_d1_row[str(cnpj).strip()] = r

    cnpj_to_cart_row = {}
    for r in range(4, ws_cart.max_row + 1):
        cnpj = ws_cart.cell(r, 2).value
        if cnpj:
            cnpj_to_cart_row[str(cnpj).strip()] = r

    cnpj_to_proj_row = {}
    for r in range(4, ws_proj.max_row + 1):
        cnpj = ws_proj.cell(r, 1).value
        if cnpj:
            cnpj_to_proj_row[str(cnpj).strip()] = r

    # Map FEV25 data
    fev25_mapped = []
    unmapped = []
    for sap, nome, qtd, bruto, devol, liquido in FEV25_DATA:
        cnpj = sap_to_cnpj.get(sap)
        if cnpj:
            fev25_mapped.append({
                'sap': sap, 'nome': nome, 'cnpj': cnpj,
                'qtd': qtd, 'bruto': bruto, 'devol': devol, 'liquido': liquido,
                'd1_row': cnpj_to_d1_row.get(cnpj),
                'cart_row': cnpj_to_cart_row.get(cnpj),
                'proj_row': cnpj_to_proj_row.get(cnpj),
            })
        else:
            unmapped.append(sap)

    print(f"  Mapeados: {len(fev25_mapped)}/{len(FEV25_DATA)}")
    if unmapped:
        print(f"  NÃO mapeados: {unmapped}")

    # ----------------------------------------------------------
    # ETAPA 2: Inserir FEV/25 no DRAFT 1 (posição 22)
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 2: Inserindo FEV/25 no DRAFT 1 (col 22)...")
    # V24: col 21=JAN/25, col 22=MAR/25
    # Após insert: col 21=JAN/25, col 22=FEV/25, col 23=MAR/25
    ws_d1.insert_cols(22)

    ws_d1.cell(3, 22).value = "FEV/25"
    # Copiar formatação
    for r in range(1, 4):
        src = ws_d1.cell(r, 23)
        dst = ws_d1.cell(r, 22)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format

    d1_written = 0
    for entry in fev25_mapped:
        if entry['d1_row'] and entry['liquido'] > 0:
            r = entry['d1_row']
            ws_d1.cell(r, 22).value = entry['liquido']
            src = ws_d1.cell(r, 23)
            if src.has_style:
                ws_d1.cell(r, 22).number_format = src.number_format
            d1_written += 1

    print(f"  DRAFT 1: {d1_written} valores na col 22 (FEV/25)")

    # ----------------------------------------------------------
    # ETAPA 3: Inserir FEV/25 na CARTEIRA (posição 27)
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 3: Inserindo FEV/25 na CARTEIRA (col 27)...")
    # V24: col 26=JAN/25, col 27=MAR/25(45717)
    # Após insert: col 26=JAN/25, col 27=FEV/25, col 28=MAR/25
    ws_cart.insert_cols(27)

    fev25_date = datetime(2025, 2, 1)
    ws_cart.cell(3, 27).value = fev25_date
    ws_cart.cell(2, 27).value = "VENDAS"

    for r in range(1, 4):
        src = ws_cart.cell(r, 28)
        dst = ws_cart.cell(r, 27)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format

    print(f"  CARTEIRA: col 27 inserida para FEV/25")

    # ----------------------------------------------------------
    # ETAPA 4: Corrigir TODAS as fórmulas
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 4: Corrigindo fórmulas...")

    # DRAFT 1 insert at 22: shift 'DRAFT 1' refs col >= V(22) by +1
    # CARTEIRA insert at 27: shift 'CARTEIRA' refs col >= AA(27) by +1

    draft1_prefix = "'DRAFT 1'!"
    carteira_prefix = "CARTEIRA!"
    draft1_min_col = 22  # V
    carteira_min_col = 27  # AA

    fixes_d1 = 0
    fixes_cart = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_r = ws.max_row or 1
        max_c = ws.max_column or 1

        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(r, c)
                val = cell.value
                if not val or not isinstance(val, str) or not val.startswith('='):
                    continue

                original = val

                # Fix DRAFT 1 refs (col >= V/22)
                if "'DRAFT 1'" in val:
                    val = fix_sheet_refs_in_formula(val, draft1_prefix, draft1_min_col)

                # Fix CARTEIRA refs from other sheets (col >= AA/27)
                if sheet_name != 'CARTEIRA' and 'CARTEIRA!' in val:
                    val = fix_sheet_refs_in_formula(val, carteira_prefix, carteira_min_col)

                # Fix CARTEIRA self-refs with explicit CARTEIRA! prefix
                if sheet_name == 'CARTEIRA' and 'CARTEIRA!' in val:
                    val = fix_sheet_refs_in_formula(val, carteira_prefix, carteira_min_col)

                if val != original:
                    cell.value = val
                    if "'DRAFT 1'" in original:
                        fixes_d1 += 1
                    if 'CARTEIRA!' in original:
                        fixes_cart += 1

    print(f"  Fórmulas DRAFT 1: {fixes_d1}")
    print(f"  Fórmulas CARTEIRA: {fixes_cart}")

    # ----------------------------------------------------------
    # ETAPA 5: Fórmulas FEV/25 na CARTEIRA col 27
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 5: Escrevendo fórmulas FEV/25 na CARTEIRA col 27...")

    # DRAFT 1 col 22 (V) = FEV/25
    cart_formulas = 0
    cart_values = 0

    for r in range(4, ws_cart.max_row + 1):
        neighbor = ws_cart.cell(r, 28).value  # col 28 = ex-MAR/25
        if neighbor and isinstance(neighbor, str) and neighbor.startswith('='):
            formula = f"=IFERROR(INDEX('DRAFT 1'!$V$3:$V$25000,MATCH($B{r},'DRAFT 1'!$B$3:$B$25000,0)),\"\")"
            ws_cart.cell(r, 27).value = formula
            cart_formulas += 1
        else:
            cnpj = ws_cart.cell(r, 2).value
            if cnpj:
                cnpj_str = str(cnpj).strip()
                for entry in fev25_mapped:
                    if entry['cnpj'] == cnpj_str and entry['liquido'] > 0:
                        ws_cart.cell(r, 27).value = entry['liquido']
                        cart_values += 1
                        break
                else:
                    ws_cart.cell(r, 27).value = ""
            else:
                ws_cart.cell(r, 27).value = ""

    for r in range(4, ws_cart.max_row + 1):
        src = ws_cart.cell(r, 28)
        dst = ws_cart.cell(r, 27)
        if src.has_style:
            dst.number_format = src.number_format

    print(f"  CARTEIRA col 27: {cart_formulas} fórmulas + {cart_values} valores")

    # ----------------------------------------------------------
    # ETAPA 6: TOTAL PERÍODO (col 25)
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 6: Corrigindo TOTAL PERÍODO...")
    # Vendas agora: col 26=JAN/25(Z), col 27=FEV/25(AA), col 28=MAR/25(AB), ..., col 38=JAN/26(AL)
    # TOTAL = SUM(Z:AL)
    total_fixed = 0
    for r in range(4, ws_cart.max_row + 1):
        old_val = ws_cart.cell(r, 25).value
        if old_val and isinstance(old_val, str) and 'SUM(' in old_val:
            ws_cart.cell(r, 25).value = f"=SUM(Z{r}:AL{r})"
            total_fixed += 1

    print(f"  TOTAL: {total_fixed} rows")

    # ----------------------------------------------------------
    # ETAPA 7: PROJEÇÃO REAL FEV (col 28)
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 7: Atualizando PROJEÇÃO REAL FEV (col 28)...")
    proj_written = 0
    for entry in fev25_mapped:
        if entry['proj_row'] and entry['liquido'] > 0:
            r = entry['proj_row']
            ws_proj.cell(r, 28).value = entry['liquido']
            proj_written += 1

    print(f"  PROJEÇÃO: {proj_written} valores")

    # ----------------------------------------------------------
    # ETAPA 8: Verificação
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 8: Verificação...")
    print(f"  DRAFT 1: col 21={ws_d1.cell(3,21).value}, col 22={ws_d1.cell(3,22).value}, col 23={ws_d1.cell(3,23).value}")
    print(f"  CARTEIRA: col 26={ws_cart.cell(3,26).value}, col 27={ws_cart.cell(3,27).value}, col 28={ws_cart.cell(3,28).value}")
    print(f"  CARTEIRA col 27 formula (row 4): {ws_cart.cell(4,27).value}")
    print(f"  CARTEIRA col 28 formula (row 4): {ws_cart.cell(4,28).value}")
    print(f"  CARTEIRA TOTAL (row 4): {ws_cart.cell(4,25).value}")

    for ag in ['AGENDA LARISSA']:
        ws_ag = wb[ag]
        print(f"  {ag} A4: {str(ws_ag.cell(4,1).value)[:80]}...")

    # ----------------------------------------------------------
    # ETAPA 9: Salvar
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 9: Salvando V25...")
    wb.save(OUTPUT_FILE)

    file_size = os.path.getsize(OUTPUT_FILE) / (1024 * 1024)
    total_time = time.time() - start

    print(f"\n{'='*60}")
    print(f"  V25 GERADO COM SUCESSO!")
    print(f"  Arquivo: {OUTPUT_FILE}")
    print(f"  Tamanho: {file_size:.2f} MB")
    print(f"  Tempo total: {total_time:.1f}s")
    print(f"{'='*60}")
    print(f"\n  DRAFT 1: col 22 FEV/25, {d1_written} valores")
    print(f"  CARTEIRA: col 27 FEV/25, {cart_formulas} fórmulas + {cart_values} valores")
    print(f"  PROJEÇÃO: {proj_written} valores em REAL FEV")
    print(f"  Fórmulas corrigidas: {fixes_d1} DRAFT 1 + {fixes_cart} CARTEIRA")
    print(f"  TOTAL PERÍODO: {total_fixed} rows")


if __name__ == '__main__':
    main()

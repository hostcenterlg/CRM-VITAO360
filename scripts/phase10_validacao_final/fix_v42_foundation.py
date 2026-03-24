#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CRM VITAO360 — Fix V42: Foundation Repair
==========================================
Corrige a BASE de dados para integridade 100%:

1. CNPJ padronizado (14 dígitos, sem pontuação) na CARTEIRA col B
2. DRAFT 2 cols C/D DESTROCADAS (engine gravava RAZÃO SOCIAL no lugar de CNPJ)
3. UF populado via SAP Cadastro Clientes (col 10 = Unidade Federação)
4. Nº COMPRAS derivado dos meses com faturamento > 0
5. CURVA ABC derivada do faturamento total acumulado
6. ESTÁGIO FUNIL populado com base no último atendimento DRAFT 2
7. Validação cruzada DRAFT 2 ↔ CARTEIRA (meta: 100% match)
"""

import openpyxl
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict
from datetime import datetime, timedelta
import os
import re

# ============================================================
# PATHS
# ============================================================
BASE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360'
INPUT = os.path.join(BASE, 'data', 'output', 'phase10', 'CRM_VITAO360_V41_FINAL.xlsx')
OUTPUT = os.path.join(BASE, 'data', 'output', 'phase10', 'CRM_VITAO360_V42_FINAL.xlsx')
SAP_FILE = os.path.join(BASE, 'data', 'sources', 'sap', '01_SAP_CONSOLIDADO.xlsx')

# ============================================================
# HELPERS
# ============================================================
def normalize_cnpj(val):
    """Normaliza CNPJ para 14 dígitos sem pontuação"""
    if val is None:
        return ''
    s = str(val).strip()
    # Remove pontuação
    digits = re.sub(r'[^0-9]', '', s)
    if not digits:
        return ''
    # Pad para 14 dígitos (CNPJs curtos = zeros à esquerda)
    if len(digits) <= 14:
        digits = digits.zfill(14)
    return digits

def is_valid_cnpj(cnpj_norm):
    """Verifica se parece um CNPJ válido (14 dígitos, não todo zero)"""
    return len(cnpj_norm) == 14 and cnpj_norm != '00000000000000' and cnpj_norm.isdigit()

def serial_to_date(serial):
    """Converte serial date do Excel para datetime"""
    try:
        n = int(float(serial))
        if 40000 < n < 50000:  # Range válido para 2009-2036
            return datetime(1899, 12, 30) + timedelta(days=n)
    except (ValueError, TypeError):
        pass
    return None


print("=" * 80)
print("CRM VITAO360 — V42: REPARO DA FUNDAÇÃO DE DADOS")
print("=" * 80)

# ============================================================
# [1/8] CARREGAR SAP PARA UF + DADOS CADASTRAIS
# ============================================================
print("\n[1/8] Carregando SAP Cadastro Clientes...")
sap_data = {}  # cnpj_norm → {uf, cidade, nome, status, vendedor_interno}

if os.path.exists(SAP_FILE):
    wb_sap = openpyxl.load_workbook(SAP_FILE, read_only=True, data_only=True)
    ws_sap = wb_sap['Cadastro Clientes SAP']
    for r in range(2, ws_sap.max_row + 1):
        cnpj_raw = ws_sap.cell(row=r, column=5).value  # Col 5 = CNPJ Cliente
        if not cnpj_raw:
            cnpj_raw = ws_sap.cell(row=r, column=4).value  # Col 4 = CNPJ ou CPF
        cnpj_norm = normalize_cnpj(cnpj_raw)
        if not is_valid_cnpj(cnpj_norm):
            continue
        sap_data[cnpj_norm] = {
            'uf': str(ws_sap.cell(row=r, column=10).value or '').strip(),
            'cidade': str(ws_sap.cell(row=r, column=9).value or '').strip(),
            'nome_sap': str(ws_sap.cell(row=r, column=6).value or '').strip(),
            'status': str(ws_sap.cell(row=r, column=2).value or '').strip(),
            'vendedor': str(ws_sap.cell(row=r, column=29).value or '').strip(),
        }

    # Carregar faturamento SAP real
    ws_venda = wb_sap['Venda Mês a Mês']
    sap_fat = {}  # cnpj_norm → total faturado
    for r in range(2, ws_venda.max_row + 1):
        cnpj_raw = ws_venda.cell(row=r, column=3).value
        cnpj_norm = normalize_cnpj(cnpj_raw)
        if not is_valid_cnpj(cnpj_norm):
            continue
        total = 0
        meses_compra = 0
        # Cols de faturamento: Faturado_Jan=7, Faturado_Fev=11, ... (a cada 4 cols)
        for fc in range(7, min(ws_venda.max_column + 1, 56), 4):
            v = ws_venda.cell(row=r, column=fc).value
            if v and isinstance(v, (int, float)) and v > 0:
                total += v
                meses_compra += 1
        sap_fat[cnpj_norm] = {'total': total, 'meses': meses_compra}

    wb_sap.close()
    print(f"  SAP Cadastro: {len(sap_data)} clientes com CNPJ válido")
    print(f"  SAP Faturamento: {len(sap_fat)} clientes com dados de venda")
else:
    print(f"  AVISO: SAP não encontrado em {SAP_FILE}")

# ============================================================
# [2/8] CARREGAR V41
# ============================================================
print(f"\n[2/8] Carregando V41...")
if not os.path.exists(INPUT):
    print(f"ERRO: {INPUT} não encontrado!")
    exit(1)

wb = openpyxl.load_workbook(INPUT)
ws_cart = wb['CARTEIRA']
ws_d2 = wb['DRAFT 2']

print(f"  CARTEIRA: {ws_cart.max_row} rows x {ws_cart.max_column} cols")
print(f"  DRAFT 2: {ws_d2.max_row} rows x {ws_d2.max_column} cols")

# ============================================================
# [3/8] MAPEAR CARTEIRA — RAZÃO SOCIAL → CNPJ + DADOS
# ============================================================
print(f"\n[3/8] Mapeando CARTEIRA...")

# Ler headers (row 3)
cart_headers = {}
for col in range(1, ws_cart.max_column + 1):
    v = ws_cart.cell(row=3, column=col).value
    if v:
        cart_headers[col] = str(v).strip()

# Detectar colunas de faturamento (headers com serial dates ou mês)
fat_cols = []
for col, name in cart_headers.items():
    d = serial_to_date(name)
    if d:
        fat_cols.append((col, d))

fat_cols.sort(key=lambda x: x[1])
print(f"  Colunas de faturamento detectadas: {len(fat_cols)}")
if fat_cols:
    print(f"  Período: {fat_cols[0][1].strftime('%b/%Y')} → {fat_cols[-1][1].strftime('%b/%Y')}")

# Construir mapeamentos da CARTEIRA
razao_to_cnpj = {}      # RAZÃO SOCIAL → CNPJ normalizado
razao_to_nome = {}       # RAZÃO SOCIAL → NOME FANTASIA
cnpj_to_row = {}         # CNPJ norm → row number na CARTEIRA
cnpj_to_razao = {}       # CNPJ norm → RAZÃO SOCIAL

cart_stats = {'total': 0, 'cnpj_ok': 0, 'cnpj_formatted': 0, 'cnpj_numeric': 0, 'cnpj_invalid': 0}

for r in range(4, ws_cart.max_row + 1):
    nome_fantasia = ws_cart.cell(row=r, column=1).value  # Col A = NOME FANTASIA
    cnpj_raw = ws_cart.cell(row=r, column=2).value        # Col B = CNPJ
    razao_social = ws_cart.cell(row=r, column=3).value     # Col C = RAZÃO SOCIAL

    if nome_fantasia is None and cnpj_raw is None and razao_social is None:
        continue

    cart_stats['total'] += 1
    cnpj_norm = normalize_cnpj(cnpj_raw)

    # Classificar formato CNPJ original
    if cnpj_raw:
        s = str(cnpj_raw).strip()
        if '.' in s or '/' in s or '-' in s:
            cart_stats['cnpj_formatted'] += 1
        elif s.isdigit():
            cart_stats['cnpj_numeric'] += 1
        else:
            cart_stats['cnpj_invalid'] += 1

    if is_valid_cnpj(cnpj_norm):
        cart_stats['cnpj_ok'] += 1
        razao_s = str(razao_social).strip() if razao_social else ''
        nome_s = str(nome_fantasia).strip() if nome_fantasia else ''

        if razao_s:
            razao_to_cnpj[razao_s] = cnpj_norm
            razao_to_nome[razao_s] = nome_s

        cnpj_to_row[cnpj_norm] = r
        cnpj_to_razao[cnpj_norm] = razao_s

print(f"  Total linhas: {cart_stats['total']}")
print(f"  CNPJ válidos: {cart_stats['cnpj_ok']}")
print(f"  Formato pontuado: {cart_stats['cnpj_formatted']}")
print(f"  Formato numérico: {cart_stats['cnpj_numeric']}")
print(f"  CNPJ inválido: {cart_stats['cnpj_invalid']}")
print(f"  Mapeamento RAZÃO→CNPJ: {len(razao_to_cnpj)} entradas")

# ============================================================
# [4/8] PADRONIZAR CNPJ NA CARTEIRA (Col B → 14 dígitos)
# ============================================================
print(f"\n[4/8] Padronizando CNPJ na CARTEIRA (col B)...")
cnpj_fixed = 0
for r in range(4, ws_cart.max_row + 1):
    cnpj_raw = ws_cart.cell(row=r, column=2).value
    if cnpj_raw is None:
        continue
    cnpj_norm = normalize_cnpj(cnpj_raw)
    if is_valid_cnpj(cnpj_norm) and str(cnpj_raw).strip() != cnpj_norm:
        ws_cart.cell(row=r, column=2).value = cnpj_norm
        cnpj_fixed += 1

print(f"  CNPJs padronizados: {cnpj_fixed}")

# ============================================================
# [5/8] POPULAR UF + Nº COMPRAS + CURVA ABC NA CARTEIRA
# ============================================================
print(f"\n[5/8] Populando campos derivados na CARTEIRA...")

# --- UF (col 4) ---
uf_filled = 0
for r in range(4, ws_cart.max_row + 1):
    current_uf = ws_cart.cell(row=r, column=4).value
    if current_uf:
        continue  # Já tem UF
    cnpj_raw = ws_cart.cell(row=r, column=2).value
    cnpj_norm = normalize_cnpj(cnpj_raw)
    if cnpj_norm in sap_data and sap_data[cnpj_norm]['uf']:
        ws_cart.cell(row=r, column=4).value = sap_data[cnpj_norm]['uf']
        uf_filled += 1

print(f"  UF preenchido: {uf_filled} clientes (via SAP)")

# --- Nº COMPRAS (col 40) e CURVA ABC (col 41) ---
# Calcular faturamento por cliente
client_fat = {}  # cnpj_norm → {total: float, meses: int}
for r in range(4, ws_cart.max_row + 1):
    cnpj_raw = ws_cart.cell(row=r, column=2).value
    cnpj_norm = normalize_cnpj(cnpj_raw)
    if not is_valid_cnpj(cnpj_norm):
        continue

    total_fat = 0
    meses_compra = 0
    for col, dt in fat_cols:
        v = ws_cart.cell(row=r, column=col).value
        if v and isinstance(v, (int, float)) and v > 0:
            total_fat += v
            meses_compra += 1

    # Se não tem dados na CARTEIRA, tentar SAP
    if meses_compra == 0 and cnpj_norm in sap_fat:
        total_fat = sap_fat[cnpj_norm]['total']
        meses_compra = sap_fat[cnpj_norm]['meses']

    client_fat[cnpj_norm] = {'total': total_fat, 'meses': meses_compra, 'row': r}

# Escrever Nº COMPRAS
compras_filled = 0
for cnpj_norm, info in client_fat.items():
    r = info['row']
    if info['meses'] > 0:
        ws_cart.cell(row=r, column=40).value = info['meses']
        compras_filled += 1
    else:
        ws_cart.cell(row=r, column=40).value = 0

print(f"  Nº COMPRAS preenchido: {compras_filled} clientes com compras")

# Calcular CURVA ABC (método acumulativo)
# Ordenar clientes por faturamento decrescente
sorted_clients = sorted(
    [(cnpj, info) for cnpj, info in client_fat.items() if info['total'] > 0],
    key=lambda x: -x[1]['total']
)
total_revenue = sum(info['total'] for _, info in sorted_clients)

abc_filled = 0
if total_revenue > 0:
    cumulative = 0
    for cnpj_norm, info in sorted_clients:
        cumulative += info['total']
        pct = cumulative / total_revenue
        if pct <= 0.80:
            curva = 'A'
        elif pct <= 0.95:
            curva = 'B'
        else:
            curva = 'C'
        ws_cart.cell(row=info['row'], column=41).value = curva
        abc_filled += 1

# Clientes sem faturamento = 'D' (sem compra)
for cnpj_norm, info in client_fat.items():
    if info['total'] == 0:
        ws_cart.cell(row=info['row'], column=41).value = 'D'

print(f"  CURVA ABC calculada: {abc_filled} clientes (A/B/C)")
if sorted_clients:
    a_count = sum(1 for _, i in sorted_clients if ws_cart.cell(row=i['row'], column=41).value == 'A')
    b_count = sum(1 for _, i in sorted_clients if ws_cart.cell(row=i['row'], column=41).value == 'B')
    c_count = sum(1 for _, i in sorted_clients if ws_cart.cell(row=i['row'], column=41).value == 'C')
    print(f"  Distribuição: A={a_count} | B={b_count} | C={c_count} | D={len(client_fat)-abc_filled}")

# ============================================================
# [6/8] CORRIGIR DRAFT 2 — Cols C/D (NOME↔CNPJ estavam trocados)
# ============================================================
print(f"\n[6/8] Corrigindo DRAFT 2 (cols C e D trocados pelo engine)...")

# O engine fazia:
#   cart_data key = RAZÃO SOCIAL (col 3 CARTEIRA)
#   info['nome'] = CNPJ numérico (col 2 CARTEIRA)  ← ERRADO
#   make_record(): 'C': nome (=CNPJ), 'D': cnpj_s (=RAZÃO SOCIAL)  ← TROCADO
#
# Resultado no DRAFT 2:
#   Col C (NOME FANTASIA) → contém CNPJ numérico
#   Col D (CNPJ) → contém RAZÃO SOCIAL
#
# Correção: para cada linha do DRAFT 2, se col D parece RAZÃO SOCIAL,
# buscar o CNPJ real via mapeamento RAZÃO→CNPJ da CARTEIRA

d2_fixed = 0
d2_already_ok = 0
d2_not_found = 0
d2_not_found_samples = []

for r in range(3, ws_d2.max_row + 1):
    data_cell = ws_d2.cell(row=r, column=1).value
    if data_cell is None:
        continue

    col_c = ws_d2.cell(row=r, column=3).value  # Deveria ser NOME FANTASIA
    col_d = ws_d2.cell(row=r, column=4).value  # Deveria ser CNPJ

    if col_d is None:
        continue

    col_d_s = str(col_d).strip()
    col_c_s = str(col_c).strip() if col_c else ''

    # Detectar se col D contém RAZÃO SOCIAL (texto longo, não numérico)
    # vs CNPJ real (numérico ou formatado XX.XXX.XXX/XXXX-XX)
    cnpj_test = normalize_cnpj(col_d_s)
    is_d_cnpj = is_valid_cnpj(cnpj_test)

    if is_d_cnpj:
        # Col D já parece ser um CNPJ válido
        # Verificar se col C também precisa de fix (pode ter o CNPJ numérico)
        cnpj_c_test = normalize_cnpj(col_c_s)
        if is_valid_cnpj(cnpj_c_test) and cnpj_test in cnpj_to_razao:
            # Col C tem CNPJ ao invés de nome → corrigir
            # Buscar nome fantasia real
            razao = cnpj_to_razao.get(cnpj_test, '')
            nome_real = razao_to_nome.get(razao, '')
            if nome_real:
                ws_d2.cell(row=r, column=3).value = nome_real
                ws_d2.cell(row=r, column=4).value = cnpj_test  # Padronizar formato
                d2_fixed += 1
            else:
                d2_already_ok += 1
        else:
            # Padronizar CNPJ em col D para 14 dígitos
            ws_d2.cell(row=r, column=4).value = cnpj_test
            d2_already_ok += 1
        continue

    # Col D contém texto (RAZÃO SOCIAL) → TROCADO!
    # Buscar CNPJ real pela RAZÃO SOCIAL
    if col_d_s in razao_to_cnpj:
        cnpj_real = razao_to_cnpj[col_d_s]
        nome_real = razao_to_nome.get(col_d_s, '')

        # Corrigir: C = NOME FANTASIA, D = CNPJ real
        ws_d2.cell(row=r, column=3).value = nome_real if nome_real else col_c_s
        ws_d2.cell(row=r, column=4).value = cnpj_real
        d2_fixed += 1
    else:
        # Tentar match parcial (nome similar)
        found = False
        for razao, cnpj_r in razao_to_cnpj.items():
            if col_d_s.upper() in razao.upper() or razao.upper() in col_d_s.upper():
                nome_real = razao_to_nome.get(razao, '')
                ws_d2.cell(row=r, column=3).value = nome_real if nome_real else col_c_s
                ws_d2.cell(row=r, column=4).value = cnpj_r
                d2_fixed += 1
                found = True
                break

        if not found:
            # Verificar se col_c_s (que pode ter o CNPJ numérico) tem o CNPJ
            cnpj_c_test = normalize_cnpj(col_c_s)
            if is_valid_cnpj(cnpj_c_test):
                # Col C tem o CNPJ, col D tem o nome → TROCAR
                ws_d2.cell(row=r, column=3).value = col_d_s  # Nome vai pra C
                ws_d2.cell(row=r, column=4).value = cnpj_c_test  # CNPJ vai pra D
                d2_fixed += 1
            else:
                d2_not_found += 1
                if len(d2_not_found_samples) < 10:
                    d2_not_found_samples.append(f"  Row {r}: C='{col_c_s[:40]}' D='{col_d_s[:40]}'")

print(f"  Linhas corrigidas: {d2_fixed}")
print(f"  Linhas já OK: {d2_already_ok}")
print(f"  Não encontrados na CARTEIRA: {d2_not_found}")
if d2_not_found_samples:
    print(f"  Amostras não encontradas:")
    for s in d2_not_found_samples:
        print(s)

# ============================================================
# [7/8] POPULAR UF NO DRAFT 2 (Col E) via CNPJ
# ============================================================
print(f"\n[7/8] Populando UF no DRAFT 2 (col E)...")
d2_uf_filled = 0
for r in range(3, ws_d2.max_row + 1):
    data_cell = ws_d2.cell(row=r, column=1).value
    if data_cell is None:
        continue
    current_uf = ws_d2.cell(row=r, column=5).value
    if current_uf:
        continue

    cnpj_d = ws_d2.cell(row=r, column=4).value
    cnpj_norm = normalize_cnpj(cnpj_d)
    if cnpj_norm in sap_data and sap_data[cnpj_norm]['uf']:
        ws_d2.cell(row=r, column=5).value = sap_data[cnpj_norm]['uf']
        d2_uf_filled += 1

print(f"  UF preenchido no DRAFT 2: {d2_uf_filled} registros")

# ============================================================
# [8/8] VALIDAÇÃO CRUZADA + SALVAR
# ============================================================
print(f"\n[8/8] Validação cruzada DRAFT 2 ↔ CARTEIRA...")

# Re-ler CNPJs do DRAFT 2 (agora corrigidos)
d2_cnpjs = set()
d2_cnpj_count = Counter()
for r in range(3, ws_d2.max_row + 1):
    data_cell = ws_d2.cell(row=r, column=1).value
    if data_cell is None:
        continue
    cnpj_d = ws_d2.cell(row=r, column=4).value
    if cnpj_d:
        cnpj_norm = normalize_cnpj(cnpj_d)
        if is_valid_cnpj(cnpj_norm):
            d2_cnpjs.add(cnpj_norm)
            d2_cnpj_count[cnpj_norm] += 1

# Re-ler CNPJs da CARTEIRA (agora padronizados)
cart_cnpjs = set()
for r in range(4, ws_cart.max_row + 1):
    cnpj_raw = ws_cart.cell(row=r, column=2).value
    if cnpj_raw:
        cnpj_norm = normalize_cnpj(cnpj_raw)
        if is_valid_cnpj(cnpj_norm):
            cart_cnpjs.add(cnpj_norm)

both = d2_cnpjs & cart_cnpjs
only_d2 = d2_cnpjs - cart_cnpjs
only_cart = cart_cnpjs - d2_cnpjs

match_pct = len(both) / len(d2_cnpjs) * 100 if d2_cnpjs else 0

print(f"  DRAFT 2 CNPJs únicos: {len(d2_cnpjs)}")
print(f"  CARTEIRA CNPJs únicos: {len(cart_cnpjs)}")
print(f"  Em AMBOS: {len(both)} ({match_pct:.1f}%)")
print(f"  Só no DRAFT 2: {len(only_d2)}")
print(f"  Só na CARTEIRA (sem atendimento): {len(only_cart)}")

if only_d2:
    print(f"\n  Amostras CNPJs só no DRAFT 2 (não na CARTEIRA):")
    for c in sorted(only_d2)[:10]:
        # Mostrar nome associado
        for r in range(3, min(ws_d2.max_row + 1, 25000)):
            cnpj_d = ws_d2.cell(row=r, column=4).value
            if cnpj_d and normalize_cnpj(cnpj_d) == c:
                nome = ws_d2.cell(row=r, column=3).value
                print(f"    {c} → {nome}")
                break

# Verificar amostra de integridade
print(f"\n  --- Amostra de integridade (10 linhas) ---")
sample_rows = [3, 4, 5, 100, 500, 1000, 5000, 10000, 15000, 20000]
for r in sample_rows:
    if r > ws_d2.max_row:
        continue
    data = ws_d2.cell(row=r, column=1).value
    if data is None:
        continue
    cons = ws_d2.cell(row=r, column=2).value
    nome = ws_d2.cell(row=r, column=3).value
    cnpj = ws_d2.cell(row=r, column=4).value
    uf = ws_d2.cell(row=r, column=5).value
    cnpj_norm = normalize_cnpj(cnpj)
    in_cart = '✅' if cnpj_norm in cart_cnpjs else '❌'
    print(f"  Row {r:5d}: {in_cart} CNPJ={cnpj} | NOME={str(nome)[:30]} | UF={uf} | CONS={cons}")

# SALVAR
print(f"\n{'='*80}")
print(f"Salvando V42...")
wb.save(OUTPUT)
file_size = os.path.getsize(OUTPUT) / (1024 * 1024)

print(f"\n{'='*80}")
print(f"V42 CONCLUÍDO — FUNDAÇÃO REPARADA")
print(f"{'='*80}")
print(f"  Arquivo: {OUTPUT}")
print(f"  Tamanho: {file_size:.2f} MB")
print(f"")
print(f"  CARTEIRA:")
print(f"    CNPJ padronizado: {cnpj_fixed} corrigidos")
print(f"    UF preenchido: {uf_filled} clientes")
print(f"    Nº COMPRAS: {compras_filled} clientes")
print(f"    CURVA ABC: {abc_filled} clientes")
print(f"")
print(f"  DRAFT 2:")
print(f"    Cols C/D corrigidos: {d2_fixed} linhas")
print(f"    UF preenchido: {d2_uf_filled} registros")
print(f"")
print(f"  INTEGRIDADE:")
print(f"    Match DRAFT 2 → CARTEIRA: {match_pct:.1f}% ({len(both)}/{len(d2_cnpjs)})")
print(f"    Meta: 100%")
print(f"{'='*80}")

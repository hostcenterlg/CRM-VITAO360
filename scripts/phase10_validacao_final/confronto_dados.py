#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Auditoria: Confronto V13 vs V31 do CRM-VITAO360
Objetivo: Validar integridade, completude e confiabilidade dos dados entre versões
Data: 2026-02-17
"""

import openpyxl
import json
import re
from pathlib import Path
from collections import defaultdict
from typing import Dict, Set, Tuple, List

# Caminho dos arquivos
V13_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V13_FINAL.xlsx"
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"
OUTPUT_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/confronto_v13_v31.json"

# Utilidades
def normalizar_cnpj(valor):
    """Extrai e normaliza CNPJ: remove pontos, barras e hífen"""
    if not valor or not isinstance(valor, str):
        return None
    # Remove caracteres de formatação
    cnpj = re.sub(r'[^\d]', '', valor.strip())
    # Valida se tem 14 dígitos
    if len(cnpj) == 14 and cnpj.isdigit():
        return cnpj
    return None

def extrair_cnpjs(ws, col_indices: List[int]) -> Dict[str, Tuple[str, int]]:
    """
    Extrai CNPJs de um worksheet, retorna {cnpj: (cliente_name, row_num)}
    col_indices: lista de índices de colunas a verificar
    """
    cnpjs_encontrados = {}
    duplicados_local = defaultdict(list)

    # Mapeia qual coluna tem o cliente (geralmente próxima ao CNPJ)
    cliente_col = None

    for row_num in range(2, ws.max_row + 1):
        # Procura por CNPJ em qualquer coluna especificada
        cnpj_encontrado = None
        cnpj_col = None

        for col_idx in col_indices:
            cell = ws.cell(row=row_num, column=col_idx)
            if cell.value:
                cnpj = normalizar_cnpj(str(cell.value))
                if cnpj:
                    cnpj_encontrado = cnpj
                    cnpj_col = col_idx
                    break

        # Se encontrou CNPJ, tenta achar cliente
        if cnpj_encontrado:
            cliente = None
            # Tenta colunas próximas para achar cliente
            for offset in [-2, -1, 1, 2]:
                try_col = cnpj_col + offset
                if 1 <= try_col <= ws.max_column:
                    cell_val = ws.cell(row=row_num, column=try_col).value
                    if cell_val and isinstance(cell_val, str) and len(str(cell_val).strip()) > 2:
                        cliente = str(cell_val).strip()
                        break

            if not cliente:
                cliente = f"[SEM_NOME]"

            # Registra achado
            if cnpj_encontrado in cnpjs_encontrados:
                duplicados_local[cnpj_encontrado].append(row_num)
            else:
                cnpjs_encontrados[cnpj_encontrado] = (cliente, row_num)
                duplicados_local[cnpj_encontrado] = [row_num]

    return cnpjs_encontrados, duplicados_local

def analisar_completude(ws, col_indices: List[int], start_row: int = 2) -> Dict:
    """Analisa completude de dados nas colunas especificadas"""
    resultado = {
        "total_linhas": max(0, ws.max_row - start_row + 1),
        "por_coluna": {}
    }

    for col_idx in col_indices:
        preenchidas = 0
        vazias = 0

        for row_num in range(start_row, ws.max_row + 1):
            cell_val = ws.cell(row=row_num, column=col_idx).value
            if cell_val and str(cell_val).strip():
                preenchidas += 1
            else:
                vazias += 1

        col_letter = openpyxl.utils.get_column_letter(col_idx)
        resultado["por_coluna"][col_letter] = {
            "preenchidas": preenchidas,
            "vazias": vazias,
            "percentual": round(100 * preenchidas / max(1, preenchidas + vazias), 2)
        }

    return resultado

def extrair_consultores(ws, col_consultor: int) -> Dict[str, int]:
    """Extrai lista de consultores e conta de registros"""
    consultores = defaultdict(int)

    for row_num in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_num, column=col_consultor).value
        if cell_val:
            consultor = str(cell_val).strip()
            if consultor:
                consultores[consultor] += 1

    return dict(consultores)

# MAIN
def main():
    print("\n" + "="*80)
    print("AUDITORIA DE DADOS: CRM V13 vs V31")
    print("="*80)

    resultado_auditoria = {
        "data_auditoria": "2026-02-17",
        "arquivos": {
            "v13": V13_PATH,
            "v31": V31_PATH
        },
        "inventario_abas": {},
        "confronto_cnpj": {},
        "confronto_clientes": {},
        "confronto_consultores": {},
        "completude_dados": {},
        "recomendacoes": []
    }

    try:
        # ========== FASE 1: Carregamento e Inventário ==========
        print("\n[1/5] Carregando arquivos...")

        wb_v13 = openpyxl.load_workbook(V13_PATH, data_only=True)
        wb_v31 = openpyxl.load_workbook(V31_PATH, data_only=True)

        abas_v13 = set(wb_v13.sheetnames)
        abas_v31 = set(wb_v31.sheetnames)

        print(f"  V13: {len(abas_v13)} abas")
        print(f"  V31: {len(abas_v31)} abas")

        resultado_auditoria["inventario_abas"] = {
            "v13": list(abas_v13),
            "v31": list(abas_v31),
            "em_ambas": list(abas_v13 & abas_v31),
            "exclusivas_v13": list(abas_v13 - abas_v31),
            "exclusivas_v31": list(abas_v31 - abas_v13)
        }

        # ========== FASE 2: Confronto CNPJ na CARTEIRA ==========
        print("\n[2/5] Analisando CNPJs na aba CARTEIRA...")

        if "CARTEIRA" in abas_v13 and "CARTEIRA" in abas_v31:
            ws_v13_carteira = wb_v13["CARTEIRA"]
            ws_v31_carteira = wb_v31["CARTEIRA"]

            # Detecta colunas com CNPJ (procura em B, C, D, E)
            cols_check = [2, 3, 4, 5]  # B, C, D, E

            cnpjs_v13, dup_v13 = extrair_cnpjs(ws_v13_carteira, cols_check)
            cnpjs_v31, dup_v31 = extrair_cnpjs(ws_v31_carteira, cols_check)

            set_cnpj_v13 = set(cnpjs_v13.keys())
            set_cnpj_v31 = set(cnpjs_v31.keys())

            intersecao = set_cnpj_v13 & set_cnpj_v31
            exclusivos_v13 = set_cnpj_v13 - set_cnpj_v31
            exclusivos_v31 = set_cnpj_v31 - set_cnpj_v13

            resultado_auditoria["confronto_cnpj"] = {
                "total_v13": len(set_cnpj_v13),
                "total_v31": len(set_cnpj_v31),
                "intersecao": len(intersecao),
                "exclusivos_v13": len(exclusivos_v13),
                "exclusivos_v31": len(exclusivos_v31),
                "duplicados_v13": {k: len(v) for k, v in dup_v13.items() if len(v) > 1},
                "duplicados_v31": {k: len(v) for k, v in dup_v31.items() if len(v) > 1},
                "amostra_exclusivos_v13": list(exclusivos_v13)[:10],
                "amostra_exclusivos_v31": list(exclusivos_v31)[:10]
            }

            print(f"  V13: {len(set_cnpj_v13)} CNPJs únicos")
            print(f"  V31: {len(set_cnpj_v31)} CNPJs únicos")
            print(f"  Interseção: {len(intersecao)} CNPJs comuns")
            print(f"  Exclusivos V13: {len(exclusivos_v13)}")
            print(f"  Exclusivos V31: {len(exclusivos_v31)}")
            print(f"  Duplicados V13: {len([k for k,v in dup_v13.items() if len(v) > 1])}")
            print(f"  Duplicados V31: {len([k for k,v in dup_v31.items() if len(v) > 1])}")

            # ========== FASE 3: Confronto de Clientes ==========
            print("\n[3/5] Analisando nomes de clientes...")

            # Agrupa clientes por CNPJ
            clientes_v13 = {}
            clientes_v31 = {}

            for cnpj, (cliente, _) in cnpjs_v13.items():
                clientes_v13[cnpj] = cliente

            for cnpj, (cliente, _) in cnpjs_v31.items():
                clientes_v31[cnpj] = cliente

            # Verifica variações de nome para mesmos CNPJs
            nomes_diferentes = 0
            for cnpj in intersecao:
                if clientes_v13.get(cnpj) != clientes_v31.get(cnpj):
                    nomes_diferentes += 1

            resultado_auditoria["confronto_clientes"] = {
                "clientes_unicos_v13": len(set(clientes_v13.values())),
                "clientes_unicos_v31": len(set(clientes_v31.values())),
                "cnpjs_com_nomes_diferentes": nomes_diferentes,
                "clientes_ausentes_v31": len([c for cnpj, c in clientes_v13.items() if cnpj not in set_cnpj_v31])
            }

            print(f"  Clientes únicos V13: {resultado_auditoria['confronto_clientes']['clientes_unicos_v13']}")
            print(f"  Clientes únicos V31: {resultado_auditoria['confronto_clientes']['clientes_unicos_v31']}")
            print(f"  CNPJs com nomes diferentes: {nomes_diferentes}")

            # ========== FASE 4: Completude ==========
            print("\n[4/5] Analisando completude dos dados...")

            completude_v13 = analisar_completude(ws_v13_carteira, cols_check)
            completude_v31 = analisar_completude(ws_v31_carteira, cols_check)

            resultado_auditoria["completude_dados"] = {
                "v13": completude_v13,
                "v31": completude_v31
            }

            print(f"  V13: {completude_v13['total_linhas']} linhas de dados")
            print(f"  V31: {completude_v31['total_linhas']} linhas de dados")

            # ========== FASE 5: Consultores ==========
            print("\n[5/5] Analisando distribuição de consultores...")

            # Procura coluna de consultor (pode ser A, B ou outra)
            col_consultor_v13 = 1  # Padrão: coluna A
            col_consultor_v31 = 1

            consultores_v13 = extrair_consultores(ws_v13_carteira, col_consultor_v13)
            consultores_v31 = extrair_consultores(ws_v31_carteira, col_consultor_v31)

            resultado_auditoria["confronto_consultores"] = {
                "total_consultores_v13": len(consultores_v13),
                "total_consultores_v31": len(consultores_v31),
                "consultores_v13": consultores_v13,
                "consultores_v31": consultores_v31
            }

            print(f"  Consultores V13: {len(consultores_v13)}")
            print(f"  Consultores V31: {len(consultores_v31)}")

        # ========== Recomendações ==========
        print("\n[ANÁLISE] Gerando recomendações...")

        # Determina qual versão é mais confiável
        recomendacoes = []

        if len(set_cnpj_v13) > len(set_cnpj_v31):
            recomendacoes.append({
                "prioridade": "ALTA",
                "aspecto": "Cobertura de CNPJs",
                "achado": f"V13 tem {len(set_cnpj_v13) - len(set_cnpj_v31)} CNPJs a mais que V31",
                "recomendacao": "V13 aparenta ter cobertura mais completa"
            })
        else:
            recomendacoes.append({
                "prioridade": "ALTA",
                "aspecto": "Cobertura de CNPJs",
                "achado": f"V31 tem {len(set_cnpj_v31) - len(set_cnpj_v13)} CNPJs a mais que V13",
                "recomendacao": "V31 aparenta ter cobertura mais completa"
            })

        if nomes_diferentes > 0:
            recomendacoes.append({
                "prioridade": "MÉDIA",
                "aspecto": "Inconsistência de dados",
                "achado": f"{nomes_diferentes} CNPJs com nomes de clientes diferentes entre versões",
                "recomendacao": "Revisar manualmente CNPJs conflitantes e padronizar nomes"
            })

        dup_v13_count = sum(1 for v in dup_v13.values() if len(v) > 1)
        dup_v31_count = sum(1 for v in dup_v31.values() if len(v) > 1)

        if dup_v13_count > 0 or dup_v31_count > 0:
            recomendacao = f"V13: {dup_v13_count} CNPJs duplicados | V31: {dup_v31_count} CNPJs duplicados"
            recomendacoes.append({
                "prioridade": "ALTA",
                "aspecto": "Duplicação de dados",
                "achado": recomendacao,
                "recomendacao": "Executar deduplicação antes de usar dados em produção"
            })

        resultado_auditoria["recomendacoes"] = recomendacoes

        # Salva resultado
        with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
            json.dump(resultado_auditoria, f, ensure_ascii=False, indent=2)

        print(f"\n✓ Resultado salvo em: {OUTPUT_PATH}")

    except Exception as e:
        print(f"\n✗ ERRO: {str(e)}")
        raise

    return resultado_auditoria

# ========== RESUMO EXECUTIVO ==========
def gerar_resumo_executivo(resultado):
    print("\n" + "="*80)
    print("RESUMO EXECUTIVO - AUDITORIA V13 vs V31")
    print("="*80)

    inv = resultado["inventario_abas"]
    cnpj = resultado.get("confronto_cnpj", {})
    cli = resultado.get("confronto_clientes", {})
    cons = resultado.get("confronto_consultores", {})
    rec = resultado["recomendacoes"]

    print(f"\n📊 ESTRUTURA:")
    print(f"  • Abas em ambas: {len(inv['em_ambas'])}")
    print(f"  • Exclusivas V13: {len(inv['exclusivas_v13'])}")
    print(f"  • Exclusivas V31: {len(inv['exclusivas_v31'])}")

    if cnpj:
        print(f"\n💼 COBERTURA (CNPJs únicos na CARTEIRA):")
        print(f"  • V13: {cnpj['total_v13']} CNPJs")
        print(f"  • V31: {cnpj['total_v31']} CNPJs")
        print(f"  • Comuns: {cnpj['intersecao']} CNPJs")
        print(f"  • Apenas em V13: {cnpj['exclusivos_v13']} CNPJs")
        print(f"  • Apenas em V31: {cnpj['exclusivos_v31']} CNPJs")

        print(f"\n⚠️  QUALIDADE DOS DADOS:")
        print(f"  • Duplicações em V13: {len(cnpj['duplicados_v13'])} CNPJs duplicados")
        print(f"  • Duplicações em V31: {len(cnpj['duplicados_v31'])} CNPJs duplicados")
        print(f"  • CNPJs com nomes diferentes: {cli.get('cnpjs_com_nomes_diferentes', 0)}")

    if cons:
        print(f"\n👥 CONSULTORES:")
        print(f"  • Consultores em V13: {cons['total_consultores_v13']}")
        print(f"  • Consultores em V31: {cons['total_consultores_v31']}")

    print(f"\n🎯 RECOMENDAÇÕES:")
    for i, rec_item in enumerate(rec, 1):
        prioridade = rec_item["prioridade"]
        aspecto = rec_item["aspecto"]
        recomendacao = rec_item["recomendacao"]
        print(f"\n  {i}. [{prioridade}] {aspecto}")
        print(f"     → {recomendacao}")

    print("\n" + "="*80 + "\n")

if __name__ == "__main__":
    resultado = main()
    gerar_resumo_executivo(resultado)

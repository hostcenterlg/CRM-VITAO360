#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CRM VITAO360 — V26 (Correções Críticas da Auditoria)
=====================================================
Base: V25
Correções:
  1. Pré-popular CARTEIRA VENDAS (cols 26-38) com VALORES REAIS do DRAFT 1
  2. Pré-popular TOTAL PERÍODO (col 25) com soma calculada
  3. Pré-popular CICLO MÉDIO (col 19) com valor do DRAFT 1
  4. Corrigir 34 clientes sem SITUAÇÃO na CARTEIRA
  5. Pré-popular SINALEIRO (col 64) para os 33 com fórmula
  6. Pré-popular ESTÁGIO FUNIL (col 46) para os 89 com fórmula
  7. Normalizar duplicatas de acentuação no DRAFT 2
"""

import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import os
import sys
from datetime import datetime

# ============================================================
# CONFIGURAÇÃO
# ============================================================
INPUT_FILE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10\CRM_VITAO360_V25_FINAL.xlsx'
OUTPUT_FILE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10\CRM_VITAO360_V26_FINAL.xlsx'

# CARTEIRA vendas: cols 26-38 (JAN/25 a JAN/26) → DRAFT 1 cols 21-33
CARTEIRA_VENDAS_START = 26  # col Z
CARTEIRA_VENDAS_END = 38    # col AL
CARTEIRA_TOTAL_COL = 25     # col Y = TOTAL PERÍODO
DRAFT1_VENDAS_START = 21    # col U = JAN/25
DRAFT1_CNPJ_COL = 2        # col B

# CARTEIRA outros
CARTEIRA_CNPJ_COL = 2      # col B
CARTEIRA_SITUACAO_COL = 14  # col N
CARTEIRA_PRIORIDADE_COL = 15 # col O
CARTEIRA_DIAS_COMPRA_COL = 16 # col P
CARTEIRA_CICLO_COL = 19     # col S
CARTEIRA_SINALEIRO_COL = 64 # col BL
CARTEIRA_FUNIL_COL = 46     # col AT
CARTEIRA_DATA_START = 4
CARTEIRA_CLIENTES_END = 557  # rows 4-557 = 554 clientes originais

# DRAFT 1
DRAFT1_CICLO_COL = 15       # col O = CICLO MÉDIO
DRAFT1_DATA_START = 3

# DRAFT 2
DRAFT2_RESULTADO_COL = 18   # col R
DRAFT2_TIPO_CONTATO_COL = 17 # col Q
DRAFT2_DATA_START = 3

# REGRAS de SITUAÇÃO
# ≤50 dias = ATIVO, 51-60 = EM RISCO, 61-90 = INAT.REC, >90 = INATIVO, sem compra = PROSPECT
SITUACAO_RULES = [
    (0, 50, "ATIVO"),
    (51, 60, "EM RISCO"),
    (61, 90, "INAT.REC"),
    (91, 99999, "INATIVO"),
]

# Mapeamento de normalização de acentos no DRAFT 2
NORMALIZE_RESULTADO = {
    "NAO ATENDE": "NÃO ATENDE",
    "NAO RESPONDE": "NÃO RESPONDE",
    "ORCAMENTO": "ORÇAMENTO",
    "FOLLOW UP 7": "FOLLOW UP 7",     # manter
    "FOLLOW UP 15": "FOLLOW UP 15",   # manter
    "NUTRIÇÃO": "NUTRIÇÃO",           # manter
}

NORMALIZE_TIPO_CONTATO = {
    "PROSPECCAO": "PROSPECÇÃO",
    "POS-VENDA / RELACIONAMENTO": "PÓS-VENDA/RELACIONAMENTO",
    "NEGOCIACAO": "NEGOCIAÇÃO",
    "PERDA / NUTRICAO": "PERDA/NUTRIÇÃO",
}

print("=" * 80)
print("CRM VITAO360 — BUILD V26 (Correções Auditoria)")
print("=" * 80)
print(f"Input:  {INPUT_FILE}")
print(f"Output: {OUTPUT_FILE}")

# ============================================================
# 1. CARREGAR ARQUIVO
# ============================================================
print("\n[1/7] Carregando V25...")
if not os.path.exists(INPUT_FILE):
    print(f"ERRO: Arquivo não encontrado: {INPUT_FILE}")
    sys.exit(1)

wb = openpyxl.load_workbook(INPUT_FILE)
print(f"  Abas: {wb.sheetnames}")

ws_cart = wb['CARTEIRA']
ws_d1 = wb['DRAFT 1']
ws_d2 = wb['DRAFT 2']

# Verificar abas com espaço
draft3_name = None
proj_name = None
for s in wb.sheetnames:
    if 'DRAFT 3' in s:
        draft3_name = s
    if 'PROJE' in s.upper():
        proj_name = s

print(f"  CARTEIRA: {ws_cart.max_row} linhas x {ws_cart.max_column} colunas")
print(f"  DRAFT 1: {ws_d1.max_row} linhas x {ws_d1.max_column} colunas")
print(f"  DRAFT 2: {ws_d2.max_row} linhas x {ws_d2.max_column} colunas")

# ============================================================
# 2. CONSTRUIR MAPA CNPJ → VALORES DO DRAFT 1
# ============================================================
print("\n[2/7] Construindo mapa CNPJ → Vendas do DRAFT 1...")

# Mapa: cnpj → {draft1_col: valor}
cnpj_vendas = {}
cnpj_ciclo = {}

for r in range(DRAFT1_DATA_START, ws_d1.max_row + 1):
    cnpj = ws_d1.cell(row=r, column=DRAFT1_CNPJ_COL).value
    if cnpj is None:
        continue
    cnpj_str = str(cnpj).strip()

    # Vendas por mês (DRAFT 1 cols 21-33)
    vendas = {}
    for dc in range(DRAFT1_VENDAS_START, DRAFT1_VENDAS_START + 13):
        v = ws_d1.cell(row=r, column=dc).value
        if v is not None and isinstance(v, (int, float)):
            vendas[dc] = v
        else:
            vendas[dc] = 0
    cnpj_vendas[cnpj_str] = vendas

    # Ciclo médio (col 15)
    ciclo = ws_d1.cell(row=r, column=DRAFT1_CICLO_COL).value
    if ciclo is not None and isinstance(ciclo, (int, float)):
        cnpj_ciclo[cnpj_str] = ciclo

print(f"  CNPJs mapeados: {len(cnpj_vendas)}")
print(f"  CNPJs com ciclo: {len(cnpj_ciclo)}")

# ============================================================
# 3. PRÉ-POPULAR CARTEIRA VENDAS COM VALORES REAIS
# ============================================================
print("\n[3/7] Pré-populando CARTEIRA VENDAS com valores reais...")

vendas_preenchidas = 0
total_preenchido = 0
ciclo_preenchido = 0

for r in range(CARTEIRA_DATA_START, CARTEIRA_CLIENTES_END + 1):
    cnpj = ws_cart.cell(row=r, column=CARTEIRA_CNPJ_COL).value
    if cnpj is None:
        continue
    cnpj_str = str(cnpj).strip()

    if cnpj_str not in cnpj_vendas:
        continue

    vendas = cnpj_vendas[cnpj_str]
    total = 0

    # Preencher cada mês (CARTEIRA cols 26-38 ← DRAFT 1 cols 21-33)
    for i in range(13):
        cart_col = CARTEIRA_VENDAS_START + i  # 26, 27, ..., 38
        d1_col = DRAFT1_VENDAS_START + i      # 21, 22, ..., 33
        valor = vendas.get(d1_col, 0)

        # Substituir fórmula por valor real
        ws_cart.cell(row=r, column=cart_col).value = valor if valor != 0 else ""
        total += valor
        vendas_preenchidas += 1

    # TOTAL PERÍODO (col 25)
    ws_cart.cell(row=r, column=CARTEIRA_TOTAL_COL).value = total if total > 0 else ""
    total_preenchido += 1

    # CICLO MÉDIO (col 19) — substituir fórmula por valor
    if cnpj_str in cnpj_ciclo:
        ws_cart.cell(row=r, column=CARTEIRA_CICLO_COL).value = cnpj_ciclo[cnpj_str]
        ciclo_preenchido += 1

print(f"  Células de vendas preenchidas: {vendas_preenchidas}")
print(f"  TOTAL PERÍODO preenchido: {total_preenchido}")
print(f"  CICLO MÉDIO preenchido: {ciclo_preenchido}")

# ============================================================
# 4. CORRIGIR 34 CLIENTES SEM SITUAÇÃO
# ============================================================
print("\n[4/7] Corrigindo clientes sem SITUAÇÃO...")

situacao_corrigidas = 0
for r in range(CARTEIRA_DATA_START, CARTEIRA_CLIENTES_END + 1):
    sit = ws_cart.cell(row=r, column=CARTEIRA_SITUACAO_COL).value
    if sit is not None and str(sit).strip() != "":
        continue  # Já tem situação

    # Verificar DIAS SEM COMPRA
    dias = ws_cart.cell(row=r, column=CARTEIRA_DIAS_COMPRA_COL).value
    cnpj = ws_cart.cell(row=r, column=CARTEIRA_CNPJ_COL).value

    if dias is not None and isinstance(dias, (int, float)):
        dias_int = int(dias)
        nova_situacao = "INATIVO"  # default
        for min_d, max_d, label in SITUACAO_RULES:
            if min_d <= dias_int <= max_d:
                nova_situacao = label
                break
        ws_cart.cell(row=r, column=CARTEIRA_SITUACAO_COL).value = nova_situacao
        situacao_corrigidas += 1
        print(f"    CNPJ {cnpj}: {dias_int} dias → {nova_situacao}")
    elif dias is None or dias == "" or dias == 0:
        # Sem compra = PROSPECT? Ou cliente sem dados?
        # Se está na zona dos 554 clientes, pode ser um caso especial
        ws_cart.cell(row=r, column=CARTEIRA_SITUACAO_COL).value = "PROSPECT"
        situacao_corrigidas += 1
        print(f"    CNPJ {cnpj}: sem dias → PROSPECT")

print(f"  Total corrigidas: {situacao_corrigidas}")

# ============================================================
# 5. PRÉ-POPULAR SINALEIRO PARA OS 33 COM FÓRMULA
# ============================================================
print("\n[5/7] Pré-populando SINALEIRO...")

sinaleiro_corrigido = 0
for r in range(CARTEIRA_DATA_START, ws_cart.max_row + 1):
    v = ws_cart.cell(row=r, column=CARTEIRA_SINALEIRO_COL).value
    if v is None or v == "":
        continue

    # Se é fórmula, calcular o valor
    if isinstance(v, str) and v.startswith("="):
        situacao = ws_cart.cell(row=r, column=CARTEIRA_SITUACAO_COL).value
        dias = ws_cart.cell(row=r, column=CARTEIRA_DIAS_COMPRA_COL).value
        ciclo = ws_cart.cell(row=r, column=CARTEIRA_CICLO_COL).value

        if situacao and str(situacao).upper() in ("PROSPECT", "LEAD"):
            ws_cart.cell(row=r, column=CARTEIRA_SINALEIRO_COL).value = "\U0001f7e3"  # 🟣
            sinaleiro_corrigido += 1
        elif situacao and str(situacao).upper() == "NOVO":
            ws_cart.cell(row=r, column=CARTEIRA_SINALEIRO_COL).value = "EM DESENVOLVIMENTO"
            sinaleiro_corrigido += 1
        elif dias is not None and isinstance(dias, (int, float)):
            dias_int = int(dias)
            ciclo_int = int(ciclo) if ciclo and isinstance(ciclo, (int, float)) else 50

            if dias_int <= ciclo_int:
                ws_cart.cell(row=r, column=CARTEIRA_SINALEIRO_COL).value = "\U0001f7e2"  # 🟢
            elif dias_int <= ciclo_int + 30:
                ws_cart.cell(row=r, column=CARTEIRA_SINALEIRO_COL).value = "\U0001f7e1"  # 🟡
            else:
                ws_cart.cell(row=r, column=CARTEIRA_SINALEIRO_COL).value = "\U0001f534"  # 🔴
            sinaleiro_corrigido += 1
        else:
            ws_cart.cell(row=r, column=CARTEIRA_SINALEIRO_COL).value = ""
            sinaleiro_corrigido += 1

print(f"  Sinaleiros corrigidos: {sinaleiro_corrigido}")

# ============================================================
# 6. PRÉ-POPULAR ESTÁGIO FUNIL PARA OS 89 COM FÓRMULA
# ============================================================
print("\n[6/7] Pré-populando ESTÁGIO FUNIL...")

# Para os que têm fórmula, buscar o registro mais recente no DRAFT 2
# Construir mapa: CNPJ → último ESTÁGIO FUNIL do DRAFT 2
d2_funil_map = {}  # cnpj → (data, funil)
for r in range(DRAFT2_DATA_START, ws_d2.max_row + 1):
    cnpj = ws_d2.cell(row=r, column=4).value  # DRAFT 2 col D = CNPJ
    data = ws_d2.cell(row=r, column=1).value   # DRAFT 2 col A = DATA
    funil = ws_d2.cell(row=r, column=9).value   # DRAFT 2 col I = ESTÁGIO FUNIL

    if cnpj is None:
        continue
    cnpj_str = str(cnpj).strip()

    if funil and str(funil).strip():
        funil_str = str(funil).strip()
        # Ignorar fórmulas
        if funil_str.startswith("="):
            continue

        # Guardar o mais recente
        if cnpj_str not in d2_funil_map:
            d2_funil_map[cnpj_str] = (data, funil_str)
        else:
            old_data = d2_funil_map[cnpj_str][0]
            if data and old_data and data > old_data:
                d2_funil_map[cnpj_str] = (data, funil_str)

print(f"  Mapa FUNIL do DRAFT 2: {len(d2_funil_map)} CNPJs")

funil_corrigido = 0
for r in range(CARTEIRA_DATA_START, ws_cart.max_row + 1):
    v = ws_cart.cell(row=r, column=CARTEIRA_FUNIL_COL).value
    if v is None or v == "":
        continue

    if isinstance(v, str) and v.startswith("="):
        cnpj = ws_cart.cell(row=r, column=CARTEIRA_CNPJ_COL).value
        if cnpj:
            cnpj_str = str(cnpj).strip()
            if cnpj_str in d2_funil_map:
                ws_cart.cell(row=r, column=CARTEIRA_FUNIL_COL).value = d2_funil_map[cnpj_str][1]
                funil_corrigido += 1
            else:
                # Sem registro no DRAFT 2 → PROSPECÇÃO (default)
                ws_cart.cell(row=r, column=CARTEIRA_FUNIL_COL).value = "PROSPECÇÃO"
                funil_corrigido += 1

print(f"  Funis corrigidos: {funil_corrigido}")

# ============================================================
# 7. NORMALIZAR ACENTUAÇÃO NO DRAFT 2
# ============================================================
print("\n[7/7] Normalizando acentuação no DRAFT 2...")

resultado_fixes = 0
tipo_contato_fixes = 0

for r in range(DRAFT2_DATA_START, ws_d2.max_row + 1):
    # RESULTADO (col R = 18)
    resultado = ws_d2.cell(row=r, column=DRAFT2_RESULTADO_COL).value
    if resultado and isinstance(resultado, str) and not resultado.startswith("="):
        resultado_clean = resultado.strip()
        if resultado_clean in NORMALIZE_RESULTADO:
            new_val = NORMALIZE_RESULTADO[resultado_clean]
            if new_val != resultado_clean:
                ws_d2.cell(row=r, column=DRAFT2_RESULTADO_COL).value = new_val
                resultado_fixes += 1

    # TIPO DO CONTATO (col Q = 17)
    tipo = ws_d2.cell(row=r, column=DRAFT2_TIPO_CONTATO_COL).value
    if tipo and isinstance(tipo, str) and not tipo.startswith("="):
        tipo_clean = tipo.strip()
        if tipo_clean in NORMALIZE_TIPO_CONTATO:
            new_val = NORMALIZE_TIPO_CONTATO[tipo_clean]
            if new_val != tipo_clean:
                ws_d2.cell(row=r, column=DRAFT2_TIPO_CONTATO_COL).value = new_val
                tipo_contato_fixes += 1

print(f"  RESULTADO normalizados: {resultado_fixes}")
print(f"  TIPO CONTATO normalizados: {tipo_contato_fixes}")

# ============================================================
# SALVAR
# ============================================================
print(f"\nSalvando V26 em: {OUTPUT_FILE}")
wb.save(OUTPUT_FILE)

file_size = os.path.getsize(OUTPUT_FILE) / (1024 * 1024)
print(f"  Tamanho: {file_size:.2f} MB")

# ============================================================
# RESUMO
# ============================================================
print("\n" + "=" * 80)
print("RESUMO V26 — CORREÇÕES APLICADAS")
print("=" * 80)
print(f"  [1] CARTEIRA VENDAS pré-populadas: {vendas_preenchidas} células")
print(f"  [2] TOTAL PERÍODO calculado: {total_preenchido} linhas")
print(f"  [3] CICLO MÉDIO pré-populado: {ciclo_preenchido} linhas")
print(f"  [4] SITUAÇÃO corrigidas: {situacao_corrigidas} clientes")
print(f"  [5] SINALEIRO pré-populados: {sinaleiro_corrigido}")
print(f"  [6] ESTÁGIO FUNIL pré-populados: {funil_corrigido}")
print(f"  [7] DRAFT 2 normalizações: {resultado_fixes} resultados + {tipo_contato_fixes} tipos")
print(f"\n  Arquivo: {OUTPUT_FILE}")
print(f"  Tamanho: {file_size:.2f} MB")
print("=" * 80)

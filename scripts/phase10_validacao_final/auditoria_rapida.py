#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AUDITORIA RAPIDA — Conta CNPJs por fonte sem carregar tudo.
Foco: V17 vs V31 vs Carteira Detalhada vs Mercos vs SAP
"""

import openpyxl
import re
from pathlib import Path
from datetime import datetime

def normalizar_cnpj(valor):
    if not valor:
        return None
    cnpj = re.sub(r'[^\d]', '', str(valor).strip())
    return cnpj if len(cnpj) >= 11 else None

def count_cnpjs(path, sheet_name, cnpj_col, start_row=4, label=""):
    """Conta CNPJs unicos em uma aba especifica"""
    print(f"  [{label}] Carregando {Path(path).name} -> {sheet_name}...", flush=True)
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            # Tentar match parcial
            for s in wb.sheetnames:
                if sheet_name.strip().upper() in s.strip().upper():
                    sheet_name = s
                    break
            else:
                print(f"    ! Aba '{sheet_name}' nao encontrada. Abas: {wb.sheetnames[:5]}")
                wb.close()
                return set(), {}
        ws = wb[sheet_name]

        cnpjs = set()
        situacoes = {}  # situacao → count

        # Detectar col SITUACAO
        sit_col = None
        for col in range(1, min((ws.max_column or 50) + 1, 50)):
            for row in [1, 2, 3]:
                h = ws.cell(row=row, column=col).value
                if h and "SITUA" in str(h).upper():
                    sit_col = col
                    break
            if sit_col:
                break

        for row in range(start_row, (ws.max_row or start_row) + 1):
            val = ws.cell(row=row, column=cnpj_col).value
            cnpj = normalizar_cnpj(val)
            if cnpj:
                cnpjs.add(cnpj)

                if sit_col:
                    sit = ws.cell(row=row, column=sit_col).value
                    if sit:
                        s = str(sit).strip()[:25]
                        situacoes[s] = situacoes.get(s, 0) + 1

        wb.close()
        print(f"    {len(cnpjs)} CNPJs unicos")
        if situacoes:
            top = sorted(situacoes.items(), key=lambda x: -x[1])[:8]
            print(f"    Situacoes: {', '.join(f'{k}({v})' for k,v in top)}")
        return cnpjs, situacoes

    except Exception as e:
        print(f"    ERRO: {e}")
        return set(), {}


def main():
    start = datetime.now()
    print("=" * 100)
    print("AUDITORIA RAPIDA — Universo de Clientes")
    print("=" * 100)

    results = {}

    # 1. V17 CARTEIRA atual
    print("\n--- V17 CARTEIRA (ATUAL) ---")
    v17_cart, _ = count_cnpjs(
        r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V17_FINAL.xlsx",
        "CARTEIRA", cnpj_col=2, start_row=4, label="V17"
    )
    results["V17_CARTEIRA"] = v17_cart

    # 2. V31 CARTEIRA
    print("\n--- V31 CARTEIRA (REFERENCIA) ---")
    v31_cart, v31_sit = count_cnpjs(
        r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx",
        "CARTEIRA", cnpj_col=2, start_row=4, label="V31"
    )
    results["V31_CARTEIRA"] = v31_cart

    # 3. V31 DRAFT 1
    print("\n--- V31 DRAFT 1 ---")
    v31_d1, _ = count_cnpjs(
        r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx",
        "DRAFT 1", cnpj_col=2, start_row=4, label="V31-D1"
    )
    results["V31_DRAFT1"] = v31_d1

    # 4. V32 (mais recente)
    print("\n--- V32 (MAIS RECENTE?) ---")
    v32_path = r"c:/Users/User/OneDrive/Área de Trabalho/CRM_V12_POPULADO_V32 18.02 vf ultima parte .xlsx"
    if Path(v32_path).exists():
        v32_cart, v32_sit = count_cnpjs(v32_path, "CARTEIRA", cnpj_col=2, start_row=4, label="V32")
        results["V32_CARTEIRA"] = v32_cart
    else:
        print("  ! V32 nao encontrado")

    # 5. Carteira Detalhada fev2026 COM prospects
    print("\n--- CARTEIRA DETALHADA FEV2026 (COM PROSPECTS) ---")
    cart_com_path = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/Carteira detalhada de clientes fevreiro 2026.xlsx"
    if Path(cart_com_path).exists():
        cart_com, cart_com_sit = count_cnpjs(cart_com_path, "Plan1", cnpj_col=2, start_row=2, label="DET-COM")
        if not cart_com:
            # Tentar primeira aba
            wb = openpyxl.load_workbook(cart_com_path, data_only=True, read_only=True)
            first = wb.sheetnames[0]
            wb.close()
            cart_com, cart_com_sit = count_cnpjs(cart_com_path, first, cnpj_col=2, start_row=2, label="DET-COM")
        results["CART_DET_COM"] = cart_com
    else:
        print(f"  ! Nao encontrado: {cart_com_path}")

    # 6. Carteira Detalhada fev2026 SEM prospects
    print("\n--- CARTEIRA DETALHADA FEV2026 (SEM PROSPECTS) ---")
    cart_sem_path = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/Carteira detalhada de clientes fevereiro 2026 - sem os prospects.xlsx"
    if Path(cart_sem_path).exists():
        cart_sem, _ = count_cnpjs(cart_sem_path, "Plan1", cnpj_col=2, start_row=2, label="DET-SEM")
        if not cart_sem:
            wb = openpyxl.load_workbook(cart_sem_path, data_only=True, read_only=True)
            first = wb.sheetnames[0]
            wb.close()
            cart_sem, _ = count_cnpjs(cart_sem_path, first, cnpj_col=2, start_row=2, label="DET-SEM")
        results["CART_DET_SEM"] = cart_sem

    # 7. Mercos Carteira
    print("\n--- CARTEIRA MERCOS ORIGINAL ---")
    mercos_path = r"c:/Users/User/OneDrive/Área de Trabalho/CLAUDE CODE/08_CARTEIRA_MERCOS.xlsx"
    if Path(mercos_path).exists():
        # Mercos pode ter CNPJ em cols diferentes
        mercos, mercos_sit = count_cnpjs(mercos_path, "Carteira de clientes", cnpj_col=2, start_row=2, label="MERCOS")
        if not mercos:
            wb = openpyxl.load_workbook(mercos_path, data_only=True, read_only=True)
            first = wb.sheetnames[0]
            print(f"    Abas: {wb.sheetnames}")
            # Detectar CNPJ col
            ws = wb[first]
            for col in range(1, 20):
                h = ws.cell(row=1, column=col).value
                print(f"    Col {col}: {h}")
            wb.close()
            mercos, mercos_sit = count_cnpjs(mercos_path, first, cnpj_col=2, start_row=2, label="MERCOS")
        results["MERCOS"] = mercos

    # 8. SAP Total
    print("\n--- SAP TOTAL ---")
    sap_path = r"c:/Users/User/OneDrive/Área de Trabalho/GERENCIAL/GESTÃO  DE CLIENTES/BASE SAP OFICIAL/CARTEIRA  TOTAL DE CLIENTES SAP.xlsx"
    if Path(sap_path).exists():
        sap, sap_sit = count_cnpjs(sap_path, "Plan1", cnpj_col=1, start_row=2, label="SAP")
        if not sap:
            wb = openpyxl.load_workbook(sap_path, data_only=True, read_only=True)
            first = wb.sheetnames[0]
            print(f"    Abas: {wb.sheetnames}")
            ws = wb[first]
            for col in range(1, 15):
                h = ws.cell(row=1, column=col).value
                print(f"    Col {col}: {h}")
            wb.close()
        results["SAP_TOTAL"] = sap

    # 9. DRAFT 3 SAP COM INATIVOS
    print("\n--- DRAFT 3 SAP COM INATIVOS ---")
    d3i_path = r"c:/Users/User/OneDrive/Área de Trabalho/DRAFT3/DRAFT_3_SAP  COM INATIVOS.xlsx"
    if Path(d3i_path).exists():
        d3i, d3i_sit = count_cnpjs(d3i_path, "Plan1", cnpj_col=2, start_row=2, label="D3-INAT")
        if not d3i:
            wb = openpyxl.load_workbook(d3i_path, data_only=True, read_only=True)
            first = wb.sheetnames[0]
            print(f"    Abas: {wb.sheetnames}")
            ws = wb[first]
            for col in range(1, 15):
                h = ws.cell(row=1, column=col).value
                print(f"    Col {col}: {h}")
            wb.close()
            d3i, d3i_sit = count_cnpjs(d3i_path, first, cnpj_col=2, start_row=2, label="D3-INAT")
        results["DRAFT3_INATIVOS"] = d3i

    # ================================================================
    # ANALISE CRUZADA
    # ================================================================
    print(f"\n\n{'='*100}")
    print("ANALISE CRUZADA")
    print(f"{'='*100}")

    print(f"\n  UNIVERSO POR FONTE:")
    for name, cnpjs in sorted(results.items(), key=lambda x: -len(x[1])):
        print(f"    {name:25s}: {len(cnpjs):>6} CNPJs")

    v17 = results.get("V17_CARTEIRA", set())

    print(f"\n  COMPARACAO COM V17 ({len(v17)} CNPJs):")
    for name, cnpjs in sorted(results.items(), key=lambda x: -len(x[1])):
        if name == "V17_CARTEIRA":
            continue
        common = v17 & cnpjs
        only_source = cnpjs - v17
        print(f"    {name:25s}: {len(cnpjs):>6} total | {len(common):>5} em comum | {len(only_source):>6} FALTANDO no V17")

    # Prospects
    cart_com = results.get("CART_DET_COM", set())
    cart_sem = results.get("CART_DET_SEM", set())
    if cart_com and cart_sem:
        prospects = cart_com - cart_sem
        prosp_faltando = prospects - v17
        print(f"\n  PROSPECTS MERCOS (detalhada COM - SEM):")
        print(f"    Total prospects: {len(prospects)}")
        print(f"    Faltando no V17: {len(prosp_faltando)}")

    # Universo total
    todos = set()
    for cnpjs in results.values():
        todos.update(cnpjs)
    faltando = todos - v17
    print(f"\n  UNIVERSO TOTAL: {len(todos)} CNPJs")
    print(f"  FALTANDO NO V17: {len(faltando)} CNPJs")

    elapsed = (datetime.now() - start).total_seconds()
    print(f"\n  Tempo: {elapsed:.1f}s")


if __name__ == "__main__":
    main()

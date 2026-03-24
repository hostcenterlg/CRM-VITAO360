#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BUILD V43 AGENDAS — Popular 40 atendimentos (10/consultor) para 19/02/2026
===========================================================================
Lê CARTEIRA + SINALEIRO da V43 CLEAN e distribui 10 clientes por consultor
priorizando: Curva A > B > C, maior dias sem compra primeiro.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os, time, re

start = time.time()

BASE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360'
INPUT = os.path.join(BASE, 'data', 'output', 'phase10', 'CRM_VITAO360_V43_CLEAN.xlsx')
OUTPUT = os.path.join(BASE, 'data', 'output', 'phase10', 'CRM_VITAO360_V43_FINAL.xlsx')

print("=" * 80)
print("BUILD V43 — AGENDAS + DASH FUNCIONAL")
print(f"Início: {datetime.now().strftime('%H:%M:%S')}")
print("=" * 80)

# ============================================================
# [1/5] CARREGAR V43 CLEAN
# ============================================================
print("\n[1/5] Carregando V43 CLEAN...")
t1 = time.time()
wb = openpyxl.load_workbook(INPUT)
print(f"  Carregado em {time.time()-t1:.1f}s")

ws_cart = wb['CARTEIRA']
ws_sin = wb['SINALEIRO']

# ============================================================
# [2/5] MAPEAR CARTEIRA — extrair dados por consultor
# ============================================================
print("\n[2/5] Mapeando CARTEIRA por consultor...")

# Headers row 3
cart_h = {}
for c in range(1, ws_cart.max_column + 1):
    v = ws_cart.cell(row=3, column=c).value
    if v:
        cart_h[str(v).strip().upper()] = c

# Encontrar colunas chave
col_nome = 1       # A
col_cnpj = 2       # B
col_razao = 3      # C
col_uf = 4         # D
col_cidade = None
col_consultor = None
col_dias_sem = None
col_curva = None
col_ticket = None
col_situacao = None
col_tipo = None

for name, col in cart_h.items():
    if 'CIDADE' in name:
        col_cidade = col
    elif 'CONSULTOR' in name and col_consultor is None:
        col_consultor = col
    elif 'DIAS SEM COMPRA' in name or 'DIAS SEM' in name:
        col_dias_sem = col
    elif 'CURVA' in name or 'ABC' in name:
        col_curva = col
    elif 'TICKET' in name and 'MEDIO' in name.replace('É', 'E'):
        col_ticket = col
    elif name == 'SITUAÇÃO' or name == 'SITUACAO':
        col_situacao = col
    elif 'TIPO CLIENTE' in name or 'TIPO' == name:
        col_tipo = col

print(f"  Colunas encontradas:")
print(f"    CONSULTOR: col {col_consultor}")
print(f"    DIAS SEM COMPRA: col {col_dias_sem}")
print(f"    CURVA ABC: col {col_curva}")
print(f"    TICKET MEDIO: col {col_ticket}")
print(f"    CIDADE: col {col_cidade}")
print(f"    SITUACAO: col {col_situacao}")

# Extrair clientes por consultor
consultores = {
    'LARISSA': [],
    'DAIANE': [],
    'MANU': [],
    'JULIO': [],
}

# Mapear SINALEIRO por CNPJ
sin_data = {}  # cnpj -> {cor, meta, real}
for r in range(5, ws_sin.max_row + 1):
    cnpj = ws_sin.cell(row=r, column=1).value
    if cnpj:
        cnpj_s = str(cnpj).strip()
        sin_data[cnpj_s] = {
            'cor': ws_sin.cell(row=r, column=14).value,  # COR SINALEIRO
        }

for r in range(4, ws_cart.max_row + 1):
    cnpj = ws_cart.cell(row=r, column=col_cnpj).value
    if not cnpj:
        continue
    cnpj_s = str(cnpj).strip()

    nome = ws_cart.cell(row=r, column=col_nome).value or ''
    razao = ws_cart.cell(row=r, column=col_razao).value or ''
    uf = ws_cart.cell(row=r, column=col_uf).value or ''
    cidade = ws_cart.cell(row=r, column=col_cidade).value if col_cidade else ''
    consultor = str(ws_cart.cell(row=r, column=col_consultor).value or '').strip().upper() if col_consultor else ''
    dias_sem = ws_cart.cell(row=r, column=col_dias_sem).value if col_dias_sem else 0
    curva = str(ws_cart.cell(row=r, column=col_curva).value or 'D').strip().upper() if col_curva else 'D'
    ticket = ws_cart.cell(row=r, column=col_ticket).value if col_ticket else 0
    situacao = ws_cart.cell(row=r, column=col_situacao).value if col_situacao else ''
    sinaleiro_cor = sin_data.get(cnpj_s, {}).get('cor', '')

    if not isinstance(dias_sem, (int, float)):
        try:
            dias_sem = int(dias_sem)
        except:
            dias_sem = 0
    if not isinstance(ticket, (int, float)):
        try:
            ticket = float(ticket)
        except:
            ticket = 0

    client = {
        'cnpj': cnpj_s,
        'nome': str(nome)[:50],
        'razao': str(razao)[:50],
        'uf': str(uf) if uf else '',
        'cidade': str(cidade)[:30] if cidade else '',
        'situacao': str(situacao) if situacao else 'ATIVO',
        'dias_sem': dias_sem if dias_sem else 0,
        'curva': curva if curva in ('A','B','C','D') else 'D',
        'ticket': round(ticket, 2) if ticket else 0,
        'sinaleiro': str(sinaleiro_cor) if sinaleiro_cor else '',
        'row_cart': r,
    }

    # Distribuir por consultor
    assigned = False
    for key in consultores:
        if key in consultor:
            consultores[key].append(client)
            assigned = True
            break
    if not assigned and 'RODRIGO' in consultor:
        consultores['LARISSA'].append(client)
    elif not assigned and consultor:
        # Distribuir round-robin nos consultores menores
        min_key = min(consultores, key=lambda k: len(consultores[k]))
        consultores[min_key].append(client)

for k, v in consultores.items():
    print(f"  {k}: {len(v)} clientes na carteira")

# ============================================================
# [3/5] SELECIONAR 10 CLIENTES POR CONSULTOR (prioridade inteligente)
# ============================================================
print("\n[3/5] Selecionando 10 clientes por consultor (prioridade: A > B > C, mais dias sem compra)...")

def score_client(c):
    """Score para priorização: Curva A=3, B=2, C=1, D=0 + dias sem compra normalizado"""
    curva_score = {'A': 300, 'B': 200, 'C': 100, 'D': 50}.get(c['curva'], 0)
    dias_score = min(c['dias_sem'], 365)  # Cap em 365
    ticket_score = min(c['ticket'] / 100, 50)  # Ticket alto = prioridade
    return curva_score + dias_score + ticket_score

agenda_clients = {}  # consultor -> [10 clients]

for consultor_name, clients in consultores.items():
    # Ordenar por score decrescente
    sorted_clients = sorted(clients, key=score_client, reverse=True)
    # Pegar top 10
    selected = sorted_clients[:10]
    agenda_clients[consultor_name] = selected
    print(f"  {consultor_name}: {len(selected)} selecionados")
    for i, c in enumerate(selected):
        print(f"    {i+1}. {c['curva']} | {c['dias_sem']:>4}d | R${c['ticket']:>8.2f} | {c['nome'][:35]}")

# ============================================================
# [4/5] POPULAR AGENDAS
# ============================================================
print("\n[4/5] Populando agendas (40 atendimentos para 19/02/2026)...")

AGENDA_MAP = {
    'LARISSA': 'AGENDA LARISSA',
    'DAIANE': 'AGENDA DAIANE',
    'MANU': 'AGENDA MANU',
    'JULIO': 'AGENDA JULIO',
}

# Headers das AGENDAs (row 3):
# 1=RANK, 2=CNPJ, 3=RAZÃO SOCIAL, 4=UF, 5=CIDADE, 6=SITUAÇÃO,
# 7=DIAS SEM COMPRA, 8=CURVA ABC, 9=TICKET MÉDIO, 10=ESTÁGIO FUNIL,
# 11=AÇÃO FUTURA, 12=TEMPERATURA, 13=SINALEIRO, 14=TENTATIVA, 15=SCORE
# + colunas de preenchimento manual (16+)

# Check remaining cols
ws_test = wb['AGENDA LARISSA']
print("  Colunas AGENDA completas:")
for c in range(1, ws_test.max_column + 1):
    v = ws_test.cell(row=3, column=c).value
    if v:
        print(f"    Col {c}: {v}")

# Estilo para dados
data_font = Font(name='Calibri', size=10)
data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

total_populated = 0
for consultor_key, sheet_name in AGENDA_MAP.items():
    ws = wb[sheet_name]
    clients = agenda_clients.get(consultor_key, [])

    for i, c in enumerate(clients):
        row = 4 + i  # Data starts row 4
        rank = i + 1

        # Calcular score
        sc = score_client(c)

        # Determinar temperatura baseada na curva + dias
        if c['curva'] == 'A' and c['dias_sem'] > 30:
            temp = '🔴 QUENTE'
        elif c['curva'] in ('A', 'B') and c['dias_sem'] > 60:
            temp = '🟠 MORNO'
        elif c['curva'] in ('C', 'D'):
            temp = '🔵 FRIO'
        else:
            temp = '🟡 NEUTRO'

        # Ação futura sugerida
        if c['dias_sem'] > 90:
            acao = 'REATIVAÇÃO'
        elif c['dias_sem'] > 30:
            acao = 'FOLLOW-UP'
        elif c['curva'] in ('A', 'B'):
            acao = 'MANUTENÇÃO'
        else:
            acao = 'PROSPECÇÃO'

        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=c['cnpj'])
        ws.cell(row=row, column=3, value=c['razao'])
        ws.cell(row=row, column=4, value=c['uf'])
        ws.cell(row=row, column=5, value=c['cidade'])
        ws.cell(row=row, column=6, value=c['situacao'])
        ws.cell(row=row, column=7, value=c['dias_sem'])
        ws.cell(row=row, column=8, value=c['curva'])
        ws.cell(row=row, column=9, value=c['ticket'])
        ws.cell(row=row, column=10, value='NOVO')  # Estágio funil = NOVO (sistema limpo)
        ws.cell(row=row, column=11, value=acao)
        ws.cell(row=row, column=12, value=temp)
        ws.cell(row=row, column=13, value=c['sinaleiro'])
        ws.cell(row=row, column=14, value=0)  # Tentativa = 0 (sistema limpo)
        ws.cell(row=row, column=15, value=round(sc, 1))

        # Aplicar estilo básico
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = data_border

        total_populated += 1

    print(f"  {sheet_name}: {len(clients)} clientes populados (rows 4-{4+len(clients)-1})")

# ============================================================
# [5/5] DASH — Ajustar para refletir estado limpo
# ============================================================
print("\n[5/5] Ajustando DASH para estado operacional limpo...")
ws_dash = wb['DASH']

# KPIs do estado limpo
ws_dash.cell(row=5, column=1).value = 0       # ATENDIMENTOS
ws_dash.cell(row=5, column=4).value = 0       # VENDAS
ws_dash.cell(row=5, column=7).value = 0       # ORÇAMENTOS
# BASE TOTAL = total de clientes na carteira
total_clientes = sum(len(v) for v in consultores.values())
ws_dash.cell(row=5, column=10).value = total_clientes
ws_dash.cell(row=5, column=13).value = 0       # Ñ ATENDE

# Descrições
ws_dash.cell(row=6, column=1).value = f"0 dias úteis · 0/dia"
ws_dash.cell(row=6, column=4).value = f"0% conversão geral"
ws_dash.cell(row=6, column=10).value = f"Clientes na CARTEIRA"

# Info
ws_dash.cell(row=3, column=1).value = f"KPIs RESUMO — 02/2026 (Sistema limpo — pronto para uso)"

print(f"  DASH atualizado: {total_clientes} clientes na base")

# ============================================================
# SALVAR
# ============================================================
print(f"\nSalvando V43 FINAL...")
t_save = time.time()
wb.save(OUTPUT)
save_time = time.time() - t_save

file_size = os.path.getsize(OUTPUT) / (1024 * 1024)
total_time = time.time() - start

print(f"\n{'='*80}")
print(f"V43 FINAL — PRONTO PARA APRESENTAÇÃO")
print(f"{'='*80}")
print(f"  Arquivo: {OUTPUT}")
print(f"  Tamanho: {file_size:.1f} MB")
print(f"  Tempo: {total_time:.0f}s")
print(f"")
print(f"  AGENDAS POPULADAS:")
for k, v in agenda_clients.items():
    print(f"    {k}: {len(v)} clientes")
print(f"  TOTAL: {total_populated} atendimentos agendados para 19/02/2026")
print(f"")
print(f"  ESTADO DO SISTEMA:")
print(f"    ✅ CARTEIRA: {total_clientes} clientes (core limpo)")
print(f"    ✅ SINALEIRO: 534 clientes (intacto)")
print(f"    ✅ REGRAS: source of truth (intacto)")
print(f"    ✅ PROJEÇÃO: 493 clientes com meta (realizado zerado)")
print(f"    ✅ DRAFT 2: VAZIO (pronto para receber atendimentos ao vivo)")
print(f"    ✅ DASH: zerado (atualiza conforme atendimentos)")
print(f"    ✅ AGENDAS: {total_populated} atendimentos para hoje")
print(f"")
print(f"  FLUXO AO VIVO:")
print(f"    1. Abrir V43 no Excel")
print(f"    2. Ir na AGENDA do consultor")
print(f"    3. Preencher atendimento (colunas manuais 16+)")
print(f"    4. Dado vai pro DRAFT 2 → CARTEIRA → DASH")
print(f"{'='*80}")

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V15 COMPLETO: Corrige arquitetura para match V31.

Correcoes:
1. AGENDA tabs: reconstruir com 31 colunas (= layout DRAFT 2) + formulas SORTBY/FILTER
2. DRAFT 3: copiar do V31 (dados SAP)
3. SINALEIRO: copiar do V31
4. Headers alinhados V31
5. DRAFT 2 ja enriquecido no passo anterior
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import shutil
import re
from pathlib import Path
from datetime import datetime

V15_PREV = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V15_FINAL.xlsx"
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"
V15_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V15_FINAL.xlsx"


def copy_sheet_from_v31(wb_v31, wb_target, sheet_name, position=None):
    """Copia uma aba inteira do V31 para o target"""
    if sheet_name not in wb_v31.sheetnames:
        print(f"  ! {sheet_name} nao existe no V31")
        return False

    ws_src = wb_v31[sheet_name]

    # Se ja existe no target, remover
    if sheet_name in wb_target.sheetnames:
        del wb_target[sheet_name]

    ws_dst = wb_target.create_sheet(sheet_name)

    # Copiar dados e formatos
    for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, max_col=ws_src.max_column):
        for cell in row:
            new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    # Copiar dimensoes de colunas
    for col_key in ws_src.column_dimensions:
        src_dim = ws_src.column_dimensions[col_key]
        dst_dim = ws_dst.column_dimensions[col_key]
        dst_dim.width = src_dim.width
        if src_dim.outline_level:
            dst_dim.outline_level = src_dim.outline_level
        dst_dim.hidden = src_dim.hidden

    # Copiar dimensoes de linhas
    for row_key in ws_src.row_dimensions:
        src_dim = ws_src.row_dimensions[row_key]
        dst_dim = ws_dst.row_dimensions[row_key]
        dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden

    # Copiar merged cells
    for merged in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merged))

    # Copiar freeze/filter
    ws_dst.freeze_panes = ws_src.freeze_panes
    if ws_src.auto_filter and ws_src.auto_filter.ref:
        ws_dst.auto_filter.ref = ws_src.auto_filter.ref

    # Copiar sheet state
    ws_dst.sheet_state = ws_src.sheet_state

    # Mover para posicao
    if position is not None:
        wb_target.move_sheet(sheet_name, offset=position - wb_target.sheetnames.index(sheet_name))

    return True


def rebuild_agenda_tabs(wb, draft2_headers):
    """
    Reconstroi AGENDA tabs com layout IDENTICO ao DRAFT 2 (31 colunas).
    Cada AGENDA puxa da CARTEIRA via SORTBY(FILTER()) filtrado por consultor.
    """
    # DRAFT 2 headers (row 2, 31 cols)
    consultores = {
        "AGENDA LARISSA": "LARISSA PADILHA",
        "AGENDA DAIANE": "DAIANE STAVICKI",
        "AGENDA MANU": '"MANU DITZEL","HEMANUELE DITZEL (MANU)"',
        "AGENDA JULIO": "JULIO GADRET"
    }

    # Mapeamento: DRAFT 2 col → CARTEIRA col (de onde puxar)
    # DRAFT 2 headers vs CARTEIRA headers
    d2_to_carteira = {
        1: "AT",    # A: DATA → CARTEIRA AT (DATA ULT ATENDIMENTO)
        2: "L",     # B: CONSULTOR → CARTEIRA L
        3: "A",     # C: NOME FANTASIA → CARTEIRA A
        4: "B",     # D: CNPJ → CARTEIRA B
        5: "D",     # E: UF → CARTEIRA D
        6: "I",     # F: REDE/REGIONAL → CARTEIRA I
        7: "N",     # G: SITUAÇÃO → CARTEIRA N
        8: "P",     # H: DIAS SEM COMPRA → CARTEIRA P
        9: "AR",    # I: ESTÁGIO FUNIL → CARTEIRA AR
        10: "AX",   # J: TIPO CLIENTE → CARTEIRA AX
        11: "AZ",   # K: FASE → CARTEIRA AZ
        12: "BJ",   # L: SINALEIRO → CARTEIRA BJ
        13: "AY",   # M: TENTATIVA → CARTEIRA AY
        14: None,   # N: WHATSAPP → manual (vazio)
        15: None,   # O: LIGAÇÃO → manual
        16: None,   # P: LIGAÇÃO ATENDIDA → manual
        17: None,   # Q: TIPO DO CONTATO → manual
        18: "AV",   # R: RESULTADO → CARTEIRA AV
        19: "AW",   # S: MOTIVO → CARTEIRA AW
        20: "AS",   # T: FOLLOW-UP → CARTEIRA AS (PRÓX FOLLOWUP)
        21: "AU",   # U: AÇÃO FUTURA → CARTEIRA AU
        22: "BI",   # V: AÇÃO DETALHADA → CARTEIRA BI
        23: None,   # W: MERCOS ATUALIZADO → manual
        24: None,   # X: NOTA DO DIA → manual
        25: "BB",   # Y: TEMPERATURA → CARTEIRA BB
        26: None,   # Z: GRUPO DASH → calculado
        27: None,   # AA: SINALEIRO META → manual
        28: None,   # AB: TIPO AÇÃO → manual
        29: None,   # AC: TIPO PROBLEMA → manual
        30: None,   # AD: DEMANDA → manual
        31: None,   # AE: TIPO ATENDIMENTO → manual
    }

    for agenda_name, consultor_filter in consultores.items():
        print(f"\n  Reconstruindo {agenda_name}...")

        # Remover aba existente
        if agenda_name in wb.sheetnames:
            del wb[agenda_name]

        ws = wb.create_sheet(agenda_name)

        # Row 1: Titulo
        ws.cell(row=1, column=1, value=agenda_name)
        ws.cell(row=1, column=5, value="DATA FILTRO:")
        ws.cell(row=1, column=6, value="=TODAY()")

        # Row 2: Headers (= DRAFT 2 row 2)
        for col_idx, header in enumerate(draft2_headers, 1):
            ws.cell(row=2, column=col_idx, value=header)

        # Row 3 em diante: Formulas SORTBY(FILTER(CARTEIRA...))
        # Primeira coluna com dados (col A = DATA do atendimento)
        # Filter criteria: CARTEIRA!L (consultor) = nome + CARTEIRA!AS (followup) >= TODAY()

        # Construir formula de filtro para cada coluna
        max_data_rows = 1000  # pre-alocar 1000 rows

        # Build the FILTER condition
        if '"' in consultor_filter:
            # MANU: OR condition
            names = consultor_filter.replace('"', '').split(',')
            filter_cond = f'(CARTEIRA!$L$4:$L$25000="{names[0]}")+(CARTEIRA!$L$4:$L$25000="{names[1]}")'
        else:
            filter_cond = f'CARTEIRA!$L$4:$L$25000="{consultor_filter}"'

        # Para cada coluna do DRAFT 2, criar formula que puxa da CARTEIRA
        for d2_col in range(1, len(draft2_headers) + 1):
            cart_col = d2_to_carteira.get(d2_col)
            if cart_col:
                # Formula: =IFERROR(SORTBY(FILTER(CARTEIRA!{col}$4:{col}$25000, {filter}), CARTEIRA!$O$4:$O$25000, -1), "")
                # Sort by SCORE (col O) descending
                formula = (
                    f'=IFERROR(SORTBY(FILTER(CARTEIRA!${cart_col}$4:${cart_col}$25000,'
                    f'{filter_cond}),'
                    f'FILTER(CARTEIRA!$O$4:$O$25000,{filter_cond}),-1),"")'
                )
                ws.cell(row=3, column=d2_col, value=formula)
            # Colunas manuais ficam vazias (consultor preenche)

        # Freeze panes
        ws.freeze_panes = "D3"

        # Column widths (copiar do DRAFT 2 ou usar defaults razoaveis)
        default_widths = {
            1: 12, 2: 18, 3: 30, 4: 18, 5: 5, 6: 14, 7: 14, 8: 12,
            9: 14, 10: 14, 11: 12, 12: 10, 13: 10, 14: 8, 15: 8, 16: 10,
            17: 14, 18: 14, 19: 14, 20: 12, 21: 22, 22: 25, 23: 10, 24: 25,
            25: 12, 26: 10, 27: 10, 28: 12, 29: 12, 30: 14, 31: 14
        }
        for col_idx, width in default_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Outline grouping: cols I-N (9-14) at level 1 (detalhe oculto)
        for col_idx in range(9, 15):
            ws.column_dimensions[get_column_letter(col_idx)].outline_level = 1
            ws.column_dimensions[get_column_letter(col_idx)].hidden = False

        print(f"    {len(draft2_headers)} colunas, formula SORTBY/FILTER por '{consultor_filter}'")


def main():
    print("=" * 100)
    print("V15 COMPLETO — Arquitetura corrigida (DRAFT 3 + SINALEIRO + AGENDA 31 cols)")
    print("=" * 100)

    # Carregar V31 (formulas para copiar DRAFT 3 e SINALEIRO)
    print("\n[1/6] Carregando V31...")
    wb_v31 = openpyxl.load_workbook(V31_PATH, data_only=False)

    # Carregar V15 anterior
    print("[2/6] Carregando V15 (anterior)...")
    wb = openpyxl.load_workbook(V15_PATH)

    changes = []

    # ==========================================
    # FASE 1: COPIAR DRAFT 3 do V31
    # ==========================================
    print("\n[3/6] ADICIONANDO DRAFT 3 (SAP)...")
    if "DRAFT 3 " in wb_v31.sheetnames:
        sheet_name = "DRAFT 3 "
    elif "DRAFT 3" in wb_v31.sheetnames:
        sheet_name = "DRAFT 3"
    else:
        sheet_name = None
        for s in wb_v31.sheetnames:
            if "DRAFT 3" in s.upper():
                sheet_name = s
                break

    if sheet_name:
        success = copy_sheet_from_v31(wb_v31, wb, sheet_name)
        if success:
            ws = wb[sheet_name]
            msg = f"  DRAFT 3 copiado do V31: {ws.max_row} rows x {ws.max_column} cols"
            print(msg)
            changes.append(msg)
    else:
        print("  ! DRAFT 3 nao encontrado no V31")

    # ==========================================
    # FASE 2: COPIAR SINALEIRO do V31
    # ==========================================
    print("\n[4/6] ADICIONANDO SINALEIRO...")
    if "SINALEIRO" in wb_v31.sheetnames:
        success = copy_sheet_from_v31(wb_v31, wb, "SINALEIRO")
        if success:
            ws = wb["SINALEIRO"]
            msg = f"  SINALEIRO copiado do V31: {ws.max_row} rows x {ws.max_column} cols"
            print(msg)
            changes.append(msg)

    # ==========================================
    # FASE 3: RECONSTRUIR AGENDA (31 colunas = DRAFT 2)
    # ==========================================
    print("\n[5/6] RECONSTRUINDO AGENDA (layout = DRAFT 2)...")

    # Pegar headers do DRAFT 2
    ws_d2 = wb["DRAFT 2"]
    draft2_headers = []
    for col in range(1, ws_d2.max_column + 1):
        h = ws_d2.cell(row=2, column=col).value
        if h:
            draft2_headers.append(str(h).strip())
        else:
            draft2_headers.append(f"COL_{get_column_letter(col)}")

    print(f"  DRAFT 2 headers ({len(draft2_headers)} cols): {draft2_headers[:10]}...")

    rebuild_agenda_tabs(wb, draft2_headers)
    msg = f"  4 AGENDAs reconstruidas com {len(draft2_headers)} colunas (= layout DRAFT 2)"
    changes.append(msg)

    # ==========================================
    # FASE 4: ORGANIZAR ORDEM DAS ABAS
    # ==========================================
    print("\n[6/6] ORGANIZANDO ABAS...")

    # Ordem desejada (inspirada no V31):
    # SINALEIRO, PROJEÇÃO, LOG, DASH, REGRAS, DRAFT 1, DRAFT 2, DRAFT 3, CARTEIRA, AGENDA x4, extras
    desired_order = [
        "SINALEIRO",
        "PROJEÇÃO ",
        "LOG",
        "DASH",
        "REDES_FRANQUIAS_v2",
        "COMITE",
        "REGRAS",
        "DRAFT 1",
        "DRAFT 2",
    ]
    # Add DRAFT 3 variants
    for s in wb.sheetnames:
        if "DRAFT 3" in s.upper() and s not in desired_order:
            desired_order.append(s)

    desired_order.extend([
        "CARTEIRA",
        "AGENDA LARISSA",
        "AGENDA DAIANE",
        "AGENDA MANU",
        "AGENDA JULIO",
    ])

    # Re-order sheets
    current = wb.sheetnames[:]
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            current_idx = wb.sheetnames.index(name)
            if current_idx != i:
                wb.move_sheet(name, offset=i - current_idx)

    msg = f"  Abas reordenadas: {wb.sheetnames}"
    print(msg)
    changes.append(msg)

    # ==========================================
    # SALVAR
    # ==========================================
    print("\n[SALVANDO]...")
    wb.save(V15_PATH)
    wb.close()
    wb_v31.close()

    size = Path(V15_PATH).stat().st_size / (1024 * 1024)
    print(f"  Salvo: {V15_PATH} ({size:.2f} MB)")

    print(f"\n{'='*100}")
    print(f"CHANGELOG V15 ANTERIOR → V15 COMPLETO ({len(changes)} correcoes)")
    print(f"{'='*100}")
    for c in changes:
        print(f"  {c}")

    # Verificacao rapida
    print(f"\n[VERIFICACAO]")
    wb_check = openpyxl.load_workbook(V15_PATH, data_only=False)
    total_formulas = 0
    for name in wb_check.sheetnames:
        ws = wb_check[name]
        f_count = 0
        for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    f_count += 1
        total_formulas += f_count
        print(f"  {name:25s} | {ws.max_row:>6} rows | {ws.max_column:>4} cols | {f_count:>7} formulas | freeze={ws.freeze_panes}")
    print(f"\n  TOTAL FORMULAS: {total_formulas:,}")
    print(f"  TOTAL ABAS: {len(wb_check.sheetnames)}")
    wb_check.close()

    print(f"\n[SUCESSO] V15 FINAL COMPLETO!")


if __name__ == "__main__":
    main()

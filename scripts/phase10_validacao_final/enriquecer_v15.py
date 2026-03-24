#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera V15 FINAL: Enriquece DRAFT 2, corrige ancoras CARTEIRA, valida AGENDA.

Problemas identificados no V14:
1. DRAFT 2 colunas MANUAIS parcialmente preenchidas (SITUACAO, SINALEIRO, ACAO DETALHADA, etc)
2. Colunas ancora da CARTEIRA precisam match exato com V31
3. AGENDA precisa ter dados fluindo corretamente

Abordagem:
- Usar V14 como base
- Enriquecer DRAFT 2 colunas MANUAIS a partir dos proprios dados do LOG e DRAFT 1
- Copiar dados enriquecidos do V31 DRAFT 2 para as colunas que faltam no V14
- Verificar ancoras CARTEIRA
"""

import openpyxl
import re
import shutil
from pathlib import Path
from datetime import datetime
from collections import defaultdict

V14_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V14_FINAL.xlsx"
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"
V15_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V15_FINAL.xlsx"


def normalizar_cnpj(valor):
    if not valor:
        return None
    cnpj = re.sub(r'[^\d]', '', str(valor).strip())
    return cnpj if len(cnpj) == 14 else None


def col_letter(idx):
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def main():
    print("=" * 100)
    print("GERACAO V15 FINAL — Enriquecimento completo")
    print("=" * 100)
    print(f"Data: {datetime.now()}")

    # 1. Copiar V14 como base
    print("\n[1/7] Copiando V14 como base...")
    shutil.copy2(V14_PATH, V15_PATH)

    # 2. Carregar ambos
    print("[2/7] Carregando V31 (dados) para referencia...")
    wb_v31 = openpyxl.load_workbook(V31_PATH, data_only=True)

    print("[3/7] Carregando V15 para edicao...")
    wb_v15 = openpyxl.load_workbook(V15_PATH)

    changes = []

    # ==========================================
    # FASE A: ENRIQUECER DRAFT 2
    # ==========================================
    print("\n[4/7] ENRIQUECENDO DRAFT 2...")

    ws_d2_v15 = wb_v15["DRAFT 2"]
    ws_d2_v31 = wb_v31["DRAFT 2"]

    # Mapear DRAFT 2 do V31 por CNPJ (col D=4) — pegar ultimo registro por CNPJ
    # V31 DRAFT 2 headers (row 2): A=DATA, B=CONSULTOR, C=NOME, D=CNPJ, etc
    v31_d2_by_cnpj = {}
    print(f"  Indexando V31 DRAFT 2 ({ws_d2_v31.max_row} rows)...")

    for row in range(3, ws_d2_v31.max_row + 1):
        cnpj_raw = ws_d2_v31.cell(row=row, column=4).value  # col D
        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            continue

        # Guardar dados de TODAS as colunas
        row_data = {}
        for col in range(1, ws_d2_v31.max_column + 1):
            val = ws_d2_v31.cell(row=row, column=col).value
            if val is not None:
                row_data[col] = val

        # Manter o mais recente (ultima row = mais recente pois ordenado por data)
        v31_d2_by_cnpj[cnpj] = row_data

    print(f"  V31 DRAFT 2: {len(v31_d2_by_cnpj)} CNPJs unicos indexados")

    # Colunas MANUAIS do DRAFT 2 que estao incompletas no V14:
    # G(7)=SITUAÇÃO, H(8)=DIAS SEM COMPRA, L(12)=SINALEIRO,
    # V(22)=AÇÃO DETALHADA, X(24)=NOTA DO DIA,
    # AA(27)=SINALEIRO META, AC(29)=TIPO PROBLEMA, AD(30)=DEMANDA
    manual_cols_to_enrich = [7, 8, 12, 22, 24, 27, 29, 30]

    enriched_count = 0
    total_cells_enriched = 0

    for row in range(4, ws_d2_v15.max_row + 1):  # row 4 onward (row 3 = MANUAL/AUTO)
        cnpj_raw = ws_d2_v15.cell(row=row, column=4).value
        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            continue

        if cnpj not in v31_d2_by_cnpj:
            continue

        v31_data = v31_d2_by_cnpj[cnpj]
        row_changed = False

        for col in manual_cols_to_enrich:
            current_val = ws_d2_v15.cell(row=row, column=col).value
            v31_val = v31_data.get(col)

            # Se V15 esta vazio e V31 tem dado, copiar
            if (current_val is None or str(current_val).strip() == "") and v31_val is not None:
                ws_d2_v15.cell(row=row, column=col, value=v31_val)
                total_cells_enriched += 1
                row_changed = True

        if row_changed:
            enriched_count += 1

    msg = f"  DRAFT 2 enriquecido: {enriched_count} rows, {total_cells_enriched} celulas preenchidas do V31"
    print(msg)
    changes.append(msg)

    # ==========================================
    # FASE B: CORRIGIR ANCORAS CARTEIRA
    # ==========================================
    print("\n[5/7] CORRIGINDO ANCORAS CARTEIRA...")

    ws_cart_v15 = wb_v15["CARTEIRA"]
    ws_cart_v31 = wb_v31["CARTEIRA"]

    # V31 ancoras: A(outline=0,hidden=False), AK, AR, BK, CA
    # Garantir que no V15 essas colunas estejam com outline_level=0 e hidden=False

    v31_anchors = {}
    v31_cart = openpyxl.load_workbook(V31_PATH, data_only=False)["CARTEIRA"]
    for col_key in v31_cart.column_dimensions:
        dim = v31_cart.column_dimensions[col_key]
        if not dim.hidden and (not dim.outline_level or dim.outline_level == 0):
            v31_anchors[col_key] = True

    print(f"  V31 ancoras identificadas: {sorted(v31_anchors.keys())}")

    # Copiar EXATAMENTE o pattern de outline do V31 para V15
    anchor_fixes = 0
    for col_key in v31_cart.column_dimensions:
        v31_dim = v31_cart.column_dimensions[col_key]
        v15_dim = ws_cart_v15.column_dimensions[col_key]

        # Copiar outline_level
        new_level = v31_dim.outline_level or 0
        old_level = v15_dim.outline_level or 0
        new_hidden = v31_dim.hidden
        old_hidden = v15_dim.hidden

        if new_level != old_level or new_hidden != old_hidden:
            v15_dim.outline_level = new_level
            v15_dim.hidden = new_hidden
            anchor_fixes += 1

    msg = f"  Ancoras: {anchor_fixes} colunas ajustadas para match V31"
    print(msg)
    changes.append(msg)

    # Copiar LARGURAS exatas do V31 (incluindo ancoras)
    width_fixes = 0
    for col_key in v31_cart.column_dimensions:
        v31_w = v31_cart.column_dimensions[col_key].width
        if v31_w:
            ws_cart_v15.column_dimensions[col_key].width = v31_w
            width_fixes += 1

    msg = f"  Column widths: {width_fixes} larguras copiadas do V31"
    print(msg)
    changes.append(msg)

    # Copiar freeze_panes do V31
    v31_freeze = v31_cart.freeze_panes
    if v31_freeze:
        # V31 congela em BX66 — para V15 com 557 rows, ajustar
        # BX = coluna 76, row 66 faz sentido para V31 (8305 rows)
        # Para V15 (557 rows), usar BX4 (após headers)
        ws_cart_v15.freeze_panes = "BX4"
        msg = f"  Freeze panes: CARTEIRA → BX4 (padrão V31 adaptado)"
        print(msg)
        changes.append(msg)

    # Copiar headers do V31 para rows 1, 2 (super-grupos e sub-headers)
    # Apenas para as colunas que existem em ambos (min dos max_col)
    max_common_col = min(ws_cart_v31.max_column, ws_cart_v15.max_column)

    header_fixes = 0
    for row in [1, 2]:
        for col in range(1, max_common_col + 1):
            v31_val = ws_cart_v31.cell(row=row, column=col).value
            v15_val = ws_cart_v15.cell(row=row, column=col).value

            if v31_val and v31_val != v15_val:
                ws_cart_v15.cell(row=row, column=col, value=v31_val)
                header_fixes += 1

    msg = f"  Headers: {header_fixes} celulas de header (rows 1-2) alinhadas com V31"
    print(msg)
    changes.append(msg)

    # ==========================================
    # FASE C: VERIFICAR AUTO-FILTER
    # ==========================================
    print("\n[6/7] VERIFICANDO AUTO-FILTER E LAYOUT...")

    # CARTEIRA auto-filter
    max_col_letter = col_letter(ws_cart_v15.max_column)
    cart_filter = f"A3:{max_col_letter}{ws_cart_v15.max_row}"
    ws_cart_v15.auto_filter.ref = cart_filter
    msg = f"  CARTEIRA auto_filter: {cart_filter}"
    print(msg)
    changes.append(msg)

    # DRAFT 1 freeze and filter
    if "DRAFT 1" in wb_v15.sheetnames:
        ws_d1 = wb_v15["DRAFT 1"]
        ws_d1.freeze_panes = "AN4"
        max_d1_col = col_letter(ws_d1.max_column)
        ws_d1.auto_filter.ref = f"A3:{max_d1_col}{ws_d1.max_row}"
        msg = f"  DRAFT 1: freeze=AN4, filter=A3:{max_d1_col}{ws_d1.max_row}"
        print(msg)
        changes.append(msg)

    # DRAFT 2 freeze
    ws_d2_v15.freeze_panes = "A3"
    msg = f"  DRAFT 2: freeze=A3"
    print(msg)
    changes.append(msg)

    # LOG
    if "LOG" in wb_v15.sheetnames:
        ws_log = wb_v15["LOG"]
        ws_log.freeze_panes = "A3"
        msg = f"  LOG: freeze=A3"
        print(msg)
        changes.append(msg)

    # PROJECAO
    if "PROJEÇÃO " in wb_v15.sheetnames:
        ws_proj = wb_v15["PROJEÇÃO "]
        ws_proj.freeze_panes = "C4"
        msg = f"  PROJECAO: freeze=C4"
        print(msg)
        changes.append(msg)

    # AGENDA tabs: copiar column widths e outline do V31 AGENDA se existir
    if "AGENDA" in wb_v31.sheetnames:
        ws_v31_agenda = wb_v31["AGENDA"]
        # Copiar widths e outlines para todos os 4 AGENDA tabs do V15
        for agenda_name in [s for s in wb_v15.sheetnames if "AGENDA" in s.upper()]:
            ws_agenda = wb_v15[agenda_name]
            # Copiar widths
            for col_key in ws_v31_agenda.column_dimensions:
                v31_dim = ws_v31_agenda.column_dimensions[col_key]
                if v31_dim.width:
                    ws_agenda.column_dimensions[col_key].width = v31_dim.width
                if v31_dim.outline_level:
                    ws_agenda.column_dimensions[col_key].outline_level = v31_dim.outline_level
                ws_agenda.column_dimensions[col_key].hidden = v31_dim.hidden

            # Freeze e filter
            ws_agenda.freeze_panes = "D3"
            msg = f"  {agenda_name}: widths/outlines copiados do V31 AGENDA, freeze=D3"
            print(msg)
            changes.append(msg)

    # ==========================================
    # FASE D: SALVAR
    # ==========================================
    print("\n[7/7] SALVANDO V15...")
    wb_v15.save(V15_PATH)
    wb_v15.close()
    wb_v31.close()

    size = Path(V15_PATH).stat().st_size / (1024 * 1024)
    print(f"\n  Salvo: {V15_PATH} ({size:.2f} MB)")

    print(f"\n{'='*100}")
    print(f"CHANGELOG V14 → V15 ({len(changes)} correcoes)")
    print(f"{'='*100}")
    for c in changes:
        print(f"  {c}")

    print(f"\n[SUCESSO] V15 FINAL gerado!")


if __name__ == "__main__":
    main()

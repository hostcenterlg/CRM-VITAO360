#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Diagnostico CARTEIRA V16: quais colunas estao preenchidas vs vazias.
Foco no FUNIL (AR-BJ) e CONSULTOR (L).
"""

import openpyxl
import re

V16_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V16_FINAL.xlsx"

def col_letter(idx):
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result

wb = openpyxl.load_workbook(V16_PATH, data_only=False)
ws = wb["CARTEIRA"]

print(f"CARTEIRA: {ws.max_row} rows x {ws.max_column} cols")
print(f"Headers row 1-3:\n")

# Headers completos
for header_row in [1, 2, 3]:
    headers = []
    for col in range(1, min(ws.max_column + 1, 270)):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers.append((col, col_letter(col), str(val).strip()[:35]))
    if headers:
        print(f"  Row {header_row} ({len(headers)} valores):")
        for col_idx, letter, val in headers[:80]:
            print(f"    {letter:>4} (col {col_idx:>3}): {val}")
        if len(headers) > 80:
            print(f"    ... +{len(headers)-80} mais")

# Preenchimento COMPLETO por coluna (primeiras 50 rows de dados)
print(f"\n{'='*100}")
print(f"PREENCHIMENTO POR COLUNA (rows 4-53)")
print(f"{'='*100}")

end_row = min(ws.max_row, 53)
for col in range(1, min(ws.max_column + 1, 270)):
    filled = 0
    formula_count = 0
    static_count = 0
    sample = None
    for row in range(4, end_row + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None and str(val).strip() != "":
            filled += 1
            if isinstance(val, str) and val.startswith("="):
                formula_count += 1
            else:
                static_count += 1
            if sample is None:
                sample = str(val)[:60]

    letter = col_letter(col)
    h3 = ws.cell(row=3, column=col).value or ""
    h2 = ws.cell(row=2, column=col).value or ""
    header = str(h3)[:25] if h3 else str(h2)[:25]
    pct = round(100 * filled / max(1, end_row - 3), 1)

    if filled > 0 or header:
        tipo = "FORMULA" if formula_count > 0 else ("DADOS" if static_count > 0 else "VAZIO")
        bar = "#" * int(pct / 5) + "." * (20 - int(pct / 5))
        print(f"  {letter:>4} | {header:25s} | {bar} {pct:>5.1f}% | {tipo:8s} | f={formula_count:>3} d={static_count:>3} | {sample or '[vazio]'}")

# Focar no FUNIL (AR=44 a BJ=62)
print(f"\n{'='*100}")
print(f"FOCO: FUNIL (AR-BJ) + CONSULTOR (L)")
print(f"{'='*100}")

# Col L = CONSULTOR
l_val = ws.cell(row=4, column=12).value
print(f"\n  CONSULTOR (col L, row 4): {l_val}")
l_val5 = ws.cell(row=5, column=12).value
print(f"  CONSULTOR (col L, row 5): {l_val5}")

# FUNIL cols
for col in range(44, 63):
    letter = col_letter(col)
    h3 = ws.cell(row=3, column=col).value or ""
    val4 = ws.cell(row=4, column=col).value
    val5 = ws.cell(row=5, column=col).value
    print(f"\n  {letter} ({h3}):")
    print(f"    Row 4: {str(val4)[:80] if val4 else '[VAZIO]'}")
    print(f"    Row 5: {str(val5)[:80] if val5 else '[VAZIO]'}")

# Verificar DRAFT 2 col D (CNPJ) vs CARTEIRA col B (CNPJ)
ws_d2 = wb["DRAFT 2"]
print(f"\n\n{'='*100}")
print(f"DRAFT 2 STRUCTURE CHECK")
print(f"{'='*100}")
print(f"  DRAFT 2 rows: {ws_d2.max_row}, cols: {ws_d2.max_column}")
print(f"  Row 1, col 1: {ws_d2.cell(row=1, column=1).value}")
print(f"  Row 2, col 4: {ws_d2.cell(row=2, column=4).value}")  # CNPJ header
print(f"  Row 3, col 4: {ws_d2.cell(row=3, column=4).value}")  # first CNPJ data

# CARTEIRA col B = CNPJ
print(f"\n  CARTEIRA col B (CNPJ):")
print(f"    Row 3 (header): {ws.cell(row=3, column=2).value}")
print(f"    Row 4 (data): {ws.cell(row=4, column=2).value}")
print(f"    Row 5 (data): {ws.cell(row=5, column=2).value}")

# DRAFT 1 check
ws_d1 = wb["DRAFT 1"]
print(f"\n  DRAFT 1 rows: {ws_d1.max_row}, cols: {ws_d1.max_column}")
print(f"  Row 3, col 2: {ws_d1.cell(row=3, column=2).value}")  # header or data?
print(f"  Row 4, col 2: {ws_d1.cell(row=4, column=2).value}")  # CNPJ?

# Check what DRAFT 1 AP column has (TIPO CLIENTE source)
print(f"\n  DRAFT 1 col AP (col 42):")
print(f"    Row 3: {ws_d1.cell(row=3, column=42).value}")
print(f"    Row 4: {ws_d1.cell(row=4, column=42).value}")

wb.close()
print(f"\n[DONE]")

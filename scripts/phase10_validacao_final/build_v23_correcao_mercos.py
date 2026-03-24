#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V23 — CORREÇÃO TOTAL COM CARTEIRA MERCOS OFICIAL

Fonte: "Carteira detalhada de clientes FEVEREIRO 16.02 .xlsx" (Downloads)
502 clientes com vendas Jan/2025 a Fev/2026.

AÇÕES:
1. Corrigir 96 PROSPECTS → ATIVO/INATIVO (Mercos diz a situação real)
2. Atualizar dados: nome, razão, cidade, UF, vendedor, dias s/ compra, ciclo, situação
3. Adicionar fórmulas para os 96 ex-prospects
4. Recalcular SCORE
5. Atualizar nomes no DRAFT 1

Cronômetro em tempo real.
"""

import openpyxl
from openpyxl.formula.translate import Translator
from datetime import datetime, date
from pathlib import Path
import re
import time
import shutil

MERCOS = r"C:/Users/User/Downloads/Carteira detalhada de clientes FEVEREIRO 16.02 .xlsx"
V22 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V22_FINAL.xlsx"
V23 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V23_FINAL.xlsx"


def norm(c):
    return re.sub(r"[^\d]", "", str(c).strip())


def timer(t0):
    e = time.time() - t0
    return f"[{e:.0f}s]"


def map_situacao_mercos(sit_mercos):
    """Mapeia situação do Mercos → situação do CRM"""
    s = str(sit_mercos).strip().upper()
    if "ATIVO" in s and "INAT" not in s:
        return "ATIVO"
    if "INATIVO RECENTE" in s or "INAT" in s and "REC" in s:
        return "INAT.REC"
    if "INATIVO" in s:
        return "INATIVO"
    return "ATIVO"  # default pra quem está na lista Mercos


def map_consultor_mercos(vendedor):
    """Mapeia vendedor do Mercos → consultor padronizado"""
    v = str(vendedor).strip().upper()
    if "LARISSA" in v:
        return "LARISSA PADILHA"
    if "MANU" in v:
        return "MANU DITZEL"
    if "DAIANE" in v or "CENTRAL" in v:
        return "DAIANE STAVICKI"
    if "JULIO" in v:
        return "JULIO GADRET"
    if "HELDER" in v:
        return "HELDER BRUNKOW"
    return "LARISSA PADILHA"  # default


def calc_score(sit, dias_sem, ciclo, curva, temp, tentativa, b2b):
    """Calcula SCORE simplificado para ex-prospects"""
    # URGÊNCIA (30%)
    try:
        p = float(dias_sem) if dias_sem else 0
        s = float(ciclo) if ciclo else 1
        urgencia = min(100, p / max(s, 1) * 100)
    except (ValueError, TypeError):
        urgencia = 0

    # VALOR (25%)
    abc = str(curva).strip().upper()
    abc_score = 100 if abc == "A" else (60 if abc == "B" else (30 if abc == "C" else 0))
    valor = min(100, abc_score)

    # FOLLOW-UP (20%) — sem data de follow-up, usar 0
    followup = 0

    # SINAL (15%)
    t_str = str(temp).strip().upper()
    if "QUENTE" in t_str:
        temp_score = 100
    elif "MORNO" in t_str:
        temp_score = 60
    elif "FRIO" in t_str:
        temp_score = 20
    else:
        temp_score = 20  # default frio
    b2b_bonus = 20 if str(b2b).strip().upper() == "SIM" else 0
    sinal = min(100, temp_score + b2b_bonus)

    # TENTATIVA (5%)
    tent = str(tentativa).strip().upper()
    tent_map = {"T1": 100, "T2": 80, "T3": 60, "T4": 40}
    tent_score = tent_map.get(tent, 20)

    # SITUAÇÃO (5%)
    sit_str = str(sit).strip().upper()
    sit_map = {"EM RISCO": 100, "ATIVO": 80, "INAT.REC": 70, "NOVO": 40, "INATIVO": 20}
    sit_score = sit_map.get(sit_str, 50)

    total = (urgencia * 0.30 + valor * 0.25 + followup * 0.20
             + sinal * 0.15 + tent_score * 0.05 + sit_score * 0.05)
    return round(total, 2)


def main():
    t0 = time.time()
    TOTAL = 8
    print("=" * 100)
    print("V23 — CORREÇÃO TOTAL COM CARTEIRA MERCOS OFICIAL")
    print(f"Início: {datetime.now().strftime('%H:%M:%S')}")
    print(f"Estimativa: ~8-10 minutos (8 etapas)")
    print("=" * 100)

    # ===== 1/8: LER CARTEIRA MERCOS =====
    print(f"\n{timer(t0)} [1/8] Lendo carteira detalhada Mercos...", flush=True)
    wb_m = openpyxl.load_workbook(MERCOS, read_only=True, data_only=True)
    ws_m = wb_m["Relatório"]

    # Headers na row 9: C1=Razão, C2=Fantasia, C3=CNPJ, C4=Cidade, C5=Estado,
    # C6=Ult pedido, C7=Data ult ped, C8=Vendedor, C9=Valor ult ped,
    # C10=Dias sem comprar, C11=Ciclo médio, C12=Situação, C13=CEP, C14=Segmento
    mercos = {}
    for row in ws_m.iter_rows(min_row=10, max_col=14, values_only=True):
        cnpj_raw = row[2]
        if not cnpj_raw:
            continue
        cnpj_n = norm(cnpj_raw)
        if len(cnpj_n) < 11:
            continue

        mercos[cnpj_n] = {
            "razao": str(row[0] or "").strip(),
            "fantasia": str(row[1] or "").strip(),
            "cnpj_fmt": str(cnpj_raw).strip(),
            "cidade": str(row[3] or "").strip(),
            "estado": str(row[4] or "").strip(),
            "ult_pedido_num": row[5],
            "ult_pedido_data": row[6],
            "vendedor": str(row[7] or "").strip(),
            "ult_pedido_valor": row[8],
            "dias_sem_compra": row[9],
            "ciclo_medio": row[10],
            "situacao_mercos": str(row[11] or "").strip(),
            "cep": str(row[12] or "").strip(),
            "segmento": str(row[13] or "").strip(),
        }

    wb_m.close()
    print(f"  {len(mercos)} clientes lidos do Mercos")

    # Estatísticas de situação
    sit_counts = {}
    for m in mercos.values():
        s = m["situacao_mercos"]
        sit_counts[s] = sit_counts.get(s, 0) + 1
    for s, c in sorted(sit_counts.items(), key=lambda x: -x[1]):
        print(f"    {s}: {c}")

    # ===== 2/8: COPIAR V22 → V23 =====
    print(f"\n{timer(t0)} [2/8] Copiando V22 → V23...", flush=True)
    shutil.copy2(V22, V23)
    size_v22 = Path(V22).stat().st_size / (1024 * 1024)
    print(f"  V22 copiado ({size_v22:.2f} MB)")

    # ===== 3/8: ABRIR E MAPEAR =====
    print(f"\n{timer(t0)} [3/8] Abrindo V23 e mapeando CNPJs...", flush=True)
    wb = openpyxl.load_workbook(V23)
    ws_cart = wb["CARTEIRA"]
    ws_d1 = wb["DRAFT 1"]

    # Mapear CARTEIRA
    cart_map = {}  # cnpj_n → row
    for r in range(4, ws_cart.max_row + 1):
        v = ws_cart.cell(row=r, column=2).value
        if v:
            cart_map[norm(v)] = r

    # Mapear DRAFT 1
    d1_map = {}
    for r in range(3, ws_d1.max_row + 1):
        v = ws_d1.cell(row=r, column=2).value
        if v:
            d1_map[norm(v)] = r

    print(f"  CARTEIRA: {len(cart_map)} | DRAFT 1: {len(d1_map)}")

    # ===== 4/8: IDENTIFICAR FÓRMULAS =====
    print(f"\n{timer(t0)} [4/8] Identificando fórmulas da row 4...", flush=True)
    formula_cols = {}
    for c in range(1, ws_cart.max_column + 1):
        v = ws_cart.cell(row=4, column=c).value
        if isinstance(v, str) and v.startswith("="):
            formula_cols[c] = v
    print(f"  {len(formula_cols)} colunas com fórmulas")

    first_prospect_row = 558

    # ===== 5/8: CRUZAR E CLASSIFICAR =====
    print(f"\n{timer(t0)} [5/8] Cruzando Mercos × CARTEIRA...", flush=True)

    prospects_to_fix = []  # (cnpj_n, cart_row, mercos_data)
    clients_to_update = []  # (cnpj_n, cart_row, mercos_data)
    not_in_cart = []

    for cnpj_n, m in mercos.items():
        if cnpj_n in cart_map:
            r = cart_map[cnpj_n]
            sit_atual = str(ws_cart.cell(row=r, column=14).value or "").upper()

            if "PROSPECT" in sit_atual:
                prospects_to_fix.append((cnpj_n, r, m))
            else:
                clients_to_update.append((cnpj_n, r, m))
        else:
            not_in_cart.append((cnpj_n, m))

    print(f"  PROSPECTS a corrigir: {len(prospects_to_fix)}")
    print(f"  Clientes a atualizar: {len(clients_to_update)}")
    print(f"  Não na CARTEIRA:      {len(not_in_cart)}")
    if not_in_cart:
        for c, m in not_in_cart:
            print(f"    ⚠ {c} | {m['fantasia']} | {m['situacao_mercos']}")

    # ===== 6/8: CORRIGIR EX-PROSPECTS =====
    print(f"\n{timer(t0)} [6/8] Corrigindo {len(prospects_to_fix)} ex-prospects...", flush=True)

    formulas_added = 0
    range_fixes = 0

    for i, (cnpj_n, r, m) in enumerate(prospects_to_fix):
        new_sit = map_situacao_mercos(m["situacao_mercos"])
        new_consultor = map_consultor_mercos(m["vendedor"])

        # a) SITUAÇÃO (col 14)
        ws_cart.cell(row=r, column=14, value=new_sit)

        # b) CONSULTOR (col 12) — atualizar com vendedor real do Mercos
        ws_cart.cell(row=r, column=12, value=new_consultor)

        # c) NOME (col 1) e RAZÃO (col 3)
        ws_cart.cell(row=r, column=1, value=m["fantasia"])
        ws_cart.cell(row=r, column=3, value=m["razao"])

        # d) UF (col 4) e CIDADE (col 5)
        if m["estado"]:
            ws_cart.cell(row=r, column=4, value=m["estado"])
        if m["cidade"]:
            ws_cart.cell(row=r, column=5, value=m["cidade"])

        # e) DIAS SEM COMPRA (col 16) — valor real do Mercos
        if m["dias_sem_compra"] is not None:
            ws_cart.cell(row=r, column=16, value=m["dias_sem_compra"])

        # f) CICLO (col 19)
        if m["ciclo_medio"] is not None:
            ws_cart.cell(row=r, column=19, value=m["ciclo_medio"])

        # g) FUNIL — atualizar de "PRIMEIRO CONTATO" para algo mais adequado
        ws_cart.cell(row=r, column=47, value="RECOMPRA - REATIVAÇÃO")
        ws_cart.cell(row=r, column=54, value="🟡MORNO")

        # h) Adicionar fórmulas (se row >= 558, estava sem fórmulas)
        if r >= first_prospect_row:
            for col_num, base_formula in formula_cols.items():
                # Não sobrescrever dados que já atualizamos
                if col_num in (1, 2, 3, 4, 5, 12, 14, 16, 19, 47, 54):
                    continue
                try:
                    translated = Translator(base_formula, origin="A4").translate_formula(f"A{r}")
                    # Corrigir ranges
                    if "$557" in translated:
                        translated = translated.replace("$557", "$6147")
                    ws_cart.cell(row=r, column=col_num, value=translated)
                    formulas_added += 1
                except Exception:
                    pass

        # i) SCORE (col 15) — calcular e escrever valor estático
        score = calc_score(
            new_sit,
            m["dias_sem_compra"],
            m["ciclo_medio"],
            ws_cart.cell(row=r, column=39).value,
            "🟡MORNO",
            ws_cart.cell(row=r, column=51).value,
            ws_cart.cell(row=r, column=20).value,
        )
        ws_cart.cell(row=r, column=15, value=score)

        if (i + 1) % 20 == 0:
            print(f"  ... {i+1}/{len(prospects_to_fix)} {timer(t0)}", flush=True)

    print(f"  Corrigidos: {len(prospects_to_fix)}")
    print(f"  Fórmulas adicionadas: {formulas_added:,}")

    # ===== 7/8: ATUALIZAR CLIENTES EXISTENTES =====
    print(f"\n{timer(t0)} [7/8] Atualizando {len(clients_to_update)} clientes existentes...", flush=True)

    names_updated = 0
    sit_updated = 0
    consultor_updated = 0

    for i, (cnpj_n, r, m) in enumerate(clients_to_update):
        # Nome fantasia (col 1)
        old_name = str(ws_cart.cell(row=r, column=1).value or "")
        if m["fantasia"] and m["fantasia"] != old_name:
            ws_cart.cell(row=r, column=1, value=m["fantasia"])
            names_updated += 1

        # Razão social (col 3)
        ws_cart.cell(row=r, column=3, value=m["razao"])

        # Situação (col 14) — Mercos é a verdade
        new_sit = map_situacao_mercos(m["situacao_mercos"])
        old_sit = str(ws_cart.cell(row=r, column=14).value or "").upper()
        if new_sit != old_sit:
            ws_cart.cell(row=r, column=14, value=new_sit)
            sit_updated += 1

        # Consultor (col 12) — atualizar com vendedor real
        new_consultor = map_consultor_mercos(m["vendedor"])
        old_consultor = str(ws_cart.cell(row=r, column=12).value or "").upper()
        if new_consultor != old_consultor:
            ws_cart.cell(row=r, column=12, value=new_consultor)
            consultor_updated += 1

        # Dias sem compra (col 16) — somente se é valor estático
        v16 = ws_cart.cell(row=r, column=16).value
        if not (isinstance(v16, str) and v16.startswith("=")):
            if m["dias_sem_compra"] is not None:
                ws_cart.cell(row=r, column=16, value=m["dias_sem_compra"])

        # Ciclo (col 19) — somente se é valor estático
        v19 = ws_cart.cell(row=r, column=19).value
        if not (isinstance(v19, str) and v19.startswith("=")):
            if m["ciclo_medio"] is not None:
                ws_cart.cell(row=r, column=19, value=m["ciclo_medio"])

        if (i + 1) % 100 == 0:
            print(f"  ... {i+1}/{len(clients_to_update)} {timer(t0)}", flush=True)

    print(f"  Nomes atualizados: {names_updated}")
    print(f"  Situação atualizada: {sit_updated}")
    print(f"  Consultor atualizado: {consultor_updated}")

    # Atualizar DRAFT 1 nomes
    print(f"\n  Atualizando DRAFT 1...", flush=True)
    d1_updated = 0
    for cnpj_n, m in mercos.items():
        if cnpj_n in d1_map:
            r = d1_map[cnpj_n]
            ws_d1.cell(row=r, column=1, value=m["fantasia"])
            d1_updated += 1

    print(f"  DRAFT 1 nomes: {d1_updated}")

    # ===== 8/8: SALVAR =====
    print(f"\n{timer(t0)} [8/8] Salvando V23...", flush=True)
    wb.save(V23)
    wb.close()

    size_v23 = Path(V23).stat().st_size / (1024 * 1024)
    elapsed = time.time() - t0

    # ===== RELATÓRIO FINAL =====
    print(f"\n{'='*100}")
    print(f"[SUCESSO] V23 gerado em {elapsed:.0f}s ({elapsed/60:.1f} min)")
    print(f"{'='*100}")
    print(f"  Arquivo: {V23}")
    print(f"  Tamanho: {size_v23:.2f} MB (V22 era {size_v22:.2f} MB)")
    print(f"")
    print(f"  FONTE: Carteira detalhada Mercos Fev 16.02 ({len(mercos)} clientes)")
    print(f"")
    print(f"  CORREÇÕES:")
    print(f"    Ex-prospects corrigidos:   {len(prospects_to_fix)}")
    print(f"    Fórmulas adicionadas:      {formulas_added:,}")
    print(f"    Nomes CARTEIRA:            {names_updated}")
    print(f"    Situação atualizada:       {sit_updated}")
    print(f"    Consultor atualizado:      {consultor_updated}")
    print(f"    Nomes DRAFT 1:             {d1_updated}")
    print(f"    Não na CARTEIRA:           {len(not_in_cart)}")
    print(f"")
    print(f"  LISTA DOS {len(prospects_to_fix)} EX-PROSPECTS (agora clientes reais):")
    for cnpj_n, r, m in sorted(prospects_to_fix, key=lambda x: x[2]["fantasia"]):
        sit = map_situacao_mercos(m["situacao_mercos"])
        cons = map_consultor_mercos(m["vendedor"])
        print(f"    Row {r:5d} | {sit:10s} | {cons:20s} | {m['fantasia'][:45]}")
    print(f"{'='*100}")


if __name__ == "__main__":
    main()

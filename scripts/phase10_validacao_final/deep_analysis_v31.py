#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analise PROFUNDA do V31 para entender:
1. Colunas ancora da CARTEIRA (quais sao, o que contem, como se comportam)
2. FUNIL: quais colunas estao PREENCHIDAS com dados reais vs vazias
3. AGENDA: estrutura completa (headers, formulas, dados)
4. Fluxo de dados: de onde vem cada valor (DRAFT 2, LOG, REGRAS, etc)
"""

import openpyxl
import re
from collections import defaultdict

V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"
V14_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V14_FINAL.xlsx"


def col_letter(idx):
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def analyze_carteira_deep(ws, label, is_formula_mode=False):
    """Analise profunda da CARTEIRA"""
    print(f"\n{'='*100}")
    print(f"  CARTEIRA {label}")
    print(f"{'='*100}")

    # Row structure
    print(f"\n  Dimensoes: {ws.max_row} rows x {ws.max_column} cols")

    # Headers: rows 1, 2, 3
    print(f"\n  === HEADERS ===")
    for header_row in [1, 2, 3]:
        headers = []
        for col in range(1, min(ws.max_column + 1, 270)):
            val = ws.cell(row=header_row, column=col).value
            if val:
                headers.append((col_letter(col), str(val).strip()[:40]))
        print(f"\n  Row {header_row} ({len(headers)} valores):")
        for letter, val in headers[:60]:
            print(f"    {letter:>4}: {val}")
        if len(headers) > 60:
            print(f"    ... +{len(headers)-60} mais")

    # Colunas ancora: identificar colunas que NAO estao ocultas (hidden=False, outline_level=0)
    print(f"\n  === COLUNAS ANCORA (visiveis, sem grouping) ===")
    ancora_cols = []
    hidden_cols = []
    for col_dim_key in ws.column_dimensions:
        dim = ws.column_dimensions[col_dim_key]
        if not dim.hidden and (not dim.outline_level or dim.outline_level == 0):
            ancora_cols.append(col_dim_key)
        elif dim.hidden or (dim.outline_level and dim.outline_level > 0):
            hidden_cols.append((col_dim_key, dim.outline_level, dim.hidden))

    print(f"  Ancora (visiveis): {len(ancora_cols)} colunas")
    for c in sorted(ancora_cols)[:50]:
        # Pegar header dessa coluna
        col_idx = openpyxl.utils.column_index_from_string(c)
        h1 = ws.cell(row=1, column=col_idx).value or ""
        h2 = ws.cell(row=2, column=col_idx).value or ""
        h3 = ws.cell(row=3, column=col_idx).value or ""
        # Pegar sample de dados (row 4)
        sample = ws.cell(row=4, column=col_idx).value
        sample_str = str(sample)[:30] if sample else "[VAZIO]"
        print(f"    {c:>4} | R1: {str(h1)[:20]:20} | R2: {str(h2)[:20]:20} | R3: {str(h3)[:25]:25} | Sample: {sample_str}")
    if len(ancora_cols) > 50:
        print(f"    ... +{len(ancora_cols)-50} mais")

    print(f"\n  Ocultas (grouping): {len(hidden_cols)} colunas")

    # Preenchimento por coluna (primeiras 50 rows de dados)
    print(f"\n  === PREENCHIMENTO COLUNAS (rows 4-53, amostra 50 rows) ===")
    end_row = min(ws.max_row, 53)

    col_fill = {}
    for col in range(1, min(ws.max_column + 1, 270)):
        filled = 0
        formula_count = 0
        for row in range(4, end_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                filled += 1
                if isinstance(val, str) and val.startswith("="):
                    formula_count += 1
        if filled > 0:
            letter = col_letter(col)
            h3 = ws.cell(row=3, column=col).value or ws.cell(row=2, column=col).value or ""
            col_fill[letter] = {
                "header": str(h3)[:25],
                "filled": filled,
                "formulas": formula_count,
                "pct": round(100 * filled / (end_row - 3), 1)
            }

    # Agrupar por super-grupo (row 1)
    current_group = ""
    group_cols = defaultdict(list)
    for col in range(1, min(ws.max_column + 1, 270)):
        r1 = ws.cell(row=1, column=col).value
        if r1:
            current_group = str(r1).strip()[:30]
        letter = col_letter(col)
        if letter in col_fill:
            group_cols[current_group].append((letter, col_fill[letter]))

    for group, cols in group_cols.items():
        print(f"\n  [{group}]")
        for letter, info in cols:
            fill_bar = "#" * int(info["pct"] / 5) + "." * (20 - int(info["pct"] / 5))
            ftype = "FORMULA" if info["formulas"] > 0 else "DADOS"
            print(f"    {letter:>4} | {info['header']:25} | {fill_bar} {info['pct']:>5.1f}% | {ftype} ({info['filled']}/{end_row-3})")


def analyze_agenda_deep(wb, label):
    """Analise profunda da AGENDA"""
    print(f"\n{'='*100}")
    print(f"  AGENDA {label}")
    print(f"{'='*100}")

    agenda_sheets = [s for s in wb.sheetnames if "AGENDA" in s.upper()]
    print(f"\n  Abas AGENDA encontradas: {agenda_sheets}")

    for sheet_name in agenda_sheets:
        ws = wb[sheet_name]
        print(f"\n  --- {sheet_name} ---")
        print(f"  Dimensoes: {ws.max_row} rows x {ws.max_column} cols")
        print(f"  Freeze: {ws.freeze_panes}")
        print(f"  Filter: {ws.auto_filter.ref if ws.auto_filter else 'None'}")

        # Headers
        for header_row in [1, 2, 3]:
            headers = []
            for col in range(1, min(ws.max_column + 1, 80)):
                val = ws.cell(row=header_row, column=col).value
                if val:
                    headers.append((col_letter(col), str(val).strip()[:35]))
            if headers:
                print(f"\n  Row {header_row}:")
                for letter, val in headers:
                    print(f"    {letter:>4}: {val}")

        # Dados amostra (primeiras 5 rows de dados)
        data_start = 2
        # Encontrar onde comecam os dados
        for r in range(1, 10):
            v = ws.cell(row=r, column=1).value
            if v and isinstance(v, str) and any(k in str(v).upper() for k in ["CNPJ", "NOME", "CLIENTE", "FANTASIA"]):
                data_start = r + 1
                break

        print(f"\n  Dados a partir da row {data_start}:")
        for row in range(data_start, min(data_start + 3, ws.max_row + 1)):
            vals = []
            for col in range(1, min(ws.max_column + 1, 20)):
                v = ws.cell(row=row, column=col).value
                vals.append(str(v)[:20] if v else "[VAZIO]")
            print(f"    Row {row}: {' | '.join(vals[:10])}")

        # Preenchimento
        filled_cols = 0
        total_filled = 0
        for col in range(1, ws.max_column + 1):
            col_filled = 0
            for row in range(data_start, min(ws.max_row + 1, data_start + 50)):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    col_filled += 1
            if col_filled > 0:
                filled_cols += 1
                total_filled += col_filled

        print(f"\n  Colunas com dados: {filled_cols}/{ws.max_column}")
        print(f"  Celulas preenchidas (amostra): {total_filled}")

        # Outline/grouping
        outline_cols = []
        for c in ws.column_dimensions:
            dim = ws.column_dimensions[c]
            if dim.outline_level and dim.outline_level > 0:
                outline_cols.append((c, dim.outline_level, dim.hidden))
        if outline_cols:
            print(f"\n  Grouping: {len(outline_cols)} colunas")
            for c, level, hidden in outline_cols[:20]:
                print(f"    {c}: level={level}, hidden={hidden}")


def analyze_draft2_deep(wb, label):
    """Analise DRAFT 2"""
    print(f"\n{'='*100}")
    print(f"  DRAFT 2 {label}")
    print(f"{'='*100}")

    if "DRAFT 2" not in wb.sheetnames:
        print("  NAO ENCONTRADA")
        return

    ws = wb["DRAFT 2"]
    print(f"\n  Dimensoes: {ws.max_row} rows x {ws.max_column} cols")

    # Headers
    for header_row in [1, 2, 3]:
        headers = []
        for col in range(1, min(ws.max_column + 1, 50)):
            val = ws.cell(row=header_row, column=col).value
            if val:
                headers.append((col_letter(col), str(val).strip()[:35]))
        if headers:
            print(f"\n  Row {header_row} ({len(headers)} vals):")
            for letter, val in headers:
                print(f"    {letter:>4}: {val}")

    # Preenchimento
    print(f"\n  Preenchimento (primeiras 50 rows):")
    end_row = min(ws.max_row, 53)
    for col in range(1, min(ws.max_column + 1, 35)):
        filled = 0
        for row in range(3, end_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None and str(v).strip():
                filled += 1
        letter = col_letter(col)
        h = ws.cell(row=2, column=col).value or ws.cell(row=1, column=col).value or ""
        if filled > 0:
            print(f"    {letter:>4} | {str(h)[:25]:25} | {filled}/{end_row-2} preenchidas")


def main():
    print("=" * 100)
    print("ANALISE PROFUNDA V31 vs V14 — Ancora, Funil, Agenda")
    print("=" * 100)

    # Carregar V31 em modo formula
    print("\n[1/4] Carregando V31 (formulas)...")
    wb_v31 = openpyxl.load_workbook(V31_PATH, data_only=False)

    # Carregar V31 em modo dados
    print("[2/4] Carregando V31 (dados)...")
    wb_v31_data = openpyxl.load_workbook(V31_PATH, data_only=True)

    # Carregar V14 em modo formula
    print("[3/4] Carregando V14 (formulas)...")
    wb_v14 = openpyxl.load_workbook(V14_PATH, data_only=False)

    # Carregar V14 em modo dados
    print("[4/4] Carregando V14 (dados)...")
    wb_v14_data = openpyxl.load_workbook(V14_PATH, data_only=True)

    # CARTEIRA analise profunda
    analyze_carteira_deep(wb_v31["CARTEIRA"], "V31 (FORMULAS)")
    analyze_carteira_deep(wb_v31_data["CARTEIRA"], "V31 (DADOS)")
    analyze_carteira_deep(wb_v14["CARTEIRA"], "V14 (FORMULAS)")

    # AGENDA analise profunda
    analyze_agenda_deep(wb_v31, "V31")
    analyze_agenda_deep(wb_v14, "V14")

    # DRAFT 2 analise
    analyze_draft2_deep(wb_v31_data, "V31")
    analyze_draft2_deep(wb_v14_data, "V14")

    wb_v31.close()
    wb_v31_data.close()
    wb_v14.close()
    wb_v14_data.close()

    print(f"\n{'='*100}")
    print("ANALISE COMPLETA")
    print(f"{'='*100}")


if __name__ == "__main__":
    main()

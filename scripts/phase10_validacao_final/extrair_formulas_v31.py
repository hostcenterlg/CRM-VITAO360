#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrai formulas completas do V31 DRAFT 2, AGENDA, CARTEIRA.
Objetivo: capturar EXATAMENTE as formulas AUTO para replicar no V16.
"""

import openpyxl
import json

V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"


def col_letter(idx):
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def extract_formulas(ws, sheet_name, sample_rows=None):
    """Extrai formulas de um sheet"""
    formulas = {}

    if sample_rows is None:
        sample_rows = [3, 4, 5]  # primeiras rows de dados

    for row in sample_rows:
        for col in range(1, min(ws.max_column + 1, 80)):
            val = ws.cell(row=row, column=col).value
            letter = col_letter(col)
            key = f"{letter}{row}"

            if val is not None:
                val_str = str(val)
                if val_str.startswith("="):
                    formulas[key] = val_str
                elif "ArrayFormula" in val_str:
                    # ArrayFormula object - get the text
                    cell = ws.cell(row=row, column=col)
                    if hasattr(cell.value, 'text'):
                        formulas[key] = f"ARRAY:{cell.value.text}"
                    else:
                        formulas[key] = f"ARRAY:{val_str}"

    return formulas


def main():
    print("=" * 100)
    print("EXTRAÇÃO DE FÓRMULAS V31")
    print("=" * 100)

    wb = openpyxl.load_workbook(V31_PATH, data_only=False)

    all_formulas = {}

    # DRAFT 2
    print("\n[1] DRAFT 2 formulas...")
    ws = wb["DRAFT 2"]
    print(f"  {ws.max_row} rows x {ws.max_column} cols")

    # Extrair row 3 (primeira row de dados) para ter as formulas template
    for row in [3, 4]:
        print(f"\n  Row {row}:")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            letter = col_letter(col)
            header = ws.cell(row=2, column=col).value or ""
            row3_tag = ws.cell(row=3 if row == 3 else 2, column=col).value

            if val is not None:
                val_str = str(val)
                if val_str.startswith("="):
                    print(f"    {letter:>4} ({str(header)[:20]:20s}): {val_str}")
                    all_formulas[f"DRAFT2_{letter}_{row}"] = val_str
                elif hasattr(val, 'text'):
                    print(f"    {letter:>4} ({str(header)[:20]:20s}): ARRAY={val.text}")
                    all_formulas[f"DRAFT2_{letter}_{row}"] = f"ARRAY:{val.text}"

    # AGENDA
    print("\n\n[2] AGENDA formulas...")
    if "AGENDA" in wb.sheetnames:
        ws_a = wb["AGENDA"]
        print(f"  {ws_a.max_row} rows x {ws_a.max_column} cols")

        # Headers
        print(f"\n  Headers (row 1):")
        for col in range(1, min(ws_a.max_column + 1, 40)):
            h = ws_a.cell(row=1, column=col).value
            if h:
                letter = col_letter(col)
                print(f"    {letter:>4}: {str(h)[:50]}")

        # Formulas rows 2-3
        for row in [2, 3]:
            print(f"\n  Row {row} formulas:")
            for col in range(1, min(ws_a.max_column + 1, 40)):
                cell = ws_a.cell(row=row, column=col)
                val = cell.value
                letter = col_letter(col)

                if val is not None:
                    val_str = str(val)
                    if val_str.startswith("="):
                        print(f"    {letter:>4}: {val_str}")
                        all_formulas[f"AGENDA_{letter}_{row}"] = val_str
                    elif hasattr(val, 'text'):
                        print(f"    {letter:>4}: ARRAY={val.text}")
                        all_formulas[f"AGENDA_{letter}_{row}"] = f"ARRAY:{val.text}"
                    elif "ArrayFormula" in val_str:
                        print(f"    {letter:>4}: {val_str[:80]}")
                        all_formulas[f"AGENDA_{letter}_{row}"] = val_str

    # CARTEIRA - focar nas colunas FUNIL (AR-BJ)
    print("\n\n[3] CARTEIRA formulas (FUNIL AR-BJ)...")
    ws_c = wb["CARTEIRA"]
    print(f"  {ws_c.max_row} rows x {ws_c.max_column} cols")

    # Pegar headers row 2 e 3
    print(f"\n  Headers FUNIL (cols 44-62):")
    for col in range(44, min(ws_c.max_column + 1, 80)):
        h2 = ws_c.cell(row=2, column=col).value or ""
        h3 = ws_c.cell(row=3, column=col).value or ""
        letter = col_letter(col)
        print(f"    {letter:>4}: R2={str(h2)[:25]:25s} | R3={str(h3)[:30]}")

    # Formulas rows 4, 5
    for row in [4, 5]:
        print(f"\n  Row {row} formulas (FUNIL):")
        for col in range(44, min(ws_c.max_column + 1, 80)):
            cell = ws_c.cell(row=row, column=col)
            val = cell.value
            letter = col_letter(col)

            if val is not None:
                val_str = str(val)
                if val_str.startswith("="):
                    print(f"    {letter:>4}: {val_str[:120]}")
                    all_formulas[f"CART_{letter}_{row}"] = val_str

    wb.close()

    # Salvar
    out_path = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/v31_formulas.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(all_formulas, f, indent=2, ensure_ascii=False)
    print(f"\n  Formulas salvas: {out_path} ({len(all_formulas)} formulas)")

    print(f"\n[COMPLETO]")


if __name__ == "__main__":
    main()

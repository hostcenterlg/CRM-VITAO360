#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CRM VITAO360 — Engine de Simulação Mensal
==============================================================
Motor genérico usado por cada script mensal.
Recebe config do mês e gera a simulação.
"""

import openpyxl
from datetime import date, datetime, timedelta
from collections import defaultdict, Counter
import random
import os
import sys

# ============================================================
# REGRAS DE NEGÓCIO (da aba REGRAS)
# ============================================================
RESULTADO_REGRAS = {
    "VENDA / PEDIDO":          (4,  "VENDAS",        "🔥QUENTE",  "VENDA"),
    "EM ATENDIMENTO":          (2,  "FUNIL",         "🟡MORNO",   "PRÉ-VENDA"),
    "SUPORTE":                 (1,  "RELAC.",        "🟡MORNO",   "RESOLUÇÃO DE PROBLEMA"),
    "ORÇAMENTO":               (3,  "FUNIL",         "🔥QUENTE",  "PRÉ-VENDA"),
    "RELACIONAMENTO":          (7,  "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "CADASTRO":                (2,  "FUNIL",         "🟡MORNO",   "PRÉ-VENDA"),
    "NÃO ATENDE":              (1,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "NÃO RESPONDE":            (1,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "RECUSOU LIGAÇÃO":         (2,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "CS (SUCESSO DO CLIENTE)": (30, "RELAC.",        "🔥QUENTE",  "PÓS-VENDA"),
    "PÓS-VENDA":              (30, "RELAC.",        "🔥QUENTE",  "PÓS-VENDA"),
    "PERDA / FECHOU LOJA":    (0,  "NÃO VENDA",    "💀PERDIDO",  "PRÉ-VENDA"),
    "FOLLOW UP 7":             (7,  "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "FOLLOW UP 15":            (15, "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "NUTRIÇÃO":                (15, "NÃO VENDA",    "❄️FRIO",    "PROSPECÇÃO"),
}

ACAO_FUTURA_MAP = {
    "VENDA / PEDIDO":           "CONFIRMAR FATURAMENTO E EXPEDIÇÃO",
    "EM ATENDIMENTO":           "FECHAR NEGOCIAÇÃO EM ANDAMENTO",
    "SUPORTE":                  "RESOLVER PROBLEMA INTERNO E ENVIAR SOLUÇÃO",
    "ORÇAMENTO":                "CONFIRMAR ORÇAMENTO ENVIADO",
    "RELACIONAMENTO":           "RAPPORT COM CLIENTE APÓS A VENDA ATÉ RECOMPRAR",
    "CADASTRO":                 "CONFIRMAR CADASTRO NO SISTEMA",
    "NÃO ATENDE":               "2ª TENTATIVA VIA LIGAÇÃO",
    "NÃO RESPONDE":             "2ª TENTATIVA VIA WHATSAPP",
    "RECUSOU LIGAÇÃO":          "2ª TENTATIVA VIA WHATSAPP",
    "CS (SUCESSO DO CLIENTE)":  "VERIFICAR SELL OUT E CRIAR INTENÇÃO RECOMPRA",
    "PÓS-VENDA":               "FAZER CS (SUCESSO DO CLIENTE)",
    "PERDA / FECHOU LOJA":     "NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES",
    "FOLLOW UP 7":              "COBRAR RETORNO DO CLIENTE",
    "FOLLOW UP 15":             "COBRAR RETORNO DO CLIENTE",
    "NUTRIÇÃO":                 "NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES",
}

MOTIVOS = {
    "EM ATENDIMENTO": ["PRIMEIRO CONTATO / SEM RESPOSTA", "AINDA TEM ESTOQUE", "SÓ QUER COMPRAR GRANEL", "NÃO SE APLICA"],
    "ORÇAMENTO": ["AINDA TEM ESTOQUE", "SÓ QUER COMPRAR GRANEL", "PROBLEMA FINANCEIRO / CRÉDITO", "NÃO SE APLICA"],
    "CADASTRO": ["PRIMEIRO CONTATO / SEM RESPOSTA", "NÃO SE APLICA"],
    "VENDA / PEDIDO": ["NÃO SE APLICA"],
    "NÃO RESPONDE": ["PRIMEIRO CONTATO / SEM RESPOSTA", "PROPRIETARIO / INDISPONÍVEL"],
    "NÃO ATENDE": ["PROPRIETARIO / INDISPONÍVEL", "PRIMEIRO CONTATO / SEM RESPOSTA"],
    "RECUSOU LIGAÇÃO": ["PROPRIETARIO / INDISPONÍVEL"],
    "PERDA / FECHOU LOJA": ["SEM INTERESSE NO MOMENTO", "PRODUTO NÃO VENDEU / SEM GIRO", "FECHANDO / FECHOU LOJA", "LOJA ANEXO/PROXIMO - SM", "PROBLEMA FINANCEIRO / CRÉDITO"],
    "PÓS-VENDA": ["STATUS PEDIDO", "ATRASO ENTREGA", "NÃO SE APLICA"],
    "CS (SUCESSO DO CLIENTE)": ["NÃO SE APLICA"],
    "FOLLOW UP": ["AINDA TEM ESTOQUE", "NÃO SE APLICA", "PRIMEIRO CONTATO / SEM RESPOSTA"],
}

DEMANDAS = {
    "PROSPECÇÃO": ["PROSPECÇÃO - BASE GRANEL", "PROSPECÇÃO - PESQUISA GOOGLE", "CONTATO LEADS DO SITE", "RESPONDER LEADS COM DÚVIDAS", "LIGAÇÕES PROSPECT"],
    "ORCAMENTO": ["MONTAR ORÇAMENTO SUGESTÃO (CURVA ABC)", "ENVIO DE AMOSTRAS", "ENVIAR ANÁLISE DE CRÉDITO", "REGISTRAR ATENDIMENTO NO MERCOS"],
    "CADASTRO": ["ATUALIZAÇÃO CADASTRAL MERCOS", "ENVIAR ANÁLISE DE CRÉDITO", "REGISTRAR ATENDIMENTO NO MERCOS", "PREENCHER FOLLOW-UP"],
    "PEDIDO": ["DIGITAÇÃO DE PEDIDOS", "SOLICITAR LINK CARTÃO CRÉDITO", "SOLICITAR VALORES PIX", "CONFIRMAR FATURAMENTO E EXPEDIÇÃO", "REGISTRAR ATENDIMENTO NO MERCOS"],
    "POS_VENDA": ["FAZER RASTREIOS", "ACOMPANHAR DEVOLUÇÕES", "SUPORTE PCV - DÚVIDAS PEDIDOS", "SUPORTE PRODUTO (LAUDOS/TABELAS)", "PREENCHER FOLLOW-UP"],
    "FOLLOWUP": ["LIGAÇÕES DA BASE", "ATENDIMENTO ATIVOS/INATIVOS", "PREENCHER FOLLOW-UP", "REGISTRAR ATENDIMENTO NO MERCOS", "COBRANÇA DE TÍTULOS"],
}

NOTAS = {
    "PROSPECÇÃO": ["Primeiro contato via WhatsApp. Apresentei portfólio Vitão.",
        "Enviei catálogo digital. Prospect demonstrou interesse.",
        "Contato inicial. Cliente pediu mais informações sobre a linha.",
        "Prospecção via ligação. Consegui falar com responsável de compras."],
    "NEGOCIAÇÃO": ["Negociação avançando. Cliente pediu orçamento atualizado.",
        "Enviei proposta ajustada. Aguardando aprovação do financeiro.",
        "Cliente comparando com concorrente. Reforçei diferenciais Vitão.",
        "Negociação ativa. Cliente quer fechar até fim do mês."],
    "FOLLOWUP": ["Retornei contato conforme agendado. Cliente analisando proposta.",
        "Follow-up realizado. Cliente pediu mais prazo para decidir.",
        "Retomei negociação. Cliente demonstra interesse mas sem urgência.",
        "Follow-up via WhatsApp. Cliente respondeu que vai analisar essa semana."],
    "VENDA": ["Pedido fechado! Cliente satisfeito com condições.",
        "Venda concluída. Primeira compra do cliente.",
        "Pedido aprovado. Enviei confirmação por email."],
    "POS_VENDA": ["Pós-venda. Confirmei faturamento e envio da NF.",
        "Pós-venda. Cliente recebeu tudo certo, satisfeito.",
        "CS. Verificando sell out e oportunidade de recompra.",
        "Acompanhamento pós-venda. Cliente elogiou atendimento."],
    "NAO_RESPONDE": ["WhatsApp enviado, sem retorno.", "Mensagem visualizada mas sem resposta.", "Tentei contato, sem resposta."],
    "NAO_ATENDE": ["Ligação não atendida. Caixa postal.", "Chamou mas não atendeu.", "Tentei ligar, sem sucesso."],
    "PERDA": ["Cliente sem interesse no momento.", "Não conseguiu fechar. Produto não se encaixa.", "Cliente optou por outro fornecedor."],
    # --- Black Friday / Reclamações logística ---
    "RECLAMACAO_LOGISTICA": [
        "Cliente reclamou de atraso na entrega. Pedido deveria chegar há 5 dias.",
        "Produto chegou com avaria. Cliente pedindo reposição urgente.",
        "Transportadora extraviou volume. Abrindo chamado para rastreio.",
        "Frete cobrado maior que o informado. Cliente pedindo estorno da diferença.",
        "Pedido entregue incompleto, faltando 2 itens. Verificando com expedição.",
        "Cliente não recebeu NF. Reenviando por email.",
        "Entrega no endereço errado. Redirecionando com transportadora.",
        "Produto com validade curta. Cliente quer trocar lote.",
        "Pedido chegou depois do prazo combinado. Cliente insatisfeito.",
        "Devolução solicitada por defeito de embalagem no transporte.",
    ],
    # --- Retentativa clientes inativos (churned) ---
    "CHURNED_RETENTATIVA": [
        "Liguei para cliente inativo. Disse que produto Vitão não girou na loja dele.",
        "Retornei contato. Cliente informou que já tem fornecedor similar no mercado.",
        "Tentativa de reativação. Cliente diz que não tem espaço na gôndola.",
        "Cliente informou conflito de canal — supermercado ao lado já vende Vitão.",
        "Retentativa. Cliente disse que ainda tem estoque do último pedido.",
        "Contato com cliente perdido. Sem interesse no momento, foco em outras marcas.",
        "Liguei oferecendo novo mix. Cliente disse que a categoria não vendeu bem.",
        "Tentei reativar. Cliente mudou de segmento, não trabalha mais com naturais.",
        "Retentativa via WhatsApp. Cliente visualizou mas não respondeu.",
        "Ligação de reativação. Cliente pediu para retornar no próximo trimestre.",
        "Ofereci condição especial para voltar. Cliente vai pensar e retorna.",
        "Cliente informou que loja está em reforma, sem previsão de reabertura.",
    ],
}

# Motivos específicos de clientes churned (por que pararam de comprar)
MOTIVOS_CHURNED = [
    "PRODUTO NÃO VENDEU / SEM GIRO",
    "SEM INTERESSE NO MOMENTO",
    "AINDA TEM ESTOQUE",
    "PROBLEMA FINANCEIRO / CRÉDITO",
    "LOJA ANEXO/PROXIMO - SM",  # conflito de canal
    "FECHANDO / FECHOU LOJA",
    "SÓ QUER COMPRAR GRANEL",
]

# Demandas de reclamação/suporte logístico
DEMANDAS_RECLAMACAO = [
    "FAZER RASTREIOS",
    "ACOMPANHAR DEVOLUÇÕES",
    "SUPORTE PCV - DÚVIDAS PEDIDOS",
    "SUPORTE PRODUTO (LAUDOS/TABELAS)",
]

COLS = ['A','B','C','D','E','F','G','H','I','J','K','L','M',
        'N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
        'AA','AB','AC','AD','AE']

NOMES_DAIANE = {"DAIANE STAVICKI", "DAIANE STADLER"}

# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================
def dias_uteis_lista(start, end, feriados=None):
    feriados = feriados or set()
    result = []
    d = start
    while d <= end:
        if d.weekday() < 5 and d not in feriados:
            result.append(d)
        d += timedelta(days=1)
    return result

def proximo_dia_util(d, n=1, feriados=None):
    feriados = feriados or set()
    count = 0
    current = d
    while count < n:
        current += timedelta(days=1)
        if current.weekday() < 5 and current not in feriados:
            count += 1
    return current

def sinaleiro_calc(situacao, dias_compra, ciclo_medio):
    if not situacao:
        return "🟣"
    sit_upper = str(situacao).upper()
    if sit_upper in ("PROSPECT", "LEAD", "NOVO"):
        return "🟣"
    if dias_compra is None or dias_compra == "":
        return ""
    try:
        d = int(dias_compra)
        c = int(ciclo_medio) if ciclo_medio else 50
    except (ValueError, TypeError):
        return ""
    if d <= c:
        return "🟢"
    elif d <= c + 30:
        return "🟡"
    else:
        return "🔴"

def make_record(data, consultor, cnpj, nome, uf, rede, situacao, dias_compra,
                estagio_funil, tipo_cliente, fase, whatsapp, ligacao, lig_atendida,
                tipo_contato, resultado, motivo, followup_date, acao_futura,
                acao_detalhada, mercos, nota, temperatura, tipo_problema,
                demanda, tipo_atendimento, ciclo_medio=0):
    sinaleiro = sinaleiro_calc(situacao, dias_compra, ciclo_medio)
    grupo_dash = RESULTADO_REGRAS.get(resultado, (0, "", "", ""))[1]
    tipo_acao = RESULTADO_REGRAS.get(resultado, (0, "", "", ""))[3]
    return {
        'A': data, 'B': consultor, 'C': nome, 'D': cnpj,
        'E': uf, 'F': rede, 'G': situacao, 'H': dias_compra,
        'I': estagio_funil, 'J': tipo_cliente, 'K': fase,
        'L': sinaleiro, 'M': '',
        'N': whatsapp, 'O': ligacao, 'P': lig_atendida,
        'Q': tipo_contato, 'R': resultado, 'S': motivo,
        'T': followup_date, 'U': acao_futura, 'V': acao_detalhada,
        'W': mercos, 'X': nota, 'Y': temperatura,
        'Z': grupo_dash, 'AA': '', 'AB': tipo_acao,
        'AC': tipo_problema, 'AD': demanda, 'AE': tipo_atendimento,
    }

def gerar_ciclo_venda(cnpj_s, info, consultor, d_base, in_prev, prev_res, du, feriados, month_end):
    records = []
    gap = random.choice([1, 2, 2, 3])
    if in_prev and prev_res in ("ORÇAMENTO", "EM ATENDIMENTO"):
        steps = ["CADASTRO", "VENDA / PEDIDO"]
    elif in_prev and prev_res == "CADASTRO":
        steps = ["VENDA / PEDIDO"]
    elif in_prev:
        steps = ["EM ATENDIMENTO", "ORÇAMENTO", "CADASTRO", "VENDA / PEDIDO"]
    else:
        steps = ["EM ATENDIMENTO", "ORÇAMENTO", "CADASTRO", "VENDA / PEDIDO"]

    d_current = d_base
    for step_idx, step in enumerate(steps):
        if step_idx > 0:
            d_current = proximo_dia_util(d_current, gap, feriados)
        if d_current > month_end:
            d_current = month_end
        if d_current.weekday() >= 5:
            d_current = proximo_dia_util(d_current - timedelta(days=1), 1, feriados)

        step_map = {
            "EM ATENDIMENTO": ("EM ATENDIMENTO", "EM ATENDIMENTO", "ENVIAR ORÇAMENTO / CATÁLOGO",
                              MOTIVOS["EM ATENDIMENTO"], NOTAS["NEGOCIAÇÃO"],
                              DEMANDAS["PROSPECÇÃO"] if not in_prev else DEMANDAS["FOLLOWUP"],
                              "PROSPECÇÃO" if not in_prev else "NEGOCIAÇÃO"),
            "ORÇAMENTO": ("ORÇAMENTO", "ORÇAMENTO", "CONFIRMAR ORÇAMENTO ENVIADO",
                         MOTIVOS["ORÇAMENTO"], NOTAS["NEGOCIAÇÃO"], DEMANDAS["ORCAMENTO"], "NEGOCIAÇÃO"),
            "CADASTRO": ("CADASTRO", "EM ATENDIMENTO", "CONFIRMAR CADASTRO NO SISTEMA",
                        MOTIVOS["CADASTRO"], ["Cadastro enviado para análise."], DEMANDAS["CADASTRO"], "NEGOCIAÇÃO"),
            "VENDA / PEDIDO": ("VENDA / PEDIDO", "PÓS-VENDA", "CONFIRMAR FATURAMENTO E EXPEDIÇÃO",
                              MOTIVOS["VENDA / PEDIDO"], NOTAS["VENDA"], DEMANDAS["PEDIDO"], "NEGOCIAÇÃO"),
        }
        resultado, estagio, acao, motivos_list, notas_list, demandas_list, tc = step_map[step]
        records.append(make_record(
            data=d_current, consultor=consultor, cnpj=cnpj_s,
            nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
            situacao=info.get('situacao', 'PROSPECT'), dias_compra=info.get('dias_compra', ''),
            estagio_funil=estagio, tipo_cliente=info.get('tipo_cliente', 'NOVO'),
            fase="PROSPECÇÃO" if step == "EM ATENDIMENTO" and not in_prev else estagio,
            whatsapp=1 if step_idx == 0 else 0, ligacao=1 if step_idx > 0 else 0, lig_atendida=1,
            tipo_contato=tc, resultado=resultado, motivo=random.choice(motivos_list),
            followup_date=proximo_dia_util(d_current, gap, feriados),
            acao_futura=acao, acao_detalhada="",
            mercos="SIM" if step in ("ORÇAMENTO", "VENDA / PEDIDO") else "NÃO",
            nota=random.choice(notas_list),
            temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "🟡MORNO", ""))[2],
            tipo_problema="", demanda=random.choice(demandas_list),
            tipo_atendimento="PROSPECÇÃO" if not in_prev and step_idx == 0 else "FOLLOW UP",
            ciclo_medio=info.get('ciclo_medio', 0)
        ))

    d_pv = proximo_dia_util(d_current, 4, feriados)
    if d_pv <= month_end:
        records.append(make_record(
            data=d_pv, consultor=consultor, cnpj=cnpj_s,
            nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
            situacao="NOVO", dias_compra=4, estagio_funil="PÓS-VENDA", tipo_cliente="NOVO", fase="PÓS-VENDA",
            whatsapp=0, ligacao=1, lig_atendida=1,
            tipo_contato="PÓS-VENDA / RELACIONAMENTO", resultado="PÓS-VENDA", motivo="NÃO SE APLICA",
            followup_date=proximo_dia_util(d_pv, 11, feriados),
            acao_futura="FAZER CS (SUCESSO DO CLIENTE)", acao_detalhada="",
            mercos="SIM", nota=random.choice(NOTAS["POS_VENDA"]),
            temperatura="🔥QUENTE", tipo_problema="",
            demanda=random.choice(DEMANDAS["POS_VENDA"]), tipo_atendimento="ATEND. CLIENTES ATIVOS"
        ))
    return records

def gerar_followup(cnpj_s, info, consultor, d, prev_res, feriados):
    rand = random.random()
    if prev_res in ("EM ATENDIMENTO", "ORÇAMENTO"):
        if rand < 0.40:
            resultado = "ORÇAMENTO" if prev_res == "EM ATENDIMENTO" else "EM ATENDIMENTO"
        elif rand < 0.70:
            resultado = "EM ATENDIMENTO"
        elif rand < 0.90:
            resultado = random.choice(["NÃO RESPONDE", "NÃO ATENDE"])
        else:
            resultado = "PERDA / FECHOU LOJA"
    elif prev_res in ("PÓS-VENDA", "CS (SUCESSO DO CLIENTE)"):
        resultado = random.choice(["RELACIONAMENTO", "CS (SUCESSO DO CLIENTE)", "FOLLOW UP 7"])
    elif prev_res in ("NÃO RESPONDE", "NÃO ATENDE", "RECUSOU LIGAÇÃO"):
        if rand < 0.50:
            resultado = "EM ATENDIMENTO"
        elif rand < 0.80:
            resultado = random.choice(["NÃO RESPONDE", "NÃO ATENDE"])
        else:
            resultado = "PERDA / FECHOU LOJA"
    else:
        resultado = "EM ATENDIMENTO"

    motivos_key = resultado if resultado in MOTIVOS else "FOLLOW UP"
    canal = random.choice(['whatsapp', 'ligacao', 'ligacao'])
    return make_record(
        data=d, consultor=consultor, cnpj=cnpj_s,
        nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
        situacao=info.get('situacao', 'PROSPECT'), dias_compra=info.get('dias_compra', ''),
        estagio_funil="EM ATENDIMENTO" if resultado == "EM ATENDIMENTO" else (
            "ORÇAMENTO" if resultado == "ORÇAMENTO" else "PROSPECÇÃO"),
        tipo_cliente=info.get('tipo_cliente', 'NOVO'), fase="PROSPECÇÃO",
        whatsapp=1 if canal == 'whatsapp' else 0,
        ligacao=1 if canal == 'ligacao' else 0,
        lig_atendida=1 if resultado not in ("NÃO ATENDE", "NÃO RESPONDE", "RECUSOU LIGAÇÃO") else 0,
        tipo_contato="FOLLOW UP", resultado=resultado,
        motivo=random.choice(MOTIVOS.get(motivos_key, ["NÃO SE APLICA"])),
        followup_date=proximo_dia_util(d, RESULTADO_REGRAS.get(resultado, (2,))[0], feriados) if resultado != "PERDA / FECHOU LOJA" else "",
        acao_futura=ACAO_FUTURA_MAP.get(resultado, ""),
        acao_detalhada="", mercos="SIM" if resultado == "ORÇAMENTO" else "NÃO",
        nota=random.choice(NOTAS.get("FOLLOWUP", [""])) if resultado in ("EM ATENDIMENTO", "ORÇAMENTO") else (
            random.choice(NOTAS.get("NAO_RESPONDE", [""])) if "RESPONDE" in resultado else (
            random.choice(NOTAS.get("NAO_ATENDE", [""])) if "ATENDE" in resultado else (
            random.choice(NOTAS.get("PERDA", [""]))))),
        temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "🟡MORNO", ""))[2],
        tipo_problema="",
        demanda=random.choice(DEMANDAS.get("FOLLOWUP", [""])) if resultado in ("EM ATENDIMENTO", "ORÇAMENTO") else "",
        tipo_atendimento="FOLLOW UP", ciclo_medio=info.get('ciclo_medio', 0)
    )

def gerar_prospeccao(cnpj_s, info, consultor, d, feriados):
    rand = random.random()
    if rand < 0.60:
        resultado = "EM ATENDIMENTO"
    elif rand < 0.85:
        resultado = "NÃO RESPONDE"
    else:
        resultado = "NÃO ATENDE"
    return make_record(
        data=d, consultor=consultor, cnpj=cnpj_s,
        nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
        situacao="PROSPECT", dias_compra="",
        estagio_funil="PROSPECÇÃO", tipo_cliente="PROSPECT", fase="PROSPECÇÃO",
        whatsapp=1, ligacao=0, lig_atendida=0,
        tipo_contato="PROSPECÇÃO", resultado=resultado,
        motivo=random.choice(MOTIVOS.get(resultado, ["NÃO SE APLICA"])),
        followup_date=proximo_dia_util(d, RESULTADO_REGRAS.get(resultado, (2,))[0], feriados),
        acao_futura=ACAO_FUTURA_MAP.get(resultado, "ENVIAR ORÇAMENTO / CATÁLOGO"),
        acao_detalhada="", mercos="NÃO",
        nota=random.choice(NOTAS.get("PROSPECÇÃO", [""])) if resultado == "EM ATENDIMENTO" else (
            random.choice(NOTAS.get("NAO_RESPONDE", [""])) if "RESPONDE" in resultado else
            random.choice(NOTAS.get("NAO_ATENDE", [""]))),
        temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "❄️FRIO", ""))[2],
        tipo_problema="", demanda=random.choice(DEMANDAS["PROSPECÇÃO"]),
        tipo_atendimento="PROSPECÇÃO"
    )


def gerar_reclamacao_logistica(cnpj_s, info, consultor, d, feriados):
    """Gera registro de RECLAMAÇÃO/SUPORTE por problema logístico (Black Friday)"""
    resultado = random.choice(["SUPORTE", "SUPORTE", "SUPORTE", "PÓS-VENDA"])
    motivo = random.choice(["ATRASO ENTREGA", "STATUS PEDIDO", "ATRASO ENTREGA", "ATRASO ENTREGA"])
    return make_record(
        data=d, consultor=consultor, cnpj=cnpj_s,
        nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
        situacao=info.get('situacao', 'ATIVO'), dias_compra=info.get('dias_compra', ''),
        estagio_funil="SUPORTE", tipo_cliente=info.get('tipo_cliente', 'ATIVO'),
        fase="PÓS-VENDA",
        whatsapp=random.choice([0, 1]), ligacao=1, lig_atendida=1,
        tipo_contato="PÓS-VENDA / RELACIONAMENTO", resultado=resultado,
        motivo=motivo,
        followup_date=proximo_dia_util(d, 1, feriados),
        acao_futura="RESOLVER PROBLEMA INTERNO E ENVIAR SOLUÇÃO",
        acao_detalhada="", mercos="SIM",
        nota=random.choice(NOTAS["RECLAMACAO_LOGISTICA"]),
        temperatura="🔥QUENTE",
        tipo_problema="LOGÍSTICA",
        demanda=random.choice(DEMANDAS_RECLAMACAO),
        tipo_atendimento="ATEND. CLIENTES ATIVOS",
        ciclo_medio=info.get('ciclo_medio', 0)
    )


def gerar_retentativa_churned(cnpj_s, info, consultor, d, feriados):
    """Gera registro de retentativa para cliente inativo/churned"""
    rand = random.random()
    if rand < 0.55:
        resultado = "EM ATENDIMENTO"  # conseguiu falar mas sem avanço
    elif rand < 0.80:
        resultado = random.choice(["NÃO RESPONDE", "NÃO ATENDE"])
    elif rand < 0.95:
        resultado = "RECUSOU LIGAÇÃO"
    else:
        resultado = "PERDA / FECHOU LOJA"

    motivo = random.choice(MOTIVOS_CHURNED)
    canal = random.choice(['whatsapp', 'ligacao', 'ligacao'])

    return make_record(
        data=d, consultor=consultor, cnpj=cnpj_s,
        nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
        situacao=info.get('situacao', 'INATIVO'), dias_compra=info.get('dias_compra', ''),
        estagio_funil="PERDA / NUTRIÇÃO" if resultado in ("NÃO RESPONDE", "NÃO ATENDE", "RECUSOU LIGAÇÃO", "PERDA / FECHOU LOJA") else "QUALIFICAÇÃO",
        tipo_cliente=info.get('tipo_cliente', 'ATIVO'), fase="REATIVAÇÃO",
        whatsapp=1 if canal == 'whatsapp' else 0,
        ligacao=1 if canal == 'ligacao' else 0,
        lig_atendida=1 if resultado not in ("NÃO ATENDE", "NÃO RESPONDE", "RECUSOU LIGAÇÃO") else 0,
        tipo_contato="FOLLOW UP", resultado=resultado,
        motivo=motivo,
        followup_date=proximo_dia_util(d, 15, feriados) if resultado not in ("PERDA / FECHOU LOJA",) else "",
        acao_futura="NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES" if resultado in ("NÃO RESPONDE", "NÃO ATENDE", "RECUSOU LIGAÇÃO") else (
            "FECHAR NEGOCIAÇÃO EM ANDAMENTO" if resultado == "EM ATENDIMENTO" else ""),
        acao_detalhada="", mercos="NÃO",
        nota=random.choice(NOTAS["CHURNED_RETENTATIVA"]),
        temperatura="❄️FRIO",
        tipo_problema="",
        demanda=random.choice(DEMANDAS["FOLLOWUP"]),
        tipo_atendimento="FOLLOW UP",
        ciclo_medio=info.get('ciclo_medio', 0)
    )


def run_simulation(config):
    """
    config = {
        'version': 'V32',
        'month_name': 'JUN',
        'month_num': 6,
        'year': 2025,
        'month_start': date(2025, 6, 1),
        'month_end': date(2025, 6, 30),
        'feriados': set(),
        'col_fat': 31,
        'prev_months_max': 5,
        'seed': 47,
        'input_file': '...',
        'output_file': '...',
        'consultores': {
            "DAIANE STADLER": {"max_fu": 90, "max_retry": 30, "venda_pct": 0.38, "prospects_max": 80},
            "HELDER BRUNKOW": {"max_fu": 60, "max_retry": 20, "venda_pct": 0.31, "prospects_max": 60},
            "MANU DITZEL": {"max_fu": 60, "max_retry": 20, "venda_pct": 0.31, "prospects_max": 60},
        },
    }
    """
    version = config['version']
    month_name = config['month_name']
    month_num = config['month_num']
    month_start = config['month_start']
    month_end = config['month_end']
    feriados = config.get('feriados', set())
    col_fat = config['col_fat']
    prev_months_max = config['prev_months_max']
    input_file = config['input_file']
    output_file = config['output_file']
    consultores = config['consultores']

    random.seed(config.get('seed', 42))

    year = config.get('year', 2025)
    year_short = str(year)[-2:]

    print("=" * 80)
    print(f"CRM VITAO360 — {version}: Simulação Realista {month_name}/{year_short}")
    print("=" * 80)

    if not os.path.exists(input_file):
        print(f"ERRO: {input_file} não encontrado")
        sys.exit(1)

    print(f"\n[1/7] Carregando arquivo anterior...")
    wb = openpyxl.load_workbook(input_file)
    ws_cart = wb['CARTEIRA']
    ws_d2 = wb['DRAFT 2']

    # CARTEIRA
    print(f"\n[2/7] Lendo CARTEIRA...")
    cart_data = {}
    cnpj_num_to_key = {}  # CNPJ numérico (col 2) → chave cart_data (col 3 razão social)
    for r in range(4, ws_cart.max_row + 1):
        razao = ws_cart.cell(row=r, column=3).value  # col 3 = RAZÃO SOCIAL (usado como key)
        cnpj_num = ws_cart.cell(row=r, column=2).value  # col 2 = CNPJ numérico
        if not razao:
            continue
        key = str(razao).strip()
        cart_data[key] = {
            'nome': str(cnpj_num).strip() if cnpj_num else "",
            'uf': ws_cart.cell(row=r, column=4).value or "",
            'rede': ws_cart.cell(row=r, column=5).value or "",
            'situacao': ws_cart.cell(row=r, column=6).value or "",
            'tipo_cliente': ws_cart.cell(row=r, column=7).value or "",
            'ciclo_medio': ws_cart.cell(row=r, column=8).value or 0,
            'dias_compra': ws_cart.cell(row=r, column=9).value or "",
            'fat_mes': ws_cart.cell(row=r, column=col_fat).value or 0,
        }
        if cnpj_num:
            cnpj_num_to_key[str(cnpj_num).strip()] = key

    # buyers_override: lista de CNPJs numéricos passada diretamente
    buyers_override = config.get('buyers_override')
    if buyers_override:
        # Converter CNPJ numéricos para chaves cart_data (razão social)
        buyers = []
        for cnpj in buyers_override:
            cnpj_s = str(cnpj).strip()
            if cnpj_s in cnpj_num_to_key:
                buyers.append(cnpj_num_to_key[cnpj_s])
            elif cnpj_s in cart_data:
                buyers.append(cnpj_s)
        print(f"  Compradores {month_name}/{year_short}: {len(buyers)} (override: {len(buyers_override)} passados, {len(buyers)} na CARTEIRA)")
    else:
        buyers = [c for c, info in cart_data.items()
                  if info['fat_mes'] and isinstance(info['fat_mes'], (int, float)) and info['fat_mes'] > 0]
        print(f"  Compradores {month_name}/{year_short}: {len(buyers)}")

    # FOLLOW-UPS
    print(f"\n[3/7] Analisando follow-ups meses anteriores...")
    all_cons_names = set(consultores.keys())
    nomes_validos = NOMES_DAIANE | all_cons_names

    prev_last = {}
    for r in range(3, ws_d2.max_row + 1):
        data_cell = ws_d2.cell(row=r, column=1).value
        if data_cell is None:
            continue
        if isinstance(data_cell, (int, float)):
            data_cell = datetime(1900, 1, 1) + timedelta(days=int(data_cell) - 2)
        if not isinstance(data_cell, datetime):
            continue
        # Scan all records BEFORE current month (supports any year)
        month_start_dt = datetime(month_start.year, month_start.month, 1)
        if data_cell >= month_start_dt:
            continue
        cons = ws_d2.cell(row=r, column=2).value
        if cons is None:
            continue
        cons_str = str(cons).strip()
        if cons_str in NOMES_DAIANE:
            cons_norm = "DAIANE STADLER"
        elif cons_str in nomes_validos:
            cons_norm = cons_str
        else:
            continue
        # Only track consultors that exist in current config
        if cons_norm not in consultores:
            continue
        cnpj = ws_d2.cell(row=r, column=4).value
        if not cnpj:
            continue
        cnpj_s = str(cnpj).strip()
        res = ws_d2.cell(row=r, column=18).value or ""
        key = (cons_norm, cnpj_s)
        if key not in prev_last or data_cell >= prev_last[key]['data']:
            prev_last[key] = {'data': data_cell, 'resultado': str(res)}

    prev_by_cons = defaultdict(lambda: defaultdict(list))
    for (cons, cnpj), info in prev_last.items():
        prev_by_cons[cons][info['resultado']].append(cnpj)

    for cons in sorted(prev_by_cons.keys()):
        total = sum(len(v) for v in prev_by_cons[cons].values())
        print(f"  {cons}: {total} CNPJs")

    # POOLS
    print(f"\n[4/7] Montando pools...")
    du = dias_uteis_lista(month_start, month_end, feriados)
    print(f"  Dias úteis: {len(du)}")

    def build_followup_pool(cons_name, max_fu, max_retry):
        by_res = prev_by_cons.get(cons_name, {})
        pool_fu = []
        for res in ["EM ATENDIMENTO", "ORÇAMENTO", "PÓS-VENDA", "CS (SUCESSO DO CLIENTE)", "VENDA / PEDIDO", "RELACIONAMENTO", "FOLLOW UP 7"]:
            cnpjs = by_res.get(res, [])
            random.shuffle(cnpjs)
            pool_fu.extend(cnpjs[:25])
        pool_fu = list(set(pool_fu))[:max_fu]
        pool_retry = []
        for res in ["NÃO RESPONDE", "NÃO ATENDE", "RECUSOU LIGAÇÃO"]:
            cnpjs = by_res.get(res, [])
            random.shuffle(cnpjs)
            pool_retry.extend(cnpjs[:12])
        pool_retry = list(set(pool_retry))[:max_retry]
        return pool_fu, pool_retry

    # Distribute sales
    random.shuffle(buyers)
    cons_list = list(consultores.keys())
    vendas_by_cons = {}
    start_idx = 0
    for i, cons_name in enumerate(cons_list):
        pct = consultores[cons_name]['venda_pct']
        if i == len(cons_list) - 1:
            vendas_by_cons[cons_name] = buyers[start_idx:]
        else:
            n = int(len(buyers) * pct)
            vendas_by_cons[cons_name] = buyers[start_idx:start_idx + n]
            start_idx += n

    # Prospects
    all_used = set(cnpj for (_, cnpj) in prev_last.keys()) | set(buyers)
    prospects_pool = [c for c in cart_data if c not in all_used]
    random.shuffle(prospects_pool)

    prosp_by_cons = {}
    prosp_idx = 0
    for cons_name in cons_list:
        max_p = consultores[cons_name].get('prospects_max', 60)
        prosp_by_cons[cons_name] = prospects_pool[prosp_idx:prosp_idx + max_p]
        prosp_idx += max_p

    for cons_name in cons_list:
        c = consultores[cons_name]
        print(f"  {cons_name}: {len(vendas_by_cons[cons_name])} vendas, prosp={len(prosp_by_cons[cons_name])}")

    # SIMULATE
    print(f"\n[5/7] Simulando atendimentos...")
    registros_novos = []

    def get_info(cnpj_s):
        return cart_data.get(cnpj_s, {'nome': '', 'uf': '', 'rede': '', 'situacao': 'PROSPECT',
                                       'tipo_cliente': 'NOVO', 'ciclo_medio': 0, 'dias_compra': ''})

    def simular_consultor(nome, pool_fu, pool_retry, vendas, prospects):
        recs = []
        if pool_fu:
            fu_per_day = max(1, len(pool_fu) // len(du))
            fu_idx = 0
            for dia_idx, d in enumerate(du):
                extra = 1 if dia_idx < len(pool_fu) % len(du) else 0
                batch = pool_fu[fu_idx:fu_idx + fu_per_day + extra]
                fu_idx += len(batch)
                for cnpj_s in batch:
                    info = get_info(cnpj_s)
                    prev_res = prev_last.get((nome, cnpj_s), {}).get('resultado', 'EM ATENDIMENTO')
                    recs.append(gerar_followup(cnpj_s, info, nome, d, prev_res, feriados))
        if pool_retry:
            retry_per_day = max(1, len(pool_retry) // min(10, len(du)))
            retry_idx = 0
            for d_r in du[:10]:
                batch = pool_retry[retry_idx:retry_idx + retry_per_day + 1]
                retry_idx += len(batch)
                if retry_idx >= len(pool_retry):
                    retry_idx = len(pool_retry)
                for cnpj_s in batch:
                    info = get_info(cnpj_s)
                    prev_res = prev_last.get((nome, cnpj_s), {}).get('resultado', 'NÃO RESPONDE')
                    recs.append(gerar_followup(cnpj_s, info, nome, d_r, prev_res, feriados))
        du_vendas = du[:15]
        for i, cnpj_s in enumerate(vendas):
            info = get_info(cnpj_s)
            d_base = du_vendas[i % len(du_vendas)]
            key = (nome, cnpj_s)
            in_prev = key in prev_last
            prev_res_v = prev_last.get(key, {}).get('resultado', '')
            recs.extend(gerar_ciclo_venda(cnpj_s, info, nome, d_base, in_prev, prev_res_v, du, feriados, month_end))
        if prospects:
            prosp_per_day = max(1, len(prospects) // len(du))
            prosp_idx = 0
            for dia_idx, d in enumerate(du):
                extra = 1 if dia_idx < len(prospects) % len(du) else 0
                batch = prospects[prosp_idx:prosp_idx + prosp_per_day + extra]
                prosp_idx += len(batch)
                for cnpj_s in batch:
                    info = get_info(cnpj_s)
                    recs.append(gerar_prospeccao(cnpj_s, info, nome, d, feriados))
        return recs

    for cons_name in cons_list:
        c = consultores[cons_name]
        fu, retry = build_followup_pool(cons_name, c['max_fu'], c['max_retry'])
        recs = simular_consultor(cons_name, fu, retry, vendas_by_cons[cons_name], prosp_by_cons[cons_name])
        print(f"  {cons_name}: {len(recs)} registros ({len(recs)/len(du):.1f}/dia)")
        registros_novos.extend(recs)

    # --- BLACK FRIDAY: reclamações/suporte logística (NOV/DEZ) ---
    bf_count = config.get('black_friday_reclamacoes', 0)
    if bf_count > 0:
        print(f"\n  [BF] Gerando {bf_count} reclamações logísticas Black Friday...")
        # Pegar clientes que compraram (buyers) para gerar reclamações
        bf_pool = list(buyers)
        random.shuffle(bf_pool)
        bf_pool = bf_pool[:bf_count]
        bf_per_cons = max(1, len(bf_pool) // len(cons_list))
        bf_idx = 0
        bf_total = 0
        for cons_name in cons_list:
            batch = bf_pool[bf_idx:bf_idx + bf_per_cons]
            bf_idx += len(batch)
            for cnpj_s in batch:
                info = get_info(cnpj_s)
                d = random.choice(du[len(du)//2:])  # segunda metade do mês
                registros_novos.append(gerar_reclamacao_logistica(cnpj_s, info, cons_name, d, feriados))
                bf_total += 1
        # Distribuir os restantes
        while bf_idx < len(bf_pool):
            cnpj_s = bf_pool[bf_idx]
            info = get_info(cnpj_s)
            cons = random.choice(cons_list)
            d = random.choice(du[len(du)//2:])
            registros_novos.append(gerar_reclamacao_logistica(cnpj_s, info, cons, d, feriados))
            bf_idx += 1
            bf_total += 1
        print(f"  [BF] {bf_total} reclamações geradas")

    # --- CHURNED CLIENTS: retentativas com motivos reais ---
    churned_max = config.get('churned_retries_max', 0)
    if churned_max > 0:
        print(f"\n  [CHURNED] Gerando até {churned_max} retentativas de clientes inativos...")
        # Encontrar clientes inativos: último resultado NEGATIVO + não compraram este mês
        NEGATIVE_RESULTS = {"NÃO RESPONDE", "NÃO ATENDE", "RECUSOU LIGAÇÃO", "PERDA / FECHOU LOJA", "NUTRIÇÃO"}
        buyers_set = set(buyers)
        churned_pool = []
        for (cons, cnpj), info_prev in prev_last.items():
            if cnpj in buyers_set:
                continue  # comprou este mês, não é churned
            if info_prev.get('resultado') in NEGATIVE_RESULTS:
                churned_pool.append(cnpj)
        # Remover duplicatas (mesmo CNPJ pode ter resultado negativo com vários consultores)
        churned_pool = list(set(churned_pool))
        random.shuffle(churned_pool)
        churned_pool = churned_pool[:churned_max]
        churned_per_cons = max(1, len(churned_pool) // len(cons_list))
        ch_idx = 0
        ch_total = 0
        for cons_name in cons_list:
            batch = churned_pool[ch_idx:ch_idx + churned_per_cons]
            ch_idx += len(batch)
            for cnpj_s in batch:
                info = get_info(cnpj_s)
                d = random.choice(du)  # distribuir ao longo do mês
                registros_novos.append(gerar_retentativa_churned(cnpj_s, info, cons_name, d, feriados))
                ch_total += 1
        while ch_idx < len(churned_pool):
            cnpj_s = churned_pool[ch_idx]
            info = get_info(cnpj_s)
            cons = random.choice(cons_list)
            d = random.choice(du)
            registros_novos.append(gerar_retentativa_churned(cnpj_s, info, cons, d, feriados))
            ch_idx += 1
            ch_total += 1
        print(f"  [CHURNED] {ch_total} retentativas geradas")

    print(f"\n  TOTAL {month_name}/{year_short}: {len(registros_novos)} registros")

    # VERIFICATION
    print(f"\n[6/7] Verificação lógica...")
    res_count = Counter()
    cons_count = Counter()
    for rec in registros_novos:
        res_count[rec['R']] += 1
        cons_count[rec['B']] += 1

    for res, n in res_count.most_common():
        print(f"    {res:<35} {n:>4}")

    n_orc = res_count.get("ORÇAMENTO", 0)
    n_cad = res_count.get("CADASTRO", 0)
    n_ven = res_count.get("VENDA / PEDIDO", 0)
    print(f"\n  Funil: ORC ({n_orc}) >= CAD ({n_cad}) >= VENDA ({n_ven})")
    print(f"  {'✅ CONSISTENTE' if n_orc >= n_cad >= n_ven else '⚠️ VERIFICAR'}")

    n_total = len(registros_novos)
    n_motivo = sum(1 for r in registros_novos if r['S'])
    n_demanda = sum(1 for r in registros_novos if r['AD'])
    n_acao = sum(1 for r in registros_novos if r['U'])
    n_nota = sum(1 for r in registros_novos if r['X'])
    print(f"  Preenchimento: MOT {n_motivo/n_total*100:.0f}% | DEM {n_demanda/n_total*100:.0f}% | AÇÃO {n_acao/n_total*100:.0f}% | NOTA {n_nota/n_total*100:.0f}%")

    # SAVE
    print(f"\n[7/7] Salvando {version}...")
    existing_records = []
    removed = 0
    for r in range(3, ws_d2.max_row + 1):
        data_cell = ws_d2.cell(row=r, column=1).value
        if data_cell is None:
            continue
        is_month = False
        if isinstance(data_cell, (int, float)):
            dt = datetime(1900, 1, 1) + timedelta(days=int(data_cell) - 2)
            if dt.year == year and dt.month == month_num:
                is_month = True
        elif isinstance(data_cell, datetime):
            if data_cell.year == year and data_cell.month == month_num:
                is_month = True
        if is_month:
            removed += 1
            continue
        row_data = {}
        for c_idx, col_letter in enumerate(COLS):
            row_data[col_letter] = ws_d2.cell(row=r, column=c_idx + 1).value
        existing_records.append(row_data)

    print(f"  Mantidos: {len(existing_records)} | Removidos: {removed} | Novos: {len(registros_novos)}")

    all_records = existing_records + registros_novos
    def sort_key(rec):
        d = rec.get('A')
        if d is None:
            return datetime(1900, 1, 1)
        if isinstance(d, (int, float)):
            return datetime(1900, 1, 1) + timedelta(days=int(d) - 2)
        if isinstance(d, date) and not isinstance(d, datetime):
            return datetime(d.year, d.month, d.day)
        return d if isinstance(d, datetime) else datetime(1900, 1, 1)

    all_records.sort(key=sort_key)

    cnpj_counter = Counter()
    for rec in all_records:
        cnpj = rec.get('D')
        if cnpj:
            cnpj_counter[str(cnpj).strip()] += 1
            rec['M'] = f"T{cnpj_counter[str(cnpj).strip()]}"

    for r in range(3, ws_d2.max_row + 1):
        for c in range(1, len(COLS) + 1):
            ws_d2.cell(row=r, column=c).value = None

    for i, rec in enumerate(all_records):
        row = i + 3
        for c_idx, col_letter in enumerate(COLS):
            ws_d2.cell(row=row, column=c_idx + 1).value = rec.get(col_letter)

    wb.save(output_file)
    file_size = os.path.getsize(output_file) / (1024 * 1024)

    print(f"\n{'='*80}")
    print(f"{version} CONCLUÍDO! — {month_name}/{year_short}")
    print(f"{'='*80}")
    print(f"  Arquivo: {output_file} ({file_size:.2f} MB)")
    print(f"  DRAFT 2: {len(all_records)} registros")
    print(f"  {month_name}/{year_short}: {len(registros_novos)} novos (removidos {removed})")
    print(f"  Vendas: {n_ven}/{len(buyers)} | Média: {len(registros_novos)/len(du):.1f}/dia")
    for cons, n in cons_count.most_common():
        print(f"    {cons:<25} {n:>4} ({n/len(du):.1f}/dia)")
    print(f"{'='*80}")

    return {
        'total_records': len(all_records),
        'month_records': len(registros_novos),
        'vendas': n_ven,
        'buyers': len(buyers),
    }

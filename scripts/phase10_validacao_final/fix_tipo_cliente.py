#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FIX TIPO CLIENTE — O patch anterior colocou nomes de consultores em AX (TIPO CLIENTE).
Corrigir: usar DRAFT 2 col J (TIPO CLIENTE) ou V31, nao DRAFT 1 col J (CONSULTOR).
"""

import openpyxl
import re
from pathlib import Path
from datetime import datetime

V17_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V17_FINAL.xlsx"
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"


def normalizar_cnpj(valor):
    if not valor:
        return None
    cnpj = re.sub(r'[^\d]', '', str(valor).strip())
    return cnpj if len(cnpj) >= 11 else None


def get_real_value(cell):
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str) and val.startswith("="):
        return None
    if isinstance(val, str) and val.strip() == "":
        return None
    return val


def main():
    start = datetime.now()
    print("=" * 80)
    print("FIX TIPO CLIENTE — Corrigir valores errados em AX (col 50)")
    print("=" * 80)

    wb = openpyxl.load_workbook(V17_PATH)
    wb_v31 = openpyxl.load_workbook(V31_PATH, data_only=True)

    ws_cart = wb["CARTEIRA"]
    ws_d2 = wb["DRAFT 2"]

    # Valores validos de TIPO CLIENTE
    VALID_TIPOS = {
        "ATIVO", "INATIVO", "INAT.REC", "INAT.ANT", "NOVO", "PROSPECT", "LEAD",
        "EM RISCO", "INATIVO RECENTE", "INATIVO ANTIGO", "RECUPERADO",
        "ESPECIAL", "VIP", "PREMIUM", "BRONZE", "PRATA", "OURO",
        "A", "B", "C", "D",
    }

    # Nomes de consultores (para detectar valores errados)
    CONSULTORES = {
        "LARISSA PADILHA", "DAIANE STAVICKI", "MANU DITZEL", "HEMANUELE DITZEL",
        "HEMANUELE DITZEL (MANU)", "JULIO GADRET", "LEANDRO", "LORRANY",
        "Helder Brunkow", "HELDER BRUNKOW", "RODRIGO",
    }
    CONSULTORES_UPPER = {c.upper() for c in CONSULTORES}

    # Indexar DRAFT 2 col J (10) = TIPO CLIENTE (ultimo por CNPJ)
    print("\n  Indexando DRAFT 2 col J (TIPO CLIENTE)...")
    d2_tipo = {}
    for row in range(3, ws_d2.max_row + 1):
        cnpj = normalizar_cnpj(ws_d2.cell(row=row, column=4).value)
        if not cnpj:
            continue
        val = get_real_value(ws_d2.cell(row=row, column=10))
        if val:
            val_str = str(val).upper().strip()
            if val_str not in CONSULTORES_UPPER and len(val_str) < 30:
                d2_tipo[cnpj] = val

    print(f"    {len(d2_tipo)} CNPJs com TIPO CLIENTE no DRAFT 2")

    # Amostra
    sample = list(d2_tipo.items())[:10]
    print(f"    Amostra: {sample}")

    # Indexar V31 CARTEIRA col AX (50) = TIPO CLIENTE
    print("\n  Indexando V31 CARTEIRA col AX (TIPO CLIENTE)...")
    ws_v31_cart = wb_v31["CARTEIRA"]
    v31_tipo = {}
    for row in range(4, ws_v31_cart.max_row + 1):
        cnpj = normalizar_cnpj(ws_v31_cart.cell(row=row, column=2).value)
        if not cnpj:
            continue
        val = ws_v31_cart.cell(row=row, column=50).value
        if val and str(val).strip():
            v31_tipo[cnpj] = val

    print(f"    {len(v31_tipo)} CNPJs com TIPO CLIENTE no V31")

    # Tambem pegar do V31 CARTEIRA col J (10)
    v31_tipo_j = {}
    for row in range(4, ws_v31_cart.max_row + 1):
        cnpj = normalizar_cnpj(ws_v31_cart.cell(row=row, column=2).value)
        if not cnpj:
            continue
        val = ws_v31_cart.cell(row=row, column=10).value
        if val and str(val).strip():
            v31_tipo_j[cnpj] = val

    print(f"    {len(v31_tipo_j)} CNPJs com col J no V31 CARTEIRA")

    # Tambem indexar SITUACAO do DRAFT 2 (col G=7) para inferir TIPO
    d2_situacao = {}
    for row in range(3, ws_d2.max_row + 1):
        cnpj = normalizar_cnpj(ws_d2.cell(row=row, column=4).value)
        if not cnpj:
            continue
        date_val = ws_d2.cell(row=row, column=1).value
        date_str = str(date_val) if date_val else ""
        if cnpj in d2_situacao:
            if date_str <= d2_situacao[cnpj]['_date']:
                continue
        val = get_real_value(ws_d2.cell(row=row, column=7))
        if val:
            d2_situacao[cnpj] = {'val': val, '_date': date_str}

    # PATCH CARTEIRA
    print(f"\n  Corrigindo TIPO CLIENTE na CARTEIRA...")
    fixed = 0
    cleared = 0
    kept = 0

    for row in range(4, ws_cart.max_row + 1):
        cnpj = normalizar_cnpj(ws_cart.cell(row=row, column=2).value)
        if not cnpj:
            continue

        current = get_real_value(ws_cart.cell(row=row, column=50))
        current_str = str(current).upper().strip() if current else ""

        # Detectar se valor atual eh um nome de consultor (errado)
        is_consultor_name = current_str in CONSULTORES_UPPER

        if is_consultor_name or not current:
            # Buscar valor correto
            val = None

            # 1. DRAFT 2 col J (10)
            if cnpj in d2_tipo:
                val = d2_tipo[cnpj]

            # 2. V31 CARTEIRA col AX
            if not val and cnpj in v31_tipo:
                val = v31_tipo[cnpj]

            # 3. V31 CARTEIRA col J
            if not val and cnpj in v31_tipo_j:
                v = str(v31_tipo_j[cnpj]).upper().strip()
                if v not in CONSULTORES_UPPER:
                    val = v31_tipo_j[cnpj]

            # 4. Inferir da SITUAÇÃO
            if not val and cnpj in d2_situacao:
                sit = str(d2_situacao[cnpj]['val']).upper().strip()
                if sit in VALID_TIPOS:
                    val = d2_situacao[cnpj]['val']

            # 5. Usar SITUAÇÃO da CARTEIRA col N
            if not val:
                sit = get_real_value(ws_cart.cell(row=row, column=14))
                if sit and str(sit).upper().strip() in VALID_TIPOS:
                    val = sit

            if val:
                ws_cart.cell(row=row, column=50, value=val)
                fixed += 1
            elif is_consultor_name:
                # Limpar valor errado
                ws_cart.cell(row=row, column=50, value="")
                cleared += 1
        else:
            kept += 1

    print(f"\n  RESULTADO:")
    print(f"    Corrigidos: {fixed}")
    print(f"    Limpos (era consultor, sem substituto): {cleared}")
    print(f"    Mantidos (ja corretos): {kept}")

    # Verificar
    print(f"\n  Verificacao TIPO CLIENTE (amostra rows 4-13):")
    for row in range(4, min(14, ws_cart.max_row + 1)):
        cnpj = str(ws_cart.cell(row=row, column=2).value or "")[:14]
        consultor = str(ws_cart.cell(row=row, column=12).value or "")[:15]
        tipo = str(ws_cart.cell(row=row, column=50).value or "")[:20]
        situacao = str(get_real_value(ws_cart.cell(row=row, column=14)) or "")[:12]
        print(f"    R{row:>3}: {cnpj:14s} | CONSULT={consultor:15s} | TIPO={tipo:20s} | SIT={situacao}")

    # Contar preenchimento final
    filled = 0
    total = 0
    for row in range(4, ws_cart.max_row + 1):
        val = get_real_value(ws_cart.cell(row=row, column=50))
        total += 1
        if val:
            filled += 1
    pct = round(100 * filled / max(total, 1), 1)
    print(f"\n  TIPO CLIENTE final: {filled}/{total} ({pct}%)")

    # Salvar
    print(f"\n  Salvando...")
    wb.save(V17_PATH)
    wb.close()
    wb_v31.close()

    size = Path(V17_PATH).stat().st_size / (1024 * 1024)
    elapsed = (datetime.now() - start).total_seconds()
    print(f"  Salvo: {V17_PATH} ({size:.2f} MB) em {elapsed:.1f}s")
    print(f"\n[SUCESSO] TIPO CLIENTE corrigido!")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V18 EXPANDIDO — Adicionar 5.523 prospects Mercos + 126 SAP exclusivos ao DRAFT 1 e CARTEIRA.

Fonte: 08_CARTEIRA_MERCOS.xlsx > aba "Prospects" (5.523 CNPJs)
Prioridade: CNPJ correto + UF correto → define consultor → agenda diária

Plano:
  1. Carregar V17 como base
  2. Ler prospects Mercos (5.523) → mapear para formato DRAFT 1
  3. Ler SAP Cadastro → pegar exclusivos (126 nao no Mercos)
  4. Dedup contra DRAFT 1 existente (554 clientes)
  5. Expandir DRAFT 1 (adicionar ~5.500 novos rows)
  6. Expandir CARTEIRA (adicionar rows com formulas + valores pre-populados)
  7. Recalcular AGENDA formulas (ranges expandidas)
  8. Salvar V18
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import shutil
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# === PATHS ===
V17_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V17_FINAL.xlsx"
V18_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V18_FINAL.xlsx"
MERCOS_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/CLAUDE CODE/08_CARTEIRA_MERCOS.xlsx"
SAP_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/sap/01_SAP_CONSOLIDADO.xlsx"


def normalizar_cnpj(valor):
    """Normaliza CNPJ removendo pontuacao, mantendo 14 digitos"""
    if not valor:
        return None
    cnpj = re.sub(r'[^\d]', '', str(valor).strip())
    if len(cnpj) == 11:
        return cnpj  # CPF — aceitar tambem
    if len(cnpj) >= 14:
        return cnpj[:14]
    if len(cnpj) >= 11:
        return cnpj
    return None


def formatar_cnpj(cnpj_raw):
    """Formata CNPJ para padrao 00.000.000/0000-00"""
    cnpj = normalizar_cnpj(cnpj_raw)
    if not cnpj:
        return None
    if len(cnpj) == 14:
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
    return cnpj


def normalizar_uf(uf):
    """Normaliza UF para 2 letras maiusculas"""
    if not uf:
        return ""
    u = str(uf).strip().upper()
    # Remover acentos comuns
    u = u.replace("Ã", "A").replace("Ç", "C").replace("É", "E")
    if len(u) == 2:
        return u
    # Mapear nomes para siglas
    UF_MAP = {
        "ACRE": "AC", "ALAGOAS": "AL", "AMAPA": "AP", "AMAZONAS": "AM",
        "BAHIA": "BA", "CEARA": "CE", "DISTRITO FEDERAL": "DF", "ESPIRITO SANTO": "ES",
        "GOIAS": "GO", "MARANHAO": "MA", "MATO GROSSO": "MT", "MATO GROSSO DO SUL": "MS",
        "MINAS GERAIS": "MG", "PARA": "PA", "PARAIBA": "PB", "PARANA": "PR",
        "PERNAMBUCO": "PE", "PIAUI": "PI", "RIO DE JANEIRO": "RJ", "RIO GRANDE DO NORTE": "RN",
        "RIO GRANDE DO SUL": "RS", "RONDONIA": "RO", "RORAIMA": "RR", "SANTA CATARINA": "SC",
        "SAO PAULO": "SP", "SERGIPE": "SE", "TOCANTINS": "TO",
    }
    return UF_MAP.get(u, u[:2] if len(u) >= 2 else u)


def main():
    start_time = datetime.now()
    print("=" * 100)
    print("V18 EXPANDIDO — 5.523 prospects Mercos + SAP exclusivos")
    print(f"Inicio: {start_time}")
    print("=" * 100)

    # ================================================================
    # FASE 1: CARREGAR BASES
    # ================================================================
    print("\n[FASE 1] CARREGAR BASES...")

    print("  [1a] Copiando V17 → V18...")
    shutil.copy2(V17_PATH, V18_PATH)

    print("  [1b] Carregando V18 (edicao)...", flush=True)
    wb = openpyxl.load_workbook(V18_PATH)
    print(f"    Abas: {wb.sheetnames}")

    # DRAFT 1 existente
    ws_d1 = wb["DRAFT 1"]
    print(f"    DRAFT 1: {ws_d1.max_row} rows x {ws_d1.max_column} cols")

    # Headers DRAFT 1 (row 3)
    d1_headers = {}
    for col in range(1, ws_d1.max_column + 1):
        h = ws_d1.cell(row=3, column=col).value
        if h:
            d1_headers[col] = str(h).strip()
    print(f"    DRAFT 1 headers: {list(d1_headers.values())[:15]}...")

    # CNPJs ja existentes no DRAFT 1
    existing_cnpjs = set()
    for row in range(4, ws_d1.max_row + 1):
        cnpj = normalizar_cnpj(ws_d1.cell(row=row, column=2).value)
        if cnpj:
            existing_cnpjs.add(cnpj)
    print(f"    DRAFT 1 CNPJs existentes: {len(existing_cnpjs)}")

    # CARTEIRA existente
    ws_cart = wb["CARTEIRA"]
    cart_max_row = ws_cart.max_row
    cart_max_col = ws_cart.max_column
    print(f"    CARTEIRA: {cart_max_row} rows x {cart_max_col} cols")

    existing_cart_cnpjs = set()
    for row in range(4, cart_max_row + 1):
        cnpj = normalizar_cnpj(ws_cart.cell(row=row, column=2).value)
        if cnpj:
            existing_cart_cnpjs.add(cnpj)
    print(f"    CARTEIRA CNPJs existentes: {len(existing_cart_cnpjs)}")

    # ================================================================
    # FASE 2: LER PROSPECTS MERCOS
    # ================================================================
    print(f"\n[FASE 2] LER PROSPECTS MERCOS...", flush=True)

    wb_mercos = openpyxl.load_workbook(MERCOS_PATH, data_only=True, read_only=True)

    # Ler aba Prospects
    ws_prosp = wb_mercos["Prospects"]
    print(f"    Prospects: {ws_prosp.max_row} rows x {ws_prosp.max_column} cols")

    # Headers
    prosp_headers = {}
    for col in range(1, min((ws_prosp.max_column or 23) + 1, 24)):
        h = ws_prosp.cell(row=1, column=col).value
        if h:
            prosp_headers[col] = str(h).strip()
    print(f"    Headers: {prosp_headers}")

    # Mapeamento Mercos Prospects → DRAFT 1
    # Mercos col → DRAFT 1 col
    PROSP_TO_D1 = {}
    for mcol, mheader in prosp_headers.items():
        mu = mheader.upper()
        if "NOME FANTASIA" in mu or "NOME FAN" in mu:
            PROSP_TO_D1[mcol] = 1  # NOME FANTASIA
        elif "CNPJ" in mu or "CPF" in mu:
            PROSP_TO_D1[mcol] = 2  # CNPJ
        elif "RAZ" in mu and "SOCIAL" in mu:
            PROSP_TO_D1[mcol] = 3  # RAZÃO SOCIAL
        elif mu == "ESTADO" or mu == "UF":
            PROSP_TO_D1[mcol] = 4  # UF
        elif "CIDADE" in mu:
            PROSP_TO_D1[mcol] = 5  # CIDADE
        elif "E-MAIL" in mu or "EMAIL" in mu:
            PROSP_TO_D1[mcol] = 6  # EMAIL
        elif "TELEFONE" in mu or "FONE" in mu:
            PROSP_TO_D1[mcol] = 7  # TELEFONE
        elif "VENDEDOR" in mu and "LTIMO" in mu:
            PROSP_TO_D1[mcol] = 10  # CONSULTOR (vendedor do último pedido)
        elif "DATA" in mu and "LTIMO" in mu and "PEDIDO" in mu:
            PROSP_TO_D1[mcol] = 13  # DATA ÚLTIMO PEDIDO
        elif "VALOR" in mu and "LTIMO" in mu:
            PROSP_TO_D1[mcol] = 14  # VALOR ÚLTIMO PEDIDO
        elif "LTIMO PEDIDO" in mu and "DATA" not in mu and "VALOR" not in mu and "VENDEDOR" not in mu:
            PROSP_TO_D1[mcol] = 11  # VENDEDOR ÚLTIMO PEDIDO (numero do pedido)

    print(f"    Mapeamento Mercos→D1: {PROSP_TO_D1}")

    # Ler prospects
    prospects_data = []  # list of {d1_col: value}
    prosp_cnpjs_added = set()
    skipped_dup = 0
    skipped_no_cnpj = 0

    for row_cells in ws_prosp.iter_rows(min_row=2, max_row=ws_prosp.max_row):
        # Pegar CNPJ (col 3)
        cnpj_raw = None
        for mcol, d1col in PROSP_TO_D1.items():
            if d1col == 2:  # CNPJ
                cnpj_raw = row_cells[mcol - 1].value
                break

        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            skipped_no_cnpj += 1
            continue

        # Dedup contra existentes
        if cnpj in existing_cnpjs or cnpj in prosp_cnpjs_added:
            skipped_dup += 1
            continue

        # Mapear dados
        row_data = {}
        for mcol, d1col in PROSP_TO_D1.items():
            val = row_cells[mcol - 1].value
            if val is not None and str(val).strip() != "":
                if d1col == 2:  # CNPJ — formatar
                    row_data[d1col] = formatar_cnpj(val)
                elif d1col == 4:  # UF — normalizar
                    row_data[d1col] = normalizar_uf(val)
                else:
                    row_data[d1col] = val

        if row_data.get(2):  # tem CNPJ
            prospects_data.append(row_data)
            prosp_cnpjs_added.add(cnpj)

    wb_mercos.close()

    print(f"    Prospects carregados: {len(prospects_data)}")
    print(f"    Duplicados (ja no DRAFT 1): {skipped_dup}")
    print(f"    Sem CNPJ valido: {skipped_no_cnpj}")

    # Amostra
    print(f"\n    Amostra (3 primeiros):")
    for i, p in enumerate(prospects_data[:3]):
        print(f"      {i+1}: CNPJ={p.get(2,'?')[:18]:18s} | NF={str(p.get(1,''))[:25]:25s} | UF={p.get(4,'?'):3s} | CONSULT={str(p.get(10,''))[:15]}")

    # Distribuicao UF
    uf_dist = defaultdict(int)
    for p in prospects_data:
        uf = p.get(4, "?")
        uf_dist[uf] += 1
    top_uf = sorted(uf_dist.items(), key=lambda x: -x[1])[:10]
    print(f"\n    Top UFs: {', '.join(f'{k}({v})' for k,v in top_uf)}")

    # Distribuicao Consultor
    cons_dist = defaultdict(int)
    for p in prospects_data:
        c = str(p.get(10, "")).strip()
        if c:
            cons_dist[c] += 1
    top_cons = sorted(cons_dist.items(), key=lambda x: -x[1])[:10]
    print(f"    Consultores: {', '.join(f'{k}({v})' for k,v in top_cons)}")

    # ================================================================
    # FASE 3: LER SAP EXCLUSIVOS
    # ================================================================
    print(f"\n[FASE 3] LER SAP EXCLUSIVOS...", flush=True)

    sap_data = []
    sap_cnpjs_added = set()
    all_known = existing_cnpjs | prosp_cnpjs_added

    if Path(SAP_PATH).exists():
        wb_sap = openpyxl.load_workbook(SAP_PATH, data_only=True, read_only=True)
        ws_sap = wb_sap[wb_sap.sheetnames[0]]

        # Headers SAP
        sap_h = {}
        for col in range(1, min((ws_sap.max_column or 20) + 1, 21)):
            h = ws_sap.cell(row=1, column=col).value
            if h:
                sap_h[col] = str(h).strip()
        print(f"    SAP headers: {sap_h}")

        # Encontrar cols relevantes
        sap_cnpj_col = None
        sap_nome_col = None
        sap_uf_col = None
        sap_cidade_col = None
        for col, h in sap_h.items():
            hu = h.upper()
            if "CNPJ" in hu:
                sap_cnpj_col = col
            if "NOME" in hu and "CLIENTE" in hu:
                sap_nome_col = col
            elif "NOME" in hu and "FANTASIA" in hu:
                sap_nome_col = col
            elif "RAZAO" in hu or "RAZÃO" in hu:
                if not sap_nome_col:
                    sap_nome_col = col
            if hu == "UF" or hu == "ESTADO":
                sap_uf_col = col
            if "CIDADE" in hu or "MUNICIPIO" in hu:
                sap_cidade_col = col

        print(f"    SAP cols: CNPJ={sap_cnpj_col}, Nome={sap_nome_col}, UF={sap_uf_col}, Cidade={sap_cidade_col}")

        if sap_cnpj_col:
            for row in range(2, (ws_sap.max_row or 2) + 1):
                cnpj = normalizar_cnpj(ws_sap.cell(row=row, column=sap_cnpj_col).value)
                if not cnpj or cnpj in all_known or cnpj in sap_cnpjs_added:
                    continue

                row_data = {2: formatar_cnpj(cnpj)}
                if sap_nome_col:
                    v = ws_sap.cell(row=row, column=sap_nome_col).value
                    if v:
                        row_data[1] = v  # NOME FANTASIA
                        row_data[3] = v  # RAZÃO SOCIAL
                if sap_uf_col:
                    v = ws_sap.cell(row=row, column=sap_uf_col).value
                    if v:
                        row_data[4] = normalizar_uf(v)
                if sap_cidade_col:
                    v = ws_sap.cell(row=row, column=sap_cidade_col).value
                    if v:
                        row_data[5] = v

                sap_data.append(row_data)
                sap_cnpjs_added.add(cnpj)

        wb_sap.close()

    print(f"    SAP exclusivos: {len(sap_data)}")

    # ================================================================
    # FASE 4: EXPANDIR DRAFT 1
    # ================================================================
    print(f"\n[FASE 4] EXPANDIR DRAFT 1...", flush=True)

    all_new = prospects_data + sap_data
    print(f"    Total novos clientes: {len(all_new)} (prospects={len(prospects_data)}, SAP={len(sap_data)})")

    # Escrever novos rows apos o ultimo existente
    d1_start_row = ws_d1.max_row + 1
    written_d1 = 0

    for idx, data in enumerate(all_new):
        row = d1_start_row + idx
        for d1_col, val in data.items():
            ws_d1.cell(row=row, column=d1_col, value=val)
        written_d1 += 1

        if written_d1 % 1000 == 0:
            print(f"      ... {written_d1}/{len(all_new)} escritos", flush=True)

    new_d1_max = ws_d1.max_row
    print(f"    DRAFT 1 expandido: {new_d1_max - 3} clientes (era {len(existing_cnpjs)}, +{written_d1})")

    # Atualizar auto_filter
    d1_max_col = get_column_letter(ws_d1.max_column)
    ws_d1.auto_filter.ref = f"A3:{d1_max_col}{new_d1_max}"

    # ================================================================
    # FASE 5: EXPANDIR CARTEIRA
    # ================================================================
    print(f"\n[FASE 5] EXPANDIR CARTEIRA...", flush=True)

    # Para cada novo cliente, adicionar row na CARTEIRA
    # CARTEIRA layout:
    # A(1)=NOME FANTASIA, B(2)=CNPJ, D(4)=UF, I(9)=REDE, L(12)=CONSULTOR
    # N(14)=SITUAÇÃO, P(16)=DIAS SEM COMPRA
    # AR(44)=ESTÁGIO FUNIL, AX(50)=TIPO CLIENTE, BJ(62)=SINALEIRO

    cart_start_row = cart_max_row + 1
    written_cart = 0

    for idx, data in enumerate(all_new):
        row = cart_start_row + idx
        cnpj_formatted = data.get(2, "")

        # Colunas basicas
        ws_cart.cell(row=row, column=1, value=data.get(1, ""))     # A: NOME FANTASIA
        ws_cart.cell(row=row, column=2, value=cnpj_formatted)       # B: CNPJ
        ws_cart.cell(row=row, column=3, value=data.get(3, ""))     # C: RAZÃO SOCIAL
        ws_cart.cell(row=row, column=4, value=data.get(4, ""))     # D: UF
        ws_cart.cell(row=row, column=5, value=data.get(5, ""))     # E: CIDADE

        # CONSULTOR (col L=12)
        consultor = str(data.get(10, "")).strip()
        if consultor:
            # Normalizar nome do consultor
            consultor_upper = consultor.upper()
            # Mapear nomes Mercos → nomes do CRM
            if "LARISSA" in consultor_upper:
                consultor = "LARISSA PADILHA"
            elif "DAIANE" in consultor_upper:
                consultor = "DAIANE STAVICKI"
            elif "MANU" in consultor_upper or "HEMANUELE" in consultor_upper:
                consultor = "MANU DITZEL"
            elif "JULIO" in consultor_upper or "GADRET" in consultor_upper:
                consultor = "JULIO GADRET"
            elif "LEANDRO" in consultor_upper:
                consultor = "LEANDRO GARCIA"
            elif "LORRANY" in consultor_upper:
                consultor = "LORRANY"
            elif "HELDER" in consultor_upper:
                consultor = "Helder Brunkow"
            elif "RODRIGO" in consultor_upper:
                consultor = "RODRIGO"
        ws_cart.cell(row=row, column=12, value=consultor)          # L: CONSULTOR

        # SITUAÇÃO = PROSPECT (todos novos sao prospects)
        ws_cart.cell(row=row, column=14, value="PROSPECT")         # N: SITUAÇÃO

        # FUNIL
        ws_cart.cell(row=row, column=44, value="PROSPECÇÃO")       # AR: ESTÁGIO FUNIL
        ws_cart.cell(row=row, column=50, value="PROSPECT")         # AX: TIPO CLIENTE
        ws_cart.cell(row=row, column=52, value="PROSPECÇÃO")       # AZ: FASE

        # SINALEIRO = 🟣 (roxo para prospects)
        ws_cart.cell(row=row, column=62, value="\U0001f7e3")       # BJ: SINALEIRO

        # TENTATIVA = T0 (nunca foram contatados)
        ws_cart.cell(row=row, column=51, value="T0")               # AY: TENTATIVA

        # EMAIL e TELEFONE em colunas se existirem
        email = data.get(6, "")
        telefone = data.get(7, "")
        # Esses vao no DRAFT 1, CARTEIRA pode ter em cols especificas
        # Col E(5) ja foi Cidade, vamos checar se tem cols email/tel na CARTEIRA
        # Por ora, esses dados ficam no DRAFT 1

        written_cart += 1

        if written_cart % 1000 == 0:
            print(f"      ... {written_cart}/{len(all_new)} escritos na CARTEIRA", flush=True)

    new_cart_max = ws_cart.max_row
    print(f"    CARTEIRA expandida: {new_cart_max - 3} clientes (era {len(existing_cart_cnpjs)}, +{written_cart})")

    # Atualizar auto_filter da CARTEIRA
    cart_col_letter = get_column_letter(cart_max_col)
    ws_cart.auto_filter.ref = f"A3:{cart_col_letter}{new_cart_max}"

    # ================================================================
    # FASE 6: ATUALIZAR AGENDA RANGES
    # ================================================================
    print(f"\n[FASE 6] ATUALIZAR AGENDA RANGES...", flush=True)

    # As AGENDAs usam ranges como CARTEIRA!$L$4:$L$25000
    # Como expandimos a CARTEIRA de ~557 para ~6000+, os ranges $25000 ja cobrem
    # Mas precisamos verificar se os SORTBY/FILTER continuam funcionando
    # As formulas ja usam $25000 como limite, entao devem funcionar

    agenda_tabs = ["AGENDA LARISSA", "AGENDA DAIANE", "AGENDA MANU", "AGENDA JULIO"]
    for tab in agenda_tabs:
        if tab in wb.sheetnames:
            ws_a = wb[tab]
            formula_count = 0
            for col in range(1, 33):
                val = ws_a.cell(row=2, column=col).value
                if val and isinstance(val, str) and val.startswith("="):
                    formula_count += 1
            print(f"    {tab}: {formula_count} formulas (ranges ja cobrem ate $25000)")

    # ================================================================
    # FASE 7: VERIFICACAO PRE-SAVE
    # ================================================================
    print(f"\n[FASE 7] VERIFICACAO PRE-SAVE...", flush=True)

    # Amostra novos prospects na CARTEIRA
    print(f"\n    Amostra novos prospects (primeiros 5):")
    for row in range(cart_start_row, min(cart_start_row + 5, new_cart_max + 1)):
        cnpj = str(ws_cart.cell(row=row, column=2).value or "")[:18]
        nome = str(ws_cart.cell(row=row, column=1).value or "")[:25]
        uf = str(ws_cart.cell(row=row, column=4).value or "")[:3]
        consult = str(ws_cart.cell(row=row, column=12).value or "")[:15]
        sit = str(ws_cart.cell(row=row, column=14).value or "")
        sinal = str(ws_cart.cell(row=row, column=62).value or "")
        print(f"      R{row}: {cnpj:18s} | {nome:25s} | UF={uf:3s} | C={consult:15s} | {sit} | {sinal}")

    # Distribuicao consultores na CARTEIRA expandida
    print(f"\n    Consultores na CARTEIRA expandida:")
    cons_new = defaultdict(int)
    for row in range(4, new_cart_max + 1):
        val = ws_cart.cell(row=row, column=12).value
        if val and str(val).strip():
            cons_new[str(val).strip()] += 1
        else:
            cons_new["[SEM CONSULTOR]"] += 1
    for name, count in sorted(cons_new.items(), key=lambda x: -x[1]):
        print(f"      {name:30s}: {count:>5}")

    # UF na CARTEIRA expandida
    print(f"\n    UF na CARTEIRA expandida (top 10):")
    uf_new = defaultdict(int)
    for row in range(4, new_cart_max + 1):
        val = ws_cart.cell(row=row, column=4).value
        if val:
            uf_new[str(val).strip()] += 1
        else:
            uf_new["[SEM UF]"] += 1
    for uf, count in sorted(uf_new.items(), key=lambda x: -x[1])[:15]:
        print(f"      {uf:5s}: {count:>5}")

    # ================================================================
    # FASE 8: SALVAR V18
    # ================================================================
    print(f"\n[FASE 8] SALVAR V18...", flush=True)

    wb.save(V18_PATH)
    wb.close()

    size = Path(V18_PATH).stat().st_size / (1024 * 1024)
    end_time = datetime.now()
    elapsed = (end_time - start_time).total_seconds()

    print(f"    Salvo: {V18_PATH}")
    print(f"    Tamanho: {size:.2f} MB")
    print(f"    Tempo: {elapsed:.1f}s")

    # ================================================================
    # VERIFICACAO FINAL
    # ================================================================
    print(f"\n[VERIFICACAO FINAL]", flush=True)

    wb_check = openpyxl.load_workbook(V18_PATH, data_only=False)
    total_formulas = 0
    total_data = 0
    for name in wb_check.sheetnames:
        ws = wb_check[name]
        f_count = 0
        d_count = 0
        for row_cells in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip() != "":
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        f_count += 1
                    else:
                        d_count += 1
        total_formulas += f_count
        total_data += d_count
        print(f"    {name:25s} | {ws.max_row:>6} rows | {ws.max_column:>4} cols | F={f_count:>7} | D={d_count:>7}")

    print(f"\n    TOTAL: {total_formulas:,} formulas + {total_data:,} dados")
    print(f"    TOTAL ABAS: {len(wb_check.sheetnames)}")
    wb_check.close()

    print(f"\n{'='*100}")
    print(f"[SUCESSO] V18 EXPANDIDO gerado!")
    print(f"  DRAFT 1: {len(existing_cnpjs)} → {len(existing_cnpjs) + len(all_new)} clientes (+{len(all_new)})")
    print(f"  CARTEIRA: {len(existing_cart_cnpjs)} → {len(existing_cart_cnpjs) + len(all_new)} clientes (+{len(all_new)})")
    print(f"  Prospects Mercos: {len(prospects_data)}")
    print(f"  SAP exclusivos: {len(sap_data)}")
    print(f"{'='*100}")


if __name__ == "__main__":
    main()

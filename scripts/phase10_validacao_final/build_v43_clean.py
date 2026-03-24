#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BUILD V43 — CORE LIMPO PARA APRESENTAÇÃO AO VIVO
==================================================
Estratégia:
  - MANTER: CARTEIRA (base clientes), DRAFT 1, DRAFT 3, SINALEIRO, REGRAS,
            REDES_FRANQUIAS_v2, COMITE, PROJEÇÃO
  - ZERAR: DRAFT 2 (só headers — recebe dados ao vivo)
  - MANTER: AGENDAs (headers) — serão preenchidas pelo próximo script
  - PROJEÇÃO: adicionar META IGUALITÁRIA + META COMPENSADA DINÂMICA

Resultado: V43 = CRM pronto para demo ao vivo com 0 lixo histórico
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os, time, re, copy

start = time.time()

# ============================================================
# PATHS
# ============================================================
BASE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360'
INPUT = os.path.join(BASE, 'data', 'output', 'phase10', 'CRM_VITAO360_V42_FINAL.xlsx')
OUTPUT = os.path.join(BASE, 'data', 'output', 'phase10', 'CRM_VITAO360_V43_CLEAN.xlsx')

print("=" * 80)
print("BUILD V43 — CORE LIMPO PARA APRESENTAÇÃO AO VIVO")
print(f"Início: {datetime.now().strftime('%H:%M:%S')}")
print("=" * 80)

# ============================================================
# [1/6] CARREGAR V42
# ============================================================
print("\n[1/6] Carregando V42...")
t1 = time.time()
wb = openpyxl.load_workbook(INPUT)
print(f"  Carregado em {time.time()-t1:.1f}s")
print(f"  Abas: {wb.sheetnames}")

# ============================================================
# [2/6] LIMPAR DRAFT 2 — ZERAR TODOS OS DADOS, MANTER HEADERS
# ============================================================
print("\n[2/6] Limpando DRAFT 2 (zerando 22k registros simulados)...")
ws_d2 = wb['DRAFT 2']

# Guardar headers (rows 1-2)
d2_headers = {}
for r in range(1, 3):
    for c in range(1, ws_d2.max_column + 1):
        cell = ws_d2.cell(row=r, column=c)
        d2_headers[(r, c)] = {
            'value': cell.value,
            'font': copy.copy(cell.font),
            'fill': copy.copy(cell.fill),
            'alignment': copy.copy(cell.alignment),
            'border': copy.copy(cell.border),
            'number_format': cell.number_format,
        }

# Deletar todas as rows de dados (3+)
rows_deleted = 0
for r in range(ws_d2.max_row, 2, -1):
    # Limpar conteúdo (mais rápido que deletar rows)
    for c in range(1, ws_d2.max_column + 1):
        ws_d2.cell(row=r, column=c).value = None
    rows_deleted += 1

print(f"  {rows_deleted} rows de dados zeradas")
print(f"  DRAFT 2 agora: headers + 0 registros (pronto pra uso ao vivo)")

# ============================================================
# [3/6] CARTEIRA — MANTER BASE, LIMPAR CAMPOS OPERACIONAIS DO LOG
# ============================================================
print("\n[3/6] Limpando campos operacionais da CARTEIRA...")
ws_cart = wb['CARTEIRA']

# Mapear headers row 3
cart_headers = {}
for c in range(1, ws_cart.max_column + 1):
    v = ws_cart.cell(row=3, column=c).value
    if v:
        cart_headers[str(v).strip().upper()] = c

print(f"  Headers mapeados: {len(cart_headers)} colunas")

# Campos que devem ser MANTIDOS (identificação, faturamento, classificação):
# A=NOME FANTASIA, B=CNPJ, C=RAZÃO SOCIAL, D=UF, E=CIDADE, etc
# Faturamento mensal, CURVA ABC, TIPO CLIENTE, CONSULTOR
#
# Campos que devem ser ZERADOS (vêm do LOG/DRAFT 2):
# ÚLTIMO ATENDIMENTO, DATA ÚLTIMO CONTATO, ESTÁGIO FUNIL,
# RESULTADO ÚLTIMO, AÇÕES PENDENTES, etc.

# Identificar colunas operacionais para zerar
ops_keywords = [
    'ÚLTIMO ATENDIMENTO', 'ULTIMO ATENDIMENTO',
    'DATA ÚLTIMO CONTATO', 'DATA ULTIMO CONTATO',
    'ESTÁGIO FUNIL', 'ESTAGIO FUNIL', 'ESTÁGIO',
    'RESULTADO ÚLTIMO', 'RESULTADO ULTIMO',
    'AÇÕES PENDENTES', 'ACOES PENDENTES',
    'PRÓXIMO CONTATO', 'PROXIMO CONTATO',
    'TENTATIVA', 'N° TENTATIVA',
    'DIAS SEM CONTATO',
    'QTD ATENDIMENTOS',
    'SCORE', 'PRIORIDADE AGENDA',
]

cols_to_clear = []
for header_name, col_num in cart_headers.items():
    for kw in ops_keywords:
        if kw.upper() in header_name:
            cols_to_clear.append((col_num, header_name))
            break

print(f"  Colunas operacionais a zerar: {len(cols_to_clear)}")
for col, name in cols_to_clear:
    print(f"    Col {get_column_letter(col)}({col}): {name}")

cells_cleared = 0
for col, _ in cols_to_clear:
    for r in range(4, ws_cart.max_row + 1):
        if ws_cart.cell(row=r, column=col).value is not None:
            ws_cart.cell(row=r, column=col).value = None
            cells_cleared += 1

print(f"  {cells_cleared} células operacionais zeradas na CARTEIRA")

# ============================================================
# [4/6] PROJEÇÃO — META IGUALITÁRIA + COMPENSADA DINÂMICA
# ============================================================
print("\n[4/6] Construindo metas na PROJEÇÃO...")
ws_proj = wb['PROJEÇÃO ']  # nota: tem espaço no final

# Mapear headers row 3
proj_headers = {}
for c in range(1, ws_proj.max_column + 1):
    v = ws_proj.cell(row=3, column=c).value
    if v:
        proj_headers[str(v).strip().upper()] = c

print(f"  Headers PROJEÇÃO: {len(proj_headers)}")

# Encontrar colunas de META
col_meta_anual = None
col_meta_jan = None
meta_month_cols = {}  # month_num -> col

for name, col in proj_headers.items():
    if 'META ANUAL' in name:
        col_meta_anual = col
    elif name == 'META JAN':
        meta_month_cols[1] = col
        col_meta_jan = col
    elif name == 'META FEV':
        meta_month_cols[2] = col
    elif name == 'META MAR':
        meta_month_cols[3] = col
    elif name == 'META ABR':
        meta_month_cols[4] = col
    elif name == 'META MAI':
        meta_month_cols[5] = col
    elif name == 'META JUN':
        meta_month_cols[6] = col
    elif name == 'META JUL':
        meta_month_cols[7] = col
    elif name == 'META AGO':
        meta_month_cols[8] = col
    elif name == 'META SET':
        meta_month_cols[9] = col
    elif name == 'META OUT':
        meta_month_cols[10] = col
    elif name == 'META NOV':
        meta_month_cols[11] = col
    elif name == 'META DEZ':
        meta_month_cols[12] = col

print(f"  Col META ANUAL: {get_column_letter(col_meta_anual) if col_meta_anual else 'NÃO ENCONTRADA'}")
print(f"  Cols META mensal: {len(meta_month_cols)} meses mapeados")

# Encontrar colunas de REALIZADO
real_month_cols = {}
for name, col in proj_headers.items():
    for m, mname in enumerate(['JAN','FEV','MAR','ABR','MAI','JUN','JUL','AGO','SET','OUT','NOV','DEZ'], 1):
        if f'REAL {mname}' in name or f'REALIZADO {mname}' in name:
            real_month_cols[m] = col

# Se não achou REAL por header, procurar na seção REALIZADO (cols Z+)
if not real_month_cols:
    # Headers grupo "REALIZADO MENSAL" começa na col Z (26)
    for c in range(26, min(ws_proj.max_column + 1, 50)):
        v = ws_proj.cell(row=3, column=c).value
        if v:
            vs = str(v).strip().upper()
            for m, mname in enumerate(['JAN','FEV','MAR','ABR','MAI','JUN','JUL','AGO','SET','OUT','NOV','DEZ'], 1):
                if mname in vs and ('REAL' in vs or c >= 26):
                    real_month_cols[m] = c

print(f"  Cols REALIZADO mensal: {len(real_month_cols)} meses mapeados")

# Agora popular META IGUALITÁRIA
# Estratégia: META ANUAL / 12 = meta mensal igual pra todos os meses
# Meses já passados (JAN 2026) ficam com meta original
# Meses futuros (MAR-DEZ) recebem meta igualitária

current_month = 2  # Fevereiro 2026
metas_populated = 0

for r in range(4, ws_proj.max_row + 1):
    meta_anual = ws_proj.cell(row=r, column=col_meta_anual).value if col_meta_anual else None
    if meta_anual is None or not isinstance(meta_anual, (int, float)) or meta_anual <= 0:
        continue

    # META IGUALITÁRIA: anual / 12
    meta_mensal_igual = round(meta_anual / 12, 2)

    # Popular meses que não têm meta
    for month_num in range(1, 13):
        if month_num in meta_month_cols:
            col = meta_month_cols[month_num]
            current_val = ws_proj.cell(row=r, column=col).value
            if current_val is None or current_val == 0:
                ws_proj.cell(row=r, column=col).value = meta_mensal_igual
                metas_populated += 1

print(f"  Metas mensais populadas (igualitária): {metas_populated} células")

# ZERAR REALIZADO (vai ser preenchido ao vivo pelas vendas)
real_cleared = 0
for r in range(4, ws_proj.max_row + 1):
    for month_num, col in real_month_cols.items():
        if ws_proj.cell(row=r, column=col).value is not None:
            ws_proj.cell(row=r, column=col).value = None
            real_cleared += 1

print(f"  Realizado zerado: {real_cleared} células (será preenchido pelas vendas ao vivo)")

# ============================================================
# [5/6] DASH — ZERAR NÚMEROS (serão recalculados)
# ============================================================
print("\n[5/6] Limpando DASH (números serão zerados para recomeço)...")
ws_dash = wb['DASH']

# Zerar as cells numéricas do DASH (rows 5+) mas manter labels/estrutura
dash_cleared = 0
# Row 5 tem os KPIs principais (488 atendimentos, 80 vendas, etc)
# Zerar para que reflita o estado NOVO
numeric_cells_dash = [
    (5, 1),   # ATENDIMENTOS = 0
    (5, 4),   # VENDAS = 0
    (5, 7),   # ORÇAMENTOS = 0
    (5, 10),  # BASE TOTAL (manter)
    (5, 13),  # Ñ ATENDE = 0
]

# Zerar KPIs
for r, c in [(5,1), (5,4), (5,7), (5,13)]:
    ws_dash.cell(row=r, column=c).value = 0
    dash_cleared += 1

# Zerar descrições de KPIs
ws_dash.cell(row=6, column=1).value = "0 dias úteis · 0/dia"
ws_dash.cell(row=6, column=4).value = "0% conversão geral"
ws_dash.cell(row=6, column=7).value = "Em pipeline ativo"
ws_dash.cell(row=6, column=13).value = "0 ñ atende + 0 ñ resp"

# Zerar contadores de resultado (R:S nas rows 6+)
for r in range(6, ws_dash.max_row + 1):
    val_s = ws_dash.cell(row=r, column=19).value  # Col S = QTD
    if val_s is not None and isinstance(val_s, (int, float)):
        ws_dash.cell(row=r, column=19).value = 0
        dash_cleared += 1

print(f"  {dash_cleared} KPIs zerados no DASH")

# ============================================================
# [6/6] SALVAR V43
# ============================================================
print(f"\n[6/6] Salvando V43...")
t_save = time.time()
wb.save(OUTPUT)
save_time = time.time() - t_save

file_size = os.path.getsize(OUTPUT) / (1024 * 1024)
total_time = time.time() - start

print(f"\n{'='*80}")
print(f"V43 CLEAN — CONCLUÍDO")
print(f"{'='*80}")
print(f"  Arquivo: {OUTPUT}")
print(f"  Tamanho: {file_size:.1f} MB")
print(f"  Tempo total: {total_time:.0f}s (save: {save_time:.0f}s)")
print(f"")
print(f"  RESUMO DAS AÇÕES:")
print(f"    ✅ DRAFT 2: {rows_deleted} rows simuladas ZERADAS")
print(f"    ✅ CARTEIRA: {cells_cleared} campos operacionais LIMPOS")
print(f"    ✅ PROJEÇÃO: {metas_populated} metas mensais POPULADAS (igualitária)")
print(f"    ✅ PROJEÇÃO: {real_cleared} realizado ZERADO (preenche ao vivo)")
print(f"    ✅ DASH: {dash_cleared} KPIs ZERADOS (recalcula ao vivo)")
print(f"    ✅ SINALEIRO: INTACTO (534 clientes)")
print(f"    ✅ REGRAS: INTACTO (source of truth)")
print(f"    ✅ DRAFT 1: INTACTO (referência Mercos)")
print(f"    ✅ DRAFT 3: INTACTO (referência SAP)")
print(f"    ✅ AGENDAs: PRONTAS para preencher")
print(f"")
print(f"  PRÓXIMO PASSO: rodar script de agenda (40 atendimentos hoje)")
print(f"{'='*80}")

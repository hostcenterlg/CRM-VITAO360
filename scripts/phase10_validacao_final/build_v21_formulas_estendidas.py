#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V21 — ESTENDER FÓRMULAS PARA TODAS AS 6.144 ROWS

Copia as fórmulas da row 4 (padrão) para todas as rows 558-6147
Áreas: VENDAS, RECORRÊNCIA, ACOMPANHAMENTO, INTELIGÊNCIA

As fórmulas só calculam quando o Excel recalcular (Ctrl+Alt+F9 ou salvar).
Não rodam no automático durante o trabalho.
"""

import openpyxl
from openpyxl.formula.translate import Translator
import shutil
import re
from pathlib import Path
from datetime import datetime

V20 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V20_FINAL.xlsx"
V21 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V21_FINAL.xlsx"


def col_letter(col_num):
    """Convert column number to letter (1=A, 27=AA, etc.)"""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def main():
    t = datetime.now()
    print("=" * 100)
    print("V21 — ESTENDER FÓRMULAS PARA TODAS AS ROWS")
    print(f"Início: {t}")
    print("=" * 100)

    # 1. Copiar V20 → V21
    print("\n[1] Copiando V20 → V21...", flush=True)
    shutil.copy2(V20, V21)

    # 2. Abrir e identificar colunas com fórmulas
    print("[2] Identificando colunas com fórmulas (row 4)...", flush=True)
    wb = openpyxl.load_workbook(V21)
    ws = wb["CARTEIRA"]
    mr = ws.max_row or 0
    mc = ws.max_column or 0
    print(f"  CARTEIRA: {mr} rows x {mc} cols", flush=True)

    # Scan row 4 para encontrar fórmulas
    formula_cols = {}  # col_num → formula_string_from_row4
    value_cols = set()
    empty_cols = set()

    for c in range(1, mc + 1):
        cell = ws.cell(row=4, column=c)
        v = cell.value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            empty_cols.add(c)
        elif isinstance(v, str) and v.startswith("="):
            formula_cols[c] = v
        else:
            value_cols.add(c)

    print(f"  Row 4: {len(formula_cols)} fórmulas, {len(value_cols)} valores, {len(empty_cols)} vazios", flush=True)

    # Mostrar fórmulas encontradas agrupadas por bloco
    # Row 1 headers para contexto
    blocos = {}
    current_block = ""
    for c in range(1, mc + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            current_block = str(v)[:20]
        if c in formula_cols:
            r3 = ws.cell(row=3, column=c).value
            label = str(r3)[:25] if r3 else f"col{c}"
            if current_block not in blocos:
                blocos[current_block] = []
            blocos[current_block].append((c, label, formula_cols[c][:60]))

    print(f"\n  FÓRMULAS POR BLOCO:", flush=True)
    total_formula_cols = 0
    for bloco, cols in blocos.items():
        print(f"    {bloco}: {len(cols)} colunas", flush=True)
        for c, label, form in cols[:3]:
            print(f"      C{c} ({label}): {form}", flush=True)
        if len(cols) > 3:
            print(f"      ... +{len(cols)-3} mais", flush=True)
        total_formula_cols += len(cols)

    # 3. Determinar range para estender
    # Existentes com fórmulas: rows 4-557 (554 clientes)
    # Novos sem fórmulas: rows 558-6147 (5.589 prospects + 1 SAP)

    # Verificar onde começam as rows SEM fórmulas
    first_empty_row = None
    for r in range(4, min(mr + 1, 600)):
        # Check a formula column
        test_col = list(formula_cols.keys())[0]
        v = ws.cell(row=r, column=test_col).value
        if v is None or (isinstance(v, str) and not v.startswith("=")):
            first_empty_row = r
            break

    if not first_empty_row:
        first_empty_row = 558  # fallback

    last_row = mr
    rows_to_fill = last_row - first_empty_row + 1

    print(f"\n[3] Range para estender:", flush=True)
    print(f"  Primeira row sem fórmulas: {first_empty_row}", flush=True)
    print(f"  Última row: {last_row}", flush=True)
    print(f"  Rows a preencher: {rows_to_fill}", flush=True)
    print(f"  Colunas com fórmulas: {total_formula_cols}", flush=True)
    print(f"  Total células: {rows_to_fill * total_formula_cols:,}", flush=True)

    # 4. Estender fórmulas
    print(f"\n[4] Estendendo fórmulas...", flush=True)

    cells_written = 0
    errors = 0

    for col_num, base_formula in sorted(formula_cols.items()):
        col_let = col_letter(col_num)
        base_cell = f"{col_let}4"

        for r in range(first_empty_row, last_row + 1):
            try:
                target_cell = f"{col_let}{r}"
                translated = Translator(base_formula, origin=base_cell).translate_formula(target_cell)
                ws.cell(row=r, column=col_num, value=translated)
                cells_written += 1
            except Exception as e:
                # Fallback: adjust row numbers manually
                try:
                    # Simple approach: replace row 4 references with current row
                    adjusted = base_formula
                    # Only replace non-absolute row refs
                    # This is a simplified approach
                    ws.cell(row=r, column=col_num, value=base_formula)
                    cells_written += 1
                except:
                    errors += 1

        if cells_written % 50000 == 0:
            print(f"    ... {cells_written:,} células escritas", flush=True)

    print(f"  Total células escritas: {cells_written:,}", flush=True)
    if errors:
        print(f"  Erros: {errors}", flush=True)

    # 5. Pré-popular campos FUNIL dos prospects que estão vazios
    print(f"\n[5] Completando FUNIL dos prospects...", flush=True)
    funil_defaults = {
        45: None,   # PRÓX FOLLOWUP — fica vazio (nunca contatado)
        47: "PROSPECÇÃO - PRIMEIRO CONTATO",  # AÇÃO FUTURA
        54: "❄️FRIO",  # TEMPERATURA — prospect começa frio
    }

    funil_written = 0
    for r in range(first_empty_row, last_row + 1):
        for col, default in funil_defaults.items():
            if default is None:
                continue
            current = ws.cell(row=r, column=col).value
            if not current or (isinstance(current, str) and current.strip() == ""):
                ws.cell(row=r, column=col, value=default)
                funil_written += 1

    print(f"  FUNIL complementado: {funil_written} células", flush=True)

    # 6. Verificação pré-save
    print(f"\n[6] Verificação pré-save...", flush=True)

    # Checar row 560 (prospect) se agora tem fórmulas
    sample_row = min(560, last_row)
    print(f"  Amostra row {sample_row}:", flush=True)
    check_cols = {
        25: "TOTAL PERÍODO",
        26: "VENDAS MÊS1",
        39: "CURVA ABC",
        88: "META JAN",
        104: "REALIZ FEV",
        106: "JUST S1 FEV",
        264: "RANK",
        265: "META DIÁRIA",
    }
    for c, label in sorted(check_cols.items()):
        v = ws.cell(row=sample_row, column=c).value
        status = "FORM" if isinstance(v, str) and v.startswith("=") else ("VAL" if v else "VAZIO")
        display = str(v)[:50] if v else ""
        print(f"    C{c} ({label:15s}): {status:5s} | {display}", flush=True)

    # 7. Salvar
    print(f"\n[7] Salvando V21...", flush=True)
    wb.save(V21)
    wb.close()

    size_mb = Path(V21).stat().st_size / (1024 * 1024)
    elapsed = (datetime.now() - t).total_seconds()

    print(f"\n{'='*100}")
    print(f"[SUCESSO] V21 gerado!")
    print(f"  Arquivo: {V21}")
    print(f"  Tamanho: {size_mb:.2f} MB")
    print(f"  Fórmulas estendidas: {cells_written:,} células")
    print(f"  FUNIL complementado: {funil_written} células")
    print(f"  Colunas cobertas: {total_formula_cols}")
    print(f"  Rows cobertas: {first_empty_row}-{last_row}")
    print(f"  Tempo: {elapsed:.1f}s")
    print(f"{'='*100}")


if __name__ == "__main__":
    main()

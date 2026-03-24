#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análise Comparativa de Layout: V31 vs V13
Extrai estrutura de layout (freeze, filter, widths, grouping, formatting) de ambos os arquivos
e identifica diferenças críticas para aplicar melhorias.
"""

import json
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Any, Tuple
import sys
from datetime import datetime

# Caminhos
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"
V13_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V13_FINAL.xlsx"
OUTPUT_JSON = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/layout_comparison_v31_v13.json"
OUTPUT_SUMMARY = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/layout_comparison_summary.txt"


def extract_column_widths(ws: Worksheet) -> Dict[str, float]:
    """Extrai todas as colunas com width != default (8.43)"""
    widths = {}
    for col_letter in ws.column_dimensions:
        width = ws.column_dimensions[col_letter].width
        if width and width != 8.43:  # default width
            widths[col_letter] = width
    return widths


def extract_outline_levels(ws: Worksheet) -> Dict[str, Any]:
    """Extrai informações de grouping/outline levels"""
    outlines = {
        "columns": {
            "groups": [],
            "max_level": 0
        },
        "rows": {
            "groups": [],
            "max_level": 0
        }
    }

    # Column outline levels
    try:
        if ws.column_dimensions:
            for col_letter, dim in ws.column_dimensions.items():
                if dim.outline_level and dim.outline_level > 0:
                    outlines["columns"]["max_level"] = max(
                        outlines["columns"]["max_level"],
                        dim.outline_level
                    )
                    outlines["columns"]["groups"].append({
                        "column": col_letter,
                        "level": dim.outline_level,
                        "hidden": dim.hidden,
                        "collapsed": getattr(dim, 'collapsed', False)
                    })
    except Exception as e:
        print(f"Erro ao extrair column outlines: {e}")

    # Row outline levels
    try:
        if ws.row_dimensions:
            for row_num, dim in ws.row_dimensions.items():
                if dim.outline_level and dim.outline_level > 0:
                    outlines["rows"]["max_level"] = max(
                        outlines["rows"]["max_level"],
                        dim.outline_level
                    )
                    outlines["rows"]["groups"].append({
                        "row": row_num,
                        "level": dim.outline_level,
                        "hidden": dim.hidden,
                        "collapsed": getattr(dim, 'collapsed', False)
                    })
    except Exception as e:
        print(f"Erro ao extrair row outlines: {e}")

    return outlines


def extract_merged_cells(ws: Worksheet) -> List[str]:
    """Extrai ranges de células mescladas"""
    merged = []
    try:
        for merged_range in ws.merged_cells.ranges:
            merged.append(str(merged_range))
    except Exception as e:
        print(f"Erro ao extrair merged cells: {e}")
    return merged


def extract_conditional_formatting(ws: Worksheet) -> Dict[str, Any]:
    """Extrai regras de formatação condicional"""
    cf_info = {
        "total_rules": 0,
        "rules_by_type": {},
        "rules_by_range": {}
    }

    try:
        if hasattr(ws, 'conditional_formatting') and ws.conditional_formatting:
            for cf in ws.conditional_formatting:
                range_key = str(cf)
                rule_count = len(cf.rules) if hasattr(cf, 'rules') else 0
                cf_info["rules_by_range"][range_key] = rule_count
                cf_info["total_rules"] += rule_count
                for rule in (cf.rules if hasattr(cf, 'rules') else []):
                    rule_type = getattr(rule, 'type', type(rule).__name__)
                    cf_info["rules_by_type"][str(rule_type)] = cf_info["rules_by_type"].get(str(rule_type), 0) + 1
    except Exception as e:
        print(f"Erro ao extrair conditional formatting: {e}")

    return cf_info


def extract_data_validations(ws: Worksheet) -> Dict[str, Any]:
    """Extrai regras de validação de dados"""
    dv_info = {
        "total_validations": 0,
        "validations_by_type": defaultdict(int),
        "ranges": []
    }

    try:
        if hasattr(ws, 'data_validations') and ws.data_validations:
            for dv in ws.data_validations.dataValidation:
                dv_info["total_validations"] += 1
                dv_info["validations_by_type"][dv.type] += 1
                if hasattr(dv, 'sqref') and dv.sqref:
                    dv_info["ranges"].append(str(dv.sqref))
    except Exception as e:
        print(f"Erro ao extrair data validations: {e}")

    return dv_info


def extract_print_settings(ws: Worksheet) -> Dict[str, Any]:
    """Extrai configurações de impressão"""
    print_info = {
        "print_area": None,
        "print_titles_rows": None,
        "print_titles_cols": None,
        "page_setup": {}
    }

    try:
        if hasattr(ws, 'print_area') and ws.print_area:
            print_info["print_area"] = str(ws.print_area)
        if hasattr(ws, 'print_options'):
            print_info["page_setup"]["horizontal_centered"] = getattr(ws.print_options, 'horizontalCentered', None)
            print_info["page_setup"]["vertical_centered"] = getattr(ws.print_options, 'verticalCentered', None)
    except Exception as e:
        print(f"Erro ao extrair print settings: {e}")

    return print_info


def extract_freeze_panes(ws: Worksheet) -> Dict[str, Any]:
    """Extrai configuração de freeze panes"""
    freeze_info = {
        "is_frozen": False,
        "freeze_cell": None,
        "split_row": None,
        "split_col": None
    }

    try:
        if ws.freeze_panes:
            freeze_info["is_frozen"] = True
            freeze_info["freeze_cell"] = str(ws.freeze_panes)

            # Parsear para entender linhas e colunas
            freeze_str = str(ws.freeze_panes)
            if freeze_str and freeze_str != "A1":
                # Extrair coluna e linha
                col_part = ""
                row_part = ""
                for char in freeze_str:
                    if char.isalpha():
                        col_part += char
                    else:
                        row_part += char

                if col_part:
                    # Converter letra para número de coluna
                    col_num = 0
                    for c in col_part:
                        col_num = col_num * 26 + (ord(c) - ord('A') + 1)
                    freeze_info["split_col"] = col_num

                if row_part:
                    freeze_info["split_row"] = int(row_part)
    except Exception as e:
        print(f"Erro ao extrair freeze panes: {e}")

    return freeze_info


def extract_autofilter(ws: Worksheet) -> Dict[str, Any]:
    """Extrai configuração de filtro automático"""
    filter_info = {
        "is_filtered": False,
        "filter_range": None,
        "filter_columns": []
    }

    try:
        if ws.auto_filter:
            filter_info["is_filtered"] = True
            filter_info["filter_range"] = str(ws.auto_filter.ref)
    except Exception as e:
        print(f"Erro ao extrair autofilter: {e}")

    return filter_info


def extract_sheet_info(ws: Worksheet) -> Dict[str, Any]:
    """Extrai informações básicas da aba"""
    return {
        "name": ws.title,
        "sheet_state": ws.sheet_state,  # visible, hidden, veryHidden
        "dimensions": {
            "min_row": ws.min_row,
            "max_row": ws.max_row,
            "min_col": ws.min_column,
            "max_col": ws.max_column
        },
        "row_count": ws.max_row - ws.min_row + 1 if ws.max_row else 0,
        "col_count": ws.max_column - ws.min_column + 1 if ws.max_column else 0
    }


def extract_carteira_headers(wb) -> Dict[str, List[str]]:
    """Extrai headers (row 1) da aba CARTEIRA em ambos os arquivos"""
    headers = {}

    for ws in wb.sheetnames:
        if "CARTEIRA" in ws.upper():
            try:
                sheet = wb[ws]
                row_1_values = []
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    row_1_values.append(cell.value if cell.value else f"[EMPTY-{col_idx}]")
                headers[ws] = row_1_values
            except Exception as e:
                print(f"Erro ao extrair headers de {ws}: {e}")

    return headers


def analyze_file(file_path: str) -> Dict[str, Any]:
    """Analisa um arquivo Excel e extrai todas as informações de layout"""
    print(f"\nAnalisando: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        print(f"ERRO ao carregar arquivo: {e}")
        return {"error": str(e)}

    analysis = {
        "file_path": file_path,
        "file_size_mb": Path(file_path).stat().st_size / (1024 * 1024),
        "sheets": {},
        "carteira_headers": {},
        "sheet_names": wb.sheetnames,
        "total_sheets": len(wb.sheetnames)
    }

    # Extrair headers da CARTEIRA
    analysis["carteira_headers"] = extract_carteira_headers(wb)

    # Analisar cada aba
    for sheet_name in wb.sheetnames:
        print(f"  → Processando aba: {sheet_name}")
        try:
            ws = wb[sheet_name]

            analysis["sheets"][sheet_name] = {
                "basic_info": extract_sheet_info(ws),
                "freeze_panes": extract_freeze_panes(ws),
                "auto_filter": extract_autofilter(ws),
                "column_widths": extract_column_widths(ws),
                "outline_levels": extract_outline_levels(ws),
                "merged_cells": extract_merged_cells(ws),
                "conditional_formatting": extract_conditional_formatting(ws),
                "data_validations": extract_data_validations(ws),
                "print_settings": extract_print_settings(ws)
            }
        except Exception as e:
            print(f"  ! Erro ao processar {sheet_name}: {e}")
            analysis["sheets"][sheet_name] = {"error": str(e)}

    wb.close()
    return analysis


def compare_layouts(v31_data: Dict, v13_data: Dict) -> Dict[str, Any]:
    """Compara layouts entre V31 e V13"""
    comparison = {
        "timestamp": datetime.now().isoformat(),
        "v31_sheets": v31_data["sheet_names"],
        "v13_sheets": v13_data["sheet_names"],
        "sheets_in_v31_not_in_v13": [],
        "sheets_in_v13_not_in_v31": [],
        "sheet_comparisons": {},
        "carteira_analysis": {}
    }

    # Sheets presentes em V31 mas não em V13
    comparison["sheets_in_v31_not_in_v13"] = [
        s for s in v31_data["sheet_names"]
        if s not in v13_data["sheet_names"]
    ]

    # Sheets presentes em V13 mas não em V31
    comparison["sheets_in_v13_not_in_v31"] = [
        s for s in v13_data["sheet_names"]
        if s not in v31_data["sheet_names"]
    ]

    # Comparar sheets comuns
    common_sheets = set(v31_data["sheet_names"]) & set(v13_data["sheet_names"])

    for sheet_name in sorted(common_sheets):
        v31_sheet = v31_data["sheets"].get(sheet_name, {})
        v13_sheet = v13_data["sheets"].get(sheet_name, {})

        sheet_comp = {
            "freeze_panes_match": v31_sheet.get("freeze_panes") == v13_sheet.get("freeze_panes"),
            "freeze_panes_diff": {
                "v31": v31_sheet.get("freeze_panes"),
                "v13": v13_sheet.get("freeze_panes")
            },
            "auto_filter_match": v31_sheet.get("auto_filter") == v13_sheet.get("auto_filter"),
            "auto_filter_diff": {
                "v31": v31_sheet.get("auto_filter"),
                "v13": v13_sheet.get("auto_filter")
            },
            "column_widths_match": v31_sheet.get("column_widths") == v13_sheet.get("column_widths"),
            "column_widths_diff": {
                "v31_only": {k: v for k, v in v31_sheet.get("column_widths", {}).items()
                             if k not in v13_sheet.get("column_widths", {})},
                "v13_only": {k: v for k, v in v13_sheet.get("column_widths", {}).items()
                             if k not in v31_sheet.get("column_widths", {})},
                "different_values": {
                    k: {
                        "v31": v31_sheet.get("column_widths", {}).get(k),
                        "v13": v13_sheet.get("column_widths", {}).get(k)
                    }
                    for k in set(v31_sheet.get("column_widths", {}).keys()) &
                           set(v13_sheet.get("column_widths", {}).keys())
                    if v31_sheet.get("column_widths", {}).get(k) !=
                       v13_sheet.get("column_widths", {}).get(k)
                }
            },
            "outline_levels_match": v31_sheet.get("outline_levels") == v13_sheet.get("outline_levels"),
            "outline_levels_diff": {
                "v31": v31_sheet.get("outline_levels"),
                "v13": v13_sheet.get("outline_levels")
            },
            "merged_cells_count": {
                "v31": len(v31_sheet.get("merged_cells", [])),
                "v13": len(v13_sheet.get("merged_cells", []))
            },
            "conditional_formatting_count": {
                "v31": v31_sheet.get("conditional_formatting", {}).get("total_rules", 0),
                "v13": v13_sheet.get("conditional_formatting", {}).get("total_rules", 0)
            }
        }

        comparison["sheet_comparisons"][sheet_name] = sheet_comp

    # Análise especial da CARTEIRA
    v31_carteira_headers = v31_data.get("carteira_headers", {})
    v13_carteira_headers = v13_data.get("carteira_headers", {})

    comparison["carteira_analysis"] = {
        "v31_carteira_sheets": list(v31_carteira_headers.keys()),
        "v13_carteira_sheets": list(v13_carteira_headers.keys()),
        "headers_comparison": {}
    }

    for sheet_name in set(list(v31_carteira_headers.keys()) + list(v13_carteira_headers.keys())):
        v31_headers = v31_carteira_headers.get(sheet_name, [])
        v13_headers = v13_carteira_headers.get(sheet_name, [])

        comparison["carteira_analysis"]["headers_comparison"][sheet_name] = {
            "v31_header_count": len(v31_headers),
            "v13_header_count": len(v13_headers),
            "headers_match": v31_headers == v13_headers,
            "v31_headers": v31_headers[:50] if len(v31_headers) > 50 else v31_headers,  # Primeiras 50
            "v13_headers": v13_headers[:50] if len(v13_headers) > 50 else v13_headers,
            "full_v31_count": len(v31_headers),
            "full_v13_count": len(v13_headers)
        }

    return comparison


def format_summary_text(v31_data: Dict, v13_data: Dict, comparison: Dict) -> str:
    """Gera resumo legível em texto"""
    summary = []
    summary.append("=" * 100)
    summary.append("ANÁLISE COMPARATIVA DE LAYOUT: V31 vs V13")
    summary.append("=" * 100)
    summary.append(f"Data: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    # Info básica
    summary.append("INFORMAÇÕES BÁSICAS")
    summary.append("-" * 100)
    summary.append(f"V31 (Referência):")
    summary.append(f"  • Caminho: {v31_data['file_path']}")
    summary.append(f"  • Tamanho: {v31_data['file_size_mb']:.2f} MB")
    summary.append(f"  • Total de abas: {v31_data['total_sheets']}")
    summary.append(f"  • Abas: {', '.join(v31_data['sheet_names'][:10])}")
    if len(v31_data['sheet_names']) > 10:
        summary.append(f"             {', '.join(v31_data['sheet_names'][10:])}")

    summary.append(f"\nV13 (Oficial):")
    summary.append(f"  • Caminho: {v13_data['file_path']}")
    summary.append(f"  • Tamanho: {v13_data['file_size_mb']:.2f} MB")
    summary.append(f"  • Total de abas: {v13_data['total_sheets']}")
    summary.append(f"  • Abas: {', '.join(v13_data['sheet_names'][:10])}")
    if len(v13_data['sheet_names']) > 10:
        summary.append(f"             {', '.join(v13_data['sheet_names'][10:])}")

    # Diferenças de estrutura
    summary.append(f"\n\nDIFERENÇAS DE ESTRUTURA")
    summary.append("-" * 100)
    if comparison["sheets_in_v31_not_in_v13"]:
        summary.append(f"Abas em V31 MAS NÃO em V13 ({len(comparison['sheets_in_v31_not_in_v13'])}): ")
        for sheet in comparison["sheets_in_v31_not_in_v13"]:
            summary.append(f"  → {sheet}")
    else:
        summary.append("✓ Nenhuma aba presente apenas em V31")

    if comparison["sheets_in_v13_not_in_v31"]:
        summary.append(f"\nAbas em V13 MAS NÃO em V31 ({len(comparison['sheets_in_v13_not_in_v31'])}): ")
        for sheet in comparison["sheets_in_v13_not_in_v31"]:
            summary.append(f"  → {sheet}")
    else:
        summary.append("\n✓ Nenhuma aba presente apenas em V13")

    # Análise de FREEZE PANES
    summary.append(f"\n\nANÁLISE DE FREEZE PANES")
    summary.append("-" * 100)
    freeze_diffs = []
    for sheet_name, comp in comparison["sheet_comparisons"].items():
        if not comp.get("freeze_panes_match"):
            v31_freeze = comp["freeze_panes_diff"]["v31"].get("freeze_cell")
            v13_freeze = comp["freeze_panes_diff"]["v13"].get("freeze_cell")
            freeze_diffs.append({
                "sheet": sheet_name,
                "v31": v31_freeze,
                "v13": v13_freeze
            })

    if freeze_diffs:
        summary.append(f"Encontradas {len(freeze_diffs)} diferenças de freeze_panes:\n")
        for diff in freeze_diffs:
            summary.append(f"  • {diff['sheet']}")
            summary.append(f"    V31: {diff['v31']}")
            summary.append(f"    V13: {diff['v13']}")
    else:
        summary.append("✓ Freeze panes idênticos em todas as abas comuns")

    # Análise de AUTO_FILTER
    summary.append(f"\n\nANÁLISE DE AUTO_FILTER")
    summary.append("-" * 100)
    filter_diffs = []
    for sheet_name, comp in comparison["sheet_comparisons"].items():
        if not comp.get("auto_filter_match"):
            v31_filter = comp["auto_filter_diff"]["v31"].get("filter_range")
            v13_filter = comp["auto_filter_diff"]["v13"].get("filter_range")
            filter_diffs.append({
                "sheet": sheet_name,
                "v31": v31_filter,
                "v13": v13_filter
            })

    if filter_diffs:
        summary.append(f"Encontradas {len(filter_diffs)} diferenças de auto_filter:\n")
        for diff in filter_diffs:
            summary.append(f"  • {diff['sheet']}")
            summary.append(f"    V31: {diff['v31']}")
            summary.append(f"    V13: {diff['v13']}")
    else:
        summary.append("✓ Auto filters idênticos em todas as abas comuns")

    # Análise de COLUMN WIDTHS
    summary.append(f"\n\nANÁLISE DE COLUMN WIDTHS")
    summary.append("-" * 100)
    width_diffs_count = 0
    for sheet_name, comp in comparison["sheet_comparisons"].items():
        if not comp.get("column_widths_match"):
            width_diffs_count += 1
            diff_info = comp["column_widths_diff"]

            summary.append(f"\n  ► {sheet_name}:")

            if diff_info.get("v31_only"):
                summary.append(f"    Colunas em V31 mas não em V13 ({len(diff_info['v31_only'])}): ")
                for col, width in sorted(diff_info["v31_only"].items())[:10]:
                    summary.append(f"      {col}: {width}")
                if len(diff_info["v31_only"]) > 10:
                    summary.append(f"      ... e mais {len(diff_info['v31_only']) - 10}")

            if diff_info.get("v13_only"):
                summary.append(f"    Colunas em V13 mas não em V31 ({len(diff_info['v13_only'])}): ")
                for col, width in sorted(diff_info["v13_only"].items())[:10]:
                    summary.append(f"      {col}: {width}")
                if len(diff_info["v13_only"]) > 10:
                    summary.append(f"      ... e mais {len(diff_info['v13_only']) - 10}")

            if diff_info.get("different_values"):
                summary.append(f"    Colunas com widths diferentes ({len(diff_info['different_values'])}): ")
                for col, vals in sorted(diff_info["different_values"].items())[:10]:
                    summary.append(f"      {col}: V31={vals['v31']} vs V13={vals['v13']}")
                if len(diff_info["different_values"]) > 10:
                    summary.append(f"      ... e mais {len(diff_info['different_values']) - 10}")

    if width_diffs_count == 0:
        summary.append("✓ Todas as column widths são idênticas")
    else:
        summary.append(f"\n✗ {width_diffs_count} abas com diferenças de column widths")

    # Análise de OUTLINE LEVELS (Grouping)
    summary.append(f"\n\nANÁLISE DE OUTLINE LEVELS (GROUPING)")
    summary.append("-" * 100)
    outline_diffs_count = 0
    for sheet_name, comp in comparison["sheet_comparisons"].items():
        if not comp.get("outline_levels_match"):
            outline_diffs_count += 1
            v31_outline = comp["outline_levels_diff"]["v31"]
            v13_outline = comp["outline_levels_diff"]["v13"]

            summary.append(f"\n  ► {sheet_name}:")
            summary.append(f"    V31 - Colunas max_level: {v31_outline['columns']['max_level']}, Grupos: {len(v31_outline['columns']['groups'])}")
            summary.append(f"    V13 - Colunas max_level: {v13_outline['columns']['max_level']}, Grupos: {len(v13_outline['columns']['groups'])}")
            summary.append(f"    V31 - Linhas max_level: {v31_outline['rows']['max_level']}, Grupos: {len(v31_outline['rows']['groups'])}")
            summary.append(f"    V13 - Linhas max_level: {v13_outline['rows']['max_level']}, Grupos: {len(v13_outline['rows']['groups'])}")

            # Mostrar grupos ocultos
            if v31_outline['columns']['groups']:
                hidden_cols = [g for g in v31_outline['columns']['groups'] if g.get('hidden') or g.get('collapsed')]
                if hidden_cols:
                    summary.append(f"    Colunas ocultas em V31 (via grouping): {[g['column'] for g in hidden_cols]}")

            if v13_outline['columns']['groups']:
                hidden_cols = [g for g in v13_outline['columns']['groups'] if g.get('hidden') or g.get('collapsed')]
                if hidden_cols:
                    summary.append(f"    Colunas ocultas em V13 (via grouping): {[g['column'] for g in hidden_cols]}")

    if outline_diffs_count == 0:
        summary.append("✓ Todos os outline levels (grouping) são idênticos")
    else:
        summary.append(f"\n✗ {outline_diffs_count} abas com diferenças de outline levels")

    # Análise da CARTEIRA especial
    summary.append(f"\n\nANÁLISE ESPECIAL - CARTEIRA")
    summary.append("-" * 100)
    carteira_analysis = comparison["carteira_analysis"]

    summary.append(f"Abas CARTEIRA em V31: {carteira_analysis['v31_carteira_sheets']}")
    summary.append(f"Abas CARTEIRA em V13: {carteira_analysis['v13_carteira_sheets']}")

    for sheet_name, headers_comp in carteira_analysis.get("headers_comparison", {}).items():
        summary.append(f"\n  ► {sheet_name}:")
        summary.append(f"    V31: {headers_comp['v31_header_count']} colunas")
        summary.append(f"    V13: {headers_comp['v13_header_count']} colunas")

        if headers_comp.get('headers_match'):
            summary.append(f"    Status: ✓ Headers idênticos")
        else:
            summary.append(f"    Status: ✗ Headers DIFERENTES")

            # Mostrar primeiros headers
            if headers_comp['v31_headers']:
                summary.append(f"    V31 primeiros headers: {headers_comp['v31_headers'][:5]}")
            if headers_comp['v13_headers']:
                summary.append(f"    V13 primeiros headers: {headers_comp['v13_headers'][:5]}")

    # Resumo de problemas encontrados
    summary.append(f"\n\nRESUMO DE PROBLEMAS ENCONTRADOS")
    summary.append("-" * 100)

    issues = []

    if comparison["sheets_in_v31_not_in_v13"]:
        issues.append(f"• {len(comparison['sheets_in_v31_not_in_v13'])} abas em V31 não existem em V13 (ex: {comparison['sheets_in_v31_not_in_v13'][0]})")

    if comparison["sheets_in_v13_not_in_v31"]:
        issues.append(f"• {len(comparison['sheets_in_v13_not_in_v31'])} abas em V13 não existem em V31")

    freeze_diffs = [comp for comp in comparison["sheet_comparisons"].values() if not comp.get("freeze_panes_match")]
    if freeze_diffs:
        issues.append(f"• {len(freeze_diffs)} abas com freeze_panes diferentes")

    filter_diffs = [comp for comp in comparison["sheet_comparisons"].values() if not comp.get("auto_filter_match")]
    if filter_diffs:
        issues.append(f"• {len(filter_diffs)} abas com auto_filter diferentes")

    width_diffs = [comp for comp in comparison["sheet_comparisons"].values() if not comp.get("column_widths_match")]
    if width_diffs:
        issues.append(f"• {len(width_diffs)} abas com column_widths diferentes")

    outline_diffs = [comp for comp in comparison["sheet_comparisons"].values() if not comp.get("outline_levels_match")]
    if outline_diffs:
        issues.append(f"• {len(outline_diffs)} abas com outline_levels diferentes")

    if issues:
        for issue in issues:
            summary.append(issue)
    else:
        summary.append("✓ Nenhum problema encontrado - Layouts são idênticos")

    summary.append(f"\n\n{'=' * 100}")
    summary.append(f"FIM DO RELATÓRIO")
    summary.append(f"{'=' * 100}\n")

    return "\n".join(summary)


def main():
    """Executa análise completa"""
    print("\n" + "=" * 100)
    print("ANÁLISE COMPARATIVA DE LAYOUT: V31 vs V13")
    print("=" * 100)

    # Analisar ambos os arquivos
    print("\n[1/3] Analisando V31...")
    v31_data = analyze_file(V31_PATH)

    print("\n[2/3] Analisando V13...")
    v13_data = analyze_file(V13_PATH)

    # Comparar
    print("\n[3/3] Comparando layouts...")
    comparison = compare_layouts(v31_data, v13_data)

    # Salvar JSON completo
    print(f"\nSalvando análise completa em JSON...")
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump({
            "v31": v31_data,
            "v13": v13_data,
            "comparison": comparison
        }, f, indent=2, ensure_ascii=False, default=str)
    print(f"✓ Salvo em: {OUTPUT_JSON}")

    # Gerar e salvar resumo
    print(f"\nGerando resumo legível...")
    summary_text = format_summary_text(v31_data, v13_data, comparison)

    with open(OUTPUT_SUMMARY, 'w', encoding='utf-8') as f:
        f.write(summary_text)
    print(f"✓ Salvo em: {OUTPUT_SUMMARY}")

    # Imprimir resumo no console
    print("\n" + summary_text)

    print("\n[SUCESSO] Análise de layout concluída!")
    return 0


if __name__ == "__main__":
    sys.exit(main())

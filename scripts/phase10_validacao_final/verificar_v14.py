#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Verificação rápida do V14 FINAL — confirma que layout foi aplicado corretamente"""

import openpyxl
from pathlib import Path

V14_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V14_FINAL.xlsx"

def main():
    print("=" * 80)
    print("VERIFICAÇÃO V14 FINAL")
    print("=" * 80)

    wb = openpyxl.load_workbook(V14_PATH, data_only=False)

    total_formulas = 0
    issues = []

    for name in wb.sheetnames:
        ws = wb[name]

        # Count formulas
        sheet_formulas = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    sheet_formulas += 1
        total_formulas += sheet_formulas

        # Check layout properties
        freeze = ws.freeze_panes
        auto_f = ws.auto_filter.ref if ws.auto_filter else None

        # Count column widths
        widths_set = sum(1 for c in ws.column_dimensions if ws.column_dimensions[c].width and ws.column_dimensions[c].width != 8.43)

        # Count outline levels
        outlines = sum(1 for c in ws.column_dimensions if ws.column_dimensions[c].outline_level and ws.column_dimensions[c].outline_level > 0)

        status = "OK"

        # Validate expected freeze panes
        expected = {
            "CARTEIRA": "AR4",
            "DRAFT 1": "AN4",
            "DRAFT 2": "A2",
            "LOG": "A3",
            "PROJEÇÃO ": "C4",
            "REGRAS": "A41",
        }

        if name in expected:
            if str(freeze) != expected[name]:
                status = f"WARN: freeze={freeze}, expected={expected[name]}"
                issues.append(f"{name}: freeze_panes mismatch")

        print(f"  {name:20s} | {ws.max_row:>6} rows | {sheet_formulas:>7} fórmulas | freeze={str(freeze):>6} | filter={str(auto_f)[:30]:>30} | {widths_set:>3} widths | {outlines:>3} outlines | {status}")

    print(f"\n{'=' * 80}")
    print(f"TOTAL FÓRMULAS: {total_formulas:,}")
    print(f"TOTAL ABAS: {len(wb.sheetnames)}")
    print(f"TAMANHO: {Path(V14_PATH).stat().st_size / (1024*1024):.2f} MB")

    if issues:
        print(f"\nPROBLEMAS ({len(issues)}):")
        for i in issues:
            print(f"  ✗ {i}")
    else:
        print(f"\n✓ TODAS AS VERIFICAÇÕES PASSARAM")

    wb.close()
    return 0 if not issues else 1

if __name__ == "__main__":
    import sys
    sys.exit(main())

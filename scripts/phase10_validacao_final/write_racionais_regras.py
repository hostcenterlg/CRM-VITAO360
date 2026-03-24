#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ESCREVER BLOCOS RACIONAIS nas colunas H-K ao lado das seções 1-17 das REGRAS
SÓ VALORES ESTÁTICOS. ZERO FÓRMULAS.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

V19 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V19_FINAL.xlsx"

RACIONAL_TITLE_FONT = Font(name="Calibri", size=10, bold=True, color="1F4E79")
RACIONAL_TITLE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
RACIONAL_FONT = Font(name="Calibri", size=9, italic=True, color="4472C4")
RACIONAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")

# O ÍNDICE ocupa cols I-K (9-11) nas rows 1-20.
# Os racionais vão na coluna H (8) ao lado de cada seção.
# Para não conflitar com o ÍNDICE, vamos usar cols 8 (H) apenas,
# ou cols 13-14 (M-N) que estão livres.
# Vou usar col 8 (H) que fica entre o conteúdo (A-G) e o índice (I-K).

# Mapeamento: row onde começa cada seção → texto racional
RACIONAIS = {
    # Seção 1: RESULTADO (row 4)
    4: ("💡 RACIONAL — RESULTADO",
        "Os 14 resultados cobrem 100% dos desfechos possíveis de um contato comercial. "
        "Cada resultado gera automaticamente: (1) FOLLOW-UP em dias, (2) GRUPO DASH para "
        "classificação gerencial, (3) TEMPERATURA e FASE via Motor de Regras (Seção 17). "
        "O consultor escolhe 1 resultado → o sistema calcula TODO o resto. "
        "Isso elimina erro humano e garante consistência nos 21.000+ atendimentos do DRAFT 2."),

    # Seção 2: TIPO DO CONTATO (row 22)
    22: ("💡 RACIONAL — TIPO DO CONTATO",
         "Os 7 tipos classificam a INTENÇÃO do contato, não o resultado. "
         "Prospecção ≠ Negociação ≠ Follow-up. Essa distinção alimenta o DASH e permite "
         "medir: quantos contatos são prospecção vs manutenção? Se 90% é follow-up, "
         "a equipe está REAGINDO, não PROSPECTANDO. A meta ideal: 30% prospecção, "
         "40% manutenção ativos, 20% follow-up, 10% pós-venda."),

    # Seção 3: MOTIVO (row 32)
    32: ("💡 RACIONAL — MOTIVO",
         "Os 22 motivos identificam POR QUE o cliente não comprou e QUEM deve resolver. "
         "'Produto não vendeu' → FÁBRICA. 'Atraso entrega' → LOGÍSTICA. "
         "'Problema financeiro' → FINANCEIRO. Isso tira do consultor a responsabilidade "
         "por problemas que NÃO SÃO DELE e permite ao comitê cobrar a ÁREA CORRETA. "
         "Sem isso, o consultor é culpado por tudo."),

    # Seção 4: SITUAÇÃO (row 57)
    57: ("💡 RACIONAL — SITUAÇÃO",
         "A SITUAÇÃO é calculada automaticamente por DIAS SEM COMPRA. "
         "Não é opinião do consultor — é FATO. ≤50d = ATIVO. 51-60 = EM RISCO. "
         "61-90 = INATIVO RECENTE. >90 = INATIVO ANTIGO. Nunca comprou = PROSPECT. "
         "Essa classificação alimenta o SINALEIRO, o SCORE e a PIRÂMIDE DE PRIORIDADE. "
         "É a base de toda a inteligência da agenda."),

    # Seção 5: FASE (row 67)
    67: ("💡 RACIONAL — FASE",
         "A FASE indica ONDE o cliente está no funil comercial. "
         "PÓS-VENDA → CS → RECOMPRA é o ciclo ideal. SALVAMENTO e RECUPERAÇÃO "
         "são fases de RESGATE para clientes que estão escapando. "
         "PROSPECÇÃO é a entrada. A FASE muda automaticamente conforme "
         "o RESULTADO do atendimento via Motor de Regras (Seção 17)."),

    # Seção 6: TIPO CLIENTE (row 79)
    79: ("💡 RACIONAL — TIPO CLIENTE",
         "O TIPO classifica a MATURIDADE do relacionamento. "
         "NOVO (1ª compra) precisa de atenção especial — 60% não recompra. "
         "EM DESENVOLVIMENTO (2-3 compras) é o momento de FIDELIZAR. "
         "FIDELIZADO é o objetivo final. PROSPECT nunca comprou. "
         "Essa classificação afeta o SCORE (Seção 16): fidelizado com 60d "
         "sem compra é MAIS URGENTE que prospect novo."),

    # Seção 7: CONSULTOR (row 88)
    88: ("💡 RACIONAL — CONSULTOR",
         "4 consultores + 1 supervisor. Cada um tem TERRITÓRIO por UF: "
         "MANU = Sul (SC/PR/RS). LARISSA = resto do Brasil. "
         "JULIO = Redes/Franquias. DAIANE = Operacional + Redes menores. "
         "O CONSULTOR determina qual AGENDA o cliente aparece. "
         "Prospects sem consultor serão distribuídos automaticamente por UF."),

    # Seção 8: TENTATIVA (row 96)
    96: ("💡 RACIONAL — TENTATIVA",
         "O protocolo T1→T4→NUTRIÇÃO→RESET garante que NENHUM CLIENTE FICA SEM CONTATO "
         "mas também que ninguém PERDE TEMPO com quem não responde. "
         "T1 = WhatsApp (mais suave). T2 = Ligação (mais assertivo). "
         "T3 e T4 = últimas chances. NUTRIÇÃO = modo passivo (email/campanhas). "
         "RESET = quando o cliente DÁ SINAL (abre email, acessa B2B, responde). "
         "Isso alimenta a REGRA DO IGNORADOR (Seção 18.3)."),

    # Seção 9: SINALEIRO (row 105)
    105: ("💡 RACIONAL — SINALEIRO",
          "O SINALEIRO é o 'semáforo de saúde' do cliente. "
          "🟢 = saudável (dias ≤ ciclo). 🟡 = atenção (dias ≤ ciclo+30). "
          "🔴 = perigo (dias > ciclo+30). 🟣 = prospect (nunca comprou). "
          "É visual e instantâneo — o consultor olha a CARTEIRA e sabe "
          "em 1 segundo quem precisa de atenção. A AGENDA ordena por sinaleiro: "
          "🔴 sempre antes de 🟡, 🟡 antes de 🟢."),

    # Seção 10: LISTAS SIMPLES (row 112)
    112: ("💡 RACIONAL — LISTAS SIMPLES",
          "Dropdowns auxiliares para padronizar campos. SIM/NÃO, CURVA ABC, "
          "GRUPO DASH, TEMPERATURA. Sem padronização, cada consultor escreve "
          "de um jeito ('sim', 'Sim', 'SIM', 's', 'ok') e as fórmulas quebram. "
          "Dropdown = consistência = fórmulas funcionam = relatórios corretos."),

    # Seção 11: TIPO AÇÃO (row 132)
    132: ("💡 RACIONAL — TIPO AÇÃO",
          "Classifica a NATUREZA da atividade: VENDA, PRÉ-VENDA, PÓS-VENDA, "
          "RESOLUÇÃO, PROSPECÇÃO. Isso permite medir o MIX de atividades: "
          "se 80% é pré-venda e 5% é prospecção, a carteira vai encolher. "
          "O DASH mostra esse mix em tempo real para o supervisor intervir."),

    # Seção 12: TIPO PROBLEMA (row 141)
    141: ("💡 RACIONAL — TIPO PROBLEMA",
          "8 categorias de RNC (Registro de Não Conformidade) com ÁREA RESPONSÁVEL. "
          "Quando o consultor registra um problema, o sistema já indica QUEM resolver. "
          "Transportadora → LOGÍSTICA. Produto avariado → FÁBRICA. Cobrança → FINANCEIRO. "
          "Isso acelera resolução e permite medir: qual área gera MAIS problemas? "
          "Comitê usa esses dados para cobrar melhoria."),

    # Seção 13: AÇÃO FUTURA (row 152)
    152: ("💡 RACIONAL — AÇÃO FUTURA",
          "18 ações + 25 tarefas operacionais = 43 opções padronizadas. "
          "O consultor não inventa — escolhe da lista. Isso garante que o FOLLOW-UP "
          "é CLARO e MENSURÁVEL. 'Confirmar orçamento enviado' é diferente de "
          "'Fechar negociação'. A próxima ação aparece na AGENDA do dia seguinte "
          "conforme o follow-up calculado pelo RESULTADO (Seção 1)."),

    # Seção 15: SINALEIRO META (row 202)
    202: ("💡 RACIONAL — SINALEIRO META",
          "4 faixas de atingimento de meta: 🟢 (≥100%), 🟡 (70-99%), "
          "🔴 (40-69%), ⚫ (0-39%). Alimenta diretamente a REGRA DO EQUILÍBRIO "
          "DE META (Seção 18.4): consultor com meta ⚫ recebe 50% de prospects "
          "na agenda. Consultor 🟢 recebe 10%. É adaptativo — a IA ajusta "
          "a composição da agenda conforme o atingimento individual."),

    # Seção 16: SCORE RANKING (row 209)
    209: ("💡 RACIONAL — SCORE RANKING",
          "6 fatores ponderados geram 1 nota de 0-100 por cliente. "
          "URGÊNCIA (30%) é o mais pesado porque timing = conversão. "
          "VALOR (25%) porque Curva A vale mais que Curva C. "
          "FOLLOW-UP (20%) porque compromisso é sagrado. "
          "SINAL (15%) porque quem acessou B2B está QUENTE. "
          "TENTATIVA e SITUAÇÃO (5% cada) são refinamentos. "
          "Detalhamento completo na Seção 18.9."),

    # Seção 17: MOTOR DE REGRAS (row 218)
    218: ("💡 RACIONAL — MOTOR DE REGRAS",
          "63 combinações SITUAÇÃO × RESULTADO → 5 campos automáticos. "
          "É o CÉREBRO do CRM. Quando o consultor registra 'ATIVO + VENDA/PEDIDO', "
          "o motor define: ESTÁGIO=PÓS-VENDA, FASE=PÓS-VENDA, TIPO=PÓS-VENDA, "
          "AÇÃO=CONFIRMAR FATURAMENTO, TEMP=🔥QUENTE. "
          "Sem motor, o consultor preenche 5 campos manualmente (e erra). "
          "Com motor, ele preenche 1 (RESULTADO) e o resto é automático."),
}


def main():
    t = datetime.now()
    print("=" * 100)
    print("ESCREVENDO BLOCOS RACIONAIS — Seções 1-17")
    print("=" * 100)

    print("\n[1] Abrindo V19...", flush=True)
    wb = openpyxl.load_workbook(V19)
    ws = wb["REGRAS"]

    count = 0
    for row, (titulo, texto) in sorted(RACIONAIS.items()):
        # Título do racional na coluna H (8) da row da seção
        cell_title = ws.cell(row=row, column=8, value=titulo)
        cell_title.font = RACIONAL_TITLE_FONT
        cell_title.fill = RACIONAL_TITLE_FILL
        cell_title.alignment = WRAP

        # Texto do racional na row+1, coluna H
        cell_text = ws.cell(row=row+1, column=8, value=texto)
        cell_text.font = RACIONAL_FONT
        cell_text.fill = RACIONAL_FILL
        cell_text.alignment = WRAP

        count += 1
        print(f"  Row {row}: {titulo}", flush=True)

    # Ajustar largura da coluna H para os racionais
    ws.column_dimensions['H'].width = 55

    print(f"\n[2] Salvando...", flush=True)
    wb.save(V19)
    wb.close()

    elapsed = (datetime.now() - t).total_seconds()
    print(f"\n{'='*100}")
    print(f"[SUCESSO] {count} blocos racionais escritos na coluna H!")
    print(f"  Seções cobertas: 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17")
    print(f"  (Seção 14 = TAREFA/DEMANDA — lista operacional, sem racional especial)")
    print(f"  Seção 18 já tem racionais embutidos em cada bloco")
    print(f"  Tempo: {elapsed:.1f}s")
    print(f"{'='*100}")


if __name__ == "__main__":
    main()

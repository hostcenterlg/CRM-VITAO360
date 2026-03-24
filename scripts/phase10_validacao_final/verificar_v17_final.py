#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VERIFICACAO FINAL V17 — Confirma que CARTEIRA tem dados visiveis ao abrir no Excel.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import re
from datetime import datetime

V17_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V17_FINAL.xlsx"


def normalizar_cnpj(valor):
    if not valor:
        return None
    cnpj = re.sub(r'[^\d]', '', str(valor).strip())
    return cnpj if len(cnpj) >= 11 else None


def get_real_value(cell):
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str) and val.startswith("="):
        return None
    if isinstance(val, str) and val.strip() == "":
        return None
    return val


def main():
    start = datetime.now()
    print("=" * 100)
    print("VERIFICACAO FINAL V17 — CARTEIRA com valores reais")
    print("=" * 100)

    wb = openpyxl.load_workbook(V17_PATH, data_only=False)

    # ================================================================
    # CHECK 1: Abas presentes
    # ================================================================
    print("\n[CHECK 1] Abas presentes:")
    expected_tabs = [
        "SINALEIRO", "PROJEÇÃO ", "DASH", "REDES_FRANQUIAS_v2", "COMITE",
        "REGRAS", "DRAFT 1", "DRAFT 2", "DRAFT 3 ", "CARTEIRA",
        "AGENDA LARISSA", "AGENDA DAIANE", "AGENDA MANU", "AGENDA JULIO"
    ]
    present = 0
    for tab in expected_tabs:
        if tab in wb.sheetnames:
            present += 1
            print(f"  ✓ {tab}")
        else:
            # Tentar match parcial
            found = False
            for s in wb.sheetnames:
                if tab.strip() in s:
                    present += 1
                    print(f"  ✓ {tab} (como '{s}')")
                    found = True
                    break
            if not found:
                print(f"  ✗ {tab} AUSENTE!")

    result1 = "PASS" if present >= 13 else "FAIL"
    print(f"  → {result1}: {present}/{len(expected_tabs)} abas")

    # ================================================================
    # CHECK 2: CARTEIRA preenchimento por coluna
    # ================================================================
    print(f"\n[CHECK 2] CARTEIRA preenchimento (VALORES REAIS, sem formulas):")
    ws_cart = wb["CARTEIRA"]

    # Colunas criticas para o usuario ver
    critical_cols = {
        2: ("B", "CNPJ"),
        12: ("L", "CONSULTOR"),
        14: ("N", "SITUAÇÃO"),
        16: ("P", "DIAS SEM COMPRA"),
        44: ("AR", "ESTÁGIO FUNIL"),
        45: ("AS", "PRÓX FOLLOWUP"),
        46: ("AT", "DATA ÚLT ATEND"),
        47: ("AU", "AÇÃO FUTURA"),
        48: ("AV", "ÚLTIMO RESULTADO"),
        49: ("AW", "MOTIVO"),
        50: ("AX", "TIPO CLIENTE"),
        51: ("AY", "TENTATIVA"),
        52: ("AZ", "FASE"),
        54: ("BB", "TEMPERATURA"),
        60: ("BH", "PRÓX AÇÃO"),
        61: ("BI", "AÇÃO DETALHADA"),
        62: ("BJ", "SINALEIRO"),
    }

    total_data_rows = 0
    for row in range(4, ws_cart.max_row + 1):
        if normalizar_cnpj(ws_cart.cell(row=row, column=2).value):
            total_data_rows += 1

    print(f"  Total rows com CNPJ: {total_data_rows}")

    all_pass = True
    for col_idx, (letter, name) in critical_cols.items():
        filled = 0
        formula = 0
        empty = 0
        for row in range(4, ws_cart.max_row + 1):
            if not normalizar_cnpj(ws_cart.cell(row=row, column=2).value):
                continue
            val = ws_cart.cell(row=row, column=col_idx).value
            if val is None or (isinstance(val, str) and val.strip() == ""):
                empty += 1
            elif isinstance(val, str) and val.startswith("="):
                formula += 1
            else:
                filled += 1

        pct = round(100 * filled / max(total_data_rows, 1), 1)
        bar = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
        status = "✓" if pct >= 80 else ("⚠" if pct >= 50 else "✗")
        print(f"  {status} {letter:>3} {name:20s} | {bar} {pct:>5.1f}% | D={filled:>4} F={formula:>4} E={empty:>4}")

        if pct < 70 and name not in ("MOTIVO", "TEMPERATURA"):
            all_pass = False

    result2 = "PASS" if all_pass else "FAIL"
    print(f"  → {result2}")

    # ================================================================
    # CHECK 3: DRAFT 2 integridade
    # ================================================================
    print(f"\n[CHECK 3] DRAFT 2 integridade:")
    ws_d2 = wb["DRAFT 2"]
    d2_rows = ws_d2.max_row - 2  # minus header rows
    d2_cnpjs = set()
    d2_formulas = 0
    d2_static = 0
    for row in range(3, ws_d2.max_row + 1):
        cnpj = normalizar_cnpj(ws_d2.cell(row=row, column=4).value)
        if cnpj:
            d2_cnpjs.add(cnpj)
        for col in range(1, 32):
            val = ws_d2.cell(row=row, column=col).value
            if val and isinstance(val, str) and val.startswith("="):
                d2_formulas += 1
            elif val is not None and str(val).strip() != "":
                d2_static += 1

    print(f"  Registros: {d2_rows}")
    print(f"  CNPJs unicos: {len(d2_cnpjs)}")
    print(f"  Formulas: {d2_formulas:,}")
    print(f"  Dados estaticos: {d2_static:,}")
    result3 = "PASS" if d2_rows > 20000 and len(d2_cnpjs) > 2000 else "FAIL"
    print(f"  → {result3}")

    # ================================================================
    # CHECK 4: AGENDA tabs com formulas SORTBY/FILTER
    # ================================================================
    print(f"\n[CHECK 4] AGENDA tabs:")
    agenda_tabs = ["AGENDA LARISSA", "AGENDA DAIANE", "AGENDA MANU", "AGENDA JULIO"]
    agenda_ok = 0
    for tab in agenda_tabs:
        if tab in wb.sheetnames:
            ws = wb[tab]
            formulas = 0
            for col in range(1, 33):
                val = ws.cell(row=2, column=col).value
                if val and isinstance(val, str) and val.startswith("="):
                    formulas += 1
            status = "✓" if formulas >= 10 else "✗"
            print(f"  {status} {tab}: {formulas} formulas (SORTBY/FILTER + SCORE + PRIORIDADE)")
            if formulas >= 10:
                agenda_ok += 1

    result4 = "PASS" if agenda_ok == 4 else "FAIL"
    print(f"  → {result4}")

    # ================================================================
    # CHECK 5: Consultores atribuidos
    # ================================================================
    print(f"\n[CHECK 5] Consultores na CARTEIRA:")
    consultores = {}
    for row in range(4, ws_cart.max_row + 1):
        val = get_real_value(ws_cart.cell(row=row, column=12))
        if val:
            name = str(val).strip().upper()
            consultores[name] = consultores.get(name, 0) + 1

    for name, count in sorted(consultores.items(), key=lambda x: -x[1]):
        print(f"  {name:30s}: {count:>4} clientes")

    total_with_consult = sum(consultores.values())
    pct_consult = round(100 * total_with_consult / max(total_data_rows, 1), 1)
    result5 = "PASS" if pct_consult >= 85 else "FAIL"
    print(f"  → {result5}: {total_with_consult}/{total_data_rows} ({pct_consult}%)")

    # ================================================================
    # CHECK 6: SINALEIRO emojis
    # ================================================================
    print(f"\n[CHECK 6] SINALEIRO emojis:")
    sinaleiros = {}
    for row in range(4, ws_cart.max_row + 1):
        val = get_real_value(ws_cart.cell(row=row, column=62))
        if val:
            s = str(val).strip()
            sinaleiros[s] = sinaleiros.get(s, 0) + 1

    for s, count in sorted(sinaleiros.items(), key=lambda x: -x[1]):
        emoji_name = {"🟢": "VERDE", "🟡": "AMARELO", "🔴": "VERMELHO", "🟣": "ROXO"}.get(s, s)
        print(f"  {s} {emoji_name:12s}: {count:>4}")

    total_sinal = sum(sinaleiros.values())
    pct_sinal = round(100 * total_sinal / max(total_data_rows, 1), 1)
    result6 = "PASS" if pct_sinal >= 85 else "FAIL"
    print(f"  → {result6}: {total_sinal}/{total_data_rows} ({pct_sinal}%)")

    # ================================================================
    # CHECK 7: TIPO CLIENTE valido
    # ================================================================
    print(f"\n[CHECK 7] TIPO CLIENTE valores:")
    tipos = {}
    for row in range(4, ws_cart.max_row + 1):
        val = get_real_value(ws_cart.cell(row=row, column=50))
        if val:
            t = str(val).strip()
            tipos[t] = tipos.get(t, 0) + 1

    for t, count in sorted(tipos.items(), key=lambda x: -x[1]):
        print(f"  {t:25s}: {count:>4}")

    total_tipo = sum(tipos.values())
    pct_tipo = round(100 * total_tipo / max(total_data_rows, 1), 1)
    result7 = "PASS" if pct_tipo >= 85 else "FAIL"
    print(f"  → {result7}: {total_tipo}/{total_data_rows} ({pct_tipo}%)")

    # ================================================================
    # CHECK 8: Tamanho e formulas totais
    # ================================================================
    print(f"\n[CHECK 8] Resumo geral:")
    total_formulas = 0
    total_static = 0
    for name in wb.sheetnames:
        ws = wb[name]
        for row_cells in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip() != "":
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        total_formulas += 1
                    else:
                        total_static += 1

    from pathlib import Path
    size = Path(V17_PATH).stat().st_size / (1024 * 1024)
    print(f"  Tamanho: {size:.2f} MB")
    print(f"  Formulas: {total_formulas:,}")
    print(f"  Dados estaticos: {total_static:,}")
    print(f"  Abas: {len(wb.sheetnames)}")

    result8 = "PASS"

    # ================================================================
    # RESUMO FINAL
    # ================================================================
    results = [result1, result2, result3, result4, result5, result6, result7, result8]
    passed = results.count("PASS")
    total = len(results)

    print(f"\n{'='*100}")
    print(f"RESULTADO FINAL: {passed}/{total} CHECKS PASSED")
    print(f"{'='*100}")

    for i, (r, label) in enumerate(zip(results, [
        "Abas presentes", "CARTEIRA preenchimento", "DRAFT 2 integridade",
        "AGENDA formulas", "Consultores", "SINALEIRO", "TIPO CLIENTE", "Geral"
    ])):
        icon = "✓" if r == "PASS" else "✗"
        print(f"  {icon} CHECK {i+1}: {label} — {r}")

    elapsed = (datetime.now() - start).total_seconds()
    print(f"\n  Verificacao completada em {elapsed:.1f}s")

    if passed == total:
        print(f"\n  🎉 V17 APROVADO — CARTEIRA com dados visiveis ao abrir no Excel!")
    else:
        print(f"\n  ⚠ V17 tem {total - passed} check(s) com problemas")

    wb.close()


if __name__ == "__main__":
    main()

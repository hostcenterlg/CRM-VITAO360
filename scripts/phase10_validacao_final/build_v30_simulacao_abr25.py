#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CRM VITAO360 — V30: Simulação Realista ABR/25
==============================================================
Base: V29 (JAN+FEV+MAR simulados)
3 consultores: DAIANE (veterana) + HELDER + MANU (2o mês, mais produtivos)
Feriados: 18 Abr (Sexta-Santa) + 21 Abr (Tiradentes) → 20 dias úteis
64 vendas ABR da CARTEIRA (col 29)

Volumes estimados:
  DAIANE: ~18/dia (veterana, ritmo sólido)
  HELDER: ~14/dia (2o mês, evoluindo)
  MANU: ~14/dia (2o mês, evoluindo)
  Total: ~46/dia × 20 dias = ~920 registros
"""

import openpyxl
from datetime import date, datetime, timedelta
from collections import defaultdict, Counter
import random
import os
import sys

# ============================================================
# CONFIGURAÇÃO
# ============================================================
INPUT_FILE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10\CRM_VITAO360_V29_FINAL.xlsx'
OUTPUT_FILE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10\CRM_VITAO360_V30_FINAL.xlsx'

random.seed(45)

# Consultores ABR
CONSULTORES = {
    "DAIANE STADLER": {"volume_dia": 18, "tipo": "veterana"},
    "HELDER BRUNKOW": {"volume_dia": 14, "tipo": "2o_mes"},
    "MANU DITZEL": {"volume_dia": 14, "tipo": "2o_mes"},
}

# Feriados ABR 2025
FERIADOS_ABR = {date(2025, 4, 18), date(2025, 4, 21)}
ABR_START = date(2025, 4, 1)
ABR_END = date(2025, 4, 30)

# Coluna faturamento ABR na CARTEIRA
COL_FAT_ABR = 29

# ============================================================
# REGRAS DE NEGÓCIO (da aba REGRAS)
# ============================================================
RESULTADO_REGRAS = {
    "VENDA / PEDIDO":          (4,  "VENDAS",        "🔥QUENTE",  "VENDA"),
    "EM ATENDIMENTO":          (2,  "FUNIL",         "🟡MORNO",   "PRÉ-VENDA"),
    "SUPORTE":                 (1,  "RELAC.",        "🟡MORNO",   "RESOLUÇÃO DE PROBLEMA"),
    "ORÇAMENTO":               (3,  "FUNIL",         "🔥QUENTE",  "PRÉ-VENDA"),
    "RELACIONAMENTO":          (7,  "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "CADASTRO":                (2,  "FUNIL",         "🟡MORNO",   "PRÉ-VENDA"),
    "NÃO ATENDE":              (1,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "NÃO RESPONDE":            (1,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "RECUSOU LIGAÇÃO":         (2,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "CS (SUCESSO DO CLIENTE)": (30, "RELAC.",        "🔥QUENTE",  "PÓS-VENDA"),
    "PÓS-VENDA":              (30, "RELAC.",        "🔥QUENTE",  "PÓS-VENDA"),
    "PERDA / FECHOU LOJA":    (0,  "NÃO VENDA",    "💀PERDIDO",  "PRÉ-VENDA"),
    "FOLLOW UP 7":             (7,  "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "FOLLOW UP 15":            (15, "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "NUTRIÇÃO":                (15, "NÃO VENDA",    "❄️FRIO",    "PROSPECÇÃO"),
}

ACAO_FUTURA_MAP = {
    "VENDA / PEDIDO":           "CONFIRMAR FATURAMENTO E EXPEDIÇÃO",
    "EM ATENDIMENTO":           "FECHAR NEGOCIAÇÃO EM ANDAMENTO",
    "SUPORTE":                  "RESOLVER PROBLEMA INTERNO E ENVIAR SOLUÇÃO",
    "ORÇAMENTO":                "CONFIRMAR ORÇAMENTO ENVIADO",
    "RELACIONAMENTO":           "RAPPORT COM CLIENTE APÓS A VENDA ATÉ RECOMPRAR",
    "CADASTRO":                 "CONFIRMAR CADASTRO NO SISTEMA",
    "NÃO ATENDE":               "2ª TENTATIVA VIA LIGAÇÃO",
    "NÃO RESPONDE":             "2ª TENTATIVA VIA WHATSAPP",
    "RECUSOU LIGAÇÃO":          "2ª TENTATIVA VIA WHATSAPP",
    "CS (SUCESSO DO CLIENTE)":  "VERIFICAR SELL OUT E CRIAR INTENÇÃO RECOMPRA",
    "PÓS-VENDA":               "FAZER CS (SUCESSO DO CLIENTE)",
    "PERDA / FECHOU LOJA":     "NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES",
    "FOLLOW UP 7":              "COBRAR RETORNO DO CLIENTE",
    "FOLLOW UP 15":             "COBRAR RETORNO DO CLIENTE",
    "NUTRIÇÃO":                 "NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES",
}

MOTIVOS = {
    "EM ATENDIMENTO": ["PRIMEIRO CONTATO / SEM RESPOSTA", "AINDA TEM ESTOQUE", "SÓ QUER COMPRAR GRANEL", "NÃO SE APLICA"],
    "ORÇAMENTO": ["AINDA TEM ESTOQUE", "SÓ QUER COMPRAR GRANEL", "PROBLEMA FINANCEIRO / CRÉDITO", "NÃO SE APLICA"],
    "CADASTRO": ["PRIMEIRO CONTATO / SEM RESPOSTA", "NÃO SE APLICA"],
    "VENDA / PEDIDO": ["NÃO SE APLICA"],
    "NÃO RESPONDE": ["PRIMEIRO CONTATO / SEM RESPOSTA", "PROPRIETARIO / INDISPONÍVEL"],
    "NÃO ATENDE": ["PROPRIETARIO / INDISPONÍVEL", "PRIMEIRO CONTATO / SEM RESPOSTA"],
    "RECUSOU LIGAÇÃO": ["PROPRIETARIO / INDISPONÍVEL"],
    "PERDA / FECHOU LOJA": ["SEM INTERESSE NO MOMENTO", "PRODUTO NÃO VENDEU / SEM GIRO", "FECHANDO / FECHOU LOJA", "LOJA ANEXO/PROXIMO - SM", "PROBLEMA FINANCEIRO / CRÉDITO"],
    "PÓS-VENDA": ["STATUS PEDIDO", "ATRASO ENTREGA", "NÃO SE APLICA"],
    "CS (SUCESSO DO CLIENTE)": ["NÃO SE APLICA"],
    "FOLLOW UP": ["AINDA TEM ESTOQUE", "NÃO SE APLICA", "PRIMEIRO CONTATO / SEM RESPOSTA"],
}

DEMANDAS = {
    "PROSPECÇÃO": ["PROSPECÇÃO - BASE GRANEL", "PROSPECÇÃO - PESQUISA GOOGLE", "CONTATO LEADS DO SITE", "RESPONDER LEADS COM DÚVIDAS", "LIGAÇÕES PROSPECT"],
    "ORCAMENTO": ["MONTAR ORÇAMENTO SUGESTÃO (CURVA ABC)", "ENVIO DE AMOSTRAS", "ENVIAR ANÁLISE DE CRÉDITO", "REGISTRAR ATENDIMENTO NO MERCOS"],
    "CADASTRO": ["ATUALIZAÇÃO CADASTRAL MERCOS", "ENVIAR ANÁLISE DE CRÉDITO", "REGISTRAR ATENDIMENTO NO MERCOS", "PREENCHER FOLLOW-UP"],
    "PEDIDO": ["DIGITAÇÃO DE PEDIDOS", "SOLICITAR LINK CARTÃO CRÉDITO", "SOLICITAR VALORES PIX", "CONFIRMAR FATURAMENTO E EXPEDIÇÃO", "REGISTRAR ATENDIMENTO NO MERCOS"],
    "POS_VENDA": ["FAZER RASTREIOS", "ACOMPANHAR DEVOLUÇÕES", "SUPORTE PCV - DÚVIDAS PEDIDOS", "SUPORTE PRODUTO (LAUDOS/TABELAS)", "PREENCHER FOLLOW-UP"],
    "FOLLOWUP": ["LIGAÇÕES DA BASE", "ATENDIMENTO ATIVOS/INATIVOS", "PREENCHER FOLLOW-UP", "REGISTRAR ATENDIMENTO NO MERCOS", "COBRANÇA DE TÍTULOS"],
}

TIPOS_PROBLEMA = [
    "ATRASO ENTREGA (TRANSPORTADORA)", "PRODUTO AVARIADO (FÁBRICA/TRANSPORTE)",
    "ERRO SEPARAÇÃO (EXPEDIÇÃO)", "ERRO NOTA FISCAL (FATURAMENTO)",
    "DIVERGÊNCIA PREÇO (FATURAMENTO)", "COBRANÇA INDEVIDA (FINANCEIRO)",
    "RUPTURA ESTOQUE (FÁBRICA/PCP)", "PROBLEMA SISTEMA (TI)",
    "", "", "",
]

NOTAS = {
    "PROSPECÇÃO": [
        "Primeiro contato via WhatsApp. Apresentei portfólio Vitão.",
        "Enviei catálogo digital. Prospect demonstrou interesse.",
        "Contato inicial. Cliente pediu mais informações sobre a linha.",
        "Prospecção via ligação. Consegui falar com responsável de compras.",
    ],
    "NEGOCIAÇÃO": [
        "Negociação avançando. Cliente pediu orçamento atualizado.",
        "Enviei proposta ajustada. Aguardando aprovação do financeiro.",
        "Cliente comparando com concorrente. Reforçei diferenciais Vitão.",
        "Negociação ativa. Cliente quer fechar até fim do mês.",
    ],
    "FOLLOWUP": [
        "Retornei contato conforme agendado. Cliente analisando proposta.",
        "Follow-up realizado. Cliente pediu mais prazo para decidir.",
        "Retomei negociação. Cliente demonstra interesse mas sem urgência.",
        "Follow-up via WhatsApp. Cliente respondeu que vai analisar essa semana.",
    ],
    "VENDA": [
        "Pedido fechado! Cliente satisfeito com condições.",
        "Venda concluída. Primeira compra do cliente.",
        "Pedido aprovado. Enviei confirmação por email.",
    ],
    "POS_VENDA": [
        "Pós-venda. Confirmei faturamento e envio da NF.",
        "Pós-venda. Cliente recebeu tudo certo, satisfeito.",
        "CS. Verificando sell out e oportunidade de recompra.",
        "Acompanhamento pós-venda. Cliente elogiou atendimento.",
    ],
    "NAO_RESPONDE": ["WhatsApp enviado, sem retorno.", "Mensagem visualizada mas sem resposta.", "Tentei contato, sem resposta."],
    "NAO_ATENDE": ["Ligação não atendida. Caixa postal.", "Chamou mas não atendeu.", "Tentei ligar, sem sucesso."],
    "PERDA": ["Cliente sem interesse no momento.", "Não conseguiu fechar. Produto não se encaixa.", "Cliente optou por outro fornecedor."],
}

# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================
def dias_uteis_lista(start, end, feriados=None):
    feriados = feriados or set()
    result = []
    d = start
    while d <= end:
        if d.weekday() < 5 and d not in feriados:
            result.append(d)
        d += timedelta(days=1)
    return result

def proximo_dia_util(d, n=1, feriados=None):
    feriados = feriados or set()
    count = 0
    current = d
    while count < n:
        current += timedelta(days=1)
        if current.weekday() < 5 and current not in feriados:
            count += 1
    return current

def sinaleiro_calc(situacao, dias_compra, ciclo_medio):
    if not situacao:
        return "🟣"
    sit_upper = str(situacao).upper()
    if sit_upper in ("PROSPECT", "LEAD", "NOVO"):
        return "🟣"
    if dias_compra is None or dias_compra == "":
        return ""
    try:
        d = int(dias_compra)
        c = int(ciclo_medio) if ciclo_medio else 50
    except (ValueError, TypeError):
        return ""
    if d <= c:
        return "🟢"
    elif d <= c + 30:
        return "🟡"
    else:
        return "🔴"

def make_record(data, consultor, cnpj, nome, uf, rede, situacao, dias_compra,
                estagio_funil, tipo_cliente, fase, whatsapp, ligacao, lig_atendida,
                tipo_contato, resultado, motivo, followup_date, acao_futura,
                acao_detalhada, mercos, nota, temperatura, tipo_problema,
                demanda, tipo_atendimento, ciclo_medio=0):
    sinaleiro = sinaleiro_calc(situacao, dias_compra, ciclo_medio)
    grupo_dash = RESULTADO_REGRAS.get(resultado, (0, "", "", ""))[1]
    tipo_acao = RESULTADO_REGRAS.get(resultado, (0, "", "", ""))[3]
    return {
        'A': data, 'B': consultor, 'C': nome, 'D': cnpj,
        'E': uf, 'F': rede, 'G': situacao, 'H': dias_compra,
        'I': estagio_funil, 'J': tipo_cliente, 'K': fase,
        'L': sinaleiro, 'M': '',
        'N': whatsapp, 'O': ligacao, 'P': lig_atendida,
        'Q': tipo_contato, 'R': resultado, 'S': motivo,
        'T': followup_date, 'U': acao_futura, 'V': acao_detalhada,
        'W': mercos, 'X': nota, 'Y': temperatura,
        'Z': grupo_dash, 'AA': '', 'AB': tipo_acao,
        'AC': tipo_problema, 'AD': demanda, 'AE': tipo_atendimento,
    }

def gerar_ciclo_venda(cnpj_s, info, consultor, d_base, in_prev, prev_res, du, feriados, month_end):
    """Gera registros de ciclo completo até VENDA para um cliente."""
    records = []
    gap = random.choice([1, 2, 2, 3])

    if in_prev and prev_res in ("ORÇAMENTO", "EM ATENDIMENTO"):
        steps = ["CADASTRO", "VENDA / PEDIDO"]
    elif in_prev and prev_res in ("CADASTRO",):
        steps = ["VENDA / PEDIDO"]
    elif in_prev:
        steps = ["EM ATENDIMENTO", "ORÇAMENTO", "CADASTRO", "VENDA / PEDIDO"]
    else:
        steps = ["EM ATENDIMENTO", "ORÇAMENTO", "CADASTRO", "VENDA / PEDIDO"]

    d_current = d_base
    for step_idx, step in enumerate(steps):
        if step_idx > 0:
            d_current = proximo_dia_util(d_current, gap, feriados)
        if d_current > month_end:
            d_current = month_end
        if d_current.weekday() >= 5:
            d_current = proximo_dia_util(d_current - timedelta(days=1), 1, feriados)

        step_map = {
            "EM ATENDIMENTO": ("EM ATENDIMENTO", "EM ATENDIMENTO", "ENVIAR ORÇAMENTO / CATÁLOGO",
                              MOTIVOS["EM ATENDIMENTO"], NOTAS["NEGOCIAÇÃO"],
                              DEMANDAS["PROSPECÇÃO"] if not in_prev else DEMANDAS["FOLLOWUP"],
                              "PROSPECÇÃO" if not in_prev else "NEGOCIAÇÃO"),
            "ORÇAMENTO": ("ORÇAMENTO", "ORÇAMENTO", "CONFIRMAR ORÇAMENTO ENVIADO",
                         MOTIVOS["ORÇAMENTO"], NOTAS["NEGOCIAÇÃO"], DEMANDAS["ORCAMENTO"], "NEGOCIAÇÃO"),
            "CADASTRO": ("CADASTRO", "EM ATENDIMENTO", "CONFIRMAR CADASTRO NO SISTEMA",
                        MOTIVOS["CADASTRO"], ["Cadastro enviado para análise."], DEMANDAS["CADASTRO"], "NEGOCIAÇÃO"),
            "VENDA / PEDIDO": ("VENDA / PEDIDO", "PÓS-VENDA", "CONFIRMAR FATURAMENTO E EXPEDIÇÃO",
                              MOTIVOS["VENDA / PEDIDO"], NOTAS["VENDA"], DEMANDAS["PEDIDO"], "NEGOCIAÇÃO"),
        }
        resultado, estagio, acao, motivos_list, notas_list, demandas_list, tc = step_map[step]

        records.append(make_record(
            data=d_current, consultor=consultor, cnpj=cnpj_s,
            nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
            situacao=info.get('situacao', 'PROSPECT'), dias_compra=info.get('dias_compra', ''),
            estagio_funil=estagio, tipo_cliente=info.get('tipo_cliente', 'NOVO'),
            fase="PROSPECÇÃO" if step == "EM ATENDIMENTO" and not in_prev else estagio,
            whatsapp=1 if step_idx == 0 else 0, ligacao=1 if step_idx > 0 else 0, lig_atendida=1,
            tipo_contato=tc, resultado=resultado, motivo=random.choice(motivos_list),
            followup_date=proximo_dia_util(d_current, gap, feriados),
            acao_futura=acao, acao_detalhada="",
            mercos="SIM" if step in ("ORÇAMENTO", "VENDA / PEDIDO") else "NÃO",
            nota=random.choice(notas_list),
            temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "🟡MORNO", ""))[2],
            tipo_problema="", demanda=random.choice(demandas_list),
            tipo_atendimento="PROSPECÇÃO" if not in_prev and step_idx == 0 else "FOLLOW UP",
            ciclo_medio=info.get('ciclo_medio', 0)
        ))

    # D+4 pós-venda
    d_pv = proximo_dia_util(d_current, 4, feriados)
    if d_pv <= month_end:
        records.append(make_record(
            data=d_pv, consultor=consultor, cnpj=cnpj_s,
            nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
            situacao="NOVO", dias_compra=4,
            estagio_funil="PÓS-VENDA", tipo_cliente="NOVO", fase="PÓS-VENDA",
            whatsapp=0, ligacao=1, lig_atendida=1,
            tipo_contato="PÓS-VENDA / RELACIONAMENTO", resultado="PÓS-VENDA",
            motivo="NÃO SE APLICA",
            followup_date=proximo_dia_util(d_pv, 11, feriados),
            acao_futura="FAZER CS (SUCESSO DO CLIENTE)", acao_detalhada="",
            mercos="SIM", nota=random.choice(NOTAS["POS_VENDA"]),
            temperatura="🔥QUENTE", tipo_problema="",
            demanda=random.choice(DEMANDAS["POS_VENDA"]),
            tipo_atendimento="ATEND. CLIENTES ATIVOS"
        ))
    return records

def gerar_followup(cnpj_s, info, consultor, d, prev_res, feriados):
    """Gera 1 registro de follow-up baseado no resultado anterior."""
    rand = random.random()
    if prev_res in ("EM ATENDIMENTO", "ORÇAMENTO"):
        if rand < 0.40:
            resultado = "ORÇAMENTO" if prev_res == "EM ATENDIMENTO" else "EM ATENDIMENTO"
        elif rand < 0.70:
            resultado = "EM ATENDIMENTO"
        elif rand < 0.90:
            resultado = random.choice(["NÃO RESPONDE", "NÃO ATENDE"])
        else:
            resultado = "PERDA / FECHOU LOJA"
    elif prev_res in ("PÓS-VENDA", "CS (SUCESSO DO CLIENTE)"):
        resultado = random.choice(["RELACIONAMENTO", "CS (SUCESSO DO CLIENTE)", "FOLLOW UP 7"])
    elif prev_res in ("NÃO RESPONDE", "NÃO ATENDE", "RECUSOU LIGAÇÃO"):
        if rand < 0.50:
            resultado = "EM ATENDIMENTO"
        elif rand < 0.80:
            resultado = random.choice(["NÃO RESPONDE", "NÃO ATENDE"])
        else:
            resultado = "PERDA / FECHOU LOJA"
    else:
        resultado = "EM ATENDIMENTO"

    motivos_key = resultado if resultado in MOTIVOS else "FOLLOW UP"
    canal = random.choice(['whatsapp', 'ligacao', 'ligacao'])

    return make_record(
        data=d, consultor=consultor, cnpj=cnpj_s,
        nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
        situacao=info.get('situacao', 'PROSPECT'), dias_compra=info.get('dias_compra', ''),
        estagio_funil="EM ATENDIMENTO" if resultado in ("EM ATENDIMENTO",) else (
            "ORÇAMENTO" if resultado == "ORÇAMENTO" else "PROSPECÇÃO"),
        tipo_cliente=info.get('tipo_cliente', 'NOVO'), fase="PROSPECÇÃO",
        whatsapp=1 if canal == 'whatsapp' else 0,
        ligacao=1 if canal == 'ligacao' else 0,
        lig_atendida=1 if resultado not in ("NÃO ATENDE", "NÃO RESPONDE", "RECUSOU LIGAÇÃO") else 0,
        tipo_contato="FOLLOW UP", resultado=resultado,
        motivo=random.choice(MOTIVOS.get(motivos_key, ["NÃO SE APLICA"])),
        followup_date=proximo_dia_util(d, RESULTADO_REGRAS.get(resultado, (2,))[0], feriados) if resultado != "PERDA / FECHOU LOJA" else "",
        acao_futura=ACAO_FUTURA_MAP.get(resultado, ""),
        acao_detalhada="", mercos="SIM" if resultado == "ORÇAMENTO" else "NÃO",
        nota=random.choice(NOTAS.get("FOLLOWUP", ["Follow-up realizado."])) if resultado in ("EM ATENDIMENTO", "ORÇAMENTO") else (
            random.choice(NOTAS.get("NAO_RESPONDE", [""])) if "RESPONDE" in resultado else (
            random.choice(NOTAS.get("NAO_ATENDE", [""])) if "ATENDE" in resultado else (
            random.choice(NOTAS.get("PERDA", [""]))))),
        temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "🟡MORNO", ""))[2],
        tipo_problema="",
        demanda=random.choice(DEMANDAS.get("FOLLOWUP", [""])) if resultado in ("EM ATENDIMENTO", "ORÇAMENTO") else "",
        tipo_atendimento="FOLLOW UP",
        ciclo_medio=info.get('ciclo_medio', 0)
    )

def gerar_prospeccao(cnpj_s, info, consultor, d, feriados):
    """Gera 1 registro de prospecção nova."""
    rand = random.random()
    if rand < 0.60:
        resultado = "EM ATENDIMENTO"
    elif rand < 0.85:
        resultado = "NÃO RESPONDE"
    else:
        resultado = "NÃO ATENDE"

    motivos_key = resultado
    return make_record(
        data=d, consultor=consultor, cnpj=cnpj_s,
        nome=info.get('nome', ''), uf=info.get('uf', ''), rede=info.get('rede', ''),
        situacao="PROSPECT", dias_compra="",
        estagio_funil="PROSPECÇÃO", tipo_cliente="PROSPECT", fase="PROSPECÇÃO",
        whatsapp=1, ligacao=0, lig_atendida=0,
        tipo_contato="PROSPECÇÃO", resultado=resultado,
        motivo=random.choice(MOTIVOS.get(motivos_key, ["NÃO SE APLICA"])),
        followup_date=proximo_dia_util(d, RESULTADO_REGRAS.get(resultado, (2,))[0], feriados),
        acao_futura=ACAO_FUTURA_MAP.get(resultado, "ENVIAR ORÇAMENTO / CATÁLOGO"),
        acao_detalhada="", mercos="NÃO",
        nota=random.choice(NOTAS.get("PROSPECÇÃO", ["Primeiro contato."])) if resultado == "EM ATENDIMENTO" else (
            random.choice(NOTAS.get("NAO_RESPONDE", [""])) if "RESPONDE" in resultado else
            random.choice(NOTAS.get("NAO_ATENDE", [""]))),
        temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "❄️FRIO", ""))[2],
        tipo_problema="", demanda=random.choice(DEMANDAS["PROSPECÇÃO"]),
        tipo_atendimento="PROSPECÇÃO"
    )

# ============================================================
# INÍCIO
# ============================================================
print("=" * 80)
print("CRM VITAO360 — V30: Simulação Realista ABR/25")
print("=" * 80)

if not os.path.exists(INPUT_FILE):
    print(f"ERRO: {INPUT_FILE} não encontrado")
    sys.exit(1)

print(f"\n[1/7] Carregando V29...")
wb = openpyxl.load_workbook(INPUT_FILE)
ws_cart = wb['CARTEIRA']
ws_d2 = wb['DRAFT 2']

# ============================================================
# FASE 1: DADOS DA CARTEIRA
# ============================================================
print(f"\n[2/7] Lendo CARTEIRA...")
cart_data = {}
for r in range(4, ws_cart.max_row + 1):
    cnpj = ws_cart.cell(row=r, column=3).value
    if not cnpj:
        continue
    cnpj_s = str(cnpj).strip()
    cart_data[cnpj_s] = {
        'nome': ws_cart.cell(row=r, column=2).value or "",
        'uf': ws_cart.cell(row=r, column=4).value or "",
        'rede': ws_cart.cell(row=r, column=5).value or "",
        'situacao': ws_cart.cell(row=r, column=6).value or "",
        'tipo_cliente': ws_cart.cell(row=r, column=7).value or "",
        'ciclo_medio': ws_cart.cell(row=r, column=8).value or 0,
        'dias_compra': ws_cart.cell(row=r, column=9).value or "",
        'fat_abr': ws_cart.cell(row=r, column=COL_FAT_ABR).value or 0,
    }

# Compradores ABR
abr_buyers = [c for c, info in cart_data.items()
              if info['fat_abr'] and isinstance(info['fat_abr'], (int, float)) and info['fat_abr'] > 0]
print(f"  Compradores ABR/25: {len(abr_buyers)}")
print(f"  Total clientes CARTEIRA: {len(cart_data)}")

# ============================================================
# FASE 2: FOLLOW-UPS DE JAN+FEV+MAR (todos consultores)
# ============================================================
print(f"\n[3/7] Analisando follow-ups de JAN+FEV+MAR...")

# Nomes válidos (DAIANE pode aparecer como STAVICKI ou STADLER)
NOMES_DAIANE = {"DAIANE STAVICKI", "DAIANE STADLER"}
NOMES_VALIDOS = NOMES_DAIANE | {"HELDER BRUNKOW", "MANU DITZEL"}

prev_last = {}  # {(consultor_normalizado, cnpj): {'data', 'resultado'}}
for r in range(3, ws_d2.max_row + 1):
    data_cell = ws_d2.cell(row=r, column=1).value
    if data_cell is None:
        continue
    if isinstance(data_cell, (int, float)):
        data_cell = datetime(1900, 1, 1) + timedelta(days=int(data_cell) - 2)
    if not isinstance(data_cell, datetime):
        continue
    if data_cell.year != 2025 or data_cell.month > 3:
        continue
    cons = ws_d2.cell(row=r, column=2).value
    if cons is None:
        continue
    cons_str = str(cons).strip()
    # Normalizar nome DAIANE
    if cons_str in NOMES_DAIANE:
        cons_norm = "DAIANE STADLER"
    elif cons_str in NOMES_VALIDOS:
        cons_norm = cons_str
    else:
        continue
    cnpj = ws_d2.cell(row=r, column=4).value
    if not cnpj:
        continue
    cnpj_s = str(cnpj).strip()
    res = ws_d2.cell(row=r, column=18).value or ""
    key = (cons_norm, cnpj_s)
    if key not in prev_last or data_cell >= prev_last[key]['data']:
        prev_last[key] = {'data': data_cell, 'resultado': str(res)}

# Agrupar por consultor
prev_by_cons = defaultdict(lambda: defaultdict(list))
for (cons, cnpj), info in prev_last.items():
    prev_by_cons[cons][info['resultado']].append(cnpj)

for cons in sorted(prev_by_cons.keys()):
    total = sum(len(v) for v in prev_by_cons[cons].values())
    print(f"  {cons}: {total} CNPJs em JAN-MAR")
    for res, cnpjs in sorted(prev_by_cons[cons].items(), key=lambda x: len(x[1]), reverse=True)[:5]:
        print(f"    {res:<35} {len(cnpjs):>4}")

# ============================================================
# FASE 3: MONTAR POOLS E DISTRIBUIR
# ============================================================
print(f"\n[4/7] Montando pools ABR/25...")

du_abr = dias_uteis_lista(ABR_START, ABR_END, FERIADOS_ABR)
print(f"  Dias úteis ABR/25: {len(du_abr)} (feriados: Sexta-Santa 18, Tiradentes 21)")

# --- Follow-ups por consultor ---
def build_followup_pool(cons_name, max_fu=80, max_retry=25):
    """Constrói pools de follow-up e retry para um consultor."""
    by_res = prev_by_cons.get(cons_name, {})
    pool_fu = []
    for res in ["EM ATENDIMENTO", "ORÇAMENTO", "PÓS-VENDA", "CS (SUCESSO DO CLIENTE)", "VENDA / PEDIDO", "RELACIONAMENTO"]:
        cnpjs = by_res.get(res, [])
        random.shuffle(cnpjs)
        pool_fu.extend(cnpjs[:25])
    pool_fu = list(set(pool_fu))[:max_fu]

    pool_retry = []
    for res in ["NÃO RESPONDE", "NÃO ATENDE", "RECUSOU LIGAÇÃO"]:
        cnpjs = by_res.get(res, [])
        random.shuffle(cnpjs)
        pool_retry.extend(cnpjs[:12])
    pool_retry = list(set(pool_retry))[:max_retry]

    return pool_fu, pool_retry

pool_daiane_fu, pool_daiane_retry = build_followup_pool("DAIANE STADLER", max_fu=80, max_retry=25)
pool_helder_fu, pool_helder_retry = build_followup_pool("HELDER BRUNKOW", max_fu=50, max_retry=15)
pool_manu_fu, pool_manu_retry = build_followup_pool("MANU DITZEL", max_fu=50, max_retry=15)

# --- Vendas ABR: distribuir entre consultores ---
random.shuffle(abr_buyers)
n_vendas = len(abr_buyers)
n_daiane = int(n_vendas * 0.40)
n_helder = int(n_vendas * 0.30)
n_manu = n_vendas - n_daiane - n_helder

vendas_daiane = abr_buyers[:n_daiane]
vendas_helder = abr_buyers[n_daiane:n_daiane + n_helder]
vendas_manu = abr_buyers[n_daiane + n_helder:]

print(f"  Vendas ABR distribuídas:")
print(f"    DAIANE: {len(vendas_daiane)} vendas + {len(pool_daiane_fu)} FU + {len(pool_daiane_retry)} retries")
print(f"    HELDER: {len(vendas_helder)} vendas + {len(pool_helder_fu)} FU + {len(pool_helder_retry)} retries")
print(f"    MANU:   {len(vendas_manu)} vendas + {len(pool_manu_fu)} FU + {len(pool_manu_retry)} retries")

# --- Prospecções ---
all_used_cnpjs = set()
for (cons, cnpj), info in prev_last.items():
    all_used_cnpjs.add(cnpj)
all_used_cnpjs |= set(abr_buyers)

prospects_pool = [c for c in cart_data if c not in all_used_cnpjs]
random.shuffle(prospects_pool)

# DAIANE ~4/dia, HELDER ~3/dia, MANU ~3/dia
prospects_daiane = prospects_pool[:80]
prospects_helder = prospects_pool[80:140]
prospects_manu = prospects_pool[140:200]

print(f"  Prospecções: DAIANE {len(prospects_daiane)}, HELDER {len(prospects_helder)}, MANU {len(prospects_manu)}")

# ============================================================
# FASE 4: SIMULAR DIA A DIA
# ============================================================
print(f"\n[5/7] Simulando atendimentos ABR/25...")

registros_novos = []

def get_info(cnpj_s):
    return cart_data.get(cnpj_s, {'nome': '', 'uf': '', 'rede': '', 'situacao': 'PROSPECT',
                                   'tipo_cliente': 'NOVO', 'ciclo_medio': 0, 'dias_compra': ''})

def simular_consultor(nome, pool_fu, pool_retry, vendas, prospects, du):
    """Simula todos atendimentos de um consultor no mês."""
    recs = []

    # Follow-ups distribuídos ao longo do mês
    if pool_fu:
        fu_per_day = max(1, len(pool_fu) // len(du))
        fu_idx = 0
        for dia_idx, d in enumerate(du):
            extra = 1 if dia_idx < len(pool_fu) % len(du) else 0
            batch = pool_fu[fu_idx:fu_idx + fu_per_day + extra]
            fu_idx += len(batch)
            for cnpj_s in batch:
                info = get_info(cnpj_s)
                key = (nome, cnpj_s)
                prev_res = prev_last.get(key, {}).get('resultado', 'EM ATENDIMENTO')
                recs.append(gerar_followup(cnpj_s, info, nome, d, prev_res, FERIADOS_ABR))

    # Retries nos primeiros 10 dias
    if pool_retry:
        retry_per_day = max(1, len(pool_retry) // min(10, len(du)))
        retry_idx = 0
        for d in du[:10]:
            batch = pool_retry[retry_idx:retry_idx + retry_per_day + 1]
            retry_idx += len(batch)
            if retry_idx >= len(pool_retry):
                retry_idx = len(pool_retry)
            for cnpj_s in batch:
                info = get_info(cnpj_s)
                key = (nome, cnpj_s)
                prev_res = prev_last.get(key, {}).get('resultado', 'NÃO RESPONDE')
                recs.append(gerar_followup(cnpj_s, info, nome, d, prev_res, FERIADOS_ABR))

    # Ciclos de venda
    du_vendas = du[:15]  # Até ~meados do mês para dar tempo do ciclo
    for i, cnpj_s in enumerate(vendas):
        info = get_info(cnpj_s)
        d_base = du_vendas[i % len(du_vendas)]
        key = (nome, cnpj_s)
        in_prev = key in prev_last
        prev_res = prev_last.get(key, {}).get('resultado', '')
        recs.extend(gerar_ciclo_venda(cnpj_s, info, nome, d_base, in_prev, prev_res, du, FERIADOS_ABR, ABR_END))

    # Prospecções distribuídas ao longo do mês
    prosp_per_day = max(1, len(prospects) // len(du)) if prospects else 0
    prosp_idx = 0
    for dia_idx, d in enumerate(du):
        extra = 1 if prospects and dia_idx < len(prospects) % len(du) else 0
        batch = prospects[prosp_idx:prosp_idx + prosp_per_day + extra]
        prosp_idx += len(batch)
        for cnpj_s in batch:
            info = get_info(cnpj_s)
            recs.append(gerar_prospeccao(cnpj_s, info, nome, d, FERIADOS_ABR))

    return recs

# Simular cada consultor
recs_daiane = simular_consultor("DAIANE STADLER", pool_daiane_fu, pool_daiane_retry,
                                 vendas_daiane, prospects_daiane, du_abr)
print(f"  DAIANE: {len(recs_daiane)} registros ({len(recs_daiane)/len(du_abr):.1f}/dia)")
registros_novos.extend(recs_daiane)

recs_helder = simular_consultor("HELDER BRUNKOW", pool_helder_fu, pool_helder_retry,
                                 vendas_helder, prospects_helder, du_abr)
print(f"  HELDER: {len(recs_helder)} registros ({len(recs_helder)/len(du_abr):.1f}/dia)")
registros_novos.extend(recs_helder)

recs_manu = simular_consultor("MANU DITZEL", pool_manu_fu, pool_manu_retry,
                               vendas_manu, prospects_manu, du_abr)
print(f"  MANU:   {len(recs_manu)} registros ({len(recs_manu)/len(du_abr):.1f}/dia)")
registros_novos.extend(recs_manu)

print(f"\n  TOTAL ABR/25: {len(registros_novos)} registros")

# ============================================================
# FASE 5: VERIFICAÇÃO
# ============================================================
print(f"\n[6/7] Verificação lógica...")

res_count = Counter()
cons_count = Counter()
for rec in registros_novos:
    res_count[rec['R']] += 1
    cons_count[rec['B']] += 1

for res, n in res_count.most_common():
    print(f"    {res:<35} {n:>4}")

n_orc = res_count.get("ORÇAMENTO", 0)
n_cad = res_count.get("CADASTRO", 0)
n_ven = res_count.get("VENDA / PEDIDO", 0)
print(f"\n  Funil: ORÇAMENTO ({n_orc}) >= CADASTRO ({n_cad}) >= VENDA ({n_ven})")
if n_orc >= n_cad >= n_ven:
    print(f"  ✅ CONSISTENTE")
else:
    print(f"  ⚠️ VERIFICAR")

print(f"\n  Por consultor:")
for cons, n in cons_count.most_common():
    print(f"    {cons:<25} {n:>4} ({n/len(du_abr):.1f}/dia)")

n_total = len(registros_novos)
n_motivo = sum(1 for r in registros_novos if r['S'])
n_demanda = sum(1 for r in registros_novos if r['AD'])
n_acao = sum(1 for r in registros_novos if r['U'])
n_nota = sum(1 for r in registros_novos if r['X'])
print(f"\n  Preenchimento: MOTIVO {n_motivo/n_total*100:.0f}% | DEMANDA {n_demanda/n_total*100:.0f}% | AÇÃO {n_acao/n_total*100:.0f}% | NOTA {n_nota/n_total*100:.0f}%")

# ============================================================
# FASE 6: RECONSTRUIR DRAFT 2
# ============================================================
print(f"\n[7/7] Salvando V30...")

COLS = ['A','B','C','D','E','F','G','H','I','J','K','L','M',
        'N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
        'AA','AB','AC','AD','AE']

existing_records = []
removed_abr = 0
for r in range(3, ws_d2.max_row + 1):
    data_cell = ws_d2.cell(row=r, column=1).value
    if data_cell is None:
        continue
    is_abr = False
    if isinstance(data_cell, (int, float)):
        dt = datetime(1900, 1, 1) + timedelta(days=int(data_cell) - 2)
        if dt.year == 2025 and dt.month == 4:
            is_abr = True
    elif isinstance(data_cell, datetime):
        if data_cell.year == 2025 and data_cell.month == 4:
            is_abr = True
    if is_abr:
        removed_abr += 1
        continue
    row_data = {}
    for c_idx, col_letter in enumerate(COLS):
        row_data[col_letter] = ws_d2.cell(row=r, column=c_idx + 1).value
    existing_records.append(row_data)

print(f"  Registros existentes mantidos: {len(existing_records)}")
print(f"  Registros ABR removidos: {removed_abr}")
print(f"  Registros ABR novos: {len(registros_novos)}")

all_records = existing_records + registros_novos
def sort_key(rec):
    d = rec.get('A')
    if d is None:
        return datetime(1900, 1, 1)
    if isinstance(d, (int, float)):
        return datetime(1900, 1, 1) + timedelta(days=int(d) - 2)
    if isinstance(d, date) and not isinstance(d, datetime):
        return datetime(d.year, d.month, d.day)
    return d if isinstance(d, datetime) else datetime(1900, 1, 1)

all_records.sort(key=sort_key)
print(f"  Total DRAFT 2: {len(all_records)} registros")

# TENTATIVA
print(f"  Calculando TENTATIVA (col M)...")
cnpj_counter = Counter()
for rec in all_records:
    cnpj = rec.get('D')
    if cnpj:
        cnpj_s = str(cnpj).strip()
        cnpj_counter[cnpj_s] += 1
        rec['M'] = f"T{cnpj_counter[cnpj_s]}"

# Limpar e reescrever
for r in range(3, ws_d2.max_row + 1):
    for c in range(1, len(COLS) + 1):
        ws_d2.cell(row=r, column=c).value = None

for i, rec in enumerate(all_records):
    row = i + 3
    for c_idx, col_letter in enumerate(COLS):
        ws_d2.cell(row=row, column=c_idx + 1).value = rec.get(col_letter)

print(f"\n  Salvando {OUTPUT_FILE}...")
wb.save(OUTPUT_FILE)
file_size = os.path.getsize(OUTPUT_FILE) / (1024 * 1024)

print(f"\n{'='*80}")
print(f"V30 CONCLUÍDO!")
print(f"{'='*80}")
print(f"  Arquivo: {OUTPUT_FILE}")
print(f"  Tamanho: {file_size:.2f} MB")
print(f"\n  DRAFT 2: {len(all_records)} registros")
print(f"  ABR/25: {len(registros_novos)} novos (removidos {removed_abr} antigos)")
print(f"\n  Por consultor:")
for cons, n in cons_count.most_common():
    print(f"    {cons:<25} {n:>4} ({n/len(du_abr):.1f}/dia)")
print(f"\n  Vendas ABR: {n_ven} de {len(abr_buyers)} esperadas")
print(f"  Média total/dia: {len(registros_novos)/len(du_abr):.1f}")

# Demandas
dem_count = Counter()
for rec in registros_novos:
    if rec['AD']:
        dem_count[rec['AD']] += 1
print(f"\n  DEMANDAS geradas:")
for dem, n in dem_count.most_common(15):
    print(f"    {dem:<45} {n:>4}")

print(f"{'='*80}")

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera CRM_VITAO360_V14_FINAL.xlsx
Aplica correções de layout aprendidas do V31 sobre a base V13.

Correções aplicadas:
1. CARTEIRA: freeze_panes corrigido para padrão V31
2. DRAFT 1: freeze_panes adicionado
3. DRAFT 2: freeze_panes adicionado
4. LOG: freeze_panes ajustado
5. CARTEIRA: auto_filter range expandido
6. Column widths: precisão restaurada do V31 para CARTEIRA
7. Todas abas: column widths recalculados onde necessário
"""

import openpyxl
import json
import shutil
from pathlib import Path
from datetime import datetime

# Caminhos
V13_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V13_FINAL.xlsx"
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"
V14_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V14_FINAL.xlsx"
LAYOUT_JSON = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/layout_comparison_v31_v13.json"

def load_v31_layout():
    """Carrega layout do V31 para referência"""
    with open(LAYOUT_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data.get("v31", {})

def apply_freeze_panes(ws, freeze_cell):
    """Aplica freeze_panes de forma segura"""
    try:
        ws.freeze_panes = freeze_cell
        return True
    except Exception as e:
        print(f"  ! Erro ao aplicar freeze_panes {freeze_cell}: {e}")
        return False

def apply_auto_filter(ws, filter_range):
    """Aplica auto_filter"""
    try:
        ws.auto_filter.ref = filter_range
        return True
    except Exception as e:
        print(f"  ! Erro ao aplicar auto_filter: {e}")
        return False

def copy_column_widths_from_v31(ws_target, v31_sheet_data):
    """Copia larguras de coluna do V31"""
    widths = v31_sheet_data.get("column_widths", {})
    count = 0
    for col_letter, width in widths.items():
        try:
            ws_target.column_dimensions[col_letter].width = width
            count += 1
        except Exception:
            pass
    return count

def copy_outline_levels_from_v31(ws_target, v31_sheet_data):
    """Copia outline levels (grouping) do V31"""
    outlines = v31_sheet_data.get("outline_levels", {})
    col_groups = outlines.get("columns", {}).get("groups", [])
    count = 0
    for group in col_groups:
        col = group.get("column")
        level = group.get("level", 0)
        hidden = group.get("hidden", False)
        if col:
            try:
                ws_target.column_dimensions[col].outline_level = level
                ws_target.column_dimensions[col].hidden = hidden
                count += 1
            except Exception:
                pass
    return count


def main():
    print("\n" + "=" * 80)
    print("GERAÇÃO V14 FINAL — Layout corrigido com aprendizados V31")
    print("=" * 80)
    print(f"Data: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # Step 1: Copiar V13 como base
    print("\n[1/6] Copiando V13 como base para V14...")
    shutil.copy2(V13_PATH, V14_PATH)
    print(f"  ✓ Copiado: {V14_PATH}")

    # Step 2: Carregar V31 layout data
    print("\n[2/6] Carregando dados de layout do V31...")
    v31_data = load_v31_layout()
    v31_sheets = v31_data.get("sheets", {})
    print(f"  ✓ Layout de {len(v31_sheets)} abas carregado")

    # Step 3: Abrir V14 para edição
    print("\n[3/6] Abrindo V14 para edição...")
    wb = openpyxl.load_workbook(V14_PATH)
    print(f"  ✓ {len(wb.sheetnames)} abas: {', '.join(wb.sheetnames)}")

    changes_log = []

    # Step 4: Aplicar correções de freeze_panes
    print("\n[4/6] Aplicando correções de freeze_panes...")

    freeze_fixes = {
        # CARTEIRA: V31 usa BX66 mas V13 tem estrutura diferente (557 rows vs 8305)
        # Vamos usar freeze sensato: congelar header (row 3-5) + coluna âncora
        # V31 congela em BX66 porque tem 8305 rows. Para V13 com 557 rows:
        # Melhor: congelar coluna AR (como antes) mas na linha 4 (após headers)
        "CARTEIRA": "AR4",
        # DRAFT 1: V31 tem AN4, V13 não tinha
        "DRAFT 1": "AN4",
        # DRAFT 2: V31 tem A13070 (absurdo, provavelmente scroll position)
        # Melhor: freeze na linha 2 para manter headers
        "DRAFT 2": "A2",
        # LOG: V31 usa A2, V13 usa A3 — usar A3 que preserva header melhor
        "LOG": "A3",
        # PROJEÇÃO: V31 usa C4, V13 usa E30 — C4 é mais útil
        "PROJEÇÃO ": "C4",
        # REGRAS: V31 usa A41, V13 não tinha
        "REGRAS": "A41",
    }

    for sheet_name, freeze_cell in freeze_fixes.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            old_freeze = ws.freeze_panes
            if apply_freeze_panes(ws, freeze_cell):
                msg = f"  ✓ {sheet_name}: freeze_panes {old_freeze} → {freeze_cell}"
                print(msg)
                changes_log.append(msg)
        else:
            print(f"  - {sheet_name}: aba não encontrada")

    # AGENDA tabs: add freeze for all 4
    for sheet_name in wb.sheetnames:
        if "AGENDA" in sheet_name.upper():
            ws = wb[sheet_name]
            if not ws.freeze_panes:
                apply_freeze_panes(ws, "A2")
                msg = f"  ✓ {sheet_name}: freeze_panes adicionado A2"
                print(msg)
                changes_log.append(msg)

    # Step 5: Aplicar auto_filter corrigido
    print("\n[5/6] Aplicando auto_filter...")

    # CARTEIRA: expandir auto_filter para cobrir todas as linhas
    if "CARTEIRA" in wb.sheetnames:
        ws = wb["CARTEIRA"]
        max_col_letter = openpyxl.utils.get_column_letter(ws.max_column)
        filter_range = f"A3:{max_col_letter}{ws.max_row}"
        old_filter = ws.auto_filter.ref if ws.auto_filter else "None"
        if apply_auto_filter(ws, filter_range):
            msg = f"  ✓ CARTEIRA: auto_filter {old_filter} → {filter_range}"
            print(msg)
            changes_log.append(msg)

    # DRAFT 1: add auto_filter
    if "DRAFT 1" in wb.sheetnames:
        ws = wb["DRAFT 1"]
        if ws.max_row and ws.max_row > 1:
            max_col = openpyxl.utils.get_column_letter(ws.max_column)
            filter_range = f"A3:{max_col}{ws.max_row}"
            if apply_auto_filter(ws, filter_range):
                msg = f"  ✓ DRAFT 1: auto_filter adicionado {filter_range}"
                print(msg)
                changes_log.append(msg)

    # LOG: keep existing auto_filter (already good in V13)

    # Step 6: Aplicar column widths do V31
    print("\n[6/6] Aplicando column widths do V31...")

    # Sheets que existem em ambos — copiar widths do V31
    common_sheets = set(wb.sheetnames) & set(v31_sheets.keys())
    for sheet_name in sorted(common_sheets):
        ws = wb[sheet_name]
        v31_sd = v31_sheets[sheet_name]

        # Copiar widths
        count_w = copy_column_widths_from_v31(ws, v31_sd)

        # Copiar outline levels (grouping) — apenas CARTEIRA
        count_o = 0
        if "CARTEIRA" in sheet_name.upper():
            count_o = copy_outline_levels_from_v31(ws, v31_sd)

        if count_w > 0 or count_o > 0:
            msg = f"  ✓ {sheet_name}: {count_w} column widths + {count_o} outline levels copiados do V31"
            print(msg)
            changes_log.append(msg)

    # Salvar
    print("\n[SALVANDO] Gerando V14 FINAL...")
    wb.save(V14_PATH)
    wb.close()

    file_size = Path(V14_PATH).stat().st_size / (1024 * 1024)
    print(f"  ✓ Salvo: {V14_PATH} ({file_size:.2f} MB)")

    # Gerar log de mudanças
    print("\n" + "=" * 80)
    print("CHANGELOG V13 → V14")
    print("=" * 80)
    for change in changes_log:
        print(change)

    print(f"\nTotal de correções: {len(changes_log)}")
    print("\n[SUCESSO] V14 FINAL gerado com layout corrigido!")

    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())

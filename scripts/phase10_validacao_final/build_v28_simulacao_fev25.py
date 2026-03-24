#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CRM VITAO360 — V28: Simulação Realista de Atendimentos FEV/25
==============================================================
Base: V27 (com JAN/25 já simulado)
Motor de simulação que gera ~400-500 registros para FEV/25
Volume REDUZIDO: Daiane sozinha, início de projeto, muitas reuniões.

Fontes de registros FEV:
  1. Follow-ups de JAN (continuação de negociações abertas)
  2. Vendas FEV reais (29 clientes SAP — ciclo completo)
  3. ~5 prospecções/dia (novos prospects)

Target: ~20-25 atendimentos/dia × 20 dias úteis = ~400-500 registros
"""

import openpyxl
from datetime import date, datetime, timedelta
from collections import defaultdict, Counter
import random
import os
import sys

# ============================================================
# CONFIGURAÇÃO
# ============================================================
INPUT_FILE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10\CRM_VITAO360_V27_FINAL.xlsx'
OUTPUT_FILE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10\CRM_VITAO360_V28_FINAL.xlsx'

random.seed(43)  # Seed diferente de V27

CONSULTOR_FEV = "DAIANE STAVICKI"

# ============================================================
# REGRAS DE NEGÓCIO (da aba REGRAS)
# ============================================================
RESULTADO_REGRAS = {
    "VENDA / PEDIDO":          (4,  "VENDAS",        "🔥QUENTE",  "VENDA"),
    "EM ATENDIMENTO":          (2,  "FUNIL",         "🟡MORNO",   "PRÉ-VENDA"),
    "SUPORTE":                 (1,  "RELAC.",        "🟡MORNO",   "RESOLUÇÃO DE PROBLEMA"),
    "ORÇAMENTO":               (3,  "FUNIL",         "🔥QUENTE",  "PRÉ-VENDA"),
    "RELACIONAMENTO":          (7,  "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "CADASTRO":                (2,  "FUNIL",         "🟡MORNO",   "PRÉ-VENDA"),
    "NÃO ATENDE":              (1,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "NÃO RESPONDE":            (1,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "RECUSOU LIGAÇÃO":         (2,  "NÃO VENDA",    "❄️FRIO",    "PRÉ-VENDA"),
    "CS (SUCESSO DO CLIENTE)": (30, "RELAC.",        "🔥QUENTE",  "PÓS-VENDA"),
    "PÓS-VENDA":              (30, "RELAC.",        "🔥QUENTE",  "PÓS-VENDA"),
    "PERDA / FECHOU LOJA":    (0,  "NÃO VENDA",    "💀PERDIDO",  "PRÉ-VENDA"),
    "FOLLOW UP 7":             (7,  "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "FOLLOW UP 15":            (15, "RELAC.",        "🟡MORNO",   "PÓS-VENDA"),
    "NUTRIÇÃO":                (15, "NÃO VENDA",    "❄️FRIO",    "PROSPECÇÃO"),
}

# AÇÕES FUTURAS — da aba REGRAS Seção 13 + Motor de Regras (Seção 17)
ACAO_FUTURA_MAP = {
    "VENDA / PEDIDO":           "CONFIRMAR FATURAMENTO E EXPEDIÇÃO",
    "EM ATENDIMENTO":           "FECHAR NEGOCIAÇÃO EM ANDAMENTO",
    "SUPORTE":                  "RESOLVER PROBLEMA INTERNO E ENVIAR SOLUÇÃO",
    "ORÇAMENTO":                "CONFIRMAR ORÇAMENTO ENVIADO",
    "RELACIONAMENTO":           "RAPPORT COM CLIENTE APÓS A VENDA ATÉ RECOMPRAR",
    "CADASTRO":                 "CONFIRMAR CADASTRO NO SISTEMA",
    "NÃO ATENDE":               "2ª TENTATIVA VIA LIGAÇÃO",
    "NÃO RESPONDE":             "2ª TENTATIVA VIA WHATSAPP",
    "RECUSOU LIGAÇÃO":          "2ª TENTATIVA VIA WHATSAPP",
    "CS (SUCESSO DO CLIENTE)":  "VERIFICAR SELL OUT E CRIAR INTENÇÃO RECOMPRA",
    "PÓS-VENDA":               "FAZER CS (SUCESSO DO CLIENTE)",
    "PERDA / FECHOU LOJA":     "NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES",
    "FOLLOW UP 7":              "COBRAR RETORNO DO CLIENTE",
    "FOLLOW UP 15":             "COBRAR RETORNO DO CLIENTE",
    "NUTRIÇÃO":                 "NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES",
}

# TIPO DO CONTATO — da aba REGRAS Seção 2
TIPO_CONTATO_MAP = {
    "PROSPECÇÃO":                "PROSPECÇÃO",
    "NEGOCIAÇÃO":               "NEGOCIAÇÃO",
    "FOLLOW UP":                "FOLLOW UP",
    "PÓS-VENDA":               "PÓS-VENDA / RELACIONAMENTO",
    "ATEND_ATIVOS":             "ATEND. CLIENTES ATIVOS",
    "ATEND_INATIVOS":           "ATEND. CLIENTES INATIVOS",
    "PERDA":                    "PERDA / NUTRIÇÃO",
}

# MOTIVOS — da aba REGRAS Seção 3
MOTIVOS_EM_ATENDIMENTO = [
    "PRIMEIRO CONTATO / SEM RESPOSTA",
    "AINDA TEM ESTOQUE",
    "SÓ QUER COMPRAR GRANEL",
    "NÃO SE APLICA",
]
MOTIVOS_ORCAMENTO = [
    "AINDA TEM ESTOQUE",
    "SÓ QUER COMPRAR GRANEL",
    "PROBLEMA FINANCEIRO / CRÉDITO",
    "NÃO SE APLICA",
]
MOTIVOS_CADASTRO = [
    "PRIMEIRO CONTATO / SEM RESPOSTA",
    "NÃO SE APLICA",
]
MOTIVOS_VENDA = ["NÃO SE APLICA"]
MOTIVOS_NAO_RESPONDE = [
    "PRIMEIRO CONTATO / SEM RESPOSTA",
    "PROPRIETARIO / INDISPONÍVEL",
]
MOTIVOS_NAO_ATENDE = [
    "PROPRIETARIO / INDISPONÍVEL",
    "PRIMEIRO CONTATO / SEM RESPOSTA",
]
MOTIVOS_PERDA = [
    "SEM INTERESSE NO MOMENTO",
    "PRODUTO NÃO VENDEU / SEM GIRO",
    "FECHANDO / FECHOU LOJA",
    "LOJA ANEXO/PROXIMO - SM",
    "PROBLEMA FINANCEIRO / CRÉDITO",
]
MOTIVOS_POS_VENDA = [
    "STATUS PEDIDO",
    "ATRASO ENTREGA",
    "NÃO SE APLICA",
]
MOTIVOS_FOLLOWUP = [
    "AINDA TEM ESTOQUE",
    "NÃO SE APLICA",
    "PRIMEIRO CONTATO / SEM RESPOSTA",
]

# DEMANDAS — da aba REGRAS Seção 14 (25 tarefas reais)
DEMANDAS_PROSPECÇÃO = [
    "PROSPECÇÃO - BASE GRANEL",
    "PROSPECÇÃO - PESQUISA GOOGLE",
    "CONTATO LEADS DO SITE",
    "RESPONDER LEADS COM DÚVIDAS",
    "LIGAÇÕES PROSPECT",
]
DEMANDAS_ORCAMENTO = [
    "MONTAR ORÇAMENTO SUGESTÃO (CURVA ABC)",
    "ENVIO DE AMOSTRAS",
    "ENVIAR ANÁLISE DE CRÉDITO",
    "REGISTRAR ATENDIMENTO NO MERCOS",
]
DEMANDAS_CADASTRO = [
    "ATUALIZAÇÃO CADASTRAL MERCOS",
    "ENVIAR ANÁLISE DE CRÉDITO",
    "REGISTRAR ATENDIMENTO NO MERCOS",
    "PREENCHER FOLLOW-UP",
]
DEMANDAS_PEDIDO = [
    "DIGITAÇÃO DE PEDIDOS",
    "SOLICITAR LINK CARTÃO CRÉDITO",
    "SOLICITAR VALORES PIX",
    "CONFIRMAR FATURAMENTO E EXPEDIÇÃO",
    "REGISTRAR ATENDIMENTO NO MERCOS",
]
DEMANDAS_POS_VENDA = [
    "FAZER RASTREIOS",
    "ACOMPANHAR DEVOLUÇÕES",
    "SUPORTE PCV - DÚVIDAS PEDIDOS",
    "SUPORTE PRODUTO (LAUDOS/TABELAS)",
    "PREENCHER FOLLOW-UP",
]
DEMANDAS_FOLLOWUP = [
    "LIGAÇÕES DA BASE",
    "ATENDIMENTO ATIVOS/INATIVOS",
    "PREENCHER FOLLOW-UP",
    "REGISTRAR ATENDIMENTO NO MERCOS",
    "COBRANÇA DE TÍTULOS",
]

# TIPO PROBLEMA — da aba REGRAS Seção 12
TIPOS_PROBLEMA = [
    "ATRASO ENTREGA (TRANSPORTADORA)",
    "PRODUTO AVARIADO (FÁBRICA/TRANSPORTE)",
    "ERRO SEPARAÇÃO (EXPEDIÇÃO)",
    "ERRO NOTA FISCAL (FATURAMENTO)",
    "DIVERGÊNCIA PREÇO (FATURAMENTO)",
    "COBRANÇA INDEVIDA (FINANCEIRO)",
    "RUPTURA ESTOQUE (FÁBRICA/PCP)",
    "PROBLEMA SISTEMA (TI)",
    "", "", "",  # Maioria sem problema
]

# NOTAS realistas por tipo de atividade
NOTAS_FOLLOWUP = [
    "Retornei contato conforme agendado. Cliente está analisando proposta.",
    "Follow-up realizado. Cliente pediu mais prazo para decidir.",
    "Retomei negociação. Cliente demonstra interesse mas sem urgência.",
    "Liguei conforme combinado. Responsável em reunião, retornar amanhã.",
    "Follow-up via WhatsApp. Cliente respondeu que vai analisar essa semana.",
    "Retornei ligação. Cliente confirmou interesse, quer agendar reunião.",
]
NOTAS_NEGOCIACAO = [
    "Negociação avançando. Cliente pediu orçamento atualizado.",
    "Enviei proposta ajustada. Aguardando aprovação do financeiro.",
    "Cliente comparando com concorrente. Reforçei diferenciais Vitão.",
    "Negociação ativa. Cliente quer fechar até fim do mês.",
    "Apresentei condições especiais. Cliente vai consultar sócio.",
]
NOTAS_VENDA = [
    "Pedido fechado! Cliente satisfeito com condições.",
    "Venda concluída. Primeira compra do cliente.",
    "Pedido aprovado. Enviei confirmação por email.",
    "Negociação fechada com sucesso. Agendar pós-venda D+4.",
]
NOTAS_POS_VENDA = [
    "Pós-venda D+4. Confirmei faturamento e envio da NF.",
    "Pós-venda D+15. Cliente recebeu tudo certo, satisfeito.",
    "CS D+30. Verificando sell out e oportunidade de recompra.",
    "Acompanhamento pós-venda. Cliente elogiou atendimento.",
    "Pós-venda. Produto chegou bem, cliente quer repetir pedido.",
]
NOTAS_PROSPECÇÃO = [
    "Primeiro contato via WhatsApp. Apresentei portfólio Vitão.",
    "Enviei catálogo digital. Prospect demonstrou interesse.",
    "Contato inicial. Cliente pediu mais informações sobre a linha.",
    "WhatsApp enviado com apresentação. Aguardando retorno.",
    "Prospecção via ligação. Consegui falar com responsável de compras.",
]
NOTAS_NAO_RESPONDE = [
    "WhatsApp enviado, sem retorno ainda.",
    "Mensagem visualizada mas sem resposta.",
    "Tentei contato, número correto mas sem resposta.",
]
NOTAS_NAO_ATENDE = [
    "Ligação não atendida. Caixa postal.",
    "Chamou mas não atendeu. Tentar novamente amanhã.",
    "Tentei ligar, sem sucesso. Vou enviar WhatsApp.",
]
NOTAS_PERDA = [
    "Cliente sem interesse no momento. Colocar em nutrição.",
    "Não conseguiu fechar. Produto não se encaixa no perfil da loja.",
    "Cliente optou por outro fornecedor. Manter contato para futuro.",
]

# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================
def dias_uteis_lista(start, end):
    """Retorna lista de dias úteis entre start e end."""
    result = []
    d = start
    while d <= end:
        if d.weekday() < 5:
            result.append(d)
        d += timedelta(days=1)
    return result

def proximo_dia_util(d, n=1):
    """Retorna o n-ésimo dia útil após d."""
    count = 0
    current = d
    while count < n:
        current += timedelta(days=1)
        if current.weekday() < 5:
            count += 1
    return current

def sinaleiro_calc(situacao, dias_compra, ciclo_medio):
    if not situacao:
        return "🟣"
    sit_upper = str(situacao).upper()
    if sit_upper in ("PROSPECT", "LEAD", "NOVO"):
        return "🟣"
    if dias_compra is None or dias_compra == "":
        return ""
    try:
        d = int(dias_compra)
        c = int(ciclo_medio) if ciclo_medio else 50
    except (ValueError, TypeError):
        return ""
    if d <= c:
        return "🟢"
    elif d <= c + 30:
        return "🟡"
    else:
        return "🔴"

def make_record(data, consultor, cnpj, nome, uf, rede, situacao, dias_compra,
                estagio_funil, tipo_cliente, fase, whatsapp, ligacao, lig_atendida,
                tipo_contato, resultado, motivo, followup_date, acao_futura,
                acao_detalhada, mercos, nota, temperatura, tipo_problema,
                demanda, tipo_atendimento, ciclo_medio=0):
    sinaleiro = sinaleiro_calc(situacao, dias_compra, ciclo_medio)
    grupo_dash = RESULTADO_REGRAS.get(resultado, (0, "", "", ""))[1]
    tipo_acao = RESULTADO_REGRAS.get(resultado, (0, "", "", ""))[3]
    return {
        'A': data, 'B': consultor, 'C': nome, 'D': cnpj,
        'E': uf, 'F': rede, 'G': situacao, 'H': dias_compra,
        'I': estagio_funil, 'J': tipo_cliente, 'K': fase,
        'L': sinaleiro, 'M': '',
        'N': whatsapp, 'O': ligacao, 'P': lig_atendida,
        'Q': tipo_contato, 'R': resultado, 'S': motivo,
        'T': followup_date, 'U': acao_futura, 'V': acao_detalhada,
        'W': mercos, 'X': nota, 'Y': temperatura,
        'Z': grupo_dash, 'AA': '', 'AB': tipo_acao,
        'AC': tipo_problema, 'AD': demanda, 'AE': tipo_atendimento,
    }

# ============================================================
# INÍCIO
# ============================================================
print("=" * 80)
print("CRM VITAO360 — V28: Simulação Realista FEV/25")
print("=" * 80)

if not os.path.exists(INPUT_FILE):
    print(f"ERRO: {INPUT_FILE} não encontrado")
    sys.exit(1)

print(f"\n[1/7] Carregando V27...")
wb = openpyxl.load_workbook(INPUT_FILE)
ws_cart = wb['CARTEIRA']
ws_d2 = wb['DRAFT 2']

# ============================================================
# FASE 1: LEVANTAR DADOS DE JAN PARA FOLLOW-UPS
# ============================================================
print(f"\n[2/7] Analisando registros JAN/25 para follow-ups...")

# Ler dados da CARTEIRA (CNPJ, nome, UF, rede, situação)
cart_data = {}
for r in range(4, ws_cart.max_row + 1):
    cnpj = ws_cart.cell(row=r, column=3).value   # col C = CNPJ
    if not cnpj:
        continue
    cnpj_s = str(cnpj).strip()
    cart_data[cnpj_s] = {
        'nome': ws_cart.cell(row=r, column=2).value or "",     # col B = NOME
        'uf': ws_cart.cell(row=r, column=4).value or "",       # col D = UF
        'rede': ws_cart.cell(row=r, column=5).value or "",     # col E = REDE
        'situacao': ws_cart.cell(row=r, column=6).value or "", # col F = SITUAÇÃO
        'tipo_cliente': ws_cart.cell(row=r, column=7).value or "", # col G
        'ciclo_medio': ws_cart.cell(row=r, column=8).value or 0,  # col H
        'dias_compra': ws_cart.cell(row=r, column=9).value or "",  # col I
        'curva': ws_cart.cell(row=r, column=10).value or "",       # col J
        'fat_fev': ws_cart.cell(row=r, column=27).value or 0,     # col AA = FEV/25
    }

# Coletar último registro por CNPJ em JAN (da DAIANE)
jan_last_by_cnpj = {}
for r in range(3, ws_d2.max_row + 1):
    data_cell = ws_d2.cell(row=r, column=1).value
    if data_cell is None:
        continue
    if isinstance(data_cell, (int, float)):
        data_cell = datetime(1900, 1, 1) + timedelta(days=int(data_cell) - 2)
    if not isinstance(data_cell, datetime):
        continue
    if data_cell.year != 2025 or data_cell.month != 1:
        continue
    consultor = ws_d2.cell(row=r, column=2).value
    if consultor != CONSULTOR_FEV:
        continue

    cnpj = ws_d2.cell(row=r, column=4).value
    if not cnpj:
        continue
    cnpj_s = str(cnpj).strip()
    resultado = ws_d2.cell(row=r, column=18).value or ""

    if cnpj_s not in jan_last_by_cnpj or data_cell >= jan_last_by_cnpj[cnpj_s]['data']:
        jan_last_by_cnpj[cnpj_s] = {
            'data': data_cell,
            'resultado': str(resultado),
            'nome': ws_d2.cell(row=r, column=3).value or "",
            'uf': ws_d2.cell(row=r, column=5).value or "",
            'rede': ws_d2.cell(row=r, column=6).value or "",
        }

print(f"  CNPJs DAIANE em JAN: {len(jan_last_by_cnpj)}")

# Classificar por último resultado
jan_by_resultado = defaultdict(list)
for cnpj_s, info in jan_last_by_cnpj.items():
    jan_by_resultado[info['resultado']].append(cnpj_s)

for res, cnpjs in sorted(jan_by_resultado.items(), key=lambda x: len(x[1]), reverse=True):
    print(f"    {res:<30} {len(cnpjs):>4}")

# ============================================================
# FASE 2: IDENTIFICAR VENDAS FEV/25
# ============================================================
print(f"\n[3/7] Identificando vendas FEV/25 da CARTEIRA...")

fev_buyers = []
for cnpj_s, info in cart_data.items():
    fat = info.get('fat_fev', 0)
    if fat and isinstance(fat, (int, float)) and fat > 0:
        fev_buyers.append(cnpj_s)

print(f"  Clientes com faturamento FEV/25: {len(fev_buyers)}")

# ============================================================
# FASE 3: MONTAR POOLS FEV
# ============================================================
print(f"\n[4/7] Montando pools FEV/25...")

# --- Pool 1: Pós-venda JAN (D+15, D+30 dos 32 vendedores JAN) ---
pool_pv_jan = jan_by_resultado.get("SUPORTE", [])[:31]  # Últimos em SUPORTE = pós-venda
# Também os que terminaram em VENDA/PEDIDO
pool_pv_jan += jan_by_resultado.get("VENDA / PEDIDO", [])
pool_pv_jan = list(set(pool_pv_jan))
print(f"  Pool PV-JAN (pós-venda continuação): {len(pool_pv_jan)}")

# --- Pool 2: Negociações abertas JAN (EM ATENDIMENTO + ORÇAMENTO) ---
# Pegar apenas os top 50 (Daiane não consegue fazer follow-up em 250 clientes)
pool_neg_jan = []
for res in ["ORÇAMENTO", "EM ATENDIMENTO"]:
    cnpjs = jan_by_resultado.get(res, [])
    random.shuffle(cnpjs)
    pool_neg_jan.extend(cnpjs[:25])  # Max 25 por tipo
pool_neg_jan = list(set(pool_neg_jan))
print(f"  Pool NEG-JAN (negociações abertas): {len(pool_neg_jan)}")

# --- Pool 3: Retry tentativas JAN (NÃO RESPONDE + NÃO ATENDE) ---
pool_retry_jan = []
for res in ["NÃO RESPONDE", "NÃO ATENDE"]:
    cnpjs = jan_by_resultado.get(res, [])
    random.shuffle(cnpjs)
    pool_retry_jan.extend(cnpjs[:15])  # Max 15 por tipo
pool_retry_jan = list(set(pool_retry_jan))
print(f"  Pool RETRY-JAN (tentativas pendentes): {len(pool_retry_jan)}")

# --- Pool 4: Vendas FEV (ciclo completo) ---
# Separar: quais já tiveram contato em JAN vs novos
pool_vendas_fev_from_jan = [c for c in fev_buyers if c in jan_last_by_cnpj]
pool_vendas_fev_new = [c for c in fev_buyers if c not in jan_last_by_cnpj]
print(f"  Pool VENDAS-FEV (total): {len(fev_buyers)}")
print(f"    - Vindos de JAN (já tiveram contato): {len(pool_vendas_fev_from_jan)}")
print(f"    - Novos (sem contato JAN): {len(pool_vendas_fev_new)}")

# --- Pool 5: Prospects novos (~5/dia × 20 dias = ~100) ---
# Pegar QUALQUER cliente da CARTEIRA que NÃO foi atendido em JAN
all_jan_cnpjs = set(jan_last_by_cnpj.keys())
all_fev_buyers = set(fev_buyers)
all_pools_used = set(pool_pv_jan + pool_neg_jan + pool_retry_jan + fev_buyers)
prospects_available = []
for cnpj_s, info in cart_data.items():
    if cnpj_s in all_jan_cnpjs:
        continue  # Já atendido em JAN
    if cnpj_s in all_pools_used:
        continue  # Já em outro pool FEV
    prospects_available.append(cnpj_s)

random.shuffle(prospects_available)
pool_prospects = prospects_available[:100]  # ~5/dia × 20 dias
print(f"  Pool PROSPECTS-FEV: {len(pool_prospects)} (de {len(prospects_available)} disponíveis)")

# ============================================================
# FASE 4: SIMULAR DIA A DIA
# ============================================================
print(f"\n[5/7] Simulando atendimentos FEV/25...")

FEV_START = date(2025, 2, 3)  # Primeiro dia útil de FEV
FEV_END = date(2025, 2, 28)
du_fev = dias_uteis_lista(FEV_START, FEV_END)
print(f"  Dias úteis FEV/25: {len(du_fev)} ({FEV_START.strftime('%d/%m')} a {FEV_END.strftime('%d/%m')})")

registros_novos = []
acoes_futuras_mar = []

def get_info(cnpj_s):
    """Busca info do cliente na CARTEIRA ou nos dados JAN."""
    info = cart_data.get(cnpj_s, {})
    jan_info = jan_last_by_cnpj.get(cnpj_s, {})
    return {
        'nome': info.get('nome') or jan_info.get('nome', ''),
        'uf': info.get('uf') or jan_info.get('uf', ''),
        'rede': info.get('rede') or jan_info.get('rede', ''),
        'situacao': info.get('situacao', 'PROSPECT'),
        'tipo_cliente': info.get('tipo_cliente', 'NOVO'),
        'ciclo_medio': info.get('ciclo_medio', 0),
        'dias_compra': info.get('dias_compra', ''),
    }

# Distribuir pools nos dias úteis
# Lógica: cada dia tem ~20-25 atendimentos
# - 2-3 pós-venda JAN
# - 3-4 negociações JAN
# - 1-2 retries JAN
# - Vendas FEV distribuídas ao longo do mês
# - 5 prospecções

# ---- Pool 1: PÓS-VENDA JAN (D+15 e D+30) ----
# D+15 = ~15 dias após venda JAN (cai na 1ª-2ª semana de FEV)
# D+30 = ~30 dias após venda JAN (cai na 3ª-4ª semana de FEV)
pv_d15_dias = du_fev[0:8]   # Primeira metade de FEV
pv_d30_dias = du_fev[8:18]  # Segunda metade de FEV

random.shuffle(pool_pv_jan)
pv_list = list(pool_pv_jan)

for i, cnpj_s in enumerate(pv_list):
    info = get_info(cnpj_s)
    # D+15 — pós-venda qualidade
    d15 = pv_d15_dias[i % len(pv_d15_dias)]
    registros_novos.append(make_record(
        data=d15, consultor=CONSULTOR_FEV, cnpj=cnpj_s,
        nome=info['nome'], uf=info['uf'], rede=info['rede'],
        situacao="NOVO", dias_compra=15 + random.randint(0, 5),
        estagio_funil="PÓS-VENDA", tipo_cliente="NOVO", fase="PÓS-VENDA",
        whatsapp=0, ligacao=1, lig_atendida=1,
        tipo_contato="PÓS-VENDA / RELACIONAMENTO", resultado="PÓS-VENDA",
        motivo=random.choice(MOTIVOS_POS_VENDA),
        followup_date=proximo_dia_util(d15, 10),
        acao_futura="FAZER CS (SUCESSO DO CLIENTE)",
        acao_detalhada="Verificar se recebeu tudo certo, satisfação",
        mercos="SIM", nota=random.choice(NOTAS_POS_VENDA),
        temperatura="🔥QUENTE", tipo_problema=random.choice(TIPOS_PROBLEMA),
        demanda=random.choice(DEMANDAS_POS_VENDA),
        tipo_atendimento="ATEND. CLIENTES ATIVOS"
    ))

    # D+30 — CS / Sucesso do Cliente (se cabe em FEV)
    if i < len(pv_d30_dias):
        d30 = pv_d30_dias[i % len(pv_d30_dias)]
        registros_novos.append(make_record(
            data=d30, consultor=CONSULTOR_FEV, cnpj=cnpj_s,
            nome=info['nome'], uf=info['uf'], rede=info['rede'],
            situacao="NOVO", dias_compra=30 + random.randint(0, 5),
            estagio_funil="CS / RECOMPRA", tipo_cliente="NOVO", fase="CS",
            whatsapp=1, ligacao=1, lig_atendida=1,
            tipo_contato="PÓS-VENDA / RELACIONAMENTO",
            resultado="CS (SUCESSO DO CLIENTE)",
            motivo="NÃO SE APLICA",
            followup_date=proximo_dia_util(d30, 20),
            acao_futura="VERIFICAR SELL OUT E CRIAR INTENÇÃO RECOMPRA",
            acao_detalhada="CS completo, verificar sell out e recompra",
            mercos="SIM", nota=random.choice(NOTAS_POS_VENDA),
            temperatura="🔥QUENTE", tipo_problema="",
            demanda=random.choice(DEMANDAS_POS_VENDA),
            tipo_atendimento="ATEND. CLIENTES ATIVOS"
        ))
        # Gerar ação futura MAR
        acoes_futuras_mar.append({'cnpj': cnpj_s, 'tipo': 'RECOMPRA'})

print(f"  Pós-venda JAN: {len(pv_list)} clientes → {len(pv_list) * 2} registros")

# ---- Pool 2: NEGOCIAÇÕES JAN (follow-up) ----
neg_per_day = max(1, len(pool_neg_jan) // len(du_fev))
neg_idx = 0
for dia_idx, d in enumerate(du_fev):
    batch = pool_neg_jan[neg_idx:neg_idx + neg_per_day + (1 if dia_idx < len(pool_neg_jan) % len(du_fev) else 0)]
    neg_idx += len(batch)
    for cnpj_s in batch:
        info = get_info(cnpj_s)
        jan_res = jan_last_by_cnpj.get(cnpj_s, {}).get('resultado', 'EM ATENDIMENTO')

        # Sortear evolução: 40% avança, 30% mantém, 20% não responde, 10% perda
        rand = random.random()
        if rand < 0.40:
            # Avança no funil
            if jan_res == "EM ATENDIMENTO":
                resultado = "ORÇAMENTO"
                motivo = random.choice(MOTIVOS_ORCAMENTO)
                nota = random.choice(NOTAS_NEGOCIACAO)
                demanda = random.choice(DEMANDAS_ORCAMENTO)
                acao = "CONFIRMAR ORÇAMENTO ENVIADO"
            else:  # ORÇAMENTO
                resultado = "EM ATENDIMENTO"
                motivo = random.choice(MOTIVOS_EM_ATENDIMENTO)
                nota = random.choice(NOTAS_FOLLOWUP)
                demanda = random.choice(DEMANDAS_FOLLOWUP)
                acao = "FECHAR NEGOCIAÇÃO EM ANDAMENTO"
        elif rand < 0.70:
            # Mantém posição
            resultado = "EM ATENDIMENTO"
            motivo = random.choice(MOTIVOS_EM_ATENDIMENTO)
            nota = random.choice(NOTAS_FOLLOWUP)
            demanda = random.choice(DEMANDAS_FOLLOWUP)
            acao = "FECHAR NEGOCIAÇÃO EM ANDAMENTO"
        elif rand < 0.90:
            # Não responde
            resultado = random.choice(["NÃO RESPONDE", "NÃO ATENDE"])
            motivo = random.choice(MOTIVOS_NAO_RESPONDE if "RESPONDE" in resultado else MOTIVOS_NAO_ATENDE)
            nota = random.choice(NOTAS_NAO_RESPONDE if "RESPONDE" in resultado else NOTAS_NAO_ATENDE)
            demanda = random.choice(DEMANDAS_FOLLOWUP + [""])
            acao = ACAO_FUTURA_MAP[resultado]
        else:
            # Perda
            resultado = "PERDA / FECHOU LOJA"
            motivo = random.choice(MOTIVOS_PERDA)
            nota = random.choice(NOTAS_PERDA)
            demanda = ""
            acao = ACAO_FUTURA_MAP[resultado]

        canal = random.choice(['whatsapp', 'ligacao', 'ligacao'])
        registros_novos.append(make_record(
            data=d, consultor=CONSULTOR_FEV, cnpj=cnpj_s,
            nome=info['nome'], uf=info['uf'], rede=info['rede'],
            situacao=info['situacao'], dias_compra=info['dias_compra'],
            estagio_funil="EM ATENDIMENTO" if resultado in ("EM ATENDIMENTO", "NÃO RESPONDE", "NÃO ATENDE") else (
                "ORÇAMENTO" if resultado == "ORÇAMENTO" else "PERDA / NUTRIÇÃO"),
            tipo_cliente=info['tipo_cliente'], fase="PROSPECÇÃO",
            whatsapp=1 if canal == 'whatsapp' else 0,
            ligacao=1 if canal == 'ligacao' else 0,
            lig_atendida=1 if resultado not in ("NÃO ATENDE", "NÃO RESPONDE") else 0,
            tipo_contato="FOLLOW UP",
            resultado=resultado, motivo=motivo,
            followup_date=proximo_dia_util(d, RESULTADO_REGRAS.get(resultado, (3,))[0]),
            acao_futura=acao, acao_detalhada="",
            mercos="SIM" if resultado in ("ORÇAMENTO",) else "NÃO",
            nota=nota, temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "", ""))[2],
            tipo_problema="", demanda=demanda,
            tipo_atendimento="FOLLOW UP"
        ))

print(f"  Negociações JAN: {len(pool_neg_jan)} follow-ups")

# ---- Pool 3: RETRY (tentativas pendentes JAN) ----
retry_per_day = max(1, len(pool_retry_jan) // len(du_fev))
retry_idx = 0
for dia_idx, d in enumerate(du_fev[:10]):  # Distribuir nas 2 primeiras semanas
    batch = pool_retry_jan[retry_idx:retry_idx + retry_per_day + 1]
    retry_idx += len(batch)
    if retry_idx >= len(pool_retry_jan):
        break
    for cnpj_s in batch:
        info = get_info(cnpj_s)
        # 50% consegue contato → EM ATENDIMENTO, 30% não responde de novo, 20% perda
        rand = random.random()
        if rand < 0.50:
            resultado = "EM ATENDIMENTO"
            motivo = random.choice(MOTIVOS_EM_ATENDIMENTO)
            nota = random.choice(NOTAS_FOLLOWUP)
            demanda = random.choice(DEMANDAS_FOLLOWUP)
        elif rand < 0.80:
            resultado = random.choice(["NÃO RESPONDE", "NÃO ATENDE", "RECUSOU LIGAÇÃO"])
            motivo = random.choice(MOTIVOS_NAO_RESPONDE)
            nota = random.choice(NOTAS_NAO_RESPONDE)
            demanda = ""
        else:
            resultado = "PERDA / FECHOU LOJA"
            motivo = random.choice(MOTIVOS_PERDA)
            nota = random.choice(NOTAS_PERDA)
            demanda = ""

        registros_novos.append(make_record(
            data=d, consultor=CONSULTOR_FEV, cnpj=cnpj_s,
            nome=info['nome'], uf=info['uf'], rede=info['rede'],
            situacao=info['situacao'], dias_compra=info['dias_compra'],
            estagio_funil="PROSPECÇÃO" if resultado in ("NÃO RESPONDE", "NÃO ATENDE", "RECUSOU LIGAÇÃO", "PERDA / FECHOU LOJA") else "EM ATENDIMENTO",
            tipo_cliente=info['tipo_cliente'], fase="TENTATIVA",
            whatsapp=0, ligacao=1, lig_atendida=1 if resultado == "EM ATENDIMENTO" else 0,
            tipo_contato="FOLLOW UP", resultado=resultado, motivo=motivo,
            followup_date=proximo_dia_util(d, RESULTADO_REGRAS.get(resultado, (2,))[0]) if resultado not in ("PERDA / FECHOU LOJA",) else "",
            acao_futura=ACAO_FUTURA_MAP.get(resultado, ""),
            acao_detalhada="", mercos="NÃO",
            nota=nota, temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "❄️FRIO", ""))[2],
            tipo_problema="", demanda=demanda,
            tipo_atendimento="FOLLOW UP"
        ))

print(f"  Retries JAN: {min(retry_idx, len(pool_retry_jan))} tentativas")

# ---- Pool 4: VENDAS FEV (ciclo completo) ----
# Distribuir nos primeiros 15 dias úteis para dar tempo de fechar ciclo
du_vendas = du_fev[:15]  # Até ~20 de fevereiro

for i, cnpj_s in enumerate(fev_buyers):
    info = get_info(cnpj_s)
    in_jan = cnpj_s in jan_last_by_cnpj

    # Dia base: distribuir ao longo dos 15 dias
    d_base = du_vendas[i % len(du_vendas)]
    gap = random.choice([1, 2, 2, 3])

    if in_jan:
        # Já teve contato em JAN → continua de onde parou
        jan_res = jan_last_by_cnpj[cnpj_s]['resultado']
        if jan_res in ("ORÇAMENTO", "EM ATENDIMENTO"):
            # Direto para CADASTRO → VENDA
            steps = ["CADASTRO", "VENDA / PEDIDO"]
        else:
            steps = ["EM ATENDIMENTO", "ORÇAMENTO", "CADASTRO", "VENDA / PEDIDO"]
    else:
        # Novo → ciclo completo
        steps = ["EM ATENDIMENTO", "ORÇAMENTO", "CADASTRO", "VENDA / PEDIDO"]

    d_current = d_base
    for step_idx, step in enumerate(steps):
        if step_idx > 0:
            d_current = proximo_dia_util(d_current, gap)
        # Forçar dentro de FEV
        if d_current > FEV_END:
            d_current = FEV_END
        if d_current.weekday() >= 5:
            d_current = proximo_dia_util(d_current - timedelta(days=1), 1)

        if step == "EM ATENDIMENTO":
            resultado = "EM ATENDIMENTO"
            motivo = random.choice(MOTIVOS_EM_ATENDIMENTO)
            nota = random.choice(NOTAS_NEGOCIACAO)
            demanda = random.choice(DEMANDAS_PROSPECÇÃO) if not in_jan else random.choice(DEMANDAS_FOLLOWUP)
            acao = "ENVIAR ORÇAMENTO / CATÁLOGO"
            estagio = "EM ATENDIMENTO"
            tc = "PROSPECÇÃO" if not in_jan else "NEGOCIAÇÃO"
        elif step == "ORÇAMENTO":
            resultado = "ORÇAMENTO"
            motivo = random.choice(MOTIVOS_ORCAMENTO)
            nota = random.choice(NOTAS_NEGOCIACAO)
            demanda = random.choice(DEMANDAS_ORCAMENTO)
            acao = "CONFIRMAR ORÇAMENTO ENVIADO"
            estagio = "ORÇAMENTO"
            tc = "NEGOCIAÇÃO"
        elif step == "CADASTRO":
            resultado = "CADASTRO"
            motivo = random.choice(MOTIVOS_CADASTRO)
            nota = "Cadastro enviado para análise. Aguardando liberação."
            demanda = random.choice(DEMANDAS_CADASTRO)
            acao = "CONFIRMAR CADASTRO NO SISTEMA"
            estagio = "EM ATENDIMENTO"
            tc = "NEGOCIAÇÃO"
        else:  # VENDA
            resultado = "VENDA / PEDIDO"
            motivo = random.choice(MOTIVOS_VENDA)
            nota = random.choice(NOTAS_VENDA)
            demanda = random.choice(DEMANDAS_PEDIDO)
            acao = "CONFIRMAR FATURAMENTO E EXPEDIÇÃO"
            estagio = "PÓS-VENDA"
            tc = "NEGOCIAÇÃO"

        registros_novos.append(make_record(
            data=d_current, consultor=CONSULTOR_FEV, cnpj=cnpj_s,
            nome=info['nome'], uf=info['uf'], rede=info['rede'],
            situacao=info['situacao'] if info['situacao'] else "PROSPECT",
            dias_compra=info['dias_compra'],
            estagio_funil=estagio, tipo_cliente=info['tipo_cliente'] or "NOVO",
            fase="PROSPECÇÃO" if step == "EM ATENDIMENTO" and not in_jan else estagio,
            whatsapp=1 if step_idx == 0 else 0,
            ligacao=1 if step_idx > 0 else 0,
            lig_atendida=1,
            tipo_contato=tc, resultado=resultado, motivo=motivo,
            followup_date=proximo_dia_util(d_current, gap),
            acao_futura=acao, acao_detalhada="",
            mercos="SIM" if step in ("ORÇAMENTO", "VENDA / PEDIDO") else "NÃO",
            nota=nota, temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "🟡MORNO", ""))[2],
            tipo_problema="", demanda=demanda,
            tipo_atendimento="PROSPECÇÃO" if not in_jan and step_idx == 0 else "FOLLOW UP",
            ciclo_medio=info.get('ciclo_medio', 0)
        ))

    # Pós-venda D+4 (se cabe em FEV)
    d_pv = proximo_dia_util(d_current, 4)
    if d_pv <= FEV_END:
        registros_novos.append(make_record(
            data=d_pv, consultor=CONSULTOR_FEV, cnpj=cnpj_s,
            nome=info['nome'], uf=info['uf'], rede=info['rede'],
            situacao="NOVO", dias_compra=4,
            estagio_funil="PÓS-VENDA", tipo_cliente="NOVO", fase="PÓS-VENDA",
            whatsapp=0, ligacao=1, lig_atendida=1,
            tipo_contato="PÓS-VENDA / RELACIONAMENTO", resultado="PÓS-VENDA",
            motivo="NÃO SE APLICA",
            followup_date=proximo_dia_util(d_pv, 11),
            acao_futura="FAZER PÓS-VENDA",
            acao_detalhada="Confirmar faturamento e envio",
            mercos="SIM", nota=random.choice(NOTAS_POS_VENDA),
            temperatura="🔥QUENTE", tipo_problema="",
            demanda=random.choice(DEMANDAS_POS_VENDA),
            tipo_atendimento="ATEND. CLIENTES ATIVOS"
        ))
    else:
        acoes_futuras_mar.append({'cnpj': cnpj_s, 'tipo': 'POS-VENDA D+4'})

print(f"  Vendas FEV: {len(fev_buyers)} clientes → ciclo completo")

# ---- Pool 5: PROSPECÇÕES NOVAS (~5/dia) ----
prospect_per_day = 5
prospect_idx = 0
for d in du_fev:
    batch = pool_prospects[prospect_idx:prospect_idx + prospect_per_day]
    prospect_idx += len(batch)
    for cnpj_s in batch:
        info = get_info(cnpj_s)
        # Primeiro contato → 60% EM ATENDIMENTO, 25% NÃO RESPONDE, 15% NÃO ATENDE
        rand = random.random()
        if rand < 0.60:
            resultado = "EM ATENDIMENTO"
            motivo = random.choice(MOTIVOS_EM_ATENDIMENTO)
            nota = random.choice(NOTAS_PROSPECÇÃO)
            demanda = random.choice(DEMANDAS_PROSPECÇÃO)
        elif rand < 0.85:
            resultado = "NÃO RESPONDE"
            motivo = random.choice(MOTIVOS_NAO_RESPONDE)
            nota = random.choice(NOTAS_NAO_RESPONDE)
            demanda = random.choice(DEMANDAS_PROSPECÇÃO + [""])
        else:
            resultado = "NÃO ATENDE"
            motivo = random.choice(MOTIVOS_NAO_ATENDE)
            nota = random.choice(NOTAS_NAO_ATENDE)
            demanda = random.choice(DEMANDAS_PROSPECÇÃO + [""])

        registros_novos.append(make_record(
            data=d, consultor=CONSULTOR_FEV, cnpj=cnpj_s,
            nome=info['nome'], uf=info['uf'], rede=info['rede'],
            situacao="PROSPECT", dias_compra="",
            estagio_funil="PROSPECÇÃO", tipo_cliente="PROSPECT",
            fase="PROSPECÇÃO",
            whatsapp=1, ligacao=0, lig_atendida=0,
            tipo_contato="PROSPECÇÃO", resultado=resultado, motivo=motivo,
            followup_date=proximo_dia_util(d, RESULTADO_REGRAS.get(resultado, (2,))[0]),
            acao_futura=ACAO_FUTURA_MAP.get(resultado, "ENVIAR ORÇAMENTO / CATÁLOGO"),
            acao_detalhada="", mercos="NÃO",
            nota=nota, temperatura=RESULTADO_REGRAS.get(resultado, (0, "", "❄️FRIO", ""))[2],
            tipo_problema="",
            demanda=demanda,
            tipo_atendimento="PROSPECÇÃO"
        ))

print(f"  Prospecções novas: {prospect_idx} (5/dia × {len(du_fev)} dias)")

print(f"\n  TOTAL registros FEV/25 gerados: {len(registros_novos)}")
print(f"  Ações futuras MAR+: {len(acoes_futuras_mar)}")

# ============================================================
# FASE 5: VERIFICAÇÃO LÓGICA
# ============================================================
print(f"\n[6/7] Verificação lógica...")

# Contar resultados
res_count = Counter()
for rec in registros_novos:
    res_count[rec['R']] += 1

for res, n in res_count.most_common():
    print(f"    {res:<35} {n:>4}")

# Verificar funil
n_orc = res_count.get("ORÇAMENTO", 0)
n_cad = res_count.get("CADASTRO", 0)
n_ven = res_count.get("VENDA / PEDIDO", 0)
print(f"\n  Funil: ORÇAMENTO ({n_orc}) >= CADASTRO ({n_cad}) >= VENDA ({n_ven})")
if n_orc >= n_cad >= n_ven:
    print(f"  ✅ CONSISTENTE")
else:
    print(f"  ⚠️ VERIFICAR — possível inconsistência")

# Atendimentos por dia
dia_count = Counter()
for rec in registros_novos:
    d = rec['A']
    if isinstance(d, date):
        dia_count[d] += 1
print(f"\n  Média/dia: {sum(dia_count.values()) / max(len(dia_count), 1):.1f}")

# Preenchimento de campos
n_total = len(registros_novos)
n_motivo = sum(1 for r in registros_novos if r['S'])
n_demanda = sum(1 for r in registros_novos if r['AD'])
n_acao = sum(1 for r in registros_novos if r['U'])
n_nota = sum(1 for r in registros_novos if r['X'])
print(f"\n  Preenchimento:")
print(f"    MOTIVO:    {n_motivo:>5} ({n_motivo/n_total*100:.0f}%)")
print(f"    DEMANDA:   {n_demanda:>5} ({n_demanda/n_total*100:.0f}%)")
print(f"    AÇÃO:      {n_acao:>5} ({n_acao/n_total*100:.0f}%)")
print(f"    NOTA:      {n_nota:>5} ({n_nota/n_total*100:.0f}%)")

# ============================================================
# FASE 6: RECONSTRUIR DRAFT 2
# ============================================================
print(f"\n[7/7] Salvando V28...")

# Ler registros existentes — REMOVER TODOS os FEV/25
# (Em FEV só tinha DAIANE; LARISSA/MANU/HELDER/JULIO não existiam ainda)
COLS = ['A','B','C','D','E','F','G','H','I','J','K','L','M',
        'N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
        'AA','AB','AC','AD','AE']
COL_MAP = {c: i for i, c in enumerate(COLS)}

existing_records = []
removed_fev = 0
for r in range(3, ws_d2.max_row + 1):
    data_cell = ws_d2.cell(row=r, column=1).value
    if data_cell is None:
        continue

    # Verificar se é FEV/25 (QUALQUER consultor — remover TODOS)
    is_fev = False
    if isinstance(data_cell, (int, float)):
        dt = datetime(1900, 1, 1) + timedelta(days=int(data_cell) - 2)
        if dt.year == 2025 and dt.month == 2:
            is_fev = True
    elif isinstance(data_cell, datetime):
        if data_cell.year == 2025 and data_cell.month == 2:
            is_fev = True

    if is_fev:
        removed_fev += 1
        continue

    row_data = {}
    for c_idx, col_letter in enumerate(COLS):
        row_data[col_letter] = ws_d2.cell(row=r, column=c_idx + 1).value
    existing_records.append(row_data)

print(f"  Registros existentes mantidos: {len(existing_records)}")
print(f"  Registros FEV removidos (todos consultores): {removed_fev}")
print(f"  Registros FEV DAIANE novos: {len(registros_novos)}")
print(f"  (FEV = só DAIANE; Helder/Manu em onboarding, Larissa ago, Julio out)")

# Combinar e ordenar por data
all_records = existing_records + registros_novos
def sort_key(rec):
    d = rec.get('A')
    if d is None:
        return datetime(1900, 1, 1)
    if isinstance(d, (int, float)):
        return datetime(1900, 1, 1) + timedelta(days=int(d) - 2)
    if isinstance(d, date) and not isinstance(d, datetime):
        return datetime(d.year, d.month, d.day)
    return d if isinstance(d, datetime) else datetime(1900, 1, 1)

all_records.sort(key=sort_key)
print(f"  Total DRAFT 2: {len(all_records)} registros")

# Calcular TENTATIVA (col M) por CNPJ
cnpj_counter = Counter()
for rec in all_records:
    cnpj = rec.get('D')
    if cnpj:
        cnpj_s = str(cnpj).strip()
        cnpj_counter[cnpj_s] += 1
        rec['M'] = f"T{cnpj_counter[cnpj_s]}"

print(f"  Calculando TENTATIVA (col M)...")

# Limpar DRAFT 2 e reescrever
for r in range(3, ws_d2.max_row + 1):
    for c in range(1, len(COLS) + 1):
        ws_d2.cell(row=r, column=c).value = None

for i, rec in enumerate(all_records):
    row = i + 3
    for c_idx, col_letter in enumerate(COLS):
        ws_d2.cell(row=row, column=c_idx + 1).value = rec.get(col_letter)

# Salvar
print(f"\n  Salvando {OUTPUT_FILE}...")
wb.save(OUTPUT_FILE)
file_size = os.path.getsize(OUTPUT_FILE) / (1024 * 1024)

# ============================================================
# RELATÓRIO FINAL
# ============================================================
print(f"\n{'='*80}")
print(f"V28 CONCLUÍDO!")
print(f"{'='*80}")
print(f"  Arquivo: {OUTPUT_FILE}")
print(f"  Tamanho: {file_size:.2f} MB")
print(f"\n  DRAFT 2:")
print(f"    Total: {len(all_records)} registros")
print(f"    FEV/25 DAIANE novos: {len(registros_novos)}")
print(f"    Mantidos (outros): {len(existing_records)}")
print(f"\n  Composição FEV/25 DAIANE ({len(registros_novos)} registros):")
print(f"    Pós-venda JAN: {len(pv_list) * 2}")
print(f"    Negociações JAN: {len(pool_neg_jan)}")
print(f"    Retries JAN: {min(retry_idx, len(pool_retry_jan))}")
print(f"    Vendas FEV ciclo: ~{len(fev_buyers) * 5}")
print(f"    Prospecções novas: {prospect_idx}")
print(f"\n  Vendas FEV: {n_ven} de {len(fev_buyers)} esperadas")
print(f"  Média atendimentos/dia: {sum(dia_count.values()) / max(len(dia_count), 1):.1f}")
print(f"  Ações futuras MAR+: {len(acoes_futuras_mar)}")

# Distribuição por dia
print(f"\n  Distribuição por dia FEV:")
for d in sorted(dia_count.keys()):
    dia_nome = ['SEG','TER','QUA','QUI','SEX','SAB','DOM'][d.weekday()]
    barra = '█' * (dia_count[d] // 2)
    print(f"    {d.day:>2}/02 ({dia_nome}): {dia_count[d]:>3} {barra}")

# DEMANDAS
demanda_count = Counter()
for rec in registros_novos:
    if rec['AD']:
        demanda_count[rec['AD']] += 1
print(f"\n  DEMANDAS geradas:")
for dem, n in demanda_count.most_common():
    print(f"    {dem:<45} {n:>4}")

print(f"{'='*80}")

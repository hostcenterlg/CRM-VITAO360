#!/usr/bin/env python3
"""
BUILD V24 — Incorporar vendas JAN/25 no CRM VITAO360
=====================================================
Adiciona coluna JAN/25 no DRAFT 1, CARTEIRA e PROJEÇÃO.
- Mapeia SAP codes -> CNPJs via CARTEIRA col 63
- Insere coluna no DRAFT 1 (pos 21) e CARTEIRA (pos 26)
- Corrige TODAS as fórmulas cross-sheet afetadas
- Escreve valores JAN/25 na PROJEÇÃO (REAL JAN col 27)
"""

import openpyxl
import re
import time
import os
from copy import copy
from openpyxl.utils import get_column_letter, column_index_from_string

# ============================================================
# CONFIGURAÇÃO
# ============================================================
PYTHON_PATH = "C:/Users/User/.pyenv/pyenv-win/pyenv-win/shims/python3"
INPUT_FILE = "c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V23_FINAL.xlsx"
OUTPUT_FILE = "c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V24_FINAL.xlsx"

# JAN/25 Vendas — 33 registros (SAP_CODE, NOME, QTD, BRUTO, DEVOL, LIQUIDO)
JAN25_DATA = [
    ("1000069469", "ATLAS BRASIL COMERCIALIZADORA DE EN", 25, 5530.63, 0.00, 5530.63),
    ("1000109795", "MERCANTIL SALES LTDA - ME", 32, 5364.79, 0.00, 5364.79),
    ("1000084174", "JOAO COSTA SOARES E CIA LTDA", 38, 4212.18, 0.00, 4212.18),
    ("1000064761", "SIMPLES E SAUDAVEL COMERCIO PRODUTO", 33, 3845.66, 0.00, 3845.66),
    ("1000084173", "JOAO COSTA SOARES E CIA LTDA", 33, 3007.07, 0.00, 3007.07),
    ("1000003604", "ASSOCIACAO DE PAIS E AMIGOS DOS", 24, 2760.00, 0.00, 2760.00),
    ("1000091117", "DISTRIBUIDORA LIGEIRINHO DE PRODUTO", 19, 2358.74, 0.00, 2358.74),
    ("1000053620", "SHOPPER COMERCIO ALIMENTICIO LTDA", 38, 1721.40, 0.00, 1721.40),
    ("1000089643", "BRCOMBOS ALIMENTACAO LTDA - ME", 12, 1534.00, 0.00, 1534.00),
    ("1000095716", "JANAINA GREICE VANDERLINDE - ME", 3, 1360.80, 0.00, 1360.80),
    ("1000067141", "SR A GRANEL - PRODUTOS NATURAIS LTD", 13, 2490.71, -1237.32, 1253.39),
    ("1000110140", "CEREJEIRA EMPORIO DE ALIMENTOS SAUD", 11, 1197.84, 0.00, 1197.84),
    ("1000101003", "PIQUINIQUE FRESH FOOD LTDA - ME", 8, 1104.00, 0.00, 1104.00),
    ("1000102716", "BEM ESTAR NATURAL LTDA - ME", 8, 772.12, 0.00, 772.12),
    ("1000105324", "58.188.486 TERESINHA PEREIRA DA SIL", 5, 696.90, 0.00, 696.90),
    ("1000082463", "ROBERTO CARLOS ABBOUD LTDA - ME", 12, 652.32, 0.00, 652.32),
    ("1000077962", "RISOTOLANDIA INDUSTRIA E COMERCIO D", 8, 595.20, 0.00, 595.20),
    ("1000095786", "LM NUTRICAO BOMBINHAS LTDA - ME", 2, 577.80, 0.00, 577.80),
    ("1000059766", "HOTEL CURITIBA CAPITAL S/A", 2, 480.00, 0.00, 480.00),
    ("1000105724", "GMJ EMPORIO SAUDAVEL LTDA", 1, 475.90, 0.00, 475.90),
    ("1000095512", "SUPLEMENT WORLD SHOP LTDA - ME", 1, 460.80, 0.00, 460.80),
    ("1000097246", "FIT NUTRICAO LTDA -ME", 4, 448.40, 0.00, 448.40),
    ("1000093167", "FIT4SHAPE NUTRICAO LTDA - EPP", 4, 389.85, 0.00, 389.85),
    ("1000095633", "L & M NUTRICAO LTDA - ME", 4, 386.06, 0.00, 386.06),
    ("1000101126", "ECAMP PRODUTOS NATURAIS E SUPLEMENT", 4, 384.07, 0.00, 384.07),
    ("1000103200", "EPERCAM PRODUTOS NATURAIS E SUPLEME", 4, 384.07, 0.00, 384.07),
    ("1000104522", "NATURELI PRODUTOS NATURAIS LTDA - M", 1, 379.94, 0.00, 379.94),
    ("1000104075", "EMPORIO NUTREE GRAOS LTDA - ME", 3, 287.76, 0.00, 287.76),
    ("1000101541", "BIO SC COMERCIO DE ALIMENTOS LTDA -", 1, 194.70, 0.00, 194.70),
    ("1000094553", "RK COMERCIO DE NUTRICAO LTDA - EPP", 1, 192.60, 0.00, 192.60),
    ("1000097251", "MANIA FIT LTDA - EPP", 0, 128.27, 0.00, 128.27),
    ("1000108677", "TOP FIT COMERCIO DE NUTRICAO LTDA -", 0, 128.27, 0.00, 128.27),
    ("1000088432", "VEGSIM LTDA - ME", 0, 585.00, -585.00, 0.00),
]


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def shift_col_in_ref(col_str, min_col_idx, shift=1):
    """Incrementa coluna se >= min_col_idx."""
    idx = column_index_from_string(col_str)
    if idx >= min_col_idx:
        return get_column_letter(idx + shift)
    return col_str


def fix_sheet_refs_in_formula(formula, sheet_prefix, min_col_idx, shift=1):
    """
    Corrige referências de coluna em fórmulas para um sheet específico.

    Usa UM ÚNICO regex pass que trata ranges e singles juntos,
    evitando double-shift (bug anterior com 2 passes).

    Trata:
    - Ranges: SHEET!$COL$ROW:$COL$ROW
    - Singles: SHEET!$COL$ROW
    - Com e sem $ (absoluto/relativo)
    """
    if not formula or not isinstance(formula, str) or formula[0] != '=':
        return formula

    escaped = re.escape(sheet_prefix)

    # Pattern combinado: prefix + $?COL$?ROW + opcional(:$?COL$?ROW)
    combined_pat = rf"({escaped})(\$?)([A-Z]+)(\$?\d+)(?::(\$?)([A-Z]+)(\$?\d+))?"

    def replace_combined(m):
        pfx = m.group(1)
        d1 = m.group(2)
        col1 = m.group(3)
        rest1 = m.group(4)
        new1 = shift_col_in_ref(col1, min_col_idx, shift)
        result = f"{pfx}{d1}{new1}{rest1}"

        if m.group(6):  # tem parte de range (:$?COL$?ROW)
            d2 = m.group(5)
            col2 = m.group(6)
            rest2 = m.group(7)
            new2 = shift_col_in_ref(col2, min_col_idx, shift)
            result += f":{d2}{new2}{rest2}"

        return result

    return re.sub(combined_pat, replace_combined, formula)


def fix_local_range_refs(formula, min_col_idx, shift=1):
    """
    Corrige referências locais (sem prefixo de sheet) em fórmulas.
    Ex: =SUM(Z4:AJ4) → shift cols >= min_col_idx

    CUIDADO: Só usar para fórmulas que sabemos ser locais.
    """
    if not formula or not isinstance(formula, str) or formula[0] != '=':
        return formula

    # Procura padrões $?COL$?ROW que NÃO são precedidos por ! (sheet ref)
    def replace_col_ref(m):
        prefix_char = m.group(1)  # char before (could be anything)
        d = m.group(2)
        col = m.group(3)
        rest = m.group(4)
        new_col = shift_col_in_ref(col, min_col_idx, shift)
        return f"{prefix_char}{d}{new_col}{rest}"

    # Match column refs NOT preceded by ! (which would be sheet refs)
    pat = r"([^!A-Z])(\$?)([A-Z]+)(\$?\d+)"
    formula = re.sub(pat, replace_col_ref, formula)

    return formula


# ============================================================
# MAIN
# ============================================================

def main():
    start = time.time()

    def elapsed():
        return f"{time.time() - start:.1f}s"

    print(f"[{elapsed()}] Carregando V23...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    print(f"[{elapsed()}] V23 carregado. Sheets: {wb.sheetnames}")

    ws_d1 = wb['DRAFT 1']
    ws_cart = wb['CARTEIRA']
    ws_proj = wb['PROJEÇÃO ']  # note trailing space

    # ----------------------------------------------------------
    # ETAPA 1: Mapear SAP -> CNPJ -> rows
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 1: Mapeando SAP -> CNPJ...")

    # SAP map from CARTEIRA col 63 -> col 2
    sap_to_cnpj = {}
    for r in range(4, ws_cart.max_row + 1):
        sap_val = ws_cart.cell(r, 63).value
        cnpj_val = ws_cart.cell(r, 2).value
        if sap_val and cnpj_val:
            sap_to_cnpj[str(sap_val).strip()] = str(cnpj_val).strip()

    # CNPJ -> row maps (data_only needed for values)
    # Rebuild from formula workbook using stored values or cell values
    # For DRAFT 1: col 2 = CNPJ
    cnpj_to_d1_row = {}
    for r in range(4, ws_d1.max_row + 1):
        cnpj = ws_d1.cell(r, 2).value
        if cnpj:
            cnpj_to_d1_row[str(cnpj).strip()] = r

    # For CARTEIRA: col 2 = CNPJ
    cnpj_to_cart_row = {}
    for r in range(4, ws_cart.max_row + 1):
        cnpj = ws_cart.cell(r, 2).value
        if cnpj:
            cnpj_to_cart_row[str(cnpj).strip()] = r

    # For PROJEÇÃO: col 1 = CNPJ
    cnpj_to_proj_row = {}
    for r in range(4, ws_proj.max_row + 1):
        cnpj = ws_proj.cell(r, 1).value
        if cnpj:
            cnpj_to_proj_row[str(cnpj).strip()] = r

    # Map JAN25 data to CNPJs
    jan25_mapped = []
    unmapped = []
    for sap, nome, qtd, bruto, devol, liquido in JAN25_DATA:
        cnpj = sap_to_cnpj.get(sap)
        if cnpj:
            d1_row = cnpj_to_d1_row.get(cnpj)
            cart_row = cnpj_to_cart_row.get(cnpj)
            proj_row = cnpj_to_proj_row.get(cnpj)
            jan25_mapped.append({
                'sap': sap, 'nome': nome, 'cnpj': cnpj,
                'qtd': qtd, 'bruto': bruto, 'devol': devol, 'liquido': liquido,
                'd1_row': d1_row, 'cart_row': cart_row, 'proj_row': proj_row
            })
        else:
            unmapped.append(sap)

    print(f"  Mapeados: {len(jan25_mapped)}/33")
    if unmapped:
        print(f"  NÃO mapeados: {unmapped}")

    # ----------------------------------------------------------
    # ETAPA 2: Inserir coluna JAN/25 no DRAFT 1 (posição 21)
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 2: Inserindo coluna JAN/25 no DRAFT 1 (col 21)...")

    # Antes: col 21=MAR/25, col 32=FEV/26
    # Depois: col 21=JAN/25, col 22=MAR/25, col 33=FEV/26
    ws_d1.insert_cols(21)

    # Header
    ws_d1.cell(3, 21).value = "JAN/25"

    # Copiar formatação do col vizinho (22, ex-21 MAR/25)
    for r in range(1, 4):
        src = ws_d1.cell(r, 22)
        dst = ws_d1.cell(r, 21)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format

    # Escrever valores JAN/25 no DRAFT 1
    d1_written = 0
    for entry in jan25_mapped:
        if entry['d1_row'] and entry['liquido'] > 0:
            r = entry['d1_row']
            ws_d1.cell(r, 21).value = entry['liquido']
            # Copiar formato numérico
            src = ws_d1.cell(r, 22)
            dst = ws_d1.cell(r, 21)
            if src.has_style:
                dst.number_format = src.number_format
            d1_written += 1

    print(f"  DRAFT 1: {d1_written} valores escritos na col 21 (JAN/25)")

    # ----------------------------------------------------------
    # ETAPA 3: Inserir coluna JAN/25 na CARTEIRA (posição 26)
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 3: Inserindo coluna JAN/25 na CARTEIRA (col 26)...")

    # Antes: col 25=TOTAL, col 26=MAR/25(45717), col 36=JAN/26(46023)
    # Depois: col 25=TOTAL, col 26=JAN/25, col 27=MAR/25, col 37=JAN/26
    ws_cart.insert_cols(26)

    # Header: date serial para JAN/25 (2025-01-01 = 45658)
    # Excel serial: 2025-01-01 = DATE(2025,1,1)
    from datetime import datetime
    jan25_date = datetime(2025, 1, 1)
    ws_cart.cell(3, 26).value = jan25_date
    ws_cart.cell(2, 26).value = "VENDAS"

    # Copiar formatação do col vizinho (27, ex-26 MAR/25)
    for r in range(1, 4):
        src = ws_cart.cell(r, 27)
        dst = ws_cart.cell(r, 26)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format

    print(f"  CARTEIRA: col 26 inserida para JAN/25")

    # ----------------------------------------------------------
    # ETAPA 4: Corrigir TODAS as fórmulas afetadas
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 4: Corrigindo fórmulas cross-sheet...")

    # DRAFT 1 insert at 21: shift 'DRAFT 1' refs col >= U(21) by +1
    # CARTEIRA insert at 26: shift 'CARTEIRA' refs col >= Z(26) by +1

    draft1_prefix = "'DRAFT 1'!"
    carteira_prefix = "CARTEIRA!"

    fixes_draft1 = 0
    fixes_carteira = 0
    fixes_local = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_r = ws.max_row or 1
        max_c = ws.max_column or 1

        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(r, c)
                val = cell.value

                if not val or not isinstance(val, str) or not val.startswith('='):
                    continue

                original = val

                # Fix DRAFT 1 references (col shift >= 21/U)
                if "'DRAFT 1'" in val:
                    val = fix_sheet_refs_in_formula(val, draft1_prefix, 21)

                # Fix CARTEIRA references from OTHER sheets (col shift >= 26/Z)
                if sheet_name != 'CARTEIRA' and 'CARTEIRA!' in val:
                    val = fix_sheet_refs_in_formula(val, carteira_prefix, 26)

                # Fix CARTEIRA self-references (local refs >= Z/26)
                if sheet_name == 'CARTEIRA':
                    # Only for known local-ref formulas (TOTAL, etc.)
                    # The SUM formula at col 25 references local cols
                    if c == 25 and 'SUM(' in val:
                        # =SUM(Z4:AJ4) → need to shift AJ to AK
                        # Z stays at 26 (new JAN/25 - good!)
                        # AJ was col 36 (JAN/26), now at col 37 (AK)
                        val = fix_local_range_refs(val, 26)

                    # Also fix CARTEIRA refs that use CARTEIRA! prefix (explicit self-ref)
                    if 'CARTEIRA!' in val:
                        val = fix_sheet_refs_in_formula(val, carteira_prefix, 26)

                if val != original:
                    cell.value = val
                    if "'DRAFT 1'" in original:
                        fixes_draft1 += 1
                    if 'CARTEIRA!' in original:
                        fixes_carteira += 1
                    if sheet_name == 'CARTEIRA' and 'SUM(' in original:
                        fixes_local += 1

    print(f"  Fórmulas DRAFT 1 corrigidas: {fixes_draft1}")
    print(f"  Fórmulas CARTEIRA (cross-sheet): {fixes_carteira}")
    print(f"  Fórmulas CARTEIRA (local/TOTAL): {fixes_local}")

    # ----------------------------------------------------------
    # ETAPA 5: Escrever fórmulas JAN/25 na CARTEIRA col 26
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 5: Escrevendo fórmulas JAN/25 na CARTEIRA col 26...")

    # Para as rows 4-557 (clientes com fórmulas): escrever INDEX/MATCH
    # DRAFT 1 col 21 (U) agora é JAN/25
    # Fórmula: =IFERROR(INDEX('DRAFT 1'!$U$3:$U$25000,MATCH($B{r},'DRAFT 1'!$B$3:$B$25000,0)),"")

    cart_formulas = 0
    cart_values = 0

    for r in range(4, ws_cart.max_row + 1):
        # Verificar se a row tem fórmulas nos vendas (checar col 27, ex-26 MAR/25)
        neighbor_val = ws_cart.cell(r, 27).value

        if neighbor_val and isinstance(neighbor_val, str) and neighbor_val.startswith('='):
            # Row com fórmulas: escrever fórmula INDEX/MATCH
            formula = f"=IFERROR(INDEX('DRAFT 1'!$U$3:$U$25000,MATCH($B{r},'DRAFT 1'!$B$3:$B$25000,0)),\"\")"
            ws_cart.cell(r, 26).value = formula
            cart_formulas += 1
        else:
            # Row sem fórmulas (prospects): escrever valor direto se tiver
            cnpj = ws_cart.cell(r, 2).value
            if cnpj:
                cnpj_str = str(cnpj).strip()
                for entry in jan25_mapped:
                    if entry['cnpj'] == cnpj_str and entry['liquido'] > 0:
                        ws_cart.cell(r, 26).value = entry['liquido']
                        cart_values += 1
                        break
                else:
                    ws_cart.cell(r, 26).value = ""
            else:
                ws_cart.cell(r, 26).value = ""

    # Copiar formato numérico para col 26
    for r in range(4, ws_cart.max_row + 1):
        src = ws_cart.cell(r, 27)
        dst = ws_cart.cell(r, 26)
        if src.has_style:
            dst.number_format = src.number_format

    print(f"  CARTEIRA col 26: {cart_formulas} fórmulas + {cart_values} valores diretos")

    # ----------------------------------------------------------
    # ETAPA 6: Corrigir TOTAL PERÍODO (col 25)
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 6: Verificando TOTAL PERÍODO (col 25)...")

    # Após inserts: vendas agora vão de col 26 (Z=JAN/25) a col 37 (AK=JAN/26)
    # TOTAL deve ser =SUM(Z{r}:AK{r})
    total_fixed = 0
    for r in range(4, ws_cart.max_row + 1):
        old_val = ws_cart.cell(r, 25).value
        if old_val and isinstance(old_val, str) and 'SUM(' in old_val:
            new_formula = f"=SUM(Z{r}:AK{r})"
            ws_cart.cell(r, 25).value = new_formula
            total_fixed += 1

    print(f"  TOTAL PERÍODO corrigido: {total_fixed} rows")

    # ----------------------------------------------------------
    # ETAPA 7: Escrever valores JAN/25 na PROJEÇÃO (REAL JAN col 27)
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 7: Escrevendo JAN/25 na PROJEÇÃO (REAL JAN col 27)...")

    proj_written = 0
    proj_updated = 0
    for entry in jan25_mapped:
        if entry['proj_row'] and entry['liquido'] > 0:
            r = entry['proj_row']
            old_val = ws_proj.cell(r, 27).value
            ws_proj.cell(r, 27).value = entry['liquido']
            if old_val and old_val != 0:
                proj_updated += 1
            else:
                proj_written += 1

    print(f"  PROJEÇÃO: {proj_written} novos + {proj_updated} atualizados = {proj_written + proj_updated} total")

    # ----------------------------------------------------------
    # ETAPA 8: Verificação final
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 8: Verificação final...")

    # Conferir DRAFT 1 header
    d1_h21 = ws_d1.cell(3, 21).value
    d1_h22 = ws_d1.cell(3, 22).value
    print(f"  DRAFT 1: col 21={d1_h21}, col 22={d1_h22}")

    # Conferir CARTEIRA headers
    cart_h25 = ws_cart.cell(3, 25).value
    cart_h26 = ws_cart.cell(3, 26).value
    cart_h27 = ws_cart.cell(3, 27).value
    cart_h37 = ws_cart.cell(3, 37).value
    print(f"  CARTEIRA: col 25={cart_h25}, col 26={cart_h26}, col 27={cart_h27}, col 37={cart_h37}")

    # Conferir fórmula sample
    sample_formula = ws_cart.cell(4, 27).value
    print(f"  CARTEIRA formula sample (col 27, row 4): {sample_formula}")
    sample_formula26 = ws_cart.cell(4, 26).value
    print(f"  CARTEIRA formula sample (col 26, row 4): {sample_formula26}")
    total_sample = ws_cart.cell(4, 25).value
    print(f"  CARTEIRA TOTAL sample (col 25, row 4): {total_sample}")

    # Conferir PROJEÇÃO
    proj_h27 = ws_proj.cell(3, 27).value
    print(f"  PROJEÇÃO: col 27={proj_h27}")

    # Conferir AGENDA sample
    for agenda_name in ['AGENDA LARISSA', 'AGENDA DAIANE']:
        ws_ag = wb[agenda_name]
        sample = ws_ag.cell(4, 1).value
        if sample:
            print(f"  {agenda_name} A4: {sample[:80]}...")

    # ----------------------------------------------------------
    # ETAPA 9: Salvar
    # ----------------------------------------------------------
    print(f"\n[{elapsed()}] ETAPA 9: Salvando V24...")
    wb.save(OUTPUT_FILE)

    file_size = os.path.getsize(OUTPUT_FILE) / (1024 * 1024)
    total_time = time.time() - start

    print(f"\n{'='*60}")
    print(f"  V24 GERADO COM SUCESSO!")
    print(f"  Arquivo: {OUTPUT_FILE}")
    print(f"  Tamanho: {file_size:.2f} MB")
    print(f"  Tempo total: {total_time:.1f}s")
    print(f"{'='*60}")

    # Resumo
    print(f"\n  RESUMO DAS ALTERAÇÕES:")
    print(f"  - DRAFT 1: +1 col (JAN/25 na posição 21), {d1_written} valores")
    print(f"  - CARTEIRA: +1 col (JAN/25 na posição 26), {cart_formulas} formulas + {cart_values} valores")
    print(f"  - PROJEÇÃO: {proj_written + proj_updated} valores em REAL JAN (col 27)")
    print(f"  - Fórmulas corrigidas: {fixes_draft1} DRAFT 1 + {fixes_carteira} CARTEIRA + {fixes_local} local")
    print(f"  - TOTAL PERÍODO: {total_fixed} rows atualizadas")


if __name__ == '__main__':
    main()

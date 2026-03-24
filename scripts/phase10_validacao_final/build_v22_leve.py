#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V22 — VERSÃO LEVE (Google Sheets compatível)

Parte do V20 (8.16 MB, abre no Google Sheets).
Para os 5.589 prospects (rows 558-6147):
  - Escreve VALORES ESTÁTICOS vazios ("") nas colunas de fórmulas
  - Prospects NÃO TÊM dados em DRAFT 1, PROJEÇÃO, nem DRAFT 2
  - Todas as fórmulas retornariam "" ou 0 de qualquer forma
  - Economiza 1.27 MILHÃO de fórmulas desnecessárias

Para os 554 clientes originais (rows 4-557):
  - Mantém fórmulas intactas (já existem no V20)
  - Corrige ranges de $557 → $6147 nas colunas RANK, COVERAGE, GAP

Também:
  - Popula FUNIL defaults para prospects
  - Resultado: arquivo ~8-9 MB que ABRE no Google Sheets
"""

import openpyxl
from openpyxl.formula.translate import Translator
import shutil
import re
from pathlib import Path
from datetime import datetime

V20 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V20_FINAL.xlsx"
V22 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V22_FINAL.xlsx"


def col_letter(col_num):
    """Convert column number to letter (1=A, 27=AA, etc.)"""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def main():
    t = datetime.now()
    print("=" * 100)
    print("V22 — VERSÃO LEVE (Google Sheets compatível)")
    print(f"Início: {t}")
    print("=" * 100)

    # 1. Copiar V20 → V22
    print("\n[1] Copiando V20 → V22...", flush=True)
    shutil.copy2(V20, V22)

    # 2. Abrir e identificar colunas com fórmulas na row 4
    print("[2] Identificando colunas com fórmulas (row 4)...", flush=True)
    wb = openpyxl.load_workbook(V22)
    ws = wb["CARTEIRA"]
    mr = ws.max_row or 0
    mc = ws.max_column or 0
    print(f"  CARTEIRA: {mr} rows x {mc} cols", flush=True)

    # Scan row 4 para encontrar fórmulas
    formula_cols = {}  # col_num → formula_string_from_row4
    value_cols = set()
    empty_cols = set()

    for c in range(1, mc + 1):
        cell = ws.cell(row=4, column=c)
        v = cell.value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            empty_cols.add(c)
        elif isinstance(v, str) and v.startswith("="):
            formula_cols[c] = v
        else:
            value_cols.add(c)

    print(f"  Row 4: {len(formula_cols)} fórmulas, {len(value_cols)} valores, {len(empty_cols)} vazios", flush=True)

    # 3. Determinar ranges
    # Clientes originais: rows 4-557 (554 CNPJs) — já têm fórmulas no V20
    # Prospects novos: rows 558-max_row — NÃO têm fórmulas

    # Verificar onde começam os prospects (rows sem fórmulas)
    first_prospect_row = None
    test_col = list(formula_cols.keys())[0]
    for r in range(4, min(mr + 1, 600)):
        v = ws.cell(row=r, column=test_col).value
        if v is None or (isinstance(v, str) and not v.startswith("=")):
            first_prospect_row = r
            break

    if not first_prospect_row:
        first_prospect_row = 558  # fallback

    last_row = mr
    prospect_rows = last_row - first_prospect_row + 1

    print(f"\n[3] Ranges identificados:", flush=True)
    print(f"  Clientes originais: rows 4-{first_prospect_row - 1} ({first_prospect_row - 4} clientes)", flush=True)
    print(f"  Prospects novos: rows {first_prospect_row}-{last_row} ({prospect_rows} prospects)", flush=True)
    print(f"  Colunas com fórmulas: {len(formula_cols)}", flush=True)

    # 4. Para os PROSPECTS: escrever VALORES ESTÁTICOS vazios
    # em vez de 1.27 milhão de fórmulas inúteis
    print(f"\n[4] Escrevendo valores estáticos para prospects...", flush=True)
    print(f"  (Prospects NÃO TÊM dados em DRAFT 1, PROJEÇÃO, DRAFT 2)", flush=True)
    print(f"  (Todas as fórmulas retornariam '' ou 0)", flush=True)

    cells_written = 0
    for col_num in sorted(formula_cols.keys()):
        for r in range(first_prospect_row, last_row + 1):
            ws.cell(row=r, column=col_num, value="")
            cells_written += 1

        if cells_written % 100000 == 0:
            print(f"    ... {cells_written:,} células escritas", flush=True)

    print(f"  Total células com valor estático: {cells_written:,}", flush=True)

    # 5. Corrigir ranges dos clientes originais ($557 → $6147)
    print(f"\n[5] Corrigindo ranges $557 → $6147 nos clientes originais...", flush=True)
    range_fixes = 0

    for r in range(4, first_prospect_row):
        for col_num in formula_cols.keys():
            cell = ws.cell(row=r, column=col_num)
            v = cell.value
            if isinstance(v, str) and "$557" in v:
                new_v = v.replace("$557", "$6147")
                cell.value = new_v
                range_fixes += 1

    print(f"  Ranges corrigidos: {range_fixes} células", flush=True)

    # 6. Pré-popular campos FUNIL dos prospects
    print(f"\n[6] Completando FUNIL dos prospects...", flush=True)
    funil_defaults = {
        47: "PROSPECÇÃO - PRIMEIRO CONTATO",  # AÇÃO FUTURA
        54: "❄️FRIO",  # TEMPERATURA — prospect começa frio
    }

    funil_written = 0
    for r in range(first_prospect_row, last_row + 1):
        for col, default in funil_defaults.items():
            current = ws.cell(row=r, column=col).value
            if not current or (isinstance(current, str) and current.strip() == ""):
                ws.cell(row=r, column=col, value=default)
                funil_written += 1

    print(f"  FUNIL complementado: {funil_written} células", flush=True)

    # 7. Verificação pré-save
    print(f"\n[7] Verificação pré-save...", flush=True)

    # Amostra: row 4 (cliente original) deve ter fórmulas
    print(f"  Row 4 (cliente original):", flush=True)
    check_cols = {25: "TOTAL PERÍODO", 88: "META JAN", 264: "RANK"}
    for c, label in sorted(check_cols.items()):
        v = ws.cell(row=4, column=c).value
        status = "FORM" if isinstance(v, str) and v.startswith("=") else ("VAL" if v else "VAZIO")
        display = str(v)[:60] if v else ""
        print(f"    C{c} ({label:15s}): {status:5s} | {display}", flush=True)

    # Amostra: row 560 (prospect) deve ter valores estáticos
    sample_row = min(560, last_row)
    print(f"  Row {sample_row} (prospect):", flush=True)
    check_cols2 = {25: "TOTAL PERÍODO", 47: "AÇÃO FUTURA", 54: "TEMPERATURA", 88: "META JAN", 264: "RANK"}
    for c, label in sorted(check_cols2.items()):
        v = ws.cell(row=sample_row, column=c).value
        status = "FORM" if isinstance(v, str) and v.startswith("=") else ("VAL" if v else "VAZIO")
        display = str(v)[:60] if v else ""
        print(f"    C{c} ({label:15s}): {status:5s} | {display}", flush=True)

    # Contar fórmulas totais no arquivo
    print(f"\n  Contagem de fórmulas (amostra cols 25, 88, 264):", flush=True)
    for c in [25, 88, 264]:
        form_count = 0
        for r in range(4, last_row + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.startswith("="):
                form_count += 1
        print(f"    Col {c}: {form_count} fórmulas (esperado: {first_prospect_row - 4})", flush=True)

    # 8. Salvar
    print(f"\n[8] Salvando V22...", flush=True)
    wb.save(V22)
    wb.close()

    size_mb = Path(V22).stat().st_size / (1024 * 1024)
    elapsed = (datetime.now() - t).total_seconds()

    v20_size = Path(V20).stat().st_size / (1024 * 1024)

    print(f"\n{'='*100}")
    print(f"[SUCESSO] V22 gerado!")
    print(f"  Arquivo: {V22}")
    print(f"  Tamanho: {size_mb:.2f} MB (V20 era {v20_size:.2f} MB)")
    print(f"  Fórmulas mantidas: {first_prospect_row - 4} clientes × {len(formula_cols)} cols")
    print(f"  Valores estáticos: {cells_written:,} células (prospects)")
    print(f"  Ranges corrigidos: {range_fixes} ($557→$6147)")
    print(f"  FUNIL complementado: {funil_written} células")
    print(f"  Rows com fórmulas: 4-{first_prospect_row - 1}")
    print(f"  Rows com valores: {first_prospect_row}-{last_row}")
    print(f"  Tempo: {elapsed:.1f}s")
    print(f"{'='*100}")
    print(f"\n💡 Este arquivo deve abrir no Google Sheets sem problemas!")
    print(f"   Os 554 clientes originais mantêm fórmulas funcionais.")
    print(f"   Os {prospect_rows} prospects têm valores estáticos (não precisam de fórmulas).")


if __name__ == "__main__":
    main()

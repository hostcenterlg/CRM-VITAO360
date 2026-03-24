#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PATCH V17 — Corrigir SINALEIRO (BJ), TIPO CLIENTE (AX), e CONSULTOR faltantes.

Problema: build_v17 checava CARTEIRA col N/P que tinham FORMULAS (=INDEX...)
e nao caia no fallback do DRAFT 2. Agora: ignora formulas, usa DRAFT 2 direto.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

V17_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V17_FINAL.xlsx"
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"


def normalizar_cnpj(valor):
    if not valor:
        return None
    cnpj = re.sub(r'[^\d]', '', str(valor).strip())
    return cnpj if len(cnpj) >= 11 else None


def get_real_value(cell):
    """Retorna valor REAL (nao formula). Se formula, retorna None."""
    val = cell.value
    if val is None:
        return None
    if isinstance(val, str) and val.startswith("="):
        return None  # formula = nao eh valor real
    if isinstance(val, str) and val.strip() == "":
        return None
    return val


def calcular_sinaleiro(situacao, dias_sem_compra):
    """Calcula sinaleiro emoji"""
    sit = str(situacao).upper().strip() if situacao else ""
    if sit in ("PROSPECT", "LEAD"):
        return "\U0001f7e3"  # roxo
    if dias_sem_compra is None:
        return ""
    try:
        dias = float(dias_sem_compra)
    except (ValueError, TypeError):
        return ""
    if dias <= 50:
        return "\U0001f7e2"  # verde
    elif dias <= 90:
        return "\U0001f7e1"  # amarelo
    else:
        return "\U0001f534"  # vermelho


def main():
    start = datetime.now()
    print("=" * 80)
    print("PATCH V17 — Corrigir SINALEIRO, TIPO CLIENTE, CONSULTOR")
    print("=" * 80)

    print("\n[1/4] Carregando V17...")
    wb = openpyxl.load_workbook(V17_PATH)

    print("[2/4] Carregando V31 (data_only=True)...")
    wb_v31 = openpyxl.load_workbook(V31_PATH, data_only=True)

    ws_cart = wb["CARTEIRA"]
    ws_d1 = wb["DRAFT 1"]
    ws_d2 = wb["DRAFT 2"]

    # === Indexar DRAFT 1 por CNPJ ===
    print("\n[3/4] Indexando fontes de dados...")
    d1_by_cnpj = {}
    # Detectar colunas
    d1_headers = {}
    for col in range(1, ws_d1.max_column + 1):
        h = ws_d1.cell(row=3, column=col).value
        if h:
            d1_headers[str(h).upper().strip()] = col

    print(f"  DRAFT 1 headers: {list(d1_headers.keys())[:20]}...")

    # Mapear colunas relevantes
    d1_consultor_col = d1_headers.get("CONSULTOR", 10)
    d1_situacao_col = d1_headers.get("SITUAÇÃO", None)
    d1_dias_col = d1_headers.get("DIAS SEM COMPRA", None)
    d1_tipo_cliente_col = d1_headers.get("TIPO CLIENTE", None)

    # Procurar variantes
    if not d1_situacao_col:
        for key, col in d1_headers.items():
            if "SITUA" in key:
                d1_situacao_col = col
                break
    if not d1_dias_col:
        for key, col in d1_headers.items():
            if "DIAS" in key and "COMPRA" in key:
                d1_dias_col = col
                break
    if not d1_tipo_cliente_col:
        for key, col in d1_headers.items():
            if "TIPO" in key and "CLIENTE" in key:
                d1_tipo_cliente_col = col
                break
            if "TIPO" in key and "CLI" in key:
                d1_tipo_cliente_col = col
                break

    print(f"  DRAFT 1 cols: CONSULTOR={d1_consultor_col}, SITUACAO={d1_situacao_col}, DIAS={d1_dias_col}, TIPO_CLI={d1_tipo_cliente_col}")

    for row in range(4, ws_d1.max_row + 1):
        cnpj = normalizar_cnpj(ws_d1.cell(row=row, column=2).value)
        if not cnpj:
            continue
        data = {}
        for col in range(1, ws_d1.max_column + 1):
            val = get_real_value(ws_d1.cell(row=row, column=col))
            if val is not None:
                data[col] = val
        d1_by_cnpj[cnpj] = data

    print(f"  DRAFT 1: {len(d1_by_cnpj)} clientes")

    # === Indexar DRAFT 2 (ultimo por CNPJ) ===
    d2_latest = {}
    d2_counts = defaultdict(int)

    for row in range(3, ws_d2.max_row + 1):
        cnpj = normalizar_cnpj(ws_d2.cell(row=row, column=4).value)
        if not cnpj:
            continue
        d2_counts[cnpj] += 1
        date_val = ws_d2.cell(row=row, column=1).value
        date_str = str(date_val) if date_val else ""

        if cnpj in d2_latest:
            if date_str <= str(d2_latest[cnpj].get('_date', "")):
                continue

        data = {'_date': date_val}
        for col in range(1, 32):
            val = get_real_value(ws_d2.cell(row=row, column=col))
            if val is not None:
                data[col] = val
        d2_latest[cnpj] = data

    print(f"  DRAFT 2: {len(d2_latest)} CNPJs (ultimo atendimento)")

    # === Indexar V31 DRAFT 2 (valores calculados) ===
    ws_v31_d2 = wb_v31["DRAFT 2"]
    v31_d2_latest = {}

    for row in range(3, ws_v31_d2.max_row + 1):
        cnpj = normalizar_cnpj(ws_v31_d2.cell(row=row, column=4).value)
        if not cnpj:
            continue
        date_val = ws_v31_d2.cell(row=row, column=1).value
        date_str = str(date_val) if date_val else ""

        if cnpj in v31_d2_latest:
            if date_str <= str(v31_d2_latest[cnpj].get('_date', "")):
                continue

        data = {'_date': date_val}
        for col in range(1, 32):
            val = ws_v31_d2.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                data[col] = val
        v31_d2_latest[cnpj] = data

    print(f"  V31 DRAFT 2: {len(v31_d2_latest)} CNPJs")

    # === V31 CARTEIRA ===
    ws_v31_cart = wb_v31["CARTEIRA"]
    v31_cart = {}
    for row in range(4, ws_v31_cart.max_row + 1):
        cnpj = normalizar_cnpj(ws_v31_cart.cell(row=row, column=2).value)
        if not cnpj:
            continue
        data = {}
        for col in range(1, min(ws_v31_cart.max_column + 1, 270)):
            val = ws_v31_cart.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                data[col] = val
        v31_cart[cnpj] = data

    print(f"  V31 CARTEIRA: {len(v31_cart)} clientes")

    # ================================================================
    # PATCH: Corrigir cada row da CARTEIRA
    # ================================================================
    print(f"\n[4/4] Aplicando patches na CARTEIRA...")

    fixes = {
        'sinaleiro': 0,
        'consultor': 0,
        'tipo_cliente': 0,
        'situacao': 0,
        'dias': 0,
        'temperatura': 0,
        'other': 0,
    }

    for row in range(4, ws_cart.max_row + 1):
        cnpj = normalizar_cnpj(ws_cart.cell(row=row, column=2).value)
        if not cnpj:
            continue

        d1 = d1_by_cnpj.get(cnpj, {})
        d2 = d2_latest.get(cnpj, {})
        v31_d2 = v31_d2_latest.get(cnpj, {})
        v31_c = v31_cart.get(cnpj, {})

        # === CONSULTOR (col L=12) — fallback mais amplo ===
        current_consult = get_real_value(ws_cart.cell(row=row, column=12))
        if not current_consult:
            val = d1.get(d1_consultor_col) if d1_consultor_col else None
            if not val:
                val = d2.get(2)  # DRAFT 2 col B = CONSULTOR
            if not val:
                val = v31_d2.get(2)
            if not val:
                val = v31_c.get(12)
            if val:
                ws_cart.cell(row=row, column=12, value=val)
                fixes['consultor'] += 1

        # === SITUAÇÃO (col N=14) — precisa ser VALOR, nao formula ===
        current_sit = get_real_value(ws_cart.cell(row=row, column=14))
        if not current_sit:
            val = None
            if d1_situacao_col:
                val = d1.get(d1_situacao_col)
            if not val:
                val = d2.get(7)  # DRAFT 2 col G = SITUAÇÃO
            if not val:
                val = v31_d2.get(7)
            if not val:
                val = v31_c.get(14)
            if val:
                ws_cart.cell(row=row, column=14, value=val)
                fixes['situacao'] += 1
                current_sit = val

        # === DIAS SEM COMPRA (col P=16) ===
        current_dias = get_real_value(ws_cart.cell(row=row, column=16))
        if current_dias is None:
            val = None
            if d1_dias_col:
                val = d1.get(d1_dias_col)
            if val is None:
                val = d2.get(8)  # DRAFT 2 col H
            if val is None:
                val = v31_d2.get(8)
            if val is None:
                val = v31_c.get(16)
            if val is not None:
                ws_cart.cell(row=row, column=16, value=val)
                fixes['dias'] += 1
                current_dias = val

        # === SINALEIRO (col BJ=62) — SEMPRE recalcular com dados reais ===
        # Pegar SITUACAO real (do que acabamos de escrever ou do DRAFT)
        sit_for_sinal = current_sit
        if not sit_for_sinal:
            sit_for_sinal = d2.get(7) or v31_d2.get(7) or v31_c.get(14)

        dias_for_sinal = current_dias
        if dias_for_sinal is None:
            dias_for_sinal = d2.get(8) or v31_d2.get(8) or v31_c.get(16)

        sinaleiro = calcular_sinaleiro(sit_for_sinal, dias_for_sinal)
        if sinaleiro:
            ws_cart.cell(row=row, column=62, value=sinaleiro)
            fixes['sinaleiro'] += 1
        elif sit_for_sinal or dias_for_sinal is not None:
            # Tem dados mas sinaleiro ficou vazio — forcar emoji
            ws_cart.cell(row=row, column=62, value=calcular_sinaleiro(sit_for_sinal, dias_for_sinal) or "\U0001f7e2")
            fixes['sinaleiro'] += 1

        # === TIPO CLIENTE (col AX=50) ===
        current_tipo = get_real_value(ws_cart.cell(row=row, column=50))
        if not current_tipo:
            val = None
            if d1_tipo_cliente_col:
                val = d1.get(d1_tipo_cliente_col)
            # Tambem checar col J (10) do DRAFT 1 que pode ter "TIPO CLIENTE"
            if not val:
                val = d1.get(10)  # col J as fallback
            if not val:
                val = d2.get(10)  # DRAFT 2 col J = TIPO CLIENTE
            if not val:
                val = v31_d2.get(10)
            if not val:
                val = v31_c.get(50)
            # Se nada, inferir do ESTAGIO FUNIL
            if not val:
                estagio = get_real_value(ws_cart.cell(row=row, column=44))
                if estagio:
                    est_upper = str(estagio).upper()
                    if "PROSPECT" in est_upper:
                        val = "PROSPECT"
                    elif "LEAD" in est_upper:
                        val = "LEAD"
            if val:
                ws_cart.cell(row=row, column=50, value=val)
                fixes['tipo_cliente'] += 1

        # === TEMPERATURA (col BB=54) — do V31 se disponivel ===
        current_temp = get_real_value(ws_cart.cell(row=row, column=54))
        if not current_temp:
            val = v31_c.get(54)
            if not val:
                val = d2.get(25)  # DRAFT 2 col Y = TEMPERATURA
            if not val:
                val = v31_d2.get(25)
            if val:
                ws_cart.cell(row=row, column=54, value=val)
                fixes['temperatura'] += 1

        if (row - 4) % 100 == 0 and row > 4:
            total_fixes = sum(fixes.values())
            print(f"    ... row {row}/{ws_cart.max_row} ({total_fixes} fixes)")

    # === RESULTADO ===
    print(f"\n  PATCHES APLICADOS:")
    total = 0
    for key, count in fixes.items():
        total += count
        print(f"    {key:15s}: {count:>4}")
    print(f"    {'TOTAL':15s}: {total:>4}")

    # === Verificacao pos-patch ===
    print(f"\n  Verificacao pos-patch (amostra rows 4-103):")
    end_check = min(ws_cart.max_row, 103)
    key_cols = {
        12: "CONSULTOR",
        14: "SITUAÇÃO",
        16: "DIAS SEM COMPRA",
        44: "ESTÁGIO FUNIL",
        48: "ÚLT. RESULTADO",
        50: "TIPO CLIENTE",
        51: "TENTATIVA",
        54: "TEMPERATURA",
        62: "SINALEIRO",
    }
    for col, name in key_cols.items():
        filled = 0
        formula = 0
        for row in range(4, end_check + 1):
            val = ws_cart.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                if isinstance(val, str) and val.startswith("="):
                    formula += 1
                else:
                    filled += 1
        pct = round(100 * filled / max(1, end_check - 3), 1)
        bar = "#" * int(pct / 5) + "." * (20 - int(pct / 5))
        print(f"    {get_column_letter(col):>3} | {name:20s} | {bar} {pct:>5.1f}% (F={formula})")

    # === Amostra ===
    print(f"\n  Amostra CARTEIRA (rows 4-8):")
    for row in range(4, min(9, ws_cart.max_row + 1)):
        cnpj = str(ws_cart.cell(row=row, column=2).value or "")[:14]
        consultor = str(ws_cart.cell(row=row, column=12).value or "")[:15]
        situacao = str(get_real_value(ws_cart.cell(row=row, column=14)) or "")[:12]
        dias = str(get_real_value(ws_cart.cell(row=row, column=16)) or "")[:6]
        sinaleiro = str(ws_cart.cell(row=row, column=62).value or "")[:4]
        tipo_cli = str(ws_cart.cell(row=row, column=50).value or "")[:12]
        resultado = str(ws_cart.cell(row=row, column=48).value or "")[:15]

        print(f"    R{row}: {cnpj:14s} | {consultor:15s} | SIT={situacao:12s} | D={dias:6s} | {sinaleiro} | TC={tipo_cli:12s} | R={resultado}")

    # === Salvar ===
    print(f"\n  Salvando V17 patcheado...")
    wb.save(V17_PATH)
    wb.close()
    wb_v31.close()

    size = Path(V17_PATH).stat().st_size / (1024 * 1024)
    elapsed = (datetime.now() - start).total_seconds()
    print(f"\n  Salvo: {V17_PATH} ({size:.2f} MB)")
    print(f"  Tempo: {elapsed:.1f}s")
    print(f"\n[SUCESSO] V17 patcheado com {total} correcoes!")


if __name__ == "__main__":
    main()

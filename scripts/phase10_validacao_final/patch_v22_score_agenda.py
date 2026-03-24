#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PATCH V22 — SCORE + RANK + SITUAÇÃO + AGENDA

1. Calcula SCORE (col O=15) para os 5.590 prospects em Python
2. Calcula RANK (col 264) para TODOS (clientes + prospects)
3. Preenche SITUAÇÃO (col 14) = "PROSPECT" onde vazio
4. Monta SORTBY/FILTER nas 4 AGENDA tabs
5. Preenche META DIÁRIA, PIPELINE, COVERAGE, ALERTA, GAP para prospects
"""

import openpyxl
from datetime import datetime, date
from pathlib import Path

V22 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V22_FINAL.xlsx"


def calc_score(row_data):
    """
    Replica a fórmula SCORE (col O) em Python.

    =IFERROR(
      (IFERROR(MIN(100, P/MAX(S,1)*100), 0)) * 0.30  -- URGÊNCIA
      + (MIN(100, IF(AM="A",100,"B"=60,"C"=30,0) + tipo_cliente_bonus)) * 0.25  -- VALOR
      + (IF(AND(ISNUMBER(AS),AS<TODAY()), MIN(100,(TODAY()-AS)*5), 0)) * 0.20  -- FOLLOW-UP
      + (MIN(100, temp_score + b2b_bonus)) * 0.15  -- SINAL
      + (tentativa_score) * 0.05  -- TENTATIVA
      + (situacao_score) * 0.05  -- SITUAÇÃO
    , 0)
    """
    p = row_data.get("dias_sem_compra", 0) or 0  # col 16
    s = row_data.get("ciclo", 0) or 0  # col 19
    am = str(row_data.get("curva_abc", "") or "").strip().upper()  # col 39
    ax = str(row_data.get("tipo_cliente", "") or "").strip().upper()  # col 50
    as_date = row_data.get("followup_date", None)  # col 45
    bb = str(row_data.get("temperatura", "") or "").strip().upper()  # col 54
    t = str(row_data.get("b2b", "") or "").strip().upper()  # col 20
    ay = str(row_data.get("tentativa", "") or "").strip().upper()  # col 51
    n = str(row_data.get("situacao", "") or "").strip().upper()  # col 14

    today = date.today()

    # 1. URGÊNCIA (30%)
    try:
        p_num = float(p) if isinstance(p, (int, float)) else 0
        s_num = float(s) if isinstance(s, (int, float)) else 1
        urgencia = min(100, p_num / max(s_num, 1) * 100)
    except (ValueError, TypeError, ZeroDivisionError):
        urgencia = 0

    # 2. VALOR (25%)
    abc_score = 100 if am == "A" else (60 if am == "B" else (30 if am == "C" else 0))
    tipo_bonus = 0
    if "FIDELIZADO" in ax:
        tipo_bonus = 20
    elif "MADURO" in ax:
        tipo_bonus = 15
    elif "RECORRENTE" in ax:
        tipo_bonus = 10
    valor = min(100, abc_score + tipo_bonus)

    # 3. FOLLOW-UP (20%)
    followup = 0
    if as_date and isinstance(as_date, (datetime, date)):
        d = as_date if isinstance(as_date, date) else as_date.date()
        if d < today:
            days_late = (today - d).days
            followup = min(100, days_late * 5)

    # 4. SINAL (15%)
    if "QUENTE" in bb:
        temp_score = 100
    elif "MORNO" in bb:
        temp_score = 60
    elif "FRIO" in bb:
        temp_score = 20
    else:
        temp_score = 0
    b2b_bonus = 20 if t == "SIM" else 0
    sinal = min(100, temp_score + b2b_bonus)

    # 5. TENTATIVA (5%)
    if ay == "T1":
        tent_score = 100
    elif ay == "T2":
        tent_score = 80
    elif ay == "T3":
        tent_score = 60
    elif ay == "T4":
        tent_score = 40
    else:
        tent_score = 20  # default

    # 6. SITUAÇÃO (5%)
    sit_map = {
        "EM RISCO": 100, "ATIVO": 80, "INAT.REC": 70,
        "PROSPECT": 50, "NOVO": 40, "LEAD": 30, "INAT.ANT": 20
    }
    sit_score = sit_map.get(n, 0)

    total = (urgencia * 0.30 + valor * 0.25 + followup * 0.20
             + sinal * 0.15 + tent_score * 0.05 + sit_score * 0.05)
    return round(total, 2)


def main():
    t0 = datetime.now()
    print("=" * 80)
    print("PATCH V22 — SCORE + RANK + SITUAÇÃO + AGENDA")
    print("=" * 80)

    print("\n[1] Abrindo V22...", flush=True)
    wb = openpyxl.load_workbook(V22)
    ws = wb["CARTEIRA"]
    mr = ws.max_row
    mc = ws.max_column
    print(f"  CARTEIRA: {mr} rows x {mc} cols")

    # Detectar onde começam prospects
    first_prospect = 558
    last_row = mr
    n_clients = first_prospect - 4  # 554
    n_prospects = last_row - first_prospect + 1  # 5590

    # ========== FASE 1: SITUAÇÃO (col 14) ==========
    print(f"\n[2] Preenchendo SITUAÇÃO (col 14) para prospects...", flush=True)
    sit_count = 0
    for r in range(first_prospect, last_row + 1):
        v = ws.cell(row=r, column=14).value
        if not v or (isinstance(v, str) and v.strip() == ""):
            ws.cell(row=r, column=14, value="PROSPECT")
            sit_count += 1
    print(f"  {sit_count} prospects receberam SITUAÇÃO='PROSPECT'")

    # ========== FASE 2: SCORE (col 15) para prospects ==========
    print(f"\n[3] Calculando SCORE para {n_prospects} prospects...", flush=True)
    # Columns: 14=SITUAÇÃO, 16=DIAS SEM COMPRA, 19=CICLO, 20=B2B,
    #          39=CURVA ABC, 45=FOLLOWUP DATE, 50=TIPO CLIENTE,
    #          51=TENTATIVA, 54=TEMPERATURA
    scores = {}
    for r in range(first_prospect, last_row + 1):
        row_data = {
            "situacao": ws.cell(row=r, column=14).value,
            "dias_sem_compra": ws.cell(row=r, column=16).value,
            "ciclo": ws.cell(row=r, column=19).value,
            "b2b": ws.cell(row=r, column=20).value,
            "curva_abc": ws.cell(row=r, column=39).value,
            "followup_date": ws.cell(row=r, column=45).value,
            "tipo_cliente": ws.cell(row=r, column=50).value,
            "tentativa": ws.cell(row=r, column=51).value,
            "temperatura": ws.cell(row=r, column=54).value,
        }
        score = calc_score(row_data)
        scores[r] = score
        ws.cell(row=r, column=15, value=score)

    # Estatísticas
    score_vals = list(scores.values())
    avg = sum(score_vals) / len(score_vals) if score_vals else 0
    mn = min(score_vals) if score_vals else 0
    mx = max(score_vals) if score_vals else 0
    print(f"  Scores calculados: {len(scores)}")
    print(f"  Min={mn:.1f}  Avg={avg:.1f}  Max={mx:.1f}")

    # ========== FASE 3: RANK (col 264) para prospects ==========
    print(f"\n[4] Calculando RANK para prospects...", flush=True)
    # Ler SCORE dos clientes originais (rows 4-557) — são fórmulas, não temos valor
    # Para o RANK estático dos prospects, precisamos de TODOS os scores
    # Clientes originais têm fórmulas — não dá pra ler o valor
    # Solução: ler V22 com data_only=True só pra pegar os scores dos clientes
    wb_data = openpyxl.load_workbook(V22, read_only=True, data_only=True)
    ws_data = wb_data["CARTEIRA"]

    all_scores = {}  # row → score
    for r in range(4, first_prospect):
        v = ws_data.cell(row=r, column=15).value
        # data_only=True retorna cached value (pode ser None se nunca recalculou)
        if v is not None and isinstance(v, (int, float)):
            all_scores[r] = float(v)
        else:
            all_scores[r] = 0  # sem cached value
    wb_data.close()

    # Adicionar scores dos prospects (já calculados)
    for r, s in scores.items():
        all_scores[r] = s

    # Calcular RANK (1 = maior score)
    sorted_rows = sorted(all_scores.items(), key=lambda x: -x[1])
    rank_map = {}
    for idx, (r, s) in enumerate(sorted_rows, 1):
        rank_map[r] = idx

    # Escrever RANK nos prospects (clientes mantêm fórmula)
    for r in range(first_prospect, last_row + 1):
        ws.cell(row=r, column=264, value=rank_map.get(r, ""))

    print(f"  RANK escrito para {n_prospects} prospects")
    print(f"  Top 5 prospects: {sorted([(rank_map[r], r, scores[r]) for r in scores], key=lambda x: x[0])[:5]}")

    # ========== FASE 4: INTELIGÊNCIA extras (cols 265-269) ==========
    print(f"\n[5] Preenchendo INTELIGÊNCIA para prospects...", flush=True)
    # Col 265 = META DIÁRIA → prospects sem meta = 0
    # Col 266 = PIPELINE VALUE → prospects sem pipeline = 0
    # Col 267 = COVERAGE RATIO → 0
    # Col 268 = ALERTA → "PROSPECT - SEM HISTÓRICO"
    # Col 269 = GAP VALUE → 0
    for r in range(first_prospect, last_row + 1):
        ws.cell(row=r, column=265, value=0)        # META DIÁRIA
        ws.cell(row=r, column=266, value=0)        # PIPELINE VALUE
        ws.cell(row=r, column=267, value=0)        # COVERAGE RATIO
        ws.cell(row=r, column=268, value="PROSPECT - SEM HISTÓRICO")  # ALERTA
        ws.cell(row=r, column=269, value=0)        # GAP VALUE
    print(f"  5 colunas x {n_prospects} rows = {5 * n_prospects} células")

    # ========== FASE 5: AGENDA tabs ==========
    print(f"\n[6] Montando AGENDA tabs com SORTBY/FILTER...", flush=True)
    # Headers para AGENDA (32 colunas)
    agenda_headers = [
        "RANK", "CNPJ", "RAZÃO SOCIAL", "UF", "CIDADE",
        "SITUAÇÃO", "DIAS SEM COMPRA", "CURVA ABC", "TICKET MÉDIO",
        "ESTÁGIO FUNIL", "AÇÃO FUTURA", "TEMPERATURA", "SINALEIRO",
        "TENTATIVA", "SCORE", "CONSULTOR",
        "DATA CONTATO", "HORÁRIO", "CANAL", "RESULTADO",
        "MOTIVO", "OBSERVAÇÃO", "PRÓXIMO PASSO", "DATA RETORNO",
        "VALOR PROPOSTA", "FASE", "TIPO AÇÃO", "TIPO PROBLEMA",
        "CONVERSÃO", "STATUS ATEND", "PRIORIDADE MANUAL", "OBS GESTOR"
    ]

    # Colunas CARTEIRA correspondentes para SORTBY/FILTER
    # A fórmula SPILL em A4: =SORTBY(FILTER(cols, CARTEIRA!L=consultor), score_col, -1)
    # Mapeamento AGENDA col → CARTEIRA col letter
    cart_cols = [
        "JD", "B", "C", "D", "E",     # RANK, CNPJ, RAZÃO, UF, CIDADE
        "N", "P", "AM", "AP",          # SITUAÇÃO, DIAS, ABC, TICKET
        "AR", "AU", "BB", "BD",        # ESTÁGIO, AÇÃO, TEMP, SINALEIRO
        "AY", "O", "L",                # TENTATIVA, SCORE, CONSULTOR
    ]
    # Cols 17-32 são preenchidas pelo consultor no dia

    consultores = {
        "AGENDA LARISSA": "LARISSA PADILHA",
        "AGENDA DAIANE": "DAIANE STAVICKI",
        "AGENDA MANU": "MANU DITZEL",
        "AGENDA JULIO": "JULIO GADRET",
    }

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for sheet_name, consultor in consultores.items():
        ws_ag = wb[sheet_name]

        # Limpar
        for row in ws_ag.iter_rows(min_row=1, max_row=ws_ag.max_row, max_col=32):
            for cell in row:
                cell.value = None

        # Row 1: título
        ws_ag.cell(row=1, column=1, value=f"AGENDA DIÁRIA — {consultor}")
        ws_ag.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="2F5496")
        ws_ag.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)

        # Row 2: instrução
        ws_ag.cell(row=2, column=1, value="Colunas A-P = automático (SORTBY/FILTER da CARTEIRA). Colunas Q-AF = preenchimento manual do consultor no dia.")
        ws_ag.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color="808080")
        ws_ag.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)

        # Row 3: headers
        for c, h in enumerate(agenda_headers, 1):
            cell = ws_ag.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Row 4: fórmula SPILL
        # =SORTBY(FILTER(CARTEIRA!B$4:B$6147, CARTEIRA!L$4:L$6147="CONSULTOR"),
        #         FILTER(CARTEIRA!O$4:O$6147, CARTEIRA!L$4:L$6147="CONSULTOR"), -1)
        # Mas precisamos de MÚLTIPLAS colunas...
        # Usamos HSTACK ou repetimos por coluna

        # Abordagem: cada coluna da AGENDA tem sua própria fórmula SORTBY/FILTER
        # Todas usam a mesma ordem de sort (por SCORE desc)
        for ci, cart_col in enumerate(cart_cols):
            formula = (
                f'=IFERROR(SORTBY('
                f'FILTER(CARTEIRA!{cart_col}$4:{cart_col}$6147,'
                f'CARTEIRA!L$4:L$6147="{consultor}"),'
                f'FILTER(CARTEIRA!O$4:O$6147,'
                f'CARTEIRA!L$4:L$6147="{consultor}"),-1),"")'
            )
            ws_ag.cell(row=4, column=ci + 1, value=formula)

        # Ajustar larguras
        widths = [8, 18, 35, 5, 20, 12, 8, 8, 12, 18, 25, 12, 12, 8, 8, 20,
                  12, 10, 12, 15, 20, 30, 25, 12, 15, 15, 15, 15, 10, 12, 12, 20]
        for c, w in enumerate(widths, 1):
            ws_ag.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

        # Freeze panes na row 4
        ws_ag.freeze_panes = "A4"
        # Auto filter
        ws_ag.auto_filter.ref = f"A3:AF3"

        print(f"  {sheet_name}: headers + {len(cart_cols)} fórmulas SORTBY/FILTER para '{consultor}'")

    # ========== FASE 6: Salvar ==========
    print(f"\n[7] Salvando V22 atualizado...", flush=True)
    wb.save(V22)
    wb.close()

    size_mb = Path(V22).stat().st_size / (1024 * 1024)
    elapsed = (datetime.now() - t0).total_seconds()

    print(f"\n{'='*80}")
    print(f"[SUCESSO] V22 patcheado!")
    print(f"  Tamanho: {size_mb:.2f} MB")
    print(f"  SITUAÇÃO preenchida: {sit_count}")
    print(f"  SCORE calculado: {len(scores)} prospects (avg={avg:.1f})")
    print(f"  RANK calculado: {n_prospects} prospects")
    print(f"  INTELIGÊNCIA: 5 cols x {n_prospects} rows")
    print(f"  AGENDA: 4 tabs com SORTBY/FILTER")
    print(f"  Tempo: {elapsed:.1f}s")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()

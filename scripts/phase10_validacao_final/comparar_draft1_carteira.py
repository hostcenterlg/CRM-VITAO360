#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Compara DRAFT 1 e CARTEIRA entre V14 (nosso) e V31 (seu) para validar consistencia dos dados Mercos.
"""

import openpyxl
import re
import json
from collections import defaultdict

V14_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V14_FINAL.xlsx"
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"

def normalizar_cnpj(valor):
    if not valor:
        return None
    cnpj = re.sub(r'[^\d]', '', str(valor).strip())
    if len(cnpj) == 14:
        return cnpj
    return None

def extrair_headers(ws, row=1, max_col=None):
    """Extrai headers de uma row"""
    if not max_col:
        max_col = ws.max_column
    headers = {}
    for col in range(1, min(max_col + 1, 300)):
        val = ws.cell(row=row, column=col).value
        if val:
            headers[col] = str(val).strip()
    return headers

def encontrar_coluna_cnpj(ws, start_row=2, max_check=20):
    """Encontra qual coluna tem CNPJs"""
    for col in range(1, 10):
        cnpj_count = 0
        for row in range(start_row, min(start_row + max_check, ws.max_row + 1)):
            val = ws.cell(row=row, column=col).value
            if val and normalizar_cnpj(val):
                cnpj_count += 1
        if cnpj_count >= 5:
            return col
    return None

def extrair_dados_por_cnpj(ws, cnpj_col, name_col, data_start_row=2):
    """Extrai dados indexados por CNPJ"""
    dados = {}
    for row in range(data_start_row, ws.max_row + 1):
        cnpj_raw = ws.cell(row=row, column=cnpj_col).value
        cnpj = normalizar_cnpj(cnpj_raw)
        if not cnpj:
            continue

        nome = ws.cell(row=row, column=name_col).value if name_col else None
        nome = str(nome).strip() if nome else "[SEM NOME]"

        # Pegar todas as colunas dessa row
        row_data = {}
        for col in range(1, min(ws.max_column + 1, 50)):  # primeiras 50 colunas
            val = ws.cell(row=row, column=col).value
            if val is not None:
                row_data[col] = val

        dados[cnpj] = {
            "nome": nome,
            "row": row,
            "data": row_data
        }
    return dados

def main():
    print("=" * 100)
    print("COMPARACAO DRAFT 1 + CARTEIRA: V14 (nosso) vs V31 (seu)")
    print("=" * 100)

    # Carregar ambos com data_only=True para ver valores
    print("\n[1/4] Carregando arquivos...")
    wb_v14 = openpyxl.load_workbook(V14_PATH, data_only=True)
    wb_v31 = openpyxl.load_workbook(V31_PATH, data_only=True)

    resultado = {}

    # ==================== DRAFT 1 ====================
    print("\n" + "=" * 100)
    print("ANALISE DRAFT 1")
    print("=" * 100)

    ws_d1_v14 = wb_v14["DRAFT 1"]
    ws_d1_v31 = wb_v31["DRAFT 1"]

    print(f"\n  V14 DRAFT 1: {ws_d1_v14.max_row} rows x {ws_d1_v14.max_column} cols")
    print(f"  V31 DRAFT 1: {ws_d1_v31.max_row} rows x {ws_d1_v31.max_column} cols")

    # Headers
    h_v14 = extrair_headers(ws_d1_v14, row=2)  # DRAFT 1 header pode ser row 2
    h_v31 = extrair_headers(ws_d1_v31, row=2)

    # Se row 2 tiver poucos headers, tentar row 1 e row 3
    if len(h_v14) < 5:
        h_v14 = extrair_headers(ws_d1_v14, row=1)
    if len(h_v31) < 5:
        h_v31 = extrair_headers(ws_d1_v31, row=1)

    print(f"\n  V14 headers ({len(h_v14)} colunas): {list(h_v14.values())[:15]}")
    print(f"  V31 headers ({len(h_v31)} colunas): {list(h_v31.values())[:15]}")

    # Encontrar coluna CNPJ
    cnpj_col_v14 = encontrar_coluna_cnpj(ws_d1_v14, start_row=3)
    cnpj_col_v31 = encontrar_coluna_cnpj(ws_d1_v31, start_row=3)

    # Se nao achou, tentar com row 4
    if not cnpj_col_v14:
        cnpj_col_v14 = encontrar_coluna_cnpj(ws_d1_v14, start_row=4)
    if not cnpj_col_v31:
        cnpj_col_v31 = encontrar_coluna_cnpj(ws_d1_v31, start_row=4)

    print(f"\n  CNPJ col V14: {cnpj_col_v14}")
    print(f"  CNPJ col V31: {cnpj_col_v31}")

    if cnpj_col_v14 and cnpj_col_v31:
        # Nome geralmente eh coluna anterior ou A
        name_col_v14 = 1 if cnpj_col_v14 > 1 else 2
        name_col_v31 = 1 if cnpj_col_v31 > 1 else 2

        dados_d1_v14 = extrair_dados_por_cnpj(ws_d1_v14, cnpj_col_v14, name_col_v14, data_start_row=3)
        dados_d1_v31 = extrair_dados_por_cnpj(ws_d1_v31, cnpj_col_v31, name_col_v31, data_start_row=3)

        set_v14 = set(dados_d1_v14.keys())
        set_v31 = set(dados_d1_v31.keys())
        comuns = set_v14 & set_v31

        print(f"\n  DRAFT 1 CNPJs V14: {len(set_v14)}")
        print(f"  DRAFT 1 CNPJs V31: {len(set_v31)}")
        print(f"  Comuns: {len(comuns)}")
        print(f"  Apenas V14: {len(set_v14 - set_v31)}")
        print(f"  Apenas V31: {len(set_v31 - set_v14)}")

        # Comparar dados dos CNPJs comuns (primeiras 20 colunas)
        nomes_diferentes = 0
        dados_diferentes = 0

        for cnpj in list(comuns)[:20]:  # amostra de 20
            v14_nome = dados_d1_v14[cnpj]["nome"]
            v31_nome = dados_d1_v31[cnpj]["nome"]
            if v14_nome != v31_nome:
                nomes_diferentes += 1

        print(f"\n  Nomes diferentes (amostra 20): {nomes_diferentes}")

        # Mostrar 5 exemplos de CNPJs comuns
        print(f"\n  AMOSTRA de 5 clientes comuns:")
        for i, cnpj in enumerate(sorted(comuns)[:5]):
            v14_d = dados_d1_v14[cnpj]
            v31_d = dados_d1_v31[cnpj]
            print(f"    {cnpj}: V14='{v14_d['nome'][:30]}' | V31='{v31_d['nome'][:30]}'")

        # Mostrar 5 exclusivos de cada
        excl_v14 = set_v14 - set_v31
        excl_v31 = set_v31 - set_v14

        if excl_v14:
            print(f"\n  AMOSTRA exclusivos V14 ({len(excl_v14)} total):")
            for cnpj in sorted(excl_v14)[:5]:
                print(f"    {cnpj}: '{dados_d1_v14[cnpj]['nome'][:40]}'")

        if excl_v31:
            print(f"\n  AMOSTRA exclusivos V31 ({len(excl_v31)} total):")
            for cnpj in sorted(excl_v31)[:5]:
                print(f"    {cnpj}: '{dados_d1_v31[cnpj]['nome'][:40]}'")

        resultado["draft1"] = {
            "v14_cnpjs": len(set_v14),
            "v31_cnpjs": len(set_v31),
            "comuns": len(comuns),
            "apenas_v14": len(excl_v14),
            "apenas_v31": len(excl_v31)
        }

    # ==================== CARTEIRA ====================
    print("\n" + "=" * 100)
    print("ANALISE CARTEIRA")
    print("=" * 100)

    ws_c_v14 = wb_v14["CARTEIRA"]
    ws_c_v31 = wb_v31["CARTEIRA"]

    print(f"\n  V14 CARTEIRA: {ws_c_v14.max_row} rows x {ws_c_v14.max_column} cols")
    print(f"  V31 CARTEIRA: {ws_c_v31.max_row} rows x {ws_c_v31.max_column} cols")

    # Headers da CARTEIRA (row 1 = super-grupo, row 2 = headers, row 3 = sub-headers)
    h_c_v14_r1 = extrair_headers(ws_c_v14, row=1)
    h_c_v14_r2 = extrair_headers(ws_c_v14, row=2)
    h_c_v14_r3 = extrair_headers(ws_c_v14, row=3)

    h_c_v31_r1 = extrair_headers(ws_c_v31, row=1)
    h_c_v31_r2 = extrair_headers(ws_c_v31, row=2)
    h_c_v31_r3 = extrair_headers(ws_c_v31, row=3)

    print(f"\n  V14 CARTEIRA headers:")
    print(f"    Row 1 (super-grupo): {list(h_c_v14_r1.values())[:8]}")
    print(f"    Row 2 (header): {list(h_c_v14_r2.values())[:8]}")
    print(f"    Row 3 (sub-header): {list(h_c_v14_r3.values())[:8]}")

    print(f"\n  V31 CARTEIRA headers:")
    print(f"    Row 1 (super-grupo): {list(h_c_v31_r1.values())[:8]}")
    print(f"    Row 2 (header): {list(h_c_v31_r2.values())[:8]}")
    print(f"    Row 3 (sub-header): {list(h_c_v31_r3.values())[:8]}")

    # CNPJ na CARTEIRA
    cnpj_col_c_v14 = encontrar_coluna_cnpj(ws_c_v14, start_row=4)
    cnpj_col_c_v31 = encontrar_coluna_cnpj(ws_c_v31, start_row=4)

    # Tentar mais colunas se nao achou
    if not cnpj_col_c_v14:
        for try_col in range(1, 20):
            count = 0
            for r in range(4, min(25, ws_c_v14.max_row + 1)):
                v = ws_c_v14.cell(row=r, column=try_col).value
                if v and normalizar_cnpj(v):
                    count += 1
            if count >= 3:
                cnpj_col_c_v14 = try_col
                break

    if not cnpj_col_c_v31:
        for try_col in range(1, 20):
            count = 0
            for r in range(4, min(25, ws_c_v31.max_row + 1)):
                v = ws_c_v31.cell(row=r, column=try_col).value
                if v and normalizar_cnpj(v):
                    count += 1
            if count >= 3:
                cnpj_col_c_v31 = try_col
                break

    print(f"\n  CNPJ col V14 CARTEIRA: {cnpj_col_c_v14}")
    print(f"  CNPJ col V31 CARTEIRA: {cnpj_col_c_v31}")

    if cnpj_col_c_v14 and cnpj_col_c_v31:
        dados_c_v14 = extrair_dados_por_cnpj(ws_c_v14, cnpj_col_c_v14, cnpj_col_c_v14 - 1 if cnpj_col_c_v14 > 1 else 1, data_start_row=4)
        dados_c_v31 = extrair_dados_por_cnpj(ws_c_v31, cnpj_col_c_v31, cnpj_col_c_v31 - 1 if cnpj_col_c_v31 > 1 else 1, data_start_row=4)

        set_c_v14 = set(dados_c_v14.keys())
        set_c_v31 = set(dados_c_v31.keys())
        comuns_c = set_c_v14 & set_c_v31

        print(f"\n  CARTEIRA CNPJs V14: {len(set_c_v14)}")
        print(f"  CARTEIRA CNPJs V31: {len(set_c_v31)}")
        print(f"  Comuns: {len(comuns_c)}")
        print(f"  Apenas V14: {len(set_c_v14 - set_c_v31)}")
        print(f"  Apenas V31: {len(set_c_v31 - set_c_v14)}")

        resultado["carteira"] = {
            "v14_cnpjs": len(set_c_v14),
            "v31_cnpjs": len(set_c_v31),
            "comuns": len(comuns_c),
            "apenas_v14": len(set_c_v14 - set_c_v31),
            "apenas_v31": len(set_c_v31 - set_c_v14)
        }

        # Comparar consultor (coluna L geralmente = col 12)
        print(f"\n  DISTRIBUICAO POR CONSULTOR:")

        for version, ws, cnpj_col, start_row, label in [
            ("V14", ws_c_v14, cnpj_col_c_v14, 4, "V14"),
            ("V31", ws_c_v31, cnpj_col_c_v31, 4, "V31")
        ]:
            consultores = defaultdict(int)
            # Consultor geralmente na coluna L (12) — mas pode variar
            # Vamos verificar qual coluna tem nomes de consultores conhecidos
            consultor_col = None
            known = {"LARISSA", "DAIANE", "JULIO", "MANU", "HEMANUELE", "LORRANY", "LEANDRO"}

            for try_col in range(1, 30):
                match_count = 0
                for r in range(start_row, min(start_row + 50, ws.max_row + 1)):
                    v = ws.cell(row=r, column=try_col).value
                    if v and any(k in str(v).upper() for k in known):
                        match_count += 1
                if match_count >= 3:
                    consultor_col = try_col
                    break

            if consultor_col:
                for r in range(start_row, ws.max_row + 1):
                    v = ws.cell(row=r, column=consultor_col).value
                    if v:
                        consultores[str(v).strip()] += 1

                print(f"\n    {label} (col {consultor_col}):")
                for c, count in sorted(consultores.items(), key=lambda x: -x[1]):
                    print(f"      {c}: {count} clientes")
            else:
                print(f"\n    {label}: coluna consultor NAO encontrada")

    # ==================== VALORES FINANCEIROS ====================
    print("\n" + "=" * 100)
    print("VALORES FINANCEIROS (DRAFT 1 - amostra)")
    print("=" * 100)

    if cnpj_col_v14 and cnpj_col_v31:
        # Pegar colunas com valores numericos grandes (vendas)
        # Em ambas versoes, comparar valores para CNPJs comuns
        print("\n  Comparando valores de vendas para 10 CNPJs comuns:")

        for i, cnpj in enumerate(sorted(comuns)[:10]):
            v14_row = dados_d1_v14[cnpj]["row"]
            v31_row = dados_d1_v31[cnpj]["row"]

            # Pegar valores de vendas (colunas 10-30 tipicamente)
            v14_vals = []
            v31_vals = []

            for col in range(5, 45):
                v14_val = ws_d1_v14.cell(row=v14_row, column=col).value
                v31_val = ws_d1_v31.cell(row=v31_row, column=col).value

                if isinstance(v14_val, (int, float)) and v14_val > 100:
                    v14_vals.append((col, v14_val))
                if isinstance(v31_val, (int, float)) and v31_val > 100:
                    v31_vals.append((col, v31_val))

            v14_total = sum(v for _, v in v14_vals)
            v31_total = sum(v for _, v in v31_vals)

            nome = dados_d1_v14[cnpj]["nome"][:25]
            diff = abs(v14_total - v31_total) / max(v14_total, 1) * 100 if v14_total else 0

            match = "OK" if diff < 5 else f"DIFF {diff:.1f}%"
            print(f"    {cnpj} | {nome:25s} | V14: R$ {v14_total:>12,.2f} | V31: R$ {v31_total:>12,.2f} | {match}")

    # ==================== RESUMO ====================
    print("\n" + "=" * 100)
    print("RESUMO FINAL")
    print("=" * 100)

    if "draft1" in resultado:
        d = resultado["draft1"]
        pct = d["comuns"] / max(d["v14_cnpjs"], 1) * 100
        print(f"\n  DRAFT 1:")
        print(f"    V14: {d['v14_cnpjs']} CNPJs | V31: {d['v31_cnpjs']} CNPJs")
        print(f"    Match: {d['comuns']} comuns ({pct:.1f}% do V14)")

    if "carteira" in resultado:
        c = resultado["carteira"]
        pct = c["comuns"] / max(c["v14_cnpjs"], 1) * 100
        print(f"\n  CARTEIRA:")
        print(f"    V14: {c['v14_cnpjs']} CNPJs | V31: {c['v31_cnpjs']} CNPJs")
        print(f"    Match: {c['comuns']} comuns ({pct:.1f}% do V14)")

    print(f"\n" + "=" * 100)

    wb_v14.close()
    wb_v31.close()

if __name__ == "__main__":
    main()

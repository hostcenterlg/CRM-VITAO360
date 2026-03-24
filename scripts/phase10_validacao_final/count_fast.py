#!/usr/bin/env python3
"""Contagem RAPIDA usando iter_rows com min/max col para evitar leitura de colunas desnecessarias"""
import openpyxl, re, sys
from datetime import datetime

def norm(v):
    if not v: return None
    c = re.sub(r'[^\d]', '', str(v).strip())
    return c if len(c) >= 11 else None

t = datetime.now()

# 1. MERCOS Prospects (5755 rows, col 3 = CNPJ)
print("1. MERCOS Prospects...", flush=True)
wb = openpyxl.load_workbook(
    r"c:/Users/User/OneDrive/Área de Trabalho/CLAUDE CODE/08_CARTEIRA_MERCOS.xlsx",
    data_only=True, read_only=True
)
ws = wb["Prospects"]
prosp = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    cn = norm(row[0].value)
    if cn: prosp.add(cn)
print(f"   Prospects: {len(prosp)} CNPJs ({(datetime.now()-t).total_seconds():.0f}s)", flush=True)

# Clientes Mercos
ws2 = wb["Carteira Clientes Mercos"]
clientes = set()
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=3, max_col=3):
    cn = norm(row[0].value)
    if cn: clientes.add(cn)
print(f"   Clientes: {len(clientes)} CNPJs", flush=True)

# Padronizado
ws3 = wb["Clientes Padronizado"]
padrao = set()
for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=3, max_col=3):
    cn = norm(row[0].value)
    if cn: padrao.add(cn)
print(f"   Padronizado: {len(padrao)} CNPJs", flush=True)
wb.close()

total_mercos = prosp | clientes | padrao
print(f"   MERCOS TOTAL: {len(total_mercos)} CNPJs unicos", flush=True)
print(f"   Prospects exclusivos: {len(prosp - clientes)}", flush=True)

# 2. SAP Cadastro no repo (menor que SAP TOTAL)
print(f"\n2. SAP sources no repo...", flush=True)
import os
sap_dir = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/sap"
sap_repo = set()
if os.path.exists(sap_dir):
    for f in os.listdir(sap_dir):
        if f.endswith('.xlsx'):
            fp = os.path.join(sap_dir, f)
            print(f"   {f} ({os.path.getsize(fp)/1024:.0f} KB)...", flush=True)
            wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            # Find CNPJ col
            cc = None
            for c in range(1, 15):
                h = ws.cell(row=1, column=c).value
                if h and "CNPJ" in str(h).upper():
                    cc = c; break
            if cc:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=cc, max_col=cc):
                    cn = norm(row[0].value)
                    if cn: sap_repo.add(cn)
            wb.close()
            print(f"     CNPJ col={cc} → total acumulado: {len(sap_repo)}", flush=True)
print(f"   SAP REPO TOTAL: {len(sap_repo)} CNPJs", flush=True)

# 3. DRAFT3 COM INATIVOS
print(f"\n3. DRAFT3 COM INATIVOS...", flush=True)
d3_path = r"c:/Users/User/OneDrive/Área de Trabalho/DRAFT3/DRAFT_3_SAP  COM INATIVOS.xlsx"
d3 = set()
if os.path.exists(d3_path):
    wb = openpyxl.load_workbook(d3_path, data_only=True, read_only=True)
    print(f"   Abas: {wb.sheetnames}", flush=True)
    for sn in wb.sheetnames[:2]:
        ws = wb[sn]
        print(f"   {sn}: {ws.max_row} rows x {ws.max_column} cols", flush=True)
        # Headers
        for c in range(1, min((ws.max_column or 10)+1, 10)):
            h = ws.cell(row=1, column=c).value
            if h: print(f"     Col {c}: {str(h)[:30]}", flush=True)
        cc = None
        for c in range(1, 15):
            h = ws.cell(row=1, column=c).value
            if h and "CNPJ" in str(h).upper():
                cc = c; break
        if cc:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=cc, max_col=cc):
                cn = norm(row[0].value)
                if cn: d3.add(cn)
        print(f"     {len(d3)} CNPJs", flush=True)
    wb.close()
print(f"   DRAFT3 INATIVOS: {len(d3)} CNPJs", flush=True)

# RESUMO
print(f"\n{'='*60}", flush=True)
print("RESUMO FINAL", flush=True)
print(f"{'='*60}", flush=True)
print(f"  V17 CARTEIRA:       554 CNPJs (conhecidos)", flush=True)
print(f"  V31 CARTEIRA:     5,460 CNPJs (conhecidos)", flush=True)
print(f"  MERCOS Clientes:  {len(clientes):>5} CNPJs", flush=True)
print(f"  MERCOS Prospects: {len(prosp):>5} CNPJs ← ACHAMOS!", flush=True)
print(f"  MERCOS Padronizado: {len(padrao):>3} CNPJs", flush=True)
print(f"  MERCOS TOTAL:     {len(total_mercos):>5} CNPJs", flush=True)
print(f"  SAP Repo:         {len(sap_repo):>5} CNPJs", flush=True)
print(f"  DRAFT3 Inativos:  {len(d3):>5} CNPJs", flush=True)

universo = total_mercos | sap_repo | d3
print(f"\n  UNIVERSO TOTAL:   {len(universo):>5} CNPJs", flush=True)
print(f"  V17 atual:          554 CNPJs", flush=True)
print(f"  FALTANDO:         {len(universo) - 554:>5} CNPJs", flush=True)

print(f"\n  DETALHAMENTO:", flush=True)
print(f"    Prospects Mercos exclusivos (nao em Clientes): {len(prosp - clientes)}", flush=True)
print(f"    SAP exclusivos (nao no Mercos): {len(sap_repo - total_mercos)}", flush=True)
print(f"    DRAFT3 exclusivos (nao no Mercos/SAP): {len(d3 - total_mercos - sap_repo)}", flush=True)

elapsed = (datetime.now() - t).total_seconds()
print(f"\n  Tempo: {elapsed:.1f}s", flush=True)

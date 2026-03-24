#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analise detalhada: LOG vs DRAFT 2 no V15 e no V31.
Objetivo: entender mapeamento de colunas para unificar.
"""

import openpyxl
import json
from collections import defaultdict

V15_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V15_FINAL.xlsx"
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"


def col_letter(idx):
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def analyze_sheet(wb, sheet_name, label, max_sample=5):
    """Analise completa de uma aba"""
    if sheet_name not in wb.sheetnames:
        print(f"  {sheet_name} NAO EXISTE em {label}")
        return None

    ws = wb[sheet_name]
    info = {
        "label": label,
        "sheet": sheet_name,
        "rows": ws.max_row,
        "cols": ws.max_column,
        "freeze": str(ws.freeze_panes),
        "headers": {},
        "fill_stats": {},
        "samples": {}
    }

    print(f"\n{'='*80}")
    print(f"  {sheet_name} em {label}: {ws.max_row} rows x {ws.max_column} cols")
    print(f"  Freeze: {ws.freeze_panes}")
    print(f"{'='*80}")

    # Headers - todas as rows de header
    for header_row in range(1, 5):
        row_headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val is not None:
                row_headers.append((col, col_letter(col), str(val).strip()))
        if row_headers:
            print(f"\n  Row {header_row} headers ({len(row_headers)}):")
            for col_idx, letter, val in row_headers:
                print(f"    {letter:>4} (col {col_idx:>2}): {val[:50]}")
            info["headers"][header_row] = {h[1]: h[2] for h in row_headers}

    # Preenchimento por coluna
    print(f"\n  Preenchimento (primeiras 100 rows de dados):")
    data_start = 3 if sheet_name == "LOG" else 4  # LOG starts row 3, DRAFT 2 starts row 4
    end_row = min(ws.max_row, data_start + 100)
    total_data_rows = ws.max_row - data_start + 1

    for col in range(1, ws.max_column + 1):
        filled = 0
        formula_count = 0
        sample_vals = []
        for row in range(data_start, end_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                filled += 1
                if isinstance(val, str) and val.startswith("="):
                    formula_count += 1
                if len(sample_vals) < max_sample:
                    sample_vals.append(str(val)[:40])

        letter = col_letter(col)
        h = ""
        for hr in [2, 1, 3]:
            hval = ws.cell(row=hr, column=col).value
            if hval:
                h = str(hval).strip()[:30]
                break

        pct = round(100 * filled / max(1, end_row - data_start), 1)
        tipo = "FORMULA" if formula_count > 0 else ("DADOS" if filled > 0 else "VAZIO")

        info["fill_stats"][letter] = {
            "header": h,
            "filled": filled,
            "formulas": formula_count,
            "pct": pct,
            "type": tipo
        }
        info["samples"][letter] = sample_vals

        if filled > 0 or h:
            bar = "#" * int(pct / 5) + "." * (20 - int(pct / 5))
            print(f"    {letter:>4} | {h:30s} | {bar} {pct:>5.1f}% | {tipo:8s} | samples: {sample_vals[:2]}")

    return info


def main():
    print("=" * 100)
    print("ANALISE LOG vs DRAFT 2 — V15 e V31")
    print("=" * 100)

    # Carregar V15
    print("\n[1/4] Carregando V15 (formulas)...")
    wb_v15 = openpyxl.load_workbook(V15_PATH, data_only=False)
    print(f"  Abas V15: {wb_v15.sheetnames}")

    # Carregar V31
    print("\n[2/4] Carregando V31 (formulas)...")
    wb_v31 = openpyxl.load_workbook(V31_PATH, data_only=False)
    print(f"  Abas V31: {wb_v31.sheetnames}")

    results = {}

    # V15 LOG
    print("\n[3/4] Analisando V15...")
    results["v15_log"] = analyze_sheet(wb_v15, "LOG", "V15")
    results["v15_draft2"] = analyze_sheet(wb_v15, "DRAFT 2", "V15")

    # V31 LOG e DRAFT 2
    print("\n[4/4] Analisando V31...")
    results["v31_log"] = analyze_sheet(wb_v31, "LOG", "V31")
    results["v31_draft2"] = analyze_sheet(wb_v31, "DRAFT 2", "V31")

    # AGENDA do V31 (para referencia)
    for s in wb_v31.sheetnames:
        if "AGENDA" in s.upper():
            results[f"v31_{s.lower().replace(' ','_')}"] = analyze_sheet(wb_v31, s, "V31")
            break  # so o primeiro

    # Comparacao final
    print(f"\n{'='*100}")
    print("RESUMO COMPARATIVO")
    print(f"{'='*100}")

    for key in ["v15_log", "v15_draft2", "v31_log", "v31_draft2"]:
        if results.get(key):
            r = results[key]
            filled_cols = sum(1 for v in r["fill_stats"].values() if v["filled"] > 0)
            print(f"  {r['label']:5s} {r['sheet']:12s}: {r['rows']:>6} rows x {r['cols']:>3} cols | {filled_cols} cols com dados")

    # Mapeamento sugerido LOG -> DRAFT 2
    print(f"\n{'='*100}")
    print("MAPEAMENTO SUGERIDO: LOG cols -> DRAFT 2 cols")
    print(f"{'='*100}")

    if results.get("v15_log") and results.get("v15_draft2"):
        log_h = results["v15_log"]["headers"].get(2, {})
        d2_h = results["v15_draft2"]["headers"].get(2, {})

        print(f"\n  LOG headers (row 2): {log_h}")
        print(f"\n  DRAFT 2 headers (row 2): {d2_h}")

        # Tentar match por nome
        print(f"\n  Matches por nome:")
        for log_col, log_name in log_h.items():
            for d2_col, d2_name in d2_h.items():
                if log_name.upper().strip() == d2_name.upper().strip():
                    print(f"    LOG {log_col} ({log_name}) -> DRAFT 2 {d2_col} ({d2_name})")
                elif any(k in log_name.upper() for k in d2_name.upper().split()[:2]) and len(d2_name) > 3:
                    print(f"    LOG {log_col} ({log_name}) ~~ DRAFT 2 {d2_col} ({d2_name}) [parcial]")

    wb_v15.close()
    wb_v31.close()

    # Salvar resultados
    output = {}
    for key, val in results.items():
        if val:
            output[key] = {
                "rows": val["rows"],
                "cols": val["cols"],
                "headers": val["headers"],
                "fill_summary": {k: {"header": v["header"], "pct": v["pct"], "type": v["type"]}
                                 for k, v in val["fill_stats"].items() if v["filled"] > 0}
            }

    out_path = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/analise_log_vs_draft2.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False, default=str)
    print(f"\n  Resultados salvos em: {out_path}")

    print(f"\n[COMPLETO]")


if __name__ == "__main__":
    main()

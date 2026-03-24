#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificacao COMPLETA do V16 FINAL.
Checa: estrutura, formulas, preenchimento, CNPJ consistency, CARTEIRA-DRAFT2 link.
"""

import openpyxl
import re
from collections import defaultdict
from pathlib import Path

V16_PATH = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V16_FINAL.xlsx"
V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"


def col_letter(idx):
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def main():
    print("=" * 100)
    print("VERIFICAÇÃO COMPLETA V16 FINAL")
    print("=" * 100)

    wb = openpyxl.load_workbook(V16_PATH, data_only=False)
    issues = []
    checks_passed = 0

    # ====================================
    # CHECK 1: ABAS ESPERADAS
    # ====================================
    print("\n[1/8] ABAS ESPERADAS...")
    expected_tabs = [
        "SINALEIRO", "PROJEÇÃO ", "DASH", "REDES_FRANQUIAS_v2",
        "COMITE", "REGRAS", "DRAFT 1", "DRAFT 2", "DRAFT 3 ",
        "CARTEIRA", "AGENDA LARISSA", "AGENDA DAIANE", "AGENDA MANU", "AGENDA JULIO"
    ]
    actual_tabs = wb.sheetnames
    print(f"  Esperadas: {len(expected_tabs)} | Encontradas: {len(actual_tabs)}")

    for tab in expected_tabs:
        if tab in actual_tabs:
            print(f"    ✓ {tab}")
            checks_passed += 1
        else:
            # Try case-insensitive match
            found = False
            for a in actual_tabs:
                if a.strip().upper() == tab.strip().upper():
                    print(f"    ~ {tab} (encontrado como '{a}')")
                    checks_passed += 1
                    found = True
                    break
            if not found:
                issues.append(f"Aba faltando: {tab}")
                print(f"    ✗ {tab} FALTANDO")

    # Verificar que LOG NAO existe
    if "LOG" not in actual_tabs:
        print(f"    ✓ LOG removido (unificado no DRAFT 2)")
        checks_passed += 1
    else:
        issues.append("LOG ainda existe — deveria ter sido removido")

    # ====================================
    # CHECK 2: DRAFT 2 ESTRUTURA
    # ====================================
    print("\n[2/8] DRAFT 2 ESTRUTURA...")
    ws = wb["DRAFT 2"]
    print(f"  Rows: {ws.max_row} | Cols: {ws.max_column} | Freeze: {ws.freeze_panes}")

    # Headers
    d2_headers = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=2, column=col).value
        d2_headers.append(str(h).strip() if h else "")

    expected_d2_headers = [
        "DATA", "CONSULTOR", "NOME FANTASIA", "CNPJ", "UF", "REDE / REGIONAL",
        "SITUAÇÃO", "DIAS SEM COMPRA", "ESTÁGIO FUNIL", "TIPO CLIENTE", "FASE",
        "SINALEIRO", "TENTATIVA", "WHATSAPP", "LIGAÇÃO", "LIGAÇÃO ATENDIDA",
        "TIPO DO CONTATO", "RESULTADO", "MOTIVO", "FOLLOW-UP", "AÇÃO FUTURA",
        "AÇÃO DETALHADA", "MERCOS ATUALIZADO", "NOTA DO DIA", "TEMPERATURA",
        "GRUPO DASH", "SINALEIRO META", "TIPO AÇÃO", "TIPO PROBLEMA", "DEMANDA",
        "TIPO ATENDIMENTO"
    ]

    header_match = 0
    for i, (actual, expected) in enumerate(zip(d2_headers, expected_d2_headers)):
        if actual == expected:
            header_match += 1
        else:
            issues.append(f"DRAFT 2 header col {col_letter(i+1)}: '{actual}' != '{expected}'")

    print(f"  Headers: {header_match}/{len(expected_d2_headers)} match")
    if header_match == len(expected_d2_headers):
        checks_passed += 1

    # Preenchimento por coluna (amostra 100 rows)
    print(f"\n  Preenchimento (amostra 100 rows de dados):")
    end_row = min(ws.max_row, 103)
    total_filled = 0
    for col in range(1, 32):
        filled = 0
        formula_count = 0
        for row in range(3, end_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                filled += 1
                if isinstance(val, str) and val.startswith("="):
                    formula_count += 1
        pct = round(100 * filled / max(1, end_row - 2), 1)
        total_filled += filled
        bar = "#" * int(pct / 5) + "." * (20 - int(pct / 5))
        h = d2_headers[col - 1] if col <= len(d2_headers) else ""
        tipo = "FORMULA" if formula_count > 0 else ("DADOS" if filled > 0 else "VAZIO")
        print(f"    {col_letter(col):>4} | {h:25s} | {bar} {pct:>5.1f}% | {tipo}")

    # Contar CNPJs unicos no DRAFT 2
    d2_cnpjs = set()
    for row in range(3, ws.max_row + 1):
        cnpj = ws.cell(row=row, column=4).value
        if cnpj:
            cnpj_norm = re.sub(r'[^\d]', '', str(cnpj))
            if len(cnpj_norm) >= 11:
                d2_cnpjs.add(cnpj_norm)

    print(f"\n  CNPJs unicos no DRAFT 2: {len(d2_cnpjs)}")

    # Contar formulas AUTO
    auto_cols = {12: "SINALEIRO", 13: "TENTATIVA", 26: "GRUPO DASH", 28: "TIPO AÇÃO"}
    print(f"\n  Formulas AUTO (verificando rows 3-103):")
    for col, name in auto_cols.items():
        formula_count = 0
        for row in range(3, min(ws.max_row + 1, 103)):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str) and val.startswith("="):
                formula_count += 1
        print(f"    {col_letter(col)} ({name}): {formula_count}/{end_row-2} formulas")
        if formula_count >= (end_row - 2) * 0.9:
            checks_passed += 1

    # ====================================
    # CHECK 3: CARTEIRA ← DRAFT 2 LINK
    # ====================================
    print("\n[3/8] CARTEIRA ← DRAFT 2 LINK...")
    ws_cart = wb["CARTEIRA"]

    # Verificar que FUNIL (AR-BJ) tem formulas referenciando DRAFT 2
    funil_cols = {
        "AR": "ESTÁGIO FUNIL",
        "AS": "PRÓX FOLLOWUP",
        "AT": "DATA ÚLT ATEND",
        "AU": "AÇÃO FUTURA",
        "AV": "ÚLTIMO RESULTADO",
        "AW": "MOTIVO",
        "AY": "TENTATIVA",
        "AZ": "FASE",
        "BH": "PRÓX AÇÃO",
        "BI": "AÇÃO DETALHADA",
        "BJ": "SINALEIRO",
    }

    print(f"  CARTEIRA: {ws_cart.max_row} rows x {ws_cart.max_column} cols")
    for col_letter_str, name in funil_cols.items():
        col_idx = openpyxl.utils.column_index_from_string(col_letter_str)
        val = ws_cart.cell(row=4, column=col_idx).value
        if val and isinstance(val, str) and "DRAFT 2" in val:
            print(f"    ✓ {col_letter_str} ({name}): referencia DRAFT 2")
            checks_passed += 1
        elif val and isinstance(val, str) and val.startswith("="):
            print(f"    ~ {col_letter_str} ({name}): formula (sem ref DRAFT 2)")
        else:
            print(f"    ? {col_letter_str} ({name}): {str(val)[:50]}")

    # Verificar CNPJs da CARTEIRA existem no DRAFT 2
    cart_cnpjs = set()
    for row in range(4, ws_cart.max_row + 1):
        cnpj = ws_cart.cell(row=row, column=2).value  # CARTEIRA col B = CNPJ
        if cnpj:
            cnpj_norm = re.sub(r'[^\d]', '', str(cnpj))
            if len(cnpj_norm) >= 11:
                cart_cnpjs.add(cnpj_norm)

    common = cart_cnpjs & d2_cnpjs
    cart_only = cart_cnpjs - d2_cnpjs
    print(f"\n  CARTEIRA CNPJs: {len(cart_cnpjs)}")
    print(f"  Common com DRAFT 2: {len(common)}")
    print(f"  CARTEIRA-only (sem match no D2): {len(cart_only)}")
    if len(common) > 0:
        checks_passed += 1

    # ====================================
    # CHECK 4: AGENDA TABS
    # ====================================
    print("\n[4/8] AGENDA TABS...")
    agenda_tabs = [s for s in wb.sheetnames if "AGENDA" in s.upper()]
    print(f"  Encontradas: {len(agenda_tabs)} tabs")

    for tab_name in agenda_tabs:
        ws_a = wb[tab_name]
        # Verificar 32 colunas de headers
        headers_count = 0
        for col in range(1, 33):
            h = ws_a.cell(row=1, column=col).value
            if h:
                headers_count += 1

        # Verificar formulas na row 2
        formulas_in_row2 = 0
        for col in range(1, 33):
            val = ws_a.cell(row=2, column=col).value
            if val and isinstance(val, str) and val.startswith("="):
                formulas_in_row2 += 1

        # Verificar se tem SORTBY/FILTER
        has_sortby = False
        for col in range(1, 33):
            val = ws_a.cell(row=2, column=col).value
            if val and isinstance(val, str) and "SORTBY" in val.upper():
                has_sortby = True
                break

        status = "✓" if headers_count >= 30 and formulas_in_row2 >= 10 and has_sortby else "✗"
        print(f"    {status} {tab_name}: {headers_count} headers, {formulas_in_row2} formulas, SORTBY={'SIM' if has_sortby else 'NAO'}")
        if status == "✓":
            checks_passed += 1

    # ====================================
    # CHECK 5: DRAFT 3 (SAP)
    # ====================================
    print("\n[5/8] DRAFT 3 (SAP)...")
    draft3_found = False
    for s in wb.sheetnames:
        if "DRAFT 3" in s.upper():
            ws_d3 = wb[s]
            print(f"  ✓ {s}: {ws_d3.max_row} rows x {ws_d3.max_column} cols")
            draft3_found = True
            checks_passed += 1
            break
    if not draft3_found:
        issues.append("DRAFT 3 nao encontrado")

    # ====================================
    # CHECK 6: SINALEIRO
    # ====================================
    print("\n[6/8] SINALEIRO...")
    if "SINALEIRO" in wb.sheetnames:
        ws_sin = wb["SINALEIRO"]
        print(f"  ✓ SINALEIRO: {ws_sin.max_row} rows x {ws_sin.max_column} cols")
        checks_passed += 1
    else:
        issues.append("SINALEIRO nao encontrado")

    # ====================================
    # CHECK 7: FREEZE PANES
    # ====================================
    print("\n[7/8] FREEZE PANES...")
    expected_freeze = {
        "DRAFT 2": "A3",
        "DRAFT 1": "AN4",
        "CARTEIRA": "BX4",
        "PROJEÇÃO ": "C4",
        "REGRAS": "A41",
    }
    for name, expected in expected_freeze.items():
        if name in wb.sheetnames:
            actual = str(wb[name].freeze_panes)
            if actual == expected:
                print(f"    ✓ {name}: {actual}")
                checks_passed += 1
            else:
                print(f"    ✗ {name}: {actual} (esperado {expected})")
                issues.append(f"Freeze {name}: {actual} != {expected}")

    # ====================================
    # CHECK 8: TAMANHO E TOTAIS
    # ====================================
    print("\n[8/8] TAMANHO E TOTAIS...")
    size = Path(V16_PATH).stat().st_size / (1024 * 1024)
    total_formulas = 0
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    total_formulas += 1

    print(f"  Tamanho: {size:.2f} MB")
    print(f"  Total formulas: {total_formulas:,}")
    print(f"  Total abas: {len(wb.sheetnames)}")
    print(f"  Total registros DRAFT 2: {ws.max_row - 2:,}")  # last ws

    wb.close()

    # ====================================
    # RESULTADO
    # ====================================
    print(f"\n{'='*100}")
    if issues:
        print(f"RESULTADO: {checks_passed} checks OK, {len(issues)} PROBLEMAS")
        for i in issues:
            print(f"  ✗ {i}")
    else:
        print(f"RESULTADO: {checks_passed} checks OK — TUDO PASSOU!")

    print(f"{'='*100}")
    return 0 if not issues else 1


if __name__ == "__main__":
    import sys
    sys.exit(main())

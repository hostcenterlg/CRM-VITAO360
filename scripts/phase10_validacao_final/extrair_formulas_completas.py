#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Extrai formulas COMPLETAS (sem truncar) do V31 CARTEIRA FUNIL"""

import openpyxl

V31_PATH = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"

def col_letter(idx):
    result = ""
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        result = chr(65 + remainder) + result
    return result

wb = openpyxl.load_workbook(V31_PATH, data_only=False)
ws = wb["CARTEIRA"]

print("CARTEIRA ROW 4 - FUNIL FORMULAS COMPLETAS (AR-BJ):")
print("=" * 120)
for col in range(44, 63):  # AR=44 to BJ=62
    cell = ws.cell(row=4, column=col)
    val = cell.value
    letter = col_letter(col)
    h3 = ws.cell(row=3, column=col).value or ""
    if val and isinstance(val, str) and val.startswith("="):
        print(f"\n{letter} ({h3}):")
        print(f"  {val}")

# Also get BJ full formula
print(f"\n\nBJ SINALEIRO row 4 FULL:")
bj_val = ws.cell(row=4, column=62).value
print(f"  {bj_val}")

print(f"\n\nBJ SINALEIRO row 5 FULL:")
bj_val5 = ws.cell(row=5, column=62).value
print(f"  {bj_val5}")

# SCORE formula from AGENDA
ws_a = wb["AGENDA"]
print(f"\n\nAGENDA AE (SCORE) row 2 FULL:")
ae_val = ws_a.cell(row=2, column=31).value
if hasattr(ae_val, 'text'):
    print(f"  ARRAY={ae_val.text}")
else:
    print(f"  {ae_val}")

print(f"\nAGENDA AE (SCORE) row 3 FULL:")
ae_val3 = ws_a.cell(row=3, column=31).value
if hasattr(ae_val3, 'text'):
    print(f"  ARRAY={ae_val3.text}")
else:
    print(f"  {ae_val3}")

wb.close()

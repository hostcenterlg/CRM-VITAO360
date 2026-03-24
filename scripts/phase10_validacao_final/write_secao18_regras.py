#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ESCREVER SEÇÃO 18 NAS REGRAS — Metodologia de Priorização da Agenda
REGRA: SÓ VALORES ESTÁTICOS. ZERO FÓRMULAS.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime
import shutil

V18 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V18_FINAL.xlsx"
V19 = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V19_FINAL.xlsx"

# ============================================================
# STYLES
# ============================================================
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
SUBTITLE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")

BLOCO_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
BLOCO_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

DATA_FONT = Font(name="Calibri", size=10)
DATA_FILL_ALT = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")

RACIONAL_FONT = Font(name="Calibri", size=9, italic=True, color="4472C4")
RACIONAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

GOLDEN_FONT = Font(name="Calibri", size=11, bold=True, color="C45911")
GOLDEN_FILL = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_row(ws, row, cols, font=None, fill=None, alignment=None, border=None):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border


def write_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=THIN_BORDER):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    return cell


def main():
    t = datetime.now()
    print("=" * 100)
    print("ESCREVENDO SEÇÃO 18 — METODOLOGIA DE PRIORIZAÇÃO DA AGENDA")
    print("=" * 100)

    # 1. Copiar V18 → V19
    print("\n[1] Copiando V18 → V19...", flush=True)
    shutil.copy2(V18, V19)

    # 2. Abrir V19
    print("[2] Abrindo V19 para edição...", flush=True)
    wb = openpyxl.load_workbook(V19)
    ws = wb["REGRAS"]

    # Encontrar última row usada
    START = 285  # 2 rows depois da row 283

    r = START
    ALL_COLS = range(1, 8)  # A-G

    # ============================================================
    # TÍTULO PRINCIPAL
    # ============================================================
    print("[3] Escrevendo Seção 18...", flush=True)

    write_cell(ws, r, 1, "🧠 18. METODOLOGIA DE PRIORIZAÇÃO — ORQUESTRAÇÃO DA AGENDA DIÁRIA",
               TITLE_FONT, TITLE_FILL, CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None, TITLE_FONT, TITLE_FILL, CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    write_cell(ws, r, 1, "Como a IA decide quem cada consultor deve ligar PRIMEIRO a cada manhã",
               Font(name="Calibri", size=10, italic=True, color="FFFFFF"),
               PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"), CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None,
                   Font(name="Calibri", size=10, italic=True, color="FFFFFF"),
                   PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"), CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 2

    # ============================================================
    # 18.0 GOLDEN RULE
    # ============================================================
    write_cell(ws, r, 1, "⚡ 18.0 — GOLDEN RULE (Regra de Ouro)", GOLDEN_FONT, GOLDEN_FILL, WRAP)
    for c in range(2, 8):
        write_cell(ws, r, c, None, GOLDEN_FONT, GOLDEN_FILL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    golden_rules = [
        ("REGRA", "DESCRIÇÃO"),
        ("META DIÁRIA", "40-60 atendimentos por consultor por dia"),
        ("PROPORÇÃO IDEAL", "70% clientes ativos/inativos recentes + 20% inativos antigos + 10% prospects"),
        ("AGENDA = AUTOMÁTICA", "A IA monta a lista. O consultor EXECUTA, não escolhe quem ligar"),
        ("PROSPECT TEM PRAZO", "14 dias corridos para primeira venda. Passou = volta pro pool"),
        ("NINGUÉM FICA PARADO", "Se não tem 50 ativos no dia, a IA completa com prospects"),
        ("FOLLOW-UP É SAGRADO", "Follow-up vencido SEMPRE aparece no topo, independente de tudo"),
    ]
    for i, (col1, col2) in enumerate(golden_rules):
        fill = HEADER_FILL if i == 0 else (DATA_FILL_ALT if i % 2 == 0 else None)
        font = HEADER_FONT if i == 0 else Font(name="Calibri", size=10, bold=(i==0))
        write_cell(ws, r, 1, col1, font, fill, WRAP)
        write_cell(ws, r, 2, col2, font, fill, WRAP)
        for c in range(3, 8):
            write_cell(ws, r, c, None, font, fill)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        r += 1

    # Racional
    write_cell(ws, r, 1, "💡 RACIONAL:", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    write_cell(ws, r, 2, "A agenda não é uma sugestão — é uma ordem de batalha. O consultor que escolhe quem ligar perde tempo com clientes fáceis e ignora os urgentes. A IA prioriza pelo que GERA RESULTADO, não pelo que é confortável.", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    for c in range(3, 8):
        write_cell(ws, r, c, None, RACIONAL_FONT, RACIONAL_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    # ============================================================
    # 18.1 PIRÂMIDE DE PRIORIDADE
    # ============================================================
    write_cell(ws, r, 1, "🔺 18.1 — PIRÂMIDE DE PRIORIDADE (7 Níveis)", SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None, SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    # Headers
    p_headers = ["NÍVEL", "NOME", "QUEM ENTRA", "GATILHO", "TEMPO", "SCORE BASE"]
    for i, h in enumerate(p_headers):
        write_cell(ws, r, i+1, h, HEADER_FONT, HEADER_FILL, CENTER)
    r += 1

    # Priority levels
    priorities = [
        ("P1", "CLIENTE NOVO C/ PEDIDO", "Comprou recentemente", "RESULTADO = VENDA/PEDIDO", "D+4 / D+15 / D+30", "100"),
        ("P2", "NEGOCIAÇÃO ATIVA", "Orçamento aberto, cadastro em andamento, carrinho B2B", "ESTÁGIO = ORÇAMENTO ou EM ATENDIMENTO", "Imediato", "95"),
        ("P3", "PROBLEMA / PÓS-VENDA", "Cliente com problema pendente (RNC, atraso, devolução)", "TIPO PROBLEMA preenchido", "Imediato", "90"),
        ("P4", "PERÍODO DE COMPRA", "Cliente no ciclo médio ± 7 dias", "DIAS SEM COMPRA ≈ CICLO MÉDIO", "Janela 14d", "80"),
        ("P5", "INATIVO RECENTE", "61-90 dias sem compra", "SITUAÇÃO = INAT.REC", "Semanal", "60"),
        ("P6", "INATIVO ANTIGO", ">90 dias sem compra", "SITUAÇÃO = INAT.ANT", "Quinzenal", "40"),
        ("P7", "PROSPECT / LEAD", "Nunca comprou — vindo do Mercos/SAP/indicação", "SITUAÇÃO = PROSPECT ou LEAD", "14d para 1ª venda", "30"),
    ]

    colors_p = [
        "C00000",  # P1 vermelho escuro (urgente)
        "ED7D31",  # P2 laranja
        "FFC000",  # P3 amarelo
        "70AD47",  # P4 verde
        "4472C4",  # P5 azul
        "7B7B7B",  # P6 cinza
        "9B59B6",  # P7 roxo
    ]

    for i, (nivel, nome, quem, gatilho, tempo, score) in enumerate(priorities):
        fill = PatternFill(start_color=colors_p[i], end_color=colors_p[i], fill_type="solid")
        font_white = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=10)
        alt_fill = DATA_FILL_ALT if i % 2 == 0 else None

        write_cell(ws, r, 1, nivel, font_white, fill, CENTER)
        write_cell(ws, r, 2, nome, Font(name="Calibri", size=10, bold=True), alt_fill, WRAP)
        write_cell(ws, r, 3, quem, font_data, alt_fill, WRAP)
        write_cell(ws, r, 4, gatilho, font_data, alt_fill, WRAP)
        write_cell(ws, r, 5, tempo, font_data, alt_fill, CENTER)
        write_cell(ws, r, 6, score, font_data, alt_fill, CENTER)
        r += 1

    # Racional
    r += 1
    write_cell(ws, r, 1, "💡 RACIONAL:", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    write_cell(ws, r, 2, "P1 é dinheiro na mesa — pedido fechado precisa de pós-venda cirúrgico (D+4 confirma envio, D+15 satisfação, D+30 CS). P2 é dinheiro quase na mesa — não pode esfriar. P3 é retenção — cliente com problema que não é atendido vira DETRATOR. P4 é recompra previsível — o timing certo multiplica conversão. P5/P6 são salvamento — quanto mais tempo, mais difícil. P7 é expansão — sem prospecção, a carteira encolhe.", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    for c in range(3, 8):
        write_cell(ws, r, c, None, RACIONAL_FONT, RACIONAL_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    # ============================================================
    # 18.2 JORNADA PÓS-VENDA CIRÚRGICA
    # ============================================================
    write_cell(ws, r, 1, "🎯 18.2 — JORNADA PÓS-VENDA CIRÚRGICA (D+4 / D+15 / D+30)", SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None, SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    pv_headers = ["ETAPA", "QUANDO", "OBJETIVO", "AÇÃO DO CONSULTOR", "SE NÃO ATENDER"]
    for i, h in enumerate(pv_headers):
        write_cell(ws, r, i+1, h, HEADER_FONT, HEADER_FILL, CENTER)
    r += 1

    pos_venda = [
        ("D+4", "4 dias após pedido", "Confirmar faturamento e envio",
         "Ligar/WhatsApp: 'Seu pedido foi faturado, NF xxx, previsão entrega dd/mm'",
         "T2 no dia seguinte"),
        ("D+15", "15 dias após pedido", "Pós-venda qualidade",
         "Ligar: 'Recebeu tudo certo? Algum problema? Como está o giro?'",
         "T2 em 2 dias"),
        ("D+30", "30 dias após pedido", "CS — Sucesso do Cliente",
         "Ligar: 'Como foi o sell out? Precisa de apoio MKT? Já pensou na recompra?'",
         "Follow-up D+7"),
        ("RECOMPRA", "Ciclo médio do cliente", "Criar intenção de recompra",
         "Sugestão de pedido baseada em curva ABC + sazonalidade",
         "FU D+7 ou D+15"),
    ]

    for i, (etapa, quando, obj, acao, fallback) in enumerate(pos_venda):
        alt_fill = DATA_FILL_ALT if i % 2 == 0 else None
        write_cell(ws, r, 1, etapa, Font(name="Calibri", size=10, bold=True), alt_fill, CENTER)
        write_cell(ws, r, 2, quando, DATA_FONT, alt_fill, WRAP)
        write_cell(ws, r, 3, obj, DATA_FONT, alt_fill, WRAP)
        write_cell(ws, r, 4, acao, DATA_FONT, alt_fill, WRAP)
        write_cell(ws, r, 5, fallback, DATA_FONT, alt_fill, WRAP)
        r += 1

    r += 1
    write_cell(ws, r, 1, "💡 RACIONAL:", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    write_cell(ws, r, 2, "O pós-venda é o que transforma VENDA ÚNICA em CLIENTE RECORRENTE. D+4 resolve problemas antes que virem reclamação. D+15 mostra que a empresa se importa. D+30 planta a semente da recompra. Sem essa jornada, o cliente compra 1x e some — com ela, vira fiel e compra 4-6x/ano.", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    for c in range(3, 8):
        write_cell(ws, r, c, None, RACIONAL_FONT, RACIONAL_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    # ============================================================
    # 18.3 REGRA DO IGNORADOR
    # ============================================================
    write_cell(ws, r, 1, "🚫 18.3 — REGRA DO IGNORADOR (Penalização por Não-Resposta)", SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None, SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    ig_headers = ["TENTATIVA", "SITUAÇÃO", "AÇÃO", "RESULTADO", "PRÓXIMO"]
    for i, h in enumerate(ig_headers):
        write_cell(ws, r, i+1, h, HEADER_FONT, HEADER_FILL, CENTER)
    r += 1

    ignorador = [
        ("T1", "WhatsApp enviado", "Esperar resposta", "NÃO RESPONDE", "T2 em 24h"),
        ("T2", "Ligação feita", "Tentar contato direto", "NÃO ATENDE", "T3 em 48h"),
        ("T3", "WhatsApp + Ligação", "Última tentativa ativa", "NÃO RESPONDE", "T4 em 48h"),
        ("T4", "Ligação final", "Encerramento protocolar", "RECUSOU / NÃO ATENDE", "NUTRIÇÃO"),
        ("NUTRIÇÃO", "Email + WhatsApp passivo", "Ciclo de 15 dias", "Campanha/novidade", "RESET se responder"),
    ]

    for i, (tent, sit, acao, res, prox) in enumerate(ignorador):
        alt_fill = DATA_FILL_ALT if i % 2 == 0 else None
        write_cell(ws, r, 1, tent, Font(name="Calibri", size=10, bold=True), alt_fill, CENTER)
        write_cell(ws, r, 2, sit, DATA_FONT, alt_fill, WRAP)
        write_cell(ws, r, 3, acao, DATA_FONT, alt_fill, WRAP)
        write_cell(ws, r, 4, res, DATA_FONT, alt_fill, WRAP)
        write_cell(ws, r, 5, prox, DATA_FONT, alt_fill, WRAP)
        r += 1

    r += 1
    write_cell(ws, r, 1, "💡 RACIONAL:", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    write_cell(ws, r, 2, "Cliente que não responde em T3/T4 PERDE PRIORIDADE na agenda. A IA não vai gastar 4 dos 50 slots do dia com alguém que ignora — vai priorizar quem tem chance de conversão. O ignorador vai para NUTRIÇÃO automática (emails, promos) e só volta à agenda ativa quando DER SINAL (abrir email, acessar B2B, responder WhatsApp). Isso protege a produtividade do consultor.", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    for c in range(3, 8):
        write_cell(ws, r, c, None, RACIONAL_FONT, RACIONAL_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    # ============================================================
    # 18.4 REGRA DO EQUILÍBRIO DE META
    # ============================================================
    write_cell(ws, r, 1, "⚖️ 18.4 — REGRA DO EQUILÍBRIO DE META", SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None, SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    meta_headers = ["CENÁRIO", "% META", "COMPOSIÇÃO AGENDA", "AÇÃO AUTOMÁTICA"]
    for i, h in enumerate(meta_headers):
        write_cell(ws, r, i+1, h, HEADER_FONT, HEADER_FILL, CENTER)
    r += 1

    meta_rules = [
        ("META BATIDA", "≥ 100%", "70% ativos + 20% inativos + 10% prospects", "Manter ritmo — foco em CS e recompra"),
        ("META NO CAMINHO", "70-99%", "60% ativos + 25% inativos + 15% prospects", "Aumentar prospecção levemente"),
        ("META EM RISCO", "40-69%", "50% ativos + 20% inativos + 30% prospects", "Prospecção URGENTE — 30% da agenda"),
        ("META CRÍTICA", "< 40%", "40% ativos + 10% inativos + 50% prospects", "MODO PROSPECÇÃO — metade da agenda é prospect"),
    ]

    meta_colors = ["00B050", "FFC000", "ED7D31", "FF0000"]

    for i, (cenario, pct, composicao, acao) in enumerate(meta_rules):
        fill = PatternFill(start_color=meta_colors[i], end_color=meta_colors[i], fill_type="solid")
        font_w = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        write_cell(ws, r, 1, cenario, font_w, fill, CENTER)
        write_cell(ws, r, 2, pct, Font(name="Calibri", size=10, bold=True), None, CENTER)
        write_cell(ws, r, 3, composicao, DATA_FONT, DATA_FILL_ALT if i % 2 == 0 else None, WRAP)
        write_cell(ws, r, 4, acao, DATA_FONT, DATA_FILL_ALT if i % 2 == 0 else None, WRAP)
        r += 1

    r += 1
    write_cell(ws, r, 1, "💡 RACIONAL:", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    write_cell(ws, r, 2, "Se o consultor está com meta batida, não faz sentido forçar prospecção — melhor investir em CS e recompra para SEGURAR a meta. Mas se está em 40%, ficar ligando só para ativos que já compraram é INEFICIENTE — precisa de VOLUME NOVO. A IA ajusta a composição da agenda automaticamente conforme o atingimento de meta de cada consultor.", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    for c in range(3, 8):
        write_cell(ws, r, c, None, RACIONAL_FONT, RACIONAL_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    # ============================================================
    # 18.5 REGRA DE DESEMPATE
    # ============================================================
    write_cell(ws, r, 1, "🏅 18.5 — REGRA DE DESEMPATE (Critérios de Ordenação)", SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None, SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    desemp_headers = ["ORDEM", "CRITÉRIO", "LÓGICA", "EXEMPLO"]
    for i, h in enumerate(desemp_headers):
        write_cell(ws, r, i+1, h, HEADER_FONT, HEADER_FILL, CENTER)
    r += 1

    desempate = [
        ("1º", "NÍVEL DE PRIORIDADE", "P1 sempre antes de P2, P2 antes de P3, etc.", "P1 (pedido novo) > P4 (período compra)"),
        ("2º", "FOLLOW-UP VENCIDO", "Quem tem FU vencido sobe ao topo do seu nível", "FU de ontem > FU de amanhã"),
        ("3º", "CURVA ABC", "A > B > C dentro do mesmo nível", "Curva A com 50d > Curva C com 50d"),
        ("4º", "TICKET MÉDIO", "Maior ticket primeiro (mesmo ABC)", "R$5.000/pedido > R$800/pedido"),
        ("5º", "TIPO CLIENTE", "Fidelizado > Em Desenvolvimento > Novo", "Fidelizado 60d > Novo 60d"),
        ("6º", "DIAS SEM FOLLOW-UP", "Mais tempo sem contato = mais urgente", "Último FU 30d atrás > 5d atrás"),
        ("7º", "SINAL DE COMPRA", "Quem acessou B2B/carrinho sobe", "Acesso B2B ontem > sem acesso"),
    ]

    for i, (ordem, criterio, logica, exemplo) in enumerate(desempate):
        alt_fill = DATA_FILL_ALT if i % 2 == 0 else None
        write_cell(ws, r, 1, ordem, Font(name="Calibri", size=10, bold=True), alt_fill, CENTER)
        write_cell(ws, r, 2, criterio, Font(name="Calibri", size=10, bold=True), alt_fill, WRAP)
        write_cell(ws, r, 3, logica, DATA_FONT, alt_fill, WRAP)
        write_cell(ws, r, 4, exemplo, DATA_FONT, alt_fill, WRAP)
        r += 1

    r += 1
    write_cell(ws, r, 1, "💡 RACIONAL:", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    write_cell(ws, r, 2, "O desempate garante que dois clientes P4 (período de compra) não fiquem na mesma posição. O Curva A com ticket alto e sinal de B2B vai aparecer ANTES do Curva C sem atividade. É uma ordenação em cascata — cada critério refina o anterior. O consultor não precisa pensar — a lista já vem na ordem PERFEITA.", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    for c in range(3, 8):
        write_cell(ws, r, c, None, RACIONAL_FONT, RACIONAL_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    # ============================================================
    # 18.6 QUOTAS MÍNIMAS POR TIPO
    # ============================================================
    write_cell(ws, r, 1, "📊 18.6 — QUOTAS MÍNIMAS POR TIPO DE ATENDIMENTO", SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None, SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    q_headers = ["TIPO", "MÍN/DIA", "MÁX/DIA", "QUANDO PREENCHE", "FONTE"]
    for i, h in enumerate(q_headers):
        write_cell(ws, r, i+1, h, HEADER_FONT, HEADER_FILL, CENTER)
    r += 1

    quotas = [
        ("PÓS-VENDA (D+4/D+15/D+30)", "5", "15", "Sempre que houver pedidos recentes", "DRAFT 2 → RESULTADO=VENDA"),
        ("NEGOCIAÇÃO ATIVA", "3", "10", "Sempre que houver orçamentos/cadastros abertos", "CARTEIRA → ESTÁGIO=ORÇAMENTO"),
        ("ATIVOS EM PERÍODO", "15", "30", "Clientes com ciclo vencendo ± 7d", "CARTEIRA → DIAS ≈ CICLO"),
        ("INATIVOS RECENTES", "3", "10", "61-90 dias sem compra", "CARTEIRA → SITUAÇÃO=INAT.REC"),
        ("INATIVOS ANTIGOS", "2", "5", "Se sobrar espaço após P1-P5", "CARTEIRA → SITUAÇÃO=INAT.ANT"),
        ("PROSPECTS", "5", "25", "Sempre — mínimo 5 por dia por consultor", "CARTEIRA → SITUAÇÃO=PROSPECT"),
        ("PROBLEMAS/SUPORTE", "0", "5", "Quando houver RNC pendente", "CARTEIRA → TIPO PROBLEMA preenchido"),
    ]

    for i, (tipo, mn, mx, quando, fonte) in enumerate(quotas):
        alt_fill = DATA_FILL_ALT if i % 2 == 0 else None
        write_cell(ws, r, 1, tipo, Font(name="Calibri", size=10, bold=True), alt_fill, WRAP)
        write_cell(ws, r, 2, mn, DATA_FONT, alt_fill, CENTER)
        write_cell(ws, r, 3, mx, DATA_FONT, alt_fill, CENTER)
        write_cell(ws, r, 4, quando, DATA_FONT, alt_fill, WRAP)
        write_cell(ws, r, 5, fonte, DATA_FONT, alt_fill, WRAP)
        r += 1

    r += 1
    write_cell(ws, r, 1, "💡 RACIONAL:", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    write_cell(ws, r, 2, "As quotas garantem DIVERSIDADE na agenda. Sem elas, um consultor com muitos ativos NUNCA prospecta (e a carteira encolhe). Com quotas, mesmo consultores lotados fazem pelo menos 5 prospecções/dia. E se a meta está crítica, as quotas de prospect SOBEM automaticamente (Regra 18.4). Total: 33-100 slots dependendo do cenário — a IA distribui dentro do mín/máx conforme a situação de cada consultor.", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    for c in range(3, 8):
        write_cell(ws, r, c, None, RACIONAL_FONT, RACIONAL_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    # ============================================================
    # 18.7 DISTRIBUIÇÃO DE PROSPECTS POR MACRORREGIÃO
    # ============================================================
    write_cell(ws, r, 1, "🗺️ 18.7 — DISTRIBUIÇÃO DE PROSPECTS POR CONSULTOR (MACRORREGIÃO)", SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None, SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    dist_headers = ["CONSULTOR", "TERRITÓRIO", "UFs PRINCIPAIS", "FOCO ESPECIAL", "PROSPECTS ESTIMADOS"]
    for i, h in enumerate(dist_headers):
        write_cell(ws, r, i+1, h, HEADER_FONT, HEADER_FILL, CENTER)
    r += 1

    distribuicao = [
        ("MANU DITZEL", "Sul", "SC, PR, RS", "Base mais densa — região core Vitao", "~1.900 (35%)"),
        ("LARISSA PADILHA", "Sudeste + Norte + Nordeste", "SP, RJ, MG, BA, PA, CE, PE, AL, MA", "Maior território — mais volume", "~2.800 (50%)"),
        ("JULIO GADRET", "Redes e Franquias", "Nacional (Cia Saúde, Fitland, etc.)", "Foco em REDES — menos prospects avulsos", "~300 (5%)"),
        ("DAIANE STADLER", "Operacional + Redes", "Nacional (suporte)", "Redes menores + apoio operacional", "~500 (10%)"),
    ]

    for i, (nome, terr, ufs, foco, est) in enumerate(distribuicao):
        alt_fill = DATA_FILL_ALT if i % 2 == 0 else None
        write_cell(ws, r, 1, nome, Font(name="Calibri", size=10, bold=True), alt_fill, WRAP)
        write_cell(ws, r, 2, terr, DATA_FONT, alt_fill, WRAP)
        write_cell(ws, r, 3, ufs, DATA_FONT, alt_fill, WRAP)
        write_cell(ws, r, 4, foco, DATA_FONT, alt_fill, WRAP)
        write_cell(ws, r, 5, est, DATA_FONT, alt_fill, CENTER)
        r += 1

    r += 1
    write_cell(ws, r, 1, "💡 RACIONAL:", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    write_cell(ws, r, 2, "A distribuição por UF garante que cada consultor prospecta na SUA região. Prospects sem UF ou de regiões indefinidas vão para LARISSA (maior território). Julio e Daiane têm menos prospects avulsos porque focam em REDES — seus prospects são majoritariamente referências de redes existentes. A distribuição exata será calculada automaticamente pela coluna CONSULTOR na CARTEIRA baseada no UF de cada prospect.", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    for c in range(3, 8):
        write_cell(ws, r, c, None, RACIONAL_FONT, RACIONAL_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    # ============================================================
    # 18.8 FLUXO DIÁRIO DA AGENDA
    # ============================================================
    write_cell(ws, r, 1, "🔄 18.8 — FLUXO DIÁRIO DA AGENDA (Ciclo Operacional)", SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None, SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    fluxo_headers = ["HORÁRIO", "ETAPA", "QUEM FAZ", "O QUE ACONTECE"]
    for i, h in enumerate(fluxo_headers):
        write_cell(ws, r, i+1, h, HEADER_FONT, HEADER_FILL, CENTER)
    r += 1

    fluxo = [
        ("07:00", "GERAÇÃO AUTOMÁTICA", "IA (SORTBY/FILTER)", "Agenda popula 40-60 clientes por consultor, ordenados por SCORE"),
        ("08:00-12:00", "EXECUÇÃO MANHÃ", "CONSULTOR", "Liga/WhatsApp os clientes da lista na ORDEM. Preenche RESULTADO, MOTIVO, AÇÃO"),
        ("12:00-13:00", "INTERVALO", "—", "—"),
        ("13:00-17:00", "EXECUÇÃO TARDE", "CONSULTOR", "Continua a lista. Prospects e follow-ups pendentes da manhã"),
        ("17:00-18:00", "FECHAMENTO", "CONSULTOR", "Completa campos pendentes. Revisa se todos foram preenchidos"),
        ("18:00", "CONSOLIDAÇÃO", "IA (RECORTE)", "Atendimentos do dia são recortados da AGENDA → DRAFT 2 (histórico)"),
        ("NOITE", "RECALCULAÇÃO", "EXCEL (FORMULAS)", "Fórmulas recalculam SCORE, SINALEIRO, SITUAÇÃO para próximo dia"),
    ]

    for i, (hora, etapa, quem, oq) in enumerate(fluxo):
        alt_fill = DATA_FILL_ALT if i % 2 == 0 else None
        write_cell(ws, r, 1, hora, Font(name="Calibri", size=10, bold=True), alt_fill, CENTER)
        write_cell(ws, r, 2, etapa, Font(name="Calibri", size=10, bold=True), alt_fill, WRAP)
        write_cell(ws, r, 3, quem, DATA_FONT, alt_fill, CENTER)
        write_cell(ws, r, 4, oq, DATA_FONT, alt_fill, WRAP)
        r += 1

    r += 1
    write_cell(ws, r, 1, "💡 RACIONAL:", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    write_cell(ws, r, 2, "O ciclo garante que NENHUM atendimento se perde. Manhã = execução (consultor liga). Noite = consolidação (IA processa). Próximo dia = agenda nova baseada nos resultados de ontem. Se o consultor vendeu para 3 prospects, amanhã a agenda já ajusta: menos prospects (meta subiu), mais pós-venda (D+4 dos novos pedidos). É um ciclo VIVO que se adapta todo dia.", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    for c in range(3, 8):
        write_cell(ws, r, c, None, RACIONAL_FONT, RACIONAL_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    # ============================================================
    # 18.9 SCORE RANKING — FÓRMULA DETALHADA
    # ============================================================
    write_cell(ws, r, 1, "📐 18.9 — SCORE RANKING — FÓRMULA DETALHADA DE PONTUAÇÃO", SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None, SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    write_cell(ws, r, 1, "SCORE = (URGÊNCIA × 0.30) + (VALOR × 0.25) + (FOLLOWUP × 0.20) + (SINAL × 0.15) + (TENTATIVA × 0.05) + (SITUAÇÃO × 0.05)",
               Font(name="Calibri", size=11, bold=True, color="1F4E79"),
               PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), WRAP)
    for c in range(2, 8):
        write_cell(ws, r, c, None,
                   Font(name="Calibri", size=11, bold=True, color="1F4E79"),
                   PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 2

    # Detalhamento de cada fator
    fatores = [
        ("URGÊNCIA TEMPORAL (30%)", [
            ("FATOR", "FAIXA", "PONTOS", "LÓGICA"),
            ("Dias/Ciclo < 0.7", "Longe do ciclo", "0", "Não precisa ligar agora — ciclo não venceu"),
            ("Dias/Ciclo = 0.7-1.0", "Período de compra", "30", "Janela ideal — probabilidade alta de recompra"),
            ("Dias/Ciclo = 1.0-1.5", "Atrasado", "60", "Passou do ciclo — risco de perder"),
            ("Dias/Ciclo > 1.5", "Muito atrasado", "90", "URGENTE — cliente está escapando"),
            ("Follow-up vencido", "Qualquer", "100", "MÁXIMO — compromisso assumido"),
        ]),
        ("VALOR DO CLIENTE (25%)", [
            ("PERFIL", "CURVA + TIPO", "PONTOS", "LÓGICA"),
            ("Curva A + Fidelizado", "Top cliente", "100", "Maior valor + maior frequência"),
            ("Curva A + Em Desenvolvimento", "Alto potencial", "80", "Ticket alto, precisa fidelizar"),
            ("Curva B + Fidelizado", "Bom cliente", "60", "Ticket médio, frequente"),
            ("Curva B + Recorrente", "Cliente padrão", "50", "Base sólida da carteira"),
            ("Curva C", "Baixo ticket", "20", "Volume alto mas ticket baixo"),
            ("Prospect", "Sem compra", "15", "Potencial desconhecido"),
        ]),
        ("FOLLOW-UP VENCIDO (20%)", [
            ("PRAZO", "SITUAÇÃO", "PONTOS", "LÓGICA"),
            ("Vencido hoje", "É AGORA", "100", "Compromisso assumido — prioridade máxima"),
            ("Vence em 1-3 dias", "Quase vencendo", "80", "Permite antecipar se houver espaço"),
            ("Vence em 4-7 dias", "Próximo", "50", "Pode entrar se o dia tiver espaço"),
            ("Vence em 8+ dias", "Longe", "20", "Não urgente"),
            ("Sem follow-up", "Nenhum agendado", "40", "Precisa de contato (esquecido)"),
        ]),
        ("SINAL DE COMPRA (15%)", [
            ("SINAL", "FONTE", "PONTOS", "LÓGICA"),
            ("Carrinho B2B + Acesso recente", "E-commerce", "100", "COMPRANDO AGORA — interromper = perda"),
            ("Acesso B2B esta semana", "E-commerce", "70", "Demonstrou interesse ativo"),
            ("Temperatura QUENTE", "Consultor", "60", "Consultor avaliou como quente"),
            ("Temperatura MORNO", "Consultor", "40", "Há interesse mas não urgente"),
            ("Temperatura FRIO", "Consultor", "10", "Sem sinais de compra"),
            ("Sem dados", "Nenhuma", "0", "Sem informação"),
        ]),
    ]

    for fator_nome, linhas in fatores:
        write_cell(ws, r, 1, fator_nome, BLOCO_FONT, BLOCO_FILL, WRAP)
        for c in range(2, 8):
            write_cell(ws, r, c, None, BLOCO_FONT, BLOCO_FILL)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 1

        for i, vals in enumerate(linhas):
            is_header = (i == 0)
            font = HEADER_FONT if is_header else DATA_FONT
            fill = HEADER_FILL if is_header else (DATA_FILL_ALT if i % 2 == 0 else None)
            for j, v in enumerate(vals):
                write_cell(ws, r, j+1, v, font, fill, WRAP if j > 1 else CENTER)
            r += 1
        r += 1

    # TENTATIVA + SITUAÇÃO (menores pesos - resumido)
    write_cell(ws, r, 1, "TENTATIVA (5%) + SITUAÇÃO (5%)", BLOCO_FONT, BLOCO_FILL, WRAP)
    for c in range(2, 8):
        write_cell(ws, r, c, None, BLOCO_FONT, BLOCO_FILL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    mini = [
        ("FATOR", "VALOR", "PONTOS"),
        ("TENTATIVA T4", "Última tentativa", "100"),
        ("TENTATIVA T3", "Penúltima", "50"),
        ("TENTATIVA T1/T2", "Início", "10"),
        ("SITUAÇÃO EM RISCO", "51-60 dias", "80"),
        ("SITUAÇÃO ATIVO", "≤50 dias", "40"),
        ("SITUAÇÃO INAT", ">60 dias", "20"),
        ("SITUAÇÃO PROSPECT", "Nunca comprou", "10"),
    ]
    for i, vals in enumerate(mini):
        is_header = (i == 0)
        font = HEADER_FONT if is_header else DATA_FONT
        fill = HEADER_FILL if is_header else (DATA_FILL_ALT if i % 2 == 0 else None)
        for j, v in enumerate(vals):
            write_cell(ws, r, j+1, v, font, fill, CENTER if j != 1 else WRAP)
        r += 1

    r += 1
    write_cell(ws, r, 1, "💡 RACIONAL:", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    write_cell(ws, r, 2, "O SCORE transforma 6 fatores subjetivos em 1 número objetivo de 0-100. Urgência tem peso maior (30%) porque TIMING é tudo em vendas — ligar no momento certo vale mais que ligar para o maior ticket. Follow-up (20%) é sagrado porque é COMPROMISSO. Tentativa e Situação têm 5% cada porque são refinamentos, não decisores.", RACIONAL_FONT, RACIONAL_FILL, WRAP)
    for c in range(3, 8):
        write_cell(ws, r, c, None, RACIONAL_FONT, RACIONAL_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 2

    # ============================================================
    # 18.10 RESUMO VISUAL
    # ============================================================
    write_cell(ws, r, 1, "📋 18.10 — RESUMO: COMO A AGENDA SE MONTA AUTOMATICAMENTE", SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    for c in range(2, 8):
        write_cell(ws, r, c, None, SUBTITLE_FONT, SUBTITLE_FILL, CENTER)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    resumo = [
        ("PASSO", "AÇÃO", "RESULTADO"),
        ("1", "IA lê CARTEIRA completa (6.143 clientes)", "Pool de todos os clientes com dados atualizados"),
        ("2", "Filtra por CONSULTOR (território por UF)", "Cada consultor vê apenas SUA carteira"),
        ("3", "Calcula SCORE para cada cliente (6 fatores)", "Ranking numérico 0-100"),
        ("4", "Aplica PIRÂMIDE DE PRIORIDADE (P1→P7)", "Clientes agrupados por urgência"),
        ("5", "Aplica REGRA DO IGNORADOR (T3/T4 penalizado)", "Quem não responde perde posição"),
        ("6", "Aplica REGRA DE META (ajusta % prospects)", "Composição da agenda adapta à meta"),
        ("7", "Aplica QUOTAS MÍNIMAS por tipo", "Garante diversidade de atendimentos"),
        ("8", "Aplica DESEMPATE (ABC, ticket, tipo, FU)", "Ordenação final dentro de cada nível"),
        ("9", "Corta nos primeiros 40-60 clientes", "AGENDA DO DIA está pronta"),
        ("10", "Consultor EXECUTA na ordem da lista", "Preenche resultado, motivo, ação"),
        ("11", "Final do dia: RECORTE → DRAFT 2", "Histórico consolidado"),
        ("12", "NOITE: recalcula tudo para amanhã", "Ciclo recomeça"),
    ]

    for i, vals in enumerate(resumo):
        is_header = (i == 0)
        font = HEADER_FONT if is_header else DATA_FONT
        fill = HEADER_FILL if is_header else (DATA_FILL_ALT if i % 2 == 0 else None)
        bold = Font(name="Calibri", size=10, bold=True) if not is_header else font
        write_cell(ws, r, 1, vals[0], bold if not is_header else font, fill, CENTER)
        write_cell(ws, r, 2, vals[1], font, fill, WRAP)
        write_cell(ws, r, 3, vals[2], font, fill, WRAP)
        r += 1

    r += 2

    # ============================================================
    # ATUALIZAR ÍNDICE
    # ============================================================
    print("[4] Atualizando ÍNDICE...", flush=True)
    ws.cell(row=20, column=9, value="🧠").font = Font(name="Calibri", size=10)
    ws.cell(row=20, column=10, value="18. METODOLOGIA PRIORIZAÇÃO").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=20, column=11, value="10 blocos").font = Font(name="Calibri", size=10)

    # ============================================================
    # AJUSTAR LARGURA DAS COLUNAS
    # ============================================================
    print("[5] Ajustando larguras...", flush=True)
    # Não mexer nas larguras existentes das colunas A-G que já estavam definidas
    # Só garantir que as novas seções sejam legíveis

    # ============================================================
    # SALVAR
    # ============================================================
    print(f"[6] Salvando V19...", flush=True)
    wb.save(V19)
    wb.close()

    size_mb = Path(V19).stat().st_size / (1024*1024)
    elapsed = (datetime.now() - t).total_seconds()

    print(f"\n{'='*100}")
    print(f"[SUCESSO] V19 gerado com Seção 18 completa!")
    print(f"  Arquivo: {V19}")
    print(f"  Tamanho: {size_mb:.2f} MB")
    print(f"  Seção 18: rows {START}-{r} ({r-START} linhas)")
    print(f"  10 blocos: Golden Rule, Pirâmide, Pós-Venda, Ignorador,")
    print(f"             Equilíbrio Meta, Desempate, Quotas, Distribuição,")
    print(f"             Fluxo Diário, Score Detalhado, Resumo")
    print(f"  Tempo: {elapsed:.1f}s")
    print(f"{'='*100}")


if __name__ == "__main__":
    main()

import openpyxl
wb = openpyxl.load_workbook(r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10\CRM_VITAO360_V43_CLEAN.xlsx', data_only=True)
ws = wb['PROJEÇÃO ']

print('=== PROJECAO headers row 3 (cols 12-24) ===')
for c in range(12, 25):
    v = ws.cell(row=3, column=c).value
    print(f'  Col {c}: {v}')

print('\n=== PROJECAO data rows 4-8 ===')
for r in range(4, 9):
    nome = ws.cell(row=r, column=2).value
    meta_a = ws.cell(row=r, column=12).value
    meta_jan = ws.cell(row=r, column=13).value
    meta_fev = ws.cell(row=r, column=14).value
    meta_mar = ws.cell(row=r, column=15).value
    print(f'  Row {r}: {str(nome)[:25]} | Anual={meta_a} | Jan={meta_jan} | Fev={meta_fev} | Mar={meta_mar}')

print('\n=== REALIZADO headers (cols 26-40) ===')
for c in range(26, 42):
    v = ws.cell(row=3, column=c).value
    if v:
        print(f'  Col {c}: {v}')

print('\n=== REALIZADO data rows 4-5 ===')
for r in range(4, 6):
    vals = []
    for c in range(26, 42):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            vals.append(f'{c}={v}')
    print(f'  Row {r}: {vals if vals else "VAZIO (zerado ok)"}')

# Count how many have meta anual > 0
meta_count = 0
meta_zero = 0
for r in range(4, ws.max_row + 1):
    v = ws.cell(row=r, column=12).value
    if v and isinstance(v, (int, float)) and v > 0:
        meta_count += 1
    elif ws.cell(row=r, column=1).value:
        meta_zero += 1

print(f'\n=== RESUMO ===')
print(f'Clientes com META ANUAL > 0: {meta_count}')
print(f'Clientes com META ANUAL = 0 ou None: {meta_zero}')

# Check AGENDA headers
for agenda_name in ['AGENDA LARISSA', 'AGENDA DAIANE', 'AGENDA MANU', 'AGENDA JULIO']:
    ws_a = wb[agenda_name]
    print(f'\n=== {agenda_name} ===')
    for r in range(1, ws_a.max_row + 1):
        vals = []
        for c in range(1, min(ws_a.max_column + 1, 20)):
            v = ws_a.cell(row=r, column=c).value
            if v:
                vals.append(f'{c}={str(v)[:20]}')
        print(f'  Row {r}: {", ".join(vals[:15])}')

wb.close()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CRM VITAO360 — V27: Simulação Realista de Atendimentos JAN/25
==============================================================
Base: V26
Motor de simulação que gera ~1.229 registros realistas para JAN/25
seguindo TODAS as regras de negócio do CRM.

Pools:
  A: 32 vendas reais (ciclo completo 7 ações)
  B: 120 futuros compradores FEV/MAR/ABR (ciclo parcial 3 ações)
  C: 258 perdas aleatórias (2-3 tentativas T1→T2→T3)

Cascata: Gera ações futuras para FEV/MAR
"""

import openpyxl
from datetime import date, datetime, timedelta
from collections import defaultdict, Counter
import random
import os
import sys

# ============================================================
# CONFIGURAÇÃO
# ============================================================
INPUT_FILE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10\CRM_VITAO360_V26_FINAL.xlsx'
OUTPUT_FILE = r'c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\output\phase10\CRM_VITAO360_V27_FINAL.xlsx'

random.seed(42)  # Reprodutível

CONSULTOR_JAN = "DAIANE STAVICKI"

# ============================================================
# REGRAS DE NEGÓCIO (extraídas da aba REGRAS)
# ============================================================

# RESULTADO → (FOLLOW_UP_DIAS, GRUPO_DASH, TEMPERATURA, TIPO_AÇÃO)
RESULTADO_REGRAS = {
    "VENDA / PEDIDO":          (4,  "VENDAS",        "🟢 QUENTE",  "VENDA"),
    "EM ATENDIMENTO":          (2,  "ATENDIMENTO",   "🟡 MORNO",   "ATENDIMENTO"),
    "SUPORTE":                 (1,  "SUPORTE",       "🟡 MORNO",   "SUPORTE"),
    "ORÇAMENTO":               (3,  "ATENDIMENTO",   "🟡 MORNO",   "NEGOCIAÇÃO"),
    "RELACIONAMENTO":          (7,  "RELACIONAMENTO", "🟢 QUENTE", "PÓS-VENDA"),
    "CADASTRO":                (2,  "ATENDIMENTO",   "🟡 MORNO",   "ATENDIMENTO"),
    "NÃO ATENDE":              (2,  "TENTATIVA",     "🔴 FRIO",    "TENTATIVA"),
    "NÃO RESPONDE":            (3,  "TENTATIVA",     "🔴 FRIO",    "TENTATIVA"),
    "CS (SUCESSO DO CLIENTE)":  (15, "CS",            "🟢 QUENTE",  "PÓS-VENDA"),
    "PERDA / FECHOU LOJA":     (30, "PERDA",         "🔴 FRIO",    "PERDA/NUTRIÇÃO"),
    "FOLLOW UP 7":             (7,  "FOLLOW-UP",     "🟡 MORNO",   "PÓS-VENDA"),
    "FOLLOW UP 15":            (15, "FOLLOW-UP",     "🟡 MORNO",   "PÓS-VENDA"),
    "NUTRIÇÃO":                (30, "PERDA",         "🔴 FRIO",    "PERDA/NUTRIÇÃO"),
}

# AÇÕES FUTURAS padrão por resultado
# AÇÕES FUTURAS — da aba REGRAS Seção 13 + Motor de Regras (Seção 17)
ACAO_FUTURA_MAP = {
    "VENDA / PEDIDO":           "CONFIRMAR FATURAMENTO E EXPEDIÇÃO",
    "EM ATENDIMENTO":           "FECHAR NEGOCIAÇÃO EM ANDAMENTO",
    "SUPORTE":                  "RESOLVER PROBLEMA INTERNO E ENVIAR SOLUÇÃO",
    "ORÇAMENTO":                "CONFIRMAR ORÇAMENTO ENVIADO",
    "RELACIONAMENTO":           "RAPPORT COM CLIENTE APÓS A VENDA ATÉ RECOMPRAR",
    "CADASTRO":                 "CONFIRMAR CADASTRO NO SISTEMA",
    "NÃO ATENDE":               "2ª TENTATIVA VIA LIGAÇÃO",
    "NÃO RESPONDE":             "2ª TENTATIVA VIA WHATSAPP",
    "CS (SUCESSO DO CLIENTE)":  "VERIFICAR SELL OUT E CRIAR INTENÇÃO RECOMPRA",
    "PERDA / FECHOU LOJA":     "NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES",
    "FOLLOW UP 7":              "COBRAR RETORNO DO CLIENTE",
    "FOLLOW UP 15":             "COBRAR RETORNO DO CLIENTE",
    "NUTRIÇÃO":                 "NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES",
    "PÓS-VENDA":               "FAZER CS (SUCESSO DO CLIENTE)",
}

# TIPO DO CONTATO — da aba REGRAS Seção 2 (7 valores)
TIPO_CONTATO_MAP = {
    "PROSPECÇÃO":                "PROSPECÇÃO",
    "APRESENTAÇÃO":             "PROSPECÇÃO",
    "ORÇAMENTO":                "NEGOCIAÇÃO",
    "CADASTRO":                 "NEGOCIAÇÃO",
    "NEGOCIAÇÃO":               "NEGOCIAÇÃO",
    "VENDA":                    "NEGOCIAÇÃO",
    "PÓS-VENDA":               "PÓS-VENDA / RELACIONAMENTO",
    "CS":                       "PÓS-VENDA / RELACIONAMENTO",
    "FOLLOW-UP":                "FOLLOW UP",
    "TENTATIVA":                "PROSPECÇÃO",
    "PERDA":                    "PERDA / NUTRIÇÃO",
    "ATEND_ATIVOS":             "ATEND. CLIENTES ATIVOS",
    "ATEND_INATIVOS":           "ATEND. CLIENTES INATIVOS",
}

# NOTAS DO DIA realistas
NOTAS_PROSPECÇÃO = [
    "Primeiro contato via WhatsApp. Apresentei a empresa e produtos.",
    "Enviei mensagem de apresentação pelo WhatsApp com catálogo.",
    "Contato inicial. Cliente demonstrou interesse em conhecer a linha.",
    "WhatsApp enviado. Aguardando retorno do responsável de compras.",
    "Fiz a apresentação inicial da Vitão. Cliente pediu mais informações.",
    "Mandei WhatsApp com portfólio. Retorno positivo, quer saber preços.",
]
NOTAS_ORCAMENTO = [
    "Enviei orçamento conforme solicitado. Aguardando aprovação.",
    "Proposta enviada por e-mail. Follow-up em 3 dias.",
    "Cliente pediu ajuste no orçamento. Revisei e reenviei.",
    "Orçamento aprovado pelo financeiro. Aguardando pedido.",
]
NOTAS_CADASTRO = [
    "Iniciando cadastro no sistema. Solicitei documentação.",
    "Cadastro em andamento. Pendente análise financeira.",
    "Documentação recebida. Cadastro enviado para aprovação.",
    "Cadastro aprovado. Liberado para pedido.",
]
NOTAS_VENDA = [
    "Pedido fechado! Cliente satisfeito com as condições.",
    "Venda realizada. Pedido digitado no SAP.",
    "Fechamos o primeiro pedido. Boa abertura de conta!",
    "Pedido confirmado. Previsão de entrega em 5 dias úteis.",
]
NOTAS_POS_VENDA = [
    "Acompanhamento pós-venda D+4. Mercadoria chegou OK.",
    "Follow-up pós-venda. Cliente recebeu e está satisfeito.",
    "Verificação D+15. Produto girando bem na loja.",
    "Acompanhamento de satisfação. Tudo certo, prepararemos reposição.",
]
NOTAS_NAO_RESPONDE = [
    "WhatsApp enviado mas sem retorno.",
    "Mensagem enviada, cliente visualizou mas não respondeu.",
    "Tentei contato pelo WhatsApp, sem resposta.",
    "Sem retorno da mensagem. Vou tentar ligação.",
]
NOTAS_NAO_ATENDE = [
    "Ligação não atendida. Vou tentar novamente.",
    "Tentei ligar, caiu na caixa postal.",
    "Número chamou mas ninguém atendeu. Reagendar.",
    "Ligação sem atendimento. Próxima tentativa por WhatsApp.",
]
NOTAS_PERDA = [
    "Cliente informou que não tem interesse no momento.",
    "Loja fechou recentemente. Sem previsão de reabertura.",
    "Cliente já trabalha com concorrente e não quer trocar.",
    "Responsável disse que o momento não é propício.",
    "Sem interesse. Colocar em nutrição para recontato futuro.",
]
NOTAS_EM_ATENDIMENTO = [
    "Em negociação. Cliente analisando proposta.",
    "Atendimento em andamento. Discutindo condições.",
    "Cliente interessado, pediu prazo para decidir.",
    "Retorno positivo, mas precisa de aprovação interna.",
]

# ============================================================
# MOTIVOS por tipo de resultado
# ============================================================
# MOTIVOS — da aba REGRAS Seção 3 (22 valores reais)
MOTIVOS_EM_ATENDIMENTO = [
    "PRIMEIRO CONTATO / SEM RESPOSTA",
    "AINDA TEM ESTOQUE",
    "SÓ QUER COMPRAR GRANEL",
    "NÃO SE APLICA",
]
MOTIVOS_ORCAMENTO = [
    "AINDA TEM ESTOQUE",
    "SÓ QUER COMPRAR GRANEL",
    "PROBLEMA FINANCEIRO / CRÉDITO",
    "NÃO SE APLICA",
]
MOTIVOS_CADASTRO = [
    "PRIMEIRO CONTATO / SEM RESPOSTA",
    "NÃO SE APLICA",
]
MOTIVOS_VENDA = [
    "NÃO SE APLICA",
]
MOTIVOS_NAO_RESPONDE = [
    "PRIMEIRO CONTATO / SEM RESPOSTA",
    "PROPRIETARIO / INDISPONÍVEL",
]
MOTIVOS_NAO_ATENDE = [
    "PROPRIETARIO / INDISPONÍVEL",
    "PRIMEIRO CONTATO / SEM RESPOSTA",
]
MOTIVOS_PERDA = [
    "SEM INTERESSE NO MOMENTO",
    "PRODUTO NÃO VENDEU / SEM GIRO",
    "FECHANDO / FECHOU LOJA",
    "LOJA ANEXO/PROXIMO - SM",
    "PROBLEMA FINANCEIRO / CRÉDITO",
    "PROBLEMA LOGÍSTICO / ENTREGA",
]
MOTIVOS_POS_VENDA = [
    "STATUS PEDIDO",
    "ATRASO ENTREGA",
    "NÃO SE APLICA",
]
MOTIVOS_SUPORTE = [
    "SEGUNDA VIA DE BOLETO",
    "SEGUNDA VIA DA NFE",
    "XML",
    "STATUS PEDIDO",
    "ATRASO ENTREGA",
    "PROBLEMA COM PEDIDOS",
    "SOLICITAÇÃO MATERIAL MKT",
    "DEVOLUÇÃO",
    "LINK PGTO 1ª TENTATIVA",
]

# ============================================================
# DEMANDAS INTERNAS — da aba REGRAS Seção 14 (25 tarefas reais)
# ============================================================
DEMANDAS_PROSPECÇÃO = [
    "PROSPECÇÃO - BASE GRANEL",
    "PROSPECÇÃO - PESQUISA GOOGLE",
    "CONTATO LEADS DO SITE",
    "RESPONDER LEADS COM DÚVIDAS",
    "LIGAÇÕES PROSPECT",
]
DEMANDAS_ORCAMENTO = [
    "MONTAR ORÇAMENTO SUGESTÃO (CURVA ABC)",
    "ENVIO DE AMOSTRAS",
    "ENVIAR ANÁLISE DE CRÉDITO",
    "REGISTRAR ATENDIMENTO NO MERCOS",
]
DEMANDAS_CADASTRO = [
    "ATUALIZAÇÃO CADASTRAL MERCOS",
    "ENVIAR ANÁLISE DE CRÉDITO",
    "REGISTRAR ATENDIMENTO NO MERCOS",
    "PREENCHER FOLLOW-UP",
]
DEMANDAS_PEDIDO = [
    "DIGITAÇÃO DE PEDIDOS",
    "SOLICITAR LINK CARTÃO CRÉDITO",
    "SOLICITAR VALORES PIX",
    "CONFIRMAR FATURAMENTO E EXPEDIÇÃO",
    "REGISTRAR ATENDIMENTO NO MERCOS",
]
DEMANDAS_POS_VENDA = [
    "FAZER RASTREIOS",
    "ACOMPANHAR DEVOLUÇÕES",
    "SUPORTE PCV - DÚVIDAS PEDIDOS",
    "SUPORTE PRODUTO (LAUDOS/TABELAS)",
    "PREENCHER FOLLOW-UP",
]
# Tarefas gerais (cross-funnel) para uso esporádico
DEMANDAS_GERAL = [
    "LIGAÇÕES DA BASE",
    "ATENDIMENTO ATIVOS/INATIVOS",
    "RESPOSTAS GRUPO ALINHAMENTO",
    "COBRANÇA DE TÍTULOS",
    "AJUSTE DE BOLETOS",
    "SOLICITAÇÃO DE NF (SEM SAP)",
]

# ============================================================
# TIPO PROBLEMA (para suporte/pós-venda)
# ============================================================
# TIPO PROBLEMA — da aba REGRAS Seção 12 (8 categorias reais)
TIPOS_PROBLEMA_POS_VENDA = [
    "ATRASO ENTREGA (TRANSPORTADORA)",
    "PRODUTO AVARIADO (FÁBRICA/TRANSPORTE)",
    "ERRO SEPARAÇÃO (EXPEDIÇÃO)",
    "ERRO NOTA FISCAL (FATURAMENTO)",
    "DIVERGÊNCIA PREÇO (FATURAMENTO)",
    "COBRANÇA INDEVIDA (FINANCEIRO)",
    "RUPTURA ESTOQUE (FÁBRICA/PCP)",
    "PROBLEMA SISTEMA (TI)",
    "",  # Nem sempre tem problema
    "",
    "",  # Maioria sem problema
]

# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================
def dias_uteis(start, end):
    """Retorna lista de dias úteis entre start e end."""
    result = []
    d = start
    while d <= end:
        if d.weekday() < 5:
            result.append(d)
        d += timedelta(days=1)
    return result

def proximo_dia_util(d, n=1):
    """Retorna o n-ésimo dia útil após d."""
    count = 0
    current = d
    while count < n:
        current += timedelta(days=1)
        if current.weekday() < 5:
            count += 1
    return current

def sinaleiro_calc(situacao, dias_compra, ciclo_medio):
    """Calcula sinaleiro baseado nas regras."""
    if not situacao:
        return "🟣"
    sit_upper = str(situacao).upper()
    if sit_upper in ("PROSPECT", "LEAD", "NOVO"):
        return "🟣"
    if dias_compra is None or dias_compra == "":
        return ""
    try:
        d = int(dias_compra)
        c = int(ciclo_medio) if ciclo_medio else 50
    except (ValueError, TypeError):
        return ""
    if d <= c:
        return "🟢"
    elif d <= c + 30:
        return "🟡"
    else:
        return "🔴"

def make_record(data, consultor, cnpj, nome, uf, rede, situacao, dias_compra,
                estagio_funil, tipo_cliente, fase, whatsapp, ligacao, lig_atendida,
                tipo_contato, resultado, motivo, followup_date, acao_futura,
                acao_detalhada, mercos, nota, temperatura, tipo_problema,
                demanda, tipo_atendimento, ciclo_medio=0):
    """Cria um registro DRAFT 2 completo (31 campos)."""
    # Campos calculados
    sinaleiro = sinaleiro_calc(situacao, dias_compra, ciclo_medio)
    grupo_dash = RESULTADO_REGRAS.get(resultado, (0, "", "", ""))[1]
    tipo_acao = RESULTADO_REGRAS.get(resultado, (0, "", "", ""))[3]

    return {
        'A': data,                     # DATA
        'B': consultor,                # CONSULTOR
        'C': nome,                     # NOME FANTASIA
        'D': cnpj,                     # CNPJ
        'E': uf,                       # UF
        'F': rede,                     # REDE / REGIONAL
        'G': situacao,                 # SITUAÇÃO
        'H': dias_compra,              # DIAS SEM COMPRA
        'I': estagio_funil,            # ESTÁGIO FUNIL
        'J': tipo_cliente,             # TIPO CLIENTE
        'K': fase,                     # FASE
        'L': sinaleiro,               # SINALEIRO
        'M': '',                       # TENTATIVA (calculado depois)
        'N': whatsapp,                 # WHATSAPP
        'O': ligacao,                  # LIGAÇÃO
        'P': lig_atendida,            # LIGAÇÃO ATENDIDA
        'Q': tipo_contato,            # TIPO DO CONTATO
        'R': resultado,               # RESULTADO
        'S': motivo,                   # MOTIVO
        'T': followup_date,           # FOLLOW-UP
        'U': acao_futura,             # AÇÃO FUTURA
        'V': acao_detalhada,          # AÇÃO DETALHADA
        'W': mercos,                   # MERCOS ATUALIZADO
        'X': nota,                     # NOTA DO DIA
        'Y': temperatura,             # TEMPERATURA
        'Z': grupo_dash,              # GRUPO DASH
        'AA': '',                      # SINALEIRO META
        'AB': tipo_acao,              # TIPO AÇÃO
        'AC': tipo_problema,          # TIPO PROBLEMA
        'AD': demanda,                # DEMANDA
        'AE': tipo_atendimento,       # TIPO ATENDIMENTO
    }

# ============================================================
# INÍCIO DO SCRIPT
# ============================================================
print("=" * 80)
print("CRM VITAO360 — V27: Simulação Realista JAN/25")
print("=" * 80)

if not os.path.exists(INPUT_FILE):
    print(f"ERRO: {INPUT_FILE} não encontrado")
    sys.exit(1)

print(f"\n[1/6] Carregando V26...")
wb = openpyxl.load_workbook(INPUT_FILE)
ws_cart = wb['CARTEIRA']
ws_d2 = wb['DRAFT 2']

# ============================================================
# FASE 1: LEVANTAR POOLS
# ============================================================
print(f"\n[2/6] Montando pools de clientes...")

# Pool A: Compradores JAN/25
pool_a = []
for r in range(4, ws_cart.max_row + 1):
    v = ws_cart.cell(row=r, column=26).value
    if v and isinstance(v, (int, float)) and v > 0:
        pool_a.append({
            'cnpj': str(ws_cart.cell(row=r, column=2).value).strip(),
            'nome': str(ws_cart.cell(row=r, column=3).value or ''),
            'uf': str(ws_cart.cell(row=r, column=4).value or 'PR'),
            'rede': str(ws_cart.cell(row=r, column=9).value or 'SEM GRUPO'),
            'valor': v,
        })

jan_cnpjs = {p['cnpj'] for p in pool_a}
print(f"  Pool A (vendas JAN): {len(pool_a)}")

# Pool B: Futuros compradores FEV/MAR/ABR
pool_b = []
for r in range(4, ws_cart.max_row + 1):
    cnpj = ws_cart.cell(row=r, column=2).value
    if cnpj is None:
        continue
    cnpj_str = str(cnpj).strip()
    if cnpj_str in jan_cnpjs:
        continue
    fev = ws_cart.cell(row=r, column=27).value
    mar = ws_cart.cell(row=r, column=28).value
    abr = ws_cart.cell(row=r, column=29).value
    valor = 0
    mes = ''
    if fev and isinstance(fev, (int, float)) and fev > 0:
        valor, mes = fev, 'FEV'
    elif mar and isinstance(mar, (int, float)) and mar > 0:
        valor, mes = mar, 'MAR'
    elif abr and isinstance(abr, (int, float)) and abr > 0:
        valor, mes = abr, 'ABR'
    if valor > 0:
        pool_b.append({
            'cnpj': cnpj_str,
            'nome': str(ws_cart.cell(row=r, column=3).value or ''),
            'uf': str(ws_cart.cell(row=r, column=4).value or 'PR'),
            'rede': str(ws_cart.cell(row=r, column=9).value or 'SEM GRUPO'),
            'valor_futuro': valor,
            'mes_compra': mes,
        })

pool_b_cnpjs = {p['cnpj'] for p in pool_b}
pool_b_selected = random.sample(pool_b, min(120, len(pool_b)))
print(f"  Pool B (futuros compradores): {len(pool_b_selected)}")

# Pool C: Perdas aleatórias
pool_c_candidates = []
for r in range(4, ws_cart.max_row + 1):
    cnpj = ws_cart.cell(row=r, column=2).value
    if cnpj is None:
        continue
    cnpj_str = str(cnpj).strip()
    if cnpj_str in jan_cnpjs or cnpj_str in pool_b_cnpjs:
        continue
    pool_c_candidates.append({
        'cnpj': cnpj_str,
        'nome': str(ws_cart.cell(row=r, column=3).value or ''),
        'uf': str(ws_cart.cell(row=r, column=4).value or 'PR'),
        'rede': str(ws_cart.cell(row=r, column=9).value or 'SEM GRUPO'),
    })

n_perdas = 410 - len(pool_a) - len(pool_b_selected)
pool_c_selected = random.sample(pool_c_candidates, min(n_perdas, len(pool_c_candidates)))
print(f"  Pool C (perdas): {len(pool_c_selected)}")

total_clientes = len(pool_a) + len(pool_b_selected) + len(pool_c_selected)
print(f"  TOTAL: {total_clientes} clientes")

# ============================================================
# FASE 2: DIAS ÚTEIS E DISTRIBUIÇÃO
# ============================================================
print(f"\n[3/6] Calculando dias úteis e distribuindo clientes...")
jan_dias = dias_uteis(date(2025, 1, 6), date(2025, 1, 31))
print(f"  Dias úteis: {len(jan_dias)} (06/01 a 31/01)")

# Distribuir clientes pelos dias (mais no início, menos no final)
# Semana 1: ~25/dia | Semana 2: ~22/dia | Semana 3-4: ~18/dia (follow-ups ocupam tempo)
distribuicao = []
todos_clientes = []

# Pool A - distribuir APENAS nos primeiros 8 dias úteis (06/01 a 15/01)
# Ciclo completo = 6 gaps de ~2 dias = 12 dias úteis
# Começando até dia 15/01, termina até 31/01
pool_a_dias = jan_dias[:8]  # Primeiros 8 dias úteis (06-15/jan)
for i, cli in enumerate(pool_a):
    dia_idx = i % len(pool_a_dias)
    cli['dia_inicio'] = pool_a_dias[dia_idx]
    cli['pool'] = 'A'
    todos_clientes.append(cli)

# Pool B - distribuir ao longo do mês
for i, cli in enumerate(pool_b_selected):
    dia_idx = int(i * len(jan_dias) / len(pool_b_selected))
    cli['dia_inicio'] = jan_dias[min(dia_idx, len(jan_dias)-1)]
    cli['pool'] = 'B'
    todos_clientes.append(cli)

# Pool C - distribuir ao longo do mês
for i, cli in enumerate(pool_c_selected):
    dia_idx = int(i * len(jan_dias) / len(pool_c_selected))
    cli['dia_inicio'] = jan_dias[min(dia_idx, len(jan_dias)-1)]
    cli['pool'] = 'C'
    todos_clientes.append(cli)

# Embaralhar levemente para naturalidade
random.shuffle(todos_clientes)

# ============================================================
# FASE 3: SIMULAÇÃO DIA A DIA
# ============================================================
print(f"\n[4/6] Simulando atendimentos dia a dia...")

registros_novos = []
acoes_futuras = []  # Ações que vão para FEV/MAR

for cli in todos_clientes:
    cnpj = cli['cnpj']
    nome = cli['nome']
    uf = cli['uf']
    rede = cli.get('rede', 'SEM GRUPO')
    pool = cli['pool']
    dia_inicio = cli['dia_inicio']

    # ================================
    # POOL A: CICLO COMPLETO DE VENDA
    # ================================
    if pool == 'A':
        valor = cli['valor']
        # Calcular gap máximo para que o ciclo caiba em JAN
        # Ciclo = 6 gaps (PROSP→APRES→ORC→CAD→NEG→VENDA) + margem pós-venda
        dias_disponiveis = len(dias_uteis(dia_inicio, date(2025, 1, 31)))
        # Precisamos de pelo menos 6 gaps + 1 dia pós-venda = 7 steps
        max_gap = max(1, (dias_disponiveis - 2) // 6)
        gap_dias = min(random.choice([1, 1, 2, 2, 2, 3]), max_gap)

        # Ação 1: PROSPECÇÃO (WhatsApp)
        d1 = dia_inicio
        registros_novos.append(make_record(
            data=d1, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
            situacao="PROSPECT", dias_compra="", estagio_funil="PROSPECÇÃO",
            tipo_cliente="NOVO", fase="CONTATO INICIAL",
            whatsapp=1, ligacao=0, lig_atendida=0,
            tipo_contato="PROSPECÇÃO", resultado="EM ATENDIMENTO",
            motivo=random.choice(MOTIVOS_EM_ATENDIMENTO),
            followup_date=proximo_dia_util(d1, gap_dias),
            acao_futura="ENVIAR ORÇAMENTO / CATÁLOGO",
            acao_detalhada="Enviar catálogo e tabela de preços",
            mercos="NÃO", nota=random.choice(NOTAS_PROSPECÇÃO),
            temperatura="🟡 MORNO", tipo_problema="",
            demanda=random.choice(DEMANDAS_PROSPECÇÃO),
            tipo_atendimento="PROSPECÇÃO"
        ))

        # Ação 2: APRESENTAÇÃO (Ligação)
        d2 = proximo_dia_util(d1, gap_dias)
        registros_novos.append(make_record(
            data=d2, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
            situacao="PROSPECT", dias_compra="", estagio_funil="QUALIFICAÇÃO",
            tipo_cliente="NOVO", fase="APRESENTAÇÃO",
            whatsapp=0, ligacao=1, lig_atendida=1,
            tipo_contato="PROSPECÇÃO", resultado="EM ATENDIMENTO",
            motivo="SOLICITOU ORÇAMENTO",
            followup_date=proximo_dia_util(d2, gap_dias),
            acao_futura="ENVIAR ORÇAMENTO / CATÁLOGO",
            acao_detalhada="Apresentação do portfólio realizada. Cliente interessado.",
            mercos="NÃO", nota="Ligação produtiva. Cliente quer receber orçamento.",
            temperatura="🟡 MORNO", tipo_problema="",
            demanda=random.choice(DEMANDAS_PROSPECÇÃO),
            tipo_atendimento="PROSPECÇÃO"
        ))

        # Ação 3: ORÇAMENTO — DEVE cair em JAN
        d3 = proximo_dia_util(d2, gap_dias)
        if d3 > date(2025, 1, 31):
            d3 = date(2025, 1, 31)
        registros_novos.append(make_record(
            data=d3, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
            situacao="PROSPECT", dias_compra="", estagio_funil="EM ATENDIMENTO",
            tipo_cliente="NOVO", fase="ORÇAMENTO",
            whatsapp=1, ligacao=0, lig_atendida=0,
            tipo_contato="NEGOCIAÇÃO", resultado="ORÇAMENTO",
            motivo=random.choice(MOTIVOS_ORCAMENTO),
            followup_date=proximo_dia_util(d3, 3),
            acao_futura="CONFIRMAR ORÇAMENTO ENVIADO",
            acao_detalhada=f"Orçamento enviado. Valor aprox. R$ {valor:,.2f}",
            mercos="SIM", nota=random.choice(NOTAS_ORCAMENTO),
            temperatura="🟡 MORNO", tipo_problema="",
            demanda=random.choice(DEMANDAS_ORCAMENTO),
            tipo_atendimento="NEGOCIAÇÃO"
        ))

        # Ação 4: CADASTRO — DEVE cair em JAN
        d4 = proximo_dia_util(d3, min(random.choice([2, 3, 3]), max_gap))
        if d4 > date(2025, 1, 31):
            d4 = date(2025, 1, 31)
        registros_novos.append(make_record(
            data=d4, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
            situacao="PROSPECT", dias_compra="", estagio_funil="CADASTRO",
            tipo_cliente="NOVO", fase="CADASTRO",
            whatsapp=1, ligacao=1, lig_atendida=1,
            tipo_contato="NEGOCIAÇÃO", resultado="CADASTRO",
            motivo=random.choice(MOTIVOS_CADASTRO),
            followup_date=proximo_dia_util(d4, 2),
            acao_futura="CONFIRMAR CADASTRO NO SISTEMA",
            acao_detalhada="Cadastro iniciado. Enviando docs para análise.",
            mercos="SIM", nota=random.choice(NOTAS_CADASTRO),
            temperatura="🟢 QUENTE", tipo_problema="",
            demanda=random.choice(DEMANDAS_CADASTRO),
            tipo_atendimento="NEGOCIAÇÃO"
        ))

        # Ação 5: NEGOCIAÇÃO / FECHAMENTO — DEVE cair em JAN
        d5 = proximo_dia_util(d4, min(random.choice([2, 3]), max_gap))
        if d5 > date(2025, 1, 31):
            d5 = date(2025, 1, 31)
        registros_novos.append(make_record(
            data=d5, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
            situacao="PROSPECT", dias_compra="", estagio_funil="FECHAMENTO",
            tipo_cliente="NOVO", fase="NEGOCIAÇÃO",
            whatsapp=0, ligacao=1, lig_atendida=1,
            tipo_contato="NEGOCIAÇÃO", resultado="EM ATENDIMENTO",
            motivo="CONDIÇÕES ACEITAS",
            followup_date=proximo_dia_util(d5, 1),
            acao_futura="CONFIRMAR FATURAMENTO E EXPEDIÇÃO",
            acao_detalhada="Negociação finalizada. Condições aprovadas pelo cliente.",
            mercos="SIM", nota="Cliente aprovou condições. Preparando pedido.",
            temperatura="🟢 QUENTE", tipo_problema="",
            demanda=random.choice(DEMANDAS_PEDIDO),
            tipo_atendimento="NEGOCIAÇÃO"
        ))

        # Ação 6: VENDA / PEDIDO — DEVE cair em JAN (forçar se necessário)
        d6 = proximo_dia_util(d5, min(random.choice([1, 1, 2]), max_gap))
        if d6 > date(2025, 1, 31):
            d6 = date(2025, 1, 31)
        registros_novos.append(make_record(
            data=d6, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
            situacao="ATIVO", dias_compra=0, estagio_funil="CS / RECOMPRA",
            tipo_cliente="NOVO", fase="VENDA",
            whatsapp=1, ligacao=0, lig_atendida=0,
            tipo_contato="NEGOCIAÇÃO", resultado="VENDA / PEDIDO",
            motivo=random.choice(MOTIVOS_VENDA),
            followup_date=proximo_dia_util(d6, 4),
            acao_futura="FAZER PÓS-VENDA",
            acao_detalhada=f"Pedido fechado! Valor: R$ {valor:,.2f}",
            mercos="SIM", nota=random.choice(NOTAS_VENDA),
            temperatura="🟢 QUENTE", tipo_problema="",
            demanda=random.choice(DEMANDAS_PEDIDO),
            tipo_atendimento="VENDA"
        ))

        # Ação 7: PÓS-VENDA D+4 (se cabe em janeiro)
        d7 = proximo_dia_util(d6, 4)
        if d7 <= date(2025, 1, 31):
            registros_novos.append(make_record(
                data=d7, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                situacao="ATIVO", dias_compra=4, estagio_funil="CS / RECOMPRA",
                tipo_cliente="NOVO", fase="PÓS-VENDA",
                whatsapp=1, ligacao=0, lig_atendida=0,
                tipo_contato="PÓS-VENDA/RELACIONAMENTO", resultado="SUPORTE",
                motivo=random.choice(MOTIVOS_POS_VENDA),
                followup_date=proximo_dia_util(d7, 11),
                acao_futura="FAZER CS (SUCESSO DO CLIENTE)",
                acao_detalhada="Acompanhamento D+4. Mercadoria entregue.",
                mercos="NÃO", nota=random.choice(NOTAS_POS_VENDA),
                temperatura="🟢 QUENTE",
                tipo_problema=random.choice(TIPOS_PROBLEMA_POS_VENDA),
                demanda=random.choice(DEMANDAS_POS_VENDA),
                tipo_atendimento="PÓS-VENDA"
            ))
            # Ação futura D+15 vai pra FEV
            d_d15 = proximo_dia_util(d6, 15)
            acoes_futuras.append({
                'data': d_d15, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
                'tipo': 'POS_VENDA_D15', 'consultor': CONSULTOR_JAN
            })
        else:
            # D+4 cai em FEV
            acoes_futuras.append({
                'data': d7, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
                'tipo': 'POS_VENDA_D4', 'consultor': CONSULTOR_JAN
            })

        # D+30 e CS sempre vão pra FEV/MAR
        d_d30 = proximo_dia_util(d6, 30)
        acoes_futuras.append({
            'data': d_d30, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
            'tipo': 'POS_VENDA_D30', 'consultor': CONSULTOR_JAN
        })
        d_cs = proximo_dia_util(d6, 45)
        acoes_futuras.append({
            'data': d_cs, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
            'tipo': 'CS_RECOMPRA', 'consultor': CONSULTOR_JAN
        })

    # ================================
    # POOL B: CICLO PARCIAL (não fecha em JAN)
    # ================================
    elif pool == 'B':
        gap = random.choice([2, 2, 3, 3, 4])

        # Ação 1: PROSPECÇÃO
        d1 = dia_inicio
        canal = random.choice(['whatsapp', 'ligacao'])
        registros_novos.append(make_record(
            data=d1, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
            situacao="PROSPECT", dias_compra="", estagio_funil="PROSPECÇÃO",
            tipo_cliente="NOVO", fase="CONTATO INICIAL",
            whatsapp=1 if canal == 'whatsapp' else 0,
            ligacao=1 if canal == 'ligacao' else 0,
            lig_atendida=1 if canal == 'ligacao' and random.random() > 0.3 else 0,
            tipo_contato="PROSPECÇÃO", resultado="EM ATENDIMENTO",
            motivo=random.choice(MOTIVOS_EM_ATENDIMENTO),
            followup_date=proximo_dia_util(d1, gap),
            acao_futura="FECHAR NEGOCIAÇÃO EM ANDAMENTO",
            acao_detalhada="Primeiro contato. Cliente demonstrou interesse inicial.",
            mercos="NÃO", nota=random.choice(NOTAS_PROSPECÇÃO),
            temperatura="🟡 MORNO", tipo_problema="",
            demanda=random.choice(DEMANDAS_PROSPECÇÃO),
            tipo_atendimento="PROSPECÇÃO"
        ))

        # Ação 2: APRESENTAÇÃO / FOLLOW-UP
        d2 = proximo_dia_util(d1, gap)
        if d2 <= date(2025, 1, 31):
            registros_novos.append(make_record(
                data=d2, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                situacao="PROSPECT", dias_compra="", estagio_funil="QUALIFICAÇÃO",
                tipo_cliente="NOVO", fase="APRESENTAÇÃO",
                whatsapp=random.choice([0, 1]), ligacao=1, lig_atendida=1,
                tipo_contato="PROSPECÇÃO", resultado="EM ATENDIMENTO",
                motivo="SOLICITOU ORÇAMENTO",
                followup_date=proximo_dia_util(d2, gap),
                acao_futura="ENVIAR ORÇAMENTO / CATÁLOGO",
                acao_detalhada="Apresentação realizada. Cliente quer orçamento.",
                mercos="NÃO",
                nota="Apresentação do portfólio. Cliente receptivo.",
                temperatura="🟡 MORNO", tipo_problema="",
                demanda=random.choice(DEMANDAS_PROSPECÇÃO),
                tipo_atendimento="PROSPECÇÃO"
            ))

            # Ação 3: ORÇAMENTO (se cabe em JAN)
            d3 = proximo_dia_util(d2, gap)
            if d3 <= date(2025, 1, 31):
                registros_novos.append(make_record(
                    data=d3, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                    situacao="PROSPECT", dias_compra="", estagio_funil="EM ATENDIMENTO",
                    tipo_cliente="NOVO", fase="ORÇAMENTO",
                    whatsapp=1, ligacao=0, lig_atendida=0,
                    tipo_contato="NEGOCIAÇÃO", resultado="ORÇAMENTO",
                    motivo=random.choice(MOTIVOS_ORCAMENTO),
                    followup_date=proximo_dia_util(d3, 5),
                    acao_futura="CONFIRMAR ORÇAMENTO ENVIADO",
                    acao_detalhada="Orçamento enviado. Aguardando retorno.",
                    mercos="SIM", nota=random.choice(NOTAS_ORCAMENTO),
                    temperatura="🟡 MORNO", tipo_problema="",
                    demanda=random.choice(DEMANDAS_ORCAMENTO),
                    tipo_atendimento="NEGOCIAÇÃO"
                ))
                # Follow-up vai pra FEV
                acoes_futuras.append({
                    'data': proximo_dia_util(d3, 5), 'cnpj': cnpj,
                    'nome': nome, 'uf': uf, 'rede': rede,
                    'tipo': 'FOLLOWUP_ORCAMENTO', 'consultor': CONSULTOR_JAN
                })
            else:
                # Ação 3 já cai em FEV
                acoes_futuras.append({
                    'data': d3, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
                    'tipo': 'CONTINUAR_APRESENTACAO', 'consultor': CONSULTOR_JAN
                })
        else:
            # Ação 2 já cai em FEV
            acoes_futuras.append({
                'data': d2, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
                'tipo': 'CONTINUAR_PROSPECÇÃO', 'consultor': CONSULTOR_JAN
            })

    # ================================
    # POOL C: 50% PERDAS + 50% ATIVOS PROGREDINDO
    # ================================
    elif pool == 'C':
        # Índice do cliente no pool pra dividir 50/50
        cli_idx = pool_c_selected.index(cli) if cli in pool_c_selected else 0
        is_perda = cli_idx < len(pool_c_selected) // 2  # Primeira metade = perdas

        if is_perda:
            # ---- METADE PERDAS: T1→T2→T3 com NÃO RESPONDE / NÃO ATENDE / PERDA ----
            rand = random.random() * 100
            if rand < 40:
                resultado_final = "NÃO RESPONDE"
                n_tentativas = random.choice([2, 2, 3, 3])
            elif rand < 70:
                resultado_final = "NÃO ATENDE"
                n_tentativas = random.choice([2, 3])
            elif rand < 90:
                resultado_final = "PERDA / FECHOU LOJA"
                n_tentativas = random.choice([1, 2])
            else:
                resultado_final = "NÃO RESPONDE"
                n_tentativas = random.choice([3, 4])

            for t in range(n_tentativas):
                gap = random.choice([2, 3, 3, 4, 5])
                d = proximo_dia_util(dia_inicio, t * gap) if t > 0 else dia_inicio
                if d > date(2025, 1, 31):
                    acoes_futuras.append({
                        'data': d, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
                        'tipo': 'TENTATIVA_CONTINUACAO', 'consultor': CONSULTOR_JAN
                    })
                    break

                # Selecionar motivo correto por resultado
                def _motivo_perda(res):
                    if "RESPONDE" in res: return random.choice(MOTIVOS_NAO_RESPONDE)
                    if "ATENDE" in res: return random.choice(MOTIVOS_NAO_ATENDE)
                    if "PERDA" in res: return random.choice(MOTIVOS_PERDA)
                    return random.choice(MOTIVOS_EM_ATENDIMENTO)

                if t == 0:
                    resultado_t = "NÃO RESPONDE" if "RESPONDE" in resultado_final else (
                        "NÃO ATENDE" if "ATENDE" in resultado_final else resultado_final)
                    res_usado = resultado_t if t < n_tentativas - 1 else resultado_final
                    registros_novos.append(make_record(
                        data=d, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                        situacao="PROSPECT", dias_compra="", estagio_funil="PROSPECÇÃO",
                        tipo_cliente="NOVO", fase="CONTATO INICIAL",
                        whatsapp=1, ligacao=0, lig_atendida=0,
                        tipo_contato="PROSPECÇÃO", resultado=res_usado,
                        motivo=_motivo_perda(res_usado),
                        followup_date=proximo_dia_util(d, gap) if t < n_tentativas - 1 else "",
                        acao_futura=ACAO_FUTURA_MAP.get(res_usado, "2ª TENTATIVA VIA LIGAÇÃO") if t < n_tentativas - 1 else (
                            ACAO_FUTURA_MAP.get(resultado_final, "NUTRIR / NOVO CONTATO") if "PERDA" in resultado_final or "NÃO" in resultado_final else ""),
                        acao_detalhada="",
                        mercos="NÃO", nota=random.choice(NOTAS_NAO_RESPONDE) if "RESPONDE" in resultado_t else (
                            random.choice(NOTAS_NAO_ATENDE) if "ATENDE" in resultado_t else
                            random.choice(NOTAS_PERDA)),
                        temperatura="🔴 FRIO", tipo_problema="",
                        demanda=random.choice(DEMANDAS_PROSPECÇÃO + [""]),
                        tipo_atendimento="PROSPECÇÃO"
                    ))
                elif t == 1:
                    atendeu = random.random() > 0.5
                    resultado_t = resultado_final if t == n_tentativas - 1 else (
                        "NÃO ATENDE" if not atendeu else "EM ATENDIMENTO")
                    registros_novos.append(make_record(
                        data=d, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                        situacao="PROSPECT", dias_compra="", estagio_funil="PROSPECÇÃO",
                        tipo_cliente="NOVO", fase="TENTATIVA",
                        whatsapp=0, ligacao=1, lig_atendida=1 if atendeu else 0,
                        tipo_contato="PROSPECÇÃO", resultado=resultado_t,
                        motivo=_motivo_perda(resultado_t),
                        followup_date=proximo_dia_util(d, gap) if t < n_tentativas - 1 else "",
                        acao_futura=ACAO_FUTURA_MAP.get(resultado_t, "COBRAR RETORNO DO CLIENTE") if t < n_tentativas - 1 else (
                            ACAO_FUTURA_MAP.get(resultado_t, "") if "NÃO" in resultado_t or "PERDA" in resultado_t else ""),
                        acao_detalhada="",
                        mercos="NÃO", nota=random.choice(NOTAS_NAO_ATENDE) if not atendeu else (
                            random.choice(NOTAS_PERDA) if "PERDA" in resultado_t else
                            random.choice(NOTAS_EM_ATENDIMENTO)),
                        temperatura="🔴 FRIO" if "NÃO" in resultado_t or "PERDA" in resultado_t else "🟡 MORNO",
                        tipo_problema="",
                        demanda=random.choice(DEMANDAS_PROSPECÇÃO + [""]),
                        tipo_atendimento="PROSPECÇÃO"
                    ))
                else:
                    canal = random.choice(['whatsapp', 'ligacao'])
                    registros_novos.append(make_record(
                        data=d, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                        situacao="PROSPECT", dias_compra="", estagio_funil="PROSPECÇÃO",
                        tipo_cliente="NOVO", fase="TENTATIVA",
                        whatsapp=1 if canal == 'whatsapp' else 0,
                        ligacao=1 if canal == 'ligacao' else 0, lig_atendida=0,
                        tipo_contato="PROSPECÇÃO", resultado=resultado_final,
                        motivo=_motivo_perda(resultado_final),
                        followup_date="",
                        acao_futura=ACAO_FUTURA_MAP.get(resultado_final, "NUTRIR ENVIANDO CAMPANHAS, PROMOÇÕES E NOVIDADES"),
                        acao_detalhada="",
                        mercos="NÃO", nota=random.choice(NOTAS_PERDA) if "PERDA" in resultado_final else (
                            random.choice(NOTAS_NAO_RESPONDE)),
                        temperatura="🔴 FRIO", tipo_problema="",
                        demanda=random.choice(DEMANDAS_PROSPECÇÃO + [""]),
                        tipo_atendimento="PROSPECÇÃO"
                    ))

        else:
            # ---- METADE ATIVOS: Progredindo no funil, empurrando pra FEV ----
            rand = random.random() * 100
            gap = random.choice([2, 3, 3, 4])

            if rand < 40:
                # EM ATENDIMENTO: Prospecção + retorno positivo → continua em FEV
                d1 = dia_inicio
                registros_novos.append(make_record(
                    data=d1, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                    situacao="PROSPECT", dias_compra="", estagio_funil="PROSPECÇÃO",
                    tipo_cliente="NOVO", fase="CONTATO INICIAL",
                    whatsapp=1, ligacao=0, lig_atendida=0,
                    tipo_contato="PROSPECÇÃO", resultado="EM ATENDIMENTO",
                    motivo=random.choice(MOTIVOS_EM_ATENDIMENTO),
                    followup_date=proximo_dia_util(d1, gap),
                    acao_futura="FECHAR NEGOCIAÇÃO EM ANDAMENTO",
                    acao_detalhada="Primeiro contato. Cliente pediu pra retornar.",
                    mercos="NÃO", nota=random.choice(NOTAS_PROSPECÇÃO),
                    temperatura="🟡 MORNO", tipo_problema="",
                    demanda=random.choice(DEMANDAS_PROSPECÇÃO),
                    tipo_atendimento="PROSPECÇÃO"
                ))
                # T2: Follow-up
                d2 = proximo_dia_util(d1, gap)
                if d2 <= date(2025, 1, 31):
                    registros_novos.append(make_record(
                        data=d2, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                        situacao="PROSPECT", dias_compra="", estagio_funil="QUALIFICAÇÃO",
                        tipo_cliente="NOVO", fase="APRESENTAÇÃO",
                        whatsapp=0, ligacao=1, lig_atendida=1,
                        tipo_contato="PROSPECÇÃO", resultado="EM ATENDIMENTO",
                        motivo="RETORNO POSITIVO, QUER MAIS INFO",
                        followup_date=proximo_dia_util(d2, gap),
                        acao_futura="ENVIAR ORÇAMENTO / CATÁLOGO",
                        acao_detalhada="Cliente receptivo. Quer conhecer a linha.",
                        mercos="NÃO", nota=random.choice(NOTAS_EM_ATENDIMENTO),
                        temperatura="🟡 MORNO", tipo_problema="",
                        demanda=random.choice(DEMANDAS_PROSPECÇÃO),
                        tipo_atendimento="PROSPECÇÃO"
                    ))
                    acoes_futuras.append({
                        'data': proximo_dia_util(d2, gap), 'cnpj': cnpj,
                        'nome': nome, 'uf': uf, 'rede': rede,
                        'tipo': 'CONTINUAR_PROSPECÇÃO', 'consultor': CONSULTOR_JAN
                    })
                else:
                    acoes_futuras.append({
                        'data': d2, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
                        'tipo': 'CONTINUAR_PROSPECÇÃO', 'consultor': CONSULTOR_JAN
                    })

            elif rand < 70:
                # PROSPECÇÃO AVANÇADA: Já na fase de apresentação, empurrando pra orçamento
                d1 = dia_inicio
                registros_novos.append(make_record(
                    data=d1, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                    situacao="PROSPECT", dias_compra="", estagio_funil="PROSPECÇÃO",
                    tipo_cliente="NOVO", fase="CONTATO INICIAL",
                    whatsapp=1, ligacao=0, lig_atendida=0,
                    tipo_contato="PROSPECÇÃO", resultado="EM ATENDIMENTO",
                    motivo=random.choice(MOTIVOS_EM_ATENDIMENTO),
                    followup_date=proximo_dia_util(d1, gap),
                    acao_futura="ENVIAR ORÇAMENTO / CATÁLOGO",
                    acao_detalhada="WhatsApp enviado. Bom retorno.",
                    mercos="NÃO", nota=random.choice(NOTAS_PROSPECÇÃO),
                    temperatura="🟡 MORNO", tipo_problema="",
                    demanda=random.choice(DEMANDAS_PROSPECÇÃO),
                    tipo_atendimento="PROSPECÇÃO"
                ))
                d2 = proximo_dia_util(d1, gap)
                if d2 <= date(2025, 1, 31):
                    registros_novos.append(make_record(
                        data=d2, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                        situacao="PROSPECT", dias_compra="", estagio_funil="QUALIFICAÇÃO",
                        tipo_cliente="NOVO", fase="APRESENTAÇÃO",
                        whatsapp=0, ligacao=1, lig_atendida=1,
                        tipo_contato="PROSPECÇÃO", resultado="EM ATENDIMENTO",
                        motivo="SOLICITOU ORÇAMENTO",
                        followup_date=proximo_dia_util(d2, gap),
                        acao_futura="ENVIAR ORÇAMENTO / CATÁLOGO",
                        acao_detalhada="Apresentação realizada. Cliente quer orçamento.",
                        mercos="NÃO",
                        nota="Apresentação completa. Cliente interessado nos produtos.",
                        temperatura="🟡 MORNO", tipo_problema="",
                        demanda=random.choice(DEMANDAS_PROSPECÇÃO),
                        tipo_atendimento="PROSPECÇÃO"
                    ))
                    d3 = proximo_dia_util(d2, gap)
                    if d3 <= date(2025, 1, 31):
                        registros_novos.append(make_record(
                            data=d3, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                            situacao="PROSPECT", dias_compra="", estagio_funil="EM ATENDIMENTO",
                            tipo_cliente="NOVO", fase="ORÇAMENTO",
                            whatsapp=1, ligacao=0, lig_atendida=0,
                            tipo_contato="NEGOCIAÇÃO", resultado="ORÇAMENTO",
                            motivo=random.choice(MOTIVOS_ORCAMENTO),
                            followup_date=proximo_dia_util(d3, 5),
                            acao_futura="CONFIRMAR ORÇAMENTO ENVIADO",
                            acao_detalhada="Orçamento enviado. Aguardando análise.",
                            mercos="SIM", nota=random.choice(NOTAS_ORCAMENTO),
                            temperatura="🟡 MORNO", tipo_problema="",
                            demanda=random.choice(DEMANDAS_ORCAMENTO),
                            tipo_atendimento="NEGOCIAÇÃO"
                        ))
                        acoes_futuras.append({
                            'data': proximo_dia_util(d3, 5), 'cnpj': cnpj,
                            'nome': nome, 'uf': uf, 'rede': rede,
                            'tipo': 'FOLLOWUP_ORCAMENTO', 'consultor': CONSULTOR_JAN
                        })
                    else:
                        acoes_futuras.append({
                            'data': d3, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
                            'tipo': 'ENVIAR_ORCAMENTO', 'consultor': CONSULTOR_JAN
                        })
                else:
                    acoes_futuras.append({
                        'data': d2, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
                        'tipo': 'CONTINUAR_APRESENTACAO', 'consultor': CONSULTOR_JAN
                    })

            else:
                # NEGOCIAÇÃO AVANÇADA: Já recebeu orçamento, está em negociação pra FEV
                d1 = dia_inicio
                registros_novos.append(make_record(
                    data=d1, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                    situacao="PROSPECT", dias_compra="", estagio_funil="PROSPECÇÃO",
                    tipo_cliente="NOVO", fase="CONTATO INICIAL",
                    whatsapp=1, ligacao=0, lig_atendida=0,
                    tipo_contato="PROSPECÇÃO", resultado="EM ATENDIMENTO",
                    motivo=random.choice(MOTIVOS_EM_ATENDIMENTO),
                    followup_date=proximo_dia_util(d1, 2),
                    acao_futura="ENVIAR ORÇAMENTO / CATÁLOGO",
                    acao_detalhada="Cliente já conhece o segmento. Ir direto.",
                    mercos="NÃO", nota=random.choice(NOTAS_PROSPECÇÃO),
                    temperatura="🟡 MORNO", tipo_problema="",
                    demanda=random.choice(DEMANDAS_PROSPECÇÃO),
                    tipo_atendimento="PROSPECÇÃO"
                ))
                d2 = proximo_dia_util(d1, 2)
                if d2 <= date(2025, 1, 31):
                    registros_novos.append(make_record(
                        data=d2, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                        situacao="PROSPECT", dias_compra="", estagio_funil="EM ATENDIMENTO",
                        tipo_cliente="NOVO", fase="ORÇAMENTO",
                        whatsapp=1, ligacao=1, lig_atendida=1,
                        tipo_contato="NEGOCIAÇÃO", resultado="ORÇAMENTO",
                        motivo=random.choice(MOTIVOS_ORCAMENTO),
                        followup_date=proximo_dia_util(d2, 3),
                        acao_futura="FECHAR NEGOCIAÇÃO EM ANDAMENTO",
                        acao_detalhada="Orçamento apresentado e enviado.",
                        mercos="SIM", nota=random.choice(NOTAS_ORCAMENTO),
                        temperatura="🟡 MORNO", tipo_problema="",
                        demanda=random.choice(DEMANDAS_ORCAMENTO),
                        tipo_atendimento="NEGOCIAÇÃO"
                    ))
                    d3 = proximo_dia_util(d2, 3)
                    if d3 <= date(2025, 1, 31):
                        registros_novos.append(make_record(
                            data=d3, consultor=CONSULTOR_JAN, cnpj=cnpj, nome=nome, uf=uf, rede=rede,
                            situacao="PROSPECT", dias_compra="", estagio_funil="EM ATENDIMENTO",
                            tipo_cliente="NOVO", fase="NEGOCIAÇÃO",
                            whatsapp=0, ligacao=1, lig_atendida=1,
                            tipo_contato="NEGOCIAÇÃO", resultado="EM ATENDIMENTO",
                            motivo="ANALISANDO PROPOSTA",
                            followup_date=proximo_dia_util(d3, 5),
                            acao_futura="COBRAR RETORNO DO CLIENTE",
                            acao_detalhada="Cliente analisando proposta. Retorno em breve.",
                            mercos="SIM", nota=random.choice(NOTAS_EM_ATENDIMENTO),
                            temperatura="🟡 MORNO", tipo_problema="",
                            demanda=random.choice(DEMANDAS_ORCAMENTO),
                            tipo_atendimento="NEGOCIAÇÃO"
                        ))
                        acoes_futuras.append({
                            'data': proximo_dia_util(d3, 5), 'cnpj': cnpj,
                            'nome': nome, 'uf': uf, 'rede': rede,
                            'tipo': 'FOLLOWUP_NEGOCIACAO', 'consultor': CONSULTOR_JAN
                        })
                    else:
                        acoes_futuras.append({
                            'data': d3, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
                            'tipo': 'FOLLOWUP_ORCAMENTO', 'consultor': CONSULTOR_JAN
                        })
                else:
                    acoes_futuras.append({
                        'data': d2, 'cnpj': cnpj, 'nome': nome, 'uf': uf, 'rede': rede,
                        'tipo': 'ENVIAR_ORCAMENTO', 'consultor': CONSULTOR_JAN
                    })

print(f"  Registros JAN/25 gerados: {len(registros_novos)}")
print(f"  Ações futuras (FEV+): {len(acoes_futuras)}")

# ============================================================
# FASE 4: RECONSTRUIR DRAFT 2
# ============================================================
print(f"\n[5/6] Reconstruindo DRAFT 2...")

# Ler registros existentes (não-JAN)
registros_existentes = []
for r in range(3, ws_d2.max_row + 1):
    data_val = ws_d2.cell(row=r, column=1).value
    if data_val is None:
        continue

    # Converter data
    if isinstance(data_val, datetime):
        data_date = data_val.date() if hasattr(data_val, 'date') else data_val
    elif isinstance(data_val, date):
        data_date = data_val
    elif isinstance(data_val, (int, float)):
        data_date = (datetime(1900, 1, 1) + timedelta(days=int(data_val) - 2)).date()
    else:
        continue

    # Pular registros de JAN/25 (vamos substituir)
    if data_date.year == 2025 and data_date.month == 1:
        continue

    row_data = {}
    for c in range(1, 32):
        row_data[c] = ws_d2.cell(row=r, column=c).value
    row_data['_date'] = data_date
    registros_existentes.append(row_data)

print(f"  Registros existentes (não-JAN): {len(registros_existentes)}")

# Converter novos registros para formato de escrita
registros_jan_write = []
col_keys = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P',
            'Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE']

for rec in registros_novos:
    row_write = {}
    for i, key in enumerate(col_keys):
        row_write[i + 1] = rec[key]
    # Converter date pra datetime pra openpyxl
    if isinstance(row_write[1], date) and not isinstance(row_write[1], datetime):
        row_write[1] = datetime(row_write[1].year, row_write[1].month, row_write[1].day)
    # FOLLOW-UP também
    if isinstance(row_write[20], date) and not isinstance(row_write[20], datetime):
        row_write[20] = datetime(row_write[20].year, row_write[20].month, row_write[20].day)
    row_write['_date'] = rec['A']
    registros_jan_write.append(row_write)

# Combinar e ordenar por data
todos_registros = registros_jan_write + registros_existentes
todos_registros.sort(key=lambda x: x['_date'])

print(f"  Total de registros (JAN novo + resto): {len(todos_registros)}")

# Limpar DRAFT 2 (manter headers linhas 1-2)
for r in range(3, ws_d2.max_row + 1):
    for c in range(1, 32):
        ws_d2.cell(row=r, column=c).value = None

# Escrever todos os registros
for i, rec in enumerate(todos_registros):
    r = i + 3  # Dados começam na linha 3
    for c in range(1, 32):
        ws_d2.cell(row=r, column=c).value = rec.get(c)

# Calcular TENTATIVA (col 13 = M): "T" + contagem de ocorrências do CNPJ até aqui
print("  Calculando TENTATIVA (col M)...")
cnpj_count = Counter()
for i, rec in enumerate(todos_registros):
    r = i + 3
    cnpj = rec.get(4)  # col D = CNPJ
    if cnpj:
        cnpj_str = str(cnpj).strip()
        cnpj_count[cnpj_str] += 1
        ws_d2.cell(row=r, column=13).value = f"T{cnpj_count[cnpj_str]}"

# Recalcular fórmulas de SINALEIRO (col L), GRUPO DASH (col Z), TIPO AÇÃO (col AB)
# (já escritos como valores nos novos, manter existentes como estão)

total_final = len(todos_registros)
print(f"  DRAFT 2 final: {total_final} registros")

# ============================================================
# FASE 5: SALVAR
# ============================================================
print(f"\n[6/6] Salvando V27...")
wb.save(OUTPUT_FILE)
file_size = os.path.getsize(OUTPUT_FILE) / (1024 * 1024)

# ============================================================
# RESUMO
# ============================================================
print(f"\n{'='*80}")
print(f"V27 CONCLUÍDO!")
print(f"{'='*80}")
print(f"  Arquivo: {OUTPUT_FILE}")
print(f"  Tamanho: {file_size:.2f} MB")
print(f"\n  DRAFT 2:")
print(f"    Total de registros: {total_final}")
print(f"    Novos JAN/25: {len(registros_novos)}")
print(f"    Mantidos (FEV-FEV/26): {len(registros_existentes)}")
print(f"\n  Ações futuras geradas (FEV+): {len(acoes_futuras)}")

# Distribuição mensal
mes_count = Counter()
for rec in todos_registros:
    d = rec['_date']
    if isinstance(d, (date, datetime)):
        mes_count[d.strftime('%Y-%m') if isinstance(d, date) else d.strftime('%Y-%m')] += 1
print(f"\n  Distribuição mensal:")
for m in sorted(mes_count.keys()):
    print(f"    {m}: {mes_count[m]:>6} registros")

# Distribuição de resultados JAN
res_count = Counter()
for rec in registros_novos:
    res_count[rec['R']] += 1
print(f"\n  Resultados JAN/25:")
for r_val, n in sorted(res_count.items(), key=lambda x: x[1], reverse=True):
    print(f"    {r_val}: {n}")

# Verificação de consistência lógica: ORÇAMENTO >= CADASTRO >= VENDA
n_orc = res_count.get('ORÇAMENTO', 0)
n_cad = res_count.get('CADASTRO', 0)
n_ven = res_count.get('VENDA / PEDIDO', 0)
print(f"\n  VERIFICAÇÃO LÓGICA (funil):")
print(f"    ORÇAMENTO ({n_orc}) >= CADASTRO ({n_cad}) >= VENDA ({n_ven})")
if n_orc >= n_cad >= n_ven:
    print(f"    ✅ CONSISTENTE — todo comprador passou por orçamento e cadastro")
else:
    print(f"    ❌ INCONSISTENTE — revisar lógica!")

# Verificar que TODAS as 32 vendas estão em JAN
jan_vendas = sum(1 for rec in registros_novos
                 if rec['R'] == 'VENDA / PEDIDO'
                 and isinstance(rec['A'], date) and rec['A'].month == 1)
print(f"    Vendas em JAN: {jan_vendas} de {len(pool_a)} esperadas")

# Estatísticas de preenchimento de campos JAN/25
n_total = len(registros_novos)
n_motivo = sum(1 for r in registros_novos if r.get('S') and str(r['S']).strip())
n_demanda = sum(1 for r in registros_novos if r.get('AD') and str(r['AD']).strip())
n_acao_fut = sum(1 for r in registros_novos if r.get('U') and str(r['U']).strip())
n_acao_det = sum(1 for r in registros_novos if r.get('V') and str(r['V']).strip())
n_nota = sum(1 for r in registros_novos if r.get('X') and str(r['X']).strip())
n_tipo_prob = sum(1 for r in registros_novos if r.get('AC') and str(r['AC']).strip())

print(f"\n  PREENCHIMENTO DE CAMPOS (JAN/25 — {n_total} registros):")
print(f"    MOTIVO:          {n_motivo:>5} ({n_motivo/n_total*100:.0f}%)")
print(f"    DEMANDA:         {n_demanda:>5} ({n_demanda/n_total*100:.0f}%)")
print(f"    AÇÃO FUTURA:     {n_acao_fut:>5} ({n_acao_fut/n_total*100:.0f}%)")
print(f"    AÇÃO DETALHADA:  {n_acao_det:>5} ({n_acao_det/n_total*100:.0f}%)")
print(f"    NOTA DO DIA:     {n_nota:>5} ({n_nota/n_total*100:.0f}%)")
print(f"    TIPO PROBLEMA:   {n_tipo_prob:>5} ({n_tipo_prob/n_total*100:.0f}%)")

# Distribuição de demandas internas
dem_count = Counter()
for r in registros_novos:
    d = r.get('AD')
    if d and str(d).strip():
        dem_count[str(d)] += 1
if dem_count:
    print(f"\n  DEMANDAS INTERNAS geradas:")
    for d_val, n in sorted(dem_count.items(), key=lambda x: x[1], reverse=True):
        print(f"    {d_val:<40} {n:>4}")

print(f"{'='*80}")

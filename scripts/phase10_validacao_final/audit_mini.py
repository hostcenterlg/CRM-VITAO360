#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AUDITORIA MINI — Só arquivos pequenos (skip V17/V31)"""
import openpyxl, re
from pathlib import Path
from datetime import datetime

def norm(v):
    if not v: return None
    c = re.sub(r'[^\d]', '', str(v).strip())
    return c if len(c) >= 11 else None

def scan(path, label):
    print(f"\n--- {label} ---")
    p = Path(path)
    if not p.exists():
        print(f"  ! NAO ENCONTRADO: {path}")
        return set()
    print(f"  Tamanho: {p.stat().st_size/1024:.0f} KB", flush=True)
    wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
    print(f"  Abas: {wb.sheetnames[:5]}", flush=True)
    ws = wb[wb.sheetnames[0]]
    mr = ws.max_row or 0
    mc = ws.max_column or 0
    print(f"  {mr} rows x {mc} cols", flush=True)
    # Show headers
    for r in range(1, min(3, mr+1)):
        h = []
        for c in range(1, min(mc+1, 15)):
            v = ws.cell(row=r, column=c).value
            if v: h.append(f"{c}:{str(v)[:20]}")
        if h: print(f"  Row{r}: {', '.join(h)}", flush=True)
    # Find CNPJ col
    cnpj_col = None
    start = 2
    for c in range(1, min(mc+1, 20)):
        for r in range(1, 4):
            v = ws.cell(row=r, column=c).value
            if v and "CNPJ" in str(v).upper():
                cnpj_col = c
                start = r + 1
                break
        if cnpj_col: break
    if not cnpj_col:
        for c in range(1, min(mc+1, 10)):
            cnt = 0
            for r in range(2, min(mr+1, 20)):
                v = ws.cell(row=r, column=c).value
                if v and 11 <= len(re.sub(r'[^\d]','',str(v))) <= 14:
                    cnt += 1
            if cnt >= 3:
                cnpj_col = c
                break
    # Count
    cnpjs = set()
    sit_col = None
    for c in range(1, min(mc+1, 30)):
        for r in range(1, 4):
            v = ws.cell(row=r, column=c).value
            if v and ("SITUA" in str(v).upper() or "STATUS" in str(v).upper()):
                sit_col = c
                break
        if sit_col: break
    sits = {}
    if cnpj_col:
        for r in range(start, mr+1):
            v = ws.cell(row=r, column=cnpj_col).value
            cn = norm(v)
            if cn:
                cnpjs.add(cn)
                if sit_col:
                    sv = ws.cell(row=r, column=sit_col).value
                    if sv:
                        s = str(sv).strip()[:25]
                        sits[s] = sits.get(s,0) + 1
    wb.close()
    print(f"  CNPJ col={cnpj_col}, start={start} → {len(cnpjs)} CNPJs", flush=True)
    if sits:
        top = sorted(sits.items(), key=lambda x:-x[1])[:8]
        print(f"  Situacoes: {', '.join(f'{k}({v})' for k,v in top)}", flush=True)
    return cnpjs

def main():
    t = datetime.now()
    print("="*80)
    print("AUDITORIA MINI")
    print("="*80)
    R = {}

    files = [
        (r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/Carteira detalhada de clientes fevreiro 2026.xlsx", "CART DET COM PROSPECTS"),
        (r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/Carteira detalhada de clientes fevereiro 2026 - sem os prospects.xlsx", "CART DET SEM PROSPECTS"),
        (r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/Carteira_detalhada_fev2026_sem_prospects.xlsx", "CART DET SEM V2"),
        (r"c:/Users/User/OneDrive/Área de Trabalho/CLAUDE CODE/08_CARTEIRA_MERCOS.xlsx", "MERCOS CARTEIRA"),
        (r"c:/Users/User/OneDrive/Área de Trabalho/GERENCIAL/GESTÃO  DE CLIENTES/BASE SAP OFICIAL/CARTEIRA  TOTAL DE CLIENTES SAP.xlsx", "SAP TOTAL"),
        (r"c:/Users/User/OneDrive/Área de Trabalho/GERENCIAL/GESTÃO  DE CLIENTES/BASE SAP OFICIAL/BASE SAP FINAL (INTERNO).xlsx", "SAP FINAL"),
        (r"c:/Users/User/OneDrive/Área de Trabalho/DRAFT3/DRAFT_3_SAP  COM INATIVOS.xlsx", "DRAFT3 COM INATIVOS"),
        (r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/CARTEIRA DE CLIENTES JANEIRO 2026.xlsx", "CART JAN 2026"),
        (r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/CARTEIRA  POR CONSULTOR .xlsx", "CART POR CONSULTOR"),
    ]

    for path, label in files:
        try:
            cnpjs = scan(path, label)
            if cnpjs:
                R[label] = cnpjs
        except Exception as e:
            print(f"  ERRO: {e}", flush=True)

    # Resumo
    print(f"\n\n{'='*80}")
    print("RESUMO")
    print(f"{'='*80}")
    print(f"\n  Dados conhecidos:")
    print(f"    V17 CARTEIRA: 554")
    print(f"    V31 CARTEIRA: ~5.460")
    print(f"\n  Fontes analisadas:")
    for name, cnpjs in sorted(R.items(), key=lambda x:-len(x[1])):
        print(f"    {name:30s}: {len(cnpjs):>6}")

    # Prospects
    com = R.get("CART DET COM PROSPECTS", set())
    sem = R.get("CART DET SEM PROSPECTS", set()) | R.get("CART DET SEM V2", set())
    if com and sem:
        prosp = com - sem
        print(f"\n  PROSPECTS (COM - SEM): {len(prosp)}")

    # Universo
    todos = set()
    for c in R.values(): todos.update(c)
    print(f"\n  UNIVERSO TOTAL (estas fontes): {len(todos)}")
    print(f"  Tempo: {(datetime.now()-t).total_seconds():.1f}s")

if __name__ == "__main__":
    main()

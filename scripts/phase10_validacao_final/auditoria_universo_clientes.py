#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AUDITORIA UNIVERSO DE CLIENTES — Onde estao os 6K+ prospects Mercos e 8K inativos SAP?

Fontes a analisar:
1. V17 CARTEIRA atual (554 clientes)
2. V31 CARTEIRA (5.460+ clientes)
3. V32 (mais recente?)
4. Carteira Mercos original (08_CARTEIRA_MERCOS)
5. Carteira detalhada fev2026 (com e sem prospects)
6. SAP Cadastro / CARTEIRA TOTAL SAP
7. DRAFT 3 SAP COM INATIVOS
"""

import openpyxl
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

def normalizar_cnpj(valor):
    if not valor:
        return None
    cnpj = re.sub(r'[^\d]', '', str(valor).strip())
    return cnpj if len(cnpj) >= 11 else None


def analyze_file(path, label, cnpj_cols=None, max_sheets=5):
    """Analisa um arquivo Excel e retorna CNPJs encontrados"""
    print(f"\n  [{label}] {Path(path).name}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f"    ERRO: {e}")
        return set(), {}

    total_cnpjs = set()
    sheet_info = {}

    for idx, name in enumerate(wb.sheetnames[:max_sheets]):
        ws = wb[name]
        rows = ws.max_row or 0
        cols = ws.max_column or 0

        if rows < 2:
            continue

        # Detectar coluna CNPJ
        cnpj_col = None
        if cnpj_cols and name in cnpj_cols:
            cnpj_col = cnpj_cols[name]
        else:
            # Auto-detectar por header
            for col in range(1, min(cols + 1, 50)):
                for row in [1, 2, 3]:
                    h = ws.cell(row=row, column=col).value
                    if h and "CNPJ" in str(h).upper():
                        cnpj_col = col
                        break
                if cnpj_col:
                    break

        if not cnpj_col:
            # Tentar encontrar coluna com padroes de CNPJ (14 digitos)
            for col in range(1, min(cols + 1, 20)):
                cnpj_count = 0
                for row in range(2, min(rows + 1, 20)):
                    val = ws.cell(row=row, column=col).value
                    if val:
                        cleaned = re.sub(r'[^\d]', '', str(val))
                        if 11 <= len(cleaned) <= 14:
                            cnpj_count += 1
                if cnpj_count >= 3:
                    cnpj_col = col
                    break

        cnpjs = set()
        if cnpj_col:
            start_row = 2
            # Detectar header rows
            for row in [1, 2, 3, 4]:
                val = ws.cell(row=row, column=cnpj_col).value
                if val and "CNPJ" in str(val).upper():
                    start_row = row + 1

            for row in range(start_row, rows + 1):
                val = ws.cell(row=row, column=cnpj_col).value
                cnpj = normalizar_cnpj(val)
                if cnpj:
                    cnpjs.add(cnpj)

        total_cnpjs.update(cnpjs)

        # Headers amostra
        headers = []
        header_row = start_row - 1 if cnpj_col else 1
        for col in range(1, min(cols + 1, 10)):
            h = ws.cell(row=header_row, column=col).value
            if h:
                headers.append(str(h)[:20])

        # Detectar se tem SITUACAO / STATUS
        situacao_col = None
        tipo_col = None
        for col in range(1, min(cols + 1, 50)):
            for row in [1, 2, 3]:
                h = ws.cell(row=row, column=col).value
                if h:
                    hu = str(h).upper()
                    if "SITUA" in hu or "STATUS" in hu:
                        situacao_col = col
                    if "TIPO" in hu and ("CLIENTE" in hu or "CLI" in hu):
                        tipo_col = col

        # Distribuicao de SITUACAO
        sit_dist = defaultdict(int)
        tipo_dist = defaultdict(int)
        if situacao_col:
            for row in range(start_row, rows + 1):
                val = ws.cell(row=row, column=situacao_col).value
                if val:
                    sit_dist[str(val).strip()[:30]] += 1
        if tipo_col:
            for row in range(start_row, rows + 1):
                val = ws.cell(row=row, column=tipo_col).value
                if val:
                    tipo_dist[str(val).strip()[:30]] += 1

        sheet_info[name] = {
            'rows': rows, 'cols': cols, 'cnpjs': len(cnpjs),
            'cnpj_col': cnpj_col, 'headers': headers,
            'situacao': dict(sit_dist), 'tipo': dict(tipo_dist)
        }

        print(f"    {name:25s} | {rows:>6} rows x {cols:>3} cols | {len(cnpjs):>5} CNPJs (col {cnpj_col or '?'})")
        if headers:
            print(f"    {'':25s} | Headers: {', '.join(headers[:6])}")
        if sit_dist:
            top3 = sorted(sit_dist.items(), key=lambda x: -x[1])[:5]
            print(f"    {'':25s} | Situacao: {', '.join(f'{k}({v})' for k,v in top3)}")
        if tipo_dist:
            top3 = sorted(tipo_dist.items(), key=lambda x: -x[1])[:5]
            print(f"    {'':25s} | Tipo: {', '.join(f'{k}({v})' for k,v in top3)}")

    wb.close()
    print(f"    TOTAL CNPJs unicos: {len(total_cnpjs)}")
    return total_cnpjs, sheet_info


def main():
    start = datetime.now()
    print("=" * 100)
    print("AUDITORIA UNIVERSO DE CLIENTES — Prospects Mercos + Inativos SAP")
    print(f"Inicio: {start}")
    print("=" * 100)

    all_sources = {}

    # ================================================================
    # 1. V17 CARTEIRA (referencia = o que temos agora)
    # ================================================================
    v17_path = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/output/phase10/CRM_VITAO360_V17_FINAL.xlsx"
    cnpjs, info = analyze_file(v17_path, "V17 ATUAL", cnpj_cols={"CARTEIRA": 2, "DRAFT 1": 2, "DRAFT 2": 4})
    all_sources["V17_CARTEIRA"] = cnpjs

    # Separar DRAFT 1 e CARTEIRA
    wb = openpyxl.load_workbook(v17_path, data_only=True, read_only=True)
    d1_cnpjs = set()
    ws = wb["DRAFT 1"]
    for row in range(4, ws.max_row + 1):
        cnpj = normalizar_cnpj(ws.cell(row=row, column=2).value)
        if cnpj:
            d1_cnpjs.add(cnpj)
    cart_cnpjs = set()
    ws = wb["CARTEIRA"]
    for row in range(4, ws.max_row + 1):
        cnpj = normalizar_cnpj(ws.cell(row=row, column=2).value)
        if cnpj:
            cart_cnpjs.add(cnpj)
    wb.close()
    all_sources["V17_DRAFT1"] = d1_cnpjs
    all_sources["V17_CART_ONLY"] = cart_cnpjs
    print(f"    V17 DRAFT 1: {len(d1_cnpjs)} CNPJs")
    print(f"    V17 CARTEIRA: {len(cart_cnpjs)} CNPJs")

    # ================================================================
    # 2. V31 (referencia com 5.460+ clientes)
    # ================================================================
    v31_path = r"c:/Users/User/OneDrive/Área de Trabalho/auditoria conversas sobre agenda atendimento draft 2/CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx"
    cnpjs_v31, _ = analyze_file(v31_path, "V31 REF", cnpj_cols={"CARTEIRA": 2, "DRAFT 1": 2, "DRAFT 2": 4})
    all_sources["V31"] = cnpjs_v31

    # V31 CARTEIRA especifico
    wb = openpyxl.load_workbook(v31_path, data_only=True, read_only=True)
    v31_cart = set()
    v31_d1 = set()
    ws = wb["CARTEIRA"]
    for row in range(4, ws.max_row + 1):
        cnpj = normalizar_cnpj(ws.cell(row=row, column=2).value)
        if cnpj:
            v31_cart.add(cnpj)
    ws = wb["DRAFT 1"]
    for row in range(4, ws.max_row + 1):
        cnpj = normalizar_cnpj(ws.cell(row=row, column=2).value)
        if cnpj:
            v31_d1.add(cnpj)
    wb.close()
    all_sources["V31_CARTEIRA"] = v31_cart
    all_sources["V31_DRAFT1"] = v31_d1
    print(f"    V31 CARTEIRA: {len(v31_cart)} CNPJs")
    print(f"    V31 DRAFT 1: {len(v31_d1)} CNPJs")

    # ================================================================
    # 3. V32 (mais recente?)
    # ================================================================
    v32_path = r"c:/Users/User/OneDrive/Área de Trabalho/CRM_V12_POPULADO_V32 18.02 vf ultima parte .xlsx"
    if Path(v32_path).exists():
        cnpjs_v32, _ = analyze_file(v32_path, "V32 (MAIS RECENTE?)", cnpj_cols={"CARTEIRA": 2, "DRAFT 1": 2})
        all_sources["V32"] = cnpjs_v32

    # ================================================================
    # 4. Carteira Mercos Original
    # ================================================================
    mercos_path = r"c:/Users/User/OneDrive/Área de Trabalho/CLAUDE CODE/08_CARTEIRA_MERCOS.xlsx"
    cnpjs_mercos, _ = analyze_file(mercos_path, "CARTEIRA MERCOS")
    all_sources["MERCOS_CARTEIRA"] = cnpjs_mercos

    # ================================================================
    # 5. Carteira detalhada fev2026 (COM prospects)
    # ================================================================
    cart_det_path = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/Carteira detalhada de clientes fevreiro 2026.xlsx"
    if Path(cart_det_path).exists():
        cnpjs_det, _ = analyze_file(cart_det_path, "CART DET FEV2026 (COM PROSPECTS)")
        all_sources["CART_DET_COM_PROSP"] = cnpjs_det

    # Sem prospects
    cart_sem_path = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/carteiras/Carteira detalhada de clientes fevereiro 2026 - sem os prospects.xlsx"
    if Path(cart_sem_path).exists():
        cnpjs_sem, _ = analyze_file(cart_sem_path, "CART DET FEV2026 (SEM PROSPECTS)")
        all_sources["CART_DET_SEM_PROSP"] = cnpjs_sem

    # ================================================================
    # 6. SAP Total
    # ================================================================
    sap_total_path = r"c:/Users/User/OneDrive/Área de Trabalho/GERENCIAL/GESTÃO  DE CLIENTES/BASE SAP OFICIAL/CARTEIRA  TOTAL DE CLIENTES SAP.xlsx"
    if Path(sap_total_path).exists():
        cnpjs_sap, _ = analyze_file(sap_total_path, "SAP TOTAL CLIENTES")
        all_sources["SAP_TOTAL"] = cnpjs_sap

    sap_final_path = r"c:/Users/User/OneDrive/Área de Trabalho/GERENCIAL/GESTÃO  DE CLIENTES/BASE SAP OFICIAL/BASE SAP FINAL (INTERNO).xlsx"
    if Path(sap_final_path).exists():
        cnpjs_sap_f, _ = analyze_file(sap_final_path, "SAP FINAL INTERNO")
        all_sources["SAP_FINAL"] = cnpjs_sap_f

    # ================================================================
    # 7. DRAFT 3 SAP COM INATIVOS
    # ================================================================
    draft3_inativo_path = r"c:/Users/User/OneDrive/Área de Trabalho/DRAFT3/DRAFT_3_SAP  COM INATIVOS.xlsx"
    if Path(draft3_inativo_path).exists():
        cnpjs_d3i, _ = analyze_file(draft3_inativo_path, "DRAFT 3 SAP COM INATIVOS")
        all_sources["DRAFT3_SAP_INATIVOS"] = cnpjs_d3i

    # ================================================================
    # 8. SAP no repositorio
    # ================================================================
    sap_repo_path = r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/sap"
    if Path(sap_repo_path).exists():
        sap_files = list(Path(sap_repo_path).glob("*.xlsx"))
        if sap_files:
            for sf in sap_files[:3]:
                cnpjs_sr, _ = analyze_file(str(sf), f"SAP REPO ({sf.name[:30]})")
                all_sources[f"SAP_REPO_{sf.stem[:20]}"] = cnpjs_sr

    # ================================================================
    # ANALISE CRUZADA
    # ================================================================
    print(f"\n\n{'='*100}")
    print("ANALISE CRUZADA — Quem esta faltando?")
    print(f"{'='*100}")

    # Resumo
    print(f"\n  UNIVERSO POR FONTE:")
    for name, cnpjs in sorted(all_sources.items(), key=lambda x: -len(x[1])):
        print(f"    {name:30s}: {len(cnpjs):>6} CNPJs")

    # V17 vs V31
    v17_c = all_sources.get("V17_CART_ONLY", set())
    v31_c = all_sources.get("V31_CARTEIRA", set())

    common_v17_v31 = v17_c & v31_c
    only_v17 = v17_c - v31_c
    only_v31 = v31_c - v17_c

    print(f"\n  V17 vs V31 CARTEIRA:")
    print(f"    Em comum: {len(common_v17_v31)}")
    print(f"    Só no V17: {len(only_v17)}")
    print(f"    Só no V31 (FALTANDO no V17): {len(only_v31)}")

    # Mercos prospects faltantes
    mercos = all_sources.get("MERCOS_CARTEIRA", set())
    cart_com = all_sources.get("CART_DET_COM_PROSP", set())
    cart_sem = all_sources.get("CART_DET_SEM_PROSP", set())

    if cart_com and cart_sem:
        prospects = cart_com - cart_sem
        print(f"\n  PROSPECTS MERCOS (detectados por diferenca com/sem):")
        print(f"    Cart COM prospects: {len(cart_com)}")
        print(f"    Cart SEM prospects: {len(cart_sem)}")
        print(f"    PROSPECTS = diferenca: {len(prospects)}")
        prosp_faltando = prospects - v17_c
        print(f"    Prospects FALTANDO no V17: {len(prosp_faltando)}")

    if mercos:
        mercos_faltando = mercos - v17_c
        print(f"\n  MERCOS FALTANDO no V17:")
        print(f"    Total Mercos: {len(mercos)}")
        print(f"    Faltando no V17: {len(mercos_faltando)}")

    # SAP faltantes
    sap = all_sources.get("SAP_TOTAL", set()) | all_sources.get("SAP_FINAL", set()) | all_sources.get("DRAFT3_SAP_INATIVOS", set())
    if sap:
        sap_faltando = sap - v17_c
        print(f"\n  SAP FALTANDO no V17:")
        print(f"    Total SAP (todas fontes): {len(sap)}")
        print(f"    Faltando no V17: {len(sap_faltando)}")

    # UNIVERSO TOTAL
    todos = set()
    for cnpjs in all_sources.values():
        todos.update(cnpjs)

    faltando_total = todos - v17_c
    print(f"\n  UNIVERSO TOTAL:")
    print(f"    Todos os CNPJs unicos: {len(todos)}")
    print(f"    No V17 CARTEIRA: {len(v17_c)}")
    print(f"    FALTANDO no V17: {len(faltando_total)}")

    # Classificar os faltantes por fonte
    print(f"\n  FALTANTES POR FONTE (onde eles existem):")
    from_v31 = faltando_total & v31_c
    from_mercos = faltando_total & mercos if mercos else set()
    from_sap = faltando_total & sap if sap else set()
    from_cart_det = faltando_total & cart_com if cart_com else set()

    print(f"    Existem no V31 CARTEIRA: {len(from_v31)}")
    print(f"    Existem no Mercos: {len(from_mercos)}")
    print(f"    Existem no SAP: {len(from_sap)}")
    print(f"    Existem na Cart Detalhada: {len(from_cart_det)}")

    elapsed = (datetime.now() - start).total_seconds()
    print(f"\n  Auditoria completada em {elapsed:.1f}s")

    # RECOMENDACAO
    print(f"\n{'='*100}")
    print("RECOMENDACAO")
    print(f"{'='*100}")

    if len(v31_c) > len(v17_c) * 2:
        print(f"\n  O V31 tem {len(v31_c)} clientes na CARTEIRA vs {len(v17_c)} no V17.")
        print(f"  RECOMENDO: Expandir DRAFT 1 e CARTEIRA usando V31 como base (ja tem os prospects).")
    elif cart_com and len(cart_com) > len(v17_c) * 2:
        print(f"\n  A Carteira Detalhada COM Prospects tem {len(cart_com)} clientes.")
        print(f"  RECOMENDO: Importar esses {len(cart_com) - len(v17_c)} clientes para DRAFT 1.")
    else:
        print(f"\n  Necessario extrair base completa Mercos (com prospects) e SAP (com inativos).")
        print(f"  Peca ao usuario para exportar e colocar na pasta data/sources/carteiras/.")


if __name__ == "__main__":
    main()

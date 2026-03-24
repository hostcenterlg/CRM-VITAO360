#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AUDIT FOCADO — Aba Prospects do Mercos + DRAFT3 COM INATIVOS + SAP (col 5 only)"""
import openpyxl, re
from pathlib import Path
from datetime import datetime

def norm(v):
    if not v: return None
    c = re.sub(r'[^\d]', '', str(v).strip())
    return c if len(c) >= 11 else None

def main():
    t = datetime.now()
    print("="*80, flush=True)
    print("AUDIT FOCADO — Prospects + SAP Inativos", flush=True)
    print("="*80, flush=True)

    # ================================================================
    # 1. MERCOS — Todas as abas
    # ================================================================
    print("\n>>> MERCOS CARTEIRA (todas abas) <<<", flush=True)
    mercos_path = r"c:/Users/User/OneDrive/Área de Trabalho/CLAUDE CODE/08_CARTEIRA_MERCOS.xlsx"
    wb = openpyxl.load_workbook(mercos_path, data_only=True, read_only=True)
    print(f"  Abas: {wb.sheetnames}", flush=True)

    mercos_all = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        mr = ws.max_row or 0
        mc = ws.max_column or 0
        print(f"\n  --- {sname} --- ({mr} rows x {mc} cols)", flush=True)

        # Headers
        headers = []
        for c in range(1, min(mc+1, 15)):
            v = ws.cell(row=1, column=c).value
            if v: headers.append(f"{c}:{str(v)[:20]}")
        print(f"    Headers: {', '.join(headers)}", flush=True)

        # Find CNPJ col
        cnpj_col = None
        for c in range(1, min(mc+1, 15)):
            h = ws.cell(row=1, column=c).value
            if h and "CNPJ" in str(h).upper():
                cnpj_col = c
                break
        if not cnpj_col:
            for c in range(1, min(mc+1, 10)):
                cnt = 0
                for r in range(2, min(mr+1, 15)):
                    v = ws.cell(row=r, column=c).value
                    if v and 11 <= len(re.sub(r'[^\d]','',str(v))) <= 14:
                        cnt += 1
                if cnt >= 3:
                    cnpj_col = c
                    break

        cnpjs = set()
        if cnpj_col:
            for r in range(2, mr+1):
                cn = norm(ws.cell(row=r, column=cnpj_col).value)
                if cn: cnpjs.add(cn)

        mercos_all[sname] = cnpjs
        print(f"    CNPJ col={cnpj_col} → {len(cnpjs)} CNPJs", flush=True)

        # Amostra
        if cnpjs and mr > 1:
            for r in range(2, min(5, mr+1)):
                vals = []
                for c in range(1, min(5, mc+1)):
                    v = ws.cell(row=r, column=c).value
                    if v: vals.append(str(v)[:25])
                print(f"    Row{r}: {', '.join(vals)}", flush=True)

    wb.close()

    # Total Mercos
    total_mercos = set()
    for cnpjs in mercos_all.values():
        total_mercos.update(cnpjs)
    print(f"\n  MERCOS TOTAL: {len(total_mercos)} CNPJs unicos", flush=True)
    prosp = mercos_all.get("Prospects", set())
    clientes = mercos_all.get("Carteira Clientes Mercos", set())
    print(f"    Clientes: {len(clientes)}", flush=True)
    print(f"    Prospects: {len(prosp)}", flush=True)
    print(f"    Overlap: {len(clientes & prosp)}", flush=True)
    print(f"    Prospects exclusivos: {len(prosp - clientes)}", flush=True)

    # ================================================================
    # 2. DRAFT 3 SAP COM INATIVOS
    # ================================================================
    print(f"\n\n>>> DRAFT 3 SAP COM INATIVOS <<<", flush=True)
    d3_path = r"c:/Users/User/OneDrive/Área de Trabalho/DRAFT3/DRAFT_3_SAP  COM INATIVOS.xlsx"
    if Path(d3_path).exists():
        wb = openpyxl.load_workbook(d3_path, data_only=True, read_only=True)
        print(f"  Abas: {wb.sheetnames}", flush=True)
        for sname in wb.sheetnames[:3]:
            ws = wb[sname]
            mr = ws.max_row or 0
            mc = ws.max_column or 0
            print(f"\n  --- {sname} --- ({mr} rows x {mc} cols)", flush=True)
            headers = []
            for c in range(1, min(mc+1, 15)):
                v = ws.cell(row=1, column=c).value
                if v: headers.append(f"{c}:{str(v)[:20]}")
            print(f"    Headers: {', '.join(headers)}", flush=True)

            cnpj_col = None
            for c in range(1, min(mc+1, 15)):
                h = ws.cell(row=1, column=c).value
                if h and "CNPJ" in str(h).upper():
                    cnpj_col = c
                    break
            cnpjs = set()
            if cnpj_col:
                for r in range(2, mr+1):
                    cn = norm(ws.cell(row=r, column=cnpj_col).value)
                    if cn: cnpjs.add(cn)
            print(f"    CNPJ col={cnpj_col} → {len(cnpjs)} CNPJs", flush=True)
        wb.close()
    else:
        print(f"  ! NAO ENCONTRADO", flush=True)

    # ================================================================
    # 3. SAP TOTAL — So contar col 5 (CNPJ) + col 6 (CPF)
    # ================================================================
    print(f"\n\n>>> SAP TOTAL (57K rows, so contando CNPJs) <<<", flush=True)
    sap_path = r"c:/Users/User/OneDrive/Área de Trabalho/GERENCIAL/GESTÃO  DE CLIENTES/BASE SAP OFICIAL/CARTEIRA  TOTAL DE CLIENTES SAP.xlsx"
    sap_cnpjs = set()
    if Path(sap_path).exists():
        wb = openpyxl.load_workbook(sap_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        print(f"  {ws.max_row} rows, reading col 5 (CNPJ)...", flush=True)
        for r in range(2, (ws.max_row or 2) + 1):
            cn = norm(ws.cell(row=r, column=5).value)
            if cn: sap_cnpjs.add(cn)
            if r % 10000 == 0:
                print(f"    ... row {r} ({len(sap_cnpjs)} CNPJs)", flush=True)
        wb.close()
        print(f"  SAP TOTAL: {len(sap_cnpjs)} CNPJs unicos", flush=True)
    else:
        print(f"  ! NAO ENCONTRADO", flush=True)

    # ================================================================
    # 4. SAP sources no repo
    # ================================================================
    print(f"\n\n>>> SAP sources no repo <<<", flush=True)
    sap_repo = Path(r"c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360/data/sources/sap")
    sap_repo_cnpjs = set()
    if sap_repo.exists():
        for f in sap_repo.glob("*.xlsx"):
            print(f"  {f.name} ({f.stat().st_size/1024:.0f} KB)...", flush=True)
            try:
                wb = openpyxl.load_workbook(str(f), data_only=True, read_only=True)
                ws = wb[wb.sheetnames[0]]
                # Find CNPJ col
                cnpj_col = None
                for c in range(1, min((ws.max_column or 15)+1, 15)):
                    h = ws.cell(row=1, column=c).value
                    if h and "CNPJ" in str(h).upper():
                        cnpj_col = c
                        break
                if cnpj_col:
                    for r in range(2, (ws.max_row or 2)+1):
                        cn = norm(ws.cell(row=r, column=cnpj_col).value)
                        if cn: sap_repo_cnpjs.add(cn)
                wb.close()
                print(f"    CNPJ col={cnpj_col} → {len(sap_repo_cnpjs)} total", flush=True)
            except Exception as e:
                print(f"    ERRO: {e}", flush=True)

    # ================================================================
    # RESUMO FINAL
    # ================================================================
    print(f"\n\n{'='*80}", flush=True)
    print("RESUMO FINAL", flush=True)
    print(f"{'='*80}", flush=True)

    print(f"\n  V17 CARTEIRA ATUAL: 554 CNPJs", flush=True)
    print(f"  V31 CARTEIRA (ref): ~5.460 CNPJs", flush=True)
    print(f"  → FALTANDO: ~4.906 CNPJs", flush=True)

    print(f"\n  FONTES DISPONIVEIS:", flush=True)
    print(f"    MERCOS Clientes:    {len(clientes):>6}", flush=True)
    print(f"    MERCOS Prospects:   {len(prosp):>6}", flush=True)
    print(f"    MERCOS Total:       {len(total_mercos):>6}", flush=True)
    print(f"    SAP TOTAL:          {len(sap_cnpjs):>6}", flush=True)
    print(f"    SAP repo:           {len(sap_repo_cnpjs):>6}", flush=True)

    # Onde os 4.906 faltantes estao?
    # V17 = 554, V31 = 5460, diff = 4906
    # Se MERCOS tem X prospects e SAP tem Y inativos...
    universo = total_mercos | sap_cnpjs | sap_repo_cnpjs
    print(f"\n  UNIVERSO TOTAL (Mercos + SAP): {len(universo):>6}", flush=True)

    # O V31 provavelmente usou o Mercos completo (com prospects)
    # Podemos reconstruir o DRAFT 1 expandido usando:
    # 1. Mercos Clientes (497) + Mercos Prospects (?)
    # 2. SAP Cadastro com inativos
    # 3. V31 CARTEIRA como referencia

    print(f"\n  CONCLUSAO:", flush=True)
    if len(prosp) > 100:
        print(f"    ✓ Aba 'Prospects' do Mercos tem {len(prosp)} prospects", flush=True)
        print(f"    → Estes precisam ir para DRAFT 1 e CARTEIRA", flush=True)
    elif len(prosp) == 0:
        print(f"    ✗ Aba 'Prospects' do Mercos esta VAZIA ou sem CNPJs", flush=True)
        print(f"    → Precisa extrair base nova do Mercos", flush=True)
    else:
        print(f"    ⚠ Aba 'Prospects' tem {len(prosp)} prospects (menos que esperado)", flush=True)

    if len(sap_cnpjs) > 1000:
        print(f"    ✓ SAP TOTAL tem {len(sap_cnpjs)} clientes (incluindo inativos)", flush=True)
        n_sap_excl = len(sap_cnpjs - total_mercos)
        print(f"    → {n_sap_excl} exclusivos SAP (nao estao no Mercos)", flush=True)
    elif len(sap_cnpjs) == 0:
        print(f"    ✗ SAP TOTAL nao processado ou sem CNPJs", flush=True)
        print(f"    → Usar data/sources/sap/ como alternativa", flush=True)

    elapsed = (datetime.now() - t).total_seconds()
    print(f"\n  Tempo: {elapsed:.1f}s", flush=True)

if __name__ == "__main__":
    main()

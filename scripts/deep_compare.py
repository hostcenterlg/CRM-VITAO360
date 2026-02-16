"""Deep comparison: Ours vs Downloaded V3 — every difference catalogued."""
import openpyxl
from openpyxl.utils import get_column_letter

OURS = r"C:\Users\User\OneDrive\Área de Trabalho\CLAUDE CODE\output\CRM_VITAO360_V3.xlsx"
THEIRS = r"C:\Users\User\Downloads\CRM_INTELIGENTE_VITAO360_V3 (1).xlsx"

wb_o = openpyxl.load_workbook(OURS, data_only=True)
wb_t = openpyxl.load_workbook(THEIRS, data_only=True)

# ══════════════════════════════════════════════════════════════
# 1. REGRAS — full comparison
# ══════════════════════════════════════════════════════════════
print("=" * 70)
print("  1. REGRAS — FULL COMPARISON")
print("=" * 70)
ws_o = wb_o["REGRAS"]
ws_t = wb_t["REGRAS"]
print(f"  Size: Ours {ws_o.max_row}r x {ws_o.max_column}c | Theirs {ws_t.max_row}r x {ws_t.max_column}c")

# Their sections
print("\n  Their REGRAS sections (col A headers):")
for r in range(1, ws_t.max_row + 1):
    a = ws_t.cell(row=r, column=1).value
    b = ws_t.cell(row=r, column=2).value
    if a and str(a).startswith('#') or (b and any(kw in str(b).upper() for kw in ['RESULTADO', 'ESTÁGIO', 'TIPO', 'MOTIVO', 'SITUAÇÃO', 'FASE', 'TENTATIVA', 'SINALEIRO', 'CONSULTOR', 'LISTA'])):
        if r < 5 or (a and not str(a).isdigit()):
            print(f"    R{r}: A={a} | B={b}")

# Check for LEAD
print("\n  Searching for LEAD in their REGRAS:")
for r in range(1, ws_t.max_row + 1):
    for c in range(1, ws_t.max_column + 1):
        v = ws_t.cell(row=r, column=c).value
        if v and str(v).strip().upper() == "LEAD":
            print(f"    FOUND LEAD at R{r}C{c}: {v}")

# Their SITUAÇÃO list
print("\n  Their SITUAÇÃO values:")
for r in range(1, ws_t.max_row + 1):
    a = ws_t.cell(row=r, column=1).value
    b = ws_t.cell(row=r, column=2).value
    if b and str(b).upper() in ['ATIVO', 'EM RISCO', 'INAT.REC', 'INAT.ANT', 'NOVO', 'PROSPECT', 'LEAD']:
        print(f"    R{r}: A={a} B={b} C={ws_t.cell(row=r, column=3).value}")

# ══════════════════════════════════════════════════════════════
# 2. DRAFT 1 — columns comparison
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 70}")
print("  2. DRAFT 1 — COLUMN COMPARISON")
print(f"{'=' * 70}")
ws_o = wb_o["DRAFT 1"]
ws_t = wb_t["DRAFT 1"]

# Find header rows
o_hdr_row = 3  # ours
t_hdr_row = 3  # theirs

print(f"  Size: Ours {ws_o.max_row}r x {ws_o.max_column}c | Theirs {ws_t.max_row}r x {ws_t.max_column}c")
print(f"\n  Col | OURS                         | THEIRS")
print(f"  " + "-" * 65)
max_c = max(ws_o.max_column, ws_t.max_column)
for c in range(1, max_c + 1):
    o_val = str(ws_o.cell(row=o_hdr_row, column=c).value or "—")[:28]
    t_val = str(ws_t.cell(row=t_hdr_row, column=c).value or "—")[:28]
    match = "✓" if o_val.strip() == t_val.strip() else "✗"
    if o_val != "—" or t_val != "—":
        print(f"  {get_column_letter(c):3} | {o_val:28} | {t_val:28} {match}")

# Their block headers (row 2)
print(f"\n  Their DRAFT 1 block headers (row 2):")
for c in range(1, ws_t.max_column + 1):
    v = ws_t.cell(row=2, column=c).value
    if v:
        print(f"    {get_column_letter(c)} (col {c}): {v}")

# ══════════════════════════════════════════════════════════════
# 3. PROJEÇÃO — columns + data
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 70}")
print("  3. PROJEÇÃO — COLUMN + DATA COMPARISON")
print(f"{'=' * 70}")
ws_o = wb_o["PROJEÇÃO"]
ws_t = wb_t["PROJEÇÃO"]
print(f"  Size: Ours {ws_o.max_row}r x {ws_o.max_column}c | Theirs {ws_t.max_row}r x {ws_t.max_column}c")

# Extra cols in theirs (>39)
print(f"\n  Their PROJEÇÃO extra cols beyond ours (col 40+):")
for c in range(40, ws_t.max_column + 1):
    v = ws_t.cell(row=3, column=c).value
    if v:
        print(f"    {get_column_letter(c)} (col {c}): {v}")

# Their block headers (row 2)
print(f"\n  Their PROJEÇÃO block headers (row 2):")
for c in range(1, ws_t.max_column + 1):
    v = ws_t.cell(row=2, column=c).value
    if v:
        print(f"    {get_column_letter(c)} (col {c}): {v}")

# Their data rows (how many redes + extra data)
print(f"\n  Their PROJEÇÃO data (col B, first 20 rows):")
for r in range(4, 24):
    b = ws_t.cell(row=r, column=2).value
    c3 = ws_t.cell(row=r, column=3).value
    if b:
        print(f"    R{r}: {b} | rede={c3}")

# ══════════════════════════════════════════════════════════════
# 4. DRAFT 2 — full column + formula comparison
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 70}")
print("  4. DRAFT 2 — FULL COLUMN COMPARISON")
print(f"{'=' * 70}")
ws_o = wb_o["DRAFT 2"]
ws_t = wb_t["DRAFT 2"]
print(f"  Size: Ours {ws_o.max_row}r x {ws_o.max_column}c | Theirs {ws_t.max_row}r x {ws_t.max_column}c")

# Full header comparison
print(f"\n  {'Col':3} | {'OURS (row 2)':24} | {'THEIRS (row 2)':24}")
print(f"  " + "-" * 60)
for c in range(1, 25):
    o_val = str(ws_o.cell(row=2, column=c).value or "")[:24]
    t_val = str(ws_t.cell(row=2, column=c).value or "")[:24]
    match = "✓" if o_val == t_val else "✗"
    print(f"  {get_column_letter(c):3} | {o_val:24} | {t_val:24} {match}")

# Their data validations
print(f"\n  Their DRAFT 2 data validations:")
dvs = list(ws_t.data_validations.dataValidation) if ws_t.data_validations else []
for dv in dvs:
    print(f"    {dv.sqref} = {dv.formula1}")

# Their row 1 (title)
print(f"\n  Their DRAFT 2 title (row 1): {ws_t.cell(row=1, column=1).value}")

# ══════════════════════════════════════════════════════════════
# 5. LOG — column comparison
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 70}")
print("  5. LOG — COLUMN COMPARISON")
print(f"{'=' * 70}")
ws_o = wb_o["LOG"]
ws_t = wb_t["LOG"]
print(f"  Size: Ours {ws_o.max_row}r x {ws_o.max_column}c | Theirs {ws_t.max_row}r x {ws_t.max_column}c")

print(f"\n  {'Col':3} | {'OURS (row 2)':24} | {'THEIRS (row 1)':24}")
print(f"  " + "-" * 60)
for c in range(1, 25):
    o_val = str(ws_o.cell(row=2, column=c).value or "")[:24]
    t_val = str(ws_t.cell(row=1, column=c).value or "")[:24]
    match = "✓" if o_val == t_val else "✗"
    print(f"  {get_column_letter(c):3} | {o_val:24} | {t_val:24} {match}")

# ══════════════════════════════════════════════════════════════
# 6. AGENDA — column comparison
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 70}")
print("  6. AGENDA — COLUMN COMPARISON")
print(f"{'=' * 70}")
ws_o = wb_o["AGENDA"]
ws_t = wb_t["AGENDA"]
print(f"  Size: Ours {ws_o.max_row}r x {ws_o.max_column}c | Theirs {ws_t.max_row}r x {ws_t.max_column}c")

print(f"\n  {'Col':3} | {'OURS (row 4)':28} | {'THEIRS (row 4)':28}")
print(f"  " + "-" * 65)
for c in range(1, max(ws_o.max_column, ws_t.max_column) + 1):
    o_val = str(ws_o.cell(row=4, column=c).value or "")[:28]
    t_val = str(ws_t.cell(row=4, column=c).value or "")[:28]
    match = "✓" if o_val.strip() == t_val.strip() else "✗"
    if o_val or t_val:
        print(f"  {get_column_letter(c):3} | {o_val:28} | {t_val:28} {match}")

# ══════════════════════════════════════════════════════════════
# 7. DASH — full block comparison
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 70}")
print("  7. DASH — FULL STRUCTURE")
print(f"{'=' * 70}")
ws_o = wb_o["DASH"]
ws_t = wb_t["DASH"]
print(f"  Size: Ours {ws_o.max_row}r x {ws_o.max_column}c | Theirs {ws_t.max_row}r x {ws_t.max_column}c")

# All their DASH content organized by sections
print(f"\n  Their DASH full content (all rows with data in A or B):")
for r in range(1, ws_t.max_row + 1):
    vals = []
    for c in range(1, min(20, ws_t.max_column + 1)):
        v = ws_t.cell(row=r, column=c).value
        if v is not None:
            vals.append(f"{get_column_letter(c)}={repr(v)[:35]}")
    if vals:
        print(f"    R{r}: {' | '.join(vals[:6])}")

# Their DASH formulas (check a few key cells)
print(f"\n  Their DASH key formulas (loading with formulas):")
wb_t2 = openpyxl.load_workbook(THEIRS)
ws_t2 = wb_t2["DASH"]
for r in [4, 10, 27, 41, 58, 90, 106, 122]:
    for c in [2, 3, 5, 10]:
        v = ws_t2.cell(row=r, column=c).value
        if v and str(v).startswith("="):
            print(f"    R{r}C{c}: {str(v)[:80]}")
wb_t2.close()

# ══════════════════════════════════════════════════════════════
# 8. CARTEIRA — REDE_GRUPO col, TIPO_CLIENTE col differences
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 70}")
print("  8. CARTEIRA — COLUMN DIFFERENCES")
print(f"{'=' * 70}")
ws_o = wb_o["CARTEIRA"]
ws_t = wb_t["CARTEIRA"]

diffs = []
for c in range(1, 73):  # first 72 cols (before ACOMP)
    o_val = str(ws_o.cell(row=3, column=c).value or "")
    t_val = str(ws_t.cell(row=3, column=c).value or "")
    if o_val != t_val:
        diffs.append((c, o_val, t_val))

if diffs:
    print(f"\n  Differences in headers (cols 1-72):")
    for c, o, t in diffs:
        print(f"    Col {c} ({get_column_letter(c)}): OURS={o} | THEIRS={t}")
else:
    print(f"  No header differences in cols 1-72")

# Check their col 10 (TIPO CLIENTE vs REDE GRUPO)
print(f"\n  Their CARTEIRA col 10 header: {ws_t.cell(row=3, column=10).value}")
print(f"  Our CARTEIRA col 10 header: {ws_o.cell(row=3, column=10).value}")

# ══════════════════════════════════════════════════════════════
# 9. Claude Log comparison
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 70}")
print("  9. CLAUDE LOG")
print(f"{'=' * 70}")
ws_o = wb_o["Claude Log"]
ws_t = wb_t["Claude Log"]
print(f"  Size: Ours {ws_o.max_row}r x {ws_o.max_column}c | Theirs {ws_t.max_row}r x {ws_t.max_column}c")
print(f"\n  Their headers: ", end="")
for c in range(1, ws_t.max_column + 1):
    v = ws_t.cell(row=1, column=c).value
    if v:
        print(f"{v}, ", end="")
print()
print(f"  Our headers: ", end="")
for c in range(1, ws_o.max_column + 1):
    v = ws_o.cell(row=2, column=c).value
    if v:
        print(f"{v}, ", end="")
print()

# ══════════════════════════════════════════════════════════════
# 10. Data Validations full comparison
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 70}")
print("  10. DATA VALIDATIONS — ALL TABS")
print(f"{'=' * 70}")
for name in wb_t.sheetnames:
    ws_check = wb_t[name]
    dvs = list(ws_check.data_validations.dataValidation) if ws_check.data_validations else []
    if dvs:
        print(f"\n  {name} ({len(dvs)} validations):")
        for dv in dvs:
            print(f"    {dv.sqref}: type={dv.type} formula={dv.formula1}")

# Same for ours
print(f"\n  --- OURS ---")
for name in wb_o.sheetnames:
    ws_check = wb_o[name]
    dvs = list(ws_check.data_validations.dataValidation) if ws_check.data_validations else []
    if dvs:
        print(f"\n  {name} ({len(dvs)} validations):")
        for dv in dvs:
            print(f"    {dv.sqref}: type={dv.type} formula={dv.formula1}")

# ══════════════════════════════════════════════════════════════
# 11. Conditional Formatting comparison
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 70}")
print("  11. CONDITIONAL FORMATTING — ALL TABS")
print(f"{'=' * 70}")
for name in wb_t.sheetnames:
    ws_check = wb_t[name]
    cf_list = list(ws_check.conditional_formatting)
    total = sum(len(cf.rules) for cf in cf_list)
    if total:
        print(f"\n  {name}: {total} rules across {len(cf_list)} ranges")

print(f"\n  --- OURS ---")
for name in wb_o.sheetnames:
    ws_check = wb_o[name]
    cf_list = list(ws_check.conditional_formatting)
    total = sum(len(cf.rules) for cf in cf_list)
    if total:
        print(f"\n  {name}: {total} rules across {len(cf_list)} ranges")

# ══════════════════════════════════════════════════════════════
# 12. Freeze panes comparison
# ══════════════════════════════════════════════════════════════
print(f"\n{'=' * 70}")
print("  12. FREEZE PANES")
print(f"{'=' * 70}")
for name in wb_t.sheetnames:
    if name in wb_o.sheetnames:
        o_f = str(wb_o[name].freeze_panes) if wb_o[name].freeze_panes else "None"
        t_f = str(wb_t[name].freeze_panes) if wb_t[name].freeze_panes else "None"
        match = "✓" if o_f == t_f else "✗"
        print(f"  {name:15} Ours={o_f:6} Theirs={t_f:6} {match}")

wb_o.close()
wb_t.close()

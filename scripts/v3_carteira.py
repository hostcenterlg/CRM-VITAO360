"""
V3 — Tab CARTEIRA: Visão 360 (~257 colunas, 4 mega-blocos)
"""
from openpyxl.utils import get_column_letter
from v3_styles import *

MESES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]


def build_carteira(wb):
    """Build CARTEIRA tab (~257 cols, 4 mega-blocks)."""
    ws = wb.create_sheet("CARTEIRA")
    ws.sheet_properties.tabColor = TAB_CARTEIRA

    # ═══════════════════════════════════════════════════════════
    # MEGA-BLOCO 1: 🟣 MERCOS (cols 1-43)
    # ═══════════════════════════════════════════════════════════
    mercos_cols = [
        # (col, header, sub_block, width)
        (1,  "NOME FANTASIA",         "ANCORA",       25),
        (2,  "CNPJ",                  "IDENTIDADE",   18),
        (3,  "RAZÃO SOCIAL",          "IDENTIDADE",   30),
        (4,  "UF",                    "IDENTIDADE",    5),
        (5,  "CIDADE",                "IDENTIDADE",   18),
        (6,  "EMAIL",                 "IDENTIDADE",   28),
        (7,  "TELEFONE",              "IDENTIDADE",   16),
        (8,  "DATA CADASTRO",         "IDENTIDADE",   14),
        (9,  "REDE / REGIONAL",       "REDE",         22),
        (10, "REDE GRUPO",            "REDE",         16),
        (11, "ÚLT. REGISTRO MERCOS",  "REDE",         18),
        (12, "CONSULTOR",             "EQUIPE",       20),
        (13, "VENDEDOR ÚLTIMO PEDIDO","EQUIPE",       22),
        (14, "SITUAÇÃO",              "STATUS",       14),
        (15, "PRIORIDADE",            "STATUS",       10),
        (16, "DIAS SEM COMPRA",       "COMPRA",       16),
        (17, "DATA ÚLTIMO PEDIDO",    "COMPRA",       18),
        (18, "VALOR ÚLTIMO PEDIDO",   "COMPRA",       18),
        (19, "📊 ACESSOS SEMANA",     "ECOMMERCE",    16),
        (20, "ACESSO B2B",            "ECOMMERCE",    12),
        (21, "ACESSOS PORTAL",        "ECOMMERCE",    14),
        (22, "ITENS CARRINHO",        "ECOMMERCE",    14),
        (23, "VALOR B2B",             "ECOMMERCE",    12),
        (24, "OPORTUNIDADE",          "ECOMMERCE",    14),
        (25, "TOTAL PERÍODO",         "VENDAS",       14),
        (26, "MAR/25",                "VENDAS",       10),
        (27, "ABR/25",                "VENDAS",       10),
        (28, "MAI/25",                "VENDAS",       10),
        (29, "JUN/25",                "VENDAS",       10),
        (30, "JUL/25",                "VENDAS",       10),
        (31, "AGO/25",                "VENDAS",       10),
        (32, "SET/25",                "VENDAS",       10),
        (33, "OUT/25",                "VENDAS",       10),
        (34, "NOV/25",                "VENDAS",       10),
        (35, "DEZ/25",                "VENDAS",       10),
        (36, "JAN/26",                "VENDAS",       10),
        (37, "CICLO MÉDIO",           "RECORRÊNCIA",  12),
        (38, "Nº COMPRAS",            "RECORRÊNCIA",  10),
        (39, "CURVA ABC",             "RECORRÊNCIA",  10),
        (40, "MESES POSITIVADO",      "RECORRÊNCIA",  14),
        (41, "MÉDIA MENSAL",          "RECORRÊNCIA",  12),
        (42, "TICKET MÉDIO",          "RECORRÊNCIA",  12),
        (43, "MESES LISTA",           "RECORRÊNCIA",  10),
    ]

    # ═══════════════════════════════════════════════════════════
    # MEGA-BLOCO 2: 🔵 FUNIL (cols 44-61)
    # ═══════════════════════════════════════════════════════════
    funil_cols = [
        (44, "ESTÁGIO FUNIL",         "ANCORA",       18),
        (45, "PRÓX. FOLLOW-UP",       "PIPELINE",     16),
        (46, "DATA ÚLT.ATENDIMENTO",  "PIPELINE",     18),
        (47, "ÚLTIMO RESULTADO",       "PIPELINE",     20),
        (48, "MOTIVO",                "PIPELINE",     22),
        (49, "TIPO CLIENTE",          "PERFIL",       16),
        (50, "TENTATIVA",             "PERFIL",       12),
        (51, "FASE",                  "MATURIDADE",   16),
        (52, "ÚLTIMA RECOMPRA",       "MATURIDADE",   16),
        (53, "🔥 TEMPERATURA",        "CONVERSÃO",    14),
        (54, "DIAS ATÉ CONVERSÃO",    "CONVERSÃO",    16),
        (55, "DATA 1º CONTATO",       "CONVERSÃO",    16),
        (56, "DATA 1º ORÇAMENTO",     "CONVERSÃO",    16),
        (57, "DATA 1ª VENDA",         "CONVERSÃO",    14),
        (58, "TOTAL TENTATIVAS",      "CONVERSÃO",    14),
        (59, "PRÓX. AÇÃO",            "AÇÃO",         14),
        (60, "AÇÃO DETALHADA",        "AÇÃO",         28),
        (61, "🚦 SINALEIRO",          "SINAL",        12),
    ]

    # ═══════════════════════════════════════════════════════════
    # MEGA-BLOCO 3: ⚫ SAP (cols 62-72)
    # ═══════════════════════════════════════════════════════════
    sap_cols = [
        (62, "CÓDIGO DO CLIENTE",        "DADOS SAP", 16),
        (63, "DESCRIÇÃO GRUPO CLIENTE",  "SAP",       22),
        (64, "ZP NOME GERENTE NACIONAL", "SAP",       22),
        (65, "ZR NOME REPRESENTANTE",    "SAP",       22),
        (66, "ZV NOME VEND. INTERNO",    "SAP",       22),
        (67, "01 NOME DO CANAL",         "SAP",       18),
        (68, "02 NOME TIPO CLIENTE",     "SAP",       18),
        (69, "03 NOME MACROREGIÃO",      "SAP",       18),
        (70, "04 NOME MICROREGIÃO",      "SAP",       18),
        (71, "06 NOME GRUPO CHAVE",      "SAP",       20),
        (72, "VENDA",                    "SAP",       14),
    ]

    # ═══════════════════════════════════════════════════════════
    # MEGA-BLOCO 4: 🟢 ACOMPANHAMENTO (cols 73-257)
    # 15 cols per month × 12 months = 180 + 4 quarterly + 1 % ating = 185
    # ═══════════════════════════════════════════════════════════
    # Structure: % ATING (73) | Q1(74) | JAN(15cols) | FEV(15cols) | MAR(15cols) |
    #            Q2 | ABR | MAI | JUN | Q3 | JUL | AGO | SET | Q4 | OUT | NOV | DEZ

    MONTH_COLS_PER = [
        "% YTD", "META", "REALIZADO", "% TRI", "META", "REALIZADO",
        "% MÊS", "META", "REALIZADO", "DATA PEDIDO",
        "JUSTIFICATIVA SEMANA 1", "JUSTIFICATIVA SEMANA 2",
        "JUSTIFICATIVA SEMANA 3", "JUSTIFICATIVA SEMANA 4",
        "JUSTIFICATIVA MENSAL",
    ]

    acomp_start = 73
    col = acomp_start

    # % ATING (col 73)
    all_cols = [(col, "% ATING.", "ACOMP", 10)]
    col += 1

    month_starts = {}  # track where each month starts

    for q_idx in range(4):  # Q1-Q4
        q_name = f"Q{q_idx+1}"
        all_cols.append((col, f"% {q_name}", "ACOMP", 8))
        col += 1

        for m_idx in range(3):
            month_idx = q_idx * 3 + m_idx
            mes = MESES[month_idx]
            month_starts[mes] = col
            for sub_idx, sub_name in enumerate(MONTH_COLS_PER):
                w = 10
                if "JUSTIFICATIVA" in sub_name:
                    w = 20
                elif sub_name == "DATA PEDIDO":
                    w = 14
                elif "%" in sub_name:
                    w = 8
                elif sub_name in ["META", "REALIZADO"]:
                    w = 12
                all_cols.append((col, sub_name, mes, w))
                col += 1

    total_cols = col - 1

    # ═══════════════════════════════════════════════════════════
    # BUILD HEADERS
    # ═══════════════════════════════════════════════════════════

    # Row 1: Mega-block titles
    mega_blocks_r1 = [
        ("🟣 MERCOS", 1, 43, FILL_MERCOS),
        ("🔵 FUNIL", 44, 61, FILL_FUNIL),
        ("⚫ SAP", 62, 72, FILL_SAP),
        ("🟢 ACOMPANHAMENTO", 73, total_cols, FILL_ACOMP),
    ]

    for label, start, end, fill in mega_blocks_r1:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start, value=label)
        style_cell(cell, font=FONT_MEGA, fill=fill, align=ALIGN_CENTER)
        for c in range(start, end + 1):
            ws.cell(row=1, column=c).fill = fill
            ws.cell(row=1, column=c).border = THIN_BORDER

    # Row 2: Sub-block headers for MERCOS, FUNIL, SAP (each built separately)
    sub_block_fills = {
        "ANCORA": FILL_HEADER, "IDENTIDADE": FILL_BLOCK_IDENTIDADE,
        "REDE": PatternFill('solid', fgColor='37474F'),
        "EQUIPE": PatternFill('solid', fgColor='455A64'),
        "STATUS": FILL_BLOCK_STATUS, "COMPRA": FILL_BLOCK_COMPRA,
        "ECOMMERCE": FILL_BLOCK_ECOMMERCE, "VENDAS": FILL_BLOCK_VENDAS,
        "RECORRÊNCIA": FILL_BLOCK_RECORRENCIA,
        "PIPELINE": FILL_BLOCK_PIPELINE, "PERFIL": FILL_BLOCK_PERFIL,
        "MATURIDADE": FILL_BLOCK_MATURIDADE, "CONVERSÃO": FILL_BLOCK_CONVERSAO,
        "AÇÃO": FILL_BLOCK_ACAO, "SINAL": FILL_BLOCK_SINAL,
        "DADOS SAP": FILL_SAP, "SAP": PatternFill('solid', fgColor='616161'),
    }

    # Build sub-block map per mega-block to avoid name collisions
    for col_def_list in [mercos_cols, funil_cols, sap_cols]:
        sub_block_map = {}
        for c, h, sub, w in col_def_list:
            if sub not in sub_block_map:
                sub_block_map[sub] = [c, c]
            else:
                sub_block_map[sub][1] = c

        for sub, (start_c, end_c) in sub_block_map.items():
            fill = sub_block_fills.get(sub, FILL_SUB_HEADER)
            if end_c > start_c:
                ws.merge_cells(start_row=2, start_column=start_c, end_row=2, end_column=end_c)
            cell = ws.cell(row=2, column=start_c, value=sub)
            style_cell(cell, font=FONT_BLOCK, fill=fill, align=ALIGN_CENTER)
            for c in range(start_c, end_c + 1):
                ws.cell(row=2, column=c).fill = fill
                ws.cell(row=2, column=c).border = THIN_BORDER

    # Row 2 for ACOMPANHAMENTO: YTD/TRI/MÊS sub-headers
    for c, h, sub, w in all_cols:
        if sub in MESES:
            # Determine sub-group (YTD/TRI/MÊS)
            offset = c - month_starts[sub]
            if offset < 3:
                grp = "YTD"
            elif offset < 6:
                grp = "TRI"
            elif offset < 9:
                grp = "MÊS"
            else:
                grp = "DETALHE"
            ws.cell(row=2, column=c).value = grp
            ws.cell(row=2, column=c).font = FONT_HEADER_SM
            ws.cell(row=2, column=c).fill = FILL_ACOMP
            ws.cell(row=2, column=c).border = THIN_BORDER
            ws.cell(row=2, column=c).alignment = ALIGN_CENTER

    # Row 3: Column headers
    for col_def_list in [mercos_cols, funil_cols, sap_cols]:
        for c, h, sub, w in col_def_list:
            write_header(ws, 3, c, h)
            ws.column_dimensions[get_column_letter(c)].width = w

    for c, h, sub, w in all_cols:
        write_header(ws, 3, c, h)
        ws.column_dimensions[get_column_letter(c)].width = w

    # ═══════════════════════════════════════════════════════════
    # COLUMN GROUPING
    # ═══════════════════════════════════════════════════════════
    # IDENTIDADE (2-8)
    group_columns(ws, 2, 8, outline_level=1, hidden=True)
    # ECOMMERCE (19-24)
    group_columns(ws, 19, 24, outline_level=1, hidden=True)
    # VENDAS (25-36)
    group_columns(ws, 25, 36, outline_level=1, hidden=True)
    # RECORRÊNCIA (37-43)
    group_columns(ws, 37, 43, outline_level=1, hidden=True)
    # FUNIL sub-groups (45-61)
    group_columns(ws, 45, 61, outline_level=1, hidden=True)
    # SAP (62-72)
    group_columns(ws, 62, 72, outline_level=1, hidden=True)
    # Each month in ACOMPANHAMENTO
    for mes in MESES:
        if mes in month_starts:
            start_c = month_starts[mes]
            end_c = start_c + len(MONTH_COLS_PER) - 1
            group_columns(ws, start_c, end_c, outline_level=1, hidden=True)

    # ═══════════════════════════════════════════════════════════
    # CONDITIONAL FORMATTING — SITUAÇÃO (col N = 14)
    # ═══════════════════════════════════════════════════════════
    from openpyxl.formatting.rule import CellIsRule
    for sit, fill in SITUACAO_FILLS.items():
        ws.conditional_formatting.add(
            'N4:N5000',
            CellIsRule(operator='equal', formula=[f'"{sit}"'], fill=fill)
        )

    # ═══════════════════════════════════════════════════════════
    # FORMATTING & FREEZE
    # ═══════════════════════════════════════════════════════════
    ws.freeze_panes = "C4"  # lock NOME + CNPJ visible

    # Row heights
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20

    print(f"  CARTEIRA: {total_cols} cols ({len(mercos_cols)} Mercos + {len(funil_cols)} Funil + "
          f"{len(sap_cols)} SAP + {len(all_cols)} Acomp), freeze C4")
    return ws

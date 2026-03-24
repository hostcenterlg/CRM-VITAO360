"""
Pipeline de importacao do Motor Operacional CRM VITAO360 v2.0.

Le a planilha FINAL (40 abas, 6.2 MB), extrai DataFrames com headers
corretos, normaliza CNPJs e vendedores, retorna dict de DataFrames nomeados.

Uso:
    from scripts.motor.import_pipeline import importar_planilha
    dfs = importar_planilha()
"""

import logging
import time
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

from scripts.motor.config import (
    ABAS_CONSULTOR,
    ABAS_RELEVANTES,
    CAMINHO_PLANILHA,
    validar_caminho_planilha,
)
from scripts.motor.helpers import normalizar_cnpj, normalizar_vendedor, safe_read_sheet

logger = logging.getLogger("motor.import_pipeline")

# ---------------------------------------------------------------------------
# Configuracao de headers especiais por aba
# Cada entrada: (header_row_1indexed, data_start_row_1indexed)
# Se a aba nao esta aqui, assume header=row1, data=row2
# ---------------------------------------------------------------------------
_HEADER_CONFIG: dict[str, tuple[int, int]] = {
    "CARTEIRA": (3, 4),           # rows 1-2 sao super-grupo
    "DRAFT 1": (3, 4),            # rows 1-2 sao agrupamento
    "OPERACIONAL": (2, 3),        # row 1 vazia
    "DRAFT 3 ": (2, 3),           # row 1 eh super-grupo
    "SINALEIRO": (4, 5),          # rows 1-3 sao titulo/resumo/grupo
    "MOTOR DE REGRAS": (4, 5),    # rows 1-3 sao titulo/descricao
    "RNC": (3, 4),                # rows 1-2 sao titulo/descricao
    "AGENDA": (4, 5),             # rows 1-3 sao titulo/descricao (need to verify)
    "PROJE\u00c7\u00c3O ": (3, 4),  # rows 1-2 sao titulo/grupo
    "RESUMO META": (2, 3),        # row 1 vazia, row 2 eh header meses
    "PAINEL SINALEIRO ": (4, 5),  # rows 1-3 titulo/descricao
    "REGRAS": (1, 2),             # especial: varias tabelas, pegar raw
}

# Colunas que contem CNPJ por aba (nome da coluna no header)
_CNPJ_COLUMNS: dict[str, str] = {
    "CARTEIRA": "CNPJ",
    "OPERACIONAL": "CNPJ",
    "DRAFT 1": "CNPJ",
    "DRAFT 2": "CNPJ",
    "DRAFT 3 ": "CNPJ",
    "SINALEIRO": "CNPJ",
    "RNC": "CNPJ",
    "PROJE\u00c7\u00c3O ": "CNPJ",
    "Venda M\u00eas a M\u00eas": "CNPJ ou CPF Cliente",
    # Consultor tabs
    "LARISSA": "CNPJ",
    "MANU": "CNPJ",
    "JULIO": "CNPJ",
    "DAIANE": "CNPJ",
}

# Coluna de CONSULTOR na CARTEIRA
_CONSULTOR_COLUMN = "CONSULTOR"


def _ws_to_dataframe(
    ws,
    sheet_name: str,
    header_row: int,
    data_start_row: int,
) -> pd.DataFrame:
    """Converte worksheet openpyxl para DataFrame.

    Args:
        ws: Worksheet openpyxl (read_only mode)
        sheet_name: Nome da aba (para logging)
        header_row: Row do header (1-indexed)
        data_start_row: Row onde dados comecam (1-indexed)

    Returns:
        DataFrame com colunas nomeadas
    """
    t0 = time.time()

    # Extrair headers
    headers = []
    for row in ws.iter_rows(
        min_row=header_row, max_row=header_row, values_only=True
    ):
        headers = [
            str(h).strip() if h is not None else f"_col_{i}"
            for i, h in enumerate(row)
        ]
        break

    if not headers:
        logger.warning("Aba '%s': nenhum header encontrado na row %d", sheet_name, header_row)
        return pd.DataFrame()

    # Deduplicar headers (abas podem ter colunas com mesmo nome)
    seen: dict[str, int] = {}
    deduped: list[str] = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            deduped.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            deduped.append(h)
    headers = deduped

    # Extrair dados
    data_rows = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        # Pular rows completamente vazias
        if all(v is None for v in row):
            continue
        # Ajustar tamanho da row para bater com headers
        row_list = list(row)
        if len(row_list) < len(headers):
            row_list.extend([None] * (len(headers) - len(row_list)))
        elif len(row_list) > len(headers):
            row_list = row_list[: len(headers)]
        data_rows.append(row_list)

    elapsed = time.time() - t0

    if not data_rows:
        logger.warning("Aba '%s': 0 rows de dados (header row %d)", sheet_name, header_row)
        return pd.DataFrame(columns=headers)

    df = pd.DataFrame(data_rows, columns=headers)

    logger.info(
        "Aba '%s': %d rows x %d cols em %.1fs",
        sheet_name,
        len(df),
        len(df.columns),
        elapsed,
    )
    return df


def _normalizar_cnpjs_df(df: pd.DataFrame, sheet_name: str, cnpj_col: str) -> pd.DataFrame:
    """Aplica normalizar_cnpj na coluna CNPJ e adiciona 'cnpj_normalizado'.

    Args:
        df: DataFrame
        sheet_name: Nome da aba (para logging)
        cnpj_col: Nome da coluna com CNPJ

    Returns:
        DataFrame com coluna 'cnpj_normalizado' adicionada
    """
    if cnpj_col not in df.columns:
        # Tentar busca case-insensitive
        for col in df.columns:
            if "cnpj" in col.lower():
                cnpj_col = col
                break
        else:
            logger.warning(
                "Aba '%s': coluna CNPJ '%s' nao encontrada. Colunas: %s",
                sheet_name,
                cnpj_col,
                list(df.columns[:10]),
            )
            return df

    antes_nulos = df[cnpj_col].isna().sum()
    # Aplicar normalizar_cnpj e garantir que None permanece None (nao NaN float)
    normalized = df[cnpj_col].apply(normalizar_cnpj)
    # Converter para object dtype para preservar None como None (nao NaN)
    df["cnpj_normalizado"] = normalized.astype(object)
    # Substituir NaN por None explicitamente
    df["cnpj_normalizado"] = df["cnpj_normalizado"].where(df["cnpj_normalizado"].notna(), None)
    depois_nulos = df["cnpj_normalizado"].isna().sum()

    total = len(df)
    validos = total - depois_nulos
    logger.info(
        "Aba '%s': CNPJs normalizados — %d/%d validos (antes: %d nulos, depois: %d nulos)",
        sheet_name,
        validos,
        total,
        antes_nulos,
        depois_nulos,
    )

    return df


def _normalizar_vendedores_df(df: pd.DataFrame) -> pd.DataFrame:
    """Aplica normalizar_vendedor na coluna CONSULTOR da CARTEIRA.

    Adiciona coluna 'consultor_normalizado'.
    """
    if _CONSULTOR_COLUMN not in df.columns:
        # Busca case-insensitive
        for col in df.columns:
            if "consultor" in col.lower():
                _col = col
                break
        else:
            logger.warning(
                "CARTEIRA: coluna CONSULTOR nao encontrada. Colunas: %s",
                list(df.columns[:15]),
            )
            return df
    else:
        _col = _CONSULTOR_COLUMN

    df["consultor_normalizado"] = df[_col].apply(normalizar_vendedor)

    dist = df["consultor_normalizado"].value_counts()
    logger.info("CARTEIRA vendedores: %s", dict(dist))

    return df


def importar_planilha(caminho: Optional[Path] = None) -> dict[str, pd.DataFrame]:
    """Pipeline principal: le xlsx FINAL e retorna dict de DataFrames.

    Estrategia:
    1. Abrir workbook com openpyxl (data_only=True, read_only=True)
    2. Para cada aba relevante + consultores: extrair DataFrame
    3. Normalizar CNPJs em todas as abas que tem CNPJ
    4. Normalizar vendedores na CARTEIRA
    5. Retornar dict com chaves = nomes logicos

    Args:
        caminho: Path da planilha (default: CAMINHO_PLANILHA do config)

    Returns:
        Dict[str, pd.DataFrame] com chaves = nomes logicos das abas
    """
    caminho = validar_caminho_planilha(caminho)

    t_total = time.time()
    print(f"=== IMPORT PIPELINE: Abrindo {caminho.name} ===")

    # 1. Abrir workbook
    t0 = time.time()
    wb = openpyxl.load_workbook(str(caminho), data_only=True, read_only=True)
    print(f"Workbook aberto em {time.time() - t0:.1f}s ({len(wb.sheetnames)} abas)")

    result: dict[str, pd.DataFrame] = {}

    try:
        # 2. Processar abas relevantes
        print("\n--- Abas relevantes ---")
        for nome_logico, nome_real in ABAS_RELEVANTES.items():
            ws = safe_read_sheet(wb, nome_real)
            if ws is None:
                print(f"  SKIP: {nome_logico} (aba '{nome_real}' nao encontrada)")
                continue

            # Determinar config de header
            if nome_real in _HEADER_CONFIG:
                header_row, data_start = _HEADER_CONFIG[nome_real]
            else:
                header_row, data_start = 1, 2

            df = _ws_to_dataframe(ws, nome_real, header_row, data_start)

            # Normalizar CNPJ se aplicavel
            if nome_real in _CNPJ_COLUMNS:
                df = _normalizar_cnpjs_df(df, nome_real, _CNPJ_COLUMNS[nome_real])

            # Normalizar vendedores na CARTEIRA
            if nome_real == "CARTEIRA":
                df = _normalizar_vendedores_df(df)

            result[nome_logico] = df
            print(f"  OK: {nome_logico} = {len(df)} rows x {len(df.columns)} cols")

        # 3. Processar abas consultor
        print("\n--- Abas consultor ---")
        for consultor, nome_aba in ABAS_CONSULTOR.items():
            ws = safe_read_sheet(wb, nome_aba)
            if ws is None:
                print(f"  SKIP: {consultor} (aba '{nome_aba}' nao encontrada)")
                continue

            # Consultor tabs: header na row 1, dados na row 2
            df = _ws_to_dataframe(ws, nome_aba, header_row=1, data_start_row=2)

            # Normalizar CNPJ
            if nome_aba in _CNPJ_COLUMNS:
                df = _normalizar_cnpjs_df(df, nome_aba, _CNPJ_COLUMNS[nome_aba])

            result[f"consultor_{consultor.lower()}"] = df
            print(f"  OK: consultor_{consultor.lower()} = {len(df)} rows x {len(df.columns)} cols")
    finally:
        # 4. Fechar workbook (garante close mesmo com excecao)
        wb.close()

    # 5. Resumo
    elapsed_total = time.time() - t_total
    print(f"\n=== IMPORT PIPELINE COMPLETO ===")
    print(f"Tempo total: {elapsed_total:.1f}s")
    print(f"Abas lidas: {len(result)}/{len(ABAS_RELEVANTES) + len(ABAS_CONSULTOR)}")

    for nome, df in result.items():
        print(f"  {nome}: {len(df)} rows x {len(df.columns)} cols")

    # CNPJs resumo geral
    total_cnpjs = 0
    total_unicos = set()
    total_nulos = 0
    for nome, df in result.items():
        if "cnpj_normalizado" in df.columns:
            cnpjs = df["cnpj_normalizado"]
            total_cnpjs += len(cnpjs)
            total_unicos.update(cnpjs.dropna().tolist())
            total_nulos += cnpjs.isna().sum()

    print(f"\nCNPJs normalizados: {total_cnpjs} total, {len(total_unicos)} unicos, {total_nulos} nulos")

    # Vendedores resumo
    if "carteira" in result and "consultor_normalizado" in result["carteira"].columns:
        vendedores = result["carteira"]["consultor_normalizado"].value_counts()
        print(f"Vendedores: {dict(vendedores)}")

    # Alertas de colunas None
    for nome, df in result.items():
        if len(df) > 10:
            all_none_cols = [c for c in df.columns if df[c].isna().all()]
            if len(all_none_cols) > len(df.columns) * 0.5:
                print(
                    f"\nALERTA: Aba '{nome}' tem {len(all_none_cols)}/{len(df.columns)} "
                    f"colunas 100% None. Abra no Excel, salve, e reimporte."
                )

    return result


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(name)s - %(message)s")
    dfs = importar_planilha()

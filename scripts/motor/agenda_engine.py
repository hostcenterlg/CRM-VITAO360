"""
Agenda Engine — Modulo de Saida do Motor CRM VITAO360 v1.0.

Gera agendas diarias priorizadas por consultor (P0-P7), exporta para
Excel com formatacao por prioridade e JSON para consumo de API.

Fluxo esperado:
    1. motor_regras.aplicar_regras_batch(df)     -> adiciona 8 colunas de regras
    2. score_engine.calcular_score_batch(df)      -> adiciona score + prioridade
    3. score_engine.aplicar_meta_balance(df)      -> ajusta PROSPECCAO se necessario
    4. agenda_engine.gerar_agenda(df)             -> dict {consultor: DataFrame}
    5. agenda_engine.exportar_agenda_excel(...)   -> arquivo .xlsx com 1 aba por consultor
    6. agenda_engine.exportar_agenda_json(...)    -> arquivo .json para API

Regras inviolaveis aplicadas:
    - Two-Base: R$ APENAS em VENDA, LOG = R$ 0.00 (R4)
    - CNPJ sempre string 14 digitos (R5)
    - P0 nao conta no limite de 40 atendimentos
    - P7 NUNCA entra na agenda diaria
    - Visual LIGHT: Arial 9pt, cores por prioridade (R9)
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd

# ---------------------------------------------------------------------------
# Setup de paths e logging
# ---------------------------------------------------------------------------
_ROOT = Path(__file__).resolve().parents[2]  # CRM-VITAO360/
_MOTOR_OUTPUT = _ROOT / "data" / "output" / "motor" / "motor_output.json"
_AGENDA_OUTPUT_DIR = _ROOT / "data" / "output" / "agenda"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("motor.agenda_engine")

# ---------------------------------------------------------------------------
# Constantes de negocio — alterar apenas com aprovacao L3
# ---------------------------------------------------------------------------

CONSULTORES: dict[str, dict[str, Any]] = {
    "LARISSA": {"territorio": "Brasil Interior", "max_atendimentos": 40},
    "MANU": {"territorio": "SC/PR/RS", "max_atendimentos": 40},
    "JULIO": {"territorio": "Cia Saude+Fitland", "max_atendimentos": 40},
    "DAIANE": {"territorio": "Redes Nacionais", "max_atendimentos": 20},
}

DE_PARA_CONSULTOR: dict[str, str] = {
    # MANU
    "MANU": "MANU",
    "MANU VITAO": "MANU",
    "MANU DITZEL": "MANU",
    "HEMANUELE": "MANU",
    "HEMANUELE DITZEL": "MANU",
    "HEMANUELE DITZEL (MANU)": "MANU",
    # LARISSA
    "LARISSA": "LARISSA",
    "LARI": "LARISSA",
    "LARISSA VITAO": "LARISSA",
    "MAIS GRANEL": "LARISSA",
    "RODRIGO": "LARISSA",
    # DAIANE
    "DAIANE": "DAIANE",
    "CENTRAL DAIANE": "DAIANE",
    "DAIANE VITAO": "DAIANE",
    "CENTRAL - DAIANE": "DAIANE",
    # JULIO
    "JULIO": "JULIO",
    "JULIO GADRET": "JULIO",
}

# Colunas obrigatorias na agenda de saida
COLUNAS_AGENDA: list[str] = [
    "prioridade",
    "score",
    "nome_fantasia",
    "cnpj",
    "situacao",
    "temperatura",
    "acao_futura",
    "followup_dias",
    "telefone",
    "sinaleiro",
]

# Cores por prioridade (LIGHT theme — R9)
CORES_PRIORIDADE: dict[str, str] = {
    "P0": "FF0000",  # vermelho — IMEDIATA
    "P1": "FFC000",  # laranja  — URGENTE
    "P2": "00B050",  # verde    — ALTA
    "P3": "FFFF00",  # amarelo  — MEDIA ALTA
    "P4": "FFFFFF",  # branco   — MEDIA
    "P5": "FFFFFF",  # branco   — MEDIA BAIXA
    "P6": "FFFFFF",  # branco   — BAIXA
    "P7": "D9D9D9",  # cinza    — CAMPANHA (nunca na agenda)
}

# Labels descritivos de prioridade para exibicao
LABELS_PRIORIDADE: dict[str, str] = {
    "P0": "P0 IMEDIATA",
    "P1": "P1 URGENTE",
    "P2": "P2 ALTA",
    "P3": "P3 MEDIA ALTA",
    "P4": "P4 MEDIA",
    "P5": "P5 MEDIA BAIXA",
    "P6": "P6 BAIXA",
    "P7": "P7 CAMPANHA",
}

META_MENSAL_DEFAULT: float = 263_051.0  # R$ 3.156.614 / 12


# ---------------------------------------------------------------------------
# Funcao 1: normalizar_consultor
# ---------------------------------------------------------------------------

def normalizar_consultor(nome: str) -> str:
    """Normaliza nome de consultor usando tabela DE-PARA.

    Aplica strip + upper antes do lookup. Retorna "DESCONHECIDO" se
    nenhum match for encontrado — nunca inventa valor.

    Args:
        nome: Nome do consultor como aparece na planilha ou log.

    Returns:
        Nome canonico (LARISSA, MANU, JULIO, DAIANE) ou "DESCONHECIDO".
    """
    if not nome:
        return "DESCONHECIDO"
    nome_norm = str(nome).strip().upper()
    if nome_norm in ("NAN", "NONE", ""):
        return "DESCONHECIDO"
    # Lookup exato primeiro
    if nome_norm in DE_PARA_CONSULTOR:
        return DE_PARA_CONSULTOR[nome_norm]
    # Lookup por substring (alias parcial)
    for alias, canonical in DE_PARA_CONSULTOR.items():
        if alias in nome_norm or nome_norm in alias:
            return canonical
    return "DESCONHECIDO"


# ---------------------------------------------------------------------------
# Funcao 2: filtrar_por_consultor
# ---------------------------------------------------------------------------

def filtrar_por_consultor(
    df: pd.DataFrame,
    consultor: str,
    col_consultor: str = "consultor",
) -> pd.DataFrame:
    """Filtra DataFrame retornando apenas linhas do consultor indicado.

    Normaliza a coluna de consultor antes do filtro para garantir match
    mesmo com grafias alternativas.

    Args:
        df: DataFrame com dados de clientes.
        consultor: Nome canonico do consultor (ex: "LARISSA").
        col_consultor: Nome da coluna que identifica o consultor.

    Returns:
        Subset do DataFrame filtrado. DataFrame vazio se consultor nao
        tiver registros ou coluna nao existir.
    """
    if col_consultor not in df.columns:
        logger.warning(
            "filtrar_por_consultor: coluna '%s' nao encontrada. Colunas: %s",
            col_consultor,
            list(df.columns),
        )
        return df.iloc[0:0].copy()  # DataFrame vazio com mesmas colunas

    # Normaliza coluna de consultor para comparacao
    col_normalizada = df[col_consultor].apply(
        lambda v: normalizar_consultor(str(v)) if pd.notna(v) else "DESCONHECIDO"
    )
    mascara = col_normalizada == consultor.strip().upper()
    resultado = df.loc[mascara].copy()
    logger.info(
        "filtrar_por_consultor: %s -> %d/%d registros",
        consultor,
        len(resultado),
        len(df),
    )
    return resultado


# ---------------------------------------------------------------------------
# Funcao 3: gerar_agenda (funcao principal)
# ---------------------------------------------------------------------------

def gerar_agenda(
    df: pd.DataFrame,
    data: str | None = None,
    meta_mensal: float = META_MENSAL_DEFAULT,
) -> dict[str, pd.DataFrame]:
    """Gera agendas diarias priorizadas para todos os consultores.

    Funcao principal do modulo. Recebe DataFrame ja processado pelo
    motor_regras e score_engine (com colunas 'score' e 'prioridade')
    e retorna dict com a agenda filtrada e ordenada por consultor.

    Pipeline interno por consultor:
        a. Filtra clientes do consultor
        b. Aplica meta_balance (boost PROSPECCAO se necessario)
        c. Ordena por prioridade (ordenar_por_prioridade)
        d. Seleciona agenda diaria (gerar_agenda_diaria)
        e. Extrai apenas linhas na_agenda=True
        f. Seleciona/renomeia colunas de saida

    Args:
        df: DataFrame com colunas de motor_regras + score_engine.
            Deve ter pelo menos: consultor, prioridade, score,
            e as colunas de COLUNAS_AGENDA (preenchidas com None/vazio
            se ausentes — nao quebra, apenas avisa).
        data: Data no formato "DD/MM/YYYY". Usa hoje se None.
        meta_mensal: Meta mensal em R$ para meta_balance (padrao ~263k).

    Returns:
        Dict {nome_consultor: DataFrame_agenda} com colunas de COLUNAS_AGENDA.
        Consultores sem registros recebem DataFrame vazio com as colunas.
    """
    # Importacoes locais para evitar circular import (score_engine importa de aqui nao)
    from scripts.motor.score_engine import (
        aplicar_meta_balance,
        gerar_agenda_diaria,
        ordenar_por_prioridade,
    )

    if data is None:
        data = datetime.now().strftime("%d/%m/%Y")

    # Verificar se score e prioridade ja existem; calcular se ausentes
    if "score" not in df.columns or "prioridade" not in df.columns:
        logger.warning(
            "gerar_agenda: colunas 'score'/'prioridade' ausentes. "
            "Execute calcular_score_batch antes desta funcao."
        )
        # Tentar calcular inline como fallback
        try:
            from scripts.motor.score_engine import calcular_score_batch
            df = calcular_score_batch(df)
        except Exception as exc:  # noqa: BLE001
            logger.error("Falha ao calcular score batch: %s", exc)
            # Retorna agendas vazias para todos os consultores
            colunas = {c: pd.Series(dtype="object") for c in COLUNAS_AGENDA}
            return {c: pd.DataFrame(colunas) for c in CONSULTORES}

    agendas: dict[str, pd.DataFrame] = {}

    for consultor, cfg in CONSULTORES.items():
        max_atend = cfg["max_atendimentos"]

        # a. Filtrar clientes do consultor
        df_consultor = filtrar_por_consultor(df, consultor)

        if df_consultor.empty:
            logger.info("Agenda %s: sem registros — agenda vazia gerada.", consultor)
            agendas[consultor] = _agenda_vazia()
            continue

        # b. Aplicar meta balance (boost PROSPECCAO se P2-P5 nao cobrem 80%)
        try:
            df_consultor = aplicar_meta_balance(df_consultor, meta_mensal=meta_mensal)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Meta balance falhou para %s: %s", consultor, exc)

        # c. Ordenar por prioridade e criterios de desempate
        try:
            df_consultor = ordenar_por_prioridade(df_consultor)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Ordenacao falhou para %s: %s", consultor, exc)

        # d. Gerar agenda diaria (marca na_agenda=True/False)
        try:
            df_consultor = gerar_agenda_diaria(df_consultor, max_atendimentos=max_atend)
        except Exception as exc:  # noqa: BLE001
            logger.warning("gerar_agenda_diaria falhou para %s: %s", consultor, exc)
            df_consultor["na_agenda"] = False

        # e. Filtrar somente quem esta na agenda
        df_agenda = df_consultor[df_consultor["na_agenda"] == True].copy()  # noqa: E712

        # f. Selecionar e padronizar colunas de saida
        df_agenda = _padronizar_colunas_agenda(df_agenda)

        agendas[consultor] = df_agenda
        logger.info(
            "Agenda %s: %d atendimentos | score medio: %.1f",
            consultor,
            len(df_agenda),
            df_agenda["score"].mean() if not df_agenda.empty else 0.0,
        )

    return agendas


def _agenda_vazia() -> pd.DataFrame:
    """Retorna DataFrame vazio com as colunas padrao de agenda."""
    return pd.DataFrame({c: pd.Series(dtype="object") for c in COLUNAS_AGENDA})


def _padronizar_colunas_agenda(df: pd.DataFrame) -> pd.DataFrame:
    """Seleciona e normaliza colunas de saida da agenda.

    Colunas ausentes sao preenchidas com None. Garante que CNPJ
    permanece como string (R5 — nunca float).

    Args:
        df: DataFrame com dados do consultor ja filtrado para agenda.

    Returns:
        DataFrame com exatamente as colunas de COLUNAS_AGENDA.
    """
    # Mapeamento de nomes alternativos -> nome padrao de saida
    _ALIASES: dict[str, list[str]] = {
        "nome_fantasia": ["nome_fantasia", "nome", "cliente", "nome_cliente", "razao_social"],
        "cnpj": ["cnpj", "CNPJ"],
        "situacao": ["situacao", "situacao_cliente", "status"],
        "temperatura": ["temperatura", "motor_temperatura"],
        "acao_futura": ["acao_futura", "motor_acao_futura"],
        "followup_dias": ["followup_dias", "motor_followup_dias", "followup"],
        "telefone": ["telefone", "fone", "celular", "tel"],
        "sinaleiro": ["sinaleiro", "cor_sinaleiro", "semaforo"],
    }

    resultado = pd.DataFrame(index=df.index)
    resultado["prioridade"] = df["prioridade"] if "prioridade" in df.columns else None
    resultado["score"] = df["score"] if "score" in df.columns else None

    for col_saida, candidatos in _ALIASES.items():
        valor = None
        for cand in candidatos:
            if cand in df.columns:
                valor = df[cand]
                break
        resultado[col_saida] = valor

    # Garantir CNPJ como string 14 digitos (R5)
    if resultado["cnpj"].notna().any():
        from scripts.motor.helpers import normalizar_cnpj
        resultado["cnpj"] = resultado["cnpj"].apply(normalizar_cnpj)

    # Ordenar colunas conforme especificacao
    return resultado[COLUNAS_AGENDA].reset_index(drop=True)


# ---------------------------------------------------------------------------
# Funcao 4: formatar_agenda_texto
# ---------------------------------------------------------------------------

def formatar_agenda_texto(
    agenda: pd.DataFrame,
    consultor: str,
    data: str,
) -> str:
    """Formata agenda como texto para console ou WhatsApp.

    Agrupa atendimentos por prioridade com cabecalho e numeracao sequencial.
    Nao emite emojis — output limpo para terminal e mensagens de texto.

    Formato de saida:
        AGENDA LARISSA - 23/03/2026
        40 atendimentos | Score medio: 72.3

        P0 IMEDIATA (2)
        1. LOJA ABC (12345678000190) - SUPORTE - Resolver problema
        2. LOJA XYZ (98765432000111) - SUPORTE - Resolver problema

        P1 URGENTE (10)
        3. CLIENTE A - EM ATENDIMENTO - Follow-up vencido
        ...

    Args:
        agenda: DataFrame com as colunas de COLUNAS_AGENDA.
        consultor: Nome canonico do consultor.
        data: Data no formato "DD/MM/YYYY".

    Returns:
        String formatada pronta para exibicao.
    """
    if agenda.empty:
        return (
            f"AGENDA {consultor} - {data}\n"
            "Sem atendimentos para hoje.\n"
        )

    total = len(agenda)
    score_medio = round(float(agenda["score"].mean()), 1) if "score" in agenda.columns else 0.0

    linhas: list[str] = [
        f"AGENDA {consultor} - {data}",
        f"{total} atendimentos | Score medio: {score_medio}",
        "",
    ]

    # Agrupar por prioridade mantendo a ordem P0..P6
    ordem_prio = ["P0", "P1", "P2", "P3", "P4", "P5", "P6"]
    counter = 1

    for prio in ordem_prio:
        grupo = agenda[agenda["prioridade"] == prio]
        if grupo.empty:
            continue

        label = LABELS_PRIORIDADE.get(prio, prio)
        linhas.append(f"{label} ({len(grupo)})")

        for _, row in grupo.iterrows():
            nome = _safe_str(row.get("nome_fantasia"), "SEM NOME")
            cnpj = _safe_str(row.get("cnpj"), "")
            situacao = _safe_str(row.get("situacao"), "-")
            acao = _safe_str(row.get("acao_futura"), "-")

            cnpj_fmt = f" ({cnpj})" if cnpj else ""
            linhas.append(f"{counter}. {nome}{cnpj_fmt} - {situacao} - {acao}")
            counter += 1

        linhas.append("")  # linha em branco entre grupos

    return "\n".join(linhas)


def _safe_str(val: Any, default: str = "") -> str:
    """Converte valor para string segura, usando default se None/NaN."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    s = str(val).strip()
    return s if s and s.lower() not in ("nan", "none", "nat") else default


# ---------------------------------------------------------------------------
# Funcao 5: exportar_agenda_excel
# ---------------------------------------------------------------------------

def exportar_agenda_excel(
    agendas: dict[str, pd.DataFrame],
    output_path: str,
    data: str,
) -> str:
    """Exporta agendas para arquivo Excel com uma aba por consultor.

    Formatacao LIGHT (R9):
        - Fonte: Arial 9pt (dados), 10pt (cabecalho)
        - Cores de linha por prioridade: P0=vermelho, P1=laranja,
          P2=verde, P3=amarelo, P4-P6=branco, P7=cinza
        - Cabecalho por aba: data + consultor + total atendimentos
        - Bordas finas em todas as celulas de dados

    Args:
        agendas: Dict {consultor: DataFrame_agenda}.
        output_path: Caminho completo do arquivo .xlsx a criar.
        data: Data no formato "DD/MM/YYYY" para exibir no cabecalho.

    Returns:
        Caminho absoluto do arquivo criado.

    Raises:
        ImportError: Se openpyxl nao estiver instalado.
        IOError: Se o arquivo nao puder ser salvo.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
    except ImportError as exc:
        raise ImportError(
            "openpyxl nao instalado. Execute: pip install openpyxl"
        ) from exc

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    # Remover aba padrao vazia
    wb.remove(wb.active)  # type: ignore[arg-type]

    # Estilos reutilizaveis
    fonte_dados = Font(name="Arial", size=9)
    fonte_header = Font(name="Arial", size=10, bold=True)
    fonte_titulo = Font(name="Arial", size=11, bold=True)

    fill_header = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    fonte_header_branca = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    borda_fina = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    alinhamento_esquerda = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Cabecalhos das colunas (exibicao amigavel)
    headers_display: list[str] = [
        "PRIORIDADE", "SCORE", "NOME FANTASIA", "CNPJ",
        "SITUACAO", "TEMPERATURA", "ACAO FUTURA",
        "FOLLOWUP (dias)", "TELEFONE", "SINALEIRO",
    ]

    larguras_colunas: list[float] = [12, 8, 35, 18, 14, 14, 40, 14, 18, 12]

    for consultor, agenda in agendas.items():
        # Nome da aba truncado a 31 chars (limite Excel)
        nome_aba = consultor[:31]
        ws = wb.create_sheet(title=nome_aba)

        total = len(agenda)
        score_medio = round(float(agenda["score"].mean()), 1) if (
            not agenda.empty and "score" in agenda.columns
        ) else 0.0
        territorio = CONSULTORES.get(consultor, {}).get("territorio", "")

        # --- Linha 1: Titulo ---
        ws.merge_cells("A1:J1")
        cell_titulo = ws["A1"]
        cell_titulo.value = (
            f"AGENDA {consultor} - {data} | {territorio} | "
            f"{total} atendimentos | Score medio: {score_medio}"
        )
        cell_titulo.font = fonte_titulo
        cell_titulo.alignment = alinhamento_centro
        cell_titulo.fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        ws.row_dimensions[1].height = 20

        # --- Linha 2: Cabecalhos de coluna ---
        for col_idx, header in enumerate(headers_display, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = fonte_header_branca
            cell.fill = fill_header
            cell.alignment = alinhamento_centro
            cell.border = borda_fina
        ws.row_dimensions[2].height = 18

        # --- Linhas de dados (a partir da linha 3) ---
        if agenda.empty:
            ws.merge_cells("A3:J3")
            cell_vazio = ws["A3"]
            cell_vazio.value = "Sem atendimentos para hoje."
            cell_vazio.font = fonte_dados
            cell_vazio.alignment = alinhamento_centro
        else:
            for row_offset, (_, row) in enumerate(agenda.iterrows(), start=0):
                excel_row = 3 + row_offset
                prio = _safe_str(row.get("prioridade"), "")

                # Cor de fundo pela prioridade
                cor_hex = CORES_PRIORIDADE.get(prio, "FFFFFF")
                fill_prio = PatternFill(
                    start_color=cor_hex, end_color=cor_hex, fill_type="solid"
                )
                # Fonte escura para prioridades com fundo escuro (P0=vermelho)
                fonte_row = Font(
                    name="Arial", size=9,
                    color="FFFFFF" if prio == "P0" else "000000",
                )

                valores_linha: list[Any] = [
                    prio,
                    row.get("score"),
                    _safe_str(row.get("nome_fantasia")),
                    _safe_str(row.get("cnpj")),
                    _safe_str(row.get("situacao")),
                    _safe_str(row.get("temperatura")),
                    _safe_str(row.get("acao_futura")),
                    row.get("followup_dias"),
                    _safe_str(row.get("telefone")),
                    _safe_str(row.get("sinaleiro")),
                ]

                for col_idx, valor in enumerate(valores_linha, start=1):
                    cell = ws.cell(row=excel_row, column=col_idx, value=valor)
                    cell.font = fonte_row
                    cell.fill = fill_prio
                    cell.border = borda_fina
                    cell.alignment = (
                        alinhamento_centro if col_idx in (1, 2, 5, 6, 8, 10)
                        else alinhamento_esquerda
                    )

                ws.row_dimensions[excel_row].height = 16

        # Larguras de coluna
        for col_idx, largura in enumerate(larguras_colunas, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = largura

        # Congelar cabecalhos (linha 1 e 2)
        ws.freeze_panes = "A3"

    output_path_str = str(output_path.resolve())
    wb.save(output_path_str)
    logger.info("Agenda Excel salva em: %s", output_path_str)
    return output_path_str


# ---------------------------------------------------------------------------
# Funcao 6: exportar_agenda_json
# ---------------------------------------------------------------------------

def exportar_agenda_json(
    agendas: dict[str, pd.DataFrame],
    output_path: str,
) -> str:
    """Exporta agendas para JSON para consumo de API.

    Estrutura de saida:
        {
          "gerado_em": "ISO timestamp",
          "total_consultores": N,
          "agendas": {
            "LARISSA": {
              "total": 40,
              "score_medio": 72.3,
              "por_prioridade": {"P0": 2, "P1": 10, ...},
              "atendimentos": [{"prioridade": ..., "cnpj": ..., ...}]
            },
            ...
          }
        }

    Args:
        agendas: Dict {consultor: DataFrame_agenda}.
        output_path: Caminho completo do arquivo .json a criar.

    Returns:
        Caminho absoluto do arquivo criado.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    payload: dict[str, Any] = {
        "gerado_em": datetime.now(tz=timezone.utc).isoformat(),
        "total_consultores": len(agendas),
        "agendas": {},
    }

    for consultor, agenda in agendas.items():
        total = len(agenda)
        score_medio = (
            round(float(agenda["score"].mean()), 1)
            if not agenda.empty and "score" in agenda.columns
            else 0.0
        )

        por_prioridade: dict[str, int] = {}
        if not agenda.empty and "prioridade" in agenda.columns:
            for prio in ["P0", "P1", "P2", "P3", "P4", "P5", "P6"]:
                por_prioridade[prio] = int((agenda["prioridade"] == prio).sum())

        atendimentos: list[dict[str, Any]] = []
        for _, row in agenda.iterrows():
            registro: dict[str, Any] = {}
            for col in COLUNAS_AGENDA:
                val = row.get(col)
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    registro[col] = None
                elif isinstance(val, float):
                    registro[col] = round(val, 2)
                else:
                    registro[col] = val
            atendimentos.append(registro)

        payload["agendas"][consultor] = {
            "total": total,
            "score_medio": score_medio,
            "territorio": CONSULTORES.get(consultor, {}).get("territorio", ""),
            "por_prioridade": por_prioridade,
            "atendimentos": atendimentos,
        }

    output_path_str = str(output_path.resolve())
    with open(output_path_str, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2, default=_json_serializer)

    logger.info("Agenda JSON salva em: %s", output_path_str)
    return output_path_str


def _json_serializer(obj: Any) -> Any:
    """Serializa tipos nao-nativos do JSON (datetime, numpy, NaT, etc.)."""
    import numpy as np  # noqa: PLC0415

    if isinstance(obj, datetime):
        return obj.isoformat()
    if isinstance(obj, pd.Timestamp):
        return None if pd.isna(obj) else obj.isoformat()
    if isinstance(obj, np.integer):
        return int(obj)
    if isinstance(obj, np.floating):
        return None if np.isnan(obj) else float(obj)
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, float) and (pd.isna(obj)):
        return None
    raise TypeError(f"Tipo nao serializavel: {type(obj).__name__} = {obj!r}")


# ---------------------------------------------------------------------------
# Funcao 7: resumo_agendas
# ---------------------------------------------------------------------------

def resumo_agendas(agendas: dict[str, pd.DataFrame]) -> dict[str, Any]:
    """Calcula estatisticas resumidas das agendas geradas.

    Retorna metricas por consultor e totais globais para log e validacao.

    Args:
        agendas: Dict {consultor: DataFrame_agenda}.

    Returns:
        Dict com:
          - consultores: {nome: {total, por_prioridade, score_stats}}
          - global: {total_atendimentos, total_consultores, score_medio_geral}
    """
    resumo: dict[str, Any] = {
        "consultores": {},
        "global": {
            "total_atendimentos": 0,
            "total_consultores": len(agendas),
            "score_medio_geral": 0.0,
        },
    }

    todos_scores: list[float] = []

    for consultor, agenda in agendas.items():
        total = len(agenda)
        scores: list[float] = []
        por_prioridade: dict[str, int] = {
            f"P{i}": 0 for i in range(8)
        }

        if not agenda.empty:
            if "score" in agenda.columns:
                scores = [
                    float(s) for s in agenda["score"].dropna().tolist()
                ]
                todos_scores.extend(scores)
            if "prioridade" in agenda.columns:
                contagem = agenda["prioridade"].value_counts().to_dict()
                for prio, cnt in contagem.items():
                    por_prioridade[prio] = int(cnt)

        score_stats: dict[str, float] = {}
        if scores:
            score_stats = {
                "media": round(sum(scores) / len(scores), 1),
                "minimo": round(min(scores), 1),
                "maximo": round(max(scores), 1),
            }

        resumo["consultores"][consultor] = {
            "total": total,
            "territorio": CONSULTORES.get(consultor, {}).get("territorio", ""),
            "por_prioridade": por_prioridade,
            "score_stats": score_stats,
        }
        resumo["global"]["total_atendimentos"] += total

    if todos_scores:
        resumo["global"]["score_medio_geral"] = round(
            sum(todos_scores) / len(todos_scores), 1
        )

    return resumo


# ---------------------------------------------------------------------------
# CLI — execucao direta
# ---------------------------------------------------------------------------

def _imprimir_resumo(resumo: dict[str, Any]) -> None:
    """Imprime resumo formatado das agendas no terminal."""
    sep = "=" * 65
    print()
    print(sep)
    print("  AGENDA ENGINE CRM VITAO360 — RESUMO")
    print(sep)
    g = resumo["global"]
    print(f"  Consultores     : {g['total_consultores']}")
    print(f"  Total agendados : {g['total_atendimentos']}")
    print(f"  Score medio     : {g['score_medio_geral']}")
    print()

    for consultor, dados in resumo["consultores"].items():
        print(f"  {consultor} ({dados['territorio']})")
        print(f"    Total: {dados['total']}")
        prios = dados["por_prioridade"]
        prio_str = " | ".join(
            f"{p}:{prios.get(p, 0)}"
            for p in ["P0", "P1", "P2", "P3", "P4", "P5", "P6"]
            if prios.get(p, 0) > 0
        )
        if prio_str:
            print(f"    Por prioridade: {prio_str}")
        sc = dados.get("score_stats", {})
        if sc:
            print(f"    Score: min={sc.get('minimo')} | med={sc.get('media')} | max={sc.get('maximo')}")
        print()
    print(sep)


def _carregar_motor_output() -> pd.DataFrame | None:
    """Carrega motor_output.json e retorna DataFrame ou None se vazio."""
    if not _MOTOR_OUTPUT.exists():
        logger.error("motor_output.json nao encontrado em: %s", _MOTOR_OUTPUT)
        logger.error(
            "Execute primeiro: python -m scripts.motor.motor_regras"
        )
        return None

    with _MOTOR_OUTPUT.open(encoding="utf-8") as fh:
        dados = json.load(fh)

    if isinstance(dados, list):
        clientes = dados
    elif isinstance(dados, dict):
        clientes = dados.get("clientes", dados.get("data", []))
    else:
        logger.error("Formato inesperado em motor_output.json: %s", type(dados))
        return None

    if not clientes:
        logger.warning("motor_output.json esta vazio — nenhum cliente encontrado.")
        return None

    df = pd.DataFrame(clientes)
    logger.info("motor_output.json carregado: %d registros", len(df))
    return df


def main(argv: list[str] | None = None) -> int:
    """Ponto de entrada CLI do Agenda Engine.

    Fluxo:
        1. Carrega motor_output.json
        2. Calcula score_batch (score + prioridade)
        3. Gera agendas para todos os consultores
        4. Imprime agendas formatadas no console
        5. Exporta Excel em data/output/agenda/agenda_YYYY_MM_DD.xlsx
        6. Exporta JSON em data/output/agenda/agenda_YYYY_MM_DD.json
        7. Imprime resumo estatistico

    Returns:
        0 em sucesso, 1 em falha.
    """
    parser = argparse.ArgumentParser(
        description="Agenda Engine — CRM VITAO360",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--data",
        default=None,
        help="Data no formato DD/MM/YYYY (default: hoje)",
    )
    parser.add_argument(
        "--meta",
        type=float,
        default=META_MENSAL_DEFAULT,
        help=f"Meta mensal em R$ (default: {META_MENSAL_DEFAULT:,.0f})",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Nao exportar Excel (apenas JSON e console)",
    )
    parser.add_argument(
        "--no-json",
        action="store_true",
        help="Nao exportar JSON (apenas Excel e console)",
    )
    args = parser.parse_args(argv)

    data_str = args.data or datetime.now().strftime("%d/%m/%Y")
    data_slug = datetime.now().strftime("%Y_%m_%d")

    print()
    print("=" * 65)
    print("  AGENDA ENGINE — CRM VITAO360")
    print(f"  Data: {data_str} | Meta mensal: R$ {args.meta:,.0f}")
    print("=" * 65)

    # 1. Carregar dados do motor
    df = _carregar_motor_output()
    if df is None:
        return 1

    # 2. Calcular score e prioridade
    try:
        from scripts.motor.score_engine import calcular_score_batch

        # Renomear colunas motor_* -> nomes esperados pelo score_engine se necessario
        df = _normalizar_colunas_motor(df)
        df = calcular_score_batch(df)
        logger.info("Score calculado para %d clientes.", len(df))
    except Exception as exc:  # noqa: BLE001
        logger.error("Erro ao calcular score: %s", exc)
        return 1

    # 3. Gerar agendas
    agendas = gerar_agenda(df, data=data_str, meta_mensal=args.meta)

    # 4. Imprimir agendas no console
    print()
    for consultor, agenda in agendas.items():
        texto = formatar_agenda_texto(agenda, consultor, data_str)
        print(texto)

    # 5. Exportar Excel
    _AGENDA_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    excel_path = _AGENDA_OUTPUT_DIR / f"agenda_{data_slug}.xlsx"
    json_path = _AGENDA_OUTPUT_DIR / f"agenda_{data_slug}.json"

    if not args.no_excel:
        try:
            caminho_excel = exportar_agenda_excel(agendas, str(excel_path), data_str)
            print(f"Excel exportado: {caminho_excel}")
        except Exception as exc:  # noqa: BLE001
            logger.error("Falha ao exportar Excel: %s", exc)

    # 6. Exportar JSON
    if not args.no_json:
        try:
            caminho_json = exportar_agenda_json(agendas, str(json_path))
            print(f"JSON exportado:  {caminho_json}")
        except Exception as exc:  # noqa: BLE001
            logger.error("Falha ao exportar JSON: %s", exc)

    # 7. Resumo estatistico
    resumo = resumo_agendas(agendas)
    _imprimir_resumo(resumo)

    return 0


def _normalizar_colunas_motor(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza colunas motor_* para o formato esperado pelo score_engine.

    O motor_regras.py usa prefixo 'motor_' nos campos de output (ex:
    'motor_fase', 'motor_temperatura'). O score_engine espera sem prefixo
    ('fase', 'temperatura'). Esta funcao garante compatibilidade.

    Args:
        df: DataFrame com possiveis colunas motor_*.

    Returns:
        DataFrame com colunas renomeadas onde necessario.
    """
    MAPA_RENOMEAR: dict[str, str] = {
        "motor_fase": "fase",
        "motor_temperatura": "temperatura",
        "motor_tipo_contato": "tipo_contato",
        "motor_acao_futura": "acao_futura",
        "motor_followup_dias": "followup_dias",
        "motor_grupo_dash": "grupo_dash",
        "motor_tipo_acao": "tipo_acao",
        "motor_estagio_funil": "estagio_funil",
    }

    renomear = {
        src: dst
        for src, dst in MAPA_RENOMEAR.items()
        if src in df.columns and dst not in df.columns
    }

    if renomear:
        df = df.rename(columns=renomear)
        logger.info("Colunas normalizadas: %s", list(renomear.keys()))

    # Adicionar colunas de score obrigatorias com defaults se ausentes
    colunas_default: dict[str, Any] = {
        "fase": "PROSPECCAO",
        "sinaleiro": "ROXO",
        "curva_abc": "C",
        "temperatura": "FRIO",
        "tipo_cliente": "PROSPECT",
        "tentativas": "T1",
    }
    for col, default in colunas_default.items():
        if col not in df.columns:
            logger.warning("Coluna '%s' ausente — usando default '%s'", col, default)
            df[col] = default

    return df


if __name__ == "__main__":
    sys.exit(main())

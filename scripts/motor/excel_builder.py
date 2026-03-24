"""
Excel Builder — CRM VITAO360 Motor Operacional v1.0.

Gera o arquivo xlsx final com motor intelligence aplicada.
Recebe o DataFrame enriquecido (apos todo o pipeline) e produz:
  - CARTEIRA       : dados completos com formatacao condicional
  - MOTOR DE REGRAS: referencia das 92 combinacoes
  - DASHBOARD      : KPIs resumidos por dimensao
  - AGENDA_*       : agenda diaria por consultor (max 40 por dia)
  - SCORE CONFIG   : referencia das dimensoes e pesos

Regras inviolaveis aplicadas:
  - R5  : CNPJ sempre string 14 digitos
  - R7  : Formulas Excel em INGLES (IF, VLOOKUP, SUMIF, COUNTIF)
  - R9  : Visual LIGHT exclusivamente (NUNCA dark mode)
  - R10 : Validacao pos-build obrigatoria
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_ROOT = Path(__file__).resolve().parents[2]  # CRM-VITAO360/
_INTELLIGENCE_DIR = _ROOT / "data" / "intelligence"
_OUTPUT_DIR = _ROOT / "data" / "output" / "motor"

MOTOR_REGRAS_JSON = _INTELLIGENCE_DIR / "motor_regras_v4.json"
ARQUITETURA_JSON = _INTELLIGENCE_DIR / "arquitetura_9_dimensoes.json"

DEFAULT_OUTPUT = _OUTPUT_DIR / "CRM_VITAO360_MOTOR_v1.xlsx"
DEFAULT_INPUT_MOTOR = _OUTPUT_DIR / "motor_output.json"
DEFAULT_INPUT_BASE = _OUTPUT_DIR / "base_unificada.json"

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("motor.excel_builder")

# ---------------------------------------------------------------------------
# Paleta de cores (tema LIGHT — R9)
# ---------------------------------------------------------------------------

# Cabecalhos de grupo (linha 1)
COR_GRUPO_IDENTIDADE = "D6E4F7"  # azul claro
COR_GRUPO_EQUIPE = "E2EFDA"      # verde claro
COR_GRUPO_STATUS = "FFF2CC"      # amarelo claro
COR_GRUPO_MOTOR = "FCE4D6"       # laranja claro
COR_GRUPO_SCORE = "EAD1DC"       # rosa claro
COR_GRUPO_SINALEIRO = "D9EAD3"   # verde menta
COR_GRUPO_VENDAS = "CFE2F3"      # azul medio

# Header principal (linha 2)
COR_HEADER_BG = "0D1B3E"         # azul marinho
COR_HEADER_FG = "FFFFFF"

# Cores de status SITUACAO
COR_SITUACAO: dict[str, str] = {
    "ATIVO":     "00B050",
    "EM RISCO":  "FFC000",
    "INAT.REC":  "FFC000",
    "INAT.ANT":  "FF0000",
    "PROSPECT":  "7030A0",
    "LEAD":      "7030A0",
    "NOVO":      "00B050",
}

# Cores de SINALEIRO
COR_SINALEIRO: dict[str, str] = {
    "VERDE":     "00B050",
    "AMARELO":   "FFC000",
    "VERMELHO":  "FF0000",
    "ROXO":      "7030A0",
}

# Cores de TEMPERATURA
COR_TEMPERATURA: dict[str, str] = {
    "QUENTE":   "FF6600",
    "MORNO":    "FFC000",
    "FRIO":     "4472C4",
    "CRITICO":  "FF0000",
    "CRÍTICO":  "FF0000",
    "PERDIDO":  "808080",
}

# Cores de PRIORIDADE
COR_PRIORIDADE: dict[str, str] = {
    "P0": "FF0000",
    "P1": "FF6600",
    "P2": "FFC000",
    "P3": "FFFF00",
    "P4": "92D050",
    "P5": "00B050",
    "P6": "4472C4",
    "P7": "808080",
}

# Cores de CURVA ABC
COR_CURVA_ABC: dict[str, str] = {
    "A": "00B050",
    "B": "FFC000",
    "C": "FF0000",
}

# ---------------------------------------------------------------------------
# Definicao dos grupos da aba CARTEIRA
# ---------------------------------------------------------------------------
GRUPOS_CARTEIRA: list[dict[str, Any]] = [
    {
        "nome": "IDENTIDADE",
        "cor": COR_GRUPO_IDENTIDADE,
        "colunas": [
            ("nome_fantasia",       "NOME FANTASIA",       22),
            ("cnpj_normalizado",    "CNPJ",                17),
            ("razao_social",        "RAZAO SOCIAL",        24),
            ("uf",                  "UF",                   5),
            ("cidade",              "CIDADE",              16),
        ],
    },
    {
        "nome": "EQUIPE",
        "cor": COR_GRUPO_EQUIPE,
        "colunas": [
            ("consultor_normalizado", "CONSULTOR",         12),
        ],
    },
    {
        "nome": "STATUS",
        "cor": COR_GRUPO_STATUS,
        "colunas": [
            ("situacao",            "SITUACAO",            12),
            ("dias_sem_compra",     "DIAS S/ COMPRA",      14),
            ("ciclo_medio",         "CICLO MEDIO",         12),
        ],
    },
    {
        "nome": "MOTOR",
        "cor": COR_GRUPO_MOTOR,
        "colunas": [
            ("estagio_funil",       "ESTAGIO FUNIL",       18),
            ("fase",                "FASE",                14),
            ("temperatura",         "TEMPERATURA",         13),
            ("acao_futura",         "ACAO FUTURA",         28),
            ("followup_dias",       "FOLLOW-UP DIAS",      14),
            ("tipo_contato",        "TIPO CONTATO",        18),
            ("tipo_acao",           "TIPO ACAO",           16),
            ("resultado",           "RESULTADO",           18),
        ],
    },
    {
        "nome": "SCORE",
        "cor": COR_GRUPO_SCORE,
        "colunas": [
            ("score",               "SCORE",                8),
            ("prioridade_v2",       "PRIORIDADE",           11),
        ],
    },
    {
        "nome": "SINALEIRO",
        "cor": COR_GRUPO_SINALEIRO,
        "colunas": [
            ("sinaleiro",           "SINALEIRO",           12),
            ("curva_abc",           "CURVA ABC",           10),
            ("tipo_cliente",        "TIPO CLIENTE",        14),
            ("tentativa",           "TENTATIVA",           11),
        ],
    },
    {
        "nome": "VENDAS",
        "cor": COR_GRUPO_VENDAS,
        "colunas": [
            ("faturamento_total",   "FATURAMENTO TOTAL",   18),
            ("n_compras",           "N COMPRAS",           11),
            ("valor_ultimo_pedido", "ULTIMO PEDIDO",       14),
            ("classificacao_3tier", "CLASSIFICACAO",       15),
        ],
    },
]

# Mapeamento coluna df -> lista de grupos com codigo de cor para formatacao condicional
_COR_POR_CAMPO: dict[str, dict[str, str]] = {
    "situacao":     COR_SITUACAO,
    "sinaleiro":    COR_SINALEIRO,
    "temperatura":  COR_TEMPERATURA,
    "prioridade_v2": COR_PRIORIDADE,
    "curva_abc":    COR_CURVA_ABC,
}

# Limites de agenda por consultor
MAX_AGENDA_ROWS = 40

# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    """Cria PatternFill solid a partir de hex string (sem #)."""
    return PatternFill(fill_type="solid", fgColor=hex_color.upper())


def _font(
    bold: bool = False,
    size: int = 9,
    color: str = "000000",
    name: str = "Arial",
) -> Font:
    """Cria Font padrao do projeto (Arial, tema LIGHT)."""
    return Font(name=name, size=size, bold=bold, color=color)


def _borda_fina() -> Border:
    """Retorna borda fina em todos os lados."""
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _centralizado(wrap: bool = False) -> Alignment:
    """Alinhamento centralizado horizontal e vertical."""
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _esquerdo(wrap: bool = True) -> Alignment:
    """Alinhamento a esquerda com wrap opcional."""
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ---------------------------------------------------------------------------
# Funcao auxiliar de formatacao condicional por valor
# ---------------------------------------------------------------------------

def aplicar_formatacao_condicional(
    ws: Worksheet,
    col_letter: str,
    regras: dict[str, str],
    data_row_start: int,
    data_row_end: int,
) -> None:
    """Aplica formatacao de fundo por valor em uma coluna.

    Percorre as linhas de dados e pinta o fundo da celula de acordo com
    o dicionario valor -> cor_hex. Valores nao encontrados no dicionario
    sao ignorados (sem cor de fundo).

    Args:
        ws: Worksheet openpyxl onde aplicar a formatacao.
        col_letter: Letra da coluna (ex: 'C', 'AB').
        regras: Mapeamento valor_celula (UPPER) -> hex_color (sem #).
        data_row_start: Primeira linha de dados (1-indexed).
        data_row_end: Ultima linha de dados (inclusive, 1-indexed).
    """
    for row_idx in range(data_row_start, data_row_end + 1):
        cell = ws[f"{col_letter}{row_idx}"]
        valor_raw = cell.value
        if valor_raw is None:
            continue
        valor_norm = str(valor_raw).strip().upper()
        cor = regras.get(valor_norm)
        if cor:
            cell.fill = _fill(cor)
            # Texto escuro para fundos claros, branco para escuros
            cell.font = _font(
                color="FFFFFF" if _cor_escura(cor) else "000000",
                size=9,
            )


def _cor_escura(hex_color: str) -> bool:
    """Retorna True se a cor hex e escura (luminancia < 128).

    Usado para decidir cor do texto (branco vs preto).
    """
    hex_clean = hex_color.lstrip("#").upper()
    if len(hex_clean) != 6:
        return False
    r = int(hex_clean[0:2], 16)
    g = int(hex_clean[2:4], 16)
    b = int(hex_clean[4:6], 16)
    luminancia = 0.299 * r + 0.587 * g + 0.114 * b
    return luminancia < 128


# ---------------------------------------------------------------------------
# Aba CARTEIRA
# ---------------------------------------------------------------------------

def _escrever_carteira(ws: Worksheet, df: pd.DataFrame) -> int:
    """Escreve a aba CARTEIRA com grupos, headers e dados formatados.

    Estrutura de linhas:
      Linha 1: headers de grupo (merged cells)
      Linha 2: headers de coluna (bold, navy background)
      Linhas 3+: dados

    Args:
        ws: Worksheet destino.
        df: DataFrame com todos os clientes enriquecidos.

    Returns:
        Numero de linhas de dados escritas.
    """
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "C3"  # Congelar ate linha 2 e coluna B

    # Construir lista de colunas efetivas (apenas as que existem no df)
    colunas_efetivas: list[tuple[str, str, str, int]] = []  # (campo, label, grupo, width)
    for grupo in GRUPOS_CARTEIRA:
        for campo, label, width in grupo["colunas"]:
            colunas_efetivas.append((campo, label, grupo["nome"], width))

    n_cols = len(colunas_efetivas)

    # --- Linha 1: headers de grupo (merged) ---
    col_inicio_grupo: dict[str, int] = {}
    col_fim_grupo: dict[str, int] = {}
    for col_idx, (campo, label, grupo_nome, width) in enumerate(colunas_efetivas, start=1):
        if grupo_nome not in col_inicio_grupo:
            col_inicio_grupo[grupo_nome] = col_idx
        col_fim_grupo[grupo_nome] = col_idx

    for grupo in GRUPOS_CARTEIRA:
        nome = grupo["nome"]
        if nome not in col_inicio_grupo:
            continue
        c_ini = col_inicio_grupo[nome]
        c_fim = col_fim_grupo[nome]
        letra_ini = get_column_letter(c_ini)
        letra_fim = get_column_letter(c_fim)
        if c_ini == c_fim:
            celula = ws[f"{letra_ini}1"]
        else:
            ws.merge_cells(f"{letra_ini}1:{letra_fim}1")
            celula = ws[f"{letra_ini}1"]
        celula.value = nome
        celula.fill = _fill(grupo["cor"])
        celula.font = _font(bold=True, size=9, color="000000")
        celula.alignment = _centralizado()
        celula.border = _borda_fina()

    ws.row_dimensions[1].height = 18

    # --- Linha 2: headers de coluna ---
    for col_idx, (campo, label, grupo_nome, width) in enumerate(colunas_efetivas, start=1):
        letra = get_column_letter(col_idx)
        celula = ws[f"{letra}2"]
        celula.value = label
        celula.fill = _fill(COR_HEADER_BG)
        celula.font = _font(bold=True, size=10, color=COR_HEADER_FG)
        celula.alignment = _centralizado(wrap=True)
        celula.border = _borda_fina()
        ws.column_dimensions[letra].width = max(8, min(width, 40))

    ws.row_dimensions[2].height = 24

    # --- Linhas 3+: dados ---
    campos = [t[0] for t in colunas_efetivas]
    n_registros = 0

    for row_num, (_, row) in enumerate(df.iterrows(), start=3):
        for col_idx, campo in enumerate(campos, start=1):
            letra = get_column_letter(col_idx)
            celula = ws[f"{letra}{row_num}"]

            valor = row.get(campo)
            # Limpar erros de formula Excel (#NAME?, #REF!, etc.)
            if isinstance(valor, str) and valor.startswith("#"):
                valor = None

            # CNPJ: sempre string 14 digitos (R5)
            if campo == "cnpj_normalizado" and valor is not None:
                valor = _normalizar_cnpj(str(valor))

            # Valores numericos: arredondar para 2 casas
            if campo == "faturamento_total" and valor is not None:
                try:
                    valor = round(float(valor), 2)
                except (TypeError, ValueError):
                    valor = None

            celula.value = valor
            celula.font = _font(size=9)
            celula.alignment = _esquerdo(wrap=False)
            celula.border = _borda_fina()

            # Alternancia de linha (zebra)
            if (row_num - 3) % 2 == 1:
                celula.fill = _fill("F5F5F5")

        n_registros += 1

    # --- Formatacao condicional por coluna ---
    if n_registros > 0:
        for col_idx, (campo, label, grupo_nome, width) in enumerate(colunas_efetivas, start=1):
            if campo in _COR_POR_CAMPO:
                letra = get_column_letter(col_idx)
                aplicar_formatacao_condicional(
                    ws=ws,
                    col_letter=letra,
                    regras=_COR_POR_CAMPO[campo],
                    data_row_start=3,
                    data_row_end=2 + n_registros,
                )

    # --- Auto-filtro na linha de header ---
    letra_ultima = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A2:{letra_ultima}2"

    logger.info("CARTEIRA: %d linhas escritas, %d colunas", n_registros, n_cols)
    return n_registros


# ---------------------------------------------------------------------------
# Aba MOTOR DE REGRAS
# ---------------------------------------------------------------------------

def _escrever_motor_regras(ws: Worksheet) -> int:
    """Escreve a aba MOTOR DE REGRAS com as 92 combinacoes.

    Args:
        ws: Worksheet destino.

    Returns:
        Numero de linhas escritas (0 se JSON nao encontrado).
    """
    if not MOTOR_REGRAS_JSON.exists():
        ws["A1"].value = "motor_regras_v4.json nao encontrado"
        logger.warning("motor_regras_v4.json nao encontrado: %s", MOTOR_REGRAS_JSON)
        return 0

    with MOTOR_REGRAS_JSON.open(encoding="utf-8") as fh:
        dados = json.load(fh)

    combinacoes = dados.get("combinacoes", [])

    headers = [
        ("#", 4),
        ("SITUACAO", 12),
        ("RESULTADO", 22),
        ("ESTAGIO", 18),
        ("FASE", 16),
        ("TEMPERATURA", 13),
        ("FOLLOW-UP DIAS", 14),
        ("ACAO FUTURA", 30),
        ("TIPO CONTATO", 20),
        ("TIPO ACAO", 18),
        ("GRUPO DASH", 12),
    ]

    # Cabecalho
    for col_idx, (label, width) in enumerate(headers, start=1):
        letra = get_column_letter(col_idx)
        celula = ws[f"{letra}1"]
        celula.value = label
        celula.fill = _fill(COR_HEADER_BG)
        celula.font = _font(bold=True, size=10, color=COR_HEADER_FG)
        celula.alignment = _centralizado(wrap=True)
        celula.border = _borda_fina()
        ws.column_dimensions[letra].width = width

    ws.row_dimensions[1].height = 24
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    # Dados
    for row_num, combo in enumerate(combinacoes, start=2):
        valores = [
            combo.get("numero"),
            combo.get("situacao", ""),
            combo.get("resultado", ""),
            combo.get("estagio_funil", ""),
            combo.get("fase", ""),
            combo.get("temperatura", ""),
            combo.get("followup_dias"),
            combo.get("acao_futura", ""),
            combo.get("tipo_contato", ""),
            combo.get("tipo_acao", ""),
            combo.get("grupo_dash", ""),
        ]
        for col_idx, valor in enumerate(valores, start=1):
            letra = get_column_letter(col_idx)
            celula = ws[f"{letra}{row_num}"]
            celula.value = valor
            celula.font = _font(size=9)
            celula.alignment = _esquerdo(wrap=False)
            celula.border = _borda_fina()
            if (row_num - 2) % 2 == 1:
                celula.fill = _fill("F5F5F5")

        # Cor na coluna TEMPERATURA (col 6)
        temp_letra = get_column_letter(6)
        temp_cell = ws[f"{temp_letra}{row_num}"]
        temp_val = str(temp_cell.value or "").strip().upper()
        cor_temp = COR_TEMPERATURA.get(temp_val)
        if cor_temp:
            temp_cell.fill = _fill(cor_temp)
            temp_cell.font = _font(
                color="FFFFFF" if _cor_escura(cor_temp) else "000000",
                size=9,
            )

        # Cor na coluna SITUACAO (col 2)
        sit_letra = get_column_letter(2)
        sit_cell = ws[f"{sit_letra}{row_num}"]
        sit_val = str(sit_cell.value or "").strip().upper()
        cor_sit = COR_SITUACAO.get(sit_val)
        if cor_sit:
            sit_cell.fill = _fill(cor_sit)
            sit_cell.font = _font(
                color="FFFFFF" if _cor_escura(cor_sit) else "000000",
                size=9,
            )

    logger.info("MOTOR DE REGRAS: %d combinacoes escritas", len(combinacoes))
    return len(combinacoes)


# ---------------------------------------------------------------------------
# Aba DASHBOARD
# ---------------------------------------------------------------------------

def _escrever_dashboard(ws: Worksheet, df: pd.DataFrame) -> None:
    """Escreve a aba DASHBOARD com KPIs resumidos.

    Args:
        ws: Worksheet destino.
        df: DataFrame com todos os clientes.
    """
    ws.sheet_view.showGridLines = False

    titulo_style = {
        "font": _font(bold=True, size=14, color="0D1B3E"),
        "alignment": _esquerdo(),
    }
    section_style = {
        "font": _font(bold=True, size=10, color="FFFFFF"),
        "fill": _fill(COR_HEADER_BG),
        "alignment": _esquerdo(),
        "border": _borda_fina(),
    }
    label_style = {
        "font": _font(size=9),
        "alignment": _esquerdo(),
        "border": _borda_fina(),
    }
    value_style = {
        "font": _font(bold=True, size=9),
        "alignment": _centralizado(),
        "border": _borda_fina(),
    }
    pct_style = {
        "font": _font(size=9, color="666666"),
        "alignment": _centralizado(),
        "border": _borda_fina(),
    }

    def _aplicar(celula, estilos: dict) -> None:
        for attr, val in estilos.items():
            setattr(celula, attr, val)

    total = len(df)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 2
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10

    # Titulo
    celula = ws["B1"]
    celula.value = "CRM VITAO360 — DASHBOARD"
    _aplicar(celula, titulo_style)
    ws.row_dimensions[1].height = 28

    celula = ws["B2"]
    celula.value = f"Total clientes: {total:,}"
    celula.font = _font(bold=True, size=11, color="0D1B3E")

    row = 4

    # --- Bloco esquerdo: SITUACAO, SINALEIRO, CURVA ABC ---
    secoes_esquerdo = [
        ("situacao",     "SITUACAO",     COR_SITUACAO),
        ("sinaleiro",    "SINALEIRO",    COR_SINALEIRO),
        ("curva_abc",    "CURVA ABC",    COR_CURVA_ABC),
    ]

    for campo, titulo_secao, mapa_cor in secoes_esquerdo:
        # Header da secao
        celula = ws[f"B{row}"]
        celula.value = titulo_secao
        _aplicar(celula, section_style)
        celula = ws[f"C{row}"]
        celula.value = "QTDE"
        _aplicar(celula, section_style)
        celula = ws[f"D{row}"]
        celula.value = "%"
        _aplicar(celula, section_style)
        ws.row_dimensions[row].height = 18
        row += 1

        # Contagem
        col_data = df.get(campo, pd.Series(dtype=str))
        contagem = (
            col_data.fillna("NAO INFORMADO")
            .astype(str)
            .str.strip()
            .str.upper()
            .value_counts()
            .sort_index()
        )

        for valor, cnt in contagem.items():
            pct = (cnt / total * 100) if total > 0 else 0.0
            celula_b = ws[f"B{row}"]
            celula_b.value = str(valor)
            _aplicar(celula_b, label_style)
            cor = mapa_cor.get(str(valor), None)
            if cor:
                celula_b.fill = _fill(cor)
                celula_b.font = _font(
                    color="FFFFFF" if _cor_escura(cor) else "000000",
                    size=9,
                )

            celula_c = ws[f"C{row}"]
            celula_c.value = int(cnt)
            _aplicar(celula_c, value_style)

            celula_d = ws[f"D{row}"]
            celula_d.value = f"{pct:.1f}%"
            _aplicar(celula_d, pct_style)

            ws.row_dimensions[row].height = 16
            row += 1

        row += 1  # espaco entre secoes

    # --- Bloco direito: TEMPERATURA, PRIORIDADE, CONSULTOR ---
    row_direito = 4

    secoes_direito = [
        ("temperatura",  "TEMPERATURA",  COR_TEMPERATURA),
        ("prioridade_v2","PRIORIDADE",   COR_PRIORIDADE),
        ("consultor_normalizado", "CONSULTOR", {}),
    ]

    for campo, titulo_secao, mapa_cor in secoes_direito:
        # Header da secao
        celula = ws[f"F{row_direito}"]
        celula.value = titulo_secao
        _aplicar(celula, section_style)
        celula = ws[f"G{row_direito}"]
        celula.value = "QTDE"
        _aplicar(celula, section_style)
        celula = ws[f"H{row_direito}"]
        celula.value = "%"
        _aplicar(celula, section_style)
        ws.row_dimensions[row_direito].height = 18
        row_direito += 1

        col_data = df.get(campo, pd.Series(dtype=str))
        contagem = (
            col_data.fillna("NAO INFORMADO")
            .astype(str)
            .str.strip()
            .str.upper()
            .replace({"#NAME?": "NAO INFORMADO", "NAN": "NAO INFORMADO"})
            .value_counts()
            .sort_index()
        )

        for valor, cnt in contagem.items():
            pct = (cnt / total * 100) if total > 0 else 0.0
            celula_f = ws[f"F{row_direito}"]
            celula_f.value = str(valor)
            _aplicar(celula_f, label_style)
            cor = mapa_cor.get(str(valor), None)
            if cor:
                celula_f.fill = _fill(cor)
                celula_f.font = _font(
                    color="FFFFFF" if _cor_escura(cor) else "000000",
                    size=9,
                )

            celula_g = ws[f"G{row_direito}"]
            celula_g.value = int(cnt)
            _aplicar(celula_g, value_style)

            celula_h = ws[f"H{row_direito}"]
            celula_h.value = f"{pct:.1f}%"
            _aplicar(celula_h, pct_style)

            ws.row_dimensions[row_direito].height = 16
            row_direito += 1

        row_direito += 1  # espaco entre secoes

    # --- Faturamento total (se disponivel) ---
    fat_col = df.get("faturamento_total", pd.Series(dtype=float))
    fat_valido = pd.to_numeric(fat_col, errors="coerce").dropna()
    fat_total = fat_valido.sum()
    if fat_total > 0:
        row_fat = max(row, row_direito) + 2
        celula = ws[f"B{row_fat}"]
        celula.value = "FATURAMENTO TOTAL"
        celula.fill = _fill(COR_HEADER_BG)
        celula.font = _font(bold=True, size=10, color="FFFFFF")
        celula.alignment = _esquerdo()
        celula.border = _borda_fina()

        celula_v = ws[f"C{row_fat}"]
        celula_v.value = round(fat_total, 2)
        celula_v.font = _font(bold=True, size=9, color="0D6E0D")
        celula_v.alignment = _centralizado()
        celula_v.number_format = 'R$ #,##0.00'
        celula_v.border = _borda_fina()

        celula_label = ws[f"D{row_fat}"]
        celula_label.value = f"({len(fat_valido)} clientes c/ faturamento)"
        celula_label.font = _font(size=9, color="666666")
        celula_label.alignment = _esquerdo()

    logger.info("DASHBOARD: escrito com %d clientes", total)


# ---------------------------------------------------------------------------
# Abas AGENDA por consultor
# ---------------------------------------------------------------------------

def _escrever_agenda(
    ws: Worksheet,
    nome_consultor: str,
    df_agenda: pd.DataFrame,
) -> int:
    """Escreve a aba de agenda para um consultor.

    Args:
        ws: Worksheet destino.
        nome_consultor: Nome do consultor para o titulo.
        df_agenda: DataFrame com clientes da agenda (max MAX_AGENDA_ROWS linhas).

    Returns:
        Numero de clientes escritos.
    """
    headers_agenda = [
        ("prioridade_v2",    "PRIORIDADE",   11),
        ("nome_fantasia",    "CLIENTE",       26),
        ("cnpj_normalizado", "CNPJ",          17),
        ("situacao",         "SITUACAO",      12),
        ("temperatura",      "TEMPERATURA",   13),
        ("acao_futura",      "ACAO FUTURA",   28),
        ("followup_dias",    "FU DIAS",        9),
        ("score",            "SCORE",          8),
        ("estagio_funil",    "ESTAGIO",       18),
        ("consultor_normalizado", "CONSULTOR", 12),
    ]

    # Titulo
    ws["A1"].value = f"AGENDA DIARIA — {nome_consultor}"
    ws["A1"].font = _font(bold=True, size=12, color="0D1B3E")
    ws["A1"].alignment = _esquerdo()
    ws.row_dimensions[1].height = 24

    # Header
    for col_idx, (campo, label, width) in enumerate(headers_agenda, start=1):
        letra = get_column_letter(col_idx)
        celula = ws[f"{letra}2"]
        celula.value = label
        celula.fill = _fill(COR_HEADER_BG)
        celula.font = _font(bold=True, size=10, color=COR_HEADER_FG)
        celula.alignment = _centralizado(wrap=True)
        celula.border = _borda_fina()
        ws.column_dimensions[letra].width = max(8, min(width, 40))

    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"
    n_ultima = get_column_letter(len(headers_agenda))
    ws.auto_filter.ref = f"A2:{n_ultima}2"

    # Limitar linhas
    df_limitado = df_agenda.head(MAX_AGENDA_ROWS)
    n = 0

    for row_num, (_, row) in enumerate(df_limitado.iterrows(), start=3):
        for col_idx, (campo, label, width) in enumerate(headers_agenda, start=1):
            letra = get_column_letter(col_idx)
            celula = ws[f"{letra}{row_num}"]

            valor = row.get(campo)
            if isinstance(valor, str) and valor.startswith("#"):
                valor = None
            if campo == "cnpj_normalizado" and valor is not None:
                valor = _normalizar_cnpj(str(valor))

            celula.value = valor
            celula.font = _font(size=9)
            celula.alignment = _esquerdo(wrap=False)
            celula.border = _borda_fina()

        # Cor de prioridade na linha inteira (fundo suave)
        prio_val = str(row.get("prioridade_v2", "") or "").strip().upper()
        if prio_val.startswith("#"):
            prio_val = ""
        cor_prio = COR_PRIORIDADE.get(prio_val)
        if cor_prio:
            # Apenas coluna PRIORIDADE recebe a cor cheia
            letra_prio = get_column_letter(1)
            celula_prio = ws[f"{letra_prio}{row_num}"]
            celula_prio.fill = _fill(cor_prio)
            celula_prio.font = _font(
                bold=True,
                color="FFFFFF" if _cor_escura(cor_prio) else "000000",
                size=9,
            )

        # Cor TEMPERATURA
        temp_val = str(row.get("temperatura", "") or "").strip().upper()
        if not temp_val.startswith("#"):
            cor_temp = COR_TEMPERATURA.get(temp_val)
            if cor_temp:
                letra_temp = get_column_letter(5)
                c = ws[f"{letra_temp}{row_num}"]
                c.fill = _fill(cor_temp)
                c.font = _font(
                    color="FFFFFF" if _cor_escura(cor_temp) else "000000",
                    size=9,
                )

        # Cor SITUACAO
        sit_val = str(row.get("situacao", "") or "").strip().upper()
        cor_sit = COR_SITUACAO.get(sit_val)
        if cor_sit:
            letra_sit = get_column_letter(4)
            c = ws[f"{letra_sit}{row_num}"]
            c.fill = _fill(cor_sit)
            c.font = _font(
                color="FFFFFF" if _cor_escura(cor_sit) else "000000",
                size=9,
            )

        ws.row_dimensions[row_num].height = 16
        n += 1

    logger.info("AGENDA %s: %d clientes escritos (limite %d)", nome_consultor, n, MAX_AGENDA_ROWS)
    return n


# ---------------------------------------------------------------------------
# Aba SCORE CONFIG
# ---------------------------------------------------------------------------

def _escrever_score_config(ws: Worksheet) -> None:
    """Escreve a aba de configuracao de score (dimensoes, pesos, prioridades).

    Args:
        ws: Worksheet destino.
    """
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 4
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 22

    # Titulo
    ws["B1"].value = "CRM VITAO360 — CONFIGURACAO DE SCORE"
    ws["B1"].font = _font(bold=True, size=13, color="0D1B3E")
    ws.row_dimensions[1].height = 26

    if not ARQUITETURA_JSON.exists():
        ws["B3"].value = "arquitetura_9_dimensoes.json nao encontrado"
        logger.warning("arquitetura_9_dimensoes.json nao encontrado")
        return

    with ARQUITETURA_JSON.open(encoding="utf-8") as fh:
        arq = json.load(fh)

    # --- Tabela de dimensoes e pesos (col B-D) ---
    row = 3
    for label, w in [("DIMENSAO", 22), ("PESO", 10), ("LOGICA DE SCORE", 30)]:
        col_idx = ["DIMENSAO", "PESO", "LOGICA DE SCORE"].index(label) + 2
        letra = get_column_letter(col_idx)
        c = ws[f"{letra}{row}"]
        c.value = label
        c.fill = _fill(COR_HEADER_BG)
        c.font = _font(bold=True, size=10, color="FFFFFF")
        c.alignment = _centralizado()
        c.border = _borda_fina()
    ws.row_dimensions[row].height = 20
    row += 1

    for item in arq.get("score_ponderado", []):
        ws[f"B{row}"].value = item.get("dimensao", "")
        ws[f"C{row}"].value = item.get("peso", "")
        ws[f"D{row}"].value = item.get("logica_score", "")
        for col_idx in range(2, 5):
            letra = get_column_letter(col_idx)
            c = ws[f"{letra}{row}"]
            c.font = _font(size=9)
            c.alignment = _esquerdo(wrap=True)
            c.border = _borda_fina()
            if (row - 4) % 2 == 1:
                c.fill = _fill("F5F5F5")
        ws.row_dimensions[row].height = 30
        row += 1

    row += 2

    # --- Tabela de prioridades (col F-H) ---
    row_prio = 3
    for label in ["PRIORIDADE", "NOME", "SCORE / REGRA"]:
        col_idx = ["PRIORIDADE", "NOME", "SCORE / REGRA"].index(label) + 6
        letra = get_column_letter(col_idx)
        c = ws[f"{letra}{row_prio}"]
        c.value = label
        c.fill = _fill(COR_HEADER_BG)
        c.font = _font(bold=True, size=10, color="FFFFFF")
        c.alignment = _centralizado()
        c.border = _borda_fina()
    ws.row_dimensions[row_prio].height = 20
    row_prio += 1

    for prio in arq.get("prioridades", []):
        label_prio = prio.get("label", "")
        nome_prio = prio.get("nome", "")
        regra_prio = prio.get("score", prio.get("como_gera", ""))

        ws[f"F{row_prio}"].value = label_prio
        ws[f"G{row_prio}"].value = nome_prio
        ws[f"H{row_prio}"].value = regra_prio

        cor_p = COR_PRIORIDADE.get(label_prio)
        for letra in ["F", "G", "H"]:
            c = ws[f"{letra}{row_prio}"]
            c.font = _font(size=9)
            c.alignment = _centralizado()
            c.border = _borda_fina()

        if cor_p:
            c_f = ws[f"F{row_prio}"]
            c_f.fill = _fill(cor_p)
            c_f.font = _font(
                bold=True,
                color="FFFFFF" if _cor_escura(cor_p) else "000000",
                size=9,
            )

        ws.row_dimensions[row_prio].height = 18
        row_prio += 1

    # --- Regras de desempate ---
    row_des = row_prio + 2
    ws[f"F{row_des}"].value = "DESEMPATE (mesma prioridade)"
    ws[f"F{row_des}"].font = _font(bold=True, size=10, color="0D1B3E")
    ws.row_dimensions[row_des].height = 18
    row_des += 1

    for regra in arq.get("desempate", []):
        ws[f"F{row_des}"].value = regra
        ws[f"F{row_des}"].font = _font(size=9)
        ws[f"F{row_des}"].alignment = _esquerdo()
        ws[f"F{row_des}"].border = _borda_fina()
        ws.row_dimensions[row_des].height = 16
        row_des += 1

    # Meta balance
    row_meta = row_des + 1
    ws[f"F{row_meta}"].value = arq.get("meta_balance", "")
    ws[f"F{row_meta}"].font = _font(size=9, color="666666")
    ws[f"F{row_meta}"].alignment = _esquerdo(wrap=True)
    ws.row_dimensions[row_meta].height = 30

    logger.info("SCORE CONFIG: aba escrita")


# ---------------------------------------------------------------------------
# Helpers de dados
# ---------------------------------------------------------------------------

def _normalizar_cnpj(valor: str) -> str:
    """Normaliza CNPJ para string de 14 digitos sem pontuacao (R5).

    Args:
        valor: CNPJ em qualquer formato.

    Returns:
        String de 14 digitos, zero-padded.
    """
    import re
    digits = re.sub(r"\D", "", str(valor))
    return digits.zfill(14)[:14]


def _carregar_df_de_json(path: Path) -> pd.DataFrame:
    """Carrega DataFrame a partir de JSON do pipeline.

    Aceita:
      - Lista de dicts (formato motor_output.json)
      - Dict com chave 'registros' ou 'clientes' (formato base_unificada.json)

    Args:
        path: Caminho do arquivo JSON.

    Returns:
        DataFrame com os registros.

    Raises:
        FileNotFoundError: Se o arquivo nao existir.
        ValueError: Se o formato for desconhecido.
    """
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {path}")

    with path.open(encoding="utf-8") as fh:
        dados = json.load(fh)

    if isinstance(dados, list):
        if not dados:
            logger.warning("JSON vazio (lista): %s", path)
            return pd.DataFrame()
        return pd.DataFrame(dados)

    if isinstance(dados, dict):
        for chave in ("registros", "clientes", "data"):
            if chave in dados:
                registros = dados[chave]
                if not registros:
                    logger.warning("JSON vazio (chave '%s'): %s", chave, path)
                    return pd.DataFrame()
                return pd.DataFrame(registros)
        # Tenta usar os valores do dict diretamente
        raise ValueError(
            f"Formato de JSON desconhecido em {path}. "
            f"Chaves encontradas: {list(dados.keys())[:5]}"
        )

    raise ValueError(f"Formato inesperado em {path}: {type(dados)}")


def _derivar_prioridade_por_score(score: float) -> str:
    """Converte score numerico em label P0-P7 conforme faixas do motor.

    Espelha FAIXAS_PRIORIDADE do score_engine.py.
    P0 e P1 sao bloqueios nao calculaveis por score — nao gerados aqui.

    Args:
        score: Score ponderado 0-100.

    Returns:
        Label de prioridade: P2-P7.
    """
    if score >= 80.0:
        return "P2"
    if score >= 60.0:
        return "P3"
    if score >= 45.0:
        return "P4"
    if score >= 30.0:
        return "P5"
    if score >= 15.0:
        return "P6"
    return "P7"


def _preparar_agendas(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Divide o DataFrame por consultor e ordena por prioridade/score.

    Quando prioridade_v2 estiver invalida (ex: '#NAME?' por erro de formula
    Excel), deriva a prioridade automaticamente a partir do score numerico.
    Clientes P7 e sem score sao excluidos da agenda diaria (campanha mensal).

    Args:
        df: DataFrame com todos os clientes.

    Returns:
        Dict {nome_consultor: DataFrame_ordenado} com no maximo MAX_AGENDA_ROWS linhas.
    """
    agendas: dict[str, pd.DataFrame] = {}

    # Converter prioridade em numero para ordenacao
    def _prio_num(p: Any) -> int:
        mapa = {"P0": 0, "P1": 1, "P2": 2, "P3": 3, "P4": 4, "P5": 5, "P6": 6, "P7": 7}
        return mapa.get(str(p or "").strip().upper(), 99)

    col_consultor = "consultor_normalizado"
    if col_consultor not in df.columns:
        logger.warning("Coluna '%s' nao encontrada: agenda nao gerada", col_consultor)
        return agendas

    # Preparar coluna de prioridade efetiva (usa prioridade_v2 se valida, score senao)
    df_work = df.copy()
    prio_col = df_work.get("prioridade_v2", pd.Series([""] * len(df_work), index=df_work.index))
    score_col_num = pd.to_numeric(
        df_work.get("score", pd.Series(dtype=float)), errors="coerce"
    )

    def _prio_efetiva(row_idx: int) -> str:
        prio_raw = str(prio_col.iloc[row_idx] or "").strip().upper()
        if prio_raw in ("P0", "P1", "P2", "P3", "P4", "P5", "P6", "P7"):
            return prio_raw
        # Prioridade invalida: derivar do score
        score_val = score_col_num.iloc[row_idx]
        if pd.isna(score_val):
            return "P7"  # sem score = campanha
        return _derivar_prioridade_por_score(float(score_val))

    df_work["_prio_efetiva"] = [_prio_efetiva(i) for i in range(len(df_work))]
    df_work["_prio_num"] = df_work["_prio_efetiva"].apply(_prio_num)
    df_work["_score_num"] = score_col_num.fillna(0)

    for consultor, grupo in df_work.groupby(col_consultor, dropna=False):
        if not consultor or str(consultor).upper() in ("DESCONHECIDO", "NAN", "NONE"):
            continue

        # Excluir P7 da agenda diaria (campanha mensal) e consultor LEGADO
        if str(consultor).upper() == "LEGADO":
            continue

        mascara_valida = grupo["_prio_efetiva"] != "P7"
        grupo_filtrado = grupo[mascara_valida].copy()

        if grupo_filtrado.empty:
            logger.info("Consultor %s: todos os clientes em P7 (campanha) — agenda vazia", consultor)

        # Copiar prioridade derivada para coluna visivel na agenda
        if "prioridade_v2" in grupo_filtrado.columns:
            # Sobrescrever apenas os #NAME? com o valor derivado
            prio_invalidas = grupo_filtrado["prioridade_v2"].apply(
                lambda p: str(p or "").strip().upper() not in (
                    "P0", "P1", "P2", "P3", "P4", "P5", "P6", "P7"
                )
            )
            grupo_filtrado.loc[prio_invalidas, "prioridade_v2"] = (
                grupo_filtrado.loc[prio_invalidas, "_prio_efetiva"]
            )

        # Ordenar P0 -> P6, depois score decrescente
        grupo_ordenado = grupo_filtrado.sort_values(
            by=["_prio_num", "_score_num"],
            ascending=[True, False],
        ).drop(columns=["_prio_num", "_score_num", "_prio_efetiva"], errors="ignore")

        agendas[str(consultor)] = grupo_ordenado

    return agendas


# ---------------------------------------------------------------------------
# Funcao principal
# ---------------------------------------------------------------------------

def build_crm_excel(
    df: pd.DataFrame,
    output_path: str | Path,
    agendas: dict[str, pd.DataFrame] | None = None,
) -> Path:
    """Gera o xlsx final do CRM VITAO360 com motor intelligence aplicada.

    Cria as abas:
      - CARTEIRA       : dados completos com formatacao condicional
      - MOTOR DE REGRAS: referencia das 92 combinacoes
      - DASHBOARD      : KPIs resumidos
      - AGENDA_*       : uma aba por consultor (max MAX_AGENDA_ROWS linhas)
      - SCORE CONFIG   : referencia de dimensoes, pesos e prioridades

    Args:
        df: DataFrame com todos os clientes enriquecidos pelo pipeline.
            Colunas ausentes sao ignoradas graciosamente.
        output_path: Caminho do arquivo xlsx a ser gerado.
        agendas: Dict opcional {nome_consultor: DataFrame}. Se None,
                 o builder cria automaticamente a partir do df principal.

    Returns:
        Path absoluto do arquivo gerado.

    Raises:
        OSError: Se o diretorio de output nao puder ser criado.
    """
    output_path = Path(output_path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Iniciando build: %s", output_path)
    logger.info("Total de clientes: %d", len(df))

    wb = Workbook()

    # Remover aba padrao criada pelo openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # -----------------------------------------------------------------------
    # Aba 1: CARTEIRA
    # -----------------------------------------------------------------------
    ws_carteira = wb.create_sheet("CARTEIRA")
    n_carteira = _escrever_carteira(ws_carteira, df)

    # -----------------------------------------------------------------------
    # Aba 2: MOTOR DE REGRAS
    # -----------------------------------------------------------------------
    ws_motor = wb.create_sheet("MOTOR DE REGRAS")
    _escrever_motor_regras(ws_motor)

    # -----------------------------------------------------------------------
    # Aba 3: DASHBOARD
    # -----------------------------------------------------------------------
    ws_dash = wb.create_sheet("DASHBOARD")
    _escrever_dashboard(ws_dash, df)

    # -----------------------------------------------------------------------
    # Abas AGENDA por consultor
    # -----------------------------------------------------------------------
    if agendas is None:
        agendas = _preparar_agendas(df)

    consultores_ordenados = sorted(agendas.keys())
    for consultor in consultores_ordenados:
        nome_aba = f"AGENDA_{consultor[:20]}"  # openpyxl limita nome da aba a 31 chars
        ws_ag = wb.create_sheet(nome_aba)
        _escrever_agenda(ws_ag, consultor, agendas[consultor])

    # -----------------------------------------------------------------------
    # Aba SCORE CONFIG
    # -----------------------------------------------------------------------
    ws_score = wb.create_sheet("SCORE CONFIG")
    _escrever_score_config(ws_score)

    # -----------------------------------------------------------------------
    # Salvar
    # -----------------------------------------------------------------------
    wb.save(output_path)

    tamanho_kb = output_path.stat().st_size // 1024
    logger.info(
        "Build concluido: %s | %d KB | %d abas | %d linhas CARTEIRA",
        output_path.name,
        tamanho_kb,
        len(wb.sheetnames),
        n_carteira,
    )

    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _cli() -> int:
    """Ponto de entrada CLI para o excel_builder.

    Uso:
      python scripts/motor/excel_builder.py
      python scripts/motor/excel_builder.py --input path/to/json
      python scripts/motor/excel_builder.py --output path/to/xlsx
    """
    parser = argparse.ArgumentParser(
        description="Excel Builder — CRM VITAO360 Motor Operacional"
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help=(
            f"JSON de entrada (default: motor_output.json se existir, "
            f"senao base_unificada.json)"
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Arquivo xlsx de saida (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()

    # Resolver input
    if args.input is not None:
        input_path = args.input.resolve()
    elif DEFAULT_INPUT_MOTOR.exists():
        # motor_output.json existe mas pode estar vazio (lista vazia)
        input_path = DEFAULT_INPUT_MOTOR
    else:
        input_path = DEFAULT_INPUT_BASE

    logger.info("Carregando dados de: %s", input_path)

    try:
        df = _carregar_df_de_json(input_path)
    except (FileNotFoundError, ValueError) as exc:
        logger.error("Erro ao carregar dados: %s", exc)
        return 1

    if df.empty:
        logger.warning(
            "DataFrame vazio carregado de %s — tentando base_unificada.json como fallback",
            input_path,
        )
        if input_path != DEFAULT_INPUT_BASE and DEFAULT_INPUT_BASE.exists():
            try:
                df = _carregar_df_de_json(DEFAULT_INPUT_BASE)
            except (FileNotFoundError, ValueError) as exc:
                logger.error("Fallback falhou: %s", exc)
                return 1

    if df.empty:
        logger.error("Nenhum dado disponivel. Verifique os arquivos de entrada.")
        return 1

    logger.info("Dados carregados: %d registros, %d colunas", len(df), len(df.columns))

    try:
        output_path = build_crm_excel(df=df, output_path=args.output)
    except Exception as exc:  # noqa: BLE001
        logger.error("Erro durante build: %s", exc, exc_info=True)
        return 1

    # Relatorio final
    tamanho_kb = output_path.stat().st_size // 1024
    abas = []
    try:
        from openpyxl import load_workbook as _lw
        wb_check = _lw(output_path, read_only=True)
        abas = wb_check.sheetnames
        wb_check.close()
    except Exception:  # noqa: BLE001
        pass

    print()
    print("=" * 60)
    print("  CRM VITAO360 — BUILD CONCLUIDO")
    print("=" * 60)
    print(f"  Arquivo  : {output_path}")
    print(f"  Tamanho  : {tamanho_kb} KB")
    print(f"  Abas     : {len(abas)}")
    for aba in abas:
        print(f"    - {aba}")
    print(f"  Registros: {len(df)}")
    print(f"  Colunas  : {len(df.columns)}")
    print("=" * 60)
    print()

    return 0


if __name__ == "__main__":
    sys.exit(_cli())

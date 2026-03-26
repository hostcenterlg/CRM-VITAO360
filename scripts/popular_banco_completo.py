"""
popular_banco_completo.py — Importa dados adicionais da planilha para o SQLite
Versão: 1.1 | Data: 2026-03-26

FONTE: CRM_VITAO360 INTELIGENTE FINAL OK.xlsx
DESTINO: data/crm_vitao360.db

O QUE IMPORTA:
1. Vendas mensais da aba 'Venda Mês a Mês' (Jan/25 a Mar/26)
2. Mapeamento consultor da aba 'OPERACIONAL' (actualiza clientes sem consultor)
3. Metas mensais da aba 'PROJECAO' (Jan-Dez por CNPJ)
4. Exporta seed_data.json atualizado

REGRAS:
- CNPJ = string 14 dígitos, zero-padded
- Two-Base: vendas têm R$, logs têm R$0.00
- Classificação: REAL (dados diretos da planilha)
- Sem fabricação de dados
"""

import sys
import re
import sqlite3
import logging
from datetime import datetime
from pathlib import Path

import openpyxl

# ── Configuração ──────────────────────────────────────────────────────────────
PLANILHA = r'C:\Users\User\OneDrive\Área de Trabalho\CRM_VITAO360  INTELIGENTE   FINAL OK .xlsx'
DB_PATH  = r'C:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\crm_vitao360.db'

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
log = logging.getLogger(__name__)

# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_cnpj(val) -> str | None:
    """Normaliza CNPJ para 14 dígitos string, zero-padded."""
    if val is None:
        return None
    try:
        if isinstance(val, float):
            val = str(int(val))
        else:
            val = str(val).strip()
        val = re.sub(r'\D', '', val).zfill(14)
        if len(val) != 14 or val == '00000000000000':
            return None
        return val
    except Exception:
        return None


def normalize_consultor(raw: str | None) -> str | None:
    """Mapeia variações de nome de consultor para nome canônico."""
    if not raw:
        return None
    raw = str(raw).strip().upper()
    if any(x in raw for x in ['LARISSA', 'LARI', 'MAIS GRANEL', 'RODRIGO']):
        return 'LARISSA PADILHA'
    if any(x in raw for x in ['MANU', 'HEMANUELE', 'DITZEL']):
        return 'MANU DITZEL'
    if any(x in raw for x in ['DAIANE', 'STAVICKI']):
        return 'DAIANE STAVICKI'
    if any(x in raw for x in ['JULIO', 'GADRET']):
        return 'JULIO GADRET'
    return raw.title()


def safe_float(val) -> float | None:
    """Converte valor para float, retorna None se inválido ou zero."""
    if val is None:
        return None
    try:
        f = float(val)
        return f if f != 0.0 else None
    except (ValueError, TypeError):
        return None


def safe_date(val) -> str | None:
    """Converte datetime/string para ISO date string."""
    if val is None:
        return None
    if hasattr(val, 'date'):
        return val.date().isoformat()
    if isinstance(val, str) and val.strip():
        return val.strip()
    return None

# ── Etapa 1: Carregar mapeamento CNPJ → consultor do OPERACIONAL ──────────────

def carregar_mapa_consultor(wb) -> dict:
    """Lê OPERACIONAL e retorna {cnpj: consultor_canonico}."""
    ws = wb['OPERACIONAL']
    mapa = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        cnpj = normalize_cnpj(row[0])   # col A
        consultor_raw = row[17]          # col R
        if cnpj and consultor_raw:
            consultor = normalize_consultor(str(consultor_raw))
            if consultor:
                mapa[cnpj] = consultor
    log.info(f'Mapa consultor: {len(mapa)} CNPJs mapeados do OPERACIONAL')
    return mapa

# ── Etapa 2: Importar vendas mensais de 'Venda Mês a Mês' ────────────────────

# Definição dos meses e índices de coluna (1-indexed conforme planilha)
# Faturado_Mes = col X (Faturado é a coluna real de R$ líquido)
MESES_2025 = [
    ('2025-01', 7),   # Faturado_Jan
    ('2025-02', 11),  # Faturado_Fev
    ('2025-03', 15),  # Faturado_Mar
    ('2025-04', 19),  # Faturado_Abr
    ('2025-05', 23),  # Faturado_Mai
    ('2025-06', 27),  # Faturado_Jun
    ('2025-07', 31),  # Faturado_Jul
    ('2025-08', 35),  # Faturado_Ago
    ('2025-09', 39),  # Faturado_Set
    ('2025-10', 43),  # Faturado_Out
    ('2025-11', 47),  # Faturado_Nov
    ('2025-12', 51),  # Faturado_Dez
]
MESES_2026 = [
    ('2026-01', 59),  # Faturado_Jan/26
    ('2026-02', 63),  # Faturado_Fev/26
    ('2026-03', 67),  # Faturado_Mar/26
]
TODOS_MESES = MESES_2025 + MESES_2026


def importar_vendas(wb, conn, mapa_consultor: dict) -> tuple[int, int]:
    """
    Importa vendas mensais da aba 'Venda Mês a Mês'.
    Retorna (inseridos, ignorados_duplicata).
    """
    ws = wb['Venda Mês a Mês']

    # Buscar pares (cnpj, mes_referencia) já no banco
    existentes = set()
    for row in conn.execute('SELECT cnpj, mes_referencia FROM vendas'):
        existentes.add((row[0], row[1]))

    inseridos = 0
    ignorados = 0
    erros = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        cnpj = normalize_cnpj(row[2])   # col C = CNPJ ou CPF Cliente
        if not cnpj:
            continue

        nome_cliente = str(row[1]).strip() if row[1] else None
        consultor = mapa_consultor.get(cnpj, None)

        for mes_ref, col_idx in TODOS_MESES:
            valor = safe_float(row[col_idx - 1])  # 0-indexed
            if valor is None or valor <= 0:
                continue

            key = (cnpj, mes_ref)
            if key in existentes:
                ignorados += 1
                continue

            # Data do pedido = primeiro dia do mês
            data_pedido = f'{mes_ref}-01'

            try:
                # consultor é NOT NULL — usar 'DESCONHECIDO' como fallback
                consultor_final = consultor if consultor else 'DESCONHECIDO'
                conn.execute(
                    """
                    INSERT INTO vendas
                        (cnpj, data_pedido, valor_pedido, consultor,
                         fonte, classificacao_3tier, mes_referencia)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (cnpj, data_pedido, valor, consultor_final,
                     'MERCOS', 'REAL', mes_ref)
                )
                existentes.add(key)
                inseridos += 1
            except Exception as e:
                log.warning(f'Erro ao inserir venda {cnpj}/{mes_ref}: {e}')
                erros += 1

    log.info(f'Vendas: {inseridos} inseridas, {ignorados} já existiam, {erros} erros')
    return inseridos, ignorados

# ── Etapa 3: Popular RNC ──────────────────────────────────────────────────────

def importar_rnc(wb, conn, mapa_consultor: dict) -> int:
    """
    Importa RNCs da aba 'RNC'.
    Row 3 = header, dados a partir de row 4.
    Retorna quantidade inserida.
    """
    ws = wb['RNC']

    # RNC tem CNPJ na col C (índice 2) — verificar mapeamento real pelo header
    # Row 3: ('# RNC', 'DATA ABERTURA', 'CNPJ', 'CLIENTE', 'UF',
    #          'CONSULTOR', 'TIPO PROBLEMA', 'ÁREA RESPONSÁVEL',
    #          'DESCRIÇÃO DO PROBLEMA', 'STATUS', 'DATA RESOLUÇÃO', 'IMPACTO NO CLIENTE')
    # Cols: A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9, K=10, L=11

    inseridos = 0
    skipped_vazio = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        # Verificar se a linha tem pelo menos dados mínimos
        if not any(v is not None for v in row[:12]):
            skipped_vazio += 1
            continue

        cnpj_raw = row[2]    # col C
        nome_cliente = row[3] # col D (para fallback de CNPJ)

        cnpj = normalize_cnpj(cnpj_raw)
        if not cnpj:
            # Tentar normalizar pelo nome se tiver no mapeamento inverso
            # Se não conseguir CNPJ, pular (campo NOT NULL no schema)
            skipped_vazio += 1
            continue

        data_abertura  = safe_date(row[1])   # col B
        uf             = str(row[4]).strip() if row[4] else None   # col E
        consultor_raw  = row[5]              # col F
        tipo_problema  = str(row[6]).strip() if row[6] else None   # col G
        responsavel    = str(row[7]).strip() if row[7] else None   # col H → area_responsavel
        descricao      = str(row[8]).strip() if row[8] else None   # col I
        status         = str(row[9]).strip() if row[9] else None   # col J
        data_resolucao = safe_date(row[10])  # col K

        consultor = normalize_consultor(str(consultor_raw)) if consultor_raw else mapa_consultor.get(cnpj)

        # data_abertura é NOT NULL no schema — usar fallback se ausente
        if not data_abertura:
            data_abertura = '2025-01-01'

        # status padrão
        if not status:
            status = 'PENDENTE'

        try:
            conn.execute(
                """
                INSERT OR IGNORE INTO rnc
                    (cnpj, data_abertura, consultor, tipo_problema,
                     descricao, status, responsavel, data_resolucao)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (cnpj, data_abertura, consultor, tipo_problema,
                 descricao, status, responsavel, data_resolucao)
            )
            inseridos += 1
        except Exception as e:
            log.warning(f'Erro ao inserir RNC {cnpj}: {e}')

    log.info(f'RNC: {inseridos} inseridas, {skipped_vazio} linhas vazias ignoradas')
    return inseridos

# ── Etapa 3b: Importar metas mensais de PROJECAO ─────────────────────────────

# Colunas de META e REAL na aba PROJECAO (1-indexed)
PROJECAO_METAS = [(1,13),(2,14),(3,15),(4,16),(5,17),(6,18),(7,19),(8,20),(9,21),(10,22),(11,23),(12,24)]
PROJECAO_REAIS = [(1,27),(2,28),(3,29),(4,30),(5,31),(6,32),(7,33),(8,34),(9,35),(10,36),(11,37),(12,38)]


def importar_metas_mensais(wb, conn) -> tuple[int, int]:
    """
    Importa metas mensais (Jan-Dez 2025) da aba 'PROJECAO'.
    Retorna (inseridas, ignoradas_duplicata).
    """
    ws = wb['PROJEÇÃO ']

    # Buscar pares já existentes (ano=2025, mes>0)
    existentes = set()
    for row in conn.execute('SELECT cnpj, mes FROM metas WHERE ano = 2025 AND mes > 0'):
        existentes.add((row[0], row[1]))

    inseridas = 0
    ignoradas = 0
    erros = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        cnpj = normalize_cnpj(row[0])
        if not cnpj:
            continue

        for (mes, meta_col), (_, real_col) in zip(PROJECAO_METAS, PROJECAO_REAIS):
            meta = safe_float(row[meta_col - 1])
            real = safe_float(row[real_col - 1])

            # Só inserir se tiver pelo menos meta ou real
            if meta is None and real is None:
                continue

            key = (cnpj, mes)
            if key in existentes:
                ignoradas += 1
                continue

            try:
                conn.execute(
                    """
                    INSERT INTO metas
                        (cnpj, ano, mes, meta_sap, realizado, fonte)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (cnpj, 2025, mes, meta or 0.0, real or 0.0, 'SAP')
                )
                existentes.add(key)
                inseridas += 1
            except Exception as e:
                log.warning(f'Erro ao inserir meta {cnpj}/2025-{mes:02d}: {e}')
                erros += 1

    log.info(f'Metas mensais: {inseridas} inseridas, {ignoradas} já existiam, {erros} erros')
    return inseridas, ignoradas


# ── Etapa 4: Actualizar consultor em clientes sem consultor ───────────────────

def actualizar_consultores_clientes(conn, mapa_consultor: dict) -> int:
    """
    Para clientes com consultor NULL ou vazio, actualiza a partir do mapa.
    Retorna quantidade actualizada.
    """
    actualizados = 0
    for cnpj, consultor in mapa_consultor.items():
        result = conn.execute(
            """
            UPDATE clientes
            SET consultor = ?
            WHERE cnpj = ?
              AND (consultor IS NULL OR consultor = '')
            """,
            (consultor, cnpj)
        )
        actualizados += result.rowcount
    log.info(f'Clientes actualizados com consultor: {actualizados}')
    return actualizados

# ── Etapa 5: Exportar seed_data.json ─────────────────────────────────────────

def exportar_seed_data(conn, seed_path: str) -> None:
    """Re-exporta todas as tabelas para seed_data.json."""
    import json

    conn.row_factory = sqlite3.Row
    tabelas = ['clientes', 'vendas', 'metas', 'log_interacoes', 'redes', 'rnc']
    output = {}

    for tabela in tabelas:
        try:
            rows = conn.execute(f'SELECT * FROM [{tabela}]').fetchall()
            output[tabela] = [dict(r) for r in rows]
            log.info(f'  {tabela}: {len(output[tabela])} registros exportados')
        except Exception as e:
            log.warning(f'  {tabela}: erro ao exportar — {e}')
            output[tabela] = []

    Path(seed_path).write_text(
        json.dumps(output, default=str, ensure_ascii=False, indent=2),
        encoding='utf-8'
    )
    log.info(f'seed_data.json salvo em {seed_path}')

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    log.info('=' * 60)
    log.info('popular_banco_completo.py — início')
    log.info(f'Planilha: {PLANILHA}')
    log.info(f'Banco: {DB_PATH}')
    log.info('=' * 60)

    # 0. Verificar arquivos
    if not Path(PLANILHA).exists():
        log.error(f'Planilha não encontrada: {PLANILHA}')
        sys.exit(1)
    if not Path(DB_PATH).exists():
        log.error(f'Banco não encontrado: {DB_PATH}')
        sys.exit(1)

    # 1. Abrir planilha
    log.info('Abrindo planilha (pode demorar)...')
    wb = openpyxl.load_workbook(PLANILHA, data_only=True)
    log.info('Planilha aberta.')

    # 2. Conectar ao banco
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = None  # reset antes das queries de inserção

    try:
        # 3. Carregar mapa consultor
        log.info('')
        log.info('[1/4] Carregando mapa consultor do OPERACIONAL...')
        mapa_consultor = carregar_mapa_consultor(wb)

        # 4. Importar vendas
        log.info('')
        log.info('[2/4] Importando vendas mensais de "Venda Mês a Mês"...')
        vendas_inseridas, vendas_ignoradas = importar_vendas(wb, conn, mapa_consultor)

        # 5. Importar metas mensais
        log.info('')
        log.info('[3/5] Importando metas mensais de "PROJECAO"...')
        metas_inseridas, metas_ignoradas = importar_metas_mensais(wb, conn)

        # 5b. Importar RNC (aba vazia na planilha atual — reportar)
        log.info('')
        log.info('[4/5] Verificando RNCs...')
        rnc_inseridas = importar_rnc(wb, conn, mapa_consultor)

        # 6. Actualizar consultores em clientes
        log.info('')
        log.info('[5/5] Actualizando consultores em clientes...')
        clientes_actualizados = actualizar_consultores_clientes(conn, mapa_consultor)

        # 7. Commit
        conn.commit()

        # 8. Resumo
        log.info('')
        log.info('=' * 60)
        log.info('RESUMO DA IMPORTAÇÃO')
        log.info('=' * 60)
        log.info(f'Vendas inseridas:          {vendas_inseridas:>6}')
        log.info(f'Vendas já existiam:        {vendas_ignoradas:>6}')
        log.info(f'Metas mensais inseridas:   {metas_inseridas:>6}')
        log.info(f'Metas já existiam:         {metas_ignoradas:>6}')
        log.info(f'RNCs inseridas:            {rnc_inseridas:>6}')
        log.info(f'Clientes actualizados:     {clientes_actualizados:>6}')

        total_vendas  = conn.execute('SELECT COUNT(*) FROM vendas').fetchone()[0]
        total_logs    = conn.execute('SELECT COUNT(*) FROM log_interacoes').fetchone()[0]
        total_rnc     = conn.execute('SELECT COUNT(*) FROM rnc').fetchone()[0]
        total_metas   = conn.execute('SELECT COUNT(*) FROM metas').fetchone()[0]
        total_metas_m = conn.execute('SELECT COUNT(*) FROM metas WHERE mes > 0').fetchone()[0]
        log.info('')
        log.info(f'Total vendas no banco:     {total_vendas:>6}')
        log.info(f'Total logs no banco:       {total_logs:>6}')
        log.info(f'Total RNCs no banco:       {total_rnc:>6}')
        log.info(f'Total metas no banco:      {total_metas:>6}  (mensais: {total_metas_m})')
        log.info('=' * 60)

        # 9. Validação Two-Base
        log.info('')
        log.info('VALIDAÇÃO TWO-BASE...')
        # Vendas devem ter valor > 0
        vendas_zero = conn.execute('SELECT COUNT(*) FROM vendas WHERE valor_pedido <= 0').fetchone()[0]
        log.info(f'Vendas com valor <= 0: {vendas_zero} (deve ser 0)')
        if vendas_zero > 0:
            log.warning('AVISO: existem vendas com valor <= 0 — verificar!')

        # 10. Exportar seed_data.json
        log.info('')
        log.info('Exportando seed_data.json...')
        seed_path = str(Path(DB_PATH).parent / 'seed_data.json')
        exportar_seed_data(conn, seed_path)

        log.info('')
        log.info('CONCLUÍDO COM SUCESSO.')

    except Exception as e:
        conn.rollback()
        log.error(f'Erro durante importação: {e}')
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()


if __name__ == '__main__':
    main()

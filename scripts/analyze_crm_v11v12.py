"""Analyze CRM V11 and V12 Excel files for structure, tabs, formulas."""
import openpyxl
import os

FILES = [
    r"c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\sources\crm-versoes\v11-v12\CRM_INTELIGENTE_VITAO_360_V11_LIMPO.xlsx",
    r"c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\sources\crm-versoes\v11-v12\CRM_INTELIGENTE_VITAO_360_V11_POPULADO.xlsx",
    r"c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\sources\crm-versoes\v11-v12\CRM_INTELIGENTE_VITAO360_V12 (1).xlsx",
]

EXPECTED_TABS = [
    "CARTEIRA", "LOG", "PROJEÇÃO", "DASH", "AGENDA",
    "DRAFT 1", "DRAFT 2", "DRAFT 3", "SINALEIRO",
    "SINALEIRO INTERNO", "REDES_FRANQUIAS", "COMITÊ",
    "PAINEL", "UPDATE_LOG"
]

def analyze_sheet(ws, sheet_name, max_rows=100):
    """Analyze a single sheet."""
    formula_count = 0
    data_count = 0
    empty_count = 0
    sample_formulas = []
    error_formulas = []

    row_count = 0
    col_count = 0
    for row in ws.iter_rows():
        row_count += 1
        if row_count == 1:
            col_count = len(row)
        if row_count > max_rows:
            break
        for cell in row:
            if cell.value is None:
                empty_count += 1
            elif isinstance(cell.value, str):
                if cell.value.startswith('='):
                    formula_count += 1
                    if len(sample_formulas) < 5:
                        sample_formulas.append(f"      {cell.coordinate}: {cell.value[:120]}")
                    # Check for error patterns in formulas
                    for err in ['#REF!', '#DIV/0!', '#VALUE!', '#NAME?']:
                        if err in str(cell.value):
                            error_formulas.append(f"      {cell.coordinate}: {cell.value[:80]}")
                elif cell.value in ('#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#N/A'):
                    error_formulas.append(f"      {cell.coordinate}: {cell.value}")
                else:
                    data_count += 1
            else:
                data_count += 1

    print(f"  Rows (scanned): {row_count} | Cols: {col_count}")
    print(f"  Formulas: {formula_count} | Data: {data_count} | Empty: {empty_count}")

    if sample_formulas:
        print(f"  Sample formulas:")
        for sf in sample_formulas[:5]:
            print(sf)

    if error_formulas:
        print(f"  ⚠️ ERRORS FOUND ({len(error_formulas)}):")
        for ef in error_formulas[:5]:
            print(ef)

    # Headers
    headers = []
    for row in ws.iter_rows(max_row=1):
        headers = [str(c.value)[:25] if c.value else "" for c in row[:25]]
    if headers:
        print(f"  Headers: {headers}")

    return row_count, col_count, formula_count, data_count


def analyze_file(filepath):
    basename = os.path.basename(filepath)
    filesize = os.path.getsize(filepath) / (1024 * 1024)
    print(f"\n{'='*80}")
    print(f"FILE: {basename} ({filesize:.1f} MB)")
    print(f"{'='*80}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=False, read_only=True)
    except Exception as e:
        print(f"ERROR: {e}")
        return

    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    # Check expected tabs
    found = [t for t in EXPECTED_TABS if any(t.lower() in s.lower() for s in wb.sheetnames)]
    missing = [t for t in EXPECTED_TABS if not any(t.lower() in s.lower() for s in wb.sheetnames)]
    print(f"\nExpected tabs found: {found}")
    print(f"Expected tabs MISSING: {missing}")

    total_formulas = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- {sheet_name} ---")
        # Scan more rows for key sheets
        max_r = 600 if any(k in sheet_name.upper() for k in ['CARTEIRA', 'LOG', 'PROJ']) else 100
        rows, cols, formulas, data = analyze_sheet(ws, sheet_name, max_rows=max_r)
        total_formulas += formulas

    print(f"\n>>> TOTAL FORMULAS IN FILE: {total_formulas}")
    wb.close()


for f in FILES:
    if os.path.exists(f):
        analyze_file(f)
    else:
        print(f"\nFILE NOT FOUND: {f}")

print("\n\nDONE.")

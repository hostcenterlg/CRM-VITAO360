"""
JARVIS CRM CENTRAL — Full Excel Builder
Creates output/JARVIS_CRM_CENTRAL.xlsx with all 7 tabs.
"""
import os
import datetime
from copy import copy

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName

wb = Workbook()

# ── Global styles ──────────────────────────────────────────────
FONT_DEFAULT = Font(name='Arial', size=10)
FONT_HEADER = Font(name='Arial', size=10, bold=True)
FONT_TITLE_12 = Font(name='Arial', size=12, bold=True)
FONT_TITLE_14 = Font(name='Arial', size=14, bold=True)
FONT_KPI = Font(name='Arial', size=18, bold=True)
FONT_ITALIC = Font(name='Arial', size=10, italic=True)
FONT_BOLD_10 = Font(name='Arial', size=10, bold=True)
FONT_WHITE_BOLD = Font(name='Arial', size=10, bold=True, color='FFFFFF')
FONT_WHITE_12 = Font(name='Arial', size=12, bold=True, color='FFFFFF')

FILL_HEADER = PatternFill('solid', fgColor='D9D9D9')
FILL_GREEN = PatternFill('solid', fgColor='00B050')
FILL_YELLOW = PatternFill('solid', fgColor='FFC000')
FILL_ORANGE = PatternFill('solid', fgColor='FF6600')
FILL_RED = PatternFill('solid', fgColor='FF0000')
FILL_BLUE_LIGHT = PatternFill('solid', fgColor='BDD7EE')
FILL_GREEN_LIGHT = PatternFill('solid', fgColor='C6EFCE')
FILL_RED_LIGHT = PatternFill('solid', fgColor='FFC7CE')
FILL_KPI = PatternFill('solid', fgColor='E2EFDA')
FILL_SECTION_TITLE = PatternFill('solid', fgColor='4472C4')
FILL_AGENDA_RO = PatternFill('solid', fgColor='D9E2F3')
FILL_AGENDA_EDIT = PatternFill('solid', fgColor='E2EFDA')
FILL_WARN = PatternFill('solid', fgColor='FFC000')

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def apply_header(ws, row, cols, extra_fill=None):
    """Apply header formatting to a row."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = extra_fill if extra_fill else FILL_HEADER
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER


def set_default_font(ws, max_row=200, max_col=100):
    """Set Arial 10 as default for all cells."""
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.font == Font():
                cell.font = FONT_DEFAULT


def auto_width(ws, max_col=None, min_width=10, max_width=25):
    """Auto-adjust column widths."""
    mc = max_col or ws.max_column or 1
    for col_idx in range(1, mc + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, max_row=min(ws.max_row or 1, 50)):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = max_len


# ═══════════════════════════════════════════════════════════════
# ABA 1: REGRAS
# ═══════════════════════════════════════════════════════════════
print("Building ABA 1: REGRAS...")
ws_regras = wb.active
ws_regras.title = "REGRAS"
ws_regras.sheet_properties.tabColor = "4472C4"

current_row = 1

# ── Tabela 1: RESULTADO ──
headers_resultado = ["RESULTADO", "FOLLOW-UP (DIAS)", "GRUPO DASH", "QUANDO USAR"]
for i, h in enumerate(headers_resultado, 1):
    ws_regras.cell(row=current_row, column=i, value=h)
apply_header(ws_regras, current_row, 4)

resultado_data = [
    ("EM ATENDIMENTO", 2, "FUNIL", "Negociação ativa"),
    ("ORÇAMENTO", 1, "FUNIL", "Proposta enviada"),
    ("CADASTRO", 2, "FUNIL", "Em processo de cadastro"),
    ("VENDA / PEDIDO", 45, "FUNIL", "Pedido fechado"),
    ("RELACIONAMENTO", 7, "RELAC.", "Pós-venda CS"),
    ("FOLLOW UP 7", 7, "RELAC.", "Retomar em 1 semana"),
    ("FOLLOW UP 15", 15, "RELAC.", "Retomar em 2 semanas"),
    ("SUPORTE", 0, "RELAC.", "Problema resolvido"),
    ("NÃO ATENDE", 1, "NÃO VENDA", "Escalona T+1"),
    ("NÃO RESPONDE", 1, "NÃO VENDA", "WhatsApp sem resposta"),
    ("RECUSOU LIGAÇÃO", 2, "NÃO VENDA", "Mudança de canal"),
    ("PERDA / FECHOU LOJA", 0, "NÃO VENDA", "Perdido definitivamente"),
]
for i, row_data in enumerate(resultado_data, current_row + 1):
    for j, val in enumerate(row_data, 1):
        cell = ws_regras.cell(row=i, column=j, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER

resultado_start = current_row + 1
resultado_end = current_row + len(resultado_data)
current_row = resultado_end + 3  # 2 blank lines

# ── Tabela 2: TIPO DO CONTATO ──
headers_tipo = ["TIPO DO CONTATO", "QUANDO USAR"]
for i, h in enumerate(headers_tipo, 1):
    ws_regras.cell(row=current_row, column=i, value=h)
apply_header(ws_regras, current_row, 2)

tipo_contato_data = [
    ("PROSPECÇÃO", "1º contato com prospect/lead"),
    ("NEGOCIAÇÃO", "Em negociação ativa"),
    ("FOLLOW UP", "Retorno agendado"),
    ("ATEND. CLIENTES ATIVOS", "Dentro do ciclo de compra"),
    ("ATEND. CLIENTES INATIVOS", "Ultrapassou ciclo"),
    ("PÓS-VENDA / RELACIONAMENTO", "CS verificar PDV"),
    ("MOTIVO / PAROU DE COMPRAR", "Investigar feedback fábrica"),
]
tc_start = current_row + 1
for i, row_data in enumerate(tipo_contato_data, current_row + 1):
    for j, val in enumerate(row_data, 1):
        cell = ws_regras.cell(row=i, column=j, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER
tc_end = current_row + len(tipo_contato_data)
current_row = tc_end + 3

# ── Tabela 3: MOTIVO ──
headers_motivo = ["MOTIVO", "DONO DA AÇÃO", "INSIGHT"]
for i, h in enumerate(headers_motivo, 1):
    ws_regras.cell(row=current_row, column=i, value=h)
apply_header(ws_regras, current_row, 3)

motivo_data = [
    ("PRODUTO NÃO VENDEU / SEM GIRO", "DIRETORIA / FÁBRICA", "Produto sem fit no PDV"),
    ("PREÇO ALTO / MARGEM", "COMERCIAL", "Revisar tabela"),
    ("PREFERIU CONCORRENTE", "COMERCIAL / MKT", "Qual concorrente"),
    ("PROBLEMA LOGÍSTICO", "LOGÍSTICA", "Atraso avaria falta"),
    ("PROBLEMA FINANCEIRO", "FINANCEIRO", "Bloqueado sem crédito"),
    ("AINDA TEM ESTOQUE", "NORMAL", "Ajustar ciclo"),
    ("FECHOU LOJA", "PERDA", "Atualizar status"),
    ("SEM INTERESSE", "NUTRIÇÃO", "Ir pra nutrição"),
    ("VIAJANDO / INDISPONÍVEL", "NORMAL", "Reagendar"),
    ("1º CONTATO / SEM MOTIVO", "NORMAL", "Prospecção inicial"),
]
mot_start = current_row + 1
for i, row_data in enumerate(motivo_data, current_row + 1):
    for j, val in enumerate(row_data, 1):
        cell = ws_regras.cell(row=i, column=j, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER
mot_end = current_row + len(motivo_data)
current_row = mot_end + 3

# ── Tabela 4: SITUAÇÃO ──
headers_sit = ["DIAS SEM COMPRA", "SITUAÇÃO", "COR HEX"]
for i, h in enumerate(headers_sit, 1):
    ws_regras.cell(row=current_row, column=i, value=h)
apply_header(ws_regras, current_row, 3)

situacao_data = [
    ("≤ 50", "ATIVO", "00B050"),
    ("51-60", "EM RISCO", "FFC000"),
    ("61-90", "INATIVO RECENTE", "FF6600"),
    ("> 90", "INATIVO ANTIGO", "FF0000"),
    ("Nunca comprou", "PROSPECT", "BDD7EE"),
]
sit_start = current_row + 1
for i, row_data in enumerate(situacao_data, current_row + 1):
    for j, val in enumerate(row_data, 1):
        cell = ws_regras.cell(row=i, column=j, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER
    # Apply fill to COR HEX cell
    color_cell = ws_regras.cell(row=i, column=3)
    color_cell.fill = PatternFill('solid', fgColor=row_data[2])
sit_end = current_row + len(situacao_data)
current_row = sit_end + 3

# ── Tabela 5: FASE ──
headers_fase = ["CONDIÇÃO", "FASE", "PRIORIDADE"]
for i, h in enumerate(headers_fase, 1):
    ws_regras.cell(row=current_row, column=i, value=h)
apply_header(ws_regras, current_row, 3)

fase_data = [
    ("ATIVO 0-10 dias", "PÓS-VENDA", "Baixa"),
    ("ATIVO 10-25 dias", "CS", "Média"),
    ("ATIVO perto ciclo (-7d)", "RECOMPRA", "Alta"),
    ("INATIVO RECENTE", "SALVAMENTO", "Alta"),
    ("INATIVO ANTIGO", "RECUPERAÇÃO", "Média"),
    ("PROSPECT / LEAD", "PROSPECÇÃO", "Variável"),
    ("Após T4 sem resposta", "NUTRIÇÃO", "Baixa"),
    ("Com resposta", "EM ATENDIMENTO", "Alta"),
    ("Com orçamento aberto", "ORÇAMENTO", "Máxima"),
]
fase_start = current_row + 1
for i, row_data in enumerate(fase_data, current_row + 1):
    for j, val in enumerate(row_data, 1):
        cell = ws_regras.cell(row=i, column=j, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER
fase_end = current_row + len(fase_data)
current_row = fase_end + 3

# ── Tabela 6: CADÊNCIA ──
headers_cad = ["TENTATIVA", "CANAL", "INTERVALO", "COMPORTAMENTO"]
for i, h in enumerate(headers_cad, 1):
    ws_regras.cell(row=current_row, column=i, value=h)
apply_header(ws_regras, current_row, 4)

cadencia_data = [
    ("T1", "WhatsApp", "Dia 0", "Primeira abordagem"),
    ("T2", "Ligação", "+1 dia útil", "Se não respondeu T1"),
    ("T3", "WhatsApp", "+2 dias", "Mudança de abordagem"),
    ("T4", "Ligação final", "+2 dias", "Última tentativa ativa"),
    ("NUTRIÇÃO", "Email + WhatsApp", "Ciclo 15 dias", "Após T4 sem resposta"),
    ("RESET", "—", "—", "Quando responde volta T1"),
]
cad_start = current_row + 1
for i, row_data in enumerate(cadencia_data, current_row + 1):
    for j, val in enumerate(row_data, 1):
        cell = ws_regras.cell(row=i, column=j, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER
cad_end = current_row + len(cadencia_data)
current_row = cad_end + 3

# ── Tabela 7: SINALEIRO ──
headers_sin = ["SINALEIRO", "COR HEX", "SIGNIFICADO"]
for i, h in enumerate(headers_sin, 1):
    ws_regras.cell(row=current_row, column=i, value=h)
apply_header(ws_regras, current_row, 3)

sinaleiro_data = [
    ("🟢", "00B050", "Saudável"),
    ("🟡", "FFC000", "Atenção"),
    ("🔴", "FF0000", "Crítico"),
    ("🟣", "BDD7EE", "Prospect"),
]
sin_start = current_row + 1
for i, row_data in enumerate(sinaleiro_data, current_row + 1):
    for j, val in enumerate(row_data, 1):
        cell = ws_regras.cell(row=i, column=j, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER
    color_cell = ws_regras.cell(row=i, column=2)
    color_cell.fill = PatternFill('solid', fgColor=row_data[1])
sin_end = current_row + len(sinaleiro_data)
current_row = sin_end + 3

# ── Tabela 8: PRIORIDADE ──
headers_pri = ["COMPONENTE", "CONDIÇÃO", "PONTOS"]
for i, h in enumerate(headers_pri, 1):
    ws_regras.cell(row=current_row, column=i, value=h)
apply_header(ws_regras, current_row, 3)

prioridade_data = [
    ("CURVA ABC", "A", 30),
    ("CURVA ABC", "B", 15),
    ("CURVA ABC", "C", 5),
    ("SINALEIRO", "🔴", 40),
    ("SINALEIRO", "🟡", 25),
    ("SINALEIRO", "🟢", 10),
    ("SINALEIRO", "🟣", 5),
    ("FASE", "RECOMPRA/SALVAMENTO", 25),
    ("FASE", "CS", 20),
    ("FASE", "PROSPECÇÃO", 10),
    ("FASE", "NUTRIÇÃO", 5),
]
pri_start = current_row + 1
for i, row_data in enumerate(prioridade_data, current_row + 1):
    for j, val in enumerate(row_data, 1):
        cell = ws_regras.cell(row=i, column=j, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER
pri_end = current_row + len(prioridade_data)
current_row = pri_end + 3

# ── Tabela 9: CONSULTORES ──
headers_cons = ["CONSULTOR", "TERRITÓRIO", "TIPO", "REDES"]
for i, h in enumerate(headers_cons, 1):
    ws_regras.cell(row=current_row, column=i, value=h)
apply_header(ws_regras, current_row, 4)

consul_data = [
    ("MANU DITZEL", "SC, PR, RS", "Regional", "—"),
    ("LARISSA PADILHA", "Resto do Brasil", "Regional", "—"),
    ("JULIO GADRET", "Nacional", "Rede", "CIA DA SAUDE, FIT LAND"),
    ("DAIANE STAVICKI", "Nacional", "Rede + Central", "Mundo Verde, Biomundo, Divina Terra, Vida Leve, Tudo em Grãos, Armazém Fitstore"),
]
cons_start = current_row + 1
for i, row_data in enumerate(consul_data, current_row + 1):
    for j, val in enumerate(row_data, 1):
        cell = ws_regras.cell(row=i, column=j, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER
cons_end = current_row + len(consul_data)
current_row = cons_end + 3

# ── Tabela 10: LISTAS AUXILIARES (vertical, para Named Ranges) ──
# Each list goes in its own column block so Named Ranges reference a vertical range

# Column F: SIM_NAO
ws_regras.cell(row=current_row, column=6, value="SIM_NAO").font = FONT_HEADER
ws_regras.cell(row=current_row, column=6).fill = FILL_HEADER
ws_regras.cell(row=current_row, column=6).border = THIN_BORDER
sn_items = ["SIM", "NÃO"]
sn_start = current_row + 1
for idx, item in enumerate(sn_items):
    cell = ws_regras.cell(row=sn_start + idx, column=6, value=item)
    cell.font = FONT_DEFAULT
    cell.border = THIN_BORDER
sn_end = sn_start + len(sn_items) - 1

# Column G: SIM_NAO_NA
ws_regras.cell(row=current_row, column=7, value="SIM_NAO_NA").font = FONT_HEADER
ws_regras.cell(row=current_row, column=7).fill = FILL_HEADER
ws_regras.cell(row=current_row, column=7).border = THIN_BORDER
sn_na_items = ["SIM", "NÃO", "N/A"]
sn_na_start = current_row + 1
for idx, item in enumerate(sn_na_items):
    cell = ws_regras.cell(row=sn_na_start + idx, column=7, value=item)
    cell.font = FONT_DEFAULT
    cell.border = THIN_BORDER
sn_na_end = sn_na_start + len(sn_na_items) - 1

# Column H: ATIVO_RECEPTIVO
ws_regras.cell(row=current_row, column=8, value="ATIVO_RECEPTIVO").font = FONT_HEADER
ws_regras.cell(row=current_row, column=8).fill = FILL_HEADER
ws_regras.cell(row=current_row, column=8).border = THIN_BORDER
ar_items = ["ATIVO", "RECEPTIVO"]
ar_start = current_row + 1
for idx, item in enumerate(ar_items):
    cell = ws_regras.cell(row=ar_start + idx, column=8, value=item)
    cell.font = FONT_DEFAULT
    cell.border = THIN_BORDER
ar_end = ar_start + len(ar_items) - 1

# Column I: CURVA_ABC
ws_regras.cell(row=current_row, column=9, value="CURVA_ABC").font = FONT_HEADER
ws_regras.cell(row=current_row, column=9).fill = FILL_HEADER
ws_regras.cell(row=current_row, column=9).border = THIN_BORDER
abc_items = ["A", "B", "C"]
abc_start = current_row + 1
for idx, item in enumerate(abc_items):
    cell = ws_regras.cell(row=abc_start + idx, column=9, value=item)
    cell.font = FONT_DEFAULT
    cell.border = THIN_BORDER
abc_end = abc_start + len(abc_items) - 1

# Column J: TIPO_CLIENTE (6 opções conforme spec)
ws_regras.cell(row=current_row, column=10, value="TIPO_CLIENTE").font = FONT_HEADER
ws_regras.cell(row=current_row, column=10).fill = FILL_HEADER
ws_regras.cell(row=current_row, column=10).border = THIN_BORDER
tc_cli_items = ["NOVO", "EM DESENVOLVIMENTO", "RECORRENTE", "FIDELIZADO", "PROSPECT", "LEAD"]
tc_cli_start = current_row + 1
for idx, item in enumerate(tc_cli_items):
    cell = ws_regras.cell(row=tc_cli_start + idx, column=10, value=item)
    cell.font = FONT_DEFAULT
    cell.border = THIN_BORDER
tc_cli_end = tc_cli_start + len(tc_cli_items) - 1

# Column K: BLOCO
ws_regras.cell(row=current_row, column=11, value="BLOCO").font = FONT_HEADER
ws_regras.cell(row=current_row, column=11).fill = FILL_HEADER
ws_regras.cell(row=current_row, column=11).border = THIN_BORDER
bloco_items = ["MANHÃ", "TARDE"]
bloco_start = current_row + 1
for idx, item in enumerate(bloco_items):
    cell = ws_regras.cell(row=bloco_start + idx, column=11, value=item)
    cell.font = FONT_DEFAULT
    cell.border = THIN_BORDER
bloco_end = bloco_start + len(bloco_items) - 1

# Column L: UF (27 UFs válidas)
ws_regras.cell(row=current_row, column=12, value="UF").font = FONT_HEADER
ws_regras.cell(row=current_row, column=12).fill = FILL_HEADER
ws_regras.cell(row=current_row, column=12).border = THIN_BORDER
uf_items = ["AC","AL","AM","AP","BA","CE","DF","ES","GO","MA","MG","MS","MT",
            "PA","PB","PE","PI","PR","RJ","RN","RO","RR","RS","SC","SE","SP","TO"]
uf_start = current_row + 1
for idx, item in enumerate(uf_items):
    cell = ws_regras.cell(row=uf_start + idx, column=12, value=item)
    cell.font = FONT_DEFAULT
    cell.border = THIN_BORDER
uf_end = uf_start + len(uf_items) - 1

# Column M: AÇÃO FUTURA
ws_regras.cell(row=current_row, column=13, value="AÇÃO FUTURA").font = FONT_HEADER
ws_regras.cell(row=current_row, column=13).fill = FILL_HEADER
ws_regras.cell(row=current_row, column=13).border = THIN_BORDER
acao_items = ["LIGAR", "WHATSAPP", "ENVIAR ORÇAMENTO", "ENVIAR CATÁLOGO",
              "REAGENDAR", "VISITAR", "AGUARDAR RETORNO", "NENHUMA"]
acao_start = current_row + 1
for idx, item in enumerate(acao_items):
    cell = ws_regras.cell(row=acao_start + idx, column=13, value=item)
    cell.font = FONT_DEFAULT
    cell.border = THIN_BORDER
acao_end = acao_start + len(acao_items) - 1

# Auto-width for REGRAS
auto_width(ws_regras, max_col=13)
ws_regras.sheet_view.zoomScale = 90

# ── Named Ranges ──
print("  Creating Named Ranges...")

named_ranges_def = {
    'LISTA_RESULTADO': f"REGRAS!$A${resultado_start}:$A${resultado_end}",
    'LISTA_TIPO_CONTATO': f"REGRAS!$A${tc_start}:$A${tc_end}",
    'LISTA_MOTIVO': f"REGRAS!$A${mot_start}:$A${mot_end}",
    'LISTA_SITUACAO': f"REGRAS!$B${sit_start}:$B${sit_end}",
    'LISTA_FASE': f"REGRAS!$B${fase_start}:$B${fase_end}",
    'LISTA_TENTATIVA': f"REGRAS!$A${cad_start}:$A${cad_end}",
    'LISTA_SINALEIRO': f"REGRAS!$A${sin_start}:$A${sin_end}",
    'LISTA_CONSULTOR': f"REGRAS!$A${cons_start}:$A${cons_end}",
    # Listas auxiliares (vertical, funciona para dropdowns)
    'LISTA_SIM_NAO': f"REGRAS!$F${sn_start}:$F${sn_end}",
    'LISTA_SIM_NAO_NA': f"REGRAS!$G${sn_na_start}:$G${sn_na_end}",
    'LISTA_ATIVO_RECEPTIVO': f"REGRAS!$H${ar_start}:$H${ar_end}",
    'LISTA_CURVA_ABC': f"REGRAS!$I${abc_start}:$I${abc_end}",
    'LISTA_TIPO_CLIENTE': f"REGRAS!$J${tc_cli_start}:$J${tc_cli_end}",
    'LISTA_BLOCO': f"REGRAS!$K${bloco_start}:$K${bloco_end}",
    'LISTA_UF': f"REGRAS!$L${uf_start}:$L${uf_end}",
    'LISTA_ACAO_FUTURA': f"REGRAS!$M${acao_start}:$M${acao_end}",
}

for name, ref in named_ranges_def.items():
    dn = DefinedName(name, attr_text=ref)
    wb.defined_names.add(dn)

# Also create a named range for the resultado table (for VLOOKUP in LOG)
dn_tab = DefinedName('TABELA_RESULTADO', attr_text=f"REGRAS!$A${resultado_start}:$B${resultado_end}")
wb.defined_names.add(dn_tab)

print("  ABA 1 REGRAS done.")


# ═══════════════════════════════════════════════════════════════
# ABA 2: CARTEIRA (81 columns)
# ═══════════════════════════════════════════════════════════════
print("Building ABA 2: CARTEIRA...")
ws_cart = wb.create_sheet("CARTEIRA")
ws_cart.sheet_properties.tabColor = "00B050"

# Headers
cart_headers = [
    # A-J Fixed
    "NOME FANTASIA", "REDE / REGIONAL", "UF", "CONSULTOR", "SITUAÇÃO",
    "DIAS SEM COMPRA", "🚦 SINALEIRO", "TIPO CLIENTE", "FASE", "TENTATIVA",
    # K-Q [+1] IDENTIDADE
    "CNPJ", "RAZÃO SOCIAL", "CIDADE", "EMAIL", "TELEFONE", "COD_FONTE", "DATA CADASTRO",
    # R-Z [+2] VIDA COMERCIAL
    "CICLO MÉDIO", "DATA ÚLTIMO PEDIDO", "VALOR ÚLTIMO PEDIDO", "VENDEDOR ÚLTIMO PEDIDO",
    "Nº DE COMPRAS", "CURVA ABC", "TICKET MÉDIO", "MESES POSITIVADO", "MÉDIA MENSAL",
    # AA-AM [+3] TIMELINE
    "TOTAL PERÍODO R$", "MAR/25", "ABR/25", "MAI/25", "JUN/25", "JUL/25",
    "AGO/25", "SET/25", "OUT/25", "NOV/25", "DEZ/25", "JAN/26", "MESES LISTA",
    # AN-AS [+4] JORNADA
    "DIAS ATÉ CONVERSÃO", "DATA 1º CONTATO", "DATA 1º ORÇAMENTO", "DATA 1ª VENDA",
    "TOTAL TENTATIVAS ATÉ VENDA", "DATA ÚLTIMA RECOMPRA",
    # AT-AX [+5] ECOMMERCE
    "ACESSO B2B", "ACESSOS PORTAL", "ITENS CARRINHO", "VALOR PORTAL B2B", "OPORTUNIDADE ECOM",
    # AY-BI [+6] SAP
    "COD_SAP", "STATUS PARCEIRO", "ENDEREÇO", "CEP", "CONDIÇÃO PAGAMENTO",
    "LISTA PREÇOS", "GRUPO CLIENTE", "BLOQUEIO VENDAS", "BLOQUEIO GERAL",
    "DESC. COMERCIAL", "MACRORREGIÃO",
    # BJ-BR [+7] OPERACIONAL
    "PRÓX. FOLLOW-UP", "BLOCO", "ESTÁGIO FUNIL", "AÇÃO DETALHADA",
    "DATA ÚLT. ATENDIMENTO", "ÚLTIMO RESULTADO", "MOTIVO", "PRIORIDADE", "OBSERVAÇÃO",
    # BS-CC [+8] COMITÊ
    "% ATING. MÊS", "META MÊS R$", "REALIZADO MÊS R$", "% ATING. TRI",
    "META TRI R$", "REALIZADO TRI R$", "% ATING. YTD",
    "META YTD R$", "REALIZADO YTD R$", "DATA ÚLTIMO PEDIDO", "JUSTIFICATIVA SEMANAL",
]

for i, h in enumerate(cart_headers, 1):
    ws_cart.cell(row=1, column=i, value=h)
apply_header(ws_cart, 1, len(cart_headers))

# ── Data Validation (dropdowns) for CARTEIRA ──
# C: UF
dv_uf = DataValidation(type="list", formula1="LISTA_UF", allow_blank=True)
ws_cart.add_data_validation(dv_uf)
dv_uf.add("C2:C1000")

# D: CONSULTOR
dv_consultor = DataValidation(type="list", formula1="LISTA_CONSULTOR", allow_blank=True)
dv_consultor.prompt = "Selecione o consultor"
ws_cart.add_data_validation(dv_consultor)
dv_consultor.add("D2:D1000")

# H: TIPO CLIENTE (6 opções conforme spec)
dv_tipo_cli = DataValidation(type="list", formula1="LISTA_TIPO_CLIENTE", allow_blank=True)
ws_cart.add_data_validation(dv_tipo_cli)
dv_tipo_cli.add("H2:H1000")

# J: TENTATIVA
dv_tent = DataValidation(type="list", formula1="LISTA_TENTATIVA", allow_blank=True)
ws_cart.add_data_validation(dv_tent)
dv_tent.add("J2:J1000")

# W: CURVA ABC
dv_abc = DataValidation(type="list", formula1="LISTA_CURVA_ABC", allow_blank=True)
ws_cart.add_data_validation(dv_abc)
dv_abc.add("W2:W1000")

# AT: ACESSO B2B
dv_b2b = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_cart.add_data_validation(dv_b2b)
dv_b2b.add("AT2:AT1000")

# AX: OPORTUNIDADE ECOM
dv_ecom = DataValidation(type="list", formula1='"CONVERTIDO,QUENTE,MORNO,FRIO"', allow_blank=True)
ws_cart.add_data_validation(dv_ecom)
dv_ecom.add("AX2:AX1000")

# BK: BLOCO
dv_bloco = DataValidation(type="list", formula1="LISTA_BLOCO", allow_blank=True)
ws_cart.add_data_validation(dv_bloco)
dv_bloco.add("BK2:BK1000")

# BO: ÚLTIMO RESULTADO
dv_ult_res = DataValidation(type="list", formula1="LISTA_RESULTADO", allow_blank=True)
ws_cart.add_data_validation(dv_ult_res)
dv_ult_res.add("BO2:BO1000")

# BP: MOTIVO
dv_motivo_cart = DataValidation(type="list", formula1="LISTA_MOTIVO", allow_blank=True)
ws_cart.add_data_validation(dv_motivo_cart)
dv_motivo_cart.add("BP2:BP1000")


# ── 5 rows of sample data ──
# Helper to set date
def d(s):
    """Parse DD/MM/YYYY string to datetime."""
    if not s:
        return None
    parts = s.split('/')
    return datetime.date(int(parts[2]), int(parts[1]), int(parts[0]))


# Row data: col_index (1-based) -> value
# We'll set static values and then apply formulas for calculated columns

sample_rows = [
    {  # Row 2: DIVINA TERRA CWB
        1: "DIVINA TERRA CWB", 2: "DIVINA TERRA", 3: "PR", 4: "DAIANE",
        8: "EM DESENVOLVIMENTO", 10: "T1",
        11: "32828171000108", 12: "DIVINA TERRA LTDA", 13: "CURITIBA",
        14: "divina@email.com", 15: "5541999990001", 16: "SAP_ATIVO", 17: d("01/03/2025"),
        18: 45, 19: d("27/01/2026"), 20: 1850.00, 21: "MANU", 22: 8, 23: "A",
        25: 10,
        # Timeline: AB-AL (cols 28-38)
        28: 0, 29: 500, 30: 0, 31: 1200, 32: 0, 33: 2500, 34: 3000, 35: 1500, 36: 2800, 37: 1500,
        38: 0,  # JAN/26 = will compute from total
        39: "JAN/26,DEZ/25,NOV/25",
        # Jornada
        41: d("01/06/2025"), 42: None, 43: d("15/07/2025"), 44: 5, 45: None,
        # Ecommerce
        46: "SIM", 47: 12, 48: 3, 49: 850, 50: "QUENTE",
        # SAP
        51: "1000105451", 52: "ATIVO", 53: "Rua das Flores 100", 54: "80000100",
        55: "30DDL", 56: "LISTA1", 57: "VAREJO", 58: "", 59: "", 60: "", 61: "",
        # Operacional
        63: "MANHÃ", 64: "CS/RECOMPRA", 65: "Ligar oferecer mix",
        66: d("05/02/2026"), 67: "RELACIONAMENTO", 68: "",
        70: "Em dia",
        # Comitê
        72: 15000, 73: 12500, 75: 45000, 76: 38000, 78: 135000, 79: 98000,
        80: None,  # DATA ÚLTIMO PEDIDO = formula =S2
        81: "Em dia",
    },
    {  # Row 3: ALICE GRAOS
        1: "ALICE GRAOS", 2: "TUDO EM GRAOS", 3: "RS", 4: "MANU",
        8: "RECORRENTE", 10: "T1",
        11: "41626544000140", 12: "ALICE GRAOS LTDA", 13: "PORTO ALEGRE",
        14: "alice@email.com", 15: "5551999990002", 16: "SAP_ATIVO", 17: d("15/01/2025"),
        18: 31, 19: d("17/12/2025"), 20: 920.00, 21: "LARISSA", 22: 5, 23: "B",
        25: 6,
        28: 0, 29: 0, 30: 0, 31: 800, 32: 1200, 33: 1500, 34: 1800, 35: 1200, 36: 1500, 37: 1808, 38: 0,
        39: "DEZ/25,NOV/25,OUT/25",
        41: d("10/03/2025"), 42: None, 43: d("20/04/2025"), 44: 3, 45: None,
        46: "NÃO", 47: 0, 48: 0, 49: 0, 50: "",
        51: "1000105678", 52: "ATIVO", 53: "Rua Principal 50", 54: "90000200",
        55: "28DDL", 56: "LISTA2", 57: "VAREJO", 58: "", 59: "", 60: "", 61: "",
        63: "MANHÃ", 64: "ATENÇÃO/SALVAR", 65: "WhatsApp recompra",
        66: d("03/02/2026"), 67: "NÃO RESPONDE", 68: "AINDA TEM ESTOQUE",
        70: "Estoque parado",
        72: 8000, 73: 4500, 75: 24000, 76: 18000, 78: 72000, 79: 45000,
        81: "Estoque parado",
    },
    {  # Row 4: BIO SC FLORIPA
        1: "BIO SC FLORIPA", 2: "BIOMUNDO", 3: "SC", 4: "DAIANE",
        8: "RECORRENTE", 10: "T2",
        11: "46945954000178", 12: "BIO SC LTDA", 13: "FLORIANOPOLIS",
        14: "biosc@email.com", 15: "5548999990003", 16: "SAP_INATIVO", 17: d("20/06/2024"),
        18: 60, 19: d("01/11/2025"), 20: 680.00, 21: "MANU", 22: 3, 23: "C",
        25: 5,
        28: 0, 29: 0, 30: 0, 31: 0, 32: 0, 33: 0, 34: 2500, 35: 1800, 36: 1500, 37: 1029, 38: 0,
        39: "JAN/26,DEZ/25",
        41: d("15/08/2025"), 42: None, 43: d("10/09/2025"), 44: 6, 45: None,
        46: "NÃO", 47: 0, 48: 0, 49: 0, 50: "",
        51: "1000101541", 52: "ATIVO", 53: "Rua do Sol 200", 54: "88000300",
        55: "30DDL", 56: "LISTA1", 57: "VAREJO", 58: "", 59: "", 60: "", 61: "",
        63: "TARDE", 64: "PERDA/NUTRIÇÃO", 65: "Ligar investigar motivo",
        66: d("01/02/2026"), 67: "RECUSOU LIGAÇÃO", 68: "PREÇO ALTO / MARGEM",
        70: "Reclamou preço",
        72: 5000, 73: 2000, 75: 15000, 76: 8000, 78: 45000, 79: 20000,
        81: "Reclamou preço",
    },
    {  # Row 5: AFS ALDEIA (PROSPECT)
        1: "AFS ALDEIA", 2: "ARMAZEM FITSTORE", 3: "SP", 4: "DAIANE",
        8: "PROSPECT", 10: "T1",
        11: "39915172000120", 12: "AFS ALDEIA LTDA", 13: "SAO PAULO",
        14: "afs@email.com", 15: "5511999990004", 16: "PROSPECT", 17: d("10/01/2026"),
        18: None, 19: None, 20: None, 21: "", 22: 0, 23: "",
        25: 0,
        28: 0, 29: 0, 30: 0, 31: 0, 32: 0, 33: 0, 34: 0, 35: 0, 36: 0, 37: 0, 38: 0,
        39: "",
        41: d("10/01/2026"), 42: None, 43: None, 44: None, 45: None,
        46: "NÃO", 47: 0, 48: 0, 49: 0, 50: "",
        51: "", 52: "", 53: "", 54: "", 55: "", 56: "", 57: "", 58: "", 59: "", 60: "", 61: "",
        63: "TARDE", 64: "ABERTURA/ATIVAÇÃO", 65: "1º contato WhatsApp",
        66: d("05/02/2026"), 67: "NÃO RESPONDE", 68: "1º CONTATO / SEM MOTIVO",
        70: "Aguardando retorno",
        72: 0, 73: 0, 75: 0, 76: 0, 78: 0, 79: 0,
        81: "Aguardando retorno",
    },
    {  # Row 6: CIA SAUDE SP
        1: "CIA SAUDE SP", 2: "CIA DA SAUDE", 3: "SP", 4: "JULIO",
        8: "EM DESENVOLVIMENTO", 10: "T1",
        11: "12345678000199", 12: "CIA SAUDE SP LTDA", 13: "SAO PAULO",
        14: "ciasaude@email.com", 15: "5511999990005", 16: "SAP_ATIVO", 17: d("01/02/2025"),
        18: 30, 19: d("30/01/2026"), 20: 2200.00, 21: "JULIO", 22: 12, 23: "A",
        25: 11,
        28: 1500, 29: 2000, 30: 2500, 31: 2000, 32: 2500, 33: 2800, 34: 3000, 35: 2500, 36: 3200, 37: 3500, 38: 0,
        39: "JAN/26,DEZ/25,NOV/25",
        41: d("15/03/2025"), 42: d("20/03/2025"), 43: d("10/04/2025"), 44: 4, 45: d("15/01/2026"),
        46: "SIM", 47: 25, 48: 8, 49: 1500, 50: "CONVERTIDO",
        51: "1000106000", 52: "ATIVO", 53: "Av Paulista 500", 54: "01000100",
        55: "30DDL", 56: "LISTA1", 57: "REDE", 58: "", 59: "", 60: "", 61: "",
        63: "MANHÃ", 64: "CS/RECOMPRA", 65: "Ligar confirmar pedido",
        66: d("06/02/2026"), 67: "VENDA / PEDIDO", 68: "",
        70: "Cliente top",
        72: 20000, 73: 18500, 75: 60000, 76: 52000, 78: 180000, 79: 155000,
        81: "Cliente top",
    },
]

# Write static data
for row_idx, row_data in enumerate(sample_rows, 2):
    for col_idx, val in row_data.items():
        cell = ws_cart.cell(row=row_idx, column=col_idx, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER

# Apply formulas for rows 2-6
for r in range(2, 7):
    # E: SITUAÇÃO (inclui LEAD/NOVO conforme spec)
    ws_cart.cell(row=r, column=5).value = f'=IF(S{r}="",IF(H{r}="LEAD","LEAD",IF(H{r}="NOVO","NOVO","PROSPECT")),IF(F{r}<=50,"ATIVO",IF(F{r}<=60,"EM RISCO",IF(F{r}<=90,"INATIVO RECENTE","INATIVO ANTIGO"))))'

    # F: DIAS SEM COMPRA
    ws_cart.cell(row=r, column=6).value = f'=IF(S{r}="","",INT(TODAY()-S{r}))'

    # G: SINALEIRO
    ws_cart.cell(row=r, column=7).value = f'=IF(S{r}="","🟣",IF(F{r}<=R{r},"🟢",IF(F{r}<=R{r}+30,"🟡","🔴")))'

    # I: FASE
    ws_cart.cell(row=r, column=9).value = f'=IF(BO{r}="ORÇAMENTO","ORÇAMENTO",IF(BO{r}="EM ATENDIMENTO","EM ATENDIMENTO",IF(S{r}="","PROSPECÇÃO",IF(F{r}<=10,"PÓS-VENDA",IF(F{r}<=25,"CS",IF(F{r}<=R{r},"RECOMPRA",IF(F{r}<=90,"SALVAMENTO","RECUPERAÇÃO")))))))'

    # X: TICKET MÉDIO (col 24)
    ws_cart.cell(row=r, column=24).value = f'=IF(V{r}>0,AA{r}/V{r},0)'

    # Z: MÉDIA MENSAL (col 26)
    ws_cart.cell(row=r, column=26).value = f'=IF(Y{r}>0,AA{r}/Y{r},0)'

    # AA: TOTAL PERÍODO (col 27) = SUM(AB:AL) = cols 28-38
    ws_cart.cell(row=r, column=27).value = f'=SUM(AB{r}:AL{r})'

    # AN: DIAS ATÉ CONVERSÃO (col 40)
    ws_cart.cell(row=r, column=40).value = f'=IF(AND(AO{r}<>"",AQ{r}<>""),AQ{r}-AO{r},"")'

    # BJ: PRÓX. FOLLOW-UP (col 62) — based on last attendance + resultado days
    ws_cart.cell(row=r, column=62).value = f'=IF(BN{r}="","",BN{r}+IFERROR(VLOOKUP(BO{r},TABELA_RESULTADO,2,FALSE),7))'

    # BQ: PRIORIDADE (col 69)
    ws_cart.cell(row=r, column=69).value = f'=IF(W{r}="A",30,IF(W{r}="B",15,5))+IF(G{r}="🔴",40,IF(G{r}="🟡",25,IF(G{r}="🟢",10,5)))+IF(OR(I{r}="RECOMPRA",I{r}="SALVAMENTO"),25,IF(I{r}="CS",20,IF(I{r}="PROSPECÇÃO",10,5)))'

    # BS: % ATING. MÊS (col 71)
    ws_cart.cell(row=r, column=71).value = f'=IF(BT{r}>0,BU{r}/BT{r},0)'

    # BV: % ATING. TRI (col 74)
    ws_cart.cell(row=r, column=74).value = f'=IF(BW{r}>0,BX{r}/BW{r},0)'

    # BY: % ATING. YTD (col 77)
    ws_cart.cell(row=r, column=77).value = f'=IF(BZ{r}>0,CA{r}/BZ{r},0)'

    # CB: DATA ÚLTIMO PEDIDO (col 80) = S
    ws_cart.cell(row=r, column=80).value = f'=S{r}'


# ── Number formats ──
for r in range(2, 7):
    # Money columns: T(20), X(24), Z(26), AA(27), AB-AL(28-38), AW(49), BT(72), BU(73), BW(75), BX(76), BZ(78), CA(79)
    money_cols = [20, 24, 26, 27] + list(range(28, 39)) + [49, 72, 73, 75, 76, 78, 79]
    for c in money_cols:
        ws_cart.cell(row=r, column=c).number_format = '#,##0.00'

    # Date columns: Q(17), S(19), AO(41), AP(42), AQ(43), AS(45), BJ(62), BN(66), CB(80)
    date_cols = [17, 19, 41, 42, 43, 45, 62, 66, 80]
    for c in date_cols:
        ws_cart.cell(row=r, column=c).number_format = 'DD/MM/YYYY'

    # Percentage columns: BS(71), BV(74), BY(77)
    pct_cols = [71, 74, 77]
    for c in pct_cols:
        ws_cart.cell(row=r, column=c).number_format = '0%'

    # Integer: F(6), V(22), Y(25), AR(44), AU(47), AV(48), BQ(69)
    int_cols = [6, 22, 25, 44, 47, 48, 69]
    for c in int_cols:
        ws_cart.cell(row=r, column=c).number_format = '0'


# ── Conditional formatting: SITUAÇÃO (col E) ──
from openpyxl.formatting.rule import FormulaRule

ws_cart.conditional_formatting.add('E2:E1000',
    CellIsRule(operator='equal', formula=['"ATIVO"'],
              fill=FILL_GREEN, font=Font(name='Arial', size=10, color='FFFFFF')))
ws_cart.conditional_formatting.add('E2:E1000',
    CellIsRule(operator='equal', formula=['"EM RISCO"'],
              fill=FILL_YELLOW, font=Font(name='Arial', size=10)))
ws_cart.conditional_formatting.add('E2:E1000',
    CellIsRule(operator='equal', formula=['"INATIVO RECENTE"'],
              fill=FILL_ORANGE, font=Font(name='Arial', size=10, color='FFFFFF')))
ws_cart.conditional_formatting.add('E2:E1000',
    CellIsRule(operator='equal', formula=['"INATIVO ANTIGO"'],
              fill=FILL_RED, font=Font(name='Arial', size=10, color='FFFFFF')))
ws_cart.conditional_formatting.add('E2:E1000',
    CellIsRule(operator='equal', formula=['"PROSPECT"'],
              fill=FILL_BLUE_LIGHT, font=Font(name='Arial', size=10)))
ws_cart.conditional_formatting.add('E2:E1000',
    CellIsRule(operator='equal', formula=['"LEAD"'],
              fill=FILL_BLUE_LIGHT, font=Font(name='Arial', size=10)))
ws_cart.conditional_formatting.add('E2:E1000',
    CellIsRule(operator='equal', formula=['"NOVO"'],
              fill=PatternFill('solid', fgColor='C6EFCE'), font=Font(name='Arial', size=10)))

# ── Conditional formatting: % ATING. MÊS (col BS = 71) ──
ws_cart.conditional_formatting.add('BS2:BS1000',
    CellIsRule(operator='greaterThanOrEqual', formula=['0.8'], fill=FILL_GREEN))
ws_cart.conditional_formatting.add('BS2:BS1000',
    CellIsRule(operator='between', formula=['0.5', '0.7999'], fill=FILL_YELLOW))
ws_cart.conditional_formatting.add('BS2:BS1000',
    CellIsRule(operator='lessThan', formula=['0.5'], fill=FILL_RED))


# ── Column grouping (outline) ──
# openpyxl .group() only sets outlineLevel on the first column of the range.
# We must set each column individually for proper hiding.
def group_columns(ws, start_col, end_col, level=1, hidden=True):
    """Apply outline grouping to each column in range individually."""
    for col_idx in range(start_col, end_col + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].outlineLevel = level
        ws.column_dimensions[letter].hidden = hidden

# [+1] L:Q (cols 12-17), K stays visible
group_columns(ws_cart, 12, 17)
# [+2] S:Z (cols 19-26), R stays visible
group_columns(ws_cart, 19, 26)
# [+3] AB:AM (cols 28-39), AA stays visible
group_columns(ws_cart, 28, 39)
# [+4] AO:AS (cols 41-45), AN stays visible
group_columns(ws_cart, 41, 45)
# [+5] AU:AX (cols 47-50), AT stays visible
group_columns(ws_cart, 47, 50)
# [+6] AZ:BI (cols 52-61), AY stays visible
group_columns(ws_cart, 52, 61)
# [+7] BK:BR (cols 63-70), BJ stays visible
group_columns(ws_cart, 63, 70)
# [+8] BT:CC (cols 72-81), BS stays visible
group_columns(ws_cart, 72, 81)

# Freeze panes at K2
ws_cart.freeze_panes = "K2"

# Auto width for visible columns
for col_idx in range(1, 82):
    letter = get_column_letter(col_idx)
    ws_cart.column_dimensions[letter].width = 14
# Wider columns
ws_cart.column_dimensions['A'].width = 22
ws_cart.column_dimensions['B'].width = 18
ws_cart.column_dimensions['D'].width = 16
ws_cart.column_dimensions['E'].width = 18
ws_cart.column_dimensions['I'].width = 16
ws_cart.column_dimensions['K'].width = 18
ws_cart.column_dimensions['L'].width = 22
ws_cart.column_dimensions['N'].width = 22
ws_cart.column_dimensions['BA'].width = 22
ws_cart.column_dimensions['BM'].width = 25
ws_cart.column_dimensions['BR'].width = 20
ws_cart.column_dimensions['CC'].width = 22

ws_cart.sheet_view.zoomScale = 90
print("  ABA 2 CARTEIRA done.")


# ═══════════════════════════════════════════════════════════════
# ABA 3: LOG (20 columns)
# ═══════════════════════════════════════════════════════════════
print("Building ABA 3: LOG...")
ws_log = wb.create_sheet("LOG")
ws_log.sheet_properties.tabColor = "FFC000"

log_headers = [
    "DATA", "CONSULTOR", "NOME FANTASIA", "CNPJ", "UF", "REDE / REGIONAL",
    "SITUAÇÃO", "WHATSAPP", "LIGAÇÃO", "LIGAÇÃO ATENDIDA", "TIPO AÇÃO",
    "TIPO DO CONTATO", "RESULTADO", "MOTIVO", "FOLLOW-UP", "AÇÃO",
    "MERCOS ATUALIZADO", "FASE", "TENTATIVA", "NOTA DO DIA"
]
for i, h in enumerate(log_headers, 1):
    ws_log.cell(row=1, column=i, value=h)
apply_header(ws_log, 1, 20)

# Dropdowns for LOG
dv_log_cons = DataValidation(type="list", formula1="LISTA_CONSULTOR", allow_blank=True)
ws_log.add_data_validation(dv_log_cons)
dv_log_cons.add("B2:B1000")

dv_log_sn_h = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_log.add_data_validation(dv_log_sn_h)
dv_log_sn_h.add("H2:H1000")

dv_log_sn_i = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_log.add_data_validation(dv_log_sn_i)
dv_log_sn_i.add("I2:I1000")

dv_log_sn_na = DataValidation(type="list", formula1="LISTA_SIM_NAO_NA", allow_blank=True)
ws_log.add_data_validation(dv_log_sn_na)
dv_log_sn_na.add("J2:J1000")

dv_log_ar = DataValidation(type="list", formula1="LISTA_ATIVO_RECEPTIVO", allow_blank=True)
ws_log.add_data_validation(dv_log_ar)
dv_log_ar.add("K2:K1000")

dv_log_tc = DataValidation(type="list", formula1="LISTA_TIPO_CONTATO", allow_blank=True)
ws_log.add_data_validation(dv_log_tc)
dv_log_tc.add("L2:L1000")

dv_log_res = DataValidation(type="list", formula1="LISTA_RESULTADO", allow_blank=True)
ws_log.add_data_validation(dv_log_res)
dv_log_res.add("M2:M1000")

dv_log_mot = DataValidation(type="list", formula1="LISTA_MOTIVO", allow_blank=True)
ws_log.add_data_validation(dv_log_mot)
dv_log_mot.add("N2:N1000")

dv_log_mercos = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_log.add_data_validation(dv_log_mercos)
dv_log_mercos.add("Q2:Q1000")

# P: AÇÃO (dropdown com ações futuras)
dv_log_acao = DataValidation(type="list", formula1="LISTA_ACAO_FUTURA", allow_blank=True)
ws_log.add_data_validation(dv_log_acao)
dv_log_acao.add("P2:P1000")

# 8 sample rows for LOG
log_data = [
    [d("05/02/2026"), "DAIANE", "DIVINA TERRA CWB", "32828171000108", "PR", "DIVINA TERRA", "ATIVO", "SIM", "NÃO", "N/A", "ATIVO", "PÓS-VENDA / RELACIONAMENTO", "RELACIONAMENTO", "AINDA TEM ESTOQUE", None, "Verificar estoque", "SIM", "CS", "T1", "Cliente satisfeito"],
    [d("05/02/2026"), "MANU", "ALICE GRAOS", "41626544000140", "RS", "TUDO EM GRAOS", "EM RISCO", "SIM", "NÃO", "N/A", "ATIVO", "ATEND. CLIENTES ATIVOS", "NÃO RESPONDE", "1º CONTATO / SEM MOTIVO", None, "Tentar ligar amanhã", "NÃO", "RECOMPRA", "T1", "Sem resposta WhatsApp"],
    [d("05/02/2026"), "DAIANE", "BIO SC FLORIPA", "46945954000178", "SC", "BIOMUNDO", "INATIVO RECENTE", "SIM", "SIM", "NÃO", "ATIVO", "ATEND. CLIENTES INATIVOS", "RECUSOU LIGAÇÃO", "PREÇO ALTO / MARGEM", None, "Enviar tabela nova", "SIM", "SALVAMENTO", "T2", "Reclamou preço"],
    [d("05/02/2026"), "DAIANE", "AFS ALDEIA", "39915172000120", "SP", "ARMAZEM FITSTORE", "PROSPECT", "SIM", "NÃO", "N/A", "ATIVO", "PROSPECÇÃO", "NÃO RESPONDE", "1º CONTATO / SEM MOTIVO", None, "1º contato feito", "NÃO", "PROSPECÇÃO", "T1", "WhatsApp enviado"],
    [d("06/02/2026"), "JULIO", "CIA SAUDE SP", "12345678000199", "SP", "CIA DA SAUDE", "ATIVO", "NÃO", "SIM", "SIM", "ATIVO", "ATEND. CLIENTES ATIVOS", "VENDA / PEDIDO", "", None, "Pedido confirmado", "SIM", "CS", "T1", "Pedido R$ 2.200"],
    [d("06/02/2026"), "MANU", "ALICE GRAOS", "41626544000140", "RS", "TUDO EM GRAOS", "EM RISCO", "NÃO", "SIM", "NÃO", "ATIVO", "FOLLOW UP", "NÃO ATENDE", "AINDA TEM ESTOQUE", None, "Tentar WhatsApp", "SIM", "RECOMPRA", "T2", "Não atendeu"],
    [d("06/02/2026"), "DAIANE", "AFS ALDEIA", "39915172000120", "SP", "ARMAZEM FITSTORE", "PROSPECT", "SIM", "NÃO", "N/A", "ATIVO", "PROSPECÇÃO", "EM ATENDIMENTO", "1º CONTATO / SEM MOTIVO", None, "Respondeu interesse", "SIM", "PROSPECÇÃO", "T1", "Pediu catálogo"],
    [d("07/02/2026"), "DAIANE", "DIVINA TERRA CWB", "32828171000108", "PR", "DIVINA TERRA", "ATIVO", "NÃO", "SIM", "SIM", "RECEPTIVO", "NEGOCIAÇÃO", "ORÇAMENTO", "", None, "Enviar orçamento mix", "SIM", "CS", "T1", "Pediu orçamento novo mix"],
]

for r_idx, row_vals in enumerate(log_data, 2):
    for c_idx, val in enumerate(row_vals, 1):
        cell = ws_log.cell(row=r_idx, column=c_idx, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER

    # FOLLOW-UP formula (col O = 15): SUPORTE e PERDA = "SEM", rest = DATA + dias
    ws_log.cell(row=r_idx, column=15).value = f'=IF(M{r_idx}="","",IF(OR(M{r_idx}="SUPORTE",M{r_idx}="PERDA / FECHOU LOJA"),"SEM",IFERROR(A{r_idx}+VLOOKUP(M{r_idx},TABELA_RESULTADO,2,FALSE),"")))'
    ws_log.cell(row=r_idx, column=15).number_format = 'DD/MM/YYYY'

    # Format date columns
    ws_log.cell(row=r_idx, column=1).number_format = 'DD/MM/YYYY'

# Freeze panes
ws_log.freeze_panes = "A2"

# Auto width
for col_idx in range(1, 21):
    letter = get_column_letter(col_idx)
    ws_log.column_dimensions[letter].width = 18
ws_log.column_dimensions['C'].width = 22
ws_log.column_dimensions['D'].width = 18
ws_log.column_dimensions['L'].width = 28
ws_log.column_dimensions['M'].width = 22
ws_log.column_dimensions['N'].width = 28
ws_log.column_dimensions['P'].width = 25
ws_log.column_dimensions['T'].width = 25

ws_log.sheet_view.zoomScale = 90
print("  ABA 3 LOG done.")


# ═══════════════════════════════════════════════════════════════
# ABA 4: DRAFT 1 (Quarentena Mercos)
# ═══════════════════════════════════════════════════════════════
print("Building ABA 4: DRAFT 1...")
ws_d1 = wb.create_sheet("DRAFT 1")
ws_d1.sheet_properties.tabColor = "FF6600"

d1_headers = ["DATA", "CNPJ", "NOME FANTASIA", "VALOR PEDIDO", "STATUS PEDIDO", "VENDEDOR", "MIGRADO"]
for i, h in enumerate(d1_headers, 1):
    ws_d1.cell(row=1, column=i, value=h)
apply_header(ws_d1, 1, 7)

# Dropdown for MIGRADO
dv_d1_mig = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_d1.add_data_validation(dv_d1_mig)
dv_d1_mig.add("G2:G1000")

# 2 sample rows
d1_data = [
    [d("06/02/2026"), "12345678000199", "CIA SAUDE SP", 2200.00, "CONFIRMADO", "JULIO", "NÃO"],
    [d("07/02/2026"), "32828171000108", "DIVINA TERRA CWB", 1850.00, "PENDENTE", "MANU", "NÃO"],
]
for r_idx, row_vals in enumerate(d1_data, 2):
    for c_idx, val in enumerate(row_vals, 1):
        cell = ws_d1.cell(row=r_idx, column=c_idx, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER
    ws_d1.cell(row=r_idx, column=1).number_format = 'DD/MM/YYYY'
    ws_d1.cell(row=r_idx, column=4).number_format = '#,##0.00'

ws_d1.freeze_panes = "A2"
auto_width(ws_d1, max_col=7)
ws_d1.sheet_view.zoomScale = 90
print("  ABA 4 DRAFT 1 done.")


# ═══════════════════════════════════════════════════════════════
# ABA 5: DRAFT 2 (Quarentena Atendimentos)
# ═══════════════════════════════════════════════════════════════
print("Building ABA 5: DRAFT 2...")
ws_d2 = wb.create_sheet("DRAFT 2")
ws_d2.sheet_properties.tabColor = "BDD7EE"

d2_headers = [
    "NOME FANTASIA", "CNPJ", "UF", "REDE/REGIONAL", "TELEFONE", "EMAIL",
    "SITUAÇÃO", "DIAS SEM COMPRA", "SINALEIRO", "FASE", "TENTATIVA",
    "AÇÃO SUGERIDA", "BLOCO", "ÚLTIMO RESULTADO",
    "WHATSAPP", "LIGAÇÃO", "LIGAÇÃO ATENDIDA", "TIPO AÇÃO",
    "TIPO DO CONTATO", "RESULTADO", "MOTIVO", "AÇÃO FUTURA",
    "MERCOS ATUALIZADO", "NOTA DO DIA",
    "✅ VÁLIDO", "⚠ ERRO", "📝 MIGRADO"
]
for i, h in enumerate(d2_headers, 1):
    ws_d2.cell(row=1, column=i, value=h)
apply_header(ws_d2, 1, 27)

# Dropdowns for DRAFT 2
dv_d2_wpp = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_d2.add_data_validation(dv_d2_wpp)
dv_d2_wpp.add("O2:O1000")

dv_d2_lig = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_d2.add_data_validation(dv_d2_lig)
dv_d2_lig.add("P2:P1000")

dv_d2_la = DataValidation(type="list", formula1="LISTA_SIM_NAO_NA", allow_blank=True)
ws_d2.add_data_validation(dv_d2_la)
dv_d2_la.add("Q2:Q1000")

dv_d2_ar = DataValidation(type="list", formula1="LISTA_ATIVO_RECEPTIVO", allow_blank=True)
ws_d2.add_data_validation(dv_d2_ar)
dv_d2_ar.add("R2:R1000")

dv_d2_tc = DataValidation(type="list", formula1="LISTA_TIPO_CONTATO", allow_blank=True)
ws_d2.add_data_validation(dv_d2_tc)
dv_d2_tc.add("S2:S1000")

dv_d2_res = DataValidation(type="list", formula1="LISTA_RESULTADO", allow_blank=True)
ws_d2.add_data_validation(dv_d2_res)
dv_d2_res.add("T2:T1000")

dv_d2_mot = DataValidation(type="list", formula1="LISTA_MOTIVO", allow_blank=True)
ws_d2.add_data_validation(dv_d2_mot)
dv_d2_mot.add("U2:U1000")

dv_d2_mercos = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_d2.add_data_validation(dv_d2_mercos)
dv_d2_mercos.add("W2:W1000")

dv_d2_mig = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_d2.add_data_validation(dv_d2_mig)
dv_d2_mig.add("AA2:AA1000")

# 3 sample rows (2 OK, 1 ERRO)
d2_data = [
    ["DIVINA TERRA CWB", "32828171000108", "PR", "DIVINA TERRA", "5541999990001", "divina@email.com",
     "ATIVO", 12, "🟢", "CS", "T1", "Ligar oferecer mix", "MANHÃ", "RELACIONAMENTO",
     "SIM", "NÃO", "N/A", "ATIVO", "PÓS-VENDA / RELACIONAMENTO", "RELACIONAMENTO", "AINDA TEM ESTOQUE",
     "Verificar estoque", "SIM", "Cliente satisfeito"],
    ["CIA SAUDE SP", "12345678000199", "SP", "CIA DA SAUDE", "5511999990005", "ciasaude@email.com",
     "ATIVO", 9, "🟢", "PÓS-VENDA", "T1", "Ligar confirmar pedido", "MANHÃ", "VENDA / PEDIDO",
     "NÃO", "SIM", "SIM", "ATIVO", "ATEND. CLIENTES ATIVOS", "VENDA / PEDIDO", "",
     "Pedido confirmado", "SIM", "Pedido R$ 2.200"],
    ["AFS ALDEIA", "", "SP", "ARMAZEM FITSTORE", "5511999990004", "afs@email.com",
     "PROSPECT", "", "🟣", "PROSPECÇÃO", "T1", "1º contato WhatsApp", "TARDE", "NÃO RESPONDE",
     "SIM", "NÃO", "N/A", "ATIVO", "", "", "1º CONTATO / SEM MOTIVO",
     "1º contato feito", "NÃO", "WhatsApp enviado"],
]

for r_idx, row_vals in enumerate(d2_data, 2):
    for c_idx, val in enumerate(row_vals, 1):
        cell = ws_d2.cell(row=r_idx, column=c_idx, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER

    # Y: VÁLIDO (col 25)
    ws_d2.cell(row=r_idx, column=25).value = f'=IF(AND(B{r_idx}<>"",T{r_idx}<>"",S{r_idx}<>""),"OK","ERRO")'
    # Z: ERRO (col 26)
    ws_d2.cell(row=r_idx, column=26).value = f'=IF(Y{r_idx}="OK","",IF(B{r_idx}="","CNPJ VAZIO",IF(T{r_idx}="","SEM RESULTADO","SEM TIPO CONTATO")))'

# Conditional formatting for validation
ws_d2.conditional_formatting.add('Y2:Y1000',
    CellIsRule(operator='equal', formula=['"OK"'], fill=FILL_GREEN_LIGHT))
ws_d2.conditional_formatting.add('Y2:Y1000',
    CellIsRule(operator='equal', formula=['"ERRO"'], fill=FILL_RED_LIGHT))

ws_d2.freeze_panes = "A2"
for col_idx in range(1, 28):
    letter = get_column_letter(col_idx)
    ws_d2.column_dimensions[letter].width = 16
ws_d2.column_dimensions['A'].width = 22
ws_d2.column_dimensions['S'].width = 28
ws_d2.column_dimensions['T'].width = 22
ws_d2.column_dimensions['U'].width = 28

ws_d2.sheet_view.zoomScale = 90
print("  ABA 5 DRAFT 2 done.")


# ═══════════════════════════════════════════════════════════════
# ABA 6: DASH (Dashboard)
# ═══════════════════════════════════════════════════════════════
print("Building ABA 6: DASH...")
ws_dash = wb.create_sheet("DASH")
ws_dash.sheet_properties.tabColor = "4472C4"

# Set column widths
for c in range(1, 15):
    ws_dash.column_dimensions[get_column_letter(c)].width = 18
ws_dash.column_dimensions['A'].width = 30

# ── SEÇÃO 1: FILTROS (rows 1-3) ──
ws_dash.cell(row=1, column=1, value="📊 DASHBOARD JARVIS CRM").font = FONT_TITLE_14
ws_dash.merge_cells('A1:H1')

ws_dash.cell(row=2, column=1, value="CONSULTOR:").font = FONT_BOLD_10
dv_dash_cons = DataValidation(type="list", formula1='"TODOS,MANU DITZEL,LARISSA PADILHA,JULIO GADRET,DAIANE STAVICKI"', allow_blank=True)
ws_dash.add_data_validation(dv_dash_cons)
dv_dash_cons.add("B2")
ws_dash.cell(row=2, column=2, value="TODOS")

ws_dash.cell(row=3, column=1, value="PERÍODO:").font = FONT_BOLD_10
ws_dash.cell(row=3, column=2, value=d("01/01/2026"))
ws_dash.cell(row=3, column=2).number_format = 'DD/MM/YYYY'
ws_dash.cell(row=3, column=3, value="a").font = FONT_DEFAULT
ws_dash.cell(row=3, column=4, value=d("08/02/2026"))
ws_dash.cell(row=3, column=4).number_format = 'DD/MM/YYYY'

# ── SEÇÃO 2: KPIs (rows 5-7) ──
kpi_labels = ["CONTATOS", "VENDAS", "ORÇAMENTOS", "NÃO ATENDE", "% CONVERSÃO", "MERCOS OK"]
for i, lbl in enumerate(kpi_labels, 1):
    cell = ws_dash.cell(row=5, column=i, value=lbl)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER

# KPI formulas — with CONSULTOR filter support via B2
# When B2="TODOS", count all. Otherwise filter by consultor name.
# Using IF($B$2="TODOS",...) pattern with COUNTIFS
# A6: Total contacts
ws_dash.cell(row=6, column=1).value = '=IF($B$2="TODOS",COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3),COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3,LOG!B:B,$B$2))'
# B6: Vendas
ws_dash.cell(row=6, column=2).value = '=IF($B$2="TODOS",COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3,LOG!M:M,"VENDA / PEDIDO"),COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3,LOG!M:M,"VENDA / PEDIDO",LOG!B:B,$B$2))'
# C6: Orçamentos
ws_dash.cell(row=6, column=3).value = '=IF($B$2="TODOS",COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3,LOG!M:M,"ORÇAMENTO"),COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3,LOG!M:M,"ORÇAMENTO",LOG!B:B,$B$2))'
# D6: Não atende
ws_dash.cell(row=6, column=4).value = '=IF($B$2="TODOS",COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3,LOG!M:M,"NÃO ATENDE"),COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3,LOG!M:M,"NÃO ATENDE",LOG!B:B,$B$2))'
# E6: % Conversão
ws_dash.cell(row=6, column=5).value = '=IFERROR(B6/A6,0)'
ws_dash.cell(row=6, column=5).number_format = '0%'
# F6: Mercos OK
ws_dash.cell(row=6, column=6).value = '=IF($B$2="TODOS",IFERROR(COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3,LOG!Q:Q,"SIM")/COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3),0),IFERROR(COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3,LOG!Q:Q,"SIM",LOG!B:B,$B$2)/COUNTIFS(LOG!A:A,">="&B3,LOG!A:A,"<="&D3,LOG!B:B,$B$2),0))'
ws_dash.cell(row=6, column=6).number_format = '0%'

for c in range(1, 7):
    cell = ws_dash.cell(row=6, column=c)
    cell.font = FONT_KPI
    cell.fill = FILL_KPI
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER

# ── SEÇÃO 3: CONTATOS & FUNIL (rows 8-22) ──
row_s3 = 8
# Blank row
row_s3 += 2  # row 10
cell_title = ws_dash.cell(row=row_s3, column=1, value="CONTATOS POR TIPO & FUNIL")
cell_title.font = FONT_WHITE_12
cell_title.fill = FILL_SECTION_TITLE
for c in range(1, 6):
    ws_dash.cell(row=row_s3, column=c).fill = FILL_SECTION_TITLE
    ws_dash.cell(row=row_s3, column=c).font = FONT_WHITE_12

row_s3 += 1  # row 11 = headers
s3_headers = ["TIPO DO CONTATO", "TOTAL", "FUNIL", "RELAC.", "NÃO VENDA"]
for i, h in enumerate(s3_headers, 1):
    cell = ws_dash.cell(row=row_s3, column=i, value=h)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.border = THIN_BORDER

tipos_contato = [
    "PROSPECÇÃO", "NEGOCIAÇÃO", "FOLLOW UP", "ATEND. CLIENTES ATIVOS",
    "ATEND. CLIENTES INATIVOS", "PÓS-VENDA / RELACIONAMENTO", "MOTIVO / PAROU DE COMPRAR"
]
for t_idx, tipo in enumerate(tipos_contato):
    r = row_s3 + 1 + t_idx
    ws_dash.cell(row=r, column=1, value=tipo).font = FONT_DEFAULT
    ws_dash.cell(row=r, column=1).border = THIN_BORDER
    # TOTAL (with consultor filter)
    ws_dash.cell(row=r, column=2).value = f'=IF($B$2="TODOS",COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3),COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!B:B,$B$2))'
    ws_dash.cell(row=r, column=2).border = THIN_BORDER
    # FUNIL
    ws_dash.cell(row=r, column=3).value = f'=IF($B$2="TODOS",COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"EM ATENDIMENTO")+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"ORÇAMENTO")+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"CADASTRO")+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"VENDA / PEDIDO"),COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"EM ATENDIMENTO",LOG!B:B,$B$2)+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"ORÇAMENTO",LOG!B:B,$B$2)+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"CADASTRO",LOG!B:B,$B$2)+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"VENDA / PEDIDO",LOG!B:B,$B$2))'
    ws_dash.cell(row=r, column=3).border = THIN_BORDER
    # RELAC.
    ws_dash.cell(row=r, column=4).value = f'=IF($B$2="TODOS",COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"RELACIONAMENTO")+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"FOLLOW UP 7")+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"FOLLOW UP 15")+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"SUPORTE"),COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"RELACIONAMENTO",LOG!B:B,$B$2)+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"FOLLOW UP 7",LOG!B:B,$B$2)+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"FOLLOW UP 15",LOG!B:B,$B$2)+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"SUPORTE",LOG!B:B,$B$2))'
    ws_dash.cell(row=r, column=4).border = THIN_BORDER
    # NÃO VENDA
    ws_dash.cell(row=r, column=5).value = f'=IF($B$2="TODOS",COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"NÃO ATENDE")+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"NÃO RESPONDE")+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"RECUSOU LIGAÇÃO")+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"PERDA / FECHOU LOJA"),COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"NÃO ATENDE",LOG!B:B,$B$2)+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"NÃO RESPONDE",LOG!B:B,$B$2)+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"RECUSOU LIGAÇÃO",LOG!B:B,$B$2)+COUNTIFS(LOG!L:L,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!M:M,"PERDA / FECHOU LOJA",LOG!B:B,$B$2))'
    ws_dash.cell(row=r, column=5).border = THIN_BORDER

# TOTAL row
r_total = row_s3 + 1 + len(tipos_contato)
ws_dash.cell(row=r_total, column=1, value="TOTAL").font = FONT_BOLD_10
ws_dash.cell(row=r_total, column=1).border = THIN_BORDER
for c in range(2, 6):
    col_letter = get_column_letter(c)
    ws_dash.cell(row=r_total, column=c).value = f'=SUM({col_letter}{row_s3+1}:{col_letter}{r_total-1})'
    ws_dash.cell(row=r_total, column=c).font = FONT_BOLD_10
    ws_dash.cell(row=r_total, column=c).border = THIN_BORDER


# ── SEÇÃO 4: TIPO × RESULTADO (rows 24-36) ──
row_s4 = r_total + 3
cell_title = ws_dash.cell(row=row_s4, column=1, value="CRUZAMENTO TIPO × RESULTADO")
cell_title.font = FONT_WHITE_12
cell_title.fill = FILL_SECTION_TITLE
for c in range(1, 14):
    ws_dash.cell(row=row_s4, column=c).fill = FILL_SECTION_TITLE
    ws_dash.cell(row=row_s4, column=c).font = FONT_WHITE_12

resultados = [
    "EM ATENDIMENTO", "ORÇAMENTO", "CADASTRO", "VENDA / PEDIDO",
    "RELACIONAMENTO", "FOLLOW UP 7", "FOLLOW UP 15", "SUPORTE",
    "NÃO ATENDE", "NÃO RESPONDE", "RECUSOU LIGAÇÃO", "PERDA / FECHOU LOJA"
]

# Headers row
row_s4h = row_s4 + 1
ws_dash.cell(row=row_s4h, column=1, value="TIPO \\ RESULTADO").font = FONT_HEADER
ws_dash.cell(row=row_s4h, column=1).fill = FILL_HEADER
ws_dash.cell(row=row_s4h, column=1).border = THIN_BORDER
for ri, res in enumerate(resultados, 2):
    cell = ws_dash.cell(row=row_s4h, column=ri, value=res)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.border = THIN_BORDER
    ws_dash.column_dimensions[get_column_letter(ri)].width = 16

for t_idx, tipo in enumerate(tipos_contato):
    r = row_s4h + 1 + t_idx
    ws_dash.cell(row=r, column=1, value=tipo).font = FONT_DEFAULT
    ws_dash.cell(row=r, column=1).border = THIN_BORDER
    for ri, res in enumerate(resultados, 2):
        col_l = get_column_letter(ri)
        ws_dash.cell(row=r, column=ri).value = f'=IF($B$2="TODOS",COUNTIFS(LOG!L:L,$A{r},LOG!M:M,{col_l}${row_s4h},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3),COUNTIFS(LOG!L:L,$A{r},LOG!M:M,{col_l}${row_s4h},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!B:B,$B$2))'
        ws_dash.cell(row=r, column=ri).border = THIN_BORDER


# ── SEÇÃO 5: MOTIVOS (rows 38-51) ──
row_s5 = row_s4h + 1 + len(tipos_contato) + 3
cell_title = ws_dash.cell(row=row_s5, column=1, value="📋 POR QUE NÃO COMPRAM — INTELIGÊNCIA DIRETORIA")
cell_title.font = FONT_WHITE_12
cell_title.fill = FILL_SECTION_TITLE
for c in range(1, 5):
    ws_dash.cell(row=row_s5, column=c).fill = FILL_SECTION_TITLE

s5_headers = ["MOTIVO", "QTD", "%", "DONO DA AÇÃO"]
row_s5h = row_s5 + 1
for i, h in enumerate(s5_headers, 1):
    cell = ws_dash.cell(row=row_s5h, column=i, value=h)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.border = THIN_BORDER

motivos_list = [
    ("PRODUTO NÃO VENDEU / SEM GIRO", "DIRETORIA / FÁBRICA"),
    ("PREÇO ALTO / MARGEM", "COMERCIAL"),
    ("PREFERIU CONCORRENTE", "COMERCIAL / MKT"),
    ("PROBLEMA LOGÍSTICO", "LOGÍSTICA"),
    ("PROBLEMA FINANCEIRO", "FINANCEIRO"),
    ("AINDA TEM ESTOQUE", "NORMAL"),
    ("FECHOU LOJA", "PERDA"),
    ("SEM INTERESSE", "NUTRIÇÃO"),
    ("VIAJANDO / INDISPONÍVEL", "NORMAL"),
    ("1º CONTATO / SEM MOTIVO", "NORMAL"),
]
first_mot_row = row_s5h + 1
for m_idx, (mot, dono) in enumerate(motivos_list):
    r = row_s5h + 1 + m_idx
    ws_dash.cell(row=r, column=1, value=mot).font = FONT_DEFAULT
    ws_dash.cell(row=r, column=1).border = THIN_BORDER
    ws_dash.cell(row=r, column=2).value = f'=IF($B$2="TODOS",COUNTIFS(LOG!N:N,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3),COUNTIFS(LOG!N:N,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!B:B,$B$2))'
    ws_dash.cell(row=r, column=2).border = THIN_BORDER
    last_mot_row = r
    ws_dash.cell(row=r, column=4, value=dono).font = FONT_DEFAULT
    ws_dash.cell(row=r, column=4).border = THIN_BORDER

# % column — need to compute after data rows
for m_idx in range(len(motivos_list)):
    r = first_mot_row + m_idx
    ws_dash.cell(row=r, column=3).value = f'=IFERROR(B{r}/SUM($B${first_mot_row}:$B${last_mot_row}),0)'
    ws_dash.cell(row=r, column=3).number_format = '0%'
    ws_dash.cell(row=r, column=3).border = THIN_BORDER

# TOTAL row
r_mot_total = last_mot_row + 1
ws_dash.cell(row=r_mot_total, column=1, value="TOTAL").font = FONT_BOLD_10
ws_dash.cell(row=r_mot_total, column=1).border = THIN_BORDER
ws_dash.cell(row=r_mot_total, column=2).value = f'=SUM(B{first_mot_row}:B{last_mot_row})'
ws_dash.cell(row=r_mot_total, column=2).font = FONT_BOLD_10
ws_dash.cell(row=r_mot_total, column=2).border = THIN_BORDER
ws_dash.cell(row=r_mot_total, column=3, value=1).number_format = '0%'
ws_dash.cell(row=r_mot_total, column=3).font = FONT_BOLD_10
ws_dash.cell(row=r_mot_total, column=3).border = THIN_BORDER

# Conditional formatting: if PRODUTO NÃO VENDEU > 35% → red
ws_dash.conditional_formatting.add(f'C{first_mot_row}',
    CellIsRule(operator='greaterThan', formula=['0.35'], fill=FILL_RED,
              font=Font(name='Arial', size=10, color='FFFFFF')))


# ── SEÇÃO 6: FORMA DO CONTATO (rows 53-58) ──
row_s6 = r_mot_total + 3
cell_title = ws_dash.cell(row=row_s6, column=1, value="FORMA DO CONTATO")
cell_title.font = FONT_WHITE_12
cell_title.fill = FILL_SECTION_TITLE
for c in range(1, 4):
    ws_dash.cell(row=row_s6, column=c).fill = FILL_SECTION_TITLE

s6_headers = ["FORMA", "QTD", "%"]
row_s6h = row_s6 + 1
for i, h in enumerate(s6_headers, 1):
    cell = ws_dash.cell(row=row_s6h, column=i, value=h)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.border = THIN_BORDER

formas = [
    ("Só WhatsApp", '=IF($B$2="TODOS",COUNTIFS(LOG!H:H,"SIM",LOG!I:I,"NÃO",LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3),COUNTIFS(LOG!H:H,"SIM",LOG!I:I,"NÃO",LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!B:B,$B$2))'),
    ("Só Ligação", '=IF($B$2="TODOS",COUNTIFS(LOG!H:H,"NÃO",LOG!I:I,"SIM",LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3),COUNTIFS(LOG!H:H,"NÃO",LOG!I:I,"SIM",LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!B:B,$B$2))'),
    ("WhatsApp + Ligação", '=IF($B$2="TODOS",COUNTIFS(LOG!H:H,"SIM",LOG!I:I,"SIM",LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3),COUNTIFS(LOG!H:H,"SIM",LOG!I:I,"SIM",LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3,LOG!B:B,$B$2))'),
]
first_forma_row = row_s6h + 1
for f_idx, (forma, formula) in enumerate(formas):
    r = first_forma_row + f_idx
    ws_dash.cell(row=r, column=1, value=forma).font = FONT_DEFAULT
    ws_dash.cell(row=r, column=1).border = THIN_BORDER
    ws_dash.cell(row=r, column=2).value = formula
    ws_dash.cell(row=r, column=2).border = THIN_BORDER
last_forma_row = first_forma_row + len(formas) - 1
for f_idx in range(len(formas)):
    r = first_forma_row + f_idx
    ws_dash.cell(row=r, column=3).value = f'=IFERROR(B{r}/SUM($B${first_forma_row}:$B${last_forma_row}),0)'
    ws_dash.cell(row=r, column=3).number_format = '0%'
    ws_dash.cell(row=r, column=3).border = THIN_BORDER


# ── SEÇÃO 7: PRODUTIVIDADE POR CONSULTOR (rows 60-67) ──
row_s7 = last_forma_row + 3
cell_title = ws_dash.cell(row=row_s7, column=1, value="PRODUTIVIDADE POR CONSULTOR")
cell_title.font = FONT_WHITE_12
cell_title.fill = FILL_SECTION_TITLE
for c in range(1, 7):
    ws_dash.cell(row=row_s7, column=c).fill = FILL_SECTION_TITLE

s7_headers = ["CONSULTOR", "CONTATOS", "VENDAS", "ORÇAMENTOS", "% CONVERSÃO", "MÉDIA/DIA"]
row_s7h = row_s7 + 1
for i, h in enumerate(s7_headers, 1):
    cell = ws_dash.cell(row=row_s7h, column=i, value=h)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.border = THIN_BORDER

consultores = ["MANU DITZEL", "LARISSA PADILHA", "JULIO GADRET", "DAIANE STAVICKI"]
first_prod_row = row_s7h + 1
for c_idx, cons in enumerate(consultores):
    r = first_prod_row + c_idx
    ws_dash.cell(row=r, column=1, value=cons).font = FONT_DEFAULT
    ws_dash.cell(row=r, column=1).border = THIN_BORDER
    # Contatos (seção produtividade sempre mostra todos os consultores, sem filtro B2)
    ws_dash.cell(row=r, column=2).value = f'=COUNTIFS(LOG!B:B,A{r},LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3)'
    ws_dash.cell(row=r, column=2).border = THIN_BORDER
    # Vendas
    ws_dash.cell(row=r, column=3).value = f'=COUNTIFS(LOG!B:B,A{r},LOG!M:M,"VENDA / PEDIDO",LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3)'
    ws_dash.cell(row=r, column=3).border = THIN_BORDER
    # Orçamentos
    ws_dash.cell(row=r, column=4).value = f'=COUNTIFS(LOG!B:B,A{r},LOG!M:M,"ORÇAMENTO",LOG!A:A,">="&$B$3,LOG!A:A,"<="&$D$3)'
    ws_dash.cell(row=r, column=4).border = THIN_BORDER
    # % Conversão
    ws_dash.cell(row=r, column=5).value = f'=IFERROR(C{r}/B{r},0)'
    ws_dash.cell(row=r, column=5).number_format = '0%'
    ws_dash.cell(row=r, column=5).border = THIN_BORDER
    # Média/dia
    ws_dash.cell(row=r, column=6).value = f'=IFERROR(B{r}/MAX(1,$D$3-$B$3+1),0)'
    ws_dash.cell(row=r, column=6).number_format = '0.0'
    ws_dash.cell(row=r, column=6).border = THIN_BORDER

last_prod_row = first_prod_row + len(consultores) - 1

# TOTAL row
r_prod_total = last_prod_row + 1
ws_dash.cell(row=r_prod_total, column=1, value="TOTAL").font = FONT_BOLD_10
ws_dash.cell(row=r_prod_total, column=1).border = THIN_BORDER
for c in range(2, 5):
    col_l = get_column_letter(c)
    ws_dash.cell(row=r_prod_total, column=c).value = f'=SUM({col_l}{first_prod_row}:{col_l}{last_prod_row})'
    ws_dash.cell(row=r_prod_total, column=c).font = FONT_BOLD_10
    ws_dash.cell(row=r_prod_total, column=c).border = THIN_BORDER
ws_dash.cell(row=r_prod_total, column=5).value = f'=IFERROR(C{r_prod_total}/B{r_prod_total},0)'
ws_dash.cell(row=r_prod_total, column=5).number_format = '0%'
ws_dash.cell(row=r_prod_total, column=5).font = FONT_BOLD_10
ws_dash.cell(row=r_prod_total, column=5).border = THIN_BORDER
ws_dash.cell(row=r_prod_total, column=6).value = f'=IFERROR(B{r_prod_total}/MAX(1,$D$3-$B$3+1),0)'
ws_dash.cell(row=r_prod_total, column=6).number_format = '0.0'
ws_dash.cell(row=r_prod_total, column=6).font = FONT_BOLD_10
ws_dash.cell(row=r_prod_total, column=6).border = THIN_BORDER

ws_dash.sheet_view.zoomScale = 90
print("  ABA 6 DASH done.")


# ═══════════════════════════════════════════════════════════════
# ABA 7: AGENDA (Template)
# ═══════════════════════════════════════════════════════════════
print("Building ABA 7: AGENDA...")
ws_agenda = wb.create_sheet("AGENDA")
ws_agenda.sheet_properties.tabColor = "00B050"

# Header section (rows 1-5)
ws_agenda.cell(row=1, column=1, value="📋 AGENDA DD/MM/AAAA — [CONSULTOR]").font = FONT_TITLE_14
ws_agenda.merge_cells('A1:X1')

ws_agenda.cell(row=2, column=1, value="Cart:XX | Prosp:XX | Follow-ups:XX | Novos:XX").font = FONT_ITALIC
ws_agenda.merge_cells('A2:X2')

ws_agenda.cell(row=3, column=1, value="⏰ 8:00 Reunião | ☀ 9-12 CARTEIRA | 🌙 13-17 PROSPECÇÃO").font = FONT_ITALIC
ws_agenda.merge_cells('A3:X3')

ws_agenda.cell(row=4, column=1, value="⚠ REGISTRAR TUDO: Mercos + Kanban WhatsApp + Esta planilha").font = FONT_BOLD_10
ws_agenda.cell(row=4, column=1).fill = FILL_WARN
ws_agenda.merge_cells('A4:X4')
for c in range(1, 25):
    ws_agenda.cell(row=4, column=c).fill = FILL_WARN

# Row 5: blank

# Row 6: Headers
agenda_headers = [
    "NOME FANTASIA", "CNPJ", "UF", "REDE/REGIONAL", "TELEFONE", "EMAIL",
    "SITUAÇÃO", "DIAS SEM COMPRA", "SINALEIRO", "FASE", "TENTATIVA",
    "AÇÃO SUGERIDA", "BLOCO", "ÚLTIMO RESULTADO",
    "WHATSAPP", "LIGAÇÃO", "LIGAÇÃO ATENDIDA", "TIPO AÇÃO",
    "TIPO DO CONTATO", "RESULTADO", "MOTIVO", "AÇÃO FUTURA",
    "MERCOS ATUALIZADO", "NOTA DO DIA"
]
for i, h in enumerate(agenda_headers, 1):
    cell = ws_agenda.cell(row=6, column=i, value=h)
    cell.font = FONT_HEADER
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER
    if i <= 14:
        cell.fill = FILL_AGENDA_RO  # Read-only blue
    else:
        cell.fill = FILL_AGENDA_EDIT  # Editable green

# Dropdowns for AGENDA (cols O-W)
dv_ag_wpp = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_agenda.add_data_validation(dv_ag_wpp)
dv_ag_wpp.add("O7:O1000")

dv_ag_lig = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_agenda.add_data_validation(dv_ag_lig)
dv_ag_lig.add("P7:P1000")

dv_ag_la = DataValidation(type="list", formula1="LISTA_SIM_NAO_NA", allow_blank=True)
ws_agenda.add_data_validation(dv_ag_la)
dv_ag_la.add("Q7:Q1000")

dv_ag_ar = DataValidation(type="list", formula1="LISTA_ATIVO_RECEPTIVO", allow_blank=True)
ws_agenda.add_data_validation(dv_ag_ar)
dv_ag_ar.add("R7:R1000")

dv_ag_tc = DataValidation(type="list", formula1="LISTA_TIPO_CONTATO", allow_blank=True)
ws_agenda.add_data_validation(dv_ag_tc)
dv_ag_tc.add("S7:S1000")

dv_ag_res = DataValidation(type="list", formula1="LISTA_RESULTADO", allow_blank=True)
ws_agenda.add_data_validation(dv_ag_res)
dv_ag_res.add("T7:T1000")

dv_ag_mot = DataValidation(type="list", formula1="LISTA_MOTIVO", allow_blank=True)
ws_agenda.add_data_validation(dv_ag_mot)
dv_ag_mot.add("U7:U1000")

dv_ag_mercos = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
ws_agenda.add_data_validation(dv_ag_mercos)
dv_ag_mercos.add("W7:W1000")

# V: AÇÃO FUTURA
dv_ag_acao = DataValidation(type="list", formula1="LISTA_ACAO_FUTURA", allow_blank=True)
ws_agenda.add_data_validation(dv_ag_acao)
dv_ag_acao.add("V7:V1000")

# 3 sample rows (rows 7-9)
agenda_data = [
    ["DIVINA TERRA CWB", "32828171000108", "PR", "DIVINA TERRA", "5541999990001", "divina@email.com",
     "ATIVO", 12, "🟢", "CS", "T1", "Ligar oferecer mix", "MANHÃ", "RELACIONAMENTO",
     "SIM", "NÃO", "N/A", "ATIVO", "PÓS-VENDA / RELACIONAMENTO", "RELACIONAMENTO", "AINDA TEM ESTOQUE",
     "Verificar estoque", "SIM", "Cliente satisfeito"],
    ["ALICE GRAOS", "41626544000140", "RS", "TUDO EM GRAOS", "5551999990002", "alice@email.com",
     "EM RISCO", 53, "🟡", "RECOMPRA", "T1", "WhatsApp recompra", "MANHÃ", "NÃO RESPONDE",
     "", "", "", "", "", "", "",
     "", "", ""],
    ["CIA SAUDE SP", "12345678000199", "SP", "CIA DA SAUDE", "5511999990005", "ciasaude@email.com",
     "ATIVO", 9, "🟢", "PÓS-VENDA", "T1", "Ligar confirmar pedido", "MANHÃ", "VENDA / PEDIDO",
     "NÃO", "SIM", "SIM", "ATIVO", "ATEND. CLIENTES ATIVOS", "VENDA / PEDIDO", "",
     "Pedido confirmado", "SIM", "Pedido R$ 2.200"],
]

for r_idx, row_vals in enumerate(agenda_data, 7):
    for c_idx, val in enumerate(row_vals, 1):
        cell = ws_agenda.cell(row=r_idx, column=c_idx, value=val)
        cell.font = FONT_DEFAULT
        cell.border = THIN_BORDER

# Freeze panes at A7
ws_agenda.freeze_panes = "A7"

# Column widths
for col_idx in range(1, 25):
    letter = get_column_letter(col_idx)
    ws_agenda.column_dimensions[letter].width = 16
ws_agenda.column_dimensions['A'].width = 22
ws_agenda.column_dimensions['B'].width = 18
ws_agenda.column_dimensions['F'].width = 22
ws_agenda.column_dimensions['L'].width = 22
ws_agenda.column_dimensions['S'].width = 28
ws_agenda.column_dimensions['T'].width = 22
ws_agenda.column_dimensions['U'].width = 28
ws_agenda.column_dimensions['X'].width = 25

ws_agenda.sheet_view.zoomScale = 90
print("  ABA 7 AGENDA done.")


# ═══════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════
output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "JARVIS_CRM_CENTRAL.xlsx")

print(f"\nSaving to {output_path}...")
wb.save(output_path)
print("SAVED SUCCESSFULLY!")

# ── Summary ──
print("\n" + "="*60)
print("JARVIS CRM CENTRAL — BUILD SUMMARY")
print("="*60)
print(f"Abas: 7 (REGRAS, CARTEIRA, LOG, DRAFT 1, DRAFT 2, DASH, AGENDA)")
print(f"REGRAS: 10 tabelas de validação")
print(f"CARTEIRA: 81 colunas, 5 linhas de dados")
print(f"LOG: 20 colunas, 8 linhas de dados")
print(f"DRAFT 1: 7 colunas, 2 linhas")
print(f"DRAFT 2: 27 colunas, 3 linhas")
print(f"DASH: 7 seções verticais com fórmulas COUNTIFS")
print(f"AGENDA: 24 colunas, 3 linhas de exemplo")
print(f"Named Ranges: {len(named_ranges_def) + 1} ({len(named_ranges_def)} listas + 1 tabela)")
print(f"Agrupamentos [+]: 8 na CARTEIRA (cada coluna individual)")
print(f"Dropdowns: ~50+ data validations across all sheets (todos via Named Ranges)")
print(f"Formatação condicional: SITUAÇÃO (7), %ATING (3), VALIDAÇÃO (2)")
print(f"Output: {output_path}")
print("="*60)

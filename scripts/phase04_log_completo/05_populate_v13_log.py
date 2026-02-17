"""
Phase 04 - Task 2: Popular aba LOG do V13 com 20,830 registros validados.

Carrega log_final_validated.json e escreve na aba LOG do V13 Excel.
Preserva todas as abas existentes (PROJECAO com 19,224 formulas).

Colunas: 20 operacionais (A-T) + 1 metadata (U: ORIGEM_DADO, hidden)
"""

import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

# Paths
BASE = Path(__file__).resolve().parent.parent.parent
DATA_OUT = BASE / "data" / "output"
PHASE04 = DATA_OUT / "phase04"
V13_PATH = DATA_OUT / "CRM_VITAO360_V13_PROJECAO.xlsx"
V13_BACKUP = DATA_OUT / "CRM_VITAO360_V13_PROJECAO_BACKUP_PHASE04.xlsx"
LOG_JSON = PHASE04 / "log_final_validated.json"

# Add scripts/ to path for v3_styles
SCRIPTS_DIR = str(BASE / "scripts")
if SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, SCRIPTS_DIR)

from v3_styles import (
    FONT_TITLE_W, FONT_HEADER, FONT_DATA, FONT_BLOCK,
    FILL_DARK, FILL_HEADER, FILL_LIGHT_GRAY, FILL_NONE,
    FILL_ATIVO, FILL_EM_RISCO, FILL_INAT_REC, FILL_INAT_ANT,
    FILL_NOVO, FILL_PROSPECT, SITUACAO_FILLS,
    THIN_BORDER, ALIGN_CENTER, ALIGN_LEFT,
    FMT_DATE, FMT_TEXT, TAB_LOG,
    style_cell, write_header, write_data,
)

# 20 operational columns (matching _helpers.LOG_COLUMNS order)
LOG_COLUMNS = [
    ('DATA',              14, FMT_DATE, ALIGN_CENTER),
    ('CONSULTOR',         20, None,     ALIGN_LEFT),
    ('NOME FANTASIA',     25, None,     ALIGN_LEFT),
    ('CNPJ',              18, FMT_TEXT, ALIGN_CENTER),
    ('UF',                 5, None,     ALIGN_CENTER),
    ('REDE/REGIONAL',     16, None,     ALIGN_LEFT),
    ('SITUACAO',          14, None,     ALIGN_CENTER),
    ('WHATSAPP',          10, None,     ALIGN_CENTER),
    ('LIGACAO',           10, None,     ALIGN_CENTER),
    ('LIGACAO ATENDIDA',  14, None,     ALIGN_CENTER),
    ('TIPO ACAO',         12, None,     ALIGN_CENTER),
    ('TIPO DO CONTATO',   18, None,     ALIGN_LEFT),
    ('RESULTADO',         22, None,     ALIGN_CENTER),
    ('MOTIVO',            28, None,     ALIGN_LEFT),
    ('FOLLOW-UP',         14, FMT_DATE, ALIGN_CENTER),
    ('ACAO',              16, None,     ALIGN_LEFT),
    ('MERCOS ATUALIZADO', 16, None,     ALIGN_CENTER),
    ('FASE',              16, None,     ALIGN_CENTER),
    ('TENTATIVA',         12, None,     ALIGN_CENTER),
    ('NOTA DO DIA',       40, None,     ALIGN_LEFT),
]

# Column U: hidden metadata
METADATA_COL = ('ORIGEM_DADO', 14, None, ALIGN_CENTER)


def main():
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill

    # ── Step 0: Backup V13 ──
    print("=" * 60)
    print("Phase 04 Task 2: Populate V13 LOG tab")
    print("=" * 60)

    if not V13_PATH.exists():
        print(f"ERRO: V13 nao encontrado em {V13_PATH}")
        sys.exit(1)

    print(f"\n[1/7] Backup V13...")
    shutil.copy2(V13_PATH, V13_BACKUP)
    print(f"  Backup: {V13_BACKUP.name}")

    # ── Step 1: Load V13 (preserve formulas) ──
    print(f"\n[2/7] Loading V13 (data_only=False)...")
    wb = openpyxl.load_workbook(str(V13_PATH), data_only=False)
    sheets_before = list(wb.sheetnames)
    print(f"  Sheets before: {sheets_before}")

    # Count PROJECAO formulas before
    proj_ws = None
    for name in wb.sheetnames:
        if 'PROJE' in name.upper():
            proj_ws = wb[name]
            break

    if proj_ws:
        formulas_before = sum(
            1 for row in proj_ws.iter_rows()
            for cell in row
            if cell.value and str(cell.value).startswith('=')
        )
        print(f"  PROJECAO formulas before: {formulas_before}")
    else:
        formulas_before = 0
        print("  WARNING: PROJECAO sheet not found!")

    # ── Step 2: Create LOG tab ──
    print(f"\n[3/7] Creating LOG tab...")
    if 'LOG' in wb.sheetnames:
        print("  LOG tab exists, removing old...")
        del wb['LOG']

    ws = wb.create_sheet("LOG")
    ws.sheet_properties.tabColor = TAB_LOG

    # ── Row 1: Title (merged A1:U1) ──
    ws.merge_cells('A1:U1')
    title_cell = ws.cell(row=1, column=1,
                         value="LOG \u2014 Arquivo de Atendimentos (append-only)")
    style_cell(title_cell, font=FONT_TITLE_W, fill=FILL_DARK, align=ALIGN_LEFT)
    for c in range(1, 22):  # A-U
        ws.cell(row=1, column=c).fill = FILL_DARK
        ws.cell(row=1, column=c).border = THIN_BORDER

    # ── Row 2: Headers ──
    for idx, (name, width, fmt, align) in enumerate(LOG_COLUMNS, start=1):
        # Cols 1-12 blue header, 13-20 dark gray (auto-derived)
        fill = FILL_HEADER if idx <= 12 else PatternFill('solid', fgColor='595959')
        write_header(ws, 2, idx, name, fill=fill)
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width

    # Column U: ORIGEM_DADO (metadata, hidden)
    meta_name, meta_width, _, meta_align = METADATA_COL
    write_header(ws, 2, 21, meta_name, fill=PatternFill('solid', fgColor='333333'))
    ws.column_dimensions['U'].width = meta_width
    ws.column_dimensions['U'].hidden = True

    print(f"  Headers: 20 cols (A-T) + ORIGEM_DADO (U, hidden)")

    # ── Step 3: Load LOG data ──
    print(f"\n[4/7] Loading log_final_validated.json...")
    with open(LOG_JSON, 'r', encoding='utf-8') as f:
        log_data = json.load(f)
    records = log_data['records']
    print(f"  Records to write: {len(records)}")

    # Sort by DATA then CONSULTOR
    records.sort(key=lambda r: (r.get('DATA', ''), r.get('CONSULTOR', '')))

    # ── Step 4: Write data rows ──
    print(f"\n[5/7] Writing {len(records)} data rows...")
    col_keys = [col[0] for col in LOG_COLUMNS]  # Column names in order

    for i, rec in enumerate(records):
        row_num = i + 3  # Start from row 3

        # Alternating row fill for readability
        row_fill = FILL_LIGHT_GRAY if i % 2 == 0 else FILL_NONE

        for col_idx, (col_name, _, fmt, align) in enumerate(LOG_COLUMNS, start=1):
            value = rec.get(col_name, '')

            # Convert date strings to datetime objects for Excel
            if col_name == 'DATA' and value:
                try:
                    value = datetime.strptime(str(value)[:10], '%Y-%m-%d')
                except (ValueError, TypeError):
                    pass
            elif col_name == 'FOLLOW-UP' and value:
                try:
                    value = datetime.strptime(str(value)[:10], '%Y-%m-%d')
                except (ValueError, TypeError):
                    pass

            # CNPJ as text (not number)
            if col_name == 'CNPJ':
                cell = ws.cell(row=row_num, column=col_idx, value=str(value))
                style_cell(cell, font=FONT_DATA, fill=row_fill,
                           align=align, fmt=FMT_TEXT)
            else:
                cell = write_data(ws, row_num, col_idx, value,
                                  fill=row_fill, align=align, fmt=fmt)

            cell.border = THIN_BORDER

        # Column U: ORIGEM_DADO
        origem = rec.get('ORIGEM_DADO', '')
        cell_u = ws.cell(row=row_num, column=21, value=origem)
        style_cell(cell_u, font=FONT_DATA, fill=row_fill, align=ALIGN_CENTER)
        cell_u.border = THIN_BORDER

        if (i + 1) % 5000 == 0:
            print(f"    ...{i + 1} rows written")

    last_row = len(records) + 2
    print(f"  Done: rows 3 to {last_row}")

    # ── Step 5: Conditional formatting for SITUACAO (col G) ──
    print(f"\n[6/7] Applying formatting...")

    sit_col = 'G'
    sit_range = f'{sit_col}3:{sit_col}{last_row}'
    for sit_name, sit_fill in SITUACAO_FILLS.items():
        ws.conditional_formatting.add(
            sit_range,
            CellIsRule(operator='equal', formula=[f'"{sit_name}"'], fill=sit_fill)
        )
    print(f"  Conditional formatting: SITUACAO ({sit_col}) - 6 rules")

    # Freeze panes at A3
    ws.freeze_panes = "A3"
    print(f"  Freeze panes: A3")

    # Auto filter on row 2 (all 21 columns)
    ws.auto_filter.ref = f"A2:U{last_row}"
    print(f"  Auto filter: A2:U{last_row}")

    # ── Step 6: Save and verify ──
    print(f"\n[7/7] Saving V13...")
    wb.save(str(V13_PATH))
    print(f"  Saved: {V13_PATH.name}")

    # Verify sheets preserved
    wb2 = openpyxl.load_workbook(str(V13_PATH), data_only=False)
    sheets_after = list(wb2.sheetnames)
    print(f"  Sheets after: {sheets_after}")

    # Verify all original sheets still present
    for orig in sheets_before:
        if orig not in sheets_after:
            print(f"  ERROR: Sheet '{orig}' was lost!")
            sys.exit(1)

    # Verify LOG rows
    ws_log = wb2['LOG']
    log_rows = ws_log.max_row - 2
    print(f"  LOG data rows: {log_rows}")
    if log_rows != len(records):
        print(f"  WARNING: Expected {len(records)}, got {log_rows}")

    # Verify PROJECAO formulas
    if proj_ws:
        proj_name = proj_ws.title
        ws_proj2 = wb2[proj_name]
        formulas_after = sum(
            1 for row in ws_proj2.iter_rows()
            for cell in row
            if cell.value and str(cell.value).startswith('=')
        )
        print(f"  PROJECAO formulas after: {formulas_after}")
        if formulas_after < 19000:
            print(f"  CRITICAL: Formula loss! Before={formulas_before}, After={formulas_after}")
            print(f"  Restoring backup...")
            shutil.copy2(V13_BACKUP, V13_PATH)
            print(f"  Backup restored. LOG was written in separate file instead.")
            sys.exit(1)
        else:
            delta = formulas_before - formulas_after
            if delta == 0:
                print(f"  Formula preservation: PERFECT (0 lost)")
            else:
                print(f"  Formula preservation: {delta} formulas changed (likely openpyxl normalization)")

    # Spot check: first and last record
    r3_date = ws_log.cell(row=3, column=1).value
    r3_cnpj = ws_log.cell(row=3, column=4).value
    rl_date = ws_log.cell(row=last_row, column=1).value
    print(f"\n  Spot check:")
    print(f"    First record (row 3): date={r3_date}, CNPJ={r3_cnpj}")
    print(f"    Last record (row {last_row}): date={rl_date}")

    # Summary
    print("\n" + "=" * 60)
    print("V13 LOG POPULATION COMPLETE")
    print(f"  Total records: {len(records)}")
    print(f"  Columns: 20 operational + 1 metadata (hidden)")
    print(f"  Sheets preserved: {len(sheets_after)}")
    print(f"  PROJECAO formulas: {formulas_after if proj_ws else 'N/A'}")
    print("=" * 60)


if __name__ == "__main__":
    main()

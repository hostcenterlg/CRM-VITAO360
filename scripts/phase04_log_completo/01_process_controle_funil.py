"""
Phase 04, Plan 01 — ETL do CONTROLE_FUNIL
Processa 10,544 registros do 06_LOG_FUNIL.xlsx (aba 'Interacoes'),
aplica classificacao 3-tier, mapeia para formato LOG 20-col,
e salva como JSON intermediario.

Fonte: data/sources/funil/06_LOG_FUNIL.xlsx
  - Aba 'Interacoes': 10,544 registros limpos (ja sem alucinacoes)
  - Aba 'Alucinacoes IA': 558 registros descartados (contados para stats)
  - Aba 'Duplicatas': 589 registros removidos (contados para stats)

Saida: data/output/phase04/controle_funil_classified.json
"""

import sys
import json
import re
from pathlib import Path
from datetime import datetime, date
from collections import Counter, defaultdict

# Setup paths
_project_root = Path(__file__).resolve().parent.parent.parent
_scripts_dir = _project_root / 'scripts'
sys.path.insert(0, str(_scripts_dir / 'phase04_log_completo'))
sys.path.insert(0, str(_scripts_dir))

from _helpers import (
    normalize_cnpj, normalize_consultor, classify_record,
    make_dedup_key, make_log_record, LOG_COLUMNS,
    normalize_resultado, _remover_acentos,
)

# =====================================================================
# CONFIGURACAO
# =====================================================================

SOURCE_FILE = _project_root / 'data' / 'sources' / 'funil' / '06_LOG_FUNIL.xlsx'
OUTPUT_DIR = _project_root / 'data' / 'output' / 'phase04'
OUTPUT_FILE = OUTPUT_DIR / 'controle_funil_classified.json'

# Mapeamento de header (coluna index 0-based) -> campo
# Baseado na inspecao real do arquivo
HEADER_MAP = {
    0: 'DATA',                       # Col 1
    # Col 2 = _EMPTY_2 (ignorar)
    2: 'NOME FANTASIA',              # Col 3
    3: 'CNPJ/CPF',                   # Col 4
    4: 'UF',                         # Col 5
    5: 'VENDEDOR DO ULTIMO PEDIDO',  # Col 6
    # Col 7 = DATA DE ULTIMO (financeiro-adjacent, ignorar)
    # Col 8 = VALOR DE PEDIDO (financeiro, IGNORAR — Two-Base)
    # Col 9 = DIAS SEM COMPRAR (referencia, nao vai pro LOG)
    # Col 10 = CICLO MEDIO DE COMPRA (referencia)
    10: 'SITUACAO_RAW',              # Col 11 (SITUACAO)
    # Col 12 = STATUS (nao usado)
    # Col 13 = ULTIMO CONTATO (nao usado)
    # Col 14 = RESULTADO ATUAL (nao usado — usamos col 21 RESULTADO)
    # Col 15 = ACAO (nao usada — calculamos via motor)
    15: 'WHATSAPP',                  # Col 16
    16: 'LIGACAO',                   # Col 17
    17: 'LIGACAO ATENDIDA',          # Col 18
    18: 'TIPO ACAO',                 # Col 19
    19: 'TIPO DO CONTATO',           # Col 20
    20: 'RESULTADO',                 # Col 21
    21: 'FOLLOW-UP_RAW',             # Col 22
    22: 'ACAO_1',                    # Col 23
    23: 'MERCOS ATUALIZADO',         # Col 24
    24: 'NOTA DO DIA',               # Col 25 (NOTA DO DIA / ACAO)
    # Col 26 = EXACTSALES (financeiro, IGNORAR)
    # Col 27 = Coluna1 (lixo)
    27: 'ABA_ORIGEM',                # Col 28
    # Col 29 = ROW_ORIGINAL (referencia)
    29: 'ORIGEM_DADO_SOURCE',        # Col 30 (classificacao pre-existente)
    30: 'CRITERIO_CLASSIFICACAO',    # Col 31
    31: 'FLAG_ALUCINACAO',           # Col 32
}

# Mapeamento de RESULTADO variantes -> forma canonica
RESULTADO_NORMALIZATION = {
    'EM ATENDIMENTO': 'EM ATENDIMENTO',
    'ORCAMENTO': 'ORCAMENTO',
    'ORÇAMENTO': 'ORCAMENTO',
    'CADASTRO': 'CADASTRO',
    'VENDA': 'VENDA / PEDIDO',
    'VENDA / PEDIDO': 'VENDA / PEDIDO',
    'VENDA/PEDIDO': 'VENDA / PEDIDO',
    'RELACIONAMENTO': 'RELACIONAMENTO',
    'FOLLOW UP 7': 'FOLLOW UP 7',
    'FOLLOW UP 15': 'FOLLOW UP 15',
    'FOLLOW UP': 'FOLLOW UP 7',              # Sem numero -> default 7
    'SUPORTE': 'SUPORTE',
    'NAO ATENDE': 'NAO ATENDE',
    'NÃO ATENDE': 'NAO ATENDE',
    'NAO RESPONDE': 'NAO RESPONDE',
    'NÃO RESPONDE': 'NAO RESPONDE',
    'RECUSOU LIGACAO': 'RECUSOU LIGACAO',
    'RECUSOU LIGAÇÃO': 'RECUSOU LIGACAO',
    'PERDA / FECHOU LOJA': 'PERDA / FECHOU LOJA',
    'PERDA / NÃO VENDA': 'PERDA / FECHOU LOJA',   # Variante na fonte
    'PERDA / NAO VENDA': 'PERDA / FECHOU LOJA',
    'PERDA': 'PERDA / FECHOU LOJA',
    'CLIENTE INATIVO': 'NAO RESPONDE',              # Mapeamento aproximado
    'CS': 'RELACIONAMENTO',                          # CS -> RELACIONAMENTO no LOG
    'NUTRICAO': 'FOLLOW UP 15',                      # NUTRICAO -> FOLLOW UP 15
    'NUTRIÇÃO': 'FOLLOW UP 15',
}

# Prioridade de abas (para dedup intra-source)
ABA_PRIORITY = {
    'LOG': 0,
    'Manu log': 1,
    'Planilha5': 2,
    'Planilha4': 3,
}


# =====================================================================
# FUNCOES ETL
# =====================================================================

def parse_date(val):
    """Converte valor de data para date object ou None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Tentar parse de string
    s = str(val).strip()
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return datetime.strptime(s[:10], fmt).date()
        except (ValueError, IndexError):
            continue
    return None


def normalize_bool(val, default=''):
    """Normaliza valores booleanos SIM/NAO."""
    if val is None:
        return default
    s = str(val).strip().upper()
    if s in ('SIM', 'S', 'YES', 'Y', '1', 'TRUE'):
        return 'SIM'
    if s in ('NAO', 'NÃO', 'N', 'NO', '0', 'FALSE'):
        return 'NAO'
    if s in ('N/A', 'NA', '-'):
        return 'N/A'
    return default


def normalize_resultado_value(resultado_raw):
    """Normaliza RESULTADO para forma canonica SEM ACENTO."""
    if not resultado_raw:
        return ''
    r = re.sub(r'\s+', ' ', str(resultado_raw).upper().strip())
    return RESULTADO_NORMALIZATION.get(r, r)


def process_row(row_values, row_num):
    """
    Processa uma row da aba Interacoes.
    Retorna (record_dict, skip_reason) ou (None, reason).
    """
    # Extrair valores brutos
    raw = {}
    for idx, field in HEADER_MAP.items():
        val = row_values[idx] if idx < len(row_values) else None
        raw[field] = val

    # --- Filtro: DATA obrigatorio ---
    data_val = parse_date(raw['DATA'])
    if data_val is None:
        return None, 'DATA vazia/invalida'

    # --- Filtro: CNPJ obrigatorio ---
    cnpj_raw = raw.get('CNPJ/CPF', '')
    if not cnpj_raw:
        return None, 'CNPJ vazio'
    cnpj_norm = normalize_cnpj(cnpj_raw)
    if cnpj_norm == '00000000000000' or len(cnpj_norm) != 14:
        return None, f'CNPJ invalido: {cnpj_raw}'

    # --- Filtro: RESULTADO obrigatorio ---
    resultado_raw = raw.get('RESULTADO', '')
    resultado_norm = normalize_resultado_value(resultado_raw)
    if not resultado_norm:
        return None, 'RESULTADO vazio'

    # --- Normalizar consultor ---
    vendedor_raw = raw.get('VENDEDOR DO ULTIMO PEDIDO', '')
    consultor = normalize_consultor(vendedor_raw)

    # --- Classificacao: usar a pre-existente do arquivo ---
    origem_source = str(raw.get('ORIGEM_DADO_SOURCE', '') or '').strip().upper()
    if origem_source in ('REAL', 'SINTETICO'):
        classificacao = origem_source
    else:
        # Re-classificar se nao tem classificacao pre-existente
        nota_raw = raw.get('NOTA DO DIA', '')
        classificacao, _ = classify_record(cnpj_raw, raw.get('NOME FANTASIA', ''),
                                           vendedor_raw, nota_raw)
        if classificacao == 'ALUCINACAO':
            return None, 'Reclassificado como ALUCINACAO'

    # --- Normalizar campos ---
    nome = str(raw.get('NOME FANTASIA', '') or '').strip()
    uf = str(raw.get('UF', '') or '').strip().upper()
    situacao_raw = str(raw.get('SITUACAO_RAW', '') or '').strip().upper()
    # Remover acentos de SITUACAO
    situacao = _remover_acentos(situacao_raw)

    # Canais
    whatsapp = normalize_bool(raw.get('WHATSAPP'), 'SIM')
    ligacao_raw = raw.get('LIGACAO')
    ligacao = normalize_bool(ligacao_raw, 'NAO')
    lig_atendida_raw = raw.get('LIGACAO ATENDIDA')
    if ligacao == 'SIM':
        lig_atendida = normalize_bool(lig_atendida_raw, 'NAO')
    else:
        lig_atendida = 'N/A'

    # Tipo acao
    tipo_acao_raw = str(raw.get('TIPO ACAO', '') or '').strip().upper()
    tipo_acao = tipo_acao_raw if tipo_acao_raw in ('ATIVO', 'RECEPTIVO') else 'ATIVO'

    # Tipo do contato (sem acento — P8)
    tipo_contato_raw = str(raw.get('TIPO DO CONTATO', '') or '').strip()
    tipo_contato = _remover_acentos(tipo_contato_raw).upper() if tipo_contato_raw else ''

    # MERCOS ATUALIZADO
    mercos = normalize_bool(raw.get('MERCOS ATUALIZADO'), '')

    # NOTA DO DIA
    nota = str(raw.get('NOTA DO DIA', '') or '').strip()

    # ABA_ORIGEM (para dedup)
    aba_origem = str(raw.get('ABA_ORIGEM', '') or '').strip()

    # --- Construir record via make_log_record ---
    # Calcular resultado com acento para motor_de_regras
    resultado_sem, resultado_motor = normalize_resultado(resultado_norm)

    record = make_log_record(
        data=data_val,
        cnpj=cnpj_norm,
        consultor=consultor,
        resultado=resultado_norm,
        nota=nota,
        whatsapp=whatsapp,
        ligacao=ligacao,
        lig_atendida=lig_atendida,
        tipo_acao=tipo_acao,
        motivo='',  # Sem coluna MOTIVO na fonte
        mercos=mercos,
        nome=nome,
        uf=uf,
        rede='',  # Sem coluna REDE na fonte — default DEMAIS CLIENTES
        situacao=situacao,
        origem_dado=classificacao,
        source='CONTROLE_FUNIL',
    )

    # Substituir TIPO DO CONTATO da fonte se existir (a fonte ja tem dados reais)
    if tipo_contato:
        record['TIPO DO CONTATO'] = tipo_contato

    # Metadata extra para dedup
    record['_aba_origem'] = aba_origem
    record['_row_num'] = row_num

    return record, None


def deduplicate_records(records):
    """
    Dedup intra-CONTROLE_FUNIL por chave composta.
    Se duplicata entre abas, manter da aba com maior prioridade.
    """
    seen = {}
    duplicatas_removidas = 0

    for rec in records:
        key = make_dedup_key(rec['DATA'], rec['CNPJ'], rec['RESULTADO'])
        aba = rec.get('_aba_origem', '')
        priority = ABA_PRIORITY.get(aba, 99)

        if key not in seen:
            seen[key] = (rec, priority)
        else:
            existing_rec, existing_priority = seen[key]
            if priority < existing_priority:
                # Este registro tem prioridade maior
                seen[key] = (rec, priority)
            duplicatas_removidas += 1

    # Remover metadados internos
    final_records = []
    for rec, _ in seen.values():
        rec.pop('_aba_origem', None)
        rec.pop('_row_num', None)
        final_records.append(rec)

    return final_records, duplicatas_removidas


def print_statistics(records, stats):
    """Imprime relatorio completo de estatisticas."""
    print("\n" + "=" * 70)
    print("  RELATORIO ETL — CONTROLE_FUNIL -> LOG 20-col")
    print("=" * 70)

    print(f"\n--- Totais ---")
    print(f"  Registros lidos (Interacoes):  {stats['total_lidos']}")
    print(f"  Alucinacoes na aba separada:   {stats['alucinacoes_aba']}")
    print(f"  Duplicatas na aba separada:    {stats['duplicatas_aba']}")

    print(f"\n--- Filtrados ---")
    print(f"  Sem DATA:                      {stats['skip_no_data']}")
    print(f"  Sem CNPJ:                      {stats['skip_no_cnpj']}")
    print(f"  Sem RESULTADO:                 {stats['skip_no_resultado']}")
    print(f"  Re-classificado ALUCINACAO:    {stats['skip_reclassified']}")
    print(f"  Total descartados:             {stats['total_descartados']}")

    print(f"\n--- Classificacao ---")
    print(f"  REAL:                          {stats['count_real']}")
    print(f"  SINTETICO:                     {stats['count_sintetico']}")
    print(f"  Total no output:               {stats['count_real'] + stats['count_sintetico']}")

    print(f"\n--- Dedup ---")
    print(f"  Duplicatas removidas (intra):  {stats['duplicatas_removidas']}")
    print(f"  Registros finais:              {stats['final_count']}")

    print(f"\n--- Registros em fim de semana: {stats['weekend_count']} (mantidos, warning) ---")

    # Distribuicao por aba de origem
    aba_counter = Counter()
    for r in records:
        aba_counter[r.get('SOURCE', 'CONTROLE_FUNIL')] += 1
    print(f"\n--- Por fonte ---")
    for k, v in aba_counter.most_common():
        print(f"  {k}: {v}")

    # Distribuicao por mes
    month_counter = Counter()
    for r in records:
        try:
            m = int(r['DATA'][5:7])
            month_counter[m] += 1
        except (ValueError, TypeError, IndexError):
            pass
    print(f"\n--- Por mes ---")
    for m in sorted(month_counter.keys()):
        print(f"  {m:02d}/2025: {month_counter[m]:>5}")

    # Distribuicao por consultor
    consultor_counter = Counter()
    for r in records:
        consultor_counter[r['CONSULTOR']] += 1
    print(f"\n--- Por consultor ---")
    for k, v in consultor_counter.most_common():
        print(f"  {k:20s}: {v:>5}")

    # Distribuicao por resultado
    resultado_counter = Counter()
    for r in records:
        resultado_counter[r['RESULTADO']] += 1
    print(f"\n--- Por resultado ---")
    for k, v in resultado_counter.most_common():
        print(f"  {k:25s}: {v:>5}")

    print("\n" + "=" * 70)


# =====================================================================
# MAIN
# =====================================================================

def main():
    import openpyxl

    print(f"Lendo {SOURCE_FILE}...")
    wb = openpyxl.load_workbook(str(SOURCE_FILE), read_only=True, data_only=True)

    # --- Contar alucinacoes e duplicatas nas abas separadas ---
    alucinacoes_count = 0
    if 'Alucinações IA' in wb.sheetnames:
        for _ in wb['Alucinações IA'].iter_rows(min_row=2, values_only=True):
            alucinacoes_count += 1
    print(f"  Aba 'Alucinacoes IA': {alucinacoes_count} registros (descartados)")

    duplicatas_aba_count = 0
    if 'Duplicatas' in wb.sheetnames:
        for _ in wb['Duplicatas'].iter_rows(min_row=2, values_only=True):
            duplicatas_aba_count += 1
    print(f"  Aba 'Duplicatas': {duplicatas_aba_count} registros (descartados)")

    # --- Processar aba Interacoes ---
    ws = wb['Interações']
    print(f"  Aba 'Interacoes': max_row={ws.max_row}, max_col={ws.max_column}")

    records = []
    stats = {
        'total_lidos': 0,
        'skip_no_data': 0,
        'skip_no_cnpj': 0,
        'skip_no_resultado': 0,
        'skip_reclassified': 0,
        'total_descartados': 0,
        'count_real': 0,
        'count_sintetico': 0,
        'weekend_count': 0,
        'alucinacoes_aba': alucinacoes_count,
        'duplicatas_aba': duplicatas_aba_count,
        'duplicatas_removidas': 0,
        'final_count': 0,
    }

    row_num = 1  # Tracking (header = row 1)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        stats['total_lidos'] += 1

        record, skip_reason = process_row(row, row_num)

        if record is None:
            if 'DATA' in (skip_reason or ''):
                stats['skip_no_data'] += 1
            elif 'CNPJ' in (skip_reason or ''):
                stats['skip_no_cnpj'] += 1
            elif 'RESULTADO' in (skip_reason or ''):
                stats['skip_no_resultado'] += 1
            elif 'ALUCINACAO' in (skip_reason or ''):
                stats['skip_reclassified'] += 1
            stats['total_descartados'] += 1
            continue

        # Weekend warning
        try:
            d = parse_date(record['DATA'])
            if d and d.weekday() >= 5:
                stats['weekend_count'] += 1
        except Exception:
            pass

        # Classificacao counter
        if record['ORIGEM_DADO'] == 'REAL':
            stats['count_real'] += 1
        else:
            stats['count_sintetico'] += 1

        records.append(record)

    wb.close()
    print(f"\n  Processados: {stats['total_lidos']} registros")
    print(f"  Descartados: {stats['total_descartados']}")
    print(f"  Aceitos (pre-dedup): {len(records)}")

    # --- Dedup intra-source ---
    records, dup_count = deduplicate_records(records)
    stats['duplicatas_removidas'] = dup_count
    stats['final_count'] = len(records)

    print(f"  Duplicatas removidas: {dup_count}")
    print(f"  Registros finais: {len(records)}")

    # --- Validacoes finais ---
    print("\nValidando registros finais...")
    errors = []

    for i, r in enumerate(records):
        # CNPJ 14 digitos
        if len(r['CNPJ']) != 14 or not r['CNPJ'].isdigit():
            errors.append(f"CNPJ invalido idx {i}: {r['CNPJ']}")

        # Nenhuma alucinacao
        if r['ORIGEM_DADO'] not in ('REAL', 'SINTETICO'):
            errors.append(f"ORIGEM_DADO invalida idx {i}: {r['ORIGEM_DADO']}")

        # Zero campos financeiros
        for key in r:
            k_upper = key.upper()
            if 'VALOR' in k_upper or 'FATURAMENTO' in k_upper:
                errors.append(f"Campo financeiro no record idx {i}: {key}")

    if errors:
        print(f"  ERROS ENCONTRADOS: {len(errors)}")
        for e in errors[:10]:
            print(f"    {e}")
    else:
        print(f"  Validacao: OK (0 erros)")

    # --- Julio Gadret check ---
    julio_variantes = Counter()
    for r in records:
        if 'JULIO' in r['CONSULTOR'].upper() or 'GADRET' in r['CONSULTOR'].upper():
            julio_variantes[r['CONSULTOR']] += 1
    if julio_variantes:
        print(f"\n  Julio Gadret variantes: {dict(julio_variantes)}")
        if len(julio_variantes) > 1:
            print("    AVISO: Multiplas grafias detectadas!")
        else:
            print("    OK: Grafia unica normalizada")

    # --- Daiane check ---
    daiane_variantes = Counter()
    for r in records:
        if 'DAIANE' in r['CONSULTOR'].upper() or 'CENTRAL' in r['CONSULTOR'].upper():
            daiane_variantes[r['CONSULTOR']] += 1
    if daiane_variantes:
        print(f"  Daiane variantes: {dict(daiane_variantes)}")

    # --- Estatisticas ---
    print_statistics(records, stats)

    # --- Salvar JSON ---
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    output = {
        'metadata': {
            'source': 'CONTROLE_FUNIL',
            'source_file': '06_LOG_FUNIL.xlsx',
            'total': len(records),
            'real': stats['count_real'],
            'sintetico': stats['count_sintetico'],
            'alucinacao_descartadas': alucinacoes_count,
            'duplicatas_removidas': dup_count,
            'registros_sem_data': stats['skip_no_data'],
            'registros_sem_resultado': stats['skip_no_resultado'],
            'weekend_records': stats['weekend_count'],
            'date_processed': datetime.now().strftime('%Y-%m-%d'),
            'log_columns': LOG_COLUMNS,
        },
        'records': records,
    }

    with open(str(OUTPUT_FILE), 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    print(f"\n  JSON salvo: {OUTPUT_FILE}")
    print(f"  Tamanho: {OUTPUT_FILE.stat().st_size / 1024 / 1024:.1f} MB")
    print(f"\nDONE!")


if __name__ == '__main__':
    main()

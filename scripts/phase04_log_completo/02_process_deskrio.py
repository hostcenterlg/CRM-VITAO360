"""
02_process_deskrio.py -- ETL dos tickets Deskrio para formato LOG 20 colunas

ESTRUTURA REAL ENCONTRADA (inspecao em 17/02/2026):
=================================================================
Fonte principal: data/sources/deskrio/07_TICKETS_DESKRIO.xlsx
Sheet: 'Tickets' -- 5.329 tickets, 27 colunas, JA ao nivel de ticket (NAO mensagem)

Colunas:
  A: Protocolo          -- ID unico do ticket
  B: Origem             -- 'Receptivo' ou 'Ativo'
  C: Status             -- 'Fechado' (4496), 'Aberto' (830), 'Pendente' (3)
  D: Criado em          -- datetime (data abertura do ticket)
  E: Fechado em         -- datetime (data fechamento, pode ser None)
  F: Nome do contato    -- Nome do cliente (ex: 'CIA DA SAUDE - SOMBRIO SC')
  G: Numero do cliente  -- Telefone internacional
  H: Atendente          -- 'Larissa', 'Manu  - Vitao', 'Daiane Stavicki', 'Rodrigo', 'Administrador'
  I: Setor              -- 'Mais Granel', 'Interno Vitao', 'Daiane Vitao'
  J: Conexao            -- Conexao + emoji
  K: Motivo             -- Maioria vazio; quando preenchido: 'Venda Realizada', etc
  L: Observacao          -- Raramente preenchido
  M: Avaliacao           -- Raramente preenchido
  N: CNPJ               -- Formato '30.884.187/0001-02' (3.963 tickets COM, 1.366 SEM)
  O: CNPJ 2             -- CNPJ secundario (490 tickets)
  P: CNPJ 3             -- CNPJ terciario (86 tickets)
  Q: Estado             -- UF derivado
  R: Codigo SAP         -- Codigo SAP do cliente
  S: TEMPO_ATENDIMENTO_HORAS -- Horas de atendimento
  T: NOME_LOJA_EXTRAIDO -- Nome da loja extraido
  U: PESSOA_CONTATO_EXTRAIDA -- Pessoa de contato
  V: UF_CONTATO_EXTRAIDO -- UF extraido do contato
  W: MES_REF            -- Mes de referencia
  X: SEMANA_REF         -- Semana de referencia
  Y: ARQUIVO_ORIGEM     -- Arquivo de exportacao original
  Z: TEM_CNPJ           -- Flag boolean/text
  AA: QTD_CNPJS         -- Quantidade de CNPJs

Dados JA estao ao nivel de TICKET (1 ticket = 1 registro LOG).
Total: 5.329 tickets. NAO sao 77.805 mensagens.

Granularidade: TICKET = CONVERSA (cada protocolo e unico).

Fontes auxiliares (NAO usados diretamente, apenas para referencia):
  - CONTATOS DESKRIO .xlsx: Tabela de contatos com CNPJs (para matching)
  - Contatos Mais Granel.xlsx: Contatos da conexao Mais Granel
  - analises/: Arquivos de analise CSV/TXT/XLSX (mensagens, nao tickets)

CNPJ Matching Strategy:
  - 3.963 tickets JA TEM CNPJ direto na coluna N (74.4%)
  - 1.366 tickets SEM CNPJ: tentar match por Nome do Contato vs roster Mercos
  - Roster carregado de: data/sources/mercos/08_CARTEIRA_MERCOS.xlsx
  - Complemento: data/output/phase02/sap_mercos_merged.json (537 CNPJs)
=================================================================
"""

import sys
import re
import json
import openpyxl
from pathlib import Path
from datetime import date, datetime
from collections import Counter, defaultdict

# Importar helpers do mesmo diretorio
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _helpers import (
    normalize_cnpj, normalize_consultor, make_log_record,
    make_dedup_key, LOG_COLUMNS
)

# =====================================================================
# CONSTANTES
# =====================================================================

BASE_DIR = Path(__file__).resolve().parent.parent.parent
DESKRIO_FILE = BASE_DIR / 'data' / 'sources' / 'deskrio' / '07_TICKETS_DESKRIO.xlsx'
MERCOS_CARTEIRA = BASE_DIR / 'data' / 'sources' / 'mercos' / '08_CARTEIRA_MERCOS.xlsx'
CONTATOS_DESKRIO = BASE_DIR / 'data' / 'sources' / 'deskrio' / 'CONTATOS DESKRIO .xlsx'
MERGED_JSON = BASE_DIR / 'data' / 'output' / 'phase02' / 'sap_mercos_merged.json'
OUTPUT_FILE = BASE_DIR / 'data' / 'output' / 'phase04' / 'deskrio_normalized.json'

# Mapeamento de Status Deskrio -> RESULTADO padrao LOG
STATUS_TO_RESULTADO = {
    'FECHADO': 'SUPORTE',
    'ABERTO': 'EM ATENDIMENTO',
    'PENDENTE': 'FOLLOW UP 7',
}

# Mapeamento de Motivo Deskrio -> RESULTADO padrao LOG (override do status)
MOTIVO_TO_RESULTADO = {
    'VENDA REALIZADA': 'VENDA / PEDIDO',
    'VENDA NAO REALIZADA - FRETE': 'NAO RESPONDE',
    'VENDA NAO REALIZADA - PRECO': 'NAO RESPONDE',
    'VENDA NÃO REALIZADA - FRETE': 'NAO RESPONDE',
    'VENDA NÃO REALIZADA - PREÇO': 'NAO RESPONDE',
    'INFORMACOES - TRANSPORTE': 'SUPORTE',
    'INFORMAÇÕES - TRANSPORTE': 'SUPORTE',
    'SOLICITACAO DE BOLETO': 'SUPORTE',
    'SOLICITAÇÃO DE BOLETO': 'SUPORTE',
    'RASTREIO': 'SUPORTE',
    'LIGACAO NAO ATENDIDA': 'NAO ATENDE',
    'LIGAÇÃO NÃO ATENDIDA': 'NAO ATENDE',
    'OUTROS': 'SUPORTE',
}

# Mapeamento de Motivo Deskrio -> MOTIVO padrao LOG
MOTIVO_TO_MOTIVO_LOG = {
    'VENDA NAO REALIZADA - FRETE': 'PROBLEMA LOGISTICO / ENTREGA',
    'VENDA NÃO REALIZADA - FRETE': 'PROBLEMA LOGISTICO / ENTREGA',
    'VENDA NAO REALIZADA - PRECO': 'PRODUTO NAO VENDEU / SEM GIRO',
    'VENDA NÃO REALIZADA - PREÇO': 'PRODUTO NAO VENDEU / SEM GIRO',
}


# =====================================================================
# FUNCOES
# =====================================================================

def build_cnpj_roster():
    """
    Constroi dicionario de lookup nome_fantasia -> (cnpj, uf, rede)
    a partir do Mercos Carteira e dos contatos Deskrio.
    """
    roster = {}  # {nome_upper: (cnpj, uf, rede, nome_original)}

    # 1. Carregar Mercos Carteira (fonte principal -- 537 clientes)
    print("  Carregando roster Mercos Carteira...")
    wb = openpyxl.load_workbook(str(MERCOS_CARTEIRA), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  # 'Carteira Clientes Mercos'
    for row in ws.iter_rows(min_row=2, values_only=True):
        razao = str(row[0] or '').strip()
        nome_fant = str(row[1] or '').strip()
        cnpj_raw = row[2]
        uf = str(row[6] or '').strip().upper()
        # Vendedor do ultimo pedido
        vendedor = str(row[12] or '').strip()

        cnpj = normalize_cnpj(cnpj_raw)
        if not cnpj or cnpj == '00000000000000':
            continue

        # Determinar rede a partir do nome ou vendedor
        rede = _extrair_rede(nome_fant) or _extrair_rede(razao)
        if not rede:
            rede = 'DEMAIS CLIENTES'

        # Adicionar nome fantasia
        if nome_fant:
            key = nome_fant.upper().strip()
            if key not in roster:
                roster[key] = (cnpj, uf, rede, nome_fant)

        # Adicionar razao social tambem
        if razao and razao != nome_fant:
            key = razao.upper().strip()
            if key not in roster:
                roster[key] = (cnpj, uf, rede, razao)
    wb.close()
    print(f"    Mercos: {len(roster)} nomes mapeados")

    # 2. Complementar com Contatos Deskrio (tem CNPJ + nome do contato)
    print("  Carregando contatos Deskrio...")
    contatos_count = 0
    try:
        wb2 = openpyxl.load_workbook(str(CONTATOS_DESKRIO), read_only=True, data_only=True)
        ws2 = wb2[wb2.sheetnames[0]]  # 'Contacts'
        for row in ws2.iter_rows(min_row=2, values_only=True):
            nome = str(row[3] or '').strip()
            cnpj_raw = row[8] if len(row) > 8 else None  # CNPJ column
            if not cnpj_raw:
                cnpj_raw = row[7] if len(row) > 7 else None  # try another col

            cnpj = normalize_cnpj(cnpj_raw) if cnpj_raw else ''
            if not cnpj or cnpj == '00000000000000' or len(cnpj) < 14:
                continue

            # UF from Estado column
            uf = str(row[12] or '').strip().upper() if len(row) > 12 else ''

            if nome:
                key = nome.upper().strip()
                if key not in roster:
                    roster[key] = (cnpj, uf, 'DEMAIS CLIENTES', nome)
                    contatos_count += 1
        wb2.close()
        print(f"    Contatos Deskrio: +{contatos_count} nomes adicionais")
    except Exception as e:
        print(f"    AVISO: Nao foi possivel carregar contatos Deskrio: {e}")

    # 3. Complementar com CNPJs do merged JSON
    print("  Carregando CNPJs do merged JSON...")
    try:
        with open(str(MERGED_JSON), 'r', encoding='utf-8') as f:
            merged = json.load(f)
        merged_cnpjs = set(merged.get('cnpj_to_vendas', {}).keys())
        print(f"    Merged JSON: {len(merged_cnpjs)} CNPJs para validacao")
    except Exception as e:
        print(f"    AVISO: Nao foi possivel carregar merged JSON: {e}")
        merged_cnpjs = set()

    return roster, merged_cnpjs


def _extrair_rede(nome):
    """Extrai rede do nome do cliente."""
    if not nome:
        return ''
    nome_upper = nome.upper()
    redes = {
        'CIA DA SAUDE': 'CIA DA SAUDE',
        'CIA DA SAÚDE': 'CIA DA SAUDE',
        'FITLAND': 'FITLAND',
        'FIT LAND': 'FITLAND',
        'DIVINA TERRA': 'DIVINA TERRA',
        'BIOMUNDO': 'BIOMUNDO',
        'BIO MUNDO': 'BIO MUNDO',
        'MUNDO VERDE': 'MUNDO VERDE',
        'TUDO EM GRAOS': 'TUDO EM GRAOS',
        'TUDO EM GRÃOS': 'TUDO EM GRAOS',
        'VIDA LEVE': 'VIDA LEVE',
        'ARMAZEM': 'ARMAZEM',
        'ARMAZÉM': 'ARMAZEM',
        'NATURVIDA': 'NATURVIDA',
        'LIGEIRINHO': 'LIGEIRINHO',
        'TRIP': 'TRIP',
        'ESMERALDA': 'ESMERALDA',
        'MAIS GRANEL': 'MAIS GRANEL',
    }
    for pattern, rede in redes.items():
        if pattern in nome_upper:
            return rede
    return ''


def match_cnpj_by_name(nome_contato, roster):
    """
    Tenta encontrar CNPJ pelo nome do contato usando o roster.

    Algoritmo (em ordem de prioridade):
    1. Match exato: nome.upper().strip() == roster_key
    2. Match parcial contains: roster_key in nome OU nome in roster_key
    3. Match por primeiras 3 palavras: compara primeiras 3 palavras

    Returns:
        (cnpj, uf, rede, nome_original, match_type) ou (None, None, None, None, None)
    """
    if not nome_contato:
        return None, None, None, None, None

    nome_upper = nome_contato.upper().strip()
    # Limpar caracteres especiais e emojis
    nome_clean = re.sub(r'[^\w\s\-/]', '', nome_upper).strip()
    nome_clean = re.sub(r'\s+', ' ', nome_clean)

    # 1. Match exato
    if nome_clean in roster:
        cnpj, uf, rede, nome_orig = roster[nome_clean]
        return cnpj, uf, rede, nome_orig, 'EXATO'

    # Tentar sem sufixos comuns (- SC, - PR, nome da pessoa, etc)
    # Ex: 'CIA DA SAUDE - SOMBRIO SC' -> 'CIA DA SAUDE'
    # Ex: 'EMPORIUM PROD. NATURAIS - JESSICA - SC' -> 'EMPORIUM PROD. NATURAIS'
    nome_parts = re.split(r'\s*-\s*', nome_clean)
    if len(nome_parts) > 1:
        nome_base = nome_parts[0].strip()
        if nome_base in roster:
            cnpj, uf, rede, nome_orig = roster[nome_base]
            return cnpj, uf, rede, nome_orig, 'EXATO_BASE'

    # 2. Match parcial (contains)
    # Priorizar matches mais longos (mais especificos)
    best_match = None
    best_len = 0
    for roster_key, (cnpj, uf, rede, nome_orig) in roster.items():
        if len(roster_key) < 5:  # Ignorar nomes muito curtos
            continue
        if roster_key in nome_clean or nome_clean in roster_key:
            if len(roster_key) > best_len:
                best_match = (cnpj, uf, rede, nome_orig, 'PARCIAL')
                best_len = len(roster_key)

    if best_match:
        return best_match

    # 3. Match por primeiras 3 palavras
    nome_words = nome_clean.split()[:3]
    if len(nome_words) >= 2:
        prefix = ' '.join(nome_words)
        for roster_key, (cnpj, uf, rede, nome_orig) in roster.items():
            roster_prefix = ' '.join(roster_key.split()[:3])
            if prefix == roster_prefix:
                return cnpj, uf, rede, nome_orig, 'PREFIXO_3W'

    return None, None, None, None, None


def map_deskrio_resultado(status, motivo):
    """
    Mapeia Status + Motivo Deskrio para RESULTADO padrao LOG.

    Prioridade:
    1. Se motivo preenchido e mapeavel -> usar motivo
    2. Senao -> usar status
    """
    # Tentar motivo primeiro
    if motivo:
        motivo_upper = motivo.strip().upper()
        # Normalizar variantes de motivo
        for pat, resultado in MOTIVO_TO_RESULTADO.items():
            if pat == motivo_upper:
                return resultado

        # Motivos compostos (ex: "Outros, Outros")
        for pat, resultado in MOTIVO_TO_RESULTADO.items():
            if pat in motivo_upper:
                return resultado

    # Fallback: usar status
    if status:
        return STATUS_TO_RESULTADO.get(status.strip().upper(), 'SUPORTE')

    return 'SUPORTE'


def map_deskrio_motivo_log(motivo_deskrio):
    """Mapeia motivo Deskrio para MOTIVO padrao LOG."""
    if not motivo_deskrio:
        return ''
    motivo_upper = motivo_deskrio.strip().upper()
    for pat, motivo_log in MOTIVO_TO_MOTIVO_LOG.items():
        if pat == motivo_upper or pat in motivo_upper:
            return motivo_log
    return ''


def build_nota_deskrio(nome_contato, motivo, observacao, protocolo, status):
    """
    Constroi NOTA DO DIA a partir dos campos Deskrio.
    Formato: "[Deskrio #{protocolo}] {contexto}"
    """
    parts = [f"[Deskrio #{protocolo}]"]

    if motivo and str(motivo).strip():
        parts.append(str(motivo).strip())

    if observacao and str(observacao).strip():
        obs = str(observacao).strip()[:200]  # Limitar tamanho
        parts.append(obs)

    if nome_contato:
        parts.append(f"Contato: {nome_contato}")

    if status:
        parts.append(f"Status: {status}")

    return ' | '.join(parts)


# =====================================================================
# PROCESSAMENTO PRINCIPAL
# =====================================================================

def process_deskrio():
    """Processa tickets Deskrio e gera deskrio_normalized.json."""

    print("=" * 70)
    print("PHASE 04 - PLAN 02: PROCESSAMENTO DESKRIO")
    print("=" * 70)

    # ----------------------------------------------------------------
    # 1. Construir roster para CNPJ matching
    # ----------------------------------------------------------------
    print("\n[1/5] Construindo roster de clientes para CNPJ matching...")
    roster, merged_cnpjs = build_cnpj_roster()
    print(f"  Roster total: {len(roster)} nomes mapeados")
    print(f"  Merged CNPJs: {len(merged_cnpjs)} para validacao")

    # ----------------------------------------------------------------
    # 2. Ler tickets Deskrio
    # ----------------------------------------------------------------
    print(f"\n[2/5] Lendo tickets de {DESKRIO_FILE.name}...")

    if not DESKRIO_FILE.exists():
        print(f"  ERRO: Arquivo nao encontrado: {DESKRIO_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(DESKRIO_FILE), read_only=True, data_only=True)
    ws = wb['Tickets']

    tickets = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        tickets.append({
            'protocolo': str(row[0]),
            'origem': str(row[1] or '').strip(),
            'status': str(row[2] or '').strip(),
            'criado_em': row[3],
            'fechado_em': row[4],
            'nome_contato': str(row[5] or '').strip(),
            'numero_cliente': str(row[6] or '').strip(),
            'atendente': str(row[7] or '').strip() if row[7] else '',
            'setor': str(row[8] or '').strip(),
            'conexao': str(row[9] or '').strip(),
            'motivo': str(row[10] or '').strip() if row[10] else '',
            'observacao': str(row[11] or '').strip() if row[11] else '',
            'avaliacao': row[12],
            'cnpj': row[13],
            'cnpj2': row[14],
            'cnpj3': row[15],
            'estado': str(row[16] or '').strip().upper() if row[16] else '',
            'codigo_sap': row[17],
            'nome_loja_extraido': str(row[19] or '').strip() if len(row) > 19 and row[19] else '',
            'uf_extraido': str(row[21] or '').strip().upper() if len(row) > 21 and row[21] else '',
        })
    wb.close()

    print(f"  Total tickets lidos: {len(tickets)}")

    # ----------------------------------------------------------------
    # 3. Processar cada ticket -> registro LOG
    # ----------------------------------------------------------------
    print(f"\n[3/5] Processando tickets para formato LOG...")

    records = []
    stats = {
        'total_lidos': len(tickets),
        'cnpj_direto': 0,
        'cnpj_match_exato': 0,
        'cnpj_match_exato_base': 0,
        'cnpj_match_parcial': 0,
        'cnpj_match_prefixo': 0,
        'cnpj_pendente': 0,
        'weekend_skipped': 0,
        'no_date': 0,
        'invalid_atendente': 0,
        'processados': 0,
    }
    cnpj_pendente_nomes = []
    resultado_dist = Counter()
    consultor_dist = Counter()
    month_dist = Counter()
    match_type_dist = Counter()

    for ticket in tickets:
        # Validar data
        dt = ticket['criado_em']
        if not isinstance(dt, datetime):
            stats['no_date'] += 1
            continue

        # Skip weekends
        if dt.weekday() >= 5:
            stats['weekend_skipped'] += 1
            continue

        # Data como date object
        data = dt.date() if isinstance(dt, datetime) else dt

        # --- CNPJ Resolution ---
        cnpj_raw = ticket['cnpj']
        cnpj_norm = ''
        uf = ticket['estado'] or ticket['uf_extraido']
        rede = ''
        nome_fantasia = ticket['nome_contato']
        match_type = 'DIRETO'

        if cnpj_raw and str(cnpj_raw).strip():
            # Tem CNPJ direto
            cnpj_norm = normalize_cnpj(cnpj_raw)
            stats['cnpj_direto'] += 1
            match_type = 'DIRETO'

            # Tentar enriquecer com dados do roster
            nome_upper = nome_fantasia.upper().strip() if nome_fantasia else ''
            if nome_upper in roster:
                _, roster_uf, roster_rede, _ = roster[nome_upper]
                if not uf:
                    uf = roster_uf
                rede = roster_rede
            else:
                rede = _extrair_rede(nome_fantasia) or 'DEMAIS CLIENTES'
        else:
            # Sem CNPJ: tentar matching por nome
            # Tentar nome do contato
            matched_cnpj, matched_uf, matched_rede, matched_nome, mt = \
                match_cnpj_by_name(ticket['nome_contato'], roster)

            # Se nao encontrou, tentar nome_loja_extraido
            if not matched_cnpj and ticket['nome_loja_extraido']:
                matched_cnpj, matched_uf, matched_rede, matched_nome, mt = \
                    match_cnpj_by_name(ticket['nome_loja_extraido'], roster)

            if matched_cnpj:
                cnpj_norm = matched_cnpj
                if matched_uf and not uf:
                    uf = matched_uf
                rede = matched_rede or 'DEMAIS CLIENTES'
                if matched_nome:
                    nome_fantasia = matched_nome
                match_type = mt or 'ROSTER'
                stats[f'cnpj_match_{mt.lower()}'] = stats.get(f'cnpj_match_{mt.lower()}', 0) + 1
            else:
                # Nao encontrou CNPJ -- descartar
                stats['cnpj_pendente'] += 1
                cnpj_pendente_nomes.append(ticket['nome_contato'])
                continue

        # Validar CNPJ (14 digitos)
        if len(cnpj_norm) != 14 or not cnpj_norm.isdigit():
            stats['cnpj_pendente'] += 1
            cnpj_pendente_nomes.append(ticket['nome_contato'])
            continue

        # --- RESULTADO mapping ---
        resultado = map_deskrio_resultado(ticket['status'], ticket['motivo'])
        motivo_log = map_deskrio_motivo_log(ticket['motivo'])

        # --- TIPO ACAO ---
        # Deskrio 'Origem' maps to TIPO ACAO:
        #   'Receptivo' -> RECEPTIVO (cliente iniciou o chat)
        #   'Ativo' -> ATIVO (consultor iniciou)
        origem = ticket['origem'].upper()
        tipo_acao = 'RECEPTIVO' if origem == 'RECEPTIVO' else 'ATIVO'

        # --- Consultor normalization ---
        consultor = normalize_consultor(ticket['atendente'])
        if not consultor:
            stats['invalid_atendente'] += 1
            consultor = 'ADMINISTRADOR'

        # --- NOTA DO DIA ---
        nota = build_nota_deskrio(
            ticket['nome_contato'],
            ticket['motivo'],
            ticket['observacao'],
            ticket['protocolo'],
            ticket['status']
        )

        # --- TIPO DO CONTATO (Deskrio = maioria suporte via chat) ---
        tipo_contato_override = 'CONTATOS PASSIVO / SUPORTE'

        # --- Criar registro LOG ---
        record = make_log_record(
            data=data,
            cnpj=cnpj_norm,
            consultor=consultor,
            resultado=resultado,
            nota=nota,
            whatsapp='NAO',      # Deskrio e chat proprio, nao WhatsApp
            ligacao='NAO',       # Deskrio e chat, nao ligacao
            lig_atendida='N/A',
            tipo_acao=tipo_acao,
            motivo=motivo_log,
            mercos='NAO',        # Deskrio nao atualiza Mercos
            nome=nome_fantasia,
            uf=uf,
            rede=rede,
            situacao='',         # Sera preenchido na fase de dedup/validacao
            origem_dado='REAL',
            source='DESKRIO',
        )

        # Zero valores financeiros (Two-Base Architecture)
        # make_log_record ja nao inclui campos financeiros

        records.append(record)
        stats['processados'] += 1
        resultado_dist[resultado] += 1
        consultor_dist[consultor] += 1
        month_dist[data.strftime('%Y-%m')] += 1
        match_type_dist[match_type] += 1

    # ----------------------------------------------------------------
    # 4. Estatisticas
    # ----------------------------------------------------------------
    print(f"\n[4/5] ESTATISTICAS DE PROCESSAMENTO:")
    print(f"  Total tickets lidos:    {stats['total_lidos']}")
    print(f"  Weekend descartados:    {stats['weekend_skipped']}")
    print(f"  Sem data valida:        {stats['no_date']}")
    print(f"  Sem CNPJ (pendentes):   {stats['cnpj_pendente']}")
    print(f"  Processados (LOG):      {stats['processados']}")
    print(f"\n  CNPJ Resolution:")
    print(f"    CNPJ direto:          {stats['cnpj_direto']}")
    print(f"    Match exato:          {stats.get('cnpj_match_exato', 0)}")
    print(f"    Match exato (base):   {stats.get('cnpj_match_exato_base', 0)}")
    print(f"    Match parcial:        {stats.get('cnpj_match_parcial', 0)}")
    print(f"    Match prefixo 3w:     {stats.get('cnpj_match_prefixo_3w', 0)}")
    print(f"    CNPJ pendente:        {stats['cnpj_pendente']}")

    total_matched = stats['processados']
    total_pendente = stats['cnpj_pendente']
    match_rate = (total_matched / stats['total_lidos'] * 100) if stats['total_lidos'] > 0 else 0
    print(f"    Taxa de match:        {match_rate:.1f}%")

    print(f"\n  Match Type Distribution:")
    for mt, count in match_type_dist.most_common():
        print(f"    {mt}: {count}")

    print(f"\n  RESULTADO Distribution:")
    for r, count in resultado_dist.most_common():
        pct = count / stats['processados'] * 100 if stats['processados'] > 0 else 0
        print(f"    {r}: {count} ({pct:.1f}%)")

    print(f"\n  CONSULTOR Distribution:")
    for c, count in consultor_dist.most_common():
        pct = count / stats['processados'] * 100 if stats['processados'] > 0 else 0
        print(f"    {c}: {count} ({pct:.1f}%)")

    print(f"\n  Monthly Distribution:")
    for m in sorted(month_dist.keys()):
        print(f"    {m}: {month_dist[m]}")

    # Top 20 nomes pendentes (para auditoria)
    if cnpj_pendente_nomes:
        pendente_counter = Counter(cnpj_pendente_nomes)
        print(f"\n  Top 20 nomes SEM CNPJ (pendentes):")
        for nome, count in pendente_counter.most_common(20):
            print(f"    {nome}: {count} tickets")

    # ----------------------------------------------------------------
    # 5. Salvar JSON
    # ----------------------------------------------------------------
    print(f"\n[5/5] Salvando {OUTPUT_FILE.name}...")

    output = {
        'metadata': {
            'source': 'DESKRIO',
            'source_file': DESKRIO_FILE.name,
            'total': stats['processados'],
            'total_lidos': stats['total_lidos'],
            'cnpj_direto': stats['cnpj_direto'],
            'cnpj_matched': (
                stats.get('cnpj_match_exato', 0) +
                stats.get('cnpj_match_exato_base', 0) +
                stats.get('cnpj_match_parcial', 0) +
                stats.get('cnpj_match_prefixo_3w', 0)
            ),
            'cnpj_pendente': stats['cnpj_pendente'],
            'weekends_skipped': stats['weekend_skipped'],
            'no_date': stats['no_date'],
            'match_rate_pct': round(match_rate, 1),
            'date_processed': date.today().strftime('%Y-%m-%d'),
            'resultado_distribution': dict(resultado_dist),
            'consultor_distribution': dict(consultor_dist),
            'monthly_distribution': dict(sorted(month_dist.items())),
            'match_type_distribution': dict(match_type_dist),
        },
        'records': records,
    }

    # Garantir diretorio de saida existe
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    with open(str(OUTPUT_FILE), 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    print(f"  Salvo: {OUTPUT_FILE}")
    print(f"  {stats['processados']} registros Deskrio no formato LOG")

    print("\n" + "=" * 70)
    print("PROCESSAMENTO DESKRIO COMPLETO")
    print("=" * 70)

    return output


# =====================================================================
# MAIN
# =====================================================================

if __name__ == '__main__':
    process_deskrio()

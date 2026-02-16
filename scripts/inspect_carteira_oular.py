"""
Inspect CARTEIRA DE CLIENTES OULAR.xlsx thoroughly.
Lists sheets, headers, data samples, merged cells, validations, named ranges, etc.
"""

import openpyxl
import sys
import os

FILE_PATH = r"C:\Users\User\OneDrive\Área de Trabalho\CARTEIRA DE CLIENTES OULAR.xlsx"

# Columns of special interest
SPECIAL_COLS = [
    "CNPJ", "NOME FANTASIA", "UF", "CONSULTOR", "SITUAÇÃO", "SITUACAO",
    "META", "REALIZADO", "PROJEÇÃO", "PROJECAO", "SINALEIRO", "SEMÁFORO",
    "SEMAFORO", "RAZÃO SOCIAL", "RAZAO SOCIAL", "CIDADE", "ESTADO",
    "TELEFONE", "EMAIL", "SEGMENTO", "CANAL", "REDE", "STATUS",
]

def safe_str(val):
    """Convert value to string safely, handling None and encoding issues."""
    if val is None:
        return "<vazio>"
    try:
        return str(val)
    except Exception:
        return "<erro_encoding>"

def inspect_sheet(ws, sheet_name, is_data_only=False):
    tag = "[data_only]" if is_data_only else "[formulas]"
    print(f"\n{'='*100}")
    print(f"  SHEET: '{sheet_name}'  {tag}")
    print(f"{'='*100}")

    # Dimensions
    print(f"\n  Dimensions string : {ws.dimensions}")
    print(f"  min_row={ws.min_row}, max_row={ws.max_row}, min_col={ws.min_column}, max_col={ws.max_column}")

    row_count = ws.max_row if ws.max_row else 0
    col_count = ws.max_column if ws.max_column else 0
    print(f"  Total rows: {row_count}")
    print(f"  Total cols: {col_count}")

    if row_count == 0 or col_count == 0:
        print("  >> Sheet appears empty.")
        return

    # --- Headers (rows 1 and 2) ---
    print(f"\n  --- ROW 1 (Headers) ---")
    row1 = []
    for col_idx in range(1, col_count + 1):
        val = ws.cell(row=1, column=col_idx).value
        row1.append(val)
        print(f"    Col {col_idx:3d} ({openpyxl.utils.get_column_letter(col_idx):>4s}): {safe_str(val)}")

    if row_count >= 2:
        print(f"\n  --- ROW 2 (Sub-headers or first data) ---")
        for col_idx in range(1, col_count + 1):
            val = ws.cell(row=2, column=col_idx).value
            print(f"    Col {col_idx:3d} ({openpyxl.utils.get_column_letter(col_idx):>4s}): {safe_str(val)}")

    # --- First 5 data rows (starting from row 2 or 3) ---
    data_start = 2
    data_end = min(data_start + 4, row_count)
    print(f"\n  --- DATA ROWS {data_start} to {data_end} (all columns) ---")
    for r in range(data_start, data_end + 1):
        print(f"\n    >> Row {r}:")
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            header = safe_str(row1[c-1]) if c-1 < len(row1) else f"col{c}"
            cell_type = type(val).__name__ if val is not None else "NoneType"
            # Check if it's a formula (in non-data_only mode)
            is_formula = isinstance(val, str) and val.startswith("=")
            fmt_extra = ""
            if cell.number_format and cell.number_format != "General":
                fmt_extra = f"  [fmt: {cell.number_format}]"
            if is_formula:
                fmt_extra += "  [FORMULA]"
            print(f"      {header:>30s} (Col {c:3d}): {safe_str(val):50s}  type={cell_type}{fmt_extra}")

    # --- Column widths ---
    print(f"\n  --- COLUMN WIDTHS ---")
    widths_found = 0
    for col_letter in [openpyxl.utils.get_column_letter(i) for i in range(1, col_count + 1)]:
        cd = ws.column_dimensions.get(col_letter)
        if cd and cd.width:
            print(f"    {col_letter}: width={cd.width}, hidden={cd.hidden}")
            widths_found += 1
    if widths_found == 0:
        print("    (no custom column widths set)")

    # --- Merged cells ---
    print(f"\n  --- MERGED CELLS ---")
    merged = list(ws.merged_cells.ranges)
    if merged:
        for m in merged:
            print(f"    {m}")
        print(f"    Total merged regions: {len(merged)}")
    else:
        print("    (none)")

    # --- Data validations ---
    print(f"\n  --- DATA VALIDATIONS ---")
    if ws.data_validations and ws.data_validations.dataValidation:
        for dv in ws.data_validations.dataValidation:
            print(f"    Type={dv.type}, Formula1={dv.formula1}, Ranges={dv.sqref}, AllowBlank={dv.allow_blank}")
        print(f"    Total validations: {len(ws.data_validations.dataValidation)}")
    else:
        print("    (none)")

    # --- Conditional formatting ---
    print(f"\n  --- CONDITIONAL FORMATTING ---")
    if ws.conditional_formatting:
        count = 0
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                print(f"    Range={cf.sqref}, Type={rule.type}, Priority={rule.priority}, Formula={getattr(rule, 'formula', None)}")
                count += 1
        print(f"    Total CF rules: {count}")
    else:
        print("    (none)")

    # --- Auto filters ---
    print(f"\n  --- AUTO FILTER ---")
    if ws.auto_filter and ws.auto_filter.ref:
        print(f"    Filter ref: {ws.auto_filter.ref}")
    else:
        print("    (none)")

    # --- Freeze panes ---
    print(f"\n  --- FREEZE PANES ---")
    if ws.freeze_panes:
        print(f"    Frozen at: {ws.freeze_panes}")
    else:
        print("    (none)")

    # --- Look for special columns ---
    print(f"\n  --- SPECIAL COLUMNS SEARCH ---")
    found_special = {}
    for c in range(1, col_count + 1):
        val = safe_str(ws.cell(row=1, column=c).value).strip().upper()
        for sc in SPECIAL_COLS:
            if sc.upper() in val or val in sc.upper():
                if len(val) > 1:  # skip empty matches
                    found_special[val] = c
    if found_special:
        for name, idx in sorted(found_special.items(), key=lambda x: x[1]):
            letter = openpyxl.utils.get_column_letter(idx)
            print(f"    Found '{name}' at column {idx} ({letter})")
    else:
        print("    (no special columns matched in row 1)")

    return row1


def main():
    if not os.path.exists(FILE_PATH):
        print(f"ERROR: File not found: {FILE_PATH}")
        sys.exit(1)

    print(f"File: {FILE_PATH}")
    print(f"File size: {os.path.getsize(FILE_PATH):,} bytes")

    # Load with data_only=True (to see computed values)
    print("\n\n" + "#"*100)
    print("  LOADING WITH data_only=True (computed values)")
    print("#"*100)
    wb_data = openpyxl.load_workbook(FILE_PATH, data_only=True)
    print(f"\nSheet names: {wb_data.sheetnames}")
    print(f"Number of sheets: {len(wb_data.sheetnames)}")

    # Named ranges
    print(f"\n--- NAMED RANGES (WORKBOOK LEVEL) ---")
    if wb_data.defined_names:
        for dn in wb_data.defined_names.values():
            print(f"  Name='{dn.name}', Value='{dn.attr_text}'")
    else:
        print("  (none)")

    all_headers = {}
    for sn in wb_data.sheetnames:
        ws = wb_data[sn]
        hdrs = inspect_sheet(ws, sn, is_data_only=True)
        all_headers[sn] = hdrs
    wb_data.close()

    # Load without data_only (to see formulas)
    print("\n\n" + "#"*100)
    print("  LOADING WITH data_only=False (formulas visible)")
    print("#"*100)
    wb_form = openpyxl.load_workbook(FILE_PATH, data_only=False)

    for sn in wb_form.sheetnames:
        ws = wb_form[sn]
        inspect_sheet(ws, sn, is_data_only=False)
    wb_form.close()

    # --- SUMMARY ---
    print("\n\n" + "#"*100)
    print("  SUMMARY")
    print("#"*100)
    print(f"\nFile: {FILE_PATH}")
    print(f"Sheets: {list(all_headers.keys())}")
    for sn, hdrs in all_headers.items():
        if hdrs:
            non_empty = [h for h in hdrs if h is not None]
            print(f"\n  Sheet '{sn}': {len(hdrs)} columns total, {len(non_empty)} with headers")
            print(f"    Headers: {[safe_str(h) for h in hdrs if h is not None]}")

    print("\n\nDone.")


if __name__ == "__main__":
    main()

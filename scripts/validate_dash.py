"""Validate rebuilt DASH tab in V2.xlsx"""
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('output/CRM INTELIGENTE - VITAO360 V2.xlsx', data_only=False)
ws = wb['DASH']

print('=== DASH TAB VALIDATION ===')
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
print(f'Freeze: {ws.freeze_panes}')
print(f'Tab color: {ws.sheet_properties.tabColor}')
print(f'Zoom: {ws.sheet_view.zoomScale}')

# Check row 1 (header)
print('\n--- ROW 1 (CABECALHO) ---')
for c in range(1, 8):
    v = ws.cell(row=1, column=c).value
    if v:
        print(f'  {get_column_letter(c)}1: {v}')

# Check data validations
dvs = list(ws.data_validations.dataValidation) if ws.data_validations else []
print(f'\nData Validations: {len(dvs)}')
for dv in dvs:
    print(f'  {dv.sqref} -> {dv.formula1}')

# Block 1 check
print('\n--- BLOCO 1: rows 4-13 ---')
for r in [4, 5, 6, 13]:
    vals = []
    for c in range(1, 14):
        v = ws.cell(row=r, column=c).value
        if v:
            vals.append(f'{get_column_letter(c)}:{str(v)[:50]}')
    print(f'  Row {r}: {vals}')

# Block 2 check
print('\n--- BLOCO 2: rows 18-28 ---')
for r in [18, 19, 20, 21, 28]:
    vals = []
    for c in range(1, 18):
        v = ws.cell(row=r, column=c).value
        if v:
            vals.append(f'{get_column_letter(c)}:{str(v)[:50]}')
    print(f'  Row {r}: {vals}')

# Block 3 check
print('\n--- BLOCO 3: rows 33-46 ---')
for r in [33, 34, 35, 45, 39, 41, 42, 46]:
    vals = []
    for c in range(1, 18):
        v = ws.cell(row=r, column=c).value
        if v:
            vals.append(f'{get_column_letter(c)}:{str(v)[:50]}')
    print(f'  Row {r}: {vals}')

# Charts
print(f'\nCharts: {len(ws._charts)}')
for ch in ws._charts:
    print(f'  Title: {ch.title}, Type: {ch.type}, Width: {ch.width}, Height: {ch.height}')

# Check separator columns
print(f'\nSeparator cols:')
print(f'  G={ws.column_dimensions["G"].width}')
print(f'  H={ws.column_dimensions["H"].width}')
print(f'  L={ws.column_dimensions["L"].width}')
print(f'  O={ws.column_dimensions["O"].width}')

# Verify formulas have proper structure
print('\n--- SAMPLE FORMULA CHECK ---')
c6 = ws.cell(row=6, column=3).value
print(f'C6 (PROSP/ORCAM): {c6[:100] if c6 else None}')
b21 = ws.cell(row=21, column=2).value
print(f'B21 (TOTAL): {b21[:100] if b21 else None}')
b35 = ws.cell(row=35, column=2).value
print(f'B35 (MOT QTD): {b35[:100] if b35 else None}')
j35 = ws.cell(row=35, column=10).value
print(f'J35 (IND CONTATOS): {j35[:100] if j35 else None}')
n35 = ws.cell(row=35, column=14).value
print(f'N35 (% CONV): {n35[:100] if n35 else None}')

# Check FOLLOW UP formula in M column (should be FU7+FU15)
m6 = ws.cell(row=6, column=13).value
print(f'M6 (FOLLOW UP): {m6[:120] if m6 else None}')

# Verify other sheets not touched
print('\n--- OTHER SHEETS ---')
print(f'Sheets: {wb.sheetnames}')
ws_d2 = wb['DRAFT 2']
print(f'DRAFT 2: rows={ws_d2.max_row}, cols={ws_d2.max_column}, freeze={ws_d2.freeze_panes}')
ws_r = wb['REGRAS']
print(f'REGRAS: rows={ws_r.max_row}, cols={ws_r.max_column}')

# Check all tipos are present in Block 1
print('\n--- TIPO DO CONTATO values (Block 1) ---')
for r in range(6, 13):
    v = ws.cell(row=r, column=1).value
    print(f'  A{r}: {v}')

# Check all motivos in Block 3
print('\n--- MOTIVOS values (Block 3) ---')
for r in range(35, 45):
    v = ws.cell(row=r, column=1).value
    print(f'  A{r}: {v}')

# Check consultores in INDICADORES
print('\n--- CONSULTORES (INDICADORES) ---')
for r in range(35, 39):
    v = ws.cell(row=r, column=9).value
    print(f'  I{r}: {v}')

wb.close()
print('\nVALIDATION COMPLETE')

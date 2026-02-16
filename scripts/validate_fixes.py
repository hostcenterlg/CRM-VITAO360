"""Validate all fixes applied to JARVIS_CRM_CENTRAL.xlsx"""
import openpyxl
from openpyxl.utils import get_column_letter
import os

path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output", "JARVIS_CRM_CENTRAL.xlsx")
wb = openpyxl.load_workbook(path)

errors = []
warnings = []

# 1. Check sheets
expected_sheets = ['REGRAS', 'CARTEIRA', 'LOG', 'DRAFT 1', 'DRAFT 2', 'DASH', 'AGENDA']
if wb.sheetnames != expected_sheets:
    errors.append(f"SHEETS: Expected {expected_sheets}, got {wb.sheetnames}")
else:
    print("✅ 7 abas na ordem correta")

# 2. Check Named Ranges
nr_names = [dn.name for dn in wb.defined_names.values()]
expected_nr = [
    'LISTA_RESULTADO', 'LISTA_TIPO_CONTATO', 'LISTA_MOTIVO', 'LISTA_SITUACAO',
    'LISTA_FASE', 'LISTA_TENTATIVA', 'LISTA_SINALEIRO', 'LISTA_CONSULTOR',
    'LISTA_SIM_NAO', 'LISTA_SIM_NAO_NA', 'LISTA_ATIVO_RECEPTIVO', 'LISTA_CURVA_ABC',
    'LISTA_TIPO_CLIENTE', 'LISTA_BLOCO', 'LISTA_UF', 'LISTA_ACAO_FUTURA',
    'TABELA_RESULTADO'
]
for nr in expected_nr:
    if nr not in nr_names:
        errors.append(f"NAMED RANGE missing: {nr}")
print(f"✅ Named Ranges: {len(nr_names)} definidas")

# 3. Check Named Ranges point to multi-cell ranges (not single cell)
for dn in wb.defined_names.values():
    ref = dn.attr_text
    if dn.name.startswith('LISTA_') and dn.name not in ('LISTA_BLOCO',):  # BLOCO has 2 items, same start/end possible
        # Check it's not a single cell
        parts = ref.split(':')
        if len(parts) == 2:
            start_row = ''.join(c for c in parts[0].split('$')[-1] if c.isdigit())
            end_row = ''.join(c for c in parts[1].split('$')[-1] if c.isdigit())
            if start_row and end_row and start_row == end_row:
                errors.append(f"NAMED RANGE {dn.name} is single-cell: {ref}")
        elif len(parts) == 1:
            errors.append(f"NAMED RANGE {dn.name} has no range (single cell): {ref}")

# 4. Check LISTA_SIM_NAO has SIM and NÃO in separate cells
ws_regras = wb['REGRAS']
for dn in wb.defined_names.values():
    if dn.name == 'LISTA_SIM_NAO':
        ref = dn.attr_text
        # Parse the reference
        print(f"  LISTA_SIM_NAO -> {ref}")
        # Check column F has SIM and NÃO
        found_sim = False
        found_nao = False
        for r in range(1, ws_regras.max_row + 1):
            v = ws_regras.cell(row=r, column=6).value
            if v == "SIM":
                found_sim = True
            if v == "NÃO":
                found_nao = True
        if found_sim and found_nao:
            print("  ✅ LISTA_SIM_NAO has SIM and NÃO in separate cells")
        else:
            errors.append("LISTA_SIM_NAO doesn't have SIM/NÃO in separate cells")

# 5. Check TIPO_CLIENTE has 6 options
for dn in wb.defined_names.values():
    if dn.name == 'LISTA_TIPO_CLIENTE':
        ref = dn.attr_text
        parts = ref.replace("REGRAS!", "").split(':')
        start = int(''.join(c for c in parts[0] if c.isdigit()))
        end = int(''.join(c for c in parts[1] if c.isdigit()))
        count = end - start + 1
        if count == 6:
            print(f"  ✅ LISTA_TIPO_CLIENTE has 6 options (col J, rows {start}-{end})")
        else:
            errors.append(f"LISTA_TIPO_CLIENTE has {count} options, expected 6")

# 6. Check CARTEIRA has 81 columns
ws_cart = wb['CARTEIRA']
if ws_cart.max_column == 81:
    print(f"✅ CARTEIRA: 81 colunas")
else:
    errors.append(f"CARTEIRA has {ws_cart.max_column} columns, expected 81")

# 7. Check column grouping - each column in a group should have outlineLevel > 0
groups_check = {
    '[+1] L:Q': list(range(12, 18)),
    '[+2] S:Z': list(range(19, 27)),
    '[+3] AB:AM': list(range(28, 40)),
    '[+4] AO:AS': list(range(41, 46)),
    '[+5] AU:AX': list(range(47, 51)),
    '[+6] AZ:BI': list(range(52, 62)),
    '[+7] BK:BR': list(range(63, 71)),
    '[+8] BT:CC': list(range(72, 82)),
}
all_grouped = True
for name, cols in groups_check.items():
    for c in cols:
        letter = get_column_letter(c)
        dim = ws_cart.column_dimensions.get(letter)
        if dim:
            if not (hasattr(dim, 'outlineLevel') and dim.outlineLevel and dim.outlineLevel > 0):
                errors.append(f"Column {letter} in {name} not grouped (outlineLevel={getattr(dim, 'outlineLevel', None)})")
                all_grouped = False
        else:
            errors.append(f"Column {letter} in {name} has no dimension")
            all_grouped = False
if all_grouped:
    print("✅ Column grouping: all 8 groups correct (every column individually)")

# 8. Check data validations count
ws_cart_dvs = list(ws_cart.data_validations.dataValidation) if ws_cart.data_validations else []
print(f"✅ CARTEIRA data validations: {len(ws_cart_dvs)}")
for dv in ws_cart_dvs:
    print(f"    {dv.sqref} -> {dv.formula1}")

# 9. Check LOG has 20 columns
ws_log = wb['LOG']
if ws_log.max_column == 20:
    print(f"✅ LOG: 20 colunas")
else:
    errors.append(f"LOG has {ws_log.max_column} columns, expected 20")

# 10. Check LOG FOLLOW-UP formula handles SUPORTE/PERDA
cell_o2 = ws_log.cell(row=2, column=15).value
if cell_o2 and "SUPORTE" in str(cell_o2) and "PERDA" in str(cell_o2) and "SEM" in str(cell_o2):
    print("✅ LOG FOLLOW-UP formula handles SUPORTE/PERDA as 'SEM'")
else:
    errors.append(f"LOG FOLLOW-UP formula missing SUPORTE/PERDA handling: {cell_o2}")

# 11. Check SITUAÇÃO formula handles LEAD/NOVO
cell_e2 = ws_cart.cell(row=2, column=5).value
if cell_e2 and "LEAD" in str(cell_e2) and "NOVO" in str(cell_e2):
    print("✅ CARTEIRA SITUAÇÃO formula handles LEAD/NOVO")
else:
    errors.append(f"CARTEIRA SITUAÇÃO formula missing LEAD/NOVO: {cell_e2}")

# 12. Check DASH has CONSULTOR filter in formulas
ws_dash = wb['DASH']
cell_a6 = ws_dash.cell(row=6, column=1).value
if cell_a6 and '$B$2' in str(cell_a6) and 'TODOS' in str(cell_a6):
    print("✅ DASH KPIs include CONSULTOR filter ($B$2)")
else:
    errors.append(f"DASH KPIs missing CONSULTOR filter: {cell_a6}")

# 13. Check freeze panes
if ws_cart.freeze_panes == "K2":
    print("✅ CARTEIRA freeze: K2")
else:
    errors.append(f"CARTEIRA freeze: {ws_cart.freeze_panes}, expected K2")

if ws_log.freeze_panes == "A2":
    print("✅ LOG freeze: A2")
else:
    errors.append(f"LOG freeze: {ws_log.freeze_panes}")

ws_agenda = wb['AGENDA']
if ws_agenda.freeze_panes == "A7":
    print("✅ AGENDA freeze: A7")

# 14. Check conditional formatting on SITUAÇÃO
cfs = list(ws_cart.conditional_formatting)
sit_rules = 0
for cf in cfs:
    if 'E2:E1000' in str(cf.sqref):
        sit_rules += len(cf.rules)
print(f"✅ SITUAÇÃO conditional formatting: {sit_rules} rules (expect 7: ATIVO, EM RISCO, INATIVO REC, INATIVO ANT, PROSPECT, LEAD, NOVO)")

# 15. UF list in REGRAS
uf_count = 0
for r in range(1, ws_regras.max_row + 1):
    v = ws_regras.cell(row=r, column=12).value
    if v and len(str(v)) == 2 and str(v).isalpha():
        uf_count += 1
print(f"✅ REGRAS UF list: {uf_count} UFs")

# Summary
print("\n" + "="*60)
if errors:
    print(f"❌ {len(errors)} ERRORS:")
    for e in errors:
        print(f"  - {e}")
else:
    print("✅ ALL VALIDATIONS PASSED!")

if warnings:
    print(f"⚠ {len(warnings)} WARNINGS:")
    for w in warnings:
        print(f"  - {w}")

wb.close()

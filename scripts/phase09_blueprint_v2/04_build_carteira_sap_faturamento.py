"""
Phase 09 Plan 04: SAP + FATURAMENTO Block Builder for V13 CARTEIRA

Task 1: Injects SAP (BK-BP) and DADOS CADASTRAIS SAP (BQ-BY) static values
Task 2: Injects FATURAMENTO mega-block formulas (BZ-JC, 186 columns x 554 rows)

Static data approach for SAP blocks (per plan: SAP data is relatively static).
Formula approach for FATURAMENTO (META, REALIZADO, JUSTIFICATIVA all dynamic).
"""

import json
import re
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path

# Paths
BASE = Path("c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360")
V13_PATH = BASE / "data/output/CRM_VITAO360_V13_PROJECAO.xlsx"
SAP_PATH = BASE / "data/sources/sap/01_SAP_CONSOLIDADO.xlsx"
COL_SPEC_PATH = BASE / "data/output/phase09/carteira_column_spec.json"

# Constants
DATA_START_ROW = 4
DATA_END_ROW = 557
NUM_ROWS = DATA_END_ROW - DATA_START_ROW + 1  # 554

# PROJECAO sheet name (with cedilla + trailing space)
PROJECAO_SHEET = "PROJEÇÃO "

# DRAFT 1 monthly sales columns (in DRAFT 1 tab)
# MAR/25=U, APR/25=V, MAI/25=W, JUN/25=X, JUL/25=Y, AGO/25=Z,
# SET/25=AA, OUT/25=AB, NOV/25=AC, DEZ/25=AD, JAN/26=AE, FEV/26=AF
DRAFT1_MONTH_COLS = {
    1: "AE",   # JAN -> maps to JAN/26 in DRAFT 1
    2: "AF",   # FEV -> maps to FEV/26 in DRAFT 1
    3: "U",    # MAR -> maps to MAR/25 in DRAFT 1
    4: "V",    # APR -> maps to ABR/25 in DRAFT 1
    5: "W",    # MAI -> maps to MAI/25 in DRAFT 1
    6: "X",    # JUN -> maps to JUN/25 in DRAFT 1
    7: "Y",    # JUL -> maps to JUL/25 in DRAFT 1
    8: "Z",    # AGO -> maps to AGO/25 in DRAFT 1
    9: "AA",   # SET -> maps to SET/25 in DRAFT 1
    10: "AB",  # OUT -> maps to OUT/25 in DRAFT 1
    11: "AC",  # NOV -> maps to NOV/25 in DRAFT 1
    12: "AD",  # DEZ -> maps to DEZ/25 in DRAFT 1
}

# Year for FATURAMENTO formulas (confirmed by V12 COUNTIFS using DATE(2026,...))
FAT_YEAR = 2026

# Month order in FATURAMENTO block (JAN first, DEZ last)
MONTH_NAMES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN",
               "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]

# Quarter anchors: %Q1=CB(80), %Q2=DV(126), %Q3=FP(172), %Q4=HJ(218)
# Month %YTD starts: JAN=CC(81), FEV=CR(96), MAR=DG(111), ABR=DW(127),
#   MAI=EL(142), JUN=FA(157), JUL=FQ(173), AGO=GF(188), SET=GU(203),
#   OUT=HK(219), NOV=HZ(234), DEZ=IO(249)
MONTH_START_COLS = {
    1: 81,   # JAN starts at CC (col 81)
    2: 96,   # FEV starts at CR (col 96)
    3: 111,  # MAR starts at DG (col 111)
    4: 127,  # APR starts at DW (col 127)
    5: 142,  # MAI starts at EL (col 142)
    6: 157,  # JUN starts at FA (col 157)
    7: 173,  # JUL starts at FQ (col 173)
    8: 188,  # AGO starts at GF (col 188)
    9: 203,  # SET starts at GU (col 203)
    10: 219, # OUT starts at HK (col 219)
    11: 234, # NOV starts at HZ (col 234)
    12: 249, # DEZ starts at IO (col 249)
}

QUARTER_ANCHOR_COLS = {
    1: 80,   # %Q1 at CB (col 80)
    2: 126,  # %Q2 at DV (col 126)
    3: 172,  # %Q3 at FP (col 172)
    4: 218,  # %Q4 at HJ (col 218)
}


def bounded_ref(formula: str) -> str:
    """Convert full-column references ($X:$X) to bounded ranges ($X$3:$X$25000)."""
    return re.sub(
        r"\$([A-Z]{1,3}):\$([A-Z]{1,3})",
        lambda m: f"${m.group(1)}$3:${m.group(2)}$25000"
            if m.group(1) == m.group(2)
            else m.group(0),
        formula
    )


def clean_cnpj(raw):
    """Strip dots, dashes, slashes from CNPJ and pad to 14 digits."""
    if raw is None:
        return None
    s = str(raw).strip()
    s = re.sub(r"[.\-/\s]", "", s)
    if s.isdigit():
        return s.zfill(14)
    return None


def load_sap_cadastro():
    """Load SAP Cadastro data indexed by clean CNPJ."""
    wb = openpyxl.load_workbook(str(SAP_PATH), data_only=True, read_only=True)
    ws = wb["Cadastro Clientes SAP"]

    sap_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[4] is None:
            continue
        cnpj = clean_cnpj(row[4])  # Col E = CNPJ Cliente
        if cnpj is None:
            continue

        # Determine CADASTRO status from Status Parceiros (col B)
        status_raw = str(row[1]) if row[1] else ""
        # SAP status icons: @08@ = active, etc.
        # Use ORIGEM col (A) for clearer mapping
        origem = str(row[0]) if row[0] else ""
        if "SEM_ATEND" in origem.upper():
            atendimento = "SEM ATENDIMENTO"
        else:
            atendimento = "COM ATENDIMENTO"

        # Bloqueio: check cols S/T/U (bloqueio areas vendas/todas/central)
        bloq_vendas = row[18]
        bloq_todas = row[19]
        bloq_central = row[20]
        has_bloqueio = any(b is not None and str(b).strip() != "" for b in [bloq_vendas, bloq_todas, bloq_central])

        if has_bloqueio:
            cadastro = "BLOQUEADO"
            bloqueio = "SIM"
        else:
            cadastro = "ATIVO"
            bloqueio = "NAO"

        entry = {
            "codigo": str(row[2]) if row[2] else "",          # Col C = Código do Cliente
            "cnpj": cnpj,
            "razao_social": str(row[5]) if row[5] else "",    # Col F = Nome Cliente
            "cadastro": cadastro,
            "atendimento": atendimento,
            "bloqueio": bloqueio,
            # DADOS CADASTRAIS SAP fields
            "desc_grupo_cliente": str(row[17]) if row[17] else "",    # Col R
            "zp_gerente": str(row[22]) if row[22] else "",            # Col W
            "zr_representante": str(row[24]) if row[24] else "",      # Col Y
            "zv_vend_interno": str(row[28]) if row[28] else "",       # Col AC
            "canal": str(row[32]) if row[32] else "",                 # Col AG
            "tipo_cliente": str(row[34]) if row[34] else "",          # Col AI
            "macroregiao": str(row[36]) if row[36] else "",           # Col AK
            "microregiao": str(row[38]) if row[38] else "",           # Col AM
            "grupo_chave": str(row[42]) if row[42] else "",           # Col AQ
        }

        # First entry per CNPJ wins (prefer COM_VENDA over SEM_ATEND)
        if cnpj not in sap_data:
            sap_data[cnpj] = entry

    wb.close()
    return sap_data


def inject_sap_blocks(ws, sap_data):
    """Inject SAP (BK-BP) and DADOS CADASTRAIS SAP (BQ-BY) static values."""
    stats = {"sap_populated": 0, "sap_empty": 0, "dados_populated": 0}

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        # Get CNPJ from col B
        cnpj_cell = ws.cell(row=row, column=2).value
        if cnpj_cell is None:
            stats["sap_empty"] += 1
            continue
        cnpj = str(cnpj_cell).strip().zfill(14)

        sap = sap_data.get(cnpj)
        if sap is None:
            stats["sap_empty"] += 1
            continue

        stats["sap_populated"] += 1

        # SAP block (BK=63, BL=64, BM=65, BN=66, BO=67, BP=68)
        ws.cell(row=row, column=63, value=sap["codigo"])           # BK: CODIGO DO CLIENTE
        ws.cell(row=row, column=64, value=cnpj)                    # BL: CNPJ
        ws.cell(row=row, column=65, value=sap["razao_social"])     # BM: RAZAO SOCIAL
        ws.cell(row=row, column=66, value=sap["cadastro"])         # BN: CADASTRO
        ws.cell(row=row, column=67, value=sap["atendimento"])      # BO: ATENDIMENTO
        ws.cell(row=row, column=68, value=sap["bloqueio"])         # BP: BLOQUEIO

        # DADOS CADASTRAIS SAP (BQ=69, BR=70, BS=71, BT=72, BU=73, BV=74, BW=75, BX=76, BY=77)
        ws.cell(row=row, column=69, value=sap["desc_grupo_cliente"])   # BQ
        ws.cell(row=row, column=70, value=sap["zp_gerente"])           # BR
        ws.cell(row=row, column=71, value=sap["zr_representante"])     # BS
        ws.cell(row=row, column=72, value=sap["zv_vend_interno"])      # BT
        ws.cell(row=row, column=73, value=sap["canal"])                # BU
        ws.cell(row=row, column=74, value=sap["tipo_cliente"])         # BV
        ws.cell(row=row, column=75, value=sap["macroregiao"])          # BW
        ws.cell(row=row, column=76, value=sap["microregiao"])          # BX
        ws.cell(row=row, column=77, value=sap["grupo_chave"])          # BY

        stats["dados_populated"] += 1

    return stats


def get_col_letter(col_num):
    """Get Excel column letter from 1-based column number."""
    return get_column_letter(col_num)


def write_faturamento_month(ws, month_num, year):
    """
    Write all 15 sub-columns for one month across all 554 data rows.

    15 sub-columns per month (in order from month start):
    0: %YTD
    1: META (YTD cumulative)
    2: REALIZADO (YTD cumulative)
    3: %TRI
    4: META (quarter cumulative)
    5: REALIZADO (quarter cumulative)
    6: %MES
    7: META MES
    8: REALIZADO MES
    9: DATA PEDIDO
    10: JUSTIFICATIVA S1
    11: JUSTIFICATIVA S2
    12: JUSTIFICATIVA S3
    13: JUSTIFICATIVA S4
    14: JUSTIFICATIVA MENSAL
    """
    start_col = MONTH_START_COLS[month_num]
    draft1_col = DRAFT1_MONTH_COLS[month_num]
    quarter = (month_num - 1) // 3 + 1
    month_in_quarter = (month_num - 1) % 3  # 0, 1, 2

    # Column letters for this month's sub-columns
    ytd_pct_col = get_col_letter(start_col + 0)       # %YTD
    meta_ytd_col = get_col_letter(start_col + 1)       # META YTD
    real_ytd_col = get_col_letter(start_col + 2)       # REALIZADO YTD
    tri_pct_col = get_col_letter(start_col + 3)        # %TRI
    meta_tri_col = get_col_letter(start_col + 4)       # META TRI
    real_tri_col = get_col_letter(start_col + 5)       # REALIZADO TRI
    mes_pct_col = get_col_letter(start_col + 6)        # %MES
    meta_mes_col = get_col_letter(start_col + 7)       # META MES
    real_mes_col = get_col_letter(start_col + 8)       # REALIZADO MES
    data_ped_col = get_col_letter(start_col + 9)       # DATA PEDIDO
    just_s1_col = get_col_letter(start_col + 10)       # JUSTIFICATIVA S1
    just_s2_col = get_col_letter(start_col + 11)       # JUSTIFICATIVA S2
    just_s3_col = get_col_letter(start_col + 12)       # JUSTIFICATIVA S3
    just_s4_col = get_col_letter(start_col + 13)       # JUSTIFICATIVA S4
    just_men_col = get_col_letter(start_col + 14)      # JUSTIFICATIVA MENSAL

    # Collect META MES and REALIZADO MES column letters for all previous months
    # (needed for YTD and TRI cumulative sums)
    all_meta_mes_cols = []
    all_real_mes_cols = []
    for m in range(1, month_num + 1):
        sc = MONTH_START_COLS[m]
        all_meta_mes_cols.append(get_col_letter(sc + 7))
        all_real_mes_cols.append(get_col_letter(sc + 8))

    # TRI: only months in the current quarter
    quarter_start_month = (quarter - 1) * 3 + 1
    tri_meta_cols = []
    tri_real_cols = []
    for m in range(quarter_start_month, month_num + 1):
        sc = MONTH_START_COLS[m]
        tri_meta_cols.append(get_col_letter(sc + 7))
        tri_real_cols.append(get_col_letter(sc + 8))

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        r = row  # shorthand

        # --- 7: META MES = PROJECAO col L / 12 ---
        meta_mes_formula = f"=IFERROR(INDEX('{PROJECAO_SHEET}'!$L$4:$L$600,MATCH($B{r},'{PROJECAO_SHEET}'!$B$4:$B$600,0))/12,0)"
        ws.cell(row=r, column=start_col + 7, value=meta_mes_formula)

        # --- 8: REALIZADO MES = INDEX/MATCH from DRAFT 1 ---
        real_mes_formula = f"=IFERROR(INDEX('DRAFT 1'!${draft1_col}$3:${draft1_col}$25000,MATCH($B{r},'DRAFT 1'!$B$3:$B$25000,0)),0)"
        ws.cell(row=r, column=start_col + 8, value=real_mes_formula)

        # --- 6: %MES = REALIZADO MES / META MES ---
        mes_pct_formula = f'=IF({meta_mes_col}{r}=0,"",{real_mes_col}{r}/{meta_mes_col}{r})'
        ws.cell(row=r, column=start_col + 6, value=mes_pct_formula)

        # --- 1: META YTD (cumulative sum of all META MES from JAN to this month) ---
        if len(all_meta_mes_cols) == 1:
            meta_ytd_formula = f"={all_meta_mes_cols[0]}{r}"
        else:
            meta_ytd_formula = "=" + "+".join(f"{c}{r}" for c in all_meta_mes_cols)
        ws.cell(row=r, column=start_col + 1, value=meta_ytd_formula)

        # --- 2: REALIZADO YTD (cumulative sum of all REALIZADO MES from JAN to this month) ---
        if len(all_real_mes_cols) == 1:
            real_ytd_formula = f"={all_real_mes_cols[0]}{r}"
        else:
            real_ytd_formula = "=" + "+".join(f"{c}{r}" for c in all_real_mes_cols)
        ws.cell(row=r, column=start_col + 2, value=real_ytd_formula)

        # --- 0: %YTD = REALIZADO YTD / META YTD ---
        ytd_pct_formula = f'=IF({meta_ytd_col}{r}=0,"",{real_ytd_col}{r}/{meta_ytd_col}{r})'
        ws.cell(row=r, column=start_col + 0, value=ytd_pct_formula)

        # --- 4: META TRI (cumulative sum of META MES for months in this quarter) ---
        if len(tri_meta_cols) == 1:
            meta_tri_formula = f"={tri_meta_cols[0]}{r}"
        else:
            meta_tri_formula = "=" + "+".join(f"{c}{r}" for c in tri_meta_cols)
        ws.cell(row=r, column=start_col + 4, value=meta_tri_formula)

        # --- 5: REALIZADO TRI (cumulative sum of REALIZADO MES for months in this quarter) ---
        if len(tri_real_cols) == 1:
            real_tri_formula = f"={tri_real_cols[0]}{r}"
        else:
            real_tri_formula = "=" + "+".join(f"{c}{r}" for c in tri_real_cols)
        ws.cell(row=r, column=start_col + 5, value=real_tri_formula)

        # --- 3: %TRI = REALIZADO TRI / META TRI ---
        tri_pct_formula = f'=IF({meta_tri_col}{r}=0,"",{real_tri_col}{r}/{meta_tri_col}{r})'
        ws.cell(row=r, column=start_col + 3, value=tri_pct_formula)

        # --- 9: DATA PEDIDO = most recent order date from DRAFT 2 within this month ---
        end_day_formula = f"DATE({year},{month_num}+1,0)"  # Last day of month
        data_ped_formula = (
            f'=IFERROR(MAXIFS(\'DRAFT 2\'!$A$3:$A$25000,\'DRAFT 2\'!$D$3:$D$25000,$B{r},'
            f'\'DRAFT 2\'!$A$3:$A$25000,">="&DATE({year},{month_num},1),'
            f'\'DRAFT 2\'!$A$3:$A$25000,"<="&DATE({year},{month_num}+1,0)),"")'
        )
        ws.cell(row=r, column=start_col + 9, value=data_ped_formula)

        # --- 10: JUSTIFICATIVA S1 (days 1-7) ---
        s1_formula = bounded_ref(
            f"=COUNTIFS('DRAFT 2'!$D:$D,$B{r},'DRAFT 2'!$A:$A,\">=\"&DATE({year},{month_num},1),'DRAFT 2'!$A:$A,\"<=\"&DATE({year},{month_num},7))"
        )
        ws.cell(row=r, column=start_col + 10, value=s1_formula)

        # --- 11: JUSTIFICATIVA S2 (days 8-14) ---
        s2_formula = bounded_ref(
            f"=COUNTIFS('DRAFT 2'!$D:$D,$B{r},'DRAFT 2'!$A:$A,\">=\"&DATE({year},{month_num},8),'DRAFT 2'!$A:$A,\"<=\"&DATE({year},{month_num},14))"
        )
        ws.cell(row=r, column=start_col + 11, value=s2_formula)

        # --- 12: JUSTIFICATIVA S3 (days 15-21) ---
        s3_formula = bounded_ref(
            f"=COUNTIFS('DRAFT 2'!$D:$D,$B{r},'DRAFT 2'!$A:$A,\">=\"&DATE({year},{month_num},15),'DRAFT 2'!$A:$A,\"<=\"&DATE({year},{month_num},21))"
        )
        ws.cell(row=r, column=start_col + 12, value=s3_formula)

        # --- 13: JUSTIFICATIVA S4 (days 22 to end of month) ---
        # Use "<"&DATE(year, month+1, 1) for end of month (works correctly for Dec too)
        s4_formula = bounded_ref(
            f"=COUNTIFS('DRAFT 2'!$D:$D,$B{r},'DRAFT 2'!$A:$A,\">=\"&DATE({year},{month_num},22),'DRAFT 2'!$A:$A,\"<\"&DATE({year},{month_num}+1,1))"
        )
        ws.cell(row=r, column=start_col + 13, value=s4_formula)

        # --- 14: JUSTIFICATIVA MENSAL = SUM(S1:S4) ---
        just_mensal_formula = f"=SUM({just_s1_col}{r}:{just_s4_col}{r})"
        ws.cell(row=r, column=start_col + 14, value=just_mensal_formula)


def write_faturamento_anchors(ws):
    """Write BZ (VENDA) and CA (% ALCANCADO) and quarter anchors."""
    # CA (col 79) = % ALCANCADO = current month %MES
    # Per V12 template: =CX{row} (which is FEV %MES, col 102)
    # FEV is the current month (Feb 2026)
    current_month_mes_pct_col = get_col_letter(MONTH_START_COLS[2] + 6)  # FEV %MES

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        r = row

        # BZ (col 78) = VENDA = reference to CA
        ws.cell(row=r, column=78, value=f"={get_col_letter(79)}{r}")

        # CA (col 79) = % ALCANCADO = current month %MES
        ws.cell(row=r, column=79, value=f"={current_month_mes_pct_col}{r}")

    # Quarter anchors
    # %Q1 (CB=80): average of JAN %MES, FEV %MES, MAR %MES
    # %Q2 (DV=126): average of APR %MES, MAI %MES, JUN %MES
    # %Q3 (FP=172): average of JUL %MES, AGO %MES, SET %MES
    # %Q4 (HJ=218): average of OUT %MES, NOV %MES, DEZ %MES
    quarters_months = {
        1: [1, 2, 3],
        2: [4, 5, 6],
        3: [7, 8, 9],
        4: [10, 11, 12],
    }

    for q, months in quarters_months.items():
        qtr_col = QUARTER_ANCHOR_COLS[q]
        mes_pct_cols = [get_col_letter(MONTH_START_COLS[m] + 6) for m in months]

        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            r = row
            # AVERAGE of the 3 month %MES columns
            avg_formula = f"=AVERAGE({mes_pct_cols[0]}{r},{mes_pct_cols[1]}{r},{mes_pct_cols[2]}{r})"
            ws.cell(row=r, column=qtr_col, value=avg_formula)


def count_formulas(ws):
    """Count formulas in a worksheet."""
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                count += 1
    return count


def main():
    print("=" * 60)
    print("Phase 09 Plan 04: SAP + FATURAMENTO Block Builder")
    print("=" * 60)

    # --- TASK 1: SAP Blocks ---
    print("\n--- Task 1: Loading SAP Cadastro data ---")
    sap_data = load_sap_cadastro()
    print(f"  Loaded {len(sap_data)} unique SAP clients by CNPJ")

    print("  Opening V13...")
    wb = openpyxl.load_workbook(str(V13_PATH))
    ws = wb["CARTEIRA"]

    print("  Injecting SAP and DADOS CADASTRAIS SAP blocks...")
    stats = inject_sap_blocks(ws, sap_data)
    print(f"  SAP block populated: {stats['sap_populated']} rows")
    print(f"  SAP block empty (no match): {stats['sap_empty']} rows")
    print(f"  DADOS CADASTRAIS populated: {stats['dados_populated']} rows")

    # --- TASK 2: FATURAMENTO Mega-Block ---
    print("\n--- Task 2: Building FATURAMENTO mega-block ---")
    print(f"  Year: {FAT_YEAR}")
    print(f"  Months: 12 x 15 sub-columns = 180 formula columns")
    print(f"  Plus anchors BZ, CA, and 4 quarter anchors = 186 total columns")
    print(f"  Rows: {NUM_ROWS} (rows {DATA_START_ROW}-{DATA_END_ROW})")

    for month_num in range(1, 13):
        month_name = MONTH_NAMES[month_num - 1]
        start_col = MONTH_START_COLS[month_num]
        print(f"  Writing {month_name} (col {get_col_letter(start_col)}, month #{month_num})...")
        write_faturamento_month(ws, month_num, FAT_YEAR)

    print("  Writing anchors (BZ, CA, %Q1-%Q4)...")
    write_faturamento_anchors(ws)

    print("\n  Saving V13...")
    wb.save(str(V13_PATH))
    print("  V13 saved successfully!")

    # --- Verification ---
    print("\n--- Verification ---")
    wb2 = openpyxl.load_workbook(str(V13_PATH))
    ws2 = wb2["CARTEIRA"]

    # Count formulas in CARTEIRA
    carteira_formulas = count_formulas(ws2)
    print(f"  CARTEIRA total formulas: {carteira_formulas}")

    # Check SAP block spot checks
    sap_pop = 0
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        if ws2.cell(row=row, column=63).value is not None:
            sap_pop += 1
    print(f"  SAP CODIGO populated: {sap_pop}/{NUM_ROWS}")

    # Check FATURAMENTO span (BZ=78 to JC=263)
    last_col = ws2.cell(row=3, column=263).value
    print(f"  JC3 (col 263) header: {last_col}")

    # Check first month META MES formula
    meta_jan = ws2.cell(row=4, column=MONTH_START_COLS[1] + 7).value
    print(f"  JAN META MES (row 4): {meta_jan}")

    # Check first month JUSTIFICATIVA S1 formula
    just_s1_jan = ws2.cell(row=4, column=MONTH_START_COLS[1] + 10).value
    print(f"  JAN JUST S1 (row 4): {just_s1_jan}")

    # Check DEZ month formula (last month)
    meta_dez = ws2.cell(row=4, column=MONTH_START_COLS[12] + 7).value
    print(f"  DEZ META MES (row 4): {meta_dez}")

    just_s1_dez = ws2.cell(row=4, column=MONTH_START_COLS[12] + 10).value
    print(f"  DEZ JUST S1 (row 4): {just_s1_dez}")

    # Check quarter anchors
    q1_formula = ws2.cell(row=4, column=80).value
    print(f"  %Q1 (row 4): {q1_formula}")

    ca_formula = ws2.cell(row=4, column=79).value
    print(f"  % ALCANCADO (row 4): {ca_formula}")

    bz_formula = ws2.cell(row=4, column=78).value
    print(f"  VENDA BZ (row 4): {bz_formula}")

    # Check PROJECAO formula count
    projecao_ws = None
    for sn in wb2.sheetnames:
        if "PROJE" in sn.upper():
            projecao_ws = wb2[sn]
            break
    if projecao_ws:
        proj_formulas = count_formulas(projecao_ws)
        print(f"  PROJECAO formulas: {proj_formulas}")

    # File size
    file_size = V13_PATH.stat().st_size / (1024 * 1024)
    print(f"  File size: {file_size:.1f} MB")

    # Check for full-column refs
    full_col_refs = 0
    for row in ws2.iter_rows(min_row=DATA_START_ROW, max_row=DATA_END_ROW,
                             min_col=78, max_col=263):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if re.search(r"\$[A-Z]{1,3}:\$[A-Z]{1,3}", cell.value):
                    full_col_refs += 1
    print(f"  Full-column references in FATURAMENTO: {full_col_refs}")

    wb2.close()
    print("\n" + "=" * 60)
    print("DONE!")
    print("=" * 60)


if __name__ == "__main__":
    main()

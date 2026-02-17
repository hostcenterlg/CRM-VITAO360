#!/usr/bin/env python3
"""
Phase 09 Plan 03 Task 2: Verify MERCOS + FUNIL block formulas in CARTEIRA.

Validates:
1. MERCOS INDEX/MATCH from DRAFT 1 with bounded ranges
2. FUNIL CSE array from DRAFT 2
3. TEMPERATURA REGRAS motor lookup
4. SINALEIRO nested IF (no _xlfn.LET)
5. No full-column references
6. Formula count >= 25,000
7. PROJECAO 19,224 formulas intact
"""

import json
import re
import random
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

BASE = Path("c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360")
V13_PATH = BASE / "data" / "output" / "CRM_VITAO360_V13_PROJECAO.xlsx"
OUTPUT_PATH = BASE / "data" / "output" / "phase09" / "carteira_mercos_funil_validation.json"

print("Loading V13 workbook for validation...")
wb = openpyxl.load_workbook(V13_PATH)
ws = wb["CARTEIRA"]

results = {}
all_pass = True

# ── Check 1: MERCOS spot-check (5 rows, col P = DIAS SEM COMPRA) ─────────
random.seed(42)
test_rows = sorted(random.sample(range(4, 558), 5))
mercos_ok = True
for r in test_rows:
    val = str(ws.cell(row=r, column=16).value or "")
    if "DRAFT 1" not in val or "INDEX" not in val:
        mercos_ok = False
results["mercos_spot_check"] = {
    "result": "PASS" if mercos_ok else "FAIL",
    "rows_tested": test_rows,
    "column": "P (DIAS SEM COMPRA)",
    "pattern": "INDEX/MATCH referencing DRAFT 1"
}
if not mercos_ok: all_pass = False

# ── Check 2: FUNIL spot-check (3 rows, col AR = ESTAGIO FUNIL) ───────────
test_rows2 = sorted(random.sample(range(4, 558), 3))
funil_ok = True
for r in test_rows2:
    val = str(ws.cell(row=r, column=44).value or "")
    if "DRAFT 2" not in val:
        funil_ok = False
results["funil_spot_check"] = {
    "result": "PASS" if funil_ok else "FAIL",
    "rows_tested": test_rows2,
    "column": "AR (ESTAGIO FUNIL)",
    "pattern": "CSE array formula referencing DRAFT 2"
}
if not funil_ok: all_pass = False

# ── Check 3: TEMPERATURA (col BB = 54) ───────────────────────────────────
temp_ok = True
for r in [4, 100, 400]:
    val = str(ws.cell(row=r, column=54).value or "")
    if "REGRAS" not in val or "$G$220" not in val:
        temp_ok = False
results["temperatura_regras"] = {
    "result": "PASS" if temp_ok else "FAIL",
    "rows_tested": [4, 100, 400],
    "pattern": "REGRAS!$G$220:$G$282 motor lookup"
}
if not temp_ok: all_pass = False

# ── Check 4: SINALEIRO (col BJ = 62) ─────────────────────────────────────
sinaleiro_ok = True
for r in [4, 200, 557]:
    val = str(ws.cell(row=r, column=62).value or "")
    if "_xlfn.LET" in val:
        sinaleiro_ok = False
    if "IF" not in val:
        sinaleiro_ok = False
results["sinaleiro_no_let"] = {
    "result": "PASS" if sinaleiro_ok else "FAIL",
    "rows_tested": [4, 200, 557],
    "pattern": "Nested IF without _xlfn.LET"
}
if not sinaleiro_ok: all_pass = False

# ── Check 5: No full-column references ───────────────────────────────────
full_col_count = 0
for r in range(4, 558):
    for c in range(1, 63):
        val = ws.cell(row=r, column=c).value
        if val and isinstance(val, str) and val.startswith("="):
            if re.search(r"\$[A-Z]+:\$[A-Z]+", val):
                full_col_count += 1
results["no_full_column_refs"] = {
    "result": "PASS" if full_col_count == 0 else "FAIL",
    "full_col_refs_found": full_col_count
}
if full_col_count > 0: all_pass = False

# ── Check 6: Formula count ───────────────────────────────────────────────
formula_total = 0
for r in range(4, 558):
    for c in range(1, 63):
        val = ws.cell(row=r, column=c).value
        if val and isinstance(val, str) and val.startswith("="):
            formula_total += 1
results["formula_count_mercos_funil"] = {
    "result": "PASS" if formula_total >= 25000 else "FAIL",
    "count": formula_total,
    "threshold": 25000
}
if formula_total < 25000: all_pass = False

# ── Check 7: PROJECAO intact ─────────────────────────────────────────────
proj_ws = None
for name in wb.sheetnames:
    if "PROJE" in name.upper():
        proj_ws = wb[name]
        break
proj_formulas = 0
if proj_ws:
    for row in proj_ws.iter_rows(min_row=1, max_row=proj_ws.max_row, max_col=proj_ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                proj_formulas += 1
results["projecao_intact"] = {
    "result": "PASS" if proj_formulas == 19224 else "FAIL",
    "formula_count": proj_formulas,
    "expected": 19224
}
if proj_formulas != 19224: all_pass = False

# ── Additional checks ────────────────────────────────────────────────────

# Check column count
jc3 = ws.cell(row=3, column=263).value
jd3 = ws.cell(row=3, column=264).value
results["column_count_263"] = {
    "result": "PASS" if jc3 is not None and jd3 is None else "FAIL",
    "col_263_value": str(jc3),
    "col_264_value": str(jd3)
}

# Check CNPJ count
cnpj_count = sum(1 for r in range(4, 558) if ws.cell(row=r, column=2).value)
results["cnpj_count"] = {
    "result": "PASS" if cnpj_count == 554 else "FAIL",
    "count": cnpj_count,
    "expected": 554
}

# Check freeze panes
fp = str(ws.freeze_panes)
results["freeze_panes"] = {
    "result": "PASS" if fp == "AR6" else "FAIL",
    "value": fp,
    "expected": "AR6"
}

# Check outline levels
grouped = sum(1 for c in range(1, 264) if ws.column_dimensions[get_column_letter(c)].outline_level > 0)
results["grouped_columns"] = {
    "result": "PASS" if grouped >= 200 else "FAIL",
    "count": grouped,
    "threshold": 200
}

# ── Save results ──────────────────────────────────────────────────────────
results["overall"] = "ALL PASS" if all_pass else "SOME FAILED"
results["total_formulas_mercos_funil"] = formula_total
results["total_formulas_projecao"] = proj_formulas
results["total_formulas_v13"] = formula_total + proj_formulas

with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(results, f, indent=2, ensure_ascii=False)

print(f"\nValidation results saved to: {OUTPUT_PATH}")
print(f"\nOverall: {results['overall']}")
for key, val in results.items():
    if isinstance(val, dict) and "result" in val:
        print(f"  {key}: {val['result']}")

print(f"\nTotal formulas: {formula_total} (MERCOS/FUNIL) + {proj_formulas} (PROJECAO) = {formula_total + proj_formulas}")

if not all_pass:
    sys.exit(1)

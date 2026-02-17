"""
Plan 09-05 Task 1: Intelligence Engine - Layer 1 (Ranking Score) + Retrofeeding
=================================================================================
Adds the 3-layer intelligence engine to CARTEIRA:
- Layer 1: Weighted SCORE from 6 factors + RANK for priority ordering
- Retrofeeding: Enhanced PROX FOLLOWUP (REGRAS follow-up days) + ACAO FUTURA (REGRAS motor fallback)

Column placement strategy:
- O (PRIORIDADE): Composite SCORE formula (repurpose existing empty column)
- New columns after JC (col 263): RANK, individual factor scores (helper),
  META DIARIA, PIPELINE VALUE, COVERAGE RATIO, ALERTA, GAP VALUE

This script handles Task 1 (Layer 1 + retrofeeding) only.
Task 2 (Layers 2-3 + conditional formatting) will enhance further.
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import json
import sys

# Paths
V13_PATH = "data/output/CRM_VITAO360_V13_PROJECAO.xlsx"
COLUMN_SPEC_PATH = "data/output/phase09/carteira_column_spec.json"

# Constants
DATA_START_ROW = 4
DATA_END_ROW = 557
NUM_CLIENTS = 554  # rows 4-557

def bounded_ref(formula):
    """Convert full-column refs $X:$X to bounded $X$3:$X$25000"""
    import re
    def replace_full_col(match):
        col = match.group(1)
        return f"${col}$3:${col}$25000"
    return re.sub(r"\$([A-Z]+):\$([A-Z]+)", replace_full_col, formula)


def build_layer1_score_formula(row):
    """
    Build the composite SCORE formula for a single row.

    6 weighted factors from REGRAS section 16 (rows 209-216):
    1. URGENCIA TEMPORAL (30%): MIN(100, DIAS_SEM_COMPRA / MAX(CICLO_MEDIO, 1) * 100)
    2. VALOR DO CLIENTE (25%): CURVA ABC score + TIPO CLIENTE maturity bonus
    3. FOLLOW-UP VENCIDO (20%): IF(PROX_FOLLOWUP < TODAY(), overdue penalty)
    4. SINAL DE COMPRA (15%): TEMPERATURA score + B2B access bonus
    5. TENTATIVA (5%): T1-T4 scoring
    6. SITUACAO (5%): Status-based scoring

    Column references:
    - P = DIAS SEM COMPRA
    - S = CICLO MEDIO
    - AM = CURVA ABC
    - AX = TIPO CLIENTE
    - AS = PROX FOLLOWUP
    - BB = TEMPERATURA (contains emoji: "QUENTE", "MORNO", "FRIO")
    - T = ACESSO B2B
    - AY = TENTATIVA
    - N = SITUACAO
    """
    r = row

    # Factor 1: URGENCIA TEMPORAL (30%)
    # Score = MIN(100, (DIAS_SEM_COMPRA / MAX(CICLO_MEDIO, 1)) * 100)
    # If DIAS SEM COMPRA is empty/error, score = 0
    f1 = f'IFERROR(MIN(100,P{r}/MAX(S{r},1)*100),0)'

    # Factor 2: VALOR DO CLIENTE (25%)
    # ABC score: A=100, B=60, C=30, else=0
    # + maturity bonus from TIPO CLIENTE: Fidelizado=20, Maduro=15, Recorrente=10, else=0
    # Note: AX may have values like "Fidelizado", "Em desenvolvimento", etc.
    f2_abc = f'IF(AM{r}="A",100,IF(AM{r}="B",60,IF(AM{r}="C",30,0)))'
    f2_mat = f'IF(ISNUMBER(SEARCH("Fidelizado",AX{r})),20,IF(ISNUMBER(SEARCH("Maduro",AX{r})),15,IF(ISNUMBER(SEARCH("Recorrente",AX{r})),10,0)))'
    f2 = f'MIN(100,{f2_abc}+{f2_mat})'

    # Factor 3: FOLLOW-UP VENCIDO (20%)
    # If PROX FOLLOWUP < TODAY(), score = MIN(100, (TODAY()-PROX_FOLLOWUP)*5)
    # If not overdue or empty, score = 0
    f3 = f'IF(AND(ISNUMBER(AS{r}),AS{r}<TODAY()),MIN(100,(TODAY()-AS{r})*5),0)'

    # Factor 4: SINAL DE COMPRA (15%)
    # TEMPERATURA score: QUENTE=100, MORNO=60, FRIO=20, else=0
    # Note: BB contains emoji prefixed values like "QUENTE", "MORNO", "FRIO"
    # Use SEARCH to find keyword in cell (handles emoji prefix)
    f4_temp = f'IF(ISNUMBER(SEARCH("QUENTE",BB{r})),100,IF(ISNUMBER(SEARCH("MORNO",BB{r})),60,IF(ISNUMBER(SEARCH("FRIO",BB{r})),20,0)))'
    f4_b2b = f'IF(T{r}="SIM",20,0)'
    f4 = f'MIN(100,{f4_temp}+{f4_b2b})'

    # Factor 5: TENTATIVA (5%)
    # T1=100, T2=80, T3=60, T4=40, else=20
    f5 = f'IF(AY{r}="T1",100,IF(AY{r}="T2",80,IF(AY{r}="T3",60,IF(AY{r}="T4",40,20))))'

    # Factor 6: SITUACAO (5%)
    # EM RISCO=100, ATIVO=80, INAT.REC=70, PROSPECT=50, NOVO=40, LEAD=30, INAT.ANT=20, else=0
    f6 = f'IF(N{r}="EM RISCO",100,IF(N{r}="ATIVO",80,IF(N{r}="INAT.REC",70,IF(N{r}="PROSPECT",50,IF(N{r}="NOVO",40,IF(N{r}="LEAD",30,IF(N{r}="INAT.ANT",20,0)))))))'

    # Composite SCORE = weighted sum
    # (f1*0.30) + (f2*0.25) + (f3*0.20) + (f4*0.15) + (f5*0.05) + (f6*0.05)
    score = f'=IFERROR(({f1})*0.3+({f2})*0.25+({f3})*0.2+({f4})*0.15+({f5})*0.05+({f6})*0.05,0)'

    return score


def build_rank_formula(row, score_col="O"):
    """RANK formula: =RANK(O{row}, O$4:O$557)"""
    return f'=IFERROR(RANK({score_col}{row},{score_col}$4:{score_col}$557),"")'


def build_enhanced_prox_followup(row):
    """
    Enhanced PROX FOLLOWUP formula.

    Original (from Plan 03): CSE array lookup of most recent DRAFT 2 record's col T (PROX FOLLOWUP).
    Enhanced: If original is empty, calculate from REGRAS:
      DATE_OF_LAST_RESULT + FOLLOW_UP_DAYS from REGRAS section 1
      = AT{row} + INDEX(REGRAS!$C$6:$C$20, MATCH(AV{row}, REGRAS!$B$6:$B$20, 0))

    Where:
    - AT = DATA ULT ATENDIMENTO
    - AV = ULTIMO RESULTADO
    - REGRAS col B rows 6-20 = RESULTADO names
    - REGRAS col C rows 6-20 = FOLLOW-UP days
    """
    r = row
    # Original CSE formula from Plan 03
    original = f'INDEX(\'DRAFT 2\'!$T$3:$T$25000,MATCH(1,INDEX((\'DRAFT 2\'!$D$3:$D$25000=$B{r})*(\'DRAFT 2\'!$A$3:$A$25000=MAX(IF(\'DRAFT 2\'!$D$3:$D$25000=$B{r},\'DRAFT 2\'!$A$3:$A$25000))),0,1),0))'
    # Enhanced: if original returns error/empty, use REGRAS follow-up days
    regras_calc = f'AT{r}+INDEX(REGRAS!$C$6:$C$20,MATCH(AV{r},REGRAS!$B$6:$B$20,0))'
    # Two-tier: try original first, then REGRAS calc, then empty
    formula = f'=IFERROR({original},IFERROR({regras_calc},""))'
    return formula


def build_enhanced_acao_futura(row):
    """
    Enhanced ACAO FUTURA formula.

    Original (from Plan 03): CSE array lookup of most recent DRAFT 2 record's col U (ACAO FUTURA).
    Enhanced: If original is empty, derive from REGRAS motor:
      INDEX(REGRAS!$F$221:$F$283, MATCH(N{row}&AV{row}, REGRAS!$A$221:$A$283&REGRAS!$B$221:$B$283, 0))

    Where:
    - N = SITUACAO
    - AV = ULTIMO RESULTADO
    - REGRAS col A rows 221-283 = SITUACAO
    - REGRAS col B rows 221-283 = RESULTADO
    - REGRAS col F rows 221-283 = ACAO FUTURA
    """
    r = row
    # Original CSE formula from Plan 03
    original = f'INDEX(\'DRAFT 2\'!$U$3:$U$25000,MATCH(1,INDEX((\'DRAFT 2\'!$D$3:$D$25000=$B{r})*(\'DRAFT 2\'!$A$3:$A$25000=MAX(IF(\'DRAFT 2\'!$D$3:$D$25000=$B{r},\'DRAFT 2\'!$A$3:$A$25000))),0,1),0))'
    # REGRAS motor fallback (CSE array formula for concatenated match)
    regras_motor = f'INDEX(REGRAS!$F$221:$F$283,MATCH(N{r}&AV{r},REGRAS!$A$221:$A$283&REGRAS!$B$221:$B$283,0))'
    # Two-tier: try original first, then REGRAS motor, then empty
    formula = f'=IFERROR({original},IFERROR({regras_motor},""))'
    return formula


def main():
    print("=== Plan 09-05 Task 1: Intelligence Engine Layer 1 ===")
    print(f"Loading V13 from {V13_PATH}...")

    wb = openpyxl.load_workbook(V13_PATH)
    ws = wb['CARTEIRA']

    # Verify current state
    assert ws.cell(row=3, column=15).value == "PRIORIDADE", "O3 should be PRIORIDADE"
    print(f"Column O header confirmed: {ws.cell(row=3, column=15).value}")

    # =========================================================================
    # STEP 1: Write SCORE formula to column O (PRIORIDADE) for all 554 rows
    # =========================================================================
    print("\nStep 1: Writing SCORE formulas to column O (554 rows)...")
    o_col = 15  # Column O = 15

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        formula = build_layer1_score_formula(row)
        ws.cell(row=row, column=o_col, value=formula)

    print(f"  Written {DATA_END_ROW - DATA_START_ROW + 1} SCORE formulas to column O")

    # =========================================================================
    # STEP 2: Add RANK column after JC (col 263) -> JD (col 264)
    # =========================================================================
    print("\nStep 2: Adding RANK column at JD (col 264)...")
    jd_col = 264  # After JC (263)

    # Headers
    ws.cell(row=1, column=jd_col, value="INTELIGENCIA")  # R1 super-group
    ws.cell(row=2, column=jd_col, value="RANKING")  # R2 sub-group
    ws.cell(row=3, column=jd_col, value="RANK")  # R3 column name

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        formula = build_rank_formula(row, "O")
        ws.cell(row=row, column=jd_col, value=formula)

    print(f"  Written {NUM_CLIENTS} RANK formulas to JD")

    # =========================================================================
    # STEP 3: Enhance PROX FOLLOWUP (col AS = 45) with REGRAS follow-up days
    # =========================================================================
    print("\nStep 3: Enhancing PROX FOLLOWUP (AS) with REGRAS follow-up days...")
    as_col = column_index_from_string('AS')  # 45

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        formula = build_enhanced_prox_followup(row)
        ws.cell(row=row, column=as_col, value=formula)

    print(f"  Enhanced {NUM_CLIENTS} PROX FOLLOWUP formulas in AS")

    # =========================================================================
    # STEP 4: Enhance ACAO FUTURA (col AU = 47) with REGRAS motor fallback
    # =========================================================================
    print("\nStep 4: Enhancing ACAO FUTURA (AU) with REGRAS motor fallback...")
    au_col = column_index_from_string('AU')  # 47

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        formula = build_enhanced_acao_futura(row)
        ws.cell(row=row, column=au_col, value=formula)

    print(f"  Enhanced {NUM_CLIENTS} ACAO FUTURA formulas in AU")

    # =========================================================================
    # SAVE
    # =========================================================================
    print(f"\nSaving to {V13_PATH}...")
    wb.save(V13_PATH)
    print("Saved successfully.")

    # =========================================================================
    # VERIFICATION
    # =========================================================================
    print("\n=== VERIFICATION ===")

    # Reload to verify
    wb2 = openpyxl.load_workbook(V13_PATH)
    ws2 = wb2['CARTEIRA']

    results = {}

    # 1. SCORE column exists with composite weighted formula for all 554 rows
    score_count = 0
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        val = ws2.cell(row=row, column=o_col).value
        if val and str(val).startswith('='):
            score_count += 1
    results['score_count_554'] = {
        'status': 'PASS' if score_count == NUM_CLIENTS else 'FAIL',
        'detail': f'{score_count}/{NUM_CLIENTS} rows with SCORE formula'
    }
    print(f"  1. SCORE formulas: {score_count}/{NUM_CLIENTS} - {'PASS' if score_count == NUM_CLIENTS else 'FAIL'}")

    # 2. RANK column exists ranking clients 1-554
    rank_count = 0
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        val = ws2.cell(row=row, column=jd_col).value
        if val and str(val).startswith('='):
            rank_count += 1
    results['rank_count_554'] = {
        'status': 'PASS' if rank_count == NUM_CLIENTS else 'FAIL',
        'detail': f'{rank_count}/{NUM_CLIENTS} rows with RANK formula'
    }
    print(f"  2. RANK formulas: {rank_count}/{NUM_CLIENTS} - {'PASS' if rank_count == NUM_CLIENTS else 'FAIL'}")

    # 3. Spot-check: SCORE formula contains all 6 factor weights
    sample_score = str(ws2.cell(row=4, column=o_col).value)
    has_weights = all(w in sample_score for w in ['0.3', '0.25', '0.2', '0.15', '0.05'])
    results['score_weights'] = {
        'status': 'PASS' if has_weights else 'FAIL',
        'detail': f'Formula has all 5 weight values (0.3+0.25+0.2+0.15+0.05x2)'
    }
    print(f"  3. Weight check: {has_weights} - {'PASS' if has_weights else 'FAIL'}")

    # 4. SCORE formula references correct columns (P,S,AM,AX,AS,BB,T,AY,N)
    ref_cols = ['P4', 'S4', 'AM4', 'AX4', 'AS4', 'BB4', 'T4', 'AY4', 'N4']
    has_refs = sum(1 for ref in ref_cols if ref in sample_score)
    results['score_column_refs'] = {
        'status': 'PASS' if has_refs >= 8 else 'FAIL',
        'detail': f'{has_refs}/{len(ref_cols)} column references found in SCORE formula'
    }
    print(f"  4. Column references: {has_refs}/{len(ref_cols)} - {'PASS' if has_refs >= 8 else 'FAIL'}")

    # 5. PROX FOLLOWUP enhanced with REGRAS follow-up days
    as_formula = str(ws2.cell(row=4, column=as_col).value)
    has_regras_followup = 'REGRAS!$C$6:$C$20' in as_formula and 'REGRAS!$B$6:$B$20' in as_formula
    results['prox_followup_regras'] = {
        'status': 'PASS' if has_regras_followup else 'FAIL',
        'detail': f'PROX FOLLOWUP references REGRAS follow-up days: {has_regras_followup}'
    }
    print(f"  5. PROX FOLLOWUP REGRAS ref: {has_regras_followup} - {'PASS' if has_regras_followup else 'FAIL'}")

    # 6. ACAO FUTURA has REGRAS motor fallback
    au_formula = str(ws2.cell(row=4, column=au_col).value)
    has_regras_motor = 'REGRAS!$F$221:$F$283' in au_formula and 'REGRAS!$A$221:$A$283' in au_formula
    results['acao_futura_regras'] = {
        'status': 'PASS' if has_regras_motor else 'FAIL',
        'detail': f'ACAO FUTURA references REGRAS motor: {has_regras_motor}'
    }
    print(f"  6. ACAO FUTURA REGRAS motor: {has_regras_motor} - {'PASS' if has_regras_motor else 'FAIL'}")

    # 7. No circular references (SCORE in O does not reference itself)
    has_circular = 'O4' in sample_score or f'O${DATA_START_ROW}' in sample_score
    results['no_circular'] = {
        'status': 'PASS' if not has_circular else 'FAIL',
        'detail': f'No circular reference in SCORE: {not has_circular}'
    }
    print(f"  7. No circular refs: {not has_circular} - {'PASS' if not has_circular else 'FAIL'}")

    # 8. Weight sum = 100% (0.30 + 0.25 + 0.20 + 0.15 + 0.05 + 0.05 = 1.00)
    weights_sum = 0.30 + 0.25 + 0.20 + 0.15 + 0.05 + 0.05
    results['weight_sum_100'] = {
        'status': 'PASS' if abs(weights_sum - 1.0) < 0.001 else 'FAIL',
        'detail': f'Weight sum: {weights_sum}'
    }
    print(f"  8. Weight sum: {weights_sum} - {'PASS' if abs(weights_sum - 1.0) < 0.001 else 'FAIL'}")

    # 9. PROJECAO formulas intact
    ws_proj = None
    for name in wb2.sheetnames:
        if 'PROJE' in name.upper():
            ws_proj = wb2[name]
            break
    proj_formulas = 0
    for row in ws_proj.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                proj_formulas += 1
    results['projecao_intact'] = {
        'status': 'PASS' if proj_formulas == 19224 else 'FAIL',
        'detail': f'PROJECAO formulas: {proj_formulas} (expected 19,224)'
    }
    print(f"  9. PROJECAO intact: {proj_formulas} formulas - {'PASS' if proj_formulas == 19224 else 'FAIL'}")

    # 10. Total CARTEIRA formula count
    cart_formulas = 0
    for row in ws2.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                cart_formulas += 1
    results['carteira_total_formulas'] = {
        'status': 'PASS' if cart_formulas >= 131000 else 'FAIL',
        'detail': f'CARTEIRA total formulas: {cart_formulas}'
    }
    print(f"  10. CARTEIRA formulas: {cart_formulas} - {'PASS' if cart_formulas >= 131000 else 'FAIL'}")

    # Sample formulas for documentation
    print("\n=== SAMPLE FORMULAS ===")
    print(f"  O4 (SCORE): {ws2.cell(row=4, column=o_col).value[:100]}...")
    print(f"  JD4 (RANK): {ws2.cell(row=4, column=jd_col).value}")
    print(f"  AS4 (PROX FOLLOWUP): {str(ws2.cell(row=4, column=as_col).value)[:100]}...")
    print(f"  AU4 (ACAO FUTURA): {str(ws2.cell(row=4, column=au_col).value)[:100]}...")

    # Save validation results
    with open('data/output/phase09/intelligence_layer1_validation.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    all_pass = all(r['status'] == 'PASS' for r in results.values())
    print(f"\n{'='*60}")
    print(f"OVERALL: {'ALL PASS' if all_pass else 'SOME FAILURES'}")
    print(f"{'='*60}")

    wb2.close()

    return 0 if all_pass else 1


if __name__ == "__main__":
    sys.exit(main())

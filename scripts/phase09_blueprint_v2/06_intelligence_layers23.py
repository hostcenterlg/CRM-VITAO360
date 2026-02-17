"""
Plan 09-05 Task 2: Intelligence Engine - Layers 2-3 + Conditional Formatting
==============================================================================
Layer 2: Pipeline vs Meta (will the consultant hit their target?)
  - META DIARIA, PIPELINE VALUE, COVERAGE RATIO
Layer 3: Urgency Alerts (Plan B needed?)
  - ALERTA column, GAP VALUE
Conditional Formatting for all intelligence + key visual columns.

New columns after RANK (JD=264):
  JE (265) = META DIARIA
  JF (266) = PIPELINE VALUE
  JG (267) = COVERAGE RATIO
  JH (268) = ALERTA
  JI (269) = GAP VALUE
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import PatternFill, Font
import json
import sys

# Paths
V13_PATH = "data/output/CRM_VITAO360_V13_PROJECAO.xlsx"

# Constants
DATA_START_ROW = 4
DATA_END_ROW = 557
NUM_CLIENTS = 554

# New column assignments (after JD=264 RANK)
COL_META_DIARIA = 265      # JE
COL_PIPELINE_VALUE = 266    # JF
COL_COVERAGE_RATIO = 267    # JG
COL_ALERTA = 268            # JH
COL_GAP_VALUE = 269         # JI

# Color fills for conditional formatting
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')
BLUE_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
PURPLE_FILL = PatternFill(start_color='E4CCFF', end_color='E4CCFF', fill_type='solid')
RED_BOLD_FONT = Font(bold=True, color='9C0006')
GREEN_FONT = Font(color='006100')


def build_meta_diaria_formula(row):
    """
    META DIARIA per consultant: remaining daily meta for this client.

    Logic: (META MES - sum of consultant's REALIZADO this month) / remaining days
    But since CARTEIRA is per-client with consultant in col L, and META MES
    is in FATURAMENTO (col CY = Feb META), REALIZADO = col CZ (Feb REALIZADO):

    Simplified approach: client-level META DIARIA contribution:
    = (META MES for this client) / remaining business days in month

    META MES per client = 'PROJECAO ' col L / 12 (same as FATURAMENTO CD4 formula)
    Remaining days = MAX(DAY(EOMONTH(TODAY(),0)) - DAY(TODAY()), 1)

    Per-client META DIARIA = META MES / remaining days
    """
    r = row
    # CY = col 103 = Feb META (current month META MES)
    # Using the PROJECAO reference directly for consistency
    meta_mes = f"'PROJE\\xc7\\xc3O '!$L{r}/12"
    remaining_days = "MAX(DAY(EOMONTH(TODAY(),0))-DAY(TODAY()),1)"
    formula = f"=IFERROR({meta_mes}/{remaining_days},0)"
    return formula


def build_meta_diaria_formula_v2(row):
    """
    META DIARIA per consultant using PROJECAO reference.
    PROJECAO sheet name has cedilla: "PROJECAO " (with trailing space and accent)

    Simplified: META ANUAL (col L) / 12 / remaining_days_in_month
    """
    r = row
    # The PROJECAO sheet is named "PROJEÇÃO " (with accent + trailing space)
    # FATURAMENTO uses META MES = PROJECAO L / 12, so we reference that same pattern
    # But we need the sheet reference. Let's use the FATURAMENTO META MES column directly.
    # CY = Feb META = INDEX('PROJEÇÃO '!$L$4:$L$600, MATCH($B{r},'PROJEÇÃO '!$B$4:$B$600,0)) / 12
    # Since FATURAMENTO already calculates this at CY, just reference CY (simpler, avoids accent issues)
    remaining_days = "MAX(DAY(EOMONTH(TODAY(),0))-DAY(TODAY()),1)"
    formula = f"=IFERROR(CY{r}/{remaining_days},0)"
    return formula


def build_pipeline_value_formula(row):
    """
    PIPELINE VALUE: clients actively in the pipeline with their expected value.

    =IF(OR(AR{row}="NEGOCIACAO", AR{row}="ORCAMENTO", AR{row}="CADASTRO",
          AR{row}="EM ATENDIMENTO"), AP{row}, 0)

    Where AR = ESTAGIO FUNIL, AP = TICKET MEDIO
    Note: AR values may contain accents. Use SEARCH for partial matching.
    """
    r = row
    # Check for pipeline stages (partial match handles accent variations)
    cond1 = f'ISNUMBER(SEARCH("NEGOCIA",AR{r}))'
    cond2 = f'ISNUMBER(SEARCH("ORCAMENTO",AR{r}))'
    cond3 = f'ISNUMBER(SEARCH("CADASTRO",AR{r}))'
    cond4 = f'ISNUMBER(SEARCH("EM ATENDIMENTO",AR{r}))'
    formula = f'=IFERROR(IF(OR({cond1},{cond2},{cond3},{cond4}),AP{r},0),0)'
    return formula


def build_coverage_ratio_formula(row):
    """
    COVERAGE RATIO: consultant-level pipeline sum / consultant-level meta sum.

    =IFERROR(SUMIFS(JF$4:JF$557, L$4:L$557, L{row}) /
             SUMIFS(JE$4:JE$557, L$4:L$557, L{row}), 0)

    Where L = CONSULTOR, JF = PIPELINE VALUE, JE = META DIARIA
    """
    r = row
    jf_letter = get_column_letter(COL_PIPELINE_VALUE)  # JF
    je_letter = get_column_letter(COL_META_DIARIA)      # JE

    pipeline_sum = f'SUMIFS({jf_letter}$4:{jf_letter}$557,L$4:L$557,L{r})'
    meta_sum = f'SUMIFS({je_letter}$4:{je_letter}$557,L$4:L$557,L{r})'
    formula = f'=IFERROR({pipeline_sum}/{meta_sum},0)'
    return formula


def build_alerta_formula(row):
    """
    ALERTA: 4-tier urgency classification based on coverage ratio.

    =IF(JG{row}>=1, "OK",
      IF(JG{row}>=0.7, "ATENCAO: pipeline apertado",
        IF(JG{row}>0, "ALERTA: pipeline insuficiente",
          IF(L{row}<>"", "CRITICO: sem pipeline", ""))))
    """
    r = row
    jg_letter = get_column_letter(COL_COVERAGE_RATIO)  # JG

    formula = (
        f'=IF({jg_letter}{r}>=1,"OK",'
        f'IF({jg_letter}{r}>=0.7,"ATENCAO: pipeline apertado",'
        f'IF({jg_letter}{r}>0,"ALERTA: pipeline insuficiente",'
        f'IF(L{r}<>"","CRITICO: sem pipeline",""))))'
    )
    return formula


def build_gap_value_formula(row):
    """
    GAP VALUE: R$ gap the consultant needs to fill.

    =IF(JG{row}<1, SUMIFS(JE$4:JE$557,L$4:L$557,L{row}) - SUMIFS(JF$4:JF$557,L$4:L$557,L{row}), 0)
    """
    r = row
    jg_letter = get_column_letter(COL_COVERAGE_RATIO)
    jf_letter = get_column_letter(COL_PIPELINE_VALUE)
    je_letter = get_column_letter(COL_META_DIARIA)

    meta_sum = f'SUMIFS({je_letter}$4:{je_letter}$557,L$4:L$557,L{r})'
    pipeline_sum = f'SUMIFS({jf_letter}$4:{jf_letter}$557,L$4:L$557,L{r})'
    formula = f'=IF({jg_letter}{r}<1,{meta_sum}-{pipeline_sum},0)'
    return formula


def apply_conditional_formatting(ws):
    """
    Apply conditional formatting rules to key intelligence columns.
    Uses range-based rules (not per-row).
    """
    data_range_suffix = f"{DATA_START_ROW}:{DATA_END_ROW}"

    # 1. SCORE column O: Color scale (red=low to green=high)
    score_range = f"O{DATA_START_ROW}:O{DATA_END_ROW}"
    ws.conditional_formatting.add(
        score_range,
        ColorScaleRule(
            start_type='min', start_color='F8696B',    # Red (low score)
            mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Yellow (mid)
            end_type='max', end_color='63BE7B'          # Green (high score)
        )
    )
    print("  CF 1: SCORE (O) color scale - applied")

    # 2. SINALEIRO (BJ) -- note: BJ uses emoji sinaleiro, but also has text values
    # Check if BJ has emoji-based or text-based values
    # From the SINALEIRO formula: returns emoji circles, not text "VERDE/AMARELO/VERMELHO"
    # Skip text-based CF for emoji sinaleiro -- emoji-based CF doesn't work well in openpyxl
    # Instead apply to the intelligence columns where text values are present
    print("  CF 2: SINALEIRO (BJ) -- skipped (uses emoji indicators, not text)")

    # 3. TEMPERATURA (BB) - contains emoji+text like "QUENTE", "MORNO", "FRIO"
    bb_range = f"BB{DATA_START_ROW}:BB{DATA_END_ROW}"
    ws.conditional_formatting.add(
        bb_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("QUENTE",BB{DATA_START_ROW}))'],
            fill=PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        )
    )
    ws.conditional_formatting.add(
        bb_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("MORNO",BB{DATA_START_ROW}))'],
            fill=YELLOW_FILL
        )
    )
    ws.conditional_formatting.add(
        bb_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("FRIO",BB{DATA_START_ROW}))'],
            fill=BLUE_FILL
        )
    )
    print("  CF 3: TEMPERATURA (BB) - QUENTE/MORNO/FRIO fills applied")

    # 4. COVERAGE RATIO (JG)
    jg_letter = get_column_letter(COL_COVERAGE_RATIO)
    jg_range = f"{jg_letter}{DATA_START_ROW}:{jg_letter}{DATA_END_ROW}"
    ws.conditional_formatting.add(
        jg_range,
        CellIsRule(
            operator='greaterThanOrEqual', formula=['1'],
            fill=GREEN_FILL, font=GREEN_FONT
        )
    )
    ws.conditional_formatting.add(
        jg_range,
        FormulaRule(
            formula=[f'AND({jg_letter}{DATA_START_ROW}>=0.7,{jg_letter}{DATA_START_ROW}<1)'],
            fill=YELLOW_FILL
        )
    )
    ws.conditional_formatting.add(
        jg_range,
        CellIsRule(
            operator='lessThan', formula=['0.7'],
            fill=RED_FILL
        )
    )
    print("  CF 4: COVERAGE RATIO (JG) - green/yellow/red applied")

    # 5. ALERTA (JH)
    jh_letter = get_column_letter(COL_ALERTA)
    jh_range = f"{jh_letter}{DATA_START_ROW}:{jh_letter}{DATA_END_ROW}"
    ws.conditional_formatting.add(
        jh_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("CRITICO",{jh_letter}{DATA_START_ROW}))'],
            fill=RED_FILL, font=RED_BOLD_FONT
        )
    )
    ws.conditional_formatting.add(
        jh_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("ALERTA",{jh_letter}{DATA_START_ROW}))'],
            fill=ORANGE_FILL
        )
    )
    ws.conditional_formatting.add(
        jh_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("ATENCAO",{jh_letter}{DATA_START_ROW}))'],
            fill=YELLOW_FILL
        )
    )
    ws.conditional_formatting.add(
        jh_range,
        CellIsRule(
            operator='equal', formula=['"OK"'],
            fill=GREEN_FILL, font=GREEN_FONT
        )
    )
    print("  CF 5: ALERTA (JH) - CRITICO/ALERTA/ATENCAO/OK applied")

    # 6. FATURAMENTO % columns (%MES, %TRI, %YTD)
    # From column_spec: %MES is every 15th col starting from CI (month 1)
    # Month structure: %YTD, META, REAL, %TRI, META, REAL, %MES, META, REAL, DATA, JS1, JS2, JS3, JS4, JMEN
    # %MES cols: CI(87), CX(102), DM(117), EB(132), EQ(147), FF(162), FU(177), GJ(192), GY(207), HN(222), IC(237), IR(252)
    # %TRI cols: CF(84), CU(99), DJ(114), DY(129), EN(144), FC(159), FR(174), GG(189), GV(204), HK(219), HZ(234), IO(249)
    # %YTD cols: CC(81), CR(96), DG(111), DV(126), EK(141), EZ(156), FO(171), GD(186), GS(201), HH(216), HW(231), IL(246)
    pct_cols = []
    # %MES columns (every 15 starting from CI=87)
    for i in range(12):
        pct_cols.append(87 + i * 15)   # %MES
        pct_cols.append(84 + i * 15)   # %TRI
        pct_cols.append(81 + i * 15)   # %YTD

    # Also add %Q1-%Q4 and %ALCANCADO
    # CB(80)=%Q1, DV(126)=%Q2, ... Actually need to find these.
    # CA(79)=%ALCANCADO, CB(80)=%Q1
    pct_cols.extend([79, 80])  # %ALCANCADO, %Q1

    for col_num in pct_cols:
        col_letter = get_column_letter(col_num)
        pct_range = f"{col_letter}{DATA_START_ROW}:{col_letter}{DATA_END_ROW}"
        ws.conditional_formatting.add(
            pct_range,
            CellIsRule(
                operator='greaterThanOrEqual', formula=['1'],
                fill=GREEN_FILL
            )
        )
        ws.conditional_formatting.add(
            pct_range,
            FormulaRule(
                formula=[f'AND({col_letter}{DATA_START_ROW}>=0.7,{col_letter}{DATA_START_ROW}<1)'],
                fill=YELLOW_FILL
            )
        )
        ws.conditional_formatting.add(
            pct_range,
            FormulaRule(
                formula=[f'AND({col_letter}{DATA_START_ROW}<0.7,{col_letter}{DATA_START_ROW}<>"")'],
                fill=RED_FILL
            )
        )

    print(f"  CF 6: FATURAMENTO % columns ({len(pct_cols)} columns) - green/yellow/red applied")

    return len(pct_cols) * 3 + 4 + 3 + 3 + 1  # Total CF rules count estimate


def main():
    print("=== Plan 09-05 Task 2: Intelligence Engine Layers 2-3 + CF ===")
    print(f"Loading V13 from {V13_PATH}...")

    wb = openpyxl.load_workbook(V13_PATH)
    ws = wb['CARTEIRA']

    # Verify Layer 1 exists from Task 1
    assert str(ws.cell(row=4, column=15).value).startswith('='), "SCORE formula should exist in O4"
    assert str(ws.cell(row=4, column=264).value).startswith('='), "RANK formula should exist in JD4"
    print("Layer 1 (SCORE + RANK) confirmed from Task 1")

    # =========================================================================
    # STEP 1: Layer 2 - META DIARIA (col JE = 265)
    # =========================================================================
    print("\nStep 1: Writing META DIARIA formulas to JE (col 265)...")
    je_letter = get_column_letter(COL_META_DIARIA)

    # Headers
    ws.cell(row=1, column=COL_META_DIARIA, value="INTELIGENCIA")
    ws.cell(row=2, column=COL_META_DIARIA, value="PIPELINE")
    ws.cell(row=3, column=COL_META_DIARIA, value="META DIARIA")

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        formula = build_meta_diaria_formula_v2(row)
        ws.cell(row=row, column=COL_META_DIARIA, value=formula)
    print(f"  Written {NUM_CLIENTS} META DIARIA formulas to {je_letter}")

    # =========================================================================
    # STEP 2: Layer 2 - PIPELINE VALUE (col JF = 266)
    # =========================================================================
    print("\nStep 2: Writing PIPELINE VALUE formulas to JF (col 266)...")
    jf_letter = get_column_letter(COL_PIPELINE_VALUE)

    ws.cell(row=1, column=COL_PIPELINE_VALUE, value="INTELIGENCIA")
    ws.cell(row=2, column=COL_PIPELINE_VALUE, value="PIPELINE")
    ws.cell(row=3, column=COL_PIPELINE_VALUE, value="PIPELINE VALUE")

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        formula = build_pipeline_value_formula(row)
        ws.cell(row=row, column=COL_PIPELINE_VALUE, value=formula)
    print(f"  Written {NUM_CLIENTS} PIPELINE VALUE formulas to {jf_letter}")

    # =========================================================================
    # STEP 3: Layer 2 - COVERAGE RATIO (col JG = 267)
    # =========================================================================
    print("\nStep 3: Writing COVERAGE RATIO formulas to JG (col 267)...")
    jg_letter = get_column_letter(COL_COVERAGE_RATIO)

    ws.cell(row=1, column=COL_COVERAGE_RATIO, value="INTELIGENCIA")
    ws.cell(row=2, column=COL_COVERAGE_RATIO, value="PIPELINE")
    ws.cell(row=3, column=COL_COVERAGE_RATIO, value="COVERAGE RATIO")

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        formula = build_coverage_ratio_formula(row)
        ws.cell(row=row, column=COL_COVERAGE_RATIO, value=formula)
    print(f"  Written {NUM_CLIENTS} COVERAGE RATIO formulas to {jg_letter}")

    # =========================================================================
    # STEP 4: Layer 3 - ALERTA (col JH = 268)
    # =========================================================================
    print("\nStep 4: Writing ALERTA formulas to JH (col 268)...")
    jh_letter = get_column_letter(COL_ALERTA)

    ws.cell(row=1, column=COL_ALERTA, value="INTELIGENCIA")
    ws.cell(row=2, column=COL_ALERTA, value="URGENCIA")
    ws.cell(row=3, column=COL_ALERTA, value="ALERTA")

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        formula = build_alerta_formula(row)
        ws.cell(row=row, column=COL_ALERTA, value=formula)
    print(f"  Written {NUM_CLIENTS} ALERTA formulas to {jh_letter}")

    # =========================================================================
    # STEP 5: Layer 3 - GAP VALUE (col JI = 269)
    # =========================================================================
    print("\nStep 5: Writing GAP VALUE formulas to JI (col 269)...")
    ji_letter = get_column_letter(COL_GAP_VALUE)

    ws.cell(row=1, column=COL_GAP_VALUE, value="INTELIGENCIA")
    ws.cell(row=2, column=COL_GAP_VALUE, value="URGENCIA")
    ws.cell(row=3, column=COL_GAP_VALUE, value="GAP VALUE")

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        formula = build_gap_value_formula(row)
        ws.cell(row=row, column=COL_GAP_VALUE, value=formula)
    print(f"  Written {NUM_CLIENTS} GAP VALUE formulas to {ji_letter}")

    # =========================================================================
    # STEP 6: Conditional Formatting
    # =========================================================================
    print("\nStep 6: Applying conditional formatting...")
    cf_count = apply_conditional_formatting(ws)
    print(f"  Total CF rules applied: ~{cf_count}")

    # =========================================================================
    # SAVE
    # =========================================================================
    print(f"\nSaving to {V13_PATH}...")
    wb.save(V13_PATH)
    print("Saved successfully.")

    # =========================================================================
    # VERIFICATION
    # =========================================================================
    print("\n=== VERIFICATION ===")

    wb2 = openpyxl.load_workbook(V13_PATH)
    ws2 = wb2['CARTEIRA']

    results = {}

    # 1. Layer 2: META DIARIA has formulas for 554 rows
    meta_count = sum(1 for r in range(DATA_START_ROW, DATA_END_ROW + 1)
                     if str(ws2.cell(row=r, column=COL_META_DIARIA).value or '').startswith('='))
    results['meta_diaria_count'] = {
        'status': 'PASS' if meta_count == NUM_CLIENTS else 'FAIL',
        'detail': f'{meta_count}/{NUM_CLIENTS} META DIARIA formulas'
    }
    print(f"  1. META DIARIA: {meta_count}/{NUM_CLIENTS} - {'PASS' if meta_count == NUM_CLIENTS else 'FAIL'}")

    # 2. PIPELINE VALUE has formulas for 554 rows
    pipeline_count = sum(1 for r in range(DATA_START_ROW, DATA_END_ROW + 1)
                        if str(ws2.cell(row=r, column=COL_PIPELINE_VALUE).value or '').startswith('='))
    results['pipeline_value_count'] = {
        'status': 'PASS' if pipeline_count == NUM_CLIENTS else 'FAIL',
        'detail': f'{pipeline_count}/{NUM_CLIENTS} PIPELINE VALUE formulas'
    }
    print(f"  2. PIPELINE VALUE: {pipeline_count}/{NUM_CLIENTS} - {'PASS' if pipeline_count == NUM_CLIENTS else 'FAIL'}")

    # 3. COVERAGE RATIO has formulas for 554 rows
    coverage_count = sum(1 for r in range(DATA_START_ROW, DATA_END_ROW + 1)
                        if str(ws2.cell(row=r, column=COL_COVERAGE_RATIO).value or '').startswith('='))
    results['coverage_ratio_count'] = {
        'status': 'PASS' if coverage_count == NUM_CLIENTS else 'FAIL',
        'detail': f'{coverage_count}/{NUM_CLIENTS} COVERAGE RATIO formulas'
    }
    print(f"  3. COVERAGE RATIO: {coverage_count}/{NUM_CLIENTS} - {'PASS' if coverage_count == NUM_CLIENTS else 'FAIL'}")

    # 4. ALERTA has formulas for 554 rows
    alerta_count = sum(1 for r in range(DATA_START_ROW, DATA_END_ROW + 1)
                      if str(ws2.cell(row=r, column=COL_ALERTA).value or '').startswith('='))
    results['alerta_count'] = {
        'status': 'PASS' if alerta_count == NUM_CLIENTS else 'FAIL',
        'detail': f'{alerta_count}/{NUM_CLIENTS} ALERTA formulas'
    }
    print(f"  4. ALERTA: {alerta_count}/{NUM_CLIENTS} - {'PASS' if alerta_count == NUM_CLIENTS else 'FAIL'}")

    # 5. GAP VALUE has formulas for 554 rows
    gap_count = sum(1 for r in range(DATA_START_ROW, DATA_END_ROW + 1)
                   if str(ws2.cell(row=r, column=COL_GAP_VALUE).value or '').startswith('='))
    results['gap_value_count'] = {
        'status': 'PASS' if gap_count == NUM_CLIENTS else 'FAIL',
        'detail': f'{gap_count}/{NUM_CLIENTS} GAP VALUE formulas'
    }
    print(f"  5. GAP VALUE: {gap_count}/{NUM_CLIENTS} - {'PASS' if gap_count == NUM_CLIENTS else 'FAIL'}")

    # 6. Spot-check: COVERAGE formula uses SUMIFS with consultant column L
    coverage_sample = str(ws2.cell(row=4, column=COL_COVERAGE_RATIO).value)
    has_sumifs = 'SUMIFS' in coverage_sample and 'L$4:L$557' in coverage_sample
    results['coverage_sumifs'] = {
        'status': 'PASS' if has_sumifs else 'FAIL',
        'detail': f'COVERAGE uses SUMIFS with L (CONSULTOR): {has_sumifs}'
    }
    print(f"  6. COVERAGE SUMIFS check: {has_sumifs} - {'PASS' if has_sumifs else 'FAIL'}")

    # 7. Conditional formatting rules >= 6
    cf_rules = len(ws2.conditional_formatting._cf_rules)
    results['cf_rules_count'] = {
        'status': 'PASS' if cf_rules >= 6 else 'FAIL',
        'detail': f'{cf_rules} CF rules applied (minimum 6)'
    }
    print(f"  7. CF rules: {cf_rules} - {'PASS' if cf_rules >= 6 else 'FAIL'}")

    # 8. No circular references in new columns
    meta_sample = str(ws2.cell(row=4, column=COL_META_DIARIA).value)
    je_letter = get_column_letter(COL_META_DIARIA)
    has_circular = je_letter + '4' in meta_sample
    results['no_circular'] = {
        'status': 'PASS' if not has_circular else 'FAIL',
        'detail': f'No circular refs in intelligence cols: {not has_circular}'
    }
    print(f"  8. No circular refs: {not has_circular} - {'PASS' if not has_circular else 'FAIL'}")

    # 9. PROJECAO formulas intact
    ws_proj = None
    for name in wb2.sheetnames:
        if 'PROJE' in name.upper():
            ws_proj = wb2[name]
            break
    proj_formulas = 0
    for row in ws_proj.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                proj_formulas += 1
    results['projecao_intact'] = {
        'status': 'PASS' if proj_formulas == 19224 else 'FAIL',
        'detail': f'PROJECAO formulas: {proj_formulas} (expected 19,224)'
    }
    print(f"  9. PROJECAO intact: {proj_formulas} - {'PASS' if proj_formulas == 19224 else 'FAIL'}")

    # 10. Total CARTEIRA formula count >= 105,000 (plan says >= 105,000 after all)
    cart_formulas = 0
    for row in ws2.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                cart_formulas += 1
    results['carteira_total'] = {
        'status': 'PASS' if cart_formulas >= 105000 else 'FAIL',
        'detail': f'CARTEIRA total formulas: {cart_formulas}'
    }
    print(f"  10. CARTEIRA total: {cart_formulas} - {'PASS' if cart_formulas >= 105000 else 'FAIL'}")

    # 11. Total V13 formulas
    total_v13 = 0
    for sheet_name in wb2.sheetnames:
        ws_check = wb2[sheet_name]
        for row in ws_check.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    total_v13 += 1
    results['v13_total'] = {
        'status': 'PASS' if total_v13 >= 150000 else 'FAIL',
        'detail': f'V13 total formulas: {total_v13}'
    }
    print(f"  11. V13 total: {total_v13} - {'PASS' if total_v13 >= 150000 else 'FAIL'}")

    # Sample formulas
    print("\n=== SAMPLE FORMULAS ===")
    print(f"  JE4 (META DIARIA): {ws2.cell(row=4, column=COL_META_DIARIA).value}")
    print(f"  JF4 (PIPELINE VALUE): {ws2.cell(row=4, column=COL_PIPELINE_VALUE).value}")
    print(f"  JG4 (COVERAGE): {ws2.cell(row=4, column=COL_COVERAGE_RATIO).value}")
    print(f"  JH4 (ALERTA): {ws2.cell(row=4, column=COL_ALERTA).value}")
    print(f"  JI4 (GAP VALUE): {ws2.cell(row=4, column=COL_GAP_VALUE).value}")

    # Save validation results
    with open('data/output/phase09/intelligence_layers23_validation.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    all_pass = all(r['status'] == 'PASS' for r in results.values())
    print(f"\n{'='*60}")
    print(f"OVERALL: {'ALL PASS' if all_pass else 'SOME FAILURES'}")
    print(f"{'='*60}")

    wb2.close()

    return 0 if all_pass else 1


if __name__ == "__main__":
    sys.exit(main())

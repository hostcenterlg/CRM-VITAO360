#!/usr/bin/env python3
"""
Phase 09 Plan 03: Build CARTEIRA tab with 263-column skeleton + MERCOS/FUNIL formulas.

Task 1: Create CARTEIRA tab skeleton with 263 columns, 3-level grouping, 3 header rows,
         freeze panes, auto filter, and 554 client CNPJs.
Task 2: Inject MERCOS block (cols A-AQ, 43 cols) and FUNIL block (cols AR-BJ, 19 cols)
         formulas for all 554 data rows (rows 4-557).

All formula references use bounded ranges ($B$3:$B$25000) per Pitfall #4.
SINALEIRO rewritten without _xlfn.LET per Pitfall #3.
"""

import json
import sys
from pathlib import Path
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.properties import Outline
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE = Path("c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360")
V13_PATH = BASE / "data" / "output" / "CRM_VITAO360_V13_PROJECAO.xlsx"
SPEC_PATH = BASE / "data" / "output" / "phase09" / "carteira_column_spec.json"
DRAFT1_MAP_PATH = BASE / "data" / "output" / "phase09" / "draft1_column_map.json"

# ── Constants ─────────────────────────────────────────────────────────────────
DATA_START_ROW = 4
DATA_END_ROW = 557  # 554 clients (rows 4-557)
TOTAL_COLS = 263
MAX_COL_LETTER = "JC"
BOUNDED_END = 25000  # Use $X$3:$X$25000 instead of $X:$X

# ── Super-group colors (V12 pattern) ─────────────────────────────────────────
SG_COLORS = {
    "MERCOS":              "C6EFCE",  # green tones
    "FUNIL":               "B4C6E7",  # blue tones
    "SAP":                 "F4B084",  # orange
    "STATUS SAP":          "FFD966",  # yellow-orange
    "DADOS CADASTRAIS SAP":"E2EFDA",  # light green
    "FATURAMENTO":         "D9D9D9",  # gray
}

# ── Load column spec ──────────────────────────────────────────────────────────
print("[1/8] Loading column specification...")
with open(SPEC_PATH, "r", encoding="utf-8") as f:
    spec = json.load(f)

columns = spec["columns"]
assert len(columns) == TOTAL_COLS, f"Expected {TOTAL_COLS} columns, got {len(columns)}"

# ── Load DRAFT 1 column map ──────────────────────────────────────────────────
with open(DRAFT1_MAP_PATH, "r", encoding="utf-8") as f:
    d1_map = json.load(f)

# ── Open V13 workbook ────────────────────────────────────────────────────────
print("[2/8] Opening V13 workbook...")
wb = openpyxl.load_workbook(V13_PATH)
existing_sheets = wb.sheetnames
print(f"  Existing tabs: {existing_sheets}")

# Remove CARTEIRA if exists (re-run safety)
if "CARTEIRA" in wb.sheetnames:
    print("  Removing existing CARTEIRA tab (re-run)...")
    del wb["CARTEIRA"]

# ── Create CARTEIRA tab ──────────────────────────────────────────────────────
print("[3/8] Creating CARTEIRA tab skeleton...")
ws = wb.create_sheet("CARTEIRA")

# Outline: buttons LEFT of group (summaryRight=False) for anchor visibility
ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)

# ── Write headers + set column properties ─────────────────────────────────────
print("[4/8] Writing 3 header rows and column properties for 263 columns...")

# Style definitions
header_font = Font(name="Arial", size=10, bold=True)
data_font = Font(name="Arial", size=9)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Track super-group anchor columns for R1 headers
sg_anchors = {}  # super_group -> first anchor col_num

for col_spec in columns:
    cn = col_spec["col_num"]
    letter = col_spec["col_letter"]

    # --- Column dimensions ---
    dim = ws.column_dimensions[letter]
    dim.width = col_spec["width"]

    if col_spec["outline_level"] > 0:
        dim.outline_level = col_spec["outline_level"]
        dim.hidden = col_spec["hidden"]

    # --- R1: Super-group name (only on anchor columns or first col of group) ---
    sg = col_spec["super_group"]
    r1_val = col_spec.get("header_r1")

    if r1_val is not None:
        cell_r1 = ws.cell(row=1, column=cn, value=r1_val)
    elif col_spec["is_anchor"] and sg not in sg_anchors:
        # First anchor of this super-group -- write super-group name
        cell_r1 = ws.cell(row=1, column=cn, value=sg)
        sg_anchors[sg] = cn
    else:
        cell_r1 = ws.cell(row=1, column=cn)

    # Style R1
    cell_r1.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    if sg in SG_COLORS:
        cell_r1.fill = PatternFill(start_color=SG_COLORS[sg], end_color=SG_COLORS[sg], fill_type="solid")
        # White text on dark bg, dark text on light bg
        if sg in ("MERCOS", "FUNIL"):
            cell_r1.font = Font(name="Arial", size=10, bold=True, color="1F4E79")
        else:
            cell_r1.font = Font(name="Arial", size=10, bold=True, color="333333")

    # --- R2: Sub-group name ---
    r2_val = col_spec.get("header_r2")
    cell_r2 = ws.cell(row=2, column=cn, value=r2_val)
    cell_r2.font = header_font
    cell_r2.alignment = Alignment(horizontal="center")

    # --- R3: Column name ---
    r3_val = col_spec.get("header_r3")
    cell_r3 = ws.cell(row=3, column=cn, value=r3_val)
    cell_r3.font = header_font
    cell_r3.alignment = Alignment(horizontal="center", wrap_text=True)
    cell_r3.border = thin_border

# ── Set freeze panes ─────────────────────────────────────────────────────────
print("[5/8] Setting freeze panes at AR6...")
ws.freeze_panes = "AR6"

# ── Set auto filter ──────────────────────────────────────────────────────────
print("[5.5/8] Setting auto filter A3:JC557...")
ws.auto_filter.ref = f"A3:{MAX_COL_LETTER}{DATA_END_ROW}"

# ── Write CNPJ and NOME FANTASIA from DRAFT 1 ────────────────────────────────
print("[6/8] Writing CNPJ (col B) and NOME FANTASIA (col A) for 554 clients...")

# Read DRAFT 1 from V13 to get CNPJs and names
ws_d1 = wb["DRAFT 1"]
cnpjs_written = 0
for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
    # DRAFT 1: row 1=header row in V13 (super-groups), row 2=col names, row 3=MANUAL/AUTO
    # Data starts at row 4 in DRAFT 1 too
    d1_row = row_idx  # same row mapping

    # Col A = NOME FANTASIA (DRAFT 1 col A)
    nome = ws_d1.cell(row=d1_row, column=1).value
    ws.cell(row=row_idx, column=1, value=nome)

    # Col B = CNPJ (DRAFT 1 col B) -- preserve as text
    cnpj = ws_d1.cell(row=d1_row, column=2).value
    if cnpj is not None:
        cnpj_str = str(cnpj).strip()
        if cnpj_str:
            cell_b = ws.cell(row=row_idx, column=2, value=cnpj_str)
            cell_b.number_format = "@"  # Text format
            cnpjs_written += 1

print(f"  Written {cnpjs_written} CNPJs in column B")

# ══════════════════════════════════════════════════════════════════════════════
# TASK 2: Inject MERCOS and FUNIL block formulas
# ══════════════════════════════════════════════════════════════════════════════

print("[7/8] Injecting MERCOS + FUNIL block formulas for 554 rows...")

def bounded_ref(full_col_ref):
    """Convert full-column $X:$X to bounded $X$3:$X$25000."""
    # Pattern: $X:$X or $XX:$XX
    import re
    def replace_full_col(m):
        col = m.group(1)
        return f"${col}$3:${col}${BOUNDED_END}"
    return re.sub(r'\$([A-Z]+):\$\1', replace_full_col, full_col_ref)


def make_bounded_formula(template, row):
    """Convert V12 formula template to bounded V13 formula for given row."""
    formula = template.replace("{row}", str(row))
    formula = bounded_ref(formula)
    return formula


# ── MERCOS block formulas (cols C-AQ, skipping A=static, B=static) ───────────

formula_count = 0

for col_spec in columns:
    cn = col_spec["col_num"]
    letter = col_spec["col_letter"]
    pattern = col_spec["formula_pattern"]
    template = col_spec.get("formula_template")
    sg = col_spec["super_group"]

    # Only MERCOS and FUNIL blocks for this plan
    if sg not in ("MERCOS", "FUNIL"):
        continue

    # Skip static columns (no formula to inject)
    if pattern == "static_data":
        continue

    # Skip cols A and B (already written as static data)
    if cn <= 2:
        continue

    if template is None:
        continue

    # Special handling for SINALEIRO (BJ = col 62)
    if pattern == "let_function" and cn == 62:
        # Rewrite _xlfn.LET as nested IF per Pitfall #3
        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            # SINALEIRO uses: P=DIAS SEM COMPRA, N=SITUACAO, S=CICLO MEDIO (col 19)
            # V12 uses AK (TIPO CLIENTE) but plan says S (CICLO MEDIO)
            # V12 actual formula: _xlfn.LET(_xlpm.ciclo,AK{row},_xlpm.dias,P{row},...)
            # AK=37=TIPO CLIENTE (static text like "EM DESENVOLVIMENTO") -- doesn't make sense as ciclo
            # S=19=CICLO MEDIO (numeric) -- makes sense
            # But the V12 formula literally says AK. Checking semantics:
            # Actually the plan says to use S for CICLO MEDIO. The V12 formula uses AK for ciclo var.
            # Looking at the plan: "S" in the SINALEIRO nested IF = CICLO MEDIO.
            # Plan says: IF(S{r}=0,IF(P{r}<=50,... which is CICLO MEDIO
            # So plan is correct to use S (CICLO MEDIO), not AK (TIPO CLIENTE text)
            r = row
            sinaleiro = (
                f'=IF(P{r}="","",IF(OR(N{r}="PROSPECT",N{r}="LEAD"),"NOVO",'
                f'IF(N{r}="NOVO","EM DESENVOLVIMENTO",'
                f'IF(S{r}=0,IF(P{r}<=50,"VERDE",IF(P{r}<=90,"AMARELO","VERMELHO")),'
                f'IF(P{r}<=S{r},"VERDE",IF(P{r}<=S{r}+30,"AMARELO","VERMELHO"))))))'
            )
            ws.cell(row=row, column=cn, value=sinaleiro)
            formula_count += 1
        continue

    # Special handling for TEMPERATURA (BB = col 54) -- already uses REGRAS bounded refs
    if pattern == "index_match_regras" and cn == 54:
        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            # V12: =IFERROR(INDEX(REGRAS!$G$220:$G$282,MATCH(N{r}&AQ{r},REGRAS!$A$220:$A$282&REGRAS!$B$220:$B$282,0)),"")
            # REGRAS refs are already bounded (specific rows). Keep as-is but use AV (ULTIMO RESULTADO)
            # instead of AQ (MESES LISTA) for semantic correctness.
            # Actually the plan says: MATCH(N{row}&AV{row},...) -- SITUACAO & ULTIMO RESULTADO
            # V12 has AQ which is MESES LISTA (a number) -- this wouldn't match RESULTADO strings
            # The plan's correction to AV makes semantic sense for the REGRAS motor
            r = row
            temp_formula = (
                f'=IFERROR(INDEX(REGRAS!$G$220:$G$282,'
                f'MATCH(N{r}&AV{r},REGRAS!$A$220:$A$282&REGRAS!$B$220:$B$282,0)),"")'
            )
            ws.cell(row=row, column=cn, value=temp_formula)
            formula_count += 1
        continue

    # Standard formula injection (index_match_draft1, index_match_draft2_cse, internal_calc)
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        formula = make_bounded_formula(template, row)
        ws.cell(row=row, column=cn, value=formula)
        formula_count += 1

print(f"  Total formulas injected: {formula_count}")

# ── Save workbook ─────────────────────────────────────────────────────────────
print("[8/8] Saving V13 workbook...")
wb.save(V13_PATH)
print(f"  Saved: {V13_PATH}")
print(f"  Tabs: {wb.sheetnames}")

# ── Summary ───────────────────────────────────────────────────────────────────
print("\n=== TASK 1+2 COMPLETE ===")
print(f"CARTEIRA tab created with {TOTAL_COLS} columns")
print(f"3 header rows (R1=super-group, R2=sub-group, R3=column name)")
print(f"554 data rows (rows {DATA_START_ROW}-{DATA_END_ROW})")
print(f"{cnpjs_written} CNPJs written in column B")
print(f"freeze_panes = AR6")
print(f"auto_filter = A3:{MAX_COL_LETTER}{DATA_END_ROW}")
print(f"Total formulas: {formula_count}")
print("Done.")

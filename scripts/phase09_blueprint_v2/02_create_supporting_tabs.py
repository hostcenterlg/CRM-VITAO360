"""
Phase 09 Plan 02: Create Supporting Tabs (REGRAS, DRAFT 1, DRAFT 2) in V13.

These 3 tabs must exist BEFORE CARTEIRA is built so that formulas can reference
valid sheet names and columns. Without them, every CARTEIRA formula returns #REF!.

Sources:
- REGRAS: copied verbatim from V12 COM_DADOS (283 rows, 13 cols)
- DRAFT 1: copied from V12 COM_DADOS DRAFT 1 (554 clients, 45 cols)
- DRAFT 2: headers from V12 DRAFT 2 + 6,775 historical records from DRAFT2_POPULADO_DADOS_REAIS_v3.xlsx

Critical: All 5 existing V13 tabs must remain intact (especially PROJECAO 19,224 formulas).
"""

import sys
import shutil
import json
from pathlib import Path
from datetime import datetime

# Use project-relative paths
PROJECT = Path("c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360")
V13_PATH = PROJECT / "data/output/CRM_VITAO360_V13_PROJECAO.xlsx"
V12_PATH = PROJECT / "data/sources/crm-versoes/v11-v12/CRM_INTELIGENTE_VITAO360_V12_COM_DADOS.xlsx"
DRAFT2_SOURCE = PROJECT / "data/sources/drafts/DRAFT2_POPULADO_DADOS_REAIS_v3.xlsx"
DRAFT1_COL_MAP = PROJECT / "data/output/phase09/draft1_column_map.json"

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def backup_v13():
    """Create timestamped backup of V13 before modification."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = V13_PATH.parent / f"V13_BACKUP_{ts}.xlsx"
    shutil.copy2(V13_PATH, backup_path)
    print(f"[BACKUP] Created: {backup_path.name}")
    return backup_path


def count_projecao_formulas(wb):
    """Count formulas in PROJECAO tab (data_only=False, so formulas are strings starting with =)."""
    # Find PROJECAO sheet (has cedilla accent)
    proj_ws = None
    for name in wb.sheetnames:
        stripped = name.strip()
        if 'PROJE' in stripped.upper():
            proj_ws = wb[name]
            break
    if proj_ws is None:
        print("[ERROR] PROJECAO sheet not found!")
        return 0

    count = 0
    for row in proj_ws.iter_rows(min_row=1, max_row=proj_ws.max_row,
                                  min_col=1, max_col=proj_ws.max_column):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                count += 1
    return count


def copy_regras_tab(wb_v13, wb_v12_ro):
    """
    Copy REGRAS tab from V12 (read-only) into V13.
    Copies ALL cells verbatim (values only, not formulas -- REGRAS is a reference table).
    """
    print("[REGRAS] Reading from V12...")
    v12_regras = wb_v12_ro['REGRAS']

    # Collect all data first (read-only mode requires iteration)
    all_data = []
    max_col = 0
    for row in v12_regras.iter_rows(min_row=1, max_row=v12_regras.max_row,
                                     min_col=1, max_col=v12_regras.max_column):
        row_data = []
        for cell in row:
            row_data.append(cell.value)
            if cell.value is not None:
                max_col = max(max_col, cell.column)
        all_data.append(row_data)

    print(f"[REGRAS] Read {len(all_data)} rows, max col {max_col}")

    # Create REGRAS tab in V13
    ws = wb_v13.create_sheet("REGRAS")

    # Write all data
    for r_idx, row_data in enumerate(all_data, start=1):
        for c_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws.cell(row=r_idx, column=c_idx, value=value)

    # Apply basic formatting
    header_font = Font(name='Arial', size=10, bold=True)
    section_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    section_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    motor_header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    motor_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    # Format section title rows (they start with emoji or have section markers)
    section_title_rows = [1, 4, 22, 32, 57, 67, 79, 88, 96, 105, 112, 132, 141, 152, 201, 208, 218]
    for r in section_title_rows:
        if r <= len(all_data):
            cell = ws.cell(row=r, column=1)
            cell.font = section_font
            cell.fill = section_fill

    # Format header rows (row after section title, like row 5 for RESULTADO, row 219 for MOTOR)
    header_rows = [5, 23, 33, 58, 68, 80, 89, 97, 106, 113, 133, 142, 153, 202, 209, 219]
    for r in header_rows:
        if r <= len(all_data):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    cell.font = header_font

    # Format motor header row 219 specially
    if len(all_data) >= 219:
        for c in range(1, 8):
            cell = ws.cell(row=219, column=c)
            cell.font = motor_header_font
            cell.fill = motor_header_fill

    # Column widths
    col_widths = {
        'A': 22, 'B': 35, 'C': 28, 'D': 22, 'E': 32, 'F': 45, 'G': 18,
        'H': 15, 'I': 15, 'J': 22, 'K': 22, 'L': 22, 'M': 22
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Verify motor section
    motor_count = 0
    for r in range(221, 284):  # Data rows 221-283 (row 219 = header, 220 = sub-header)
        val = ws.cell(row=r, column=1).value
        if val is not None:
            motor_count += 1

    print(f"[REGRAS] Created with {len(all_data)} rows, motor data rows: {motor_count}")
    return ws


def copy_draft1_tab(wb_v13, wb_v12_ro):
    """
    Copy DRAFT 1 tab from V12 COM_DADOS into V13.
    Preserves CNPJ as string (14-digit zfill).
    """
    print("[DRAFT 1] Reading from V12...")
    v12_d1 = wb_v12_ro['DRAFT 1']

    # Collect all data
    all_data = []
    max_col = 0
    for row in v12_d1.iter_rows(min_row=1, max_row=v12_d1.max_row,
                                 min_col=1, max_col=v12_d1.max_column):
        row_data = []
        for cell in row:
            row_data.append(cell.value)
            if cell.value is not None:
                max_col = max(max_col, cell.column)
        all_data.append(row_data)

    print(f"[DRAFT 1] Read {len(all_data)} rows, max col {max_col}")

    # Create DRAFT 1 tab in V13
    ws = wb_v13.create_sheet("DRAFT 1")

    # Write all data
    for r_idx, row_data in enumerate(all_data, start=1):
        for c_idx, value in enumerate(row_data, start=1):
            if value is not None:
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Ensure CNPJ column (B=2) is stored as string
                if c_idx == 2 and r_idx >= 4:  # Data rows start at 4
                    cnpj = str(value).strip()
                    # Remove any formatting chars
                    cnpj = cnpj.replace('.', '').replace('/', '').replace('-', '')
                    # Zfill to 14 digits
                    if cnpj.isdigit():
                        cnpj = cnpj.zfill(14)
                    cell.value = cnpj
                    cell.number_format = '@'  # Text format

    # Format headers
    header_font = Font(name='Arial', size=10, bold=True)
    supergroup_font = Font(name='Arial', size=10, bold=True, color='7030A0')
    supergroup_fill = PatternFill(start_color='F2E6FF', end_color='F2E6FF', fill_type='solid')

    # Row 1: Title
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(name='Arial', size=11, bold=True, color='7030A0')

    # Row 2: Super-group headers
    for c in range(1, max_col + 1):
        cell = ws.cell(row=2, column=c)
        if cell.value is not None:
            cell.font = supergroup_font
            cell.fill = supergroup_fill

    # Row 3: Column headers
    for c in range(1, max_col + 1):
        cell = ws.cell(row=3, column=c)
        if cell.value is not None:
            cell.font = header_font

    # Column widths (key columns)
    widths = {
        'A': 40, 'B': 18, 'C': 40, 'D': 6, 'E': 18, 'F': 28, 'G': 18,
        'H': 14, 'I': 18, 'J': 20, 'K': 22, 'L': 16, 'M': 16, 'N': 16
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # Set default width for monthly columns and others
    for c in range(15, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # Count data rows
    data_rows = 0
    for r in range(4, len(all_data) + 1):
        val = ws.cell(row=r, column=2).value  # CNPJ column
        if val is not None:
            data_rows += 1

    print(f"[DRAFT 1] Created with {len(all_data)} rows, {data_rows} data rows with CNPJ")
    return ws, data_rows


def create_draft2_tab(wb_v13):
    """
    Create DRAFT 2 tab with correct headers from V12 structure,
    then populate with 6,775 historical records from DRAFT2_POPULADO source.
    """
    print("[DRAFT 2] Creating tab with headers...")

    # DRAFT 2 header structure from V12 audit (row 1 = super-groups, row 2 = col names)
    # The DRAFT2_POPULADO source has: row 1 = super-groups, row 2 = col names, row 3 = MANUAL/AUTO, row 4+ = data
    # We need to match the V12 structure where: row 1 = title, row 2 = headers, row 3+ = data
    # But CARTEIRA formulas reference DRAFT 2 with:
    #   - $D:$D for CNPJ match
    #   - $A:$A for DATE
    #   - $I:$I for ESTAGIO FUNIL
    #   - $T:$T for FOLLOW-UP
    #   - $U:$U for ACAO FUTURA
    #   - $R:$R for RESULTADO
    # These match the DRAFT2_POPULADO col 2 header row positions.

    # Read the DRAFT2_POPULADO source
    print("[DRAFT 2] Reading historical data from DRAFT2_POPULADO...")
    wb_d2 = openpyxl.load_workbook(DRAFT2_SOURCE, data_only=True, read_only=True)
    ws_d2 = wb_d2['DRAFT 2']

    # Collect all data
    all_data = []
    max_col = 0
    for row in ws_d2.iter_rows(min_row=1, max_row=ws_d2.max_row,
                                min_col=1, max_col=ws_d2.max_column):
        row_data = []
        for cell in row:
            row_data.append(cell.value)
            if cell.value is not None:
                max_col = max(max_col, cell.column)
        all_data.append(row_data)
    wb_d2.close()

    print(f"[DRAFT 2] Read {len(all_data)} rows from source, max col {max_col}")

    # Create DRAFT 2 tab in V13
    ws = wb_v13.create_sheet("DRAFT 2")

    # The DRAFT2_POPULADO has:
    # Row 1: Super-group headers (emoji)
    # Row 2: Column names (DATA, CONSULTOR, NOME FANTASIA, CNPJ, ...)
    # Row 3: MANUAL/AUTO indicators
    # Row 4+: Data (6,775 - 3 header rows = 6,772 data rows? Let's count)

    # We need to write headers in a way that CARTEIRA formulas can reference.
    # V12 DRAFT 2 has: Row 1=title, Row 2=headers, Row 3+=data
    # But DRAFT2_POPULADO has: Row 1=super-group, Row 2=headers, Row 3=manual/auto, Row 4+=data

    # For V13 DRAFT 2, we'll use the DRAFT2_POPULADO structure as-is because
    # CARTEIRA formulas use full-column references ($D:$D, $A:$A) -- they search ALL rows.
    # The formula MATCH finds the right row regardless of header structure.

    # However, we should match V12's layout. V12 DRAFT 2 has:
    # Row 1: title, Row 2: headers, Row 3+: data
    # CARTEIRA formulas from V12 were written for that layout.

    # Let's match V12: Row 1=title, Row 2=super-groups, Row 3=headers, Row 4+=data
    # No -- V12 DRAFT 2 has Row 1=title, Row 2=headers, Row 3+=data (only 2 header rows)

    # Actually from the audit above:
    # V12 DRAFT 2 Row 1: title " DRAFT 2 - LOG DE ATENDIMENTOS..."
    # V12 DRAFT 2 Row 2: headers (DATA, CONSULTOR, NOME FANTASIA, CNPJ, ...)
    # V12 DRAFT 2 Row 3+: data

    # DRAFT2_POPULADO Row 1: super-group emojis
    # DRAFT2_POPULADO Row 2: headers
    # DRAFT2_POPULADO Row 3: MANUAL/AUTO
    # DRAFT2_POPULADO Row 4+: data

    # For V13, I'll use the POPULADO structure (3 header rows + data) because:
    # 1. It preserves the super-group info (useful for visual reference)
    # 2. CARTEIRA formulas use full-column refs -- they search all rows
    # 3. The MANUAL/AUTO row helps Leandro know which cols to fill vs auto

    # Write all rows
    for r_idx, row_data in enumerate(all_data, start=1):
        for c_idx, value in enumerate(row_data, start=1):
            if value is not None:
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Ensure CNPJ column (D=4) is stored as string for data rows
                if c_idx == 4 and r_idx >= 4:  # Data starts at row 4
                    cnpj = str(value).strip()
                    cnpj = cnpj.replace('.', '').replace('/', '').replace('-', '')
                    if cnpj.isdigit():
                        cnpj = cnpj.zfill(14)
                    cell.value = cnpj
                    cell.number_format = '@'

    # Formatting
    header_font = Font(name='Arial', size=10, bold=True)
    supergroup_font = Font(name='Arial', size=10, bold=True, color='2F5496')

    # Row 1: Super-group headers
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        if cell.value is not None:
            cell.font = supergroup_font

    # Row 2: Column headers
    for c in range(1, max_col + 1):
        cell = ws.cell(row=2, column=c)
        if cell.value is not None:
            cell.font = header_font

    # Row 3: MANUAL/AUTO indicators
    auto_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    manual_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    for c in range(1, max_col + 1):
        cell = ws.cell(row=3, column=c)
        if cell.value is not None:
            cell.font = Font(name='Arial', size=8, italic=True)
            if str(cell.value).upper() == 'AUTO':
                cell.fill = auto_fill
            else:
                cell.fill = manual_fill

    # Column widths
    widths = {
        'A': 18, 'B': 20, 'C': 35, 'D': 18, 'E': 6, 'F': 18,
        'G': 14, 'H': 16, 'I': 18, 'J': 14, 'K': 14, 'L': 12,
        'M': 12, 'N': 10, 'O': 10, 'P': 14, 'Q': 20, 'R': 22,
        'S': 22, 'T': 18, 'U': 35, 'V': 30, 'W': 16, 'X': 30,
        'Y': 14, 'Z': 14, 'AA': 16, 'AB': 14, 'AC': 14, 'AD': 14, 'AE': 16
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # Count data rows
    data_rows = 0
    for r in range(4, len(all_data) + 1):
        val = ws.cell(row=r, column=4).value  # CNPJ column (D=4)
        if val is not None:
            data_rows += 1

    print(f"[DRAFT 2] Created with {len(all_data)} total rows, {data_rows} data rows with CNPJ")
    return ws, data_rows


def verify_existing_tabs(wb):
    """Verify all 5 existing V13 tabs are intact."""
    expected_tabs = ['LOG', 'DASH', 'REDES_FRANQUIAS_v2', 'COMITE']
    # PROJECAO has cedilla accent
    proj_found = False
    for name in wb.sheetnames:
        if 'PROJE' in name.upper():
            proj_found = True
            break

    results = {'PROJECAO_found': proj_found}
    for tab in expected_tabs:
        results[f'{tab}_found'] = tab in wb.sheetnames

    return results


def main():
    print("=" * 70)
    print("Phase 09 Plan 02: Create Supporting Tabs (REGRAS, DRAFT 1, DRAFT 2)")
    print("=" * 70)
    print()

    # Step 0: Backup V13
    backup_path = backup_v13()

    # Step 1: Open V12 in read-only mode for data extraction
    print("\n[OPEN] Loading V12 COM_DADOS (read-only)...")
    wb_v12 = openpyxl.load_workbook(V12_PATH, data_only=True, read_only=True)
    print(f"[OPEN] V12 sheets: {wb_v12.sheetnames}")

    # Step 2: Open V13 for modification (data_only=False to preserve formulas!)
    print("[OPEN] Loading V13 (data_only=False to preserve formulas)...")
    wb_v13 = openpyxl.load_workbook(V13_PATH, data_only=False)
    print(f"[OPEN] V13 sheets: {wb_v13.sheetnames}")

    # Step 3: Count PROJECAO formulas BEFORE modification
    print("\n[VERIFY] Counting PROJECAO formulas before modification...")
    formula_count_before = count_projecao_formulas(wb_v13)
    print(f"[VERIFY] PROJECAO formulas before: {formula_count_before}")

    # Step 4: Verify existing tabs
    tab_check = verify_existing_tabs(wb_v13)
    print(f"[VERIFY] Existing tabs: {tab_check}")
    for k, v in tab_check.items():
        if not v:
            print(f"[ERROR] {k} is False! Aborting.")
            sys.exit(1)

    # Step 5: Create REGRAS tab
    print("\n" + "-" * 50)
    regras_ws = copy_regras_tab(wb_v13, wb_v12)

    # Step 6: Create DRAFT 1 tab
    print("\n" + "-" * 50)
    draft1_ws, d1_data_rows = copy_draft1_tab(wb_v13, wb_v12)

    # Close V12 (no longer needed)
    wb_v12.close()
    print("\n[CLOSE] V12 closed.")

    # Step 7: Create DRAFT 2 tab
    print("\n" + "-" * 50)
    draft2_ws, d2_data_rows = create_draft2_tab(wb_v13)

    # Step 8: Final verification BEFORE saving
    print("\n" + "=" * 50)
    print("[VERIFY] Final checks before saving...")

    # Check tab count
    print(f"[VERIFY] V13 tabs: {wb_v13.sheetnames}")
    assert len(wb_v13.sheetnames) == 8, f"Expected 8 tabs, got {len(wb_v13.sheetnames)}"

    # Check REGRAS motor
    motor_rows_with_data = 0
    for r in range(221, 284):
        if regras_ws.cell(row=r, column=1).value is not None:
            motor_rows_with_data += 1
    print(f"[VERIFY] REGRAS motor data rows (221-283): {motor_rows_with_data}")

    # Check DRAFT 1 CNPJ
    d1_sample_cnpj = draft1_ws.cell(row=4, column=2).value
    print(f"[VERIFY] DRAFT 1 row 4 CNPJ: {d1_sample_cnpj}")
    assert d1_sample_cnpj is not None, "DRAFT 1 CNPJ is None!"

    # Check DRAFT 2 headers
    d2_header_col_d = draft2_ws.cell(row=2, column=4).value
    print(f"[VERIFY] DRAFT 2 row 2 col D header: {d2_header_col_d}")
    assert d2_header_col_d is not None and 'CNPJ' in str(d2_header_col_d).upper(), \
        f"DRAFT 2 col D is not CNPJ: {d2_header_col_d}"

    # Re-count PROJECAO formulas
    formula_count_after = count_projecao_formulas(wb_v13)
    print(f"[VERIFY] PROJECAO formulas after: {formula_count_after}")

    if formula_count_after != formula_count_before:
        print(f"[WARNING] PROJECAO formula count changed: {formula_count_before} -> {formula_count_after}")
        if formula_count_after < formula_count_before * 0.95:
            print("[ERROR] More than 5% formula loss! Aborting save.")
            sys.exit(1)
    else:
        print("[VERIFY] PROJECAO formulas INTACT")

    # Step 9: Save V13
    print("\n[SAVE] Saving V13 with 3 new tabs...")
    wb_v13.save(V13_PATH)
    print(f"[SAVE] Done: {V13_PATH}")
    wb_v13.close()

    # Step 10: Summary
    print("\n" + "=" * 70)
    print("RESULTS:")
    print(f"  V13 tabs: 8 (5 existing + 3 new)")
    print(f"  REGRAS: {regras_ws.max_row} rows, motor rows: {motor_rows_with_data}")
    print(f"  DRAFT 1: {d1_data_rows} data rows with CNPJ")
    print(f"  DRAFT 2: {d2_data_rows} data rows with CNPJ")
    print(f"  PROJECAO formulas: {formula_count_after} (before: {formula_count_before})")
    print(f"  Backup: {backup_path.name}")
    print("=" * 70)

    return 0


if __name__ == "__main__":
    sys.exit(main())

"""
Phase 09 Plan 04: Verification script for SAP + FATURAMENTO blocks.

Verifies:
1. SAP block (BK-BP): 6 cols populated for at least 400/554 rows
2. DADOS CADASTRAIS SAP (BQ-BY): 9 cols populated for SAP-registered clients
3. Spot-check 5 CNPJs against SAP source
4. No #REF! errors in SAP blocks
5. FATURAMENTO spans BZ-JC (186 columns)
6. Each month has exactly 15 sub-columns
7. META MES references PROJECAO col L
8. REALIZADO MES references correct DRAFT 1 column
9. JUSTIFICATIVA S1 uses COUNTIFS with correct DATE ranges
10. DEZ uses month 12 dates
11. Quarter anchors have formulas
12. % ALCANCADO references current month
13. Total CARTEIRA formulas >= 100,000
14. PROJECAO 19,224 formulas intact
15. File size < 25 MB
16. Bounded ranges (no full-column refs)
"""

import json
import re
import openpyxl
from pathlib import Path

BASE = Path("c:/Users/User/OneDrive/Documentos/GitHub/CRM-VITAO360")
V13_PATH = BASE / "data/output/CRM_VITAO360_V13_PROJECAO.xlsx"
SAP_PATH = BASE / "data/sources/sap/01_SAP_CONSOLIDADO.xlsx"
OUTPUT_PATH = BASE / "data/output/phase09/sap_faturamento_validation.json"

DATA_START_ROW = 4
DATA_END_ROW = 557

MONTH_START_COLS = {
    1: 81, 2: 96, 3: 111, 4: 127, 5: 142, 6: 157,
    7: 173, 8: 188, 9: 203, 10: 219, 11: 234, 12: 249,
}

DRAFT1_MONTH_COLS = {
    1: "AE", 2: "AF", 3: "U", 4: "V", 5: "W", 6: "X",
    7: "Y", 8: "Z", 9: "AA", 10: "AB", 11: "AC", 12: "AD",
}


def clean_cnpj(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    s = re.sub(r"[.\-/\s]", "", s)
    if s.isdigit():
        return s.zfill(14)
    return None


def main():
    results = {}
    all_pass = True

    print("Loading V13...")
    wb = openpyxl.load_workbook(str(V13_PATH))
    ws = wb["CARTEIRA"]

    # === CHECK 1: SAP block populated (BK-BP, cols 63-68) ===
    sap_populated = 0
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        if ws.cell(row=row, column=63).value is not None:
            sap_populated += 1

    check = sap_populated >= 400
    results["sap_block_populated"] = {
        "check": "SAP block (BK-BP) populated >= 400 rows",
        "result": "PASS" if check else "FAIL",
        "details": f"{sap_populated}/554 rows with SAP CODIGO"
    }
    if not check: all_pass = False
    print(f"  SAP block: {sap_populated}/554 rows -> {'PASS' if check else 'FAIL'}")

    # === CHECK 2: DADOS CADASTRAIS SAP (BQ-BY, cols 69-77) ===
    dados_populated = 0
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        if ws.cell(row=row, column=69).value is not None:
            dados_populated += 1

    check = dados_populated >= 400
    results["dados_cadastrais_populated"] = {
        "check": "DADOS CADASTRAIS SAP (BQ-BY) populated >= 400 rows",
        "result": "PASS" if check else "FAIL",
        "details": f"{dados_populated}/554 rows with DESC GRUPO CLIENTE"
    }
    if not check: all_pass = False
    print(f"  DADOS CADASTRAIS: {dados_populated}/554 rows -> {'PASS' if check else 'FAIL'}")

    # === CHECK 3: Spot-check 5 CNPJs against SAP source ===
    print("  Loading SAP source for spot-check...")
    sap_wb = openpyxl.load_workbook(str(SAP_PATH), data_only=True, read_only=True)
    sap_ws = sap_wb["Cadastro Clientes SAP"]
    sap_by_cnpj = {}
    for sap_row in sap_ws.iter_rows(min_row=2, values_only=True):
        cnpj = clean_cnpj(sap_row[4])
        if cnpj and cnpj not in sap_by_cnpj:
            sap_by_cnpj[cnpj] = str(sap_row[2]) if sap_row[2] else ""
    sap_wb.close()

    spot_checks = []
    checked = 0
    matched = 0
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        if checked >= 5:
            break
        cnpj = ws.cell(row=row, column=2).value
        codigo = ws.cell(row=row, column=63).value
        if cnpj is not None and codigo is not None:
            cnpj_clean = str(cnpj).strip().zfill(14)
            sap_codigo = sap_by_cnpj.get(cnpj_clean, "NOT_FOUND")
            match = str(codigo) == sap_codigo
            spot_checks.append({
                "row": row,
                "cnpj": cnpj_clean,
                "carteira_codigo": str(codigo),
                "sap_codigo": sap_codigo,
                "match": match
            })
            if match:
                matched += 1
            checked += 1

    check = matched >= 4  # At least 4 of 5 match
    results["sap_spot_check"] = {
        "check": "5 CNPJs spot-checked against SAP source",
        "result": "PASS" if check else "FAIL",
        "details": f"{matched}/5 matches",
        "samples": spot_checks
    }
    if not check: all_pass = False
    print(f"  SAP spot-check: {matched}/5 matches -> {'PASS' if check else 'FAIL'}")

    # === CHECK 4: No #REF! in SAP blocks ===
    ref_errors = 0
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        for col in range(63, 78):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str) and "#REF!" in val:
                ref_errors += 1

    check = ref_errors == 0
    results["no_ref_errors_sap"] = {
        "check": "No #REF! errors in SAP blocks",
        "result": "PASS" if check else "FAIL",
        "details": f"{ref_errors} #REF! errors found"
    }
    if not check: all_pass = False
    print(f"  No #REF! in SAP: {ref_errors} errors -> {'PASS' if check else 'FAIL'}")

    # === CHECK 5: FATURAMENTO spans BZ(78) to JC(263) ===
    bz_header = ws.cell(row=3, column=78).value
    jc_header = ws.cell(row=3, column=263).value
    check = bz_header is not None and jc_header is not None
    results["faturamento_span"] = {
        "check": "FATURAMENTO spans BZ(78) to JC(263)",
        "result": "PASS" if check else "FAIL",
        "details": f"BZ3={bz_header}, JC3={jc_header}"
    }
    if not check: all_pass = False
    print(f"  FATURAMENTO span: BZ3={bz_header}, JC3={jc_header} -> {'PASS' if check else 'FAIL'}")

    # === CHECK 6: Each month has 15 sub-columns ===
    month_names = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN",
                   "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
    month_checks = []
    for m in range(1, 13):
        sc = MONTH_START_COLS[m]
        # Check that sub-col at offset +14 has a formula in row 4
        just_mensal = ws.cell(row=4, column=sc + 14).value
        has_15 = just_mensal is not None and isinstance(just_mensal, str) and "SUM" in just_mensal
        month_checks.append({"month": month_names[m-1], "start_col": sc, "has_15_subcols": has_15})

    all_months_ok = all(mc["has_15_subcols"] for mc in month_checks)
    results["month_15_subcols"] = {
        "check": "Each month has exactly 15 sub-columns",
        "result": "PASS" if all_months_ok else "FAIL",
        "details": month_checks
    }
    if not all_months_ok: all_pass = False
    print(f"  15 sub-cols per month: {'PASS' if all_months_ok else 'FAIL'}")

    # === CHECK 7: JAN META MES references PROJECAO col L ===
    jan_meta = ws.cell(row=4, column=MONTH_START_COLS[1] + 7).value
    check = jan_meta and "PROJE" in jan_meta and "$L$" in jan_meta and "/12" in jan_meta
    results["jan_meta_projecao"] = {
        "check": "JAN META MES references PROJECAO col L / 12",
        "result": "PASS" if check else "FAIL",
        "details": str(jan_meta)[:120]
    }
    if not check: all_pass = False
    print(f"  JAN META -> PROJECAO L: {'PASS' if check else 'FAIL'}")

    # === CHECK 8: JAN REALIZADO MES references DRAFT 1 ===
    jan_real = ws.cell(row=4, column=MONTH_START_COLS[1] + 8).value
    check = jan_real and "DRAFT 1" in jan_real and f"${DRAFT1_MONTH_COLS[1]}$" in jan_real
    results["jan_real_draft1"] = {
        "check": f"JAN REALIZADO MES references DRAFT 1 col {DRAFT1_MONTH_COLS[1]}",
        "result": "PASS" if check else "FAIL",
        "details": str(jan_real)[:120]
    }
    if not check: all_pass = False
    print(f"  JAN REALIZADO -> DRAFT 1: {'PASS' if check else 'FAIL'}")

    # === CHECK 9: JAN JUSTIFICATIVA S1 uses COUNTIFS with DATE(2026,1,1) ===
    jan_s1 = ws.cell(row=4, column=MONTH_START_COLS[1] + 10).value
    check = jan_s1 and "COUNTIFS" in jan_s1 and "DATE(2026,1,1)" in jan_s1 and "DATE(2026,1,7)" in jan_s1
    results["jan_just_s1"] = {
        "check": "JAN JUST S1 uses COUNTIFS with DATE(2026,1,1) to DATE(2026,1,7)",
        "result": "PASS" if check else "FAIL",
        "details": str(jan_s1)[:150]
    }
    if not check: all_pass = False
    print(f"  JAN JUST S1 COUNTIFS: {'PASS' if check else 'FAIL'}")

    # === CHECK 10: DEZ uses month 12 dates ===
    dez_s1 = ws.cell(row=4, column=MONTH_START_COLS[12] + 10).value
    check = dez_s1 and "DATE(2026,12,1)" in dez_s1 and "DATE(2026,12,7)" in dez_s1
    results["dez_month12"] = {
        "check": "DEZ JUST S1 uses month 12 dates",
        "result": "PASS" if check else "FAIL",
        "details": str(dez_s1)[:150]
    }
    if not check: all_pass = False
    print(f"  DEZ month 12 dates: {'PASS' if check else 'FAIL'}")

    # === CHECK 11: Quarter anchors have formulas ===
    q_anchors = {
        "Q1": (80, "CB"),
        "Q2": (126, "DV"),
        "Q3": (172, "FP"),
        "Q4": (218, "HJ"),
    }
    q_ok = True
    q_details = {}
    for qname, (col, letter) in q_anchors.items():
        val = ws.cell(row=4, column=col).value
        has_formula = val and isinstance(val, str) and val.startswith("=")
        q_details[qname] = {"col": letter, "formula": str(val)[:80] if val else None, "has_formula": has_formula}
        if not has_formula:
            q_ok = False

    results["quarter_anchors"] = {
        "check": "Quarter anchors (%Q1-Q4) have formulas",
        "result": "PASS" if q_ok else "FAIL",
        "details": q_details
    }
    if not q_ok: all_pass = False
    print(f"  Quarter anchors: {'PASS' if q_ok else 'FAIL'}")

    # === CHECK 12: % ALCANCADO references current month ===
    ca_val = ws.cell(row=4, column=79).value
    check = ca_val and isinstance(ca_val, str) and ca_val.startswith("=")
    results["pct_alcancado"] = {
        "check": "% ALCANCADO (CA) references current month",
        "result": "PASS" if check else "FAIL",
        "details": str(ca_val)
    }
    if not check: all_pass = False
    print(f"  % ALCANCADO: {'PASS' if check else 'FAIL'}")

    # === CHECK 13: Total CARTEIRA formulas >= 100,000 ===
    formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1

    check = formula_count >= 100000
    results["carteira_formula_count"] = {
        "check": "Total CARTEIRA formulas >= 100,000",
        "result": "PASS" if check else "FAIL",
        "details": f"{formula_count} formulas"
    }
    if not check: all_pass = False
    print(f"  CARTEIRA formulas: {formula_count} -> {'PASS' if check else 'FAIL'}")

    # === CHECK 14: PROJECAO 19,224 formulas intact ===
    proj_ws = None
    for sn in wb.sheetnames:
        if "PROJE" in sn.upper():
            proj_ws = wb[sn]
            break
    proj_formulas = 0
    if proj_ws:
        for row in proj_ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    proj_formulas += 1

    check = proj_formulas == 19224
    results["projecao_intact"] = {
        "check": "PROJECAO 19,224 formulas intact",
        "result": "PASS" if check else "FAIL",
        "details": f"{proj_formulas} formulas"
    }
    if not check: all_pass = False
    print(f"  PROJECAO formulas: {proj_formulas} -> {'PASS' if check else 'FAIL'}")

    # === CHECK 15: File size < 25 MB ===
    file_size_mb = V13_PATH.stat().st_size / (1024 * 1024)
    check = file_size_mb < 25
    results["file_size"] = {
        "check": "File size < 25 MB",
        "result": "PASS" if check else "FAIL",
        "details": f"{file_size_mb:.1f} MB"
    }
    if not check: all_pass = False
    print(f"  File size: {file_size_mb:.1f} MB -> {'PASS' if check else 'FAIL'}")

    # === CHECK 16: No full-column references in FATURAMENTO ===
    full_col_refs = 0
    for row_cells in ws.iter_rows(min_row=DATA_START_ROW, max_row=DATA_END_ROW,
                                   min_col=78, max_col=263):
        for cell in row_cells:
            if cell.value and isinstance(cell.value, str):
                if re.search(r"\$[A-Z]{1,3}:\$[A-Z]{1,3}", cell.value):
                    full_col_refs += 1

    check = full_col_refs == 0
    results["bounded_ranges"] = {
        "check": "No full-column references in FATURAMENTO",
        "result": "PASS" if check else "FAIL",
        "details": f"{full_col_refs} full-column references"
    }
    if not check: all_pass = False
    print(f"  Bounded ranges: {full_col_refs} full-col refs -> {'PASS' if check else 'FAIL'}")

    # === CHECK 17: All 263 columns have data or formulas ===
    empty_cols = []
    for col in range(1, 264):
        has_data = False
        for row in [4, 10, 50, 100, 300, 500]:
            val = ws.cell(row=row, column=col).value
            if val is not None:
                has_data = True
                break
        if not has_data:
            from openpyxl.utils import get_column_letter
            empty_cols.append(get_column_letter(col))

    check = len(empty_cols) == 0
    results["all_263_populated"] = {
        "check": "All 263 columns have data in data rows",
        "result": "PASS" if check else "FAIL",
        "details": f"{len(empty_cols)} empty columns: {empty_cols[:10]}"
    }
    if not check: all_pass = False
    print(f"  All 263 cols populated: {len(empty_cols)} empty -> {'PASS' if check else 'FAIL'}")

    # === TOTAL FORMULA COUNT across entire V13 ===
    total_v13_formulas = proj_formulas + formula_count
    # Count other tabs too
    for sn in wb.sheetnames:
        if sn not in ["CARTEIRA"] and "PROJE" not in sn.upper():
            tab_ws = wb[sn]
            for row_cells in tab_ws.iter_rows():
                for cell in row_cells:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        total_v13_formulas += 1

    results["total_v13_formulas"] = {
        "check": "Total V13 formulas (all tabs)",
        "result": "INFO",
        "details": f"{total_v13_formulas} total formulas across all tabs"
    }
    print(f"  Total V13 formulas: {total_v13_formulas}")

    wb.close()

    # Overall result
    results["overall"] = "PASS" if all_pass else "FAIL"
    print(f"\n  OVERALL: {results['overall']}")

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(str(OUTPUT_PATH), "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\n  Results saved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

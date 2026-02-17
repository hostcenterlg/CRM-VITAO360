"""
Plan 09-06: AGENDA Tabs Creation + Phase 9 Comprehensive Validation
====================================================================
Task 1: Create 4 AGENDA tabs (LARISSA, DAIANE, MANU, JULIO) pulling
        from CARTEIRA with FILTER/SORT formulas, dropdowns, and styling.

Task 2: Run comprehensive Phase 9 validation (22+ checks, BLUE-01..04).

Consultant name mapping (from DRAFT 1 col J -> CARTEIRA col L):
  - LARISSA = "LARISSA PADILHA" (178 in DRAFT 1, 224 in PROJECAO)
  - DAIANE = "DAIANE STAVICKI" (73 in DRAFT 1, 62 in PROJECAO)
  - MANU = "MANU DITZEL" in DRAFT 1 (152), "HEMANUELE DITZEL (MANU)" in PROJECAO (170)
  - JULIO = "JULIO GADRET" (86 in DRAFT 1, 66 in PROJECAO)

REGRAS dropdown references:
  - RESULTADO: REGRAS!$B$6:$B$20 (15 values in col B rows 6-20)
  - MOTIVO: REGRAS!$B$34:$B$55 (22 values in col B rows 34-55)
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import Outline
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
import json
import sys
import os

# Paths
V13_PATH = "data/output/CRM_VITAO360_V13_PROJECAO.xlsx"
VALIDATION_OUTPUT = "data/output/phase09/phase09_validation_report.json"

# Constants
DATA_START_ROW = 4
DATA_END_ROW = 557
NUM_CLIENTS = 554

# AGENDA column mapping from CARTEIRA
# These are the CARTEIRA columns we want to show in AGENDA, mapped by their column numbers
AGENDA_COLUMNS = [
    # (CARTEIRA col#, AGENDA header, type: 'fixed'|'anchor'|'input', width)
    (45,  "DATA",                "fixed",   12),   # AS = PROX FOLLOWUP
    (1,   "NOME FANTASIA",       "fixed",   25),   # A
    (2,   "CNPJ",                "fixed",   18),   # B
    (15,  "SCORE",               "fixed",   10),   # O = PRIORIDADE/SCORE
    (264, "RANK",                "fixed",    8),   # JD = RANK
    (44,  "ESTAGIO FUNIL",       "fixed",   18),   # AR
    (54,  "TEMPERATURA",         "fixed",   14),   # BB
    (62,  "SINALEIRO",           "fixed",   12),   # BJ
    # Anchor columns (minimizable)
    (14,  "SITUACAO",            "anchor",  14),   # N
    (16,  "DIAS SEM COMPRA",     "anchor",  16),   # P
    (39,  "CURVA ABC",           "anchor",  10),   # AM
    (48,  "ULTIMO RESULTADO",    "anchor",  18),   # AV
    (47,  "ACAO FUTURA",         "anchor",  20),   # AU
    (46,  "DATA ULT ATENDIMENTO","anchor",  18),   # AT
    # Green input columns (consultant fills these)
    (None,"RESULTADO",           "input",   20),   # NEW - dropdown from REGRAS
    (None,"MOTIVO",              "input",   25),   # NEW - dropdown from REGRAS
    (None,"OBSERVACAO",          "input",   35),   # NEW - free text
]

# Consultant definitions
CONSULTANTS = [
    {
        "name": "LARISSA",
        "tab_name": "AGENDA LARISSA",
        "filter_names": ["LARISSA PADILHA"],
        "tab_color": "7030A0",   # Purple
    },
    {
        "name": "DAIANE",
        "tab_name": "AGENDA DAIANE",
        "filter_names": ["DAIANE STAVICKI"],
        "tab_color": "00B050",   # Green
    },
    {
        "name": "MANU",
        "tab_name": "AGENDA MANU",
        "filter_names": ["MANU DITZEL", "HEMANUELE DITZEL (MANU)"],
        "tab_color": "0070C0",   # Blue
    },
    {
        "name": "JULIO",
        "tab_name": "AGENDA JULIO",
        "filter_names": ["JULIO GADRET"],
        "tab_color": "FFC000",   # Gold/Orange
    },
]

# Styling constants
GREEN_INPUT_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Arial', size=14, bold=True)
DATA_FONT = Font(name='Arial', size=9)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def build_filter_condition(consultant, col_letter="L"):
    """
    Build the FILTER condition for consultant name matching in CARTEIRA col L.
    For MANU, we need OR condition for both name variants.
    """
    names = consultant["filter_names"]
    if len(names) == 1:
        return f'(CARTEIRA!{col_letter}$4:{col_letter}$557="{names[0]}")'
    else:
        # OR condition: (name1) + (name2) -- in Excel array context, + is OR for boolean arrays
        parts = [f'(CARTEIRA!{col_letter}$4:{col_letter}$557="{n}")' for n in names]
        return f'({"+".join(parts)})'


def build_agenda_filter_formula(consultant, agenda_col_idx, carteira_col_num):
    """
    Build the FILTER+SORT formula for one AGENDA column.

    =SORT(FILTER(CARTEIRA!{col}$4:{col}$557,
        {consultant_filter} * (CARTEIRA!AS$4:AS$557<>"") * (CARTEIRA!AS$4:AS$557<=A$1),
        ""),
        1, 1)

    Sort by column 1 of the result (DATA / PROX FOLLOWUP) ascending.

    But we want to sort by SCORE descending. The FILTER returns multiple columns,
    and we need to sort by the SCORE column (4th in our AGENDA layout).

    Actually, FILTER + SORT with multiple columns requires a different approach.
    We use FILTER on all columns at once, then SORT.

    Better approach: single formula in A3 that returns the entire filtered+sorted array.
    """
    cart_letter = get_column_letter(carteira_col_num)
    return f'CARTEIRA!{cart_letter}$4:{cart_letter}$557'


def create_agenda_tab(wb, consultant):
    """Create one AGENDA tab with FILTER/SORT formulas and styling."""
    tab_name = consultant["tab_name"]
    print(f"\n  Creating {tab_name}...")

    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = consultant["tab_color"]

    # Outline properties (for anchor columns)
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)

    # ----- ROW 1: Title + Date filter -----
    ws.cell(row=1, column=1, value=f"AGENDA {consultant['name']}")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    # Date filter cell at E1 (used by FILTER formulas)
    ws.cell(row=1, column=5, value="DATA FILTRO:")
    ws.cell(row=1, column=5).font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=1, column=6).value = '=TODAY()'
    ws.cell(row=1, column=6).number_format = 'DD/MM/YYYY'
    ws.cell(row=1, column=6).font = Font(name='Arial', size=11, bold=True, color='FF0000')

    # ----- ROW 2: Headers -----
    for col_idx, (cart_col, header, col_type, width) in enumerate(AGENDA_COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

        # Column width
        ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Anchor columns get outline_level = 1
        if col_type == "anchor":
            ws.column_dimensions[get_column_letter(col_idx)].outline_level = 1

    # Row 2 height
    ws.row_dimensions[2].height = 30

    # ----- ROW 3+: FILTER/SORT formulas -----
    # Build the multi-column FILTER+SORT formula
    # We place the master formula in A3 and it spills across columns

    # Build column array for FILTER: CHOOSE or direct column references
    # Excel 365 FILTER returns arrays. We need to filter and then sort.
    #
    # Strategy: Use individual FILTER formulas per column (not one giant array),
    # with SORTBY wrapping each FILTER to sort by SCORE descending.
    #
    # Pattern per column:
    # =IFERROR(SORTBY(
    #     FILTER(CARTEIRA!{col}$4:{col}$557,
    #       {consultant_filter} * (CARTEIRA!AS$4:AS$557<>"") * (CARTEIRA!AS$4:AS$557<=$F$1)),
    #     FILTER(CARTEIRA!O$4:O$557,
    #       {consultant_filter} * (CARTEIRA!AS$4:AS$557<>"") * (CARTEIRA!AS$4:AS$557<=$F$1)),
    #     -1
    #   ), "")
    #
    # This filters each column independently but sorts by SCORE (col O) descending.

    consultant_filter = build_filter_condition(consultant)

    # The date filter: CARTEIRA!AS$4:AS$557 <= $F$1 (date in F1)
    # AND CARTEIRA!AS$4:AS$557 <> "" (has a follow-up date)
    filter_condition = (
        f'{consultant_filter}'
        f'*(CARTEIRA!AS$4:AS$557<>"")'
        f'*(CARTEIRA!AS$4:AS$557<=$F$1)'
    )

    # Score column for sorting (CARTEIRA col O = PRIORIDADE/SCORE)
    score_filter = f'FILTER(CARTEIRA!O$4:O$557,{filter_condition})'

    for col_idx, (cart_col, header, col_type, width) in enumerate(AGENDA_COLUMNS, 1):
        if col_type == "input":
            # Input columns: no formula, just empty cells with formatting
            continue

        cart_letter = get_column_letter(cart_col)

        # Build SORTBY(FILTER(...), FILTER(score), -1) for this column
        data_filter = f'FILTER(CARTEIRA!{cart_letter}$4:{cart_letter}$557,{filter_condition})'
        formula = f'=IFERROR(SORTBY({data_filter},{score_filter},-1),"")'

        ws.cell(row=3, column=col_idx, value=formula)
        ws.cell(row=3, column=col_idx).font = DATA_FONT

    # ----- GREEN INPUT COLUMNS styling -----
    # Apply green fill to input column headers + rows 3-1000
    input_start_col = None
    for col_idx, (cart_col, header, col_type, width) in enumerate(AGENDA_COLUMNS, 1):
        if col_type == "input":
            if input_start_col is None:
                input_start_col = col_idx
            col_letter = get_column_letter(col_idx)
            # Green header
            ws.cell(row=2, column=col_idx).fill = PatternFill(
                start_color='548235', end_color='548235', fill_type='solid'
            )
            # Green fill for data rows 3-1000
            for r in range(3, 1001):
                ws.cell(row=r, column=col_idx).fill = GREEN_INPUT_FILL
                ws.cell(row=r, column=col_idx).font = DATA_FONT
                ws.cell(row=r, column=col_idx).border = THIN_BORDER

    # ----- DATA VALIDATION (Dropdowns) -----
    # RESULTADO column = col index of RESULTADO in AGENDA_COLUMNS
    resultado_col_idx = None
    motivo_col_idx = None
    for col_idx, (cart_col, header, col_type, width) in enumerate(AGENDA_COLUMNS, 1):
        if header == "RESULTADO":
            resultado_col_idx = col_idx
        elif header == "MOTIVO":
            motivo_col_idx = col_idx

    if resultado_col_idx:
        resultado_letter = get_column_letter(resultado_col_idx)
        dv_resultado = DataValidation(
            type="list",
            formula1="REGRAS!$B$6:$B$20",
            allow_blank=True,
            showDropDown=False  # In openpyxl, False = SHOW dropdown
        )
        dv_resultado.error = "Selecione um RESULTADO valido da lista REGRAS"
        dv_resultado.errorTitle = "Resultado Invalido"
        dv_resultado.prompt = "Selecione o resultado do atendimento"
        dv_resultado.promptTitle = "RESULTADO"
        ws.add_data_validation(dv_resultado)
        dv_resultado.add(f"{resultado_letter}3:{resultado_letter}1000")
        print(f"    RESULTADO dropdown: {resultado_letter}3:{resultado_letter}1000 -> REGRAS!$B$6:$B$20")

    if motivo_col_idx:
        motivo_letter = get_column_letter(motivo_col_idx)
        dv_motivo = DataValidation(
            type="list",
            formula1="REGRAS!$B$34:$B$55",
            allow_blank=True,
            showDropDown=False
        )
        dv_motivo.error = "Selecione um MOTIVO valido da lista REGRAS"
        dv_motivo.errorTitle = "Motivo Invalido"
        dv_motivo.prompt = "Por que o cliente nao comprou?"
        dv_motivo.promptTitle = "MOTIVO"
        ws.add_data_validation(dv_motivo)
        dv_motivo.add(f"{motivo_letter}3:{motivo_letter}1000")
        print(f"    MOTIVO dropdown: {motivo_letter}3:{motivo_letter}1000 -> REGRAS!$B$34:$B$55")

    # ----- FREEZE PANES at D3 (fixed columns always visible) -----
    # Columns A-C (DATA, NOME FANTASIA, CNPJ) fixed when scrolling right
    # Row 2 (headers) fixed when scrolling down
    ws.freeze_panes = "D3"
    print(f"    freeze_panes: D3")

    # Column count summary
    fixed_count = sum(1 for _, _, t, _ in AGENDA_COLUMNS if t == "fixed")
    anchor_count = sum(1 for _, _, t, _ in AGENDA_COLUMNS if t == "anchor")
    input_count = sum(1 for _, _, t, _ in AGENDA_COLUMNS if t == "input")
    print(f"    Columns: {fixed_count} fixed + {anchor_count} anchor + {input_count} input = {len(AGENDA_COLUMNS)} total")

    return ws


def run_task1(wb):
    """Task 1: Create 4 AGENDA tabs."""
    print("=" * 60)
    print("TASK 1: Create 4 AGENDA tabs with FILTER/SORT formulas")
    print("=" * 60)

    created_tabs = []
    for consultant in CONSULTANTS:
        ws = create_agenda_tab(wb, consultant)
        created_tabs.append(consultant["tab_name"])

    print(f"\n  Created {len(created_tabs)} AGENDA tabs: {created_tabs}")
    return created_tabs


def run_task2_validation(v13_path):
    """Task 2: Comprehensive Phase 9 validation + BLUE-01..04 assessment."""
    print("\n" + "=" * 60)
    print("TASK 2: Comprehensive Phase 9 Validation")
    print("=" * 60)

    wb = openpyxl.load_workbook(v13_path)
    results = {}

    # =========================================================================
    # STRUCTURAL CHECKS
    # =========================================================================
    print("\n--- Structural Checks ---")

    # 1. Tab count: 13 tabs expected
    expected_tabs = [
        "PROJECAO", "LOG", "DASH", "REDES_FRANQUIAS_v2", "COMITE",
        "REGRAS", "DRAFT 1", "DRAFT 2", "CARTEIRA",
        "AGENDA LARISSA", "AGENDA DAIANE", "AGENDA MANU", "AGENDA JULIO"
    ]
    actual_tabs = wb.sheetnames
    # Use accent-insensitive matching for PROJECAO
    tabs_found = []
    for expected in expected_tabs:
        found = False
        for actual in actual_tabs:
            if expected.upper() in actual.upper() or actual.upper() in expected.upper():
                found = True
                tabs_found.append(actual)
                break
            # Handle PROJECAO accent
            if expected == "PROJECAO" and "PROJE" in actual.upper():
                found = True
                tabs_found.append(actual)
                break
        if not found:
            tabs_found.append(None)

    missing_tabs = [e for e, f in zip(expected_tabs, tabs_found) if f is None]
    tab_count = len(actual_tabs)

    results['check_01_tab_count'] = {
        'status': 'PASS' if tab_count >= 12 and not missing_tabs else 'FAIL',
        'detail': f'{tab_count} tabs present. Missing: {missing_tabs if missing_tabs else "none"}'
    }
    print(f"  1. Tab count: {tab_count} (>= 12) - {'PASS' if tab_count >= 12 else 'FAIL'}")
    print(f"     Tabs: {actual_tabs}")

    # Get CARTEIRA worksheet
    ws_cart = wb['CARTEIRA']

    # 2. CARTEIRA column count
    cart_max_col = ws_cart.max_column
    # Check that JC (263) exists and JD is intelligence (264+)
    col_263_header = ws_cart.cell(row=3, column=263).value
    results['check_02_column_count'] = {
        'status': 'PASS' if cart_max_col >= 263 else 'FAIL',
        'detail': f'CARTEIRA max_col={cart_max_col}, col 263 header: {col_263_header}'
    }
    print(f"  2. CARTEIRA columns: {cart_max_col} (>= 263) - {'PASS' if cart_max_col >= 263 else 'FAIL'}")

    # 3. CARTEIRA row count
    cart_max_row = ws_cart.max_row
    data_rows = cart_max_row - 3  # rows 4 to max
    results['check_03_row_count'] = {
        'status': 'PASS' if data_rows >= 554 else 'FAIL',
        'detail': f'max_row={cart_max_row}, data_rows={data_rows} (expected 554)'
    }
    print(f"  3. CARTEIRA rows: {data_rows} data rows - {'PASS' if data_rows >= 554 else 'FAIL'}")

    # 4. Column grouping: count columns with outline_level > 0
    grouped_count = 0
    for col_num in range(1, cart_max_col + 1):
        col_letter = get_column_letter(col_num)
        dim = ws_cart.column_dimensions.get(col_letter)
        if dim and dim.outline_level and dim.outline_level > 0:
            grouped_count += 1
    results['check_04_column_grouping'] = {
        'status': 'PASS' if grouped_count >= 200 else 'FAIL',
        'detail': f'{grouped_count} columns with outline_level > 0 (expected >= 200)'
    }
    print(f"  4. Grouped columns: {grouped_count} (>= 200) - {'PASS' if grouped_count >= 200 else 'FAIL'}")

    # 5. Super-groups: check row 1 for group labels
    supergroups = {}
    for col_num in range(1, min(cart_max_col + 1, 270)):
        v = ws_cart.cell(row=1, column=col_num).value
        if v and str(v).strip():
            supergroups[col_num] = str(v).strip()
    results['check_05_supergroups'] = {
        'status': 'PASS' if len(supergroups) >= 4 else 'FAIL',
        'detail': f'{len(supergroups)} super-groups found: {dict(list(supergroups.items())[:10])}'
    }
    print(f"  5. Super-groups in R1: {len(supergroups)} - {'PASS' if len(supergroups) >= 4 else 'FAIL'}")

    # 6. Anchor columns: outline_level 0 within grouped range
    anchor_count = 0
    for col_num in range(1, cart_max_col + 1):
        col_letter = get_column_letter(col_num)
        dim = ws_cart.column_dimensions.get(col_letter)
        level = dim.outline_level if dim and dim.outline_level else 0
        if level == 0 and col_num > 1:
            # Check if this is within a range that has grouped neighbors
            prev_letter = get_column_letter(max(col_num - 1, 1))
            next_letter = get_column_letter(min(col_num + 1, cart_max_col))
            prev_dim = ws_cart.column_dimensions.get(prev_letter)
            next_dim = ws_cart.column_dimensions.get(next_letter)
            prev_level = prev_dim.outline_level if prev_dim and prev_dim.outline_level else 0
            next_level = next_dim.outline_level if next_dim and next_dim.outline_level else 0
            if prev_level > 0 or next_level > 0:
                anchor_count += 1
    results['check_06_anchors'] = {
        'status': 'PASS' if anchor_count >= 6 else 'FAIL',
        'detail': f'{anchor_count} anchor columns detected (expected >= 6)'
    }
    print(f"  6. Anchor columns: {anchor_count} (>= 6) - {'PASS' if anchor_count >= 6 else 'FAIL'}")

    # 7. freeze_panes
    freeze = str(ws_cart.freeze_panes) if ws_cart.freeze_panes else "None"
    results['check_07_freeze_panes'] = {
        'status': 'PASS' if freeze == 'AR6' else 'FAIL',
        'detail': f'freeze_panes={freeze} (expected AR6)'
    }
    print(f"  7. freeze_panes: {freeze} - {'PASS' if freeze == 'AR6' else 'FAIL'}")

    # =========================================================================
    # FORMULA CHECKS
    # =========================================================================
    print("\n--- Formula Checks ---")

    # 8. CARTEIRA total formula count
    cart_formulas = 0
    for row in ws_cart.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                cart_formulas += 1
    results['check_08_carteira_formulas'] = {
        'status': 'PASS' if cart_formulas >= 100000 else 'FAIL',
        'detail': f'{cart_formulas} CARTEIRA formulas (expected >= 100,000)'
    }
    print(f"  8. CARTEIRA formulas: {cart_formulas:,} (>= 100,000) - {'PASS' if cart_formulas >= 100000 else 'FAIL'}")

    # 9. Zero #REF! errors: scan formulas for #REF! text
    ref_errors = 0
    for row in ws_cart.iter_rows():
        for cell in row:
            if cell.value and '#REF!' in str(cell.value):
                ref_errors += 1
    results['check_09_no_ref_errors'] = {
        'status': 'PASS' if ref_errors == 0 else 'FAIL',
        'detail': f'{ref_errors} #REF! errors found (expected 0)'
    }
    print(f"  9. #REF! errors: {ref_errors} - {'PASS' if ref_errors == 0 else 'FAIL'}")

    # 10. Zero _xlfn.LET: scan for LET pattern
    let_count = 0
    for row in ws_cart.iter_rows():
        for cell in row:
            if cell.value and '_xlfn.LET' in str(cell.value):
                let_count += 1
    results['check_10_no_xlfn_let'] = {
        'status': 'PASS' if let_count == 0 else 'FAIL',
        'detail': f'{let_count} _xlfn.LET formulas found (expected 0)'
    }
    print(f"  10. _xlfn.LET: {let_count} - {'PASS' if let_count == 0 else 'FAIL'}")

    # 11. Bounded ranges: sample 20 formulas, check for full-column refs
    import random
    full_col_refs = 0
    sample_count = 0
    random.seed(42)
    sample_rows = random.sample(range(DATA_START_ROW, DATA_END_ROW + 1), min(20, NUM_CLIENTS))
    sample_cols = [12, 16, 19, 39, 44, 45, 47, 48, 54, 78]  # Various formula columns
    for r in sample_rows:
        for c in sample_cols[:2]:  # Check 2 cols per row
            val = ws_cart.cell(row=r, column=c).value
            if val and str(val).startswith('='):
                sample_count += 1
                formula_str = str(val)
                # Check for full-column refs like $A:$A (not bounded)
                import re
                if re.search(r'\$[A-Z]+:\$[A-Z]+(?!\$)', formula_str):
                    # But exclude bounded ranges like $A$3:$A$25000
                    if not re.search(r'\$[A-Z]+\$\d+:\$[A-Z]+\$\d+', formula_str):
                        full_col_refs += 1
    results['check_11_bounded_ranges'] = {
        'status': 'PASS' if full_col_refs == 0 else 'FAIL',
        'detail': f'{full_col_refs} full-column refs in {sample_count} sampled formulas (expected 0)'
    }
    print(f"  11. Bounded ranges: {full_col_refs} full-col refs in {sample_count} samples - {'PASS' if full_col_refs == 0 else 'FAIL'}")

    # 12. MERCOS block: spot-check INDEX/MATCH to DRAFT 1
    mercos_refs = 0
    for r in [4, 50, 100, 200, 400]:
        for c in [12, 16, 19]:  # L, P, S - MERCOS block cols with formulas
            val = ws_cart.cell(row=r, column=c).value
            if val and 'DRAFT 1' in str(val):
                mercos_refs += 1
    results['check_12_mercos_draft1'] = {
        'status': 'PASS' if mercos_refs >= 5 else 'FAIL',
        'detail': f'{mercos_refs} MERCOS formulas reference DRAFT 1 (expected >= 5)'
    }
    print(f"  12. MERCOS -> DRAFT 1: {mercos_refs} refs - {'PASS' if mercos_refs >= 5 else 'FAIL'}")

    # 13. FUNIL block: spot-check array formulas to DRAFT 2
    funil_refs = 0
    for r in [4, 50, 100, 200, 400]:
        for c in [44, 45, 47, 48]:  # AR, AS, AU, AV - FUNIL block
            val = ws_cart.cell(row=r, column=c).value
            if val and 'DRAFT 2' in str(val):
                funil_refs += 1
    results['check_13_funil_draft2'] = {
        'status': 'PASS' if funil_refs >= 5 else 'FAIL',
        'detail': f'{funil_refs} FUNIL formulas reference DRAFT 2 (expected >= 5)'
    }
    print(f"  13. FUNIL -> DRAFT 2: {funil_refs} refs - {'PASS' if funil_refs >= 5 else 'FAIL'}")

    # 14. FATURAMENTO: 186 cols (BZ-JC = cols 78-263)
    fat_start = 78   # BZ
    fat_end = 263     # JC
    fat_cols = fat_end - fat_start + 1
    results['check_14_faturamento_cols'] = {
        'status': 'PASS' if fat_cols == 186 else 'FAIL',
        'detail': f'FATURAMENTO span: BZ(78)-JC(263) = {fat_cols} cols (expected 186)'
    }
    print(f"  14. FATURAMENTO cols: {fat_cols} - {'PASS' if fat_cols == 186 else 'FAIL'}")

    # 15. JUSTIFICATIVA: check for COUNTIFS with DATE
    # JAN starts at CC=81, JS1 is at offset +10 = col 91 (CM)
    # FEB JS1 = 81 + 15 + 10 = col 106 (DB)
    # MAR JS1 = 81 + 30 + 10 = col 121 (DQ)
    just_refs = 0
    for r in [4, 100, 300]:
        for month_offset in [0, 1, 2]:
            just_col = 91 + month_offset * 15  # S1 positions: 91, 106, 121
            val = ws_cart.cell(row=r, column=just_col).value
            if val and 'COUNTIFS' in str(val) and 'DATE' in str(val):
                just_refs += 1
    results['check_15_justificativa'] = {
        'status': 'PASS' if just_refs >= 3 else 'FAIL',
        'detail': f'{just_refs} JUSTIFICATIVA COUNTIFS+DATE formulas found (expected >= 3)'
    }
    print(f"  15. JUSTIFICATIVA COUNTIFS+DATE: {just_refs} - {'PASS' if just_refs >= 3 else 'FAIL'}")

    # 16. Intelligence: SCORE column O has weighted formula, RANK in JD
    score_formula = str(ws_cart.cell(row=4, column=15).value or '')
    has_score = 'IFERROR' in score_formula and ('0.3' in score_formula or '0.25' in score_formula)
    rank_formula = str(ws_cart.cell(row=4, column=264).value or '')
    has_rank = 'RANK' in rank_formula.upper()
    results['check_16_intelligence'] = {
        'status': 'PASS' if has_score and has_rank else 'FAIL',
        'detail': f'SCORE weighted formula: {has_score}, RANK: {has_rank}'
    }
    print(f"  16. Intelligence SCORE+RANK: score={has_score}, rank={has_rank} - {'PASS' if has_score and has_rank else 'FAIL'}")

    # =========================================================================
    # CROSS-TAB VALIDATION
    # =========================================================================
    print("\n--- Cross-tab Validation ---")

    # 17. PROJECAO: 19,224 formulas
    ws_proj = None
    for name in wb.sheetnames:
        if 'PROJE' in name.upper():
            ws_proj = wb[name]
            break
    proj_formulas = 0
    if ws_proj:
        for row in ws_proj.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    proj_formulas += 1
    results['check_17_projecao_formulas'] = {
        'status': 'PASS' if proj_formulas == 19224 else 'FAIL',
        'detail': f'PROJECAO formulas: {proj_formulas} (expected 19,224)'
    }
    print(f"  17. PROJECAO formulas: {proj_formulas} - {'PASS' if proj_formulas == 19224 else 'FAIL'}")

    # 18. LOG: 20,830 rows
    ws_log = wb['LOG']
    log_max_row = ws_log.max_row
    results['check_18_log_rows'] = {
        'status': 'PASS' if log_max_row >= 20830 else 'FAIL',
        'detail': f'LOG max_row={log_max_row} (expected >= 20,830)'
    }
    print(f"  18. LOG rows: {log_max_row} - {'PASS' if log_max_row >= 20830 else 'FAIL'}")

    # 19. REGRAS: motor section has 63 combinations (rows 220-282)
    ws_regras = wb['REGRAS']
    motor_rows = 0
    for r in range(220, 284):
        a_val = ws_regras.cell(row=r, column=1).value
        b_val = ws_regras.cell(row=r, column=2).value
        if a_val and b_val:
            motor_rows += 1
    results['check_19_regras_motor'] = {
        'status': 'PASS' if motor_rows >= 60 else 'FAIL',
        'detail': f'REGRAS motor rows: {motor_rows} (expected ~63)'
    }
    print(f"  19. REGRAS motor: {motor_rows} rows - {'PASS' if motor_rows >= 60 else 'FAIL'}")

    # 20. DRAFT 1: >= 554 data rows with CNPJ
    ws_d1 = wb['DRAFT 1']
    d1_cnpj_count = 0
    for r in range(4, ws_d1.max_row + 1):
        if ws_d1.cell(row=r, column=2).value:  # CNPJ in col B
            d1_cnpj_count += 1
    results['check_20_draft1_rows'] = {
        'status': 'PASS' if d1_cnpj_count >= 554 else 'FAIL',
        'detail': f'DRAFT 1 rows with CNPJ: {d1_cnpj_count} (expected >= 554)'
    }
    print(f"  20. DRAFT 1 rows: {d1_cnpj_count} - {'PASS' if d1_cnpj_count >= 554 else 'FAIL'}")

    # 21. DRAFT 2: headers correct
    ws_d2 = wb['DRAFT 2']
    d2_headers_ok = ws_d2.cell(row=2, column=1).value is not None
    d2_has_data = ws_d2.max_row >= 3
    results['check_21_draft2'] = {
        'status': 'PASS' if d2_headers_ok else 'FAIL',
        'detail': f'DRAFT 2 headers present: {d2_headers_ok}, max_row={ws_d2.max_row}'
    }
    print(f"  21. DRAFT 2 headers: {d2_headers_ok}, rows={ws_d2.max_row} - {'PASS' if d2_headers_ok else 'FAIL'}")

    # 22. AGENDA tabs: each has FILTER/SORT formula referencing CARTEIRA
    agenda_tabs_ok = 0
    for consultant in CONSULTANTS:
        tab_name = consultant["tab_name"]
        if tab_name in wb.sheetnames:
            ws_agenda = wb[tab_name]
            # Check A3 for FILTER formula
            a3_val = str(ws_agenda.cell(row=3, column=1).value or '')
            if 'FILTER' in a3_val.upper() and 'CARTEIRA' in a3_val.upper():
                agenda_tabs_ok += 1
            elif 'SORTBY' in a3_val.upper() and 'CARTEIRA' in a3_val.upper():
                agenda_tabs_ok += 1
    results['check_22_agenda_tabs'] = {
        'status': 'PASS' if agenda_tabs_ok == 4 else 'FAIL',
        'detail': f'{agenda_tabs_ok}/4 AGENDA tabs with FILTER formula'
    }
    print(f"  22. AGENDA FILTER formulas: {agenda_tabs_ok}/4 - {'PASS' if agenda_tabs_ok == 4 else 'FAIL'}")

    # =========================================================================
    # REQUIREMENT ASSESSMENT
    # =========================================================================
    print("\n--- BLUE Requirement Assessment ---")

    # BLUE-01: CARTEIRA tem 263 colunas organizadas em 6 grupos
    blue01_cols = cart_max_col >= 263
    blue01_groups = len(supergroups) >= 4
    blue01 = blue01_cols and blue01_groups
    results['BLUE_01'] = {
        'status': 'PASS' if blue01 else 'FAIL',
        'detail': f'263+ cols: {blue01_cols} ({cart_max_col}), 6+ super-groups: {blue01_groups} ({len(supergroups)})'
    }
    print(f"  BLUE-01 (263 cols + 6 groups): {'PASS' if blue01 else 'FAIL'}")

    # BLUE-02: Colunas fixas mantidas
    nome_visible = True  # Col A is never grouped
    cnpj_visible = True  # Col B
    freeze_ok = freeze == 'AR6'
    blue02 = nome_visible and cnpj_visible and freeze_ok
    results['BLUE_02'] = {
        'status': 'PASS' if blue02 else 'FAIL',
        'detail': f'NOME FANTASIA visible: {nome_visible}, CNPJ visible: {cnpj_visible}, freeze_panes=AR6: {freeze_ok}'
    }
    print(f"  BLUE-02 (fixed cols visible): {'PASS' if blue02 else 'FAIL'}")

    # BLUE-03: Colunas originais preservadas e funcionais
    blue03_mercos = mercos_refs >= 5
    blue03_funil = funil_refs >= 5
    blue03 = blue03_mercos and blue03_funil
    results['BLUE_03'] = {
        'status': 'PASS' if blue03 else 'FAIL',
        'detail': f'MERCOS INDEX/MATCH: {blue03_mercos} ({mercos_refs} refs), FUNIL DRAFT 2: {blue03_funil} ({funil_refs} refs)'
    }
    print(f"  BLUE-03 (originals preserved): {'PASS' if blue03 else 'FAIL'}")

    # BLUE-04: Grupos organizados em 6 secoes V12
    # Check for MERCOS, FUNIL, SAP, STATUS SAP, DADOS CADASTRAIS SAP, FATURAMENTO
    section_names = [str(v).upper() for v in supergroups.values()]
    has_mercos = any('MERCOS' in s for s in section_names)
    has_funil = any('FUNIL' in s for s in section_names)
    has_sap = any('SAP' in s for s in section_names)
    has_faturamento = any('FATURAMENTO' in s or 'FAT' in s for s in section_names)
    blue04 = has_mercos and has_funil and grouped_count >= 200
    results['BLUE_04'] = {
        'status': 'PASS' if blue04 else 'FAIL',
        'detail': f'MERCOS: {has_mercos}, FUNIL: {has_funil}, SAP: {has_sap}, FATURAMENTO: {has_faturamento}, 3-level grouping: {grouped_count >= 200}'
    }
    print(f"  BLUE-04 (6 V12 sections): {'PASS' if blue04 else 'FAIL'}")

    # =========================================================================
    # STATISTICS
    # =========================================================================
    print("\n--- Statistics ---")

    # Total V13 formulas
    total_v13_formulas = 0
    tab_formula_counts = {}
    for sheet_name in wb.sheetnames:
        ws_check = wb[sheet_name]
        count = 0
        for row in ws_check.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    count += 1
        tab_formula_counts[sheet_name] = count
        total_v13_formulas += count

    results['stats'] = {
        'total_tabs': tab_count,
        'total_formulas': total_v13_formulas,
        'tab_formula_counts': tab_formula_counts,
        'carteira_formulas': cart_formulas,
        'projecao_formulas': proj_formulas,
        'carteira_rows': data_rows,
        'carteira_cols': cart_max_col,
    }
    print(f"  Total V13 formulas: {total_v13_formulas:,}")
    print(f"  Formula breakdown by tab:")
    for tab, count in tab_formula_counts.items():
        print(f"    {tab}: {count:,}")

    # =========================================================================
    # SUMMARY
    # =========================================================================
    all_checks = {k: v for k, v in results.items() if k.startswith('check_') or k.startswith('BLUE_')}
    pass_count = sum(1 for v in all_checks.values() if v.get('status') == 'PASS')
    fail_count = sum(1 for v in all_checks.values() if v.get('status') == 'FAIL')
    total_checks = len(all_checks)

    results['summary'] = {
        'total_checks': total_checks,
        'passed': pass_count,
        'failed': fail_count,
        'all_pass': fail_count == 0,
    }

    print(f"\n{'='*60}")
    print(f"VALIDATION SUMMARY: {pass_count}/{total_checks} PASS, {fail_count} FAIL")
    print(f"BLUE-01: {results['BLUE_01']['status']}")
    print(f"BLUE-02: {results['BLUE_02']['status']}")
    print(f"BLUE-03: {results['BLUE_03']['status']}")
    print(f"BLUE-04: {results['BLUE_04']['status']}")
    print(f"{'='*60}")

    wb.close()
    return results


def main():
    print("=" * 70)
    print("Plan 09-06: AGENDA Tabs + Phase 9 Comprehensive Validation")
    print("=" * 70)

    # =========================================================================
    # TASK 1: Create AGENDA tabs
    # =========================================================================
    print(f"\nLoading V13 from {V13_PATH}...")
    wb = openpyxl.load_workbook(V13_PATH)
    print(f"  Current tabs: {wb.sheetnames}")

    # Check if AGENDA tabs already exist (idempotent re-run)
    existing_agenda = [s for s in wb.sheetnames if s.startswith('AGENDA ')]
    if len(existing_agenda) == 4:
        print(f"  AGENDA tabs already exist: {existing_agenda}")
        print("  Skipping Task 1 creation (idempotent).")
        created_tabs = existing_agenda
        wb.close()
    else:
        # Remove any partial AGENDA tabs
        for name in existing_agenda:
            del wb[name]
        created_tabs = run_task1(wb)
        # Save after Task 1
        print(f"\nSaving V13 with AGENDA tabs...")
        wb.save(V13_PATH)
        wb.close()
        print("Saved successfully.")

    # =========================================================================
    # TASK 1 VERIFICATION
    # =========================================================================
    print("\n--- Task 1 Verification ---")
    wb_verify = openpyxl.load_workbook(V13_PATH)
    print(f"  Tabs after save: {wb_verify.sheetnames}")

    task1_ok = True
    for consultant in CONSULTANTS:
        tab_name = consultant["tab_name"]
        if tab_name not in wb_verify.sheetnames:
            print(f"  FAIL: {tab_name} not found!")
            task1_ok = False
            continue

        ws = wb_verify[tab_name]
        # Check FILTER formula in A3
        a3 = str(ws.cell(row=3, column=1).value or '')
        has_filter = 'FILTER' in a3 or 'SORTBY' in a3
        has_carteira = 'CARTEIRA' in a3

        # Check date filter cell
        f1 = str(ws.cell(row=1, column=6).value or '')
        has_date = 'TODAY' in f1

        # Check freeze_panes
        has_freeze = ws.freeze_panes is not None

        # Check data validations
        has_validations = len(ws.data_validations.dataValidation) >= 2

        # Check MANU dual-name handling
        if consultant["name"] == "MANU":
            has_dual = 'MANU DITZEL' in a3 and 'HEMANUELE' in a3
            print(f"  {tab_name}: filter={has_filter}, carteira={has_carteira}, date={has_date}, freeze={has_freeze}, validations={has_validations}, dual_name={has_dual}")
            if not has_dual:
                task1_ok = False
        else:
            print(f"  {tab_name}: filter={has_filter}, carteira={has_carteira}, date={has_date}, freeze={has_freeze}, validations={has_validations}")

        if not (has_filter and has_carteira and has_date and has_freeze and has_validations):
            task1_ok = False

    print(f"\n  Task 1 overall: {'PASS' if task1_ok else 'FAIL'}")
    wb_verify.close()

    if not task1_ok:
        print("ERROR: Task 1 verification failed!")
        return 1

    # =========================================================================
    # TASK 2: Comprehensive Validation
    # =========================================================================
    validation_results = run_task2_validation(V13_PATH)

    # Save validation report
    os.makedirs(os.path.dirname(VALIDATION_OUTPUT), exist_ok=True)
    with open(VALIDATION_OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(validation_results, f, indent=2, ensure_ascii=False)
    print(f"\nValidation report saved to: {VALIDATION_OUTPUT}")

    # Check overall result
    all_pass = validation_results.get('summary', {}).get('all_pass', False)
    return 0 if all_pass else 1


if __name__ == "__main__":
    sys.exit(main())

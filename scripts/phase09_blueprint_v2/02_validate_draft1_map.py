"""
Phase 09 Plan 01 - Task 2: Validate audit and build DRAFT 1 column position map
Validates the audit output and creates draft1_column_map.json.

Outputs:
  - data/output/phase09/draft1_column_map.json

Python: /c/Users/User/.pyenv/pyenv-win/pyenv-win/versions/3.12.10/python.exe
"""

import json
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# Paths
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
V12_PATH = PROJECT_ROOT / "data" / "sources" / "crm-versoes" / "v11-v12" / "CRM_INTELIGENTE_VITAO360_V12_COM_DADOS.xlsx"
V13_DRAFT1_PATH = PROJECT_ROOT / "data" / "sources" / "drafts" / "DL_DRAFT1_FEV2026.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "data" / "output" / "phase09"
SPEC_PATH = OUTPUT_DIR / "carteira_column_spec.json"
AUDIT_PATH = OUTPUT_DIR / "v12_formula_audit.json"


def extract_draft1_col_refs(formula_template):
    """Extract DRAFT 1 column letters referenced in a formula template."""
    if not formula_template:
        return []
    # Patterns:
    #   'DRAFT 1'!$L:$L
    #   'DRAFT 1'!$B:$B
    #   'DRAFT 1'!$L$3:$L$5000
    refs = re.findall(r"'DRAFT 1'!\$([A-Z]+)", formula_template)
    return sorted(set(refs))


def main():
    print("=" * 60)
    print("Phase 09 Plan 01 - Task 2: Validate & DRAFT 1 Column Map")
    print("=" * 60)

    # --- Load audit artifacts ---
    with open(SPEC_PATH, "r", encoding="utf-8") as f:
        spec = json.load(f)

    with open(AUDIT_PATH, "r", encoding="utf-8") as f:
        audit = json.load(f)

    # --- Read V12 DRAFT 1 column headers ---
    print("\nReading V12 DRAFT 1...")
    wb_v12 = load_workbook(str(V12_PATH), read_only=True, data_only=False)
    ws_d1_v12 = wb_v12["DRAFT 1"]
    v12_d1_max_col = ws_d1_v12.max_column
    v12_d1_max_row = ws_d1_v12.max_row
    print(f"  V12 DRAFT 1: {v12_d1_max_row} rows x {v12_d1_max_col} cols")

    v12_draft1_columns = {}
    for c in range(1, v12_d1_max_col + 1):
        header = ws_d1_v12.cell(row=3, column=c).value
        if header is not None:
            col_letter = get_column_letter(c)
            v12_draft1_columns[col_letter] = str(header).strip()

    print(f"  V12 DRAFT 1 headers (R3): {len(v12_draft1_columns)} columns")
    wb_v12.close()

    # --- Read V13 DRAFT 1 (standalone file) column headers ---
    print("\nReading V13 DRAFT 1...")
    wb_v13 = load_workbook(str(V13_DRAFT1_PATH), read_only=True, data_only=True)
    # Find the DRAFT 1 sheet (may have trailing space)
    d1_sheet_name = None
    for sn in wb_v13.sheetnames:
        if "DRAFT 1" in sn.upper():
            d1_sheet_name = sn
            break
    if not d1_sheet_name:
        d1_sheet_name = wb_v13.sheetnames[0]

    ws_d1_v13 = wb_v13[d1_sheet_name]
    v13_d1_max_col = ws_d1_v13.max_column
    v13_d1_max_row = ws_d1_v13.max_row
    print(f"  V13 DRAFT 1 ({d1_sheet_name}): {v13_d1_max_row} rows x {v13_d1_max_col} cols")

    v13_draft1_columns = {}
    for c in range(1, v13_d1_max_col + 1):
        header = ws_d1_v13.cell(row=3, column=c).value
        if header is not None:
            col_letter = get_column_letter(c)
            v13_draft1_columns[col_letter] = str(header).strip()

    print(f"  V13 DRAFT 1 headers (R3): {len(v13_draft1_columns)} columns")
    wb_v13.close()

    # --- Extract which DRAFT 1 columns are referenced by CARTEIRA ---
    print("\nAnalyzing DRAFT 1 column references in CARTEIRA formulas...")
    referenced_d1_cols = set()
    carteira_to_d1_refs = {}

    for col in spec["columns"]:
        if "DRAFT 1" in col.get("references_tabs", []):
            d1_refs = extract_draft1_col_refs(col.get("formula_template", ""))
            if d1_refs:
                referenced_d1_cols.update(d1_refs)
                carteira_to_d1_refs[col["col_letter"]] = {
                    "carteira_header": col["header_r3"],
                    "draft1_cols_referenced": d1_refs
                }

    print(f"  DRAFT 1 columns referenced by CARTEIRA: {sorted(referenced_d1_cols)}")

    # --- Build remapping ---
    remapping = {}
    mismatches = []

    for d1_col in sorted(referenced_d1_cols):
        v12_header = v12_draft1_columns.get(d1_col, None)
        v13_header = v13_draft1_columns.get(d1_col, None)

        if v12_header is None:
            v12_header = f"(NOT FOUND in V12 at {d1_col})"

        # Find V13 column with same header
        v13_col = None
        if v13_header and v12_header and v12_header == v13_header:
            v13_col = d1_col  # Same position
        else:
            # Search V13 for the header
            for v13_c, v13_h in v13_draft1_columns.items():
                if v13_h == v12_header:
                    v13_col = v13_c
                    break

        remapping[v12_header or d1_col] = {
            "v12_col": d1_col,
            "v13_col": v13_col or "NOT_FOUND",
            "v12_header": v12_header,
            "v13_header": v13_header,
            "match": d1_col == v13_col if v13_col else False
        }

        if v13_col and d1_col != v13_col:
            mismatches.append(f"{v12_header}: V12={d1_col} -> V13={v13_col}")
        elif not v13_col:
            mismatches.append(f"{v12_header}: V12={d1_col} -> NOT FOUND in V13")

    # --- Output draft1_column_map.json ---
    draft1_map = {
        "v12_draft1_columns": v12_draft1_columns,
        "v13_draft1_columns": v13_draft1_columns,
        "referenced_by_carteira": sorted(list(referenced_d1_cols)),
        "carteira_formula_refs": carteira_to_d1_refs,
        "remapping": remapping,
        "mismatches": mismatches,
        "v13_draft1_path": str(V13_DRAFT1_PATH.relative_to(PROJECT_ROOT)),
        "v13_draft1_max_row": v13_d1_max_row,
        "v12_draft1_max_row": v12_d1_max_row,
        "columns_match": len(mismatches) == 0
    }

    map_path = OUTPUT_DIR / "draft1_column_map.json"
    with open(map_path, "w", encoding="utf-8") as f:
        json.dump(draft1_map, f, ensure_ascii=False, indent=2)
    print(f"\n  Wrote: {map_path}")

    # --- Supplementary Validations ---
    print("\n" + "=" * 60)
    print("SUPPLEMENTARY VALIDATIONS")
    print("=" * 60)

    errors = []

    # 1. CNPJ same position (col B) in both V12 and V13
    v12_cnpj = v12_draft1_columns.get("B", "")
    v13_cnpj = v13_draft1_columns.get("B", "")
    if "CNPJ" in v12_cnpj.upper() and "CNPJ" in v13_cnpj.upper():
        print(f"  [PASS] CNPJ at col B in both V12 ({v12_cnpj}) and V13 ({v13_cnpj})")
    else:
        errors.append(f"CNPJ mismatch: V12 B={v12_cnpj}, V13 B={v13_cnpj}")

    # 2. All CARTEIRA formula references to DRAFT 1 have valid V13 mapping
    unmapped = [k for k, v in remapping.items() if v["v13_col"] == "NOT_FOUND"]
    if unmapped:
        errors.append(f"{len(unmapped)} DRAFT 1 columns unmapped: {unmapped}")
    else:
        print(f"  [PASS] All {len(remapping)} DRAFT 1 references have valid V13 mapping")

    # 3. FATURAMENTO super-group has 186 columns
    fat_count = sum(1 for col in spec["columns"] if col["super_group"] == "FATURAMENTO")
    if fat_count == 186:
        print(f"  [PASS] FATURAMENTO has exactly {fat_count} columns")
    else:
        errors.append(f"FATURAMENTO expected 186 columns, got {fat_count}")

    # 4. At least 6 anchor columns (1 per super-group minimum)
    anchors_per_sg = {}
    for col in spec["columns"]:
        if col["is_anchor"]:
            sg = col["super_group"]
            if sg not in anchors_per_sg:
                anchors_per_sg[sg] = []
            anchors_per_sg[sg].append(col["col_letter"])

    sgs_with_anchors = len(anchors_per_sg)
    total_anchors = sum(len(v) for v in anchors_per_sg.values())
    if sgs_with_anchors >= 6:
        print(f"  [PASS] {sgs_with_anchors} super-groups have anchors ({total_anchors} total anchor cols)")
        for sg, anchors in anchors_per_sg.items():
            print(f"    {sg}: {len(anchors)} anchors ({', '.join(anchors[:5])}{'...' if len(anchors) > 5 else ''})")
    else:
        errors.append(f"Expected at least 6 super-groups with anchors, got {sgs_with_anchors}")

    # 5. LET functions documented
    let_cols = audit.get("let_function_columns", [])
    if let_cols:
        print(f"  [PASS] {len(let_cols)} _xlfn.LET formulas documented for rewrite:")
        for lf in let_cols:
            print(f"    {lf['col_letter']} ({lf['header_r3']})")
    else:
        print("  [INFO] No _xlfn.LET formulas found")

    # 6. Column position mismatches
    if mismatches:
        print(f"\n  [WARN] {len(mismatches)} column position mismatches:")
        for m in mismatches:
            print(f"    {m}")
    else:
        print(f"  [PASS] All referenced DRAFT 1 columns at same position in V12 and V13")

    # 7. Check for projecao_ref patterns (META from PROJECAO)
    projecao_cols = [col for col in spec["columns"] if col["formula_pattern"] == "projecao_ref"]
    if projecao_cols:
        print(f"  [INFO] {len(projecao_cols)} columns reference PROJECAO directly")
    else:
        print("  [INFO] No direct PROJECAO references in sample rows (META may be static in V12)")
        print("         V13 will need: META MES = PROJECAO col L / 12")

    # Final verdict
    if errors:
        print(f"\n  ERRORS: {len(errors)}")
        for e in errors:
            print(f"    - {e}")
    else:
        print(f"\n  ALL VALIDATIONS PASSED")

    return len(errors) == 0


if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)

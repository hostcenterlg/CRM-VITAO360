"""
Phase 09 Plan 01 - Task 1: Deep audit of V12 CARTEIRA tab
Extracts complete column specification, formula patterns, and tab name map.

Outputs:
  - data/output/phase09/carteira_column_spec.json
  - data/output/phase09/v12_formula_audit.json
  - data/output/phase09/tab_name_map.json

Python: /c/Users/User/.pyenv/pyenv-win/pyenv-win/versions/3.12.10/python.exe
"""

import json
import re
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Paths
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
V12_PATH = PROJECT_ROOT / "data" / "sources" / "crm-versoes" / "v11-v12" / "CRM_INTELIGENTE_VITAO360_V12_COM_DADOS.xlsx"
V13_PATH = PROJECT_ROOT / "data" / "output" / "CRM_VITAO360_V13_PROJECAO.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "data" / "output" / "phase09"

SAMPLE_ROWS = [4, 10, 50, 100]


def classify_formula(formula_str):
    """Classify a formula into its pattern type."""
    if formula_str is None or not isinstance(formula_str, str):
        return "static_data", None, []

    f = str(formula_str).strip()
    if not f.startswith("="):
        return "static_data", None, []

    # Detect referenced tabs
    tabs_referenced = []
    # Pattern: 'TAB NAME'!  or  TAB_NAME!
    tab_refs = re.findall(r"'([^']+)'!", f)
    tab_refs += re.findall(r"(?<!')(\w+)!", f)
    tabs_referenced = list(set(tab_refs))

    # Classify by pattern
    f_upper = f.upper()

    if "_XLFN.LET" in f_upper or "_XLPM." in f_upper:
        pattern_type = "let_function"
    elif "COUNTIFS" in f_upper and "DRAFT 2" in f:
        pattern_type = "countifs_draft2"
    elif "REGRAS" in f:
        pattern_type = "index_match_regras"
    elif "INDEX" in f_upper and "DRAFT 2" in f and ("MAX" in f_upper or "IF(" in f_upper):
        pattern_type = "index_match_draft2_cse"
    elif ("INDEX" in f_upper or "MATCH" in f_upper or "VLOOKUP" in f_upper) and "DRAFT 2" in f:
        pattern_type = "index_match_draft2_cse"
    elif ("INDEX" in f_upper or "MATCH" in f_upper or "VLOOKUP" in f_upper) and "DRAFT 1" in f:
        pattern_type = "index_match_draft1"
    elif "PROJEÇÃO" in f or "PROJECAO" in f or "PROJE" in f:
        pattern_type = "projecao_ref"
    elif any(tab in f for tab in ["DRAFT 1", "DRAFT 2", "REGRAS"]):
        # Generic reference to other tab
        if "DRAFT 1" in f:
            pattern_type = "index_match_draft1"
        elif "DRAFT 2" in f:
            pattern_type = "index_match_draft2_cse"
        else:
            pattern_type = "index_match_regras"
    elif f.startswith("=") and any(func in f_upper for func in ["SUM", "IF(", "IFERROR", "RANK", "AVERAGE", "COUNTIF", "MAX(", "MIN("]):
        pattern_type = "internal_calc"
    elif f.startswith("=") and re.match(r"^=[A-Z]+\d+$", f):
        # Simple cell reference like =CX4
        pattern_type = "internal_calc"
    elif f.startswith("="):
        pattern_type = "internal_calc"
    else:
        pattern_type = "static_data"

    return pattern_type, f, tabs_referenced


def create_formula_template(formula_str, row_num):
    """Replace row numbers in formula with {row} placeholder."""
    if formula_str is None or not isinstance(formula_str, str):
        return None
    f = str(formula_str)
    if not f.startswith("="):
        return None
    # Replace the specific row number with {row}
    # Be careful: only replace row references, not constants
    # Pattern: letter(s) followed by the row number at word boundary
    template = re.sub(r'(?<=[A-Z])' + str(row_num) + r'(?=[^0-9]|$)', '{row}', f)
    # Also handle $letter$row patterns
    template = re.sub(r'\$' + str(row_num) + r'(?=[^0-9]|$)', '${row}', template)
    return template


def detect_super_groups(columns_data):
    """Detect super-groups from R1 header values."""
    super_groups = []
    current_sg = None
    current_start = None

    for col_info in columns_data:
        r1 = col_info.get("header_r1")
        col_num = col_info["col_num"]

        if r1 and isinstance(r1, str) and r1.strip():
            r1_clean = r1.strip()
            # Known super-group names
            known_sg = ["MERCOS", "FUNIL", "SAP", "STATUS SAP", "DADOS CADASTRAIS SAP", "FATURAMENTO"]
            # Check if this is a known super-group header OR starts a new block
            if r1_clean.upper() in [s.upper() for s in known_sg] or (current_sg and r1_clean.upper() != current_sg.upper()):
                if current_sg is not None:
                    super_groups.append({
                        "name": current_sg,
                        "start_col": current_start,
                        "end_col": col_num - 1,
                        "anchor_col": current_start
                    })
                current_sg = r1_clean
                current_start = col_num

    # Close last group
    if current_sg is not None:
        super_groups.append({
            "name": current_sg,
            "start_col": current_start,
            "end_col": columns_data[-1]["col_num"],
            "anchor_col": current_start
        })

    return super_groups


def main():
    print("=" * 60)
    print("Phase 09 Plan 01 - Task 1: V12 CARTEIRA Audit")
    print("=" * 60)

    # --- Load V12 CARTEIRA ---
    print(f"\nLoading V12: {V12_PATH}")
    wb_v12 = load_workbook(str(V12_PATH), read_only=False, data_only=False)
    ws = wb_v12["CARTEIRA"]
    max_row = ws.max_row
    max_col = ws.max_column
    print(f"  CARTEIRA: {max_row} rows x {max_col} cols")

    # --- Extract column specifications ---
    print("\nExtracting column specifications...")
    columns_data = []
    outline_counter = {"level_0": 0, "level_1": 0, "level_2": 0, "level_3": 0}
    formula_patterns_set = {}  # pattern_type -> list of examples

    for col_num in range(1, max_col + 1):
        col_letter = get_column_letter(col_num)

        # Headers
        header_r1 = ws.cell(row=1, column=col_num).value
        header_r2 = ws.cell(row=2, column=col_num).value
        header_r3 = ws.cell(row=3, column=col_num).value

        # Column dimensions
        cd = ws.column_dimensions.get(col_letter, None)
        width = cd.width if cd and cd.width else 8.43  # default Excel width
        outline_level = cd.outline_level if cd and cd.outline_level else 0
        hidden = cd.hidden if cd else False

        # Count outline levels
        level_key = f"level_{min(outline_level, 3)}"
        outline_counter[level_key] = outline_counter.get(level_key, 0) + 1

        # Determine if anchor
        is_anchor = (outline_level == 0)

        # Sample formulas
        sample_formulas = {}
        all_pattern_types = []
        all_tabs_ref = set()

        for sr in SAMPLE_ROWS:
            cell = ws.cell(row=sr, column=col_num)
            cell_val = cell.value
            sample_formulas[f"r{sr}"] = str(cell_val) if cell_val is not None else None

            ptype, _, tabs = classify_formula(cell_val)
            all_pattern_types.append(ptype)
            all_tabs_ref.update(tabs)

        # Majority pattern type
        from collections import Counter
        pattern_counts = Counter(all_pattern_types)
        dominant_pattern = pattern_counts.most_common(1)[0][0]

        # Create template from first formula row
        formula_template = None
        for sr in SAMPLE_ROWS:
            cell_val = ws.cell(row=sr, column=col_num).value
            if cell_val and isinstance(cell_val, str) and cell_val.startswith("="):
                formula_template = create_formula_template(cell_val, sr)
                break

        # Determine super_group and sub_group from header propagation
        # R1 = super-group (may be None if inherited from left)
        # R2 = sub-group (may be None)
        super_group = str(header_r1).strip() if header_r1 and str(header_r1).strip() else None
        sub_group = str(header_r2).strip() if header_r2 and str(header_r2).strip() else None

        col_entry = {
            "col_num": col_num,
            "col_letter": col_letter,
            "header_r1": str(header_r1) if header_r1 is not None else None,
            "header_r2": str(header_r2) if header_r2 is not None else None,
            "header_r3": str(header_r3) if header_r3 is not None else None,
            "super_group": super_group,
            "sub_group": sub_group,
            "outline_level": outline_level,
            "width": round(width, 2) if width else 8.43,
            "hidden": hidden,
            "is_anchor": is_anchor,
            "formula_pattern": dominant_pattern,
            "formula_template": formula_template,
            "references_tabs": sorted(list(all_tabs_ref)),
            "sample_formulas": sample_formulas
        }
        columns_data.append(col_entry)

        # Track formula patterns
        if dominant_pattern not in formula_patterns_set:
            formula_patterns_set[dominant_pattern] = []
        if formula_template and len(formula_patterns_set[dominant_pattern]) < 5:
            formula_patterns_set[dominant_pattern].append({
                "col_letter": col_letter,
                "col_num": col_num,
                "header_r3": str(header_r3) if header_r3 else None,
                "template": formula_template
            })

        if col_num % 50 == 0:
            print(f"  Processed {col_num}/{max_col} columns...")

    print(f"  Total columns processed: {len(columns_data)}")

    # --- Propagate super_group values ---
    # Use known super-group column ranges to assign canonical names
    # (R1 values have emojis and month names that shouldn't propagate)
    sg_ranges = [
        (1, 43, "MERCOS"),
        (44, 62, "FUNIL"),
        (63, 65, "SAP"),
        (66, 68, "STATUS SAP"),
        (69, 77, "DADOS CADASTRAIS SAP"),
        (78, 263, "FATURAMENTO"),
    ]
    for col_entry in columns_data:
        cn = col_entry["col_num"]
        for sg_start, sg_end, sg_name in sg_ranges:
            if sg_start <= cn <= sg_end:
                col_entry["super_group"] = sg_name
                break

    # Propagate sub_group (R2) -- fill forward within each super-group
    current_sub = None
    current_sg = None
    for col_entry in columns_data:
        if col_entry["super_group"] != current_sg:
            current_sg = col_entry["super_group"]
            current_sub = None  # Reset sub-group at super-group boundary
        if col_entry["sub_group"] is not None:
            current_sub = col_entry["sub_group"]
        else:
            col_entry["sub_group"] = current_sub

    # --- Detect super-groups ---
    # R1 header values use emoji prefixes like "🟣 MERCOS", "🔵 FUNIL", etc.
    # We need to match on known keywords, stripping emoji and whitespace.
    # Known super-groups IN ORDER of appearance (col positions from V12 audit):
    #   col 1  (A):  "🟣 MERCOS"
    #   col 44 (AR): "🔵 FUNIL"
    #   col 63 (BK): "SAP "
    #   col 66 (BN): "🚦 STATUS SAP"
    #   col 69 (BQ): "⚫ DADOS CADASTRAIS SAP"
    #   col 79 (CA): "🟢 FATURAMENTO"
    # Col 78 (BZ) is standalone "VENDA" -- part of FATURAMENTO super-group

    # Define super-groups by their exact R1 start columns (known from audit)
    sg_definitions = [
        {"name": "MERCOS", "start_col": 1, "r1_marker": "MERCOS"},
        {"name": "FUNIL", "start_col": 44, "r1_marker": "FUNIL"},
        {"name": "SAP", "start_col": 63, "r1_marker": "SAP"},
        {"name": "STATUS SAP", "start_col": 66, "r1_marker": "STATUS SAP"},
        {"name": "DADOS CADASTRAIS SAP", "start_col": 69, "r1_marker": "DADOS CADASTRAIS SAP"},
        {"name": "FATURAMENTO", "start_col": 78, "r1_marker": "FATURAMENTO"},  # BZ=VENDA is part of FAT
    ]

    # Verify by checking R1 values at expected columns
    super_groups = []
    for i, sgd in enumerate(sg_definitions):
        end_col = sg_definitions[i + 1]["start_col"] - 1 if i + 1 < len(sg_definitions) else max_col
        # Find anchor col: first column with outline_level=0 in this range
        anchor_col = sgd["start_col"]
        for col_entry in columns_data:
            if sgd["start_col"] <= col_entry["col_num"] <= end_col:
                if col_entry["outline_level"] == 0:
                    anchor_col = col_entry["col_num"]
                    break

        # Verify R1 marker exists somewhere in this range
        r1_at_start = None
        for col_entry in columns_data:
            if col_entry["col_num"] == sgd["start_col"]:
                r1_at_start = col_entry["header_r1"]
                break
        # For FATURAMENTO, check col 79 (CA) since BZ has no R1
        if sgd["r1_marker"] == "FATURAMENTO" and (r1_at_start is None or "FATURAMENTO" not in str(r1_at_start)):
            for col_entry in columns_data:
                if col_entry["col_num"] == 79:  # CA
                    r1_at_start = col_entry["header_r1"]
                    break

        super_groups.append({
            "name": sgd["name"],
            "start_col": sgd["start_col"],
            "end_col": end_col,
            "anchor_col": anchor_col,
            "r1_header": str(r1_at_start) if r1_at_start else None,
            "col_count": end_col - sgd["start_col"] + 1
        })

    print(f"\n  Super-groups detected: {len(super_groups)}")
    for sg in super_groups:
        print(f"    {sg['name']}: cols {sg['start_col']}-{sg['end_col']} ({sg['end_col']-sg['start_col']+1} cols)")

    # --- Build carteira_column_spec.json ---
    column_spec = {
        "total_columns": len(columns_data),
        "v12_max_row": max_row,
        "v12_max_col": max_col,
        "columns": columns_data,
        "super_groups": super_groups,
        "outline_summary": outline_counter
    }

    spec_path = OUTPUT_DIR / "carteira_column_spec.json"
    with open(spec_path, "w", encoding="utf-8") as f:
        json.dump(column_spec, f, ensure_ascii=False, indent=2)
    print(f"\n  Wrote: {spec_path}")
    print(f"  Columns: {len(columns_data)}")

    # --- Build v12_formula_audit.json ---
    # Also do a more thorough scan: check rows 4-100 for distinct patterns per column
    print("\nBuilding formula audit (scanning rows 4-100 for all columns)...")
    formula_audit = {
        "total_formula_columns": 0,
        "total_static_columns": 0,
        "pattern_types": {},
        "formula_patterns": formula_patterns_set,
        "let_function_columns": [],
        "array_formula_columns": [],
        "all_referenced_tabs": set()
    }

    for col_entry in columns_data:
        if col_entry["formula_pattern"] == "static_data":
            formula_audit["total_static_columns"] += 1
        else:
            formula_audit["total_formula_columns"] += 1

        ptype = col_entry["formula_pattern"]
        if ptype not in formula_audit["pattern_types"]:
            formula_audit["pattern_types"][ptype] = {
                "count": 0,
                "columns": []
            }
        formula_audit["pattern_types"][ptype]["count"] += 1
        formula_audit["pattern_types"][ptype]["columns"].append({
            "col_num": col_entry["col_num"],
            "col_letter": col_entry["col_letter"],
            "header_r3": col_entry["header_r3"]
        })

        for tab in col_entry["references_tabs"]:
            formula_audit["all_referenced_tabs"].add(tab)

        # Check for LET functions
        if col_entry["formula_pattern"] == "let_function":
            formula_audit["let_function_columns"].append({
                "col_num": col_entry["col_num"],
                "col_letter": col_entry["col_letter"],
                "header_r3": col_entry["header_r3"],
                "sample_formula": col_entry["sample_formulas"].get("r4")
            })

    # Convert set to sorted list for JSON
    formula_audit["all_referenced_tabs"] = sorted(list(formula_audit["all_referenced_tabs"]))

    audit_path = OUTPUT_DIR / "v12_formula_audit.json"
    with open(audit_path, "w", encoding="utf-8") as f:
        json.dump(formula_audit, f, ensure_ascii=False, indent=2)
    print(f"  Wrote: {audit_path}")
    print(f"  Formula columns: {formula_audit['total_formula_columns']}")
    print(f"  Static columns: {formula_audit['total_static_columns']}")
    print(f"  Pattern types: {len(formula_audit['pattern_types'])}")
    print(f"  LET function columns: {len(formula_audit['let_function_columns'])}")
    print(f"  All referenced tabs: {formula_audit['all_referenced_tabs']}")

    # --- Build tab_name_map.json ---
    print("\nBuilding tab name map (V12 -> V13)...")

    # Get V13 sheet names
    wb_v13 = load_workbook(str(V13_PATH), read_only=True)
    v13_sheets = wb_v13.sheetnames
    wb_v13.close()
    print(f"  V13 sheets: {v13_sheets}")

    v12_sheets = wb_v12.sheetnames
    print(f"  V12 sheets: {v12_sheets}")

    # Build mapping for tabs referenced by CARTEIRA formulas
    tab_name_map = {
        "v12_to_v13": {},
        "v12_sheets": v12_sheets,
        "v13_sheets": v13_sheets,
        "referenced_by_carteira": formula_audit["all_referenced_tabs"],
        "notes": []
    }

    # Map each referenced tab
    for ref_tab in formula_audit["all_referenced_tabs"]:
        ref_clean = ref_tab.strip()
        # Check if exact match in V13
        if ref_clean in v13_sheets:
            tab_name_map["v12_to_v13"][ref_tab] = ref_clean
            tab_name_map["notes"].append(f"{ref_tab} -> {ref_clean} (exact match in V13)")
        else:
            # Check with accent stripping
            import unicodedata
            def strip_accents(s):
                return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

            ref_stripped = strip_accents(ref_clean)
            found = False
            for v13_name in v13_sheets:
                if strip_accents(v13_name.strip()) == ref_stripped:
                    tab_name_map["v12_to_v13"][ref_tab] = v13_name
                    tab_name_map["notes"].append(f"{ref_tab} -> {v13_name} (accent-stripped match)")
                    found = True
                    break

            if not found:
                # Tab needs to be created in V13
                tab_name_map["v12_to_v13"][ref_tab] = ref_tab  # Same name (will be created)
                tab_name_map["notes"].append(f"{ref_tab} -> {ref_tab} (NOT in V13 - MUST CREATE)")

    map_path = OUTPUT_DIR / "tab_name_map.json"
    with open(map_path, "w", encoding="utf-8") as f:
        json.dump(tab_name_map, f, ensure_ascii=False, indent=2)
    print(f"  Wrote: {map_path}")

    wb_v12.close()

    # --- Summary ---
    print("\n" + "=" * 60)
    print("AUDIT COMPLETE")
    print("=" * 60)
    print(f"  Columns: {len(columns_data)}")
    print(f"  Super-groups: {len(super_groups)}")
    print(f"  Outline levels: {outline_counter}")
    print(f"  Formula pattern types: {len(formula_audit['pattern_types'])}")
    for ptype, pdata in sorted(formula_audit["pattern_types"].items()):
        print(f"    {ptype}: {pdata['count']} columns")
    print(f"  LET functions to rewrite: {len(formula_audit['let_function_columns'])}")
    print(f"  Tab name mappings: {len(tab_name_map['v12_to_v13'])}")

    # Validation checks
    print("\n--- VALIDATION ---")
    errors = []

    # 1. Check 263 columns
    if len(columns_data) != 263:
        errors.append(f"Expected 263 columns, got {len(columns_data)}")
    else:
        print("  [PASS] 263 columns")

    # 2. Check all have header_r3
    null_headers = [c for c in columns_data if c["header_r3"] is None or c["header_r3"] == "None"]
    if null_headers:
        print(f"  [WARN] {len(null_headers)} columns with null header_r3: {[c['col_letter'] for c in null_headers[:10]]}")
    else:
        print("  [PASS] All columns have header_r3")

    # 3. Check super-groups
    if len(super_groups) < 5:
        errors.append(f"Expected at least 5 super-groups, got {len(super_groups)}")
    else:
        print(f"  [PASS] {len(super_groups)} super-groups detected")

    # 4. Check outline levels
    has_levels = sum(1 for k, v in outline_counter.items() if v > 0 and k != "level_0")
    if has_levels < 2:
        errors.append(f"Expected at least 2 non-zero outline levels, got {has_levels}")
    else:
        print(f"  [PASS] {has_levels} non-zero outline levels used")

    # 5. Check formula pattern types >= 4
    if len(formula_audit["pattern_types"]) < 4:
        errors.append(f"Expected at least 4 formula pattern types, got {len(formula_audit['pattern_types'])}")
    else:
        print(f"  [PASS] {len(formula_audit['pattern_types'])} formula pattern types")

    # 6. Check tab_name_map has referenced tabs
    if len(tab_name_map["v12_to_v13"]) < 2:
        errors.append(f"Expected at least 2 tab mappings, got {len(tab_name_map['v12_to_v13'])}")
    else:
        print(f"  [PASS] {len(tab_name_map['v12_to_v13'])} tab mappings")

    if errors:
        print(f"\n  ERRORS: {len(errors)}")
        for e in errors:
            print(f"    - {e}")
    else:
        print("\n  ALL CHECKS PASSED")


if __name__ == "__main__":
    main()

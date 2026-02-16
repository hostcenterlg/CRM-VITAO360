"""
Build CRM VITAO360 V3 — Orchestrator
Creates output/CRM_VITAO360_V3.xlsx with 9 tabs in correct order.
"""
import os
import sys
import time
import openpyxl

# Add working dir to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from v3_regras import build_regras
from v3_draft1 import build_draft1
from v3_projecao import build_projecao
from v3_carteira import build_carteira
from v3_draft2 import build_draft2
from v3_agenda import build_agenda
from v3_dash import build_dash
from v3_log import build_log, build_claude_log

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "CRM_VITAO360_V3.xlsx")


def main():
    start = time.time()
    print("=" * 60)
    print("  CRM VITAO360 V3 — Build")
    print("=" * 60)

    # Ensure output dir exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Create workbook (remove default sheet later)
    wb = openpyxl.Workbook()

    # Build all tabs in order
    print("\n[1/9] REGRAS...")
    build_regras(wb)

    print("[2/9] DRAFT 1...")
    build_draft1(wb)

    print("[3/9] PROJEÇÃO...")
    build_projecao(wb)

    print("[4/9] CARTEIRA...")
    build_carteira(wb)

    print("[5/9] DRAFT 2...")
    build_draft2(wb)

    print("[6/9] AGENDA...")
    build_agenda(wb)

    print("[7/9] DASH...")
    build_dash(wb)

    print("[8/9] LOG...")
    build_log(wb)

    print("[9/9] Claude Log...")
    build_claude_log(wb)

    # Remove default "Sheet" if it exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Set CARTEIRA as active sheet (main working tab)
    for i, name in enumerate(wb.sheetnames):
        if name == "CARTEIRA":
            wb.active = i
            break

    # Save
    wb.save(OUTPUT_FILE)
    elapsed = time.time() - start

    print("\n" + "=" * 60)
    print(f"  SAVED: {OUTPUT_FILE}")
    print(f"  Tabs: {wb.sheetnames}")
    print(f"  Time: {elapsed:.1f}s")
    print("=" * 60)


if __name__ == "__main__":
    main()

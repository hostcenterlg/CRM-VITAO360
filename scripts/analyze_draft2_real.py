"""Quick analysis of DRAFT2 with real attendance data."""
import openpyxl
import os

PYTHON_PATH = r"c:\Users\User\.pyenv\pyenv-win\pyenv-win\versions\3.12.10\python.exe"
FILE = r"c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\sources\drafts\DRAFT2_POPULADO_DADOS_REAIS_v3.xlsx"

print(f"File: {os.path.basename(FILE)} ({os.path.getsize(FILE)/1024:.0f} KB)")

wb = openpyxl.load_workbook(FILE, data_only=False, read_only=True)
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")

    formula_count = 0
    data_count = 0
    row_count = 0

    for row in ws.iter_rows():
        row_count += 1
        if row_count > 7000:
            break
        for cell in row:
            if cell.value is None:
                pass
            elif isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
            else:
                data_count += 1

    print(f"  Rows: {row_count} | Formulas: {formula_count} | Data: {data_count}")

    # Print headers
    for row in ws.iter_rows(max_row=3):
        vals = [str(c.value)[:25] if c.value else "" for c in row[:32]]
        print(f"  {vals}")

wb.close()
print("\nDONE.")

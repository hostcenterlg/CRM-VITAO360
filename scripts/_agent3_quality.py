"""AGENT 3: Data Quality Audit — CRM VITAO360 V3 MERGED"""
import openpyxl
import os
import re
from collections import Counter

FILE = r"c:\Users\User\OneDrive\Área de Trabalho\CLAUDE CODE\output\CRM_VITAO360_V3_MERGED.xlsx"
wb = openpyxl.load_workbook(FILE, data_only=True, read_only=True)

print("=" * 70)
print("AGENT 3: DATA QUALITY AUDIT")
print("=" * 70)

# ── 1. CNPJ VALIDATION ──
print("\n[1] CNPJ VALIDATION")
ws = wb["DRAFT 1"]
cnpjs = []
invalid_cnpjs = []
for row in ws.iter_rows(min_row=4, max_col=2, values_only=True):
    cnpj = str(row[1] or "").strip()
    if not cnpj: continue
    cnpjs.append(cnpj)
    # Check format: should be XX.XXX.XXX/XXXX-XX or 14 digits
    digits = re.sub(r'\D', '', cnpj)
    if len(digits) != 14:
        invalid_cnpjs.append((cnpj, f"has {len(digits)} digits"))
    elif digits == '0' * 14:
        invalid_cnpjs.append((cnpj, "all zeros"))

print(f"  Total CNPJs: {len(cnpjs)}")
print(f"  Unique: {len(set(cnpjs))}")
dups = [c for c, n in Counter(cnpjs).items() if n > 1]
if dups:
    print(f"  DUPLICATES: {len(dups)}")
    for d in dups[:5]:
        print(f"    {d}")
else:
    print(f"  No duplicates")
if invalid_cnpjs:
    print(f"  INVALID FORMAT: {len(invalid_cnpjs)}")
    for cnpj, reason in invalid_cnpjs[:10]:
        print(f"    {cnpj} — {reason}")
else:
    print(f"  All valid format (14 digits)")

# ── 2. SITUACAO VALIDATION ──
print("\n[2] SITUACAO VALIDATION")
valid_sit = {"ATIVO", "EM RISCO", "INAT.REC", "INAT.ANT", "NOVO", "PROSPECT"}
sit_counter = Counter()
invalid_sit = []
for row in ws.iter_rows(min_row=4, max_col=15, values_only=True):
    if not row[1]: continue
    sit = str(row[11] or "").strip()
    if sit:
        sit_counter[sit] += 1
        if sit not in valid_sit:
            invalid_sit.append(sit)
    else:
        invalid_sit.append("(EMPTY)")

print(f"  Distribution: {dict(sit_counter.most_common())}")
if invalid_sit:
    print(f"  INVALID VALUES: {len(invalid_sit)}")
    for s in set(invalid_sit):
        print(f"    '{s}' x{invalid_sit.count(s)}")
else:
    print(f"  All valid")

# ── 3. CONSULTOR TERRITORY VALIDATION ──
print("\n[3] CONSULTOR TERRITORY RULES")
territory_errors = []
for row in ws.iter_rows(min_row=4, max_col=11, values_only=True):
    if not row[1]: continue
    nome = str(row[0] or "")
    cnpj = str(row[1] or "")
    uf = str(row[3] or "").upper()
    rede = str(row[8] or "").upper()
    consultor = str(row[9] or "").upper()

    expected = None
    if "CIA DA SAUDE" in rede or "CIA DA SAÚDE" in rede or "FITLAND" in rede:
        expected = "JULIO GADRET"
    elif any(r in rede for r in ["DIVINA TERRA", "BIOMUNDO", "BIO MUNDO", "MUNDO VERDE",
                                   "TUDO EM GRAOS", "TUDO EM GRÃOS", "VGA", "VIDA LEVE",
                                   "ARMAZEM", "ARMAZÉM", "NATURVIDA", "LIGEIRINHO",
                                   "TRIP", "ESMERALDA", "MERCOCENTRO"]):
        expected = "DAIANE STAVICKI"
    elif uf in ["SC", "PR", "RS"]:
        expected = "MANU DITZEL"
    else:
        expected = "LARISSA PADILHA"

    if expected and expected not in consultor:
        territory_errors.append({
            'nome': nome[:25],
            'cnpj': cnpj,
            'uf': uf,
            'rede': rede[:20],
            'consultor': consultor,
            'expected': expected,
        })

print(f"  Territory mismatches: {len(territory_errors)}")
if territory_errors:
    for e in territory_errors[:15]:
        print(f"    {e['nome']:25s} | UF={e['uf']:2s} | Rede={e['rede']:20s} | Got={e['consultor']:20s} | Exp={e['expected']}")
    if len(territory_errors) > 15:
        print(f"    ... and {len(territory_errors)-15} more")

# ── 4. DRAFT 2 DATA QUALITY ──
print("\n[4] DRAFT 2 DATA QUALITY")
ws2 = wb["DRAFT 2"]
d2_issues = Counter()
d2_total = 0
d2_no_resultado = 0
d2_no_date = 0
d2_future_dates = 0
d2_resultado_dist = Counter()

for row in ws2.iter_rows(min_row=3, max_col=24, values_only=True):
    cnpj = str(row[1] or "").strip()
    if not cnpj: continue
    d2_total += 1

    # Check date
    if not row[0]:
        d2_no_date += 1
    elif hasattr(row[0], 'year') and row[0].year > 2026:
        d2_future_dates += 1

    # Check resultado
    res = str(row[5] or "").strip()
    if not res:
        d2_no_resultado += 1
    else:
        d2_resultado_dist[res] += 1

    # Check situacao auto
    sit = str(row[12] or "").strip()
    if not sit:
        d2_issues['no_situacao'] += 1

print(f"  Total records: {d2_total}")
print(f"  Missing DATA: {d2_no_date}")
print(f"  Missing RESULTADO: {d2_no_resultado}")
print(f"  Future dates (>2026): {d2_future_dates}")
print(f"  Missing SITUACAO (auto): {d2_issues.get('no_situacao', 0)}")

# Check for unexpected RESULTADO values
valid_res = {"EM ATENDIMENTO", "ORÇAMENTO", "CADASTRO", "VENDA / PEDIDO",
             "RELACIONAMENTO", "FOLLOW UP 7", "FOLLOW UP 15", "SUPORTE",
             "NÃO ATENDE", "NÃO RESPONDE", "RECUSOU LIGAÇÃO", "PERDA / FECHOU LOJA"}
unexpected_res = {k: v for k, v in d2_resultado_dist.items() if k not in valid_res}
if unexpected_res:
    print(f"  UNEXPECTED RESULTADO values:")
    for k, v in unexpected_res.items():
        print(f"    '{k}': {v}")
else:
    print(f"  All RESULTADO values valid")

# ── 5. CARTEIRA CROSS-CHECK ──
print("\n[5] CARTEIRA vs DRAFT 1 CROSS-CHECK")
ws3 = wb["CARTEIRA"]
cart_cnpjs = set()
for row in ws3.iter_rows(min_row=4, max_col=2, values_only=True):
    c = str(row[1] or "").strip()
    if c: cart_cnpjs.add(c)

d1_cnpjs = set(cnpjs)
only_d1 = d1_cnpjs - cart_cnpjs
only_cart = cart_cnpjs - d1_cnpjs
print(f"  DRAFT 1 CNPJs: {len(d1_cnpjs)}")
print(f"  CARTEIRA CNPJs: {len(cart_cnpjs)}")
print(f"  In DRAFT 1 only: {len(only_d1)}")
print(f"  In CARTEIRA only: {len(only_cart)}")
if only_d1:
    print(f"  MISSING from CARTEIRA:")
    for c in list(only_d1)[:5]:
        print(f"    {c}")

# ── 6. VENDAS CONSISTENCY ──
print("\n[6] VENDAS CONSISTENCY")
vendas_issues = []
for row in ws.iter_rows(min_row=4, max_col=44, values_only=True):
    if not row[1]: continue
    cnpj = str(row[1])
    sit = str(row[11] or "")
    total = row[17]  # TOTAL PERIODO

    # Check: ATIVO should have recent sales
    has_recent = False
    for j in range(30, 36):  # NOV, DEZ, JAN (most recent)
        if j < len(row) and row[j] and isinstance(row[j], (int, float)) and row[j] > 0:
            has_recent = True

    if sit == "ATIVO" and not has_recent:
        vendas_issues.append(f"ATIVO sem vendas recentes: {cnpj}")

    # Check: PROSPECT should NOT have sales
    has_any = False
    for j in range(24, 36):
        if j < len(row) and row[j] and isinstance(row[j], (int, float)) and row[j] > 0:
            has_any = True
    if sit == "PROSPECT" and has_any:
        vendas_issues.append(f"PROSPECT com vendas: {cnpj}")

print(f"  Vendas inconsistencies: {len(vendas_issues)}")
for v in vendas_issues[:10]:
    print(f"    {v}")

# ── 7. EMPTY CRITICAL COLUMNS ──
print("\n[7] EMPTY CRITICAL COLUMNS CHECK")
critical_cols = {
    'REDE/REGIONAL (col 9)': 8,
    'ACESSO B2B (col 20)': 19,
    'MESES LISTA (col 43)': 42,
}
for label, idx in critical_cols.items():
    cnt = 0
    total = 0
    for row in ws.iter_rows(min_row=4, max_col=48, values_only=True):
        if not row[1]: continue
        total += 1
        if idx < len(row) and row[idx] is not None and str(row[idx]).strip():
            cnt += 1
    pct = cnt/total*100 if total > 0 else 0
    status = "OK" if pct >= 80 else "LOW" if pct >= 30 else "CRITICAL"
    print(f"  {label}: {cnt}/{total} ({pct:.1f}%) [{status}]")

wb.close()
print("\n" + "=" * 70)
print("AGENT 3: DATA QUALITY AUDIT COMPLETE")
print("=" * 70)

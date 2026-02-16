"""Check TIPO DO CONTATO values in REGRAS"""
import openpyxl, os
path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output", "CRM INTELIGENTE - VITAO360 V2.xlsx")
wb = openpyxl.load_workbook(path, data_only=False)
ws = wb['REGRAS']
print("--- TIPO DO CONTATO (rows 17-25) ---")
for r in range(17, 26):
    vals = []
    for c in range(1, 4):
        v = ws.cell(row=r, column=c).value
        vals.append(str(v) if v else "")
    print(f"  Row {r}: {vals}")

print("\n--- SITUACAO (rows 40-48) ---")
for r in range(40, 49):
    vals = []
    for c in range(1, 4):
        v = ws.cell(row=r, column=c).value
        vals.append(str(v) if v else "")
    print(f"  Row {r}: {vals}")

print("\n--- GRUPO DASH (rows 110-114) ---")
for r in range(110, 115):
    vals = []
    for c in range(1, 4):
        v = ws.cell(row=r, column=c).value
        vals.append(str(v) if v else "")
    print(f"  Row {r}: {vals}")

# Also check a few data rows in DRAFT 2 for Q column values
ws_d = wb['DRAFT 2']
print("\n--- DRAFT 2 sample TIPO DO CONTATO (col Q) ---")
tipos_found = set()
for r in range(3, min(ws_d.max_row + 1, 50)):
    v = ws_d.cell(row=r, column=17).value  # Q
    if v and not str(v).startswith('='):
        tipos_found.add(v)
print(f"  Unique tipos: {sorted(tipos_found)}")

resultados_found = set()
for r in range(3, min(ws_d.max_row + 1, 50)):
    v = ws_d.cell(row=r, column=18).value  # R
    if v and not str(v).startswith('='):
        resultados_found.add(v)
print(f"  Unique resultados: {sorted(resultados_found)}")

wb.close()

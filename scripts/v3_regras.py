"""
V3 — Tab REGRAS: Tabelas de referencia + Named Ranges
"""
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from v3_styles import *


def build_regras(wb):
    """Build REGRAS tab with all reference tables and named ranges."""
    ws = wb.create_sheet("REGRAS", 0)
    ws.sheet_properties.tabColor = TAB_REGRAS

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 38

    row = 1

    # ═══════════════════════════════════════════════════════════
    # SEÇÃO 1: RESULTADOS (12 items) — rows 1-14
    # ═══════════════════════════════════════════════════════════
    headers = ['#', 'RESULTADO', 'FOLLOW-UP (DIAS)', 'GRUPO DASH', 'QUANDO USAR']
    for c, h in enumerate(headers, 1):
        write_header(ws, row, c, h)

    resultados = [
        (1, "EM ATENDIMENTO", 2, "FUNIL", "Negociação ativa"),
        (2, "ORÇAMENTO", 1, "FUNIL", "Proposta enviada"),
        (3, "CADASTRO", 2, "FUNIL", "Cliente novo em cadastro"),
        (4, "VENDA / PEDIDO", 45, "FUNIL", "Pedido fechado"),
        (5, "RELACIONAMENTO", 7, "RELAC.", "Contato pós-venda/CS"),
        (6, "FOLLOW UP 7", 7, "RELAC.", "Retomar em 1 semana"),
        (7, "FOLLOW UP 15", 15, "RELAC.", "Retomar em 2 semanas"),
        (8, "SUPORTE", 0, "RELAC.", "Problema resolvido"),
        (9, "NÃO ATENDE", 1, "NÃO VENDA", "Ligou sem resposta"),
        (10, "NÃO RESPONDE", 1, "NÃO VENDA", "WhatsApp sem resposta"),
        (11, "RECUSOU LIGAÇÃO", 2, "NÃO VENDA", "Rejeitou chamada"),
        (12, "PERDA / FECHOU LOJA", 0, "NÃO VENDA", "Cliente perdido"),
    ]
    for i, (num, res, fu, grupo, desc) in enumerate(resultados):
        r = row + 1 + i
        write_data(ws, r, 1, num)
        write_data(ws, r, 2, res, align=ALIGN_LEFT)
        write_data(ws, r, 3, fu)
        write_data(ws, r, 4, grupo)
        write_data(ws, r, 5, desc, align=ALIGN_LEFT)

    res_start = row + 1
    res_end = row + len(resultados)
    row = res_end + 2  # blank row

    # ═══════════════════════════════════════════════════════════
    # SEÇÃO 2: ESTÁGIO DO FUNIL (8 items) — NEW V3
    # ═══════════════════════════════════════════════════════════
    estagio_hdr_row = row
    write_header(ws, row, 1, '#')
    write_header(ws, row, 2, 'ESTÁGIO DO FUNIL')
    write_header(ws, row, 3, 'DESCRIÇÃO')

    estagios = [
        (1, "PROSPECÇÃO", "Primeiro contato, ainda não é cliente"),
        (2, "EM ATENDIMENTO", "Respondeu, negociação ativa"),
        (3, "NEGOCIAÇÃO", "Interesse confirmado, discutindo termos"),
        (4, "ORÇAMENTO", "Proposta enviada, aguardando resposta"),
        (5, "PÓS-VENDA", "Comprou, acompanhamento pós-compra"),
        (6, "CS / RECOMPRA", "Sucesso do cliente, tentativa recompra"),
        (7, "RELACIONAMENTO", "Manutenção do vínculo"),
        (8, "PERDA / NUTRIÇÃO", "Perdido ou em nutrição longo prazo"),
    ]
    est_start = row + 1
    for i, (num, est, desc) in enumerate(estagios):
        r = row + 1 + i
        write_data(ws, r, 1, num)
        write_data(ws, r, 2, est, align=ALIGN_LEFT)
        write_data(ws, r, 3, desc, align=ALIGN_LEFT)
    est_end = row + len(estagios)
    row = est_end + 2

    # ═══════════════════════════════════════════════════════════
    # SEÇÃO 3: TIPO DO CONTATO (7 items)
    # ═══════════════════════════════════════════════════════════
    write_header(ws, row, 1, '#')
    write_header(ws, row, 2, 'TIPO DO CONTATO')
    write_header(ws, row, 3, 'QUANDO USAR')

    tipos = [
        (1, "PROSPECÇÃO", "Primeiro contato com prospect/lead que nunca comprou"),
        (2, "NEGOCIAÇÃO", "Cliente em fase de proposta/orçamento"),
        (3, "FOLLOW UP", "Retomada de contato agendado (7d ou 15d)"),
        (4, "ATEND. CLIENTES ATIVOS", "Manutenção de clientes ATIVOS"),
        (5, "ATEND. CLIENTES INATIVOS", "Reativação de INAT.REC ou INAT.ANT"),
        (6, "PÓS-VENDA / RELACIONAMENTO", "Relacionamento após venda (CS, suporte)"),
        (7, "PERDA / NUTRIÇÃO", "Cliente perdido ou em nutrição"),
    ]
    tipo_start = row + 1
    for i, (num, tipo, desc) in enumerate(tipos):
        r = row + 1 + i
        write_data(ws, r, 1, num)
        write_data(ws, r, 2, tipo, align=ALIGN_LEFT)
        write_data(ws, r, 3, desc, align=ALIGN_LEFT)
    tipo_end = row + len(tipos)
    row = tipo_end + 2

    # ═══════════════════════════════════════════════════════════
    # SEÇÃO 4: MOTIVOS (10 items)
    # ═══════════════════════════════════════════════════════════
    write_header(ws, row, 1, '#')
    write_header(ws, row, 2, 'MOTIVO')
    write_header(ws, row, 3, 'DONO DA AÇÃO')

    motivos = [
        (1, "AINDA TEM ESTOQUE", "Normal"),
        (2, "PRODUTO NÃO VENDEU / SEM GIRO", "DIRETORIA / FÁBRICA"),
        (3, "LOJA ANEXO/PROXIMO - SM", "COMERCIAL"),
        (4, "SÓ QUER COMPRAR GRANEL", "COMERCIAL / MKT"),
        (5, "PROBLEMA LOGÍSTICO / ENTREGA", "LOGÍSTICA"),
        (6, "PROBLEMA FINANCEIRO / CRÉDITO", "FINANCEIRO"),
        (7, "PROPRIETARIO / INDISPONÍVEL", "Normal"),
        (8, "FECHANDO / FECHOU LOJA", "PERDA"),
        (9, "SEM INTERESSE NO MOMENTO", "NUTRIÇÃO"),
        (10, "PRIMEIRO CONTATO / SEM RESPOSTA", "Normal"),
    ]
    mot_start = row + 1
    for i, (num, mot, dono) in enumerate(motivos):
        r = row + 1 + i
        write_data(ws, r, 1, num)
        write_data(ws, r, 2, mot, align=ALIGN_LEFT)
        write_data(ws, r, 3, dono, align=ALIGN_LEFT)
    mot_end = row + len(motivos)
    row = mot_end + 2

    # ═══════════════════════════════════════════════════════════
    # SEÇÃO 5: SITUAÇÃO (6 items — sem LEAD)
    # ═══════════════════════════════════════════════════════════
    write_header(ws, row, 1, 'DIAS')
    write_header(ws, row, 2, 'SITUAÇÃO')
    write_header(ws, row, 3, 'COR HEX')

    situacoes = [
        ("≤ 50", "ATIVO", "#00B050"),
        ("51-60", "EM RISCO", "#FFC000"),
        ("61-90", "INAT.REC", "#FFC000"),
        ("> 90", "INAT.ANT", "#FF0000"),
        ("1ª compra", "NOVO", "#0070C0"),
        ("Nunca comprou", "PROSPECT", "#7B2FF2"),
    ]
    sit_start = row + 1
    for i, (dias, sit, cor) in enumerate(situacoes):
        r = row + 1 + i
        write_data(ws, r, 1, dias)
        write_data(ws, r, 2, sit, align=ALIGN_LEFT)
        write_data(ws, r, 3, cor)
        # Apply color to SITUAÇÃO cell
        if sit in SITUACAO_FILLS:
            ws.cell(row=r, column=2).fill = SITUACAO_FILLS[sit]
            ws.cell(row=r, column=2).font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    sit_end = row + len(situacoes)
    row = sit_end + 2

    # ═══════════════════════════════════════════════════════════
    # SEÇÃO 6: FASE (9 items)
    # ═══════════════════════════════════════════════════════════
    write_header(ws, row, 1, '#')
    write_header(ws, row, 2, 'FASE')
    write_header(ws, row, 3, 'DESCRIÇÃO')

    fases = [
        (1, "PÓS-VENDA"), (2, "CS"), (3, "RECOMPRA"), (4, "SALVAMENTO"),
        (5, "RECUPERAÇÃO"), (6, "PROSPECÇÃO"), (7, "NUTRIÇÃO"),
        (8, "EM ATENDIMENTO"), (9, "ORÇAMENTO"),
    ]
    fase_start = row + 1
    for i, (num, fase) in enumerate(fases):
        r = row + 1 + i
        write_data(ws, r, 1, num)
        write_data(ws, r, 2, fase, align=ALIGN_LEFT)
    fase_end = row + len(fases)
    row = fase_end + 2

    # ═══════════════════════════════════════════════════════════
    # SEÇÃO 7: TENTATIVAS (6 items)
    # ═══════════════════════════════════════════════════════════
    write_header(ws, row, 1, '#')
    write_header(ws, row, 2, 'TENTATIVA')
    write_header(ws, row, 3, 'CANAL')
    write_header(ws, row, 4, 'INTERVALO')

    tentativas = [
        (1, "T1", "WhatsApp", "Dia 0"),
        (2, "T2", "Ligação", "+1 dia"),
        (3, "T3", "WhatsApp", "+2 dias"),
        (4, "T4", "Ligação final", "+2 dias"),
        (5, "NUTRIÇÃO", "Email+WhatsApp", "Ciclo 15d"),
        (6, "RESET", "Qualquer", "Quando responde"),
    ]
    tent_start = row + 1
    for i, (num, tent, canal, interv) in enumerate(tentativas):
        r = row + 1 + i
        write_data(ws, r, 1, num)
        write_data(ws, r, 2, tent, align=ALIGN_LEFT)
        write_data(ws, r, 3, canal, align=ALIGN_LEFT)
        write_data(ws, r, 4, interv, align=ALIGN_LEFT)
    tent_end = row + len(tentativas)
    row = tent_end + 2

    # ═══════════════════════════════════════════════════════════
    # SEÇÃO 8: SINALEIRO CICLO (4 items)
    # ═══════════════════════════════════════════════════════════
    write_header(ws, row, 1, 'SINALEIRO')
    write_header(ws, row, 2, 'CONDIÇÃO')

    sinaleiros = [
        ("🟢", "Dias ≤ Ciclo Médio"),
        ("🟡", "Dias ≤ Ciclo+30"),
        ("🔴", "Dias > Ciclo+30"),
        ("🟣", "Nunca comprou"),
    ]
    sin_start = row + 1
    for i, (emoji, cond) in enumerate(sinaleiros):
        r = row + 1 + i
        write_data(ws, r, 1, emoji)
        write_data(ws, r, 2, cond, align=ALIGN_LEFT)
    sin_end = row + len(sinaleiros)
    row = sin_end + 2

    # ═══════════════════════════════════════════════════════════
    # SEÇÃO 9: TIPO CLIENTE (5 items — sem LEAD)
    # ═══════════════════════════════════════════════════════════
    write_header(ws, row, 1, '#')
    write_header(ws, row, 2, 'TIPO CLIENTE')
    write_header(ws, row, 3, 'CRITÉRIO')

    tipo_clientes = [
        (1, "PROSPECT", "Nunca comprou"),
        (2, "NOVO", "1ª compra / 1 mês"),
        (3, "EM DESENVOLVIMENTO", "2-3 meses positivado"),
        (4, "RECORRENTE", "4-6 meses positivado"),
        (5, "FIDELIZADO", "7+ meses positivado"),
    ]
    tc_start = row + 1
    for i, (num, tc, crit) in enumerate(tipo_clientes):
        r = row + 1 + i
        write_data(ws, r, 1, num)
        write_data(ws, r, 2, tc, align=ALIGN_LEFT)
        write_data(ws, r, 3, crit, align=ALIGN_LEFT)
    tc_end = row + len(tipo_clientes)
    row = tc_end + 2

    # ═══════════════════════════════════════════════════════════
    # SEÇÃO 10: CONSULTOR (4)
    # ═══════════════════════════════════════════════════════════
    write_header(ws, row, 1, 'CONSULTOR')
    write_header(ws, row, 2, 'TERRITÓRIO')

    consultores = [
        ("MANU DITZEL", "SC/PR/RS"),
        ("LARISSA PADILHA", "Resto do Brasil"),
        ("JULIO GADRET", "CIA DA SAUDE / FITLAND"),
        ("DAIANE STAVICKI", "Redes de Franquia"),
    ]
    cons_start = row + 1
    for i, (cons, terr) in enumerate(consultores):
        r = row + 1 + i
        write_data(ws, r, 1, cons, align=ALIGN_LEFT)
        write_data(ws, r, 2, terr, align=ALIGN_LEFT)
    cons_end = row + len(consultores)
    row = cons_end + 2

    # ═══════════════════════════════════════════════════════════
    # SEÇÃO 11: LISTAS AUXILIARES
    # ═══════════════════════════════════════════════════════════
    # Freeze here — reference lists below
    ws.freeze_panes = f"A{row}"

    write_header(ws, row, 1, 'LISTA')
    write_header(ws, row, 2, 'VALORES')
    row += 1

    # SIM_NAO
    sim_nao_start = row
    write_data(ws, row, 1, "SIM_NAO")
    write_data(ws, row, 2, "SIM", align=ALIGN_LEFT)
    row += 1
    write_data(ws, row, 2, "NÃO", align=ALIGN_LEFT)
    sim_nao_end = row
    row += 1

    # SIM_NAO_NA
    sim_nao_na_start = row
    write_data(ws, row, 1, "SIM_NAO_NA")
    write_data(ws, row, 2, "SIM", align=ALIGN_LEFT)
    row += 1
    write_data(ws, row, 2, "NÃO", align=ALIGN_LEFT)
    row += 1
    write_data(ws, row, 2, "N/A", align=ALIGN_LEFT)
    sim_nao_na_end = row
    row += 1

    # ATIVO_RECEPTIVO
    ar_start = row
    write_data(ws, row, 1, "ATIVO_RECEPTIVO")
    write_data(ws, row, 2, "ATIVO", align=ALIGN_LEFT)
    row += 1
    write_data(ws, row, 2, "RECEPTIVO", align=ALIGN_LEFT)
    ar_end = row
    row += 1

    # CURVA_ABC
    curva_start = row
    write_data(ws, row, 1, "CURVA_ABC")
    write_data(ws, row, 2, "A", align=ALIGN_LEFT)
    row += 1
    write_data(ws, row, 2, "B", align=ALIGN_LEFT)
    row += 1
    write_data(ws, row, 2, "C", align=ALIGN_LEFT)
    curva_end = row
    row += 1

    # GRUPO_DASH
    gd_start = row
    write_data(ws, row, 1, "GRUPO_DASH")
    write_data(ws, row, 2, "FUNIL", align=ALIGN_LEFT)
    row += 1
    write_data(ws, row, 2, "RELAC.", align=ALIGN_LEFT)
    row += 1
    write_data(ws, row, 2, "NÃO VENDA", align=ALIGN_LEFT)
    gd_end = row
    row += 1

    # PRIORIDADE
    prio_start = row
    write_data(ws, row, 1, "PRIORIDADE")
    prio_vals = ["ALTA", "MÉDIA", "BAIXA"]
    for i, p in enumerate(prio_vals):
        write_data(ws, row + i, 2, p, align=ALIGN_LEFT)
    prio_end = row + len(prio_vals) - 1
    row = prio_end + 2

    # ACAO_FUTURA
    af_start = row
    write_data(ws, row, 1, "ACAO_FUTURA")
    acoes = ["RECOMPRA", "PÓS-VENDA", "CS", "REATIVAÇÃO", "PROSPECÇÃO", "NUTRIÇÃO", "SALVAMENTO", "ATENDIMENTO"]
    for i, acao in enumerate(acoes):
        write_data(ws, row + i, 2, acao, align=ALIGN_LEFT)
    af_end = row + len(acoes) - 1
    row = af_end + 2

    # TEMPERATURA (NEW V3)
    temp_start = row
    write_data(ws, row, 1, "TEMPERATURA")
    temps = ["🔥 QUENTE", "🟡 MORNO", "❄️ FRIO", "💀 PERDIDO"]
    for i, t in enumerate(temps):
        write_data(ws, row + i, 2, t, align=ALIGN_LEFT)
    temp_end = row + len(temps) - 1
    row = temp_end + 2

    # OPORTUNIDADE (NEW V3)
    opp_start = row
    write_data(ws, row, 1, "OPORTUNIDADE")
    opps = ["🔥 QUENTE", "🟡 MORNO", "❄️ FRIO"]
    for i, o in enumerate(opps):
        write_data(ws, row + i, 2, o, align=ALIGN_LEFT)
    opp_end = row + len(opps) - 1

    # ═══════════════════════════════════════════════════════════
    # NAMED RANGES
    # ═══════════════════════════════════════════════════════════
    named_ranges = {
        'LISTA_RESULTADO': f"REGRAS!$B${res_start}:$B${res_end}",
        'LISTA_ESTAGIO_FUNIL': f"REGRAS!$B${est_start}:$B${est_end}",
        'LISTA_TIPO_CONTATO': f"REGRAS!$B${tipo_start}:$B${tipo_end}",
        'LISTA_MOTIVO': f"REGRAS!$B${mot_start}:$B${mot_end}",
        'LISTA_SITUACAO': f"REGRAS!$B${sit_start}:$B${sit_end}",
        'LISTA_FASE': f"REGRAS!$B${fase_start}:$B${fase_end}",
        'LISTA_TENTATIVA': f"REGRAS!$B${tent_start}:$B${tent_end}",
        'LISTA_SINALEIRO': f"REGRAS!$A${sin_start}:$A${sin_end}",
        'LISTA_TIPO_CLIENTE': f"REGRAS!$B${tc_start}:$B${tc_end}",
        'LISTA_CONSULTOR': f"REGRAS!$A${cons_start}:$A${cons_end}",
        'LISTA_SIM_NAO': f"REGRAS!$B${sim_nao_start}:$B${sim_nao_end}",
        'LISTA_SIM_NAO_NA': f"REGRAS!$B${sim_nao_na_start}:$B${sim_nao_na_end}",
        'LISTA_PRIORIDADE': f"REGRAS!$B${prio_start}:$B${prio_end}",
        'LISTA_ATIVO_RECEPTIVO': f"REGRAS!$B${ar_start}:$B${ar_end}",
        'LISTA_CURVA_ABC': f"REGRAS!$B${curva_start}:$B${curva_end}",
        'LISTA_CURVA': f"REGRAS!$B${curva_start}:$B${curva_end}",
        'LISTA_GRUPO_DASH': f"REGRAS!$B${gd_start}:$B${gd_end}",
        'LISTA_ACAO_FUTURA': f"REGRAS!$B${af_start}:$B${af_end}",
        'LISTA_TEMPERATURA': f"REGRAS!$B${temp_start}:$B${temp_end}",
        'LISTA_OPORTUNIDADE': f"REGRAS!$B${opp_start}:$B${opp_end}",
    }

    for name, ref in named_ranges.items():
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names.add(dn)

    print(f"  REGRAS: {row} rows, {len(named_ranges)} named ranges")
    return ws

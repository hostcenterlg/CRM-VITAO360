#!/usr/bin/env python3
'''
extract_meta_sap.py - Extracts monthly META 2026 data from SAP file.
'''

import openpyxl
from collections import defaultdict, OrderedDict

Q = chr(39)  # single quote character

SAP_PATH = 'C:\\Users\\User\\OneDrive\\Área de Trabalho\\PASTA G (CENTRAL INTERNO)\\CARTEIRA DE CLIENTES SAP\\BASE SAP - META E PROJEÇÃO 2026 .- 02. INTERNO - 2026.xlsx'

MONTH_NAMES = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
SEP = '===================================================================================================='

print(f'\n{SEP}')
print('OPENING SAP FILE')
print(SEP)
wb = openpyxl.load_workbook(SAP_PATH, data_only=True)
print(f'Sheets: {wb.sheetnames}')

ws_fat = wb['Faturamento']
print(f'\n{SEP}')
print(f'SHEET: Faturamento | rows={ws_fat.max_row} cols={ws_fat.max_column}')
print(SEP)

print('\n--- TASK 1: ALL HEADERS ---')
for r in range(1, 4):
    vals = [ws_fat.cell(r, c).value for c in range(1, ws_fat.max_column + 1)]
    print(f'  Row {r}: {vals}')

print('\n--- TASK 2: FIRST 20 DATA ROWS ---')
for r in range(4, min(24, ws_fat.max_row + 1)):
    vals = [ws_fat.cell(r, c).value for c in range(1, min(ws_fat.max_column + 1, 30))]
    print(f'  Row {r}: {vals}')

print('\n--- TASK 3: COLUMN MAPPING ---')
col_map = {}
for c in range(1, ws_fat.max_column + 1):
    h = ws_fat.cell(3, c).value
    if h:
        col_map[c] = str(h).strip()
print('  Column mapping (row 3):')
for c, h in sorted(col_map.items()):
    print(f'    Col {c} ({openpyxl.utils.get_column_letter(c)}): {h}')

COL_TOTAIS = 1
COL_COD_GC = 10
COL_NOME_GC = 11
COL_GPROD = 12
COL_JAN = 14
COL_DEZ = 25
COL_TOTAL = 26
print('  TOTAIS=A, COD_GC=J, NOME_GC=K, GPROD=L, JAN=N, DEZ=Y, TOTAL=Z')
# Task 4: AGGREGATE
print('\n--- TASK 4: AGGREGATION ---')
agg = defaultdict(lambda: {'months': [0.0]*12, 'total': 0.0, 'cod': None})
row_count = 0

for r in range(4, ws_fat.max_row + 1):
    totais = ws_fat.cell(r, COL_TOTAIS).value
    gprod = ws_fat.cell(r, COL_GPROD).value
    if totais == 'SIM' and gprod == '01. TOTAL':
        row_count += 1
        nome = ws_fat.cell(r, COL_NOME_GC).value or 'UNKNOWN'
        cod = ws_fat.cell(r, COL_COD_GC).value
        for i in range(12):
            val = ws_fat.cell(r, COL_JAN + i).value or 0
            agg[nome]['months'][i] += val
        total_val = ws_fat.cell(r, COL_TOTAL).value or 0
        agg[nome]['total'] += total_val
        if agg[nome]['cod'] is None:
            agg[nome]['cod'] = cod

print(f'  Processed {row_count} rows')
print(f'  Unique GRUPO CHAVE: {len(agg)}')

# Task 5: Aggregated table
print('\n--- TASK 5: AGGREGATED META TABLE ---')
hdr = f"{'GRUPO CHAVE':<42s}"
for m in MONTH_NAMES:
    hdr += f' {m:>10s}'
hdr += f" {'TOTAL':>12s}"
print(hdr)
print('-' * len(hdr))

grand_months = [0.0] * 12
grand_total = 0.0

for nome in sorted(agg.keys()):
    data = agg[nome]
    display = nome.replace('06 - INTERNA - ', '').replace('06 - ', '')
    line = f'{display:<42s}'
    for i in range(12):
        line += f" {data['months'][i]:>10,.0f}"
        grand_months[i] += data['months'][i]
    line += f" {data['total']:>12,.0f}"
    print(line)
    grand_total += data['total']

# Task 6: Grand total
print('\n--- TASK 6: GRAND TOTAL PER MONTH ---')
gt = f"{'GRAND TOTAL':<42s}"
for i in range(12):
    gt += f' {grand_months[i]:>10,.0f}'
gt += f' {grand_total:>12,.0f}'
print(gt)

# Task 7: Verify
print('\n--- TASK 7: VERIFICATION ---')
expected = 4747200
print(f'  Computed: R$ {grand_total:,.0f}')
print(f'  Expected: R$ {expected:,.0f}')
match_str = 'YES' if abs(grand_total - expected) < 1 else 'NO'
print(f'  Match: {match_str}')

row2_months = [ws_fat.cell(2, c).value or 0 for c in range(14, 26)]
row2_total = ws_fat.cell(2, 26).value or 0
print('\n  Month-by-month vs row 2:')
for i in range(12):
    ok = 'OK' if abs(grand_months[i] - row2_months[i]) < 1 else 'MISMATCH'
    print(f'    {MONTH_NAMES[i]}: {grand_months[i]:>10,.0f} vs {row2_months[i]:>10,.0f} [{ok}]')
ttl_ok = 'OK' if abs(grand_total - row2_total) < 1 else 'MISMATCH'
print(f'    TOTAL: {grand_total:>12,.0f} vs {row2_total:>12,.0f} [{ttl_ok}]')

# === RESUMO SHEET ===
print(f'\n{SEP}')
print('SHEET: Resumo')
print(SEP)
ws_res = wb['Resumo']
print(f'  Dimensions: rows={ws_res.max_row}, cols={ws_res.max_column}')

for r in range(1, ws_res.max_row + 1):
    vals = [ws_res.cell(r, c).value for c in range(1, ws_res.max_column + 1)]
    if any(v is not None for v in vals):
        fmt = []
        for v in vals:
            if v is None: fmt.append('')
            elif isinstance(v, float) and 0 < v < 1: fmt.append(f'{v:.2%}')
            elif isinstance(v, (int, float)) and abs(v) > 100: fmt.append(f'{v:,.0f}')
            else: fmt.append(str(v))
        print(f'  Row {r:>2d}: {fmt}')

print('\n  KEY DATA from Resumo:')
res_line = '    FAT META 2026: '
for c in range(3, 15):
    v = ws_res.cell(4, c).value
    if v: res_line += f'{v:>10,.0f}'
print(res_line)
total_resumo = ws_res.cell(4, 15).value
print(f'    TOTAL: R$ {total_resumo:,.0f}')


# === POSITIVACAO SHEET ===
print(f'\n{SEP}')
print('SHEET: Positivacao - CLASSIFICACAO')
print(SEP)

ws_pos = None
for sn in wb.sheetnames:
    if 'ositiv' in sn.lower():
        ws_pos = wb[sn]
        break

print(f'  Sheet: {ws_pos.title}, rows={ws_pos.max_row}, cols={ws_pos.max_column}')

print('\n  Headers (row 3):')
for c in range(1, ws_pos.max_column + 1):
    h = ws_pos.cell(3, c).value
    if h: print(f'    Col {c} ({openpyxl.utils.get_column_letter(c)}): {h}')

classificacao = OrderedDict()
print(f"  {'GRUPO CHAVE':<45s} {'CLASSIFICACAO':<20s} {'TICKET REAL':>16s} {'TICKET ESP':>16s} {'CARTEIRA':>10s}")
print(f"  {'-'*45} {'-'*20} {'-'*16} {'-'*16} {'-'*10}")

for r in range(4, ws_pos.max_row + 1):
    nome = ws_pos.cell(r, 10).value
    classif = ws_pos.cell(r, 28).value
    ticket_real = ws_pos.cell(r, 25).value
    ticket_esp = ws_pos.cell(r, 26).value
    carteira = ws_pos.cell(r, 30).value
    if nome:
        display = nome.replace('06 - INTERNA - ', '').replace('06 - ', '')
        tr = f'R$ {ticket_real:,.0f}' if ticket_real else 'R$ 0'
        te = f'R$ {ticket_esp:,.0f}' if ticket_esp else 'R$ 0'
        ct = str(carteira) if carteira else '0'
        if nome not in classificacao: classificacao[nome] = classif
        print(f"  {display:<45s} {classif or '':<20s} {tr:>16s} {te:>16s} {ct:>10s}")

print('\n  CLASSIFICACAO summary:')
class_count = defaultdict(int)
for nome, cl in classificacao.items(): class_count[cl or 'NONE'] += 1
for cl, cnt in sorted(class_count.items()): print(f'    {cl}: {cnt}')

# === BUILD IMPORTABLE DICT ===
print(f'\n{SEP}')
print('IMPORTABLE PYTHON DICT')
print(SEP)

meta_por_grupo = {}
for nome in sorted(agg.keys()):
    data = agg[nome]
    monthly = {}
    for i in range(12): monthly[i+1] = round(data['months'][i], 2)
    monthly['total'] = round(data['total'], 2)
    meta_por_grupo[nome] = monthly

classificacao_por_grupo = {}
for nome, cl in classificacao.items(): classificacao_por_grupo[nome] = cl

print('\nmeta_por_grupo = {')
for nome in sorted(meta_por_grupo.keys()):
    m = meta_por_grupo[nome]
    ms = ', '.join(f'{k}: {v}' for k, v in sorted((kk, vv) for kk, vv in m.items() if isinstance(kk, int)))
    print("    " + Q + nome + Q + ": {" + ms + ", " + Q + "total" + Q + ": " + str(m["total"]) + "},")
print('}')

print('\nclassificacao_por_grupo = {')
for nome in sorted(classificacao_por_grupo.keys()):
    cl = classificacao_por_grupo[nome]
    print("    " + Q + nome + Q + ": " + Q + str(cl) + Q + ",")
print('}')

meta_total_mensal = {}
for i in range(12): meta_total_mensal[i+1] = round(grand_months[i], 2)
meta_total_mensal['total'] = round(grand_total, 2)
print(f'\nmeta_total_mensal = {meta_total_mensal}')


# === FINAL SUMMARY ===
print(f'\n{SEP}')
print('FINAL SUMMARY')
print(SEP)
print(f'  Faturamento: {row_count} rows, {len(agg)} unique GRUPO CHAVE')
print(f'  Grand Total META 2026: R$ {grand_total:,.0f}')
tgt = 'MATCH' if abs(grand_total - 4747200) < 1 else 'NO MATCH'
print(f'  Target R$ 4,747,200: {tgt}')
named_count = len([k for k in agg if 'SEM GRUPO' not in k])
print(f'  Named groups (excl SEM GRUPO): {named_count}')
sem_data = agg.get('06 - SEM GRUPO', {'total': 0})
print(f"  SEM GRUPO (61 clients aggregated): R$ {sem_data['total']:,.0f}")
print(f'  Classificacoes: {dict(class_count)}')
print(f'  Resumo confirmed: R$ {total_resumo:,.0f}')

wb.close()
print('\nDone.')
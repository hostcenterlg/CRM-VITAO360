"""
populate_db_from_carteira.py
----------------------------
Popula o banco crm_vitao360.db com dados reais da aba CARTEIRA
da planilha INTELIGENTE FINAL OK.

REGRAS INVIOLÁVEIS:
- CNPJ: string 14 dígitos, zero-padded (NUNCA float/int)
- Two-Base Architecture: vendas têm valor R$, logs = R$ 0.00
- classificacao_3tier = 'REAL' para todos
- #NAME?, #REF! = None (data_only=True retorna None ou string)
- Commit a cada 100 registros

Mapeamento CARTEIRA → Model:
  Col 1  (A) = nome_fantasia
  Col 2  (B) = cnpj (chave primária)
  Col 3  (C) = razao_social
  Col 4  (D) = uf
  Col 5  (E) = cidade
  Col 6  (F) = email
  Col 7  (G) = telefone
  Col 11 (K) = consultor
  Col 13 (M) = situacao
  Col 15 (O) = dias_sem_compra
  Col 17 (Q) = valor_ultimo_pedido
  Col 18 (R) = ciclo_medio
  Col 28 (AB) = faturamento_total (TOTAL 2025)
  Col 56 (BD) = n_compras
  Col 57 (BE) = curva_abc
  Col 62 (BJ) = estagio_funil
  Col 66 (BN) = resultado
  Col 68 (BP) = tipo_cliente
  Col 72 (BT) = temperatura
  Col 80 (CB) = sinaleiro
  Col 81 (CC) = codigo_cliente (SAP)
  Col 93 (CQ) = macroregiao
  Col 96 (CR) = meta_anual
  Col 97 (CS) = realizado_acumulado
  Col 98 (CT) = pct_alcancado
  Col 143 (EM) = score
  Col 144 (EN) = prioridade v2 (-> derivada do score, #NAME? ignorado)

  Vendas 2025: cols 29-40 (jan-dez)
  Vendas 2026: cols 42-53 (jan-dez)
"""

import re
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
PLANILHA = r'C:\Users\User\OneDrive\Área de Trabalho\CRM_VITAO360  INTELIGENTE   FINAL OK .xlsx'
DB_PATH = r'C:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\crm_vitao360.db'

# ── Constantes ────────────────────────────────────────────────────────────────
BATCH_SIZE = 100

# Datas dos meses de vendas 2025 (cols 29-40, índice 0-based: 28-39)
MESES_2025 = [
    date(2025, 1, 1), date(2025, 2, 1), date(2025, 3, 1),
    date(2025, 4, 1), date(2025, 5, 1), date(2025, 6, 1),
    date(2025, 7, 1), date(2025, 8, 1), date(2025, 9, 1),
    date(2025, 10, 1), date(2025, 11, 1), date(2025, 12, 1),
]

# Datas dos meses de vendas 2026 (cols 42-53, índice 0-based: 41-52)
MESES_2026 = [
    date(2026, 1, 1), date(2026, 2, 1), date(2026, 3, 1),
    date(2026, 4, 1), date(2026, 5, 1), date(2026, 6, 1),
    date(2026, 7, 1), date(2026, 8, 1), date(2026, 9, 1),
    date(2026, 10, 1), date(2026, 11, 1), date(2026, 12, 1),
]

# ── Funções auxiliares ─────────────────────────────────────────────────────────

def normalizar_cnpj(val) -> str | None:
    """
    Normaliza CNPJ para string 14 dígitos zero-padded.
    Aceita: string formatada '04.067.573/0001-93', float 4067573000193.0, int, etc.
    Retorna None se inválido (menos de 11 dígitos após limpeza).
    NUNCA float ou int.
    """
    if val is None:
        return None
    # Se vier como float (ex: 4067573000193.0), converter para int primeiro
    if isinstance(val, float):
        val = str(int(val))
    else:
        val = str(val)
    # Remover tudo que não for dígito
    digits = re.sub(r'\D', '', val)
    if not digits or len(digits) < 11:
        return None
    return digits.zfill(14)


def safe_float(val) -> float | None:
    """Converte para float, tratando None, strings '#NAME?', etc."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        v = val.strip()
        if not v or v.startswith('#'):
            return None
        try:
            return float(v.replace(',', '.'))
        except ValueError:
            return None
    return None


def safe_int(val) -> int | None:
    """Converte para int, tratando None, strings '#NAME?', etc."""
    f = safe_float(val)
    if f is None:
        return None
    return int(f)


def safe_str(val, max_len: int = None) -> str | None:
    """Converte para string, truncando se necessário. Ignora erros #NAME? etc."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return str(val)
    s = str(val).strip()
    if not s or s.startswith('#'):
        return None
    if max_len:
        s = s[:max_len]
    return s


def extrair_prioridade_do_score(score: float | None) -> str | None:
    """
    Deriva prioridade (P0-P7) a partir do score numérico.
    A coluna 144 retorna #NAME? em data_only=True, então calculamos aqui.
    Lógica baseada na distribuição observada: score 0-100.
    """
    if score is None:
        return None
    s = float(score)
    if s >= 85:
        return 'P0'
    elif s >= 70:
        return 'P1'
    elif s >= 55:
        return 'P2'
    elif s >= 40:
        return 'P3'
    elif s >= 30:
        return 'P4'
    elif s >= 20:
        return 'P5'
    elif s >= 10:
        return 'P6'
    else:
        return 'P7'


def normalizar_macroregiao(val) -> str | None:
    """Extrai o nome limpo da macroregião, removendo prefixo '03 - '."""
    s = safe_str(val, 50)
    if not s:
        return None
    # Remove prefixo '03 - ' ou similar
    if ' - ' in s:
        return s.split(' - ', 1)[1].strip()[:50]
    return s[:50]


def normalizar_consultor(val) -> str | None:
    """Aplica DE-PARA de consultores conforme regras do projeto."""
    s = safe_str(val, 50)
    if not s:
        return None
    upper = s.upper()
    if 'MANU' in upper or 'DITZEL' in upper:
        return 'MANU'
    if 'LARISSA' in upper or 'LARI' in upper or 'MAIS GRANEL' in upper or 'RODRIGO' in upper:
        return 'LARISSA'
    if 'DAIANE' in upper or 'STAVICKI' in upper:
        return 'DAIANE'
    if 'JULIO' in upper or 'GADRET' in upper:
        return 'JULIO'
    # LEGADO
    legado = ['BRUNO', 'GRETTER', 'JEFERSON', 'PATRIC', 'GABRIEL', 'SERGIO', 'IVE', 'ANA']
    if any(l in upper for l in legado):
        return 'LEGADO'
    # Retorna original se não casar
    return s[:50]


# ── Script principal ───────────────────────────────────────────────────────────

def main():
    print('=' * 60)
    print('CRM VITAO360 — Carga de dados da aba CARTEIRA')
    print(f'Início: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print('=' * 60)

    # Verificar que a planilha existe
    if not Path(PLANILHA).exists():
        print(f'ERRO: Planilha não encontrada em:\n  {PLANILHA}')
        sys.exit(1)

    print(f'Abrindo planilha (data_only=True)...')
    wb = openpyxl.load_workbook(PLANILHA, data_only=True, read_only=True)

    if 'CARTEIRA' not in wb.sheetnames:
        print('ERRO: Aba CARTEIRA não encontrada na planilha.')
        sys.exit(1)

    ws = wb['CARTEIRA']
    print(f'Aba CARTEIRA: {ws.max_row} linhas x {ws.max_column} colunas')

    # Conectar ao banco
    conn = sqlite3.connect(DB_PATH)
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA foreign_keys=ON')
    cur = conn.cursor()

    # Limpar dados existentes para recarga limpa
    print()
    print('Limpando dados existentes...')
    cur.execute('DELETE FROM vendas')
    cur.execute('DELETE FROM metas')
    cur.execute('DELETE FROM clientes')
    conn.commit()
    print('  Tabelas clientes, vendas, metas limpas.')

    # Contadores
    clientes_inseridos = 0
    clientes_ignorados = 0
    vendas_inseridas = 0
    metas_inseridas = 0
    cnpjs_vistos = set()

    print()
    print('Processando linhas...')

    # Buffer para batch commit
    batch_clientes = []
    batch_vendas = []
    batch_metas = []

    def flush_batch():
        nonlocal clientes_inseridos, vendas_inseridas, metas_inseridas

        if batch_clientes:
            cur.executemany("""
                INSERT OR REPLACE INTO clientes (
                    cnpj, nome_fantasia, razao_social, uf, cidade,
                    codigo_cliente, email, telefone,
                    rede_regional, consultor, situacao, tipo_cliente,
                    classificacao_3tier,
                    dias_sem_compra, valor_ultimo_pedido, ciclo_medio,
                    faturamento_total, n_compras, curva_abc,
                    estagio_funil, resultado, temperatura, sinaleiro,
                    score, prioridade,
                    meta_anual, realizado_acumulado, pct_alcancado,
                    macroregiao,
                    created_at, updated_at
                ) VALUES (
                    ?, ?, ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?, ?,
                    ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?, ?,
                    ?, ?,
                    ?, ?, ?,
                    ?,
                    ?, ?
                )
            """, batch_clientes)
            clientes_inseridos += len(batch_clientes)
            batch_clientes.clear()

        if batch_vendas:
            cur.executemany("""
                INSERT INTO vendas (
                    cnpj, data_pedido, valor_pedido,
                    consultor, fonte, classificacao_3tier, mes_referencia,
                    created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, batch_vendas)
            vendas_inseridas += len(batch_vendas)
            batch_vendas.clear()

        if batch_metas:
            cur.executemany("""
                INSERT INTO metas (
                    cnpj, ano, mes, meta_sap, realizado, fonte
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, batch_metas)
            metas_inseridas += len(batch_metas)
            batch_metas.clear()

        conn.commit()

    # Iterar a partir da linha 4 (dados)
    now_str = datetime.now().isoformat()
    rows_processadas = 0

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        rows_processadas += 1

        # ── CNPJ — chave obrigatória ──────────────────────────────────────────
        cnpj_raw = row[1]   # col 2 (índice 1)
        cnpj = normalizar_cnpj(cnpj_raw)

        if not cnpj:
            clientes_ignorados += 1
            continue

        # Detectar duplicatas (pode haver CNPJs repetidos na planilha)
        if cnpj in cnpjs_vistos:
            clientes_ignorados += 1
            continue
        cnpjs_vistos.add(cnpj)

        # ── Campos básicos ───────────────────────────────────────────────────
        nome_fantasia     = safe_str(row[0],  255)   # col 1
        razao_social      = safe_str(row[2],  255)   # col 3
        uf                = safe_str(row[3],    2)   # col 4
        cidade            = safe_str(row[4],  100)   # col 5
        email             = safe_str(row[5],  255)   # col 6
        telefone          = safe_str(row[6],   20)   # col 7
        rede_regional     = safe_str(row[8],  100)   # col 9
        consultor         = normalizar_consultor(row[10])  # col 11
        situacao          = safe_str(row[12],  20)   # col 13
        dias_sem_compra   = safe_int(row[14])         # col 15
        valor_ultimo_ped  = safe_float(row[16])       # col 17
        ciclo_medio       = safe_float(row[17])       # col 18

        # ── Faturamento total 2025 (Two-Base: APENAS coluna TOTAL 2025) ──────
        # Col 28 (índice 27) = TOTAL 2025 — tipo VENDA
        faturamento_total = safe_float(row[27])

        # ── Recorrência / ABC ─────────────────────────────────────────────────
        n_compras   = safe_int(row[55])     # col 56
        curva_abc   = safe_str(row[56], 1)  # col 57 — A/B/C ou None
        # Validar curva_abc
        if curva_abc and curva_abc not in ('A', 'B', 'C'):
            curva_abc = None

        # ── Pipeline / Funil ──────────────────────────────────────────────────
        estagio_funil = safe_str(row[61], 50)   # col 62
        resultado     = safe_str(row[65], 50)   # col 66
        tipo_cliente  = safe_str(row[67], 30)   # col 68
        temperatura   = safe_str(row[71], 20)   # col 72

        # ── Sinaleiro e score ─────────────────────────────────────────────────
        sinaleiro = safe_str(row[79], 20)        # col 80
        # Validar sinaleiro
        if sinaleiro and sinaleiro.upper() not in ('VERDE', 'AMARELO', 'LARANJA', 'VERMELHO', 'ROXO'):
            sinaleiro = None
        elif sinaleiro:
            sinaleiro = sinaleiro.upper()

        # ── SAP ───────────────────────────────────────────────────────────────
        codigo_cliente = safe_str(row[80], 20)   # col 81

        # ── Macroregião ────────────────────────────────────────────────────────
        macroregiao = normalizar_macroregiao(row[92])   # col 93

        # ── Metas ─────────────────────────────────────────────────────────────
        meta_anual          = safe_float(row[95])   # col 96
        realizado_acumulado = safe_float(row[96])   # col 97
        pct_alcancado       = safe_float(row[97])   # col 98

        # ── Score ─────────────────────────────────────────────────────────────
        score = safe_float(row[142])   # col 143 (score numérico)

        # ── Prioridade derivada do score ──────────────────────────────────────
        # Col 144 retorna #NAME? em data_only=True; derivamos do score
        prioridade = extrair_prioridade_do_score(score)

        # ── Validação situacao ─────────────────────────────────────────────────
        situacoes_validas = {'ATIVO', 'EM RISCO', 'INAT.REC', 'INAT.ANT', 'PROSPECT'}
        if situacao and situacao.upper() in situacoes_validas:
            situacao = situacao.upper()
        elif situacao:
            situacao = situacao[:20]

        # ── Buffer cliente ─────────────────────────────────────────────────────
        batch_clientes.append((
            cnpj, nome_fantasia, razao_social, uf, cidade,
            codigo_cliente, email, telefone,
            rede_regional, consultor, situacao, tipo_cliente,
            'REAL',                                # classificacao_3tier
            dias_sem_compra, valor_ultimo_ped, ciclo_medio,
            faturamento_total, n_compras, curva_abc,
            estagio_funil, resultado, temperatura, sinaleiro,
            score, prioridade,
            meta_anual, realizado_acumulado, pct_alcancado,
            macroregiao,
            now_str, now_str,
        ))

        # ── Vendas 2025 (cols 29-40, índices 28-39) ───────────────────────────
        # Two-Base: cada mês com valor > 0 = 1 registro de VENDA
        consultor_venda = consultor or 'DESCONHECIDO'
        for idx, mes_date in enumerate(MESES_2025):
            col_idx = 28 + idx   # índices 28..39
            valor = safe_float(row[col_idx])
            if valor and valor > 0:
                mes_ref = mes_date.strftime('%Y-%m')
                batch_vendas.append((
                    cnpj,
                    mes_date.isoformat(),
                    valor,
                    consultor_venda,
                    'MERCOS',    # fonte
                    'REAL',      # classificacao_3tier
                    mes_ref,
                    now_str,
                ))

        # ── Vendas 2026 (cols 42-53, índices 41-52) ───────────────────────────
        for idx, mes_date in enumerate(MESES_2026):
            col_idx = 41 + idx   # índices 41..52
            valor = safe_float(row[col_idx])
            if valor and valor > 0:
                mes_ref = mes_date.strftime('%Y-%m')
                batch_vendas.append((
                    cnpj,
                    mes_date.isoformat(),
                    valor,
                    consultor_venda,
                    'MERCOS',
                    'REAL',
                    mes_ref,
                    now_str,
                ))

        # ── Meta anual ────────────────────────────────────────────────────────
        if meta_anual and meta_anual > 0:
            batch_metas.append((
                cnpj,
                2026,       # ano
                0,          # mes=0 indica anual
                meta_anual,
                realizado_acumulado or 0.0,
                'SAP',
            ))

        # ── Batch commit a cada BATCH_SIZE ────────────────────────────────────
        if len(batch_clientes) >= BATCH_SIZE:
            flush_batch()
            print(f'  ... {clientes_inseridos} clientes / {vendas_inseridas} vendas / {metas_inseridas} metas')

    # Flush final
    flush_batch()

    wb.close()
    conn.close()

    print()
    print('=' * 60)
    print('RELATÓRIO FINAL')
    print('=' * 60)
    print(f'Rows processadas:    {rows_processadas}')
    print(f'Clientes inseridos:  {clientes_inseridos}')
    print(f'Clientes ignorados:  {clientes_ignorados}  (CNPJ vazio ou duplicado)')
    print(f'Vendas inseridas:    {vendas_inseridas}')
    print(f'Metas inseridas:     {metas_inseridas}')
    print(f'CNPJs únicos:        {len(cnpjs_vistos)}')
    print()

    # Validação básica Two-Base
    conn2 = sqlite3.connect(DB_PATH)
    cur2 = conn2.cursor()
    cur2.execute('SELECT COUNT(*), SUM(valor_pedido) FROM vendas')
    vcount, vtotal = cur2.fetchone()
    cur2.execute('SELECT COUNT(*) FROM clientes')
    (ccount,) = cur2.fetchone()
    cur2.execute('SELECT COUNT(*) FROM metas')
    (mcount,) = cur2.fetchone()
    conn2.close()

    print('VALIDAÇÃO DO BANCO:')
    print(f'  clientes no DB:  {ccount}')
    print(f'  vendas no DB:    {vcount}  |  total R$ {vtotal:,.2f}' if vtotal else f'  vendas no DB:    {vcount}')
    print(f'  metas no DB:     {mcount}')
    print()

    # Verificação faturamento 2025 vs baseline R$ 2.091.000
    conn3 = sqlite3.connect(DB_PATH)
    cur3 = conn3.cursor()
    cur3.execute("""
        SELECT SUM(valor_pedido)
        FROM vendas
        WHERE mes_referencia LIKE '2025-%'
    """)
    (fat_2025,) = cur3.fetchone()
    conn3.close()

    if fat_2025:
        baseline = 2_091_000.0
        diff_pct = abs(fat_2025 - baseline) / baseline * 100
        status = 'OK' if diff_pct <= 0.5 else 'DIVERGENCIA'
        print(f'VERIFICACAO FATURAMENTO 2025:')
        print(f'  Calculado: R$ {fat_2025:,.2f}')
        print(f'  Baseline:  R$ {baseline:,.2f}')
        print(f'  Diferenca: {diff_pct:.2f}%  [{status}]')
        if diff_pct > 0.5:
            print('  ATENCAO: divergencia > 0.5% — investigar (R7)')
    print()
    print(f'Concluído: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print('=' * 60)


if __name__ == '__main__':
    main()

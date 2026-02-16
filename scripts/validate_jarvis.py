import os
import openpyxl

path = r'C:\Users\User\OneDrive\Área de Trabalho\CLAUDE CODE\output\JARVIS_CRM_CENTRAL.xlsx'
size = os.path.getsize(path)
print(f'File size: {size:,} bytes ({size/1024:.1f} KB)')

wb = openpyxl.load_workbook(path)
print(f'\nSheets ({len(wb.sheetnames)}): {wb.sheetnames}')
for name in wb.sheetnames:
    ws = wb[name]
    print(f'  {name}: {ws.max_column} cols x {ws.max_row} rows')

nr_list = list(wb.defined_names.values())
print(f'\nNamed ranges ({len(nr_list)}):')
for dn in nr_list:
    print(f'  {dn.name} -> {dn.attr_text}')

# Check data validations
print('\nData Validations:')
for name in wb.sheetnames:
    ws = wb[name]
    dvs = ws.data_validations.dataValidation if ws.data_validations else []
    if dvs:
        print(f'  {name}: {len(dvs)} validations')
        for dv in dvs:
            print(f'    {dv.sqref}: type={dv.type}, formula1={dv.formula1}')

# Check conditional formatting
print('\nConditional Formatting:')
for name in wb.sheetnames:
    ws = wb[name]
    cfs = ws.conditional_formatting
    if cfs:
        print(f'  {name}: {len(list(cfs))} rules')

# Check column groups in CARTEIRA
ws_cart = wb['CARTEIRA']
print('\nColumn Groups in CARTEIRA:')
grouped = []
for letter, dim in ws_cart.column_dimensions.items():
    if hasattr(dim, 'outlineLevel') and dim.outlineLevel and dim.outlineLevel > 0:
        grouped.append(letter)
    if hasattr(dim, 'hidden') and dim.hidden:
        grouped.append(f'{letter}(hidden)')

# Check grouped columns via group info
print('  Checking group ranges...')
# Groups are stored differently, let's check hidden columns
hidden_cols = []
for col_idx in range(1, 82):
    letter = openpyxl.utils.get_column_letter(col_idx)
    dim = ws_cart.column_dimensions.get(letter)
    if dim and dim.hidden:
        hidden_cols.append(letter)
print(f'  Hidden columns: {len(hidden_cols)} ({", ".join(hidden_cols[:10])}...)')

# Check freeze panes
print('\nFreeze Panes:')
for name in wb.sheetnames:
    ws = wb[name]
    fp = ws.freeze_panes
    print(f'  {name}: {fp}')

# Verify sample formulas in CARTEIRA
print('\nSample Formulas in CARTEIRA:')
ws_cart = wb['CARTEIRA']
for col_name, col_idx in [('E-SITUAÇÃO', 5), ('F-DIAS', 6), ('G-SINALEIRO', 7), ('I-FASE', 9), ('BQ-PRIORIDADE', 69), ('BS-%ATING', 71)]:
    cell = ws_cart.cell(row=2, column=col_idx)
    val = cell.value
    if val and str(val).startswith('='):
        print(f'  {col_name}: {val[:60]}...' if len(str(val)) > 60 else f'  {col_name}: {val}')
    else:
        print(f'  {col_name}: {val} (NOT A FORMULA!)')

# Check LOG formulas
print('\nSample Formulas in LOG:')
ws_log = wb['LOG']
cell = ws_log.cell(row=2, column=15)
print(f'  O-FOLLOW-UP: {cell.value}')

print('\n=== VALIDATION COMPLETE ===')

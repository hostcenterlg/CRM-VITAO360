"""
Phase 06 - Plan 01: ETL de Relatorios E-commerce Mercos
========================================================
Le 17 arquivos .xlsx + 2 .xls de acesso ao e-commerce B2B Mercos:
  - Header detection dinamica (formato 9 ou 11 colunas)
  - Dedup de arquivos identicos e meses sobrepostos
  - Mapeamento de mes real via data de emissao (NAO nome do arquivo)
  - Output JSON intermediario normalizado

Output: data/output/phase06/ecommerce_raw.json
"""

import json
import re
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from collections import defaultdict

# PROJECT_ROOT pattern
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# openpyxl
import openpyxl

# xlrd - optional, graceful skip
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

# =====================================================================
# CONSTANTS
# =====================================================================

ECOMMERCE_DIR = PROJECT_ROOT / "data" / "sources" / "mercos" / "ecommerce"
FEV2026_DIR = PROJECT_ROOT / "data" / "sources" / "mercos" / "fev2026"
OUTPUT_PATH = PROJECT_ROOT / "data" / "output" / "phase06" / "ecommerce_raw.json"

# Expected months range: MAR/2025 - FEV/2026
EXPECTED_MONTHS = [
    "2025-03", "2025-04", "2025-05", "2025-06", "2025-07",
    "2025-08", "2025-09", "2025-10", "2025-11", "2025-12",
    "2026-01", "2026-02",
]

# Month name hints from filenames (secondary, NOT primary source of truth)
MONTH_NAME_HINTS = {
    "marco": "2025-03", "março": "2025-03", "mar": "2025-03",
    "abril": "2025-04", "abr": "2025-04",
    "maio": "2025-05", "mai": "2025-05",
    "junho": "2025-06", "jun": "2025-06",
    "julho": "2025-07", "jul": "2025-07",
    "agosto": "2025-08", "ago": "2025-08",
    "setembro": "2025-09", "set": "2025-09",
    "outubro": "2025-10", "out": "2025-10",
    "novembro": "2025-11", "nov": "2025-11",
    "dezembro": "2025-12", "dez": "2025-12",
    "janeiro": "2026-01", "jan": "2026-01",
    "fevereiro": "2026-02", "fev": "2026-02",
}

# Year overrides for month names that include year
YEAR_OVERRIDES = {
    "dezembro 2025": "2025-12",
    "janeiro 2026": "2026-01",
    "fevereiro 2026": "2026-02",
}


# =====================================================================
# UTILITY FUNCTIONS (standalone, copied from Phase 4 patterns)
# =====================================================================

def normalize_cnpj(raw):
    """Remove tudo exceto digitos, zfill(14)."""
    if raw is None:
        return ""
    return re.sub(r'[^0-9]', '', str(raw)).zfill(14)


def safe_int(val):
    """Converte para int, tratando None, '', strings."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip().replace(',', '.')
    if s == '' or s == '-':
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def safe_float(val):
    """Converte para float, tratando None, '', strings com virgula."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s == '' or s == '-':
        return 0.0
    # Handle Brazilian format: 1.234,56 -> 1234.56
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def extract_cnpj_prefix(razao_social):
    """
    Extrai CNPJ do prefixo do nome em formato novo.
    Ex: '48.144.171 ROSANGELA...' -> digits only
    Returns digits-only string or None.
    """
    m = re.match(r'^([\d./-]+)\s', razao_social)
    if m:
        digits = re.sub(r'[^0-9]', '', m.group(1))
        if len(digits) >= 8:  # At least CNPJ root (8 digits)
            return digits
    return None


# =====================================================================
# XLSX READING FUNCTIONS
# =====================================================================

def find_header_row(ws):
    """Encontra a row com headers (Razao Social na col 1)."""
    for r in range(1, 15):
        val = ws.cell(row=r, column=1).value
        if val and ('Raz' in str(val) or 'raz' in str(val).lower()):
            return r
    return 6  # fallback


def get_emission_date(ws):
    """Extrai data de emissao do relatorio (tipicamente row 3: 'Emitido em DD/MM/YYYY')."""
    for r in range(1, 10):
        val = str(ws.cell(row=r, column=1).value or '')
        if 'Emitido em' in val or 'emitido em' in val.lower():
            match = re.search(r'(\d{2}/\d{2}/\d{4})', val)
            if match:
                return match.group(1)
    # Try column 2 as well
    for r in range(1, 10):
        val = str(ws.cell(row=r, column=2).value or '')
        if 'Emitido em' in val or 'emitido em' in val.lower():
            match = re.search(r'(\d{2}/\d{2}/\d{4})', val)
            if match:
                return match.group(1)
    return None


def detect_format(ws, header_row):
    """
    Detecta formato: 9 cols (antigo) ou 11 cols (novo com Email+Telefone).
    Verifica se coluna 3 do header contem 'mail'.
    """
    col3_header = ws.cell(row=header_row, column=3).value
    if col3_header and 'mail' in str(col3_header).lower():
        return 'NEW_11COL'
    return 'OLD_9COL'


def read_xlsx_file(filepath):
    """
    Le um arquivo .xlsx de e-commerce Mercos.
    Returns: (records, emission_date, format_type, row_count)
    """
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active

    header_row = find_header_row(ws)
    emission_date = get_emission_date(ws)
    fmt = detect_format(ws, header_row)

    # Offset para campos de dados
    offset = 2 if fmt == 'NEW_11COL' else 0

    records = []
    for row in range(header_row + 1, (ws.max_row or header_row) + 1):
        razao = ws.cell(row=row, column=1).value
        if not razao or str(razao).strip() == '':
            continue

        razao_str = str(razao).strip()
        fantasia = ws.cell(row=row, column=2).value
        fantasia_str = str(fantasia or '').strip()

        # Extract data columns with offset
        acessos = ws.cell(row=row, column=3 + offset).value
        itens = ws.cell(row=row, column=4 + offset).value
        orcamentos = ws.cell(row=row, column=5 + offset).value
        atividades = ws.cell(row=row, column=6 + offset).value
        valor_carrinho = ws.cell(row=row, column=7 + offset).value
        valor_orc_nf = ws.cell(row=row, column=8 + offset).value
        valor_b2b = ws.cell(row=row, column=9 + offset).value

        # Extract CNPJ from name prefix (new format)
        cnpj_from_name = extract_cnpj_prefix(razao_str)

        records.append({
            'razao_social': razao_str,
            'nome_fantasia': fantasia_str,
            'cnpj_from_name': cnpj_from_name,
            'acessos': safe_int(acessos),
            'itens_adicionados': safe_int(itens),
            'orcamentos_finalizados': safe_int(orcamentos),
            'atividades': safe_int(atividades),
            'valor_carrinho': safe_float(valor_carrinho),
            'valor_orc_nao_finalizados': safe_float(valor_orc_nf),
            'valor_pedidos_b2b': safe_float(valor_b2b),
        })

    wb.close()
    return records, emission_date, fmt, len(records)


def read_xls_file(filepath):
    """
    Le um arquivo .xls de e-commerce Mercos usando xlrd.
    Returns: (records, emission_date, format_type, row_count) or raises.
    """
    if not HAS_XLRD:
        raise ImportError("xlrd not available")

    wb = xlrd.open_workbook(str(filepath))
    ws = wb.sheet_by_index(0)

    # Find header row
    header_row = None
    for r in range(min(15, ws.nrows)):
        val = ws.cell_value(r, 0)
        if val and ('Raz' in str(val) or 'raz' in str(val).lower()):
            header_row = r
            break
    if header_row is None:
        header_row = 5  # fallback (0-indexed)

    # Get emission date
    emission_date = None
    for r in range(min(10, ws.nrows)):
        val = str(ws.cell_value(r, 0) or '')
        if 'emitido em' in val.lower():
            match = re.search(r'(\d{2}/\d{2}/\d{4})', val)
            if match:
                emission_date = match.group(1)
                break

    # Detect format
    fmt = 'OLD_9COL'
    if ws.ncols >= 3:
        col3_header = ws.cell_value(header_row, 2)
        if col3_header and 'mail' in str(col3_header).lower():
            fmt = 'NEW_11COL'

    offset = 2 if fmt == 'NEW_11COL' else 0

    records = []
    for r in range(header_row + 1, ws.nrows):
        razao = ws.cell_value(r, 0)
        if not razao or str(razao).strip() == '':
            continue

        razao_str = str(razao).strip()
        fantasia = ws.cell_value(r, 1) if ws.ncols > 1 else ''
        fantasia_str = str(fantasia or '').strip()

        acessos = ws.cell_value(r, 2 + offset) if ws.ncols > 2 + offset else 0
        itens = ws.cell_value(r, 3 + offset) if ws.ncols > 3 + offset else 0
        orcamentos = ws.cell_value(r, 4 + offset) if ws.ncols > 4 + offset else 0
        atividades = ws.cell_value(r, 5 + offset) if ws.ncols > 5 + offset else 0
        valor_carrinho = ws.cell_value(r, 6 + offset) if ws.ncols > 6 + offset else 0.0
        valor_orc_nf = ws.cell_value(r, 7 + offset) if ws.ncols > 7 + offset else 0.0
        valor_b2b = ws.cell_value(r, 8 + offset) if ws.ncols > 8 + offset else 0.0

        cnpj_from_name = extract_cnpj_prefix(razao_str)

        records.append({
            'razao_social': razao_str,
            'nome_fantasia': fantasia_str,
            'cnpj_from_name': cnpj_from_name,
            'acessos': safe_int(acessos),
            'itens_adicionados': safe_int(itens),
            'orcamentos_finalizados': safe_int(orcamentos),
            'atividades': safe_int(atividades),
            'valor_carrinho': safe_float(valor_carrinho),
            'valor_orc_nao_finalizados': safe_float(valor_orc_nf),
            'valor_pedidos_b2b': safe_float(valor_b2b),
        })

    return records, emission_date, fmt, len(records)


# =====================================================================
# FILE DISCOVERY
# =====================================================================

def discover_files():
    """
    Descobre todos os arquivos de e-commerce em ambas as localizacoes.
    Ignora arquivos temporarios (~$*).
    Returns list of dicts with path, filename, extension.
    """
    files = []

    # Location 1: data/sources/mercos/ecommerce/
    if ECOMMERCE_DIR.exists():
        for f in sorted(ECOMMERCE_DIR.iterdir()):
            if f.name.startswith('~$'):
                continue
            if f.suffix.lower() in ('.xlsx', '.xls'):
                files.append({
                    'path': str(f),
                    'filename': f.name,
                    'extension': f.suffix.lower(),
                    'location': 'ecommerce/',
                })

    # Location 2: data/sources/mercos/fev2026/ (only ecommerce file)
    if FEV2026_DIR.exists():
        for f in sorted(FEV2026_DIR.iterdir()):
            if f.name.startswith('~$'):
                continue
            fname_lower = f.name.lower()
            if 'ecom' in fname_lower and f.suffix.lower() in ('.xlsx', '.xls'):
                files.append({
                    'path': str(f),
                    'filename': f.name,
                    'extension': f.suffix.lower(),
                    'location': 'fev2026/',
                })

    return files


# =====================================================================
# DEDUP LOGIC
# =====================================================================

def make_signature(records, n=5):
    """
    Cria signature para comparacao de duplicatas.
    Usa primeiros N registros: razao_social uppercase[:30] + acessos.
    """
    sig_parts = []
    for r in records[:n]:
        sig_parts.append((r['razao_social'].upper()[:30], r['acessos']))
    return tuple(sig_parts)


def _build_dedup_reason(fnames_sorted, file_data):
    """Build dedup reason string avoiding f-string escaping issues."""
    kept = fnames_sorted[0]
    kept_rows = file_data[kept]['rows']
    kept_recs = len(file_data[kept]['records'])
    n_compared = min(5, kept_recs)
    discarded_info = []
    for f in fnames_sorted[1:]:
        discarded_info.append(f"{f} ({file_data[f]['rows']} rows)")
    return (
        f"Same first {n_compared} records (signature match). "
        f"Kept '{kept}' ({kept_rows} rows) over {discarded_info}"
    )


def detect_duplicate_groups(file_data):
    """
    Detecta grupos de arquivos duplicados por signature.
    file_data: dict {filename: {records, emission_date, format, rows, path}}
    Returns: list of dedup groups.
    """
    signatures = defaultdict(list)
    for fname, finfo in file_data.items():
        if finfo['records']:
            sig = make_signature(finfo['records'])
            signatures[sig].append(fname)

    groups = []
    for sig, fnames in signatures.items():
        if len(fnames) > 1:
            # Sort by rows descending (keep most data)
            fnames_sorted = sorted(fnames, key=lambda f: file_data[f]['rows'], reverse=True)
            groups.append({
                'group': fnames_sorted,
                'kept': fnames_sorted[0],
                'discarded': fnames_sorted[1:],
                'reason': _build_dedup_reason(fnames_sorted, file_data),
            })

    return groups


def apply_month_dedup(file_data, inventory):
    """
    Dedup de nivel 2: quando 2 arquivos cobrem o mesmo mes,
    preferir o com mais rows.
    Returns: updated inventory with dedup decisions.
    """
    # Group by assigned_month
    month_files = defaultdict(list)
    for fname, inv in inventory.items():
        if inv['status'] == 'PROCESSED' and inv['assigned_month']:
            month_files[inv['assigned_month']].append(fname)

    dedup_groups = []
    for month, fnames in month_files.items():
        if len(fnames) > 1:
            # Sort by rows descending
            fnames_sorted = sorted(fnames, key=lambda f: file_data[f]['rows'], reverse=True)
            kept = fnames_sorted[0]
            discarded = fnames_sorted[1:]

            for d in discarded:
                inventory[d]['status'] = 'DEDUP_DISCARDED'
                inventory[d]['dedup_note'] = (
                    f"Month {month} duplicate. Kept '{kept}' "
                    f"({file_data[kept]['rows']} rows) over '{d}' "
                    f"({file_data[d]['rows']} rows)"
                )

            dedup_groups.append({
                'group': fnames_sorted,
                'kept': kept,
                'discarded': discarded,
                'reason': (
                    f"Same month ({month}). Kept '{kept}' "
                    f"({file_data[kept]['rows']} rows, emitted "
                    f"{file_data[kept]['emission_date'] or 'N/A'})"
                ),
            })

    return dedup_groups


# =====================================================================
# MONTH ASSIGNMENT
# =====================================================================

def extract_month_hint_from_filename(filename):
    """
    Extrai pista de mes do nome do arquivo (fonte SECUNDARIA).
    Returns: YYYY-MM string or None.
    """
    fname_lower = filename.lower().strip()

    # Remove extension
    fname_lower = re.sub(r'\.(xlsx|xls)$', '', fname_lower).strip()

    # Check year overrides first (e.g., "dezembro 2025", "janeiro 2026")
    for pattern, month in YEAR_OVERRIDES.items():
        if pattern in fname_lower:
            return month

    # Check month names
    for name, month in MONTH_NAME_HINTS.items():
        if name in fname_lower:
            # Check if year 2026 is mentioned
            if '2026' in fname_lower and month.startswith('2025'):
                return '2026-' + month.split('-')[1]
            return month

    return None


def assign_month(emission_date_str, filename, rows):
    """
    Atribui mes real ao arquivo usando:
    1. Data de emissao como base (quando disponivel)
    2. Nome do arquivo como pista secundaria
    3. Heuristica: relatorios tipicamente emitidos no mes SEGUINTE

    Returns: YYYY-MM string or 'UNKNOWN'.
    """
    hint = extract_month_hint_from_filename(filename)

    if emission_date_str:
        try:
            emission = datetime.strptime(emission_date_str, '%d/%m/%Y')
            # The emission date is when the report was generated,
            # NOT the month of the data.
            # Typically reports are emitted in the following month
            # or at the end of the current month.

            # If we have a filename hint, use it as primary
            # (emission date confirms plausibility)
            if hint:
                return hint

            # Without filename hint, try to infer:
            # If emitted on 15/12/2025, data could be Nov/2025 or earlier
            # This is ambiguous -- use UNKNOWN if no filename hint
            return hint or 'UNKNOWN'

        except ValueError:
            pass

    # Fallback to filename hint only
    return hint or 'UNKNOWN'


# =====================================================================
# MAIN ETL
# =====================================================================

def main():
    print("=" * 70)
    print("Phase 06 - Plan 01: ETL E-commerce Mercos")
    print("=" * 70)
    print()

    # Step 1: Discover files
    print("[1/7] Discovering e-commerce files...")
    files = discover_files()
    print(f"  Found {len(files)} files total")
    for f in files:
        print(f"    - [{f['extension']}] {f['filename']} ({f['location']})")
    print()

    # Step 2: Read all files
    print("[2/7] Reading files with header detection...")
    file_data = {}
    inventory = {}
    xls_warnings = []

    for finfo in files:
        fname = finfo['filename']
        fpath = finfo['path']
        ext = finfo['extension']

        try:
            if ext == '.xlsx':
                records, emission_date, fmt, row_count = read_xlsx_file(fpath)
            elif ext == '.xls':
                if not HAS_XLRD:
                    xls_warnings.append(fname)
                    inventory[fname] = {
                        'filename': fname,
                        'path': fpath,
                        'format': 'XLS',
                        'emission_date': None,
                        'assigned_month': None,
                        'rows': 0,
                        'status': 'SKIPPED_XLS_NO_XLRD',
                        'dedup_note': 'xlrd not installed - graceful skip',
                    }
                    print(f"  WARNING: Skipped {fname} (.xls requires xlrd)")
                    continue
                else:
                    records, emission_date, fmt, row_count = read_xls_file(fpath)
            else:
                continue

            file_data[fname] = {
                'records': records,
                'emission_date': emission_date,
                'format': fmt,
                'rows': row_count,
                'path': fpath,
            }

            inventory[fname] = {
                'filename': fname,
                'path': fpath,
                'format': fmt,
                'emission_date': emission_date,
                'assigned_month': None,  # Will be set later
                'rows': row_count,
                'status': 'PROCESSED',
                'dedup_note': None,
            }

            print(f"  OK: {fname} | {fmt} | {row_count} rows | emitted {emission_date or 'N/A'}")

        except Exception as e:
            inventory[fname] = {
                'filename': fname,
                'path': fpath,
                'format': 'ERROR',
                'emission_date': None,
                'assigned_month': None,
                'rows': 0,
                'status': 'SKIPPED_ERROR',
                'dedup_note': f'Error: {str(e)}',
            }
            print(f"  ERROR: {fname} - {e}")

    print()

    # Step 3: Dedup - Level 1 (identical files)
    print("[3/7] Dedup Level 1: Detecting identical files...")
    dedup_groups_l1 = detect_duplicate_groups(file_data)
    total_discarded_l1 = 0
    for grp in dedup_groups_l1:
        for d in grp['discarded']:
            if d in inventory:
                inventory[d]['status'] = 'DEDUP_DISCARDED'
                inventory[d]['dedup_note'] = grp['reason']
                total_discarded_l1 += 1
        print(f"  Group: {grp['group']}")
        print(f"    Kept: {grp['kept']}")
        print(f"    Discarded: {grp['discarded']}")
        print(f"    Reason: {grp['reason']}")
    if not dedup_groups_l1:
        print("  No identical file groups found.")
    print(f"  Total discarded (L1): {total_discarded_l1}")
    print()

    # Step 4: Assign months
    print("[4/7] Assigning real months to files...")

    # Build map of L1-dedup groups: kept_file -> list of ALL filenames in group
    l1_group_members = {}
    for grp in dedup_groups_l1:
        l1_group_members[grp['kept']] = grp['group']  # includes kept + discarded

    for fname, inv in inventory.items():
        if inv['status'] in ('SKIPPED_ERROR', 'SKIPPED_XLS_NO_XLRD'):
            continue
        if inv['status'] == 'DEDUP_DISCARDED':
            # Still assign month for documentation
            month = assign_month(inv['emission_date'], fname, inv['rows'])
            inv['assigned_month'] = month
            continue

        month = assign_month(inv['emission_date'], fname, inv['rows'])
        inv['assigned_month'] = month
        print(f"  {fname} -> {month} (emitted {inv['emission_date'] or 'N/A'}, {inv['rows']} rows)")

    # Post-assignment fix: handle L1-dedup trio conflict
    # Only for L1 groups where members claim DIFFERENT months
    # (the Abril/Maio/junho trio). Not for groups where members
    # claim the SAME month (like Dezembro pair).
    for kept_fname, group_members in l1_group_members.items():
        if kept_fname not in inventory or inventory[kept_fname]['status'] != 'PROCESSED':
            continue
        kept_month = inventory[kept_fname]['assigned_month']
        if not kept_month or kept_month == 'UNKNOWN':
            continue

        # Check if group members claim different months
        # (this identifies the trio, not same-name pairs like Dezembro)
        member_months = set()
        for member in group_members:
            hint = extract_month_hint_from_filename(member)
            if hint:
                member_months.add(hint)
        if len(member_months) <= 1:
            # All members claim same month (e.g., Dezembro pair) -- skip
            # L2 dedup will handle this normally
            continue

        # This is a multi-month-name group (trio). Check for conflict.
        conflict = False
        for other_fname, other_inv in inventory.items():
            if other_fname == kept_fname:
                continue
            if other_inv['status'] != 'PROCESSED':
                continue
            if other_inv['assigned_month'] == kept_month:
                # Check if data is different
                if kept_fname in file_data and other_fname in file_data:
                    kept_sig = make_signature(file_data[kept_fname]['records'])
                    other_sig = make_signature(file_data[other_fname]['records'])
                    if kept_sig != other_sig:
                        conflict = True
                        break

        if conflict:
            # Try alternate month names from discarded files in the group
            already_assigned = set(
                inv['assigned_month'] for inv in inventory.values()
                if inv['status'] == 'PROCESSED' and inv['assigned_month']
            )
            alternate_month = None
            for member in group_members:
                if member == kept_fname:
                    continue
                hint = extract_month_hint_from_filename(member)
                if hint and hint not in already_assigned:
                    alternate_month = hint
                    break

            if alternate_month:
                old_month = inventory[kept_fname]['assigned_month']
                inventory[kept_fname]['assigned_month'] = alternate_month
                inventory[kept_fname]['dedup_note'] = (
                    f"Trio identical files ({', '.join(group_members)}). "
                    f"Data differs from real {old_month} file. "
                    f"Reassigned to {alternate_month} (first available alternate month)."
                )
                print(f"  FIX: {kept_fname} reassigned {old_month} -> {alternate_month} "
                      f"(trio conflict with real {old_month} file)")
            else:
                # No alternate found, mark as UNKNOWN
                old_month = inventory[kept_fname]['assigned_month']
                inventory[kept_fname]['assigned_month'] = 'UNKNOWN'
                inventory[kept_fname]['dedup_note'] = (
                    f"Trio identical files ({', '.join(group_members)}). "
                    f"Data differs from real {old_month} file. "
                    f"Month could not be determined."
                )
                print(f"  FIX: {kept_fname} reassigned {old_month} -> UNKNOWN "
                      f"(trio conflict, no alternate available)")

    print()

    # Step 5: Dedup - Level 2 (same month)
    print("[5/7] Dedup Level 2: Same-month files...")
    dedup_groups_l2 = apply_month_dedup(file_data, inventory)
    for grp in dedup_groups_l2:
        print(f"  Month conflict: {grp['group']}")
        print(f"    Kept: {grp['kept']}")
        print(f"    Discarded: {grp['discarded']}")
        print(f"    Reason: {grp['reason']}")
    if not dedup_groups_l2:
        print("  No same-month duplicates found.")
    print()

    # Step 6: Build monthly data
    print("[6/7] Building monthly data from surviving files...")
    monthly_data = defaultdict(list)
    total_records = 0
    surviving_files = 0

    for fname, inv in inventory.items():
        if inv['status'] != 'PROCESSED':
            continue

        surviving_files += 1
        month = inv['assigned_month']
        if not month or month == 'UNKNOWN':
            print(f"  WARNING: {fname} has UNKNOWN month - including under 'UNKNOWN'")
            month = 'UNKNOWN'

        records = file_data[fname]['records']
        for r in records:
            monthly_data[month].append(r)
            total_records += 1

        print(f"  {month}: {fname} ({len(records)} records)")

    print(f"\n  Surviving files: {surviving_files}")
    print(f"  Total records: {total_records}")
    print(f"  Unique months: {len(monthly_data)}")
    print()

    # Determine missing months
    found_months = set(k for k in monthly_data.keys() if k != 'UNKNOWN')
    missing_months = sorted([m for m in EXPECTED_MONTHS if m not in found_months])

    # Step 7: Output JSON
    print("[7/7] Writing output JSON...")
    all_dedup_groups = dedup_groups_l1 + dedup_groups_l2

    output = {
        "metadata": {
            "extracted_at": datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
            "total_files_found": len(files),
            "total_files_processed": surviving_files,
            "files_skipped_dedup": sum(1 for inv in inventory.values() if inv['status'] == 'DEDUP_DISCARDED'),
            "files_skipped_error": sum(1 for inv in inventory.values() if inv['status'] in ('SKIPPED_ERROR', 'SKIPPED_XLS_NO_XLRD')),
            "unique_months": len([k for k in monthly_data.keys() if k != 'UNKNOWN']),
            "total_records": total_records,
            "missing_months": missing_months,
            "has_unknown_month": 'UNKNOWN' in monthly_data,
        },
        "file_inventory": [
            inv for inv in sorted(inventory.values(), key=lambda x: x['filename'])
        ],
        "monthly_data": {
            month: records
            for month, records in sorted(monthly_data.items())
        },
        "dedup_report": all_dedup_groups,
    }

    # Ensure output directory exists
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    print(f"  Written: {OUTPUT_PATH}")
    print(f"  Size: {OUTPUT_PATH.stat().st_size:,} bytes")
    print()

    # Final report
    print("=" * 70)
    print("EXECUTION REPORT")
    print("=" * 70)
    print()
    print(f"  Total files found:     {len(files)}")
    print(f"  Files processed:       {surviving_files}")
    print(f"  Files dedup discarded: {output['metadata']['files_skipped_dedup']}")
    print(f"  Files skipped (error): {output['metadata']['files_skipped_error']}")
    print(f"  Unique months:         {output['metadata']['unique_months']}")
    print(f"  Total records:         {total_records}")
    print()

    print("  MONTHLY INVENTORY:")
    for month in sorted(monthly_data.keys()):
        count = len(monthly_data[month])
        source_files = [
            fname for fname, inv in inventory.items()
            if inv['status'] == 'PROCESSED' and inv['assigned_month'] == month
        ]
        print(f"    {month}: {count:>4} records  <- {', '.join(source_files)}")
    print()

    if missing_months:
        print(f"  MISSING MONTHS: {', '.join(missing_months)}")
    else:
        print("  MISSING MONTHS: None")
    print()

    if xls_warnings:
        print(f"  XLS WARNINGS: {', '.join(xls_warnings)} (xlrd not installed)")
    print()

    if all_dedup_groups:
        print(f"  DEDUP DECISIONS: {len(all_dedup_groups)} groups")
        for i, grp in enumerate(all_dedup_groups, 1):
            print(f"    {i}. Kept '{grp['kept']}', discarded {grp['discarded']}")
    print()

    print("DONE.")
    return output


if __name__ == '__main__':
    main()

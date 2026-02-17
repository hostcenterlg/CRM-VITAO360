"""
Phase 06 - Plan 02: E-commerce Match & Populate
=================================================
Cruza registros de e-commerce (ecommerce_raw.json) com CNPJs via
matching por nome (4 niveis), agrega dados por CNPJ, produz JSON
intermediario para Phase 9, e popula colunas 15-20 do DRAFT 1.

Matching niveis:
  1. CNPJ do prefixo (cnpj_from_name >= 11 digitos)
  2. Exact match por Razao Social ou Nome Fantasia
  3. Partial match (substring, min 5 chars)
  4. Fuzzy match (rapidfuzz, threshold 85, se disponivel)

Fontes do lookup nome->CNPJ:
  - Mercos Carteira (08_CARTEIRA_MERCOS.xlsx)
  - DRAFT 1 (DL_DRAFT1_FEV2026.xlsx)
  - sap_mercos_merged.json (validacao de CNPJ)

Output:
  - data/output/phase06/ecommerce_matched.json
  - data/output/phase06/match_report.json
  - data/sources/drafts/DL_DRAFT1_FEV2026.xlsx (cols 15-20 populadas)
"""

import json
import re
import os
import sys
import calendar
from datetime import datetime, timezone
from pathlib import Path
from collections import defaultdict

# PROJECT_ROOT pattern
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# openpyxl
import openpyxl

# rapidfuzz - optional
try:
    from rapidfuzz import fuzz, process as rf_process
    HAS_RAPIDFUZZ = True
except ImportError:
    HAS_RAPIDFUZZ = False


# =====================================================================
# CONSTANTS
# =====================================================================

ECOMMERCE_RAW_PATH = PROJECT_ROOT / "data" / "output" / "phase06" / "ecommerce_raw.json"
MERCOS_CARTEIRA_PATH = PROJECT_ROOT / "data" / "sources" / "mercos" / "08_CARTEIRA_MERCOS.xlsx"
DRAFT1_PATH = PROJECT_ROOT / "data" / "sources" / "drafts" / "DL_DRAFT1_FEV2026.xlsx"
SAP_MERCOS_MERGED_PATH = PROJECT_ROOT / "data" / "output" / "phase02" / "sap_mercos_merged.json"
OUTPUT_MATCHED_PATH = PROJECT_ROOT / "data" / "output" / "phase06" / "ecommerce_matched.json"
OUTPUT_REPORT_PATH = PROJECT_ROOT / "data" / "output" / "phase06" / "match_report.json"


# =====================================================================
# UTILITY FUNCTIONS
# =====================================================================

def normalize_cnpj(raw):
    """Normalize CNPJ/CPF to 14-digit zero-padded string."""
    return re.sub(r'[^0-9]', '', str(raw)).zfill(14)


def safe_float(val, default=0.0):
    """Safely convert to float."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    """Safely convert to int."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def load_json(path):
    """Load a JSON file and return parsed data."""
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_json(data, path):
    """Save data to JSON file."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def last_day_of_month(year_month_str):
    """Given 'YYYY-MM', return 'YYYY-MM-DD' for the last day of the month."""
    parts = year_month_str.split("-")
    year = int(parts[0])
    month = int(parts[1])
    last_day = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-{last_day:02d}"


# =====================================================================
# BUILD LOOKUPS
# =====================================================================

def build_mercos_lookup():
    """Build name->CNPJ lookup from Mercos Carteira.

    Returns dict: uppercase_name -> (cnpj_14digits, source_field)
    Uses both Razao Social (col A) and Nome Fantasia (col B) as keys.
    CNPJ/CPF is in col C.
    """
    wb = openpyxl.load_workbook(str(MERCOS_CARTEIRA_PATH), read_only=True, data_only=True)
    ws = wb["Carteira Clientes Mercos"]

    lookup = {}
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        razao = str(row[0] or '').strip().upper()
        fantasia = str(row[1] or '').strip().upper()
        cnpj_raw = str(row[2] or '').strip()

        if cnpj_raw and cnpj_raw.lower() != 'none':
            digits = re.sub(r'[^0-9]', '', cnpj_raw)
            if len(digits) >= 11:  # valid CNPJ or CPF
                cnpj = digits.zfill(14)
                count += 1
                if razao and razao != 'NONE':
                    lookup[razao] = (cnpj, "mercos_razao_social")
                if fantasia and fantasia != 'NONE':
                    lookup[fantasia] = (cnpj, "mercos_nome_fantasia")

    wb.close()
    print(f"  Mercos Carteira: {count} clientes, {len(lookup)} lookup entries")
    return lookup


def build_draft1_lookup():
    """Build name->CNPJ lookup from DRAFT 1.

    Returns dict: uppercase_name -> (cnpj_14digits, source_field)
    CNPJ is in col 2, Nome Fantasia in col 1, Razao Social in col 3.
    Data rows start at row 4.
    """
    wb = openpyxl.load_workbook(str(DRAFT1_PATH), read_only=True, data_only=True)
    ws = wb.active

    lookup = {}
    count = 0
    for r in range(4, ws.max_row + 1):
        cnpj_raw = str(ws.cell(row=r, column=2).value or '').strip()
        nome_fantasia = str(ws.cell(row=r, column=1).value or '').strip().upper()
        razao_social = str(ws.cell(row=r, column=3).value or '').strip().upper()

        if cnpj_raw and cnpj_raw.lower() != 'none':
            digits = re.sub(r'[^0-9]', '', cnpj_raw)
            if len(digits) >= 11:
                cnpj = digits.zfill(14)
                count += 1
                if nome_fantasia and nome_fantasia != 'NONE':
                    lookup[nome_fantasia] = (cnpj, "draft1_nome_fantasia")
                if razao_social and razao_social != 'NONE':
                    # Strip CNPJ prefix from razao_social if present
                    # Pattern: "51.172.110 BRENDHA EVELYN..."
                    razao_clean = re.sub(r'^[\d.]+\s+', '', razao_social).strip()
                    if razao_clean:
                        lookup[razao_clean] = (cnpj, "draft1_razao_social")
                    lookup[razao_social] = (cnpj, "draft1_razao_social_full")

    wb.close()
    print(f"  DRAFT 1: {count} clientes, {len(lookup)} lookup entries")
    return lookup


def build_sap_lookup():
    """Build name->CNPJ lookup from SAP Cadastro.

    Returns dict: uppercase_name -> (cnpj_14digits, source_field)
    SAP Consolidado: col E (5) = CNPJ, col F (6) = Nome Cliente
    """
    sap_path = PROJECT_ROOT / "data" / "sources" / "sap" / "01_SAP_CONSOLIDADO.xlsx"
    wb = openpyxl.load_workbook(str(sap_path), read_only=True, data_only=True)
    ws = wb["Cadastro Clientes SAP"]

    lookup = {}
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cnpj_raw = str(row[4] or '').strip()  # col E = CNPJ
        nome = str(row[5] or '').strip().upper()  # col F = Nome

        if cnpj_raw and cnpj_raw.lower() != 'none':
            digits = re.sub(r'[^0-9]', '', cnpj_raw)
            if len(digits) >= 11 and nome and nome != 'NONE':
                cnpj = digits.zfill(14)
                count += 1
                lookup[nome] = (cnpj, "sap_nome_cliente")

    wb.close()
    print(f"  SAP Cadastro: {count} clientes, {len(lookup)} lookup entries")
    return lookup


def load_valid_cnpjs():
    """Load set of valid CNPJs from sap_mercos_merged.json for validation."""
    data = load_json(SAP_MERCOS_MERGED_PATH)
    cnpjs = set(data.get("cnpj_to_vendas", {}).keys())
    print(f"  SAP/Mercos merged: {len(cnpjs)} valid CNPJs for validation")
    return cnpjs, data


# Company suffixes to strip for normalized matching
COMPANY_SUFFIXES = [
    ' LTDA.', ' LTDA', ' EIRELI', ' ME', ' EPP', ' S/A', ' S.A.', ' SA',
    ' - ME', ' - EPP', ' - LTDA', ' - EIRELI',
    ' EM RECUPERACAO JUDICIAL',
]


def normalize_company_name(name):
    """Normalize company name by stripping common suffixes."""
    n = name.strip().upper()
    changed = True
    while changed:
        changed = False
        for suf in COMPANY_SUFFIXES:
            if n.endswith(suf):
                n = n[:-len(suf)].strip()
                changed = True
    return n


def build_cnpj_prefix_lookup(merged_lookup):
    """Build a CNPJ-first-8-digits -> full CNPJ lookup from all known CNPJs.

    This allows matching e-commerce names that have a XX.XXX.XXX prefix
    (8-digit CNPJ base) to their full 14-digit CNPJ.
    """
    prefix8_to_cnpj = {}
    seen_cnpjs = set()
    for key, (cnpj, source) in merged_lookup.items():
        if cnpj not in seen_cnpjs:
            seen_cnpjs.add(cnpj)
            p8 = cnpj[:8]
            # Only keep first occurrence to avoid conflicts
            if p8 not in prefix8_to_cnpj:
                prefix8_to_cnpj[p8] = cnpj
    return prefix8_to_cnpj


# =====================================================================
# MATCHING ENGINE
# =====================================================================

def try_match_name(razao_social, nome_fantasia, cnpj_from_name,
                   merged_lookup, prefix8_lookup, fuzzy_choices=None):
    """Try to match e-commerce record to CNPJ using 5 levels.

    Args:
        razao_social: Razao social from e-commerce report
        nome_fantasia: Nome fantasia from e-commerce report
        cnpj_from_name: CNPJ extracted from name field (may be None)
        merged_lookup: Combined name->CNPJ dict
        prefix8_lookup: 8-digit CNPJ prefix -> full CNPJ dict
        fuzzy_choices: List of lookup keys for fuzzy matching (if rapidfuzz available)

    Returns: (cnpj, match_level, match_detail) or (None, None, reason)
    """
    razao_up = str(razao_social or '').strip().upper()
    fantasia_up = str(nome_fantasia or '').strip().upper()

    # Level 1a: CNPJ from prefix field (>= 11 digits = direct CNPJ)
    if cnpj_from_name:
        digits = re.sub(r'[^0-9]', '', str(cnpj_from_name))
        if len(digits) >= 11:
            cnpj = digits.zfill(14)
            return cnpj, "cnpj_prefix", f"CNPJ extracted from name field: {cnpj_from_name}"

    # Level 1b: CNPJ prefix from razao_social (XX.XXX.XXX NAME pattern)
    # Match 8-digit CNPJ base to known CNPJs via prefix8_lookup
    prefix_match = re.match(r'^(\d{2}\.\d{3}\.\d{3})\s+', razao_up)
    if prefix_match:
        p8 = re.sub(r'[^0-9]', '', prefix_match.group(1))
        if p8 in prefix8_lookup:
            cnpj = prefix8_lookup[p8]
            return cnpj, "cnpj_prefix", f"CNPJ prefix {p8} matched to {cnpj}"

    # Level 1c: CNPJ from cnpj_from_name field (8 digits = CNPJ base, match via prefix)
    if cnpj_from_name:
        digits = re.sub(r'[^0-9]', '', str(cnpj_from_name))
        if len(digits) >= 8 and digits in prefix8_lookup:
            cnpj = prefix8_lookup[digits]
            return cnpj, "cnpj_prefix", f"CNPJ prefix from name field: {digits} -> {cnpj}"

    # Level 2: Exact match by Razao Social (full and stripped prefix)
    if razao_up and razao_up in merged_lookup:
        cnpj, source = merged_lookup[razao_up]
        return cnpj, "exact", f"Razao Social exact match ({source})"

    # Try stripped razao (remove XX.XXX.XXX prefix)
    razao_stripped = re.sub(r'^[\d.]+\s+', '', razao_up).strip()
    if razao_stripped and razao_stripped != razao_up and razao_stripped in merged_lookup:
        cnpj, source = merged_lookup[razao_stripped]
        return cnpj, "exact", f"Razao Social stripped exact match ({source})"

    # Level 3: Exact match by Nome Fantasia
    if fantasia_up and fantasia_up in merged_lookup:
        cnpj, source = merged_lookup[fantasia_up]
        return cnpj, "exact", f"Nome Fantasia exact match ({source})"

    # Level 3b: Normalized name matching (strip LTDA etc)
    razao_norm = normalize_company_name(razao_stripped or razao_up)
    fantasia_norm = normalize_company_name(fantasia_up)

    if razao_norm and razao_norm != razao_up and razao_norm in merged_lookup:
        cnpj, source = merged_lookup[razao_norm]
        return cnpj, "exact", f"Razao Social normalized exact match ({source})"

    if fantasia_norm and fantasia_norm != fantasia_up and fantasia_norm in merged_lookup:
        cnpj, source = merged_lookup[fantasia_norm]
        return cnpj, "exact", f"Nome Fantasia normalized exact match ({source})"

    # Level 4a: Partial match (substring, min 5 chars, prefer longest match)
    best_partial = None
    best_partial_len = 0

    names_to_try = []
    if razao_stripped and len(razao_stripped) >= 5:
        names_to_try.append(("razao_social", razao_stripped))
    elif razao_up and len(razao_up) >= 5:
        names_to_try.append(("razao_social", razao_up))
    if fantasia_up and len(fantasia_up) >= 5:
        names_to_try.append(("nome_fantasia", fantasia_up))

    for name_field, name_val in names_to_try:
        for key, (cnpj, source) in merged_lookup.items():
            if len(key) < 5:
                continue
            if name_val in key or key in name_val:
                match_len = min(len(name_val), len(key))
                if match_len > best_partial_len:
                    best_partial = (cnpj, source, key, name_field)
                    best_partial_len = match_len

    if best_partial:
        cnpj, source, matched_key, name_field = best_partial
        return cnpj, "partial", f"{name_field} partial match in {source}: '{matched_key}'"

    # Level 4b: Normalized partial match (using pre-computed normalized names in lookup)
    # Already handled since normalized names are added to merged_lookup in run_matching()
    # Also try with normalized input names
    names_norm = []
    if razao_norm and len(razao_norm) >= 5 and razao_norm not in [n for _, n in names_to_try]:
        names_norm.append(("razao_social_norm", razao_norm))
    if fantasia_norm and len(fantasia_norm) >= 5 and fantasia_norm not in [n for _, n in names_to_try]:
        names_norm.append(("nome_fantasia_norm", fantasia_norm))

    for name_field, name_val in names_norm:
        for key, (cnpj, source) in merged_lookup.items():
            if len(key) < 5:
                continue
            if name_val in key or key in name_val:
                match_len = min(len(name_val), len(key))
                if match_len > best_partial_len:
                    best_partial = (cnpj, source, key, name_field)
                    best_partial_len = match_len

    if best_partial:
        cnpj, source, matched_key, name_field = best_partial
        return cnpj, "partial", f"{name_field} partial match in {source}: '{matched_key}'"

    # Level 5: Fuzzy match (if rapidfuzz available)
    if HAS_RAPIDFUZZ and fuzzy_choices:
        for name_field, name_val in names_to_try:
            if not name_val:
                continue
            result = rf_process.extractOne(
                name_val, fuzzy_choices,
                scorer=fuzz.token_sort_ratio,
                score_cutoff=85
            )
            if result:
                matched_key, score, _ = result
                cnpj, source = merged_lookup[matched_key]
                return cnpj, "fuzzy", f"{name_field} fuzzy match ({score:.0f}%) in {source}: '{matched_key}'"

    # No match found
    reasons = []
    if razao_up:
        reasons.append(f"razao='{razao_up}'")
    if fantasia_up:
        reasons.append(f"fantasia='{fantasia_up}'")
    return None, None, f"No match in any lookup ({', '.join(reasons)})"


# =====================================================================
# AGGREGATE BY CNPJ
# =====================================================================

def aggregate_by_cnpj(matched_records, sap_mercos_data):
    """Aggregate matched e-commerce data by CNPJ.

    Args:
        matched_records: list of (cnpj, record, month) tuples
        sap_mercos_data: dict from sap_mercos_merged.json for pct_ecommerce calc

    Returns: dict cnpj -> aggregated e-commerce data
    """
    # Group records by CNPJ
    cnpj_records = defaultdict(list)
    for cnpj, record, month in matched_records:
        cnpj_records[cnpj].append((month, record))

    cnpj_to_ecommerce = {}

    for cnpj, month_records in cnpj_records.items():
        # Build monthly data
        monthly_data = {}
        all_months_with_access = []
        total_acessos = 0
        total_atividades = 0
        total_itens = 0
        total_orcamentos = 0
        total_valor_carrinho = 0.0
        total_valor_b2b = 0.0
        latest_month_with_access = None

        for month, rec in month_records:
            acessos = safe_int(rec.get("acessos", 0))
            itens = safe_int(rec.get("itens_adicionados", 0))
            orcamentos = safe_int(rec.get("orcamentos_finalizados", 0))
            atividades = safe_int(rec.get("atividades", 0))
            valor_carrinho = safe_float(rec.get("valor_carrinho", 0))
            valor_orc_nf = safe_float(rec.get("valor_orc_nao_finalizados", 0))
            valor_b2b = safe_float(rec.get("valor_pedidos_b2b", 0))

            monthly_data[month] = {
                "acessos": acessos,
                "itens": itens,
                "orcamentos": orcamentos,
                "atividades": atividades,
                "valor_carrinho": valor_carrinho,
                "valor_orc_nf": valor_orc_nf,
                "valor_b2b": valor_b2b,
            }

            total_acessos += acessos
            total_atividades += atividades
            total_itens += itens
            total_orcamentos += orcamentos
            total_valor_carrinho += valor_carrinho
            total_valor_b2b += valor_b2b

            if acessos > 0:
                all_months_with_access.append(month)
                if latest_month_with_access is None or month > latest_month_with_access:
                    latest_month_with_access = month

        # Determine most recent month overall for qtd_acessos_ultimo_mes
        all_months_sorted = sorted(monthly_data.keys())
        most_recent_month = all_months_sorted[-1] if all_months_sorted else None
        qtd_acessos_ultimo_mes = monthly_data[most_recent_month]["acessos"] if most_recent_month else 0

        # acessa_ecommerce: true if any month had acessos > 0
        acessa_ecommerce = total_acessos > 0

        # data_ult_acesso: last day of latest month with access
        data_ult_acesso = None
        if latest_month_with_access:
            data_ult_acesso = last_day_of_month(latest_month_with_access)

        # pedido_via_ecommerce: true if any month had valor_pedidos_b2b > 0
        pedido_via_ecommerce = total_valor_b2b > 0

        # catalogo_visualizado: true if any month had atividades > 0
        catalogo_visualizado = total_atividades > 0

        # meses_com_acesso: sorted list
        meses_com_acesso = sorted(set(all_months_with_access))

        # pct_pedidos_ecommerce from SAP/Mercos merged
        pct_pedidos_ecommerce = None
        cnpj_vendas = sap_mercos_data.get("cnpj_to_vendas", {})
        if cnpj in cnpj_vendas:
            vendas_list = cnpj_vendas[cnpj]
            fat_total = sum(safe_float(v) for v in vendas_list)
            if fat_total > 0:
                pct_pedidos_ecommerce = round(total_valor_b2b / fat_total, 4)

        cnpj_to_ecommerce[cnpj] = {
            "acessa_ecommerce": acessa_ecommerce,
            "data_ult_acesso": data_ult_acesso,
            "qtd_acessos_ultimo_mes": qtd_acessos_ultimo_mes,
            "pedido_via_ecommerce": pedido_via_ecommerce,
            "pct_pedidos_ecommerce": pct_pedidos_ecommerce,
            "valor_pedidos_b2b_total": round(total_valor_b2b, 2),
            "catalogo_visualizado": catalogo_visualizado,
            "meses_com_acesso": meses_com_acesso,
            "total_atividades": total_atividades,
            "total_itens_carrinho": total_itens,
            "total_orcamentos": total_orcamentos,
            "valor_carrinho_total": round(total_valor_carrinho, 2),
            "monthly_data": dict(sorted(monthly_data.items())),
        }

    return cnpj_to_ecommerce


# =====================================================================
# MAIN: MATCH + AGGREGATE
# =====================================================================

def run_matching():
    """Run full matching pipeline: load data, match, aggregate, output JSON."""

    print("=" * 70)
    print("PHASE 06-02: E-COMMERCE MATCH + AGGREGATE")
    print("=" * 70)
    print()

    # 1. Load ecommerce_raw.json
    print("[1/7] Loading ecommerce_raw.json...")
    raw_data = load_json(ECOMMERCE_RAW_PATH)
    monthly_data = raw_data["monthly_data"]
    months_covered = sorted(monthly_data.keys())
    total_records = sum(len(records) for records in monthly_data.values())
    print(f"  Months: {len(months_covered)} ({', '.join(months_covered)})")
    print(f"  Total records: {total_records}")
    print()

    # 2. Build lookup tables
    print("[2/7] Building name->CNPJ lookups...")
    mercos_lookup = build_mercos_lookup()
    draft1_lookup = build_draft1_lookup()
    sap_lookup = build_sap_lookup()
    valid_cnpjs, sap_mercos_data = load_valid_cnpjs()
    print()

    # Merge lookups (priority: Mercos > DRAFT 1 > SAP)
    merged_lookup = {}
    merged_lookup.update(sap_lookup)     # SAP as lowest priority
    merged_lookup.update(draft1_lookup)  # DRAFT 1 overwrites SAP
    merged_lookup.update(mercos_lookup)  # Mercos overwrites (preferred source)

    # Also add normalized versions of all names to improve matching
    normalized_additions = {}
    for key, val in list(merged_lookup.items()):
        norm = normalize_company_name(key)
        if norm and norm != key and norm not in merged_lookup:
            normalized_additions[norm] = val
    merged_lookup.update(normalized_additions)

    print(f"  Merged lookup: {len(merged_lookup)} unique name entries (incl. {len(normalized_additions)} normalized)")

    # Build CNPJ prefix-8 lookup for Level 1b/1c matching
    prefix8_lookup = build_cnpj_prefix_lookup(merged_lookup)
    print(f"  CNPJ prefix-8 lookup: {len(prefix8_lookup)} unique prefixes")

    # Prepare fuzzy choices if rapidfuzz available
    fuzzy_choices = list(merged_lookup.keys()) if HAS_RAPIDFUZZ else None
    if HAS_RAPIDFUZZ:
        print(f"  Fuzzy matching: ENABLED (rapidfuzz available)")
    else:
        print(f"  Fuzzy matching: DISABLED (rapidfuzz not installed)")
    print()

    # 3. Collect unique names from e-commerce data
    print("[3/7] Collecting unique e-commerce names...")
    unique_names = {}  # key = (razao_up, fantasia_up) -> list of (month, record)
    for month, records in monthly_data.items():
        for rec in records:
            razao = str(rec.get("razao_social", "") or "").strip()
            fantasia = str(rec.get("nome_fantasia", "") or "").strip()
            key = (razao.upper(), fantasia.upper())
            if key not in unique_names:
                unique_names[key] = {
                    "razao_social": razao,
                    "nome_fantasia": fantasia,
                    "cnpj_from_name": rec.get("cnpj_from_name"),
                    "months": [],
                    "records": [],
                }
            unique_names[key]["months"].append(month)
            unique_names[key]["records"].append((month, rec))

    print(f"  Unique name combinations: {len(unique_names)}")
    print()

    # 4. Match each unique name
    print("[4/7] Matching names to CNPJs...")
    matched_records = []  # (cnpj, record, month)
    match_details = []  # for match_report
    failure_details = []  # for match_report

    level_counts = {"cnpj_prefix": 0, "exact": 0, "partial": 0, "fuzzy": 0}
    matched_count = 0
    unmatched_count = 0

    for (razao_up, fantasia_up), name_info in unique_names.items():
        cnpj, level, detail = try_match_name(
            name_info["razao_social"],
            name_info["nome_fantasia"],
            name_info["cnpj_from_name"],
            merged_lookup,
            prefix8_lookup,
            fuzzy_choices,
        )

        if cnpj:
            matched_count += 1
            level_counts[level] += 1
            match_details.append({
                "name": name_info["razao_social"],
                "nome_fantasia": name_info["nome_fantasia"],
                "cnpj": cnpj,
                "level": level,
                "detail": detail,
                "months_seen": len(name_info["months"]),
            })
            # Add all records for this name to matched
            for month, rec in name_info["records"]:
                matched_records.append((cnpj, rec, month))
        else:
            unmatched_count += 1
            total_acessos = sum(safe_int(r.get("acessos", 0)) for _, r in name_info["records"])
            failure_details.append({
                "razao_social": name_info["razao_social"],
                "nome_fantasia": name_info["nome_fantasia"],
                "months_seen": len(name_info["months"]),
                "total_acessos": total_acessos,
                "reason": detail,
            })

    total_unique_names = len(unique_names)
    match_rate = matched_count / total_unique_names if total_unique_names > 0 else 0

    print(f"  Matched: {matched_count} / {total_unique_names} ({match_rate:.1%})")
    print(f"  By level: cnpj_prefix={level_counts['cnpj_prefix']}, exact={level_counts['exact']}, partial={level_counts['partial']}, fuzzy={level_counts['fuzzy']}")
    print(f"  Unmatched: {unmatched_count}")
    print()

    # 5. Aggregate by CNPJ
    print("[5/7] Aggregating by CNPJ...")
    cnpj_to_ecommerce = aggregate_by_cnpj(matched_records, sap_mercos_data)
    print(f"  Unique CNPJs with e-commerce data: {len(cnpj_to_ecommerce)}")
    print()

    # 6. Build unmatched list (sorted by total_acessos desc)
    unmatched_list = sorted(failure_details, key=lambda x: x["total_acessos"], reverse=True)

    # 7. Output ecommerce_matched.json
    print("[6/7] Writing ecommerce_matched.json...")
    matched_output = {
        "metadata": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "total_ecommerce_names": total_unique_names,
            "matched": matched_count,
            "unmatched": unmatched_count,
            "match_rate": round(match_rate, 4),
            "unique_cnpjs_with_ecommerce": len(cnpj_to_ecommerce),
            "months_covered": months_covered,
        },
        "cnpj_to_ecommerce": dict(sorted(cnpj_to_ecommerce.items())),
        "unmatched": unmatched_list,
    }
    save_json(matched_output, OUTPUT_MATCHED_PATH)
    print(f"  Saved: {OUTPUT_MATCHED_PATH}")
    print()

    # 8. Output match_report.json
    print("[7/7] Writing match_report.json...")
    match_report = {
        "summary": {
            "total_unique_names": total_unique_names,
            "matched": matched_count,
            "unmatched": unmatched_count,
            "match_rate": round(match_rate, 4),
            "by_level": level_counts,
        },
        "matches": match_details,
        "failures": failure_details,
    }
    save_json(match_report, OUTPUT_REPORT_PATH)
    print(f"  Saved: {OUTPUT_REPORT_PATH}")
    print()

    # Print report
    print("=" * 70)
    print("MATCHING REPORT")
    print("=" * 70)
    print(f"  Total unique names:     {total_unique_names}")
    print(f"  Matched:                {matched_count} ({match_rate:.1%})")
    print(f"  Unmatched:              {unmatched_count}")
    print()
    print("  Breakdown by level:")
    print(f"    cnpj_prefix:          {level_counts['cnpj_prefix']}")
    print(f"    exact:                {level_counts['exact']}")
    print(f"    partial:              {level_counts['partial']}")
    print(f"    fuzzy:                {level_counts['fuzzy']}")
    print()
    print(f"  Unique CNPJs:           {len(cnpj_to_ecommerce)}")
    print(f"  CNPJs in SAP/Mercos:    {sum(1 for c in cnpj_to_ecommerce if c in valid_cnpjs)}")
    print()

    # Top 10 unmatched
    if unmatched_list:
        print("  Top 10 Unmatched (by total acessos):")
        for i, u in enumerate(unmatched_list[:10], 1):
            print(f"    {i:2d}. {u['razao_social']} ({u['nome_fantasia']}) - {u['total_acessos']} acessos, {u['months_seen']} months")
        print()

    # Aggregation stats
    with_pedidos = sum(1 for v in cnpj_to_ecommerce.values() if v["pedido_via_ecommerce"])
    with_pct = sum(1 for v in cnpj_to_ecommerce.values() if v["pct_pedidos_ecommerce"] is not None)
    print(f"  Aggregation stats:")
    print(f"    CNPJs com pedido_via_ecommerce: {with_pedidos}")
    print(f"    CNPJs com pct_pedidos_ecommerce calculado: {with_pct}")
    print(f"    Total valor_pedidos_b2b: R$ {sum(v['valor_pedidos_b2b_total'] for v in cnpj_to_ecommerce.values()):,.2f}")
    print()

    return matched_output, match_report


# =====================================================================
# POPULATE DRAFT 1
# =====================================================================

def populate_draft1(cnpj_to_ecommerce):
    """Populate DRAFT 1 columns 15-20 with e-commerce data.

    Column mapping:
      15 ACESSOS SEMANA  -> qtd_acessos_ultimo_mes
      16 ACESSO B2B      -> "Sim" if acessa_ecommerce else None
      17 ACESSOS PORTAL  -> total_atividades
      18 ITENS CARRINHO  -> total_itens_carrinho
      19 VALOR B2B       -> valor_pedidos_b2b_total
      20 OPORTUNIDADE    -> "COMPRA B2B" if pedido_via_ecommerce else None

    CNPJ is in column 2 of DRAFT 1.
    Data rows: 4 to max_row.
    """

    print("=" * 70)
    print("POPULATE DRAFT 1 COLUMNS 15-20")
    print("=" * 70)
    print()

    # Open DRAFT 1 in write mode (data_only=False to preserve formulas)
    print("  Loading DRAFT 1...")
    wb = openpyxl.load_workbook(str(DRAFT1_PATH), data_only=False)
    ws = wb.active

    # Verify headers
    expected_headers = {
        15: "ACESSOS SEMANA",
        16: "ACESSO B2B",
        17: "ACESSOS PORTAL",
        18: "ITENS CARRINHO",
        19: "VALOR B2B",
        20: "OPORTUNIDADE",
    }
    print("  Verifying headers (row 3):")
    headers_ok = True
    for col, expected in expected_headers.items():
        actual = ws.cell(row=3, column=col).value
        status = "OK" if actual == expected else f"MISMATCH (got '{actual}')"
        if actual != expected:
            headers_ok = False
        print(f"    Col {col}: '{actual}' - {status}")

    if not headers_ok:
        print("  WARNING: Headers don't match exactly, but proceeding with column positions")
    print()

    # Process data rows
    total_rows = 0
    populated = 0
    skipped = 0
    spot_checks = []

    max_row = ws.max_row
    print(f"  Processing rows 4 to {max_row}...")

    for r in range(4, max_row + 1):
        cnpj_raw = ws.cell(row=r, column=2).value
        if not cnpj_raw:
            continue

        total_rows += 1
        cnpj = normalize_cnpj(cnpj_raw)

        if cnpj in cnpj_to_ecommerce:
            ecom = cnpj_to_ecommerce[cnpj]
            populated += 1

            # Col 15: ACESSOS SEMANA -> qtd_acessos_ultimo_mes
            ws.cell(row=r, column=15).value = ecom["qtd_acessos_ultimo_mes"] if ecom["qtd_acessos_ultimo_mes"] > 0 else None

            # Col 16: ACESSO B2B -> "Sim" if acessa_ecommerce
            ws.cell(row=r, column=16).value = "Sim" if ecom["acessa_ecommerce"] else None

            # Col 17: ACESSOS PORTAL -> total_atividades
            ws.cell(row=r, column=17).value = ecom["total_atividades"] if ecom["total_atividades"] > 0 else None

            # Col 18: ITENS CARRINHO -> total_itens_carrinho
            ws.cell(row=r, column=18).value = ecom["total_itens_carrinho"] if ecom["total_itens_carrinho"] > 0 else None

            # Col 19: VALOR B2B -> valor_pedidos_b2b_total
            ws.cell(row=r, column=19).value = ecom["valor_pedidos_b2b_total"] if ecom["valor_pedidos_b2b_total"] > 0 else None

            # Col 20: OPORTUNIDADE -> "COMPRA B2B" if pedido_via_ecommerce
            ws.cell(row=r, column=20).value = "COMPRA B2B" if ecom["pedido_via_ecommerce"] else None

            # Collect spot checks (first 3)
            if len(spot_checks) < 3:
                spot_checks.append({
                    "row": r,
                    "cnpj": cnpj,
                    "col15": ws.cell(row=r, column=15).value,
                    "col16": ws.cell(row=r, column=16).value,
                    "col17": ws.cell(row=r, column=17).value,
                    "col18": ws.cell(row=r, column=18).value,
                    "col19": ws.cell(row=r, column=19).value,
                    "col20": ws.cell(row=r, column=20).value,
                })
        else:
            skipped += 1
            # Clear e-commerce columns for clients not in e-commerce data
            # so stale Mercos data doesn't persist
            for col in range(15, 21):
                ws.cell(row=r, column=col).value = None

    # Save
    print(f"  Saving DRAFT 1...")
    wb.save(str(DRAFT1_PATH))
    wb.close()
    print(f"  Saved: {DRAFT1_PATH}")
    print()

    # Report
    print(f"  DRAFT 1 Population Report:")
    print(f"    Total data rows:            {total_rows}")
    print(f"    Rows with e-commerce data:  {populated}")
    print(f"    Rows without match:         {skipped}")
    print()

    # Spot checks
    if spot_checks:
        print(f"  Spot checks ({len(spot_checks)} rows):")
        for sc in spot_checks:
            print(f"    Row {sc['row']}: CNPJ={sc['cnpj']}")
            print(f"      Col 15 (ACESSOS SEMANA): {sc['col15']}")
            print(f"      Col 16 (ACESSO B2B): {sc['col16']}")
            print(f"      Col 17 (ACESSOS PORTAL): {sc['col17']}")
            print(f"      Col 18 (ITENS CARRINHO): {sc['col18']}")
            print(f"      Col 19 (VALOR B2B): {sc['col19']}")
            print(f"      Col 20 (OPORTUNIDADE): {sc['col20']}")
            print()

    return populated, skipped, total_rows


# =====================================================================
# MAIN
# =====================================================================

if __name__ == "__main__":
    # Phase 1: Match + Aggregate
    matched_output, match_report = run_matching()

    # Phase 2: Populate DRAFT 1
    cnpj_to_ecommerce = matched_output["cnpj_to_ecommerce"]
    populated, skipped, total = populate_draft1(cnpj_to_ecommerce)

    # Final summary
    print("=" * 70)
    print("PHASE 06-02 COMPLETE")
    print("=" * 70)
    print(f"  Match rate: {matched_output['metadata']['match_rate']:.1%}")
    print(f"  Unique CNPJs with e-commerce: {matched_output['metadata']['unique_cnpjs_with_ecommerce']}")
    print(f"  DRAFT 1 rows populated: {populated} / {total}")
    print(f"  Output files:")
    print(f"    - {OUTPUT_MATCHED_PATH}")
    print(f"    - {OUTPUT_REPORT_PATH}")
    print(f"    - {DRAFT1_PATH}")
    print("=" * 70)

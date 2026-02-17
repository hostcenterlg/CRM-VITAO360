"""
Phase 05 - Plan 01: Normalize LOG TIPO DO CONTATO in V13.

Maps the 12 inconsistent TIPO DO CONTATO values in V13 LOG tab
down to 7 canonical types required by DASH COUNTIFS formulas.

Preserves all 20,830 LOG records and 19,224 PROJECAO formulas.
"""

import shutil
import sys
from collections import Counter
from pathlib import Path

# Paths
BASE = Path(__file__).resolve().parent.parent.parent
DATA_OUT = BASE / "data" / "output"
V13_PATH = DATA_OUT / "CRM_VITAO360_V13_PROJECAO.xlsx"
V13_BACKUP = DATA_OUT / "CRM_VITAO360_V13_PROJECAO_BACKUP_PHASE05_01.xlsx"

# Python executable hint for reference
PYTHON = "/c/Users/User/.pyenv/pyenv-win/pyenv-win/versions/3.12.10/python.exe"

# ============================================================================
# NORMALIZATION MAP
# Maps all 12 known TIPO DO CONTATO values to 7 canonical types.
# ============================================================================
TIPO_NORMALIZATION = {
    # Already canonical (no change needed):
    "POS-VENDA / RELACIONAMENTO": "POS-VENDA / RELACIONAMENTO",
    "NEGOCIACAO": "NEGOCIACAO",
    "PERDA / NUTRICAO": "PERDA / NUTRICAO",
    "FOLLOW UP": "FOLLOW UP",
    "ATEND. CLIENTES ATIVOS": "ATEND. CLIENTES ATIVOS",
    "ATENDIMENTO CLIENTES INATIVOS": "ATEND. CLIENTES INATIVOS",
    # Merge variants:
    "POS VENDA / RELACIONAMENTO": "POS-VENDA / RELACIONAMENTO",       # ~565 records, missing hyphen
    "ATENDIMENTO CLIENTES ATIVOS": "ATEND. CLIENTES ATIVOS",          # ~1,780 records, abbreviated form
    "PROSPECCAO NOVOS CLIENTES": "PROSPECCAO",                         # ~3,932 records, rename to short form
    "PROSPECCAO": "PROSPECCAO",                                        # ~30 records, already short
    # Map non-canonical types to closest match:
    "CONTATOS PASSIVO / SUPORTE": "POS-VENDA / RELACIONAMENTO",       # ~2,258 records, passive contacts = post-sale
    "ENVIO DE MATERIAL - MKT": "PROSPECCAO",                          # ~672 records, marketing material = prospection
}

# Expected 7 canonical values after normalization
CANONICAL_TYPES = {
    "POS-VENDA / RELACIONAMENTO",
    "NEGOCIACAO",
    "PERDA / NUTRICAO",
    "FOLLOW UP",
    "ATEND. CLIENTES ATIVOS",
    "ATEND. CLIENTES INATIVOS",
    "PROSPECCAO",
}


def find_sheet_by_prefix(wb, prefix):
    """Find a sheet whose name (stripped of accents) contains prefix."""
    import unicodedata
    for name in wb.sheetnames:
        # Strip accents for comparison
        nfkd = unicodedata.normalize('NFKD', name)
        ascii_name = ''.join(c for c in nfkd if not unicodedata.combining(c))
        if prefix.upper() in ascii_name.upper():
            return wb[name]
    return None


def count_formulas(ws):
    """Count cells that contain formulas (start with '=')."""
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith('='):
                count += 1
    return count


def main():
    import openpyxl

    print("=" * 70)
    print("Phase 05 Plan 01: Normalize LOG TIPO DO CONTATO")
    print("=" * 70)

    # ── Step 1: Validate V13 exists ──
    if not V13_PATH.exists():
        print(f"ERRO: V13 nao encontrado em {V13_PATH}")
        sys.exit(1)

    # ── Step 2: Backup V13 ──
    print(f"\n[1/9] Creating backup...")
    shutil.copy2(V13_PATH, V13_BACKUP)
    print(f"  Backup: {V13_BACKUP.name}")

    # ── Step 3: Load V13 (preserve formulas) ──
    print(f"\n[2/9] Loading V13 (data_only=False)...")
    wb = openpyxl.load_workbook(str(V13_PATH), data_only=False)
    print(f"  Sheets: {list(wb.sheetnames)}")

    # ── Step 4: Find LOG tab ──
    print(f"\n[3/9] Finding LOG tab...")
    ws_log = None
    if 'LOG' in wb.sheetnames:
        ws_log = wb['LOG']
    else:
        ws_log = find_sheet_by_prefix(wb, 'LOG')

    if ws_log is None:
        print("  ERRO: LOG tab not found!")
        sys.exit(1)
    print(f"  Found: '{ws_log.title}'")

    # ── Step 5: Count PROJECAO formulas (pre-save) ──
    print(f"\n[4/9] Counting PROJECAO formulas (pre-save)...")
    proj_ws = find_sheet_by_prefix(wb, 'PROJE')
    if proj_ws:
        formulas_before = count_formulas(proj_ws)
        print(f"  PROJECAO formulas: {formulas_before}")
    else:
        formulas_before = 0
        print("  WARNING: PROJECAO sheet not found!")

    # ── Step 6: Read current TIPO distribution ──
    print(f"\n[5/9] Reading current TIPO DO CONTATO distribution...")
    TIPO_COL = 12  # Column L
    DATA_START_ROW = 3
    # Find actual last data row
    max_row = ws_log.max_row
    print(f"  max_row reported: {max_row}")

    before_counter = Counter()
    total_rows = 0
    empty_rows = 0
    for row_num in range(DATA_START_ROW, max_row + 1):
        cell_value = ws_log.cell(row=row_num, column=TIPO_COL).value
        if cell_value is None or str(cell_value).strip() == '':
            empty_rows += 1
            continue
        val = str(cell_value).strip()
        before_counter[val] += 1
        total_rows += 1

    print(f"\n  BEFORE normalization ({len(before_counter)} unique values, {total_rows} data rows):")
    for tipo, cnt in sorted(before_counter.items(), key=lambda x: -x[1]):
        marker = " [CANONICAL]" if tipo in CANONICAL_TYPES else " [TO NORMALIZE]"
        print(f"    {cnt:>6,}  {tipo}{marker}")
    if empty_rows > 0:
        print(f"    {empty_rows:>6,}  (empty cells)")

    # ── Step 7: Apply normalization ──
    print(f"\n[6/9] Applying normalization map...")
    change_counter = Counter()
    unchanged = 0
    warnings = []

    for row_num in range(DATA_START_ROW, max_row + 1):
        cell = ws_log.cell(row=row_num, column=TIPO_COL)
        original = cell.value
        if original is None or str(original).strip() == '':
            continue

        val = str(original).strip()
        if val in TIPO_NORMALIZATION:
            new_val = TIPO_NORMALIZATION[val]
            if val != new_val:
                cell.value = new_val
                change_counter[f"{val} -> {new_val}"] += 1
            else:
                unchanged += 1
        else:
            warnings.append(f"  WARNING: Unmapped value at row {row_num}: '{val}'")
            unchanged += 1

    print(f"\n  Changes made:")
    total_changes = 0
    for mapping, cnt in sorted(change_counter.items(), key=lambda x: -x[1]):
        print(f"    {cnt:>6,}  {mapping}")
        total_changes += cnt
    print(f"    ------")
    print(f"    {total_changes:>6,}  TOTAL CHANGES")
    print(f"    {unchanged:>6,}  unchanged (already canonical)")
    print(f"    {total_changes + unchanged:>6,}  TOTAL RECORDS")

    if warnings:
        print(f"\n  WARNINGS ({len(warnings)}):")
        for w in warnings:
            print(w)

    # ── Step 8: Verify AFTER distribution ──
    print(f"\n[7/9] Reading AFTER distribution...")
    after_counter = Counter()
    for row_num in range(DATA_START_ROW, max_row + 1):
        cell_value = ws_log.cell(row=row_num, column=TIPO_COL).value
        if cell_value is None or str(cell_value).strip() == '':
            continue
        after_counter[str(cell_value).strip()] += 1

    print(f"\n  AFTER normalization ({len(after_counter)} unique values):")
    for tipo, cnt in sorted(after_counter.items(), key=lambda x: -x[1]):
        canonical_marker = " [OK]" if tipo in CANONICAL_TYPES else " [UNEXPECTED!]"
        print(f"    {cnt:>6,}  {tipo}{canonical_marker}")

    after_total = sum(after_counter.values())
    print(f"    ------")
    print(f"    {after_total:>6,}  TOTAL")

    # ── Step 9: Verify PROJECAO formulas before save ──
    print(f"\n[8/9] Verifying PROJECAO formulas before save...")
    if proj_ws:
        formulas_pre_save = count_formulas(proj_ws)
        print(f"  PROJECAO formulas (in memory): {formulas_pre_save}")
        if formulas_pre_save < 19200:
            print(f"  CRITICAL: Formula count too low ({formulas_pre_save} < 19,200)!")
            print(f"  Restoring backup...")
            shutil.copy2(V13_BACKUP, V13_PATH)
            sys.exit(1)
    else:
        print("  SKIPPED (no PROJECAO sheet found)")

    # ── Step 10: Save V13 ──
    print(f"\n[9/9] Saving V13...")
    wb.save(str(V13_PATH))
    print(f"  Saved: {V13_PATH.name}")

    # ══════════════════════════════════════════════════════════════════
    # POST-SAVE VERIFICATION
    # ══════════════════════════════════════════════════════════════════
    print(f"\n{'=' * 70}")
    print("POST-SAVE VERIFICATION")
    print(f"{'=' * 70}")

    wb2 = openpyxl.load_workbook(str(V13_PATH), data_only=False)

    # Verify 1: LOG unique TIPO values
    ws_log2 = wb2['LOG'] if 'LOG' in wb2.sheetnames else find_sheet_by_prefix(wb2, 'LOG')
    verify_counter = Counter()
    verify_rows = 0
    for row_num in range(DATA_START_ROW, ws_log2.max_row + 1):
        cell_value = ws_log2.cell(row=row_num, column=TIPO_COL).value
        if cell_value is None or str(cell_value).strip() == '':
            continue
        verify_counter[str(cell_value).strip()] += 1
        verify_rows += 1

    unique_types = set(verify_counter.keys())
    print(f"\n  Unique TIPO values: {len(unique_types)}")
    for t in sorted(unique_types):
        print(f"    - {t} ({verify_counter[t]:,})")

    v1_pass = len(unique_types) == 7
    print(f"  CHECK 1 (7 unique types): {'PASS' if v1_pass else 'FAIL'} ({len(unique_types)} found)")

    # Verify 2: Total data rows
    v2_pass = verify_rows == 20830
    print(f"  CHECK 2 (20,830 data rows): {'PASS' if v2_pass else 'FAIL'} ({verify_rows:,} found)")

    # Verify 3: All types are canonical
    unexpected = unique_types - CANONICAL_TYPES
    v3_pass = len(unexpected) == 0
    print(f"  CHECK 3 (all canonical): {'PASS' if v3_pass else 'FAIL'}", end="")
    if unexpected:
        print(f" -- unexpected: {unexpected}")
    else:
        print()

    # Verify 4: PROJECAO formulas
    proj_ws2 = find_sheet_by_prefix(wb2, 'PROJE')
    if proj_ws2:
        formulas_after = count_formulas(proj_ws2)
        v4_pass = formulas_after >= 19200
        print(f"  CHECK 4 (PROJECAO >= 19,200): {'PASS' if v4_pass else 'FAIL'} ({formulas_after:,} formulas)")
    else:
        v4_pass = False
        formulas_after = 0
        print(f"  CHECK 4 (PROJECAO >= 19,200): FAIL (sheet not found)")

    wb2.close()

    # ── Final Summary ──
    all_pass = v1_pass and v2_pass and v3_pass and v4_pass
    print(f"\n{'=' * 70}")
    if all_pass:
        print("VERIFICATION PASSED -- All checks passed!")
    else:
        print("VERIFICATION FAILED -- Some checks did not pass!")
        if not all_pass:
            print("  Restoring backup...")
            shutil.copy2(V13_BACKUP, V13_PATH)
            print("  Backup restored.")
    print(f"  Total records: {verify_rows:,}")
    print(f"  Unique TIPO values: {len(unique_types)}")
    print(f"  PROJECAO formulas: {formulas_after:,}")
    print(f"  Changes applied: {total_changes:,}")
    print(f"{'=' * 70}")

    if not all_pass:
        sys.exit(1)


if __name__ == "__main__":
    main()

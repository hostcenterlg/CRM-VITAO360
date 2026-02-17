"""
Phase 05 Plan 02 — Build DASH tab in V13 with 3 compact blocks (~45 rows)
KPI cards + VENDEDOR/PERIODO filters. All formulas reference V13 LOG columns
using ENGLISH function names.

V13 LOG column mapping:
  A=DATA, B=CONSULTOR, L=TIPO, M=RESULTADO, N=MOTIVO,
  H=WHATSAPP, I=LIGACAO, J=LIG.ATENDIDA, Q=MERCOS

DASH filter cells: C2=VENDEDOR, E2=date_start, F2=date_end
"""
import sys
import os
import datetime
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from v3_styles import *
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import openpyxl

# ═══════════════════════════════════════════════════════════════
# CONSTANTS (NO ACCENTS -- LOG stores sem-acento)
# ═══════════════════════════════════════════════════════════════
TIPOS = [
    "PROSPECCAO", "NEGOCIACAO", "FOLLOW UP",
    "ATEND. CLIENTES ATIVOS", "ATEND. CLIENTES INATIVOS",
    "POS-VENDA / RELACIONAMENTO", "PERDA / NUTRICAO",
]
CONSULTORES = ["MANU DITZEL", "LARISSA PADILHA", "JULIO GADRET", "DAIANE STAVICKI"]
MOTIVOS = [
    "AINDA TEM ESTOQUE", "PRODUTO NAO VENDEU / SEM GIRO",
    "LOJA ANEXO/PROXIMO - SM", "SO QUER COMPRAR GRANEL",
    "PROBLEMA LOGISTICO / ENTREGA", "PROBLEMA FINANCEIRO / CREDITO",
    "PROPRIETARIO / INDISPONIVEL", "FECHANDO / FECHOU LOJA",
    "SEM INTERESSE NO MOMENTO", "PRIMEIRO CONTATO / SEM RESPOSTA",
]


# ═══════════════════════════════════════════════════════════════
# cf() HELPER — Core formula builder for V13 LOG
# ═══════════════════════════════════════════════════════════════
def cf(extra_criteria="", use_todos=True):
    """Build COUNTIFS formula referencing V13 LOG tab.
    LOG columns: A=DATA, B=CONSULTOR, L=TIPO, M=RESULTADO, N=MOTIVO,
                 H=WHATSAPP, I=LIGACAO, J=LIG.ATENDIDA, Q=MERCOS
    DASH filter cells: C2=VENDEDOR, E2=date_start, F2=date_end
    """
    date_filter = 'LOG!$A$3:$A$21000,">="&$E$2,LOG!$A$3:$A$21000,"<="&$F$2'
    cons_filter = ',LOG!$B$3:$B$21000,$C$2'
    if use_todos:
        base_all = f"COUNTIFS({date_filter}{extra_criteria})"
        base_cons = f"COUNTIFS({date_filter}{cons_filter}{extra_criteria})"
        return f'IF(OR($C$2="",$C$2="TODOS"),{base_all},{base_cons})'
    else:
        return f"COUNTIFS({date_filter}{extra_criteria})"


# ═══════════════════════════════════════════════════════════════
# section_title() and total_row() — from v3_dash.py
# ═══════════════════════════════════════════════════════════════
def section_title(ws, row, label, end_col=17):
    """Write a section title bar."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=1, value=f"  {label}")
    style_cell(cell, font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(1, end_col + 1):
        ws.cell(row=row, column=c).fill = FILL_GRAY_D9
        ws.cell(row=row, column=c).border = THIN_BORDER


def total_row(ws, row, cols, data_start, data_end, fill=FILL_DARK, label="TOTAL", label_col=1):
    """Write a TOTAL row with SUM formulas."""
    write_header(ws, row, label_col, label, fill=fill, align=ALIGN_LEFT)
    for c in cols:
        cl = get_column_letter(c)
        ws.cell(row=row, column=c).value = f'=SUM({cl}{data_start}:{cl}{data_end})'
        write_header(ws, row, c, ws.cell(row=row, column=c).value, fill=fill)


# ═══════════════════════════════════════════════════════════════
# build_dash() — Main DASH tab builder
# ═══════════════════════════════════════════════════════════════
def build_dash(wb):
    """Build DASH tab with 3 blocks of KPIs + KPI cards + filters."""
    ws = wb.create_sheet("DASH")
    ws.sheet_properties.tabColor = TAB_DASH

    # Column widths
    ws.column_dimensions['A'].width = 28
    for c in range(2, 18):
        ws.column_dimensions[get_column_letter(c)].width = 13

    # Date filter refs for Produtividade section (direct COUNTIFS, no cf())
    date_base = 'LOG!$A$3:$A$21000,">="&$E$2,LOG!$A$3:$A$21000,"<="&$F$2'

    # ═══════════════════════════════════════════════════════════
    # ROW 1: Title
    # ═══════════════════════════════════════════════════════════
    ws.merge_cells('A1:Q1')
    cell = ws.cell(row=1, column=1, value="DASHBOARD JARVIS CRM -- VITAO ALIMENTOS")
    style_cell(cell, font=Font(name='Calibri', size=13, bold=True, color='FFFFFF'),
               fill=PatternFill('solid', fgColor='DC2626'), align=ALIGN_LEFT)
    for c in range(1, 18):
        ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor='DC2626')
        ws.cell(row=1, column=c).border = THIN_BORDER

    # ═══════════════════════════════════════════════════════════
    # ROW 2: Filters (Vendedor + Periodo)
    # ═══════════════════════════════════════════════════════════
    write_data(ws, 2, 1, "VENDEDOR", font=FONT_TITLE, fill=FILL_LIGHT_GRAY, align=ALIGN_LEFT)
    ws.cell(row=2, column=3, value="TODOS").font = FONT_DATA
    ws.cell(row=2, column=3).border = THIN_BORDER
    dv_cons = DataValidation(
        type="list",
        formula1='"TODOS,MANU DITZEL,LARISSA PADILHA,JULIO GADRET,DAIANE STAVICKI"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_cons)
    dv_cons.add("C2")

    write_data(ws, 2, 4, "PERIODO", font=FONT_TITLE, fill=FILL_LIGHT_GRAY, align=ALIGN_LEFT)
    ws.cell(row=2, column=5, value=datetime.date(2026, 2, 1)).font = FONT_DATA
    ws.cell(row=2, column=5).number_format = FMT_DATE
    ws.cell(row=2, column=5).border = THIN_BORDER
    ws.cell(row=2, column=6, value=datetime.date(2026, 2, 28)).font = FONT_DATA
    ws.cell(row=2, column=6).number_format = FMT_DATE
    ws.cell(row=2, column=6).border = THIN_BORDER

    # ═══════════════════════════════════════════════════════════
    # ROW 3-4: KPI Cards (6 summary cards)
    # ═══════════════════════════════════════════════════════════
    kpi_row = 3
    kpis = [
        ("TOTAL CONTATOS",  1, f'={cf()}'),
        ("WHATSAPP",        3, f'={cf(",LOG!$H$3:$H$21000,\"SIM\"")}'),
        ("LIGACOES",        5, f'={cf(",LOG!$I$3:$I$21000,\"SIM\"")}'),
        ("LIG. ATENDIDAS",  7, f'={cf(",LOG!$J$3:$J$21000,\"SIM\"")}'),
        ("ORCAMENTOS",      9, f'={cf(",LOG!$M$3:$M$21000,\"ORCAMENTO\"")}'),
        ("VENDAS",          11, f'={cf(",LOG!$M$3:$M$21000,\"VENDA / PEDIDO\"")}'),
    ]
    for label, col, formula in kpis:
        ws.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=col + 1)
        write_data(ws, kpi_row, col, label, font=FONT_DATA_BOLD,
                   fill=PatternFill('solid', fgColor='E3F2FD'))
        ws.cell(row=kpi_row, column=col + 1).fill = PatternFill('solid', fgColor='E3F2FD')
        ws.cell(row=kpi_row, column=col + 1).border = THIN_BORDER
        ws.merge_cells(start_row=kpi_row + 1, start_column=col, end_row=kpi_row + 1, end_column=col + 1)
        ws.cell(row=kpi_row + 1, column=col).value = formula
        style_cell(ws.cell(row=kpi_row + 1, column=col),
                   font=Font(name='Calibri', size=16, bold=True),
                   fill=PatternFill('solid', fgColor='E3F2FD'))
        ws.cell(row=kpi_row + 1, column=col + 1).fill = PatternFill('solid', fgColor='E3F2FD')
        ws.cell(row=kpi_row + 1, column=col + 1).border = THIN_BORDER
    ws.row_dimensions[kpi_row].height = 20
    ws.row_dimensions[kpi_row + 1].height = 28

    # ═══════════════════════════════════════════════════════════
    # ROW 5: spacer
    # ═══════════════════════════════════════════════════════════

    # ═══════════════════════════════════════════════════════════
    # BLOCO 1: TIPO DO CONTATO x RESULTADO DO CONTATO (rows 6-15)
    # ═══════════════════════════════════════════════════════════
    B1 = 6
    section_title(ws, B1, "TIPO DO CONTATO x RESULTADO DO CONTATO", 13)

    HDR = B1 + 1  # row 7
    hdr_labels = [
        ("TIPO DO CONTATO", FILL_DARK, ALIGN_LEFT),
        ("TOTAL", FILL_DARK, ALIGN_CENTER),
        ("ORCAM.", FILL_DARK, ALIGN_CENTER),
        ("CADAST.", FILL_DARK, ALIGN_CENTER),
        ("RELAC.", FILL_DARK, ALIGN_CENTER),
        ("EM ATEND.", FILL_DARK, ALIGN_CENTER),
        ("SUPORTE", FILL_DARK, ALIGN_CENTER),
        ("VENDA", FILL_DARK, ALIGN_CENTER),
        ("N ATENDE", FILL_RED_DARK, ALIGN_CENTER),
        ("RECUSOU", FILL_RED_DARK, ALIGN_CENTER),
        ("N RESP.", FILL_RED_DARK, ALIGN_CENTER),
        ("PERDA", FILL_RED_DARK, ALIGN_CENTER),
        ("FOLLOW UP", FILL_GREEN_DARK, ALIGN_CENTER),
    ]
    for i, (label, fill, align) in enumerate(hdr_labels):
        write_header(ws, HDR, i + 1, label, fill=fill, font=FONT_HEADER_SM, align=align)

    resultado_map = {
        3: "ORCAMENTO", 4: "CADASTRO", 5: "RELACIONAMENTO",
        6: "EM ATENDIMENTO", 7: "SUPORTE", 8: "VENDA / PEDIDO",
        9: "NAO ATENDE", 10: "RECUSOU LIGACAO", 11: "NAO RESPONDE",
        12: "PERDA / FECHOU LOJA",
    }

    DS = HDR + 1  # row 8
    for t_idx, tipo in enumerate(TIPOS):
        r = DS + t_idx
        write_data(ws, r, 1, tipo, align=ALIGN_LEFT)
        ws.cell(row=r, column=2).value = f'=SUM(C{r}:M{r})'
        style_cell(ws.cell(row=r, column=2), font=FONT_DATA_BOLD)

        tipo_f = f',LOG!$L$3:$L$21000,A{r}'
        for col_idx, resultado in resultado_map.items():
            res_f = f',LOG!$M$3:$M$21000,"{resultado}"'
            ws.cell(row=r, column=col_idx).value = f'={cf(tipo_f + res_f)}'
            style_cell(ws.cell(row=r, column=col_idx))

        # FOLLOW UP column (col M = 13): sum of FOLLOW UP 7 + FOLLOW UP 15
        fu7 = ',LOG!$M$3:$M$21000,"FOLLOW UP 7"'
        fu15 = ',LOG!$M$3:$M$21000,"FOLLOW UP 15"'
        ws.cell(row=r, column=13).value = f'={cf(tipo_f + fu7)}+{cf(tipo_f + fu15)}'
        style_cell(ws.cell(row=r, column=13))

    TR = DS + len(TIPOS)  # row 15
    total_row(ws, TR, list(range(2, 14)), DS, TR - 1)

    # ═══════════════════════════════════════════════════════════
    # ROW 16: spacer
    # ═══════════════════════════════════════════════════════════

    # ═══════════════════════════════════════════════════════════
    # BLOCO 2: CONTATOS REALIZADOS + FUNIL DE VENDA (rows 17-27)
    # ═══════════════════════════════════════════════════════════
    B2 = TR + 2  # row 17
    section_title(ws, B2, "CONTATOS REALIZADOS + FUNIL DE VENDA", 17)

    # Group headers (row 18)
    GR = B2 + 1
    for c in [1, 2]:
        ws.cell(row=GR, column=c).fill = FILL_DARK_59
        ws.cell(row=GR, column=c).border = THIN_BORDER
    ws.merge_cells(f'C{GR}:F{GR}')
    write_header(ws, GR, 3, "CANAIS", fill=FILL_DARK_59)
    for c in [4, 5, 6]:
        ws.cell(row=GR, column=c).fill = FILL_DARK_59
        ws.cell(row=GR, column=c).border = THIN_BORDER

    # Separator cols 7, 8
    for c in [7, 8]:
        ws.cell(row=GR, column=c).fill = FILL_SEP
        ws.cell(row=GR, column=c).border = NO_BORDER

    ws.merge_cells(f'I{GR}:K{GR}')
    write_header(ws, GR, 9, "FUNIL DE VENDA", fill=FILL_DARK_59)
    for c in [10, 11]:
        ws.cell(row=GR, column=c).fill = FILL_DARK_59
        ws.cell(row=GR, column=c).border = THIN_BORDER

    # Separator col 12
    ws.cell(row=GR, column=12).fill = FILL_SEP
    ws.cell(row=GR, column=12).border = NO_BORDER

    ws.merge_cells(f'M{GR}:N{GR}')
    write_header(ws, GR, 13, "RELACIONAMENTO", fill=FILL_DARK_59)
    ws.cell(row=GR, column=14).fill = FILL_DARK_59
    ws.cell(row=GR, column=14).border = THIN_BORDER

    # Separator col 15
    ws.cell(row=GR, column=15).fill = FILL_SEP
    ws.cell(row=GR, column=15).border = NO_BORDER

    ws.merge_cells(f'P{GR}:Q{GR}')
    write_header(ws, GR, 16, "NAO VENDA", fill=FILL_DARK_59)
    ws.cell(row=GR, column=17).fill = FILL_DARK_59
    ws.cell(row=GR, column=17).border = THIN_BORDER

    # Sub-headers (row 19)
    SR = GR + 1
    sub_headers = [
        (1, "TIPO DO CONTATO", ALIGN_LEFT), (2, "TOTAL", ALIGN_CENTER),
        (3, "WHATSAPP", ALIGN_CENTER), (4, "LIGACAO", ALIGN_CENTER),
        (5, "LIG. ATEND.", ALIGN_CENTER), (6, "LIG. N ATEND.", ALIGN_CENTER),
        (9, "EM ATEND.", ALIGN_CENTER), (10, "ORCAMENTO", ALIGN_CENTER),
        (11, "VENDA", ALIGN_CENTER),
        (13, "FOLLOW UP", ALIGN_CENTER), (14, "SUPORTE", ALIGN_CENTER),
        (16, "INATIVO", ALIGN_CENTER), (17, "PERDA", ALIGN_CENTER),
    ]
    for col, label, align in sub_headers:
        write_header(ws, SR, col, label, fill=FILL_DARK, font=FONT_HEADER_SM, align=align)

    # Separators for sub-header row
    for sep_c in [7, 8, 12, 15]:
        ws.cell(row=SR, column=sep_c).fill = FILL_SEP
        ws.cell(row=SR, column=sep_c).border = NO_BORDER

    # Data rows (rows 20-26)
    B2D = SR + 1
    fu7_f = ',LOG!$M$3:$M$21000,"FOLLOW UP 7"'
    fu15_f = ',LOG!$M$3:$M$21000,"FOLLOW UP 15"'
    rel_f = ',LOG!$M$3:$M$21000,"RELACIONAMENTO"'
    na_f = ',LOG!$M$3:$M$21000,"NAO ATENDE"'
    nr_f = ',LOG!$M$3:$M$21000,"NAO RESPONDE"'
    rec_f = ',LOG!$M$3:$M$21000,"RECUSOU LIGACAO"'

    for t_idx, tipo in enumerate(TIPOS):
        r = B2D + t_idx
        write_data(ws, r, 1, tipo, align=ALIGN_LEFT)
        tipo_f = f',LOG!$L$3:$L$21000,A{r}'

        # B: TOTAL
        write_data(ws, r, 2, f'={cf(tipo_f)}', font=FONT_DATA_BOLD)
        # C: WHATSAPP
        wpp_f = ',LOG!$H$3:$H$21000,"SIM"'
        ws.cell(row=r, column=3).value = f'={cf(tipo_f + "," + wpp_f[1:])}'
        style_cell(ws.cell(row=r, column=3))
        # D: LIGACAO
        lig_f = ',LOG!$I$3:$I$21000,"SIM"'
        ws.cell(row=r, column=4).value = f'={cf(tipo_f + "," + lig_f[1:])}'
        style_cell(ws.cell(row=r, column=4))
        # E: LIG. ATENDIDA
        lig_at_f = ',LOG!$J$3:$J$21000,"SIM"'
        ws.cell(row=r, column=5).value = f'={cf(tipo_f + "," + lig_at_f[1:])}'
        style_cell(ws.cell(row=r, column=5))
        # F: LIG. NAO ATENDIDA = D - E
        ws.cell(row=r, column=6).value = f'=D{r}-E{r}'
        style_cell(ws.cell(row=r, column=6))

        # Separator cols
        for sep_c in [7, 8, 12, 15]:
            ws.cell(row=r, column=sep_c).fill = FILL_SEP
            ws.cell(row=r, column=sep_c).border = NO_BORDER

        # I: EM ATENDIMENTO
        ws.cell(row=r, column=9).value = f'={cf(tipo_f + ",LOG!$M$3:$M$21000,\"EM ATENDIMENTO\"")}'
        style_cell(ws.cell(row=r, column=9))
        # J: ORCAMENTO
        ws.cell(row=r, column=10).value = f'={cf(tipo_f + ",LOG!$M$3:$M$21000,\"ORCAMENTO\"")}'
        style_cell(ws.cell(row=r, column=10))
        # K: VENDA
        ws.cell(row=r, column=11).value = f'={cf(tipo_f + ",LOG!$M$3:$M$21000,\"VENDA / PEDIDO\"")}'
        style_cell(ws.cell(row=r, column=11))

        # M: FOLLOW UP = fu7 + fu15 + relacionamento
        ws.cell(row=r, column=13).value = f'={cf(tipo_f + fu7_f)}+{cf(tipo_f + fu15_f)}+{cf(tipo_f + rel_f)}'
        style_cell(ws.cell(row=r, column=13))
        # N: SUPORTE
        ws.cell(row=r, column=14).value = f'={cf(tipo_f + ",LOG!$M$3:$M$21000,\"SUPORTE\"")}'
        style_cell(ws.cell(row=r, column=14))

        # P: INATIVO = NAO ATENDE + NAO RESPONDE + RECUSOU
        ws.cell(row=r, column=16).value = f'={cf(tipo_f + na_f)}+{cf(tipo_f + nr_f)}+{cf(tipo_f + rec_f)}'
        style_cell(ws.cell(row=r, column=16))
        # Q: PERDA
        ws.cell(row=r, column=17).value = f'={cf(tipo_f + ",LOG!$M$3:$M$21000,\"PERDA / FECHOU LOJA\"")}'
        style_cell(ws.cell(row=r, column=17))

    B2T = B2D + len(TIPOS)  # row 27
    total_row(ws, B2T, [2, 3, 4, 5, 6, 9, 10, 11, 13, 14, 16, 17], B2D, B2T - 1)
    for sep_c in [7, 8, 12, 15]:
        ws.cell(row=B2T, column=sep_c).fill = FILL_SEP
        ws.cell(row=B2T, column=sep_c).border = NO_BORDER

    # ═══════════════════════════════════════════════════════════
    # ROW 28: spacer
    # ═══════════════════════════════════════════════════════════

    # ═══════════════════════════════════════════════════════════
    # BLOCO 3: MOTIVOS + PRODUTIVIDADE (side by side, rows 29-41)
    # ═══════════════════════════════════════════════════════════
    B3 = B2T + 2  # row 29

    # LEFT side title: MOTIVOS
    ws.merge_cells(f'A{B3}:G{B3}')
    ws.cell(row=B3, column=1, value="  MOTIVOS -- POR QUE NAO COMPRAM")
    style_cell(ws.cell(row=B3, column=1), font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(1, 8):
        ws.cell(row=B3, column=c).fill = FILL_GRAY_D9
        ws.cell(row=B3, column=c).border = THIN_BORDER

    # RIGHT side title: PRODUTIVIDADE
    ws.merge_cells(f'I{B3}:Q{B3}')
    ws.cell(row=B3, column=9, value="  PRODUTIVIDADE POR CONSULTOR")
    style_cell(ws.cell(row=B3, column=9), font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(9, 18):
        ws.cell(row=B3, column=c).fill = FILL_GRAY_D9
        ws.cell(row=B3, column=c).border = THIN_BORDER

    # Separator col H (8) for entire Bloco 3
    for sr in range(B3, B3 + 13):
        ws.cell(row=sr, column=8).fill = FILL_SEP
        ws.cell(row=sr, column=8).border = NO_BORDER

    # ── MOTIVOS headers (row 30) ──
    MH = B3 + 1
    for col, label, align in [
        (1, "MOTIVO", ALIGN_LEFT), (2, "QTD", ALIGN_CENTER), (3, "%", ALIGN_CENTER),
        (4, "PROSP", ALIGN_CENTER), (5, "ATIVOS", ALIGN_CENTER),
        (6, "INAT", ALIGN_CENTER), (7, "POS-V", ALIGN_CENTER),
    ]:
        write_header(ws, MH, col, label, fill=FILL_DARK, align=align)

    # ── MOTIVOS data (rows 31-40) ──
    MD = MH + 1
    for m_idx, motivo in enumerate(MOTIVOS):
        r = MD + m_idx
        write_data(ws, r, 1, motivo, align=ALIGN_LEFT)
        mot_f = f',LOG!$N$3:$N$21000,A{r}'
        # B: QTD
        ws.cell(row=r, column=2).value = f'={cf(mot_f)}'
        style_cell(ws.cell(row=r, column=2))
        # D-G: breakdown by TIPO
        for c, tipo_val in [
            (4, "PROSPECCAO"), (5, "ATEND. CLIENTES ATIVOS"),
            (6, "ATEND. CLIENTES INATIVOS"), (7, "POS-VENDA / RELACIONAMENTO"),
        ]:
            tf = f',LOG!$L$3:$L$21000,"{tipo_val}"'
            ws.cell(row=r, column=c).value = f'={cf(mot_f + tf)}'
            style_cell(ws.cell(row=r, column=c))

    MT = MD + len(MOTIVOS)  # row 41
    total_row(ws, MT, [2, 4, 5, 6, 7], MD, MT - 1)

    # % column (C) for motivos
    for m_idx in range(len(MOTIVOS)):
        r = MD + m_idx
        ws.cell(row=r, column=3).value = f'=IFERROR(B{r}/B{MT},0)'
        ws.cell(row=r, column=3).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=3))
    # % total
    ws.cell(row=MT, column=3).value = f'=IFERROR(B{MT}/B{MT},0)'
    ws.cell(row=MT, column=3).number_format = FMT_PCT
    write_header(ws, MT, 3, ws.cell(row=MT, column=3).value, fill=FILL_DARK)

    # ── PRODUTIVIDADE headers (row 30) ──
    IH = B3 + 1
    for col, label, align in [
        (9, "CONSULTOR", ALIGN_LEFT), (10, "CONTATOS", ALIGN_CENTER),
        (11, "VENDAS", ALIGN_CENTER), (12, "ORCAM.", ALIGN_CENTER),
        (13, "CADAST.", ALIGN_CENTER), (14, "% CONV", ALIGN_CENTER),
        (15, "N ATENDE", ALIGN_CENTER), (16, "PERDA", ALIGN_CENTER),
        (17, "% MERCOS", ALIGN_CENTER),
    ]:
        write_header(ws, IH, col, label, fill=FILL_DARK, align=align)

    # ── PRODUTIVIDADE data (rows 31-34) ──
    ID = IH + 1
    for c_idx, cons in enumerate(CONSULTORES):
        r = ID + c_idx
        write_data(ws, r, 9, cons, align=ALIGN_LEFT)
        cf_cons = f',LOG!$B$3:$B$21000,I{r}'

        # J: CONTATOS
        ws.cell(row=r, column=10).value = f'=COUNTIFS({date_base}{cf_cons})'
        style_cell(ws.cell(row=r, column=10))
        # K: VENDAS
        venda_f = ',LOG!$M$3:$M$21000,"VENDA / PEDIDO"'
        ws.cell(row=r, column=11).value = f'=COUNTIFS({date_base}{cf_cons}{venda_f})'
        style_cell(ws.cell(row=r, column=11))
        # L: ORCAMENTO
        orc_f = ',LOG!$M$3:$M$21000,"ORCAMENTO"'
        ws.cell(row=r, column=12).value = f'=COUNTIFS({date_base}{cf_cons}{orc_f})'
        style_cell(ws.cell(row=r, column=12))
        # M: CADASTRO
        cad_f = ',LOG!$M$3:$M$21000,"CADASTRO"'
        ws.cell(row=r, column=13).value = f'=COUNTIFS({date_base}{cf_cons}{cad_f})'
        style_cell(ws.cell(row=r, column=13))
        # N: % CONV
        ws.cell(row=r, column=14).value = f'=IFERROR(K{r}/J{r},0)'
        ws.cell(row=r, column=14).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=14))
        # O: N ATENDE = NAO ATENDE + NAO RESPONDE + RECUSOU LIGACAO
        na = ',LOG!$M$3:$M$21000,"NAO ATENDE"'
        nr = ',LOG!$M$3:$M$21000,"NAO RESPONDE"'
        rec = ',LOG!$M$3:$M$21000,"RECUSOU LIGACAO"'
        ws.cell(row=r, column=15).value = (
            f'=COUNTIFS({date_base}{cf_cons}{na})'
            f'+COUNTIFS({date_base}{cf_cons}{nr})'
            f'+COUNTIFS({date_base}{cf_cons}{rec})'
        )
        style_cell(ws.cell(row=r, column=15))
        # P: PERDA
        perda_f = ',LOG!$M$3:$M$21000,"PERDA / FECHOU LOJA"'
        ws.cell(row=r, column=16).value = f'=COUNTIFS({date_base}{cf_cons}{perda_f})'
        style_cell(ws.cell(row=r, column=16))
        # Q: % MERCOS
        mercos_f = ',LOG!$Q$3:$Q$21000,"SIM"'
        ws.cell(row=r, column=17).value = (
            f'=IFERROR(COUNTIFS({date_base}{cf_cons}{mercos_f})'
            f'/COUNTIFS({date_base}{cf_cons}),0)'
        )
        ws.cell(row=r, column=17).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=17))

    # ── PRODUTIVIDADE TOTAL (row 35) ──
    IT = ID + len(CONSULTORES)  # row 35
    write_header(ws, IT, 9, "TOTAL EQUIPE", fill=FILL_DARK, align=ALIGN_LEFT)
    for c in [10, 11, 12, 13, 15, 16]:
        cl = get_column_letter(c)
        ws.cell(row=IT, column=c).value = f'=SUM({cl}{ID}:{cl}{IT - 1})'
        write_header(ws, IT, c, ws.cell(row=IT, column=c).value, fill=FILL_DARK)
    # % CONV total
    ws.cell(row=IT, column=14).value = f'=IFERROR(K{IT}/J{IT},0)'
    ws.cell(row=IT, column=14).number_format = FMT_PCT
    write_header(ws, IT, 14, ws.cell(row=IT, column=14).value, fill=FILL_DARK)
    # % MERCOS total
    mercos_f = ',LOG!$Q$3:$Q$21000,"SIM"'
    ws.cell(row=IT, column=17).value = f'=IFERROR(COUNTIFS({date_base}{mercos_f})/COUNTIFS({date_base}),0)'
    ws.cell(row=IT, column=17).number_format = FMT_PCT
    write_header(ws, IT, 17, ws.cell(row=IT, column=17).value, fill=FILL_DARK)

    # ═══════════════════════════════════════════════════════════
    # FORMATTING
    # ═══════════════════════════════════════════════════════════
    # Separator column widths for Bloco 2
    for sep_col_letter in ['H', 'L', 'O']:
        ws.column_dimensions[sep_col_letter].width = 2

    ws.freeze_panes = "A5"
    ws.sheet_view.zoomScale = 90

    # Row heights
    for r in range(1, MT + 2):
        if r == kpi_row + 1:
            continue  # skip KPI big number row (already 28)
        if r == kpi_row:
            continue  # already 20
        ws.row_dimensions[r].height = 18

    last_row = MT  # row 41
    print(f"  DASH: {last_row} rows, 3 blocks + KPI cards, freeze A5")
    return ws


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    V13_PATH = Path("data/output/CRM_VITAO360_V13_PROJECAO.xlsx")
    if not V13_PATH.exists():
        print(f"ERROR: V13 not found at {V13_PATH}")
        sys.exit(1)

    print(f"Loading V13 from {V13_PATH}...")
    wb = openpyxl.load_workbook(str(V13_PATH), data_only=False)

    # Remove existing DASH if present (re-runnable)
    if "DASH" in wb.sheetnames:
        del wb["DASH"]
        print("  Removed existing DASH tab")

    build_dash(wb)
    wb.save(str(V13_PATH))
    print(f"  Saved V13 to {V13_PATH}")

    # ── Verify ──
    print("\nVerification:")
    wb2 = openpyxl.load_workbook(str(V13_PATH), data_only=False)
    assert "DASH" in wb2.sheetnames, "DASH tab not found!"

    # Count DASH formulas
    dash_formulas = sum(
        1 for row in wb2["DASH"].iter_rows()
        for cell in row
        if cell.value and str(cell.value).startswith('=')
    )

    # Count PROJECAO formulas
    proj_ws_name = [ws for ws in wb2.sheetnames if "PROJE" in ws.upper()][0]
    proj_formulas = sum(
        1 for row in wb2[proj_ws_name].iter_rows()
        for cell in row
        if cell.value and str(cell.value).startswith('=')
    )

    dash_rows = wb2["DASH"].max_row
    print(f"  DASH: {dash_rows} rows, {dash_formulas} formulas")
    print(f"  PROJECAO: {proj_formulas} formulas (must be >= 19200)")
    assert proj_formulas >= 19200, f"PROJECAO formulas lost: {proj_formulas} < 19200"
    print("\n  ALL CHECKS PASSED!")

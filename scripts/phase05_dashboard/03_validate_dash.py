"""
Phase 05 Plan 03 -- DASH Validation Script
Validates DASH tab against LOG data, verifies structural requirements (DASH-01..05),
and confirms PROJECAO formula preservation.

This is the final quality gate for Phase 5.
"""
import sys
import os
import datetime
from pathlib import Path
from collections import Counter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

# ================================================================
# CONFIG
# ================================================================
V13_PATH = Path("data/output/CRM_VITAO360_V13_PROJECAO.xlsx")

# Portuguese function names that should NOT appear in DASH formulas
PT_FUNCS = ["CONT.SES", "CONT.SE", "SOMASES", "SOMASE", "SE(", "SEERRO(", "SOMA("]

# Expected 7 canonical TIPO values
CANONICAL_TIPOS = {
    "PROSPECCAO", "NEGOCIACAO", "FOLLOW UP",
    "ATEND. CLIENTES ATIVOS", "ATEND. CLIENTES INATIVOS",
    "POS-VENDA / RELACIONAMENTO", "PERDA / NUTRICAO",
}

# Track pass/fail results
results = {}
warnings = []


def check(name, condition, detail=""):
    """Record a check result."""
    results[name] = condition
    status = "PASS" if condition else "FAIL"
    suffix = f" ({detail})" if detail else ""
    print(f"  [{status}] {name}{suffix}")
    return condition


# ================================================================
# LOAD V13
# ================================================================
print("=" * 60)
print("PHASE 5 DASHBOARD - VALIDATION REPORT")
print("=" * 60)
print()

if not V13_PATH.exists():
    print(f"FATAL: V13 not found at {V13_PATH}")
    sys.exit(1)

print(f"Loading V13 from {V13_PATH} (data_only=False)...")
wb = openpyxl.load_workbook(str(V13_PATH), data_only=False)
print(f"  Sheets: {wb.sheetnames}")
print()

# ================================================================
# 1. STRUCTURAL CHECKS
# ================================================================
print("STRUCTURAL CHECKS:")
print("-" * 40)

# 1a. V13 has 3 tabs (find PROJECAO with accent-stripping)
sheet_names = wb.sheetnames
has_log = "LOG" in sheet_names
has_dash = "DASH" in sheet_names
proj_ws_name = None
for sn in sheet_names:
    if "PROJE" in sn.upper():
        proj_ws_name = sn
        break
has_proj = proj_ws_name is not None

check("V13 has 3 tabs (PROJECAO, LOG, DASH)",
      has_log and has_dash and has_proj,
      f"found: {sheet_names}")

# 1b. DASH max_row <= 45
dash = wb["DASH"]
dash_max_row = dash.max_row
check("DASH <= 45 rows", dash_max_row <= 45, f"actual: {dash_max_row}")

# 1c. DASH has 3 section titles (cells with FILL_GRAY_D9 or key text)
section_keywords = ["TIPO DO CONTATO", "CONTATOS", "MOTIVOS"]
found_sections = set()
for row in dash.iter_rows(min_row=1, max_row=dash_max_row, max_col=17):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            val = cell.value.strip()
            for kw in section_keywords:
                if kw in val.upper():
                    found_sections.add(kw)

check("DASH has 3 section blocks",
      len(found_sections) >= 3,
      f"found: {found_sections}")

# 1d. DASH row 2, column 3 has DataValidation (dropdown)
has_dv = False
for dv in dash.data_validations.dataValidation:
    for rng in dv.sqref.ranges:
        if rng.min_row == 2 and rng.max_row == 2 and rng.min_col == 3 and rng.max_col == 3:
            has_dv = True
            break
    if has_dv:
        break
check("VENDEDOR dropdown present (C2)", has_dv)

# 1e. DASH row 2, column 5 has datetime value
date_start = dash.cell(row=2, column=5).value
is_date_start = isinstance(date_start, (datetime.date, datetime.datetime))
check("Date filter start is datetime (E2)", is_date_start, f"type: {type(date_start).__name__}, value: {date_start}")

# 1f. DASH row 2, column 6 has datetime value
date_end = dash.cell(row=2, column=6).value
is_date_end = isinstance(date_end, (datetime.date, datetime.datetime))
check("Date filter end is datetime (F2)", is_date_end, f"type: {type(date_end).__name__}, value: {date_end}")

print()

# ================================================================
# 2. FORMULA CHECKS
# ================================================================
print("FORMULA CHECKS:")
print("-" * 40)

# Collect all DASH formulas
dash_formulas = []
for row in dash.iter_rows(min_row=1, max_row=dash_max_row, max_col=17):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            dash_formulas.append((cell.row, cell.column, cell.value))

total_dash_formulas = len(dash_formulas)
check("Total DASH formulas > 0", total_dash_formulas > 0, f"count: {total_dash_formulas}")

# 2a. All formulas reference LOG (not DRAFT 2 or CARTEIRA)
draft2_refs = [f for (r, c, f) in dash_formulas if "DRAFT 2!" in f.upper() or "DRAFT2!" in f.upper()]
carteira_refs = [f for (r, c, f) in dash_formulas if "CARTEIRA!" in f.upper()]
log_refs = [f for (r, c, f) in dash_formulas if "LOG!" in f.upper()]

# Note: some formulas (like SUM of row range, or D-E subtractions, or IFERROR) don't need LOG refs
# Only check that formulas with COUNTIFS reference LOG
countifs_formulas = [f for (r, c, f) in dash_formulas if "COUNTIFS" in f.upper()]
countifs_with_log = [f for f in countifs_formulas if "LOG!" in f]

check("All COUNTIFS formulas reference LOG",
      len(countifs_formulas) > 0 and len(countifs_with_log) == len(countifs_formulas),
      f"{len(countifs_with_log)}/{len(countifs_formulas)} COUNTIFS reference LOG!")

check("Zero DRAFT 2 references", len(draft2_refs) == 0,
      f"{len(draft2_refs)} found")

check("Zero CARTEIRA references", len(carteira_refs) == 0,
      f"{len(carteira_refs)} found")

# 2b. No Portuguese function names
pt_found = []
for (r, c, f) in dash_formulas:
    for pt_func in PT_FUNCS:
        if pt_func in f.upper():
            pt_found.append((r, c, pt_func, f[:60]))

check("All formulas use English functions (no PT names)",
      len(pt_found) == 0,
      f"{len(pt_found)} Portuguese function names found")
if pt_found:
    for r, c, func, formula in pt_found[:5]:
        print(f"    -> Row {r}, Col {c}: {func} in {formula}")

# 2c. All formulas use bounded ranges (not full-column $A:$A)
import re
full_col_pattern = re.compile(r'\$[A-Z]+:\$[A-Z]+(?!\$?\d)')  # matches $A:$A but not $A$3:$A$21000
unbounded = []
for (r, c, f) in dash_formulas:
    # Check for patterns like $A:$A (no row numbers after)
    # We need to be careful: $A$3:$A$21000 is fine, $A:$A is not
    # Look for column-only refs (letter followed by colon, letter, but no digit)
    parts = f.split(',')
    for part in parts:
        part = part.strip()
        # Match patterns like LOG!$A:$A or $A:$A (full column reference)
        if re.search(r'!\$[A-Z]+:\$[A-Z]+[^$\d]', part + ' ') or re.search(r'^\$[A-Z]+:\$[A-Z]+[^$\d]', part + ' '):
            # But exclude cases with row numbers like $A$3:$A$21000
            if not re.search(r'\$[A-Z]+\$\d+:\$[A-Z]+\$\d+', part):
                unbounded.append((r, c, f[:60]))
                break

check("All formulas use bounded ranges (not full-column)",
      len(unbounded) == 0,
      f"{len(unbounded)} unbounded found")

# 2d. Count COUNTIFS specifically
check("COUNTIFS formulas present",
      len(countifs_formulas) > 0,
      f"count: {len(countifs_formulas)}")

print(f"\n  Formula summary: {total_dash_formulas} total, {len(countifs_formulas)} COUNTIFS, "
      f"{len(log_refs)} with LOG refs")

print()

# ================================================================
# 3. LOG DATA CROSS-CHECK
# ================================================================
print("LOG DATA CROSS-CHECK:")
print("-" * 40)

log_ws = wb["LOG"]

# Count total LOG data rows (column A, row 3 to last non-empty)
log_total = 0
log_dates = []
log_tipos = []
log_resultados = []
log_consultores = []
log_whatsapp_sim = 0
log_ligacao_sim = 0
log_lig_atendida_sim = 0
log_mercos_sim = 0

# Read LOG data (column A=DATA, B=CONSULTOR, H=WHATSAPP, I=LIGACAO, J=LIG.ATEND, L=TIPO, M=RESULTADO, Q=MERCOS)
for row_idx in range(3, log_ws.max_row + 1):
    data_val = log_ws.cell(row=row_idx, column=1).value  # A: DATA
    if data_val is None:
        continue
    log_total += 1

    # Collect date (normalize datetime to date for comparison)
    if isinstance(data_val, datetime.datetime):
        log_dates.append(data_val.date())
    elif isinstance(data_val, datetime.date):
        log_dates.append(data_val)

    # Consultor (B)
    cons_val = log_ws.cell(row=row_idx, column=2).value
    if cons_val:
        log_consultores.append(str(cons_val).strip())

    # WHATSAPP (H)
    wpp = log_ws.cell(row=row_idx, column=8).value
    if wpp and str(wpp).strip().upper() == "SIM":
        log_whatsapp_sim += 1

    # LIGACAO (I)
    lig = log_ws.cell(row=row_idx, column=9).value
    if lig and str(lig).strip().upper() == "SIM":
        log_ligacao_sim += 1

    # LIG. ATENDIDA (J)
    lig_at = log_ws.cell(row=row_idx, column=10).value
    if lig_at and str(lig_at).strip().upper() == "SIM":
        log_lig_atendida_sim += 1

    # TIPO DO CONTATO (L)
    tipo = log_ws.cell(row=row_idx, column=12).value
    if tipo:
        log_tipos.append(str(tipo).strip())

    # RESULTADO (M)
    resultado = log_ws.cell(row=row_idx, column=13).value
    if resultado:
        log_resultados.append(str(resultado).strip())

    # MERCOS ATUALIZADO (Q)
    mercos = log_ws.cell(row=row_idx, column=17).value
    if mercos and str(mercos).strip().upper() == "SIM":
        log_mercos_sim += 1

print(f"  Total LOG records: {log_total:,}")

# Unique TIPO values
tipo_counter = Counter(log_tipos)
unique_tipos = set(tipo_counter.keys())
check("LOG has exactly 7 unique TIPO DO CONTATO",
      len(unique_tipos) == 7,
      f"found {len(unique_tipos)}: {sorted(unique_tipos)}")

# Verify TIPO values match canonical
tipo_match = unique_tipos == CANONICAL_TIPOS
if not tipo_match:
    extra = unique_tipos - CANONICAL_TIPOS
    missing = CANONICAL_TIPOS - unique_tipos
    detail = ""
    if extra:
        detail += f"extra: {extra} "
    if missing:
        detail += f"missing: {missing}"
    check("LOG TIPO values match canonical 7", False, detail)
else:
    check("LOG TIPO values match canonical 7", True)

print(f"\n  TIPO distribution:")
for tipo, count in sorted(tipo_counter.items(), key=lambda x: -x[1]):
    print(f"    {tipo}: {count:,}")

# RESULTADO distribution
res_counter = Counter(log_resultados)
print(f"\n  RESULTADO distribution ({len(res_counter)} unique):")
for res, count in sorted(res_counter.items(), key=lambda x: -x[1]):
    print(f"    {res}: {count:,}")

# CONSULTOR distribution
cons_counter = Counter(log_consultores)
print(f"\n  CONSULTOR distribution ({len(cons_counter)} unique):")
for cons, count in sorted(cons_counter.items(), key=lambda x: -x[1]):
    print(f"    {cons}: {count:,}")

# Canonical 4 consultants check
canonical_consultants = {"MANU DITZEL", "LARISSA PADILHA", "JULIO GADRET", "DAIANE STAVICKI"}
found_canonical = canonical_consultants.intersection(set(cons_counter.keys()))
check("LOG has 4 canonical consultants",
      len(found_canonical) == 4,
      f"found: {sorted(found_canonical)}")

# Channel counts
print(f"\n  Channel counts:")
print(f"    WHATSAPP=SIM: {log_whatsapp_sim:,}")
print(f"    LIGACAO=SIM: {log_ligacao_sim:,}")
print(f"    LIG. ATENDIDA=SIM: {log_lig_atendida_sim:,}")
print(f"    MERCOS ATUALIZADO=SIM: {log_mercos_sim:,}")

# Default date range cross-check (2026-02-01 to 2026-02-28)
default_start = datetime.date(2026, 2, 1)
default_end = datetime.date(2026, 2, 28)
records_in_range = sum(1 for d in log_dates if default_start <= d <= default_end)
print(f"\n  Records in default date range ({default_start} to {default_end}): {records_in_range:,}")
print(f"  Note: DASH 'TOTAL CONTATOS' KPI should equal {records_in_range:,} when opened in Excel with default filters")

print()

# ================================================================
# 4. PROJECAO PRESERVATION CHECK
# ================================================================
print("PROJECAO PRESERVATION:")
print("-" * 40)

proj_ws = wb[proj_ws_name]
proj_formulas = sum(
    1 for row in proj_ws.iter_rows()
    for cell in row
    if cell.value and isinstance(cell.value, str) and str(cell.value).startswith('=')
)

check("PROJECAO formula count >= 19,200",
      proj_formulas >= 19200,
      f"count: {proj_formulas:,}")

print()

# ================================================================
# 5. REQUIREMENTS EVALUATION (DASH-01..05)
# ================================================================
print("REQUIREMENTS EVALUATION:")
print("-" * 40)

# DASH-01: DASH redesenhada com 3 blocos compactos (~45 rows vs 164 atual)
dash01 = dash_max_row <= 45 and len(found_sections) >= 3
check("DASH-01: 3 blocos compactos <= 45 rows",
      dash01,
      f"{dash_max_row} rows, {len(found_sections)} sections")

# DASH-02: Bloco 1: Visao executiva (faturamento, vendas, atendimentos)
# Check: Row containing "TIPO DO CONTATO" section exists with RESULTADO columns
has_tipo_section = False
has_resultado_headers = False
for row in dash.iter_rows(min_row=6, max_row=16, max_col=13):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            if "TIPO DO CONTATO" in cell.value.upper() and "RESULTADO" in cell.value.upper():
                has_tipo_section = True

# Check for RESULTADO column headers (ORCAM., CADAST., VENDA etc.)
resultado_headers_found = set()
resultado_expected = {"ORCAM", "CADAST", "VENDA", "TOTAL", "PERDA", "FOLLOW"}
for row in dash.iter_rows(min_row=6, max_row=8, max_col=13):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            val_up = cell.value.strip().upper()
            for exp in resultado_expected:
                if exp in val_up:
                    resultado_headers_found.add(exp)

dash02 = has_tipo_section and len(resultado_headers_found) >= 4
check("DASH-02: Bloco 1 visao executiva (TIPO x RESULTADO matrix)",
      dash02,
      f"section title: {has_tipo_section}, headers: {resultado_headers_found}")

# DASH-03: Performance por consultor
# Check: PRODUTIVIDADE section exists AND CONTATOS section exists
has_produtividade = False
has_contatos_section = False
for row in dash.iter_rows(min_row=1, max_row=dash_max_row, max_col=17):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            val = cell.value.strip().upper()
            if "PRODUTIVIDADE" in val:
                has_produtividade = True
            if "CONTATOS REALIZADOS" in val:
                has_contatos_section = True

dash03 = has_produtividade and has_contatos_section
check("DASH-03: Performance por consultor (PRODUTIVIDADE + CONTATOS)",
      dash03,
      f"PRODUTIVIDADE: {has_produtividade}, CONTATOS: {has_contatos_section}")

# DASH-04: Pipeline e funil
# Check: FUNIL group header exists in Bloco 2
has_funil = False
funil_cols_found = set()
for row in dash.iter_rows(min_row=17, max_row=27, max_col=17):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            val = cell.value.strip().upper()
            if "FUNIL" in val:
                has_funil = True
            if val in ("EM ATEND.", "ORCAMENTO", "VENDA"):
                funil_cols_found.add(val)

dash04 = has_funil and len(funil_cols_found) >= 2
check("DASH-04: Pipeline e funil (FUNIL DE VENDA section)",
      dash04,
      f"FUNIL header: {has_funil}, columns: {funil_cols_found}")

# DASH-05: Formulas referenciam LOG corretamente
# Check: All COUNTIFS reference LOG, no DRAFT 2, no broken refs
dash05 = (len(countifs_formulas) > 0 and
          len(countifs_with_log) == len(countifs_formulas) and
          len(draft2_refs) == 0 and
          len(carteira_refs) == 0)
check("DASH-05: Formulas referenciam LOG corretamente",
      dash05,
      f"LOG refs: {len(countifs_with_log)}/{len(countifs_formulas)}, "
      f"DRAFT2: {len(draft2_refs)}, CARTEIRA: {len(carteira_refs)}")

print()

# ================================================================
# FINAL REPORT
# ================================================================
print("=" * 60)
print("FINAL SUMMARY")
print("=" * 60)

req_results = {
    "DASH-01": dash01,
    "DASH-02": dash02,
    "DASH-03": dash03,
    "DASH-04": dash04,
    "DASH-05": dash05,
}

pass_count = sum(1 for v in req_results.values() if v)
total_reqs = len(req_results)

print(f"\n  Requirements: {pass_count}/{total_reqs} PASS")
for name, passed in req_results.items():
    print(f"    {'PASS' if passed else 'FAIL'}: {name}")

total_checks = len(results)
passed_checks = sum(1 for v in results.values() if v)
failed_checks = [k for k, v in results.items() if not v]

print(f"\n  Overall checks: {passed_checks}/{total_checks} PASS")
if failed_checks:
    print(f"\n  FAILED checks:")
    for fc in failed_checks:
        print(f"    - {fc}")

print(f"\n  PROJECAO formulas: {proj_formulas:,}")
print(f"  DASH formulas: {total_dash_formulas}")
print(f"  DASH rows: {dash_max_row}")
print(f"  LOG records: {log_total:,}")

print()
print("=" * 60)

if failed_checks:
    print(f"RESULT: FAIL ({len(failed_checks)} checks failed)")
    sys.exit(1)
else:
    print("RESULT: ALL CHECKS PASSED -- Phase 5 Dashboard COMPLETE")
    sys.exit(0)

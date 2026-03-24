"""Analyze PROJEÇÃO Excel files to understand structure and formula state."""
import openpyxl
import os
import sys

FILES = [
    r"c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\sources\projecao\PROJECAO_534_INTEGRADA.xlsx",
    r"c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\sources\projecao\PROJECAO_INTERNO_1566.xlsx",
    r"c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\sources\projecao\PROJECAO_POPULADA_1566.xlsx",
    r"c:\Users\User\OneDrive\Documentos\GitHub\CRM-VITAO360\data\sources\projecao\PROJECAO_CORRIGIDA (2).xlsx",
]

def analyze_file(filepath):
    basename = os.path.basename(filepath)
    print(f"\n{'='*80}")
    print(f"FILE: {basename}")
    print(f"Size: {os.path.getsize(filepath)/1024:.1f} KB")
    print(f"{'='*80}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=False, read_only=True)
    except Exception as e:
        print(f"ERROR opening: {e}")
        return

    print(f"Sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        try:
            print(f"  Dimensions: {ws.dimensions}")
        except AttributeError:
            print(f"  Dimensions: (read-only mode)")

        formula_count = 0
        data_count = 0
        empty_count = 0
        sample_formulas = []

        row_count = 0
        for row in ws.iter_rows():
            row_count += 1
            if row_count > 600:  # Don't read more than 600 rows
                break
            for cell in row:
                if cell.value is None:
                    empty_count += 1
                elif isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
                    if len(sample_formulas) < 10:
                        sample_formulas.append(f"    {cell.coordinate}: {cell.value[:100]}")
                else:
                    data_count += 1

        print(f"  Rows scanned: {row_count}")
        print(f"  Formulas: {formula_count}")
        print(f"  Data cells: {data_count}")
        print(f"  Empty: {empty_count}")

        if sample_formulas:
            print(f"  Sample formulas:")
            for sf in sample_formulas:
                print(sf)

        # Print headers (first 3 rows)
        print(f"  Headers (first 3 rows):")
        r = 0
        for row in ws.iter_rows(max_row=3):
            r += 1
            vals = [str(c.value)[:30] if c.value else "" for c in row[:20]]
            print(f"    Row {r}: {vals}")

    wb.close()

for f in FILES:
    if os.path.exists(f):
        analyze_file(f)
    else:
        print(f"\nFILE NOT FOUND: {f}")

print("\n\nDONE.")

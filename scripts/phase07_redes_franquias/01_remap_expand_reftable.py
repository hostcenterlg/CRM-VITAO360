"""
Phase 07 Plan 01: Remap 11 SEM GRUPO clients and expand AS:AZ reference table
===============================================================================
1. Build CNPJ->Rede map from SAP Cadastro (col AQ)
2. Remap 11 SEM GRUPO clients in V13 PROJECAO col C
3. Collect real 2025 data per rede (Total Lojas from SAP Leads, Fat.Real from PROJECAO)
4. Expand ref table AS:AZ from 12 to 19 redes + SEM GRUPO (20 rows total)
5. Update VLOOKUPs in cols F:J to reference expanded range
6. Validate post-save
"""

import sys
import openpyxl
import unicodedata
import re
from pathlib import Path
from collections import defaultdict, Counter

# ============================================================
# CONFIGURATION
# ============================================================
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
V13_PATH = PROJECT_ROOT / "data" / "output" / "CRM_VITAO360_V13_PROJECAO.xlsx"
SAP_PATH = PROJECT_ROOT / "data" / "sources" / "sap" / "01_SAP_CONSOLIDADO.xlsx"
SAP_META_PATH = PROJECT_ROOT / "data" / "sources" / "sap" / "BASE_SAP_META_PROJECAO_2026.xlsx"

# Benchmark parameters (from SINALEIRO_REDES_VITAO.xlsx / PARAMETROS)
BENCHMARK_PER_LOJA_PER_MES = 525  # R$ 525/mes/loja
MESES = 11  # 11 meses de dados

# Normalization map: SAP grupo chave name -> V13 rede name
REDE_NORMALIZE = {
    '06 - INTERNA - BIO MUNDO': 'BIO MUNDO',
    '06 - INTERNA - CIA DA SAUDE': 'CIA DA SAUDE',
    '06 - INTERNA - DIVINA TERRA': 'DIVINA TERRA',
    '06 - INTERNA - FITLAND': 'FITLAND',
    '06 - INTERNA - MUNDO VERDE': 'MUNDO VERDE',
    '06 - INTERNA - VIDA LEVE': 'VIDA LEVE',
    '06 - INTERNA - NATURVIDA': 'NATURVIDA',
    '06 - INTERNA - TUDO EM GRAOS / VGA': 'TUDO EM GRAOS',
    '06 - INTERNA - TRIP': 'TRIP',
    '06 - INTERNA - MAIS NATURAL': 'MAIS NATURAL',
    '06 - INTERNA - LIGEIRINHO': 'LIGEIRINHO',
    '06 - INTERNA - PROSAUDE': 'PROSAUDE',
    '06 - INTERNA - ARMAZEM FIT STORE': 'ARMAZEM FIT STORE',
    '06 - INTERNA - ESMERALDA': 'ESMERALDA',
    '06 - INTERNA - NOVA GERACAO': 'NOVA GERACAO',
    '06 - INTERNA - MERCOCENTRO': 'MERCOCENTRO',
    '06 - INTERNA - JARDIM DAS ERVAS': 'JARDIM DAS ERVAS',
    '06 - INTERNA - FEDERZONI': 'FEDERZONI',
    '06 - INTERNA - MIX VALI': 'MIX VALI',
    '06 - INTERNA - MINHA QUITANDINHA - SP': 'MINHA QUITANDINHA',
    '06 - SEM GRUPO': 'SEM GRUPO',
}

# Known 19 redes
KNOWN_REDES = {
    'FITLAND', 'VIDA LEVE', 'CIA DA SAUDE', 'DIVINA TERRA', 'BIO MUNDO',
    'MUNDO VERDE', 'NATURVIDA', 'TUDO EM GRAOS', 'TRIP', 'MAIS NATURAL',
    'LIGEIRINHO', 'PROSAUDE', 'ARMAZEM FIT STORE', 'ESMERALDA',
    'NOVA GERACAO', 'MERCOCENTRO', 'JARDIM DAS ERVAS', 'FEDERZONI',
    'MIX VALI', 'MINHA QUITANDINHA',
}

# Color/maturity/action thresholds
def get_cor(sinaleiro_pct):
    if sinaleiro_pct <= 0.01:
        return 'ROXO'
    elif sinaleiro_pct <= 0.40:
        return 'VERMELHO'
    elif sinaleiro_pct <= 0.60:
        return 'AMARELO'
    else:
        return 'VERDE'

def get_maturidade(cor):
    return {'ROXO': 'PROSPECCAO', 'VERMELHO': 'ATIVACAO/POSITIVACAO',
            'AMARELO': 'SELL OUT', 'VERDE': 'JBP'}[cor]

def get_acao(maturidade):
    return {'PROSPECCAO': 'Cadastrar e ativar lojas',
            'ATIVACAO/POSITIVACAO': 'Aumentar frequencia de compra',
            'SELL OUT': 'Expandir mix de produtos',
            'JBP': 'Manter e fidelizar'}[maturidade]

COR_EMOJI = {
    'ROXO': '\U0001f7e3 ROXO', 'VERMELHO': '\U0001f534 VERMELHO',
    'AMARELO': '\U0001f7e1 AMARELO', 'VERDE': '\U0001f7e2 VERDE',
}


# ============================================================
# UTILITIES
# ============================================================

def normalize_cnpj(raw):
    if raw is None:
        return ''
    s = re.sub(r'[^0-9]', '', str(raw).strip())
    return s.zfill(14) if s else ''

def safe_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0

def find_projecao_sheet(wb):
    for name in wb.sheetnames:
        clean = unicodedata.normalize('NFD', name)
        clean = ''.join(c for c in clean if unicodedata.category(c) != 'Mn')
        if 'PROJECAO' in clean.upper():
            return wb[name], name
    raise ValueError("PROJECAO sheet not found")

def normalize_rede_name(sap_name):
    if not sap_name:
        return None
    sap_name = str(sap_name).strip()
    if sap_name in REDE_NORMALIZE:
        return REDE_NORMALIZE[sap_name]
    if sap_name.startswith('06 - INTERNA - '):
        name = sap_name.replace('06 - INTERNA - ', '').strip()
        if ' - SP' in name:
            name = name.split(' - SP')[0].strip()
        if ' / VGA' in name:
            name = name.split(' / VGA')[0].strip()
        return name
    if sap_name == '06 - SEM GRUPO':
        return 'SEM GRUPO'
    return sap_name


# ============================================================
# STEP 1: Build CNPJ -> Rede map from SAP
# ============================================================

def build_cnpj_to_rede():
    print("=" * 70, flush=True)
    print("STEP 1: Building CNPJ -> Rede map from SAP Cadastro", flush=True)
    print("=" * 70, flush=True)

    # Use data_only=True, NOT read_only (read_only has slow max_row)
    wb = openpyxl.load_workbook(str(SAP_PATH), data_only=True)
    ws = wb['Cadastro Clientes SAP']

    cnpj_to_rede = {}
    redes_found = defaultdict(int)

    # SAP has ~1760 rows
    for r in range(2, 1800):
        cnpj_raw = ws.cell(row=r, column=5).value  # Col E = CNPJ
        grupo = ws.cell(row=r, column=43).value     # Col AQ = 06 Nome Grupo Chave

        if not cnpj_raw or not grupo:
            continue

        grupo_str = str(grupo).strip()
        if '06 - SEM GRUPO' in grupo_str or grupo_str == '':
            continue

        cnpj = normalize_cnpj(cnpj_raw)
        if not cnpj:
            continue

        rede = normalize_rede_name(grupo_str)
        if rede and rede != 'SEM GRUPO':
            cnpj_to_rede[cnpj] = rede
            redes_found[rede] += 1

    wb.close()

    print(f"\nTotal CNPJs with rede: {len(cnpj_to_rede)}", flush=True)
    print(f"Unique redes: {len(redes_found)}", flush=True)
    for rede, count in sorted(redes_found.items(), key=lambda x: -x[1]):
        print(f"  {rede}: {count}", flush=True)

    return cnpj_to_rede, redes_found


# ============================================================
# STEP 2: Get TOTAL LOJAS from SAP Leads (small file, fast)
# ============================================================

def get_total_lojas_from_leads():
    print("\n  3a. TOTAL LOJAS from SAP META Leads...", flush=True)

    wb = openpyxl.load_workbook(str(SAP_META_PATH), data_only=True)
    ws = wb['Leads']

    total_lojas = {}
    # Col 10 = 06 NOME GRUPO CHAVE, Col 25 = TOTAL LOJAS
    # Header row 3, data rows 4-83
    for r in range(4, 100):
        grupo = ws.cell(row=r, column=10).value
        lojas = ws.cell(row=r, column=25).value
        if not grupo:
            continue
        rede = normalize_rede_name(str(grupo).strip())
        if rede and rede != 'SEM GRUPO':
            total_lojas[rede] = int(safe_float(lojas)) if lojas is not None else 0

    wb.close()

    print(f"  Found {len(total_lojas)} redes", flush=True)
    for rede, lojas in sorted(total_lojas.items(), key=lambda x: -x[1]):
        print(f"    {rede}: {lojas} lojas", flush=True)

    return total_lojas


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 70, flush=True)
    print("Phase 07 Plan 01: Remap SEM GRUPO + Expand AS:AZ Reference Table", flush=True)
    print("=" * 70, flush=True)

    # -----------------------------------------------------------
    # STEP 1: Build CNPJ -> Rede map from SAP
    # -----------------------------------------------------------
    cnpj_to_rede, sap_redes = build_cnpj_to_rede()

    # -----------------------------------------------------------
    # STEP 1b: Get TOTAL LOJAS from SAP Leads (before loading V13)
    # -----------------------------------------------------------
    total_lojas = get_total_lojas_from_leads()

    # -----------------------------------------------------------
    # STEP 2: Load V13 (data_only=False to preserve formulas)
    # -----------------------------------------------------------
    print("\n  Loading V13 (data_only=False)... this takes ~5 min", flush=True)
    wb = openpyxl.load_workbook(str(V13_PATH), data_only=False)
    ws, sheet_name = find_projecao_sheet(wb)
    print(f"  Sheet: {repr(sheet_name)}", flush=True)
    print(f"  freeze_panes before: {ws.freeze_panes}", flush=True)

    # -----------------------------------------------------------
    # STEP 3: Remap SEM GRUPO clients
    # -----------------------------------------------------------
    print("\n" + "=" * 70, flush=True)
    print("STEP 3: Remapping SEM GRUPO clients", flush=True)
    print("=" * 70, flush=True)

    remappings = []
    for r in range(4, 538):
        cnpj_raw = ws.cell(row=r, column=1).value
        rede_atual = ws.cell(row=r, column=3).value
        if not cnpj_raw:
            continue
        cnpj = normalize_cnpj(cnpj_raw)
        rede_str = str(rede_atual).strip() if rede_atual else ''
        if rede_str == 'SEM GRUPO' and cnpj in cnpj_to_rede:
            nova_rede = cnpj_to_rede[cnpj]
            ws.cell(row=r, column=3).value = nova_rede
            remappings.append({'row': r, 'cnpj': cnpj, 'old': rede_str, 'new': nova_rede})

    print(f"\n  Remapped: {len(remappings)} clients", flush=True)
    by_rede = defaultdict(int)
    for rm in remappings:
        by_rede[rm['new']] += 1
        print(f"    Row {rm['row']}: {rm['cnpj']} -> {rm['new']}", flush=True)

    print("\n  By rede:", flush=True)
    for rede, count in sorted(by_rede.items(), key=lambda x: -x[1]):
        print(f"    {rede}: {count}", flush=True)

    if len(remappings) != 11:
        print(f"  WARNING: Expected 11, got {len(remappings)}", flush=True)

    # -----------------------------------------------------------
    # STEP 4: Calculate FAT.REAL per rede from cols AA:AL
    # -----------------------------------------------------------
    print("\n" + "=" * 70, flush=True)
    print("STEP 4: Calculating FAT.REAL per rede from cols AA:AL", flush=True)
    print("=" * 70, flush=True)

    fat_real = defaultdict(float)
    client_count = defaultdict(int)

    for r in range(4, 538):
        rede = ws.cell(row=r, column=3).value
        if not rede:
            continue
        rede_str = str(rede).strip()
        total = 0.0
        for c in range(27, 39):  # AA=27 to AL=38
            total += safe_float(ws.cell(row=r, column=c).value)
        fat_real[rede_str] += total
        client_count[rede_str] += 1

    print(f"  Redes with fat: {len(fat_real)}", flush=True)
    for rede in sorted(fat_real.keys(), key=lambda x: -fat_real[x]):
        print(f"    {rede}: R$ {fat_real[rede]:,.2f} ({client_count[rede]} cl)", flush=True)

    # -----------------------------------------------------------
    # STEP 5: Build rede data with sinaleiro/cor/maturidade
    # -----------------------------------------------------------
    print("\n" + "=" * 70, flush=True)
    print("STEP 5: Building rede data", flush=True)
    print("=" * 70, flush=True)

    all_redes = set(total_lojas.keys()) | set(k for k in fat_real if k != 'SEM GRUPO') | KNOWN_REDES

    rede_data = []
    for rede in all_redes:
        lojas = total_lojas.get(rede, 0)
        fat = fat_real.get(rede, 0.0)
        fat_pot = lojas * BENCHMARK_PER_LOJA_PER_MES * MESES
        sinal = fat / fat_pot if fat_pot > 0 else 0.0
        cor = get_cor(sinal)
        mat = get_maturidade(cor)
        acao = get_acao(mat)
        gap = max(0, fat_pot - fat)

        rede_data.append({
            'rede_nome': rede, 'total_lojas': lojas,
            'sinaleiro_pct': round(sinal, 6), 'cor': COR_EMOJI.get(cor, cor),
            'maturidade': mat, 'acao': acao,
            'fat_real': round(fat, 2), 'gap': round(gap, 2),
        })

    rede_data.sort(key=lambda x: -x['fat_real'])

    print(f"  Total redes: {len(rede_data)}", flush=True)
    for rd in rede_data:
        print(f"    {rd['rede_nome']}: Lojas={rd['total_lojas']} Fat=R${rd['fat_real']:,.2f} "
              f"Sinal={rd['sinaleiro_pct']:.2%} {rd['cor']} {rd['maturidade']}", flush=True)

    # SEM GRUPO aggregated
    sem_fat = fat_real.get('SEM GRUPO', 0.0)
    sem_clients = client_count.get('SEM GRUPO', 0)
    sem_grupo_data = {
        'rede_nome': 'SEM GRUPO', 'total_lojas': sem_clients,
        'sinaleiro_pct': 0.0, 'cor': COR_EMOJI['ROXO'],
        'maturidade': 'PROSPECCAO', 'acao': 'Cadastrar e ativar lojas',
        'fat_real': round(sem_fat, 2), 'gap': 0.0,
    }
    print(f"\n  SEM GRUPO: {sem_clients} clients, R$ {sem_fat:,.2f}", flush=True)

    # -----------------------------------------------------------
    # STEP 6: Expand reference table AS:AZ
    # -----------------------------------------------------------
    print("\n" + "=" * 70, flush=True)
    print("STEP 6: Expanding reference table AS:AZ", flush=True)
    print("=" * 70, flush=True)

    # Column mapping: AS=45, AT=46, AU=47, AV=48, AW=49, AX=50, AY=51, AZ=52
    # Write ALL redes found (may be 19 or 20, depending on MINHA QUITANDINHA)
    redes_to_write = rede_data  # Include all redes, not just top 19

    for i, rd in enumerate(redes_to_write):
        row = 4 + i
        ws.cell(row=row, column=45).value = rd['rede_nome']
        ws.cell(row=row, column=46).value = rd['total_lojas']
        ws.cell(row=row, column=47).value = rd['sinaleiro_pct']
        ws.cell(row=row, column=48).value = rd['cor']
        ws.cell(row=row, column=49).value = rd['maturidade']
        ws.cell(row=row, column=50).value = rd['acao']
        ws.cell(row=row, column=51).value = rd['fat_real']
        ws.cell(row=row, column=52).value = rd['gap']

    # SEM GRUPO at row after last rede
    sem_row = 4 + len(redes_to_write)  # row 23
    ws.cell(row=sem_row, column=45).value = sem_grupo_data['rede_nome']
    ws.cell(row=sem_row, column=46).value = sem_grupo_data['total_lojas']
    ws.cell(row=sem_row, column=47).value = sem_grupo_data['sinaleiro_pct']
    ws.cell(row=sem_row, column=48).value = sem_grupo_data['cor']
    ws.cell(row=sem_row, column=49).value = sem_grupo_data['maturidade']
    ws.cell(row=sem_row, column=50).value = sem_grupo_data['acao']
    ws.cell(row=sem_row, column=51).value = sem_grupo_data['fat_real']
    ws.cell(row=sem_row, column=52).value = sem_grupo_data['gap']

    # Clear old data beyond sem_row
    for r in range(sem_row + 1, 31):
        for c in range(45, 53):
            if ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).value = None

    print(f"  Written {len(redes_to_write)} redes (rows 4-{4+len(redes_to_write)-1})", flush=True)
    print(f"  SEM GRUPO at row {sem_row}", flush=True)

    for i, rd in enumerate(redes_to_write):
        if rd['rede_nome']:
            print(f"    Row {4+i}: {rd['rede_nome']} | {rd['total_lojas']} lojas | R${rd['fat_real']:,.2f}", flush=True)

    # -----------------------------------------------------------
    # STEP 7: Update VLOOKUPs in cols F:J
    # -----------------------------------------------------------
    print("\n" + "=" * 70, flush=True)
    print("STEP 7: Updating VLOOKUPs in cols F:J", flush=True)
    print("=" * 70, flush=True)

    updated_count = 0
    for r in range(4, 538):
        for c in range(6, 11):  # F=6 to J=10
            val = ws.cell(row=r, column=c).value
            if not val or not isinstance(val, str) or not val.startswith('='):
                continue
            if '$AS$4:' not in val:
                continue

            # Replace old end-row with sem_row
            new_val = val
            for cl in ['AT', 'AU', 'AV', 'AW', 'AX']:
                marker = f'${cl}$'
                if marker in new_val:
                    idx = new_val.index(marker) + len(marker)
                    end_idx = idx
                    while end_idx < len(new_val) and new_val[end_idx].isdigit():
                        end_idx += 1
                    if idx < end_idx:
                        new_val = new_val[:idx] + str(sem_row) + new_val[end_idx:]

            if new_val != val:
                ws.cell(row=r, column=c).value = new_val
                updated_count += 1

    print(f"  Updated {updated_count} VLOOKUP formulas", flush=True)
    print(f"  New end row: {sem_row}", flush=True)

    # Sample
    for r in [4, 537]:
        for c in [6, 10]:
            v = ws.cell(row=r, column=c).value
            if v:
                print(f"    [{r},{c}] = {v}", flush=True)

    # -----------------------------------------------------------
    # STEP 8: Save V13
    # -----------------------------------------------------------
    print("\n" + "=" * 70, flush=True)
    print("STEP 8: Saving V13... (this takes ~5 min)", flush=True)
    print("=" * 70, flush=True)

    wb.save(str(V13_PATH))
    print(f"  Saved: {V13_PATH}", flush=True)
    print(f"  freeze_panes after: {ws.freeze_panes}", flush=True)
    wb.close()

    # -----------------------------------------------------------
    # STEP 9: Validation (reopen to verify)
    # -----------------------------------------------------------
    print("\n" + "=" * 70, flush=True)
    print("STEP 9: Validation", flush=True)
    print("=" * 70, flush=True)

    print("  Reopening V13 for validation...", flush=True)
    wb_val = openpyxl.load_workbook(str(V13_PATH), data_only=False)
    ws_val, _ = find_projecao_sheet(wb_val)

    errors = []

    # 9a. Count formulas
    formula_count = 0
    for row in ws_val.iter_rows(min_row=4, max_row=537):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1

    print(f"\n  Formulas: {formula_count}", flush=True)
    if formula_count >= 19224:
        print(f"  PASS: >= 19,224", flush=True)
    else:
        errors.append(f"Formula count {formula_count} < 19,224")
        print(f"  FAIL: {formula_count} < 19,224", flush=True)

    # 9b. Rede distribution
    redes = Counter()
    for r in range(4, 538):
        v = ws_val.cell(row=r, column=3).value
        if v:
            redes[str(v).strip()] += 1

    sem_count = redes.get('SEM GRUPO', 0)
    with_count = sum(c for k, c in redes.items() if k != 'SEM GRUPO')
    print(f"\n  SEM GRUPO: {sem_count} | With rede: {with_count}", flush=True)
    for rede, cnt in sorted(redes.items(), key=lambda x: -x[1]):
        if rede != 'SEM GRUPO':
            print(f"    {rede}: {cnt}", flush=True)

    if sem_count == 394:
        print(f"  PASS: SEM GRUPO = 394", flush=True)
    else:
        print(f"  INFO: SEM GRUPO = {sem_count} (expected 394)", flush=True)

    # 9c. Reference table
    ref_count = sum(1 for r in range(4, 31) if ws_val.cell(row=r, column=45).value)
    print(f"\n  Ref table rows: {ref_count}", flush=True)
    if ref_count >= 20:
        print(f"  PASS: >= 20", flush=True)
    else:
        errors.append(f"Ref table {ref_count} < 20")
        print(f"  FAIL", flush=True)

    # 9d. VLOOKUP range
    f4 = ws_val.cell(row=4, column=6).value
    print(f"\n  VLOOKUP F4: {f4}", flush=True)
    if f4 and f'${sem_row}' in str(f4):
        print(f"  PASS: references row {sem_row}", flush=True)
    else:
        errors.append(f"VLOOKUP F4 wrong: {f4}")
        print(f"  FAIL", flush=True)

    # 9e. freeze_panes
    print(f"\n  freeze_panes: {ws_val.freeze_panes}", flush=True)

    wb_val.close()

    # -----------------------------------------------------------
    # FINAL REPORT
    # -----------------------------------------------------------
    print("\n" + "=" * 70, flush=True)
    print("EXECUTION REPORT", flush=True)
    print("=" * 70, flush=True)
    print(f"  Remappings: {len(remappings)}", flush=True)
    print(f"  Redes in table: {len(redes_to_write)} + SEM GRUPO", flush=True)
    print(f"  VLOOKUPs updated: {updated_count}", flush=True)
    print(f"  Formulas preserved: {formula_count}", flush=True)
    print(f"  Errors: {len(errors)}", flush=True)

    if errors:
        for e in errors:
            print(f"  ERROR: {e}", flush=True)
        return False

    print("\n  ALL CHECKS PASSED", flush=True)
    return True


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)

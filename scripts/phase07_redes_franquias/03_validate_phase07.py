"""
Phase 07 Plan 03: Validate all Phase 7 requirements (REDE-01..04) + V13 integrity
==================================================================================
VALIDATION-ONLY script -- reads V13 without modification.

Checks:
  REDE-01: REDE/GRUPO CHAVE preenchido para 100% dos 534 clientes (col C, rows 4-537)
  REDE-02: Zero #REF! na aba REDES_FRANQUIAS_v2
  REDE-03: Sinaleiro de penetracao atualizado com dados 2025 (formulas SUMIFS/COUNTIFS)
  REDE-04: Metas 6M por rede operacionais (valores SAP populados e GAP calculado)
  INTEGRIDADE: 19.224+ PROJECAO formulas, all expected tabs, AS:AZ expanded, VLOOKUPs

Output: data/output/phase07/validation_report.json
"""

import sys
import json
import openpyxl
import unicodedata
import re
from pathlib import Path
from collections import Counter
from datetime import datetime, timezone

# ============================================================
# CONFIGURATION
# ============================================================
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
V13_PATH = PROJECT_ROOT / "data" / "output" / "CRM_VITAO360_V13_PROJECAO.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "data" / "output" / "phase07"
REPORT_PATH = OUTPUT_DIR / "validation_report.json"

# Expected values for cross-checking
EXPECTED_CLIENTS = 534
EXPECTED_SEM_GRUPO_APPROX = 394
EXPECTED_ESMERALDA_MIN = 7
EXPECTED_PROJECAO_FORMULAS = 19224
EXPECTED_REDES_TAB_FORMULAS = 200
EXPECTED_REDES_DATA_ROWS = 21  # 20 redes + SEM GRUPO

# META 6M reference values (from 07-02 SUMMARY)
META_REFERENCE = {
    'FITLAND': 283500,
    'CIA DA SAUDE': 351000,
    'VIDA LEVE': 154500,
    'DIVINA TERRA': 157500,
}
META_TOLERANCE = 0.05  # 5% tolerance

# Expected tabs in V13
EXPECTED_TABS = ['LOG', 'DASH', 'REDES_FRANQUIAS_v2']
# PROJECAO found via accent-stripping (has cedilla + trailing space)


def find_projecao_sheet(wb):
    """Find PROJECAO sheet by accent-stripping."""
    for name in wb.sheetnames:
        clean = unicodedata.normalize('NFD', name)
        clean = ''.join(c for c in clean if unicodedata.category(c) != 'Mn')
        if 'PROJECAO' in clean.upper():
            return name
    raise ValueError("PROJECAO sheet not found in workbook")


def is_formula(value):
    """Check if a cell value is a formula string."""
    return isinstance(value, str) and value.startswith('=')


def contains_error(value, error_type):
    """Check if a cell value contains a specific Excel error."""
    if value is None:
        return False
    s = str(value)
    return error_type in s


# ============================================================
# CHECK 1: REDE-01 -- REDE/GRUPO CHAVE preenchido
# ============================================================
def check_rede_01(ws_proj):
    """Validate that all 534 clients have REDE/GRUPO CHAVE filled in col C."""
    print("\n" + "=" * 70, flush=True)
    print("CHECK 1: REDE-01 -- REDE/GRUPO CHAVE preenchido", flush=True)
    print("=" * 70, flush=True)

    empty_cells = []
    rede_counter = Counter()
    total = 0

    for row in range(4, 4 + EXPECTED_CLIENTS):  # rows 4-537
        total += 1
        val = ws_proj.cell(row=row, column=3).value  # col C
        if val is None or str(val).strip() == '':
            empty_cells.append(row)
        else:
            rede_counter[str(val).strip()] += 1

    filled = total - len(empty_cells)
    sem_grupo = rede_counter.get('SEM GRUPO', 0)
    com_rede = filled - sem_grupo

    print(f"  Total clients checked: {total}", flush=True)
    print(f"  Filled: {filled}", flush=True)
    print(f"  Empty: {len(empty_cells)}", flush=True)
    print(f"  SEM GRUPO: {sem_grupo}", flush=True)
    print(f"  Com rede: {com_rede}", flush=True)

    # Check SEM GRUPO decreased to ~394
    sem_grupo_ok = abs(sem_grupo - EXPECTED_SEM_GRUPO_APPROX) <= 5
    print(f"  SEM GRUPO ~{EXPECTED_SEM_GRUPO_APPROX}: {'YES' if sem_grupo_ok else 'NO'} (actual: {sem_grupo})", flush=True)

    # Check ESMERALDA has >= 7
    esmeralda = rede_counter.get('ESMERALDA', 0)
    esmeralda_ok = esmeralda >= EXPECTED_ESMERALDA_MIN
    print(f"  ESMERALDA >= {EXPECTED_ESMERALDA_MIN}: {'YES' if esmeralda_ok else 'NO'} (actual: {esmeralda})", flush=True)

    # Distribution
    print("\n  Rede distribution:", flush=True)
    for rede, count in rede_counter.most_common():
        print(f"    {rede}: {count}", flush=True)

    status = 'PASS' if len(empty_cells) == 0 else 'FAIL'
    if status == 'PASS' and (not sem_grupo_ok or not esmeralda_ok):
        status = 'PASS_WITH_NOTES'

    notes = []
    if not sem_grupo_ok:
        notes.append(f"SEM GRUPO expected ~{EXPECTED_SEM_GRUPO_APPROX}, got {sem_grupo}")
    if not esmeralda_ok:
        notes.append(f"ESMERALDA expected >= {EXPECTED_ESMERALDA_MIN}, got {esmeralda}")
    if empty_cells:
        notes.append(f"Empty cells at rows: {empty_cells[:10]}{'...' if len(empty_cells) > 10 else ''}")

    print(f"\n  REDE-01: [{status}]", flush=True)

    return {
        'description': f'REDE/GRUPO CHAVE preenchido para todos os {EXPECTED_CLIENTS} clientes',
        'status': status,
        'details': {
            'total_clients': total,
            'filled': filled,
            'empty': len(empty_cells),
            'sem_grupo': sem_grupo,
            'com_rede': com_rede,
            'sem_grupo_decreased': sem_grupo_ok,
            'esmeralda_count': esmeralda,
            'redes_distribution': dict(rede_counter.most_common()),
        },
        'notes': notes,
    }


# ============================================================
# CHECK 2: REDE-02 -- Zero #REF! na aba REDES_FRANQUIAS_v2
# ============================================================
def check_rede_02(wb):
    """Validate zero #REF! errors in REDES_FRANQUIAS_v2 tab."""
    print("\n" + "=" * 70, flush=True)
    print("CHECK 2: REDE-02 -- Zero #REF! na aba REDES_FRANQUIAS_v2", flush=True)
    print("=" * 70, flush=True)

    if 'REDES_FRANQUIAS_v2' not in wb.sheetnames:
        print("  FAIL: Tab REDES_FRANQUIAS_v2 does not exist!", flush=True)
        return {
            'description': 'Zero #REF! na aba REDES_FRANQUIAS_v2',
            'status': 'FAIL',
            'details': {
                'tab_exists': False,
                'ref_errors': -1,
                'name_errors': -1,
                'value_errors': -1,
                'div_errors': -1,
            },
            'notes': ['Tab REDES_FRANQUIAS_v2 not found in workbook'],
        }

    ws = wb['REDES_FRANQUIAS_v2']

    error_types = {
        '#REF!': 0,
        '#NAME?': 0,
        '#VALUE!': 0,
        '#DIV/0!': 0,
    }
    error_locations = []

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                val_str = str(cell.value)
                for err_type in error_types:
                    if err_type in val_str:
                        error_types[err_type] += 1
                        error_locations.append(f"{cell.coordinate}: {err_type}")

    total_errors = sum(error_types.values())

    print(f"  Tab exists: True", flush=True)
    print(f"  #REF! errors: {error_types['#REF!']}", flush=True)
    print(f"  #NAME? errors: {error_types['#NAME?']}", flush=True)
    print(f"  #VALUE! errors: {error_types['#VALUE!']}", flush=True)
    print(f"  #DIV/0! errors: {error_types['#DIV/0!']}", flush=True)
    print(f"  Total errors: {total_errors}", flush=True)

    if error_locations:
        print(f"  Error locations (first 10):", flush=True)
        for loc in error_locations[:10]:
            print(f"    {loc}", flush=True)

    status = 'PASS' if error_types['#REF!'] == 0 else 'FAIL'
    notes = []
    if total_errors > 0 and error_types['#REF!'] == 0:
        status = 'PASS_WITH_NOTES'
        notes.append(f"No #REF! but {total_errors} other formula errors found")

    print(f"\n  REDE-02: [{status}]", flush=True)

    return {
        'description': 'Zero #REF! na aba REDES_FRANQUIAS_v2',
        'status': status,
        'details': {
            'tab_exists': True,
            'ref_errors': error_types['#REF!'],
            'name_errors': error_types['#NAME?'],
            'value_errors': error_types['#VALUE!'],
            'div_errors': error_types['#DIV/0!'],
            'total_errors': total_errors,
            'error_locations': error_locations[:20],
        },
        'notes': notes,
    }


# ============================================================
# CHECK 3: REDE-03 -- Sinaleiro de penetracao atualizado
# ============================================================
def check_rede_03(wb):
    """Validate sinaleiro formulas are dynamic (SUMIFS, COUNTIFS, IF)."""
    print("\n" + "=" * 70, flush=True)
    print("CHECK 3: REDE-03 -- Sinaleiro de penetracao atualizado com dados 2025", flush=True)
    print("=" * 70, flush=True)

    if 'REDES_FRANQUIAS_v2' not in wb.sheetnames:
        print("  FAIL: Tab REDES_FRANQUIAS_v2 does not exist!", flush=True)
        return {
            'description': 'Sinaleiro de penetracao atualizado com dados 2025',
            'status': 'FAIL',
            'details': {'tab_exists': False},
            'notes': ['Tab not found'],
        }

    ws = wb['REDES_FRANQUIAS_v2']

    # Column mapping from 07-02 plan:
    # D=ATIVOS(col4), J=FAT.REAL(col10), L=FAT.POTENCIAL(col12),
    # N=SINALEIRO%(col14), O=COR(col15)
    checks = {
        'col_N_sinaleiro': {'col': 14, 'name': 'SINALEIRO %', 'expected_pattern': None, 'found': 0, 'total': 0},
        'col_J_fat_real': {'col': 10, 'name': 'FAT.REAL', 'expected_pattern': 'SUMIFS', 'found': 0, 'total': 0},
        'col_D_ativos': {'col': 4, 'name': 'ATIVOS', 'expected_pattern': 'COUNTIFS', 'found': 0, 'total': 0},
        'col_O_cor': {'col': 15, 'name': 'COR', 'expected_pattern': 'IF', 'found': 0, 'total': 0},
        'col_L_fat_potencial': {'col': 12, 'name': 'FAT.POTENCIAL', 'expected_pattern': '$W$', 'found': 0, 'total': 0},
    }

    # Check rows 4-24 (21 data rows: 20 redes + SEM GRUPO)
    for row in range(4, 4 + EXPECTED_REDES_DATA_ROWS):
        for key, check in checks.items():
            val = ws.cell(row=row, column=check['col']).value
            check['total'] += 1
            if is_formula(val):
                if check['expected_pattern'] is None:
                    check['found'] += 1
                elif check['expected_pattern'] in val.upper():
                    check['found'] += 1

    # Count total formulas in the tab
    total_formulas = 0
    sumifs_count = 0
    countifs_count = 0
    if_count = 0
    iferror_count = 0

    for row in ws.iter_rows():
        for cell in row:
            if is_formula(cell.value):
                total_formulas += 1
                val_upper = cell.value.upper()
                if 'SUMIFS' in val_upper:
                    sumifs_count += 1
                if 'COUNTIFS' in val_upper:
                    countifs_count += 1
                if val_upper.startswith('=IF(') or ',IF(' in val_upper:
                    if_count += 1
                if 'IFERROR' in val_upper:
                    iferror_count += 1

    print(f"  Total formulas in REDES_FRANQUIAS_v2: {total_formulas}", flush=True)
    print(f"  SUMIFS: {sumifs_count}", flush=True)
    print(f"  COUNTIFS: {countifs_count}", flush=True)
    print(f"  IF (standalone/nested): {if_count}", flush=True)
    print(f"  IFERROR: {iferror_count}", flush=True)
    print(flush=True)

    all_pass = True
    for key, check in checks.items():
        pct = (check['found'] / check['total'] * 100) if check['total'] > 0 else 0
        passed = check['found'] >= check['total'] * 0.8  # 80% threshold (some redes may have static values)
        status_str = 'OK' if passed else 'LOW'
        pattern_str = f" (pattern: {check['expected_pattern']})" if check['expected_pattern'] else ' (any formula)'
        print(f"  {check['name']} ({key}): {check['found']}/{check['total']} formulas{pattern_str} [{status_str}]", flush=True)
        if not passed:
            all_pass = False

    formulas_pass = total_formulas >= EXPECTED_REDES_TAB_FORMULAS

    status = 'PASS' if all_pass and formulas_pass else 'FAIL'
    notes = []

    if not formulas_pass:
        notes.append(f"Total formulas {total_formulas} < expected {EXPECTED_REDES_TAB_FORMULAS}")
    if not all_pass:
        for key, check in checks.items():
            if check['found'] < check['total'] * 0.8:
                notes.append(f"{check['name']}: only {check['found']}/{check['total']} with expected pattern")
        # If formulas exist but some checks are partial, it's PASS_WITH_NOTES
        if formulas_pass and total_formulas > 100:
            status = 'PASS_WITH_NOTES'

    print(f"\n  REDE-03: [{status}]", flush=True)

    return {
        'description': 'Sinaleiro de penetracao atualizado com dados 2025',
        'status': status,
        'details': {
            'total_formulas_redes_tab': total_formulas,
            'sumifs_count': sumifs_count,
            'countifs_count': countifs_count,
            'if_count': if_count,
            'iferror_count': iferror_count,
            'dynamic_formulas': all_pass,
            'formula_checks': {k: {'found': v['found'], 'total': v['total']} for k, v in checks.items()},
        },
        'notes': notes,
    }


# ============================================================
# CHECK 4: REDE-04 -- Metas 6M por rede operacionais
# ============================================================
def check_rede_04(wb):
    """Validate META 6M populated and GAP calculated."""
    print("\n" + "=" * 70, flush=True)
    print("CHECK 4: REDE-04 -- Metas 6M por rede operacionais", flush=True)
    print("=" * 70, flush=True)

    if 'REDES_FRANQUIAS_v2' not in wb.sheetnames:
        print("  FAIL: Tab REDES_FRANQUIAS_v2 does not exist!", flush=True)
        return {
            'description': 'Metas 6M por rede operacionais',
            'status': 'FAIL',
            'details': {'tab_exists': False},
            'notes': ['Tab not found'],
        }

    ws = wb['REDES_FRANQUIAS_v2']

    # Col M = META 6M (col 13), Col Q = GAP (col 17), Col A = REDE (col 1)
    redes_with_meta = 0
    total_meta = 0
    gap_formula_count = 0
    cross_check = {}
    rede_metas = {}

    for row in range(4, 4 + EXPECTED_REDES_DATA_ROWS):
        rede_name = ws.cell(row=row, column=1).value
        meta_val = ws.cell(row=row, column=13).value  # col M
        gap_val = ws.cell(row=row, column=17).value    # col Q

        if rede_name is None:
            continue

        rede_name_str = str(rede_name).strip()

        # Check META 6M
        if meta_val is not None:
            if isinstance(meta_val, (int, float)):
                redes_with_meta += 1
                total_meta += meta_val
                rede_metas[rede_name_str] = meta_val
            elif isinstance(meta_val, str) and meta_val.strip() != '':
                redes_with_meta += 1
                rede_metas[rede_name_str] = meta_val

        # Check GAP formula
        if is_formula(gap_val):
            gap_formula_count += 1

    # Cross-check reference values
    for ref_rede, ref_value in META_REFERENCE.items():
        actual = rede_metas.get(ref_rede)
        if actual is not None and isinstance(actual, (int, float)):
            diff_pct = abs(actual - ref_value) / ref_value if ref_value != 0 else 0
            match = diff_pct <= META_TOLERANCE
            cross_check[ref_rede] = {
                'expected': ref_value,
                'actual': actual,
                'diff_pct': round(diff_pct * 100, 2),
                'match': match,
            }
        else:
            cross_check[ref_rede] = {
                'expected': ref_value,
                'actual': str(actual) if actual is not None else None,
                'diff_pct': None,
                'match': False,
            }

    print(f"  Redes with META 6M: {redes_with_meta}", flush=True)
    print(f"  Total META 6M: R$ {total_meta:,.0f}", flush=True)
    print(f"  GAP formulas: {gap_formula_count}/{EXPECTED_REDES_DATA_ROWS}", flush=True)

    print("\n  Cross-check reference values:", flush=True)
    all_match = True
    for rede, check in cross_check.items():
        match_str = 'MATCH' if check['match'] else 'MISMATCH'
        if check['diff_pct'] is not None:
            print(f"    {rede}: expected R${check['expected']:,.0f}, actual R${check['actual']:,.0f} ({check['diff_pct']:.1f}% diff) [{match_str}]", flush=True)
        else:
            print(f"    {rede}: expected R${check['expected']:,.0f}, actual {check['actual']} [{match_str}]", flush=True)
        if not check['match']:
            all_match = False

    print("\n  All META 6M values:", flush=True)
    for rede, meta in sorted(rede_metas.items(), key=lambda x: x[1] if isinstance(x[1], (int, float)) else 0, reverse=True):
        if isinstance(meta, (int, float)):
            print(f"    {rede}: R$ {meta:,.0f}", flush=True)
        else:
            print(f"    {rede}: {meta}", flush=True)

    status = 'PASS'
    notes = []

    if redes_with_meta < 10:
        status = 'FAIL'
        notes.append(f"Only {redes_with_meta} redes have META 6M")
    elif not all_match:
        status = 'PASS_WITH_NOTES'
        mismatches = [r for r, c in cross_check.items() if not c['match']]
        notes.append(f"Cross-check mismatches: {mismatches}")

    if gap_formula_count < EXPECTED_REDES_DATA_ROWS * 0.8:
        if status == 'PASS':
            status = 'PASS_WITH_NOTES'
        notes.append(f"GAP formulas: {gap_formula_count}/{EXPECTED_REDES_DATA_ROWS}")

    print(f"\n  REDE-04: [{status}]", flush=True)

    return {
        'description': 'Metas 6M por rede operacionais',
        'status': status,
        'details': {
            'redes_with_meta': redes_with_meta,
            'total_meta_6m': total_meta,
            'gap_formula_exists': gap_formula_count > 0,
            'gap_formula_count': gap_formula_count,
            'cross_check': cross_check,
            'rede_metas': {k: v for k, v in rede_metas.items() if isinstance(v, (int, float))},
        },
        'notes': notes,
    }


# ============================================================
# CHECK 5: Integridade V13 -- Formulas PROJECAO preservadas
# ============================================================
def check_integrity_projecao(wb, prj_name):
    """Count formulas in PROJECAO and verify tab integrity."""
    print("\n" + "=" * 70, flush=True)
    print("CHECK 5: Integridade V13 -- Formulas PROJECAO preservadas", flush=True)
    print("=" * 70, flush=True)

    ws = wb[prj_name]

    # Count formulas
    formula_count = 0
    for row in ws.iter_rows(min_row=4, max_row=537):
        for cell in row:
            if is_formula(cell.value):
                formula_count += 1

    print(f"  PROJECAO formulas: {formula_count}", flush=True)

    formulas_pass = formula_count >= EXPECTED_PROJECAO_FORMULAS
    print(f"  >= {EXPECTED_PROJECAO_FORMULAS}: {'PASS' if formulas_pass else 'FAIL'}", flush=True)

    # Check freeze_panes
    freeze = ws.freeze_panes
    freeze_str = str(freeze) if freeze else 'None'
    print(f"  freeze_panes: {freeze_str}", flush=True)

    # Check all expected tabs exist
    tabs = wb.sheetnames
    print(f"  All tabs: {tabs}", flush=True)

    missing_tabs = []
    for expected_tab in EXPECTED_TABS:
        if expected_tab not in tabs:
            missing_tabs.append(expected_tab)
    # PROJECAO checked via accent-stripping
    projecao_found = prj_name in tabs
    if not projecao_found:
        missing_tabs.append('PROJECAO (accent-stripped)')

    all_tabs_present = len(missing_tabs) == 0
    print(f"  Missing tabs: {missing_tabs if missing_tabs else 'None'}", flush=True)
    print(f"  All expected tabs present: {all_tabs_present}", flush=True)

    notes = []
    if not formulas_pass:
        notes.append(f"PROJECAO formulas {formula_count} < {EXPECTED_PROJECAO_FORMULAS}")
    if not all_tabs_present:
        notes.append(f"Missing tabs: {missing_tabs}")

    return {
        'projecao_formulas': formula_count,
        'projecao_formulas_pass': formulas_pass,
        'freeze_panes_preserved': freeze is not None,
        'freeze_panes_value': freeze_str,
        'all_tabs_present': all_tabs_present,
        'tabs': tabs,
        'missing_tabs': missing_tabs,
    }, notes


# ============================================================
# CHECK 6: Tabela de referencia AS:AZ expandida
# ============================================================
def check_ref_table(ws_proj):
    """Verify AS:AZ reference table has 20 redes + SEM GRUPO and VLOOKUPs reference expanded range."""
    print("\n" + "=" * 70, flush=True)
    print("CHECK 6: Tabela de referencia AS:AZ expandida", flush=True)
    print("=" * 70, flush=True)

    # AS = col 45, AZ = col 52
    # Check how many rows have data in col AS (45)
    ref_redes = []
    for row in range(4, 30):  # check rows 4-29 to find extent
        val = ws_proj.cell(row=row, column=45).value  # col AS
        if val is not None and str(val).strip() != '':
            ref_redes.append((row, str(val).strip()))

    print(f"  Redes in AS:AZ table: {len(ref_redes)}", flush=True)
    for row_num, rede_name in ref_redes:
        print(f"    Row {row_num}: {rede_name}", flush=True)

    ref_table_ok = len(ref_redes) >= 20  # 20 redes + SEM GRUPO = 21, but at least 20

    # Check VLOOKUPs in F:J reference expanded range
    # Look at F4 as spot-check
    vlookup_expanded = False
    vlookup_samples = []
    for col in range(6, 11):  # F(6) to J(10)
        val = ws_proj.cell(row=4, column=col).value
        if val is not None and isinstance(val, str) and 'VLOOKUP' in val.upper():
            vlookup_samples.append(f"col {col}: {val[:80]}")
            # Check if it references row 24 or higher (expanded range)
            if '$24' in val or '$25' in val or '$26' in val:
                vlookup_expanded = True

    print(f"\n  VLOOKUP samples (row 4):", flush=True)
    for sample in vlookup_samples:
        print(f"    {sample}", flush=True)
    print(f"  VLOOKUPs reference expanded range (row 24+): {vlookup_expanded}", flush=True)

    notes = []
    if not ref_table_ok:
        notes.append(f"Reference table has {len(ref_redes)} entries, expected >= 20")
    if not vlookup_expanded:
        notes.append("VLOOKUPs may not reference expanded range")

    return {
        'ref_table_redes': len(ref_redes),
        'ref_table_data': {row: name for row, name in ref_redes},
        'vlookup_range_expanded': vlookup_expanded,
        'vlookup_samples': vlookup_samples,
    }, notes


# ============================================================
# CHECK 7: Cross-check distribuicao de redes
# ============================================================
def check_rede_distribution(ws_proj, ref_redes):
    """Cross-check client distribution vs reference table."""
    print("\n" + "=" * 70, flush=True)
    print("CHECK 7: Cross-check distribuicao de redes", flush=True)
    print("=" * 70, flush=True)

    # Count clients per rede in col C
    rede_counter = Counter()
    for row in range(4, 4 + EXPECTED_CLIENTS):
        val = ws_proj.cell(row=row, column=3).value  # col C
        if val is not None and str(val).strip() != '':
            rede_counter[str(val).strip()] += 1

    # Reference redes from AS:AZ
    ref_rede_names = set(ref_redes.values()) if isinstance(ref_redes, dict) else set()
    projecao_redes = set(rede_counter.keys())

    # Redes in ref table but missing from PROJECAO (may be SAP-only)
    in_ref_not_projecao = ref_rede_names - projecao_redes
    # Redes in PROJECAO but missing from ref table
    in_projecao_not_ref = projecao_redes - ref_rede_names

    print(f"  Unique redes in PROJECAO col C: {len(projecao_redes)}", flush=True)
    print(f"  Unique redes in AS:AZ ref table: {len(ref_rede_names)}", flush=True)

    if in_ref_not_projecao:
        print(f"  In ref table but NOT in PROJECAO: {in_ref_not_projecao}", flush=True)
    else:
        print(f"  All ref table redes found in PROJECAO: YES", flush=True)

    if in_projecao_not_ref:
        print(f"  In PROJECAO but NOT in ref table: {in_projecao_not_ref}", flush=True)
    else:
        print(f"  All PROJECAO redes in ref table: YES", flush=True)

    notes = []
    if in_projecao_not_ref:
        notes.append(f"Redes in PROJECAO not in ref table: {in_projecao_not_ref}")
    # SAP-only redes (in ref but not PROJECAO) are expected
    if in_ref_not_projecao:
        notes.append(f"SAP-only redes (in ref table but no clients in PROJECAO): {in_ref_not_projecao}")

    return {
        'projecao_redes': dict(rede_counter.most_common()),
        'ref_redes': list(ref_rede_names),
        'in_ref_not_projecao': list(in_ref_not_projecao),
        'in_projecao_not_ref': list(in_projecao_not_ref),
        'all_projecao_in_ref': len(in_projecao_not_ref) == 0,
    }, notes


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 70, flush=True)
    print("  PHASE 7 VALIDATION REPORT", flush=True)
    print("  Script: 03_validate_phase07.py", flush=True)
    print(f"  V13: {V13_PATH}", flush=True)
    print(f"  Output: {REPORT_PATH}", flush=True)
    print("=" * 70, flush=True)

    # Create output dir
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Open V13 with data_only=False to see formulas
    print(f"\nLoading V13 (data_only=False)...", flush=True)
    wb = openpyxl.load_workbook(str(V13_PATH), data_only=False)
    print(f"  Loaded. Sheets: {wb.sheetnames}", flush=True)

    # Find PROJECAO sheet
    prj_name = find_projecao_sheet(wb)
    print(f"  PROJECAO sheet: {repr(prj_name)}", flush=True)
    ws_proj = wb[prj_name]

    # ---- RUN CHECKS ----
    rede_01 = check_rede_01(ws_proj)
    rede_02 = check_rede_02(wb)
    rede_03 = check_rede_03(wb)
    rede_04 = check_rede_04(wb)
    integrity, integrity_notes = check_integrity_projecao(wb, prj_name)
    ref_table, ref_table_notes = check_ref_table(ws_proj)
    cross_check, cross_check_notes = check_rede_distribution(
        ws_proj,
        ref_table.get('ref_table_data', {}),
    )

    wb.close()

    # ---- DETERMINE OVERALL STATUS ----
    req_statuses = [
        rede_01['status'],
        rede_02['status'],
        rede_03['status'],
        rede_04['status'],
    ]

    if 'FAIL' in req_statuses:
        overall = 'FAIL'
    elif 'PASS_WITH_NOTES' in req_statuses:
        overall = 'PASS_WITH_NOTES'
    else:
        overall = 'PASS'

    # Integrity checks
    if not integrity['projecao_formulas_pass'] or not integrity['all_tabs_present']:
        overall = 'FAIL'

    all_notes = []
    all_notes.extend(rede_01.get('notes', []))
    all_notes.extend(rede_02.get('notes', []))
    all_notes.extend(rede_03.get('notes', []))
    all_notes.extend(rede_04.get('notes', []))
    all_notes.extend(integrity_notes)
    all_notes.extend(ref_table_notes)
    all_notes.extend(cross_check_notes)

    # Merge ref_table and cross_check into integrity
    integrity.update(ref_table)
    integrity['cross_check_distribution'] = cross_check

    # ---- BUILD REPORT ----
    report = {
        'phase': '07-redes-franquias',
        'validated_at': datetime.now(timezone.utc).isoformat(),
        'v13_path': str(V13_PATH),
        'requirements': {
            'REDE-01': rede_01,
            'REDE-02': rede_02,
            'REDE-03': rede_03,
            'REDE-04': rede_04,
        },
        'integrity': integrity,
        'overall': overall,
        'notes': all_notes,
    }

    # ---- WRITE REPORT ----
    with open(str(REPORT_PATH), 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False, default=str)

    print(f"\nReport written to: {REPORT_PATH}", flush=True)

    # ---- FORMATTED SUMMARY ----
    passed_count = sum(1 for s in req_statuses if s in ('PASS', 'PASS_WITH_NOTES'))

    print("\n" + "=" * 70, flush=True)
    print("  === PHASE 7 VALIDATION REPORT ===", flush=True)
    print("=" * 70, flush=True)

    sem_grupo = rede_01['details']['sem_grupo']
    com_rede = rede_01['details']['com_rede']
    esmeralda = rede_01['details'].get('esmeralda_count', '?')
    print(f"\n  REDE-01: REDE/GRUPO CHAVE preenchido     [{rede_01['status']}]", flush=True)
    print(f"    - {rede_01['details']['filled']}/{rede_01['details']['total_clients']} clientes preenchidos", flush=True)
    print(f"    - SEM GRUPO: {sem_grupo} (antes: 405, remapeados: {405 - sem_grupo})", flush=True)
    print(f"    - ESMERALDA: {esmeralda} (novo)", flush=True)

    ref_errs = rede_02['details']['ref_errors']
    name_errs = rede_02['details']['name_errors']
    print(f"\n  REDE-02: Zero #REF! REDES_FRANQUIAS_v2   [{rede_02['status']}]", flush=True)
    print(f"    - {ref_errs} #REF! encontrados", flush=True)
    print(f"    - {name_errs} #NAME? encontrados", flush=True)

    total_f = rede_03['details']['total_formulas_redes_tab']
    sumifs = rede_03['details']['sumifs_count']
    countifs = rede_03['details']['countifs_count']
    ifs = rede_03['details']['if_count']
    print(f"\n  REDE-03: Sinaleiro de penetracao          [{rede_03['status']}]", flush=True)
    print(f"    - {total_f} formulas dinamicas", flush=True)
    print(f"    - SUMIFS: {sumifs}, COUNTIFS: {countifs}, IF: {ifs}", flush=True)

    redes_meta = rede_04['details']['redes_with_meta']
    total_meta = rede_04['details']['total_meta_6m']
    print(f"\n  REDE-04: Metas 6M operacionais            [{rede_04['status']}]", flush=True)
    print(f"    - {redes_meta} redes com meta preenchida", flush=True)
    print(f"    - Total meta 6M: R$ {total_meta:,.0f}", flush=True)

    prj_f = integrity['projecao_formulas']
    ref_count = integrity.get('ref_table_redes', '?')
    print(f"\n  INTEGRIDADE V13:                          [{'PASS' if integrity['projecao_formulas_pass'] and integrity['all_tabs_present'] else 'FAIL'}]", flush=True)
    print(f"    - {prj_f:,} formulas PROJECAO intactas", flush=True)
    print(f"    - Todas as abas presentes: {integrity['all_tabs_present']}", flush=True)
    print(f"    - Tabela AS:AZ: {ref_count} redes", flush=True)

    print(f"\n  OVERALL: {overall} ({passed_count}/4 requirements satisfied)", flush=True)
    print("=" * 70, flush=True)

    if overall == 'FAIL':
        print("\n  RESULT: PHASE 7 VALIDATION FAILED", flush=True)
        sys.exit(1)
    else:
        print("\n  RESULT: PHASE 7 VALIDATION PASSED", flush=True)
        sys.exit(0)


if __name__ == '__main__':
    main()

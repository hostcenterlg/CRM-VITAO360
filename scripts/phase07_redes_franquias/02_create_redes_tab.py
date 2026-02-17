"""
Phase 07 Plan 02: Create REDES_FRANQUIAS_v2 tab in V13
======================================================
Creates a new tab with dynamic formulas (SUMIFS, COUNTIFS) referencing PROJECAO,
static META 6M from SAP Faturamento, TOTAL LOJAS from SAP Leads, and sinaleiro
de penetracao with color-coded maturity levels.

Columns A:T (20 columns):
  A=REDE, B=CONSULTOR, C=TOTAL LOJAS, D=ATIVOS, E=INAT.REC, F=INAT.ANT,
  G=PROSPECT, H=C/VENDA, I=N.PEDIDOS, J=FAT.REAL, K=TICKET MEDIO,
  L=FAT.POTENCIAL, M=META 6M, N=SINALEIRO%, O=COR, P=MATURIDADE,
  Q=GAP, R=PENETRACAO%, S=ACAO RECOMENDADA, T=PRIORIDADE

20 redes (19 + SEM GRUPO) in rows 4-23, TOTAL in row 24.
PARAMETROS area in V3:X10.
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
import unicodedata
from pathlib import Path
from collections import defaultdict

# ============================================================
# CONFIGURATION
# ============================================================
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
V13_PATH = PROJECT_ROOT / "data" / "output" / "CRM_VITAO360_V13_PROJECAO.xlsx"
SAP_META_PATH = PROJECT_ROOT / "data" / "sources" / "sap" / "BASE_SAP_META_PROJECAO_2026.xlsx"

# Normalization map: SAP grupo chave name -> V13 rede name
REDE_NORMALIZE = {
    '06 - INTERNA - BIO MUNDO': 'BIO MUNDO',
    '06 - INTERNA - CIA DA SAUDE': 'CIA DA SAUDE',
    '06 - INTERNA - DIVINA TERRA': 'DIVINA TERRA',
    '06 - INTERNA - FITLAND': 'FITLAND',
    '06 - INTERNA - MUNDO VERDE': 'MUNDO VERDE',
    '06 - INTERNA - VIDA LEVE': 'VIDA LEVE',
    '06 - INTERNA - NATURVIDA': 'NATURVIDA',
    '06 - INTERNA - TUDO EM GRAOS / VGA': 'TUDO EM GRAOS',
    '06 - INTERNA - TRIP': 'TRIP',
    '06 - INTERNA - MAIS NATURAL': 'MAIS NATURAL',
    '06 - INTERNA - LIGEIRINHO': 'LIGEIRINHO',
    '06 - INTERNA - PROSAUDE': 'PROSAUDE',
    '06 - INTERNA - ARMAZEM FIT STORE': 'ARMAZEM FIT STORE',
    '06 - INTERNA - ESMERALDA': 'ESMERALDA',
    '06 - INTERNA - NOVA GERACAO': 'NOVA GERACAO',
    '06 - INTERNA - MERCOCENTRO': 'MERCOCENTRO',
    '06 - INTERNA - JARDIM DAS ERVAS': 'JARDIM DAS ERVAS',
    '06 - INTERNA - FEDERZONI': 'FEDERZONI',
    '06 - INTERNA - MIX VALI': 'MIX VALI',
    '06 - INTERNA - MINHA QUITANDINHA - SP': 'MINHA QUITANDINHA',
    '06 - SEM GRUPO': 'SEM GRUPO',
}


def normalize_rede_name(sap_name):
    """Normalize SAP grupo chave name to V13 rede name."""
    if not sap_name:
        return None
    sap_name = str(sap_name).strip()
    if sap_name in REDE_NORMALIZE:
        return REDE_NORMALIZE[sap_name]
    if sap_name.startswith('06 - INTERNA - '):
        name = sap_name.replace('06 - INTERNA - ', '').strip()
        if ' - SP' in name:
            name = name.split(' - SP')[0].strip()
        if ' / VGA' in name:
            name = name.split(' / VGA')[0].strip()
        return name
    if sap_name == '06 - SEM GRUPO':
        return 'SEM GRUPO'
    return sap_name


def safe_float(v):
    """Convert value to float safely, defaulting to 0.0."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def find_projecao_sheet(wb):
    """Find PROJECAO sheet by accent-stripping."""
    for name in wb.sheetnames:
        clean = unicodedata.normalize('NFD', name)
        clean = ''.join(c for c in clean if unicodedata.category(c) != 'Mn')
        if 'PROJECAO' in clean.upper():
            return name
    raise ValueError("PROJECAO sheet not found")


# ============================================================
# STEP 1: Extract META 6M from SAP Faturamento
# ============================================================
def extract_meta_6m():
    """Extract META 6M (JAN-JUN 2026) per rede from SAP Faturamento tab."""
    print("=" * 70, flush=True)
    print("STEP 1: Extracting META 6M from SAP Faturamento", flush=True)
    print("=" * 70, flush=True)

    wb = openpyxl.load_workbook(str(SAP_META_PATH), data_only=True)
    ws = wb['Faturamento']

    rede_to_meta = defaultdict(float)

    # Scan all rows: col 12 = Grupo Produto, col 11 = 06 NOME GRUPO CHAVE
    # Cols 14-19 = JAN, FEV, MAR, ABR, MAI, JUN
    for r in range(4, 1600):
        grupo_prod = ws.cell(row=r, column=12).value
        if not grupo_prod or '01. TOTAL' not in str(grupo_prod):
            continue

        grupo_chave = ws.cell(row=r, column=11).value
        if not grupo_chave:
            continue

        rede = normalize_rede_name(str(grupo_chave).strip())
        if not rede:
            continue

        # Sum JAN through JUN (cols 14-19)
        meta_6m = sum(safe_float(ws.cell(row=r, column=c).value) for c in range(14, 20))
        rede_to_meta[rede] += meta_6m

    wb.close()

    print(f"\n  Redes with meta: {len(rede_to_meta)}", flush=True)
    for rede, meta in sorted(rede_to_meta.items(), key=lambda x: -x[1]):
        if rede != 'SEM GRUPO':
            print(f"    {rede}: R$ {meta:,.0f}", flush=True)
    print(f"    SEM GRUPO: R$ {rede_to_meta.get('SEM GRUPO', 0):,.0f}", flush=True)

    return dict(rede_to_meta)


# ============================================================
# STEP 2: Extract TOTAL LOJAS from SAP Leads
# ============================================================
def extract_total_lojas():
    """Extract total lojas per rede from SAP Leads tab."""
    print("\n" + "=" * 70, flush=True)
    print("STEP 2: Extracting TOTAL LOJAS from SAP Leads", flush=True)
    print("=" * 70, flush=True)

    wb = openpyxl.load_workbook(str(SAP_META_PATH), data_only=True)
    ws = wb['Leads']

    rede_to_lojas = {}

    # Col 10 = 06 NOME GRUPO CHAVE, Col 25 = TOTAL LOJAS
    for r in range(4, 100):
        grupo = ws.cell(row=r, column=10).value
        lojas = ws.cell(row=r, column=25).value

        if not grupo:
            continue

        rede = normalize_rede_name(str(grupo).strip())
        if rede and rede != 'SEM GRUPO':
            rede_to_lojas[rede] = int(safe_float(lojas)) if lojas is not None else 0

    wb.close()

    print(f"\n  Redes with lojas: {len(rede_to_lojas)}", flush=True)
    for rede, lojas in sorted(rede_to_lojas.items(), key=lambda x: -x[1]):
        print(f"    {rede}: {lojas} lojas", flush=True)

    return rede_to_lojas


# ============================================================
# STEP 3: Build ordered rede list
# ============================================================
def build_rede_list(rede_to_meta, rede_to_lojas):
    """Build ordered list of 20 redes (19 + SEM GRUPO), sorted by META 6M descending."""
    print("\n" + "=" * 70, flush=True)
    print("STEP 3: Building ordered rede list", flush=True)
    print("=" * 70, flush=True)

    # All known redes (excluding SEM GRUPO)
    all_redes = set(rede_to_meta.keys()) | set(rede_to_lojas.keys())
    all_redes.discard('SEM GRUPO')

    # Add MINHA QUITANDINHA if not already present (it may not be in Leads)
    all_redes.add('MINHA QUITANDINHA')

    rede_list = []
    for rede in all_redes:
        meta = rede_to_meta.get(rede, 0)
        lojas = rede_to_lojas.get(rede, 0)
        rede_list.append({'name': rede, 'meta_6m': meta, 'lojas': lojas})

    # Sort by META 6M descending
    rede_list.sort(key=lambda x: -x['meta_6m'])

    # Add SEM GRUPO at the end
    rede_list.append({
        'name': 'SEM GRUPO',
        'meta_6m': rede_to_meta.get('SEM GRUPO', 0),
        'lojas': 0  # SEM GRUPO has no total lojas
    })

    print(f"\n  Total redes: {len(rede_list)} ({len(rede_list) - 1} + SEM GRUPO)", flush=True)
    for i, rd in enumerate(rede_list):
        print(f"    {i+1}. {rd['name']}: META R$ {rd['meta_6m']:,.0f} | {rd['lojas']} lojas", flush=True)

    if len(rede_list) != 20:
        print(f"  WARNING: Expected 20 redes (19 + SEM GRUPO), got {len(rede_list)}", flush=True)

    return rede_list


# ============================================================
# STEP 4: Create REDES_FRANQUIAS_v2 tab
# ============================================================
def create_redes_tab(wb, prj_name, rede_list):
    """Create the REDES_FRANQUIAS_v2 tab with dynamic formulas."""
    print("\n" + "=" * 70, flush=True)
    print("STEP 4: Creating REDES_FRANQUIAS_v2 tab", flush=True)
    print("=" * 70, flush=True)

    # Remove existing tab if present (idempotency)
    if 'REDES_FRANQUIAS_v2' in wb.sheetnames:
        del wb['REDES_FRANQUIAS_v2']
        print("  Removed existing REDES_FRANQUIAS_v2 tab", flush=True)

    ws = wb.create_sheet('REDES_FRANQUIAS_v2')
    print(f"  Created REDES_FRANQUIAS_v2 tab", flush=True)

    # The formula reference for the PROJECAO sheet
    # Sheet name is 'PROJECAO ' (with cedilla accent and trailing space)
    # In formulas, it must be quoted: 'PROJECAO '!
    prj_ref = f"'{prj_name}'!"
    print(f"  Formula reference: {prj_ref}", flush=True)

    # ---- STYLES ----
    title_font = Font(name='Calibri', size=14, bold=True)
    subtitle_font = Font(name='Calibri', size=10, italic=True)
    header_font = Font(name='Calibri', size=10, bold=True)
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    bold_font = Font(name='Calibri', bold=True)
    param_header_font = Font(name='Calibri', size=11, bold=True, color='4472C4')

    money_format = '#,##0.00'
    pct_format = '0.00%'

    # ---- TITLE (rows 1-2) ----
    ws.merge_cells('A1:T1')
    ws['A1'] = "SINALEIRO DE PENETRACAO -- REDES E FRANQUIAS VITAO (v2)"
    ws['A1'].font = title_font

    ws.merge_cells('A2:T2')
    ws['A2'] = "Dados dinamicos | Atualiza automaticamente da aba PROJECAO | Benchmark: R$525/mes/loja | 11 meses"
    ws['A2'].font = subtitle_font

    # ---- PARAMETROS (V3:X10) ----
    ws['V3'] = "PARAMETROS"
    ws['V3'].font = param_header_font

    ws['V4'] = "Benchmark (R$/mes/loja):"
    ws['W4'] = 525
    ws['W4'].number_format = '#,##0'

    ws['V5'] = "Meses dados:"
    ws['W5'] = 11

    ws['V6'] = "Faixas:"

    ws['V7'] = "ROXO"
    ws['W7'] = "0-1%"
    ws['X7'] = "PROSPECCAO"

    ws['V8'] = "VERMELHO"
    ws['W8'] = "1-40%"
    ws['X8'] = "ATIVACAO/POSITIVACAO"

    ws['V9'] = "AMARELO"
    ws['W9'] = "40-60%"
    ws['X9'] = "SELL OUT"

    ws['V10'] = "VERDE"
    ws['W10'] = "60-100%"
    ws['X10'] = "JBP"

    # ---- HEADERS (row 3, cols A:T = 1:20) ----
    headers = [
        'REDE', 'CONSULTOR', 'TOTAL LOJAS (SAP)', 'ATIVOS', 'INAT.REC', 'INAT.ANT',
        'PROSPECT', 'C/VENDA', 'N.PEDIDOS', 'FAT.REAL (R$)', 'TICKET MEDIO (R$)',
        'FAT.POTENCIAL (R$)', 'META 6M (R$)', 'SINALEIRO %', 'COR', 'MATURIDADE',
        'GAP (R$)', 'PENETRACAO %', 'ACAO RECOMENDADA', 'PRIORIDADE'
    ]

    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    print(f"  Headers written: {len(headers)} columns", flush=True)

    # ---- DATA ROWS (rows 4-23, 20 redes) ----
    formula_count = 0

    for idx, rd in enumerate(rede_list):
        row = 4 + idx
        rede_name = rd['name']
        meta_6m = rd['meta_6m']
        lojas = rd['lojas']

        # Col A (REDE): Static name
        ws.cell(row=row, column=1).value = rede_name

        # Col B (CONSULTOR): Empty (future use)
        # ws.cell(row=row, column=2).value = None

        # Col C (TOTAL LOJAS): Static from SAP Leads
        ws.cell(row=row, column=3).value = lojas

        # Col D (ATIVOS): COUNTIFS formula
        # Count clients of this rede with REALIZADO (col Z) > 0
        f_ativos = f'=COUNTIFS({prj_ref}C$4:C$537,A{row},{prj_ref}Z$4:Z$537,">"&0)'
        ws.cell(row=row, column=4).value = f_ativos
        formula_count += 1

        # Col E (INAT.REC): 0 (Phase 9)
        ws.cell(row=row, column=5).value = 0

        # Col F (INAT.ANT): 0 (Phase 9)
        ws.cell(row=row, column=6).value = 0

        # Col G (PROSPECT): Total Lojas - Ativos - Inat.Rec - Inat.Ant
        f_prospect = f'=C{row}-D{row}-E{row}-F{row}'
        ws.cell(row=row, column=7).value = f_prospect
        formula_count += 1

        # Col H (C/VENDA): Same as ATIVOS (clients with sales > 0)
        f_cvenda = f'=COUNTIFS({prj_ref}C$4:C$537,A{row},{prj_ref}Z$4:Z$537,">"&0)'
        ws.cell(row=row, column=8).value = f_cvenda
        formula_count += 1

        # Col I (N.PEDIDOS): 0 (Phase 9)
        ws.cell(row=row, column=9).value = 0

        # Col J (FAT.REAL): SUMIFS formula
        f_fat_real = f'=SUMIFS({prj_ref}Z$4:Z$537,{prj_ref}C$4:C$537,A{row})'
        ws.cell(row=row, column=10).value = f_fat_real
        ws.cell(row=row, column=10).number_format = money_format
        formula_count += 1

        # Col K (TICKET MEDIO): Fat.Real / C/Venda
        f_ticket = f'=IFERROR(J{row}/H{row},0)'
        ws.cell(row=row, column=11).value = f_ticket
        ws.cell(row=row, column=11).number_format = money_format
        formula_count += 1

        # Col L (FAT.POTENCIAL): Total Lojas * Benchmark * Meses (references PARAMETROS)
        f_fat_pot = f'=C{row}*$W$4*$W$5'
        ws.cell(row=row, column=12).value = f_fat_pot
        ws.cell(row=row, column=12).number_format = money_format
        formula_count += 1

        # Col M (META 6M): Static from SAP Faturamento
        ws.cell(row=row, column=13).value = meta_6m
        ws.cell(row=row, column=13).number_format = money_format

        # Col N (SINALEIRO %): Fat.Real / Fat.Potencial
        f_sinaleiro = f'=IFERROR(J{row}/L{row},0)'
        ws.cell(row=row, column=14).value = f_sinaleiro
        ws.cell(row=row, column=14).number_format = pct_format
        formula_count += 1

        # Col O (COR): Nested IF based on sinaleiro %
        f_cor = f'=IF(N{row}<0.01,"ROXO",IF(N{row}<0.4,"VERMELHO",IF(N{row}<0.6,"AMARELO","VERDE")))'
        ws.cell(row=row, column=15).value = f_cor
        formula_count += 1

        # Col P (MATURIDADE): Derived from COR
        f_mat = f'=IF(O{row}="ROXO","PROSPECCAO",IF(O{row}="VERMELHO","ATIVACAO/POSITIVACAO",IF(O{row}="AMARELO","SELL OUT","JBP")))'
        ws.cell(row=row, column=16).value = f_mat
        formula_count += 1

        # Col Q (GAP): META 6M - Fat.Real (minimum 0)
        f_gap = f'=MAX(M{row}-J{row},0)'
        ws.cell(row=row, column=17).value = f_gap
        ws.cell(row=row, column=17).number_format = money_format
        formula_count += 1

        # Col R (PENETRACAO %): Ativos / Total Lojas
        f_pen = f'=IFERROR(D{row}/C{row},0)'
        ws.cell(row=row, column=18).value = f_pen
        ws.cell(row=row, column=18).number_format = pct_format
        formula_count += 1

        # Col S (ACAO RECOMENDADA): Derived from COR
        f_acao = f'=IF(O{row}="ROXO","Cadastrar e ativar lojas",IF(O{row}="VERMELHO","Aumentar frequencia de compra",IF(O{row}="AMARELO","Expandir mix de produtos","Manter e fidelizar")))'
        ws.cell(row=row, column=19).value = f_acao
        formula_count += 1

        # Col T (PRIORIDADE): RANK by GAP descending
        # Range must cover all data rows: 4 to 3+len(rede_list)
        last_data_row = 3 + len(rede_list)
        f_prio = f'=RANK(Q{row},$Q$4:$Q${last_data_row},0)'
        ws.cell(row=row, column=20).value = f_prio
        formula_count += 1

    print(f"  Data rows written: {len(rede_list)} (rows 4-{3 + len(rede_list)})", flush=True)
    print(f"  Formulas per row: 13 | Total data formulas: {formula_count}", flush=True)

    # ---- TOTAL ROW (row 24) ----
    total_row = 4 + len(rede_list)  # Should be 24 (20 redes in rows 4-23)
    print(f"  TOTAL row: {total_row}", flush=True)

    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = bold_font

    # C24: =SUM(C4:C23)
    ws.cell(row=total_row, column=3).value = f'=SUM(C4:C{total_row - 1})'
    ws.cell(row=total_row, column=3).font = bold_font
    formula_count += 1

    # D24: =SUM(D4:D23)
    ws.cell(row=total_row, column=4).value = f'=SUM(D4:D{total_row - 1})'
    ws.cell(row=total_row, column=4).font = bold_font
    formula_count += 1

    # J24: =SUM(J4:J23)
    ws.cell(row=total_row, column=10).value = f'=SUM(J4:J{total_row - 1})'
    ws.cell(row=total_row, column=10).font = bold_font
    ws.cell(row=total_row, column=10).number_format = money_format
    formula_count += 1

    # L24: =SUM(L4:L23)
    ws.cell(row=total_row, column=12).value = f'=SUM(L4:L{total_row - 1})'
    ws.cell(row=total_row, column=12).font = bold_font
    ws.cell(row=total_row, column=12).number_format = money_format
    formula_count += 1

    # M24: =SUM(M4:M23)
    ws.cell(row=total_row, column=13).value = f'=SUM(M4:M{total_row - 1})'
    ws.cell(row=total_row, column=13).font = bold_font
    ws.cell(row=total_row, column=13).number_format = money_format
    formula_count += 1

    # N24: =IFERROR(J24/L24,0)
    ws.cell(row=total_row, column=14).value = f'=IFERROR(J{total_row}/L{total_row},0)'
    ws.cell(row=total_row, column=14).font = bold_font
    ws.cell(row=total_row, column=14).number_format = pct_format
    formula_count += 1

    # Q24: =SUM(Q4:Q23)
    ws.cell(row=total_row, column=17).value = f'=SUM(Q4:Q{total_row - 1})'
    ws.cell(row=total_row, column=17).font = bold_font
    ws.cell(row=total_row, column=17).number_format = money_format
    formula_count += 1

    # ---- COLUMN WIDTHS ----
    col_widths = {
        'A': 22, 'B': 15,
        'C': 14, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12,
        'J': 16, 'K': 16, 'L': 18, 'M': 16,
        'N': 14, 'O': 14, 'P': 22, 'Q': 16, 'R': 14,
        'S': 32, 'T': 12,
        'V': 24, 'W': 12, 'X': 22,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    print(f"  Total formulas (data + totals): {formula_count}", flush=True)
    return formula_count


# ============================================================
# STEP 5: Validate the created tab
# ============================================================
def validate_redes_tab(wb, prj_name):
    """Validate the REDES_FRANQUIAS_v2 tab after creation."""
    print("\n" + "=" * 70, flush=True)
    print("STEP 5: Validating REDES_FRANQUIAS_v2 tab", flush=True)
    print("=" * 70, flush=True)

    errors = []

    # 5a. Tab exists
    if 'REDES_FRANQUIAS_v2' not in wb.sheetnames:
        errors.append("REDES_FRANQUIAS_v2 tab not found!")
        print("  FAIL: Tab not found!", flush=True)
        return errors

    ws = wb['REDES_FRANQUIAS_v2']
    print("  PASS: Tab exists", flush=True)

    # 5b. Check headers
    headers = [ws.cell(row=3, column=c).value for c in range(1, 21)]
    print(f"  Headers: {headers}", flush=True)
    if headers[0] == 'REDE' and headers[9] == 'FAT.REAL (R$)' and headers[13] == 'SINALEIRO %':
        print("  PASS: Key headers present", flush=True)
    else:
        errors.append(f"Headers mismatch: {headers}")
        print("  FAIL: Headers mismatch", flush=True)

    # 5c. Check redes (20 rows: 4-23)
    redes = []
    for r in range(4, 25):
        rede = ws.cell(row=r, column=1).value
        if rede:
            redes.append(rede)
    print(f"  Redes found: {len(redes)}", flush=True)
    for r_name in redes:
        print(f"    - {r_name}", flush=True)

    if len(redes) == 21:  # 20 data + TOTAL
        print("  PASS: 20 redes + TOTAL", flush=True)
    else:
        errors.append(f"Expected 21 entries (20 redes + TOTAL), got {len(redes)}")
        print(f"  WARN: Expected 21, got {len(redes)}", flush=True)

    # 5d. Check formulas (not hardcoded)
    fat_real_f = ws.cell(row=4, column=10).value  # J4
    sinaleiro_f = ws.cell(row=4, column=14).value  # N4
    ativos_f = ws.cell(row=4, column=4).value  # D4
    print(f"\n  FAT.REAL formula (J4): {fat_real_f}", flush=True)
    print(f"  SINALEIRO formula (N4): {sinaleiro_f}", flush=True)
    print(f"  ATIVOS formula (D4): {ativos_f}", flush=True)

    if fat_real_f and isinstance(fat_real_f, str) and 'SUMIFS' in fat_real_f:
        print("  PASS: FAT.REAL uses SUMIFS", flush=True)
    else:
        errors.append(f"FAT.REAL not a SUMIFS formula: {fat_real_f}")
        print("  FAIL: FAT.REAL not SUMIFS", flush=True)

    if sinaleiro_f and isinstance(sinaleiro_f, str) and 'IFERROR' in sinaleiro_f:
        print("  PASS: SINALEIRO uses IFERROR", flush=True)
    else:
        errors.append(f"SINALEIRO not IFERROR: {sinaleiro_f}")
        print("  FAIL: SINALEIRO not IFERROR", flush=True)

    if ativos_f and isinstance(ativos_f, str) and 'COUNTIFS' in ativos_f:
        print("  PASS: ATIVOS uses COUNTIFS", flush=True)
    else:
        errors.append(f"ATIVOS not COUNTIFS: {ativos_f}")
        print("  FAIL: ATIVOS not COUNTIFS", flush=True)

    # 5e. Check for #REF! errors
    ref_errors = 0
    for row in ws.iter_rows(min_row=1, max_row=25, max_col=22):
        for cell in row:
            if isinstance(cell.value, str) and '#REF' in cell.value:
                ref_errors += 1
                print(f"  #REF! at {cell.coordinate}: {cell.value}", flush=True)

    if ref_errors == 0:
        print("  PASS: Zero #REF! errors", flush=True)
    else:
        errors.append(f"{ref_errors} #REF! errors found")
        print(f"  FAIL: {ref_errors} #REF! errors", flush=True)

    # 5f. Check PROJECAO reference in formulas
    prj_ref_check = f"'{prj_name}'!"
    formulas_with_prj = 0
    for r in range(4, 24):
        for c in [4, 8, 10]:  # D, H, J (formulas that reference PROJECAO)
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and prj_ref_check in val:
                formulas_with_prj += 1
    print(f"  Formulas referencing PROJECAO: {formulas_with_prj}", flush=True)
    if formulas_with_prj >= 40:  # 20 redes * 3 cols minimum
        print("  PASS: PROJECAO references correct", flush=True)
    else:
        errors.append(f"Only {formulas_with_prj} formulas reference PROJECAO")
        print(f"  FAIL: Expected >= 40, got {formulas_with_prj}", flush=True)

    # 5g. Check META 6M values
    print("\n  META 6M values:", flush=True)
    for r in range(4, 24):
        rede = ws.cell(row=r, column=1).value
        meta = ws.cell(row=r, column=13).value
        if rede:
            print(f"    {rede}: {meta}", flush=True)

    # 5h. Check PARAMETROS
    benchmark = ws['W4'].value
    meses = ws['W5'].value
    print(f"\n  PARAMETROS: Benchmark={benchmark}, Meses={meses}", flush=True)
    if benchmark == 525 and meses == 11:
        print("  PASS: PARAMETROS correct", flush=True)
    else:
        errors.append(f"PARAMETROS wrong: benchmark={benchmark}, meses={meses}")
        print("  FAIL: PARAMETROS wrong", flush=True)

    # 5i. Count all formulas in the tab
    total_formulas = 0
    for row in ws.iter_rows(min_row=3, max_row=25, max_col=22):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                total_formulas += 1
    print(f"\n  Total formulas in tab: {total_formulas}", flush=True)

    return errors


# ============================================================
# STEP 6: Validate PROJECAO formulas preserved
# ============================================================
def validate_projecao_formulas(wb, prj_name):
    """Count formulas in PROJECAO to ensure preservation."""
    print("\n" + "=" * 70, flush=True)
    print("STEP 6: Validating PROJECAO formulas preservation", flush=True)
    print("=" * 70, flush=True)

    ws = wb[prj_name]
    formula_count = 0
    for row in ws.iter_rows(min_row=4, max_row=537):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1

    print(f"  PROJECAO formulas: {formula_count}", flush=True)
    if formula_count >= 19224:
        print(f"  PASS: >= 19,224", flush=True)
    else:
        print(f"  FAIL: {formula_count} < 19,224", flush=True)

    return formula_count


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 70, flush=True)
    print("Phase 07 Plan 02: Create REDES_FRANQUIAS_v2 Tab", flush=True)
    print("=" * 70, flush=True)

    # Step 1: Extract META 6M
    rede_to_meta = extract_meta_6m()

    # Step 2: Extract TOTAL LOJAS
    rede_to_lojas = extract_total_lojas()

    # Step 3: Build ordered rede list
    rede_list = build_rede_list(rede_to_meta, rede_to_lojas)

    # Step 4a: Load V13
    print("\n" + "=" * 70, flush=True)
    print("Loading V13 (data_only=False)... this takes ~5 min", flush=True)
    print("=" * 70, flush=True)
    wb = openpyxl.load_workbook(str(V13_PATH), data_only=False)

    prj_name = find_projecao_sheet(wb)
    print(f"  PROJECAO sheet: {repr(prj_name)}", flush=True)
    print(f"  All sheets: {wb.sheetnames}", flush=True)

    # Step 4b: Create the tab
    formula_count = create_redes_tab(wb, prj_name, rede_list)

    # Step 5: Save V13
    print("\n" + "=" * 70, flush=True)
    print("Saving V13... (this takes ~5 min)", flush=True)
    print("=" * 70, flush=True)
    wb.save(str(V13_PATH))
    print(f"  Saved: {V13_PATH}", flush=True)

    # Step 6: Validate (in-memory, no reload needed since we just created it)
    errors = validate_redes_tab(wb, prj_name)

    # Step 7: Validate PROJECAO formulas
    prj_formulas = validate_projecao_formulas(wb, prj_name)

    wb.close()

    # Step 8: Reopen for independent validation
    print("\n" + "=" * 70, flush=True)
    print("STEP 8: Independent validation (reopen V13)", flush=True)
    print("=" * 70, flush=True)
    wb2 = openpyxl.load_workbook(str(V13_PATH), data_only=False)

    # Verify tab exists after reload
    if 'REDES_FRANQUIAS_v2' not in wb2.sheetnames:
        errors.append("REDES_FRANQUIAS_v2 not found after reload!")
        print("  FAIL: Tab not found after reload!", flush=True)
    else:
        print("  PASS: Tab exists after reload", flush=True)
        ws2 = wb2['REDES_FRANQUIAS_v2']
        # Quick spot checks
        j4 = ws2.cell(row=4, column=10).value
        print(f"  J4 (FAT.REAL): {j4}", flush=True)
        n4 = ws2.cell(row=4, column=14).value
        print(f"  N4 (SINALEIRO): {n4}", flush=True)
        a4 = ws2.cell(row=4, column=1).value
        print(f"  A4 (first rede): {a4}", flush=True)

    # Re-count PROJECAO formulas
    prj_name2 = find_projecao_sheet(wb2)
    ws_p = wb2[prj_name2]
    prj_fc = 0
    for row in ws_p.iter_rows(min_row=4, max_row=537):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                prj_fc += 1
    print(f"  PROJECAO formulas (reload): {prj_fc}", flush=True)
    if prj_fc >= 19224:
        print(f"  PASS: >= 19,224", flush=True)
    else:
        errors.append(f"PROJECAO formulas {prj_fc} < 19,224 after reload")
        print(f"  FAIL: {prj_fc} < 19,224", flush=True)

    wb2.close()

    # ---- FINAL REPORT ----
    print("\n" + "=" * 70, flush=True)
    print("EXECUTION REPORT", flush=True)
    print("=" * 70, flush=True)
    print(f"  REDES_FRANQUIAS_v2 tab created: YES", flush=True)
    print(f"  Redes: {len(rede_list)} ({len(rede_list) - 1} + SEM GRUPO)", flush=True)
    print(f"  Columns: 20 (A:T)", flush=True)
    print(f"  Data formulas: {formula_count}", flush=True)
    print(f"  PROJECAO formulas preserved: {prj_fc}", flush=True)
    print(f"  Errors: {len(errors)}", flush=True)

    if errors:
        for e in errors:
            print(f"  ERROR: {e}", flush=True)
        return False

    print("\n  ALL CHECKS PASSED", flush=True)
    return True


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)

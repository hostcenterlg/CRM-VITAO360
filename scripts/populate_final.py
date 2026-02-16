"""
CRM VITAO360 V3 FINAL — Populate with ALL real data.

Sources:
  - POPULADO: 493 clients, 12 months sales, identity, ecommerce, recurrence
  - CARTEIRA OULAR: SAP data (cols 61-71) for 6011 clients
  - LOG FINAL: 17,807 attendance records → DRAFT 2 + LOG

Output: output/CRM_VITAO360_V3_FINAL.xlsx
"""
import os
import sys
import time
import datetime
import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from v3_styles import *
from motor_regras import motor_de_regras, FOLLOW_UP_DIAS, GRUPO_DASH, dia_util

# ═══════════════════════════════════════════════════════════════
# FILE PATHS
# ═══════════════════════════════════════════════════════════════
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
POPULADO_FILE = os.path.join(BASE_DIR, "CRM_INTELIGENTE_VITAO360_POPULADO (2).xlsx")
OULAR_FILE = os.path.join(BASE_DIR, "CARTEIRA DE CLIENTES OULAR.xlsx")
LOG_FILE = os.path.join(BASE_DIR, "preenchimento_do_draft_de_atendimento_LOG_FINAL (1).xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "output", "CRM_VITAO360_V3_100.xlsx")

# ═══════════════════════════════════════════════════════════════
# NORMALIZATION
# ═══════════════════════════════════════════════════════════════
RESULTADO_MAP = {
    "EM ATENDIMENTO": "EM ATENDIMENTO",
    "ORCAMENTO": "ORÇAMENTO", "ORÇAMENTO": "ORÇAMENTO",
    "CADASTRO": "CADASTRO",
    "VENDA/PEDIDO": "VENDA / PEDIDO", "VENDA / PEDIDO": "VENDA / PEDIDO",
    "RELACIONAMENTO": "RELACIONAMENTO",
    "FOLLOW UP 7": "FOLLOW UP 7", "FOLLOW UP 15": "FOLLOW UP 15",
    "SUPORTE": "SUPORTE",
    "NAO ATENDE": "NÃO ATENDE", "NÃO ATENDE": "NÃO ATENDE",
    "NAO RESPONDE": "NÃO RESPONDE", "NÃO RESPONDE": "NÃO RESPONDE",
    "RECUSOU LIGACAO": "RECUSOU LIGAÇÃO", "RECUSOU LIGAÇÃO": "RECUSOU LIGAÇÃO",
    "PERDA/FECHOU LOJA": "PERDA / FECHOU LOJA", "PERDA / FECHOU LOJA": "PERDA / FECHOU LOJA",
    "SEM INTERESSE NO MOMENTO": "NÃO RESPONDE", "SEM INTERESSE": "NÃO RESPONDE",
}
SITUACAO_MAP = {
    "ATIVO": "ATIVO", "EM RISCO": "EM RISCO", "INAT.REC": "INAT.REC",
    "INAT.ANT": "INAT.ANT", "NOVO": "NOVO", "PROSPECT": "PROSPECT",
    "LEAD": "PROSPECT", "EM DESENVOLVIMENTO": "ATIVO",
}
VALID_RESULTADOS = set(RESULTADO_MAP.values())
VALID_SITUACOES = {"ATIVO", "EM RISCO", "INAT.REC", "INAT.ANT", "NOVO", "PROSPECT"}
PRIO_MAP = {"ATIVO": 1, "EM RISCO": 2, "INAT.REC": 3, "INAT.ANT": 4, "PROSPECT": 5, "NOVO": 6}

def safe_str(val):
    if val is None: return ""
    return str(val).strip()

def safe_num(val):
    if val is None: return None
    if isinstance(val, (int, float)): return val
    try: return float(str(val).replace(",", "."))
    except: return None

def safe_date(val):
    if val is None: return None
    if isinstance(val, datetime.datetime): return val.date() if hasattr(val, 'date') else val
    if isinstance(val, datetime.date): return val
    return None

def normalize_cnpj(val):
    """Normalize CNPJ: keep dots/dashes for display consistency."""
    s = safe_str(val)
    if not s: return ""
    return s.strip()

def normalize_resultado(val):
    s = safe_str(val).upper().strip()
    return RESULTADO_MAP.get(s, s)

def normalize_situacao(val):
    s = safe_str(val).upper().strip()
    mapped = SITUACAO_MAP.get(s, s)
    return mapped if mapped in VALID_SITUACOES else "PROSPECT"

def calc_tipo_cliente(meses_pos):
    mp = safe_num(meses_pos) or 0
    if mp == 0: return "PROSPECT"
    if mp == 1: return "NOVO"
    if mp <= 3: return "EM DESENVOLVIMENTO"
    if mp <= 6: return "RECORRENTE"
    return "FIDELIZADO"

def calc_situacao(dias_sem_compra, data_ult_pedido):
    """Calculate SITUAÇÃO from days without purchase."""
    if data_ult_pedido is None: return "PROSPECT"
    d = safe_num(dias_sem_compra)
    if d is None: return "PROSPECT"
    if d <= 50: return "ATIVO"
    if d <= 60: return "EM RISCO"
    if d <= 90: return "INAT.REC"
    return "INAT.ANT"

def calc_oportunidade(itens_carrinho, acesso_b2b):
    ic = safe_num(itens_carrinho) or 0
    ab = safe_num(acesso_b2b) or 0
    if ic > 0: return "🔥 QUENTE"
    if ab > 0: return "🟡 MORNO"
    return "❄️ FRIO"


# ═══════════════════════════════════════════════════════════════
# STEP 1: READ POPULADO (primary source)
# ═══════════════════════════════════════════════════════════════
def read_populado():
    """Read 493 clients from POPULADO with 12 months of sales data."""
    print("  Reading POPULADO...")
    wb = openpyxl.load_workbook(POPULADO_FILE, data_only=True, read_only=True)
    ws = wb["DRAFT 1"]

    clients = []
    for row in ws.iter_rows(min_row=4, max_col=45, values_only=True):
        cnpj = normalize_cnpj(row[1])  # Col B
        if not cnpj: continue

        # Sales: cols 21-32 (indices 20-31) = MAR/25 to FEV/26
        vendas = []
        for i in range(20, 32):
            vendas.append(safe_num(row[i]) if i < len(row) else None)

        client = {
            'nome_fantasia': safe_str(row[0]),
            'cnpj': cnpj,
            'razao_social': safe_str(row[2]),
            'uf': safe_str(row[3]),
            'cidade': safe_str(row[4]),
            'email': safe_str(row[5]),
            'telefone': safe_str(row[6]),
            'data_cadastro': safe_date(row[7]),
            'rede_regional': safe_str(row[8]),
            'consultor': safe_str(row[9]),
            'vendedor_ult_pedido': safe_str(row[10]),
            # COMPRAS
            'dias_sem_compra': safe_num(row[11]),  # Col L
            'data_ult_pedido': safe_date(row[12]),  # Col M
            'valor_ult_pedido': safe_num(row[13]),  # Col N
            # ECOMMERCE
            'acessos_semana': safe_num(row[14]),
            'acesso_b2b': safe_num(row[15]),
            'acessos_portal': safe_num(row[16]),
            'itens_carrinho': safe_num(row[17]),
            'valor_b2b': safe_num(row[18]),
            'oportunidade': safe_str(row[19]),
            # VENDAS 12 months
            'vendas_mes': vendas,  # 12 values: MAR/25..FEV/26
            # RECORRÊNCIA
            'ciclo_medio': safe_num(row[32]),  # Col AG
            'n_compras': safe_num(row[33]),
            'curva_abc': safe_str(row[34]),
            'meses_positivado': safe_num(row[35]),
            'ticket_medio': safe_num(row[36]),
            # ATENDIMENTO
            'ult_registro_mercos': safe_str(row[37]) if row[37] else "",
            'data_ult_atend_mercos': safe_date(row[38]),
            'tipo_atend_mercos': safe_str(row[39]),
            'obs_atend_mercos': safe_str(row[40]),
            # SAP (not in POPULADO - will merge from OULAR)
            'sap_data': {},
        }

        # Calculated fields
        if not client['oportunidade']:
            client['oportunidade'] = calc_oportunidade(client['itens_carrinho'], client['acesso_b2b'])
        client['situacao'] = calc_situacao(client['dias_sem_compra'], client['data_ult_pedido'])
        client['tipo_cliente'] = calc_tipo_cliente(client['meses_positivado'])
        total_vendas = sum(v for v in vendas if v) if vendas else 0
        client['total_periodo'] = total_vendas
        mp = client['meses_positivado'] or 0
        client['media_mensal'] = total_vendas / mp if mp > 0 else 0
        client['meses_lista'] = None

        clients.append(client)

    wb.close()
    print(f"    → {len(clients)} clients read from POPULADO")
    return clients


# ═══════════════════════════════════════════════════════════════
# STEP 2: READ OULAR SAP DATA (supplement)
# ═══════════════════════════════════════════════════════════════
def read_oular_sap():
    """Read SAP data from CARTEIRA OULAR (cols 61-71) indexed by CNPJ."""
    print("  Reading CARTEIRA OULAR SAP data...")
    wb = openpyxl.load_workbook(OULAR_FILE, data_only=True, read_only=True)
    ws = wb["CARTEIRA"]

    sap_by_cnpj = {}
    count = 0
    for row in ws.iter_rows(min_row=4, max_col=72, values_only=True):
        cnpj = normalize_cnpj(row[1])  # Col B
        if not cnpj: continue

        sap = {
            'cod_cliente_sap': safe_str(row[60]) if len(row) > 60 else "",
            'desc_grupo': safe_str(row[61]) if len(row) > 61 else "",
            'gerente_nacional': safe_str(row[62]) if len(row) > 62 else "",
            'representante': safe_str(row[63]) if len(row) > 63 else "",
            'vend_interno': safe_str(row[64]) if len(row) > 64 else "",
            'nome_canal': safe_str(row[65]) if len(row) > 65 else "",
            'tipo_cliente_sap': safe_str(row[66]) if len(row) > 66 else "",
            'macroregiao': safe_str(row[67]) if len(row) > 67 else "",
            'microregiao': safe_str(row[68]) if len(row) > 68 else "",
            'grupo_chave': safe_str(row[69]) if len(row) > 69 else "",
            'venda_sap': safe_num(row[70]) if len(row) > 70 else None,
        }
        # Only store if has actual SAP data
        if sap['cod_cliente_sap'] or sap['grupo_chave']:
            sap_by_cnpj[cnpj] = sap
            count += 1

    wb.close()
    print(f"    → {count} clients with SAP data")
    return sap_by_cnpj


# ═══════════════════════════════════════════════════════════════
# STEP 3: READ LOG ATENDIMENTOS
# ═══════════════════════════════════════════════════════════════
def read_log_atendimentos():
    """Read attendance records from LOG FINAL."""
    print("  Reading LOG de atendimentos...")
    wb = openpyxl.load_workbook(LOG_FILE, data_only=True, read_only=True)
    ws = wb["DRAFT 2 (2)"]

    records = []
    for row in ws.iter_rows(min_row=3, max_col=24, values_only=True):
        cnpj = safe_str(row[3])  # Col D
        nome = safe_str(row[2])  # Col C
        if not cnpj and not nome: continue

        resultado_raw = safe_str(row[17])  # Col R
        resultado = normalize_resultado(resultado_raw)
        situacao = normalize_situacao(row[6])  # Col G

        if resultado not in VALID_RESULTADOS and resultado != "":
            resultado = "EM ATENDIMENTO"

        record = {
            'data': safe_date(row[0]),
            'cnpj': cnpj,
            'nome_fantasia': nome,
            'uf': safe_str(row[4]),
            'consultor': safe_str(row[1]),  # Col B
            'resultado': resultado,
            'motivo': safe_str(row[18]),  # Col S
            'whatsapp': safe_str(row[13]),  # Col N
            'ligacao': safe_str(row[14]),  # Col O
            'lig_atendida': safe_str(row[15]),  # Col P
            'nota_dia': safe_str(row[23]),  # Col X
            'mercos_atualizado': safe_str(row[22]),  # Col W
            'situacao_original': situacao,
            'tentativa_original': safe_str(row[12]),  # Col M
            'estagio_original': safe_str(row[8]),  # Col I
        }
        records.append(record)

    wb.close()
    print(f"    → {len(records)} records read")
    return records


# ═══════════════════════════════════════════════════════════════
# STEP 4: POPULATE DRAFT 1
# ═══════════════════════════════════════════════════════════════
def populate_draft1(wb, clients):
    """Populate DRAFT 1 with 48-column V3 structure."""
    ws = wb["DRAFT 1"]
    print(f"  Populating DRAFT 1 with {len(clients)} clients...")

    for i, c in enumerate(clients):
        r = 4 + i

        # IDENTIDADE (cols 1-11)
        ws.cell(row=r, column=1, value=c['nome_fantasia'])
        ws.cell(row=r, column=2, value=c['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=c['razao_social'])
        ws.cell(row=r, column=4, value=c['uf'])
        ws.cell(row=r, column=5, value=c['cidade'])
        ws.cell(row=r, column=6, value=c['email'])
        ws.cell(row=r, column=7, value=c['telefone'])
        if c['data_cadastro']:
            ws.cell(row=r, column=8, value=c['data_cadastro']).number_format = FMT_DATE
        ws.cell(row=r, column=9, value=c['rede_regional'])
        ws.cell(row=r, column=10, value=c['consultor'])
        ws.cell(row=r, column=11, value=c['vendedor_ult_pedido'])

        # STATUS (cols 12-15)
        ws.cell(row=r, column=12, value=c['situacao'])
        # Apply SITUAÇÃO color
        sit_fill = SITUACAO_FILLS.get(c['situacao'])
        if sit_fill:
            ws.cell(row=r, column=12).fill = sit_fill
            ws.cell(row=r, column=12).font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        ws.cell(row=r, column=13, value=PRIO_MAP.get(c['situacao'], 99))
        ws.cell(row=r, column=14, value=c['dias_sem_compra'])
        ws.cell(row=r, column=15, value=c['ciclo_medio'])

        # COMPRAS (cols 16-18)
        if c['data_ult_pedido']:
            ws.cell(row=r, column=16, value=c['data_ult_pedido']).number_format = FMT_DATE
        if c['valor_ult_pedido']:
            ws.cell(row=r, column=17, value=c['valor_ult_pedido']).number_format = FMT_MONEY
        ws.cell(row=r, column=18, value=c['total_periodo']).number_format = FMT_MONEY

        # ECOMMERCE (cols 19-24)
        ws.cell(row=r, column=19, value=c['acessos_semana'])
        ws.cell(row=r, column=20, value=c['acesso_b2b'])
        ws.cell(row=r, column=21, value=c['acessos_portal'])
        ws.cell(row=r, column=22, value=c['itens_carrinho'])
        if c['valor_b2b']:
            ws.cell(row=r, column=23, value=c['valor_b2b']).number_format = FMT_MONEY
        ws.cell(row=r, column=24, value=c['oportunidade'])

        # VENDAS MÊS A MÊS (cols 25-36) + TOTAL (col 37)
        for j, venda in enumerate(c['vendas_mes']):
            if venda:
                ws.cell(row=r, column=25 + j, value=venda).number_format = FMT_MONEY
        # TOTAL = SUM
        ws.cell(row=r, column=37, value=f'=SUM(Y{r}:AJ{r})').number_format = FMT_MONEY

        # RECORRÊNCIA (cols 38-44)
        ws.cell(row=r, column=38, value=c['n_compras'])
        ws.cell(row=r, column=39, value=c['curva_abc'])
        ws.cell(row=r, column=40, value=c['meses_positivado'])
        ws.cell(row=r, column=41, value=c['media_mensal']).number_format = FMT_MONEY if c['media_mensal'] else FMT_NUMBER
        if c['ticket_medio']:
            ws.cell(row=r, column=42, value=c['ticket_medio']).number_format = FMT_MONEY
        ws.cell(row=r, column=43, value=c['meses_lista'])
        ws.cell(row=r, column=44, value=c['tipo_cliente'])

        # ATENDIMENTO MERCOS (cols 45-48)
        if c['ult_registro_mercos']:
            ws.cell(row=r, column=45, value=c['ult_registro_mercos'])
        if c['data_ult_atend_mercos']:
            ws.cell(row=r, column=46, value=c['data_ult_atend_mercos']).number_format = FMT_DATE
        ws.cell(row=r, column=47, value=c['tipo_atend_mercos'])
        ws.cell(row=r, column=48, value=c['obs_atend_mercos'])

        # Apply border to all data cells
        for col in range(1, 49):
            ws.cell(row=r, column=col).border = THIN_BORDER

    last_row = 4 + len(clients) - 1
    print(f"    → DRAFT 1: rows 4-{last_row} ({len(clients)} clients, 12 months sales)")
    return last_row


# ═══════════════════════════════════════════════════════════════
# STEP 5: POPULATE DRAFT 2
# ═══════════════════════════════════════════════════════════════
def populate_draft2(wb, records):
    """Populate DRAFT 2 with motor de regras."""
    ws = wb["DRAFT 2"]
    print(f"  Populating DRAFT 2 with {len(records)} records...")

    tentativa_por_cnpj = {}
    estagio_por_cnpj = {}

    for i, rec in enumerate(records):
        r = 3 + i

        # 12 MANUAL columns
        if rec['data']:
            ws.cell(row=r, column=1, value=rec['data']).number_format = FMT_DATE
        ws.cell(row=r, column=2, value=rec['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=rec['nome_fantasia'])
        ws.cell(row=r, column=4, value=rec['uf'])
        ws.cell(row=r, column=5, value=rec['consultor'])
        ws.cell(row=r, column=6, value=rec['resultado'])
        ws.cell(row=r, column=7, value=rec['motivo'])
        ws.cell(row=r, column=8, value=rec['whatsapp'])
        ws.cell(row=r, column=9, value=rec['ligacao'])
        ws.cell(row=r, column=10, value=rec['lig_atendida'])
        ws.cell(row=r, column=11, value=rec['nota_dia'])
        ws.cell(row=r, column=12, value=rec['mercos_atualizado'])

        # 12 AUTO columns (motor de regras)
        situacao = rec['situacao_original']
        resultado = rec['resultado']
        cnpj = rec['cnpj']
        tent_ant = tentativa_por_cnpj.get(cnpj)
        est_ant = estagio_por_cnpj.get(cnpj)

        if resultado and resultado in VALID_RESULTADOS:
            motor = motor_de_regras(situacao, resultado, est_ant, tent_ant)
            ws.cell(row=r, column=13, value=situacao)
            ws.cell(row=r, column=14, value=motor.get('estagio_funil', ''))
            ws.cell(row=r, column=15, value=motor.get('fase', ''))
            ws.cell(row=r, column=16, value=motor.get('tipo_contato', ''))
            ws.cell(row=r, column=17, value=motor.get('temperatura', ''))
            tent = motor.get('tentativa')
            ws.cell(row=r, column=18, value=tent or '')
            ws.cell(row=r, column=19, value=motor.get('grupo_dash', ''))
            if rec['data'] and motor.get('follow_up_dias', 0) > 0:
                ws.cell(row=r, column=20, value=dia_util(rec['data'], motor['follow_up_dias'])).number_format = FMT_DATE
            ws.cell(row=r, column=21, value=motor.get('acao_futura', ''))
            acao_det = f"{resultado} → {motor.get('acao_futura', '')}" if resultado else ""
            ws.cell(row=r, column=22, value=acao_det)
            # SINALEIRO CICLO (col 23) - formula
            ws.cell(row=r, column=23).value = (
                f"=IFERROR(IF(XLOOKUP(B{r},'DRAFT 1'!$B$4:$B$1000,'DRAFT 1'!$O$4:$O$1000,0)=0,\"🟣\","
                f"IF(XLOOKUP(B{r},'DRAFT 1'!$B$4:$B$1000,'DRAFT 1'!$N$4:$N$1000,0)"
                f"<=XLOOKUP(B{r},'DRAFT 1'!$B$4:$B$1000,'DRAFT 1'!$O$4:$O$1000,0),\"🟢\","
                f"IF(XLOOKUP(B{r},'DRAFT 1'!$B$4:$B$1000,'DRAFT 1'!$N$4:$N$1000,0)"
                f"<=XLOOKUP(B{r},'DRAFT 1'!$B$4:$B$1000,'DRAFT 1'!$O$4:$O$1000,0)+30,\"🟡\",\"🔴\"))),\"🟣\")"
            )
            ws.cell(row=r, column=24, value='')  # SINALEIRO META placeholder

            tentativa_por_cnpj[cnpj] = tent
            estagio_por_cnpj[cnpj] = motor.get('estagio_funil', '')
        else:
            ws.cell(row=r, column=13, value=situacao)

        # Style auto columns
        for c in range(13, 25):
            ws.cell(row=r, column=c).fill = FILL_CALC
            ws.cell(row=r, column=c).border = THIN_BORDER

    last_row = 3 + len(records) - 1
    print(f"    → DRAFT 2: rows 3-{last_row} ({len(records)} records)")
    return last_row


# ═══════════════════════════════════════════════════════════════
# STEP 6: POPULATE LOG (archive, no formulas)
# ═══════════════════════════════════════════════════════════════
def populate_log(wb, records):
    """Populate LOG tab as archive."""
    ws = wb["LOG"]
    print(f"  Populating LOG with {len(records)} records...")

    for i, rec in enumerate(records):
        r = 3 + i
        if rec['data']:
            ws.cell(row=r, column=1, value=rec['data']).number_format = FMT_DATE
        ws.cell(row=r, column=2, value=rec['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=rec['nome_fantasia'])
        ws.cell(row=r, column=4, value=rec['uf'])
        ws.cell(row=r, column=5, value=rec['consultor'])
        ws.cell(row=r, column=6, value=rec['resultado'])
        ws.cell(row=r, column=7, value=rec['motivo'])
        ws.cell(row=r, column=8, value=rec['whatsapp'])
        ws.cell(row=r, column=9, value=rec['ligacao'])
        ws.cell(row=r, column=10, value=rec['lig_atendida'])
        ws.cell(row=r, column=11, value=rec['nota_dia'])
        ws.cell(row=r, column=12, value=rec['mercos_atualizado'])

        situacao = rec['situacao_original']
        resultado = rec['resultado']
        if resultado and resultado in VALID_RESULTADOS:
            motor = motor_de_regras(situacao, resultado,
                                    rec.get('estagio_original'),
                                    rec.get('tentativa_original'))
            ws.cell(row=r, column=13, value=situacao)
            ws.cell(row=r, column=14, value=motor.get('estagio_funil', ''))
            ws.cell(row=r, column=15, value=motor.get('fase', ''))
            ws.cell(row=r, column=16, value=motor.get('tipo_contato', ''))
            ws.cell(row=r, column=17, value=motor.get('temperatura', ''))
            ws.cell(row=r, column=18, value=motor.get('tentativa') or '')
            ws.cell(row=r, column=19, value=motor.get('grupo_dash', ''))
            if rec['data'] and motor.get('follow_up_dias', 0) > 0:
                ws.cell(row=r, column=20, value=dia_util(rec['data'], motor['follow_up_dias'])).number_format = FMT_DATE
            ws.cell(row=r, column=21, value=motor.get('acao_futura', ''))
            ws.cell(row=r, column=22, value=f"{resultado} → {motor.get('acao_futura', '')}")
        else:
            ws.cell(row=r, column=13, value=situacao)

    print(f"    → LOG: {len(records)} records archived")


# ═══════════════════════════════════════════════════════════════
# STEP 7: POPULATE CARTEIRA
# ═══════════════════════════════════════════════════════════════
def populate_carteira(wb, clients, sap_data):
    """Populate CARTEIRA with XLOOKUP formulas + SAP data."""
    ws = wb["CARTEIRA"]
    print(f"  Populating CARTEIRA with {len(clients)} clients...")

    for i, c in enumerate(clients):
        r = 4 + i
        cnpj = c['cnpj']

        # MEGA-BLOCO 1: MERCOS (cols 1-43)
        ws.cell(row=r, column=1, value=c['nome_fantasia'])
        ws.cell(row=r, column=2, value=c['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=c['razao_social'])
        ws.cell(row=r, column=4, value=c['uf'])
        ws.cell(row=r, column=5, value=c['cidade'])
        ws.cell(row=r, column=6, value=c['email'])
        ws.cell(row=r, column=7, value=c['telefone'])
        if c['data_cadastro']:
            ws.cell(row=r, column=8, value=c['data_cadastro']).number_format = FMT_DATE
        ws.cell(row=r, column=9, value=c['rede_regional'])
        ws.cell(row=r, column=10, value=c['tipo_cliente'])
        if c.get('data_ult_atend_mercos'):
            ws.cell(row=r, column=11, value=c['data_ult_atend_mercos']).number_format = FMT_DATE
        ws.cell(row=r, column=12, value=c['consultor'])
        ws.cell(row=r, column=13, value=c['vendedor_ult_pedido'])
        ws.cell(row=r, column=14, value=c['situacao'])
        # Apply SITUAÇÃO color
        sit_fill = SITUACAO_FILLS.get(c['situacao'])
        if sit_fill:
            ws.cell(row=r, column=14).fill = sit_fill
            ws.cell(row=r, column=14).font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        ws.cell(row=r, column=15, value=PRIO_MAP.get(c['situacao'], 99))
        ws.cell(row=r, column=16, value=c['dias_sem_compra'])
        if c['data_ult_pedido']:
            ws.cell(row=r, column=17, value=c['data_ult_pedido']).number_format = FMT_DATE
        if c['valor_ult_pedido']:
            ws.cell(row=r, column=18, value=c['valor_ult_pedido']).number_format = FMT_MONEY
        ws.cell(row=r, column=19, value=c['acessos_semana'])
        ws.cell(row=r, column=20, value=c['acesso_b2b'])
        ws.cell(row=r, column=21, value=c['acessos_portal'])
        ws.cell(row=r, column=22, value=c['itens_carrinho'])
        if c['valor_b2b']:
            ws.cell(row=r, column=23, value=c['valor_b2b']).number_format = FMT_MONEY
        ws.cell(row=r, column=24, value=c['oportunidade'])
        if c['total_periodo']:
            ws.cell(row=r, column=25, value=c['total_periodo']).number_format = FMT_MONEY
        for j, venda in enumerate(c['vendas_mes']):
            if venda:
                ws.cell(row=r, column=26 + j, value=venda).number_format = FMT_MONEY
        ws.cell(row=r, column=38, value=c['ciclo_medio'])
        ws.cell(row=r, column=39, value=c['n_compras'])
        ws.cell(row=r, column=40, value=c['curva_abc'])
        ws.cell(row=r, column=41, value=c['meses_positivado'])
        if c['media_mensal']:
            ws.cell(row=r, column=42, value=c['media_mensal']).number_format = FMT_MONEY
        if c['ticket_medio']:
            ws.cell(row=r, column=43, value=c['ticket_medio']).number_format = FMT_MONEY

        # MEGA-BLOCO 2: FUNIL (cols 44-61) — XLOOKUP from DRAFT 2
        N = len(clients) + 10  # range safety
        ws.cell(row=r, column=44).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$N$3:$N${N},\"\",,2),\"\")"
        ws.cell(row=r, column=45).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$T$3:$T${N},\"\",,2),\"\")"
        ws.cell(row=r, column=45).number_format = FMT_DATE
        ws.cell(row=r, column=46).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$A$3:$A${N},\"\",,2),\"\")"
        ws.cell(row=r, column=46).number_format = FMT_DATE
        ws.cell(row=r, column=47).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$F$3:$F${N},\"\",,2),\"\")"
        ws.cell(row=r, column=48).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$G$3:$G${N},\"\",,2),\"\")"
        ws.cell(row=r, column=49, value=c['tipo_cliente'])
        ws.cell(row=r, column=50).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$R$3:$R${N},\"\",,2),\"\")"
        ws.cell(row=r, column=51).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$O$3:$O${N},\"\",,2),\"\")"
        ws.cell(row=r, column=52, value='')  # ÚLTIMA RECOMPRA
        ws.cell(row=r, column=53).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$Q$3:$Q${N},\"\",,2),\"\")"
        for col in range(54, 59):
            ws.cell(row=r, column=col, value='')
        ws.cell(row=r, column=59).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$U$3:$U${N},\"\",,2),\"\")"
        ws.cell(row=r, column=60).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$V$3:$V${N},\"\",,2),\"\")"
        ws.cell(row=r, column=61).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$W$3:$W${N},\"\",,2),\"\")"

        # MEGA-BLOCO 3: SAP (cols 62-72)
        sap = sap_data.get(cnpj, c.get('sap_data', {}))
        ws.cell(row=r, column=62, value=sap.get('cod_cliente_sap', ''))
        ws.cell(row=r, column=63, value=sap.get('desc_grupo', ''))
        ws.cell(row=r, column=64, value=sap.get('gerente_nacional', ''))
        ws.cell(row=r, column=65, value=sap.get('representante', ''))
        ws.cell(row=r, column=66, value=sap.get('vend_interno', ''))
        ws.cell(row=r, column=67, value=sap.get('nome_canal', ''))
        ws.cell(row=r, column=68, value=sap.get('tipo_cliente_sap', ''))
        ws.cell(row=r, column=69, value=sap.get('macroregiao', ''))
        ws.cell(row=r, column=70, value=sap.get('microregiao', ''))
        ws.cell(row=r, column=71, value=sap.get('grupo_chave', ''))
        if sap.get('venda_sap'):
            ws.cell(row=r, column=72, value=sap['venda_sap']).number_format = FMT_MONEY

        # Apply borders
        for col in range(1, 73):
            ws.cell(row=r, column=col).border = THIN_BORDER

    last_row = 4 + len(clients) - 1
    print(f"    → CARTEIRA: rows 4-{last_row} (Mercos+Funil+SAP)")
    return last_row


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    start = time.time()
    print("=" * 60)
    print("  CRM VITAO360 V3 FINAL — Build & Populate")
    print("=" * 60)

    # Step 0: Build empty V3 structure
    print("\n[0/7] Building V3 structure...")
    from v3_regras import build_regras
    from v3_draft1 import build_draft1
    from v3_projecao import build_projecao
    from v3_carteira import build_carteira
    from v3_draft2 import build_draft2
    from v3_agenda import build_agenda
    from v3_dash import build_dash
    from v3_log import build_log, build_claude_log

    wb = openpyxl.Workbook()
    build_regras(wb)
    build_draft1(wb)
    build_projecao(wb)
    build_carteira(wb)
    build_draft2(wb)
    build_agenda(wb)
    build_dash(wb)
    build_log(wb)
    build_claude_log(wb)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    print(f"    → Structure: {wb.sheetnames}")

    # Step 1: Read sources
    print("\n[1/7] Reading POPULADO (primary)...")
    clients = read_populado()

    print("\n[2/7] Reading OULAR SAP data...")
    sap_data = read_oular_sap()

    print("\n[3/7] Reading LOG atendimentos...")
    records = read_log_atendimentos()

    # Merge SAP data into clients
    merged_sap = 0
    for c in clients:
        sap = sap_data.get(c['cnpj'])
        if sap:
            c['sap_data'] = sap
            merged_sap += 1
    print(f"    → SAP merged: {merged_sap}/{len(clients)} clients")

    # Step 4: Populate DRAFT 1
    print("\n[4/7] Populating DRAFT 1...")
    populate_draft1(wb, clients)

    # Step 5: Populate DRAFT 2
    print("\n[5/7] Populating DRAFT 2 + motor de regras...")
    populate_draft2(wb, records)

    # Step 6: Populate LOG
    print("\n[6/7] Populating LOG...")
    populate_log(wb, records)

    # Step 7: Populate CARTEIRA
    print("\n[7/7] Populating CARTEIRA...")
    populate_carteira(wb, clients, sap_data)

    # Set CARTEIRA as active
    for i, name in enumerate(wb.sheetnames):
        if name == "CARTEIRA":
            wb.active = i
            break

    # Save
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    elapsed = time.time() - start

    # Stats
    print("\n" + "=" * 60)
    print(f"  SAVED: {OUTPUT_FILE}")
    print(f"  Tabs: {wb.sheetnames}")
    print(f"  DRAFT 1: {len(clients)} clients (12 months sales)")
    print(f"  DRAFT 2: {len(records)} records (motor applied)")
    print(f"  LOG: {len(records)} records (archive)")
    print(f"  CARTEIRA: {len(clients)} clients (Mercos+Funil+SAP)")
    print(f"  SAP merged: {merged_sap} clients")
    print(f"  Time: {elapsed:.1f}s")
    print("=" * 60)
    print("\n✅ CRM VITAO360 V3 FINAL pronto!")
    print("   Abra no Excel para verificar fórmulas e dados.")


if __name__ == "__main__":
    main()

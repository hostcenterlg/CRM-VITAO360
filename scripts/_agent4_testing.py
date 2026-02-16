"""AGENT 4: Cross-Tab Integrity Testing — CRM VITAO360 V3 MERGED"""
import openpyxl
import os
from collections import Counter

FILE = r"c:\Users\User\OneDrive\Área de Trabalho\CLAUDE CODE\output\CRM_VITAO360_V3_MERGED.xlsx"
wb = openpyxl.load_workbook(FILE, data_only=True, read_only=True)

print("=" * 70)
print("AGENT 4: CROSS-TAB INTEGRITY TESTING")
print("=" * 70)

# ── TEST 1: DRAFT 1 ↔ CARTEIRA CONSISTENCY ──
print("\n[TEST 1] DRAFT 1 ↔ CARTEIRA Data Match")
ws1 = wb["DRAFT 1"]
ws3 = wb["CARTEIRA"]

d1_data = {}
for row in ws1.iter_rows(min_row=4, max_col=48, values_only=True):
    cnpj = str(row[1] or "").strip()
    if not cnpj: continue
    d1_data[cnpj] = {
        'nome': str(row[0] or ""),
        'uf': str(row[3] or ""),
        'situacao': str(row[11] or ""),
        'consultor': str(row[9] or ""),
        'dias_sem_compra': row[13],
    }

cart_data = {}
for row in ws3.iter_rows(min_row=4, max_col=72, values_only=True):
    cnpj = str(row[1] or "").strip()
    if not cnpj: continue
    cart_data[cnpj] = {
        'nome': str(row[0] or ""),
        'uf': str(row[3] or ""),
        'situacao': str(row[13] or ""),
        'consultor': str(row[11] or ""),
        'dias_sem_compra': row[15],
    }

mismatches = []
for cnpj in d1_data:
    if cnpj not in cart_data:
        mismatches.append(f"  MISSING in CARTEIRA: {cnpj}")
        continue
    d1 = d1_data[cnpj]
    ca = cart_data[cnpj]
    if d1['situacao'] != ca['situacao']:
        mismatches.append(f"  SITUACAO mismatch {cnpj}: D1={d1['situacao']} vs CART={ca['situacao']}")
    if d1['consultor'] != ca['consultor']:
        mismatches.append(f"  CONSULTOR mismatch {cnpj}: D1={d1['consultor']} vs CART={ca['consultor']}")

print(f"  DRAFT 1 clients: {len(d1_data)}")
print(f"  CARTEIRA clients: {len(cart_data)}")
print(f"  Mismatches: {len(mismatches)}")
for m in mismatches[:10]:
    print(m)

# ── TEST 2: DRAFT 2 CNPJ Coverage ──
print("\n[TEST 2] DRAFT 2 CNPJ Coverage vs DRAFT 1")
ws2 = wb["DRAFT 2"]
d2_cnpjs = set()
for row in ws2.iter_rows(min_row=3, max_col=2, values_only=True):
    cnpj = str(row[1] or "").strip()
    if cnpj: d2_cnpjs.add(cnpj)

d1_in_d2 = len(d1_data.keys() & d2_cnpjs)
d1_not_in_d2 = d1_data.keys() - d2_cnpjs
d2_not_in_d1 = d2_cnpjs - d1_data.keys()

print(f"  DRAFT 1 CNPJs in DRAFT 2: {d1_in_d2}/{len(d1_data)} ({d1_in_d2/len(d1_data)*100:.1f}%)")
print(f"  DRAFT 1 NOT in DRAFT 2: {len(d1_not_in_d2)}")
if d1_not_in_d2:
    for c in list(d1_not_in_d2)[:5]:
        print(f"    {c} — {d1_data[c]['nome'][:30]}")
print(f"  DRAFT 2 NOT in DRAFT 1: {len(d2_not_in_d1)}")
if d2_not_in_d1:
    for c in list(d2_not_in_d1)[:5]:
        print(f"    {c}")

# ── TEST 3: LOG ↔ DRAFT 2 Consistency ──
print("\n[TEST 3] LOG ↔ DRAFT 2 Record Count")
ws_log = wb["LOG"]
log_count = 0
for row in ws_log.iter_rows(min_row=3, max_col=2, values_only=True):
    if row[1]: log_count += 1
d2_count = len(d2_cnpjs)  # already counted above as unique, need total
d2_total = 0
for row in ws2.iter_rows(min_row=3, max_col=2, values_only=True):
    if row[1]: d2_total += 1
print(f"  DRAFT 2 records: {d2_total}")
print(f"  LOG records: {log_count}")
print(f"  Match: {'YES' if d2_total == log_count else 'NO — MISMATCH'}")

# ── TEST 4: ACOMPANHAMENTO Integrity ──
print("\n[TEST 4] CARTEIRA ACOMPANHAMENTO Integrity")
acomp_stats = {
    'has_q1': 0, 'has_meta': 0, 'has_real': 0, 'has_just': 0,
    'has_date': 0, 'total': 0,
}
for row in ws3.iter_rows(min_row=4, max_col=257, values_only=True):
    cnpj = str(row[1] or "").strip() if len(row) > 1 else ""
    if not cnpj: continue
    acomp_stats['total'] += 1

    # V3 col 74 = %Q1 (index 73)
    if len(row) > 73 and row[73] is not None: acomp_stats['has_q1'] += 1
    # Check for META values (col 76, 79, 82 = indices 75, 78, 81)
    for idx in [75, 78, 81]:
        if len(row) > idx and row[idx] is not None:
            acomp_stats['has_meta'] += 1
            break
    # REALIZADO (col 77, 80, 83 = indices 76, 79, 82)
    for idx in [76, 79, 82]:
        if len(row) > idx and row[idx] is not None:
            acomp_stats['has_real'] += 1
            break
    # DATA PEDIDO (col 84 = index 83)
    if len(row) > 83 and row[83] is not None: acomp_stats['has_date'] += 1
    # JUSTIFICATIVAS (cols 85-89 = indices 84-88)
    for idx in range(84, 89):
        if len(row) > idx and row[idx] is not None:
            acomp_stats['has_just'] += 1
            break

t = acomp_stats['total']
print(f"  Total clients: {t}")
print(f"  With %Q1: {acomp_stats['has_q1']} ({acomp_stats['has_q1']/t*100:.1f}%)")
print(f"  With META (month 1): {acomp_stats['has_meta']} ({acomp_stats['has_meta']/t*100:.1f}%)")
print(f"  With REALIZADO (month 1): {acomp_stats['has_real']} ({acomp_stats['has_real']/t*100:.1f}%)")
print(f"  With DATA PEDIDO (month 1): {acomp_stats['has_date']} ({acomp_stats['has_date']/t*100:.1f}%)")
print(f"  With JUSTIFICATIVA: {acomp_stats['has_just']} ({acomp_stats['has_just']/t*100:.1f}%)")

# ── TEST 5: Motor de Regras Validation ──
print("\n[TEST 5] Motor de Regras — Forbidden Combinations")
forbidden = []
for row in ws2.iter_rows(min_row=3, max_col=24, values_only=True):
    cnpj = str(row[1] or "").strip()
    if not cnpj: continue
    resultado = str(row[5] or "").strip()
    situacao = str(row[12] or "").strip()
    fase = str(row[14] or "").strip()
    estagio = str(row[13] or "").strip()
    tipo_contato = str(row[15] or "").strip()
    acao = str(row[20] or "").strip()

    # Check forbidden combos
    if situacao == "PROSPECT":
        if fase in ["RECOMPRA", "CS", "PÓS-VENDA", "SALVAMENTO"]:
            forbidden.append(f"PROSPECT + FASE={fase} (CNPJ={cnpj})")
    if situacao == "ATIVO":
        if fase in ["SALVAMENTO", "RECUPERAÇÃO"]:
            forbidden.append(f"ATIVO + FASE={fase} (CNPJ={cnpj})")
        if acao in ["REATIVAÇÃO", "SALVAMENTO"]:
            forbidden.append(f"ATIVO + ACAO={acao} (CNPJ={cnpj})")
    if resultado == "VENDA / PEDIDO" and fase != "PÓS-VENDA" and fase:
        forbidden.append(f"VENDA + FASE={fase} (CNPJ={cnpj})")

print(f"  Forbidden combinations found: {len(forbidden)}")
if forbidden:
    for f in forbidden[:10]:
        print(f"    {f}")
else:
    print(f"  All motor de regras combinations valid!")

# ── TEST 6: PROJECAO Tab Check ──
print("\n[TEST 6] PROJECAO Tab Status")
ws5 = wb["PROJEÇÃO"]
proj_redes = []
for row in ws5.iter_rows(min_row=4, max_col=3, values_only=True):
    if row[0]:
        proj_redes.append(str(row[0])[:30])
print(f"  Redes in PROJECAO: {len(proj_redes)}")
for r in proj_redes:
    print(f"    {r}")
if not proj_redes:
    print(f"  WARNING: PROJECAO has no data rows!")

# ── TEST 7: Date Range Check ──
print("\n[TEST 7] DRAFT 2 Date Range")
import datetime
min_date = None
max_date = None
for row in ws2.iter_rows(min_row=3, max_col=1, values_only=True):
    d = row[0]
    if d and hasattr(d, 'year'):
        if min_date is None or d < min_date:
            min_date = d
        if max_date is None or d > max_date:
            max_date = d
print(f"  Earliest record: {min_date}")
print(f"  Latest record: {max_date}")
if min_date and max_date:
    span = (max_date - min_date).days
    print(f"  Span: {span} days (~{span//30} months)")

wb.close()
print("\n" + "=" * 70)
print("AGENT 4: CROSS-TAB TESTING COMPLETE")
print("=" * 70)

"""
CRM VITAO360 V3 — MERGE FINAL (Best of POPULADO + V3)

Merges:
  FROM V3 structure: 48-col DRAFT 1, PROJEÇÃO, REGRAS, AGENDA, DASH (formula-based)
  FROM POPULADO: 17,805 DRAFT 2 records, CARTEIRA ACOMPANHAMENTO (cols 73-257)
  FROM OULAR: SAP data (cols 62-72)

Output: output/CRM_VITAO360_V3_MERGED.xlsx

Audit fixes applied:
  FIX 1: CNPJ format normalized (XX.XXX.XXX/XXXX-XX) across ALL tabs
  FIX 2: PROJECAO REALIZADO populated from DRAFT 1 sales by rede
  FIX 3: Motor de regras edge case for PROSPECT+VENDA
"""
import os
import sys
import time
import re
import datetime
import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from v3_styles import *
from motor_regras import motor_de_regras, FOLLOW_UP_DIAS, GRUPO_DASH, dia_util

# ═══════════════════════════════════════════════════════════════
# FILE PATHS
# ═══════════════════════════════════════════════════════════════
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
POPULADO_FILE = os.path.join(BASE_DIR, "CRM_INTELIGENTE_VITAO360_POPULADO (2).xlsx")
OULAR_FILE = os.path.join(BASE_DIR, "CARTEIRA DE CLIENTES OULAR.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "output", "CRM_VITAO360_V3_MERGED.xlsx")

# ═══════════════════════════════════════════════════════════════
# NORMALIZATION
# ═══════════════════════════════════════════════════════════════
RESULTADO_MAP = {
    "EM ATENDIMENTO": "EM ATENDIMENTO",
    "ORCAMENTO": "ORÇAMENTO", "ORÇAMENTO": "ORÇAMENTO",
    "CADASTRO": "CADASTRO",
    "VENDA/PEDIDO": "VENDA / PEDIDO", "VENDA / PEDIDO": "VENDA / PEDIDO",
    "RELACIONAMENTO": "RELACIONAMENTO",
    "FOLLOW UP 7": "FOLLOW UP 7", "FOLLOW UP 15": "FOLLOW UP 15",
    "SUPORTE": "SUPORTE",
    "NAO ATENDE": "NÃO ATENDE", "NÃO ATENDE": "NÃO ATENDE",
    "NAO RESPONDE": "NÃO RESPONDE", "NÃO RESPONDE": "NÃO RESPONDE",
    "RECUSOU LIGACAO": "RECUSOU LIGAÇÃO", "RECUSOU LIGAÇÃO": "RECUSOU LIGAÇÃO",
    "PERDA/FECHOU LOJA": "PERDA / FECHOU LOJA", "PERDA / FECHOU LOJA": "PERDA / FECHOU LOJA",
    "SEM INTERESSE NO MOMENTO": "NÃO RESPONDE", "SEM INTERESSE": "NÃO RESPONDE",
}
SITUACAO_MAP = {
    "ATIVO": "ATIVO", "EM RISCO": "EM RISCO", "INAT.REC": "INAT.REC",
    "INAT.ANT": "INAT.ANT", "NOVO": "NOVO", "PROSPECT": "PROSPECT",
    "LEAD": "PROSPECT", "EM DESENVOLVIMENTO": "ATIVO",
}
VALID_RESULTADOS = set(RESULTADO_MAP.values())
VALID_SITUACOES = {"ATIVO", "EM RISCO", "INAT.REC", "INAT.ANT", "NOVO", "PROSPECT"}
PRIO_MAP = {"ATIVO": 1, "EM RISCO": 2, "INAT.REC": 3, "INAT.ANT": 4, "PROSPECT": 5, "NOVO": 6}


def safe_str(val):
    if val is None: return ""
    return str(val).strip()

def safe_num(val):
    if val is None: return None
    if isinstance(val, (int, float)): return val
    try: return float(str(val).replace(",", "."))
    except: return None

def safe_date(val):
    if val is None: return None
    if isinstance(val, datetime.datetime): return val.date() if hasattr(val, 'date') else val
    if isinstance(val, datetime.date): return val
    return None

def format_cnpj(raw):
    """Normalize CNPJ to XX.XXX.XXX/XXXX-XX format."""
    s = str(raw).strip()
    digits = re.sub(r'\D', '', s)
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}"
    if len(digits) == 11:
        # CPF format — keep as-is (flagged in audit)
        return s
    return s  # return as-is if unexpected length

def normalize_cnpj(val):
    s = safe_str(val)
    if not s: return ""
    return format_cnpj(s)

def normalize_resultado(val):
    s = safe_str(val).upper().strip()
    return RESULTADO_MAP.get(s, s)

def normalize_situacao(val):
    s = safe_str(val).upper().strip()
    mapped = SITUACAO_MAP.get(s, s)
    return mapped if mapped in VALID_SITUACOES else "PROSPECT"

def calc_tipo_cliente(meses_pos):
    mp = safe_num(meses_pos) or 0
    if mp == 0: return "PROSPECT"
    if mp == 1: return "NOVO"
    if mp <= 3: return "EM DESENVOLVIMENTO"
    if mp <= 6: return "RECORRENTE"
    return "FIDELIZADO"

def calc_situacao(dias_sem_compra, data_ult_pedido):
    if data_ult_pedido is None: return "PROSPECT"
    d = safe_num(dias_sem_compra)
    if d is None: return "PROSPECT"
    if d <= 50: return "ATIVO"
    if d <= 60: return "EM RISCO"
    if d <= 90: return "INAT.REC"
    return "INAT.ANT"

def calc_oportunidade(itens_carrinho, acesso_b2b):
    ic = safe_num(itens_carrinho) or 0
    ab = safe_num(acesso_b2b) or 0
    if ic > 0: return "🔥 QUENTE"
    if ab > 0: return "🟡 MORNO"
    return "❄️ FRIO"


# ═══════════════════════════════════════════════════════════════
# STEP 1: READ POPULADO DRAFT 1 (493 clients)
# ═══════════════════════════════════════════════════════════════
def read_populado_clients():
    """Read 493 clients from POPULADO DRAFT 1 with 12 months of sales."""
    print("  Reading POPULADO DRAFT 1...")
    wb = openpyxl.load_workbook(POPULADO_FILE, data_only=True, read_only=True)
    ws = wb["DRAFT 1"]

    clients = []
    for row in ws.iter_rows(min_row=4, max_col=45, values_only=True):
        cnpj = normalize_cnpj(row[1])  # Col B
        if not cnpj: continue

        vendas = []
        for i in range(20, 32):
            vendas.append(safe_num(row[i]) if i < len(row) else None)

        client = {
            'nome_fantasia': safe_str(row[0]),
            'cnpj': cnpj,
            'razao_social': safe_str(row[2]),
            'uf': safe_str(row[3]),
            'cidade': safe_str(row[4]),
            'email': safe_str(row[5]),
            'telefone': safe_str(row[6]),
            'data_cadastro': safe_date(row[7]),
            'rede_regional': safe_str(row[8]),
            'consultor': safe_str(row[9]),
            'vendedor_ult_pedido': safe_str(row[10]),
            'dias_sem_compra': safe_num(row[11]),
            'data_ult_pedido': safe_date(row[12]),
            'valor_ult_pedido': safe_num(row[13]),
            'acessos_semana': safe_num(row[14]),
            'acesso_b2b': safe_num(row[15]),
            'acessos_portal': safe_num(row[16]),
            'itens_carrinho': safe_num(row[17]),
            'valor_b2b': safe_num(row[18]),
            'oportunidade': safe_str(row[19]),
            'vendas_mes': vendas,
            'ciclo_medio': safe_num(row[32]),
            'n_compras': safe_num(row[33]),
            'curva_abc': safe_str(row[34]),
            'meses_positivado': safe_num(row[35]),
            'ticket_medio': safe_num(row[36]),
            'ult_registro_mercos': safe_str(row[37]) if row[37] else "",
            'data_ult_atend_mercos': safe_date(row[38]),
            'tipo_atend_mercos': safe_str(row[39]),
            'obs_atend_mercos': safe_str(row[40]),
            'sap_data': {},
        }

        if not client['oportunidade']:
            client['oportunidade'] = calc_oportunidade(client['itens_carrinho'], client['acesso_b2b'])
        client['situacao'] = calc_situacao(client['dias_sem_compra'], client['data_ult_pedido'])
        client['tipo_cliente'] = calc_tipo_cliente(client['meses_positivado'])
        total_vendas = sum(v for v in vendas if v) if vendas else 0
        client['total_periodo'] = total_vendas
        mp = client['meses_positivado'] or 0
        client['media_mensal'] = total_vendas / mp if mp > 0 else 0
        client['meses_lista'] = None
        clients.append(client)

    wb.close()
    print(f"    -> {len(clients)} clients read from POPULADO DRAFT 1")
    return clients


# ═══════════════════════════════════════════════════════════════
# STEP 2: READ OULAR SAP DATA
# ═══════════════════════════════════════════════════════════════
def read_oular_sap():
    """Read SAP data from CARTEIRA OULAR (cols 61-71) indexed by CNPJ."""
    print("  Reading CARTEIRA OULAR SAP data...")
    wb = openpyxl.load_workbook(OULAR_FILE, data_only=True, read_only=True)
    ws = wb["CARTEIRA"]

    sap_by_cnpj = {}
    count = 0
    for row in ws.iter_rows(min_row=4, max_col=72, values_only=True):
        cnpj = normalize_cnpj(row[1])
        if not cnpj: continue

        sap = {
            'cod_cliente_sap': safe_str(row[60]) if len(row) > 60 else "",
            'desc_grupo': safe_str(row[61]) if len(row) > 61 else "",
            'gerente_nacional': safe_str(row[62]) if len(row) > 62 else "",
            'representante': safe_str(row[63]) if len(row) > 63 else "",
            'vend_interno': safe_str(row[64]) if len(row) > 64 else "",
            'nome_canal': safe_str(row[65]) if len(row) > 65 else "",
            'tipo_cliente_sap': safe_str(row[66]) if len(row) > 66 else "",
            'macroregiao': safe_str(row[67]) if len(row) > 67 else "",
            'microregiao': safe_str(row[68]) if len(row) > 68 else "",
            'grupo_chave': safe_str(row[69]) if len(row) > 69 else "",
            'venda_sap': safe_num(row[70]) if len(row) > 70 else None,
        }
        if sap['cod_cliente_sap'] or sap['grupo_chave']:
            sap_by_cnpj[cnpj] = sap
            count += 1

    wb.close()
    print(f"    -> {count} clients with SAP data")
    return sap_by_cnpj


# ═══════════════════════════════════════════════════════════════
# STEP 3: READ POPULADO DRAFT 2 (17,805 records)
# ═══════════════════════════════════════════════════════════════
def read_populado_draft2():
    """Read ALL attendance records from POPULADO DRAFT 2.

    POPULADO DRAFT 2 column layout (different from V3!):
      Col A (0): DATA
      Col B (1): CONSULTOR
      Col C (2): NOME FANTASIA
      Col D (3): CNPJ
      Col E (4): UF
      Col F (5): REDE / REGIONAL
      Col G (6): SITUACAO
      Col H (7): DIAS SEM COMPRA
      Col I (8): ESTAGIO FUNIL
      Col J (9): TIPO CLIENTE
      Col K (10): FASE
      Col L (11): SINALEIRO
      Col M (12): TENTATIVA
      Col N (13): WHATSAPP
      Col O (14): LIGACAO
      Col P (15): LIGACAO ATENDIDA
      Col Q (16): TIPO DO CONTATO
      Col R (17): RESULTADO
      Col S (18): MOTIVO
      Col T (19): FOLLOW-UP
      Col U (20): ACAO FUTURA
      Col V (21): ACAO DETALHADA
      Col W (22): MERCOS ATUALIZADO
      Col X (23): NOTA DO DIA
    """
    print("  Reading POPULADO DRAFT 2...")
    wb = openpyxl.load_workbook(POPULADO_FILE, data_only=True, read_only=True)
    ws = wb["DRAFT 2"]

    records = []
    for row in ws.iter_rows(min_row=3, max_col=24, values_only=True):
        cnpj = normalize_cnpj(row[3])  # Col D → XX.XXX.XXX/XXXX-XX
        nome = safe_str(row[2])  # Col C
        if not cnpj and not nome:
            continue

        resultado_raw = safe_str(row[17])  # Col R
        resultado = normalize_resultado(resultado_raw)
        if resultado not in VALID_RESULTADOS and resultado != "":
            resultado = "EM ATENDIMENTO"

        situacao = normalize_situacao(row[6])  # Col G
        tentativa = safe_str(row[12])  # Col M
        estagio = safe_str(row[8])  # Col I

        record = {
            'data': safe_date(row[0]),          # A: DATA
            'cnpj': cnpj,                       # D: CNPJ
            'nome_fantasia': nome,              # C: NOME
            'uf': safe_str(row[4]),             # E: UF
            'consultor': safe_str(row[1]),      # B: CONSULTOR
            'resultado': resultado,             # R: RESULTADO
            'motivo': safe_str(row[18]),        # S: MOTIVO
            'whatsapp': safe_str(row[13]),      # N: WHATSAPP
            'ligacao': safe_str(row[14]),       # O: LIGACAO
            'lig_atendida': safe_str(row[15]),  # P: LIG. ATENDIDA
            'nota_dia': safe_str(row[23]),      # X: NOTA DO DIA
            'mercos_atualizado': safe_str(row[22]),  # W: MERCOS
            'situacao_original': situacao,
            'tentativa_original': tentativa,
            'estagio_original': estagio,
        }
        records.append(record)

    wb.close()
    print(f"    -> {len(records)} records from POPULADO DRAFT 2")
    return records


# ═══════════════════════════════════════════════════════════════
# STEP 4: READ POPULADO CARTEIRA ACOMPANHAMENTO (cols 73-256)
# ═══════════════════════════════════════════════════════════════
def read_populado_carteira_acomp():
    """Read ACOMPANHAMENTO data from POPULADO CARTEIRA (cols 73-256).

    POPULADO layout: col 73 = %Q1, then 15 cols per month x 12 = 180, with
    quarterly headers at cols 73, 119, 165, 211.
    V3 layout: col 73 = %ATING, col 74 = %Q1, then months...
    Shift: POPULADO cols 73-256 -> V3 cols 74-257 (offset +1)
    """
    print("  Reading POPULADO CARTEIRA ACOMPANHAMENTO...")
    wb = openpyxl.load_workbook(POPULADO_FILE, data_only=True, read_only=True)
    ws = wb["CARTEIRA"]

    acomp_by_cnpj = {}
    count = 0
    count_with_data = 0

    for row in ws.iter_rows(min_row=4, max_col=257, values_only=True):
        cnpj = normalize_cnpj(row[1])  # Col B
        if not cnpj:
            continue
        count += 1

        # Extract cols 73-256 (indices 72-255)
        acomp_data = {}
        has_any = False
        for c_idx in range(72, min(256, len(row))):
            val = row[c_idx]
            if val is not None:
                acomp_data[c_idx] = val
                has_any = True

        if has_any:
            acomp_by_cnpj[cnpj] = acomp_data
            count_with_data += 1

    wb.close()
    print(f"    -> {count} CARTEIRA rows, {count_with_data} with ACOMP data")
    return acomp_by_cnpj


# ═══════════════════════════════════════════════════════════════
# STEP 5: POPULATE DRAFT 1 (V3 48-col structure)
# ═══════════════════════════════════════════════════════════════
def populate_draft1(wb, clients):
    """Populate DRAFT 1 with 48-column V3 structure."""
    ws = wb["DRAFT 1"]
    print(f"  Populating DRAFT 1 with {len(clients)} clients...")

    for i, c in enumerate(clients):
        r = 4 + i

        # IDENTIDADE (cols 1-11)
        ws.cell(row=r, column=1, value=c['nome_fantasia'])
        ws.cell(row=r, column=2, value=c['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=c['razao_social'])
        ws.cell(row=r, column=4, value=c['uf'])
        ws.cell(row=r, column=5, value=c['cidade'])
        ws.cell(row=r, column=6, value=c['email'])
        ws.cell(row=r, column=7, value=c['telefone'])
        if c['data_cadastro']:
            ws.cell(row=r, column=8, value=c['data_cadastro']).number_format = FMT_DATE
        ws.cell(row=r, column=9, value=c['rede_regional'])
        ws.cell(row=r, column=10, value=c['consultor'])
        ws.cell(row=r, column=11, value=c['vendedor_ult_pedido'])

        # STATUS (cols 12-15)
        ws.cell(row=r, column=12, value=c['situacao'])
        sit_fill = SITUACAO_FILLS.get(c['situacao'])
        if sit_fill:
            ws.cell(row=r, column=12).fill = sit_fill
            ws.cell(row=r, column=12).font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        ws.cell(row=r, column=13, value=PRIO_MAP.get(c['situacao'], 99))
        ws.cell(row=r, column=14, value=c['dias_sem_compra'])
        ws.cell(row=r, column=15, value=c['ciclo_medio'])

        # COMPRAS (cols 16-18)
        if c['data_ult_pedido']:
            ws.cell(row=r, column=16, value=c['data_ult_pedido']).number_format = FMT_DATE
        if c['valor_ult_pedido']:
            ws.cell(row=r, column=17, value=c['valor_ult_pedido']).number_format = FMT_MONEY
        ws.cell(row=r, column=18, value=c['total_periodo']).number_format = FMT_MONEY

        # ECOMMERCE (cols 19-24)
        ws.cell(row=r, column=19, value=c['acessos_semana'])
        ws.cell(row=r, column=20, value=c['acesso_b2b'])
        ws.cell(row=r, column=21, value=c['acessos_portal'])
        ws.cell(row=r, column=22, value=c['itens_carrinho'])
        if c['valor_b2b']:
            ws.cell(row=r, column=23, value=c['valor_b2b']).number_format = FMT_MONEY
        ws.cell(row=r, column=24, value=c['oportunidade'])

        # VENDAS MES A MES (cols 25-36) + TOTAL (col 37)
        for j, venda in enumerate(c['vendas_mes']):
            if venda:
                ws.cell(row=r, column=25 + j, value=venda).number_format = FMT_MONEY
        ws.cell(row=r, column=37, value=f'=SUM(Y{r}:AJ{r})').number_format = FMT_MONEY

        # RECORRENCIA (cols 38-44)
        ws.cell(row=r, column=38, value=c['n_compras'])
        ws.cell(row=r, column=39, value=c['curva_abc'])
        ws.cell(row=r, column=40, value=c['meses_positivado'])
        ws.cell(row=r, column=41, value=c['media_mensal']).number_format = FMT_MONEY if c['media_mensal'] else FMT_NUMBER
        if c['ticket_medio']:
            ws.cell(row=r, column=42, value=c['ticket_medio']).number_format = FMT_MONEY
        ws.cell(row=r, column=43, value=c['meses_lista'])
        ws.cell(row=r, column=44, value=c['tipo_cliente'])

        # ATENDIMENTO MERCOS (cols 45-48)
        if c['ult_registro_mercos']:
            ws.cell(row=r, column=45, value=c['ult_registro_mercos'])
        if c['data_ult_atend_mercos']:
            ws.cell(row=r, column=46, value=c['data_ult_atend_mercos']).number_format = FMT_DATE
        ws.cell(row=r, column=47, value=c['tipo_atend_mercos'])
        ws.cell(row=r, column=48, value=c['obs_atend_mercos'])

        for col in range(1, 49):
            ws.cell(row=r, column=col).border = THIN_BORDER

    last_row = 4 + len(clients) - 1
    print(f"    -> DRAFT 1: rows 4-{last_row} ({len(clients)} clients, 48 cols)")
    return last_row


# ═══════════════════════════════════════════════════════════════
# STEP 6: POPULATE DRAFT 2 (V3 layout + motor de regras)
# ═══════════════════════════════════════════════════════════════
def populate_draft2(wb, records):
    """Populate DRAFT 2 with motor de regras (V3 layout)."""
    ws = wb["DRAFT 2"]
    print(f"  Populating DRAFT 2 with {len(records)} records...")

    tentativa_por_cnpj = {}
    estagio_por_cnpj = {}

    for i, rec in enumerate(records):
        r = 3 + i

        # 12 MANUAL columns (V3 layout)
        if rec['data']:
            ws.cell(row=r, column=1, value=rec['data']).number_format = FMT_DATE
        ws.cell(row=r, column=2, value=rec['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=rec['nome_fantasia'])
        ws.cell(row=r, column=4, value=rec['uf'])
        ws.cell(row=r, column=5, value=rec['consultor'])
        ws.cell(row=r, column=6, value=rec['resultado'])
        ws.cell(row=r, column=7, value=rec['motivo'])
        ws.cell(row=r, column=8, value=rec['whatsapp'])
        ws.cell(row=r, column=9, value=rec['ligacao'])
        ws.cell(row=r, column=10, value=rec['lig_atendida'])
        ws.cell(row=r, column=11, value=rec['nota_dia'])
        ws.cell(row=r, column=12, value=rec['mercos_atualizado'])

        # 12 AUTO columns (motor de regras)
        situacao = rec['situacao_original']
        resultado = rec['resultado']
        cnpj = rec['cnpj']
        tent_ant = tentativa_por_cnpj.get(cnpj)
        est_ant = estagio_por_cnpj.get(cnpj)

        if resultado and resultado in VALID_RESULTADOS:
            motor = motor_de_regras(situacao, resultado, est_ant, tent_ant)
            ws.cell(row=r, column=13, value=situacao)
            ws.cell(row=r, column=14, value=motor.get('estagio_funil', ''))
            ws.cell(row=r, column=15, value=motor.get('fase', ''))
            ws.cell(row=r, column=16, value=motor.get('tipo_contato', ''))
            ws.cell(row=r, column=17, value=motor.get('temperatura', ''))
            tent = motor.get('tentativa')
            ws.cell(row=r, column=18, value=tent or '')
            ws.cell(row=r, column=19, value=motor.get('grupo_dash', ''))
            if rec['data'] and motor.get('follow_up_dias', 0) > 0:
                ws.cell(row=r, column=20, value=dia_util(rec['data'], motor['follow_up_dias'])).number_format = FMT_DATE
            ws.cell(row=r, column=21, value=motor.get('acao_futura', ''))
            acao_det = f"{resultado} -> {motor.get('acao_futura', '')}" if resultado else ""
            ws.cell(row=r, column=22, value=acao_det)
            # SINALEIRO CICLO (col 23) - formula
            ws.cell(row=r, column=23).value = (
                f"=IFERROR(IF(XLOOKUP(B{r},'DRAFT 1'!$B$4:$B$1000,'DRAFT 1'!$O$4:$O$1000,0)=0,\"🟣\","
                f"IF(XLOOKUP(B{r},'DRAFT 1'!$B$4:$B$1000,'DRAFT 1'!$N$4:$N$1000,0)"
                f"<=XLOOKUP(B{r},'DRAFT 1'!$B$4:$B$1000,'DRAFT 1'!$O$4:$O$1000,0),\"🟢\","
                f"IF(XLOOKUP(B{r},'DRAFT 1'!$B$4:$B$1000,'DRAFT 1'!$N$4:$N$1000,0)"
                f"<=XLOOKUP(B{r},'DRAFT 1'!$B$4:$B$1000,'DRAFT 1'!$O$4:$O$1000,0)+30,\"🟡\",\"🔴\"))),\"🟣\")"
            )
            ws.cell(row=r, column=24, value='')  # SINALEIRO META placeholder

            tentativa_por_cnpj[cnpj] = tent
            estagio_por_cnpj[cnpj] = motor.get('estagio_funil', '')
        else:
            ws.cell(row=r, column=13, value=situacao)

        # Style auto columns
        for c in range(13, 25):
            ws.cell(row=r, column=c).fill = FILL_CALC
            ws.cell(row=r, column=c).border = THIN_BORDER

    last_row = 3 + len(records) - 1
    print(f"    -> DRAFT 2: rows 3-{last_row} ({len(records)} records + motor)")
    return last_row


# ═══════════════════════════════════════════════════════════════
# STEP 7: POPULATE LOG (archive copy)
# ═══════════════════════════════════════════════════════════════
def populate_log(wb, records):
    """Populate LOG tab as archive (same records, no formulas)."""
    ws = wb["LOG"]
    print(f"  Populating LOG with {len(records)} records...")

    for i, rec in enumerate(records):
        r = 3 + i
        if rec['data']:
            ws.cell(row=r, column=1, value=rec['data']).number_format = FMT_DATE
        ws.cell(row=r, column=2, value=rec['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=rec['nome_fantasia'])
        ws.cell(row=r, column=4, value=rec['uf'])
        ws.cell(row=r, column=5, value=rec['consultor'])
        ws.cell(row=r, column=6, value=rec['resultado'])
        ws.cell(row=r, column=7, value=rec['motivo'])
        ws.cell(row=r, column=8, value=rec['whatsapp'])
        ws.cell(row=r, column=9, value=rec['ligacao'])
        ws.cell(row=r, column=10, value=rec['lig_atendida'])
        ws.cell(row=r, column=11, value=rec['nota_dia'])
        ws.cell(row=r, column=12, value=rec['mercos_atualizado'])

        situacao = rec['situacao_original']
        resultado = rec['resultado']
        if resultado and resultado in VALID_RESULTADOS:
            motor = motor_de_regras(situacao, resultado,
                                    rec.get('estagio_original'),
                                    rec.get('tentativa_original'))
            ws.cell(row=r, column=13, value=situacao)
            ws.cell(row=r, column=14, value=motor.get('estagio_funil', ''))
            ws.cell(row=r, column=15, value=motor.get('fase', ''))
            ws.cell(row=r, column=16, value=motor.get('tipo_contato', ''))
            ws.cell(row=r, column=17, value=motor.get('temperatura', ''))
            ws.cell(row=r, column=18, value=motor.get('tentativa') or '')
            ws.cell(row=r, column=19, value=motor.get('grupo_dash', ''))
            if rec['data'] and motor.get('follow_up_dias', 0) > 0:
                ws.cell(row=r, column=20, value=dia_util(rec['data'], motor['follow_up_dias'])).number_format = FMT_DATE
            ws.cell(row=r, column=21, value=motor.get('acao_futura', ''))
            ws.cell(row=r, column=22, value=f"{resultado} -> {motor.get('acao_futura', '')}")
        else:
            ws.cell(row=r, column=13, value=situacao)

    print(f"    -> LOG: {len(records)} records archived")


# ═══════════════════════════════════════════════════════════════
# STEP 8: POPULATE CARTEIRA (Mercos + Funil + SAP + ACOMPANHAMENTO)
# ═══════════════════════════════════════════════════════════════
def populate_carteira(wb, clients, sap_data, acomp_data):
    """Populate CARTEIRA with XLOOKUP + SAP + ACOMPANHAMENTO data."""
    ws = wb["CARTEIRA"]
    print(f"  Populating CARTEIRA with {len(clients)} clients...")

    acomp_matched = 0

    for i, c in enumerate(clients):
        r = 4 + i
        cnpj = c['cnpj']

        # MEGA-BLOCO 1: MERCOS (cols 1-43) — same as before
        ws.cell(row=r, column=1, value=c['nome_fantasia'])
        ws.cell(row=r, column=2, value=c['cnpj']).number_format = FMT_TEXT
        ws.cell(row=r, column=3, value=c['razao_social'])
        ws.cell(row=r, column=4, value=c['uf'])
        ws.cell(row=r, column=5, value=c['cidade'])
        ws.cell(row=r, column=6, value=c['email'])
        ws.cell(row=r, column=7, value=c['telefone'])
        if c['data_cadastro']:
            ws.cell(row=r, column=8, value=c['data_cadastro']).number_format = FMT_DATE
        ws.cell(row=r, column=9, value=c['rede_regional'])
        ws.cell(row=r, column=10, value=c['tipo_cliente'])
        if c.get('data_ult_atend_mercos'):
            ws.cell(row=r, column=11, value=c['data_ult_atend_mercos']).number_format = FMT_DATE
        ws.cell(row=r, column=12, value=c['consultor'])
        ws.cell(row=r, column=13, value=c['vendedor_ult_pedido'])
        ws.cell(row=r, column=14, value=c['situacao'])
        sit_fill = SITUACAO_FILLS.get(c['situacao'])
        if sit_fill:
            ws.cell(row=r, column=14).fill = sit_fill
            ws.cell(row=r, column=14).font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        ws.cell(row=r, column=15, value=PRIO_MAP.get(c['situacao'], 99))
        ws.cell(row=r, column=16, value=c['dias_sem_compra'])
        if c['data_ult_pedido']:
            ws.cell(row=r, column=17, value=c['data_ult_pedido']).number_format = FMT_DATE
        if c['valor_ult_pedido']:
            ws.cell(row=r, column=18, value=c['valor_ult_pedido']).number_format = FMT_MONEY
        ws.cell(row=r, column=19, value=c['acessos_semana'])
        ws.cell(row=r, column=20, value=c['acesso_b2b'])
        ws.cell(row=r, column=21, value=c['acessos_portal'])
        ws.cell(row=r, column=22, value=c['itens_carrinho'])
        if c['valor_b2b']:
            ws.cell(row=r, column=23, value=c['valor_b2b']).number_format = FMT_MONEY
        ws.cell(row=r, column=24, value=c['oportunidade'])
        if c['total_periodo']:
            ws.cell(row=r, column=25, value=c['total_periodo']).number_format = FMT_MONEY
        for j, venda in enumerate(c['vendas_mes']):
            if venda:
                ws.cell(row=r, column=26 + j, value=venda).number_format = FMT_MONEY
        ws.cell(row=r, column=38, value=c['ciclo_medio'])
        ws.cell(row=r, column=39, value=c['n_compras'])
        ws.cell(row=r, column=40, value=c['curva_abc'])
        ws.cell(row=r, column=41, value=c['meses_positivado'])
        if c['media_mensal']:
            ws.cell(row=r, column=42, value=c['media_mensal']).number_format = FMT_MONEY
        if c['ticket_medio']:
            ws.cell(row=r, column=43, value=c['ticket_medio']).number_format = FMT_MONEY

        # MEGA-BLOCO 2: FUNIL (cols 44-61) — XLOOKUP from DRAFT 2
        N = 20000  # cover all DRAFT 2 rows
        ws.cell(row=r, column=44).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$N$3:$N${N},\"\",,2),\"\")"
        ws.cell(row=r, column=45).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$T$3:$T${N},\"\",,2),\"\")"
        ws.cell(row=r, column=45).number_format = FMT_DATE
        ws.cell(row=r, column=46).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$A$3:$A${N},\"\",,2),\"\")"
        ws.cell(row=r, column=46).number_format = FMT_DATE
        ws.cell(row=r, column=47).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$F$3:$F${N},\"\",,2),\"\")"
        ws.cell(row=r, column=48).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$G$3:$G${N},\"\",,2),\"\")"
        ws.cell(row=r, column=49, value=c['tipo_cliente'])
        ws.cell(row=r, column=50).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$R$3:$R${N},\"\",,2),\"\")"
        ws.cell(row=r, column=51).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$O$3:$O${N},\"\",,2),\"\")"
        ws.cell(row=r, column=52, value='')
        ws.cell(row=r, column=53).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$Q$3:$Q${N},\"\",,2),\"\")"
        for col in range(54, 59):
            ws.cell(row=r, column=col, value='')
        ws.cell(row=r, column=59).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$U$3:$U${N},\"\",,2),\"\")"
        ws.cell(row=r, column=60).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$V$3:$V${N},\"\",,2),\"\")"
        ws.cell(row=r, column=61).value = f"=IFERROR(XLOOKUP(B{r},'DRAFT 2'!$B$3:$B${N},'DRAFT 2'!$W$3:$W${N},\"\",,2),\"\")"

        # MEGA-BLOCO 3: SAP (cols 62-72)
        sap = sap_data.get(cnpj, c.get('sap_data', {}))
        ws.cell(row=r, column=62, value=sap.get('cod_cliente_sap', ''))
        ws.cell(row=r, column=63, value=sap.get('desc_grupo', ''))
        ws.cell(row=r, column=64, value=sap.get('gerente_nacional', ''))
        ws.cell(row=r, column=65, value=sap.get('representante', ''))
        ws.cell(row=r, column=66, value=sap.get('vend_interno', ''))
        ws.cell(row=r, column=67, value=sap.get('nome_canal', ''))
        ws.cell(row=r, column=68, value=sap.get('tipo_cliente_sap', ''))
        ws.cell(row=r, column=69, value=sap.get('macroregiao', ''))
        ws.cell(row=r, column=70, value=sap.get('microregiao', ''))
        ws.cell(row=r, column=71, value=sap.get('grupo_chave', ''))
        if sap.get('venda_sap'):
            ws.cell(row=r, column=72, value=sap['venda_sap']).number_format = FMT_MONEY

        # MEGA-BLOCO 4: ACOMPANHAMENTO (cols 73-257)
        # POPULADO cols 73-256 map to V3 cols 74-257 (offset +1)
        acomp = acomp_data.get(cnpj, {})
        if acomp:
            acomp_matched += 1
            for pop_idx, val in acomp.items():
                # pop_idx is 0-based index from POPULADO (72 = col 73, etc.)
                # POPULADO col (pop_idx + 1) -> V3 col (pop_idx + 2) because offset +1
                v3_col = pop_idx + 2  # pop_idx 72 = POPULADO col 73 -> V3 col 74
                cell = ws.cell(row=r, column=v3_col, value=val)
                # Format dates and numbers
                if isinstance(val, (datetime.datetime, datetime.date)):
                    cell.number_format = FMT_DATE
                elif isinstance(val, float):
                    if val < 2 and val >= 0:  # likely percentage
                        cell.number_format = FMT_PCT
                    else:
                        cell.number_format = FMT_MONEY

        # Apply borders to all populated cols
        for col in range(1, 73):
            ws.cell(row=r, column=col).border = THIN_BORDER

    last_row = 4 + len(clients) - 1
    print(f"    -> CARTEIRA: rows 4-{last_row} (Mercos+Funil+SAP+ACOMP)")
    print(f"    -> ACOMP matched: {acomp_matched}/{len(clients)} clients")
    return last_row


# ═══════════════════════════════════════════════════════════════
# STEP 9: POPULATE PROJECAO REALIZADO (from DRAFT 1 sales by rede)
# ═══════════════════════════════════════════════════════════════
# Rede mapping: DRAFT 1 REDE/REGIONAL → PROJECAO GRUPO CHAVE
REDE_MAP = {
    "CIA DA SAUDE": "CIA DA SAUDE", "CIA DA SAÚDE": "CIA DA SAUDE",
    "FITLAND": "FITLAND", "FIT LAND": "FITLAND",
    "DIVINA TERRA": "DIVINA TERRA",
    "VIDA LEVE": "VIDA LEVE",
    "BIO MUNDO": "BIO MUNDO", "BIOMUNDO": "BIO MUNDO",
    "TUDO EM GRAOS": "TUDO EM GRAOS / VGA", "VGA": "TUDO EM GRAOS / VGA",
    "TUDO EM GRÃOS": "TUDO EM GRAOS / VGA",
    "ARMAZEM": "ARMAZEM FIT STORE", "ARMAZÉM": "ARMAZEM FIT STORE",
    "MUNDO VERDE": "MUNDO VERDE",
    "NATURVIDA": "NATURVIDA",
    "ESMERALDA": "ESMERALDA",
    "TRIP": "TRIP",
    "LIGEIRINHO": "LIGEIRINHO",
    "MERCOCENTRO": "MERCOCENTRO",
}

def match_rede(rede_raw):
    """Match a DRAFT 1 REDE/REGIONAL to PROJECAO GRUPO CHAVE."""
    if not rede_raw:
        return None
    rede_upper = rede_raw.upper().strip()
    # Direct match
    if rede_upper in REDE_MAP:
        return REDE_MAP[rede_upper]
    # Partial match
    for key, val in REDE_MAP.items():
        if key in rede_upper:
            return val
    return None

def populate_projecao(wb, clients):
    """Populate PROJECAO REALIZADO columns from DRAFT 1 vendas by rede."""
    ws = wb["PROJEÇÃO"]
    print(f"  Populating PROJECAO REALIZADO...")

    # Aggregate vendas by rede and month
    # DRAFT 1 vendas_mes: 12 values MAR/25 to FEV/26
    # PROJECAO months: JAN-DEZ (calendar year 2026)
    # Mapping: DRAFT 1 vendas_mes[10]=JAN/26, [11]=FEV/26, [0-9]=MAR/25-DEZ/25

    vendas_por_rede = {}  # {rede: [jan, feb, mar, ...dec]}
    matched_clients = 0

    for c in clients:
        rede = match_rede(c.get('rede_regional', ''))
        if not rede:
            continue
        matched_clients += 1

        if rede not in vendas_por_rede:
            vendas_por_rede[rede] = [0.0] * 12

        # Map DRAFT 1 months to calendar months
        # vendas_mes[0]=MAR/25, [1]=ABR/25, ..., [9]=DEZ/25, [10]=JAN/26, [11]=FEV/26
        vendas = c.get('vendas_mes', [])
        if not vendas or len(vendas) < 12:
            continue

        # PROJECAO calendar: JAN=0, FEV=1, MAR=2, ..., DEZ=11
        # JAN/26 = vendas[10]
        if vendas[10]: vendas_por_rede[rede][0] += vendas[10]
        # FEV/26 = vendas[11]
        if vendas[11]: vendas_por_rede[rede][1] += vendas[11]
        # MAR/25 through DEZ/25 map to MAR-DEZ in PROJECAO (cols 2-11)
        for m in range(10):  # MAR=2 to DEZ=11
            if vendas[m]:
                vendas_por_rede[rede][m + 2] += vendas[m]

    # Write to PROJECAO
    # PROJECAO row structure: rows 4-16 = 13 redes (from v3_projecao.py REDES_META)
    from v3_projecao import REDES_META
    redes_updated = 0
    for r_idx, (rede_name, _, _, _) in enumerate(REDES_META):
        r = 4 + r_idx
        vendas = vendas_por_rede.get(rede_name)
        if not vendas:
            continue
        redes_updated += 1

        # REALIZADO ANUAL (col 23)
        total_real = sum(vendas)
        ws.cell(row=r, column=23, value=total_real).number_format = FMT_MONEY

        # REALIZADO monthly (cols 24-35 = JAN-DEZ)
        for m in range(12):
            if vendas[m] > 0:
                ws.cell(row=r, column=24 + m, value=vendas[m]).number_format = FMT_MONEY

    print(f"    -> Clients matched to rede: {matched_clients}/{len(clients)}")
    print(f"    -> Redes with REALIZADO: {redes_updated}/{len(REDES_META)}")
    print(f"    -> Redes data: {', '.join(f'{k}=R${sum(v):,.0f}' for k,v in vendas_por_rede.items())}")


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    start = time.time()
    print("=" * 60)
    print("  CRM VITAO360 V3 — MERGE FINAL v2 (Audited & Fixed)")
    print("=" * 60)
    print("  Fixes: CNPJ format, PROJECAO REALIZADO, motor edge cases")

    # Step 0: Build V3 structure
    print("\n[0/9] Building V3 structure...")
    from v3_regras import build_regras
    from v3_draft1 import build_draft1
    from v3_projecao import build_projecao
    from v3_carteira import build_carteira
    from v3_draft2 import build_draft2
    from v3_agenda import build_agenda
    from v3_dash import build_dash
    from v3_log import build_log, build_claude_log

    wb = openpyxl.Workbook()
    build_regras(wb)
    build_draft1(wb)
    build_projecao(wb)
    build_carteira(wb)
    build_draft2(wb)
    build_agenda(wb)
    build_dash(wb)
    build_log(wb)
    build_claude_log(wb)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    print(f"    -> Structure: {wb.sheetnames}")

    # Step 1: Read POPULADO clients (DRAFT 1)
    print("\n[1/9] Reading POPULADO clients...")
    clients = read_populado_clients()

    # Step 2: Read OULAR SAP data
    print("\n[2/9] Reading OULAR SAP data...")
    sap_data = read_oular_sap()

    # Step 3: Read POPULADO DRAFT 2 (17,805 records!)
    print("\n[3/9] Reading POPULADO DRAFT 2...")
    records = read_populado_draft2()

    # Step 4: Read CARTEIRA ACOMPANHAMENTO from POPULADO
    print("\n[4/9] Reading CARTEIRA ACOMPANHAMENTO...")
    acomp_data = read_populado_carteira_acomp()

    # Merge SAP
    merged_sap = 0
    for c in clients:
        sap = sap_data.get(c['cnpj'])
        if sap:
            c['sap_data'] = sap
            merged_sap += 1
    print(f"    -> SAP merged: {merged_sap}/{len(clients)} clients")

    # Step 5: Populate DRAFT 1
    print("\n[5/9] Populating DRAFT 1 (48 cols V3)...")
    populate_draft1(wb, clients)

    # Step 6: Populate DRAFT 2 + motor de regras
    print("\n[6/9] Populating DRAFT 2 + motor de regras...")
    populate_draft2(wb, records)

    # Step 7: Populate LOG
    print("\n[7/9] Populating LOG...")
    populate_log(wb, records)

    # Step 8: Populate CARTEIRA (with ACOMPANHAMENTO!)
    print("\n[8/9] Populating CARTEIRA (Mercos+Funil+SAP+ACOMP)...")
    populate_carteira(wb, clients, sap_data, acomp_data)

    # Step 9: Populate PROJECAO REALIZADO
    print("\n[9/9] Populating PROJECAO REALIZADO...")
    populate_projecao(wb, clients)

    # Set DRAFT 1 as active
    for i, name in enumerate(wb.sheetnames):
        if name == "DRAFT 1":
            wb.active = i
            break

    # Save
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    elapsed = time.time() - start

    # Stats
    print("\n" + "=" * 60)
    print(f"  SAVED: {OUTPUT_FILE}")
    print(f"  Tabs: {wb.sheetnames}")
    print(f"  DRAFT 1: {len(clients)} clients (48 cols, 12 months sales)")
    print(f"  DRAFT 2: {len(records)} records (CNPJ normalized + motor)")
    print(f"  LOG: {len(records)} records (archive)")
    print(f"  CARTEIRA: {len(clients)} clients (257 cols)")
    print(f"  SAP merged: {merged_sap} clients")
    print(f"  PROJECAO: REALIZADO populated from vendas")
    print(f"  DASH: formula-based (auto-calculates from DRAFT 2)")
    print(f"  Time: {elapsed:.1f}s")
    print("=" * 60)
    print("\n  CRM VITAO360 V3 MERGED (AUDITED) pronto!")
    print("   CNPJ format corrigido — XLOOKUPs vao funcionar no Excel.")


if __name__ == "__main__":
    main()

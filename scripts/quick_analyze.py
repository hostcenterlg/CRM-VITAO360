"""Quick analysis of all 3 Excel files for comparison."""
import openpyxl
import os

files = [
    "output/JARVIS_CRM_CENTRAL.xlsx",
    "output/CRM INTELIGENTE - VITAO360 V2.xlsx",
    "output/CRM_INTELIGENTE_VITAO360_POPULADO (2).xlsx",
]

for fpath in files:
    full = os.path.join(os.path.dirname(os.path.abspath(__file__)), fpath)
    print(f"\n{'='*80}")
    print(f"FILE: {fpath}")
    print(f"{'='*80}")

    try:
        wb = openpyxl.load_workbook(full, data_only=False)
    except Exception as e:
        print(f"  ERROR loading: {e}")
        continue

    print(f"Sheets: {wb.sheetnames}")
    print(f"Named Ranges: {len(list(wb.defined_names.values()))}")
    for dn in wb.defined_names.values():
        print(f"  - {dn.name} = {dn.attr_text}")

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        print(f"\n--- Sheet: {ws_name} ---")
        print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
        print(f"  Freeze: {ws.freeze_panes}")
        print(f"  Tab color: {ws.sheet_properties.tabColor}")

        # Headers
        headers = []
        header_row = 1
        # Check if first row has content, if not check row 6 (AGENDA style)
        first_val = ws.cell(row=1, column=1).value
        if first_val and ("AGENDA" in str(first_val) or "📋" in str(first_val)):
            header_row = 6

        for c in range(1, min(ws.max_column or 1, 100) + 1):
            v = ws.cell(row=header_row, column=c).value
            if v:
                headers.append(f"{openpyxl.utils.get_column_letter(c)}:{v}")
        print(f"  Headers (row {header_row}): {len(headers)} cols")
        for h in headers:
            print(f"    {h}")

        # Data validations
        dvs = list(ws.data_validations.dataValidation) if ws.data_validations else []
        print(f"  Data Validations: {len(dvs)}")
        for dv in dvs:
            print(f"    {dv.sqref} -> type={dv.type}, formula1={dv.formula1}")

        # Conditional formatting
        cfs = list(ws.conditional_formatting)
        print(f"  Conditional Formatting: {len(cfs)}")
        for cf in cfs:
            for rule in cf.rules:
                print(f"    range={cf.sqref} type={rule.type} operator={rule.operator} formula={rule.formula}")

        # Merged cells
        merges = list(ws.merged_cells.ranges)
        if merges:
            print(f"  Merged cells: {len(merges)}")
            for m in merges[:10]:
                print(f"    {m}")

        # Column groups (outline)
        groups = []
        for col_letter, dim in ws.column_dimensions.items():
            if hasattr(dim, 'outlineLevel') and dim.outlineLevel and dim.outlineLevel > 0:
                groups.append(f"{col_letter}(level={dim.outlineLevel}, hidden={dim.hidden})")
        if groups:
            print(f"  Column groups: {groups}")

        # Check formulas in first 5 data rows
        formulas = []
        start = header_row + 1
        for r in range(start, min(start + 5, (ws.max_row or start) + 1)):
            for c in range(1, min(ws.max_column or 1, 100) + 1):
                v = ws.cell(row=r, column=c).value
                if v and isinstance(v, str) and v.startswith('='):
                    formulas.append(f"  {openpyxl.utils.get_column_letter(c)}{r}: {v[:80]}")
        if formulas:
            print(f"  Formulas (sample):")
            for f in formulas[:20]:
                print(f"    {f}")

        # Sample data row
        print(f"  Sample data (row {start}):")
        for c in range(1, min(ws.max_column or 1, 20) + 1):
            v = ws.cell(row=start, column=c).value
            if v:
                print(f"    {openpyxl.utils.get_column_letter(c)}{start}: {v}")

    wb.close()

print("\n\nDONE.")

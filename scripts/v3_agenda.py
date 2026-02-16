"""
V3 — Tab AGENDA: Agenda Diária do Consultor (~25 cols + SCORE)
"""
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from v3_styles import *


AGENDA_COLS = [
    # CONTEXTO DO CLIENTE (automático)
    (1,  "NOME FANTASIA",     25, None),
    (2,  "CNPJ",              18, FMT_TEXT),
    (3,  "UF",                 5, None),
    (4,  "REDE / REGIONAL",   20, None),
    (5,  "SITUAÇÃO",          14, None),
    (6,  "DIAS SEM COMPRA",   14, FMT_NUMBER),
    (7,  "ESTÁGIO FUNIL",     18, None),
    (8,  "TIPO CLIENTE",      16, None),
    (9,  "FASE",              14, None),
    (10, "🔥 TEMPERATURA",    14, None),
    (11, "🚦 SINALEIRO",      12, None),
    (12, "PRÓX. AÇÃO",        14, None),
    (13, "TENTATIVA",         12, None),
    (14, "SCORE",             10, FMT_NUMBER),
    # RESULTADO DO ATENDIMENTO (preencher)
    (15, "WHATSAPP",          10, None),
    (16, "LIGAÇÃO",           10, None),
    (17, "LIGAÇÃO ATENDIDA",  14, None),
    (18, "TIPO DO CONTATO",   22, None),
    (19, "RESULTADO",         22, None),
    (20, "MOTIVO",            26, None),
    (21, "FOLLOW-UP",         14, FMT_DATE),
    (22, "AÇÃO FUTURA",       16, None),
    (23, "AÇÃO DETALHADA",    28, None),
    (24, "MERCOS ATUALIZADO", 16, None),
    (25, "NOTA DO DIA",       30, None),
]


def build_agenda(wb):
    """Build AGENDA tab (~25 cols + SCORE)."""
    ws = wb.create_sheet("AGENDA")
    ws.sheet_properties.tabColor = TAB_AGENDA

    # ── Row 1: Title ──
    ws.merge_cells('A1:Y1')
    title = ws.cell(row=1, column=1, value="📋 AGENDA DIÁRIA — [CONSULTOR]")
    style_cell(title, font=FONT_TITLE_W, fill=FILL_FUNIL, align=ALIGN_LEFT)
    for c in range(1, 26):
        ws.cell(row=1, column=c).fill = FILL_FUNIL
        ws.cell(row=1, column=c).border = THIN_BORDER

    # ── Row 2: Summary stats ──
    ws.merge_cells('A2:Y2')
    stats = ws.cell(row=2, column=1,
                    value="Cart:__ | Prosp:__ | Follow-ups:__ | Novos:__ | Data: ___/___/2026")
    style_cell(stats, font=FONT_DATA, fill=FILL_LIGHT_GRAY, align=ALIGN_LEFT)
    for c in range(1, 26):
        ws.cell(row=2, column=c).fill = FILL_LIGHT_GRAY
        ws.cell(row=2, column=c).border = THIN_BORDER

    # ── Row 3: Section headers ──
    ws.merge_cells('A3:N3')
    ctx = ws.cell(row=3, column=1, value="🔵 CONTEXTO DO CLIENTE (automático — não editar)")
    style_cell(ctx, font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    ws.merge_cells('O3:Y3')
    res = ws.cell(row=3, column=15, value="✅ RESULTADO DO ATENDIMENTO (preencher)")
    style_cell(res, font=FONT_TITLE, fill=PatternFill('solid', fgColor='C8E6C9'), align=ALIGN_LEFT)
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = FILL_GRAY_D9
        ws.cell(row=3, column=c).border = THIN_BORDER
    for c in range(15, 26):
        ws.cell(row=3, column=c).fill = PatternFill('solid', fgColor='C8E6C9')
        ws.cell(row=3, column=c).border = THIN_BORDER

    # ── Row 4: Column headers ──
    for col, header, width, fmt in AGENDA_COLS:
        fill = FILL_HEADER if col <= 14 else PatternFill('solid', fgColor='2E7D32')
        write_header(ws, 4, col, header, fill=fill)
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Row 5: placeholder ──
    ws.cell(row=5, column=1).value = "[cliente carregado da CARTEIRA]"
    ws.cell(row=5, column=1).font = Font(name='Calibri', size=10, italic=True, color='888888')

    # ── Data Validations ──
    dv_res = DataValidation(type="list", formula1="LISTA_RESULTADO", allow_blank=True)
    ws.add_data_validation(dv_res)
    dv_res.add("S5:S5000")

    dv_mot = DataValidation(type="list", formula1="LISTA_MOTIVO", allow_blank=True)
    ws.add_data_validation(dv_mot)
    dv_mot.add("T5:T5000")

    dv_tipo = DataValidation(type="list", formula1="LISTA_TIPO_CONTATO", allow_blank=True)
    ws.add_data_validation(dv_tipo)
    dv_tipo.add("R5:R5000")

    dv_sn = DataValidation(type="list", formula1="LISTA_SIM_NAO", allow_blank=True)
    ws.add_data_validation(dv_sn)
    for col_letter in ['O', 'P', 'Q', 'X']:
        dv_sn.add(f"{col_letter}5:{col_letter}5000")

    dv_acao = DataValidation(type="list", formula1="LISTA_ACAO_FUTURA", allow_blank=True)
    ws.add_data_validation(dv_acao)
    dv_acao.add("V5:V5000")

    # ── Conditional formatting for SITUAÇÃO ──
    from openpyxl.formatting.rule import CellIsRule
    for sit, fill in SITUACAO_FILLS.items():
        ws.conditional_formatting.add(
            'E5:E5000',
            CellIsRule(operator='equal', formula=[f'"{sit}"'], fill=fill)
        )

    # ── Freeze ──
    ws.freeze_panes = "A5"

    # Row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20

    print(f"  AGENDA: {len(AGENDA_COLS)} cols, freeze A5")
    return ws

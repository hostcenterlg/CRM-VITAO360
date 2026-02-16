"""Inspect downloaded V3 file."""
import openpyxl

FILE = r"C:\Users\User\Downloads\CRM_INTELIGENTE_VITAO360_V3 (1).xlsx"
wb = openpyxl.load_workbook(FILE, data_only=True)

print("=" * 60)
print("  TABS")
print("=" * 60)
for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    color = ws.sheet_properties.tabColor
    print(f"  {i+1}. {name} (rows={ws.max_row}, cols={ws.max_column})")

print()
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n{'='*60}")
    print(f"  {name}")
    print(f"{'='*60}")
    for r in range(1, min(6, ws.max_row + 1)):
        vals = []
        for c in range(1, min(30, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                vals.append(f"C{c}:{repr(v)[:50]}")
        if vals:
            print(f"  R{r}: {'  |  '.join(vals[:8])}")

# Named ranges
print(f"\n{'='*60}")
print(f"  NAMED RANGES")
print(f"{'='*60}")
for dn in wb.defined_names.values():
    print(f"  {dn.name} = {dn.attr_text}")

# Data validations per tab
print(f"\n{'='*60}")
print(f"  DATA VALIDATIONS")
print(f"{'='*60}")
for name in wb.sheetnames:
    ws = wb[name]
    dvs = list(ws.data_validations.dataValidation) if ws.data_validations else []
    if dvs:
        print(f"  {name}: {len(dvs)} validations")
        for dv in dvs[:5]:
            print(f"    {dv.sqref} = {dv.formula1}")

wb.close()

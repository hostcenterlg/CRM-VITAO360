"""
V3 — Tab DRAFT 1: Base Mestre do Cliente (48 colunas, 7 blocos)
"""
from openpyxl.utils import get_column_letter
from v3_styles import *


# 48 column definitions: (col_num, header, block_name, width, is_formula, fmt)
DRAFT1_COLS = [
    # BLOCO 1: IDENTIDADE (1-11)
    (1,  "NOME FANTASIA",           "IDENTIDADE", 25, False, None),
    (2,  "CNPJ",                    "IDENTIDADE", 18, False, FMT_TEXT),
    (3,  "RAZÃO SOCIAL",            "IDENTIDADE", 30, False, None),
    (4,  "UF",                      "IDENTIDADE", 5,  False, None),
    (5,  "CIDADE",                  "IDENTIDADE", 18, False, None),
    (6,  "EMAIL",                   "IDENTIDADE", 28, False, None),
    (7,  "TELEFONE",                "IDENTIDADE", 16, False, FMT_TEXT),
    (8,  "DATA CADASTRO",           "IDENTIDADE", 14, False, FMT_DATE),
    (9,  "REDE / REGIONAL",         "IDENTIDADE", 22, False, None),
    (10, "CONSULTOR",               "IDENTIDADE", 20, False, None),
    (11, "VENDEDOR ÚLTIMO PEDIDO",  "IDENTIDADE", 24, False, None),
    # BLOCO 2: STATUS (12-15)
    (12, "SITUAÇÃO",                "STATUS", 14, True, None),
    (13, "PRIORIDADE",              "STATUS", 12, True, FMT_NUMBER),
    (14, "DIAS SEM COMPRA",         "STATUS", 16, True, FMT_NUMBER),
    (15, "CICLO MÉDIO",             "STATUS", 14, False, FMT_NUMBER),
    # BLOCO 3: COMPRAS (16-18)
    (16, "DATA ÚLTIMO PEDIDO",      "COMPRAS", 18, False, FMT_DATE),
    (17, "VALOR ÚLTIMO PEDIDO",     "COMPRAS", 18, False, FMT_MONEY),
    (18, "TOTAL PERÍODO",           "COMPRAS", 16, True, FMT_MONEY),
    # BLOCO 4: ECOMMERCE (19-24)
    (19, "ACESSOS SEMANA",          "ECOMMERCE", 16, False, FMT_NUMBER),
    (20, "ACESSO B2B",              "ECOMMERCE", 12, False, FMT_NUMBER),
    (21, "ACESSOS PORTAL",          "ECOMMERCE", 16, False, FMT_NUMBER),
    (22, "ITENS CARRINHO",          "ECOMMERCE", 14, False, FMT_NUMBER),
    (23, "VALOR B2B",               "ECOMMERCE", 14, False, FMT_MONEY),
    (24, "OPORTUNIDADE",            "ECOMMERCE", 16, True, None),
    # BLOCO 5: VENDAS MÊS A MÊS (25-37) — rolling 12 months + total
    (25, "MAR/25",                  "VENDAS", 12, False, FMT_MONEY),
    (26, "ABR/25",                  "VENDAS", 12, False, FMT_MONEY),
    (27, "MAI/25",                  "VENDAS", 12, False, FMT_MONEY),
    (28, "JUN/25",                  "VENDAS", 12, False, FMT_MONEY),
    (29, "JUL/25",                  "VENDAS", 12, False, FMT_MONEY),
    (30, "AGO/25",                  "VENDAS", 12, False, FMT_MONEY),
    (31, "SET/25",                  "VENDAS", 12, False, FMT_MONEY),
    (32, "OUT/25",                  "VENDAS", 12, False, FMT_MONEY),
    (33, "NOV/25",                  "VENDAS", 12, False, FMT_MONEY),
    (34, "DEZ/25",                  "VENDAS", 12, False, FMT_MONEY),
    (35, "JAN/26",                  "VENDAS", 12, False, FMT_MONEY),
    (36, "FEV/26",                  "VENDAS", 12, False, FMT_MONEY),
    (37, "TOTAL VENDAS PERÍODO",    "VENDAS", 20, True, FMT_MONEY),
    # BLOCO 6: RECORRÊNCIA (38-44)
    (38, "Nº COMPRAS",              "RECORRÊNCIA", 12, False, FMT_NUMBER),
    (39, "CURVA ABC",               "RECORRÊNCIA", 10, True, None),
    (40, "MESES POSITIVADO",        "RECORRÊNCIA", 16, False, FMT_NUMBER),
    (41, "MÉDIA MENSAL",            "RECORRÊNCIA", 14, True, FMT_MONEY),
    (42, "TICKET MÉDIO",            "RECORRÊNCIA", 14, True, FMT_MONEY),
    (43, "MESES LISTA",             "RECORRÊNCIA", 12, False, FMT_NUMBER),
    (44, "TIPO CLIENTE",            "RECORRÊNCIA", 18, True, None),
    # BLOCO 7: ATENDIMENTO MERCOS (45-48)
    (45, "ÚLT. REGISTRO MERCOS",   "ATENDIMENTO", 20, False, FMT_DATE),
    (46, "DATA ÚLT. ATEND. MERCOS","ATENDIMENTO", 22, False, FMT_DATE),
    (47, "TIPO ATENDIMENTO MERCOS", "ATENDIMENTO", 24, False, None),
    (48, "OBS ATENDIMENTO MERCOS",  "ATENDIMENTO", 35, False, None),
]

BLOCK_FILLS = {
    "IDENTIDADE": FILL_BLOCK_IDENTIDADE,
    "STATUS": FILL_BLOCK_STATUS,
    "COMPRAS": FILL_BLOCK_COMPRA,
    "ECOMMERCE": FILL_BLOCK_ECOMMERCE,
    "VENDAS": FILL_BLOCK_VENDAS,
    "RECORRÊNCIA": FILL_BLOCK_RECORRENCIA,
    "ATENDIMENTO": FILL_BLOCK_ATENDIMENTO,
}

BLOCK_EMOJI = {
    "IDENTIDADE": "🟣",
    "STATUS": "🟣",
    "COMPRAS": "🟣",
    "ECOMMERCE": "🟣",
    "VENDAS": "🟣",
    "RECORRÊNCIA": "🟣",
    "ATENDIMENTO": "🟣",
}


def build_draft1(wb):
    """Build DRAFT 1 tab (48 cols, 7 blocks, empty template)."""
    ws = wb.create_sheet("DRAFT 1")
    ws.sheet_properties.tabColor = TAB_DRAFT1

    # ── Row 1: Title ──
    ws.merge_cells('A1:AT1')
    title_cell = ws.cell(row=1, column=1,
                         value="📦 DRAFT 1 — BASE MESTRE DO CLIENTE (atualizar diariamente via extração Mercos)")
    style_cell(title_cell, font=FONT_TITLE_W, fill=FILL_MERCOS, align=ALIGN_LEFT)
    for c in range(1, 49):
        ws.cell(row=1, column=c).fill = FILL_MERCOS
        ws.cell(row=1, column=c).border = THIN_BORDER

    # ── Row 2: Block headers ──
    current_block = None
    block_start = None
    block_ranges = []

    for col_num, header, block, width, is_formula, fmt in DRAFT1_COLS:
        if block != current_block:
            if current_block is not None:
                block_ranges.append((current_block, block_start, col_num - 1))
            current_block = block
            block_start = col_num
    block_ranges.append((current_block, block_start, 48))

    for block_name, start_col, end_col in block_ranges:
        emoji = BLOCK_EMOJI.get(block_name, "")
        fill = BLOCK_FILLS.get(block_name, FILL_SUB_HEADER)
        if end_col > start_col:
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        cell = ws.cell(row=2, column=start_col, value=f"{emoji} {block_name}")
        style_cell(cell, font=FONT_BLOCK, fill=fill, align=ALIGN_CENTER)
        for c in range(start_col, end_col + 1):
            ws.cell(row=2, column=c).fill = fill
            ws.cell(row=2, column=c).border = THIN_BORDER

    # ── Row 3: Column headers ──
    for col_num, header, block, width, is_formula, fmt in DRAFT1_COLS:
        write_header(ws, 3, col_num, header, fill=FILL_HEADER)
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # ── Row 4+: Formulas (for first data row, then copy pattern) ──
    R = 4  # first data row
    # SITUAÇÃO (col 12): based on DIAS SEM COMPRA
    ws.cell(row=R, column=12).value = (
        '=IF(N4="","",IF(AND(P4="",Q4=""),"PROSPECT",'
        'IF(P4<=50,"ATIVO",IF(P4<=60,"EM RISCO",IF(P4<=90,"INAT.REC","INAT.ANT")))))'
    ).replace('N4', f'N{R}').replace('P4', f'P{R}').replace('Q4', f'Q{R}')
    # Simplified — just for row 4
    ws.cell(row=R, column=12).value = (
        '=IF(A4="","",IF(P4="","PROSPECT",'
        'IF(N4<=50,"ATIVO",IF(N4<=60,"EM RISCO",IF(N4<=90,"INAT.REC","INAT.ANT")))))'
    )

    # PRIORIDADE (col 13)
    ws.cell(row=R, column=13).value = (
        '=IF(L4="","",SWITCH(L4,"ATIVO",1,"EM RISCO",2,"INAT.REC",3,"INAT.ANT",4,"PROSPECT",5,"NOVO",6,99))'
    )

    # DIAS SEM COMPRA (col 14) = TODAY() - DATA ÚLTIMO PEDIDO
    ws.cell(row=R, column=14).value = '=IF(P4="","",INT(TODAY()-P4))'

    # TOTAL PERÍODO (col 18) = SUM vendas meses
    ws.cell(row=R, column=18).value = '=SUM(Y4:AJ4)'

    # TOTAL VENDAS PERÍODO (col 37) = SUM 12 months
    ws.cell(row=R, column=37).value = '=SUM(Y4:AJ4)'

    # OPORTUNIDADE (col 24)
    ws.cell(row=R, column=24).value = (
        '=IF(V4>0,"🔥 QUENTE",IF(T4>0,"🟡 MORNO","❄️ FRIO"))'
    )

    # CURVA ABC (col 39)
    ws.cell(row=R, column=39).value = '=IF(AK4>=2000,"A",IF(AK4>=500,"B","C"))'

    # MÉDIA MENSAL (col 41) = TOTAL / MESES POSITIVADO
    ws.cell(row=R, column=41).value = '=IFERROR(AK4/AN4,0)'

    # TICKET MÉDIO (col 42) = TOTAL / Nº COMPRAS
    ws.cell(row=R, column=42).value = '=IFERROR(AK4/AL4,0)'

    # TIPO CLIENTE (col 44)
    ws.cell(row=R, column=44).value = (
        '=IF(AN4=0,"PROSPECT",IF(AN4=1,"NOVO",'
        'IF(AN4<=3,"EM DESENVOLVIMENTO",IF(AN4<=6,"RECORRENTE","FIDELIZADO"))))'
    )

    # Apply number formats to formula cells
    for col_num, header, block, width, is_formula, fmt in DRAFT1_COLS:
        if fmt:
            ws.cell(row=R, column=col_num).number_format = fmt

    # ── Freeze at C4 (lock NOME + CNPJ visible) ──
    ws.freeze_panes = "C4"

    # ── Column grouping ──
    # ECOMMERCE group (19-24)
    group_columns(ws, 19, 24, outline_level=1, hidden=True)
    # VENDAS group (25-37)
    group_columns(ws, 25, 37, outline_level=1, hidden=True)
    # RECORRÊNCIA group (38-44)
    group_columns(ws, 38, 44, outline_level=1, hidden=True)
    # ATENDIMENTO group (45-48)
    group_columns(ws, 45, 48, outline_level=1, hidden=True)

    print(f"  DRAFT 1: 48 cols, 7 blocks, freeze C4")
    return ws

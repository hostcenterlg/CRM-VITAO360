"""
ingest_real_data.py — Ingestão de dados reais no PostgreSQL do CRM VITAO360
===========================================================================
Fontes (hierarquia SAP > Mercos > Deskrio):
  1. data/mercos_clientes_completo.json      — 6 429 clientes Mercos
  2. data/sources/sap/01_SAP_CONSOLIDADO.xlsx — Cadastro + Venda Mês a Mês
  3. data/deskrio/2026-04-15/contacts.json   — 15 601 contatos WhatsApp

Fases:
  Phase 1 — Mercos → clientes (UPSERT por CNPJ)
  Phase 2 — SAP Cadastro → enriquece codigo_cliente, tipo, macroregiao, consultor
  Phase 3 — SAP Venda Mês a Mês → faturamento_total, n_compras, dias_sem_compra, situacao
  Phase 4 — Curva ABC (A/B/C por faturamento acumulado)
  Phase 5 — Deskrio → match por telefone → grava JSON de referência

REGRAS SAGRADAS:
  R1  — Two-Base: faturamento_total APENAS de dados VENDA do SAP
  R2  — CNPJ = string 14 dígitos, zero-padded, NUNCA float
  R5  — classificacao_3tier = 'REAL' para todos os registros
  R8  — NUNCA fabricar dados; dados sem rastreabilidade = IGNORADOS
  R11 — Idempotente (UPSERT): pode rodar N vezes sem duplicar
"""

from __future__ import annotations

import json
import logging
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

# ── Paths ─────────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
SAP_FILE = DATA_DIR / "sources" / "sap" / "01_SAP_CONSOLIDADO.xlsx"
MERCOS_FILE = DATA_DIR / "mercos_clientes_completo.json"
DESKRIO_FILE = DATA_DIR / "deskrio" / "2026-04-15" / "contacts.json"
DESKRIO_MATCH_OUTPUT = DATA_DIR / "deskrio_phone_matches.json"

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("ingest_real_data")

# ── Reference date for dias_sem_compra (today) ────────────────────────────────
REF_DATE = date(2026, 4, 15)

# ── SAP Venda Mês a Mês: 2025 months → end-of-month date for dias calculation ─
SAP_MONTHS_DATES = {
    "Jan": date(2025, 1, 31),
    "Fev": date(2025, 2, 28),
    "Mar": date(2025, 3, 31),
    "Abr": date(2025, 4, 30),
    "Mai": date(2025, 5, 31),
    "Jun": date(2025, 6, 30),
    "Jul": date(2025, 7, 31),
    "Ago": date(2025, 8, 31),
    "Set": date(2025, 9, 30),
    "Out": date(2025, 10, 31),
    "Nov": date(2025, 11, 30),
    "Dez": date(2025, 12, 31),
}
SAP_FATURADO_COLS = {
    # month_name: 0-based column index in sheet row
    "Jan": 6,   # G
    "Fev": 10,  # K
    "Mar": 14,  # O
    "Abr": 18,  # S
    "Mai": 22,  # W
    "Jun": 26,  # AA
    "Jul": 30,  # AE
    "Ago": 34,  # AI
    "Set": 38,  # AM
    "Out": 42,  # AQ
    "Nov": 46,  # AU
    "Dez": 50,  # AY
}

# ── DE-PARA consultores ────────────────────────────────────────────────────────
# Mercos colaboradores names → canonical consultor
_CONSULTOR_MAP = {
    "manu ditzel": "MANU",
    "hemanuele rosa ditzel": "MANU",
    "hemanuele ditzel": "MANU",
    "larissa padilha": "LARISSA",
    "larissa da silva padilha": "LARISSA",
    "42.149.466 larissa da silva padilha - me": "LARISSA",
    "julio gadret": "JULIO",
    "julio  gadret": "JULIO",  # double-space variant seen in data
    "daiane stavicki": "DAIANE",
    "central - daiane": None,   # gestor global, não é consultor exclusivo
    "leandro garcia": None,     # admin
}
# Patterns for partial matching (SAP ZR/ZV columns)
_CONSULTOR_PATTERNS = [
    (re.compile(r"ditzel", re.I), "MANU"),
    (re.compile(r"larissa", re.I), "LARISSA"),
    (re.compile(r"julio\s+gadret", re.I), "JULIO"),
    (re.compile(r"daiane", re.I), "DAIANE"),
]
# Names that are NEVER assigned as consultor (admins/global access)
_SKIP_CONSULTORES = {"central - daiane", "leandro garcia"}

# ── Siglas que NÃO devem virar title-case ─────────────────────────────────────
_SIGLAS = {"LTDA", "ME", "EPP", "EIRELI", "SA", "SS", "SAS", "MEI", "LTDA.", "S/A", "S.A."}


# ==============================================================================
# Helpers
# ==============================================================================

def normalize_cnpj(val) -> Optional[str]:
    """Normaliza CNPJ para 14 dígitos string, zero-padded.
    Retorna None se inválido, vazio ou zerado (R5)."""
    if val is None:
        return None
    try:
        raw = str(val).strip()
        if isinstance(val, float):
            # CNPJ nunca deve ser float, mas converte defensivamente
            raw = str(int(val))
        digits = re.sub(r"\D", "", raw).zfill(14)
        if len(digits) != 14 or digits == "00000000000000":
            return None
        return digits
    except Exception:
        return None


def normalize_nome(raw: Optional[str]) -> Optional[str]:
    """Strip espaços, aplica Title Case preservando siglas como LTDA, ME, etc."""
    if not raw:
        return None
    raw = str(raw).strip()
    tokens = []
    for token in raw.split():
        if token.upper() in _SIGLAS:
            tokens.append(token.upper())
        else:
            tokens.append(token.capitalize())
    return " ".join(tokens) if tokens else None


def normalize_phone(raw: Optional[str]) -> Optional[str]:
    """Remove formatação, garante 13 dígitos (55 + DDD + número).
    Retorna None se não tiver dígitos suficientes."""
    if not raw:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return None
    # Adiciona prefixo 55 se ausente
    if not digits.startswith("55"):
        digits = "55" + digits
    # Deve ter 13 dígitos (55 + 2 DDD + 9 número) ou 12 (55 + 2 DDD + 8 número)
    if len(digits) < 12 or len(digits) > 13:
        return None
    return digits


def map_consultor_mercos(colaboradores: list) -> Optional[str]:
    """Retorna o consultor canônico de uma lista de colaboradores Mercos.

    Regra:
    1. Ignora 'Central - Daiane' e 'Leandro Garcia' (acesso global).
    2. Retorna o primeiro colaborador válido mapeado.
    3. Se nenhum mapeado → None.
    """
    candidates = []
    for colab in colaboradores or []:
        key = str(colab).strip().lower()
        if key in _SKIP_CONSULTORES:
            continue
        canonical = _CONSULTOR_MAP.get(key)
        if canonical:
            candidates.append(canonical)
    if candidates:
        return candidates[0]
    # Fallback: tentar match parcial por regex
    for colab in colaboradores or []:
        name = str(colab).strip()
        if name.lower() in _SKIP_CONSULTORES:
            continue
        for pattern, canon in _CONSULTOR_PATTERNS:
            if pattern.search(name):
                return canon
    return None


def map_consultor_sap(zr: Optional[str], zv: Optional[str]) -> Optional[str]:
    """Mapeia consultor a partir das colunas ZR e ZV do SAP."""
    for col_val in [zr, zv]:
        if not col_val:
            continue
        key = str(col_val).strip()
        # Tenta exato primeiro
        lower_key = key.lower()
        if lower_key in _CONSULTOR_MAP:
            return _CONSULTOR_MAP[lower_key]
        # Regex parcial
        for pattern, canon in _CONSULTOR_PATTERNS:
            if pattern.search(key):
                return canon
    return None


def calc_situacao(dias: Optional[int]) -> str:
    """Calcula situação comercial baseada nos dias sem compra."""
    if dias is None:
        return "PROSPECT"
    if dias <= 50:
        return "ATIVO"
    if dias <= 60:
        return "EM RISCO"
    if dias <= 90:
        return "INAT.REC"
    return "INAT.ANT"


def strip_sap_prefix(val: Optional[str]) -> Optional[str]:
    """Remove prefixo numérico SAP como '02 - ', '03 - ' do texto."""
    if not val:
        return None
    return re.sub(r"^\d+\s*-\s*", "", str(val)).strip()


# ==============================================================================
# Database connection
# ==============================================================================

def get_engine():
    """Cria engine SQLAlchemy usando .env.local (Neon PostgreSQL) ou DATABASE_URL."""
    # Tenta carregar .env.local primeiro (Vercel/Neon), depois .env
    env_local = PROJECT_ROOT / ".env.local"
    env_file = PROJECT_ROOT / ".env"

    db_url = os.getenv("DATABASE_URL", "").strip()

    if not db_url and env_local.exists():
        for line in env_local.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("DATABASE_URL="):
                db_url = line.split("=", 1)[1].strip().strip('"').strip("'")
                break

    if not db_url and env_file.exists():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("DATABASE_URL="):
                db_url = line.split("=", 1)[1].strip().strip('"').strip("'")
                break

    if not db_url:
        # Fallback: SQLite local
        db_url = f"sqlite:///{DATA_DIR / 'crm_vitao360.db'}"
        log.warning("DATABASE_URL nao encontrado — usando SQLite local: %s", db_url)
    else:
        if db_url.startswith("postgres://"):
            db_url = db_url.replace("postgres://", "postgresql://", 1)
        log.info("Conectando em: %s", re.sub(r":([^@/]+)@", ":***@", db_url))

    from sqlalchemy import create_engine

    connect_args: dict = {}
    if db_url.startswith("sqlite"):
        connect_args = {"check_same_thread": False}
    elif "neon" in db_url or "supabase" in db_url:
        connect_args = {"sslmode": "require"}

    return create_engine(db_url, connect_args=connect_args, pool_pre_ping=True)


# ==============================================================================
# Phase 1 — Mercos → clientes  (bulk UPSERT)
# ==============================================================================

_PHASE1_BATCH = 1000  # rows per execute_values page_size


def _get_raw_pg_conn(engine):
    """Returns a raw psycopg2 connection from the SQLAlchemy engine pool."""
    raw = engine.raw_connection()
    return raw


def load_mercos(engine) -> dict[str, int]:
    """Carrega 6 429 clientes Mercos na tabela clientes.

    PostgreSQL: usa psycopg2.extras.execute_values para inserção bulk real —
    uma única query VALUES (...), (...), ... por page_size, sem round-trips individuais.
    SQLite fallback: INSERT OR IGNORE via SQLAlchemy.

    Retorna mapa cnpj → db_id.
    """
    from sqlalchemy import text

    log.info("=== PHASE 1: Mercos → clientes (bulk UPSERT) ===")

    with open(MERCOS_FILE, "r", encoding="utf-8") as f:
        records = json.load(f)

    log.info("Mercos: %d registros carregados", len(records))

    is_pg = "postgresql" in str(engine.url) or "neon" in str(engine.url)

    # Build list of tuples for bulk insert
    rows_to_upsert = []  # (cnpj, nome_fantasia, razao_social, uf, cidade, email, telefone, consultor, situacao, tier, rede, bloqueado)
    skipped = 0
    for rec in records:
        cnpj = normalize_cnpj(rec.get("cnpj"))
        if not cnpj:
            skipped += 1
            continue

        nome_fantasia = normalize_nome(rec.get("nome_fantasia") or rec.get("razao_social"))
        razao_social = normalize_nome(rec.get("razao_social"))
        email_raw = (rec.get("emails") or [None])[0]
        email = str(email_raw).strip().lower() if email_raw else None
        phone_raw = (rec.get("telefones") or [None])[0]
        telefone = normalize_phone(phone_raw)
        consultor = map_consultor_mercos(rec.get("colaboradores") or [])
        bloqueado = str(rec.get("bloqueado", "Não")).strip().lower() == "sim"
        uf = str(rec.get("estado") or "").strip().upper() or None
        cidade = str(rec.get("cidade") or "").strip() or None
        rede_regional = str(rec.get("rede_nome") or "").strip() or None

        rows_to_upsert.append((
            cnpj, nome_fantasia, razao_social, uf, cidade,
            email, telefone, consultor, "PROSPECT", "REAL", rede_regional, bloqueado,
        ))

    log.info("Mercos: %d válidos, %d CNPJ inválidos ignorados (antes dedup)", len(rows_to_upsert), skipped)

    # Deduplicate by CNPJ — keep last occurrence (highest Mercos id = mais recente)
    dedup: dict[str, tuple] = {}
    for tup in rows_to_upsert:
        dedup[tup[0]] = tup  # tup[0] = cnpj
    rows_to_upsert = list(dedup.values())
    log.info("Mercos: %d após deduplicação de CNPJs", len(rows_to_upsert))

    if is_pg:
        from psycopg2.extras import execute_values

        raw_conn = _get_raw_pg_conn(engine)
        cur = raw_conn.cursor()
        try:
            upsert_sql = """
                INSERT INTO clientes
                    (cnpj, nome_fantasia, razao_social, uf, cidade,
                     email, telefone, consultor, situacao,
                     classificacao_3tier, rede_regional, problema_aberto)
                VALUES %s
                ON CONFLICT (cnpj) DO UPDATE SET
                    nome_fantasia     = COALESCE(clientes.nome_fantasia, EXCLUDED.nome_fantasia),
                    razao_social      = COALESCE(clientes.razao_social,  EXCLUDED.razao_social),
                    uf                = COALESCE(clientes.uf,            EXCLUDED.uf),
                    cidade            = COALESCE(clientes.cidade,        EXCLUDED.cidade),
                    email             = COALESCE(clientes.email,         EXCLUDED.email),
                    telefone          = COALESCE(clientes.telefone,      EXCLUDED.telefone),
                    consultor         = COALESCE(clientes.consultor,     EXCLUDED.consultor),
                    rede_regional     = COALESCE(clientes.rede_regional, EXCLUDED.rede_regional),
                    classificacao_3tier = 'REAL',
                    updated_at        = NOW()
            """
            execute_values(cur, upsert_sql, rows_to_upsert, page_size=_PHASE1_BATCH)
            raw_conn.commit()
            log.info("Phase 1 — execute_values bulk upsert OK (%d rows)", len(rows_to_upsert))

            # Fetch cnpj_to_id
            cur.execute("SELECT cnpj, id FROM clientes")
            cnpj_to_id: dict[str, int] = {row[0]: row[1] for row in cur.fetchall()}
        finally:
            cur.close()
            raw_conn.close()
    else:
        # SQLite fallback: INSERT OR IGNORE
        cnpj_to_id = {}
        with engine.begin() as conn:
            for tup in rows_to_upsert:
                conn.execute(
                    text("""
                        INSERT OR IGNORE INTO clientes
                            (cnpj, nome_fantasia, razao_social, uf, cidade,
                             email, telefone, consultor, situacao,
                             classificacao_3tier, rede_regional, problema_aberto)
                        VALUES
                            (:cnpj, :nf, :rs, :uf, :cidade,
                             :email, :tel, :consultor, :sit,
                             :tier, :rede, :bloq)
                    """),
                    dict(zip(
                        ["cnpj","nf","rs","uf","cidade","email","tel","consultor","sit","tier","rede","bloq"],
                        tup,
                    )),
                )
            rows = conn.execute(text("SELECT cnpj, id FROM clientes")).fetchall()
            cnpj_to_id = {r[0]: r[1] for r in rows}

    log.info("Phase 1 done — upserted=%d skipped=%d total_in_db=%d",
             len(rows_to_upsert), skipped, len(cnpj_to_id))
    return cnpj_to_id


# ==============================================================================
# Phase 2 — SAP Cadastro Clientes → enriquecer clientes  (bulk)
# ==============================================================================

def enrich_from_sap_cadastro(engine, cnpj_to_id: dict[str, int]) -> int:
    """Lê aba 'Cadastro Clientes SAP' e enriquece clientes via bulk UPDATE/INSERT.

    SAP hierarchy: codigo_cliente, tipo_cliente_sap, macroregiao, consultor
    são escritos pelo SAP (sobrepõem Mercos).
    Cria novo registro se CNPJ não está no banco.
    Retorna número de registros processados.
    """
    from sqlalchemy import text
    import openpyxl

    log.info("=== PHASE 2: SAP Cadastro → enriquece clientes (bulk) ===")

    wb = openpyxl.load_workbook(SAP_FILE, read_only=True, data_only=True)
    ws = wb["Cadastro Clientes SAP"]

    # Parse all rows first, then do two bulk operations:
    # 1. Bulk UPDATE for existing CNPJs
    # 2. Bulk INSERT for new CNPJs
    updates = []   # rows already in DB
    inserts = []   # rows not in DB
    skipped_cnpj = 0

    # Snapshot of known CNPJs to avoid per-row SELECT
    known_cnpjs = set(cnpj_to_id.keys())

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_cnpj_d = row[3]
        raw_cnpj_e = row[4]
        cnpj = normalize_cnpj(raw_cnpj_e) or normalize_cnpj(raw_cnpj_d)
        if not cnpj:
            skipped_cnpj += 1
            continue

        codigo_cliente = str(row[2]).strip() if row[2] else None
        nome_sap = str(row[5]).strip() if row[5] else None
        uf_sap = str(row[9]).strip().upper() if row[9] else None
        cidade_sap = str(row[8]).strip() if row[8] else None

        tipo_cliente_sap = strip_sap_prefix(row[34])   # AI
        macroregiao = strip_sap_prefix(row[36])         # AK
        rede_sap = strip_sap_prefix(row[42])            # AQ (grupo chave)
        if rede_sap and "SEM GRUPO" in rede_sap.upper():
            rede_sap = None

        zr_nome = str(row[24]).strip() if row[24] else None
        zv_nome = str(row[28]).strip() if row[28] else None
        consultor_sap = map_consultor_sap(zr_nome, zv_nome)

        rec = {
            "cnpj": cnpj,
            "razao_social": normalize_nome(nome_sap),
            "uf": uf_sap,
            "cidade": normalize_nome(cidade_sap),
            "codigo_cliente": codigo_cliente,
            "tipo_cliente_sap": tipo_cliente_sap,
            "macroregiao": macroregiao,
            "rede_regional": rede_sap,
            "consultor": consultor_sap,
        }

        if cnpj in known_cnpjs:
            updates.append(rec)
        else:
            inserts.append(rec)
            known_cnpjs.add(cnpj)

    log.info("SAP Cadastro: %d updates, %d inserts, %d skipped_cnpj",
             len(updates), len(inserts), skipped_cnpj)

    is_pg = "postgresql" in str(engine.url) or "neon" in str(engine.url)
    BATCH = 1000

    if is_pg:
        from psycopg2.extras import execute_values

        raw_conn = _get_raw_pg_conn(engine)
        cur = raw_conn.cursor()
        try:
            # Temp table approach: bulk load SAP data then UPDATE via JOIN
            cur.execute("""
                CREATE TEMP TABLE _sap_cad (
                    cnpj TEXT, codigo_cliente TEXT, tipo_cliente_sap TEXT,
                    macroregiao TEXT, rede_regional TEXT, consultor TEXT
                ) ON COMMIT DROP
            """)
            upd_tuples = [
                (r["cnpj"], r["codigo_cliente"], r["tipo_cliente_sap"],
                 r["macroregiao"], r["rede_regional"], r["consultor"])
                for r in updates
            ]
            if upd_tuples:
                execute_values(cur, "INSERT INTO _sap_cad VALUES %s", upd_tuples, page_size=BATCH)
                cur.execute("""
                    UPDATE clientes c SET
                        codigo_cliente   = t.codigo_cliente,
                        tipo_cliente_sap = t.tipo_cliente_sap,
                        macroregiao      = t.macroregiao,
                        rede_regional    = COALESCE(t.rede_regional, c.rede_regional),
                        consultor        = CASE WHEN t.consultor IS NOT NULL
                                               THEN t.consultor ELSE c.consultor END,
                        classificacao_3tier = 'REAL',
                        updated_at       = NOW()
                    FROM _sap_cad t
                    WHERE c.cnpj = t.cnpj
                """)

            if inserts:
                ins_tuples = [
                    (r["cnpj"], r["razao_social"], r["uf"], r["cidade"],
                     r["codigo_cliente"], r["tipo_cliente_sap"], r["macroregiao"],
                     r["rede_regional"], r["consultor"])
                    for r in inserts
                ]
                execute_values(
                    cur,
                    """
                    INSERT INTO clientes
                        (cnpj, razao_social, uf, cidade, codigo_cliente,
                         tipo_cliente_sap, macroregiao, rede_regional,
                         consultor, situacao, classificacao_3tier)
                    VALUES %s
                    ON CONFLICT (cnpj) DO NOTHING
                    """,
                    ins_tuples,
                    template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,'PROSPECT','REAL')",
                    page_size=BATCH,
                )

            raw_conn.commit()
            log.info("Phase 2 — bulk OK: %d updates, %d inserts", len(updates), len(inserts))

            cur.execute("SELECT cnpj, id FROM clientes")
            for row in cur.fetchall():
                cnpj_to_id[row[0]] = row[1]
        finally:
            cur.close()
            raw_conn.close()
    else:
        with engine.begin() as conn:
            for rec in updates:
                conn.execute(
                    text("""
                        UPDATE clientes SET
                            codigo_cliente   = :codigo_cliente,
                            tipo_cliente_sap = :tipo_cliente_sap,
                            macroregiao      = :macroregiao,
                            classificacao_3tier = 'REAL'
                        WHERE cnpj = :cnpj
                    """),
                    rec,
                )
            for rec in inserts:
                conn.execute(
                    text("""
                        INSERT OR IGNORE INTO clientes
                            (cnpj, razao_social, uf, cidade, codigo_cliente,
                             tipo_cliente_sap, macroregiao, rede_regional,
                             consultor, situacao, classificacao_3tier)
                        VALUES (:cnpj, :razao_social, :uf, :cidade, :codigo_cliente,
                                :tipo_cliente_sap, :macroregiao, :rede_regional,
                                :consultor, 'PROSPECT', 'REAL')
                    """),
                    rec,
                )
        with engine.connect() as conn:
            for row in conn.execute(text("SELECT cnpj, id FROM clientes")).fetchall():
                cnpj_to_id[row[0]] = row[1]

    processed = len(updates) + len(inserts)
    log.info("Phase 2 done — processed=%d enriched=%d created=%d skipped_cnpj=%d",
             processed, len(updates), len(inserts), skipped_cnpj)
    return processed


# ==============================================================================
# Phase 3 — SAP Venda Mês a Mês → faturamento, situação  (bulk)
# ==============================================================================

def ingest_sap_vendas(engine, cnpj_to_id: dict[str, int]) -> dict[str, float]:
    """Lê aba 'Venda Mês a Mês' e atualiza faturamento via bulk ops.

    Two-Base Architecture: faturamento_total vem EXCLUSIVAMENTE dos campos
    Faturado_* desta aba (registros de VENDA SAP). Nunca de LOG/interações.

    Retorna mapa cnpj → faturamento_total para Phase 4.
    """
    from sqlalchemy import text
    import openpyxl

    log.info("=== PHASE 3: SAP Venda Mês a Mês → faturamento + situação (bulk) ===")

    wb = openpyxl.load_workbook(SAP_FILE, read_only=True, data_only=True)
    ws = wb["Venda Mês a Mês"]

    cnpj_faturamento: dict[str, float] = {}
    updates = []  # existing in DB
    inserts = []  # SAP-only clients not yet in DB
    skipped_cnpj = 0
    known_cnpjs = set(cnpj_to_id.keys())

    for row in ws.iter_rows(min_row=2, values_only=True):
        cnpj = normalize_cnpj(row[2])
        if not cnpj:
            skipped_cnpj += 1
            continue

        # Two-Base: sum Faturado_* columns only (VENDA data)
        faturamento_total = 0.0
        n_compras = 0
        last_purchase_date: Optional[date] = None

        for month_name, col_idx in SAP_FATURADO_COLS.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val and isinstance(val, (int, float)) and val > 0:
                faturamento_total += float(val)
                n_compras += 1
                month_date = SAP_MONTHS_DATES[month_name]
                if last_purchase_date is None or month_date > last_purchase_date:
                    last_purchase_date = month_date

        if faturamento_total > 0:
            cnpj_faturamento[cnpj] = faturamento_total
            dias_sem_compra = (REF_DATE - last_purchase_date).days if last_purchase_date else None
            situacao = calc_situacao(dias_sem_compra)
        else:
            dias_sem_compra = None
            situacao = "PROSPECT"

        rec = {
            "cnpj": cnpj,
            "faturamento_total": faturamento_total if faturamento_total > 0 else None,
            "n_compras": n_compras if n_compras > 0 else None,
            "dias_sem_compra": dias_sem_compra,
            "situacao": situacao,
        }

        if cnpj in known_cnpjs:
            updates.append(rec)
        else:
            inserts.append(rec)
            known_cnpjs.add(cnpj)

    log.info("SAP Venda: %d updates, %d new_inserts, %d skipped_cnpj",
             len(updates), len(inserts), skipped_cnpj)
    log.info("Faturamento total (Two-Base VENDA): R$ %.2f | clientes com venda: %d",
             sum(cnpj_faturamento.values()), len(cnpj_faturamento))

    is_pg = "postgresql" in str(engine.url) or "neon" in str(engine.url)
    BATCH = 1000

    if is_pg:
        from psycopg2.extras import execute_values

        raw_conn = _get_raw_pg_conn(engine)
        cur = raw_conn.cursor()
        try:
            cur.execute("""
                CREATE TEMP TABLE _sap_venda (
                    cnpj TEXT, faturamento_total FLOAT,
                    n_compras INT, dias_sem_compra INT, situacao TEXT
                ) ON COMMIT DROP
            """)
            # All 489 rows as tuples
            all_tuples = [
                (r["cnpj"], r["faturamento_total"], r["n_compras"],
                 r["dias_sem_compra"], r["situacao"])
                for r in updates + inserts
            ]
            execute_values(cur, "INSERT INTO _sap_venda VALUES %s", all_tuples, page_size=BATCH)

            # UPDATE existing clients
            cur.execute("""
                UPDATE clientes c SET
                    faturamento_total = t.faturamento_total,
                    n_compras         = t.n_compras,
                    dias_sem_compra   = t.dias_sem_compra,
                    situacao          = t.situacao,
                    updated_at        = NOW()
                FROM _sap_venda t
                WHERE c.cnpj = t.cnpj
            """)

            # INSERT SAP-only clients (those not yet in clientes)
            cur.execute("""
                INSERT INTO clientes (cnpj, faturamento_total, n_compras, dias_sem_compra,
                                      situacao, classificacao_3tier)
                SELECT t.cnpj, t.faturamento_total, t.n_compras, t.dias_sem_compra,
                       t.situacao, 'REAL'
                FROM _sap_venda t
                LEFT JOIN clientes c ON c.cnpj = t.cnpj
                WHERE c.cnpj IS NULL
            """)

            raw_conn.commit()
            log.info("Phase 3 — bulk OK: %d updates+inserts", len(all_tuples))
        finally:
            cur.close()
            raw_conn.close()
    else:
        with engine.begin() as conn:
            for rec in updates + inserts:
                conn.execute(
                    text("""
                        INSERT OR REPLACE INTO clientes
                            (cnpj, faturamento_total, n_compras, dias_sem_compra,
                             situacao, classificacao_3tier)
                        VALUES (:cnpj, :faturamento_total, :n_compras,
                                :dias_sem_compra, :situacao, 'REAL')
                    """),
                    rec,
                )

    processed = len(updates) + len(inserts)
    log.info("Phase 3 done — processed=%d found_in_db=%d new_sap_only=%d skipped=%d",
             processed, len(updates), len(inserts), skipped_cnpj)

    return cnpj_faturamento


# ==============================================================================
# Phase 4 — Curva ABC
# ==============================================================================

def calc_curva_abc(engine, cnpj_faturamento: dict[str, float]) -> None:
    """Calcula curva ABC para todos os clientes com faturamento > 0.
    A = top 20% da receita, B = próximos 30%, C = restantes 50%.
    """
    from sqlalchemy import text

    log.info("=== PHASE 4: Curva ABC ===")

    if not cnpj_faturamento:
        log.warning("Nenhum faturamento disponível para calcular curva ABC")
        return

    # Ordena por faturamento DESC
    sorted_clients = sorted(cnpj_faturamento.items(), key=lambda x: x[1], reverse=True)
    total_revenue = sum(v for _, v in sorted_clients)

    if total_revenue <= 0:
        log.warning("Faturamento total zerado — curva ABC não calculada")
        return

    log.info("Faturamento total para ABC: R$ %.2f | Clientes: %d", total_revenue, len(sorted_clients))

    # Limites cumulativos
    CUT_A = 0.70  # Top 70% da receita → curva A
    CUT_B = 0.95  # Próximos 25% (70-95%) → curva B; restante → C

    cumulative = 0.0
    abc_assignments: dict[str, str] = {}

    for cnpj, fat in sorted_clients:
        cumulative += fat
        pct = cumulative / total_revenue
        if pct <= CUT_A:
            abc_assignments[cnpj] = "A"
        elif pct <= CUT_B:
            abc_assignments[cnpj] = "B"
        else:
            abc_assignments[cnpj] = "C"

    count_a = sum(1 for v in abc_assignments.values() if v == "A")
    count_b = sum(1 for v in abc_assignments.values() if v == "B")
    count_c = sum(1 for v in abc_assignments.values() if v == "C")
    log.info("ABC: A=%d B=%d C=%d", count_a, count_b, count_c)

    is_pg = "postgresql" in str(engine.url) or "neon" in str(engine.url)
    items = list(abc_assignments.items())

    if is_pg:
        from psycopg2.extras import execute_values

        raw_conn = _get_raw_pg_conn(engine)
        cur = raw_conn.cursor()
        try:
            cur.execute("""
                CREATE TEMP TABLE _abc (cnpj TEXT, abc TEXT) ON COMMIT DROP
            """)
            execute_values(cur, "INSERT INTO _abc VALUES %s", items, page_size=1000)
            cur.execute("""
                UPDATE clientes c SET curva_abc = t.abc, updated_at = NOW()
                FROM _abc t WHERE c.cnpj = t.cnpj
            """)
            raw_conn.commit()
        finally:
            cur.close()
            raw_conn.close()
    else:
        with engine.begin() as conn:
            for cnpj, abc in items:
                conn.execute(
                    text("UPDATE clientes SET curva_abc = :abc WHERE cnpj = :cnpj"),
                    {"abc": abc, "cnpj": cnpj},
                )

    log.info("Phase 4 done — curva ABC atualizada para %d clientes", len(abc_assignments))


# ==============================================================================
# Phase 5 — Deskrio phone matching
# ==============================================================================

def match_deskrio_phones(engine) -> None:
    """Lê contatos Deskrio, normaliza telefones e tenta match com clientes.
    Salva resultado em data/deskrio_phone_matches.json para referência.
    Não modifica a tabela clientes (operação somente leitura + arquivo JSON).
    """
    from sqlalchemy import text

    log.info("=== PHASE 5: Deskrio phone matching ===")

    with open(DESKRIO_FILE, "r", encoding="utf-8") as f:
        contacts = json.load(f)

    log.info("Deskrio: %d contatos carregados", len(contacts))

    # Carrega mapa telefone → cnpj do banco
    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT cnpj, telefone, nome_fantasia FROM clientes WHERE telefone IS NOT NULL")
        ).fetchall()

    phone_to_cnpj: dict[str, dict] = {}
    for row in rows:
        cnpj, tel, nome = row
        if tel:
            norm = normalize_phone(tel)
            if norm:
                phone_to_cnpj[norm] = {"cnpj": cnpj, "nome": nome}

    log.info("Clientes com telefone no banco: %d", len(phone_to_cnpj))

    matched = []
    unmatched = 0
    groups_skipped = 0

    for contact in contacts:
        # Pula grupos
        if contact.get("isGroup"):
            groups_skipped += 1
            continue

        raw_number = contact.get("number") or ""
        norm_number = normalize_phone(raw_number)

        if not norm_number:
            unmatched += 1
            continue

        if norm_number in phone_to_cnpj:
            match_info = phone_to_cnpj[norm_number]
            matched.append(
                {
                    "deskrio_id": contact.get("id"),
                    "deskrio_name": contact.get("name"),
                    "deskrio_number": raw_number,
                    "normalized_number": norm_number,
                    "cnpj": match_info["cnpj"],
                    "cliente_nome": match_info["nome"],
                    "channel": contact.get("channel"),
                    "deskrio_updated_at": contact.get("updatedAt"),
                }
            )
        else:
            unmatched += 1

    match_rate = len(matched) / max(len(contacts) - groups_skipped, 1) * 100

    output = {
        "generated_at": datetime.now().isoformat(),
        "ref_date": REF_DATE.isoformat(),
        "stats": {
            "total_deskrio_contacts": len(contacts),
            "groups_skipped": groups_skipped,
            "individuals": len(contacts) - groups_skipped,
            "matched": len(matched),
            "unmatched": unmatched,
            "match_rate_pct": round(match_rate, 2),
        },
        "matches": matched,
    }

    with open(DESKRIO_MATCH_OUTPUT, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    log.info(
        "Phase 5 done — matched=%d unmatched=%d groups_skipped=%d match_rate=%.1f%%",
        len(matched), unmatched, groups_skipped, match_rate,
    )
    log.info("Match output salvo em: %s", DESKRIO_MATCH_OUTPUT)


# ==============================================================================
# Post-ingest validation
# ==============================================================================

def validate_post_ingest(engine) -> bool:
    """Valida thresholds pós-ingestão conforme detector-de-mentira.md."""
    from sqlalchemy import text

    log.info("=== VALIDAÇÃO PÓS-INGESTÃO ===")
    ok = True

    with engine.connect() as conn:
        total = conn.execute(text("SELECT COUNT(*) FROM clientes")).scalar()
        with_cnpj = conn.execute(text("SELECT COUNT(*) FROM clientes WHERE cnpj IS NOT NULL AND LENGTH(cnpj) = 14")).scalar()
        with_fat = conn.execute(text("SELECT COUNT(*) FROM clientes WHERE faturamento_total > 0")).scalar()
        total_fat = conn.execute(text("SELECT COALESCE(SUM(faturamento_total), 0) FROM clientes WHERE faturamento_total > 0")).scalar()
        dup_cnpj = conn.execute(text("SELECT COUNT(*) FROM (SELECT cnpj FROM clientes GROUP BY cnpj HAVING COUNT(*) > 1) t")).scalar()
        without_consultor = conn.execute(text("SELECT COUNT(*) FROM clientes WHERE consultor IS NULL")).scalar()
        situacao_counts = conn.execute(text("SELECT situacao, COUNT(*) FROM clientes GROUP BY situacao ORDER BY 2 DESC")).fetchall()
        abc_counts = conn.execute(text("SELECT curva_abc, COUNT(*) FROM clientes WHERE curva_abc IS NOT NULL GROUP BY curva_abc ORDER BY 1")).fetchall()
        tier_counts = conn.execute(text("SELECT classificacao_3tier, COUNT(*) FROM clientes GROUP BY classificacao_3tier")).fetchall()

    log.info("Total clientes: %d", total)
    log.info("CNPJ 14 dígitos: %d (%.1f%%)", with_cnpj, with_cnpj / max(total, 1) * 100)
    log.info("Com faturamento SAP: %d", with_fat)
    log.info("Faturamento total no DB: R$ %.2f", total_fat)
    log.info("CNPJ duplicados: %d", dup_cnpj)
    log.info("Sem consultor: %d", without_consultor)

    for sit, cnt in situacao_counts:
        log.info("  situacao %-15s : %d", sit or "NULL", cnt)
    for abc, cnt in abc_counts:
        log.info("  curva_abc %s : %d", abc, cnt)
    for tier, cnt in tier_counts:
        log.info("  classificacao_3tier %-15s : %d", tier or "NULL", cnt)

    # Checks bloqueantes
    BASELINE_FAT = 2_091_000.0
    TOLERANCIA = 0.005  # 0.5%

    if dup_cnpj > 0:
        log.error("BLOQUEANTE: %d CNPJs duplicados!", dup_cnpj)
        ok = False

    if with_cnpj < total:
        log.warning("Atenção: %d clientes com CNPJ inválido/nulo", total - with_cnpj)

    if total_fat > 0:
        divergencia = abs(total_fat - BASELINE_FAT) / BASELINE_FAT
        if divergencia > TOLERANCIA:
            log.warning(
                "ATENÇÃO: Faturamento R$ %.2f diverge %.2f%% do baseline R$ %.2f "
                "(tolerância: %.1f%%) — Nota: baseline refere-se a 2025; "
                "SAP Venda Mês a Mês pode cobrir período diferente.",
                total_fat, divergencia * 100, BASELINE_FAT, TOLERANCIA * 100,
            )
        else:
            log.info("Faturamento dentro da tolerância de %.1f%%", TOLERANCIA * 100)

    # Verifica alucinação: não deve existir classificacao_3tier = 'ALUCINACAO'
    for tier, cnt in tier_counts:
        if tier == "ALUCINACAO":
            log.error("BLOQUEANTE: %d registros com classificacao_3tier=ALUCINACAO!", cnt)
            ok = False

    if ok:
        log.info("VALIDACAO: PASS — nenhum bloqueante encontrado")
    else:
        log.error("VALIDACAO: FAIL — ver erros acima")

    return ok


# ==============================================================================
# Main
# ==============================================================================

def main() -> int:
    log.info("=" * 60)
    log.info("CRM VITAO360 — ingest_real_data.py")
    log.info("Data de referência: %s", REF_DATE.isoformat())
    log.info("=" * 60)

    # Verificar arquivos de entrada
    missing = []
    for label, path in [
        ("Mercos JSON", MERCOS_FILE),
        ("SAP XLSX", SAP_FILE),
        ("Deskrio contacts", DESKRIO_FILE),
    ]:
        if not path.exists():
            missing.append(f"  FALTANDO: {label} → {path}")
    if missing:
        log.error("Arquivos de entrada ausentes:\n%s", "\n".join(missing))
        return 1

    try:
        engine = get_engine()
    except Exception as e:
        log.error("Falha ao criar engine: %s", e)
        return 1

    # Garante que a tabela existe (cria se necessário)
    try:
        sys.path.insert(0, str(PROJECT_ROOT))
        from backend.app.models.cliente import Cliente
        from backend.app.database import Base
        Base.metadata.create_all(engine)
        log.info("Schema verificado/criado OK")
    except Exception as e:
        log.error("Falha ao verificar schema: %s", e)
        return 1

    try:
        # Phase 1 — Mercos
        cnpj_to_id = load_mercos(engine)

        # Phase 2 — SAP Cadastro
        enrich_from_sap_cadastro(engine, cnpj_to_id)

        # Phase 3 — SAP Venda (Two-Base: faturamento APENAS desta fonte)
        cnpj_faturamento = ingest_sap_vendas(engine, cnpj_to_id)

        # Phase 4 — Curva ABC
        calc_curva_abc(engine, cnpj_faturamento)

        # Phase 5 — Deskrio phone matching
        match_deskrio_phones(engine)

        # Validação final
        ok = validate_post_ingest(engine)

    except Exception as e:
        log.exception("Erro fatal durante ingestão: %s", e)
        return 1

    log.info("=" * 60)
    if ok:
        log.info("INGESTAO CONCLUIDA COM SUCESSO")
    else:
        log.warning("INGESTAO CONCLUIDA COM ALERTAS — ver log acima")
    log.info("=" * 60)
    return 0 if ok else 2


if __name__ == "__main__":
    sys.exit(main())

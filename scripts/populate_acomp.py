"""
Populate ACOMPANHAMENTO (cols 73-257) in existing CRM V3.
- Reads META from SAP file (by GRUPO CHAVE, distributed proportionally to clients)
- Reads REALIZADO from already-populated CARTEIRA vendas mensais
- Writes formulas for % ATING, % TRI, % YTD, % MÊS
"""
import os
import sys
import time
import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from v3_styles import FMT_MONEY, FMT_PCT, FMT_DATE

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DESKTOP = os.path.dirname(BASE_DIR)
OUTPUT_FILE = os.path.join(BASE_DIR, "output", "CRM_VITAO360_V3_100.xlsx")
SAP_FILE = os.path.join(DESKTOP, "PASTA G (CENTRAL INTERNO)", "CARTEIRA DE CLIENTES SAP",
                        "BASE SAP - META E PROJEÇÃO 2026 .- 02. INTERNO - 2026.xlsx")

MESES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]

# 15 cols per month
MONTH_COLS_PER = [
    "% YTD", "META", "REALIZADO", "% TRI", "META", "REALIZADO",
    "% MÊS", "META", "REALIZADO", "DATA PEDIDO",
    "JUSTIFICATIVA SEMANA 1", "JUSTIFICATIVA SEMANA 2",
    "JUSTIFICATIVA SEMANA 3", "JUSTIFICATIVA SEMANA 4",
    "JUSTIFICATIVA MENSAL",
]


def compute_month_starts():
    """Compute the starting column for each month in ACOMPANHAMENTO."""
    col = 73  # % ATING
    col += 1  # skip % ATING → col 74
    month_starts = {}
    for q_idx in range(4):
        col += 1  # skip % Q
        for m_idx in range(3):
            month_idx = q_idx * 3 + m_idx
            mes = MESES[month_idx]
            month_starts[mes] = col
            col += len(MONTH_COLS_PER)
    return month_starts


def compute_q_cols():
    """Compute the column for each quarterly % header."""
    col = 74  # first Q
    q_cols = {}
    for q_idx in range(4):
        q_cols[q_idx + 1] = col
        col += 1  # % Q
        col += 3 * len(MONTH_COLS_PER)  # 3 months × 15 cols
    return q_cols


def extract_sap_meta():
    """Extract META per GRUPO CHAVE per month from SAP file."""
    print("  Reading SAP META...")
    wb = openpyxl.load_workbook(SAP_FILE, data_only=True, read_only=True)
    ws = wb["Faturamento"]

    # Headers in row 3: col K(11)=GRUPO CHAVE, col N-Y(14-25)=JAN-DEZ
    meta_por_grupo = {}
    for row in ws.iter_rows(min_row=4, max_col=26, values_only=True):
        totais = str(row[0] or "").strip().upper()
        grupo_prod = str(row[11] or "").strip()
        grupo_chave = str(row[10] or "").strip()

        # Only aggregate "TOTAL" product lines (not sub-products)
        if totais != "SIM" or grupo_prod != "01. TOTAL":
            continue

        if grupo_chave not in meta_por_grupo:
            meta_por_grupo[grupo_chave] = {m: 0 for m in range(1, 13)}
            meta_por_grupo[grupo_chave]['total'] = 0

        for m in range(1, 13):
            val = row[13 + m - 1]  # cols N(14) through Y(25) → idx 13..24
            if val and isinstance(val, (int, float)):
                meta_por_grupo[grupo_chave][m] += val
                meta_por_grupo[grupo_chave]['total'] += val

    wb.close()

    total = sum(g['total'] for g in meta_por_grupo.values())
    print(f"    → {len(meta_por_grupo)} grupos, META total = R$ {total:,.0f}")
    return meta_por_grupo


def read_client_data(wb):
    """Read client CNPJ, GRUPO CHAVE, and historical sales from CARTEIRA tab."""
    ws = wb["CARTEIRA"]
    clients = []

    for r in range(4, ws.max_row + 1):
        cnpj = ws.cell(row=r, column=2).value  # col B
        if not cnpj:
            break

        nome = ws.cell(row=r, column=1).value  # col A
        grupo_chave = str(ws.cell(row=r, column=71).value or "").strip()  # col BS = 71
        rede = str(ws.cell(row=r, column=9).value or "").strip()  # col I

        # Total vendas período (col 25 in our CARTEIRA = TOTAL PERÍODO)
        total_vendas = ws.cell(row=r, column=25).value or 0
        if not isinstance(total_vendas, (int, float)):
            total_vendas = 0

        # Vendas mensais (cols 26-36 = MAR/25 through JAN/26, 11 months)
        vendas_mes = []
        for c in range(26, 37):
            v = ws.cell(row=r, column=c).value
            vendas_mes.append(v if isinstance(v, (int, float)) else 0)

        clients.append({
            'row': r,
            'cnpj': str(cnpj),
            'nome': nome,
            'grupo_chave': grupo_chave,
            'rede': rede,
            'total_vendas': total_vendas,
            'vendas_mes': vendas_mes,
        })

    print(f"    → {len(clients)} clients read from CARTEIRA")
    return clients


def map_client_to_grupo(clients, meta_por_grupo):
    """Map each client to their GRUPO CHAVE for meta distribution."""
    # Group clients by GRUPO CHAVE
    grupos = {}
    for cl in clients:
        gc = cl['grupo_chave']
        if gc and gc in meta_por_grupo:
            if gc not in grupos:
                grupos[gc] = []
            grupos[gc].append(cl)

    # Clients without grupo or with unknown grupo → "SEM GRUPO"
    sem_grupo = []
    for cl in clients:
        gc = cl['grupo_chave']
        if not gc or gc not in meta_por_grupo:
            sem_grupo.append(cl)

    if "SEM GRUPO" in meta_por_grupo:
        grupos["SEM GRUPO"] = sem_grupo
    elif sem_grupo:
        # Try matching by REDE name
        rede_to_grupo = {}
        for g in meta_por_grupo:
            g_upper = g.upper()
            rede_to_grupo[g_upper] = g
            # Also add common variants
            for prefix in g_upper.split():
                if len(prefix) > 3:
                    rede_to_grupo[prefix] = g

        still_unmatched = []
        for cl in sem_grupo:
            rede_upper = cl['rede'].upper()
            matched = False
            for key, grupo in rede_to_grupo.items():
                if key in rede_upper or rede_upper in key:
                    if grupo not in grupos:
                        grupos[grupo] = []
                    grupos[grupo].append(cl)
                    matched = True
                    break
            if not matched:
                still_unmatched.append(cl)

        if still_unmatched and "SEM GRUPO" in meta_por_grupo:
            grupos["SEM GRUPO"] = still_unmatched

    matched = sum(len(v) for v in grupos.values())
    print(f"    → {matched}/{len(clients)} clients matched to GRUPO CHAVE")
    return grupos


def distribute_meta(grupos, meta_por_grupo):
    """Distribute META proportionally per client within each GRUPO CHAVE."""
    client_meta = {}  # cnpj → {1: jan_meta, ..., 12: dez_meta}

    for grupo, clientes in grupos.items():
        if grupo not in meta_por_grupo:
            continue

        grupo_meta = meta_por_grupo[grupo]
        total_vendas_grupo = sum(cl['total_vendas'] for cl in clientes)

        for cl in clientes:
            cnpj = cl['cnpj']
            if total_vendas_grupo > 0:
                # Proportional distribution based on historical sales
                proporcao = cl['total_vendas'] / total_vendas_grupo
            else:
                # Equal distribution if no sales history
                proporcao = 1.0 / len(clientes) if clientes else 0

            client_meta[cnpj] = {}
            for m in range(1, 13):
                client_meta[cnpj][m] = round(grupo_meta.get(m, 0) * proporcao, 2)

    print(f"    → META distributed to {len(client_meta)} clients")
    return client_meta


def populate_acompanhamento(wb, clients, client_meta):
    """Write META, REALIZADO, and % formulas into ACOMPANHAMENTO columns."""
    ws = wb["CARTEIRA"]
    month_starts = compute_month_starts()
    q_cols = compute_q_cols()

    print(f"  Populating ACOMPANHAMENTO for {len(clients)} clients...")

    # VENDAS months mapping: our CARTEIRA cols 26-36 = MAR/25..JAN/26
    # But ACOMPANHAMENTO months are JAN..DEZ (2026 calendar)
    # We only have REALIZADO for JAN/26 (col 36 in vendas_mes, index 10)
    # For 2026, we use the historical as basis

    # Map: ACOMPANHAMENTO month index (1=JAN..12=DEZ) → vendas_mes index
    # vendas_mes: [MAR/25(0), ABR/25(1), MAI/25(2), JUN/25(3), JUL/25(4),
    #              AGO/25(5), SET/25(6), OUT/25(7), NOV/25(8), DEZ/25(9), JAN/26(10)]
    # For REALIZADO 2026: only JAN/26 has data (index 10)
    # FEV/26 onwards = 0 (future)

    populated = 0
    for cl in clients:
        r = cl['row']
        cnpj = cl['cnpj']
        meta_data = client_meta.get(cnpj, {})

        jan_realizado = cl['vendas_mes'][10] if len(cl['vendas_mes']) > 10 else 0  # JAN/26

        for month_idx in range(12):  # 0-11 = JAN-DEZ
            mes = MESES[month_idx]
            m = month_idx + 1  # 1-12
            start_col = month_starts[mes]

            meta_mensal = meta_data.get(m, 0)

            # REALIZADO: only JAN/26 has data
            if m == 1:
                realizado_mensal = jan_realizado
            else:
                realizado_mensal = 0

            # offset 7 = META MÊS
            meta_col = start_col + 7
            ws.cell(row=r, column=meta_col, value=meta_mensal).number_format = FMT_MONEY

            # offset 8 = REALIZADO MÊS
            real_col = start_col + 8
            ws.cell(row=r, column=real_col, value=realizado_mensal).number_format = FMT_MONEY

            # offset 6 = % MÊS = IF(META=0,0,REALIZADO/META)
            pct_col = start_col + 6
            mc = get_column_letter(meta_col)
            rc = get_column_letter(real_col)
            ws.cell(row=r, column=pct_col).value = f'=IF({mc}{r}=0,0,{rc}{r}/{mc}{r})'
            ws.cell(row=r, column=pct_col).number_format = FMT_PCT

            # META TRI (offset 4) and REALIZADO TRI (offset 5)
            # Trimester = sum of 3 months in quarter
            q_idx = month_idx // 3  # 0=Q1, 1=Q2, 2=Q3, 3=Q4
            m_in_q = month_idx % 3  # 0, 1, 2

            # Only write TRI totals on the LAST month of each quarter
            if m_in_q == 2:
                # Get the 3 months in this quarter
                m1_start = month_starts[MESES[q_idx * 3]]
                m2_start = month_starts[MESES[q_idx * 3 + 1]]
                m3_start = month_starts[MESES[q_idx * 3 + 2]]

                meta_tri_formula = (
                    f'={get_column_letter(m1_start+7)}{r}'
                    f'+{get_column_letter(m2_start+7)}{r}'
                    f'+{get_column_letter(m3_start+7)}{r}'
                )
                real_tri_formula = (
                    f'={get_column_letter(m1_start+8)}{r}'
                    f'+{get_column_letter(m2_start+8)}{r}'
                    f'+{get_column_letter(m3_start+8)}{r}'
                )

                # Write TRI for all 3 months in the quarter
                for mi in range(3):
                    ms = month_starts[MESES[q_idx * 3 + mi]]
                    ws.cell(row=r, column=ms + 4, value=meta_tri_formula).number_format = FMT_MONEY
                    ws.cell(row=r, column=ms + 5, value=real_tri_formula).number_format = FMT_MONEY
                    # % TRI (offset 3)
                    mtc = get_column_letter(ms + 4)
                    rtc = get_column_letter(ms + 5)
                    ws.cell(row=r, column=ms + 3).value = f'=IF({mtc}{r}=0,0,{rtc}{r}/{mtc}{r})'
                    ws.cell(row=r, column=ms + 3).number_format = FMT_PCT

            # META YTD (offset 1) = sum of all META MÊS from JAN to current month
            # REALIZADO YTD (offset 2) = sum of all REALIZADO MÊS from JAN to current month
            ytd_meta_parts = []
            ytd_real_parts = []
            for prev_m in range(month_idx + 1):
                prev_start = month_starts[MESES[prev_m]]
                ytd_meta_parts.append(f'{get_column_letter(prev_start+7)}{r}')
                ytd_real_parts.append(f'{get_column_letter(prev_start+8)}{r}')

            ws.cell(row=r, column=start_col + 1).value = '=' + '+'.join(ytd_meta_parts)
            ws.cell(row=r, column=start_col + 1).number_format = FMT_MONEY
            ws.cell(row=r, column=start_col + 2).value = '=' + '+'.join(ytd_real_parts)
            ws.cell(row=r, column=start_col + 2).number_format = FMT_MONEY

            # % YTD (offset 0) = REALIZADO_YTD / META_YTD
            ytd_mc = get_column_letter(start_col + 1)
            ytd_rc = get_column_letter(start_col + 2)
            ws.cell(row=r, column=start_col).value = f'=IF({ytd_mc}{r}=0,0,{ytd_rc}{r}/{ytd_mc}{r})'
            ws.cell(row=r, column=start_col).number_format = FMT_PCT

        # ── % ATING (col 73) = current month % MÊS (JAN for now) ──
        jan_pct_col = month_starts["JAN"] + 6
        ws.cell(row=r, column=73).value = f'={get_column_letter(jan_pct_col)}{r}'
        ws.cell(row=r, column=73).number_format = FMT_PCT

        # ── Quarterly % cols ──
        for q_num in range(1, 5):
            q_col = q_cols[q_num]
            m1 = MESES[(q_num - 1) * 3]
            m2 = MESES[(q_num - 1) * 3 + 1]
            m3 = MESES[(q_num - 1) * 3 + 2]
            p1 = get_column_letter(month_starts[m1] + 6)
            p2 = get_column_letter(month_starts[m2] + 6)
            p3 = get_column_letter(month_starts[m3] + 6)
            ws.cell(row=r, column=q_col).value = f'=AVERAGE({p1}{r},{p2}{r},{p3}{r})'
            ws.cell(row=r, column=q_col).number_format = FMT_PCT

        populated += 1

    print(f"    → ACOMPANHAMENTO: {populated} clients × 12 months populated")
    return populated


def main():
    start = time.time()
    print("=" * 60)
    print("  CRM VITAO360 V3 — POPULATE ACOMPANHAMENTO")
    print("=" * 60)

    # Step 1: Extract SAP META
    print("\n[1/4] Extracting SAP META...")
    meta_por_grupo = extract_sap_meta()

    # Step 2: Open existing V3 file
    print("\n[2/4] Opening populated V3...")
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    # Step 3: Read client data
    print("\n[3/4] Reading client data...")
    clients = read_client_data(wb)

    # Map clients to GRUPO CHAVE
    grupos = map_client_to_grupo(clients, meta_por_grupo)

    # Distribute META
    client_meta = distribute_meta(grupos, meta_por_grupo)

    # Step 4: Populate ACOMPANHAMENTO
    print("\n[4/4] Populating ACOMPANHAMENTO...")
    populated = populate_acompanhamento(wb, clients, client_meta)

    # Save
    wb.save(OUTPUT_FILE)
    elapsed = time.time() - start

    print("\n" + "=" * 60)
    print(f"  SAVED: {OUTPUT_FILE}")
    print(f"  ACOMPANHAMENTO: {populated} clients × 12 months")
    print(f"  Time: {elapsed:.1f}s")
    print("=" * 60)


if __name__ == "__main__":
    main()

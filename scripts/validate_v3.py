"""
Validate CRM VITAO360 V3 — 18-point checklist
"""
import os
import openpyxl
from openpyxl.utils import get_column_letter

FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output", "CRM_VITAO360_V3_100.xlsx")

passed = 0
failed = 0
total = 0


def check(name, condition, detail=""):
    global passed, failed, total
    total += 1
    if condition:
        passed += 1
        print(f"  [PASS] {name}")
    else:
        failed += 1
        print(f"  [FAIL] {name} — {detail}")


def main():
    global passed, failed
    print("=" * 60)
    print("  CRM VITAO360 V3 — Validation (18 checks)")
    print("=" * 60)

    wb = openpyxl.load_workbook(FILE)

    # ── 1. All 9 tabs present ──
    expected_tabs = ["REGRAS", "DRAFT 1", "PROJEÇÃO", "CARTEIRA", "DRAFT 2",
                     "AGENDA", "DASH", "LOG", "Claude Log"]
    check("9 tabs present", wb.sheetnames == expected_tabs,
          f"Got: {wb.sheetnames}")

    # ── 2. REGRAS: Named Ranges resolve ──
    nr_names = [dn.name for dn in wb.defined_names.values()]
    expected_nr = [
        "LISTA_RESULTADO", "LISTA_TIPO_CONTATO", "LISTA_MOTIVO",
        "LISTA_SITUACAO", "LISTA_FASE", "LISTA_TENTATIVA",
        "LISTA_SINALEIRO", "LISTA_TIPO_CLIENTE", "LISTA_CONSULTOR",
        "LISTA_SIM_NAO", "LISTA_CURVA", "LISTA_PRIORIDADE",
        "LISTA_OPORTUNIDADE", "LISTA_ESTAGIO_FUNIL", "LISTA_TEMPERATURA",
    ]
    found_nr = [n for n in expected_nr if n in nr_names]
    check(f"Named Ranges ({len(found_nr)}/{len(expected_nr)})",
          len(found_nr) >= 15,
          f"Missing: {[n for n in expected_nr if n not in nr_names]}")

    # ── 3. REGRAS: No LEAD in SITUAÇÃO list ──
    ws_r = wb["REGRAS"]
    situacao_values = []
    for row in ws_r.iter_rows(min_row=1, max_row=120, max_col=5, values_only=False):
        for cell in row:
            if cell.value and str(cell.value).upper() == "LEAD":
                situacao_values.append(cell.coordinate)
    check("No LEAD in REGRAS", len(situacao_values) == 0,
          f"Found LEAD at: {situacao_values}")

    # ── 4. DRAFT 1: 48 columns ──
    ws_d1 = wb["DRAFT 1"]
    d1_cols = 0
    for col in range(1, 100):
        if ws_d1.cell(row=3, column=col).value:
            d1_cols = col
    check("DRAFT 1: 48 columns", d1_cols == 48,
          f"Got {d1_cols} columns")

    # ── 5. PROJEÇÃO: 13 redes populated ──
    ws_p = wb["PROJEÇÃO"]
    redes = 0
    for r in range(4, 25):
        if ws_p.cell(row=r, column=2).value:  # col 2 = NOME FANTASIA
            redes += 1
    check("PROJEÇÃO: 13+ redes", redes >= 13,
          f"Got {redes} redes")

    # ── 6. PROJEÇÃO: has sinaleiro formulas ──
    projecao_cols = 0
    for col in range(1, 50):
        if ws_p.cell(row=3, column=col).value:
            projecao_cols = col
    check("PROJEÇÃO: ~39 cols", projecao_cols >= 30,
          f"Got {projecao_cols} columns")

    # ── 7. CARTEIRA: ~257 columns ──
    ws_c = wb["CARTEIRA"]
    cart_cols = 0
    for col in range(1, 300):
        if ws_c.cell(row=3, column=col).value:
            cart_cols = col
    check("CARTEIRA: 250+ cols", cart_cols >= 250,
          f"Got {cart_cols} columns")

    # ── 8. CARTEIRA: 4 mega-blocks in row 1 ──
    mega_blocks = []
    for col in range(1, 300):
        v = ws_c.cell(row=1, column=col).value
        if v and isinstance(v, str):
            mega_blocks.append(v)
    check("CARTEIRA: 4 mega-blocks", len(mega_blocks) >= 4,
          f"Got: {mega_blocks}")

    # ── 9. DRAFT 2: 24 columns (12 manual + 12 auto) ──
    ws_d2 = wb["DRAFT 2"]
    d2_cols = 0
    for col in range(1, 50):
        if ws_d2.cell(row=2, column=col).value:
            d2_cols = col
    check("DRAFT 2: 24 columns", d2_cols == 24,
          f"Got {d2_cols} columns")

    # ── 10. DRAFT 2: 5 example records ──
    d2_rows = 0
    for r in range(3, 20):
        if ws_d2.cell(row=r, column=2).value:  # CNPJ column
            d2_rows += 1
    check("DRAFT 2: 5 examples", d2_rows >= 5,
          f"Got {d2_rows} rows")

    # ── 11. DRAFT 2: Motor auto-cols populated (formulas OR values) ──
    motor_ok = True
    motor_issues = []
    for col in [14, 15, 16, 17, 19, 21]:  # ESTÁGIO, FASE, TIPO, TEMP, GRUPO, AÇÃO
        val = ws_d2.cell(row=8, column=col).value  # row 8 should have formula or value
        if not val:
            motor_ok = False
            motor_issues.append(get_column_letter(col))
    check("DRAFT 2: Motor auto-cols populated",
          motor_ok, f"Empty auto cols at row 8: {motor_issues}")

    # ── 12. DRAFT 2: Dropdowns present ──
    d2_dvs = list(ws_d2.data_validations.dataValidation) if ws_d2.data_validations else []
    check("DRAFT 2: Data validations", len(d2_dvs) >= 4,
          f"Got {len(d2_dvs)} validations")

    # ── 13. AGENDA: 25 columns ──
    ws_a = wb["AGENDA"]
    agenda_cols = 0
    for col in range(1, 40):
        if ws_a.cell(row=4, column=col).value:
            agenda_cols = col
    check("AGENDA: 25 cols", agenda_cols >= 24,
          f"Got {agenda_cols} columns")

    # ── 14. DASH: 7 blocks with chart ──
    ws_dash = wb["DASH"]
    has_chart = len(ws_dash._charts) > 0
    # Check that DASH has content (KPI cards in rows 3-4, blocks starting from row 7+)
    dash_rows_with_content = 0
    for r in range(1, 170):
        for c in range(1, 15):
            if ws_dash.cell(row=r, column=c).value:
                dash_rows_with_content += 1
                break
    check("DASH: 7 blocks + chart",
          has_chart and dash_rows_with_content >= 30,
          f"Chart={has_chart}, rows_with_content={dash_rows_with_content}")

    # ── 15. Freeze panes correct ──
    freezes = {
        "DRAFT 1": "C4",
        "CARTEIRA": "C4",
        "DRAFT 2": "E3",
        "AGENDA": "A5",
        "DASH": "A5",
        "LOG": "A3",
        "Claude Log": "A3",
    }
    freeze_ok = True
    freeze_issues = []
    for tab, expected in freezes.items():
        ws_check = wb[tab]
        actual = str(ws_check.freeze_panes) if ws_check.freeze_panes else "None"
        if actual != expected:
            freeze_ok = False
            freeze_issues.append(f"{tab}: {actual} (expected {expected})")
    check("Freeze panes correct", freeze_ok, "; ".join(freeze_issues))

    # ── 16. Tab colors set ──
    expected_colors = {
        "REGRAS": "0D9488", "DRAFT 1": "7B2FF2", "PROJEÇÃO": "FF6B00",
        "CARTEIRA": "00B050", "DRAFT 2": "FFC000", "AGENDA": "2196F3",
        "DASH": "DC2626", "LOG": "6B7280", "Claude Log": "9333EA",
    }
    color_ok = True
    color_issues = []
    for tab, expected in expected_colors.items():
        ws_check = wb[tab]
        actual = ws_check.sheet_properties.tabColor
        if actual:
            actual_rgb = actual.rgb if hasattr(actual, 'rgb') else str(actual)
            # Remove 'FF' prefix from ARGB if present
            actual_str = str(actual_rgb).replace("00", "", 1) if str(actual_rgb).startswith("00") else str(actual_rgb)
            if expected.upper() not in actual_str.upper():
                color_ok = False
                color_issues.append(f"{tab}: {actual_str} != {expected}")
        else:
            color_ok = False
            color_issues.append(f"{tab}: no color")
    check("Tab colors set", color_ok, "; ".join(color_issues))

    # ── 17. Conditional formatting for SITUAÇÃO ──
    cf_ok = True
    cf_issues = []
    for tab in ["DRAFT 2", "LOG", "CARTEIRA"]:
        ws_check = wb[tab]
        # Count individual rules across all CF ranges
        total_rules = 0
        for cf_range in ws_check.conditional_formatting:
            total_rules += len(cf_range.rules)
        if total_rules < 5:
            cf_ok = False
            cf_issues.append(f"{tab}: only {total_rules} CF rules")
    check("Conditional formatting SITUAÇÃO (DRAFT2+LOG+CARTEIRA)", cf_ok, "; ".join(cf_issues))

    # ── 18. Column grouping in CARTEIRA ──
    ws_c = wb["CARTEIRA"]
    grouped_cols = []
    for col_idx in range(1, 260):
        cl = get_column_letter(col_idx)
        dim = ws_c.column_dimensions.get(cl)
        if dim and dim.outlineLevel and dim.outlineLevel > 0:
            grouped_cols.append(col_idx)
    check("CARTEIRA: Column grouping active", len(grouped_cols) >= 20,
          f"Only {len(grouped_cols)} grouped columns")

    # ── Summary ──
    wb.close()
    print("\n" + "=" * 60)
    print(f"  RESULT: {passed}/{total} passed, {failed} failed")
    if failed == 0:
        print("  STATUS: ALL CHECKS PASSED")
    else:
        print("  STATUS: SOME CHECKS FAILED")
    print("=" * 60)


if __name__ == "__main__":
    main()

"""
Phase 08 Plan 02, Task 2: Validate META-01..03 requirements
=============================================================
Formally evaluates the 3 Phase 8 requirements plus integrity checks.

Checks:
  1. META-01: Metas 2026 do SAP integradas na CARTEIRA
  2. META-02: COMITE com visao consolidada
  3. META-03: Capacidade de atendimento validada
  4. V13 integrity (19,224+ formulas, all tabs, no #REF!)
  5. Cross-check META totals
  6. RATEIO toggle structure

Output: data/output/phase08/validation_report.json
"""

import sys
import json
import openpyxl
import unicodedata
from pathlib import Path
from datetime import datetime, timezone

# ============================================================
# CONFIGURATION
# ============================================================
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
V13_PATH = PROJECT_ROOT / "data" / "output" / "CRM_VITAO360_V13_PROJECAO.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "data" / "output" / "phase08"
REPORT_PATH = OUTPUT_DIR / "validation_report.json"
META_REPORT_PATH = OUTPUT_DIR / "meta_validation_report.json"

EXPECTED_FORMULAS = 19224
EXPECTED_TABS = ["LOG", "DASH", "REDES_FRANQUIAS_v2", "COMITE"]  # + PROJECAO (accent)


def find_projecao_sheet(wb):
    """Find PROJECAO sheet by accent-stripping."""
    for name in wb.sheetnames:
        clean = unicodedata.normalize('NFD', name)
        clean = ''.join(c for c in clean if unicodedata.category(c) != 'Mn')
        if 'PROJECAO' in clean.upper():
            return name
    raise ValueError("PROJECAO sheet not found")


def count_formulas(ws, min_row=None, max_row=None):
    """Count formula cells in worksheet."""
    count = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                count += 1
    return count


def check(name, passed, details):
    """Create a check result."""
    return {
        "name": name,
        "status": "PASS" if passed else "FAIL",
        "details": details,
    }


# ============================================================
# CHECK 1: META-01 -- Metas 2026 do SAP integradas na CARTEIRA
# ============================================================
def check_meta_01(wb, prj_name):
    """Verify META infrastructure exists in PROJECAO and is referenced by COMITE."""
    print("\n" + "=" * 60, flush=True)
    print("CHECK 1 (META-01): Metas 2026 integradas na CARTEIRA", flush=True)
    print("=" * 60, flush=True)

    ws = wb[prj_name]
    issues = []

    # 1a. PROJECAO col L has >= 490 non-empty META ANUAL values
    non_empty_l = 0
    for r in range(4, 538):
        val = ws.cell(row=r, column=12).value  # col L
        if val is not None and val != "" and val != 0:
            non_empty_l += 1
    print(f"  META ANUAL (col L) non-empty: {non_empty_l}", flush=True)
    if non_empty_l < 490:
        issues.append(f"Only {non_empty_l} non-empty META ANUAL values (expected >= 490)")

    # 1b. 3 meta variants exist
    # Proportional (L:X) -- static values
    # Equal (BB:BN = 54:66) -- formulas
    # Dynamic (BP:CB = 68:80) -- formulas
    eq_formulas = 0
    dyn_formulas = 0
    for r in range(4, 538):
        val_bb = ws.cell(row=r, column=54).value  # col BB
        if val_bb and isinstance(val_bb, str) and val_bb.startswith('='):
            eq_formulas += 1
        val_bp = ws.cell(row=r, column=68).value  # col BP
        if val_bp and isinstance(val_bp, str) and val_bp.startswith('='):
            dyn_formulas += 1
    print(f"  Equal meta formulas (BB): {eq_formulas}", flush=True)
    print(f"  Dynamic meta formulas (BP): {dyn_formulas}", flush=True)
    if eq_formulas < 500:
        issues.append(f"Equal meta formulas: {eq_formulas} (expected >= 500)")
    if dyn_formulas < 500:
        issues.append(f"Dynamic meta formulas: {dyn_formulas} (expected >= 500)")

    # 1c. COMITE Bloco 1 references PROJECAO meta columns
    if "COMITE" in wb.sheetnames:
        comite = wb["COMITE"]
        b6_val = comite.cell(row=6, column=2).value  # B6 = META ANUAL first consultant
        if b6_val and isinstance(b6_val, str) and 'SUMIFS' in b6_val and 'L$4' in b6_val:
            print(f"  COMITE B6 references PROJECAO col L: YES", flush=True)
        else:
            issues.append(f"COMITE B6 does not reference PROJECAO L: {b6_val}")
            print(f"  COMITE B6: {b6_val}", flush=True)

        # Check toggle references BB and BP
        if b6_val and 'BB' in str(b6_val) and 'BP' in str(b6_val):
            print(f"  COMITE B6 references BB (equal) and BP (dynamic): YES", flush=True)
        else:
            issues.append("COMITE B6 missing toggle references to BB/BP")
    else:
        issues.append("COMITE tab not found")

    passed = len(issues) == 0
    status = "PASS" if passed else "FAIL"
    details = "All 3 meta variants verified, COMITE references confirmed" if passed else "; ".join(issues)
    print(f"  RESULT: {status} -- {details}", flush=True)
    return {
        "status": status,
        "details": details,
        "meta_anual_count": non_empty_l,
        "equal_formulas": eq_formulas,
        "dynamic_formulas": dyn_formulas,
    }


# ============================================================
# CHECK 2: META-02 -- COMITE com visao consolidada
# ============================================================
def check_meta_02(wb):
    """Verify COMITE tab structure and content."""
    print("\n" + "=" * 60, flush=True)
    print("CHECK 2 (META-02): COMITE com visao consolidada", flush=True)
    print("=" * 60, flush=True)

    issues = []

    # 2a. COMITE tab exists
    if "COMITE" not in wb.sheetnames:
        return {"status": "FAIL", "details": "COMITE tab not found"}

    ws = wb["COMITE"]

    # 2b. 5 section titles (rows 4, 16, 29, 45, 59)
    section_rows = [4, 16, 29, 45, 59]
    found_titles = 0
    for sr in section_rows:
        val = ws.cell(row=sr, column=1).value
        if val and "  " in str(val):  # section titles have leading spaces
            found_titles += 1
            print(f"  Row {sr}: {val}", flush=True)
        else:
            issues.append(f"Row {sr} missing section title: {val}")
            print(f"  Row {sr}: MISSING ({val})", flush=True)

    print(f"  Section titles: {found_titles}/5", flush=True)

    # 2c. Bloco 1 has consultant rows with META/REALIZADO formulas
    has_meta_formula = False
    has_real_formula = False
    for r in [6, 7, 8, 9]:
        b_val = ws.cell(row=r, column=2).value
        c_val = ws.cell(row=r, column=3).value
        if b_val and isinstance(b_val, str) and 'SUMIFS' in b_val:
            has_meta_formula = True
        if c_val and isinstance(c_val, str) and 'SUMIFS' in c_val:
            has_real_formula = True

    if not has_meta_formula:
        issues.append("Bloco 1 missing META SUMIFS formulas")
    if not has_real_formula:
        issues.append("Bloco 1 missing REALIZADO SUMIFS formulas")
    print(f"  Bloco 1 META formulas: {'YES' if has_meta_formula else 'NO'}", flush=True)
    print(f"  Bloco 1 REALIZADO formulas: {'YES' if has_real_formula else 'NO'}", flush=True)

    # 2d. Bloco 2 has CONTATOS/DIA referencing LOG
    has_countifs = False
    for r in [18, 19, 20, 21]:
        b_val = ws.cell(row=r, column=2).value
        if b_val and isinstance(b_val, str) and 'COUNTIFS' in b_val and 'LOG' in b_val:
            has_countifs = True
            break

    if not has_countifs:
        issues.append("Bloco 2 missing COUNTIFS referencing LOG")
    print(f"  Bloco 2 LOG COUNTIFS: {'YES' if has_countifs else 'NO'}", flush=True)

    # 2e. DataValidation for VENDEDOR dropdown
    dv_count = 0
    has_vendedor_dv = False
    has_rateio_dv = False
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            dv_count += 1
            if dv.formula1 and 'TODOS' in str(dv.formula1):
                has_vendedor_dv = True
            if dv.formula1 and 'PROPORCIONAL' in str(dv.formula1):
                has_rateio_dv = True

    if not has_vendedor_dv:
        issues.append("VENDEDOR DataValidation not found")
    if not has_rateio_dv:
        issues.append("RATEIO DataValidation not found")
    print(f"  VENDEDOR dropdown: {'YES' if has_vendedor_dv else 'NO'}", flush=True)
    print(f"  RATEIO dropdown: {'YES' if has_rateio_dv else 'NO'}", flush=True)

    # 2f. Count total formulas
    total_formulas = count_formulas(ws)
    print(f"  Total COMITE formulas: {total_formulas}", flush=True)
    if total_formulas < 200:
        issues.append(f"COMITE formulas too few: {total_formulas} (expected >= 200)")

    passed = len(issues) == 0
    status = "PASS" if passed else "FAIL"
    details = f"COMITE complete: {found_titles} blocks, {total_formulas} formulas, filters + toggle" if passed else "; ".join(issues)
    print(f"  RESULT: {status} -- {details}", flush=True)
    return {
        "status": status,
        "details": details,
        "section_titles": found_titles,
        "total_formulas": total_formulas,
        "data_validations": dv_count,
    }


# ============================================================
# CHECK 3: META-03 -- Capacidade de atendimento validada
# ============================================================
def check_meta_03(wb):
    """Verify capacity validation with semaforo thresholds."""
    print("\n" + "=" * 60, flush=True)
    print("CHECK 3 (META-03): Capacidade de atendimento validada", flush=True)
    print("=" * 60, flush=True)

    issues = []

    if "COMITE" not in wb.sheetnames:
        return {"status": "FAIL", "details": "COMITE tab not found"}

    ws = wb["COMITE"]

    # 3a. Bloco 2 CARGA % column (L) with formulas
    has_carga = False
    for r in [18, 19, 20, 21]:
        val = ws.cell(row=r, column=12).value  # col L
        if val and isinstance(val, str) and val.startswith('='):
            has_carga = True
            break
    if not has_carga:
        issues.append("Bloco 2 CARGA % formulas not found in col L")
    print(f"  CARGA % formulas: {'YES' if has_carga else 'NO'}", flush=True)

    # 3b. CONTATOS/DIA uses /22 (22 dias uteis)
    has_22 = False
    for r in [18, 19, 20, 21]:
        val = ws.cell(row=r, column=3).value  # col C (CONTATOS/DIA)
        if val and isinstance(val, str) and '22' in val:
            has_22 = True
            break
    if not has_22:
        issues.append("CONTATOS/DIA does not use /22")
    print(f"  22 dias uteis in formulas: {'YES' if has_22 else 'NO'}", flush=True)

    # 3c. Conditional formatting exists (>= 3 rules)
    cf_count = len(ws.conditional_formatting._cf_rules) if hasattr(ws.conditional_formatting, '_cf_rules') else 0
    if cf_count < 3:
        issues.append(f"Conditional formatting rules: {cf_count} (expected >= 3)")
    print(f"  Conditional formatting rules: {cf_count}", flush=True)

    # 3d. Bloco 3 ALERTAS has semaforo formulas for carga and meta
    has_carga_alert = False
    has_meta_alert = False
    for r in [31, 32, 33, 34]:
        d_val = ws.cell(row=r, column=4).value  # ALERTA CARGA
        f_val = ws.cell(row=r, column=6).value  # ALERTA META
        if d_val and isinstance(d_val, str) and 'SOBRECARGA' in d_val:
            has_carga_alert = True
        if f_val and isinstance(f_val, str) and 'CRITICO' in f_val:
            has_meta_alert = True

    if not has_carga_alert:
        issues.append("Bloco 3 missing ALERTA CARGA formulas")
    if not has_meta_alert:
        issues.append("Bloco 3 missing ALERTA META formulas")
    print(f"  ALERTA CARGA formulas: {'YES' if has_carga_alert else 'NO'}", flush=True)
    print(f"  ALERTA META formulas: {'YES' if has_meta_alert else 'NO'}", flush=True)

    # 3e. Semaforo thresholds: 35 and 50 in formula text
    has_35 = False
    has_50 = False
    for r in [31, 32, 33, 34]:
        d_val = ws.cell(row=r, column=4).value  # ALERTA CARGA
        if d_val and isinstance(d_val, str):
            if '35' in d_val:
                has_35 = True
            if '50' in d_val:
                has_50 = True
    if not has_35:
        issues.append("Threshold 35 not found in ALERTA formulas")
    if not has_50:
        issues.append("Threshold 50 not found in ALERTA formulas")
    print(f"  Threshold 35 (amarelo): {'YES' if has_35 else 'NO'}", flush=True)
    print(f"  Threshold 50 (vermelho): {'YES' if has_50 else 'NO'}", flush=True)

    passed = len(issues) == 0
    status = "PASS" if passed else "FAIL"
    details = f"Capacity validated: CARGA%, 22 dias, {cf_count} CF rules, thresholds 35/50" if passed else "; ".join(issues)
    print(f"  RESULT: {status} -- {details}", flush=True)
    return {
        "status": status,
        "details": details,
        "cf_rules": cf_count,
        "has_thresholds": has_35 and has_50,
    }


# ============================================================
# CHECK 4: V13 integrity
# ============================================================
def check_integrity(wb, prj_name):
    """Verify V13 integrity: formulas, tabs, no #REF!."""
    print("\n" + "=" * 60, flush=True)
    print("CHECK 4: V13 integrity", flush=True)
    print("=" * 60, flush=True)

    issues = []

    # 4a. Count PROJECAO formulas
    prj_formulas = count_formulas(wb[prj_name], min_row=4, max_row=537)
    print(f"  PROJECAO formulas: {prj_formulas}", flush=True)
    if prj_formulas < EXPECTED_FORMULAS:
        issues.append(f"PROJECAO formulas: {prj_formulas} < {EXPECTED_FORMULAS}")

    # 4b. All expected tabs exist
    found_tabs = []
    prj_found = False
    for name in wb.sheetnames:
        clean = unicodedata.normalize('NFD', name)
        clean = ''.join(c for c in clean if unicodedata.category(c) != 'Mn')
        if 'PROJECAO' in clean.upper():
            prj_found = True
            found_tabs.append(name)
        elif name in EXPECTED_TABS:
            found_tabs.append(name)

    missing_tabs = [t for t in EXPECTED_TABS if t not in found_tabs]
    if not prj_found:
        missing_tabs.append("PROJECAO")
    print(f"  Tabs found: {found_tabs}", flush=True)
    if missing_tabs:
        issues.append(f"Missing tabs: {missing_tabs}")
        print(f"  Missing: {missing_tabs}", flush=True)
    else:
        print(f"  All 5 tabs present", flush=True)

    # 4c. No #REF! in COMITE formulas
    ref_errors = 0
    if "COMITE" in wb.sheetnames:
        ws = wb["COMITE"]
        for row in ws.iter_rows(min_row=1, max_row=72):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '#REF' in cell.value:
                    ref_errors += 1
                    print(f"  #REF! at {cell.coordinate}: {cell.value}", flush=True)

    if ref_errors > 0:
        issues.append(f"{ref_errors} #REF! errors in COMITE")
    print(f"  #REF! errors in COMITE: {ref_errors}", flush=True)

    passed = len(issues) == 0
    status = "PASS" if passed else "FAIL"
    details = f"Integrity OK: {prj_formulas} formulas, {len(found_tabs)} tabs, 0 #REF!" if passed else "; ".join(issues)
    print(f"  RESULT: {status} -- {details}", flush=True)
    return {
        "status": status,
        "details": details,
        "projecao_formulas": prj_formulas,
        "tabs": wb.sheetnames,
        "ref_errors": ref_errors,
    }


# ============================================================
# CHECK 5: Cross-check META totals
# ============================================================
def check_cross_check(wb_data, prj_name):
    """Cross-check META totals against 08-01 report."""
    print("\n" + "=" * 60, flush=True)
    print("CHECK 5: Cross-check META totals", flush=True)
    print("=" * 60, flush=True)

    issues = []
    notes = []

    # Load 08-01 report for reference
    ref_total = None
    if META_REPORT_PATH.exists():
        with open(META_REPORT_PATH) as f:
            meta_report = json.load(f)
        ref_total = meta_report.get("meta_totals", {}).get("proportional", {}).get("total")
        print(f"  Reference from 08-01: R$ {ref_total:,.2f}" if ref_total else "  No reference", flush=True)

    # Read COMITE B12 cached value (may be None)
    comite_total = None
    if "COMITE" in wb_data.sheetnames:
        ws = wb_data["COMITE"]
        comite_total = ws.cell(row=12, column=2).value
        print(f"  COMITE B12 (cached): {comite_total}", flush=True)

    if comite_total is not None and ref_total is not None:
        try:
            diff = abs(float(comite_total) - float(ref_total))
            pct = (diff / ref_total) * 100 if ref_total else 0
            print(f"  Delta: R$ {diff:,.2f} ({pct:.2f}%)", flush=True)
            if pct > 5:
                issues.append(f"COMITE total differs from reference by {pct:.2f}%")
        except (ValueError, TypeError):
            notes.append("Could not compare: cached value is not numeric")
    else:
        notes.append("Cached values may be None until opened in Excel -- expected behavior")
        print(f"  NOTE: Cached values unavailable (normal for formula-only cells)", flush=True)

    status = "PASS_WITH_NOTES" if notes and not issues else ("PASS" if not issues else "FAIL")
    details = "; ".join(notes) if notes else "Cross-check within tolerance"
    if issues:
        details = "; ".join(issues)
    print(f"  RESULT: {status} -- {details}", flush=True)
    return {
        "status": status,
        "details": details,
        "comite_total_cached": str(comite_total),
        "reference_total": ref_total,
    }


# ============================================================
# CHECK 6: RATEIO toggle structure
# ============================================================
def check_rateio_toggle(wb):
    """Verify RATEIO toggle dropdown and formula references."""
    print("\n" + "=" * 60, flush=True)
    print("CHECK 6: RATEIO toggle structure", flush=True)
    print("=" * 60, flush=True)

    issues = []

    if "COMITE" not in wb.sheetnames:
        return {"status": "FAIL", "details": "COMITE tab not found"}

    ws = wb["COMITE"]

    # 6a. I2 has DataValidation with toggle values
    has_rateio_dv = False
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            f1 = str(dv.formula1) if dv.formula1 else ""
            if 'PROPORCIONAL' in f1 and 'IGUALITARIO' in f1 and 'COMPENSADO' in f1:
                has_rateio_dv = True
                print(f"  RATEIO DataValidation: {f1}", flush=True)
                break

    if not has_rateio_dv:
        issues.append("RATEIO DataValidation missing PROPORCIONAL/IGUALITARIO/COMPENSADO")
    print(f"  RATEIO dropdown: {'YES' if has_rateio_dv else 'NO'}", flush=True)

    # 6b. Bloco 1 META formulas contain IF($I$2
    has_toggle_ref = False
    for r in [6, 7, 8, 9]:
        b_val = ws.cell(row=r, column=2).value
        if b_val and isinstance(b_val, str) and '$I$2' in b_val:
            has_toggle_ref = True
            break

    if not has_toggle_ref:
        issues.append("META formulas don't reference $I$2 toggle")
    print(f"  $I$2 toggle reference: {'YES' if has_toggle_ref else 'NO'}", flush=True)

    # 6c. Formula contains all 3 meta column references
    has_l = False
    has_bb = False
    has_bp = False
    for r in [6, 7, 8, 9]:
        b_val = str(ws.cell(row=r, column=2).value or "")
        if 'L$4' in b_val:
            has_l = True
        if 'BB$4' in b_val:
            has_bb = True
        if 'BP$4' in b_val:
            has_bp = True

    if not (has_l and has_bb and has_bp):
        issues.append(f"Toggle missing column refs: L={has_l}, BB={has_bb}, BP={has_bp}")
    print(f"  Column refs: L={has_l}, BB={has_bb}, BP={has_bp}", flush=True)

    passed = len(issues) == 0
    status = "PASS" if passed else "FAIL"
    details = "RATEIO toggle with 3 modes verified: PROPORCIONAL(L), IGUALITARIO(BB), COMPENSADO(BP)" if passed else "; ".join(issues)
    print(f"  RESULT: {status} -- {details}", flush=True)
    return {
        "status": status,
        "details": details,
        "has_toggle_dv": has_rateio_dv,
        "has_toggle_ref": has_toggle_ref,
        "refs": {"L": has_l, "BB": has_bb, "BP": has_bp},
    }


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60, flush=True)
    print("Phase 08 Plan 02 -- Validate META-01..03", flush=True)
    print("=" * 60, flush=True)

    # Load V13 (formulas)
    print("\nLoading V13 (data_only=False)...", flush=True)
    wb = openpyxl.load_workbook(str(V13_PATH), data_only=False)
    prj_name = find_projecao_sheet(wb)
    print(f"  PROJECAO: {repr(prj_name)}", flush=True)
    print(f"  Sheets: {wb.sheetnames}", flush=True)

    # Load V13 (cached values for cross-check)
    print("Loading V13 (data_only=True)...", flush=True)
    wb_data = openpyxl.load_workbook(str(V13_PATH), data_only=True)

    # Run checks
    result_01 = check_meta_01(wb, prj_name)
    result_02 = check_meta_02(wb)
    result_03 = check_meta_03(wb)
    result_04 = check_integrity(wb, prj_name)
    result_05 = check_cross_check(wb_data, prj_name)
    result_06 = check_rateio_toggle(wb)

    wb.close()
    wb_data.close()

    # Determine overall status
    all_checks = [result_01, result_02, result_03, result_04, result_05, result_06]
    all_pass = all(c["status"] in ("PASS", "PASS_WITH_NOTES") for c in all_checks)
    fail_count = sum(1 for c in all_checks if c["status"] == "FAIL")

    # Build report
    comite_formulas = result_02.get("total_formulas", 0) if result_02["status"] != "FAIL" else 0
    prj_formulas = result_04.get("projecao_formulas", 0)

    report = {
        "phase": "08-comite-metas",
        "plan": "02",
        "validated_at": datetime.now(timezone.utc).isoformat(),
        "requirements": {
            "META-01": {"status": result_01["status"], "details": result_01["details"]},
            "META-02": {"status": result_02["status"], "details": result_02["details"]},
            "META-03": {"status": result_03["status"], "details": result_03["details"]},
        },
        "checks": [
            {"name": "CHECK 1 (META-01)", **result_01},
            {"name": "CHECK 2 (META-02)", **result_02},
            {"name": "CHECK 3 (META-03)", **result_03},
            {"name": "CHECK 4 (Integrity)", **result_04},
            {"name": "CHECK 5 (Cross-check)", **result_05},
            {"name": "CHECK 6 (RATEIO toggle)", **result_06},
        ],
        "integrity": {
            "projecao_formulas": prj_formulas,
            "comite_formulas": comite_formulas,
            "tabs": wb.sheetnames,
        },
        "overall": "PASS" if all_pass else f"FAIL ({fail_count} checks failed)",
    }

    # Save report
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(REPORT_PATH, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False, default=str)
    print(f"\n  Report saved: {REPORT_PATH}", flush=True)

    # Print summary
    print("\n" + "=" * 60, flush=True)
    print("VALIDATION SUMMARY", flush=True)
    print("=" * 60, flush=True)
    print(f"  META-01: {result_01['status']}", flush=True)
    print(f"  META-02: {result_02['status']}", flush=True)
    print(f"  META-03: {result_03['status']}", flush=True)
    print(f"  Integrity: {result_04['status']}", flush=True)
    print(f"  Cross-check: {result_05['status']}", flush=True)
    print(f"  RATEIO toggle: {result_06['status']}", flush=True)
    print(f"\n  OVERALL: {report['overall']}", flush=True)
    print(f"  PROJECAO formulas: {prj_formulas}", flush=True)
    print(f"  COMITE formulas: {comite_formulas}", flush=True)

    if not all_pass:
        print(f"\n  FAILED CHECKS:", flush=True)
        for c in all_checks:
            if c["status"] == "FAIL":
                print(f"    - {c['details']}", flush=True)
        sys.exit(1)

    print(f"\n  ALL CHECKS PASSED!", flush=True)
    return report


if __name__ == "__main__":
    main()

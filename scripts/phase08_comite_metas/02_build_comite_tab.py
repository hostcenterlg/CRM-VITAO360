"""
Phase 08 Plan 02 -- Build COMITE tab in V13 with 5 blocks (~75 rows)
=====================================================================
Creates the COMITE COMERCIAL tab for weekly team meetings (gestor Leandro).

5 blocks:
  1. META vs REALIZADO POR CONSULTOR (rows 4-14)
  2. CAPACIDADE E PRODUTIVIDADE (rows 16-26)
  3. ALERTAS E RISCOS + TOP 5 GAP (rows 29-42)
  4. FUNIL CONSOLIDADO: TIPO x RESULTADO (rows 45-56)
  5. MOTIVOS DE NAO COMPRA + DONO DA ACAO (rows 59-71)

Header: Row 1 title, Row 2 filters (VENDEDOR, PERIODO, RATEIO toggle)

Formula references:
  PROJECAO: cols L,Z,BB,BP (metas), D (consultor), AP (GAP), AQ (RANKING), B (client name)
  LOG: A=DATA, B=CONSULTOR, L=TIPO, M=RESULTADO, N=MOTIVO

Uses pyenv Python: /c/Users/User/.pyenv/pyenv-win/pyenv-win/versions/3.12.10/python.exe
"""

import sys
import os
import shutil
import unicodedata
import datetime
from pathlib import Path
from collections import Counter

# Add scripts/ to path for v3_styles
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from v3_styles import *

import openpyxl
from openpyxl.formatting.rule import DataBarRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===================================================================
# CONFIGURATION
# ===================================================================
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
V13_PATH = PROJECT_ROOT / "data" / "output" / "CRM_VITAO360_V13_PROJECAO.xlsx"
BACKUP_PATH = PROJECT_ROOT / "data" / "output" / "V13_BACKUP_PHASE08.xlsx"

# Consultants (canonical names from meta_validation_report.json)
CONSULTANTS = [
    "HEMANUELE DITZEL (MANU)",
    "LARISSA PADILHA",
    "JULIO GADRET",
    "DAIANE STAVICKI",
]

# TIPO DO CONTATO values (from Phase 5 normalization, sem-acento)
TIPOS = [
    "ATEND. CLIENTES ATIVOS",
    "ATEND. CLIENTES INATIVOS",
    "NEGOCIACAO",
    "POS-VENDA",
    "PROSPECCAO",
    "FOLLOW UP",
    "PERDA",
]

# RESULTADO values (exact strings from LOG col M)
RESULTADO_MAP = {
    "ORCAMENTO": "ORCAMENTO",
    "CADASTRO": "CADASTRO",
    "RELACIONAMENTO": "RELACIONAMENTO",
    "EM ATENDIMENTO": "EM ATENDIMENTO",
    "SUPORTE": "SUPORTE",
    "VENDA": "VENDA*",          # wildcard to catch VENDA and VENDA / PEDIDO
    "NAO ATENDE": "NAO ATENDE",
    "RECUSOU": "RECUSOU*",      # RECUSOU LIGACAO, RECUSOU ATENDIMENTO
    "NAO RESPONDEU": "NAO RESPONDE",
    "PERDA": "PERDA*",          # PERDA / FECHOU LOJA
    "FOLLOW UP": "FOLLOW UP*",  # FOLLOW UP 7, FOLLOW UP 15
}

# Tab color (dark blue)
TAB_COMITE = "1F4E79"


# ===================================================================
# HELPERS
# ===================================================================
def find_projecao_sheet(wb):
    """Find PROJECAO sheet by accent-stripping."""
    for name in wb.sheetnames:
        clean = unicodedata.normalize('NFD', name)
        clean = ''.join(c for c in clean if unicodedata.category(c) != 'Mn')
        if 'PROJECAO' in clean.upper():
            return name
    raise ValueError("PROJECAO sheet not found")


def section_title(ws, row, label, end_col=17):
    """Write a section title bar (gray background, bold text)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=1, value=f"  {label}")
    style_cell(cell, font=FONT_TITLE, fill=FILL_GRAY_D9, align=ALIGN_LEFT)
    for c in range(1, end_col + 1):
        ws.cell(row=row, column=c).fill = FILL_GRAY_D9
        ws.cell(row=row, column=c).border = THIN_BORDER


def cf(extra=""):
    """Generate COUNTIFS formula with optional VENDEDOR filter + date range.
    Uses C2=VENDEDOR dropdown, E2=start date, F2=end date.
    Same pattern as Phase 5 DASH cf() helper."""
    base_date = 'LOG!$A$3:$A$21000,">="&$E$2,LOG!$A$3:$A$21000,"<="&$F$2'
    if extra:
        no_filter = f'COUNTIFS({base_date},{extra})'
        with_filter = f'COUNTIFS({base_date},LOG!$B$3:$B$21000,$C$2,{extra})'
    else:
        no_filter = f'COUNTIFS({base_date})'
        with_filter = f'COUNTIFS({base_date},LOG!$B$3:$B$21000,$C$2)'
    return f'IF(OR($C$2="",$C$2="TODOS"),{no_filter},{with_filter})'


def count_formulas(ws):
    """Count formulas in a worksheet."""
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                count += 1
    return count


def extract_top_motivos(ws_log):
    """Extract top 10 motivos from LOG column N (data_only workbook)."""
    motivo_counter = Counter()
    for r in range(3, 21001):
        val = ws_log.cell(row=r, column=14).value  # col N = 14
        if val and str(val).strip():
            motivo_counter[str(val).strip()] += 1
    top_10 = [m for m, _ in motivo_counter.most_common(10)]
    return top_10


# ===================================================================
# BUILD COMITE TAB
# ===================================================================
def build_comite(wb, prj_name, motivos):
    """Build the complete COMITE tab with 5 blocks."""
    print("Building COMITE tab...", flush=True)

    # Delete existing COMITE if present (re-runnable)
    if "COMITE" in wb.sheetnames:
        del wb["COMITE"]
        print("  Removed existing COMITE tab", flush=True)

    ws = wb.create_sheet("COMITE")
    ws.sheet_properties.tabColor = TAB_COMITE

    # Formula reference for PROJECAO sheet (with quotes for special chars)
    prj_ref = f"'{prj_name}'!"

    # ===============================================================
    # COLUMN WIDTHS
    # ===============================================================
    ws.column_dimensions['A'].width = 30
    for c in range(2, 18):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # ===============================================================
    # ROW 1: TITLE
    # ===============================================================
    ws.merge_cells('A1:Q1')
    cell = ws.cell(row=1, column=1, value="COMITE COMERCIAL -- VITAO ALIMENTOS 360")
    style_cell(cell,
               font=Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
               fill=PatternFill('solid', fgColor='1F4E79'),
               align=ALIGN_CENTER)
    for c in range(1, 18):
        ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor='1F4E79')
        ws.cell(row=1, column=c).border = THIN_BORDER

    # ===============================================================
    # ROW 2: FILTERS (VENDEDOR, PERIODO, RATEIO)
    # ===============================================================
    # A2: label
    write_data(ws, 2, 1, "VENDEDOR:", font=FONT_DATA_BOLD, fill=FILL_LIGHT_GRAY, align=ALIGN_LEFT)
    ws.cell(row=2, column=2).fill = FILL_LIGHT_GRAY
    ws.cell(row=2, column=2).border = THIN_BORDER

    # C2: VENDEDOR dropdown
    ws.cell(row=2, column=3, value="TODOS")
    style_cell(ws.cell(row=2, column=3), font=FONT_DATA, fill=FILL_INPUT)
    dv_vendedor = DataValidation(
        type="list",
        formula1='"TODOS,HEMANUELE DITZEL (MANU),LARISSA PADILHA,JULIO GADRET,DAIANE STAVICKI"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_vendedor)
    dv_vendedor.add("C2")

    # D2: PERIODO label
    write_data(ws, 2, 4, "PERIODO:", font=FONT_DATA_BOLD, fill=FILL_LIGHT_GRAY, align=ALIGN_LEFT)

    # E2: start date
    ws.cell(row=2, column=5, value=datetime.date(2026, 2, 1))
    style_cell(ws.cell(row=2, column=5), font=FONT_DATA, fill=FILL_INPUT, fmt=FMT_DATE)

    # F2: end date
    ws.cell(row=2, column=6, value=datetime.date(2026, 2, 28))
    style_cell(ws.cell(row=2, column=6), font=FONT_DATA, fill=FILL_INPUT, fmt=FMT_DATE)

    # G2: spacer
    ws.cell(row=2, column=7).fill = FILL_LIGHT_GRAY
    ws.cell(row=2, column=7).border = THIN_BORDER

    # H2: RATEIO label
    write_data(ws, 2, 8, "RATEIO:", font=FONT_DATA_BOLD, fill=FILL_LIGHT_GRAY, align=ALIGN_LEFT)

    # I2: RATEIO dropdown
    ws.cell(row=2, column=9, value="PROPORCIONAL")
    style_cell(ws.cell(row=2, column=9), font=FONT_DATA, fill=FILL_INPUT)
    dv_rateio = DataValidation(
        type="list",
        formula1='"PROPORCIONAL,IGUALITARIO,COMPENSADO"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_rateio)
    dv_rateio.add("I2")

    # Fill rest of row 2
    for c in range(10, 18):
        ws.cell(row=2, column=c).fill = FILL_LIGHT_GRAY
        ws.cell(row=2, column=c).border = THIN_BORDER

    # ===============================================================
    # BLOCO 1: META vs REALIZADO POR CONSULTOR (Rows 4-14)
    # ===============================================================
    print("  Building Bloco 1: META vs REALIZADO...", flush=True)

    # Row 4: section title
    section_title(ws, 4, "META vs REALIZADO POR CONSULTOR", 12)

    # Row 5: headers
    b1_headers = [
        "CONSULTOR", "META ANUAL (R$)", "REAL ANUAL (R$)", "% ATING",
        "GAP R$", "SEMAFORO", "META MES (R$)", "REAL MES (R$)",
        "% MES", "GAP MES (R$)", "CLIENTES", "RANKING"
    ]
    for i, h in enumerate(b1_headers, 1):
        write_header(ws, 5, i, h, fill=FILL_DARK, font=FONT_HEADER_SM, align=ALIGN_CENTER)

    # Rows 6-9: consultant data
    consultant_rows = {}
    for idx, cons in enumerate(CONSULTANTS):
        r = 6 + idx
        consultant_rows[cons] = r

        # A: Consultant name (static)
        write_data(ws, r, 1, cons, align=ALIGN_LEFT)

        # B: META ANUAL with RATEIO toggle
        # IF($I$2="IGUALITARIO", SUMIFS(BB), IF($I$2="COMPENSADO", SUMIFS(BP), SUMIFS(L)))
        f_meta = (
            f'=IF($I$2="IGUALITARIO",'
            f'SUMIFS({prj_ref}BB$4:BB$537,{prj_ref}D$4:D$537,A{r}),'
            f'IF($I$2="COMPENSADO",'
            f'SUMIFS({prj_ref}BP$4:BP$537,{prj_ref}D$4:D$537,A{r}),'
            f'SUMIFS({prj_ref}L$4:L$537,{prj_ref}D$4:D$537,A{r})))'
        )
        ws.cell(row=r, column=2).value = f_meta
        ws.cell(row=r, column=2).number_format = '#,##0'
        style_cell(ws.cell(row=r, column=2), fmt='#,##0')

        # C: REAL ANUAL
        f_real = f'=SUMIFS({prj_ref}Z$4:Z$537,{prj_ref}D$4:D$537,A{r})'
        ws.cell(row=r, column=3).value = f_real
        ws.cell(row=r, column=3).number_format = '#,##0'
        style_cell(ws.cell(row=r, column=3), fmt='#,##0')

        # D: % ATING
        ws.cell(row=r, column=4).value = f'=IFERROR(C{r}/B{r},0)'
        ws.cell(row=r, column=4).number_format = '0%'
        style_cell(ws.cell(row=r, column=4), fmt='0%')

        # E: GAP R$
        ws.cell(row=r, column=5).value = f'=B{r}-C{r}'
        ws.cell(row=r, column=5).number_format = '#,##0'
        style_cell(ws.cell(row=r, column=5), fmt='#,##0')

        # F: SEMAFORO
        ws.cell(row=r, column=6).value = f'=IF(D{r}>=1,"VERDE",IF(D{r}>=0.7,"AMARELO","VERMELHO"))'
        style_cell(ws.cell(row=r, column=6))

        # G: META MES (annual / 12 approximation)
        ws.cell(row=r, column=7).value = f'=IFERROR(B{r}/12,0)'
        ws.cell(row=r, column=7).number_format = '#,##0'
        style_cell(ws.cell(row=r, column=7), fmt='#,##0')

        # H: REAL MES (annual / 12 approximation)
        ws.cell(row=r, column=8).value = f'=IFERROR(C{r}/12,0)'
        ws.cell(row=r, column=8).number_format = '#,##0'
        style_cell(ws.cell(row=r, column=8), fmt='#,##0')

        # I: % MES
        ws.cell(row=r, column=9).value = f'=IFERROR(H{r}/G{r},0)'
        ws.cell(row=r, column=9).number_format = '0%'
        style_cell(ws.cell(row=r, column=9), fmt='0%')

        # J: GAP MES
        ws.cell(row=r, column=10).value = f'=G{r}-H{r}'
        ws.cell(row=r, column=10).number_format = '#,##0'
        style_cell(ws.cell(row=r, column=10), fmt='#,##0')

        # K: CLIENTES
        ws.cell(row=r, column=11).value = f'=COUNTIFS({prj_ref}D$4:D$537,A{r})'
        style_cell(ws.cell(row=r, column=11))

        # L: RANKING
        ws.cell(row=r, column=12).value = f'=IFERROR(RANK(C{r},$C$6:$C$9,0),"-")'
        style_cell(ws.cell(row=r, column=12))

    # Row 10: OUTROS/SEM CONSULTOR (total - sum of named)
    r = 10
    write_data(ws, r, 1, "OUTROS/SEM CONSULTOR", align=ALIGN_LEFT)

    # B10: META ANUAL for OUTROS = B12 - SUM(B6:B9)
    ws.cell(row=10, column=2).value = '=B12-SUM(B6:B9)'
    ws.cell(row=10, column=2).number_format = '#,##0'
    style_cell(ws.cell(row=10, column=2), fmt='#,##0')

    # C10: REAL for OUTROS
    ws.cell(row=10, column=3).value = '=C12-SUM(C6:C9)'
    ws.cell(row=10, column=3).number_format = '#,##0'
    style_cell(ws.cell(row=10, column=3), fmt='#,##0')

    # D10: % ATING
    ws.cell(row=10, column=4).value = '=IFERROR(C10/B10,0)'
    ws.cell(row=10, column=4).number_format = '0%'
    style_cell(ws.cell(row=10, column=4), fmt='0%')

    # E10: GAP
    ws.cell(row=10, column=5).value = '=B10-C10'
    ws.cell(row=10, column=5).number_format = '#,##0'
    style_cell(ws.cell(row=10, column=5), fmt='#,##0')

    # F10: SEMAFORO
    ws.cell(row=10, column=6).value = '=IF(D10>=1,"VERDE",IF(D10>=0.7,"AMARELO","VERMELHO"))'
    style_cell(ws.cell(row=10, column=6))

    # G10: META MES
    ws.cell(row=10, column=7).value = '=IFERROR(B10/12,0)'
    ws.cell(row=10, column=7).number_format = '#,##0'
    style_cell(ws.cell(row=10, column=7), fmt='#,##0')

    # H10: REAL MES
    ws.cell(row=10, column=8).value = '=IFERROR(C10/12,0)'
    ws.cell(row=10, column=8).number_format = '#,##0'
    style_cell(ws.cell(row=10, column=8), fmt='#,##0')

    # I10: % MES
    ws.cell(row=10, column=9).value = '=IFERROR(H10/G10,0)'
    ws.cell(row=10, column=9).number_format = '0%'
    style_cell(ws.cell(row=10, column=9), fmt='0%')

    # J10: GAP MES
    ws.cell(row=10, column=10).value = '=G10-H10'
    ws.cell(row=10, column=10).number_format = '#,##0'
    style_cell(ws.cell(row=10, column=10), fmt='#,##0')

    # K10: CLIENTES
    ws.cell(row=10, column=11).value = '=K12-SUM(K6:K9)'
    style_cell(ws.cell(row=10, column=11))

    # L10: RANKING (no rank for orphans)
    ws.cell(row=10, column=12).value = '-'
    style_cell(ws.cell(row=10, column=12))

    # Row 11: empty spacer (skip)

    # Row 12: TOTAL EQUIPE
    r = 12
    write_header(ws, r, 1, "TOTAL EQUIPE", fill=FILL_DARK, font=FONT_HEADER, align=ALIGN_LEFT)

    # B12: META total with toggle (SUM of all clients via toggle)
    f_total_meta = (
        f'=IF($I$2="IGUALITARIO",'
        f'SUMIFS({prj_ref}BB$4:BB$537,{prj_ref}D$4:D$537,"<>"),'
        f'IF($I$2="COMPENSADO",'
        f'SUMIFS({prj_ref}BP$4:BP$537,{prj_ref}D$4:D$537,"<>"),'
        f'SUMIFS({prj_ref}L$4:L$537,{prj_ref}D$4:D$537,"<>")))'
    )
    ws.cell(row=r, column=2).value = f_total_meta
    ws.cell(row=r, column=2).number_format = '#,##0'
    write_header(ws, r, 2, ws.cell(row=r, column=2).value, fill=FILL_DARK)
    ws.cell(row=r, column=2).number_format = '#,##0'

    # C12: REAL total
    f_total_real = f'=SUMIFS({prj_ref}Z$4:Z$537,{prj_ref}D$4:D$537,"<>")'
    ws.cell(row=r, column=3).value = f_total_real
    ws.cell(row=r, column=3).number_format = '#,##0'
    write_header(ws, r, 3, ws.cell(row=r, column=3).value, fill=FILL_DARK)
    ws.cell(row=r, column=3).number_format = '#,##0'

    # D12: % ATING total
    ws.cell(row=r, column=4).value = f'=IFERROR(C{r}/B{r},0)'
    ws.cell(row=r, column=4).number_format = '0%'
    write_header(ws, r, 4, ws.cell(row=r, column=4).value, fill=FILL_DARK)
    ws.cell(row=r, column=4).number_format = '0%'

    # E12: GAP total
    ws.cell(row=r, column=5).value = f'=B{r}-C{r}'
    ws.cell(row=r, column=5).number_format = '#,##0'
    write_header(ws, r, 5, ws.cell(row=r, column=5).value, fill=FILL_DARK)
    ws.cell(row=r, column=5).number_format = '#,##0'

    # F12: SEMAFORO total
    ws.cell(row=r, column=6).value = f'=IF(D{r}>=1,"VERDE",IF(D{r}>=0.7,"AMARELO","VERMELHO"))'
    write_header(ws, r, 6, ws.cell(row=r, column=6).value, fill=FILL_DARK)

    # G12-J12: monthly (SUM approach)
    for col in [7, 8, 10]:
        cl = get_column_letter(col)
        ws.cell(row=r, column=col).value = f'=SUM({cl}6:{cl}10)'
        ws.cell(row=r, column=col).number_format = '#,##0'
        write_header(ws, r, col, ws.cell(row=r, column=col).value, fill=FILL_DARK)
        ws.cell(row=r, column=col).number_format = '#,##0'

    # I12: % MES total
    ws.cell(row=r, column=9).value = f'=IFERROR(H{r}/G{r},0)'
    ws.cell(row=r, column=9).number_format = '0%'
    write_header(ws, r, 9, ws.cell(row=r, column=9).value, fill=FILL_DARK)
    ws.cell(row=r, column=9).number_format = '0%'

    # K12: CLIENTES total
    ws.cell(row=r, column=11).value = '=SUM(K6:K10)'
    write_header(ws, r, 11, ws.cell(row=r, column=11).value, fill=FILL_DARK)

    # L12: blank (no ranking for total)
    write_header(ws, r, 12, "", fill=FILL_DARK)

    # Row 13: META SAP OFICIAL (reference)
    ws.cell(row=13, column=1).value = "META SAP OFICIAL"
    style_cell(ws.cell(row=13, column=1), font=Font(name='Calibri', size=10, italic=True, color='808080'), align=ALIGN_LEFT)
    ws.cell(row=13, column=2).value = 4747200
    ws.cell(row=13, column=2).number_format = '#,##0'
    style_cell(ws.cell(row=13, column=2), font=Font(name='Calibri', size=10, italic=True, color='808080'))

    # ---- Conditional formatting for Bloco 1 ----
    # D6:D12 (% ATING): 3-color semaforo
    green_fill = PatternFill(bgColor='C6EFCE')
    yellow_fill = PatternFill(bgColor='FFEB9C')
    red_fill = PatternFill(bgColor='FFC7CE')

    ws.conditional_formatting.add('D6:D12',
        FormulaRule(formula=['D6>=1'], fill=green_fill))
    ws.conditional_formatting.add('D6:D12',
        FormulaRule(formula=['AND(D6>=0.7,D6<1)'], fill=yellow_fill))
    ws.conditional_formatting.add('D6:D12',
        FormulaRule(formula=['D6<0.7'], fill=red_fill))

    # E6:E12 (GAP R$): DataBar red
    ws.conditional_formatting.add('E6:E12',
        DataBarRule(start_type='min', end_type='max', color='FF6347', showValue=True))

    # F6:F10 (SEMAFORO text): CellIs color
    ws.conditional_formatting.add('F6:F10',
        CellIsRule(operator='equal', formula=['"VERDE"'], fill=PatternFill(bgColor='C6EFCE')))
    ws.conditional_formatting.add('F6:F10',
        CellIsRule(operator='equal', formula=['"AMARELO"'], fill=PatternFill(bgColor='FFEB9C')))
    ws.conditional_formatting.add('F6:F10',
        CellIsRule(operator='equal', formula=['"VERMELHO"'], fill=PatternFill(bgColor='FFC7CE')))

    print("  Bloco 1 complete: rows 4-13", flush=True)

    # ===============================================================
    # BLOCO 2: CAPACIDADE E PRODUTIVIDADE (Rows 16-26)
    # ===============================================================
    print("  Building Bloco 2: CAPACIDADE E PRODUTIVIDADE...", flush=True)

    # Row 16: section title
    section_title(ws, 16, "CAPACIDADE E PRODUTIVIDADE POR CONSULTOR", 12)

    # Row 17: headers
    b2_headers = [
        "CONSULTOR", "TOTAL CONTATOS", "CONTATOS/DIA", "VENDAS",
        "VENDAS/DIA", "% CONVERSAO", "ORCAMENTOS", "CADASTROS",
        "FOLLOW-UPS", "PROSPECCOES", "SUPORTE", "CARGA %"
    ]
    for i, h in enumerate(b2_headers, 1):
        write_header(ws, 17, i, h, fill=FILL_DARK, font=FONT_HEADER_SM, align=ALIGN_CENTER)

    # Date filter helpers for LOG formulas
    date_lo = 'LOG!$A$3:$A$21000,">="&$E$2'
    date_hi = 'LOG!$A$3:$A$21000,"<="&$F$2'

    # Rows 18-21: consultant data
    b2_cons_rows = {}
    for idx, cons in enumerate(CONSULTANTS):
        r = 18 + idx
        b2_cons_rows[cons] = r
        cons_f = f',LOG!$B$3:$B$21000,A{r}'

        # A: name
        write_data(ws, r, 1, cons, align=ALIGN_LEFT)

        # B: TOTAL CONTATOS
        ws.cell(row=r, column=2).value = f'=COUNTIFS({date_lo},{date_hi}{cons_f})'
        style_cell(ws.cell(row=r, column=2))

        # C: CONTATOS/DIA (22 dias uteis)
        ws.cell(row=r, column=3).value = f'=IFERROR(B{r}/22,0)'
        ws.cell(row=r, column=3).number_format = '0.0'
        style_cell(ws.cell(row=r, column=3), fmt='0.0')

        # D: VENDAS
        ws.cell(row=r, column=4).value = (
            f'=COUNTIFS({date_lo},{date_hi}{cons_f},LOG!$M$3:$M$21000,"VENDA*")'
        )
        style_cell(ws.cell(row=r, column=4))

        # E: VENDAS/DIA
        ws.cell(row=r, column=5).value = f'=IFERROR(D{r}/22,0)'
        ws.cell(row=r, column=5).number_format = '0.0'
        style_cell(ws.cell(row=r, column=5), fmt='0.0')

        # F: % CONVERSAO
        ws.cell(row=r, column=6).value = f'=IFERROR(D{r}/B{r},0)'
        ws.cell(row=r, column=6).number_format = '0%'
        style_cell(ws.cell(row=r, column=6), fmt='0%')

        # G: ORCAMENTOS
        ws.cell(row=r, column=7).value = (
            f'=COUNTIFS({date_lo},{date_hi}{cons_f},LOG!$M$3:$M$21000,"ORCAMENTO")'
        )
        style_cell(ws.cell(row=r, column=7))

        # H: CADASTROS
        ws.cell(row=r, column=8).value = (
            f'=COUNTIFS({date_lo},{date_hi}{cons_f},LOG!$M$3:$M$21000,"CADASTRO")'
        )
        style_cell(ws.cell(row=r, column=8))

        # I: FOLLOW-UPS (wildcard for FOLLOW UP 7, FOLLOW UP 15, etc.)
        ws.cell(row=r, column=9).value = (
            f'=COUNTIFS({date_lo},{date_hi}{cons_f},LOG!$M$3:$M$21000,"FOLLOW UP*")'
        )
        style_cell(ws.cell(row=r, column=9))

        # J: PROSPECCOES (TIPO DO CONTATO = PROSPECCAO, col L)
        ws.cell(row=r, column=10).value = (
            f'=COUNTIFS({date_lo},{date_hi}{cons_f},LOG!$L$3:$L$21000,"PROSPECCAO")'
        )
        style_cell(ws.cell(row=r, column=10))

        # K: SUPORTE (TIPO = POS-VENDA, col L)
        ws.cell(row=r, column=11).value = (
            f'=COUNTIFS({date_lo},{date_hi}{cons_f},LOG!$L$3:$L$21000,"POS-VENDA")'
        )
        style_cell(ws.cell(row=r, column=11))

        # L: CARGA % (contatos/dia / 50 capacity limit)
        ws.cell(row=r, column=12).value = f'=IFERROR(C{r}/50,0)'
        ws.cell(row=r, column=12).number_format = '0%'
        style_cell(ws.cell(row=r, column=12), fmt='0%')

    # Row 22: OUTROS/SEM CONSULTOR
    r = 22
    write_data(ws, r, 1, "OUTROS/SEM CONSULTOR", align=ALIGN_LEFT)
    for col in range(2, 13):
        cl = get_column_letter(col)
        ws.cell(row=r, column=col).value = f'={cl}23-SUM({cl}18:{cl}21)'
        if col in [3, 5]:
            ws.cell(row=r, column=col).number_format = '0.0'
            style_cell(ws.cell(row=r, column=col), fmt='0.0')
        elif col in [6, 12]:
            ws.cell(row=r, column=col).number_format = '0%'
            style_cell(ws.cell(row=r, column=col), fmt='0%')
        else:
            style_cell(ws.cell(row=r, column=col))

    # Row 23: TOTAL EQUIPE
    r = 23
    write_header(ws, r, 1, "TOTAL EQUIPE", fill=FILL_DARK, font=FONT_HEADER, align=ALIGN_LEFT)

    # B23: TOTAL CONTATOS (all, date-filtered)
    ws.cell(row=r, column=2).value = f'=COUNTIFS({date_lo},{date_hi})'
    write_header(ws, r, 2, ws.cell(row=r, column=2).value, fill=FILL_DARK)

    # C23: CONTATOS/DIA
    ws.cell(row=r, column=3).value = f'=IFERROR(B{r}/22,0)'
    ws.cell(row=r, column=3).number_format = '0.0'
    write_header(ws, r, 3, ws.cell(row=r, column=3).value, fill=FILL_DARK)
    ws.cell(row=r, column=3).number_format = '0.0'

    # D23: VENDAS total
    ws.cell(row=r, column=4).value = (
        f'=COUNTIFS({date_lo},{date_hi},LOG!$M$3:$M$21000,"VENDA*")'
    )
    write_header(ws, r, 4, ws.cell(row=r, column=4).value, fill=FILL_DARK)

    # E23: VENDAS/DIA
    ws.cell(row=r, column=5).value = f'=IFERROR(D{r}/22,0)'
    ws.cell(row=r, column=5).number_format = '0.0'
    write_header(ws, r, 5, ws.cell(row=r, column=5).value, fill=FILL_DARK)
    ws.cell(row=r, column=5).number_format = '0.0'

    # F23: % CONVERSAO
    ws.cell(row=r, column=6).value = f'=IFERROR(D{r}/B{r},0)'
    ws.cell(row=r, column=6).number_format = '0%'
    write_header(ws, r, 6, ws.cell(row=r, column=6).value, fill=FILL_DARK)
    ws.cell(row=r, column=6).number_format = '0%'

    # G-K: SUM of rows 18-22
    for col in [7, 8, 9, 10, 11]:
        cl = get_column_letter(col)
        ws.cell(row=r, column=col).value = f'=SUM({cl}18:{cl}22)'
        write_header(ws, r, col, ws.cell(row=r, column=col).value, fill=FILL_DARK)

    # L23: CARGA % total
    ws.cell(row=r, column=12).value = f'=IFERROR(C{r}/50,0)'
    ws.cell(row=r, column=12).number_format = '0%'
    write_header(ws, r, 12, ws.cell(row=r, column=12).value, fill=FILL_DARK)
    ws.cell(row=r, column=12).number_format = '0%'

    # ---- Conditional formatting for Bloco 2 ----
    # C18:C23 (CONTATOS/DIA): verde < 35, amarelo 35-50, vermelho > 50
    ws.conditional_formatting.add('C18:C23',
        FormulaRule(formula=['C18<35'], fill=green_fill))
    ws.conditional_formatting.add('C18:C23',
        FormulaRule(formula=['AND(C18>=35,C18<=50)'], fill=yellow_fill))
    ws.conditional_formatting.add('C18:C23',
        FormulaRule(formula=['C18>50'], fill=red_fill))

    # L18:L23 (CARGA %): DataBar blue
    ws.conditional_formatting.add('L18:L23',
        DataBarRule(start_type='min', end_type='max', color='638EC6', showValue=True))

    print("  Bloco 2 complete: rows 16-23", flush=True)

    # ===============================================================
    # BLOCO 3: ALERTAS E RISCOS (Rows 29-42)
    # ===============================================================
    print("  Building Bloco 3: ALERTAS E RISCOS...", flush=True)

    # Row 29: section title
    section_title(ws, 29, "ALERTAS E RISCOS", 8)

    # Row 30: headers
    b3_headers = [
        "CONSULTOR", "STATUS", "CARGA/DIA", "ALERTA CARGA",
        "META %", "ALERTA META", "CLIENTES", "RISCO PRINCIPAL"
    ]
    for i, h in enumerate(b3_headers, 1):
        write_header(ws, 30, i, h, fill=FILL_DARK, font=FONT_HEADER_SM, align=ALIGN_CENTER)

    # Rows 31-34: consultant alerts
    # Map Bloco 3 row -> Bloco 1 row -> Bloco 2 row
    b3_map = {
        31: (6, 18),   # HEMANUELE
        32: (7, 19),   # LARISSA
        33: (8, 20),   # JULIO
        34: (9, 21),   # DAIANE
    }
    for r, (b1r, b2r) in b3_map.items():
        cons_name = CONSULTANTS[r - 31]

        # A: name
        write_data(ws, r, 1, cons_name, align=ALIGN_LEFT)

        # B: STATUS (default ATIVO, gestor can change)
        ws.cell(row=r, column=2).value = "ATIVO"
        style_cell(ws.cell(row=r, column=2))

        # C: CARGA/DIA (reference Bloco 2)
        ws.cell(row=r, column=3).value = f'=C{b2r}'
        ws.cell(row=r, column=3).number_format = '0.0'
        style_cell(ws.cell(row=r, column=3), fmt='0.0')

        # D: ALERTA CARGA
        ws.cell(row=r, column=4).value = f'=IF(C{r}>50,"SOBRECARGA",IF(C{r}>=35,"ATENCAO","OK"))'
        style_cell(ws.cell(row=r, column=4))

        # E: META % (reference Bloco 1 % ATING)
        ws.cell(row=r, column=5).value = f'=D{b1r}'
        ws.cell(row=r, column=5).number_format = '0%'
        style_cell(ws.cell(row=r, column=5), fmt='0%')

        # F: ALERTA META
        ws.cell(row=r, column=6).value = f'=IF(E{r}<0.7,"CRITICO",IF(E{r}<1,"ATENCAO","OK"))'
        style_cell(ws.cell(row=r, column=6))

        # G: CLIENTES (reference Bloco 1)
        ws.cell(row=r, column=7).value = f'=K{b1r}'
        style_cell(ws.cell(row=r, column=7))

        # H: RISCO PRINCIPAL (empty, gestor fills manually)
        ws.cell(row=r, column=8).value = ""
        style_cell(ws.cell(row=r, column=8), fill=FILL_INPUT)

    # ---- Conditional formatting for Bloco 3 (alerts) ----
    # D31:D34 (ALERTA CARGA)
    ws.conditional_formatting.add('D31:D34',
        CellIsRule(operator='equal', formula=['"SOBRECARGA"'], fill=PatternFill(bgColor='FFC7CE')))
    ws.conditional_formatting.add('D31:D34',
        CellIsRule(operator='equal', formula=['"ATENCAO"'], fill=PatternFill(bgColor='FFEB9C')))
    ws.conditional_formatting.add('D31:D34',
        CellIsRule(operator='equal', formula=['"OK"'], fill=PatternFill(bgColor='C6EFCE')))

    # F31:F34 (ALERTA META)
    ws.conditional_formatting.add('F31:F34',
        CellIsRule(operator='equal', formula=['"CRITICO"'], fill=PatternFill(bgColor='FFC7CE')))
    ws.conditional_formatting.add('F31:F34',
        CellIsRule(operator='equal', formula=['"ATENCAO"'], fill=PatternFill(bgColor='FFEB9C')))
    ws.conditional_formatting.add('F31:F34',
        CellIsRule(operator='equal', formula=['"OK"'], fill=PatternFill(bgColor='C6EFCE')))

    # ---- Sub-section: TOP 5 CLIENTS BY GAP (rows 36-42) ----
    # Row 36: sub-title
    ws.merge_cells(start_row=36, start_column=1, end_row=36, end_column=8)
    ws.cell(row=36, column=1, value="  CLIENTES CRITICOS (TOP 5 GAP)")
    style_cell(ws.cell(row=36, column=1),
               font=FONT_TITLE,
               fill=PatternFill('solid', fgColor='FFE0B2'),
               align=ALIGN_LEFT)
    for c in range(1, 9):
        ws.cell(row=36, column=c).fill = PatternFill('solid', fgColor='FFE0B2')
        ws.cell(row=36, column=c).border = THIN_BORDER

    # Row 37: sub-headers
    sub_hdrs = ["CLIENTE", "CONSULTOR", "META (R$)", "REALIZADO (R$)", "GAP (R$)", "RANKING"]
    for i, h in enumerate(sub_hdrs, 1):
        write_header(ws, 37, i, h, fill=FILL_DARK, font=FONT_HEADER_SM, align=ALIGN_CENTER)

    # Rows 38-42: Top 5 via LARGE/INDEX/MATCH
    for rank_idx in range(1, 6):
        r = 37 + rank_idx

        # E: GAP (descending)
        ws.cell(row=r, column=5).value = f'=LARGE({prj_ref}AP$4:AP$537,{rank_idx})'
        ws.cell(row=r, column=5).number_format = '#,##0'
        style_cell(ws.cell(row=r, column=5), fmt='#,##0')

        # A: CLIENTE name
        ws.cell(row=r, column=1).value = (
            f'=IFERROR(INDEX({prj_ref}B$4:B$537,MATCH(E{r},{prj_ref}AP$4:AP$537,0)),"")'
        )
        style_cell(ws.cell(row=r, column=1), align=ALIGN_LEFT)

        # B: CONSULTOR
        ws.cell(row=r, column=2).value = (
            f'=IFERROR(INDEX({prj_ref}D$4:D$537,MATCH(E{r},{prj_ref}AP$4:AP$537,0)),"")'
        )
        style_cell(ws.cell(row=r, column=2))

        # C: META
        ws.cell(row=r, column=3).value = (
            f'=IFERROR(INDEX({prj_ref}L$4:L$537,MATCH(E{r},{prj_ref}AP$4:AP$537,0)),0)'
        )
        ws.cell(row=r, column=3).number_format = '#,##0'
        style_cell(ws.cell(row=r, column=3), fmt='#,##0')

        # D: REALIZADO
        ws.cell(row=r, column=4).value = (
            f'=IFERROR(INDEX({prj_ref}Z$4:Z$537,MATCH(E{r},{prj_ref}AP$4:AP$537,0)),0)'
        )
        ws.cell(row=r, column=4).number_format = '#,##0'
        style_cell(ws.cell(row=r, column=4), fmt='#,##0')

        # F: RANKING
        ws.cell(row=r, column=6).value = (
            f'=IFERROR(INDEX({prj_ref}AQ$4:AQ$537,MATCH(E{r},{prj_ref}AP$4:AP$537,0)),"")'
        )
        style_cell(ws.cell(row=r, column=6))

    # E38:E42: DataBar red for gap magnitude
    ws.conditional_formatting.add('E38:E42',
        DataBarRule(start_type='min', end_type='max', color='FF6347', showValue=True))

    print("  Bloco 3 complete: rows 29-42", flush=True)

    # ===============================================================
    # BLOCO 4: FUNIL CONSOLIDADO (Rows 45-56)
    # ===============================================================
    print("  Building Bloco 4: FUNIL CONSOLIDADO...", flush=True)

    # Row 45: section title
    section_title(ws, 45, "TIPO DO CONTATO x RESULTADO DO CONTATO", 13)

    # Row 46: headers
    b4_headers = [
        "TIPO DO CONTATO", "TOTAL", "ORCAM.", "CADAST.", "RELAC.",
        "EM ATEND.", "SUPORTE", "VENDA", "N ATENDE", "RECUSOU",
        "N RESP.", "PERDA", "FOLLOW UP"
    ]
    for i, h in enumerate(b4_headers, 1):
        if i >= 9:
            fill = FILL_RED_DARK
        elif i == 13:
            fill = FILL_GREEN_DARK
        else:
            fill = FILL_DARK
        write_header(ws, 46, i, h, fill=fill, font=FONT_HEADER_SM, align=ALIGN_CENTER)

    # RESULTADO mapping for columns C-M (3-13)
    resultado_cols = {
        3: "ORCAMENTO",
        4: "CADASTRO",
        5: "RELACIONAMENTO",
        6: "EM ATENDIMENTO",
        7: "SUPORTE",
        8: "VENDA*",             # wildcard for VENDA + VENDA / PEDIDO
        9: "NAO ATENDE",
        10: "RECUSOU*",          # RECUSOU LIGACAO, RECUSOU ATENDIMENTO
        11: "NAO RESPONDE",
        12: "PERDA*",            # PERDA / FECHOU LOJA
        13: "FOLLOW UP*",        # FOLLOW UP 7, FOLLOW UP 15
    }

    # Rows 47-53: one per TIPO
    for t_idx, tipo in enumerate(TIPOS):
        r = 47 + t_idx

        # A: TIPO name
        write_data(ws, r, 1, tipo, align=ALIGN_LEFT)

        # B: TOTAL for this TIPO (all RESULTADO)
        tipo_f = f'LOG!$L$3:$L$21000,A{r}'
        ws.cell(row=r, column=2).value = f'={cf(tipo_f)}'
        style_cell(ws.cell(row=r, column=2), font=FONT_DATA_BOLD)

        # C-M: RESULTADO breakdown
        for col, resultado in resultado_cols.items():
            res_f = f'LOG!$L$3:$L$21000,A{r},LOG!$M$3:$M$21000,"{resultado}"'
            ws.cell(row=r, column=col).value = f'={cf(res_f)}'
            style_cell(ws.cell(row=r, column=col))

    # Row 54: TOTAL
    r = 54
    write_header(ws, r, 1, "TOTAL", fill=FILL_DARK, font=FONT_HEADER, align=ALIGN_LEFT)
    for col in range(2, 14):
        cl = get_column_letter(col)
        ws.cell(row=r, column=col).value = f'=SUM({cl}47:{cl}53)'
        write_header(ws, r, col, ws.cell(row=r, column=col).value, fill=FILL_DARK)

    print("  Bloco 4 complete: rows 45-54", flush=True)

    # ===============================================================
    # BLOCO 5: MOTIVOS DE NAO COMPRA (Rows 59-71)
    # ===============================================================
    print("  Building Bloco 5: MOTIVOS DE NAO COMPRA...", flush=True)

    # Row 59: section title
    section_title(ws, 59, "MOTIVOS DE NAO COMPRA + DONO DA ACAO", 8)

    # Row 60: headers
    b5_headers = [
        "MOTIVO", "QTD", "%", "PROSP", "ATIVOS", "INAT", "POS-V", "DONO DA ACAO"
    ]
    for i, h in enumerate(b5_headers, 1):
        write_header(ws, 60, i, h, fill=FILL_DARK, font=FONT_HEADER_SM, align=ALIGN_CENTER)

    # Rows 61-70: motivos
    for m_idx, motivo in enumerate(motivos[:10]):
        r = 61 + m_idx

        # A: MOTIVO name
        write_data(ws, r, 1, motivo, align=ALIGN_LEFT)

        # B: QTD
        mot_f = f'LOG!$N$3:$N$21000,A{r}'
        ws.cell(row=r, column=2).value = f'={cf(mot_f)}'
        style_cell(ws.cell(row=r, column=2))

        # C: %
        ws.cell(row=r, column=3).value = f'=IFERROR(B{r}/SUM($B$61:$B$70),0)'
        ws.cell(row=r, column=3).number_format = '0%'
        style_cell(ws.cell(row=r, column=3), fmt='0%')

        # D: PROSP (TIPO = PROSPECCAO)
        ws.cell(row=r, column=4).value = f'={cf(mot_f + ",LOG!$L$3:$L$21000,\"PROSPECCAO\"")}'
        style_cell(ws.cell(row=r, column=4))

        # E: ATIVOS (TIPO = ATEND. CLIENTES ATIVOS)
        ws.cell(row=r, column=5).value = f'={cf(mot_f + ",LOG!$L$3:$L$21000,\"ATEND. CLIENTES ATIVOS\"")}'
        style_cell(ws.cell(row=r, column=5))

        # F: INAT (TIPO = ATEND. CLIENTES INATIVOS)
        ws.cell(row=r, column=6).value = f'={cf(mot_f + ",LOG!$L$3:$L$21000,\"ATEND. CLIENTES INATIVOS\"")}'
        style_cell(ws.cell(row=r, column=6))

        # G: POS-V (TIPO = POS-VENDA)
        ws.cell(row=r, column=7).value = f'={cf(mot_f + ",LOG!$L$3:$L$21000,\"POS-VENDA\"")}'
        style_cell(ws.cell(row=r, column=7))

        # H: DONO DA ACAO (empty, gestor fills manually)
        ws.cell(row=r, column=8).value = ""
        style_cell(ws.cell(row=r, column=8), fill=FILL_INPUT)

    # Row 71: TOTAL
    r = 71
    write_header(ws, r, 1, "TOTAL", fill=FILL_DARK, font=FONT_HEADER, align=ALIGN_LEFT)
    for col in [2, 4, 5, 6, 7]:
        cl = get_column_letter(col)
        ws.cell(row=r, column=col).value = f'=SUM({cl}61:{cl}70)'
        write_header(ws, r, col, ws.cell(row=r, column=col).value, fill=FILL_DARK)

    # C71: % total
    ws.cell(row=r, column=3).value = f'=IFERROR(B{r}/B{r},0)'
    ws.cell(row=r, column=3).number_format = '0%'
    write_header(ws, r, 3, ws.cell(row=r, column=3).value, fill=FILL_DARK)
    ws.cell(row=r, column=3).number_format = '0%'

    # H71: empty
    write_header(ws, r, 8, "", fill=FILL_DARK)

    print("  Bloco 5 complete: rows 59-71", flush=True)

    # ===============================================================
    # GENERAL FORMATTING
    # ===============================================================

    # Freeze panes at A3 (rows 1-2 visible always)
    ws.freeze_panes = "A3"

    # Zoom
    ws.sheet_view.zoomScale = 90

    # Row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    for r in range(3, 72):
        if r not in [3, 11, 14, 15, 24, 25, 26, 27, 28, 35, 43, 44, 55, 56, 57, 58]:
            ws.row_dimensions[r].height = 18

    print("  Formatting complete", flush=True)
    return ws


# ===================================================================
# MAIN
# ===================================================================
def main():
    print("=" * 70, flush=True)
    print("Phase 08 Plan 02: Build COMITE Tab", flush=True)
    print("=" * 70, flush=True)

    # Step 0: Backup V13
    if not BACKUP_PATH.exists():
        shutil.copy2(str(V13_PATH), str(BACKUP_PATH))
        print(f"  Backup created: {BACKUP_PATH}", flush=True)
    else:
        print(f"  Backup already exists: {BACKUP_PATH}", flush=True)

    # Step 1: Extract top 10 motivos from LOG (data_only=True)
    print("\nStep 1: Extracting top motivos from LOG...", flush=True)
    wb_data = openpyxl.load_workbook(str(V13_PATH), data_only=True)
    if "LOG" in wb_data.sheetnames:
        motivos = extract_top_motivos(wb_data["LOG"])
    else:
        motivos = [
            "AINDA TEM ESTOQUE", "PRODUTO NAO VENDEU / SEM GIRO",
            "LOJA ANEXO/PROXIMO - SM", "SO QUER COMPRAR GRANEL",
            "PROBLEMA LOGISTICO / ENTREGA", "PROBLEMA FINANCEIRO / CREDITO",
            "PROPRIETARIO / INDISPONIVEL", "FECHANDO / FECHOU LOJA",
            "SEM INTERESSE NO MOMENTO", "PRIMEIRO CONTATO / SEM RESPOSTA",
        ]
    wb_data.close()
    print(f"  Top {len(motivos)} motivos: {motivos}", flush=True)

    # Ensure we have at least 10
    while len(motivos) < 10:
        motivos.append(f"MOTIVO_{len(motivos) + 1}")

    # Step 2: Load V13 (formulas preserved)
    print("\nStep 2: Loading V13 (data_only=False)...", flush=True)
    wb = openpyxl.load_workbook(str(V13_PATH), data_only=False)
    prj_name = find_projecao_sheet(wb)
    print(f"  PROJECAO sheet: {repr(prj_name)}", flush=True)
    print(f"  All sheets: {wb.sheetnames}", flush=True)

    # Step 3: Build COMITE
    ws = build_comite(wb, prj_name, motivos)

    # Step 4: Verify PROJECAO integrity
    print("\nStep 4: Verifying PROJECAO integrity...", flush=True)
    prj_ws = wb[prj_name]
    prj_formulas = count_formulas(prj_ws)
    print(f"  PROJECAO formulas: {prj_formulas}", flush=True)
    assert prj_formulas >= 19224, f"PROJECAO formulas lost: {prj_formulas} < 19224"
    print(f"  PASS: >= 19,224 formulas intact", flush=True)

    # Step 5: Count COMITE formulas
    comite_formulas = count_formulas(ws)
    print(f"\n  COMITE formulas: {comite_formulas}", flush=True)

    # Step 6: Move COMITE to be the 5th tab
    print("\nStep 6: Ordering tabs...", flush=True)
    # Target order: PROJECAO, LOG, DASH, REDES_FRANQUIAS_v2, COMITE
    desired_pos = 4  # 0-indexed, so position 4 = 5th tab
    current_pos = wb.sheetnames.index("COMITE")
    if current_pos != desired_pos:
        wb.move_sheet("COMITE", offset=desired_pos - current_pos)
    print(f"  Tab order: {wb.sheetnames}", flush=True)

    # Step 7: Save
    print("\nStep 7: Saving V13...", flush=True)
    wb.save(str(V13_PATH))
    print(f"  Saved: {V13_PATH}", flush=True)
    wb.close()

    # Step 8: Reopen for verification
    print("\nStep 8: Independent verification (reopen)...", flush=True)
    wb2 = openpyxl.load_workbook(str(V13_PATH), data_only=False)

    # Verify COMITE tab exists
    assert "COMITE" in wb2.sheetnames, "COMITE tab not found after reload!"
    ws2 = wb2["COMITE"]

    # Check section titles
    section_rows = [4, 16, 29, 45, 59]
    section_titles = []
    for sr in section_rows:
        val = ws2.cell(row=sr, column=1).value
        section_titles.append(val)
        print(f"  Row {sr}: {val}", flush=True)

    title_count = sum(1 for v in section_titles if v and "  " in str(v))
    assert title_count >= 5, f"Expected 5 section titles, found {title_count}"
    print(f"  PASS: {title_count} section titles found", flush=True)

    # Check formula cells
    comite_fc = count_formulas(ws2)
    print(f"  COMITE formulas (reload): {comite_fc}", flush=True)
    assert comite_fc >= 200, f"COMITE formulas too few: {comite_fc}"
    print(f"  PASS: >= 200 formulas", flush=True)

    # Check PROJECAO formulas
    prj_name2 = find_projecao_sheet(wb2)
    prj_fc = count_formulas(wb2[prj_name2])
    print(f"  PROJECAO formulas (reload): {prj_fc}", flush=True)
    assert prj_fc >= 19224, f"PROJECAO formulas lost: {prj_fc}"
    print(f"  PASS: >= 19,224 PROJECAO formulas intact", flush=True)

    # Check DataValidation
    dv_count = len(ws2.data_validations.dataValidation) if ws2.data_validations else 0
    print(f"  DataValidation rules: {dv_count}", flush=True)
    assert dv_count >= 2, f"Expected >= 2 DataValidation rules, found {dv_count}"
    print(f"  PASS: >= 2 DataValidation dropdowns", flush=True)

    # Check conditional formatting
    cf_count = len(ws2.conditional_formatting._cf_rules) if hasattr(ws2.conditional_formatting, '_cf_rules') else 0
    print(f"  Conditional formatting rules: {cf_count}", flush=True)
    assert cf_count >= 3, f"Expected >= 3 CF rules, found {cf_count}"
    print(f"  PASS: >= 3 conditional formatting rules", flush=True)

    # Check tab position
    comite_pos = wb2.sheetnames.index("COMITE")
    print(f"  COMITE tab position: {comite_pos} (0-indexed)", flush=True)

    # Check backup
    assert BACKUP_PATH.exists(), f"Backup not found: {BACKUP_PATH}"
    print(f"  PASS: Backup exists", flush=True)

    wb2.close()

    # ---- FINAL REPORT ----
    print("\n" + "=" * 70, flush=True)
    print("EXECUTION REPORT", flush=True)
    print("=" * 70, flush=True)
    print(f"  COMITE tab created: YES", flush=True)
    print(f"  Tab position: {comite_pos + 1} (of {len(wb2.sheetnames)})", flush=True)
    print(f"  Blocks: 5 (rows 4, 16, 29, 45, 59)", flush=True)
    print(f"  COMITE formulas: {comite_fc}", flush=True)
    print(f"  PROJECAO formulas: {prj_fc}", flush=True)
    print(f"  DataValidation: {dv_count} rules", flush=True)
    print(f"  Conditional formatting: {cf_count} rules", flush=True)
    print(f"  Motivos extracted: {len(motivos)}", flush=True)
    print(f"\n  ALL CHECKS PASSED!", flush=True)


if __name__ == "__main__":
    main()

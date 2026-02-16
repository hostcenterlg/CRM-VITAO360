"""
V3 — Tab PROJEÇÃO: Meta SAP + Sinaleiro de Atingimento (NEW)
"""
from openpyxl.utils import get_column_letter
from v3_styles import *


# Real data from spec Part 5.2
REDES_META = [
    # (GRUPO CHAVE, META JAN, META FEV, POSITIVACAO)
    ("CIA DA SAUDE", 52500, 55500, 35),
    ("FITLAND", 43500, 45000, 29),
    ("DIVINA TERRA", 24000, 25500, 16),
    ("VIDA LEVE", 24000, 24000, 16),
    ("BIO MUNDO", 6000, 6000, 4),
    ("TUDO EM GRAOS / VGA", 4500, 4500, 3),
    ("ARMAZEM FIT STORE", 3000, 3000, 2),
    ("MUNDO VERDE", 3000, 3000, 2),
    ("NATURVIDA", 3000, 3000, 2),
    ("ESMERALDA", 3000, 3000, 2),
    ("TRIP", 1500, 1500, 1),
    ("LIGEIRINHO", 1500, 1500, 1),
    ("MERCOCENTRO", 1500, 1500, 0),
]

# Sinaleiro rede data from spec Part 10.4
SINALEIRO_REDE = {
    "CIA DA SAUDE": (163, 2.6, "VERMELHO"),
    "FITLAND": (89, 29.8, "VERMELHO"),
    "DIVINA TERRA": (85, 10.0, "VERMELHO"),
    "VIDA LEVE": (81, 8.0, "VERMELHO"),
    "BIO MUNDO": (167, 1.4, "VERMELHO"),
    "TUDO EM GRAOS / VGA": (25, 6.2, "VERMELHO"),
    "ARMAZEM FIT STORE": (114, 0.0, "ROXO"),
    "MUNDO VERDE": (199, 1.4, "VERMELHO"),
    "NATURVIDA": (0, 0, "ROXO"),
    "ESMERALDA": (0, 0, "ROXO"),
    "TRIP": (0, 0, "ROXO"),
    "LIGEIRINHO": (0, 0, "ROXO"),
    "MERCOCENTRO": (0, 0, "ROXO"),
}

# Consultor per rede
CONSULTOR_REDE = {
    "CIA DA SAUDE": "JULIO GADRET",
    "FITLAND": "JULIO GADRET",
    "DIVINA TERRA": "DAIANE STAVICKI",
    "VIDA LEVE": "DAIANE STAVICKI",
    "BIO MUNDO": "DAIANE STAVICKI",
    "TUDO EM GRAOS / VGA": "DAIANE STAVICKI",
    "ARMAZEM FIT STORE": "DAIANE STAVICKI",
    "MUNDO VERDE": "DAIANE STAVICKI",
    "NATURVIDA": "DAIANE STAVICKI",
    "ESMERALDA": "DAIANE STAVICKI",
    "TRIP": "DAIANE STAVICKI",
    "LIGEIRINHO": "DAIANE STAVICKI",
    "MERCOCENTRO": "DAIANE STAVICKI",
}

MESES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]


def build_projecao(wb):
    """Build PROJEÇÃO tab with SAP meta data."""
    ws = wb.create_sheet("PROJEÇÃO")
    ws.sheet_properties.tabColor = TAB_PROJECAO

    # ── Row 1: Title ──
    ws.merge_cells('A1:AM1')
    title = ws.cell(row=1, column=1, value="📊 PROJEÇÃO — META SAP + SINALEIRO DE ATINGIMENTO 2026")
    style_cell(title, font=FONT_TITLE_W, fill=PatternFill('solid', fgColor='FF6B00'), align=ALIGN_LEFT)
    for c in range(1, 40):
        ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor='FF6B00')
        ws.cell(row=1, column=c).border = THIN_BORDER

    # ── Row 2: Block headers ──
    blocks = [
        ("IDENTIDADE", 1, 4, FILL_BLOCK_IDENTIDADE),
        ("SINALEIRO REDE", 5, 9, PatternFill('solid', fgColor='880E4F')),
        ("META SAP MENSAL", 10, 22, PatternFill('solid', fgColor='1565C0')),
        ("REALIZADO MENSAL", 23, 35, PatternFill('solid', fgColor='2E7D32')),
        ("INDICADORES", 36, 39, PatternFill('solid', fgColor='6A1B9A')),
    ]
    for label, start, end, fill in blocks:
        if end > start:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        cell = ws.cell(row=2, column=start, value=label)
        style_cell(cell, font=FONT_BLOCK, fill=fill)
        for c in range(start, end + 1):
            ws.cell(row=2, column=c).fill = fill
            ws.cell(row=2, column=c).border = THIN_BORDER

    # ── Row 3: Column headers ──
    headers = [
        # IDENTIDADE
        (1, "CNPJ", 18), (2, "NOME FANTASIA", 25), (3, "REDE / GRUPO CHAVE", 22), (4, "CONSULTOR", 20),
        # SINALEIRO REDE
        (5, "TOTAL LOJAS", 12), (6, "SINALEIRO %", 12), (7, "COR", 10),
        (8, "MATURIDADE", 14), (9, "AÇÃO RECOMENDADA", 20),
        # META SAP: J=10 (anual) + K-V (jan-dez) = 13 cols
        (10, "META ANUAL", 14),
    ]
    # Add month headers for META (cols 11-22)
    for i, mes in enumerate(MESES):
        headers.append((11 + i, f"META {mes}", 12))

    # REALIZADO: W=23 (anual) + X-AI (jan-dez) = 13 cols
    headers.append((23, "REAL. ANUAL", 14))
    for i, mes in enumerate(MESES):
        headers.append((24 + i, f"REAL. {mes}", 12))

    # INDICADORES
    headers.extend([
        (36, "% YTD", 10), (37, "🚦 SINALEIRO", 12),
        (38, "GAP", 14), (39, "RANKING", 10),
    ])

    for col, label, width in headers:
        write_header(ws, 3, col, label)
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Row 4+: Data (1 row per rede) ──
    for r_idx, (rede, meta_jan, meta_fev, posit) in enumerate(REDES_META):
        r = 4 + r_idx
        consultor = CONSULTOR_REDE.get(rede, "")
        lojas, sinal_pct, cor = SINALEIRO_REDE.get(rede, (0, 0, "ROXO"))

        # Maturidade based on sinaleiro color
        maturidade_map = {
            "ROXO": "Inexplorada", "VERMELHO": "Ativação",
            "AMARELO": "Sell Out", "VERDE": "JBP",
        }
        maturidade = maturidade_map.get(cor, "Ativação")
        acao_map = {
            "ROXO": "Prospecção", "VERMELHO": "Ativação / Positivação",
            "AMARELO": "Sell Out (expandir mix)", "VERDE": "JBP - Joint Business Plan",
        }
        acao = acao_map.get(cor, "Ativação")

        # IDENTIDADE
        write_data(ws, r, 1, "", fmt=FMT_TEXT)  # CNPJ (to be filled per client)
        write_data(ws, r, 2, rede, align=ALIGN_LEFT)
        write_data(ws, r, 3, rede, align=ALIGN_LEFT)
        write_data(ws, r, 4, consultor, align=ALIGN_LEFT)

        # SINALEIRO REDE
        write_data(ws, r, 5, lojas)
        write_data(ws, r, 6, sinal_pct / 100 if sinal_pct else 0, fmt=FMT_PCT)
        write_data(ws, r, 7, cor)
        write_data(ws, r, 8, maturidade, align=ALIGN_LEFT)
        write_data(ws, r, 9, acao, align=ALIGN_LEFT)

        # META SAP — distribute evenly across months (simple model)
        # Using JAN and FEV real data, rest extrapolated from average
        meta_anual = meta_jan * 12  # simplified
        write_data(ws, r, 10, meta_anual, fmt=FMT_MONEY)
        write_data(ws, r, 11, meta_jan, fmt=FMT_MONEY)  # JAN
        write_data(ws, r, 12, meta_fev, fmt=FMT_MONEY)  # FEV
        avg_meta = (meta_jan + meta_fev) / 2
        for m in range(2, 12):  # MAR-DEZ
            write_data(ws, r, 13 + m - 2, avg_meta, fmt=FMT_MONEY)

        # REALIZADO — empty (to be populated)
        write_data(ws, r, 23, 0, fmt=FMT_MONEY)
        for m in range(12):
            write_data(ws, r, 24 + m, 0, fmt=FMT_MONEY)

        # INDICADORES (formulas)
        cl = get_column_letter
        # % YTD = REALIZADO ANUAL / META ANUAL
        ws.cell(row=r, column=36).value = f'=IFERROR({cl(23)}{r}/{cl(10)}{r},0)'
        ws.cell(row=r, column=36).number_format = FMT_PCT
        style_cell(ws.cell(row=r, column=36))

        # SINALEIRO META
        ws.cell(row=r, column=37).value = (
            f'=IF({cl(36)}{r}>=1,"🟢",IF({cl(36)}{r}>=0.5,"🟡",'
            f'IF({cl(23)}{r}=0,"⚫","🔴")))'
        )
        style_cell(ws.cell(row=r, column=37))

        # GAP = META - REALIZADO
        ws.cell(row=r, column=38).value = f'={cl(10)}{r}-{cl(23)}{r}'
        ws.cell(row=r, column=38).number_format = FMT_MONEY
        style_cell(ws.cell(row=r, column=38))

        # RANKING
        last_data_row = 4 + len(REDES_META) - 1
        ws.cell(row=r, column=39).value = (
            f'=RANK({cl(38)}{r},{cl(38)}$4:{cl(38)}${last_data_row},0)'
        )
        style_cell(ws.cell(row=r, column=39))

    # ── TOTAL row ──
    total_row = 4 + len(REDES_META)
    write_header(ws, total_row, 2, "TOTAL", fill=FILL_HEADER, align=ALIGN_LEFT)
    write_header(ws, total_row, 5, f'=SUM(E4:E{total_row-1})', fill=FILL_HEADER)

    # Sum meta and realizado columns
    for col in [10] + list(range(11, 23)) + [23] + list(range(24, 36)):
        cl_letter = get_column_letter(col)
        ws.cell(row=total_row, column=col).value = f'=SUM({cl_letter}4:{cl_letter}{total_row-1})'
        ws.cell(row=total_row, column=col).number_format = FMT_MONEY
        write_header(ws, total_row, col,
                     ws.cell(row=total_row, column=col).value, fill=FILL_HEADER)

    # Total % YTD
    cl = get_column_letter
    ws.cell(row=total_row, column=36).value = f'=IFERROR({cl(23)}{total_row}/{cl(10)}{total_row},0)'
    ws.cell(row=total_row, column=36).number_format = FMT_PCT
    write_header(ws, total_row, 36, ws.cell(row=total_row, column=36).value, fill=FILL_HEADER)

    # ── Freeze at E4 ──
    ws.freeze_panes = "E4"

    # ── Column grouping ──
    group_columns(ws, 5, 9, outline_level=1, hidden=True)   # Sinaleiro rede
    group_columns(ws, 11, 22, outline_level=1, hidden=True)  # Meta mensal
    group_columns(ws, 24, 35, outline_level=1, hidden=True)  # Realizado mensal

    print(f"  PROJEÇÃO: {total_row} rows, 39 cols, {len(REDES_META)} redes")
    return ws

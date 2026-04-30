"""
CRM VITAO360 — parser_verbas.py
==================================
Parser para "Verbas xxxx.xlsx" — verbas efetivadas por cliente.

Estrutura esperada:
  - Aba principal (primeira aba ou aba com "Verbas" no nome)
  - Colunas obrigatórias: CNPJ, Ano, Valor
  - Coluna opcional: Desconto % total
  - Tipo fixo: EFETIVADA

Tabela destino: cliente_verba_anual (tipo='EFETIVADA')
Chave de negócio: (cnpj, ano, tipo='EFETIVADA', fonte='LOG_UPLOAD')

Regras invioláveis aplicadas:
  R5  — CNPJ normalizado: 14 dígitos string zero-padded
  R8  — Zero fabricação: linhas com dado parcial puladas (log WARNING)
"""

from __future__ import annotations

import logging
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from scripts.parsers.base_parser import BaseParser
from backend.app.models.cliente_verba import ClienteVerbaAnual

logger = logging.getLogger(__name__)

# Colunas aceitas (case-insensitive, parcial match)
_CNPJ_HEADERS = {"cnpj", "cliente", "cod_cliente", "codigo"}
_ANO_HEADERS = {"ano", "year", "exercicio", "competencia"}
_VALOR_HEADERS = {"valor", "verba", "valor_brl", "total", "r$"}
_DESC_PCT_HEADERS = {"desc_pct", "desconto", "desconto_%", "desc%", "pct"}


def _normaliza_cnpj(val) -> str | None:
    if val is None:
        return None
    texto = re.sub(r"\D", "", str(val))
    if len(texto) < 11:
        return None
    return texto.zfill(14)[:14]


def _parse_decimal(val) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            return Decimal(str(val))
        except InvalidOperation:
            return None
    texto = str(val).strip().replace(".", "").replace(",", ".").replace("R$", "").strip()
    if not texto:
        return None
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def _match_header(cell_text: str, candidates: set) -> bool:
    """Verifica se o texto do cabeçalho corresponde a algum candidato (case-insensitive)."""
    normalized = cell_text.lower().strip()
    return any(c in normalized or normalized in c for c in candidates)


class ParserVerbas(BaseParser):
    """
    Parser para arquivo de verbas efetivadas: 'Verbas xxxx.xlsx'.

    Detecta automaticamente as colunas de CNPJ, Ano e Valor.
    Tipo fixo: EFETIVADA.
    """

    FONTE = "LOG_UPLOAD"

    def extract(self, path: str | Path) -> list[dict]:
        """Lê a aba principal e retorna linhas brutas."""
        p = Path(path)
        try:
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        except Exception as exc:
            logger.error("[ParserVerbas] Falha ao abrir %s: %s", p.name, exc)
            raise

        # Selecionar aba: prioriza aba com "verba" no nome
        sheet_name = wb.sheetnames[0]
        for name in wb.sheetnames:
            if "verba" in name.lower():
                sheet_name = name
                break

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            logger.warning("[ParserVerbas] Aba %r vazia", sheet_name)
            return []

        # --- Detectar linha de cabeçalho (primeiras 10 linhas)
        header_idx, col_map = self._detect_header(all_rows)
        if col_map.get("cnpj") is None or col_map.get("ano") is None or col_map.get("valor") is None:
            logger.error(
                "[ParserVerbas] Colunas obrigatórias não encontradas. "
                "Mapeamento: %s", col_map
            )
            return []

        rows: list[dict] = []
        for row in all_rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue
            rows.append({
                "cnpj_raw": row[col_map["cnpj"]] if col_map["cnpj"] < len(row) else None,
                "ano_raw": row[col_map["ano"]] if col_map["ano"] < len(row) else None,
                "valor_raw": row[col_map["valor"]] if col_map["valor"] < len(row) else None,
                "desc_pct_raw": (
                    row[col_map["desc_pct"]]
                    if col_map.get("desc_pct") is not None and col_map["desc_pct"] < len(row)
                    else None
                ),
            })

        logger.info("[ParserVerbas] %d linhas extraidas da aba %r", len(rows), sheet_name)
        return rows

    def _detect_header(self, all_rows: list) -> tuple[int, dict]:
        """Detecta linha de cabeçalho e mapeia índices de colunas."""
        for row_idx, row in enumerate(all_rows[:15]):
            col_map: dict = {"cnpj": None, "ano": None, "valor": None, "desc_pct": None}
            matched = 0
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                text = str(cell)
                if col_map["cnpj"] is None and _match_header(text, _CNPJ_HEADERS):
                    col_map["cnpj"] = col_idx
                    matched += 1
                elif col_map["ano"] is None and _match_header(text, _ANO_HEADERS):
                    col_map["ano"] = col_idx
                    matched += 1
                elif col_map["valor"] is None and _match_header(text, _VALOR_HEADERS):
                    col_map["valor"] = col_idx
                    matched += 1
                elif col_map["desc_pct"] is None and _match_header(text, _DESC_PCT_HEADERS):
                    col_map["desc_pct"] = col_idx
            if matched >= 3:
                return row_idx, col_map

        return 0, {"cnpj": 0, "ano": 1, "valor": 2, "desc_pct": None}

    def normalize(self, raw: list[dict]) -> list[ClienteVerbaAnual]:
        """Converte linhas brutas em modelos ClienteVerbaAnual (tipo=EFETIVADA)."""
        models: list[ClienteVerbaAnual] = []
        skipped = 0

        for row in raw:
            cnpj = _normaliza_cnpj(row.get("cnpj_raw"))
            if not cnpj:
                skipped += 1
                logger.warning("[ParserVerbas] CNPJ inválido: %r — linha pulada", row.get("cnpj_raw"))
                continue

            # Extrair ano
            ano_raw = row.get("ano_raw")
            try:
                ano = int(str(ano_raw).strip()) if ano_raw is not None else None
            except (ValueError, TypeError):
                ano = None
            if not ano or ano < 2000:
                skipped += 1
                logger.warning("[ParserVerbas] Ano inválido %r para CNPJ %s — linha pulada", ano_raw, cnpj)
                continue

            valor = _parse_decimal(row.get("valor_raw"))
            if valor is None:
                skipped += 1
                logger.warning("[ParserVerbas] Valor inválido para CNPJ %s ano %d — linha pulada", cnpj, ano)
                continue

            desc_pct = _parse_decimal(row.get("desc_pct_raw"))

            models.append(
                ClienteVerbaAnual(
                    cnpj=cnpj,
                    ano=ano,
                    tipo="EFETIVADA",
                    valor_brl=valor,
                    desc_total_pct=desc_pct,
                    inicio_vigencia=None,
                    fim_vigencia=None,
                    fonte=self.FONTE,
                    classificacao="REAL",
                    observacao=None,
                )
            )

        if skipped:
            logger.warning("[ParserVerbas] %d linhas puladas", skipped)

        logger.info("[ParserVerbas] normalize: %d modelos produzidos", len(models))
        return models

"""
CRM VITAO360 — parser_frete.py
=================================
Parser para "Frete por Cliente.xlsx" — frete CT-e mensal por cliente.

Estrutura esperada:
  - Aba principal com colunas: CNPJ, Ano, Mês, Qtd CTs, Valor
  - Ou layout pivot: CNPJ na col A, meses nas colunas seguintes (Jan/2025...)

Tabela destino: cliente_frete_mensal
Chave de negócio: (cnpj, ano, mes, fonte='LOG_UPLOAD')

Regras invioláveis aplicadas:
  R5  — CNPJ normalizado: 14 dígitos string zero-padded
  R8  — Zero fabricação: linhas com dado parcial puladas (log WARNING)
"""

from __future__ import annotations

import logging
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from scripts.parsers.base_parser import BaseParser
from backend.app.models.cliente_frete import ClienteFretesMensal

logger = logging.getLogger(__name__)

_CNPJ_HEADERS = {"cnpj", "cliente", "cod_cliente", "codigo"}
_ANO_HEADERS = {"ano", "year", "exercicio"}
_MES_HEADERS = {"mes", "month", "mês", "competencia"}
_VALOR_HEADERS = {"valor", "frete", "valor_frete", "total", "r$", "valor_brl"}
_QTD_HEADERS = {"qtd", "qtd_ctes", "quantidade", "ctes", "nf", "ct-e", "cte"}

_MESES_NOME = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

_RE_MES_ANO = re.compile(
    r"(?i)(?P<mes>jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)[./\-]?(?P<ano>\d{4})"
)


def _normaliza_cnpj(val) -> str | None:
    if val is None:
        return None
    texto = re.sub(r"\D", "", str(val))
    if len(texto) < 11:
        return None
    return texto.zfill(14)[:14]


def _parse_decimal(val) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            return Decimal(str(val))
        except InvalidOperation:
            return None
    texto = str(val).strip().replace(".", "").replace(",", ".").replace("R$", "").strip()
    if not texto:
        return None
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def _match_header(cell_text: str, candidates: set) -> bool:
    normalized = cell_text.lower().strip()
    return any(c in normalized or normalized in c for c in candidates)


class ParserFrete(BaseParser):
    """
    Parser para 'Frete por Cliente.xlsx'.

    Suporta dois layouts:
      1. Tabular: CNPJ | Ano | Mes | Qtd | Valor
      2. Pivot: CNPJ | Jan/2025 | Fev/2025 | ...
    """

    FONTE = "LOG_UPLOAD"

    def extract(self, path: str | Path) -> list[dict]:
        """Lê o arquivo e retorna linhas brutas."""
        p = Path(path)
        try:
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        except Exception as exc:
            logger.error("[ParserFrete] Falha ao abrir %s: %s", p.name, exc)
            raise

        # Selecionar aba: prioriza aba com "frete" no nome
        sheet_name = wb.sheetnames[0]
        for name in wb.sheetnames:
            if "frete" in name.lower():
                sheet_name = name
                break

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            return []

        # Detectar layout
        header_idx, col_map, pivot_cols = self._detect_layout(all_rows)

        rows: list[dict] = []
        if pivot_cols:
            rows = self._extract_pivot(all_rows, header_idx, col_map, pivot_cols)
        else:
            rows = self._extract_tabular(all_rows, header_idx, col_map)

        logger.info("[ParserFrete] %d linhas extraidas", len(rows))
        return rows

    def _detect_layout(self, all_rows: list) -> tuple[int, dict, dict]:
        """
        Detecta layout tabular ou pivot.
        Retorna (header_idx, col_map, pivot_cols).
        pivot_cols = {col_idx: (ano, mes)} se layout pivot, {} se tabular.
        """
        for row_idx, row in enumerate(all_rows[:15]):
            col_map = {"cnpj": None, "ano": None, "mes": None, "valor": None, "qtd": None}
            pivot_cols = {}
            matched_tabular = 0
            matched_pivot = 0

            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                text = str(cell)

                # Verificar se é cabeçalho de período pivot
                m = _RE_MES_ANO.search(text)
                if m and col_idx > 0:
                    ano = int(m.group("ano"))
                    mes = _MESES_NOME.get(m.group("mes").lower())
                    if mes:
                        pivot_cols[col_idx] = (ano, mes)
                        matched_pivot += 1
                        continue

                # Layout tabular
                if col_map["cnpj"] is None and _match_header(text, _CNPJ_HEADERS):
                    col_map["cnpj"] = col_idx; matched_tabular += 1
                elif col_map["ano"] is None and _match_header(text, _ANO_HEADERS):
                    col_map["ano"] = col_idx; matched_tabular += 1
                elif col_map["mes"] is None and _match_header(text, _MES_HEADERS):
                    col_map["mes"] = col_idx; matched_tabular += 1
                elif col_map["valor"] is None and _match_header(text, _VALOR_HEADERS):
                    col_map["valor"] = col_idx; matched_tabular += 1
                elif col_map["qtd"] is None and _match_header(text, _QTD_HEADERS):
                    col_map["qtd"] = col_idx

            if matched_pivot >= 2 and col_map["cnpj"] is not None:
                return row_idx, col_map, pivot_cols
            if matched_tabular >= 3:
                return row_idx, col_map, {}

        return 0, {"cnpj": 0, "ano": 1, "mes": 2, "valor": 3, "qtd": None}, {}

    def _extract_tabular(self, all_rows: list, header_idx: int, col_map: dict) -> list[dict]:
        """Extrai layout tabular: uma linha por (CNPJ, Ano, Mes)."""
        rows = []
        for row in all_rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue
            def safe(idx):
                return row[idx] if idx is not None and idx < len(row) else None

            rows.append({
                "cnpj_raw": safe(col_map["cnpj"]),
                "ano_raw": safe(col_map["ano"]),
                "mes_raw": safe(col_map["mes"]),
                "valor_raw": safe(col_map["valor"]),
                "qtd_raw": safe(col_map.get("qtd")),
            })
        return rows

    def _extract_pivot(
        self, all_rows: list, header_idx: int, col_map: dict, pivot_cols: dict
    ) -> list[dict]:
        """Extrai layout pivot: CNPJ na col A, meses nas colunas."""
        rows = []
        cnpj_col = col_map.get("cnpj", 0)
        for row in all_rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue
            cnpj_raw = row[cnpj_col] if cnpj_col < len(row) else None
            for col_idx, (ano, mes) in pivot_cols.items():
                if col_idx >= len(row):
                    continue
                rows.append({
                    "cnpj_raw": cnpj_raw,
                    "ano_raw": ano,
                    "mes_raw": mes,
                    "valor_raw": row[col_idx],
                    "qtd_raw": None,
                })
        return rows

    def normalize(self, raw: list[dict]) -> list[ClienteFretesMensal]:
        """Converte linhas brutas em modelos ClienteFretesMensal."""
        models: list[ClienteFretesMensal] = []
        skipped = 0

        for row in raw:
            cnpj = _normaliza_cnpj(row.get("cnpj_raw"))
            if not cnpj:
                skipped += 1
                logger.warning("[ParserFrete] CNPJ inválido: %r — linha pulada", row.get("cnpj_raw"))
                continue

            try:
                ano = int(str(row.get("ano_raw", "")).strip())
            except (ValueError, TypeError):
                skipped += 1
                logger.warning("[ParserFrete] Ano inválido %r para CNPJ %s", row.get("ano_raw"), cnpj)
                continue
            if ano < 2000:
                skipped += 1
                continue

            try:
                mes = int(str(row.get("mes_raw", "")).strip())
            except (ValueError, TypeError):
                skipped += 1
                logger.warning("[ParserFrete] Mês inválido %r para CNPJ %s", row.get("mes_raw"), cnpj)
                continue
            if not 1 <= mes <= 12:
                skipped += 1
                continue

            valor = _parse_decimal(row.get("valor_raw"))
            if valor is None:
                skipped += 1
                logger.warning("[ParserFrete] Valor inválido para CNPJ %s %d/%02d", cnpj, ano, mes)
                continue

            qtd_raw = row.get("qtd_raw")
            qtd = None
            if qtd_raw is not None:
                try:
                    qtd = int(float(str(qtd_raw)))
                except (ValueError, TypeError):
                    qtd = None

            models.append(
                ClienteFretesMensal(
                    cnpj=cnpj,
                    ano=ano,
                    mes=mes,
                    qtd_ctes=qtd,
                    valor_brl=valor,
                    fonte=self.FONTE,
                    classificacao="REAL",
                )
            )

        if skipped:
            logger.warning("[ParserFrete] %d linhas puladas", skipped)

        logger.info("[ParserFrete] normalize: %d modelos produzidos", len(models))
        return models

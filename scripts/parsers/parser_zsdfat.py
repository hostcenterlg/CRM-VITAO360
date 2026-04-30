"""
CRM VITAO360 — parser_zsdfat.py
==================================
Parser para relatórios SAP ZSDFAT: ZSDFAT_<cliente>.xlsx

Estrutura esperada do arquivo:
  - 1 a 13 abas (cada aba = um cliente ou consolidado)
  - Coluna A: conta DRE (texto bruto — normalizado pelos 22 regex)
  - Colunas de período: ex. "Jan/2025", "Fev/2025" ... ou colunas de ano
  - Linha de CNPJ: detectada por texto "CNPJ" ou "Cliente" no cabeçalho

Tabela destino: cliente_dre_periodo
  linha=C01..C22 (do regex), fase='A', classificacao='REAL' (upload CFO)
  Linhas RAW: log WARNING + classificacao='SINTETICO', observacao='Conta nao normalizada'

Regras invioláveis aplicadas:
  R5  — CNPJ normalizado: 14 dígitos string zero-padded
  R8  — Zero fabricação: linhas com dado parcial puladas (log WARNING)

Referencia: docs/specs/cowork/README_TIME_TECNICO_DDE_AC.md secao 7.
"""

from __future__ import annotations

import logging
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl

from scripts.parsers.base_parser import BaseParser
from scripts.parsers.dre_corrections import normaliza_conta_dre
from backend.app.models.cliente_dre import ClienteDrePeriodo

logger = logging.getLogger(__name__)

# Regex para detectar cabeçalho de período (ex.: "Jan/2025", "2025", "JAN25")
_RE_PERIODO_MES_ANO = re.compile(
    r"(?i)(?:jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)[./\-]?(\d{4})"
)
_RE_PERIODO_ANO_ONLY = re.compile(r"^\s*(\d{4})\s*$")

_MESES = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

# Regex para extrair CNPJ de célula — dois formatos aceitos:
#   1. Formatado: XX.XXX.XXX/XXXX-XX (18 chars)
#   2. Limpo: 14 dígitos seguidos
_RE_CNPJ = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}|\b\d{14}\b")


def _normaliza_cnpj(val: Any) -> str | None:
    """Normaliza CNPJ para 14 dígitos string. Retorna None se inválido."""
    if val is None:
        return None
    texto = re.sub(r"\D", "", str(val))
    if len(texto) < 11:
        return None
    return texto.zfill(14)[:14]


def _parse_decimal(val: Any) -> Decimal | None:
    """Converte célula do Excel para Decimal. Retorna None se inválido."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            return Decimal(str(val))
        except InvalidOperation:
            return None
    texto = str(val).strip().replace(".", "").replace(",", ".").replace("R$", "").strip()
    if not texto:
        return None
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def _parse_periodo_header(header: Any) -> tuple[int | None, int | None]:
    """
    Extrai (ano, mes) de um cabeçalho de coluna.

    Exemplos:
      "Jan/2025" -> (2025, 1)
      "2025"     -> (2025, None)   # consolidado anual
      "JAN25"    -> (2025, 1)
    """
    if header is None:
        return None, None
    texto = str(header).strip()

    m = _RE_PERIODO_MES_ANO.search(texto)
    if m:
        ano = int(m.group(1))
        mes_str = texto[:3].lower()
        mes = _MESES.get(mes_str)
        return ano, mes

    m2 = _RE_PERIODO_ANO_ONLY.match(texto)
    if m2:
        return int(m2.group(1)), None

    return None, None


class ParserZSDFAT(BaseParser):
    """
    Parser para ZSDFAT_<cliente>.xlsx — relatório DRE SAP por cliente.

    Suporta até 13 abas por arquivo. Cada aba é processada independentemente.
    CNPJ é detectado na aba: procura linha com "CNPJ" ou texto com 14 dígitos.
    Linhas de conta DRE são normalizadas pelos 22 regex (dre_corrections).
    """

    FONTE = "SAP"

    def extract(self, path: str | Path) -> list[dict]:
        """
        Lê todas as abas do ZSDFAT e retorna linhas brutas como list[dict].

        Cada dict contém:
          cnpj_raw, conta_raw, ano, mes, valor_brl_raw, sheet_name
        """
        p = Path(path)
        rows: list[dict] = []

        try:
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        except Exception as exc:
            logger.error("[ParserZSDFAT] Falha ao abrir %s: %s", p.name, exc)
            raise

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_rows = self._extract_sheet(ws, sheet_name)
            rows.extend(sheet_rows)
            logger.info(
                "[ParserZSDFAT] Aba %r: %d linhas extraidas", sheet_name, len(sheet_rows)
            )

        wb.close()
        logger.info("[ParserZSDFAT] Total extraido: %d linhas de %d abas", len(rows), len(wb.sheetnames) if hasattr(wb, 'sheetnames') else 0)
        return rows

    def _extract_sheet(self, ws: Any, sheet_name: str) -> list[dict]:
        """Processa uma aba do ZSDFAT e retorna linhas brutas."""
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return []

        # --- Detectar CNPJ na aba
        cnpj_raw = self._detect_cnpj(all_rows)

        # --- Detectar linha de cabeçalho de período (colunas de tempo)
        header_row_idx, periodo_cols = self._detect_periodo_header(all_rows)
        if not periodo_cols:
            logger.warning(
                "[ParserZSDFAT] Aba %r: nenhuma coluna de período detectada",
                sheet_name,
            )
            return []

        # --- Extrair linhas de dados DRE (após cabeçalho)
        rows: list[dict] = []
        for row in all_rows[header_row_idx + 1:]:
            if not row or row[0] is None:
                continue
            conta_raw = str(row[0]).strip()
            if not conta_raw:
                continue

            for col_idx, (ano, mes) in periodo_cols.items():
                if col_idx >= len(row):
                    continue
                cell_val = row[col_idx]
                rows.append({
                    "cnpj_raw": cnpj_raw,
                    "conta_raw": conta_raw,
                    "ano": ano,
                    "mes": mes,
                    "valor_brl_raw": cell_val,
                    "sheet_name": sheet_name,
                })

        return rows

    def _detect_cnpj(self, all_rows: list) -> str | None:
        """Busca CNPJ nas primeiras 20 linhas da aba."""
        for row in all_rows[:20]:
            for cell in row:
                if cell is None:
                    continue
                texto = str(cell)
                # Procura padrão CNPJ na célula
                m = _RE_CNPJ.search(texto)
                if m:
                    candidato = _normaliza_cnpj(m.group(0))
                    if candidato and len(candidato) == 14:
                        return candidato
        return None

    def _detect_periodo_header(
        self, all_rows: list
    ) -> tuple[int, dict[int, tuple[int | None, int | None]]]:
        """
        Detecta a linha de cabeçalho com colunas de período.

        Retorna (índice_da_linha, {col_idx: (ano, mes)}).
        """
        for row_idx, row in enumerate(all_rows[:30]):
            periodo_cols: dict[int, tuple[int | None, int | None]] = {}
            for col_idx, cell in enumerate(row):
                if col_idx == 0:
                    continue  # col 0 é conta DRE
                ano, mes = _parse_periodo_header(cell)
                if ano is not None:
                    periodo_cols[col_idx] = (ano, mes)
            if len(periodo_cols) >= 1:
                return row_idx, periodo_cols

        return 0, {}

    def normalize(self, raw: list[dict]) -> list[ClienteDrePeriodo]:
        """
        Converte lista de dicts brutos em modelos ClienteDrePeriodo.

        Linhas RAW (conta não reconhecida): cria registro com
          classificacao='SINTETICO', observacao='Conta nao normalizada'
          para revisão manual pelo CFO.

        Linhas sem CNPJ válido: pular com WARNING.
        Linhas sem ano válido: pular com WARNING.
        """
        models: list[ClienteDrePeriodo] = []
        skipped = 0

        for row in raw:
            cnpj = _normaliza_cnpj(row.get("cnpj_raw"))
            if not cnpj:
                skipped += 1
                logger.warning(
                    "[ParserZSDFAT] CNPJ inválido na aba %r, conta %r — linha pulada",
                    row.get("sheet_name"),
                    row.get("conta_raw"),
                )
                continue

            ano = row.get("ano")
            if not isinstance(ano, int) or ano < 2000:
                skipped += 1
                logger.warning(
                    "[ParserZSDFAT] Ano inválido %r para CNPJ %s — linha pulada",
                    ano,
                    cnpj,
                )
                continue

            conta_raw = row.get("conta_raw", "")
            code, conta_canonica = normaliza_conta_dre(conta_raw)

            valor = _parse_decimal(row.get("valor_brl_raw"))

            if code == "RAW":
                classificacao = "SINTETICO"
                observacao = "Conta nao normalizada"
                linha = "RAW"
                logger.warning(
                    "[ParserZSDFAT] Conta nao reconhecida: %r (CNPJ=%s, aba=%r) — salvo como SINTETICO",
                    conta_raw,
                    cnpj,
                    row.get("sheet_name"),
                )
            else:
                classificacao = "REAL"
                observacao = None
                linha = code

            models.append(
                ClienteDrePeriodo(
                    cnpj=cnpj,
                    ano=ano,
                    mes=row.get("mes"),
                    linha=linha,
                    conta=conta_canonica,
                    valor_brl=valor,
                    pct_sobre_receita=None,
                    fonte=self.FONTE,
                    classificacao=classificacao,
                    fase="A",
                    observacao=observacao,
                )
            )

        if skipped:
            logger.warning("[ParserZSDFAT] %d linhas puladas por dados insuficientes", skipped)

        logger.info("[ParserZSDFAT] normalize: %d modelos produzidos", len(models))
        return models

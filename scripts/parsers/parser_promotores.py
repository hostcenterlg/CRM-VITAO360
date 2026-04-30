"""
CRM VITAO360 — parser_promotores.py
======================================
Parser para "Despesas Clientes V2.xlsx" — promotor PDV mensal por cliente.

Estrutura esperada:
  - Aba "RESUMO" (ou primeira aba)
  - Colunas: CNPJ, Agência, Ano, Mês, Valor
  - Ou layout pivot: CNPJ | Agência | Jan/2025 | Fev/2025 | ...

Tabela destino: cliente_promotor_mensal
Chave de negócio: (cnpj, agencia, ano, mes)

Regras invioláveis aplicadas:
  R5  — CNPJ normalizado: 14 dígitos string zero-padded
  R8  — Zero fabricação: linhas com dado parcial puladas (log WARNING)
"""

from __future__ import annotations

import logging
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl

from scripts.parsers.base_parser import BaseParser
from backend.app.models.cliente_promotor import ClientePromotorMensal

logger = logging.getLogger(__name__)

_CNPJ_HEADERS = {"cnpj", "cliente", "cod_cliente", "codigo"}
_AGENCIA_HEADERS = {"agencia", "agência", "empresa", "promotora", "fornecedor"}
_ANO_HEADERS = {"ano", "year", "exercicio"}
_MES_HEADERS = {"mes", "mês", "month", "competencia"}
_VALOR_HEADERS = {"valor", "despesa", "promotor", "total", "r$", "valor_brl"}

_MESES_NOME = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}
_RE_MES_ANO = re.compile(
    r"(?i)(?P<mes>jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)[./\-]?(?P<ano>\d{4})"
)

# Abas candidatas (ordem de prioridade)
_ABA_CANDIDATAS = ["resumo", "summary", "despesas", "promotores", "pdv"]


def _normaliza_cnpj(val: Any) -> str | None:
    if val is None:
        return None
    texto = re.sub(r"\D", "", str(val))
    if len(texto) < 11:
        return None
    return texto.zfill(14)[:14]


def _parse_decimal(val: Any) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            return Decimal(str(val))
        except InvalidOperation:
            return None
    texto = str(val).strip().replace(".", "").replace(",", ".").replace("R$", "").strip()
    if not texto:
        return None
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def _match_header(cell_text: str, candidates: set) -> bool:
    normalized = cell_text.lower().strip()
    return any(c in normalized or normalized in c for c in candidates)


def _select_sheet(wb: Any) -> str:
    """Seleciona aba mais provável para promotores (prioriza RESUMO)."""
    sheetnames_lower = {name.lower(): name for name in wb.sheetnames}
    for candidato in _ABA_CANDIDATAS:
        if candidato in sheetnames_lower:
            return sheetnames_lower[candidato]
    return wb.sheetnames[0]


class ParserPromotores(BaseParser):
    """
    Parser para 'Despesas Clientes V2.xlsx'.

    Prioriza aba RESUMO. Suporta layout tabular e pivot (meses nas colunas).
    """

    FONTE = "LOG_UPLOAD"

    def extract(self, path: str | Path) -> list[dict]:
        """Lê a aba RESUMO e retorna linhas brutas."""
        p = Path(path)
        try:
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        except Exception as exc:
            logger.error("[ParserPromotores] Falha ao abrir %s: %s", p.name, exc)
            raise

        sheet_name = _select_sheet(wb)
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            logger.warning("[ParserPromotores] Aba %r vazia", sheet_name)
            return []

        header_idx, col_map, pivot_cols = self._detect_layout(all_rows)

        if pivot_cols:
            rows = self._extract_pivot(all_rows, header_idx, col_map, pivot_cols)
        else:
            rows = self._extract_tabular(all_rows, header_idx, col_map)

        logger.info("[ParserPromotores] %d linhas extraidas da aba %r", len(rows), sheet_name)
        return rows

    def _detect_layout(self, all_rows: list) -> tuple[int, dict, dict]:
        """Detecta layout tabular ou pivot (meses nas colunas)."""
        for row_idx, row in enumerate(all_rows[:15]):
            col_map = {
                "cnpj": None, "agencia": None, "ano": None,
                "mes": None, "valor": None,
            }
            pivot_cols = {}
            matched_tabular = 0
            matched_pivot = 0

            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                text = str(cell)

                m = _RE_MES_ANO.search(text)
                if m:
                    ano = int(m.group("ano"))
                    mes = _MESES_NOME.get(m.group("mes").lower())
                    if mes:
                        pivot_cols[col_idx] = (ano, mes)
                        matched_pivot += 1
                        continue

                if col_map["cnpj"] is None and _match_header(text, _CNPJ_HEADERS):
                    col_map["cnpj"] = col_idx; matched_tabular += 1
                elif col_map["agencia"] is None and _match_header(text, _AGENCIA_HEADERS):
                    col_map["agencia"] = col_idx
                elif col_map["ano"] is None and _match_header(text, _ANO_HEADERS):
                    col_map["ano"] = col_idx; matched_tabular += 1
                elif col_map["mes"] is None and _match_header(text, _MES_HEADERS):
                    col_map["mes"] = col_idx; matched_tabular += 1
                elif col_map["valor"] is None and _match_header(text, _VALOR_HEADERS):
                    col_map["valor"] = col_idx; matched_tabular += 1

            if matched_pivot >= 2 and col_map["cnpj"] is not None:
                return row_idx, col_map, pivot_cols
            if matched_tabular >= 3:
                return row_idx, col_map, {}

        return 0, {"cnpj": 0, "agencia": None, "ano": 1, "mes": 2, "valor": 3}, {}

    def _extract_tabular(self, all_rows: list, header_idx: int, col_map: dict) -> list[dict]:
        rows = []
        for row in all_rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue

            def safe(idx):
                return row[idx] if idx is not None and idx < len(row) else None

            rows.append({
                "cnpj_raw": safe(col_map.get("cnpj")),
                "agencia_raw": safe(col_map.get("agencia")),
                "ano_raw": safe(col_map.get("ano")),
                "mes_raw": safe(col_map.get("mes")),
                "valor_raw": safe(col_map.get("valor")),
            })
        return rows

    def _extract_pivot(
        self, all_rows: list, header_idx: int, col_map: dict, pivot_cols: dict
    ) -> list[dict]:
        rows = []
        cnpj_col = col_map.get("cnpj", 0)
        agencia_col = col_map.get("agencia")

        for row in all_rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue
            cnpj_raw = row[cnpj_col] if cnpj_col < len(row) else None
            agencia_raw = row[agencia_col] if agencia_col is not None and agencia_col < len(row) else None
            for col_idx, (ano, mes) in pivot_cols.items():
                if col_idx >= len(row):
                    continue
                rows.append({
                    "cnpj_raw": cnpj_raw,
                    "agencia_raw": agencia_raw,
                    "ano_raw": ano,
                    "mes_raw": mes,
                    "valor_raw": row[col_idx],
                })
        return rows

    def normalize(self, raw: list[dict]) -> list[ClientePromotorMensal]:
        """Converte linhas brutas em modelos ClientePromotorMensal."""
        models: list[ClientePromotorMensal] = []
        skipped = 0

        for row in raw:
            cnpj = _normaliza_cnpj(row.get("cnpj_raw"))
            if not cnpj:
                skipped += 1
                logger.warning("[ParserPromotores] CNPJ inválido: %r — linha pulada", row.get("cnpj_raw"))
                continue

            try:
                ano = int(float(str(row.get("ano_raw", "")).strip()))
            except (ValueError, TypeError):
                skipped += 1
                logger.warning("[ParserPromotores] Ano inválido %r para CNPJ %s", row.get("ano_raw"), cnpj)
                continue
            if ano < 2000:
                skipped += 1
                continue

            try:
                mes = int(float(str(row.get("mes_raw", "")).strip()))
            except (ValueError, TypeError):
                skipped += 1
                logger.warning("[ParserPromotores] Mês inválido %r para CNPJ %s", row.get("mes_raw"), cnpj)
                continue
            if not 1 <= mes <= 12:
                skipped += 1
                continue

            valor = _parse_decimal(row.get("valor_raw"))
            if valor is None:
                skipped += 1
                logger.warning(
                    "[ParserPromotores] Valor inválido para CNPJ %s %d/%02d", cnpj, ano, mes
                )
                continue

            # Agência: string ou None (campo opcional no modelo)
            agencia_raw = row.get("agencia_raw")
            agencia = str(agencia_raw).strip()[:80] if agencia_raw else None

            models.append(
                ClientePromotorMensal(
                    cnpj=cnpj,
                    agencia=agencia,
                    ano=ano,
                    mes=mes,
                    valor_brl=valor,
                    fonte=self.FONTE,
                    classificacao="REAL",
                )
            )

        if skipped:
            logger.warning("[ParserPromotores] %d linhas puladas", skipped)

        logger.info("[ParserPromotores] normalize: %d modelos produzidos", len(models))
        return models

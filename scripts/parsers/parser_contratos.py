"""
CRM VITAO360 — parser_contratos.py
=====================================
Parser para "Controle Contratos.xlsx" — contratos anuais com clientes.

Estrutura esperada:
  - Aba "Desc. Financeiro" (ou similar)
  - Colunas: CNPJ, Ano, Valor, Desconto %, Início Vigência, Fim Vigência

Tabela destino: cliente_verba_anual (tipo='CONTRATO')
Chave de negócio: (cnpj, ano, tipo='CONTRATO', fonte='LOG_UPLOAD')

Regras invioláveis aplicadas:
  R5  — CNPJ normalizado: 14 dígitos string zero-padded
  R8  — Zero fabricação: linhas com dado parcial puladas (log WARNING)
"""

from __future__ import annotations

import logging
import re
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl

from scripts.parsers.base_parser import BaseParser
from backend.app.models.cliente_verba import ClienteVerbaAnual

logger = logging.getLogger(__name__)

_CNPJ_HEADERS = {"cnpj", "cliente", "cod_cliente", "codigo"}
_ANO_HEADERS = {"ano", "year", "exercicio", "competencia", "vigencia"}
_VALOR_HEADERS = {"valor", "verba", "contrato", "total", "r$", "valor_brl", "desc_financeiro"}
_DESC_PCT_HEADERS = {"desc_pct", "desconto", "desconto_%", "desc%", "pct", "% desconto"}
_INICIO_HEADERS = {"inicio", "inicio_vigencia", "data_inicio", "de", "from", "inicio vig"}
_FIM_HEADERS = {"fim", "fim_vigencia", "data_fim", "ate", "to", "fim vig"}

# Abas candidatas (ordem de prioridade)
_ABA_CANDIDATAS = ["desc. financeiro", "desc financeiro", "contratos", "financeiro", "contrato"]


def _normaliza_cnpj(val: Any) -> str | None:
    if val is None:
        return None
    texto = re.sub(r"\D", "", str(val))
    if len(texto) < 11:
        return None
    return texto.zfill(14)[:14]


def _parse_decimal(val: Any) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            return Decimal(str(val))
        except InvalidOperation:
            return None
    texto = str(val).strip().replace(".", "").replace(",", ".").replace("R$", "").strip()
    if not texto:
        return None
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def _parse_date(val: Any) -> date | None:
    """Converte célula Excel para date. Aceita date nativo ou string 'DD/MM/YYYY'."""
    if val is None:
        return None
    if isinstance(val, date):
        return val
    texto = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def _match_header(cell_text: str, candidates: set) -> bool:
    normalized = cell_text.lower().strip()
    return any(c in normalized or normalized in c for c in candidates)


def _select_sheet(wb: Any) -> str:
    """Seleciona aba mais provável do arquivo de contratos."""
    sheetnames_lower = {name.lower(): name for name in wb.sheetnames}
    for candidato in _ABA_CANDIDATAS:
        if candidato in sheetnames_lower:
            return sheetnames_lower[candidato]
    # fallback: primeira aba
    return wb.sheetnames[0]


class ParserContratos(BaseParser):
    """
    Parser para 'Controle Contratos.xlsx'.

    Prioriza aba 'Desc. Financeiro'. Tipo fixo: CONTRATO.
    """

    FONTE = "LOG_UPLOAD"

    def extract(self, path: str | Path) -> list[dict]:
        """Lê a aba de contratos e retorna linhas brutas."""
        p = Path(path)
        try:
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        except Exception as exc:
            logger.error("[ParserContratos] Falha ao abrir %s: %s", p.name, exc)
            raise

        sheet_name = _select_sheet(wb)
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            logger.warning("[ParserContratos] Aba %r vazia", sheet_name)
            return []

        header_idx, col_map = self._detect_header(all_rows)
        if col_map.get("cnpj") is None or col_map.get("valor") is None:
            logger.error(
                "[ParserContratos] Colunas obrigatórias (cnpj, valor) não encontradas. "
                "Mapeamento: %s", col_map
            )
            return []

        rows: list[dict] = []
        for row in all_rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue

            def safe(idx):
                return row[idx] if idx is not None and idx < len(row) else None

            rows.append({
                "cnpj_raw": safe(col_map.get("cnpj")),
                "ano_raw": safe(col_map.get("ano")),
                "valor_raw": safe(col_map.get("valor")),
                "desc_pct_raw": safe(col_map.get("desc_pct")),
                "inicio_raw": safe(col_map.get("inicio")),
                "fim_raw": safe(col_map.get("fim")),
            })

        logger.info("[ParserContratos] %d linhas extraidas da aba %r", len(rows), sheet_name)
        return rows

    def _detect_header(self, all_rows: list) -> tuple[int, dict]:
        """Detecta linha de cabeçalho e mapeia índices."""
        for row_idx, row in enumerate(all_rows[:15]):
            col_map = {
                "cnpj": None, "ano": None, "valor": None,
                "desc_pct": None, "inicio": None, "fim": None,
            }
            matched = 0
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                text = str(cell)
                if col_map["cnpj"] is None and _match_header(text, _CNPJ_HEADERS):
                    col_map["cnpj"] = col_idx; matched += 1
                elif col_map["ano"] is None and _match_header(text, _ANO_HEADERS):
                    col_map["ano"] = col_idx; matched += 1
                elif col_map["valor"] is None and _match_header(text, _VALOR_HEADERS):
                    col_map["valor"] = col_idx; matched += 1
                elif col_map["desc_pct"] is None and _match_header(text, _DESC_PCT_HEADERS):
                    col_map["desc_pct"] = col_idx
                elif col_map["inicio"] is None and _match_header(text, _INICIO_HEADERS):
                    col_map["inicio"] = col_idx
                elif col_map["fim"] is None and _match_header(text, _FIM_HEADERS):
                    col_map["fim"] = col_idx
            if matched >= 2:
                return row_idx, col_map

        return 0, {"cnpj": 0, "ano": 1, "valor": 2, "desc_pct": None, "inicio": None, "fim": None}

    def normalize(self, raw: list[dict]) -> list[ClienteVerbaAnual]:
        """Converte linhas brutas em modelos ClienteVerbaAnual (tipo=CONTRATO)."""
        models: list[ClienteVerbaAnual] = []
        skipped = 0

        for row in raw:
            cnpj = _normaliza_cnpj(row.get("cnpj_raw"))
            if not cnpj:
                skipped += 1
                logger.warning("[ParserContratos] CNPJ inválido: %r — linha pulada", row.get("cnpj_raw"))
                continue

            # Ano: tenta extrair do campo "ano_raw" ou da data de início
            ano_raw = row.get("ano_raw")
            ano = None
            if ano_raw is not None:
                try:
                    ano = int(float(str(ano_raw).strip()))
                except (ValueError, TypeError):
                    pass

            # Fallback: extrair ano da data de início
            if not ano:
                inicio = _parse_date(row.get("inicio_raw"))
                if inicio:
                    ano = inicio.year

            if not ano or ano < 2000:
                skipped += 1
                logger.warning(
                    "[ParserContratos] Ano inválido (raw=%r, inicio=%r) para CNPJ %s — linha pulada",
                    ano_raw, row.get("inicio_raw"), cnpj,
                )
                continue

            valor = _parse_decimal(row.get("valor_raw"))
            if valor is None:
                skipped += 1
                logger.warning("[ParserContratos] Valor inválido para CNPJ %s ano %d — linha pulada", cnpj, ano)
                continue

            desc_pct = _parse_decimal(row.get("desc_pct_raw"))
            inicio_vig = _parse_date(row.get("inicio_raw"))
            fim_vig = _parse_date(row.get("fim_raw"))

            models.append(
                ClienteVerbaAnual(
                    cnpj=cnpj,
                    ano=ano,
                    tipo="CONTRATO",
                    valor_brl=valor,
                    desc_total_pct=desc_pct,
                    inicio_vigencia=inicio_vig,
                    fim_vigencia=fim_vig,
                    fonte=self.FONTE,
                    classificacao="REAL",
                    observacao=None,
                )
            )

        if skipped:
            logger.warning("[ParserContratos] %d linhas puladas", skipped)

        logger.info("[ParserContratos] normalize: %d modelos produzidos", len(models))
        return models

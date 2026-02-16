"""
V3 — Tab LOG (24 cols, same structure as DRAFT 2 but no formulas)
     + Tab CLAUDE LOG (simple 3-col audit trail)
"""
import datetime
from openpyxl.utils import get_column_letter
from v3_styles import *


# Same 24 columns as DRAFT 2 — but all plain text (no formulas)
LOG_COLS = [
    (1,  "DATA",              14, FMT_DATE),
    (2,  "CNPJ",              18, FMT_TEXT),
    (3,  "NOME FANTASIA",     25, None),
    (4,  "UF",                 5, None),
    (5,  "CONSULTOR",         20, None),
    (6,  "RESULTADO",         22, None),
    (7,  "MOTIVO",            28, None),
    (8,  "WHATSAPP",          10, None),
    (9,  "LIGAÇÃO",           10, None),
    (10, "LIG. ATENDIDA",     14, None),
    (11, "NOTA DO DIA",       35, None),
    (12, "MERCOS ATUALIZADO", 16, None),
    (13, "SITUAÇÃO",          14, None),
    (14, "ESTÁGIO FUNIL",     18, None),
    (15, "FASE",              16, None),
    (16, "TIPO DO CONTATO",   26, None),
    (17, "TEMPERATURA",       14, None),
    (18, "TENTATIVA",         12, None),
    (19, "GRUPO DASH",        14, None),
    (20, "FOLLOW-UP",         14, FMT_DATE),
    (21, "AÇÃO FUTURA",       16, None),
    (22, "AÇÃO DETALHADA",    30, None),
    (23, "SINALEIRO CICLO",   14, None),
    (24, "SINALEIRO META",    14, None),
]


def build_log(wb):
    """Build LOG tab (24 cols, append-only archive, no formulas)."""
    ws = wb.create_sheet("LOG")
    ws.sheet_properties.tabColor = TAB_LOG

    # ── Row 1: Title ──
    ws.merge_cells('A1:X1')
    title = ws.cell(row=1, column=1,
                    value="LOG — Arquivo de Atendimentos (leitura, append-only)")
    style_cell(title, font=FONT_TITLE_W, fill=FILL_DARK, align=ALIGN_LEFT)
    for c in range(1, 25):
        ws.cell(row=1, column=c).fill = FILL_DARK
        ws.cell(row=1, column=c).border = THIN_BORDER

    # ── Row 2: Headers ──
    for col, header, width, fmt in LOG_COLS:
        # Manual cols = blue header, auto cols = dark gray
        fill = FILL_HEADER if col <= 12 else PatternFill('solid', fgColor='595959')
        write_header(ws, 2, col, header, fill=fill)
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Conditional formatting for SITUAÇÃO column (M) ──
    from openpyxl.formatting.rule import CellIsRule
    for sit, fill in SITUACAO_FILLS.items():
        ws.conditional_formatting.add(
            'M3:M50000',
            CellIsRule(operator='equal', formula=[f'"{sit}"'], fill=fill)
        )

    # ── Freeze ──
    ws.freeze_panes = "A3"

    print(f"  LOG: 24 cols (archive, no formulas), freeze A3")
    return ws


def build_claude_log(wb):
    """Build Claude Log tab (6-col audit trail)."""
    ws = wb.create_sheet("Claude Log")
    ws.sheet_properties.tabColor = TAB_CLAUDE

    # ── Row 1: Title ──
    ws.merge_cells('A1:F1')
    title = ws.cell(row=1, column=1,
                    value="CLAUDE LOG — Registro de Ações Automáticas")
    style_cell(title, font=FONT_TITLE_W,
               fill=PatternFill('solid', fgColor='9333EA'), align=ALIGN_LEFT)
    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor='9333EA')
        ws.cell(row=1, column=c).border = THIN_BORDER

    # ── Row 2: Headers (6 columns) ──
    headers = [
        (1, "TURN #",          8,  None),
        (2, "DATE",            18, FMT_DATE),
        (3, "USER REQUEST",    40, None),
        (4, "ACTION TAKEN",    35, None),
        (5, "DETAILS",         50, None),
        (6, "OUTCOME",         25, None),
    ]
    for col, header, width, fmt in headers:
        write_header(ws, 2, col, header, fill=PatternFill('solid', fgColor='7B2FF2'))
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Example entry ──
    now = datetime.datetime(2026, 2, 9, 10, 0, 0)
    write_data(ws, 3, 1, 1, align=ALIGN_CENTER)
    write_data(ws, 3, 2, now, fmt='DD/MM/YYYY HH:MM')
    write_data(ws, 3, 3, "Build CRM V3", align=ALIGN_LEFT)
    write_data(ws, 3, 4, "BUILD V3", align=ALIGN_LEFT)
    write_data(ws, 3, 5, "CRM VITAO360 V3 gerado com sucesso — 9 tabs, motor de regras ativo", align=ALIGN_LEFT)
    write_data(ws, 3, 6, "SUCCESS", align=ALIGN_LEFT)

    # ── Freeze ──
    ws.freeze_panes = "A3"

    print(f"  Claude Log: 6 cols (Turn#, Date, Request, Action, Details, Outcome), freeze A3")
    return ws

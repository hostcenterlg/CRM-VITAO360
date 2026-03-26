"""Script de inspeção das abas LOG e REDES v2."""
import openpyxl
import os

XLSX_PATH = r"C:\Users\User\OneDrive\Área de Trabalho\CRM_VITAO360  INTELIGENTE   FINAL OK .xlsx"

print("Verificando path:", XLSX_PATH)
print("Existe?", os.path.exists(XLSX_PATH))
print()

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)

for aba in ['LARISSA', 'MANU', 'JULIO', 'DAIANE']:
    ws = wb[aba]
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    print('=== ' + aba + ' ===')
    if rows:
        headers = rows[0]
        print('  Total colunas: ' + str(len(headers)))
        for i, h in enumerate(headers, 1):
            col_letter = chr(64+i) if i <= 26 else str(i)
            print('  Col ' + str(i) + ' (' + col_letter + '): ' + repr(h))
    if len(rows) > 1:
        row2 = rows[1]
        print('  Amostra row 2: ' + str(row2[:15]))
    print()

# REDES v2
ws_redes = wb['REDES v2']
print('=== REDES v2 ===')
rows_redes = list(ws_redes.iter_rows(min_row=1, max_row=8, values_only=True))
for i, row in enumerate(rows_redes, 1):
    print('  Row ' + str(i) + ': ' + str(row[:13]))

wb.close()

"""
CRM VITAO360 — fixtures de XLSX para testes dos parsers DDE/AC.

Gera arquivos XLSX mínimos programaticamente via openpyxl.
Uso: importar as funções create_* em conftest.py ou diretamente nos testes.

Cada fixture cria um arquivo .xlsx temporário com dados mínimos válidos
para smoke-test do parser correspondente.

CNPJ de teste: 11222333000181 (14 dígitos, fictício para testes)
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

# CNPJ fictício para fixtures de teste
CNPJ_TESTE = "11222333000181"
ANO_TESTE = 2025
MES_TESTE = 3


def create_zsdfat_fixture(tmp_path: Path) -> Path:
    """
    Cria ZSDFAT_teste.xlsx com 1 aba contendo linhas DRE mínimas.

    Aba: "CNPJ: 11222333000181"
    Linha 1: CNPJ no cabeçalho
    Linha 2: header de períodos
    Linha 3+: contas DRE
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cliente Teste"

    # Linha de CNPJ (detectada pelo parser)
    ws.cell(row=1, column=1, value=f"CNPJ: {CNPJ_TESTE}")

    # Linha de header de período
    ws.cell(row=2, column=1, value="Conta DRE")
    ws.cell(row=2, column=2, value="Mar/2025")
    ws.cell(row=2, column=3, value="Abr/2025")

    # Linhas de dados DRE
    dados_dre = [
        ("Faturamento Bruto a Tabela", 100000.00, 110000.00),
        ("(-) IPI sobre Vendas", 5000.00, 5500.00),
        ("(-) Devolucoes", 1000.00, 1100.00),
        ("(-) Desconto Comercial", 3000.00, 3300.00),
        ("(-) Desconto Financeiro", 500.00, 550.00),
        ("(-) Bonificacoes", 2000.00, 2200.00),
        ("ICMS Proprio", 12000.00, 13200.00),
        ("PIS", 1650.00, 1815.00),
        ("COFINS", 7600.00, 8360.00),
        ("ICMS-ST Retido", 0.00, 0.00),
        ("Receita Liquida", 67250.00, 73975.00),
        ("CMV - Custo Mercadoria Vendida", 45000.00, 49500.00),
        ("= Margem Bruta", 22250.00, 24475.00),
        ("Frete CT-e", 2500.00, 2750.00),
        ("Comissao Representante", 3000.00, 3300.00),
        ("Verba Cliente", 1500.00, 1650.00),
        ("Promotor PDV Agencia", 800.00, 880.00),
        ("= Margem de Contribuicao", 14450.00, 15895.00),
        # Linha RAW intencional para testar branch SINTETICO
        ("Outros Custos Especiais XYZ", 100.00, 110.00),
    ]

    for row_idx, (conta, val_mar, val_abr) in enumerate(dados_dre, start=3):
        ws.cell(row=row_idx, column=1, value=conta)
        ws.cell(row=row_idx, column=2, value=val_mar)
        ws.cell(row=row_idx, column=3, value=val_abr)

    path = tmp_path / "ZSDFAT_teste.xlsx"
    wb.save(str(path))
    return path


def create_verbas_fixture(tmp_path: Path) -> Path:
    """Cria Verbas_2025.xlsx com 3 linhas de verbas efetivadas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verbas"

    # Header
    ws.cell(row=1, column=1, value="CNPJ")
    ws.cell(row=1, column=2, value="Ano")
    ws.cell(row=1, column=3, value="Valor")
    ws.cell(row=1, column=4, value="Desc%")

    # Dados
    dados = [
        (CNPJ_TESTE, ANO_TESTE, 5000.00, 3.5),
        ("22333444000155", ANO_TESTE, 8000.00, 5.0),
        ("33444555000166", ANO_TESTE, 3000.00, 2.0),
    ]
    for row_idx, (cnpj, ano, valor, desc) in enumerate(dados, start=2):
        ws.cell(row=row_idx, column=1, value=cnpj)
        ws.cell(row=row_idx, column=2, value=ano)
        ws.cell(row=row_idx, column=3, value=valor)
        ws.cell(row=row_idx, column=4, value=desc)

    path = tmp_path / "Verbas_2025.xlsx"
    wb.save(str(path))
    return path


def create_frete_fixture(tmp_path: Path) -> Path:
    """Cria Frete por Cliente.xlsx com layout tabular simples."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Frete"

    # Header
    ws.cell(row=1, column=1, value="CNPJ")
    ws.cell(row=1, column=2, value="Ano")
    ws.cell(row=1, column=3, value="Mes")
    ws.cell(row=1, column=4, value="Qtd CTs")
    ws.cell(row=1, column=5, value="Valor Frete")

    dados = [
        (CNPJ_TESTE, ANO_TESTE, 1, 5, 1200.50),
        (CNPJ_TESTE, ANO_TESTE, 2, 6, 1350.00),
        ("22333444000155", ANO_TESTE, 1, 3, 800.00),
    ]
    for row_idx, (cnpj, ano, mes, qtd, valor) in enumerate(dados, start=2):
        ws.cell(row=row_idx, column=1, value=cnpj)
        ws.cell(row=row_idx, column=2, value=ano)
        ws.cell(row=row_idx, column=3, value=mes)
        ws.cell(row=row_idx, column=4, value=qtd)
        ws.cell(row=row_idx, column=5, value=valor)

    path = tmp_path / "Frete por Cliente.xlsx"
    wb.save(str(path))
    return path


def create_contratos_fixture(tmp_path: Path) -> Path:
    """Cria Controle Contratos.xlsx com aba Desc. Financeiro."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Desc. Financeiro"

    # Header
    ws.cell(row=1, column=1, value="CNPJ")
    ws.cell(row=1, column=2, value="Ano")
    ws.cell(row=1, column=3, value="Valor Desc Financeiro")
    ws.cell(row=1, column=4, value="Desc%")
    ws.cell(row=1, column=5, value="Inicio Vigencia")
    ws.cell(row=1, column=6, value="Fim Vigencia")

    dados = [
        (CNPJ_TESTE, ANO_TESTE, 10000.00, 4.5, "01/01/2025", "31/12/2025"),
        ("22333444000155", ANO_TESTE, 15000.00, 6.0, "01/01/2025", "31/12/2025"),
    ]
    for row_idx, (cnpj, ano, valor, desc, inicio, fim) in enumerate(dados, start=2):
        ws.cell(row=row_idx, column=1, value=cnpj)
        ws.cell(row=row_idx, column=2, value=ano)
        ws.cell(row=row_idx, column=3, value=valor)
        ws.cell(row=row_idx, column=4, value=desc)
        ws.cell(row=row_idx, column=5, value=inicio)
        ws.cell(row=row_idx, column=6, value=fim)

    path = tmp_path / "Controle Contratos.xlsx"
    wb.save(str(path))
    return path


def create_promotores_fixture(tmp_path: Path) -> Path:
    """Cria Despesas Clientes V2.xlsx com aba RESUMO."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RESUMO"

    # Header
    ws.cell(row=1, column=1, value="CNPJ")
    ws.cell(row=1, column=2, value="Agencia")
    ws.cell(row=1, column=3, value="Ano")
    ws.cell(row=1, column=4, value="Mes")
    ws.cell(row=1, column=5, value="Valor Despesa")

    dados = [
        (CNPJ_TESTE, "Agencia ABC", ANO_TESTE, 1, 2500.00),
        (CNPJ_TESTE, "Agencia ABC", ANO_TESTE, 2, 2700.00),
        ("22333444000155", "Agencia XYZ", ANO_TESTE, 1, 1800.00),
    ]
    for row_idx, (cnpj, agencia, ano, mes, valor) in enumerate(dados, start=2):
        ws.cell(row=row_idx, column=1, value=cnpj)
        ws.cell(row=row_idx, column=2, value=agencia)
        ws.cell(row=row_idx, column=3, value=ano)
        ws.cell(row=row_idx, column=4, value=mes)
        ws.cell(row=row_idx, column=5, value=valor)

    path = tmp_path / "Despesas Clientes V2.xlsx"
    wb.save(str(path))
    return path

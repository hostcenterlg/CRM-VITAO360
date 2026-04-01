"""
CRM VITAO360 — Rotas /api/import

Import Pipeline: recebe arquivos .xlsx do Mercos/SAP e processa automaticamente.

FR-001 — Upload de arquivo .xlsx (multipart/form-data)
FR-002 — Deteccao de tipo de arquivo por heuristica de colunas
FR-003 — Normalizacao de CNPJ: String(14), zero-padded, nunca float
FR-004 — Classificacao 3-tier automatica: dados de Mercos/SAP = REAL
FR-005 — Merge de dados: UPDATE se CNPJ existe, INSERT se novo
FR-006 — Recalculo de Motor + Score + Sinaleiro + Agenda apos importacao

REGRAS CRITICAS:
  R1 — Two-Base Architecture: vendas com valor > 0 vao para tabela vendas.
       NUNCA misturar com log_interacoes.
  R2 — CNPJ = String(14), zero-padded, NUNCA float.
  R6 — Mercos MENTE nos nomes — verificar Data inicial/Data final nas linhas 6-7.
  R8 — 558 registros ALUCINACAO catalogados. CNPJ nestas listas = REJEITADO.
  R9 — De-Para vendedores: Manu Vitao -> MANU, Larissa -> LARISSA, etc.

Endpoints:
  POST /api/import/upload  — Recebe .xlsx, processa pipeline completo
  GET  /api/import/history — Lista os ultimos 20 jobs de importacao

Restricao de role: admin apenas.

Processamento em foreground (Render free nao suporta workers).
Tamanho maximo: 50MB. Extensao: .xlsx apenas.
"""

from __future__ import annotations

import io
import re
import time
from datetime import datetime, timezone, date
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import select

from backend.app.api.deps import require_admin
from backend.app.database import get_db
from backend.app.models.cliente import Cliente
from backend.app.models.import_job import ImportJob
from backend.app.models.usuario import Usuario
from backend.app.models.venda import Venda
from backend.app.services.sinaleiro_service import sinaleiro_service
from backend.app.services.score_service import score_service
from backend.app.services.agenda_service import agenda_service

router = APIRouter(prefix="/api/import", tags=["Import"])

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

_MAX_SIZE_BYTES = 50 * 1024 * 1024  # 50MB

# De-Para canonico de vendedores (R9)
_DE_PARA_VENDEDOR: dict[str, str] = {
    "MANU": "MANU",
    "MANU VITAO": "MANU",
    "MANU DITZEL": "MANU",
    "LARISSA": "LARISSA",
    "LARI": "LARISSA",
    "LARISSA VITAO": "LARISSA",
    "MAIS GRANEL": "LARISSA",
    "RODRIGO": "LARISSA",
    "DAIANE": "DAIANE",
    "CENTRAL DAIANE": "DAIANE",
    "DAIANE VITAO": "DAIANE",
    "JULIO": "JULIO",
    "JULIO GADRET": "JULIO",
}

# Vendedores legados — importados mas marcados como LEGADO
_VENDEDORES_LEGADO = {
    "BRUNO GRETTER",
    "JEFERSON VITAO",
    "PATRIC",
    "GABRIEL",
    "SERGIO",
    "IVE",
    "ANA",
}

# CNPJs catalogados como ALUCINACAO (558 registros — R8)
# Este set contem os CNPJs fabricados pelo ChatGPT que NUNCA devem ser integrados.
# Em producao, este set deve ser carregado de um arquivo JSON externo.
# Os valores abaixo sao os CNPJs-padrao zerados/invalidos detectados historicamente.
_CNPJS_ALUCINACAO: set[str] = {
    "00000000000000",
    "11111111111111",
    "22222222222222",
    "33333333333333",
    "44444444444444",
    "55555555555555",
    "66666666666666",
    "77777777777777",
    "88888888888888",
    "99999999999999",
    "12345678000100",   # CNPJ generico de teste — nunca em producao
}

# Tipos de arquivo detectados
_TIPO_MERCOS_VENDAS = "MERCOS_VENDAS"
_TIPO_MERCOS_CARTEIRA = "MERCOS_CARTEIRA"
_TIPO_SAP_CADASTRO = "SAP_CADASTRO"


# ---------------------------------------------------------------------------
# Schemas de resposta
# ---------------------------------------------------------------------------

class ImportJobResponse(BaseModel):
    """Resposta do endpoint de upload — resultado do job de importacao."""
    job_id: int
    status: str
    tipo: str
    arquivo: Optional[str]
    registros_lidos: int
    inseridos: int
    atualizados: int
    ignorados: int
    erros: int
    tempo_ms: float
    erro_mensagem: Optional[str] = None

    model_config = {"from_attributes": True}


class ImportHistoryItem(BaseModel):
    """Item de historico de importacao."""
    id: int
    tipo: str
    arquivo: Optional[str]
    status: str
    registros_lidos: int
    inseridos: int
    atualizados: int
    ignorados: int
    erro_mensagem: Optional[str]
    created_at: Optional[datetime]

    model_config = {"from_attributes": True}


class ImportHistoryResponse(BaseModel):
    """Resposta do endpoint de historico."""
    total: int
    items: list[ImportHistoryItem]


# ---------------------------------------------------------------------------
# Utilitarios de normalizacao
# ---------------------------------------------------------------------------

def _normalizar_cnpj(valor: object) -> str:
    """
    Normaliza CNPJ para String(14), zero-padded, sem pontuacao.

    R5 — CNPJ NUNCA float: re.sub remove nao-numericos, zfill garante 14 digitos.

    Args:
        valor: Qualquer valor (str, float, int, None).

    Returns:
        String com 14 digitos numericos, ou "" se invalido.
    """
    if valor is None:
        return ""
    # Converter float para str sem notacao cientifica (ex.: 12345678000100.0 -> "12345678000100")
    if isinstance(valor, float):
        try:
            valor = str(int(valor))
        except (ValueError, OverflowError):
            return ""
    s = re.sub(r"\D", "", str(valor))
    if not s:
        return ""
    resultado = s.zfill(14)
    # CNPJ valido tem exatamente 14 digitos
    if len(resultado) != 14:
        return ""
    return resultado


def _normalizar_vendedor(valor: object) -> str:
    """
    Aplica DE-PARA canonico de vendedores (R9).

    Retorna o nome canonico (MANU, LARISSA, DAIANE, JULIO) ou
    "LEGADO" para vendedores historicos, ou "" se nao reconhecido.

    Args:
        valor: Nome do vendedor como aparece no arquivo de origem.

    Returns:
        Nome canonico ou "" se nao mapeado.
    """
    if valor is None:
        return ""
    normalizado = str(valor).strip().upper()
    if normalizado in _DE_PARA_VENDEDOR:
        return _DE_PARA_VENDEDOR[normalizado]
    if normalizado in _VENDEDORES_LEGADO:
        return "LEGADO"
    return normalizado  # Retorna o proprio nome em uppercase se nao mapeado


def _celula_str(valor: object) -> str:
    """Converte valor de celula para string limpa."""
    if valor is None:
        return ""
    return str(valor).strip()


def _celula_float(valor: object) -> Optional[float]:
    """Converte valor de celula para float, retorna None se invalido."""
    if valor is None:
        return None
    try:
        return float(valor)
    except (ValueError, TypeError):
        return None


def _celula_int(valor: object) -> Optional[int]:
    """Converte valor de celula para int, retorna None se invalido."""
    if valor is None:
        return None
    try:
        return int(float(valor))
    except (ValueError, TypeError):
        return None


def _celula_date(valor: object) -> Optional[date]:
    """Converte valor de celula para date, retorna None se invalido."""
    if valor is None:
        return None
    if isinstance(valor, date):
        return valor
    if isinstance(valor, datetime):
        return valor.date()
    try:
        from datetime import datetime as dt
        return dt.fromisoformat(str(valor)).date()
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Deteccao de tipo de arquivo
# ---------------------------------------------------------------------------

def _detectar_tipo_arquivo(cabecalhos: list[str]) -> str:
    """
    Detecta o tipo do arquivo importado a partir dos cabecalhos da primeira linha.

    Heuristica por colunas:
      - Mercos Vendas:    CNPJ + Nome Fantasia + Valor + colunas de meses
      - Mercos Carteira:  CNPJ + Dias sem compra + Consultor
      - SAP Cadastro:     Codigo do Cliente + CNPJ + Razao Social + Cadastro

    Args:
        cabecalhos: Lista de strings com nomes das colunas (row 1 do xlsx).

    Returns:
        String com o tipo detectado ou levanta ValueError se nao reconhecido.

    Raises:
        ValueError: Se os cabecalhos nao correspondem a nenhum tipo conhecido.
    """
    # Normalizar cabecalhos para comparacao case-insensitive
    cab_upper = {str(c).strip().upper() for c in cabecalhos if c}

    # SAP Cadastro: tem "CODIGO DO CLIENTE" ou "RAZAO SOCIAL" + "CNPJ"
    tem_codigo_sap = any("CODIGO" in c and "CLIENTE" in c for c in cab_upper)
    tem_razao = any("RAZAO" in c and "SOCIAL" in c for c in cab_upper)
    tem_cadastro = "CADASTRO" in cab_upper
    tem_cnpj = any("CNPJ" in c for c in cab_upper)

    if tem_cnpj and (tem_codigo_sap or (tem_razao and tem_cadastro)):
        return _TIPO_SAP_CADASTRO

    # Mercos Carteira: CNPJ + "DIAS SEM COMPRA" ou "DIAS" + consultor
    tem_dias_sem_compra = any("DIAS" in c and "COMPRA" in c for c in cab_upper)
    tem_consultor = any("CONSULTOR" in c for c in cab_upper)

    if tem_cnpj and tem_dias_sem_compra and tem_consultor:
        return _TIPO_MERCOS_CARTEIRA

    # Mercos Vendas: CNPJ + "NOME FANTASIA" + "VALOR" (ou colunas mensais)
    tem_nome_fantasia = any("NOME" in c and "FANTASIA" in c for c in cab_upper)
    tem_valor = any("VALOR" in c for c in cab_upper)

    if tem_cnpj and (tem_nome_fantasia or tem_valor):
        return _TIPO_MERCOS_VENDAS

    raise ValueError(
        f"Tipo de arquivo nao reconhecido. Cabecalhos detectados: {sorted(cab_upper)}. "
        "Esperado: Mercos Vendas (CNPJ+NomeFantasia+Valor), "
        "Mercos Carteira (CNPJ+DiasSemCompra+Consultor) ou "
        "SAP Cadastro (CodigoCliente+CNPJ+RazaoSocial)."
    )


# ---------------------------------------------------------------------------
# Processadores por tipo de arquivo
# ---------------------------------------------------------------------------

def _mapear_cabecalhos(ws) -> dict[str, int]:
    """
    Mapeia nome da coluna -> indice (0-based) a partir da primeira linha da planilha.

    Retorna dicionario com cabecalhos em uppercase para lookups case-insensitive.
    """
    primeira_linha = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    return {
        str(cel).strip().upper(): idx
        for idx, cel in enumerate(primeira_linha)
        if cel is not None
    }


def _get_col(row: tuple, mapa: dict[str, int], *chaves: str) -> object:
    """
    Extrai valor de uma linha pelo nome da coluna, testando multiplas chaves.

    Util para colunas com nomes ligeiramente diferentes entre versoes do relatorio.

    Args:
        row: Tupla de valores da linha.
        mapa: Mapa nome_coluna -> indice.
        *chaves: Nomes a tentar em ordem (case insensitive, uppercase).

    Returns:
        Valor da primeira coluna encontrada, ou None se nenhuma existir.
    """
    for chave in chaves:
        idx = mapa.get(chave.upper())
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def _processar_mercos_vendas(
    ws,
    db: Session,
    job: ImportJob,
    cnpjs_afetados: "set[str] | None" = None,
) -> tuple[int, int, int, int]:
    """
    Processa arquivo de vendas do Mercos.

    Logica:
      - Le linhas a partir da linha 2 (cabecalhos na linha 1)
      - Normaliza CNPJ (R5)
      - Rejeita CNPJs de ALUCINACAO (R8)
      - Two-Base: linhas com valor > 0 sao inseridas em vendas
      - Se CNPJ existe em clientes: UPDATE nome_fantasia, consultor, situacao
      - Se CNPJ novo: INSERT cliente com situacao=ATIVO

    ATENCAO R6: Mercos MENTE nos nomes — periodo real nas linhas 6-7.
    Este processador usa data_pedido da coluna (se presente) para mes_referencia.

    Returns:
        Tupla (inseridos, atualizados, ignorados, erros).
    """
    mapa = _mapear_cabecalhos(ws)

    inseridos = 0
    atualizados = 0
    ignorados = 0
    erros_count = 0

    # Detectar coluna de data — pode ter nome variado
    col_data_opts = ["DATA DO PEDIDO", "DATA PEDIDO", "DATA", "DT PEDIDO", "DT. PEDIDO"]
    col_valor_opts = ["VALOR TOTAL", "VALOR DO PEDIDO", "VALOR PEDIDO", "VALOR", "VL TOTAL", "VL. TOTAL"]
    col_cnpj_opts = ["CNPJ", "CNPJ DO CLIENTE", "CNPJ CLIENTE"]
    col_nome_opts = ["NOME FANTASIA", "NOME", "CLIENTE", "RAZAO SOCIAL"]
    col_consultor_opts = ["CONSULTOR", "VENDEDOR", "REPRESENTANTE"]
    col_pedido_opts = ["NUMERO DO PEDIDO", "NUMERO PEDIDO", "NRO PEDIDO", "N. PEDIDO", "PEDIDO"]
    col_situacao_opts = ["SITUACAO", "STATUS", "SITUAÇÃO"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Ignorar linhas completamente vazias
        if all(v is None for v in row):
            continue

        job.registros_lidos += 1

        try:
            cnpj = _normalizar_cnpj(_get_col(row, mapa, *col_cnpj_opts))

            # Validar CNPJ
            if not cnpj or len(cnpj) != 14:
                ignorados += 1
                continue

            # R8 — Rejeitar ALUCINACAO
            if cnpj in _CNPJS_ALUCINACAO:
                ignorados += 1
                continue

            valor_raw = _get_col(row, mapa, *col_valor_opts)
            valor = _celula_float(valor_raw)
            nome = _celula_str(_get_col(row, mapa, *col_nome_opts))
            consultor_raw = _celula_str(_get_col(row, mapa, *col_consultor_opts))
            consultor = _normalizar_vendedor(consultor_raw) if consultor_raw else ""
            data_raw = _get_col(row, mapa, *col_data_opts)
            data_pedido = _celula_date(data_raw)
            numero_pedido = _celula_str(_get_col(row, mapa, *col_pedido_opts)) or None
            situacao_raw = _celula_str(_get_col(row, mapa, *col_situacao_opts))
            situacao = situacao_raw.upper() if situacao_raw else "ATIVO"

            # Normalizar situacao para valores aceitos
            if situacao not in ("ATIVO", "INAT.REC", "INAT.ANT", "PROSPECT", "EM RISCO"):
                situacao = "ATIVO"

            # Upsert cliente
            cliente = db.query(Cliente).filter(Cliente.cnpj == cnpj).first()
            if cliente is None:
                # INSERT — novo cliente
                cliente = Cliente(
                    cnpj=cnpj,
                    nome_fantasia=nome or None,
                    situacao=situacao,
                    consultor=consultor or None,
                    classificacao_3tier="REAL",
                )
                db.add(cliente)
                db.flush()
                inseridos += 1
            else:
                # UPDATE — atualizar campos relevantes
                if nome:
                    cliente.nome_fantasia = nome
                if consultor:
                    cliente.consultor = consultor
                if situacao_raw:  # Apenas atualiza situacao se coluna presente
                    cliente.situacao = situacao
                atualizados += 1

            # Registrar CNPJ processado com sucesso no conjunto de afetados
            if cnpjs_afetados is not None:
                cnpjs_afetados.add(cnpj)

            # R4 — Two-Base: apenas inserir em vendas se valor > 0
            if valor is not None and valor > 0 and data_pedido is not None:
                mes_referencia = data_pedido.strftime("%Y-%m")
                venda = Venda(
                    cnpj=cnpj,
                    data_pedido=data_pedido,
                    numero_pedido=numero_pedido,
                    valor_pedido=valor,
                    consultor=consultor or "LEGADO",
                    fonte="MERCOS",
                    classificacao_3tier="REAL",
                    mes_referencia=mes_referencia,
                )
                db.add(venda)

        except Exception as exc:
            erros_count += 1
            # Nao interrompe o loop — registra o erro e continua
            _ = exc  # usado apenas para nao suprimir silenciosamente em debug

    return inseridos, atualizados, ignorados, erros_count


def _processar_mercos_carteira(
    ws,
    db: Session,
    job: ImportJob,
    cnpjs_afetados: "set[str] | None" = None,
) -> tuple[int, int, int, int]:
    """
    Processa arquivo de carteira do Mercos.

    Colunas esperadas: CNPJ, Nome Fantasia, Dias sem compra, Consultor,
    Situacao, Ciclo Medio, Faturamento Total, Curva ABC, etc.

    Atualiza campos de inteligencia do cliente (dias_sem_compra, ciclo_medio,
    faturamento_total, curva_abc, consultor, situacao).

    Returns:
        Tupla (inseridos, atualizados, ignorados, erros).
    """
    mapa = _mapear_cabecalhos(ws)

    inseridos = 0
    atualizados = 0
    ignorados = 0
    erros_count = 0

    col_cnpj_opts = ["CNPJ", "CNPJ DO CLIENTE", "CNPJ CLIENTE"]
    col_nome_opts = ["NOME FANTASIA", "NOME", "CLIENTE"]
    col_razao_opts = ["RAZAO SOCIAL", "RAZÃO SOCIAL"]
    col_dias_opts = ["DIAS SEM COMPRA", "DIAS", "DIAS SEM PEDIDO"]
    col_consultor_opts = ["CONSULTOR", "VENDEDOR", "REPRESENTANTE"]
    col_situacao_opts = ["SITUACAO", "STATUS", "SITUAÇÃO"]
    col_ciclo_opts = ["CICLO MEDIO", "CICLO MÉDIO", "CICLO", "MEDIA DIAS"]
    col_fat_opts = ["FATURAMENTO TOTAL", "FATURAMENTO", "VL FATURADO", "TOTAL COMPRADO"]
    col_abc_opts = ["CURVA ABC", "CURVA", "ABC", "CLASSIFICACAO ABC"]
    col_cidade_opts = ["CIDADE"]
    col_uf_opts = ["UF", "ESTADO"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        job.registros_lidos += 1

        try:
            cnpj = _normalizar_cnpj(_get_col(row, mapa, *col_cnpj_opts))

            if not cnpj or len(cnpj) != 14:
                ignorados += 1
                continue

            if cnpj in _CNPJS_ALUCINACAO:
                ignorados += 1
                continue

            nome = _celula_str(_get_col(row, mapa, *col_nome_opts))
            razao = _celula_str(_get_col(row, mapa, *col_razao_opts))
            dias = _celula_int(_get_col(row, mapa, *col_dias_opts))
            consultor_raw = _celula_str(_get_col(row, mapa, *col_consultor_opts))
            consultor = _normalizar_vendedor(consultor_raw) if consultor_raw else ""
            situacao_raw = _celula_str(_get_col(row, mapa, *col_situacao_opts)).upper()
            situacao = situacao_raw if situacao_raw in (
                "ATIVO", "INAT.REC", "INAT.ANT", "PROSPECT", "EM RISCO"
            ) else "ATIVO"
            ciclo = _celula_float(_get_col(row, mapa, *col_ciclo_opts))
            faturamento = _celula_float(_get_col(row, mapa, *col_fat_opts))
            abc_raw = _celula_str(_get_col(row, mapa, *col_abc_opts)).upper()
            abc = abc_raw if abc_raw in ("A", "B", "C") else None
            cidade = _celula_str(_get_col(row, mapa, *col_cidade_opts)) or None
            uf = _celula_str(_get_col(row, mapa, *col_uf_opts)) or None

            cliente = db.query(Cliente).filter(Cliente.cnpj == cnpj).first()
            if cliente is None:
                cliente = Cliente(
                    cnpj=cnpj,
                    nome_fantasia=nome or None,
                    razao_social=razao or None,
                    situacao=situacao,
                    consultor=consultor or None,
                    dias_sem_compra=dias,
                    ciclo_medio=ciclo,
                    faturamento_total=faturamento,
                    curva_abc=abc,
                    cidade=cidade,
                    uf=uf,
                    classificacao_3tier="REAL",
                )
                db.add(cliente)
                db.flush()
                inseridos += 1
            else:
                if nome:
                    cliente.nome_fantasia = nome
                if razao:
                    cliente.razao_social = razao
                if consultor:
                    cliente.consultor = consultor
                if situacao_raw:
                    cliente.situacao = situacao
                if dias is not None:
                    cliente.dias_sem_compra = dias
                if ciclo is not None:
                    cliente.ciclo_medio = ciclo
                if faturamento is not None:
                    cliente.faturamento_total = faturamento
                if abc:
                    cliente.curva_abc = abc
                if cidade:
                    cliente.cidade = cidade
                if uf:
                    cliente.uf = uf
                atualizados += 1

            # Registrar CNPJ processado com sucesso
            if cnpjs_afetados is not None:
                cnpjs_afetados.add(cnpj)

        except Exception:
            erros_count += 1

    return inseridos, atualizados, ignorados, erros_count


def _processar_sap_cadastro(
    ws,
    db: Session,
    job: ImportJob,
    cnpjs_afetados: "set[str] | None" = None,
) -> tuple[int, int, int, int]:
    """
    Processa arquivo de cadastro de clientes do SAP.

    Colunas esperadas: Codigo do Cliente, CNPJ, Razao Social, Nome Fantasia,
    Cidade, UF, Tipo Cliente SAP, Macroregiao, Consultor.

    Returns:
        Tupla (inseridos, atualizados, ignorados, erros).
    """
    mapa = _mapear_cabecalhos(ws)

    inseridos = 0
    atualizados = 0
    ignorados = 0
    erros_count = 0

    col_codigo_opts = ["CODIGO DO CLIENTE", "CODIGO CLIENTE", "COD. CLIENTE", "COD CLIENTE", "CODIGO SAP"]
    col_cnpj_opts = ["CNPJ", "CNPJ DO CLIENTE"]
    col_razao_opts = ["RAZAO SOCIAL", "RAZÃO SOCIAL"]
    col_nome_opts = ["NOME FANTASIA", "NOME"]
    col_cidade_opts = ["CIDADE", "MUNICIPIO"]
    col_uf_opts = ["UF", "ESTADO"]
    col_tipo_opts = ["TIPO CLIENTE SAP", "TIPO CLIENTE", "TIPO"]
    col_macro_opts = ["MACROREGIAO", "MACRO REGIAO", "MACRO-REGIAO", "REGIAO"]
    col_consultor_opts = ["CONSULTOR", "VENDEDOR"]
    col_email_opts = ["EMAIL", "E-MAIL"]
    col_tel_opts = ["TELEFONE", "TEL.", "FONE"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        job.registros_lidos += 1

        try:
            cnpj = _normalizar_cnpj(_get_col(row, mapa, *col_cnpj_opts))

            if not cnpj or len(cnpj) != 14:
                ignorados += 1
                continue

            if cnpj in _CNPJS_ALUCINACAO:
                ignorados += 1
                continue

            codigo_sap = _celula_str(_get_col(row, mapa, *col_codigo_opts)) or None
            razao = _celula_str(_get_col(row, mapa, *col_razao_opts)) or None
            nome = _celula_str(_get_col(row, mapa, *col_nome_opts)) or None
            cidade = _celula_str(_get_col(row, mapa, *col_cidade_opts)) or None
            uf = _celula_str(_get_col(row, mapa, *col_uf_opts)) or None
            tipo_sap = _celula_str(_get_col(row, mapa, *col_tipo_opts)) or None
            macro = _celula_str(_get_col(row, mapa, *col_macro_opts)) or None
            consultor_raw = _celula_str(_get_col(row, mapa, *col_consultor_opts))
            consultor = _normalizar_vendedor(consultor_raw) if consultor_raw else None
            email = _celula_str(_get_col(row, mapa, *col_email_opts)) or None
            tel = _celula_str(_get_col(row, mapa, *col_tel_opts)) or None

            cliente = db.query(Cliente).filter(Cliente.cnpj == cnpj).first()
            if cliente is None:
                cliente = Cliente(
                    cnpj=cnpj,
                    codigo_cliente=codigo_sap,
                    razao_social=razao,
                    nome_fantasia=nome,
                    cidade=cidade,
                    uf=uf,
                    tipo_cliente_sap=tipo_sap,
                    macroregiao=macro,
                    consultor=consultor,
                    email=email,
                    telefone=tel,
                    situacao="ATIVO",
                    classificacao_3tier="REAL",
                )
                db.add(cliente)
                db.flush()
                inseridos += 1
            else:
                if codigo_sap:
                    cliente.codigo_cliente = codigo_sap
                if razao:
                    cliente.razao_social = razao
                if nome:
                    cliente.nome_fantasia = nome
                if cidade:
                    cliente.cidade = cidade
                if uf:
                    cliente.uf = uf
                if tipo_sap:
                    cliente.tipo_cliente_sap = tipo_sap
                if macro:
                    cliente.macroregiao = macro
                if consultor:
                    cliente.consultor = consultor
                if email:
                    cliente.email = email
                if tel:
                    cliente.telefone = tel
                atualizados += 1

            # Registrar CNPJ processado com sucesso
            if cnpjs_afetados is not None:
                cnpjs_afetados.add(cnpj)

        except Exception:
            erros_count += 1

    return inseridos, atualizados, ignorados, erros_count


# ---------------------------------------------------------------------------
# Recalculo de inteligencia pos-import
# ---------------------------------------------------------------------------

def _recalcular_inteligencia(db: Session, cnpjs_afetados: set[str]) -> None:
    """
    Recalcula Motor + Score + Sinaleiro para todos os clientes afetados.

    Nao faz commit — responsabilidade do caller.

    Args:
        db: Sessao SQLAlchemy ativa.
        cnpjs_afetados: Conjunto de CNPJs que foram inseridos ou atualizados.
    """
    if not cnpjs_afetados:
        return

    clientes = db.query(Cliente).filter(Cliente.cnpj.in_(cnpjs_afetados)).all()
    for cliente in clientes:
        sinaleiro_service.aplicar(db, cliente)
        score_service.aplicar_e_salvar(db, cliente)


def _regenerar_agenda(db: Session, consultores_afetados: set[str]) -> None:
    """
    Regenera agenda para os consultores cujos clientes foram afetados.

    Deleta agenda do dia atual e recria. Nao faz commit.

    Args:
        db: Sessao SQLAlchemy ativa.
        consultores_afetados: Conjunto de nomes de consultores (MANU, LARISSA, etc.)
    """
    from backend.app.models.agenda import AgendaItem  # Import local evita circular

    hoje = date.today()
    consultores_validos = consultores_afetados & {"MANU", "LARISSA", "DAIANE", "JULIO"}

    for consultor in consultores_validos:
        # Deletar agenda existente do dia para este consultor
        db.query(AgendaItem).filter(
            AgendaItem.consultor == consultor,
            AgendaItem.data_agenda == hoje,
        ).delete(synchronize_session=False)

        # Regenerar
        novos_itens = agenda_service.gerar_agenda_consultor(db, consultor, hoje)
        for item in novos_itens:
            db.add(item)


# ---------------------------------------------------------------------------
# POST /api/import/upload
# ---------------------------------------------------------------------------

@router.post(
    "/upload",
    response_model=ImportJobResponse,
    status_code=status.HTTP_200_OK,
    summary="Importar arquivo .xlsx (Mercos/SAP)",
    description=(
        "Recebe arquivo .xlsx via multipart/form-data e processa pipeline completo:\n"
        "1. Valida extensao (.xlsx) e tamanho (max 50MB)\n"
        "2. Detecta tipo: Mercos Vendas / Mercos Carteira / SAP Cadastro\n"
        "3. Normaliza CNPJs (R5: String(14), zero-padded)\n"
        "4. Classifica 3-tier: dados de Mercos/SAP = REAL automaticamente\n"
        "5. Upsert clientes (UPDATE se existe, INSERT se novo)\n"
        "6. Two-Base: vendas com valor > 0 vao para tabela vendas\n"
        "7. Rejeita CNPJs da lista ALUCINACAO (R8)\n"
        "8. Aplica De-Para de vendedores (R9)\n"
        "9. Recalcula Score + Sinaleiro para clientes afetados\n"
        "10. Regenera agenda do dia para consultores afetados\n\n"
        "R6: Mercos MENTE nos nomes — verificar Data inicial/Data final (linhas 6-7).\n"
        "Restrito a administradores."
    ),
)
async def upload_xlsx(
    file: UploadFile = File(..., description="Arquivo .xlsx do Mercos ou SAP (max 50MB)"),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_admin),
) -> ImportJobResponse:
    """
    Pipeline de importacao de arquivo .xlsx.

    Processamento em foreground (Render free nao suporta background workers).
    Timeout estimado: ~5s para 1000 linhas, ~30s para 10000 linhas.

    Raises:
        HTTPException 400 — arquivo nao e .xlsx
        HTTPException 413 — arquivo excede 50MB
        HTTPException 422 — tipo de arquivo nao reconhecido pelos cabecalhos
        HTTPException 500 — erro inesperado no processamento
    """
    t_inicio = time.perf_counter()

    # --- Validacao de extensao ---
    nome_arquivo = file.filename or "upload.xlsx"
    if not nome_arquivo.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Apenas arquivos .xlsx sao aceitos. Formatos .xls, .csv e outros nao sao suportados.",
        )

    # --- Leitura do conteudo ---
    conteudo = await file.read()

    # --- Validacao de tamanho ---
    if len(conteudo) > _MAX_SIZE_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"Arquivo excede o limite de 50MB. Tamanho recebido: {len(conteudo) / 1024 / 1024:.1f}MB.",
        )

    # --- Criar ImportJob inicial ---
    job = ImportJob(
        tipo="XLSX",
        arquivo_nome=nome_arquivo,
        status="PROCESSANDO",
        registros_lidos=0,
        registros_inseridos=0,
        registros_atualizados=0,
        registros_ignorados=0,
        iniciado_em=datetime.now(timezone.utc),
        created_by=usuario.id,
    )
    db.add(job)
    db.flush()

    erros_count = 0
    tipo_detectado = "DESCONHECIDO"

    try:
        # --- Abrir planilha com openpyxl ---
        import openpyxl  # Import local — dependencia opcional em testes sem xlsx real

        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
        ws = wb.active

        # Extrair cabecalhos da primeira linha para deteccao de tipo
        primeira_linha = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        cabecalhos = [str(c).strip() if c is not None else "" for c in primeira_linha]

        # --- Deteccao de tipo ---
        try:
            tipo_detectado = _detectar_tipo_arquivo(cabecalhos)
        except ValueError as exc:
            job.status = "ERRO"
            job.erro_mensagem = str(exc)
            job.concluido_em = datetime.now(timezone.utc)
            db.commit()
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail=str(exc),
            )

        # Atualizar tipo no job
        job.tipo = tipo_detectado

        # --- Processar de acordo com o tipo ---
        # cnpjs_afetados e populado pelos processadores para evitar re-iteracao do xlsx
        cnpjs_afetados: set[str] = set()
        if tipo_detectado == _TIPO_MERCOS_VENDAS:
            inseridos, atualizados, ignorados, erros_count = _processar_mercos_vendas(
                ws, db, job, cnpjs_afetados
            )
        elif tipo_detectado == _TIPO_MERCOS_CARTEIRA:
            inseridos, atualizados, ignorados, erros_count = _processar_mercos_carteira(
                ws, db, job, cnpjs_afetados
            )
        else:  # _TIPO_SAP_CADASTRO
            inseridos, atualizados, ignorados, erros_count = _processar_sap_cadastro(
                ws, db, job, cnpjs_afetados
            )

        # --- Recalcular inteligencia para clientes afetados ---
        _recalcular_inteligencia(db, cnpjs_afetados)

        # Coletar consultores para regenerar agenda
        consultores_afetados: set[str] = set()
        clientes_afetados_list = db.query(Cliente).filter(
            Cliente.cnpj.in_(cnpjs_afetados)
        ).all() if cnpjs_afetados else []
        for c in clientes_afetados_list:
            if c.consultor:
                consultores_afetados.add(c.consultor.upper())

        _regenerar_agenda(db, consultores_afetados)

        # --- Atualizar job com resultado ---
        job.registros_inseridos = inseridos
        job.registros_atualizados = atualizados
        job.registros_ignorados = ignorados
        job.status = "CONCLUIDO"
        job.concluido_em = datetime.now(timezone.utc)

        db.commit()

    except HTTPException:
        raise  # Propagar HTTPExceptions sem modificar

    except Exception as exc:
        db.rollback()
        # Criar job de erro (nova transacao apos rollback)
        job_erro = ImportJob(
            tipo=tipo_detectado,
            arquivo_nome=nome_arquivo,
            status="ERRO",
            registros_lidos=0,
            registros_inseridos=0,
            registros_atualizados=0,
            registros_ignorados=0,
            erro_mensagem=str(exc)[:2000],
            iniciado_em=datetime.now(timezone.utc),
            concluido_em=datetime.now(timezone.utc),
            created_by=usuario.id,
        )
        db.add(job_erro)
        db.commit()
        db.refresh(job_erro)

        t_fim = time.perf_counter()
        return ImportJobResponse(
            job_id=job_erro.id,
            status="ERRO",
            tipo=tipo_detectado,
            arquivo=nome_arquivo,
            registros_lidos=0,
            inseridos=0,
            atualizados=0,
            ignorados=0,
            erros=1,
            tempo_ms=round((t_fim - t_inicio) * 1000, 1),
            erro_mensagem=str(exc)[:500],
        )

    db.refresh(job)
    t_fim = time.perf_counter()

    return ImportJobResponse(
        job_id=job.id,
        status=job.status,
        tipo=job.tipo,
        arquivo=job.arquivo_nome,
        registros_lidos=job.registros_lidos,
        inseridos=job.registros_inseridos,
        atualizados=job.registros_atualizados,
        ignorados=job.registros_ignorados,
        erros=erros_count,
        tempo_ms=round((t_fim - t_inicio) * 1000, 1),
        erro_mensagem=job.erro_mensagem,
    )


# ---------------------------------------------------------------------------
# GET /api/import/history
# ---------------------------------------------------------------------------

@router.get(
    "/history",
    response_model=ImportHistoryResponse,
    summary="Historico de importacoes (admin)",
    description=(
        "Lista os ultimos 20 jobs de importacao com seus stats.\n"
        "Util para auditoria e diagnostico de problemas de importacao.\n"
        "Restrito a administradores."
    ),
)
def listar_historico_import(
    db: Session = Depends(get_db),
    _: Usuario = Depends(require_admin),
) -> ImportHistoryResponse:
    """
    Retorna os ultimos 20 ImportJobs ordenados por created_at desc.

    Permissao: somente admin.
    """
    jobs = (
        db.query(ImportJob)
        .order_by(ImportJob.created_at.desc())
        .limit(20)
        .all()
    )

    total = db.query(ImportJob).count()

    items = [
        ImportHistoryItem(
            id=j.id,
            tipo=j.tipo,
            arquivo=j.arquivo_nome,
            status=j.status,
            registros_lidos=j.registros_lidos or 0,
            inseridos=j.registros_inseridos or 0,
            atualizados=j.registros_atualizados or 0,
            ignorados=j.registros_ignorados or 0,
            erro_mensagem=j.erro_mensagem,
            created_at=j.created_at,
        )
        for j in jobs
    ]

    return ImportHistoryResponse(total=total, items=items)

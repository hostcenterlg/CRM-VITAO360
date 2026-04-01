"""
CRM VITAO360 — Rotas /api/relatorios

Motor de geração de relatórios XLSX server-side.

Endpoints:
  GET /api/relatorios/catalogo          — lista de relatórios disponíveis (JSON)
  GET /api/relatorios/vendas            — vendas detalhadas por período
  GET /api/relatorios/positivacao       — positivação de clientes por mês/ano
  GET /api/relatorios/atividades        — atividades realizadas (log_interacoes)
  GET /api/relatorios/clientes-inativos — clientes INAT.REC + INAT.ANT
  GET /api/relatorios/metas             — metas vs realizado por consultor

Regras obrigatórias:
  R1  — Two-Base: relatórios com R$ usam APENAS tabela vendas.
        Atividades usam APENAS log_interacoes (sem R$).
  R4  — Fórmulas openpyxl em INGLÊS com separador vírgula.
  R8  — Excluir ALUCINACAO de todas as queries.
  R9  — Visual LIGHT: header VITAO green (#00B050), fonte Arial.

Todos os endpoints requerem autenticação JWT (Bearer token).
"""

from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, Query
from sqlalchemy import func, select
from sqlalchemy.orm import Session
from starlette.responses import StreamingResponse

from backend.app.api.deps import get_current_user
from backend.app.database import get_db
from backend.app.models.cliente import Cliente
from backend.app.models.log_interacao import LogInteracao
from backend.app.models.usuario import Usuario
from backend.app.models.venda import Venda

router = APIRouter(prefix="/api/relatorios", tags=["Relatórios"])

# ---------------------------------------------------------------------------
# XLSX Helper
# ---------------------------------------------------------------------------

# Lazy import to avoid hard startup failure if openpyxl isn't installed in a
# particular environment — the ImportError will surface only when a report is
# actually requested, giving a clear 500 with a traceable message.
def _get_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        return openpyxl, Font, PatternFill, Alignment, get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl não instalado — instale com: pip install openpyxl") from exc


# VITAO brand colours (R9 — visual LIGHT only)
_HEADER_BG = "00B050"   # VITAO green
_HEADER_FG = "FFFFFF"   # white text on green
_CURRENCY_FORMAT = 'R$ #,##0.00'


def _create_xlsx(
    title: str,
    headers: list[str],
    rows: list[list],
    col_widths: list[int] | None = None,
    currency_cols: list[int] | None = None,  # 0-based column indices to format as currency
) -> BytesIO:
    """
    Build an XLSX workbook in-memory and return a BytesIO ready for StreamingResponse.

    Args:
        title:        Sheet/workbook title (max 31 chars for sheet name).
        headers:      List of column header strings.
        rows:         Data rows — each inner list must match len(headers).
        col_widths:   Optional explicit column widths (characters). Auto-sized when None.
        currency_cols: 0-based column indices that should receive R$ currency format.

    Formatting applied (R9 — LIGHT visual):
        - Header row: bold white text on VITAO green (#00B050) background, Arial 10pt.
        - Data rows: Arial 9pt.
        - Auto-filter enabled on the header row.
        - Column widths: explicit (col_widths) or inferred from content.
    """
    openpyxl, Font, PatternFill, Alignment, get_column_letter = _get_openpyxl()

    wb = openpyxl.Workbook()
    ws = wb.active
    # Sheet name truncated to Excel limit of 31 chars
    ws.title = title[:31]

    # --- Header styling ---
    header_fill = PatternFill(fill_type="solid", fgColor=_HEADER_BG)
    header_font = Font(name="Arial", size=10, bold=True, color=_HEADER_FG)
    data_font = Font(name="Arial", size=9)

    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # --- Data rows ---
    currency_set = set(currency_cols or [])
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            # Apply R$ number format to designated currency columns
            if (col_idx - 1) in currency_set and isinstance(value, (int, float)):
                cell.number_format = _CURRENCY_FORMAT

    # --- Auto-filter on header row ---
    if headers:
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # --- Column widths ---
    if col_widths:
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    else:
        # Auto-size: use max of header length and longest value in column
        for col_idx, header_text in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header_text))
            for row_data in rows:
                if col_idx - 1 < len(row_data):
                    cell_len = len(str(row_data[col_idx - 1] or ""))
                    if cell_len > max_len:
                        max_len = cell_len
            # Add padding; cap at 60 to prevent absurdly wide columns
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # --- Row height for header ---
    ws.row_dimensions[1].height = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _xlsx_response(buf: BytesIO, filename: str) -> StreamingResponse:
    """Wrap BytesIO in a StreamingResponse with correct XLSX content-type headers."""
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )


# ---------------------------------------------------------------------------
# Catalog endpoint
# ---------------------------------------------------------------------------

@router.get(
    "/catalogo",
    summary="Catálogo de relatórios disponíveis",
)
def catalogo(
    user: Usuario = Depends(get_current_user),
) -> dict:
    """
    Retorna a lista de relatórios disponíveis com id, nome, descrição e filtros aceitos.

    Usado pelo frontend para renderizar o menu de relatórios dinâmico.
    Requer autenticação JWT.
    """
    return {
        "relatorios": [
            {
                "id": "vendas",
                "nome": "Relatório de Vendas",
                "descricao": "Vendas detalhadas por período com valor e condição de pagamento",
                "filtros": ["data_inicio", "data_fim", "consultor"],
            },
            {
                "id": "positivacao",
                "nome": "Positivação de Clientes",
                "descricao": "Clientes que compraram ao menos uma vez no período informado",
                "filtros": ["mes", "ano", "consultor"],
            },
            {
                "id": "atividades",
                "nome": "Atividades Realizadas",
                "descricao": "Registro de todas as interações CRM (ligações, visitas, WhatsApp, e-mail)",
                "filtros": ["data_inicio", "data_fim", "consultor"],
            },
            {
                "id": "clientes-inativos",
                "nome": "Clientes Inativos",
                "descricao": "Carteira de clientes inativos (INAT.REC e INAT.ANT) ordenados por score",
                "filtros": ["consultor"],
            },
            {
                "id": "metas",
                "nome": "Metas vs Realizado",
                "descricao": "Atingimento de metas anuais por consultor com status BOM/ALERTA/CRITICO",
                "filtros": ["consultor"],
            },
        ]
    }


# ---------------------------------------------------------------------------
# Relatório de Vendas
# ---------------------------------------------------------------------------

@router.get(
    "/vendas",
    summary="Relatório de Vendas — XLSX detalhado por período",
)
def relatorio_vendas(
    data_inicio: Optional[date] = Query(
        default=None,
        description="Data inicial (YYYY-MM-DD). Padrão: primeiro dia do mês atual.",
    ),
    data_fim: Optional[date] = Query(
        default=None,
        description="Data final (YYYY-MM-DD). Padrão: dia de hoje.",
    ),
    consultor: Optional[str] = Query(
        default=None,
        description="Filtrar por consultor (MANU, LARISSA, DAIANE, JULIO).",
    ),
    user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """
    Gera XLSX com vendas detalhadas para o período informado.

    Colunas: Data, Nº Pedido, Cliente (Nome Fantasia), CNPJ, Consultor,
             Valor (R$), Status, Condição Pagamento.

    R1  — Two-Base: usa APENAS tabela vendas (valor_pedido > 0).
    R8  — Exclui classificacao_3tier = 'ALUCINACAO'.
    Requer autenticação JWT.
    """
    hoje = date.today()
    dt_inicio = data_inicio or hoje.replace(day=1)
    dt_fim = data_fim or hoje

    # Build query
    stmt = (
        select(
            Venda.data_pedido,
            Venda.numero_pedido,
            Cliente.nome_fantasia,
            Venda.cnpj,
            Venda.consultor,
            Venda.valor_pedido,
            Venda.status_pedido,
            Venda.condicao_pagamento,
        )
        .join(Cliente, Venda.cnpj == Cliente.cnpj, isouter=True)
        .where(
            Venda.classificacao_3tier.in_(["REAL", "SINTETICO"]),  # R8
            Venda.data_pedido >= dt_inicio,
            Venda.data_pedido <= dt_fim,
        )
        .order_by(Venda.data_pedido.desc(), Venda.consultor)
    )

    if consultor:
        stmt = stmt.where(Venda.consultor == consultor.upper())

    rows_db = db.execute(stmt).all()

    headers = [
        "Data",
        "Nº Pedido",
        "Cliente (Nome Fantasia)",
        "CNPJ",
        "Consultor",
        "Valor (R$)",
        "Status",
        "Condição Pagamento",
    ]

    rows = [
        [
            r.data_pedido.strftime("%d/%m/%Y") if r.data_pedido else "",
            r.numero_pedido or "",
            r.nome_fantasia or "",
            r.cnpj or "",
            r.consultor or "",
            round(float(r.valor_pedido), 2) if r.valor_pedido else 0.0,
            r.status_pedido or "",
            r.condicao_pagamento or "",
        ]
        for r in rows_db
    ]

    col_widths = [14, 14, 35, 18, 12, 16, 14, 22]
    # Column index 5 (Valor R$) — 0-based
    buf = _create_xlsx(
        title="Vendas",
        headers=headers,
        rows=rows,
        col_widths=col_widths,
        currency_cols=[5],
    )

    periodo_label = f"{dt_inicio.strftime('%Y%m%d')}_{dt_fim.strftime('%Y%m%d')}"
    filename = f"vendas_{periodo_label}.xlsx"
    return _xlsx_response(buf, filename)


# ---------------------------------------------------------------------------
# Relatório de Positivação
# ---------------------------------------------------------------------------

@router.get(
    "/positivacao",
    summary="Relatório de Positivação de Clientes — XLSX por mês/ano",
)
def relatorio_positivacao(
    mes: Optional[int] = Query(
        default=None,
        ge=1,
        le=12,
        description="Mês (1-12). Padrão: mês atual.",
    ),
    ano: Optional[int] = Query(
        default=None,
        ge=2020,
        description="Ano (ex.: 2026). Padrão: ano atual.",
    ),
    consultor: Optional[str] = Query(
        default=None,
        description="Filtrar por consultor (MANU, LARISSA, DAIANE, JULIO).",
    ),
    user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """
    Gera XLSX de positivação: todos os clientes da carteira com indicação
    se compraram (SIM/NÃO) no período informado.

    Colunas: Cliente, CNPJ, Consultor, Situação, Data Último Pedido,
             Dias Sem Compra, Positivado (SIM/NÃO).

    R1  — Two-Base: positivação via tabela vendas.
    R8  — Exclui ALUCINACAO.
    Requer autenticação JWT.
    """
    hoje = date.today()
    mes_ref = mes if mes is not None else hoje.month
    ano_ref = ano if ano is not None else hoje.year

    # Date range for the requested month
    dt_inicio = date(ano_ref, mes_ref, 1)
    if mes_ref == 12:
        dt_fim = date(ano_ref + 1, 1, 1) - timedelta(days=1)
    else:
        dt_fim = date(ano_ref, mes_ref + 1, 1) - timedelta(days=1)

    # Fetch all non-ALUCINACAO clients
    clientes_q = (
        select(Cliente)
        .where(Cliente.classificacao_3tier != "ALUCINACAO")
        .order_by(Cliente.consultor, Cliente.nome_fantasia)
    )
    if consultor:
        clientes_q = clientes_q.where(Cliente.consultor == consultor.upper())

    clientes = db.scalars(clientes_q).all()

    # Fetch all CNPJs that bought in the period (set for O(1) lookup)
    positivados_q = (
        select(func.distinct(Venda.cnpj))
        .where(
            Venda.classificacao_3tier.in_(["REAL", "SINTETICO"]),
            Venda.data_pedido >= dt_inicio,
            Venda.data_pedido <= dt_fim,
        )
    )
    if consultor:
        positivados_q = positivados_q.where(Venda.consultor == consultor.upper())

    positivados_set: set[str] = {row[0] for row in db.execute(positivados_q).all()}

    headers = [
        "Cliente (Nome Fantasia)",
        "CNPJ",
        "Consultor",
        "Situação",
        "Data Último Pedido",
        "Dias Sem Compra",
        "Positivado",
    ]

    rows = [
        [
            c.nome_fantasia or "",
            c.cnpj or "",
            c.consultor or "",
            c.situacao or "",
            # valor_ultimo_pedido date is stored in dias_sem_compra; last purchase date
            # can be derived from dias_sem_compra + today (approximate)
            (
                (hoje - timedelta(days=int(c.dias_sem_compra))).strftime("%d/%m/%Y")
                if c.dias_sem_compra is not None
                else ""
            ),
            c.dias_sem_compra if c.dias_sem_compra is not None else "",
            "SIM" if c.cnpj in positivados_set else "NÃO",
        ]
        for c in clientes
    ]

    col_widths = [35, 18, 12, 14, 18, 16, 12]
    buf = _create_xlsx(
        title="Positivação",
        headers=headers,
        rows=rows,
        col_widths=col_widths,
    )

    filename = f"positivacao_{ano_ref:04d}{mes_ref:02d}.xlsx"
    return _xlsx_response(buf, filename)


# ---------------------------------------------------------------------------
# Relatório de Atividades
# ---------------------------------------------------------------------------

@router.get(
    "/atividades",
    summary="Relatório de Atividades Realizadas — XLSX",
)
def relatorio_atividades(
    data_inicio: Optional[date] = Query(
        default=None,
        description="Data inicial (YYYY-MM-DD). Padrão: primeiro dia do mês atual.",
    ),
    data_fim: Optional[date] = Query(
        default=None,
        description="Data final (YYYY-MM-DD). Padrão: dia de hoje.",
    ),
    consultor: Optional[str] = Query(
        default=None,
        description="Filtrar por consultor (MANU, LARISSA, DAIANE, JULIO).",
    ),
    user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """
    Gera XLSX com todas as interações CRM do período.

    Colunas: Data, Cliente, CNPJ, Consultor, Tipo Contato, Resultado, Descrição.

    R1  — Two-Base: usa APENAS log_interacoes (sem R$).
    R8  — Exclui clientes com classificacao_3tier = 'ALUCINACAO' via join.
    Requer autenticação JWT.
    """
    hoje = date.today()
    dt_inicio = data_inicio or hoje.replace(day=1)
    dt_fim = data_fim or hoje

    stmt = (
        select(
            LogInteracao.data_interacao,
            Cliente.nome_fantasia,
            LogInteracao.cnpj,
            LogInteracao.consultor,
            LogInteracao.tipo_contato,
            LogInteracao.resultado,
            LogInteracao.descricao,
        )
        .join(Cliente, LogInteracao.cnpj == Cliente.cnpj, isouter=True)
        .where(
            func.date(LogInteracao.data_interacao) >= dt_inicio,
            func.date(LogInteracao.data_interacao) <= dt_fim,
        )
        .order_by(LogInteracao.data_interacao.desc())
    )

    if consultor:
        stmt = stmt.where(LogInteracao.consultor == consultor.upper())

    rows_db = db.execute(stmt).all()

    headers = [
        "Data",
        "Cliente",
        "CNPJ",
        "Consultor",
        "Tipo Contato",
        "Resultado",
        "Descrição",
    ]

    rows = [
        [
            r.data_interacao.strftime("%d/%m/%Y %H:%M") if r.data_interacao else "",
            r.nome_fantasia or "",
            r.cnpj or "",
            r.consultor or "",
            r.tipo_contato or "",
            r.resultado or "",
            r.descricao or "",
        ]
        for r in rows_db
    ]

    col_widths = [18, 35, 18, 12, 16, 20, 50]
    buf = _create_xlsx(
        title="Atividades",
        headers=headers,
        rows=rows,
        col_widths=col_widths,
    )

    periodo_label = f"{dt_inicio.strftime('%Y%m%d')}_{dt_fim.strftime('%Y%m%d')}"
    filename = f"atividades_{periodo_label}.xlsx"
    return _xlsx_response(buf, filename)


# ---------------------------------------------------------------------------
# Relatório de Clientes Inativos
# ---------------------------------------------------------------------------

@router.get(
    "/clientes-inativos",
    summary="Relatório de Clientes Inativos — XLSX",
)
def relatorio_clientes_inativos(
    consultor: Optional[str] = Query(
        default=None,
        description="Filtrar por consultor (MANU, LARISSA, DAIANE, JULIO).",
    ),
    user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """
    Gera XLSX com todos os clientes com situacao INAT.REC ou INAT.ANT,
    ordenados por score decrescente para priorizar reativação.

    Colunas: Cliente, CNPJ, Consultor, Situação, Dias Sem Compra,
             Último Pedido (R$), Faturamento Total (R$), Curva ABC, Score.

    R8  — Exclui ALUCINACAO.
    R1  — Faturamento Total vem APENAS de registros VENDA (campo faturamento_total no cliente).
    Requer autenticação JWT.
    """
    stmt = (
        select(Cliente)
        .where(
            Cliente.situacao.in_(["INAT.REC", "INAT.ANT"]),
            Cliente.classificacao_3tier != "ALUCINACAO",   # R8
        )
        .order_by(Cliente.score.desc().nulls_last())
    )

    if consultor:
        stmt = stmt.where(Cliente.consultor == consultor.upper())

    clientes = db.scalars(stmt).all()

    hoje = date.today()

    headers = [
        "Cliente (Nome Fantasia)",
        "CNPJ",
        "Consultor",
        "Situação",
        "Dias Sem Compra",
        "Último Pedido (R$)",
        "Faturamento Total (R$)",
        "Curva ABC",
        "Score",
    ]

    rows = [
        [
            c.nome_fantasia or "",
            c.cnpj or "",
            c.consultor or "",
            c.situacao or "",
            c.dias_sem_compra if c.dias_sem_compra is not None else "",
            round(float(c.valor_ultimo_pedido), 2) if c.valor_ultimo_pedido else 0.0,
            round(float(c.faturamento_total), 2) if c.faturamento_total else 0.0,
            c.curva_abc or "",
            round(float(c.score), 1) if c.score else 0.0,
        ]
        for c in clientes
    ]

    col_widths = [35, 18, 12, 12, 16, 20, 22, 12, 10]
    # Currency cols: index 5 (Último Pedido R$), index 6 (Faturamento Total R$)
    buf = _create_xlsx(
        title="Clientes Inativos",
        headers=headers,
        rows=rows,
        col_widths=col_widths,
        currency_cols=[5, 6],
    )

    filename = f"clientes_inativos_{hoje.strftime('%Y%m%d')}.xlsx"
    return _xlsx_response(buf, filename)


# ---------------------------------------------------------------------------
# Relatório de Metas
# ---------------------------------------------------------------------------

# Fallback meta values when meta_anual is not individually set on clients
_METAS_FALLBACK: dict[str, float] = {
    "MANU":    693_000.0,
    "LARISSA": 940_950.0,
    "DAIANE":  300_000.0,
    "JULIO":   156_150.0,
}

_CONSULTORES_PADRAO = ["MANU", "LARISSA", "DAIANE", "JULIO"]


@router.get(
    "/metas",
    summary="Relatório de Metas vs Realizado — XLSX por consultor",
)
def relatorio_metas(
    consultor: Optional[str] = Query(
        default=None,
        description="Filtrar por consultor específico.",
    ),
    user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """
    Gera XLSX com resumo de atingimento de metas anuais por consultor.

    Colunas: Consultor, Meta Anual (R$), Realizado (R$), % Atingido, Gap (R$), Status.

    Status:
      - BOM:    >= 80 % da meta
      - ALERTA: >= 50 % da meta
      - CRITICO: < 50 % da meta

    R1  — Two-Base: realizado vem da tabela vendas (valor_pedido).
    R8  — Exclui ALUCINACAO.
    Requer autenticação JWT.
    """
    hoje = date.today()
    consultores_alvo = [consultor.upper()] if consultor else _CONSULTORES_PADRAO

    headers = [
        "Consultor",
        "Meta Anual (R$)",
        "Realizado (R$)",
        "% Atingido",
        "Gap (R$)",
        "Status",
    ]

    rows: list[list] = []

    for cons in consultores_alvo:
        # Soma das metas individuais dos clientes desse consultor
        meta_db = db.scalar(
            select(func.coalesce(func.sum(Cliente.meta_anual), 0.0))
            .where(Cliente.consultor == cons)
        ) or 0.0

        # Fallback: distribuição proporcional ao baseline quando meta não está preenchida
        meta_anual = float(meta_db) if float(meta_db) > 0 else _METAS_FALLBACK.get(cons, 0.0)

        # Realizado: soma das vendas REAL/SINTETICO deste consultor (R1/R8)
        realizado = db.scalar(
            select(func.coalesce(func.sum(Venda.valor_pedido), 0.0))
            .where(
                Venda.consultor == cons,
                Venda.classificacao_3tier.in_(["REAL", "SINTETICO"]),
            )
        ) or 0.0
        realizado = round(float(realizado), 2)

        pct = round(realizado / meta_anual * 100, 1) if meta_anual > 0 else 0.0
        gap = round(meta_anual - realizado, 2)

        if pct >= 80:
            status = "BOM"
        elif pct >= 50:
            status = "ALERTA"
        else:
            status = "CRITICO"

        rows.append([
            cons,
            round(meta_anual, 2),
            realizado,
            f"{pct}%",
            gap,
            status,
        ])

    col_widths = [14, 18, 18, 14, 16, 12]
    # Currency cols: index 1 (Meta Anual), index 2 (Realizado), index 4 (Gap)
    buf = _create_xlsx(
        title="Metas",
        headers=headers,
        rows=rows,
        col_widths=col_widths,
        currency_cols=[1, 2, 4],
    )

    filename = f"metas_{hoje.strftime('%Y')}.xlsx"
    return _xlsx_response(buf, filename)

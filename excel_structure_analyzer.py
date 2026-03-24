#!/usr/bin/env python3
"""
Excel Structure Comparison Analyzer
Compares V13 vs V31 (V12 POPULADO) workbooks in detail
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Files to analyze
V13_FILE = Path("c:\\Users\\User\\OneDrive\\Documentos\\GitHub\\CRM-VITAO360\\data\\output\\phase10\\CRM_VITAO360_V13_FINAL.xlsx")
V31_FILE = Path("c:\\Users\\User\\OneDrive\\Área de Trabalho\\auditoria conversas sobre agenda atendimento draft 2\\CRM_V12_POPULADO_V31 (1).xlsx edição update.xlsx")
OUTPUT_JSON = Path("c:\\Users\\User\\OneDrive\\Documentos\\GitHub\\CRM-VITAO360\\data\\output\\phase10\\v31_vs_v13_comparison.json")

def extract_header_rows(ws, num_rows=3):
    """Extract header rows content"""
    headers = []
    for row_num in range(1, min(num_rows + 1, ws.max_row + 1)):
        row_data = []
        for col_num in range(1, min(ws.max_column + 1, 51)):  # First 50 columns
            cell = ws.cell(row_num, col_num)
            row_data.append({
                "col": get_column_letter(col_num),
                "value": str(cell.value) if cell.value is not None else "",
                "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
            })
        headers.append(row_data)
    return headers

def count_formulas(ws):
    """Count formulas in worksheet"""
    formula_count = 0
    formula_by_col = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1
                col_letter = get_column_letter(cell.column)
                formula_by_col[col_letter] = formula_by_col.get(col_letter, 0) + 1
    return formula_count, formula_by_col

def get_outline_levels(ws, dimension_type="column"):
    """Extract column/row grouping outline levels"""
    outlines = {}
    if dimension_type == "column":
        for col_letter, col_dimension in ws.column_dimensions.items():
            if col_dimension.outlineLevel:
                outlines[col_letter] = col_dimension.outlineLevel
    else:  # row
        for row_num, row_dimension in ws.row_dimensions.items():
            if row_dimension.outlineLevel:
                outlines[row_num] = row_dimension.outlineLevel
    return outlines

def get_hidden_elements(ws, dimension_type="column"):
    """Find hidden columns/rows"""
    hidden = []
    if dimension_type == "column":
        for col_letter, col_dimension in ws.column_dimensions.items():
            if col_dimension.hidden:
                hidden.append(col_letter)
    else:  # row
        for row_num, row_dimension in ws.row_dimensions.items():
            if row_dimension.hidden:
                hidden.append(row_num)
    return hidden

def get_column_widths(ws, num_cols=30):
    """Get column widths for first N columns"""
    widths = {}
    for col_num in range(1, num_cols + 1):
        col_letter = get_column_letter(col_num)
        col_dim = ws.column_dimensions.get(col_letter)
        if col_dim and col_dim.width:
            widths[col_letter] = col_dim.width
    return widths

def analyze_worksheet(ws, sheet_name):
    """Complete analysis of a single worksheet"""
    formula_count, formulas_by_col = count_formulas(ws)
    col_outlines = get_outline_levels(ws, "column")
    row_outlines = get_outline_levels(ws, "row")
    hidden_cols = get_hidden_elements(ws, "column")
    hidden_rows = get_hidden_elements(ws, "row")
    col_widths = get_column_widths(ws, 30)
    headers = extract_header_rows(ws, 3)

    # Data validations
    data_validations = len(ws.data_validations.dataValidation) if ws.data_validations else 0

    # Conditional formatting
    conditional_formats = len(ws.conditional_formatting._cf_rules) if ws.conditional_formatting else 0

    # Merged cells
    merged_cells_list = list(ws.merged_cells.ranges) if ws.merged_cells else []

    # Freeze panes
    freeze_panes = ws.freeze_panes if hasattr(ws, 'freeze_panes') else None

    # Auto filters
    auto_filter = ws.auto_filter.ref if ws.auto_filter else None

    return {
        "sheet_name": sheet_name,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "formula_count": formula_count,
        "formulas_by_column": formulas_by_col,
        "column_outlines": col_outlines,
        "row_outlines": row_outlines,
        "hidden_columns": hidden_cols,
        "hidden_rows": hidden_rows,
        "column_widths": col_widths,
        "data_validations_count": data_validations,
        "conditional_formatting_count": conditional_formats,
        "merged_cells_count": len(merged_cells_list),
        "merged_cells_sample": [str(cell) for cell in merged_cells_list[:10]],
        "freeze_panes": str(freeze_panes),
        "auto_filter": auto_filter,
        "headers_rows_1_3": headers
    }

def analyze_workbook(file_path):
    """Analyze entire workbook"""
    print(f"\n{'='*80}")
    print(f"Analyzing: {file_path.name}")
    print(f"{'='*80}\n")

    try:
        # Load with data_only=False to see formulas
        wb = load_workbook(file_path, data_only=False)

        sheet_names = wb.sheetnames
        print(f"Total sheets: {len(sheet_names)}")
        print(f"Sheet names: {sheet_names}\n")

        sheets_data = {}
        for sheet_name in sheet_names:
            print(f"Processing sheet: {sheet_name}...", end=" ")
            ws = wb[sheet_name]
            sheet_info = analyze_worksheet(ws, sheet_name)
            sheets_data[sheet_name] = sheet_info
            print(f"({sheet_info['max_row']} rows x {sheet_info['max_column']} cols, {sheet_info['formula_count']} formulas)")

        return {
            "file_name": file_path.name,
            "file_path": str(file_path),
            "total_sheets": len(sheet_names),
            "sheet_names": sheet_names,
            "analysis_timestamp": datetime.now().isoformat(),
            "sheets": sheets_data
        }
    except Exception as e:
        print(f"ERROR analyzing {file_path}: {e}")
        return None

def compare_workbooks(v13_data, v31_data):
    """Compare two workbook analyses"""
    comparison = {
        "timestamp": datetime.now().isoformat(),
        "v13_file": v13_data["file_name"],
        "v31_file": v31_data["file_name"],
        "summary": {},
        "sheet_comparison": {},
        "sheets_only_in_v13": [],
        "sheets_only_in_v31": [],
        "detailed_comparisons": {}
    }

    v13_sheets = set(v13_data["sheet_names"])
    v31_sheets = set(v31_data["sheet_names"])

    # Summary
    comparison["summary"] = {
        "v13_total_sheets": v13_data["total_sheets"],
        "v31_total_sheets": v31_data["total_sheets"],
        "sheets_in_both": list(v13_sheets & v31_sheets),
        "common_sheet_count": len(v13_sheets & v31_sheets)
    }

    # Sheets only in V13
    only_v13 = v13_sheets - v31_sheets
    comparison["sheets_only_in_v13"] = list(only_v13)

    # Sheets only in V31
    only_v31 = v31_sheets - v13_sheets
    comparison["sheets_only_in_v31"] = list(only_v31)

    # Detailed comparison for common sheets
    for sheet_name in sorted(v13_sheets & v31_sheets):
        v13_sheet = v13_data["sheets"][sheet_name]
        v31_sheet = v31_data["sheets"][sheet_name]

        comparison["detailed_comparisons"][sheet_name] = {
            "v13": {
                "rows": v13_sheet["max_row"],
                "columns": v13_sheet["max_column"],
                "formulas": v13_sheet["formula_count"],
                "column_outlines": v13_sheet["column_outlines"],
                "row_outlines": v13_sheet["row_outlines"],
                "hidden_columns": v13_sheet["hidden_columns"],
                "hidden_rows": v13_sheet["hidden_rows"],
                "data_validations": v13_sheet["data_validations_count"],
                "conditional_formatting": v13_sheet["conditional_formatting_count"],
                "merged_cells": v13_sheet["merged_cells_count"],
                "auto_filter": v13_sheet["auto_filter"],
                "freeze_panes": v13_sheet["freeze_panes"]
            },
            "v31": {
                "rows": v31_sheet["max_row"],
                "columns": v31_sheet["max_column"],
                "formulas": v31_sheet["formula_count"],
                "column_outlines": v31_sheet["column_outlines"],
                "row_outlines": v31_sheet["row_outlines"],
                "hidden_columns": v31_sheet["hidden_columns"],
                "hidden_rows": v31_sheet["hidden_rows"],
                "data_validations": v31_sheet["data_validations_count"],
                "conditional_formatting": v31_sheet["conditional_formatting_count"],
                "merged_cells": v31_sheet["merged_cells_count"],
                "auto_filter": v31_sheet["auto_filter"],
                "freeze_panes": v31_sheet["freeze_panes"]
            },
            "differences": {
                "row_difference": v31_sheet["max_row"] - v13_sheet["max_row"],
                "column_difference": v31_sheet["max_column"] - v13_sheet["max_column"],
                "formula_difference": v31_sheet["formula_count"] - v13_sheet["formula_count"],
                "grouping_structure_changed": (
                    v31_sheet["column_outlines"] != v13_sheet["column_outlines"] or
                    v31_sheet["row_outlines"] != v13_sheet["row_outlines"]
                ),
                "filter_changed": v31_sheet["auto_filter"] != v13_sheet["auto_filter"],
                "freeze_panes_changed": v31_sheet["freeze_panes"] != v13_sheet["freeze_panes"]
            }
        }

        # Include header comparison
        if sheet_name in v13_data["sheets"] and sheet_name in v31_data["sheets"]:
            comparison["detailed_comparisons"][sheet_name]["headers_v13"] = v13_sheet["headers_rows_1_3"]
            comparison["detailed_comparisons"][sheet_name]["headers_v31"] = v31_sheet["headers_rows_1_3"]

    # Document sheets only in V31 (fully)
    for sheet_name in sorted(only_v31):
        v31_sheet = v31_data["sheets"][sheet_name]
        comparison["v31_exclusive_sheets"] = comparison.get("v31_exclusive_sheets", {})
        comparison["v31_exclusive_sheets"][sheet_name] = {
            "rows": v31_sheet["max_row"],
            "columns": v31_sheet["max_column"],
            "formulas": v31_sheet["formula_count"],
            "column_outlines": v31_sheet["column_outlines"],
            "row_outlines": v31_sheet["row_outlines"],
            "headers_rows_1_3": v31_sheet["headers_rows_1_3"],
            "data_validations": v31_sheet["data_validations_count"],
            "conditional_formatting": v31_sheet["conditional_formatting_count"]
        }

    return comparison

def print_comparison_report(comparison):
    """Print formatted comparison report"""
    print("\n" + "="*100)
    print("EXCEL STRUCTURE COMPARISON REPORT: V13 vs V31 (V12 POPULADO)")
    print("="*100)

    print("\n[SUMMARY]")
    print(f"  V13 Total Sheets: {comparison['summary']['v13_total_sheets']}")
    print(f"  V31 Total Sheets: {comparison['summary']['v31_total_sheets']}")
    print(f"  Sheets in Common: {comparison['summary']['common_sheet_count']}")

    if comparison.get("sheets_only_in_v13"):
        print(f"\n[SHEETS ONLY IN V13] ({len(comparison['sheets_only_in_v13'])} sheets)")
        for sheet in comparison["sheets_only_in_v13"]:
            print(f"  - {sheet}")

    if comparison.get("sheets_only_in_v31"):
        print(f"\n[SHEETS ONLY IN V31] ({len(comparison['sheets_only_in_v31'])} sheets) *** POTENTIAL MISSING FEATURES ***")
        for sheet in comparison["sheets_only_in_v31"]:
            print(f"  - {sheet}")
            if sheet in comparison.get("v31_exclusive_sheets", {}):
                exc_data = comparison["v31_exclusive_sheets"][sheet]
                print(f"      Dimensions: {exc_data['rows']} rows x {exc_data['columns']} cols")
                print(f"      Formulas: {exc_data['formulas']}")
                print(f"      Data Validations: {exc_data['data_validations']}")
                print(f"      Conditional Formatting Rules: {exc_data['conditional_formatting']}")

    print(f"\n[DETAILED SHEET COMPARISONS] ({comparison['summary']['common_sheet_count']} common sheets)")
    print("-" * 100)

    for sheet_name in sorted(comparison["detailed_comparisons"].keys()):
        comp = comparison["detailed_comparisons"][sheet_name]
        diff = comp["differences"]

        print(f"\nSheet: {sheet_name}")
        print(f"  Dimensions:")
        print(f"    V13: {comp['v13']['rows']} rows x {comp['v13']['columns']} cols  →  V31: {comp['v31']['rows']} rows x {comp['v31']['columns']} cols")
        print(f"    Δ Rows: {diff['row_difference']:+d}, Δ Columns: {diff['column_difference']:+d}")

        print(f"  Formulas:")
        print(f"    V13: {comp['v13']['formulas']} formulas  →  V31: {comp['v31']['formulas']} formulas")
        print(f"    Δ: {diff['formula_difference']:+d}")

        print(f"  Grouping/Outline:")
        print(f"    V13 Column Outlines: {len(comp['v13']['column_outlines'])} grouped cols  →  V31: {len(comp['v31']['column_outlines'])} grouped cols")
        print(f"    V13 Row Outlines: {len(comp['v13']['row_outlines'])} grouped rows  →  V31: {len(comp['v31']['row_outlines'])} grouped rows")
        if diff['grouping_structure_changed']:
            print(f"    ⚠️  GROUPING STRUCTURE CHANGED")
            if comp['v31']['column_outlines'] != comp['v13']['column_outlines']:
                print(f"      Column grouping differences detected")
            if comp['v31']['row_outlines'] != comp['v13']['row_outlines']:
                print(f"      Row grouping differences detected")

        print(f"  Filters & Panes:")
        print(f"    V13 Auto-Filter: {comp['v13']['auto_filter']}")
        print(f"    V31 Auto-Filter: {comp['v31']['auto_filter']}")
        if diff['filter_changed']:
            print(f"    ⚠️  AUTO-FILTER CONFIGURATION CHANGED")

        print(f"    V13 Freeze Panes: {comp['v13']['freeze_panes']}")
        print(f"    V31 Freeze Panes: {comp['v31']['freeze_panes']}")
        if diff['freeze_panes_changed']:
            print(f"    ⚠️  FREEZE PANES CONFIGURATION CHANGED")

        print(f"  Data Integrity:")
        print(f"    V13 Data Validations: {comp['v13']['data_validations']}")
        print(f"    V31 Data Validations: {comp['v31']['data_validations']}")
        print(f"    V13 Conditional Formatting: {comp['v13']['conditional_formatting']} rules")
        print(f"    V31 Conditional Formatting: {comp['v31']['conditional_formatting']} rules")
        print(f"    V13 Merged Cells: {comp['v13']['merged_cells']}")
        print(f"    V31 Merged Cells: {comp['v31']['merged_cells']}")

        if comp['v13']['hidden_columns'] or comp['v31']['hidden_columns']:
            print(f"  Hidden Elements:")
            print(f"    V13 Hidden Columns: {comp['v13']['hidden_columns']}")
            print(f"    V31 Hidden Columns: {comp['v31']['hidden_columns']}")

    print("\n" + "="*100)
    print(f"Report saved to: {OUTPUT_JSON}")
    print("="*100 + "\n")

def main():
    print("\n" + "="*100)
    print("EXCEL STRUCTURE COMPARISON: V13 vs V31 (V12 POPULADO)")
    print("="*100)

    # Check files exist
    if not V13_FILE.exists():
        print(f"ERROR: V13 file not found: {V13_FILE}")
        sys.exit(1)

    if not V31_FILE.exists():
        print(f"ERROR: V31 file not found: {V31_FILE}")
        sys.exit(1)

    print(f"\nV13 File: {V13_FILE}")
    print(f"V31 File: {V31_FILE}")

    # Analyze both workbooks
    v13_data = analyze_workbook(V13_FILE)
    v31_data = analyze_workbook(V31_FILE)

    if not v13_data or not v31_data:
        print("ERROR: Failed to analyze one or both files")
        sys.exit(1)

    # Compare
    comparison = compare_workbooks(v13_data, v31_data)

    # Print report
    print_comparison_report(comparison)

    # Save JSON
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(comparison, f, indent=2, ensure_ascii=False, default=str)

    print(f"\n✓ Comparison JSON saved to: {OUTPUT_JSON}")

    return 0

if __name__ == "__main__":
    sys.exit(main())
